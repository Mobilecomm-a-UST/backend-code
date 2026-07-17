from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse
from .models import Employee, DriveTestSurvey
from .serializers import (EmployeeSerializer,
                           DriveTestSurveySerializer,
                           DailyUpdateSerializer,
                           EmployeeDatewiseSerializer)
import openpyxl
from datetime import datetime
from io import BytesIO
import io
import os
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL


class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all().order_by('sr_no')
    serializer_class = EmployeeSerializer


class DriveTestSurveyViewSet(viewsets.ModelViewSet):
    queryset = DriveTestSurvey.objects.all().order_by('employee__sr_no', 'date')
    serializer_class = DriveTestSurveySerializer


class DailyTrackingViewSet(viewsets.ViewSet):

    # GET /api/daily/?date=2026-06-15
    def list(self, request):
        date = request.query_params.get('date', None)

        if not date:
            return Response(
                {"error": "Please provide a date. Example: /api/daily/?date=2026-06-15"},
                status=status.HTTP_400_BAD_REQUEST
            )

        surveys = DriveTestSurvey.objects.filter(
            date=date
        ).order_by('employee__sr_no')

        serializer = DailyUpdateSerializer(surveys, many=True)
        return Response({
            "date": date,
            "total_employees": surveys.count(),
            "data": serializer.data
        })

    # GET /api/daily/by_employee/?emp_code=MCT1979
    @action(detail=False, methods=['get'])
    def by_employee(self, request):
        emp_code = request.query_params.get('emp_code', None)

        if not emp_code:
            return Response(
                {"error": "Please provide emp_code. Example: /api/daily/by_employee/?emp_code=MCT1979"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            employee = Employee.objects.get(emp_code=emp_code)
        except Employee.DoesNotExist:
            return Response(
                {"error": f"Employee {emp_code} not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        surveys = DriveTestSurvey.objects.filter(
            employee=employee
        ).order_by('date')

        serializer = DailyUpdateSerializer(surveys, many=True)
        return Response({
            "emp_code": emp_code,
            "emp_name": employee.emp_name,
            "total_dates": surveys.count(),
            "data": serializer.data
        })

    # POST /api/daily/add/
    @action(detail=False, methods=['post'])
    def add(self, request):
        emp_code = request.data.get('emp_code')
        date = request.data.get('date')

        if not emp_code or not date:
            return Response(
                {"error": "emp_code and date are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            employee = Employee.objects.get(emp_code=emp_code)
        except Employee.DoesNotExist:
            return Response(
                {"error": f"Employee {emp_code} not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        survey, created = DriveTestSurvey.objects.update_or_create(
            employee=employee,
            date=date,
            defaults={
                'working_status': request.data.get('working_status', ''),
                'project': request.data.get('project', ''),
                'activity_assigned': request.data.get('activity_assigned', ''),
                'site_id': request.data.get('site_id', ''),
                'ssid': request.data.get('ssid', ''),
                'activity_status': request.data.get('activity_status', ''),
                'detailed_remarks': request.data.get('detailed_remarks', ''),
                'owner': request.data.get('owner', ''),
            }
        )

        serializer = DailyUpdateSerializer(survey)
        return Response({
            "message": "Created" if created else "Updated",
            "data": serializer.data
        }, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)

    # GET /api/daily/new_employees/?date=2026-06-15
    @action(detail=False, methods=['get'])
    def new_employees(self, request):
        date = request.query_params.get('date', None)

        if not date:
            return Response(
                {"error": "Please provide a date. Example: /api/daily/new_employees/?date=2026-06-15"},
                status=status.HTTP_400_BAD_REQUEST
            )

        surveys = DriveTestSurvey.objects.filter(
            date=date
        ).order_by('employee__sr_no')

        employees = [s.employee for s in surveys]
        serializer = EmployeeSerializer(employees, many=True)

        return Response({
            "date": date,
            "total": len(employees),
            "employees": serializer.data
        })


class EmployeeDatewiseViewSet(viewsets.ViewSet):

    # GET /api/employeedatewise/
    def list(self, request):
        employees = Employee.objects.all().order_by('sr_no')
        serializer = EmployeeDatewiseSerializer(employees, many=True)
        return Response(serializer.data)


class UploadFileView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        # Check if file was uploaded
        if 'file' not in request.FILES:
            return Response(
                {"error": "No file uploaded. Please upload an Excel file."},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES['file']

        # Check if it's an Excel file
        if not file.name.endswith('.xlsx'):
            return Response(
                {"error": "Only .xlsx files are allowed."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()))

            # Handle both sheet names
            if 'Details' in wb.sheetnames:
                ws = wb['Details']
            elif 'Sheet1' in wb.sheetnames:
                ws = wb['Sheet1']
            else:
                return Response(
                    {"error": f"Sheet not found. Available sheets: {wb.sheetnames}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Find date columns from row 1
            row1 = [cell.value for cell in ws[1]]
            date_columns = {}
            for i, val in enumerate(row1):
                if val and isinstance(val, datetime):
                    date_columns[i] = val.date()

            if not date_columns:
                return Response(
                    {"error": "No date columns found in row 1 of the Excel file."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Track results
            new_employees     = 0
            updated_employees = 0
            new_surveys       = 0
            updated_surveys   = 0
            skipped_rows      = 0

            # Loop through data rows from row 3
            for row in ws.iter_rows(min_row=3, values_only=True):
                sr_no             = row[0]
                circle            = row[1]
                emp_code          = str(row[2]).strip() if row[2] else None
                ust_id            = str(row[3]).strip() if row[3] else None
                emp_name          = row[4]
                contact_no        = str(row[5]).strip()[:15] if row[5] else ''
                designation_name  = row[6]
                department_name   = row[7]
                reporting_manager = row[8]
                skill_set         = row[11] if len(row) > 11 else None

                # Skip empty rows
                if not emp_code or emp_code == 'None':
                    skipped_rows += 1
                    continue

                # Skip header rows
                if not sr_no or not str(sr_no).isdigit() and not isinstance(sr_no, int):
                    skipped_rows += 1
                    continue

                try:
                    employee, created = Employee.objects.update_or_create(
                        emp_code=emp_code,
                        defaults={
                            'sr_no': int(sr_no) if sr_no else 0,
                            'circle': str(circle) if circle else '',
                            'ust_id': ust_id or '',
                            'emp_name': str(emp_name) if emp_name else '',
                            'contact_no': contact_no or '',
                            'designation_name': str(designation_name) if designation_name else '',
                            'department_name': str(department_name) if department_name else '',
                            'reporting_manager': str(reporting_manager) if reporting_manager else '',
                            'skill_set': str(skill_set) if skill_set else '',
                        }
                    )

                    if created:
                        new_employees += 1
                    else:
                        updated_employees += 1

                except Exception as e:
                    print(f"Employee save error for {emp_code}: {str(e)}")
                    skipped_rows += 1
                    continue

                # Add survey data for each date
                for col_index, survey_date in date_columns.items():
                    try:
                        working_status    = row[col_index] if len(row) > col_index else None
                        project           = row[col_index + 1] if len(row) > col_index + 1 else None
                        activity_assigned = row[col_index + 2] if len(row) > col_index + 2 else None
                        site_id           = row[col_index + 3] if len(row) > col_index + 3 else None
                        ssid              = row[col_index + 4] if len(row) > col_index + 4 else None
                        activity_status   = row[col_index + 5] if len(row) > col_index + 5 else None
                        detailed_remarks  = row[col_index + 6] if len(row) > col_index + 6 else None
                        owner             = row[col_index + 7] if len(row) > col_index + 7 else None
                    except IndexError:
                        continue

                    if not working_status:
                        continue

                    try:
                        survey, s_created = DriveTestSurvey.objects.update_or_create(
                            employee=employee,
                            date=survey_date,
                            defaults={
                                'working_status': str(working_status).strip() if working_status else '',
                                'project': str(project).strip() if project else '',
                                'activity_assigned': str(activity_assigned).strip() if activity_assigned else '',
                                'site_id': str(site_id).strip() if site_id else '',
                                'ssid': str(ssid).strip() if ssid else '',
                                'activity_status': str(activity_status).strip() if activity_status else '',
                                'detailed_remarks': str(detailed_remarks).strip() if detailed_remarks else '',
                                'owner': str(owner).strip() if owner else '',
                            }
                        )

                        if s_created:
                            new_surveys += 1
                        else:
                            updated_surveys += 1

                    except Exception as e:
                        print(f"Survey save error for {emp_code} on {survey_date}: {str(e)}")
                        continue

            return Response({
                "message": "File uploaded and processed successfully!",
                "employees": {
                    "new": new_employees,
                    "updated": updated_employees,
                    "total": new_employees + updated_employees
                },
                "surveys": {
                    "new": new_surveys,
                    "updated": updated_surveys,
                    "total": new_surveys + updated_surveys
                },
                "dates_found": [str(d) for d in date_columns.values()],
                "skipped_rows": skipped_rows
            }, status=status.HTTP_200_OK)

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {"error": f"Something went wrong: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
from django.http import HttpResponse
import io

class ExportExcelView(APIView):

    def get(self, request):
        # Get all employees ordered by sr_no
        employees = Employee.objects.all().order_by('sr_no')

        # Get all unique dates ordered
        dates = DriveTestSurvey.objects.values_list(
            'date', flat=True
        ).distinct().order_by('date')
        dates = list(dates)

        if not dates:
            return Response(
                {"error": "No survey data found in database."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Create Excel file in memory
        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Details"

        # ── Row 1 — Date headers ──────────────────────────────
        # First 12 columns are employee info (no date header)
        # Then for each date, merge across 7 columns
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Employee columns count = 12 (A to L)
        emp_col_count = 12
        survey_col_count = 7  # working_status, project, activity_assigned, site_id, ssid, activity_status, detailed_remarks

        # Write date headers in row 1
        for date_idx, date in enumerate(dates):
            start_col = emp_col_count + 1 + (date_idx * survey_col_count)
            end_col = start_col + survey_col_count - 1
            ws.cell(row=1, column=start_col, value=date)
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col
            )
            # Style the date header
            date_cell = ws.cell(row=1, column=start_col)
            date_cell.font = Font(bold=True, color="FFFFFF")
            date_cell.fill = PatternFill("solid", fgColor="1565C0")
            date_cell.alignment = Alignment(horizontal="center")

        # ── Row 2 — Column headers ────────────────────────────
        emp_headers = [
            'SR. No.', 'Circle', 'Emp. Code', 'UST ID',
            'Emp Name', 'Contact No', 'Designation Name',
            'Department Name', 'Reporting Manager Name',
            'Project Code', 'Project Name',
            'Skill Set (Survey/MW INC/RAN INC/Integration/Drive Test)'
        ]

        survey_headers = [
            'Working Status (Idle/Working/Leave/Week off)',
            'Project(Relocation/MW/Degrow/New Tower/Upgrade/IBS/Survey)',
            'Activity Assigned',
            'Site Id',
            'SSID',
            'Activity Status',
            'Detailed Remarks (Idle Reasons,etc.)'
        ]

        # Write employee headers
        for col_idx, header in enumerate(emp_headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1565C0")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Write survey headers for each date
        for date_idx, date in enumerate(dates):
            start_col = emp_col_count + 1 + (date_idx * survey_col_count)
            for col_idx, header in enumerate(survey_headers):
                cell = ws.cell(row=2, column=start_col + col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1565C0")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # ── Rows 3+ — Data ────────────────────────────────────
        for row_idx, employee in enumerate(employees, 3):
            # Write employee info
            emp_data = [
                employee.sr_no,
                employee.circle,
                employee.emp_code,
                employee.ust_id,
                employee.emp_name,
                employee.contact_no,
                employee.designation_name,
                employee.department_name,
                employee.reporting_manager,
                '',  # Project Code (not stored)
                '',  # Project Name (not stored)
                employee.skill_set,
            ]

            for col_idx, value in enumerate(emp_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

            # Write survey data for each date
            for date_idx, date in enumerate(dates):
                start_col = emp_col_count + 1 + (date_idx * survey_col_count)
                try:
                    survey = DriveTestSurvey.objects.get(
                        employee=employee,
                        date=date
                    )
                    survey_data = [
                        survey.working_status,
                        survey.project,
                        survey.activity_assigned,
                        survey.site_id,
                        survey.ssid,
                        survey.activity_status,
                        survey.detailed_remarks,
                    ]
                except DriveTestSurvey.DoesNotExist:
                    # No data for this employee on this date
                    survey_data = ['', '', '', '', '', '', '']

                for col_idx, value in enumerate(survey_data):
                    ws.cell(row=row_idx, column=start_col + col_idx, value=value)

        # ── Auto fit column widths ────────────────────────────
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

        # Save to memory
        output_dir = os.path.join(MEDIA_ROOT, "field_resource_tracking")
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"field_resource_tracking_export_{timestamp}.xlsx"

        output_path = os.path.join(output_dir, filename)

        workbook.save(output_path)

        relative_path = os.path.relpath(output_path, MEDIA_ROOT)

        download_url = request.build_absolute_uri(
            os.path.join(MEDIA_URL, relative_path).replace("\\", "/")
        )

        return Response({
            "status": True,
            "message": "File exported successfully.",
            "download_url": download_url
        }) 


class ExportDateRangeView(APIView):

    def post(self, request):
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        if not start_date or not end_date:
            return Response(
                {"error": "Please provide start_date and end_date. Example: {'start_date': '2026-06-15', 'end_date': '2026-06-17'}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get all dates in the range
        dates = DriveTestSurvey.objects.filter(
            date__gte=start_date,
            date__lte=end_date
        ).values_list('date', flat=True).distinct().order_by('date')
        dates = list(dates)

        if not dates:
            return Response(
                {"error": f"No data found between {start_date} and {end_date}"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get all employees
        employees = Employee.objects.all().order_by('sr_no')

        # Create Excel file
        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Details"

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        emp_col_count   = 12
        survey_col_count = 7

        # Row 1 — date headers
        for date_idx, date in enumerate(dates):
            start_col = emp_col_count + 1 + (date_idx * survey_col_count)
            end_col   = start_col + survey_col_count - 1
            ws.cell(row=1, column=start_col, value=date)
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col
            )
            date_cell = ws.cell(row=1, column=start_col)
            date_cell.font      = Font(bold=True, color="FFFFFF")
            date_cell.fill      = PatternFill("solid", fgColor="1565C0")
            date_cell.alignment = Alignment(horizontal="center")

        # Row 2 — column headers
        emp_headers = [
            'SR. No.', 'Circle', 'Emp. Code', 'UST ID',
            'Emp Name', 'Contact No', 'Designation Name',
            'Department Name', 'Reporting Manager Name',
            'Project Code', 'Project Name',
            'Skill Set (Survey/MW INC/RAN INC/Integration/Drive Test)'
        ]

        survey_headers = [
            'Working Status (Idle/Working/Leave/Week off)',
            'Project(Relocation/MW/Degrow/New Tower/Upgrade/IBS/Survey)',
            'Activity Assigned',
            'Site Id',
            'SSID',
            'Activity Status',
            'Detailed Remarks (Idle Reasons,etc.)'
        ]

        for col_idx, header in enumerate(emp_headers, 1):
            cell            = ws.cell(row=2, column=col_idx, value=header)
            cell.font       = Font(bold=True, color="FFFFFF")
            cell.fill       = PatternFill("solid", fgColor="1565C0")
            cell.alignment  = Alignment(horizontal="center", wrap_text=True)

        for date_idx, date in enumerate(dates):
            start_col = emp_col_count + 1 + (date_idx * survey_col_count)
            for col_idx, header in enumerate(survey_headers):
                cell           = ws.cell(row=2, column=start_col + col_idx, value=header)
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="1565C0")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Row 3+ — data
        for row_idx, employee in enumerate(employees, 3):
            emp_data = [
                employee.sr_no,
                employee.circle,
                employee.emp_code,
                employee.ust_id,
                employee.emp_name,
                employee.contact_no,
                employee.designation_name,
                employee.department_name,
                employee.reporting_manager,
                '',
                '',
                employee.skill_set,
            ]

            for col_idx, value in enumerate(emp_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

            for date_idx, date in enumerate(dates):
                start_col = emp_col_count + 1 + (date_idx * survey_col_count)
                try:
                    survey = DriveTestSurvey.objects.get(
                        employee=employee,
                        date=date
                    )
                    survey_data = [
                        survey.working_status,
                        survey.project,
                        survey.activity_assigned,
                        survey.site_id,
                        survey.ssid,
                        survey.activity_status,
                        survey.detailed_remarks,
                    ]
                except DriveTestSurvey.DoesNotExist:
                    survey_data = ['', '', '', '', '', '', '']

                for col_idx, value in enumerate(survey_data):
                    ws.cell(row=row_idx, column=start_col + col_idx, value=value)

        # Auto fit columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

        output_dir = os.path.join(MEDIA_ROOT, "field_resource_tracking")
        os.makedirs(output_dir, exist_ok=True)

        # Optional: delete previous exported files
        # delete_existing_files(output_dir)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{start_date}_to_{end_date}_{timestamp}.xlsx"

        output_path = os.path.join(output_dir, filename)

        # Save workbook
        workbook.save(output_path)

        # Generate download URL
        relative_path = os.path.relpath(output_path, MEDIA_ROOT)

        download_url = request.build_absolute_uri(
            os.path.join(MEDIA_URL, relative_path).replace("\\", "/")
        )

        return Response({
            "status": True,
            "message": "File exported successfully.",
            "download_url": download_url
        })
    
from django.db.models import Count, Q
from django.db.models.functions import TruncDate

class AnalyticsView(APIView):
    def get(self, request):
        date = request.query_params.get('date', None)

        surveys = DriveTestSurvey.objects.all()
        if date:
            surveys = surveys.filter(date=date)

        # 1. Working status breakdown
        status_breakdown = surveys.values(
            'working_status'
        ).annotate(count=Count('id')).order_by('-count')

        # 2. Project breakdown
        project_breakdown = surveys.values(
            'project'
        ).annotate(count=Count('id')).order_by('-count')

        # 3. Circle wise breakdown
        circle_breakdown = surveys.values(
            'employee__circle'
        ).annotate(count=Count('id')).order_by('-count')

        # 4. Department wise breakdown
        dept_breakdown = surveys.values(
            'employee__department_name'
        ).annotate(count=Count('id')).order_by('-count')

        # 5. Skill set breakdown
        skill_breakdown = Employee.objects.values(
            'skill_set'
        ).annotate(count=Count('emp_code')).order_by('-count')

        # 6. Date wise working status trend
        date_trend = DriveTestSurvey.objects.values(
            'date', 'working_status'
        ).annotate(count=Count('id')).order_by('date')

        # 7. Overall stats
        total_employees = Employee.objects.count()
        total_surveys   = surveys.count()
        working_count   = surveys.filter(working_status='Working').count()
        idle_count      = surveys.filter(working_status='Idle').count()
        leave_count     = surveys.filter(working_status='Leave').count()
        weekoff_count   = surveys.filter(working_status='Week off').count()
        cnr_count       = surveys.filter(working_status='Call Not Respond').count()

        return Response({
            "overall": {
                "total_employees":  total_employees,
                "total_surveys":    total_surveys,
                "working":          working_count,
                "idle":             idle_count,
                "leave":            leave_count,
                "week_off":         weekoff_count,
                "call_not_respond": cnr_count,
            },
            "status_breakdown":     list(status_breakdown),
            "project_breakdown":    list(project_breakdown),
            "circle_breakdown":     list(circle_breakdown),
            "department_breakdown": list(dept_breakdown),
            "skill_breakdown":      list(skill_breakdown),
            "date_trend":           list(date_trend),
        })


class DateSummaryView(APIView):
    # GET /api/date-summary/
    # Returns summary for each date
    def get(self, request):
        dates = DriveTestSurvey.objects.values_list(
            'date', flat=True
        ).distinct().order_by('date')

        result = []
        for date in dates:
            surveys = DriveTestSurvey.objects.filter(date=date)
            result.append({
                "date":             str(date),
                "total":            surveys.count(),
                "working":          surveys.filter(working_status='Working').count(),
                "idle":             surveys.filter(working_status='Idle').count(),
                "leave":            surveys.filter(working_status='Leave').count(),
                "week_off":         surveys.filter(working_status='Week off').count(),
                "call_not_respond": surveys.filter(working_status='Call Not Respond').count(),
            })

        return Response({
            "total_dates": len(result),
            "dates": result
        })


class CircleSummaryView(APIView):
    # GET /api/circle-summary/
    # Returns summary per circle
    def get(self, request):
        date = request.query_params.get('date', None)

        circles = Employee.objects.values_list(
            'circle', flat=True
        ).distinct().order_by('circle')

        result = []
        for circle in circles:
            employees = Employee.objects.filter(circle=circle)
            emp_codes  = employees.values_list('emp_code', flat=True)

            surveys = DriveTestSurvey.objects.filter(
                employee__emp_code__in=emp_codes
            )
            if date:
                surveys = surveys.filter(date=date)

            result.append({
                "circle":           circle,
                "total_employees":  employees.count(),
                "total_surveys":    surveys.count(),
                "working":          surveys.filter(working_status='Working').count(),
                "idle":             surveys.filter(working_status='Idle').count(),
                "leave":            surveys.filter(working_status='Leave').count(),
                "week_off":         surveys.filter(working_status='Week off').count(),
                "call_not_respond": surveys.filter(working_status='Call Not Respond').count(),
            })

        return Response({
            "total_circles": len(result),
            "circles": result
        })


class DepartmentSummaryView(APIView):
    # GET /api/dept-summary/
    # Returns summary per department
    def get(self, request):
        date = request.query_params.get('date', None)

        depts = Employee.objects.values_list(
            'department_name', flat=True
        ).distinct().order_by('department_name')

        result = []
        for dept in depts:
            employees = Employee.objects.filter(department_name=dept)
            emp_codes  = employees.values_list('emp_code', flat=True)

            surveys = DriveTestSurvey.objects.filter(
                employee__emp_code__in=emp_codes
            )
            if date:
                surveys = surveys.filter(date=date)

            result.append({
                "department":       dept,
                "total_employees":  employees.count(),
                "total_surveys":    surveys.count(),
                "working":          surveys.filter(working_status='Working').count(),
                "idle":             surveys.filter(working_status='Idle').count(),
                "leave":            surveys.filter(working_status='Leave').count(),
                "week_off":         surveys.filter(working_status='Week off').count(),
                "call_not_respond": surveys.filter(working_status='Call Not Respond').count(),
            })

        return Response({
            "total_departments": len(result),
            "departments": result
        })


class SkillSummaryView(APIView):
    # GET /api/skill-summary/
    # Returns summary per skill set
    def get(self, request):
        date = request.query_params.get('date', None)

        skills = Employee.objects.values_list(
            'skill_set', flat=True
        ).distinct().order_by('skill_set')

        result = []
        for skill in skills:
            employees = Employee.objects.filter(skill_set=skill)
            emp_codes  = employees.values_list('emp_code', flat=True)

            surveys = DriveTestSurvey.objects.filter(
                employee__emp_code__in=emp_codes
            )
            if date:
                surveys = surveys.filter(date=date)

            result.append({
                "skill_set":        skill or 'Unknown',
                "total_employees":  employees.count(),
                "total_surveys":    surveys.count(),
                "working":          surveys.filter(working_status='Working').count(),
                "idle":             surveys.filter(working_status='Idle').count(),
                "leave":            surveys.filter(working_status='Leave').count(),
                "week_off":         surveys.filter(working_status='Week off').count(),
                "call_not_respond": surveys.filter(working_status='Call Not Respond').count(),
            })

        return Response({
            "total_skills": len(result),
            "skills": result
        })


class ProjectSummaryView(APIView):
    # GET /api/project-summary/
    # Returns summary per project
    def get(self, request):
        date = request.query_params.get('date', None)

        surveys = DriveTestSurvey.objects.all()
        if date:
            surveys = surveys.filter(date=date)

        projects = surveys.values_list(
            'project', flat=True
        ).distinct().order_by('project')

        result = []
        for project in projects:
            proj_surveys = surveys.filter(project=project)
            result.append({
                "project":          project or 'None',
                "total":            proj_surveys.count(),
                "working":          proj_surveys.filter(working_status='Working').count(),
                "idle":             proj_surveys.filter(working_status='Idle').count(),
                "leave":            proj_surveys.filter(working_status='Leave').count(),
                "call_not_respond": proj_surveys.filter(working_status='Call Not Respond').count(),
            })

        return Response({
            "total_projects": len(result),
            "projects": result
        })


class EmployeeSearchView(APIView):
    # GET /api/search/?q=Raju
    # Search employees by name, emp_code, circle, department
    def get(self, request):
        query = request.query_params.get('q', None)

        if not query:
            return Response(
                {"error": "Please provide a search query. Example: /api/search/?q=Raju"},
                status=status.HTTP_400_BAD_REQUEST
            )

        employees = Employee.objects.filter(
            Q(emp_name__icontains=query) |
            Q(emp_code__icontains=query) |
            Q(circle__icontains=query)   |
            Q(department_name__icontains=query) |
            Q(designation_name__icontains=query)
        ).order_by('sr_no')

        result = []
        for emp in employees:
            surveys = DriveTestSurvey.objects.filter(
                employee=emp
            ).order_by('date')

            result.append({
                "emp_code":         emp.emp_code,
                "emp_name":         emp.emp_name,
                "circle":           emp.circle,
                "department":       emp.department_name,
                "designation":      emp.designation_name,
                "skill_set":        emp.skill_set,
                "total_dates":      surveys.count(),
                "surveys":          DailyUpdateSerializer(surveys, many=True).data
            })

        return Response({
            "query":        query,
            "total_found":  len(result),
            "employees":    result
        })


class ReportingManagerSummaryView(APIView):
    # GET /api/manager-summary/?date=2026-06-15
    # Returns summary per reporting manager
    def get(self, request):
        date = request.query_params.get('date', None)

        managers = Employee.objects.values_list(
            'reporting_manager', flat=True
        ).distinct().order_by('reporting_manager')

        result = []
        for manager in managers:
            employees = Employee.objects.filter(reporting_manager=manager)
            emp_codes  = employees.values_list('emp_code', flat=True)

            surveys = DriveTestSurvey.objects.filter(
                employee__emp_code__in=emp_codes
            )
            if date:
                surveys = surveys.filter(date=date)

            result.append({
                "reporting_manager": manager,
                "total_employees":   employees.count(),
                "working":           surveys.filter(working_status='Working').count(),
                "idle":              surveys.filter(working_status='Idle').count(),
                "leave":             surveys.filter(working_status='Leave').count(),
                "call_not_respond":  surveys.filter(working_status='Call Not Respond').count(),
            })

        return Response({
            "total_managers": len(result),
            "managers": result
        })


class IdleEmployeesView(APIView):
    # GET /api/idle/?date=2026-06-15
    # Returns all idle employees for a date
    def get(self, request):
        date = request.query_params.get('date', None)

        if not date:
            return Response(
                {"error": "Please provide a date. Example: /api/idle/?date=2026-06-15"},
                status=status.HTTP_400_BAD_REQUEST
            )

        surveys = DriveTestSurvey.objects.filter(
            date=date,
            working_status='Idle'
        ).order_by('employee__sr_no')

        result = []
        for s in surveys:
            result.append({
                "emp_code":       s.employee.emp_code,
                "emp_name":       s.employee.emp_name,
                "circle":         s.employee.circle,
                "department":     s.employee.department_name,
                "detailed_remarks": s.detailed_remarks,
            })

        return Response({
            "date":          date,
            "total_idle":    len(result),
            "employees":     result
        })


class WorkingEmployeesView(APIView):
    # GET /api/working/?date=2026-06-15
    # Returns all working employees for a date
    def get(self, request):
        date = request.query_params.get('date', None)

        if not date:
            return Response(
                {"error": "Please provide a date. Example: /api/working/?date=2026-06-15"},
                status=status.HTTP_400_BAD_REQUEST
            )

        surveys = DriveTestSurvey.objects.filter(
            date=date,
            working_status='Working'
        ).order_by('employee__sr_no')

        result = []
        for s in surveys:
            result.append({
                "emp_code":           s.employee.emp_code,
                "emp_name":           s.employee.emp_name,
                "circle":             s.employee.circle,
                "department":         s.employee.department_name,
                "project":            s.project,
                "activity_assigned":  s.activity_assigned,
                "site_id":            s.site_id,
                "activity_status":    s.activity_status,
            })

        return Response({
            "date":           date,
            "total_working":  len(result),
            "employees":      result
        })


class AvailableDatesView(APIView):
    # GET /api/dates/
    # Returns all available dates in the database
    def get(self, request):
        dates = DriveTestSurvey.objects.values_list(
            'date', flat=True
        ).distinct().order_by('date')

        return Response({
            "total_dates": len(dates),
            "dates": [str(d) for d in dates]
        })
from django.shortcuts import render

class DashboardView(APIView):
    def get(self, request):
        return render(request, 'dashboard.html')
    
class DateRangeView(APIView):

    def post(self, request):

        start_date = request.data.get("start_date")
        end_date = request.data.get("end_date")

        if not start_date or not end_date:
            return Response(
                {"error": "start_date and end_date are required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        queryset = DriveTestSurvey.objects.filter(
            date__range=[start_date, end_date]
        ).order_by("date", "employee")

        serializer = DailyUpdateSerializer(queryset, many=True)

        return Response({
            "total_records": queryset.count(),
            "data": serializer.data
        })