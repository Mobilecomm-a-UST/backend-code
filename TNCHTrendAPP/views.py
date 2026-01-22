import sys
import pandas as pd
from django.contrib import messages
# import json
from django.contrib.auth.decorators import login_required
from mcom_website.settings import MEDIA_URL,BASE_DIR,ALLOWED_HOSTS
import numpy as np
import os
import datetime
from rest_framework.decorators import api_view,parser_classes
from rest_framework.response import Response
from django.http import JsonResponse



from django.forms.models import model_to_dict

from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import authentication_classes,permission_classes
from django.core.files.storage import FileSystemStorage
from mcom_website.settings import MEDIA_ROOT
from .models import *

import openpyxl
from openpyxl.styles import PatternFill

from openpyxl import Workbook,load_workbook
from statistics import mean

from django.db.models import Sum


import pandas as pd
import matplotlib  as mpl
import numpy as np

from datetime import date, timedelta

from commom_utilities.utils import *






@api_view(["POST"])
# @authentication_classes([TokenAuthentication])
# @permission_classes([IsAuthenticated])
def old_tnch_Trend(request):
    try:
        kpi=[   "MV_RRC SETUP SUCCESS RATE [CDBH]",
                "MV_ERAB SETUP SUCCESS RATE [CDBH]",
                "MV_PS DROP CALL RATE % [CDBH]",
                "MV_DL USER THROUGHPUT_MBPS [CDBH]",
                "MV_UL USER THROUGHPUT_MBPS [CDBH]",
                "MV_PS HANDOVER SUCCESS RATE [LTE INTRA SYSTEM] [CDBH]",
                "MV_PS HANDOVER SUCCESS RATE [LTE INTER SYSTEM] [CDBH]", 
                "MV_CSFB REDIRECTION SUCCESS RATE [CDBH]",
                "MV_E-UTRAN AVERAGE CQI [CDBH]",
                "UL RSSI PUCCH",
                "MV_VOLTE ERAB SETUP SUCCESS RATE [CBBH]",
                "MV_VOLTE DCR [CBBH]",
                "MV_VOLTE PACKET LOSS DL [CBBH]",
                "MV_VOLTE PACKET LOSS UL [CBBH]",
                "MV_VOLTE INTRAF HOSR EXEC [CBBH]",
                "MV_4G DATA VOLUME_GB",
                "MV_VOLTE INTERF HOSR EXEC [CBBH]", 
                "MV_AVERAGE NUMBER OF USED DL PRBS [CDBH]",
                "MV_VOLTE TRAFFIC",
                "PS HANDOVER SUCCESS RATE_NOM [LTE INTRA SYSTEM] [CDBH]",
                "PS HANDOVER SUCCESS RATE_DENOM [LTE INTRA SYSTEM] [CDBH]",
                "PS HANDOVER SUCCESS RATE_NOM [LTE INTER SYSTEM] [CDBH]",
                "PS HANDOVER SUCCESS RATE_DENOM [LTE INTER SYSTEM] [CDBH]",
                "VOLTE DCR_NOM [CBBH]",
                "PS DROP CALL RATE_NOM [CDBH]",
                "VOLTE ERAB SETUP SUCCESS RATE_NOM [CBBH]",
                "VOLTE ERAB SETUP SUCCESS RATE_DENOM [CBBH]",
                "VOLTE INTRAF HOSR EXEC_NOM [CBBH]",
                "VOLTE INTRAF HOSR EXEC_DENOM [CBBH]",
                "VOLTE INTERF HOSR EXEC_NOM [CBBH]",
                "VOLTE INTERF HOSR EXEC_DENOM [CBBH]",
                "MV_RADIO NW AVAILABILITY",
                "MV_CSFB REDIRECTION SUCCESS RATE_NOM [CDBH]",
                "MV_CSFB REDIRECTION SUCCESS RATE_DENOM [CDBH]",
                "LTE SgNb addition Success Rate %".upper(),
                "LTE SgNb addition Success Rate_Nom".upper(),
                "LTE SgNb addition Success Rate_Denom".upper(),

            
       ]
        kpiStrToZero=[ "MV_RRC SETUP SUCCESS RATE [CDBH]",
                "MV_ERAB SETUP SUCCESS RATE [CDBH]",
                "MV_PS DROP CALL RATE % [CDBH]",
                "MV_DL USER THROUGHPUT_KBPS [CDBH]",
                "MV_UL USER THROUGHPUT_KBPS [CDBH]",
                "MV_PS HANDOVER SUCCESS RATE [LTE INTRA SYSTEM] [CDBH]",
                "MV_PS HANDOVER SUCCESS RATE [LTE INTER SYSTEM] [CDBH]", 
                "MV_CSFB REDIRECTION SUCCESS RATE [CDBH]",
                "MV_E-UTRAN AVERAGE CQI [CDBH]",
                "UL RSSI PUCCH",
                "MV_VOLTE ERAB SETUP SUCCESS RATE [CBBH]",
                "MV_VOLTE DCR [CBBH]",
                "MV_VOLTE PACKET LOSS DL [CBBH]",
                "MV_VOLTE PACKET LOSS UL [CBBH]",
                "MV_VOLTE INTRAF HOSR EXEC [CBBH]",
                "MV_4G DATA VOLUME_GB",
                "MV_VOLTE INTERF HOSR EXEC [CBBH]", 
                "MV_AVERAGE NUMBER OF USED DL PRBS [CDBH]",
                "MV_VOLTE TRAFFIC",
                "PS HANDOVER SUCCESS RATE_NOM [LTE INTRA SYSTEM] [CDBH]",
                "PS HANDOVER SUCCESS RATE_DENOM [LTE INTRA SYSTEM] [CDBH]",
                "PS HANDOVER SUCCESS RATE_NOM [LTE INTER SYSTEM] [CDBH]",
                "PS HANDOVER SUCCESS RATE_DENOM [LTE INTER SYSTEM] [CDBH]",
                "VOLTE DCR_NOM [CBBH]",
                "PS DROP CALL RATE_NOM [CDBH]",
                "VOLTE ERAB SETUP SUCCESS RATE_NOM [CBBH]",
                "VOLTE ERAB SETUP SUCCESS RATE_DENOM [CBBH]",
                "VOLTE INTRAF HOSR EXEC_NOM [CBBH]",
                "VOLTE INTRAF HOSR EXEC_DENOM [CBBH]",
                "VOLTE INTERF HOSR EXEC_NOM [CBBH]",
                "VOLTE INTERF HOSR EXEC_DENOM [CBBH]",
                "MV_RADIO NW AVAILABILITY",
                "MV_CSFB REDIRECTION SUCCESS RATE_NOM [CDBH]",
                "MV_CSFB REDIRECTION SUCCESS RATE_DENOM [CDBH]",
                "LTE SgNb addition Success Rate %".upper(),
                "LTE SgNb addition Success Rate_Nom".upper(),
                "LTE SgNb addition Success Rate_Denom".upper(), 
                

            
       ]

        raw_kpi = request.FILES["raw_kpi"] if 'raw_kpi' in request.FILES else None
        
        #-------------------------- call for column check -------------------------------

        required_cols=[ "MV_RRC Setup Success Rate [CDBH]",
                        "MV_ERAB Setup Success Rate [CDBH]",
                        "MV_PS Drop Call Rate % [CDBH]",
                        "MV_DL User Throughput_Kbps [CDBH]",
                        "MV_UL User Throughput_Kbps [CDBH]",
                        "MV_PS handover success rate [LTE Intra System] [CDBH]",
                        "MV_CSFB Redirection Success Rate [CDBH]",
                        "MV_E-UTRAN Average CQI [CDBH]",
                        "UL RSSI PUCCH",
                        "MV_VoLTE DCR [CBBH]",
                        "MV_PS handover success rate [LTE Inter System] [CDBH]",
                        "MV_VoLTE ERAB Setup Success Rate [CBBH]",
                        "MV_VoLTE IntraF HOSR Exec [CBBH]",
                        "MV_VoLTE InterF HOSR Exec [CBBH]",
                        "MV_VoLTE Packet Loss DL [CBBH]",
                        "MV_VoLTE Packet Loss UL [CBBH]",
                        "MV_4G Data Volume_GB",
                        "MV_VoLTE Traffic",
                        "MV_Average number of used DL PRBs [CDBH]",
                        "UL RSSI PUCCH",
                        "CSFB Redirection Success Rate [CDBH]",
                        "PS handover success rate_Nom [LTE Intra System] [CDBH]",
                        "PS handover success rate_Denom [LTE Intra System] [CDBH]",
                        "PS handover success rate_Nom [LTE Inter System] [CDBH]",
                        "PS handover success rate_Denom [LTE Inter System] [CDBH]",
                        "VoLTE DCR_Nom [CBBH]",
                        "PS Drop Call Rate_Nom [CDBH]",
                        "VoLTE ERAB Setup Success Rate_NOM [CBBH]",
                        "VoLTE ERAB Setup Success Rate_Denom [CBBH]",
                        "VoLTE IntraF HOSR Exec_Nom [CBBH]",
                        "VoLTE IntraF HOSR Exec_Denom [CBBH]",
                        "VoLTE InterF HOSR Exec_Nom [CBBH]",
                        "VoLTE InterF HOSR Exec_Denom [CBBH]",
                        "VoLTE DCR_Denom [CBBH]",
                        "PS Drop Call Rate_Denom [CDBH]",
                        "MV_Radio NW Availability",
                        "MV_Radio NW Availability_Denom",
                        "MV_Radio NW Availability_Nom",
                        "MV_CSFB Redirection Success Rate_Nom [CDBH]",
                        "MV_CSFB Redirection Success Rate_Denom [CDBH]",
                        "LTE SgNb addition Success Rate %",
                        "LTE SgNb addition Success Rate_Nom",
                        "LTE SgNb addition Success Rate_Denom" 
        ]
        sts,response=required_col_check(raw_kpi,required_cols)
        if sts:
            return Response(response)
        

        df_raw_kpi=pd.read_excel(raw_kpi)
      #-----------------------------------*****************------------------------------------


      #----------------------------------- site list----------------------------------------------
        # site_list_file = request.FILES["site_list"] if 'site_list' in request.FILES else None
        # str_site_list=request.POST.get("str_site_list") if "str_site_list" in request.POST else None
        
        # if site_list_file:
        #     sts,response=required_col_check(site_list_file,["2G ID"])
        #     if sts:
        #        return Response(response)
        #     df_site_list = pd.read_excel(site_list_file)
        #     site_list=list(df_site_list["2G ID"])
        
        # elif str_site_list:
           
        #     site_list=str_site_list.split(",")
        #     site_list=[site.strip() for site in site_list]
        #     ##print("site_list.........................",site_list)
        # else:
        #     print("true")
        #     return Response({"status":False,"message":"please provide site list"})
        response,s_l= site_list_handler(request)
        if s_l:
            site_list=s_l
        if response:
            return Response(response)
        
        #---------------------------------------****************-------------------------------------------
        
        
        #-------------------------------- Offered Date -----------------------------------------------
        offered_date = request.POST.get("offered_date")
        ##print(offered_date)
        offered_date=datetime.datetime.strptime(offered_date,"%Y-%m-%d").date()
        ##print("offered_date:",offered_date)
        #------------------------------ ****************---------------------------------------------
       
       
        door_root= os.path.join(MEDIA_ROOT,"Original_trend","Tnch","Trend")
        path_of_blnk_temp=os.path.join(door_root,"template","template6.xlsx")
        trend_wb=load_workbook(path_of_blnk_temp)
        ##print(trend_wb.sheetnames)
        ##print("##################################################################################")
        

        df_raw_kpi.columns=df_raw_kpi.columns.str.upper()
             
        ########################## the below code is to replace every string to zero from numeric columns ##################

        for x in kpiStrToZero:
             df_raw_kpi[x] = df_raw_kpi[x].replace(to_replace='.*', value=0, regex=True)
        ######################################  * * * * * * * * * * * * * * ################################################
       
        # df_raw_kpi.columns=df_raw_kpi.columns.str.upper()

        ##print("__________________raw KPI after converting str to zero_______________")
        ##print(df_raw_kpi)

        df_raw_kpi["SHORT NAME"].fillna( inplace=True, method="ffill")
        df_raw_kpi["SHORT NAME"] =df_raw_kpi["SHORT NAME"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..


        df_raw_kpi.rename( columns={'UNNAMED: 1':'DATE'}, inplace=True )

        df_raw_kpi["MV_DL USER THROUGHPUT_KBPS [CDBH]"] = (df_raw_kpi["MV_DL USER THROUGHPUT_KBPS [CDBH]"]/1024)
        df_raw_kpi.rename(columns={"MV_DL USER THROUGHPUT_KBPS [CDBH]" :"MV_DL USER THROUGHPUT_MBPS [CDBH]" } ,inplace = True )

        df_raw_kpi["MV_UL USER THROUGHPUT_KBPS [CDBH]"] = (df_raw_kpi["MV_UL USER THROUGHPUT_KBPS [CDBH]"]/1024)
        df_raw_kpi.rename(columns={"MV_UL USER THROUGHPUT_KBPS [CDBH]" :"MV_UL USER THROUGHPUT_MBPS [CDBH]" } ,inplace = True )

        lis=list(df_raw_kpi["SHORT NAME"])
        sit_id_lis=[]
        cell_id_lis=[]
        for item in lis:
            if("_" in item):
                cell_id=item.split("_")[-2]
                ln=len(item.split("_")[-1])
            
                sit_id=item.split("_")[-2][:-ln]
            else:
                cell_id=item
                sit_id=item
            cell_id_lis.append(cell_id)
            sit_id_lis.append(sit_id)


        df_raw_kpi.insert(1, "SITE_ID", sit_id_lis)
        df_raw_kpi.insert(2, "CELL_ID", cell_id_lis)

        message=site_comparision(sit_id_lis,site_list) # site_comparision_call 
    
       

        df_raw_kpi.rename(columns={"SHORT NAME" :"SHORTNAME" } ,inplace = True )
        df_raw_kpi.fillna(value=0,inplace=True)

        process_op_path=os.path.join(door_root,"process_outputs")
        
        savepath=os.path.join(process_op_path,"desired input.xlsx")
        # df_raw_kpi.to_excel(savepath)
    
        # date1=date(2023,2,27)
        date1=offered_date
        # date1=date.today()
        dt1 = date1 - timedelta(1)
        dt2 = date1 - timedelta(2)
        dt3 = date1 - timedelta(3)
        dt4 = date1 - timedelta(4)
        dt5 = date1 - timedelta(5)
        ls=[dt1,dt2,dt3,dt4,dt5]
        only_site_fil = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.DATE.isin(ls))]
        
        savepath=os.path.join(process_op_path,"only_site_date_filtered_input.xlsx")
        # only_site_fil.to_excel(savepath)

        def perticular_tech( tech,site_list):
            # df_filtered = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.date.isin(ls))]
            df_filtered = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.DATE.isin(ls)) & (df_raw_kpi.SHORTNAME.str.contains('|'.join(tech)))]
            
            ##print(df_filtered)
            if not df_filtered.empty:
                address="last_filtered_input"  + str(tech) + ".xlsx"
                savepath=os.path.join(process_op_path,address)
                # df_filtered.to_excel(savepath)
                df_pivoted = df_filtered.pivot_table(index=["SITE_ID","SHORTNAME","CELL_ID"], columns="DATE")
                ##print("technology:",tech)
                ##print(df_pivoted)
                address_pivot="pivoted_input"  +str(tech)+ ".xlsx"
                savepath=os.path.join(process_op_path,address_pivot)
                # df_pivoted.to_excel(savepath)
                return df_pivoted
            
            return df_filtered


      
        alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        def num_hash(num):
                if num < 26:
                    return alpha[num-1]
                else:
                    q, r = num//26, num % 26
                    if r == 0:
                        if q == 1:
                            return alpha[r-1]
                        else:
                            return num_hash(q-1) + alpha[r-1]
                    else:
                        return num_hash(q) + alpha[r-1]
    
        # Driver code

        # ##printString(27906)

        def titleToNumber(s):
        # This process is similar to binary-to-
        # decimal conversion
            result = 0
            for B in range(len(s)):
                result *= 26
                result += ord(s[B]) - ord('A') + 1
            return result


        def overwrite(df_pivoted,kpi_name,coln1,trend_ws):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            ##print(kpi_name)
            index_pivot=df_pivoted.index.to_list()
            ##print("index ;###############################",index_pivot)
            ##print(len(index_pivot))
            ##print("index of pivoted table: ",index_pivot)
           
            dr=df_pivoted[kpi_name]
           

            ##print("columns of dr table",dr.columns)
            cl=dr.columns.to_list()
            ##print("column list",cl)
            
            # site_id=dr["SITE_ID"].to_list() 
            # cell_id=dr["CELL_ID"].to_list()    
            col1=dr[str(cl[0])].to_list()
            col2=dr[str(cl[1])].to_list()
            col3=dr[str(cl[2])].to_list()
            col4=dr[str(cl[3])].to_list()
            col5=dr[str(cl[4])].to_list()

            trend_ws[coln1+"3"].value=cl[0]
            trend_ws[coln2+"3"].value=cl[1]
            trend_ws[coln3+"3"].value=cl[2]
            trend_ws[coln4+"3"].value=cl[3]
            trend_ws[coln5+"3"].value=cl[4]

            # me=column_index_from_string(coln5)+1
            # me=get_column_letter(me)
            for i,value in enumerate(index_pivot):
                j=i+4
                trend_ws["A"+str(j)].value=index_pivot[i][1]
                trend_ws["E"+str(j)].value=index_pivot[i][0] 
                # trend_ws["C"+str(j)].value=index_pivot[i][2]
                # trend_ws["K"+str(j)].value=date1
                
                
                trend_ws[coln1+str(j)].value=col1[i]
                trend_ws[coln2+str(j)].value=col2[i]
                trend_ws[coln3+str(j)].value=col3[i]
                trend_ws[coln4+str(j)].value=col4[i]
                trend_ws[coln5+str(j)].value=col5[i]
                
                # trend_ws[me+str(j)].value='=COUNTIF(P5:T5,">=98.5")'

        # def overwrite_g2(df_pivoted,kpi_name,coln1,coln2,coln3,coln4,coln5,trend_ws):
        #     index_pivot=df_pivoted.index.to_list()
        #     ##print("index ;###############################",index_pivot)
        #     ##print(len(index_pivot))
        #     ##print("index of pivoted table: ",index_pivot)
        #     dr=df_pivoted[kpi_name]
        #     ##print("columns of dr table",dr.columns)
        #     cl=dr.columns.to_list()
        #     ##print("column list",cl)
            
        #     # site_id=dr["SITE_ID"].to_list() 
        #     # cell_id=dr["CELL_ID"].to_list()    
        #     col1=dr[str(cl[0])].to_list()
        #     col2=dr[str(cl[1])].to_list()
        #     col3=dr[str(cl[2])].to_list()
        #     col4=dr[str(cl[3])].to_list()
        #     col5=dr[str(cl[4])].to_list()

        #     trend_ws[coln1+"4"].value=cl[0]
        #     trend_ws[coln2+"4"].value=cl[1]
        #     trend_ws[coln3+"4"].value=cl[2]
        #     trend_ws[coln4+"4"].value=cl[3]
        #     trend_ws[coln5+"4"].value=cl[4]

        #     # me=column_index_from_string(coln5)+1
        #     # me=get_column_letter(me)
        #     for i,value in enumerate(index_pivot):
        #         j=i+5
        #         trend_ws["B"+str(j)].value=index_pivot[i][0]
        #         trend_ws["C"+str(j)].value=index_pivot[i][1]
        #         trend_ws["G"+str(j)].value=date1
                
                
                
        #         trend_ws[coln1+str(j)].value=col1[i]
        #         trend_ws[coln2+str(j)].value=col2[i]
        #         trend_ws[coln3+str(j)].value=col3[i]
        #         trend_ws[coln4+str(j)].value=col4[i]
        #         trend_ws[coln5+str(j)].value=col5[i]


        print("processing started........")
        # for fdd
        pivot_fdd=perticular_tech(["_F3_"],site_list)
        trend_ws=trend_wb["L1800"]
        
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                
                if(kpi_name=="MV_RRC SETUP SUCCESS RATE [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AK",trend_ws)

                if(kpi_name=="MV_ERAB SETUP SUCCESS RATE [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AR",trend_ws)

                if(kpi_name=="MV_PS DROP CALL RATE % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AY",trend_ws)

                if(kpi_name=="MV_DL USER THROUGHPUT_MBPS [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BF",trend_ws)

                if(kpi_name=="MV_UL USER THROUGHPUT_MBPS [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BM",trend_ws)

                if(kpi_name=="MV_PS HANDOVER SUCCESS RATE [LTE INTRA SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BT",trend_ws)
                
                if(kpi_name=="MV_PS HANDOVER SUCCESS RATE [LTE INTER SYSTEM] [CDBH]"): 
                    overwrite(pivot_fdd,kpi_name,"CA",trend_ws)
                
                # if(kpi_name=="MV_CSFB REDIRECTION SUCCESS RATE [CDBH]"):
                #     overwrite(pivot_fdd,kpi_name,"CH",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN AVERAGE CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CO",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CV",trend_ws)
                
                if(kpi_name=="MV_VOLTE ERAB SETUP SUCCESS RATE [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DC",trend_ws)
                
                if(kpi_name=="MV_VOLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DJ",trend_ws)
                
                if(kpi_name=="MV_VOLTE PACKET LOSS DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DQ",trend_ws)
                
                if(kpi_name=="MV_VOLTE PACKET LOSS UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DX",trend_ws)
                
                if(kpi_name=="MV_VOLTE INTRAF HOSR EXEC [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"EE",trend_ws)

                if(kpi_name=="MV_4G DATA VOLUME_GB"):
                    overwrite(pivot_fdd,kpi_name,"EL",trend_ws)

                if(kpi_name=="MV_VOLTE INTERF HOSR EXEC [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ES",trend_ws)

                if(kpi_name=="MV_AVERAGE NUMBER OF USED DL PRBS [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EZ",trend_ws)
                if(kpi_name=="MV_VOLTE TRAFFIC"):
                    overwrite(pivot_fdd,kpi_name,"FG",trend_ws)
                if(kpi_name=="MV_RADIO NW AVAILABILITY"):
                    overwrite(pivot_fdd,kpi_name,"FN",trend_ws)
                       

               
                #################################### nom/Denom #########################
                if(kpi_name=="PS HANDOVER SUCCESS RATE_NOM [LTE INTRA SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"FU",trend_ws)
                
                if(kpi_name=="PS HANDOVER SUCCESS RATE_DENOM [LTE INTRA SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GB",trend_ws)
                
                if(kpi_name=="PS HANDOVER SUCCESS RATE_NOM [LTE INTER SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GI",trend_ws)
                
                if(kpi_name=="PS HANDOVER SUCCESS RATE_DENOM [LTE INTER SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GP",trend_ws)
                
                if(kpi_name=="VOLTE DCR_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"GW",trend_ws)
                
                if(kpi_name=="PS DROP CALL RATE_NOM [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"HD",trend_ws)
                
                if(kpi_name=="VOLTE ERAB SETUP SUCCESS RATE_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HK",trend_ws)
            
                if(kpi_name=="VOLTE ERAB SETUP SUCCESS RATE_DENOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HR",trend_ws)
            
                if(kpi_name=="VOLTE INTRAF HOSR EXEC_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HY",trend_ws)

                if(kpi_name=="VOLTE INTRAF HOSR EXEC_DENOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IF",trend_ws)

                if(kpi_name=="VOLTE INTERF HOSR EXEC_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IM",trend_ws)

                if(kpi_name=="VOLTE INTERF HOSR EXEC_DENOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IT",trend_ws)

                if(kpi_name=="MV_CSFB REDIRECTION SUCCESS RATE_NOM [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"JA",trend_ws)
                if(kpi_name=="MV_CSFB REDIRECTION SUCCESS RATE_DENOM [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"JG",trend_ws)  

                ###################### mixed ###############################

                if(kpi_name=="LTE SgNb addition Success Rate %".upper()):
                    overwrite(pivot_fdd,kpi_name,"JM",trend_ws)   
                
                if(kpi_name=="LTE SgNb addition Success Rate_Nom".upper()):
                    overwrite(pivot_fdd,kpi_name,"JT",trend_ws)

                if(kpi_name=="LTE SgNb addition Success Rate_Denom".upper()):
                    overwrite(pivot_fdd,kpi_name,"JZ",trend_ws)
            
        

        # for tdd
        pivot_fdd=perticular_tech(["_T1_","_T2_"],site_list)
        trend_ws=trend_wb["L2300"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC SETUP SUCCESS RATE [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AK",trend_ws)

                if(kpi_name=="MV_ERAB SETUP SUCCESS RATE [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AR",trend_ws)

                if(kpi_name=="MV_PS DROP CALL RATE % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AY",trend_ws)

                if(kpi_name=="MV_DL USER THROUGHPUT_MBPS [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BF",trend_ws)

                if(kpi_name=="MV_UL USER THROUGHPUT_MBPS [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BM",trend_ws)

                if(kpi_name=="MV_PS HANDOVER SUCCESS RATE [LTE INTRA SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BT",trend_ws)
                
                if(kpi_name=="MV_PS HANDOVER SUCCESS RATE [LTE INTER SYSTEM] [CDBH]"): 
                    overwrite(pivot_fdd,kpi_name,"CA",trend_ws)
                
                # if(kpi_name=="MV_CSFB REDIRECTION SUCCESS RATE [CDBH]"):
                #     overwrite(pivot_fdd,kpi_name,"CH",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN AVERAGE CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CO",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CV",trend_ws)
                
                if(kpi_name=="MV_VOLTE ERAB SETUP SUCCESS RATE [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DC",trend_ws)
                
                if(kpi_name=="MV_VOLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DJ",trend_ws)
                
                if(kpi_name=="MV_VOLTE PACKET LOSS DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DQ",trend_ws)
                
                if(kpi_name=="MV_VOLTE PACKET LOSS UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DX",trend_ws)
                
                if(kpi_name=="MV_VOLTE INTRAF HOSR EXEC [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"EE",trend_ws)

                if(kpi_name=="MV_4G DATA VOLUME_GB"):
                    overwrite(pivot_fdd,kpi_name,"EL",trend_ws)

                if(kpi_name=="MV_VOLTE INTERF HOSR EXEC [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ES",trend_ws)

                if(kpi_name=="MV_AVERAGE NUMBER OF USED DL PRBS [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EZ",trend_ws)
                if(kpi_name=="MV_VOLTE TRAFFIC"):
                    overwrite(pivot_fdd,kpi_name,"FG",trend_ws)
                if(kpi_name=="MV_RADIO NW AVAILABILITY"):
                    overwrite(pivot_fdd,kpi_name,"FN",trend_ws)
                       

               
                #################################### nom/Denom #########################
                if(kpi_name=="PS HANDOVER SUCCESS RATE_NOM [LTE INTRA SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"FU",trend_ws)
                
                if(kpi_name=="PS HANDOVER SUCCESS RATE_DENOM [LTE INTRA SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GB",trend_ws)
                
                if(kpi_name=="PS HANDOVER SUCCESS RATE_NOM [LTE INTER SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GI",trend_ws)
                
                if(kpi_name=="PS HANDOVER SUCCESS RATE_DENOM [LTE INTER SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GP",trend_ws)
                
                if(kpi_name=="VOLTE DCR_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"GW",trend_ws)
                
                if(kpi_name=="PS DROP CALL RATE_NOM [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"HD",trend_ws)
                
                if(kpi_name=="VOLTE ERAB SETUP SUCCESS RATE_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HK",trend_ws)
            
                if(kpi_name=="VOLTE ERAB SETUP SUCCESS RATE_DENOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HR",trend_ws)
            
                if(kpi_name=="VOLTE INTRAF HOSR EXEC_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HY",trend_ws)

                if(kpi_name=="VOLTE INTRAF HOSR EXEC_DENOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IF",trend_ws)

                if(kpi_name=="VOLTE INTERF HOSR EXEC_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IM",trend_ws)

                if(kpi_name=="VOLTE INTERF HOSR EXEC_DENOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IT",trend_ws)

                if(kpi_name=="MV_CSFB REDIRECTION SUCCESS RATE_NOM [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"JA",trend_ws)
                if(kpi_name=="MV_CSFB REDIRECTION SUCCESS RATE_DENOM [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"JG",trend_ws)  

                ###################### mixed ###############################

                if(kpi_name=="LTE SgNb addition Success Rate %".upper()):
                    overwrite(pivot_fdd,kpi_name,"JM",trend_ws)   
                
                if(kpi_name=="LTE SgNb addition Success Rate_Nom".upper()):
                    overwrite(pivot_fdd,kpi_name,"JT",trend_ws)

                if(kpi_name=="LTE SgNb addition Success Rate_Denom".upper()):
                    overwrite(pivot_fdd,kpi_name,"JZ",trend_ws)

        pivot_fdd=perticular_tech(["_F8_"],site_list)
        trend_ws=trend_wb["L900"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC SETUP SUCCESS RATE [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AK",trend_ws)

                if(kpi_name=="MV_ERAB SETUP SUCCESS RATE [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AR",trend_ws)

                if(kpi_name=="MV_PS DROP CALL RATE % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AY",trend_ws)

                if(kpi_name=="MV_DL USER THROUGHPUT_MBPS [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BF",trend_ws)

                if(kpi_name=="MV_UL USER THROUGHPUT_MBPS [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BM",trend_ws)

                if(kpi_name=="MV_PS HANDOVER SUCCESS RATE [LTE INTRA SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BT",trend_ws)
                
                if(kpi_name=="MV_PS HANDOVER SUCCESS RATE [LTE INTER SYSTEM] [CDBH]"): 
                    overwrite(pivot_fdd,kpi_name,"CA",trend_ws)
                
                # if(kpi_name=="MV_CSFB REDIRECTION SUCCESS RATE [CDBH]"):
                #     overwrite(pivot_fdd,kpi_name,"CH",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN AVERAGE CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CO",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CV",trend_ws)
                
                if(kpi_name=="MV_VOLTE ERAB SETUP SUCCESS RATE [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DC",trend_ws)
                
                if(kpi_name=="MV_VOLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DJ",trend_ws)
                
                if(kpi_name=="MV_VOLTE PACKET LOSS DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DQ",trend_ws)
                
                if(kpi_name=="MV_VOLTE PACKET LOSS UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DX",trend_ws)
                
                if(kpi_name=="MV_VOLTE INTRAF HOSR EXEC [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"EE",trend_ws)

                if(kpi_name=="MV_4G DATA VOLUME_GB"):
                    overwrite(pivot_fdd,kpi_name,"EL",trend_ws)

                if(kpi_name=="MV_VOLTE INTERF HOSR EXEC [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ES",trend_ws)

                if(kpi_name=="MV_AVERAGE NUMBER OF USED DL PRBS [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EZ",trend_ws)
                if(kpi_name=="MV_VOLTE TRAFFIC"):
                    overwrite(pivot_fdd,kpi_name,"FG",trend_ws)
                if(kpi_name=="MV_RADIO NW AVAILABILITY"):
                    overwrite(pivot_fdd,kpi_name,"FN",trend_ws)
                       

               
                #################################### nom/Denom #########################
                if(kpi_name=="PS HANDOVER SUCCESS RATE_NOM [LTE INTRA SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"FU",trend_ws)
                
                if(kpi_name=="PS HANDOVER SUCCESS RATE_DENOM [LTE INTRA SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GB",trend_ws)
                
                if(kpi_name=="PS HANDOVER SUCCESS RATE_NOM [LTE INTER SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GI",trend_ws)
                
                if(kpi_name=="PS HANDOVER SUCCESS RATE_DENOM [LTE INTER SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GP",trend_ws)
                
                if(kpi_name=="VOLTE DCR_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"GW",trend_ws)
                
                if(kpi_name=="PS DROP CALL RATE_NOM [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"HD",trend_ws)
                
                if(kpi_name=="VOLTE ERAB SETUP SUCCESS RATE_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HK",trend_ws)
            
                if(kpi_name=="VOLTE ERAB SETUP SUCCESS RATE_DENOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HR",trend_ws)
            
                if(kpi_name=="VOLTE INTRAF HOSR EXEC_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HY",trend_ws)

                if(kpi_name=="VOLTE INTRAF HOSR EXEC_DENOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IF",trend_ws)

                if(kpi_name=="VOLTE INTERF HOSR EXEC_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IM",trend_ws)

                if(kpi_name=="VOLTE INTERF HOSR EXEC_DENOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IT",trend_ws)

                if(kpi_name=="MV_CSFB REDIRECTION SUCCESS RATE_NOM [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"JA",trend_ws)
                if(kpi_name=="MV_CSFB REDIRECTION SUCCESS RATE_DENOM [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"JG",trend_ws)  

                ###################### mixed ###############################

                if(kpi_name=="LTE SgNb addition Success Rate %".upper()):
                    overwrite(pivot_fdd,kpi_name,"JM",trend_ws)   
                
                if(kpi_name=="LTE SgNb addition Success Rate_Nom".upper()):
                    overwrite(pivot_fdd,kpi_name,"JT",trend_ws)

                if(kpi_name=="LTE SgNb addition Success Rate_Denom".upper()):
                    overwrite(pivot_fdd,kpi_name,"JZ",trend_ws)




        pivot_fdd=perticular_tech(["_F1_"],site_list)
        trend_ws=trend_wb["L2100"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC SETUP SUCCESS RATE [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AK",trend_ws)

                if(kpi_name=="MV_ERAB SETUP SUCCESS RATE [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AR",trend_ws)

                if(kpi_name=="MV_PS DROP CALL RATE % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AY",trend_ws)

                if(kpi_name=="MV_DL USER THROUGHPUT_MBPS [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BF",trend_ws)

                if(kpi_name=="MV_UL USER THROUGHPUT_MBPS [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BM",trend_ws)

                if(kpi_name=="MV_PS HANDOVER SUCCESS RATE [LTE INTRA SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BT",trend_ws)
                
                if(kpi_name=="MV_PS HANDOVER SUCCESS RATE [LTE INTER SYSTEM] [CDBH]"): 
                    overwrite(pivot_fdd,kpi_name,"CA",trend_ws)
                
                # if(kpi_name=="MV_CSFB REDIRECTION SUCCESS RATE [CDBH]"):
                #     overwrite(pivot_fdd,kpi_name,"CH",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN AVERAGE CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CO",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CV",trend_ws)
                
                if(kpi_name=="MV_VOLTE ERAB SETUP SUCCESS RATE [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DC",trend_ws)
                
                if(kpi_name=="MV_VOLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DJ",trend_ws)
                
                if(kpi_name=="MV_VOLTE PACKET LOSS DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DQ",trend_ws)
                
                if(kpi_name=="MV_VOLTE PACKET LOSS UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DX",trend_ws)
                
                if(kpi_name=="MV_VOLTE INTRAF HOSR EXEC [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"EE",trend_ws)

                if(kpi_name=="MV_4G DATA VOLUME_GB"):
                    overwrite(pivot_fdd,kpi_name,"EL",trend_ws)

                if(kpi_name=="MV_VOLTE INTERF HOSR EXEC [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ES",trend_ws)

                if(kpi_name=="MV_AVERAGE NUMBER OF USED DL PRBS [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EZ",trend_ws)
                if(kpi_name=="MV_VOLTE TRAFFIC"):
                    overwrite(pivot_fdd,kpi_name,"FG",trend_ws)
                if(kpi_name=="MV_RADIO NW AVAILABILITY"):
                    overwrite(pivot_fdd,kpi_name,"FN",trend_ws)
                       

               
                #################################### nom/Denom #########################
                if(kpi_name=="PS HANDOVER SUCCESS RATE_NOM [LTE INTRA SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"FU",trend_ws)
                
                if(kpi_name=="PS HANDOVER SUCCESS RATE_DENOM [LTE INTRA SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GB",trend_ws)
                
                if(kpi_name=="PS HANDOVER SUCCESS RATE_NOM [LTE INTER SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GI",trend_ws)
                
                if(kpi_name=="PS HANDOVER SUCCESS RATE_DENOM [LTE INTER SYSTEM] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GP",trend_ws)
                
                if(kpi_name=="VOLTE DCR_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"GW",trend_ws)
                
                if(kpi_name=="PS DROP CALL RATE_NOM [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"HD",trend_ws)
                
                if(kpi_name=="VOLTE ERAB SETUP SUCCESS RATE_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HK",trend_ws)
            
                if(kpi_name=="VOLTE ERAB SETUP SUCCESS RATE_DENOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HR",trend_ws)
            
                if(kpi_name=="VOLTE INTRAF HOSR EXEC_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HY",trend_ws)

                if(kpi_name=="VOLTE INTRAF HOSR EXEC_DENOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IF",trend_ws)

                if(kpi_name=="VOLTE INTERF HOSR EXEC_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IM",trend_ws)

                if(kpi_name=="VOLTE INTERF HOSR EXEC_DENOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IT",trend_ws)

                if(kpi_name=="MV_CSFB REDIRECTION SUCCESS RATE_NOM [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"JA",trend_ws)
                if(kpi_name=="MV_CSFB REDIRECTION SUCCESS RATE_DENOM [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"JG",trend_ws)  

                ###################### mixed ###############################

                if(kpi_name=="LTE SgNb addition Success Rate %".upper()):
                    overwrite(pivot_fdd,kpi_name,"JM",trend_ws)   
                
                if(kpi_name=="LTE SgNb addition Success Rate_Nom".upper()):
                    overwrite(pivot_fdd,kpi_name,"JT",trend_ws)

                if(kpi_name=="LTE SgNb addition Success Rate_Denom".upper()):
                    overwrite(pivot_fdd,kpi_name,"JZ",trend_ws)


        # # ##print(g2_tech(g2_site)[1])
        # pivote_g2=g2_tech(g2_site)[1]
        # # pivote_g2.to_excel("g2_pivoted.xlsx")
        # trend_ws=trend_wb["2G"]



        # for kpi_name in g2_kpi:
        #     if(kpi_name=="SDCCH Blocking Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"L","M","N","O","P",trend_ws)
            
        #     if(kpi_name=="SDCCH Drop Call Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"S","T","U","V","W",trend_ws)
            
        #     if(kpi_name=="TCH Blocking Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"Z","AA","AB","AC","AD",trend_ws)

        #     if(kpi_name=="Drop Call Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"AG","AH","AI","AJ","AK",trend_ws)
            
        #     if(kpi_name=="Handover Success Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"AO","AP","AQ","AR","AS",trend_ws)
            

        #     if(kpi_name=="RX Quality"):
        #         overwrite_g2(pivote_g2,kpi_name,"AV","AW","AX","AY","AZ",trend_ws)
    
        #     if(kpi_name=="RACH Success Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"BC","BD","BE","BF","BG",trend_ws)

            
        #     if(kpi_name=="Total Voice Traffic [BBH]"):
        #         overwrite_g2(pivote_g2,kpi_name,"BJ","BK","BL","BM","BN",trend_ws)

            
        #     if(kpi_name=="TCH Assignment Success Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"BP","BQ","BR","BS","BT",trend_ws)

            


        print("processing finished")
        print("saving file....")
        output_path=os.path.join(door_root,"output","Tnch_trend_output.xlsx")
        trend_wb.save(output_path)
        print("file saved")
        # load_excel_data()
        # directory = "Tnch_OUTPUT_trend"
    
        # # Parent Directory path
        # parent_dir = "C:/Users/dell7480/Desktop/" 
        # # Path
        # path = os.path.join(parent_dir, directory)
        # ##print(path)
        # #Create the directory
        # if(not (os.path.isdir(path))):
        #     os.makedirs(path)
        #     ##print("Directory '% s' created" % directory)
        # path= path + "/trend_output.xlsx"
        # trend_wb.save(path)
        download_path=os.path.join(MEDIA_URL,"Original_trend","Tnch","Trend","output","Tnch_trend_output.xlsx")
        
        return Response({"status":True,"message":"Succesfully uploaded","missing_sites":message, "Download_url":download_path})
    
    except Exception as e:
        
        return Response({
            'status':'undefind',
            'message':str(e),
        })
