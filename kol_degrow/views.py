from tkinter import filedialog, messagebox, ttk
from openpyxl import workbook,load_workbook
import pandas as pd
import numpy as np
from datetime import date, timedelta
from datetime import datetime
from openpyxl import Workbook
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.core.files.storage import FileSystemStorage
from datetime import date
import glob
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
import os

@api_view(['POST'])
def kol_degrow(request):
    pre_kpi=request.FILES['PRE'] if 'PRE' in request.FILES else None
    print("______________pre kpi______--",pre_kpi)
    if pre_kpi:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        print("___________location________",location)
        fs=FileSystemStorage(location=location)
        file=fs.save(pre_kpi.name,pre_kpi)
        file_path=fs.path(file)
        df_pre=pd.read_excel(file_path)
        print(df_pre)
        os.remove(path=file_path)
    post_kpi=request.FILES['POST'] if 'POST' in request.FILES else None
    print("______________post kpi______--",post_kpi)
    if post_kpi:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        print("___________location________",location)
        fs=FileSystemStorage(location=location)
        file=fs.save(post_kpi.name,post_kpi)
        file_path=fs.path(file)
        df_post=pd.read_excel(file_path)
        print(df_post)
        os.remove(path=file_path)
    site_list=request.FILES['site'] if 'site' in request.FILES else None
    print("______________site______--",site_list)
    if site_list:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        print("___________location________",location)
        fs=FileSystemStorage(location=location)
        file_sheet=fs.save(site_list.name,site_list)
        file_path=fs.path(file_sheet)
        site_list=pd.read_excel(file_path)
        print(site_list)
        os.remove(path=file_path) 


    door_path=os.path.join(MEDIA_ROOT,'trends','koldegrow')

    # df_pre=pd.read_excel("actual input/19 JUNE_WB_PRE_KPI.xlsx")
    df_pre['Short name']=df_pre['Short name'].fillna(method='ffill')
    #df.fillna(method='ffill',inplace=true)
    df_pre.columns.values[1]='Date'
    df_pre['band2']=[ band.split('_')[2] for band in df_pre['Short name']]
    df_pre_site_split=[site.split('_')[-2][:-1] for site in df_pre['Short name']]
    df_pre.insert(0,'site_id',df_pre_site_split)

    df_pre['band2']=df_pre['band2'].replace(['F1','F3','F8'],['L2100','L1800','L900'])
    # df_pre['band2']=df_pre['band2'].fillna(method='ffill')

    tech=[]
    for band1 in df_pre['band2']:
        if('T1' in band1 or 'T2' in band1):
            if('T1' in band1):
                band='TDDC2'
            if('T2' in band1):
                band='TDDC1'
            tech.append(band) 
        else:
            band=band1
            tech.append(band)
    df_pre.insert(1,'band1',tech) 

    df_pre=df_pre.drop('band2',axis=1)############### drop ######

    Pre_path1=os.path.join(door_path,'process output','desired_pre.xlsx')
    df_pre.to_excel(Pre_path1,index=False)
    ########################## site ##########################
    site_path=os.path.join(door_path,'project file','pre_post site.xlsx')
    site_list=site_path
    site_list=pd.read_excel(site_list)
    print("_____________________site",site_list)

    ######################## filter ##################################
    filter_pre=df_pre[(df_pre.site_id.isin(list(site_list['site_id'])))]
    Pre_path2=os.path.join(door_path,'process output','filter_pre.xlsx')
    filter_pre.to_excel(Pre_path2,index=False)
    print("_________________filter_______",filter_pre)

    # filtered_df1=pd.read_excel('filter_pre.xlsx')
    ############################### pivot #################################
    kpi=["MV_4G Data Volume_GB","MV_DL User Throughput_Kbps [CDBH]","RRC Setup Success Rate [CDBH]","MV_VoLTE DCR [CBBH]","ERAB Setup Success Rate [CDBH]",
        "MV_Max Connected User","VoLTE Traffic [CBBH]","DL PRB Utilisation [CDBH]","MV_UL User Throughput_Kbps [CDBH]"]

    ###################################################################################################################

    pivot_pre_vol=filter_pre.pivot_table(columns="band1",index=['Date'],aggfunc={'MV_4G Data Volume_GB':'sum',"VoLTE Traffic [CBBH]":'sum','MV_DL User Throughput_Kbps [CDBH]':'mean',
                                                                                'RRC Setup Success Rate [CDBH]':'mean','MV_VoLTE DCR [CBBH]':'mean','ERAB Setup Success Rate [CDBH]':'mean',
                                                                                'MV_Max Connected User':'mean','DL PRB Utilisation [CDBH]':'mean','MV_UL User Throughput_Kbps [CDBH]':'mean'})
        

    vol=pivot_pre_vol.T

    print(vol)
    pre_path3=os.path.join(door_path,'process output','pivoted_pre.xlsx')
    vol.to_excel(pre_path3)

    ##########################################################################################################
    # df_post=pd.read_excel("actual input/19 JUNE WB_POST_KPI.xlsx")
    df_post['Short name']=df_pre['Short name'].fillna(method='ffill')
    #df.fillna(method='ffill',inplace=true)
    df_post.columns.values[1]='Date'
    df_post['band2']=[ band.split('_')[2] for band in df_pre['Short name']]
    df_post_site_split=[site.split('_')[-2][:-1] for site in df_pre['Short name']]
    df_post.insert(0,'site_id',df_post_site_split)

    df_post['band2']=df_post['band2'].replace(['F1','F3','F8'],['L2100','L1800','L900'])
    # df_pre['band2']=df_pre['band2'].fillna(method='ffill')

    tech=[]
    for band1 in df_post['band2']:
        if('T1' in band1 or 'T2' in band1):
            if('T1' in band1):
                band='TDDC2'
            if('T2' in band1):
                band='TDDC1'
            tech.append(band) 
        else:
            band=band1
            tech.append(band)
    df_post.insert(1,'band1',tech) 

    df_post=df_post.drop('band2',axis=1)############### drop ######
    post_path1=os.path.join(door_path,'process output','desired_post.xlsx')
    df_post.to_excel(post_path1,index=False)

    ######################## filter ##################################
    filter_post=df_post[(df_post.site_id.isin(list(site_list['site_id'])))]
    post_path2=os.path.join(door_path,'process output','filter_post.xlsx')
    filter_post.to_excel(post_path2,index=False)
    print("_________________filter_______",filter_post)

    ############################### pivot #################################
    # kpi=["MV_4G Data Volume_GB","MV_DL User Throughput_Kbps [CDBH]","RRC Setup Success Rate [CDBH]","MV_VoLTE DCR [CBBH]","ERAB Setup Success Rate [CDBH]",
    #     "MV_Max Connected User","VoLTE Traffic [CBBH]","DL PRB Utilisation [CDBH]","MV_UL User Throughput_Kbps [CDBH]"]

    ###################################################################################################################

    pivot_post_vol=filter_post.pivot_table(columns="band1",index=['Date'],aggfunc={'MV_4G Data Volume_GB':'sum',"VoLTE Traffic [CBBH]":'sum','MV_DL User Throughput_Kbps [CDBH]':'mean',
                                                                                'RRC Setup Success Rate [CDBH]':'mean','MV_VoLTE DCR [CBBH]':'mean','ERAB Setup Success Rate [CDBH]':'mean',
                                                                                'MV_Max Connected User':'mean','DL PRB Utilisation [CDBH]':'mean','MV_UL User Throughput_Kbps [CDBH]':'mean'})

    vol_post=pivot_post_vol.T
    print(vol)
    # vol_post=vol_post.select_dtypes(include='number')

    vol_post['Pre Avg']=round(vol.mean(axis=1),2)  
    post_path3=os.path.join(door_path,'process output','pivoted_post.xlsx')
    vol_post.to_excel(post_path3)
        
    #######################################################################
    concat_pre_post = pd.concat([vol, vol_post],axis=1) 

    concat_pre_post['Post Avg']=round(concat_pre_post.iloc[:,[5,6,7,8,9]].mean(axis=1),2)############
    concat_pre_post['Delta']=round(concat_pre_post['Post Avg']-vol_post['Pre Avg'],2)
    concat_pre_post['Change%']=round(concat_pre_post['Delta']/vol_post['Pre Avg'],2)
        
    concat_pre_post.replace([np.inf, -np.inf], 'NA', inplace=True)
    concat_pre_post['Change%'].fillna(0,inplace=True)
    concat_path=os.path.join(door_path,'process output','concat_pre_post.xlsx')
    concat_pre_post.to_excel(concat_path)


    df1=pd.read_excel(concat_path)

    ######################################################################

    df1.rename(columns={'Unnamed: 0':'KPI'}, inplace=True )
    df1.rename(columns={'band1':'TECH'},inplace=True)
    df1.set_index(['KPI','TECH'])
    df1['Change%'].fillna(0,inplace=True)
    df2=round(df1,2)
    concat_path2=os.path.join(door_path,'process output','concat_pre_post_add.xlsx')
    df2.to_excel(concat_path2,index=False,header=True,startrow=1)####### TO CHIFT A HEADER USE START ROW ###
    print("_____________concat_______")
    print(df1)
        

    df3=concat_path2

    ################################## try excel code ####################################     
    # try:
    #     df2='process output/concat_pre_post2.xlsx'
    # except Exception as e:
    #     print(f'Error:{e}') 
    #     exit()   

    ########################################### TO MERGE CELL #######################################################
    # STR='template/degrow.xlsx'
    wb=openpyxl.load_workbook(df3)
    ws=wb.active
    ######################################### INSERT A ROW #################################################
    new_row_data=['']
    ws.append(new_row_data)
    insert_index=8
    insert_index1=14
    insert_index2=20
    insert_index3=26
    insert_index4=32
    insert_index5=38
    insert_index6=44
    insert_index7=50
    insert_index8=56


    ws.insert_rows(insert_index, amount=1)
    ws.insert_rows(insert_index1, amount=1)
    ws.insert_rows(insert_index2, amount=1)
    ws.insert_rows(insert_index3, amount=1)
    ws.insert_rows(insert_index4, amount=1)
    ws.insert_rows(insert_index5, amount=1)
    ws.insert_rows(insert_index6, amount=1)
    ws.insert_rows(insert_index7, amount=1)
    ws.insert_rows(insert_index8, amount=1)

    ###################################### FONT SIZE OF ALL SHEET #########################################################################

    font=Font(size=9)
    alignment = Alignment(horizontal='center', vertical='center')
    for col_cells in ws.columns:
        for cell in col_cells:
            cell.font=font
            cell.alignment=alignment 
    ###################################### MERGE A CELL ######################################
    ws.merge_cells('A1:B1')
    ws.merge_cells('A3:A8')
    ws.merge_cells('A9:A14')
    ws.merge_cells('A15:A20')
    ws.merge_cells('A21:A26')
    ws.merge_cells('A27:A32')
    ws.merge_cells('A33:A38')
    ws.merge_cells('A39:A44')
    ws.merge_cells('A45:A50')
    ws.merge_cells('A51:A56')



    ws.merge_cells('C1:G1')
    ws.merge_cells('H1:L1')
    ws.merge_cells('M1:M2')
    ws.merge_cells('N1:N2')
    ws.merge_cells('O1:O2')
    ws.merge_cells('P1:P2')
    ws.merge_cells('Q1:Q2')
    ws.merge_cells('Q3:Q8')
    ws.merge_cells('Q9:Q14')
    ws.merge_cells('Q15:Q20')
    ws.merge_cells('Q21:Q26')
    ws.merge_cells('Q27:Q32')
    ws.merge_cells('Q33:Q38')
    ws.merge_cells('Q39:Q44')
    ws.merge_cells('Q45:Q50')
    ws.merge_cells('Q51:Q56')


    ws['A1']=''                    #################### PUT NAME OF COLUMNS ##################
    ws['C1']='Pre'
    ws['H1']='Post'
    ws['M1']='Pre Avg'
    ws['N1']='Post Avg'
    ws['O1']='Delta'
    ws['P1']='Change%'
    ws['Q1']='Remarks (Working Days KPIs)'
    ws['Q3']='Overll sector traffic should be maintained'
    ws['Q9']='Overll sector traffic should be maintained'
    ws['Q15']='Degrow sector cells should not degrade <3 Mbps'
    ws['Q21']='Degrow sector cells should not degrade < 500 Kbps'
    ws['Q27']=''
    ws['Q33']=''
    ws['Q39']='Overll sector KPI should be maintained'
    ws['Q45']='Overll sector KPI should be maintained'
    ws['Q51']='Overll sector KPI should be maintained'
    ws['B8']='NW'
    ws['B14']='NW'
    ws['B8']='NW'
    ws['B20']='NW'
    ws['B26']='NW'
    ws['B32']='NW'
    ws['B38']='NW'
    ws['B44']='NW'
    ws['B50']='NW'
    ws['B56']='NW'
    ###################################### MEAN IN WOKSHEET OF NW #######################
    column_names=[2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    # column_names=['C']
    start_row=3
    end_row=7
    output_columns=['D','E','F','G','H','I','J','K','L','M','N','O','P','Q']
    for i, column_name in enumerate(column_names):
        column_values = [ws.cell(row=row, column=column_name + 1).value for row in range(start_row, end_row + 1) if ws.cell(row=row, column=column_name+ 1).value is not None] 
        mean_value = round(sum(column_values)/ len(column_values),2)
        output_column = output_columns[i]
        print('_______________mean_values____________',mean_value)
        ws.cell(row=end_row+1 , column=ord(output_column) - ord('A') , value=mean_value)
    #######################################################################
    column_names=[2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    # column_names=['C']
    start_row=9
    end_row=13
    output_columns=['D','E','F','G','H','I','J','K','L','M','N','O','P','Q']
    for i, column_name in enumerate(column_names):
        column_values = [ws.cell(row=row, column=column_name + 1).value for row in range(start_row, end_row + 1) if ws.cell(row=row, column=column_name+ 1).value is not None] 
        mean_value = round(sum(column_values)/ len(column_values),2)
        output_column = output_columns[i]
        print('_______________mean_values____________',mean_value)
        ws.cell(row=end_row+1 , column=ord(output_column) - ord('A') , value=mean_value) 
    ############################################################################
    column_names=[2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    # column_names=['C']
    start_row=15
    end_row=19
    output_columns=['D','E','F','G','H','I','J','K','L','M','N','O','P','Q']
    for i, column_name in enumerate(column_names):
        column_values = [ws.cell(row=row, column=column_name + 1).value for row in range(start_row, end_row + 1) if ws.cell(row=row, column=column_name+ 1).value is not None] 
        mean_value = round(sum(column_values)/ len(column_values),2)
        output_column = output_columns[i]
        print('_______________mean_values____________',mean_value)
        ws.cell(row=end_row+1 , column=ord(output_column) - ord('A') , value=mean_value)              
    ########################################################################
    column_names=[2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    # column_names=['C']
    start_row=21
    end_row=25
    output_columns=['D','E','F','G','H','I','J','K','L','M','N','O','P','Q']
    for i, column_name in enumerate(column_names):
        column_values = [ws.cell(row=row, column=column_name + 1).value for row in range(start_row, end_row + 1) if ws.cell(row=row, column=column_name+ 1).value is not None] 
        mean_value = round(sum(column_values)/ len(column_values),2)
        output_column = output_columns[i]
        print('_______________mean_values____________',mean_value)
        ws.cell(row=end_row+1 , column=ord(output_column) - ord('A') , value=mean_value)       
    ##################################################################
    column_names=[2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    # column_names=['C']
    start_row=27
    end_row=31
    output_columns=['D','E','F','G','H','I','J','K','L','M','N','O','P','Q']
    for i, column_name in enumerate(column_names):
        column_values = [ws.cell(row=row, column=column_name + 1).value for row in range(start_row, end_row + 1) if ws.cell(row=row, column=column_name+ 1).value is not None] 
        mean_value = round(sum(column_values)/ len(column_values),2)
        output_column = output_columns[i]
        print('_______________mean_values____________',mean_value)
        ws.cell(row=end_row+1 , column=ord(output_column) - ord('A') , value=mean_value)
    #################################################################
    column_names=[2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    # column_names=['C']
    start_row=33
    end_row=37
    output_columns=['D','E','F','G','H','I','J','K','L','M','N','O','P','Q']
    for i, column_name in enumerate(column_names):
        column_values = [ws.cell(row=row, column=column_name + 1).value for row in range(start_row, end_row + 1) if ws.cell(row=row, column=column_name+ 1).value is not None] 
        mean_value = round(sum(column_values)/ len(column_values))
        output_column = output_columns[i]
        print('_______________mean_values____________',mean_value)
        ws.cell(row=end_row+1 , column=ord(output_column) - ord('A') , value=mean_value)
    ###################################################################
    column_names=[2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    # column_names=['C']
    start_row=39
    end_row=43
    output_columns=['D','E','F','G','H','I','J','K','L','M','N','O','P','Q']
    for i, column_name in enumerate(column_names):
        column_values = [ws.cell(row=row, column=column_name + 1).value for row in range(start_row, end_row + 1) if ws.cell(row=row, column=column_name+ 1).value is not None] 
        mean_value = round(sum(column_values)/ len(column_values),2)
        output_column = output_columns[i]
        print('_______________mean_values____________',mean_value)
        ws.cell(row=end_row+1 , column=ord(output_column) - ord('A') , value=mean_value)
    #######################################################
    column_names=[2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    # column_names=['C']
    start_row=45
    end_row=49
    output_columns=['D','E','F','G','H','I','J','K','L','M','N','O','P','Q']
    for i, column_name in enumerate(column_names):
        column_values = [ws.cell(row=row, column=column_name + 1).value for row in range(start_row, end_row + 1) if ws.cell(row=row, column=column_name+ 1).value is not None] 
        mean_value = round(sum(column_values)/ len(column_values))
        output_column = output_columns[i]
        print('_______________mean_values____________',mean_value)
        ws.cell(row=end_row+1 , column=ord(output_column) - ord('A') , value=mean_value) 
    ##############################################################
    column_names=[2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    # column_names=['C']
    start_row=51
    end_row=55
    output_columns=['D','E','F','G','H','I','J','K','L','M','N','O','P','Q']
    for i, column_name in enumerate(column_names):
        column_values = [ws.cell(row=row, column=column_name + 1).value for row in range(start_row, end_row + 1) if ws.cell(row=row, column=column_name+ 1).value is not None] 
        mean_value = round(sum(column_values)/ len(column_values))
        output_column = output_columns[i]
        print('_______________mean_values____________',mean_value)
        ws.cell(row=end_row+1 , column=ord(output_column) - ord('A') , value=mean_value)               

    #################################### IN CENTER HEADER #####################################    
    merged_cell=ws['C1']
    merged_cell1=ws['H1']
    merged_cell2=ws['M1']
    merged_cell3=ws['N1']
    merged_cell4=ws['O1']
    merged_cell5=ws['P1']
    # merged_cell6=ws['Q1']
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell1.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell2.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell3.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell4.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell5.alignment = Alignment(horizontal='center', vertical='center')
    # merged_cell6.alignment = Alignment(horizontal='center', vertical='center')



    font_header_1 = Font(bold=True, size=9) 
    font_header_2 = Font(bold=True, size=9,color='FFFFFF') 

    # font=Font(size=9)
    # alignment = Alignment(horizontal='center', vertical='center')
        

    # for cell1 in ws['A2:A2']:
    #     cell1[0].font = font_header_2
    for cell1 in ws['A3:A56']:
        cell1[0].font = font_header_1    
    for cell2 in ws['B2:B56']:
        cell2[0].font = font_header_1 
    for cell3 in ws['C1:G1'] :
        cell3[0].font = font_header_1  
    for cell4 in ws['H1:L1'] :
        cell4[0].font = font_header_1  
    for cell5 in ws['M1:M1'] :
        cell5[0].font = font_header_1  
    for cell6 in ws['N1:N1'] :
        cell6[0].font = font_header_1 
    for cell7 in ws['O1:O1'] :
        cell7[0].font = font_header_2
    for cell8 in ws['P1:P1'] :
        cell8[0].font = font_header_1 
    for cell9 in ws['Q1:Q56'] :
        cell9[0].font = font_header_1 


    for cell in ws['Q1:Q56']:
        cell[0].alignment = Alignment(horizontal='center', vertical='center')
    for cell1 in ws['A3:A47']:
        cell1[0].alignment = Alignment(horizontal='center', vertical='center') 
    for cell2 in ws['B3:B47']:
        cell2[0].alignment = Alignment(horizontal='center', vertical='center')

    ############################ to change row wise color change ####################################
    row_number=2
    fillcolor='FFFFFF'

    for cell in ws[row_number]:
        cell.font = Font(color=fillcolor,bold=True,size=9)
    ######################################################################

    PURPLE="7030A0"
    CYAN="00B0F0"
    ORANGE="FFC000"
    light_blue='BDD7EE'

    ws["A1"].fill=PatternFill(patternType='solid',fgColor=CYAN)
    ws['C1'].fill=PatternFill(patternType='solid',fgColor=ORANGE)
    ws["H1"].fill=PatternFill(patternType='solid',fgColor=CYAN)
    ws["A2"].fill=PatternFill(patternType='solid',fgColor=PURPLE)
    ws["B2"].fill=PatternFill(patternType='solid',fgColor=PURPLE)
    ws["C2"].fill=PatternFill(patternType='solid',fgColor=PURPLE)  
    ws["D2"].fill=PatternFill(patternType='solid',fgColor=PURPLE) 
    ws["E2"].fill=PatternFill(patternType='solid',fgColor=PURPLE) 
    ws["F2"].fill=PatternFill(patternType='solid',fgColor=PURPLE) 
    ws["G2"].fill=PatternFill(patternType='solid',fgColor=PURPLE) 
    ws["H2"].fill=PatternFill(patternType='solid',fgColor=PURPLE) 
    ws["I2"].fill=PatternFill(patternType='solid',fgColor=PURPLE) 
    ws["J2"].fill=PatternFill(patternType='solid',fgColor=PURPLE) 
    ws["K2"].fill=PatternFill(patternType='solid',fgColor=PURPLE) 
    ws["L2"].fill=PatternFill(patternType='solid',fgColor=PURPLE)
    ws['M1'].fill=PatternFill(patternType='solid',fgColor=ORANGE)
    ws["N1"].fill=PatternFill(patternType='solid',fgColor=CYAN)
    ws["O1"].fill=PatternFill(patternType='solid',fgColor=PURPLE)
    ws["P1"].fill=PatternFill(patternType='solid',fgColor=light_blue)
    ws["Q1"].fill=PatternFill(patternType='solid',fgColor=light_blue)

    ####################################### to change border thin ###############################################
    start_row = 1
    end_row = 56
    start_column = 'A'
    end_column = 'Q'
    border_color = '3A3838'  # Replace with the desired RGB value for the brighter color
    border = Border(left=Side(border_style='thin', color=border_color),
                    right=Side(border_style='thin', color=border_color),
                    top=Side(border_style='thin', color=border_color),
                    bottom=Side(border_style='thin', color=border_color))
    start_column_index = openpyxl.utils.column_index_from_string(start_column)
    end_column_index = openpyxl.utils.column_index_from_string(end_column)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column_index, max_col=end_column_index):
        for cell in row:
            cell.border = border    
    #########################################################################################################
        
    saveoutput=os.path.join(door_path,'output','Degrow_output.xlsx')
    wb.save(saveoutput)  
    download_path=os.path.join(MEDIA_URL,'trends','koldegrow','output','Degrow_output.xlsx')
    return Response({"Status":True,"Message":"Successfully",'download_url':download_path})
        



    

