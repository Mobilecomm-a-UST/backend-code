from django.shortcuts import render

# Create your views here.

import os
import pandas as pd
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.http import FileResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
#from project_pat.settings import MEDIA_ROOT, MEDIA_URL
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from django.shortcuts import render



# ══════════════════════════════════════════════════════════════════════════════
# FOLDER SETUP
# ══════════════════════════════════════════════════════════════════════════════

main_folder  = os.path.join(settings.MEDIA_ROOT, "performance_tat")
input_path   = os.path.join(main_folder, "input")
output_path  = os.path.join(main_folder, "output")

os.makedirs(input_path,  exist_ok=True)
os.makedirs(output_path, exist_ok=True)


# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

LAYERS_4G_EXCLUDE   = {"3500", "L3500", "N3500", "N2600"}
LAYERS_5G_INCLUDE   = {"3500", "L3500", "N3500"}
LAYERS_4G5G_EXCLUDE = {"N2600"}

TAT_BUCKETS = [
    ("<7",   0,  7),
    
    ("8-12",  8, 12),
    ("13-21", 13, 21),
    (">21",   21, None),
]

CIRCLE_COMBINE = {
    'TNCH':    ['CN', 'TN'],
    'NESA':    ['AS', 'NE'],
    'HPHP':    ['HR', 'PB'],
    'WB/KOL':  ['WB', 'KO'],
}


HEADER_BG      = "1F4E79"
HEADER_FG      = "FFFFFF"
SUBHEADER_BG   = "2E75B6"
SUBHEADER_FG   = "FFFFFF"
GRAND_TOTAL_BG = "D6E4F0"
GRAND_TOTAL_FG = "1F4E79"
ALT_ROW_BG     = "EBF3FB"
TOTAL_COL_BG   = "D9E1F2"


# ══════════════════════════════════════════════════════════════════════════════
# FILE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _get_input_file():
    """Return full path of input file, or None if not found."""
    files = os.listdir(input_path) if os.path.exists(input_path) else []
    return os.path.join(input_path, files[0]) if files else None



def _clear_files(return_deleted=False):
    deleted = []
    for folder in [input_path, output_path]:
        if os.path.exists(folder):
            for f in os.listdir(folder):
                os.remove(os.path.join(folder, f))
                if return_deleted:
                    deleted.append(f)
    return deleted if return_deleted else None

def _save_uploaded_file(uploaded):
    """Save uploaded file to input folder with its original name."""
    save_path = os.path.join(input_path, uploaded.name)
    with open(save_path, 'wb+') as destination:
        for chunk in uploaded.chunks():
            destination.write(chunk)
    return save_path


def _validate_file(filepath):
    """
    Read uploaded file and check all required columns exist.
    Required: Circle, Offered Layer, On Air Date, Performance AT Status
    Returns (df, error_message). If valid, error_message is None.
    Called only once at upload time.
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls'):
        df = pd.read_excel(filepath)
    elif ext == '.csv':
        df = pd.read_csv(filepath)
    else:
        return None, f"Unsupported file type: {ext}"

    df.rename(columns={c: c.strip() for c in df.columns}, inplace=True)
    lower_map = {c.lower(): c for c in df.columns}

    required = {
        'circle':                'Circle',
        'offered layer':         'Offered Layer',
        'on air date':           'On Air Date',
        'performance at status': 'Performance AT Status',
    }

    rename = {}
    for key, canonical in required.items():
        if canonical in df.columns:
            rename[canonical] = canonical
        elif key in lower_map:
            rename[lower_map[key]] = canonical
        else:
            match = next((c for c in df.columns if key in c.lower()), None)
            if match:
                rename[match] = canonical
            else:
                return None, (
                    f"Column '{canonical}' not found. "
                    f"Available columns: {list(df.columns)}"
                )

    df.rename(columns=rename, inplace=True)
    df['On Air Date'] = pd.to_datetime(df['On Air Date'], errors='coerce').dt.date
    df.dropna(subset=['On Air Date'], inplace=True)

    return df, None


def _read_input_file(filepath):
    """
    Read already-validated input file.
    Called in generate — skips column validation since upload already confirmed it.
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls'):
        df = pd.read_excel(filepath)
    else:
        df = pd.read_csv(filepath)

    df.rename(columns={c: c.strip() for c in df.columns}, inplace=True)
    df['On Air Date'] = pd.to_datetime(df['On Air Date'], errors='coerce').dt.date
    df.dropna(subset=['On Air Date'], inplace=True)
    return df


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL STYLE HELPER
# ══════════════════════════════════════════════════════════════════════════════

def _get_thin_border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def _style_header_cell(cell, value, font_size, bg_color, fg_color=HEADER_FG):
    """Apply standard header styling to a merged heading cell."""
    cell.value     = value
    cell.font      = Font(name="Arial", bold=True, size=font_size, color=fg_color)
    cell.fill      = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ══════════════════════════════════════════════════════════════════════════════
# DATE HELPER
# ══════════════════════════════════════════════════════════════════════════════

def get_date_range_for_month(year: int, month: int):
    """
    26th of previous month to 25th of current month.
    Example: Jan 2026 → 2025-12-26 to 2026-01-25
    """
    end_date = date(year, month, 25)
    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year
    start_date = date(prev_year, prev_month, 26)
    return start_date, end_date


def _resolve_dates(request):
    """
    Shared date resolver for all generate APIs.
    Accepts either (year + month) or (start_date + end_date).
    Returns (start_date, end_date, error_response).
    If error_response is not None, return it immediately.
    """
    year_str  = request.data.get("year")
    month_str = request.data.get("month")
    start_str = request.data.get("start_date")
    end_str   = request.data.get("end_date")

    if year_str and month_str:
        try:
            year  = int(year_str)
            month = int(month_str)
            if not (1 <= month <= 12):
                raise ValueError
        except ValueError:
            return None, None, Response(
                {'status': False, 'message': 'Invalid year or month.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        start_date, end_date = get_date_range_for_month(year, month)
        return start_date, end_date, None

    elif start_str and end_str:
        try:
            start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
            end_date   = datetime.strptime(end_str,   "%Y-%m-%d").date()
        except ValueError:
            return None, None, Response(
                {'status': False, 'message': 'Invalid date format. Use YYYY-MM-DD.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if start_date > end_date:
            return None, None, Response(
                {'status': False, 'message': 'start_date must be before or equal to end_date.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        return start_date, end_date, None

    else:
        return None, None, Response(
            {'status': False, 'message': 'Provide either (year + month) or (start_date + end_date).'},
            status=status.HTTP_400_BAD_REQUEST
        )
#_________________________________________________

def _apply_circle_combine(df):
    """
    Replace individual circle codes with combined labels.
    e.g. CN and TN both become TNCH.
    """
    reverse_map = {}
    for combined, members in CIRCLE_COMBINE.items():
        for member in members:
            reverse_map[member] = combined
    df["Circle"] = df["Circle"].map(lambda x: reverse_map.get(str(x).strip(), x))
    return df

# ══════════════════════════════════════════════════════════════════════════════
# TAT LOGIC FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def assign_tat_bucket(tat_days):
    """Map number of days to a TAT bucket label."""
    
    for label, low, high in TAT_BUCKETS:
        if high is None:
            if tat_days > low:
                return label
        else:
            if low <= tat_days <= high:
                return label
    return None


def build_pivot_table(df, today):
    """
    Compute TAT days, assign buckets, return pivot:
        Rows    → circles sorted A to Z + Grand Total
        Columns → 3-5 | 5-7 | 7-12 | 12-21 | >21 | Total
   
    """
    df = df.copy()
    #df["On Air Date"] = pd.to_datetime(df["On Air Date"]).dt.date
    df["tat_days"]    = df["On Air Date"].apply(lambda d: (today - d).days)
    df["bucket"]      = df["tat_days"].apply(assign_tat_bucket)
    df = df.dropna(subset=["bucket"])

    bucket_order = [b[0] for b in TAT_BUCKETS]

    if df.empty:
        return pd.DataFrame(columns=["Circle"] + bucket_order + ["Total"])

    pivot = (
        df.groupby(["Circle", "bucket"])
        .size()
        .unstack(fill_value=0)
        .reindex(columns=bucket_order, fill_value=0)
        .reset_index()
        .rename(columns={"Circle": "Circle"})
        .sort_values("Circle")
        .reset_index(drop=True)
    )
    pivot["Total"] = pivot[bucket_order].sum(axis=1)

    grand = pivot[bucket_order + ["Total"]].sum().to_dict()
    grand["Circle"] = "Grand Total"
    pivot = pd.concat([pivot, pd.DataFrame([grand])], ignore_index=True)

    return pivot


def build_all_pivots(df, today, start_date, end_date):
    """
    Filter by date range + exclude Accepted status.
    Apply layer filters and return 3 pivot tables.

    Sheet rules:
        4G    → exclude 3500, L3500, N3500, N2600
        5G    → include only 3500, L3500, N3500
        4G+5G → exclude N2600 only
    """
    df = df.copy()
    df["On Air Date"] = pd.to_datetime(df["On Air Date"]).dt.date

    df = df[
        (df["On Air Date"] >= start_date) &
        (df["On Air Date"] <= end_date)
    ]
    df = df[df["Performance AT Status"].str.strip().str.lower() != "accepted"]

    empty = pd.DataFrame(columns=["Circle", "Offered Layer", "On Air Date"])

    if df.empty:
        return {
            "4G":    build_pivot_table(empty, today),
            "5G":    build_pivot_table(empty, today),
            "4G+5G": build_pivot_table(empty, today),
        }

    return {
        "4G":    build_pivot_table(df[~df["Offered Layer"].isin(LAYERS_4G_EXCLUDE)],   today),
        "5G":    build_pivot_table(df[df["Offered Layer"].isin(LAYERS_5G_INCLUDE)],    today),
        "4G+5G": build_pivot_table(df[~df["Offered Layer"].isin(LAYERS_4G5G_EXCLUDE)], today),
    }


def write_excel_sheet(ws, pivot, sheet_title, start_date, end_date):
    """Write one fully styled TAT sheet into an openpyxl worksheet."""
    bucket_cols   = [b[0] for b in TAT_BUCKETS]
    all_data_cols = bucket_cols + ["Total"]
    total_cols    = 1 + len(all_data_cols)
    headers       = ["Circle"] + all_data_cols
    thin_border   = _get_thin_border()

    # Row 1 — main heading
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    _style_header_cell(ws.cell(row=1, column=1), "Count of TAT", 14, HEADER_BG)
    ws.row_dimensions[1].height = 28

    # Row 2 — sheet title + date range
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    _style_header_cell(ws.cell(row=2, column=1), f"{sheet_title}  |  {start_date} to {end_date}", 11, SUBHEADER_BG)
    ws.row_dimensions[2].height = 20

    # Row 3 — column headers
    for col_idx, hdr in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col_idx, value=hdr)
        c.font      = Font(name="Arial", bold=True, size=10, color=SUBHEADER_FG)
        c.fill      = PatternFill("solid", fgColor=SUBHEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border
    ws.row_dimensions[3].height = 18

    if pivot.empty:
        ws.cell(row=4, column=1, value="No data for the selected date range.")
        return

    # Data rows
    for row_offset, (_, row) in enumerate(pivot.iterrows()):
        excel_row = row_offset + 4
        is_grand  = str(row.get("Circle", "")).strip() == "Grand Total"
        is_alt    = (row_offset % 2 == 1) and not is_grand

        row_bg = GRAND_TOTAL_BG if is_grand else (ALT_ROW_BG if is_alt else "FFFFFF")
        row_fg = GRAND_TOTAL_FG if is_grand else "000000"

        for col_idx, col_name in enumerate(headers, start=1):
            raw_val = row.get(col_name, "")
            try:
                val = int(raw_val) if col_name != "Circle" else str(raw_val)
            except (ValueError, TypeError):
                val = raw_val

            c = ws.cell(row=excel_row, column=col_idx, value=val)
            c.font      = Font(name="Arial", bold=is_grand, size=10, color=row_fg)
            c.alignment = Alignment(
                horizontal="left" if col_name == "Circle" else "center",
                vertical="center"
            )
            c.border = thin_border
            c.fill   = PatternFill(
                "solid",
                fgColor=TOTAL_COL_BG if (col_name == "Total" and not is_grand) else row_bg
            )
        ws.row_dimensions[excel_row].height = 16

    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10
    ws.freeze_panes = "B4"


def build_excel_workbook(pivot_data, start_date, end_date):
    """Build 3-sheet workbook, save to output folder, return filename."""
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name in ["4G", "5G", "4G+5G"]:
        ws = wb.create_sheet(title=sheet_name)
        write_excel_sheet(ws, pivot_data[sheet_name], sheet_name, start_date, end_date)

    out_filename = f"output_TAT_Report_{start_date}_to_{end_date}.xlsx"
    wb.save(os.path.join(output_path, out_filename))
    return out_filename



# ══════════════════════════════════════════════════════════════════════════════
# APIs
# ══════════════════════════════════════════════════════════════════════════════

@api_view(['POST', 'GET', 'DELETE'])
def upload_file(request):
    """
    POST   → upload input file (xlsx / xls / csv), validate columns
    GET    → check what input/output files currently exist
    DELETE → clear all input and output files
    """
    try:
        # ── POST: upload ───────────────────────────────────────────────────
        if request.method == 'POST':
            files = request.FILES.getlist('file')
            if not files:
                return Response(
                    {'status': False, 'error': 'No files uploaded.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            uploaded = files[0]
            ext      = os.path.splitext(uploaded.name)[1].lower()
            if ext not in ('.xlsx', '.xls', '.csv'):
                return Response(
                    {'status': False, 'error': f"Unsupported type '{ext}'. Use .xlsx, .xls, or .csv."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            _clear_files()
            save_path = _save_uploaded_file(uploaded)
            df, error = _validate_file(save_path)

            if error:
                os.remove(save_path)
                return Response(
                    {'status': False, 'error': error},
                    status=status.HTTP_422_UNPROCESSABLE_ENTITY
                )

            return Response({
                'status':    True,
                'message':   'File uploaded and validated successfully.',
                'filename':  uploaded.name,
                'rows_read': len(df),
            }, status=status.HTTP_200_OK)

        # ── GET: check files ───────────────────────────────────────────────
        elif request.method == 'GET':
            input_file = _get_input_file()
            input_file_info = (
                {'filename': os.path.basename(input_file), 'status': 'ready'}
                if input_file else None
            )

            output_files = []
            if os.path.exists(output_path):
                for filename in os.listdir(output_path):
                    if filename.endswith('.xlsx'):
                        output_files.append({
                            'filename':     filename,
                            'download_url': request.build_absolute_uri(
                                f"{settings.MEDIA_URL.rstrip('/')}/performance_tat/output/{filename}"
                            ),
                        })

            return Response({
                'status':       True,
                'message':      'Files found.' if input_file_info else 'No input file found.',
                'input_file':   input_file_info,
                'output_files': output_files,
            }, status=status.HTTP_200_OK)

        # ── DELETE: clear files ────────────────────────────────────────────
        elif request.method == 'DELETE':
            deleted_files = _clear_files(return_deleted=True)
            return Response({
                'status':        True,
                'message':       'Files deleted successfully.',
                'deleted_files': deleted_files,
            }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def generate_tat_report(request):
    """
    POST /api/performance_tat/generate/

    Body:
        { "start_date": "YYYY-MM-DD", "end_date": "YYYY-MM-DD" }

    Reads the already-uploaded input file, applies filters,
    builds TAT pivot for 4G / 5G / 4G+5G, saves Excel, returns download URL.
    """
    try:
         # ── Resolve dates ──────────────────────────────────────────────────
        start_date, end_date, err = _resolve_dates(request)
        if err:
            return err

        # ── Check input file exists ────────────────────────────────────────
        input_file = _get_input_file()
        if not input_file:
            return Response(
                {'status': False, 'message': 'No input file found. Please upload a file first.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ── Read, process, export ──────────────────────────────────────────
        today        = date.today()
        df           = _read_input_file(input_file)
        df = _apply_circle_combine(df)
        pivot_data   = build_all_pivots(df, today, start_date, end_date)
        out_filename = build_excel_workbook(pivot_data, start_date, end_date)

        download_url = request.build_absolute_uri(
            f"{settings.MEDIA_URL.rstrip('/')}/performance_tat/output/{out_filename}"
        )
        

        return Response({
            'status':       True,
            'message':      'TAT report generated successfully.',
            'date_range':   f"{start_date} to {end_date}",
            'data': {
            '4G': pivot_data['4G'].to_dict(orient='records'),
            '5G': pivot_data['5G'].to_dict(orient='records'),
            '4G+5G': pivot_data['4G+5G'].to_dict(orient='records')},
            'download_url': download_url,
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
#________________________________________________________________________________________________________________________________
    
@api_view(['POST'])
def generate_aging_softat_report(request):
    """
    POST /api/performance/scft/generate/

    Body:
        { "start_date": "YYYY-MM-DD", "end_date": "YYYY-MM-DD" }

    Filters:
        - Offered Layer     → only 3500, L3500, N3500  (5G)
        - On Air Date       → between start_date and end_date
        - Soft AT Status    → only 'Accepted'
        - TAT               → Performance AT Status Date - Soft AT Status Date

    TAT Buckets:
        <=12 days | 13-21 days | 22-30 days | >30 days | Total | <=12% | 13-21% | 22-30% | >30% | Total %
    """
    # ── Validate dates ─────────────────────────────────────────────────
    try:
        # ── Resolve dates ──────────────────────────────────────────────────
        start_date, end_date, err = _resolve_dates(request)
        if err:
            return err

        # ── Check input file ───────────────────────────────────────────────
        input_file = _get_input_file()
        if not input_file:
            return Response(
                {'status': False, 'message': 'No input file found. Please upload a file first.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ── Read file ──────────────────────────────────────────────────────
        df = _read_input_file(input_file)
        df = _apply_circle_combine(df)
        
        # ── Apply filters ──────────────────────────────────────────────────
        df = df[df["Offered Layer"].isin(LAYERS_5G_INCLUDE)]
        df = df[
            (df["On Air Date"] >= start_date) &
            (df["On Air Date"] <= end_date)
        ]
        df = df[df["Soft AT Status"].str.strip().str.lower() == "accepted"]
        

        if df.empty:
            return Response(
                {'status': False, 'message': 'No data found for the selected filters and date range.'},
                status=status.HTTP_200_OK
            )

        df["Performance AT Status Date"] = pd.to_datetime(
        df["Performance AT Status Date"], format="%d-%b-%Y", errors="coerce")
        df["Soft AT Status Date"] = pd.to_datetime(df["Soft AT Status Date"], format="%d-%b-%Y", errors="coerce")
        df["is_pending"] = df["Performance AT Status Date"].isna()

        df["TAT"] = (df["Performance AT Status Date"] - df["Soft AT Status Date"]).map(lambda x: x.days if pd.notna(x) else 0).astype(int)
       
        
        

        # ── Assign bucket ──────────────────────────────────────────────────
        def assign_bucket(tat):
            if tat <= 12:
                return "<=12 days"
            elif tat <= 21:
                return "13-21 days"
            elif tat <= 30:
                return "22-30 days"
            else:
                return ">30 days"

        df["bucket"] = df["TAT"].apply(assign_bucket)
       
        

        # ── Build combined pivot ───────────────────────────────────────────
        bucket_order = ["<=12 days", "13-21 days", "22-30 days", ">30 days"]
        pct_order    = ["<=12%", "13-21%", "22-30%", ">30%"]
        all_pct_cols = pct_order + ["Pending%"]

        pending = (
            df[df["is_pending"]]
            .groupby("Circle")
            .size()
            .reset_index(name="Pending")
            )

        pivot = (
            df[~df["is_pending"]]
            .groupby(["Circle", "bucket"])
            .size()
            .unstack(fill_value=0)
            .reindex(columns=bucket_order, fill_value=0)
            .reset_index()
            .reset_index(drop=True)
            )

        pivot = pivot.merge(pending, on="Circle", how="outer")
        pivot["Pending"] = pivot["Pending"].fillna(0).astype(int)
        for col in bucket_order:
            pivot[col] = pivot[col].fillna(0).astype(int)

        pivot["TAT_Total"] = pivot[bucket_order].sum(axis=1)
        pivot["Total"] = pivot["TAT_Total"] + pivot["Pending"]
        pivot = pivot.sort_values("Circle").reset_index(drop=True)

        
        # Cumulative % based on Total (includes Pending)
        for i, (bucket, pct_col) in enumerate(zip(bucket_order, pct_order)):
            cumulative_buckets = bucket_order[:i+1]
            pivot[pct_col] = pivot.apply(
                lambda row, cols=cumulative_buckets: round(
                    (sum(row[b] for b in cols) / row["Total"]) * 100, 1
                    ) if row["Total"] > 0 else 0.0,
                    axis=1
                    )

        # Pending% = Pending / Total * 100
        pivot["Pending%"] = pivot.apply(
            lambda row: round((row["Pending"] / row["Total"]) * 100, 1) if row["Total"] > 0 else 0.0,
            axis=1
        )

        grand = {col: pivot[col].sum() for col in bucket_order + ["Pending", "Total"]}
        for i, (bucket, pct_col) in enumerate(zip(bucket_order, pct_order)):
            cumulative_buckets = bucket_order[:i+1]
            grand[pct_col] = round(
                (sum(grand[b] for b in cumulative_buckets) / grand["Total"]) * 100, 1
            ) if grand["Total"] > 0 else 0.0
        grand["Pending%"] = round((grand["Pending"] / grand["Total"]) * 100, 1) if grand["Total"] > 0 else 0.0
        grand["Circle"] = "Grand Total"
        pivot = pd.concat([pivot, pd.DataFrame([grand])], ignore_index=True)

        # ── Build Excel ────────────────────────────────────────────────────
        wb          = Workbook()
        ws          = wb.active
        ws.title    = "5G Aging from Soft AT Report"
        thin_border = _get_thin_border()

        headers    = ["Circle"] + bucket_order + ["Pending"] + ["Total"] + pct_order + ["Pending%"]
        total_cols = len(headers)

        # Row 1 — main heading
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        _style_header_cell(ws.cell(row=1, column=1), "5G Aging from Soft AT Report", 14, HEADER_BG)
        ws.row_dimensions[1].height = 28

        # Row 2 — date range
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        _style_header_cell(ws.cell(row=2, column=1), f"5G  |  {start_date} to {end_date}", 11, SUBHEADER_BG)
        ws.row_dimensions[2].height = 20

        # Row 3 — column headers
        for col_idx, hdr in enumerate(headers, start=1):
            c = ws.cell(row=3, column=col_idx, value=hdr)
            c.font      = Font(name="Arial", bold=True, size=10, color=SUBHEADER_FG)
            c.fill      = PatternFill("solid", fgColor=SUBHEADER_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin_border
            ws.row_dimensions[3].height = 18

        # Data rows
        current_row = 4
        for row_offset, (_, row) in enumerate(pivot.iterrows()):
            is_grand = str(row.get("Circle", "")).strip() == "Grand Total"
            is_alt   = (row_offset % 2 == 1) and not is_grand
            row_bg   = GRAND_TOTAL_BG if is_grand else (ALT_ROW_BG if is_alt else "FFFFFF")
            row_fg   = GRAND_TOTAL_FG if is_grand else "000000"

            for col_idx, col_name in enumerate(headers, start=1):
                raw_val = row.get(col_name, "")

                
                if col_name == "Circle":
                    val = str(raw_val)
                    
                elif col_name in all_pct_cols:
                    try:
                        val = round(float(raw_val), 1) if raw_val != '' and pd.notna(raw_val) else 0.0
                    except (ValueError, TypeError):
                        val = 0.0
                else:
                    try:
                        val = int(float(raw_val)) if raw_val != '' and pd.notna(raw_val) else 0
                    except (ValueError, TypeError):
                        val = 0

                c = ws.cell(row=current_row, column=col_idx, value=val)
                c.font      = Font(name="Arial", bold=is_grand, size=10, color=row_fg)
                c.alignment = Alignment(
                    horizontal="left" if col_name == "Circle" else "center",
                    vertical="center"
                )
                c.border = thin_border
                c.fill   = PatternFill(
                    "solid",
                    fgColor=TOTAL_COL_BG if (col_name in ("Total", "Total %") and not is_grand) else row_bg
                )
                if col_name in all_pct_cols :
                    c.number_format = '0.0'

            ws.row_dimensions[current_row].height = 16
            current_row += 1

        # ── Column widths + freeze ─────────────────────────────────────────
        ws.column_dimensions["A"].width = 22
        for col_idx in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        ws.freeze_panes = "B4"

        # ── Save Excel ─────────────────────────────────────────────────────
        main_folder = os.path.join(settings.MEDIA_ROOT, "performance_tat")
        output_path = os.path.join(main_folder, "output")
        os.makedirs(output_path, exist_ok=True)

        out_filename = f"output_Aging_Soft_AT_{start_date}_to_{end_date}.xlsx"
        wb.save(os.path.join(output_path, out_filename))

        pivot_response = pivot.drop(columns=["TAT_Total", "bucket"] if "TAT_Total" in pivot.columns else [], errors="ignore")
        pivot_response = pivot_response.fillna(0)

        download_url = request.build_absolute_uri(
            f"{settings.MEDIA_URL.rstrip('/')}/performance_tat/output/{out_filename}"
        )

        return Response({
            'status':       True,
            'message':      '5G Aging from Soft AT report generated successfully.',
            'date_range':   f"{start_date} to {end_date}",
            'data':         pivot_response.to_dict(orient='records'),
            'download_url': download_url,
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)