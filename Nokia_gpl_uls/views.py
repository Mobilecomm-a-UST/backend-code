import pandas as pd
import os
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST 
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from rest_framework import status
import xml.etree.ElementTree as ET
import re
import math
import gzip
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side,PatternFill
from openpyxl.utils import get_column_letter



main_folder=os.path.join(MEDIA_ROOT, "Nokia_Slicing")
output_path = os.path.join(main_folder, "Final_Output_uls")
dump_data_path=os.path.join(main_folder,"Dump_data_uls")
os.makedirs(output_path, exist_ok=True)
os.makedirs(dump_data_path, exist_ok=True)

def format_excel(file_path, sheet_name="Slicing"):
    wb = load_workbook(file_path)

    ws = wb[sheet_name] if sheet_name else wb.active
    header_fill = PatternFill(start_color="31869B",end_color="31869B",fill_type="solid")

    header_font = Font(bold=True)
    center_align = Alignment(
        vertical="center",
        horizontal="center",
        wrap_text=True
    )

    
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Header row height
    ws.row_dimensions[1].height =24

    # ---------- Data formatting ----------
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center",horizontal="center", wrap_text=True)
            cell.border = border
            
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 3, 45)
    wb.save(file_path)

#for remark--
def extract_last_mo(val):
    if pd.isna(val):
        return []

    val = str(val).lower().strip()
    parts = re.split(r'[;,]', val)

    cleaned = []

    for p in parts:
        p = p.strip()
        if not p:
            continue

        if "/" in p:
            cleaned.extend(x.strip() for x in p.split("/") if x.strip())
        else:
            cleaned.append(p)

    return sorted(set(cleaned))


def remark(internal, external):

    if pd.isna(internal) and pd.isna(external):
        return "No Changes in value"

    if pd.isna(internal) or pd.isna(external):
        return "Changes in value"

    i = str(internal).strip().lower()
    e = str(external).strip().lower()

    if i == e:
        return "No Changes in value"

    ni_all = re.findall(r'\d+', i)
    ne_all = re.findall(r'\d+', e)

    if ni_all and ne_all:
        if any(n in ne_all for n in ni_all):
            return "No Changes in value"
  

    i_parts = extract_last_mo(internal)
    e_parts = extract_last_mo(external)

    if set(i_parts) & set(e_parts):
        return "No Changes in value"

    return "Changes in value"


# change t/f-> 0,1--
def tf_to_01(val):
    if val is None:
        return val
    val = str(val).strip().lower()
    if val == "true":
        return 1
    if val == "false":
        return 0
    return val

# normalize function-----
def normalize_id(val):
    if pd.isna(val):
        return pd.NA
    parts = [p.strip() for p in str(val).split(",") if p.strip().isdigit()]
    return ",".join(map(str, sorted(map(int, parts)))) if parts else pd.NA


def normalize(val):
    if pd.isna(val):
        return val
    v = str(val).strip()
    try:
        return int(v)
    except:
        try:
            return float(v)
        except:
            return v.lower()
        


@api_view(['POST', 'GET' , 'DELETE'])
def upload_fix_parameter(request):

    folder = os.path.join(main_folder, 'Nokia_Slicing_Fixpara_uls')
    os.makedirs(folder, exist_ok=True)

    if request.method == 'POST':
        file = request.FILES.get('fix_para')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)

        path = os.path.join(folder, file.name)

        with open(path, 'wb+') as f:
            for chunk in file.chunks():
                f.write(chunk)

        return Response({'status': True, 'file': file.name})


    if request.method == 'GET':
        return Response({'files': os.listdir(folder)})


    if request.method == 'DELETE':
        for f in os.listdir(folder):
            os.remove(os.path.join(folder, f))
        return Response({'status': True, 'message': 'All files deleted'})



@api_view(["POST"])
def nokia_slicing_dump(request):
    excel_file=os.path.join(main_folder, 'Nokia_Slicing_Fixpara_uls')
    if not os.path.exists(excel_file):
        return Response({"error": "Fix_Parameter folder not found"}, status=400)

    xml_files = request.FILES.getlist("xml_file")
    if not xml_files:
        return Response({"error": "Please upload XML files"}, status=HTTP_400_BAD_REQUEST)
    
    for file in xml_files:

        file_name = file.name.lower()
        if file_name.endswith(".gz"):
            xml_bytes = gzip.decompress(file.read())
        else:
            xml_bytes = file.read()
            
        root = ET.fromstring(xml_bytes)


        m = re.match(r'\{(.*)\}', root.tag)
        ns_url = m.group(1) if m else root.attrib.get("xmlns", "")
        ns = {"ns": ns_url} if ns_url else {}

        print(f"File: {file.name}, Namespace: {ns_url}")

        # ---- Try both namespace & no namespace ----
    dumy_data=[]
    managed_objects = root.findall(".//ns:managedObject", ns) if ns else []
    if not managed_objects:
        managed_objects = root.findall(".//managedObject")

    for mo in managed_objects:

        mo_class = mo.attrib.get("class", "")
        dist_name = mo.attrib.get("distName", "")

        # for NRBTS class--------

        if mo_class == "com.nokia.srbts.nrbts:NRBTS":

            required_in_nrbts = { 
                "actSliceAwareScheduler", 
                "actHighSliceWeightFactor", 
                "actSliceAwareSchedulerUlAndEnh", 
                "actSliceSwitchToDefault", 
                "actAdditionalSliceSupport", 
                "actSliceNumExt",
                "actExtSchedWeightAndPrefWrrAlg", 
                "actDDDSPeriodOptimization",
                "actBeamControlImprovement1",
                "actBeamControlImprovement2", 
                "actCellTraceReport", 
                "n310","n311","t310","t311",
                "actConflictConfiguration", 
                "actCoordinated4g5gPowerSaving", 
                "actEnhanceIntraNrIntraFreqAnr", 
                "actEnhancedLinkAdaptation", 
                "actEnhancedX2LinkSupervision", 
                "actFdmEnhancedScheduling", 
                "actFdmScheduling", 
                "actInactDetNSAUe", 
                "actIpThpt", 
                "actMacUserThpt", 
                "actMobilityRetryToSecondBestCell", 
                "actMultiDrbNSA", 
                "actNgcFlexSaMode", 
                "actNonGbrServiceDiff", 
                "actNrLBPowerSaving", 
                "actNsaUeBasedAnrNr",
                "actOverheatingAssistance", 
                "actPaging", 
                "actPdcpRlcBufCongestionMechsm", 
                "actPerCellMacUserThpt", 
                "actRtwpMeasurement", 
                "actSACallProcessingDU", 
                "actSaHoDeltaMvInterOp", 
                "actSaIntraFreqAnrDrxProf", 
                "actSaUeBasedAnrNr", 
                "actSecDataUsageRep", 
                "actThpDist", 
                "actTimeAlignExtension", 
                "actUeInitRlf", 
                "actX2ConfigTransfer", 
                "actXnSecCapNewMode", 
                "actDataDuplicationForMobility", 
                "activityNotificationLevelSA", 
                "actnrpmqapprofiles",
                "drxdefaultpaging",
                "gNbCuType", 
                "maxNumOfNgSetupRequestRetries", 
                "maxNumX2Links", 
                "maxNumXnLinks", 
                "maxnumnranrmoiallowed", 
                "nsaActivityNotificationLevel", 
                "nsaInactivityTimer", 
                "pduSessionEndMarkerTimer", 
                "periodicalDataUsageReportTimer", 
                "rrcReestabTypeSA", 
                "sgnbRelForNoHandoverEnabled", 
                "srb3SupportEnabled", 
                "tDCoverall", 
                "tRLFindForDUSA", 
                "trlfindfordu",
                "tWaitingRlRecover", 
                "tWaitingRlRecoverSA", 
                "thpHistDownlinkMaxRange", 
                "thpHistDownlinkMinRange", 
                "thpHistScale", 
                "thpHistUplinkMaxRange", 
                "thpHistUplinkMinRange", 
                "thpTimeCalc", 
                "timerHOGuard", 
                "timerOverheatIndProhibit", 
                "timerPostOverheating", 
                "timerRRCGuard", 
                "timerX2UeProcGuard", 
                "x2EndMarkerTimer", 
                "x2linkSupervisionTmr",
                "xnLinkReestabTmr" ,
                "xnLinkReestabTmr",
                "acta3a5mrduringdrxsetup",
                "actbbpoolextensiontddfr1",
                "actbbpooltddfr1",
                "actbbpooltddfr1algenh",
                "actbbpooltddfr1enh",
                "actdrxretranstimersupport",
                "actenhrrcreconfulprbrandom",
                "actexcludingeutranbandcheck",
                "actintergnbrrepreparedcell",
                "actinterrathodirectdldatafwd",
                "actlargebbpooltddfr1",
                "actlocationreportingsa",
                "actnormvcfallbackenh",
                "actoverallocavoidforcellstrprb",
                "actpdcprlcbufcongestionmechsm",
                "actrohc",
                "actseparatepdurelforgtpuerror",
                "actseparatepdurelforrlcerror",
                "actserviceawarepdcpconfig",
                "actspatialstreampooling",
                "acttddcsiimrm",
                "acttddcsirsopt",
                "acttddprepooling",
                "acttransphaseulprbnonrandom",
                "actvoicefallbackduringhoprep",
                "actvonrpmcount",
                "actvonrprioduringhoprep",
                "allowslowprbpoolingstatndmsrs",
                "drbinactivitytimersa",
                "nrecidmeassupervisiontimer",
                "timerngsignalingconnguard",
                "timerrejectwait",
                "timerxndatafwdguard",
                "tmpactfeat1",
                "tngrelocoverall",
                "tngrelocoveralln",
                "trlfindforduvonr",
                "uespeccsirsslotsratio",
                "uplanehointerruptmodrange",
                "uplanehointerrupttgtrange",
                "vonrfwddelaymodrange",
                "vonrfwddelaytgtrange",
                "vonrmutedurthresh",
                "vonrsilentdurthresh",
                "vonrthreshfactor",
                "actdrxretranstimersupport",
                "actdynamicfirewall",
                "actpdcprlcbufcongestionmechsm",
                "actvoicefallbackduringhoprep",
                "actvonrprioduringhoprep",
                "trlfindforduvonr",
                "uespeccsirsslotsratio",
                "actforcehoinsuffuecap",
                "actpersistentsyssetmoi",
                "actPdcpRlcBufCongOvlMech",

                


                }
            required_in_nrbts_lower = {
                    x.lower() for x in required_in_nrbts
                }

            # -------- simple <p> params --------
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")
                if name.strip().lower() in required_in_nrbts_lower:
                    dumy_data.append({
                        "MO": "NRBTS",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

            # -------- list param: --------
            for item in (
                mo.findall(".//ns:list[@name='cbtsFlowControlProf']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='cbtsFlowControlProf']/item")
            ):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    if p.attrib.get("name") == "dddsPeriod":
                        dumy_data.append({
                            "MO": "NRBTS",
                            "DistName": dist_name,
                            "Parameter": "NRBTS.cbtsFlowControlProf.dddsPeriod",
                            "value": tf_to_01(p.text)
                        })


            # -------- powerSavingBwpSwitching --------
            for item in (
                mo.findall(".//ns:list[@name='powerSavingBwpSwitching']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='powerSavingBwpSwitching']/item")
            ):

                required_params = [
                    "dlBufferEmptyRatioFromPsBwp",
                    "dlBufferEmptyRatioToPsBwp",
                    "dlRateThresholdFromPsBwp",
                    "dlRateThresholdToPsBwp",
                    "measurementDurationFromPsBwp",
                    "measurementDurationToPsBwp",
                    "ulBufferEmptyRatioFromPsBwp",
                    "ulBufferEmptyRatioToPsBwp"
                ]

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in required_params:
                        dumy_data.append({
                            "MO": "NRBTS",
                            "DistName": dist_name,
                            "Parameter": f"powerSavingBwpSwitching@{param_name}",
                            "value": tf_to_01(p.text)
                        })


            # -------- xxFlowControlProfForCu --------
            for item in (
                mo.findall(".//ns:list[@name='xxFlowControlProfForCu']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='xxFlowControlProfForCu']/item")
            ):

                required_params = [
                    "dddsPeriodXx",
                    "maxTransferDelayXx",
                    "minThFlowCtrlXx"
                ]

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in required_params:
                        dumy_data.append({
                            "MO": "NRBTS",
                            "DistName": dist_name,
                            "Parameter": f"xxFlowControlProfForCu@{param_name}",
                            "value": tf_to_01(p.text)
                        })


            # -------- single parameters --------
            single_params = [
                "thpHistDownlinkMaxRange",
                "thpHistDownlinkMinRange",
                "thpHistScale",
                "thpHistUplinkMaxRange",
                "thpHistUplinkMinRange",
                "x2linkSupervisionTmr"
            ]

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                param_name = p.attrib.get("name")

                if param_name in single_params:
                    dumy_data.append({
                        "MO": "NRBTS",
                        "DistName": dist_name,
                        "Parameter": param_name,
                        "value": tf_to_01(p.text)
                    })

            # print("NRBTS FOUND ------------------")
            
            # -------- actVoNRPrioDuringHOPrep --------
            for lst in (
                mo.findall(".//ns:list[@name='actVoNRPrioDuringHOPrep']", ns)
                if ns else
                mo.findall(".//list[@name='actVoNRPrioDuringHOPrep']")
            ):
                for p in lst.findall("ns:p", ns) if ns else lst.findall("p"):
                    dumy_data.append({
                        "MO": "NRBTS",
                        "DistName": dist_name,
                        "Parameter": "actVoNRPrioDuringHOPrep",
                        "value": tf_to_01(p.text)
                    })


            # -------- actVoiceFallbackDuringHOPrep --------
            for lst in (
                mo.findall(".//ns:list[@name='actVoiceFallbackDuringHOPrep']", ns)
                if ns else
                mo.findall(".//list[@name='actVoiceFallbackDuringHOPrep']")
            ):
                for p in lst.findall("ns:p", ns) if ns else lst.findall("p"):
                    dumy_data.append({
                        "MO": "NRBTS",
                        "DistName": dist_name,
                        "Parameter": "actVoiceFallbackDuringHOPrep",
                        "value": tf_to_01(p.text)
                    })
        



        elif mo_class == "com.nokia.srbts.nrbts:NRCELLGRP":
            required_in_nrcellgrp = {
                "numberOfTransmittedSsBlocks",
                "maxNumOfUsers",
                "maxNumOfNonGBRBearers",
                "addNumOfHoUsers",
                "addNumOfNonGBRBearersHo",

                # New Parameters
                "addNumOfNonGBRBearersCGRPSAHo",
                "maxHarqMsg3Tx",
                "maxNumOfNonGBRBearersCGRPSA",
                "maxNumOfRrcCGRPSA",
                "ssBurstSetPeriod",
            }

            required_in_nrcellgrp_lower = {
                x.lower() for x in required_in_nrcellgrp
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.strip().lower() in required_in_nrcellgrp_lower:
                    dumy_data.append({
                        "MO": "NRCELLGRP",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })

            # print("NRCELLGR FOUND ------------------")
        
        elif mo_class == "com.nokia.srbts.nrbts:NRANR":
            required_in_nranrgrp = {
                "actautonrnbremoval",
                "consecHoFailThres"

            }

            required_in_nranrgrp_lower = {
                x.lower() for x in required_in_nranrgrp
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.strip().lower() in required_in_nranrgrp_lower:
                    dumy_data.append({
                        "MO": "NRANR",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

            # print("NRANR FOUND ------------------")

        elif mo_class == "com.nokia.srbts.nrbts:NRANRPR":
            required_in_nranrPRgrp = {
                "consecRecheckForNrNbRemoval",
                "anrThresRsrpNbCell",
                "cellDepMode",
                "nrarfcn",
                "maxNumAnrNrrelAllowed",
                "idletimefornrnbremoval",

            }

            required_in_nranrPRgrp_lower = {
                x.lower() for x in required_in_nranrPRgrp
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.strip().lower() in required_in_nranrPRgrp_lower:
                    dumy_data.append({
                        "MO": "NRANRPR",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

            # print("NRANRPR FOUND ------------------")  

        elif mo_class == "com.nokia.srbts.nrbts:NRMTRACEDU":
            required_in_NRMTRACEDU = {
               'actRIReporting',
                'riReportInterval',

            }

            required_in_NRMTRACEDU_lower = {
                x.lower() for x in required_in_NRMTRACEDU
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.strip().lower() in required_in_NRMTRACEDU_lower:
                    dumy_data.append({
                        "MO": "NRMTRACEDU",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    }) 



        elif mo_class == "com.nokia.srbts.nrbts:NRDU":
            required_in_NRDU = {
              'maxNumOfRrcConUEsPerDU',

            }

            required_in_NRDU_lower = {
                x.lower() for x in required_in_NRDU
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.strip().lower() in required_in_NRDU_lower:
                    dumy_data.append({
                        "MO": "NRDU",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

        # for NRCELL Class-----------
        elif mo_class == "com.nokia.srbts.nrbts:NRCELL":
            required_in_nrcell = {
                "actUlTxSkip",
                "srPeriodicityMin",
                "actdrxretranstimersupport",
                "adaptiveSrLoadThresholdUp",
                "adaptiveSrLoadThresholdDown",
                "adaptiveSrBlockingMinimization",
                "maxNbrTrafficLimit",
                "preferredVoNrSrPeriod",
                "actNbrForNonGbrBearers",
                "congDetectPeriod",
                "congWeightAlg",
                "maxNumOfNbrBearers",
                "maxPrbsPerNbrUe",
                "nbrPdcchCongHandlingDl",
                "nbrPdcchCongHandlingUl",
                "nbrPdschCongHandling",
                "nbrPuschCongHandling",
                "nrResourceGroupProfileDN",
                "actadaptiveretxresmcs",
                "actadaptvoiptbs",
                "actapercsi",
                "actbeamforming",
                "actccesplit",
                "actcdrx",
                "actcdrxvonr",
                "actcsirsmultiplexing",
                "actdataforwardingforvonr",
                "actdl256qam",
                "actdldatadmrsfdm",
                "actdlmumimo",
                "actdlmumimoenh",
                "actdlptrs",
                "actdlrbgrandomization",
                "actdlsrsbm",
                "actdynpdcchtdmode",
                "actdynulresalloc",
                "actemergencycall",
                "actenhancedvonr",
                "actfullpwrmode",
                "actgbradmissioncontrol",
                "actgbrractdenhresourcepriority",
                "actincreasedfdmusersperslot",
                "actlongpucchphase2",
                "actmcsprbreductionfdmul",
                "actnbrfornongbrbearers",
                "actnrrim",
                "actpdschatssbslots",
                "actpdschrmcsirsfortracking",
                "actpercsireportbundlingtddfr1",
                "actpowersavingbwp",
                "actprachmultiplexing",
                "actproactulscheduling",
                "actsib1fornsacell",
                "acttcpboostpug",
                "actuecsirscoexistcellcsirs",
                "actul256qam",
                "actulclosedlooppwrctrl",
                "actuldatadmrsfdm",
                "actuldftsofdm",
                "actulptrs",
                "actulsrsbm",
                "actultxskip",
                "actvoippacketaggr",
                "actvonr",
                "actvonrmobilitythresholds",
                "adaptivesrblockingminimization",
                "adaptivesrloadthresholddown",
                "adaptivesrloadthresholdup",
                "adjustprachthresholdoffset",
                "cbrapreamblesperssb",
                "cellbarred",
                "celldeptype",
                "cellreservedforoperatoruse",
                "cellreservedforotheruse",
                "congdetectperiod",
                "congweightalg",
                "dedicatedsib1",
                "dldmrsadditionalposition",
                "dllablertarget",
                "dlladeltacqimax",
                "dlladeltacqimin",
                "dlladeltacqistepdown",
                "dllainimcs",
                "dlmimomode",
                "dlmulowsinrthreshold",
                "dlmumaxnumpairedues",
                "dlmuneighborthreshold",
                "dlmupairingthreshold",
                "dmrstypeaposition",
                "drxwactdlenabled",
                "expectedcellsize",
                "freqbandindicatornr",
                "initialpreamblereceivedtargetpower",
                "maxnbrtrafficlimit",
                "maxnumberuespowersaving",
                "maxnumofgbrbearerssa",
                "maxnumofnbrbearers",
                "maxnumofrrcsa",
                "maxnumofuserspercell",
                "maxnumofuserspernrcell",
                "maxnumpdschallocationperslot",
                "maxnumpuschallocationperslot",
                "maxprbspernbrue",
                "minpctresourcereservedmcsdl",
                "minpctresourcereservedmcsul",
                "mmimoantarraymode",
                "msg1frequencystart",
                "msg3deltapreamble",
                "n310",
                "n310vonr",
                "n311",
                "nbrpdcchconghandlingdl",
                "nbrpdcchconghandlingul",
                "nbrpdschconghandling",
                "nbrpuschconghandling",
                "nrcelltype",
                "nrresourcegroupprofiledn",
                "numberofrxbeamforming",
                "numofpagingframes",
                "numpagingoccsnpagingframe",
                "pagingoffset",
                "pmax",
                "pmaxnrowncell",
                "powerrampingstep",
                "prachconfigurationindex",
                "preambletransmax",
                "preferredvonrsrperiod",
                "pucchf3maxcoderate",
                "pucchf3modulationscheme",
                "pucchmodeselect",
                "puschmappingtype",
                "racontentionresolutiontmr",
                "raresponsewindow",
                "reqminnumpuschprb",
                "rlpdetcsibsithreshold",
                "rlpdetdlharqthreshold",
                "rlpreccsibsithreshold",
                "rlprecdlharqthreshold",
                "rsrpthresholdssb",
                "srperiodicity",
                "srperiodicitymin",
                "srstypeallocfordlbm",
                "sspbchblockpower",
                "t300",
                "t301",
                "t310",
                "t310vonr",
                "t311",
                "targetpctresourcereservedmcsdl",
                "targetpctresourcereservedmcsul",
                "totalnumberofrapreambles",
                "type0coresetconfigurationindex",
                "type0searchspaceconfigurationindex",
                "ueconnectionestablishmentmode",
                "uldmrsadditionalposition",
                "ullablertarget",
                "ulladeltasinrmax",
                "ulladeltasinrmin",
                "ulladeltasinrstepdown",
                "ullainimcs",
                "ullimitedmcsstepupsizebler",
                "ulmcsdowngrademaxreduction",
                "ulptrssampledensitythresnrb0",
                "ulptrssampledensitythresnrb1",
                "ulptrssampledensitythresnrb2",
                "ulptrssampledensitythresnrb3",
                "ulptrssampledensitythresnrb4",
                "ulschedtimeinterval",
                "ultransprecodeoffsinrthresh",
                "ultransprecodeonsinrthresh",
                "ultransprecodepi2bpsk",
                "ultransprepi2bpskpwrboost",
                "zerocorrelationzoneconfig",
                "actadaptivesr",
                "actaugsrbprioboost",
                "actcsirsmuting",
                "actdldynswitchalloc",
                "actdlforcerankdowng",
                "actdlpowerboost",
                "actdlrankadpandsrb1mcslimdurho",
                "actdlvonrlimmaxmcsandlayer",
                "actearlydrxsetup",
                "actforceddftsofdmvonrul",
                "actmaxinitdeltacqiclampvonr",
                "actnonvonrgbrserviceenh",
                "actpdcchpowerboost",
                "actrat1alloc",
                "actrrcreconfprocessingdelay",
                "actserviceawareconfig",
                "actshortbsrlimitadaptation",
                "actsmallestrrcsetup",
                "actsubtrsratematch",
                "acttargetblervonr",
                "acttrspdschavgcqiratio",
                "actulaudiogapreduction",
                "actullaollaforretxindft",
                "actulmintbsforvonr",
                "actulnrecidposmethod",
                "actulnrecidtadv",
                "actultranspreenhho",
                "actultranspreenhrankswitchiwk",
                "actultranspreenhsinrmetric",
                "actvonrinccodinggain",
                "addnumofgbrbearers",
                "adjustpowercontroloffset",
                "allowcsirsoutsidedrx",
                "dllatargetblervonr",
                "dlmcsdroplimitforvonr",
                "enhrankselectircgain",
                "enhrankselectircgainmethod",
                "enhrankselectse",
                "fpdcchmonoccsnpo",
                "maxgbrtrafficlimit",
                "maxinitdeltacqiclampvonr",
                "maxmcsfordlsrb",
                "maxmcsfordlvonrctrl",
                "maxmcsvonrrankandmcsrestrict",
                "maxnumbervonrpowersaving",
                "maxnumofnongbrbearerssa",
                "maxprbsincludingpowerboostprbs",
                "nbrofssbperrachoccasion",
                "numtxthreshforpowerboost",
                "pdcchblertarget",
                "percredsedlrat1vonr",
                "powerboostscalingfactor",
                "ullatargetblervonr",
                "ultransprecodeofftotpwrthresh",
                "ultransprecodeontotpwrthresh",
                "vonrreservation",
                "nrsysinfoprofiledn",
                "arpPrioEmergencyCall",
                
               
               
              
               
               
               
               



            }

            required_in_nrcell_lower = {x.lower() for x in required_in_nrcell}
            params_nrcell = {}

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.lower() in required_in_nrcell_lower:
                    params_nrcell[name.lower()] = p.text

                    dumy_data.append({
                        "MO": "NRCELL",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })
            # beamSet
            for item in mo.findall(".//ns:list[@name='beamSet']/ns:item", ns) if ns else mo.findall(".//list[@name='beamSet']/item"):

                param_map = {
                    "basicBeamSet": "beamset@basicbeamset",
                    "leftEdgeAngle": "beamset@leftedgeangle",
                    "lowerEdgeAngle": "beamset@loweredgeangle",
                    "nrBtsBeamRefinementP2": "beamset@nrbtsbeamrefinementp2",
                    "rightEdgeAngle": "beamset@rightedgeangle",
                    "upperEdgeAngle": "beamset@upperedgeangle"
                }

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })


            # cellBwpList
            for item in mo.findall(".//ns:list[@name='cellBwpList']/ns:item", ns) if ns else mo.findall(".//list[@name='cellBwpList']/item"):

                param_map = {
                    "pdcchBlockingThresholdDL": "cellbwplist@pdcchblockingthresholddl",
                    "pdcchBlockingThresholdUL": "cellbwplist@pdcchblockingthresholdul",
                    "pdcchFreeThresholdDL": "cellbwplist@pdcchfreethresholddl",
                    "pdcchFreeThresholdUL": "cellbwplist@pdcchfreethresholdul",
                    "actUpperThrAckNackFiltering": "cellbwplist@actupperthracknackfiltering"
                }
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })


            # csirsBeamMgmt
            for item in mo.findall(".//ns:list[@name='csirsBeamMgmt']/ns:item", ns) if ns else mo.findall(".//list[@name='csirsBeamMgmt']/item"):

                param_map = {
                    "csirsBmMgmtDensity": "csirsbeammgmt@csirsbmmgmtdensity",
                    "csirsBmMgmtReIndex": "csirsbeammgmt@csirsbmmgmtreindex",
                    "csirsBmMgmtSubband": "csirsbeammgmt@csirsbmmgmtsubband"
                }

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })


            # csirsForTracking
            for item in mo.findall(".//ns:list[@name='csirsForTracking']/ns:item", ns) if ns else mo.findall(".//list[@name='csirsForTracking']/item"):

                param_map = {
                    "csirsTrackingPeriod": "csirsfortracking@csirstrackingperiod",
                    "firstRE": "csirsfortracking@firstre",
                    "rbAllocation": "csirsfortracking@rballocation",
                    "startRB": "csirsfortracking@startrb"
                }

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })


            # periodicCsiAcquisition
            for item in mo.findall(".//ns:list[@name='periodicCsiAcquisition']/ns:item", ns) if ns else mo.findall(".//list[@name='periodicCsiAcquisition']/item"):

                param_map = {
                    "csiReportPeriodicity": "periodiccsiacquisition@csireportperiodicity",
                    "csirsNumberWithinReportPeriod_Set1": "periodiccsiacquisition@csirsnumberwithinreportperiod_set1",
                    "csirsPower": "periodiccsiacquisition@csirspower",
                    "maxCsirsNumber": "periodiccsiacquisition@maxcsirsnumber"
                }

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })


            # srsForBeamMgmt
            for item in mo.findall(".//ns:list[@name='srsForBeamMgmt']/ns:item", ns) if ns else mo.findall(".//list[@name='srsForBeamMgmt']/item"):

                param_map = {
                    "srsUePeriodicity": "srsforbeammgmt.srsueperiodicity",
                    "bSRS": "srsforbeammgmt@bsrs",
                    "cSRS": "srsforbeammgmt@csrs",
                    "srsPowerThresholdForDlBm": "srsforbeammgmt@srspowerthresholdfordlbm",
                    "srsTypeAllocForDlBm": "srstypeallocfordlbm"
                }

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })


            # tddFrameStructure
            for item in mo.findall(".//ns:list[@name='tddFrameStructure']/ns:item", ns) if ns else mo.findall(".//list[@name='tddFrameStructure']/item"):

                param_map = {
                    "frameStructureType": "tddframestructure@framestructuretype",
                    "guardPeriodLength": "tddframestructure@guardperiodlength",
                    "lteToNrFrameShift": "tddframestructure@ltetonrframeshift",
                    "tdLteUlDlConfig": "tddframestructure@tdlteuldlconfig",
                    "ulDlDataSlotRatio": "tddframestructure@uldldataslotratio"
                }

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })


            # ulPowerControlCommon
            for item in mo.findall(".//ns:list[@name='ulPowerControlCommon']/ns:item", ns) if ns else mo.findall(".//list[@name='ulPowerControlCommon']/item"):

                param_map = {
                    "actUlOpenLoopPwrCtrl": "ulpowercontrolcommon@actulopenlooppwrctrl",
                    "alpha": "ulpowercontrolcommon@alpha",
                    "alphaSrs": "ulpowercontrolcommon@alphasrs",
                    "p0NominalPucch": "ulpowercontrolcommon@p0nominalpucch",
                    "p0NominalPucchF1F3": "ulpowercontrolcommon@p0nominalpucchf1f3",
                    "p0NominalPusch": "ulpowercontrolcommon@p0nominalpusch",
                    "p0NominalSrs": "ulpowercontrolcommon@p0nominalsrs",
                    "pucchF1DeltaF": "ulpowercontrolcommon@pucchf1deltaf",
                    "pucchF3DeltaF": "ulpowercontrolcommon@pucchf3deltaf"
                }

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })


            # ulPtrsForTransPrecoding
            for item in mo.findall(".//ns:list[@name='ulPtrsForTransPrecoding']/ns:item", ns) if ns else mo.findall(".//list[@name='ulPtrsForTransPrecoding']/item"):

                param_map = {
                    "ulPtrsSampleDensityThresNrb0": "ulptrssampledensitythresnrb0",
                    "ulPtrsSampleDensityThresNrb1": "ulptrssampledensitythresnrb1",
                    "ulPtrsSampleDensityThresNrb2": "ulptrssampledensitythresnrb2",
                    "ulPtrsSampleDensityThresNrb3": "ulptrssampledensitythresnrb3",
                    "ulPtrsSampleDensityThresNrb4": "ulptrssampledensitythresnrb4"
                }

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })
            # ulSuMimoEnh
            for item in mo.findall(".//ns:list[@name='ulSuMimoEnh']/ns:item", ns) if ns else mo.findall(".//list[@name='ulSuMimoEnh']/item"):

                param_map = {
                    "enhRankSelectIrcGain": "enhrankselectircgain",
                    "enhRankSelectIrcGainMethod": "enhrankselectircgainmethod",
                    "enhRankSelectSe": "enhrankselectse"
                }

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })            


            # ulTransPrecodeCtl
            for item in mo.findall(".//ns:list[@name='ulTransPrecodeCtl']/ns:item", ns) if ns else mo.findall(".//list[@name='ulTransPrecodeCtl']/item"):

                param_map = {
                    "ulTransPrecodeOffSinrThresh": "ultransprecodeoffsinrthresh",
                    "ulTransPrecodeOnSinrThresh": "ultransprecodeonsinrthresh",
                    "ulTransPrecodePi2Bpsk": "ultransprecodepi2bpsk",
                    "ulTransPrePi2BpskPwrBoost": "ultransprepi2bpskpwrboost",
                    "ulTransPrecodeOffTotPwrThresh": "ultransprecodeofftotpwrthresh",
                    "ulTransPrecodeOnTotPwrThresh": "ultransprecodeontotpwrthresh"
                }

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })


            # dlDataDmrsFdm
            for item in mo.findall(".//ns:list[@name='dlDataDmrsFdm']/ns:item", ns) if ns else mo.findall(".//list[@name='dlDataDmrsFdm']/item"):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    if p.attrib.get("name") == "actDlDataDmrsFdm":
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": "actdldatadmrsfdm",
                            "value": tf_to_01(p.text)
                        })


            # dlPtrs
            for item in mo.findall(".//ns:list[@name='dlPtrs']/ns:item", ns) if ns else mo.findall(".//list[@name='dlPtrs']/item"):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    if p.attrib.get("name") == "actDlPtrs":
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": "actdlptrs",
                            "value": tf_to_01(p.text)
                        })

            # dlForceRankDowngCtl
            for item in mo.findall(".//ns:list[@name='dlForceRankDowngCtl']/ns:item", ns) if ns else mo.findall(".//list[@name='dlForceRankDowngCtl']/item"):

                param_map = {
                    "maxMcsForDlSrb": "maxmcsfordlsrb",
                    "maxMcsForDlVoNRCtrl": "maxmcsfordlvonrctrl"
                }

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })

            # dlRetxPowerBoostCtl
            for item in mo.findall(".//ns:list[@name='dlRetxPowerBoostCtl']/ns:item", ns) if ns else mo.findall(".//list[@name='dlRetxPowerBoostCtl']/item"):

                param_map = {
                    "maxPrbsIncludingPowerBoostPrbs": "maxprbsincludingpowerboostprbs",
                    "numTxThreshForPowerBoost": "numtxthreshforpowerboost",
                    "powerBoostScalingFactor": "powerboostscalingfactor"
                }

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param_name = p.attrib.get("name")
                    if param_name in param_map:
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(p.text)
                        })            

            # ulDataDmrsFdm
            for item in mo.findall(".//ns:list[@name='ulDataDmrsFdm']/ns:item", ns) if ns else mo.findall(".//list[@name='ulDataDmrsFdm']/item"):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    if p.attrib.get("name") == "actUlDataDmrsFdm":
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": "actuldatadmrsfdm",
                            "value": tf_to_01(p.text)
                        })


            # ulPtrs
            for item in mo.findall(".//ns:list[@name='ulPtrs']/ns:item", ns) if ns else mo.findall(".//list[@name='ulPtrs']/item"):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    if p.attrib.get("name") == "actUlPtrs":
                        dumy_data.append({
                            "MO": "NRCELL",
                            "DistName": dist_name,
                            "Parameter": "actulptrs",
                            "value": tf_to_01(p.text)
                        })         

            # print("NRCELL FOUND ----------------")
            # print("Class:", mo_class)
            # print("DistName:", dist_name)
            # print("Params_NRCELL:", params_nrcell)

        elif mo_class == "com.nokia.srbts.nrbts:NRULCLPC_PROFILE":
            dist_name = mo.attrib.get("distName", "")
            print("\n========== NRULCLPC_PROFILE FOUND ==========")
            print("DistName :", dist_name)

            items = (
                mo.findall(".//ns:list[@name='clpcPusch']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='clpcPusch']/item")
            )

            print("Total clpcPusch Items :", len(items))

            param_map = {
                "maxTargSinr": "maxtargsinr",
                "minTargSinr": "mintargsinr",
                "targetSinrOffset": "targetsinroffset",
            }

            for i, item in enumerate(items):
                print(f"\nItem {i+1}")

                params = item.findall("ns:p", ns) if ns else item.findall("p")
                print("Total Parameters :", len(params))

                for p in params:
                    param_name = p.attrib.get("name")
                    value = p.text

                    print("XML Param :", param_name, "=", value)

                    if param_name in param_map:
                        print("Matched ->", param_name)

                        dumy_data.append({
                            "MO": "NRULCLPC_PROFILE",
                            "DistName": dist_name,
                            "Parameter": param_map[param_name],
                            "value": tf_to_01(value)
                        })

            print("====================================")
        
    # yaha upar wala clpcPusch code

        elif mo_class == "com.nokia.srbts.nrbts:NRPGRP":
            dist_name = mo.attrib.get("distName", "")

            # ---------- Simple parameters ----------
            required_params = {
                "lbpsMaxLoad",
                "lbpsMinLoad",
                "lbpsLastCellMinLoad",
                "lbpsLastCellSOEnabled",
                "lbpsPdcchLoadOffset",
                "lbpsCSONCtlEnabled"
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name in required_params:
                    dumy_data.append({
                        "MO": "NRPGRP",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

            # ---------- lbpsCellList ----------
            for item in (
                mo.findall(".//ns:list[@name='lbpsCellList']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='lbpsCellList']/item")
            ):

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):

                    if p.attrib.get("name") == "lbpsCellSOOrder":
                        dumy_data.append({
                            "MO": "NRPGRP",
                            "DistName": dist_name,
                            "Parameter": "lbpsCellSOOrder",
                            "value": tf_to_01(p.text)
                        })

            # ---------- lbpsPeriodList ----------
            required_period_params = {
                "lbpsDayOfWeek",
                "lbpsStartTimeHour",
                "lbpsStartTimeMinute",
                "lbpsDuration",
                "lbpsSuspended"
            }

            period_values = {param: [] for param in required_period_params}

            for item in (
                mo.findall(".//ns:list[@name='lbpsPeriodList']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='lbpsPeriodList']/item")
            ):

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):

                    name = p.attrib.get("name")

                    if name in required_period_params and p.text is not None:
                        period_values[name].append(tf_to_01(p.text))

            for param, values in period_values.items():

                if values:
                    dumy_data.append({
                        "MO": "NRPGRP",
                        "DistName": dist_name,
                        "Parameter": f"lbpsPeriodList@{param}",
                        "value": ",".join(map(str, values))
                    })


        elif mo_class == "com.nokia.srbts.nrbts:NRDLMUMIMO":
            required_in_NRDLMUMIMO = {
                "dlMuMaxNumLayerPerMuRbg",                                                                                                          
                "dlMuMimoCorrThd",
                "dlMuMimoCqiThd",
                "dlMuMimoRankAdaption",
                "dlMuMimoSpectralEffThd",
                "dlMuMimoTaperingH",
                "dlMuMimoTaperingV",
                "nrDlMuMimoId",

            }

            required_in_NRDLMUMIMO_lower = {
                x.lower() for x in required_in_NRDLMUMIMO
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.strip().lower() in required_in_NRDLMUMIMO_lower:
                    dumy_data.append({
                        "MO": "NRDLMUMIMO",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

            # print("NRDLMUMIMO found")

         # For NRDRB class------------------
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_id = int(dist_name.split("NRDRB-")[-1])
            except:
                continue

            group_1_ids = {5, 6, 11, 12, 21, 22, 25, 26, 47}
            group_2_ids = {21, 22, 25, 26, 47}
            group_3_ids = {6}
            group_4_ids = {5, 6, 7, 8, 9, 11, 12, 21, 22, 25, 47}
            group_5_ids = {26}
            group_6_ids = {11, 21}
            group_7_ids = {12, 22}
            group_8_ids = {5, 25}
            group_9_ids = {6, 26}
            group_10_ids = {7, 47}
            group_11_ids = {8, 9}
            group_12_ids = {47}
            group_13_ids = {11}
            group_14_ids = {12}
            group_1_params = {
                "pdcpStatRepWaitTimerOffset",
                "pdcpStatRepWaitTimer"
            }

            group_2_params = {
                "actDddsReduction",
                "actDlTxResumeOnPdcpStatRep",
                "actLostPduFastRetx",
                "actOptDataFlowRateEst",
                "reTxPrioritizationType"
            }

            group_3_params = {
                "inactPeriodX2PdcpDuplication"
            }

            group_4_params = {
                "schedulWeight"
            }

            group_5_params = {
                "schedulWeight",
                "nrDrbMacDN"
            }

            group_6_params = {"priorityLevel"}
            group_7_params = {"priorityLevel"}
            group_8_params = {"priorityLevel"}
            group_9_params = {"priorityLevel"}
            group_10_params = {"priorityLevel"}
            group_11_params = {"priorityLevel"}
            group_12_params = {"nrDrbMacDN"}
            group_13_params = {
                                "priorityLevel",
                                "discardTimer",
                                "discardTimerUl",
                                "maxDataRateGbrDl",
                                "maxDataRateGbrUl",
                                "queuingDelaySduDiscardThreshold",
                            }
            group_14_params = {
                        "nrDrbMacDN",
                        "maxDataRateGbrDl",
                        "maxDataRateGbrUl",
                    }

            # -------- All NRDRB parameters --------
            common_nrdrb_params = {
                "queuingDelayDiscarding",
                "volThresDiscardDl",
                "volThresDiscardUl",
                "durationThresDiscardDl",
                "durationThresDiscardUl",
                "durationThresDiscardDlMacSdu",
                "volThresDiscardDlMacSdu",
                "nrDrbTcpBoostDN",
                "allowedPdcchPowerBoostALDueCce",
                "allowedPdcchPowerBoostALDueCfg",
                "numSduDiscard",
                "sduDiscardIntervalTmr",
            }

            # deep search for all <p>
            for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):
                name = p.attrib.get("name")

                if nrdrb_id in group_1_ids and name in group_1_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_2_ids and name in group_2_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_3_ids and name in group_3_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_4_ids and name in group_4_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_5_ids and name in group_5_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_6_ids and name in group_6_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_7_ids and name in group_7_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_8_ids and name in group_8_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_9_ids and name in group_9_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_11_ids and name in group_11_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                # -------- Common parameters for all NRDRB IDs --------
                if name in common_nrdrb_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

            # Group-10 merged IDs
            param_values = {}

            if nrdrb_id in group_10_ids:
                for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):
                    name = p.attrib.get("name")

                    if name in group_10_params:
                        param_values[name] = p.text

            for param, val in param_values.items():
                dumy_data.append({
                    "MO": "NRDRB",
                    "DistName": dist_name,
                    "ID": ",".join(map(str, sorted(group_10_ids))),
                    "Parameter": param,
                    "value": tf_to_01(val)
                })

            # Group-12
            if nrdrb_id in group_12_ids:
                for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):
                    name = p.attrib.get("name")

                    if name in group_12_params:
                        dumy_data.append({
                            "MO": "NRDRB",
                            "DistName": dist_name,
                            "ID": nrdrb_id,
                            "Parameter": name,
                            "value": tf_to_01(p.text)
                        })

            if nrdrb_id in group_13_ids and name in group_13_params:
                dumy_data.append({
                    "MO": "NRDRB",
                    "DistName": dist_name,
                    "ID": nrdrb_id,
                    "Parameter": name,
                    "value": tf_to_01(p.text)
                }) 

            if nrdrb_id in group_14_ids and name in group_14_params:
                dumy_data.append({
                    "MO": "NRDRB",
                    "DistName": dist_name,
                    "ID": nrdrb_id,
                    "Parameter": name,
                    "value": tf_to_01(p.text)
                })               
                        # -------- nsaDrbParam --------
            for item in (mo.findall(".//ns:list[@name='nsaDrbParam']/ns:item", ns)
                        if ns else
                        mo.findall(".//list[@name='nsaDrbParam']/item")):

                param_map = {
                    "actPrioritizeX2AfterInactivity": "nsadrbparam@actprioritizex2afterinactivity",
                    "dlDataSplitGainThreshold": "dldatasplitgainthreshold",
                    "dlDataSplitMode": "dldatasplitmode",
                    "initialDLTrafficRouting": "initialdltrafficrouting",
                }

                for p in (item.findall("ns:p", ns) if ns else item.findall("p")):
                    pname = p.attrib.get("name")

                    if pname in param_map:
                        dumy_data.append({
                            "MO": "NRDRB",
                            "DistName": dist_name,
                            "Parameter": param_map[pname].lower(),
                            "value": tf_to_01(p.text)
                        })            
        

        elif mo_class == "NOKLTE:NRMEASDPR":
            dist_name = mo.attrib.get("distName", "")

            required_params = {
                "measgapenablednrmeas",
                "retrytimerb1nr",
                "supervisiontimerb1nr",
            }

            # -------- Direct Parameters --------
            for p in (mo.findall("ns:p", ns) if ns else mo.findall("p")):
                name = p.attrib.get("name", "").lower()

                if name in required_params:
                    dumy_data.append({
                        "MO": "NRMEASDPR",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

            # -------- nrQuantityConfig List --------
            for item in (mo.findall(".//ns:list[@name='nrQuantityConfig']/ns:item", ns)
                        if ns else
                        mo.findall(".//list[@name='nrQuantityConfig']/item")):

                for p in (item.findall("ns:p", ns) if ns else item.findall("p")):
                    name = p.attrib.get("name", "").lower()

                    if name == "filtercoefficientrsrp":
                        dumy_data.append({
                            "MO": "NRMEASDPR",
                            "DistName": dist_name,
                            "Parameter": "filtercoefficientrsrp",
                            "value": tf_to_01(p.text)
                        })

                        
        # for RDRB_5QI class---            
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB_5QI":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_5qi_id = int(dist_name.rsplit("NRDRB_5QI-", 1)[-1])
            except:
                continue

            nrdrb_5qi_param_map = {
                21: {"snssaiDN"},
                22: {"snssaiDN"},
                25: {"snssaiDN"},
                26: {"snssaiDN"},
                47: {"snssaiDN"},
                5: {"snssaiDN"},
                6: {"snssaiDN"},
                7: {"snssaiDN"},
                8: {"snssaiDN"},
                9: {"snssaiDN"},
                1: {"snssaiDN"},
                2: {"snssaiDN"},
            }

            allowed_params = nrdrb_5qi_param_map.get(nrdrb_5qi_id)
            if not allowed_params:
                continue

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                if p.attrib.get("name") == "dataForwarding":
                    dumy_data.append({
                        "MO": "NRDRB_5QI",
                        "DistName": dist_name,
                        "ID": nrdrb_5qi_id,
                        "Parameter": "dataForwarding",
                        "value": tf_to_01(p.text)
                    }) 

                        # -------- snssaiDN --------
            for item in (
                mo.findall(".//ns:list[@name='snssaiList']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='snssaiList']/item")
            ):
                p = item.find("ns:p", ns) if ns else item.find("p")
                if p is None:
                    continue

                name = p.attrib.get("name")
                if name in allowed_params:
                    dumy_data.append({
                        "MO": "NRDRB_5QI",
                        "DistName": dist_name,
                        "ID": nrdrb_5qi_id,
                        "Parameter": "Item-snssaiList-snssaiDN",
                        "value": p.text
                    })
            # -------- Grouped snssaiDN --------

            group_1 = {1, 2, 5, 6, 7, 8, 9}
            group_2 = {21, 22, 25, 26}
            group_3 = {47}

            for item in (
                mo.findall(".//ns:list[@name='snssaiList']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='snssaiList']/item")
            ):
                for p in (item.findall("ns:p", ns) if ns else item.findall("p")):
                    if p.attrib.get("name") != "snssaiDN":
                        continue

                    if nrdrb_5qi_id in group_1:
                        group_id = "1,2,5,6,7,8,9"

                    elif nrdrb_5qi_id in group_2:
                        group_id = "21,22,25,26"

                    elif nrdrb_5qi_id in group_3:
                        group_id = "47"

                    else:
                        continue

                    dumy_data.append({
                        "MO": "NRDRB_5QI",
                        "DistName": dist_name,
                        "ID": group_id,
                        "Parameter": "snssaidn",
                        "value": tf_to_01(p.text)
                    })
            # -------- fiveqiValueList --------
            fiveqi_values = []

            for p in (
                mo.findall(".//ns:list[@name='fiveqiValueList']/ns:p", ns)
                if ns else
                mo.findall(".//list[@name='fiveqiValueList']/p")
            ):
                if p.text:
                    fiveqi_values.append(tf_to_01(p.text))

            if fiveqi_values:
                dumy_data.append({
                    "MO": "NRDRB_5QI",
                    "DistName": dist_name,
                    "ID": nrdrb_5qi_id,
                    "Parameter": "fiveqiValueList",
                    "value": ",".join(map(str, fiveqi_values))
                })  
        #
        # for NRDRB_MAC classs---
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB_MAC":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_mac_id = int(dist_name.rsplit("NRDRB_MAC-", 1)[-1])
            except:
                continue

            group_1_ids={1}
            group_2_ids={2}
            group_3_ids={3}
            group_4_ids={4}
            group_5_ids={4}

            group_1_params={"schedulBSD"}
            group_2_params={"schedulBSD","maxUlHarqTxDrb"}
            group_3_params={"schedulBSD"}
            group_4_params={"schedulBSD"}
            group_5_params={'lcgid','maxDlHarqTxDrb','prioritisedBitRate','schedulPrio','nbrDl','nbrUl'}
            


            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

        
                if nrdrb_mac_id in group_1_ids and name in group_1_params:
                    dumy_data.append({
                        "MO": "NRDRB_MAC",
                        "DistName": dist_name,
                        "ID": nrdrb_mac_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

             
                if nrdrb_mac_id in group_2_ids and name in group_2_params:
                    dumy_data.append({
                        "MO": "NRDRB_MAC",
                        "DistName": dist_name,
                        "ID": nrdrb_mac_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })
             
                if nrdrb_mac_id in group_3_ids and name in group_3_params:
                    dumy_data.append({
                        "MO": "NRDRB_MAC",
                        "DistName": dist_name,
                        "ID": nrdrb_mac_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })  
            
                if nrdrb_mac_id in group_4_ids and name in group_4_params:
                    dumy_data.append({
                        "MO": "NRDRB_MAC",
                        "DistName": dist_name,
                        "ID": nrdrb_mac_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })  
               
                if nrdrb_mac_id in group_5_ids and name in group_5_params:
                    dumy_data.append({
                        "MO": "NRDRB_MAC",
                        "DistName": dist_name,
                        "ID": nrdrb_mac_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })


        # for NRDRB_PDCP class----------
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB_PDCP":
            dist_name = mo.attrib.get("distName", "")

            try:
                nrdrb_pdcp_id = int(dist_name.rsplit("NRDRB_PDCP-", 1)[-1])
            except:
                continue

            nrdrb_pdcp_param_map = {
                1: {
                    "maxCidAllowed",
                    "pdcpSNLength"
                },
                2: {
                    "tReorderingDl",
                    "tReorderingUl",
                    "maxDlHarqTxDrb",
                    "maxUlHarqTxDrb",
                    "pdcpSNLength"
                }
            }

            allowed_params = nrdrb_pdcp_param_map.get(nrdrb_pdcp_id)
            if not allowed_params:
                continue

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name in allowed_params:
                    dumy_data.append({
                        "MO": "NRDRB_PDCP",
                        "DistName": dist_name,
                        "ID": nrdrb_pdcp_id,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    }) 
        

        elif mo_class == "NOKLTE:NRDCDPR":
            dist_name = mo.attrib.get("distName", "")

            required_in_nrdcdpr = {
                "method",
                "actB1NrBeamMeas",
                "allowedBcSelectMethod",
                "dlPathlossChg",
                "enDCpMaxEUTRApowerOffset",
                "enDCpMaxNRpowerOffset",
                "tPeriodicPhr",
                "tProhibitPhr",
            }

            required_in_nrdcdpr_lower = {x.lower() for x in required_in_nrdcdpr}

            # Direct parameters
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.lower() in required_in_nrdcdpr_lower:
                    dumy_data.append({
                        "MO": "NRDCDPR",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })

    
            for lst in (mo.findall("ns:list", ns) if ns else mo.findall("list")):

                list_name = lst.attrib.get("name", "")

                for item in (lst.findall("ns:item", ns) if ns else lst.findall("item")):

                    for p in (item.findall("ns:p", ns) if ns else item.findall("p")):

                        name = p.attrib.get("name")
                        if not name:
                            continue

                        pname = name.lower()

                        if pname not in required_in_nrdcdpr_lower:
                            continue

                        # Special mapping
                        if list_name == "dynTriggerLteNrDcConfList" and pname == "method":
                            parameter = "dyntriggerltenrdcconflist@method"
                        else:
                            parameter = pname

                        dumy_data.append({
                            "MO": "NRDCDPR",
                            "DistName": dist_name,
                            "Parameter": parameter,
                            "value": tf_to_01(p.text)
                        })

        elif mo_class == "com.nokia.srbts.nrbts:NRSYSINFO_PROFILE_NSA":
            dist_name = mo.attrib.get("distName", "")

            required_in = {
                "systemInformationTargetRate",
            }

            required_in_lower = {x.lower() for x in required_in}

            for p in (mo.findall("ns:p", ns) if ns else mo.findall("p")):
                name = p.attrib.get("name", "").lower()

                if name in required_in_lower:
                    dumy_data.append({
                        "MO": "NRSYSINFO_PROFILE_NSA",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })


        #  for classs NRDRB_RLC_AM-----   
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB_RLC_AM":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_rlc_id = int(dist_name.rsplit("NRDRB_RLC_AM-", 1)[-1])
            except:
                continue


            nrdrb_rlc_param_map = {
                1:{'dlMaxRetxThreshold','dlPollByte',
                    'ulMaxRetxThreshold','ulPollByte',
                     'dlTPollRetr','ulTPollRetr'}
                    }

            allowed_params = nrdrb_rlc_param_map.get(nrdrb_rlc_id)
            if not allowed_params:
                continue                
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")
                if name in allowed_params:
                    dumy_data.append({
                        "MO": "NRDRB_RLC_AM",
                        "DistName": dist_name,
                        "ID":nrdrb_rlc_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })  
       

        # for classs NRDRB_RLC_UM-----   
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB_RLC_UM":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_rlcum_id = int(dist_name.rsplit("NRDRB_RLC_UM-", 1)[-1])
            except:
                continue


            nrdrb_rlcum_param_map = {
                1:{'rlcUMDrbSNLength',
                    'ulTReassemblyUm'}
                    }

            allowed_params = nrdrb_rlcum_param_map.get(nrdrb_rlcum_id)
            if not allowed_params:
                continue   
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):       
                name = p.attrib.get("name")
                if name in allowed_params:
                    dumy_data.append({
                        "MO": "NRDRB_RLC_UM",
                        "DistName": dist_name,
                        "ID":nrdrb_rlcum_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    }) 

        # for NRDRB_TCP_BOOST  
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB_TCP_BOOST":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_tcp_id = int(dist_name.rsplit("NRDRB_TCP_BOOST-", 1)[-1])
            except:
                continue


            nrdrb_tcp_param_map = {
                1:{'tcpBoostPugMinSrPeriodicity',
                    'tcpBoostPugDuration',
                    'tcpBoostPugInactiveTime',
                    'tcpBoostPugMinDataSize',
                    'tcpBoostPugRampUpDuration',
                    'tcpBoostPugRampUpInterval',
                    'tcpBoostPugStableInterval',
                    'tcpBoostPugTbs',
                    'tcpBoostPugUlMinSinrThresh',
}
                    }

            allowed_params = nrdrb_tcp_param_map.get(nrdrb_tcp_id)
            if not allowed_params:
                continue                
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")
                if name in allowed_params:
                    dumy_data.append({
                        "MO": "NRDRB_TCP_BOOST",
                        "DistName": dist_name,
                        "ID":nrdrb_tcp_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    }) 


        # for NRPMRNL class---           
        elif mo_class == "com.nokia.srbts.nrbts:NRPMRNL":

            required_in_nrpmrl = {
                    "miNrCellUtilPerNrg",
                    "miNrInterRATMobilitySaPmqap",
                    "miNrSaCuNrpmqap",
                    "miNrNgInterfaceNrpmqap",
                    "miNrPdcpCellNrpmqap",
                    "miNrCellUtilization",
                    "miNrHighRlcCellNrpmqap",
                    "miNrUeStatNrpmqap",
                    "miNrPdcpLatNrpmqap",
                    "miNrHighRlcCell",
                    "miNrPDCPCell",
                    "minrpdcpcellnrpmqap",
                    "minrsacunrpmqap",
                    "minrx2ccnrpmqap",
                    "minrhighrlccellnrpmqap",
                    "minrnsaaccunrpmqap",
                    "minrnginterfacenrpmqap",
                    "miNrPRBthroughputpmqap",
                    "miNrHighRlcLatNrpmqap",
                    "miNrLowRlcReport",
                    "miNrMacSduReport",
                    "miNrRim",
                    "miNrNeighborLteNrCuNsa",
                    "miNrNeighborNrNrCuNsa",
                    "miNrRtwp",
                    "miNrPaging",
                    "minrbeamforming",
                    "minrbeamforming2",
                    "minrbfconfig",
                    "minrcacu",
                    "minrcadu",
                    "minrccd",
                    "minrcellavailability",
                    "minrcellutilperplmn",
                    "minrdlharq",
                    "minrdlsignalquality",
                    "minrdlsignalquality2",
                    "minre1interfacebts",
                    "minre1interfacecell",
                    "minre1interfacecucpnsa",
                    "minre1interfacecucpsa",
                    "minre1interfacecuup",
                    "minre1interfacecuupnsa",
                    "minre1interfacecuupsa",
                    "minrecpriupfcp",
                    "minreirp",
                    "minrf1cc",
                    "minrf1cd",
                    "minrf1ucinterface",
                    "minrf1udinterface",
                    "minrf1uinterfacecupmqap",
                    "minrf1uu",
                    "minrgnss",
                    "minrhighrlcreport",
                    "minrhointerruptcuup",
                    "minrhointerruptcuupnrpmqap",
                    "minrifpscellchange",
                    "minrintergnbmobilitycellcunsa",
                    "minrintergnbmobilitycellcusa",
                    "minrintergnbngmobilitysapmqap",
                    "minrintergnbxnmobilitysapmqap",
                    "minrintragnbmobilitycellcusa",
                    "minrintragnbmobilitysapmqap",
                    "minrl2hicubearerpmqap",
                    "minrltemgpercell",
                    "minrltemobilityreport",
                    "minrmacpdutput",
                    "minrmgpercell",
                    "minrmobilityreport",
                    "minrmobilityreportcu",
                    "minrmobilityreportdu",
                    "minrnaecl",
                    "minrngcc",
                    "minrngcflexinterface",
                    "minrnginterface",
                    "minrngureport",
                    "minrnremo",
                    "minrntraf",
                    "minrpdcch",
                    "minrpdcch2",
                    "minrpdcp1report",
                    "minrpdcp2report",
                    "minrpdcpc",
                    "minrpscellchange",
                    "minrraccu",
                    "minrracdu",
                    "minrrachusage",
                    "minrradbearer",
                    "minrradqualpmqap",
                    "minrrans",
                    "minrrelreportcu",
                    "minrresourceutilizationpercell",
                    "minrrip",
                    "minrrlfpercell",
                    "minrrrcc",
                    "minrrrcd",
                    "minrrsacu",
                    "minrrsadu",
                    "minrrsanb",
                    "minrrsup",
                    "minrs1uinterface",
                    "minrsaacdunrpmqap",
                    "minrsfpbbmod",
                    "minrsfprmod",
                    "minrsfpsmod",
                    "minrslicingcucp",
                    "minrslicingcuup",
                    "minrsparebts",
                    "minrsrb3c",
                    "minrsrb3d",
                    "minrsrsquality",
                    "minrsrvcavastats",
                    "minrssbpm",
                    "minrtimingadvance",
                    "minruestate",
                    "minrulharq",
                    "minrulharqone",
                    "minrulsignalquality",
                    "minrulsignalquality2",
                    "minrx2cb",
                    "minrx2cc",
                    "minrx2uinterface",
                    "minrx2uinterfacepmqap",
                    "minrxnccinterface",
                    "minrxninterface",
                }

            required_in_nrpmrl_lower = {
                x.lower() for x in required_in_nrpmrl
            }

            params_nrpmrl = {}

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.strip().lower() in required_in_nrpmrl_lower:
                    params_nrpmrl[name.lower()] = p.text

                    dumy_data.append({
                        "MO": "NRPMRNL",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })  
                    
        elif mo_class == "com.nokia.srbts.nrbts:NRRIM_PROFILE":
            dist_name = mo.attrib.get("distName", "")

            required_in_NRRIM_PROFILE = {
                "nrRiThresBase",
                "nrRiMonitorDownhillStartSymbol",
                "nrRiMonitorDownhillEndSymbol",
                "nrRiThresDropHeight",
                "nrRiThresDurationRatio",
                "nrRiMonitorWindow",
                "nrRimPowerRampingStep",
                "nrRimRarTpcCommand",
                "nrRimMitigationScheme",
                "rsTransTrigger",
                "rsNearFarConfig",
                "numRbForRs",
                "numRsType1SetId",
                "repetitionRsType1",
                "rsPattern1Period",
                "rsSetId",
                "factorRsSequence",
                "rsPowerOffset",
                "offsetRsSequence",
                "rimDlMutingMaxRange",

                # New Parameters
                "actRiDetectTriggerDlMitigation",
                "actRimAsymmetricIndication",
                "actRimPdcchMutingOnSpecialSlot",
                "actRsEnoughIndication",
                "factorRsSequenceRs2",
                "globalF0ForRs",
                "minDlMitigationDurationWoRi",
                "nrRiDisableMonitorPowerOfUSlot",
                "nrRiMonitorInSSlotEnabled",
                "nrRiThresDropHeightSpecialSlot",
                "numDetectedSetIdForMitigation",
                "numRsType2SetId",
                "numWinRiDetectForDlMitigation",
                "offsetRsSequenceRs2",
                "repetitionRsType2",
                "rimAsymmetricDetectionWindow",
                "rimDlMitigationManualTrigger",
                "rimDynamicSrsSlotOffsetEnabled",
                "rsDetectionWindow",
                "rsSetIdDisplayInFault",
            }

            required_in_NRRIM_PROFILE_lower = {
                x.lower() for x in required_in_NRRIM_PROFILE
            }

            # ---------- Direct Parameters ----------
            for p in (mo.findall("ns:p", ns) if ns else mo.findall("p")):
                name = p.attrib.get("name")

                if name and name.lower() in required_in_NRRIM_PROFILE_lower:
                    dumy_data.append({
                        "MO": "NRRIM_PROFILE",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })

            # ---------- rsType1NScId ----------
            rs_values = []

            for p in (
                mo.findall(".//ns:list[@name='rsType1NScId']/ns:p", ns)
                if ns else
                mo.findall(".//list[@name='rsType1NScId']/p")
            ):
                rs_values.append(tf_to_01(p.text))

            if rs_values:
                dumy_data.append({
                    "MO": "NRRIM_PROFILE",
                    "DistName": dist_name,
                    "Parameter": "rstype1nscid",
                    "value": ",".join(map(str, rs_values))
                })

            # ---------- rsType2NScId ----------
            rs2_values = []

            for p in (
                mo.findall(".//ns:list[@name='rsType2NScId']/ns:p", ns)
                if ns else
                mo.findall(".//list[@name='rsType2NScId']/p")
            ):
                rs2_values.append(tf_to_01(p.text))

            if rs2_values:
                dumy_data.append({
                    "MO": "NRRIM_PROFILE",
                    "DistName": dist_name,
                    "Parameter": "rstype2nscid",
                    "value": ",".join(map(str, rs2_values))
                })

            # ---------- startFreqOffsetForRs ----------
            start_rb_values = []

            for p in (
                mo.findall(".//ns:list[@name='startFreqOffsetForRs']/ns:p", ns)
                if ns else
                mo.findall(".//list[@name='startFreqOffsetForRs']/p")
            ):
                start_rb_values.append(tf_to_01(p.text))

            if start_rb_values:
                dumy_data.append({
                    "MO": "NRRIM_PROFILE",
                    "DistName": dist_name,
                    "Parameter": "startrbforrs",
                    "value": ",".join(map(str, start_rb_values))
                })
                
            start_rb_values = []

            for p in (
                mo.findall(".//ns:list[@name='startFreqOffsetForRs']/ns:p", ns)
                if ns else
                mo.findall(".//list[@name='startFreqOffsetForRs']/p")
            ):
                start_rb_values.append(tf_to_01(p.text))

            if start_rb_values:
                dumy_data.append({
                    "MO": "NRRIM_PROFILE",
                    "DistName": dist_name,
                    "Parameter": "startrbforrs",
                    "value": ",".join(map(str, start_rb_values))
                })    
                
        elif mo_class == "NOKLTE:SIB":
            dist_name = mo.attrib.get("distName", "")

            required_in_sib = {
                "primPlmnUpperLayerIndicationR15",
            }

            required_in_sib_lower = {x.lower() for x in required_in_sib}

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.lower() in required_in_sib_lower:
                    dumy_data.append({
                        "MO": "SIB",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })


        elif mo_class == "NOKLTE:PMRNL":
            dist_name = mo.attrib.get("distName", "")

            required_in_pm = {
                "mtx2gnbpercell",
            }

            required_in_pm_lower = {x.lower() for x in required_in_pm}

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.lower() in required_in_pm_lower:
                    dumy_data.append({
                        "MO": "PMRNL",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    }) 


        # for SNSSAI class---
        elif mo_class == "com.nokia.srbts.nrbts:SNSSAI":
            dist_name = mo.attrib.get("distName", "")
            try:
                snssai_id = int(dist_name.rsplit("SNSSAI-", 1)[-1])
            except:
                continue  

            group_1_ids={1}
            group_2_ids={2}
            group_3_ids={3}

            group_1_params={'sd','sst','userLabel'}
            group_2_params={'sd','sst','userLabel'}
            group_3_params={'sd','sst','userLabel'}
          

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

        
                if snssai_id in group_1_ids and name in group_1_params:
                    dumy_data.append({
                        "MO": "SNSSAI",
                        "DistName": dist_name,
                        "ID": snssai_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

             
                if snssai_id in group_2_ids and name in group_2_params:
                    dumy_data.append({
                        "MO": "SNSSAI",
                        "DistName": dist_name,
                        "ID": snssai_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    }) 

                if snssai_id in group_3_ids and name in group_3_params:
                    dumy_data.append({
                        "MO": "SNSSAI",
                        "DistName": dist_name,
                        "ID": snssai_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })     

            

        # for TRACKINGAREA class---
        elif mo_class == "com.nokia.srbts.nrbts:TRACKINGAREA":
            dist_name = mo.attrib.get("distName", "")

            # fiveGsTac
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                if p.attrib.get("name") == "fiveGsTac":
                    dumy_data.append({
                        "MO": "TRACKINGAREA",
                        "DistName": dist_name,
                        "Parameter": "fiveGsTac",
                        "value": tf_to_01(p.text)
                    })

            # snssaiList
            snssai_values = []

            for item in (
                mo.findall(".//ns:list[@name='snssaiList']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='snssaiList']/item")
            ):
                p = item.find("ns:p", ns) if ns else item.find("p")

                if p is not None and p.attrib.get("name") == "snssaiDN":
                    snssai_values.append(p.text)

            if snssai_values:
                dumy_data.append({
                    "MO": "TRACKINGAREA",
                    "DistName": dist_name,
                    "Parameter": "snssaiDN",
                    "value": ";".join(snssai_values)
                })
            

        #for NRPMQAP class-----------------------------
        elif mo_class == "com.nokia.srbts.nrbts:NRPMQAP":

            dist_name = mo.attrib.get("distName", "")
            try:
                nrpmqap_id = int(dist_name.split("NRPMQAP-")[-1])
            except:
                continue

            group_1_ids = {11, 12, 15, 16, 17, 18, 19}
            group_2_ids = {21, 22, 25, 26}
            group_3_ids = {11,12,15,16,17,18,19,21,22,25,26,47}
            group_4_ids={47}
            group_5_ids = {21, 22, 25, 26,47}
            group_6_ids = {6, 7, 8, 9}
            group_7_ids = {6, 7, 8, 9, 11, 12, 15, 16, 17, 18, 19}
            group_8_ids = {11, 12, 15, 16, 17, 18, 19}

            if nrpmqap_id not in (
                group_1_ids
                | group_2_ids
                | group_3_ids
                | group_4_ids
                | group_5_ids
                | group_6_ids
                | group_7_ids
                | group_8_ids
            ):
                continue
              
        
            if nrpmqap_id in group_1_ids:

                for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):
                    if p.attrib.get("name") == "thpHistScale":
                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "DistName": dist_name,
                            "ID": nrpmqap_id,
                            "Parameter": "thpHistScale",
                            "value": tf_to_01(p.text)
                        })

                # cfgSliceId marker
                if mo.findall(".//ns:list[@name='cfgSliceId']", ns) if ns else mo.findall(".//list[@name='cfgSliceId']"):
                    dumy_data.append({
                        "MO": "NRPMQAP",
                        "DistName": dist_name,
                        "ID": nrpmqap_id,
                        "Parameter": "cfgSliceId",
                        "value": "List"
                    })

                # sd / sst
                for item in mo.findall(".//ns:list[@name='cfgSliceId']/ns:item", ns) \
                        if ns else mo.findall(".//list[@name='cfgSliceId']/item"):
                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                        if p.attrib.get("name") in {"sd", "sst"}:
                            dumy_data.append({
                                "MO": "NRPMQAP",
                                "DistName": dist_name,
                                "ID": nrpmqap_id,
                                "Parameter": f"cfgplmnid@{p.attrib.get('name').lower()}",
                                "value": tf_to_01(p.text)
                            })


            if nrpmqap_id in group_2_ids:

                simple_params = {
                    "cfg5qiRange",
                }

                for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):
                    name = p.attrib.get("name")
                    if name in simple_params:
                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "ID": nrpmqap_id,
                            "Parameter": "cfgqcirange",
                            "value": tf_to_01(p.text)   
                        })

                # cfgPlmnId
                if mo.findall(".//ns:list[@name='cfgPlmnId']", ns) if ns else mo.findall(".//list[@name='cfgPlmnId']"):
                    dumy_data.append({
                        "MO": "NRPMQAP",
                        "ID": nrpmqap_id,
                        "Parameter": "cfgPlmnId",
                        "value": "List"               
                    })

                for item in mo.findall(".//ns:list[@name='cfgPlmnId']/ns:item", ns) \
                        if ns else mo.findall(".//list[@name='cfgPlmnId']/item"):
                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "ID": nrpmqap_id,
                            "Parameter": f"cfgsliceid@{p.attrib.get('name').lower()}",
                            "value": tf_to_01(p.text)  
                        })

                # cfgSliceId
                if mo.findall(".//ns:list[@name='cfgSliceId']", ns) if ns else mo.findall(".//list[@name='cfgSliceId']"):
                    dumy_data.append({
                        "MO": "NRPMQAP",
                        "ID": nrpmqap_id,
                        "Parameter": "cfgSliceId",
                        "value": "List"               
                    })

                for item in mo.findall(".//ns:list[@name='cfgSliceId']/ns:item", ns) \
                        if ns else mo.findall(".//list[@name='cfgSliceId']/item"):
                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                        if p.attrib.get("name") in {"sd", "sst"}:
                            dumy_data.append({
                                "MO": "NRPMQAP",
                                "ID": nrpmqap_id,
                                "Parameter": f"cfgsliceid@{p.attrib.get('name').lower()}",
                                "value": tf_to_01(p.text)
                            })

            param_values = {}
            if nrpmqap_id in group_3_ids:
                simple_params = {
                    'actUPlaneGroup1Counters',
                    'actUPlaneGroup2Counters',
                }
  
                for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):
                    name = p.attrib.get("name")

                    if name in simple_params:
                        param_values[name] = tf_to_01(p.text)

                for param, val in param_values.items():
                    dumy_data.append({
                        "MO": "NRPMQAP",
                        "ID": ",".join(map(str, sorted(group_3_ids))), 
                        "Parameter": param,
                        "value": val
                    })



            if nrpmqap_id in group_4_ids:

                simple_params = {
                    'cfg5qiRange',
                }

                # ✅ only direct <p>
                for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                    name = p.attrib.get("name")
                    if name in simple_params:
                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "ID": nrpmqap_id,
                            "Parameter": "cfgqcirange",
                            "value": tf_to_01(p.text)
                        })

                # ✅ cfgPlmnId
                if mo.findall(".//ns:list[@name='cfgPlmnId']", ns) if ns else mo.findall(".//list[@name='cfgPlmnId']"):
                    dumy_data.append({
                        "MO": "NRPMQAP",
                        "ID": nrpmqap_id,
                        "Parameter": "cfgPlmnId",
                        "value": "List"
                    })

                for item in mo.findall(".//ns:list[@name='cfgPlmnId']/ns:item", ns) \
                        if ns else mo.findall(".//list[@name='cfgPlmnId']/item"):

                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "ID": nrpmqap_id,
                            "Parameter": f"Item-cfgPlmnId-{p.attrib.get('name')}",
                            "value": tf_to_01(p.text)
                        })

                # ✅ ADD THIS (MISSING PART - IMPORTANT 🔥)
                for item in mo.findall(".//ns:list[@name='cfgSliceId']/ns:item", ns) \
                        if ns else mo.findall(".//list[@name='cfgSliceId']/item"):

                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                        if p.attrib.get("name") in {"sd", "sst"}:
                            dumy_data.append({
                                "MO": "NRPMQAP",
                                "ID": nrpmqap_id,
                                "Parameter": f"cfgsliceid@{p.attrib.get('name').lower()}",
                                "value": tf_to_01(p.text)
                            })

            if nrpmqap_id in group_5_ids:

                simple_params = {
                    "actCplaneCounters",
                    "actL2Counters",
                    "actPacketSchedulerCounters",
                    "cfgProfType",
                    "thpHistDownlinkMaxRange",
                    "thpHistDownlinkMinRange",
                    "thpHistScale",
                    "thpHistUplinkMaxRange",
                    "thpHistUplinkMinRange",
                }

                for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):
                    name = p.attrib.get("name")
                    if name in simple_params:
                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "ID": nrpmqap_id,
                            "Parameter": name,
                            "value": tf_to_01(p.text)   
                        })                  
         
            if nrpmqap_id in group_6_ids:

                for item in (
                    mo.findall(".//ns:list[@name='cfgPlmnId']/ns:item", ns)
                    if ns else
                    mo.findall(".//list[@name='cfgPlmnId']/item")
                ):

                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):

                        name = p.attrib.get("name")

                        if name in {"mcc", "mnc", "mncLength"}:

                            dumy_data.append({
                                "MO": "NRPMQAP",
                                "ID": nrpmqap_id,
                                "Parameter": f"cfgplmnid@{name.lower()}",
                                "value": tf_to_01(p.text)
                            })

            if nrpmqap_id in group_7_ids:

                simple_params = {
                    "cfgProfType",
                    "thpHistDownlinkMaxRange",
                    "thpHistDownlinkMinRange",
                    "thpHistScale",
                    "thpHistUplinkMaxRange",
                    "thpHistUplinkMinRange"
                }

                for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):

                    name = p.attrib.get("name")

                    if name in simple_params:

                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "ID": nrpmqap_id,
                            "Parameter": name,
                            "value": tf_to_01(p.text)
                        })
            # -------- Grouped cfg5qiRange --------

            cfg5qi_group_map = {
                "11,21": {11, 21},
                "12,22": {12, 22},
                "15,25": {15, 25},
                "16,26": {16, 26},
                "17": {17},
                "18": {18},
                "19": {19},
                "47": {47},
            }

            for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):

                if p.attrib.get("name") != "cfg5qiRange":
                    continue

                for group_id, ids in cfg5qi_group_map.items():

                    if nrpmqap_id in ids:

                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "ID": group_id,
                            "Parameter": "cfg5qirange",
                            "value": tf_to_01(p.text)
                        })

                        break            
                                    
            if nrpmqap_id in group_8_ids:

                for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):

                    if p.attrib.get("name") == "cfg5qiRange":

                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "ID": nrpmqap_id,
                            "Parameter": "cfgqcirange",
                            "value": tf_to_01(p.text)
                        })


            group_nrbts_ids = {6, 7, 8, 9, 11, 12, 15, 16, 17, 18, 19}
            if nrpmqap_id in group_nrbts_ids:

                simple_params = {
                    "thpHistDownlinkMaxRange",
                    "thpHistDownlinkMinRange",
                    "thpHistScale",
                    "thpHistUplinkMaxRange",
                    "thpHistUplinkMinRange"
                }

                for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):

                    name = p.attrib.get("name")
                    if name in simple_params:
                        dumy_data.append({
                            "MO": "NRBTS",
                            "DistName": dist_name,
                            "Parameter": name.lower(),
                            "value": tf_to_01(p.text)
                        })
            
                


        #for ---------- NRRESOURCEGROUP ----------
        elif mo_class == "com.nokia.srbts.nrbts:NRRESOURCEGROUP":

            dist_name = mo.attrib.get("distName", "")
            try:
                rg_id = int(dist_name.split("NRRESOURCEGROUP-")[-1])
            except:
                continue

            # dumy_data.append({
            #     "MO": "NRRESOURCEGROUP",
            #     "ID": rg_id,
            #     "Parameter": "NRRESOURCEGROUP",
            #     "value": rg_id
            # })

            resource_dns = []

            for item in mo.findall(".//ns:list[@name='nrResourceList']/ns:item", ns) \
                    if ns else mo.findall(".//list[@name='nrResourceList']/item"):

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    if p.attrib.get("name") == "resourceDN":
                        resource_dns.append(p.text)

            if resource_dns:
                dumy_data.append({
                    "MO": "NRRESOURCEGROUP",
                    "ID": rg_id,
                    "Parameter": "resourcedn",
                    "value": ";".join(resource_dns)
                })

            # ---- schedulerParams ----
            sched_params = {}

            for item in mo.findall(".//ns:list[@name='schedulerParams']/ns:item", ns) \
                    if ns else mo.findall(".//list[@name='schedulerParams']/item"):

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    sched_params[p.attrib.get("name")] = p.text

            if sched_params:
                for k, v in sched_params.items():
                    dumy_data.append({
                        "MO": "NRRESOURCEGROUP",
                        "ID": rg_id,
                        "Parameter": f"schedulerparams@{k.lower()}",
                        "value": tf_to_01(v)
                    })


            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                if p.attrib.get("name") == "userLabel":
                    dumy_data.append({
                        "MO": "NRRESOURCEGROUP",
                        "ID": rg_id,
                        "Parameter": "userLabel",
                        "value": p.text
                    })

#---------
        elif mo_class == "com.nokia.srbts.nrbts:NRDRX":
            dist_name = mo.attrib.get("distName", "")

            try:
                nrdrx_id = int(dist_name.rsplit("NRDRX-", 1)[-1])
            except:
                continue

            # -------- NRDRX-2 --------
            if nrdrx_id == 2:

                required_params = {
                    "drxHarqRttTimerDl",
                    "drxHarqRttTimerUl",
                    "drxInactivityTimer",
                    "drxLongCycle",
                    "drxOnDurationTimer",
                    "drxPriority",
                    "drxRetransTimerDl",
                    "drxRetransTimerUl",
                    "userLabel"
                }

            # -------- NRDRX-3 & NRDRX-4 (Old Logic) --------
            elif nrdrx_id in {3, 4}:

                required_params = {
                    "drxInactivityTimer",
                    "drxLongCycle",
                    "drxOnDurationTimer",
                    "drxRetransTimerDl",
                    "drxRetransTimerUl",
                    "drxPriority"
                }

            else:
                continue

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name in required_params:
                    dumy_data.append({
                        "MO": "NRDRX",
                        "DistName": dist_name,
                        "ID": nrdrx_id,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })

        # -------- NRRESOURCEGROUP_PROFILE --------
        elif mo_class == "com.nokia.srbts.nrbts:NRRESOURCEGROUP_PROFILE":

            dist_name = mo.attrib.get("distName", "")
            try:
                profile_id = int(dist_name.split("NRRESOURCEGROUP_PROFILE-")[-1])
            except:
                continue

            # MO row
            dumy_data.append({
                "MO": "NRRESOURCEGROUP_PROFILE",
                "ID": profile_id,
                "Parameter": "NRRESOURCEGROUP_PROFILE",
                "value": profile_id
            })

            # simple p
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                if p.attrib.get("name") == "applyCellRacForPrivilegedUE":
                    dumy_data.append({
                        "MO": "NRRESOURCEGROUP_PROFILE",
                        "ID": profile_id,
                        "Parameter": "applyCellRacForPrivilegedUE",
                        "value": tf_to_01(p.text)
                    })

            # -------- defaultRgNsaList --------
            if mo.findall(".//ns:list[@name='defaultRgNsaList']", ns) \
                if ns else mo.findall(".//list[@name='defaultRgNsaList']"):

                dumy_data.append({
                    "MO": "NRRESOURCEGROUP_PROFILE",
                    "ID": profile_id,
                    "Parameter": "defaultRgNsaList",
                    "value": "List"
                })

                for item in mo.findall(".//ns:list[@name='defaultRgNsaList']/ns:item", ns) \
                        if ns else mo.findall(".//list[@name='defaultRgNsaList']/item"):
                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                        dumy_data.append({
                            "MO": "NRRESOURCEGROUP_PROFILE",
                            "ID": profile_id,
                            "Parameter": f"Item-defaultRgNsaList-{p.attrib.get('name')}",
                            "value": p.text
                        })

            # -------- defaultRgSaList --------
            if mo.findall(".//ns:list[@name='defaultRgSaList']", ns) \
                if ns else mo.findall(".//list[@name='defaultRgSaList']"):

                dumy_data.append({
                    "MO": "NRRESOURCEGROUP_PROFILE",
                    "ID": profile_id,
                    "Parameter": "defaultRgSaList",
                    "value": "List"
                })

                for item in mo.findall(".//ns:list[@name='defaultRgSaList']/ns:item", ns) \
                        if ns else mo.findall(".//list[@name='defaultRgSaList']/item"):
                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                        dumy_data.append({
                            "MO": "NRRESOURCEGROUP_PROFILE",
                            "ID": profile_id,
                            "Parameter": f"Item-defaultRgSaList-{p.attrib.get('name')}",
                            "value": p.text
        
                        })

        elif mo_class == "com.nokia.srbts.nrbts:NRSYSINFO_PROFILE":
            dist_name = mo.attrib.get("distName", "")
            # -------- Root level parameters --------
            required_root = {
                "systeminformationtargetrate",
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name", "").lower()

                if name in required_root:
                    dumy_data.append({
                        "MO": "NRSYSINFO_PROFILE",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

            # -------- sibSchedulingList parameters --------
            required_sib = {
                "sibperiodicity",
                "sibtype",
            }

            for item in (
                mo.findall(".//ns:list[@name='sibSchedulingList']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='sibSchedulingList']/item")
            ):

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):

                    name = p.attrib.get("name", "").lower()

                    if name in required_sib:
                        dumy_data.append({
                            "MO": "NRSYSINFO_PROFILE",
                            "DistName": dist_name,
                            "Parameter": f"sibschedulinglist@{name}",
                            "value": tf_to_01(p.text)
                        })

        elif mo_class == "com.nokia.srbts.nrbts:NRSYSINFO_PROFILE":
            dist_name = mo.attrib.get("distName", "")

            required_params = {
                "SystemInformationTargetRate",
            }

            required_params_lower = {x.lower() for x in required_params}

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.lower() in required_params_lower:
                    dumy_data.append({
                        "MO": "NRSYSINFO_PROFILE",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })
        elif mo_class == "com.nokia.srbts.nrbts:NRSYSINFO_PROFILE_NSA":
            dist_name = mo.attrib.get("distName", "")

            required_NRSYSINFO_PROFILE_NSA = {
                "systemInformationTargetRate"
            }

            required_in_NRSYSINFO_PROFILE_NSA_lower = {
                x.lower() for x in required_NRSYSINFO_PROFILE_NSA
            }

            params_NRSYSINFO_PROFILE_NSA = {}

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.strip().lower() in required_in_NRSYSINFO_PROFILE_NSA_lower:
                    params_NRSYSINFO_PROFILE_NSA[name.lower()] = p.text

                    dumy_data.append({
                        "MO": "NRSYSINFO_PROFILE_NSA",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })
        
        elif mo_class == "com.nokia.srbts.nrbts:PDCCH":
            dist_name = mo.attrib.get("distName", "")
            required_pdcch = {
                "alSelection"
            }
            required_dynamic_agg = {
                "aggregationLevelListHR": "dynamicAggregationLevelSet@aggregationLevelList",
                "cqiDciCssAl1HR": "dynamicAggregationLevelSet@cqiDciCssAl1",
                "cqiDciCssAl2HR": "dynamicAggregationLevelSet@cqiDciCssAl2",
                "cqiDciCssAl4HR": "dynamicAggregationLevelSet@cqiDciCssAl4",
                "cqiDciCssAl8HR": "dynamicAggregationLevelSet@cqiDciCssAl8",
                "cqiDciCssAl16HR": "dynamicAggregationLevelSet@cqiDciCssAl16"
            }
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name in required_pdcch:
                    dumy_data.append({
                        "MO": "PDCCH",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })

            # dynamicAggregationLevelHRSet parameters
            for item in (
                mo.findall(".//ns:list[@name='dynamicAggregationLevelHRSet']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='dynamicAggregationLevelHRSet']/item")
            ):

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    name = p.attrib.get("name")

                    if name in required_dynamic_agg:
                        dumy_data.append({
                            "MO": "PDCCH",
                            "DistName": dist_name,
                            "Parameter": required_dynamic_agg[name].lower(),
                            "value": tf_to_01(p.text)
                        })
        
        
        elif mo_class == "com.nokia.srbts.nrbts:NRPMCCP":
            dist_name = mo.attrib.get("distName", "")

            # -------- Direct Parameters --------
            required_params = {
                "actnrsctpcounters",
                "cfgdlairintdelaycounters",
                "cfgtacounters",
                "cfgthptdistupdate",
                "cfgulairintdelaycounters",
                "puschsinrbinmaxthresh",
                "puschsinrbinminthresh",
                "rrctabinfarthresh",
                "rrctabinnearthresh"
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name", "").strip()

                if name.lower() in required_params:
                    dumy_data.append({
                        "MO": "NRPMCCP",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })

            # -------- actRadioQualityMeasCount --------
            for lst in (
                mo.findall(".//ns:list[@name='actRadioQualityMeasCount']", ns)
                if ns else
                mo.findall(".//list[@name='actRadioQualityMeasCount']")
            ):
                for p in lst.findall("ns:p", ns) if ns else lst.findall("p"):
                    dumy_data.append({
                        "MO": "NRPMCCP",
                        "DistName": dist_name,
                        "Parameter": "actradioqualitymeascount",
                        "value": tf_to_01(p.text)
                    })
        elif mo_class == "com.nokia.srbts.mnl:PMMNL":
            dist_name = mo.attrib.get("distName", "")

            required_in_pmmnl = {
                "mtBtsEnergyMonitoring",
                "mtSBTSRfmEnergyMonitoring",
                "mtSBTSSmEnergyMonitoring",

            }

            required_in_pmmnl_lower = {
                x.lower() for x in required_in_pmmnl
            }

            params_pmmnl = {}
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.lower() in required_in_pmmnl_lower:
                    params_pmmnl[name.lower()] = p.text

                    dumy_data.append({
                        "MO": "PMMNL",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })
        

        elif mo_class == "com.nokia.srbts.eqm:RMOD":
            dist_name = mo.attrib.get("distName", "")

            required_in_rmod = {
                "energySavingMode"
            }

            required_in_rmod_lower = {
                x.lower() for x in required_in_rmod
            }

            params_rmod = {}
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.lower() in required_in_rmod_lower:
                    params_rmod[name.lower()] = p.text

                    dumy_data.append({
                        "MO": "RMOD",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })

        elif mo_class == "com.nokia.srbts.mnl:SLCELLDRCONF":
            dist_name = mo.attrib.get("distName", "")

            required_in_SLCELLDRCONF = {
                   "mtSBTSSmEnergyMonitoring",
                    "nrDegAdmCtrlSlCellDetHChFrq",
                    "nrDegAdmCtrlSlCellDetMinPrcThr",
                    "nrDegAdmCtrlSlCellDetThr",
                    "nrDegAdmCtrlSlCellRecPh1Tmr",
                    "nrDegAdmCtrlSlCellRecPh3Tmr",
                    "nrDegAdmCtrlSlCellRecPhAct",
                    "nrUeCtxStpRtSlCellDetHChFrq",
                    "nrUeCtxStpRtSlCellDetMinReqThr",
                    "nrUeCtxStpRtSlCellDetThr",
                    "nrUeCtxStpRtSlCellRecPh1Tmr",
                    "nrUeCtxStpRtSlCellRecPh2Tmr",
                    "nrUeCtxStpRtSlCellRecPh3Tmr",
                    "nrUeCtxStpRtSlCellRecPhAct",
                    "nrSgnbAddCplSlCellDetHChFrq",
                    "nrSgnbAddCplSlCellDetMinResThr",
                    "nrSgnbAddCplSlCellDetThr",
                    "nrSgnbAddCplSlCellRecPh1Tmr",
                    "nrSgnbAddCplSlCellRecPh2Tmr",
                    "nrSgnbAddCplSlCellRecPh3Tmr",
                    "nrSgnbAddCplSlCellRecPhAct",
                    "nrRxPwrSlCellDetHChFrq",
                    "nrRxPwrSlCellDetHighPwrThr",
                    "nrRxPwrSlCellDetLowPwrThr",
                    "nrRxPwrSlCellRecPh1Tmr",
                    "nrRxPwrSlCellRecPh2Tmr",
                    "nrRxPwrSlCellRecPh3Tmr",
                    "nrRxPwrSlCellRecPhAct",
                    "nrCntRachStpSlCellDetHChFrq",
                    "nrCntRachStpSlCellDetMsg3Thr",
                    "nrCntRachStpSlCellDetThr",
                    "nrCntRachStpSlCellRecPh1Tmr",
                    "nrCntRachStpSlCellRecPh2Tmr",
                    "nrCntRachStpSlCellRecPh3Tmr",
                    "nrCntRachStpSlCellRecPhAct",
                    "nrTxPwrSlCellDetHChFrq",
                    "nrTxPwrSlCellDetThr",
                    "nrTxPwrSlCellRecPh1Tmr",
                    "nrTxPwrSlCellRecPh2Tmr",
                    "nrTxPwrSlCellRecPh3Tmr",
                    "nrTxPwrSlCellRecPhAct"

            }

            required_in_SLCELLDRCONF_lower = {
                x.lower() for x in required_in_SLCELLDRCONF
            }

            params_SLCELLDRCONF = {}
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.lower() in required_in_SLCELLDRCONF_lower:
                    params_SLCELLDRCONF[name.lower()] = p.text

                    dumy_data.append({
                        "MO": "SLCELLDRCONF",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })            
        

        elif mo_class == "com.nokia.srbts.nrbts:BWP_PROFILE":
            dist_name = mo.attrib.get("distName", "")

            try:
                bwp_profile_id = int(dist_name.split("BWP_PROFILE-")[-1])
            except:
                continue

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                if p.attrib.get("name") == "bwpType":
                    dumy_data.append({
                        "MO": "BWP_PROFILE",
                        "DistName": dist_name,
                        "ID": bwp_profile_id,
                        "Parameter": "bwpType",
                        "value": tf_to_01(p.text)
                    })

            for item in (
                mo.findall(".//ns:list[@name='locationAndBandwidthDl']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='locationAndBandwidthDl']/item")
            ):

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):

                    if p.attrib.get("name") in {"cRbSizeDl", "cRbStartDl"}:
                        dumy_data.append({
                            "MO": "BWP_PROFILE",
                            "DistName": dist_name,
                            "ID": bwp_profile_id,
                            "Parameter": p.attrib.get("name"),
                            "value": tf_to_01(p.text)
                        })
        


        elif mo_class == "com.nokia.srbts.tnl:DSCP2PCPMAP":
            dist_name = mo.attrib.get("distName", "")

            required_in_DSCP2PCPMAP = {
                "vLanPrioForDscp00",
                "vLanPrioForDscp01",
                "vLanPrioForDscp02",
                "vLanPrioForDscp03",
                "vLanPrioForDscp04",
                "vLanPrioForDscp05",
                "vLanPrioForDscp06",
                "vLanPrioForDscp07",
                "vLanPrioForDscp08",
                "vLanPrioForDscp09",
                "vLanPrioForDscp10",
                "vLanPrioForDscp11",
                "vLanPrioForDscp12",
                "vLanPrioForDscp13",
                "vLanPrioForDscp14",
                "vLanPrioForDscp15",
                "vLanPrioForDscp16",
                "vLanPrioForDscp17",
                "vLanPrioForDscp18",
                "vLanPrioForDscp19",
                "vLanPrioForDscp20",
                "vLanPrioForDscp21",
                "vLanPrioForDscp22",
                "vLanPrioForDscp23",
                "vLanPrioForDscp24",
                "vLanPrioForDscp25",
                "vLanPrioForDscp26",
                "vLanPrioForDscp27",
                "vLanPrioForDscp28",
                "vLanPrioForDscp29",
                "vLanPrioForDscp30",
                "vLanPrioForDscp31",
                "vLanPrioForDscp32",
                "vLanPrioForDscp33",
                "vLanPrioForDscp34",
                "vLanPrioForDscp35",
                "vLanPrioForDscp36",
                "vLanPrioForDscp37",
                "vLanPrioForDscp38",
                "vLanPrioForDscp39",
                "vLanPrioForDscp40",
                "vLanPrioForDscp41",
                "vLanPrioForDscp42",
                "vLanPrioForDscp43",
                "vLanPrioForDscp44",
                "vLanPrioForDscp45",
                "vLanPrioForDscp46",
                "vLanPrioForDscp47",
                "vLanPrioForDscp48",
                "vLanPrioForDscp49",
                "vLanPrioForDscp50",
                "vLanPrioForDscp51",
                "vLanPrioForDscp52",
                "vLanPrioForDscp53",
                "vLanPrioForDscp54",
                "vLanPrioForDscp55",
                "vLanPrioForDscp56",
                "vLanPrioForDscp57",
                "vLanPrioForDscp58",
                "vLanPrioForDscp59",
                "vLanPrioForDscp60",
                "vLanPrioForDscp61",
                "vLanPrioForDscp62",
                "vLanPrioForDscp63",

                  

            }

            required_in_DSCP2PCPMAP_lower = {
                x.lower() for x in required_in_DSCP2PCPMAP
            }

            params_DSCP2PCPMAP = {}
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.lower() in required_in_DSCP2PCPMAP_lower:
                    params_DSCP2PCPMAP[name.lower()] = p.text

                    dumy_data.append({
                        "MO": "DSCP2PCPMAP",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })            
        

        elif mo_class == "com.nokia.srbts.tnl:DSCP2QMAP":
            dist_name = mo.attrib.get("distName", "")

            required_in_DSCP2QMAP = {
                    "queueForDscp00",
                    "queueForDscp01",
                    "queueForDscp02",
                    "queueForDscp03",
                    "queueForDscp04",
                    "queueForDscp05",
                    "queueForDscp06",
                    "queueForDscp07",
                    "queueForDscp08",
                    "queueForDscp09",
                    "queueForDscp10",
                    "queueForDscp11",
                    "queueForDscp12",
                    "queueForDscp13",
                    "queueForDscp14",
                    "queueForDscp15",
                    "queueForDscp16",
                    "queueForDscp17",
                    "queueForDscp18",
                    "queueForDscp19",
                    "queueForDscp20",
                    "queueForDscp21",
                    "queueForDscp22",
                    "queueForDscp23",
                    "queueForDscp24",
                    "queueForDscp25",
                    "queueForDscp26",
                    "queueForDscp27",
                    "queueForDscp28",
                    "queueForDscp29",
                    "queueForDscp30",
                    "queueForDscp31",
                    "queueForDscp32",
                    "queueForDscp33",
                    "queueForDscp34",
                    "queueForDscp35",
                    "queueForDscp36",
                    "queueForDscp37",
                    "queueForDscp38",
                    "queueForDscp39",
                    "queueForDscp40",
                    "queueForDscp41",
                    "queueForDscp42",
                    "queueForDscp43",
                    "queueForDscp44",
                    "queueForDscp45",
                    "queueForDscp46",
                    "queueForDscp47",
                    "queueForDscp48",
                    "queueForDscp49",
                    "queueForDscp50",
                    "queueForDscp51",
                    "queueForDscp52",
                    "queueForDscp53",
                    "queueForDscp54",
                    "queueForDscp55",
                    "queueForDscp56",
                    "queueForDscp57",
                    "queueForDscp58",
                    "queueForDscp59",
                    "queueForDscp60",
                    "queueForDscp61",
                    "queueForDscp62",
                    "queueForDscp63",
                }

            required_in_DSCP2QMAP_lower = {
                x.lower() for x in required_in_DSCP2QMAP
            }

            params_DSCP2QMAP = {}
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.lower() in required_in_DSCP2QMAP_lower:
                    params_DSCP2QMAP[name.lower()] = p.text

                    dumy_data.append({
                        "MO": "DSCP2QMAP",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })

        elif mo_class == "com.nokia.srbts.nrbts:NRPMQAP":
            dist_name = mo.attrib.get("distName", "")

            required_in_NRPMQAP = {
                "thpHistDownlinkMaxRange",
                "thpHistDownlinkMinRange",
                "thpHistScale",
                "thpHistUplinkMaxRange",
                "thpHistUplinkMinRange"
            }

            required_in_NRPMQAP_lower = {
                x.lower() for x in required_in_NRPMQAP
            }

            params_NRPMQAP = {}

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):

                name = p.attrib.get("name", "")

                if name.lower() in required_in_NRPMQAP_lower:

                    params_NRPMQAP[name.lower()] = p.text

                    dumy_data.append({
                        "MO": "NRPMQAP",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    }) 

        
        elif mo_class == "com.nokia.srbts.mnl:FEATCADM":
            dist_name = mo.attrib.get("distName", "")

            required_in_FEATCADM = {
                    "actPowerMeter",
                    "actNrAutDegAdmCtrlSlCellDetRec",
                    "actNrAutUeCtxStpRtSlCellDetRec",
                    "actNrAutSgnbAddCplSlCellDetRec",
                    "actNrAutRxPwrSlCellDetRec",
                    "actNrAutCntRachStpSlCellDetRec",
                    "actNrAutTxPwrSlCellDetRec",

                  
                }

            required_in_FEATCADM_lower = {
                x.lower() for x in required_in_FEATCADM
            }

            params_FEATCADM = {}
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

                if name and name.lower() in required_in_FEATCADM_lower:
                    params_FEATCADM[name.lower()] = p.text

                    dumy_data.append({
                        "MO": "FEATCADM",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })
        
        elif mo_class == "NOKLTE:LNCEL":
            dist_name = mo.attrib.get("distName", "")

            required_in_lncel = {
                "actflowctrlenh",
                "actperiodicb1nrmeas",
                "dlcaminpcellcqiqci1",
                "ulcaminpcellsinrqci1",
                "actmicrodtx"
            }

            required_in_lncel_lower = {x.lower() for x in required_in_lncel}

            # -------- Simple Parameters --------
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name", "").strip()

                if name.lower() in required_in_lncel_lower:
                    dumy_data.append({
                        "MO": "LNCEL",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

            # -------- drxProfile105 --------
            drx_required = {
                "drxinactivityt",
                "drxlongcycle",
                "drxonduratt",
                "drxprofileindex",
                "drxprofilepriority",
                "drxretranst"
            }

            for item in (
                mo.findall(".//ns:list[@name='drxProfile105']/ns:item", ns)
                if ns
                else mo.findall(".//list[@name='drxProfile105']/item")
            ):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):

                    param = p.attrib.get("name", "").strip()

                    if param.lower() in drx_required:
                        dumy_data.append({
                            "MO": "LNCEL",
                            "DistName": dist_name,
                            "Parameter": f"drxProfile105@{param}",
                            "value": tf_to_01(p.text)
                        })

        elif mo_class == "NOKLTE:ANR":
            required_anr = {
                "actautonrneighremoval",
                "anrnrcoverageconfirmtimer",
                "anrnrinactivitytimer",
                "anrnrnbrrelremguardtime",
            }

            required_anr_lower = {x.lower() for x in required_anr}

            # -------- simple <p> parameters --------
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name", "").strip()

                if name.lower() in required_anr_lower:
                    dumy_data.append({
                        "MO": "ANR",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

            # -------- anrNrMeasConfList --------
            meas_required = {
                "parammeasconfid",
                "thresholdnrrsrp",
                "thresholdnrrsrq",
            }

            for item in (
                mo.findall(".//ns:list[@name='anrNrMeasConfList']/ns:item", ns)
                if ns
                else mo.findall(".//list[@name='anrNrMeasConfList']/item")
            ):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param = p.attrib.get("name", "").strip()

                    if param.lower() in meas_required:
                        dumy_data.append({
                            "MO": "ANR",
                            "DistName": dist_name,
                            "Parameter": f"anrNrMeasConfList@{param}",
                            "value": tf_to_01(p.text)
                        })

            # -------- anrNrParamList --------
            param_required = {
                "anrdefreldcallowed",
                "frqnrcarselect",
                "gnbidlength",
                "parammeasconfid",
                "paramplmnid",
            }

            for item in (
                mo.findall(".//ns:list[@name='anrNrParamList']/ns:item", ns)
                if ns
                else mo.findall(".//list[@name='anrNrParamList']/item")
            ):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param = p.attrib.get("name", "").strip()

                    if param.lower() in param_required:
                        dumy_data.append({
                            "MO": "ANR",
                            "DistName": dist_name,
                            "Parameter": f"anrNrParamList@{param}",
                            "value": tf_to_01(p.text)
                        })

            # -------- anrNrParamPlmnList --------
            plmn_required = {
                "mcc",
                "mnc",
                "mnclength",
                "paramplmnid",
            }

            for item in (
                mo.findall(".//ns:list[@name='anrNrParamPlmnList']/ns:item", ns)
                if ns
                else mo.findall(".//list[@name='anrNrParamPlmnList']/item")
            ):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    param = p.attrib.get("name", "").strip()

                    if param.lower() in plmn_required:
                        dumy_data.append({
                            "MO": "ANR",
                            "DistName": dist_name,
                            "Parameter": f"anrNrParamPlmnList@{param}",
                            "value": tf_to_01(p.text)
                        })
        
        elif mo_class == "NOKLTE:ENDCDMEASCONF":
            required_endcdmeasconf = {
                "b1nrbeamreportquantity",
                "b1nrbeamthreshrsrp",
                "b1nrmaxnumbeams",
                "b1nrreportamount",
                "b1nrreportinterval",
                "endcnrdropthresh",
                "endcnrdropthreshtimer",
                "endcnrdropwaittimer",
                "nrpcinbrreladdhyst",
                "nrpcinbrreladdthresh",
                "nrpcinbrrelremthresh",
                "ssbduration",
                "ssboffset",
                "ssbperiodicity",
                "ssbsubcarrierspacing"
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name", "").strip()

                if name.lower() in required_endcdmeasconf:
                    dumy_data.append({
                        "MO": "ENDCDMEASCONF",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

        elif mo_class == "NOKLTE:LNBTS":
            required_lnbts = {
                "actb1extmeasnr",
                "actcaggrltenrdualconnectivity",
                "actcoordinated4g5gpowersaving",
                "actdrxcoordendc",
                "actendcautox2setup",
                "actendcblockpercausevalue",
                "actendcdropmonitor",
                "actendcltedlmimooptbcsel",
                "actendcltedltputbcsel",
                "actendcnrfeatsetidbcsel",
                "actenhflexcawithendc",
                "actevtsecratrep",
                "actltenrdualconnectivity",
                "actltenrfastnbrreladd",
                "actmeasgapcoordendc",
                "actmeasgapifendc",
                "actmultiscgsplit",
                "actmultiscgsplitwoelcid",
                "actrcrendc",
                "acts1inducedx2closure",
                "actsecratrep",
                "actstepaddbearendc",
                "actuebasedanrnr",
                "actuetrigendctakeback",
                "actul256qamforendc",
                "actulcaforendc",
                "actx2gnb128",
                "enablebwcomsetchkendc",
                "nrcelldeactstatindendcho",
                "nrrestrictionoverride",
                "tltenrdualconnectprep",
                "ts1erabmodind",
                "tx2sgnbreldata",
                ""
            }

            # ---------- Direct Parameters ----------
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name", "").strip()

                if name.lower() in required_lnbts:
                    dumy_data.append({
                        "MO": "LNBTS",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

            # ---------- pdcpProf1001 ----------
            pdcp_required = {
                "pdcpprofileid",
                "snsizedl",
                "snsizeul",
                "statusrepreq",
                "tdiscard"
            }

            for item in (
                mo.findall(".//ns:list[@name='pdcpProf1001']/ns:item", ns)
                if ns
                else mo.findall(".//list[@name='pdcpProf1001']/item")
            ):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):

                    param = p.attrib.get("name", "").strip()

                    if param.lower() in pdcp_required:
                        dumy_data.append({
                            "MO": "LNBTS",
                            "DistName": dist_name,
                            "Parameter": f"pdcpProf1001@{param}",
                            "value": tf_to_01(p.text)
                        })

            # ---------- qciTab1 ----------
            for item in (
                mo.findall(".//ns:list[@name='qciTab1']/ns:item", ns)
                if ns
                else mo.findall(".//list[@name='qciTab1']/item")
            ):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    if p.attrib.get("name", "").lower() == "dscp":
                        dumy_data.append({
                            "MO": "LNBTS",
                            "DistName": dist_name,
                            "Parameter": "qciTab1@dscp",
                            "value": tf_to_01(p.text)
                        })

            # ---------- qciTab2 ----------
            for item in (
                mo.findall(".//ns:list[@name='qciTab2']/ns:item", ns)
                if ns
                else mo.findall(".//list[@name='qciTab2']/item")
            ):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    if p.attrib.get("name", "").lower() == "dscp":
                        dumy_data.append({
                            "MO": "LNBTS",
                            "DistName": dist_name,
                            "Parameter": "qciTab2@dscp",
                            "value": tf_to_01(p.text)
                        })

            # ---------- qciTab3 ----------
            for item in (
                mo.findall(".//ns:list[@name='qciTab3']/ns:item", ns)
                if ns
                else mo.findall(".//list[@name='qciTab3']/item")
            ):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    if p.attrib.get("name", "").lower() == "dscp":
                        dumy_data.append({
                            "MO": "LNBTS",
                            "DistName": dist_name,
                            "Parameter": "qciTab3@dscp",
                            "value": tf_to_01(p.text)
                        })

            # ---------- qciTab4 ----------
            for item in (
                mo.findall(".//ns:list[@name='qciTab4']/ns:item", ns)
                if ns
                else mo.findall(".//list[@name='qciTab4']/item")
            ):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    if p.attrib.get("name", "").lower() == "dscp":
                        dumy_data.append({
                            "MO": "LNBTS",
                            "DistName": dist_name,
                            "Parameter": "qciTab4@dscp",
                            "value": tf_to_01(p.text)
                        })

            # ---------- qciTab5 ----------
            qci5_required = {
                "dscp",
                "ltenrdualconnectsupport"
            }

            for item in (
                mo.findall(".//ns:list[@name='qciTab5']/ns:item", ns)
                if ns
                else mo.findall(".//list[@name='qciTab5']/item")
            ):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):

                    param = p.attrib.get("name", "").strip()

                    if param.lower() in qci5_required:
                        dumy_data.append({
                            "MO": "LNBTS",
                            "DistName": dist_name,
                            "Parameter": f"qciTab5@{param}",
                            "value": tf_to_01(p.text)
                        })

            # ---------- qciTab6/7/8/9 ----------
            qci_required = {
                "arpendcmask",
                "dscp",
                "ltenrdualconnectsupport",
                "nrpdcpprofidx"
            }

            for tab in ["qciTab6", "qciTab7", "qciTab8", "qciTab9"]:

                for item in (
                    mo.findall(f".//ns:list[@name='{tab}']/ns:item", ns)
                    if ns
                    else mo.findall(f".//list[@name='{tab}']/item")
                ):

                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):

                        param = p.attrib.get("name", "").strip()

                        if param.lower() in qci_required:
                            dumy_data.append({
                                "MO": "LNBTS",
                                "DistName": dist_name,
                                "Parameter": f"{tab}@{param}",
                                "value": tf_to_01(p.text)
                            })  

            qci_map = {
                "qciTab1": {"dscp"},
                "qciTab2": {"dscp"},
                "qciTab3": {"dscp"},
                "qciTab4": {"dscp"},
                "qciTab5": {
                    "dscp",
                    "ltenrdualconnectsupport",
                },
                "qciTab6": {
                    "arpendcmask",
                    "dscp",
                    "ltenrdualconnectsupport",
                    "nrpdcpprofidx",
                },
                "qciTab7": {
                    "arpendcmask",
                    "dscp",
                    "ltenrdualconnectsupport",
                    "nrpdcpprofidx",
                },
                "qciTab8": {
                    "arpendcmask",
                    "dscp",
                    "ltenrdualconnectsupport",
                    "nrpdcpprofidx",
                },
                "qciTab9": {
                    "arpendcmask",
                    "dscp",
                    "ltenrdualconnectsupport",
                    "nrpdcpprofidx",
                },
            }

            for tab_name, required_params in qci_map.items():

                for item in (
                    mo.findall(f".//ns:list[@name='{tab_name}']/ns:item", ns)
                    if ns else
                    mo.findall(f".//list[@name='{tab_name}']/item")
                ):

                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):

                        param = p.attrib.get("name", "").strip()

                        if param.lower() in required_params:
                            dumy_data.append({
                                "MO": "LNBTS",
                                "DistName": dist_name,
                                "Parameter": f"{tab_name}{param}",
                                "value": tf_to_01(p.text)
                            })                
            
            pdcp_required = {
                "pdcpprofileid",
                "snsizedl",
                "snsizeul",
                "statusrepreq",
                "tdiscard"
            }

            for item in (
                mo.findall(".//ns:list[@name='pdcpProf1001']/ns:item", ns)
                if ns
                else mo.findall(".//list[@name='pdcpProf1001']/item")
            ):
                for p in item.findall("ns:p", ns) if ns else item.findall("p"):

                    param = p.attrib.get("name", "").strip()

                    if param.lower() in pdcp_required:
                        dumy_data.append({
                            "MO": "LNBTS",
                            "DistName": dist_name,
                            "Parameter": f"pdcpProf1001{param}",
                            "value": tf_to_01(p.text)
                        })

        elif mo_class == "NOKLTE:LNBTS_FDD":
            required_lnbts_fdd = {
                "numtxwithhighnongbr"
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name", "").strip()

                if name.lower() in required_lnbts_fdd:
                    dumy_data.append({
                        "MO": "LNBTS_FDD",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
        
                    })  

        elif mo_class == "com.nokia.srbts.nrbts:NRANRPRE":
            dist_name = mo.attrib.get("distName", "")

            required_nranrpre = {
                "iratdataforwardingmethod",
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name", "").strip()

                if name.lower() in required_nranrpre:
                    dumy_data.append({
                        "MO": "NRANRPRE",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })  

        elif mo_class == "com.nokia.srbts.nrbts:NRADJECELL":
            dist_name = mo.attrib.get("distName", "")

            required_nradjecell = {

                 "numtxwithhighnongbr"
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name", "").strip()

                if name.lower() in required_nradjecell:
                    dumy_data.append({
                        "MO": "NRADJECELL",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })            

        elif mo_class == "NOKLTE:LNBTS_TDD":
            required_lnbts_tdd = {
                "numtxwithhighnongbr"
            }

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name", "").strip()

                if name.lower() in required_lnbts_tdd:
                    dumy_data.append({
                        "MO": "LNBTS_TDD",
                        "DistName": dist_name,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })    

        elif mo_class == "NOKLTE:NRDCDPR":
            dist_name = mo.attrib.get("distName", "")

            required_in = {
                "tProhibitPhr",
            }

            required_in_lower = {
                x.lower() for x in required_in
            }

            params = {}

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name", "")

                if name and name.lower() in required_in_lower:
                    params[name.lower()] = p.text

                    dumy_data.append({
                        "MO": "MRBTS",
                        "DistName": dist_name,
                        "Parameter": name.lower(),
                        "value": tf_to_01(p.text)
                    })

#---------------data read in fix paratmeter---------      
        df = pd.DataFrame(dumy_data)
        for c in ["MO", "ID", "Parameter", "value"]:
            if c not in df:
                df[c] = ""

        # Clean ID
        def clean_id(x):
            if pd.isna(x) or x == "":
                return ""
            
            x = str(x)
            
            if "," in x:  # multiple IDs
                return ",".join(str(int(float(i))) for i in x.split(","))
            
            return str(int(float(x)))

        df["ID"] = df["ID"].apply(clean_id).astype(str)

        

        df_47 = df[df["ID"] == "47"].copy()
        df_rest = df[df["ID"] != "47"].copy()


        separate_params = ["bwpType", "cRbSizeDl", "cRbStartDl"]


      
        normal_df = (
            df_rest[
                (df_rest["Parameter"] != "cfg5qiRange")
                &
                ~(
                    (df_rest["MO"] == "BWP_PROFILE")
                    &
                    (df_rest["Parameter"].isin(separate_params))
                )
            ]
            .groupby(["MO", "Parameter", "value"], as_index=False)
            .agg({
                "ID": lambda x: ",".join(sorted(set(i for i in x if i)))
            })
        )


        # BWP_PROFILE parameters → keep ID-wise
        bwp_df = (
            df_rest[
                (df_rest["MO"] == "BWP_PROFILE")
                &
                (df_rest["Parameter"].isin(separate_params))
            ]
            .groupby(["MO", "ID", "Parameter"], as_index=False)
            .agg({
                "value": "first"
            })
        )


        # cfg5qiRange special logic
        cfg_df = df_rest[df_rest["Parameter"] == "cfg5qiRange"].copy()

        cfg_df["pair"] = cfg_df.apply(
            lambda r: f"{r['ID']}-{r['value']}",
            axis=1
        )

        cfg_df = (
            cfg_df.groupby(["MO", "Parameter"], as_index=False)
            .agg({
                "ID": lambda x: ",".join(sorted(set(x))),
                "pair": lambda x: ",".join(sorted(set(x)))
            })
            .rename(columns={"pair": "value"})
        )


        # Combine everything
        grouped_df = pd.concat(
            [normal_df, bwp_df, cfg_df],
            ignore_index=True
        )


        # Add ID=47 records back
        data_df = pd.concat(
            [grouped_df, df_47],
            ignore_index=True
        )

        # Final formatting
        data_df = data_df[["MO", "ID", "Parameter", "value"]]
        data_df.rename(columns={"value": "value(External)"}, inplace=True)
        file_name_1 = "Nokia_Slicing_dump_data.xlsx"
        dump_output_path=os.path.join(dump_data_path, file_name_1)
        data_df.to_excel(dump_output_path, index=False, engine="openpyxl")
       
        data_df["Parameter"] = (
            data_df["Parameter"]
            .astype(str)
            .str.strip()
            .str.lower()
        )
        data_df["MO"] = data_df["MO"].astype(str).str.strip().str.upper()

    #matincting-----
    fixpara_folder = os.path.join(main_folder, 'Nokia_Slicing_Fixpara_uls')
    excel_dfs = []

    for file in os.listdir(fixpara_folder):

        if file.startswith('~$'):  # ignore temp excel
            continue

        if file.endswith(('.xlsx', '.xls')):
            file_path = os.path.join(fixpara_folder, file)
            print("Reading FixPara:", file_path)

            df = pd.read_excel(file_path, engine="openpyxl")
            df["source_file"] = file
            excel_dfs.append(df)

    # merge all fix parameter files
    if excel_dfs:
        excel_df = pd.concat(excel_dfs, ignore_index=True)
    else:
        excel_df = pd.DataFrame()


    excel_df["ID"] = excel_df["ID"].apply(normalize_id)
    excel_df["MO"] = excel_df["MO"].astype(str).str.strip().str.upper()
    excel_df["Parameter"] = (
            excel_df["Parameter"]
            .astype(str)
            .str.strip()
            .str.lower()
        )
    data_df["ID"] = data_df["ID"].apply(normalize_id)


    finaldf = excel_df.merge(
        data_df[["MO", "ID", "Parameter","value(External)"]],
        on=["MO", "ID", "Parameter"],
        how="left"
    )

    if "5G/LTE" in finaldf.columns:
        col = finaldf.pop("5G/LTE")
        finaldf.insert(finaldf.columns.get_loc("ID") + 1, "5G/LTE", col)
        
    
    if "Audit if CellDepType =" in finaldf.columns:
        col = finaldf.pop("Audit if CellDepType =")
        finaldf.insert(
            finaldf.columns.get_loc("5G/LTE") + 1,
            "Audit if CellDepType =",
            col
        )

    fallback_map = (
        data_df
        .dropna(subset=["value(External)"])
        .groupby(["MO", "Parameter"])["value(External)"]
        .first()
        .to_dict()
    )

    mask_na = finaldf["value(External)"].isna()

    finaldf.loc[mask_na, "value(External)"] = finaldf.loc[mask_na].apply(
        lambda r: fallback_map.get((r["MO"], r["Parameter"])),
        axis=1
    )

    finaldf["value(External)"] = finaldf["value(External)"].apply(normalize)
    finaldf["Value(Internal)"] = finaldf["Value(Internal)"].apply(normalize)

    finaldf["Final_Value"] = finaldf["Value(Internal)"]
    finaldf["Remark"] = "No Change"

    finaldf.loc[finaldf["value(External)"].isna(), "Remark"] = "Not Found"

    mask_change = (
        finaldf["value(External)"].notna() &
        (finaldf["value(External)"] != finaldf["Value(Internal)"])
    )

    finaldf.loc[mask_change, "Final_Value"] = finaldf.loc[mask_change, "value(External)"]
    finaldf.loc[mask_change, "Remark"] = "Value Changed"

    finaldf = finaldf.sort_values(["MO", "ID", "Parameter"])
    finaldf.drop(columns=["Final_Value", "Remark","source_file"], inplace=True)
    finaldf["Remarks"] = finaldf.apply(
    lambda r: remark(r["Value(Internal)"], r["value(External)"]),
    axis=1
)  

    #for change in xml---- 
    changed_df = finaldf[
        finaldf["Remarks"] == "Changes in value"
    ].copy()

    changed_df["MO"] = changed_df["MO"].astype(str).str.upper()
    changed_df["Parameter"] = changed_df["Parameter"].astype(str).str.lower()
    changed_df["ID"] = changed_df["ID"].astype(str)
    

    file_name = "5G__UPE_ULS_output.xlsx"
    final_output_path=os.path.join(output_path, file_name)
    finaldf.drop_duplicates(inplace=True)
    finaldf.to_excel(final_output_path, index=False, engine="openpyxl",sheet_name="Slicing")
    format_excel(final_output_path,"Slicing")

    #for change in xml---- 
    change_df = finaldf[
    finaldf["Remarks"] == "Changes in value"
    ].copy()

    change_dict = {}
    for _, row in change_df.iterrows():
        mo = str(row["MO"]).upper()
        parameter = str(row["Parameter"]).lower()
        new_value = str(row["Value(Internal)"])
        try:
            id_ = int(str(row["ID"]).split(",")[0])
        except:
            continue
        change_dict[(mo, id_, parameter)] = new_value

    for mo in root.findall(".//ns:managedObject", ns) if ns else root.findall(".//managedObject"):
        mo_class = mo.attrib.get("class", "")
        dist_name = mo.attrib.get("distName", "")
        mo_name = mo_class.split(":")[-1].upper()

        try:
            mo_id = int(dist_name.rsplit("-", 1)[-1])
        except:
            continue

        # direct p tags
        for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):

            parameter = p.attrib.get("name", "").lower()

            key = (mo_name, mo_id, parameter)

            if key in change_dict:
                p.text = change_dict[key]

        # list items
        for item in (
            mo.findall(".//ns:item", ns)
            if ns else
            mo.findall(".//item")
        ):

            for p in item.findall("ns:p", ns) if ns else item.findall("p"):

                parameter = p.attrib.get("name", "").lower()

                key = (mo_name, mo_id, parameter)

                if key in change_dict:
                    p.text = change_dict[key]

    updated_file = os.path.join(output_path, "Updated_XML.xml")
    xml_data = ET.tostring(
        root,
        encoding="utf-8",
        xml_declaration=True
    )
    with open(updated_file, "wb") as f:
        f.write(xml_data)                
  

  
    relative_xml_path = os.path.relpath(updated_file,MEDIA_ROOT).replace("\\", "/")
    xml_download_url = request.build_absolute_uri(MEDIA_URL + relative_xml_path)
    relative_path = os.path.relpath(final_output_path, MEDIA_ROOT).replace("\\", "/")
    download_url = request.build_absolute_uri(MEDIA_URL + relative_path)



    print('UPE ULS GPL Process is Completed-' )
    return Response({
        "status": True,
        "message": "UPE ULS Parsed Successfully",
        "download_url": download_url,
        "xml_dowload":xml_download_url

     
    }, status=HTTP_200_OK)









