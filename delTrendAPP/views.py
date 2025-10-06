##updated Del Trend code-----
import pandas as pd
import datetime
from datetime import date, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.core.files.storage import FileSystemStorage
from zipfile import ZipFile
from tkinter import *
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl as xl
from commom_utilities.utils import *
from zipfile import ZipFile, ZIP_DEFLATED
from django.conf import settings
import os
import shutil
import zipfile

def mark_rows_vectorized(df):
    specified_columns = [
        "Short name",
        "date",
        "MV Freq_Band",
        "Freq_Bandwidth",
        "4G_ECGI"
    ]
    complete_empty_mask = (
        df[specified_columns].isnull()
        | (df[specified_columns] == 0)
        | (df[specified_columns] == "")
    ).all(axis=1)
    non_specified_columns = df.columns.difference(specified_columns)
    partial_empty_mask = (
        df[non_specified_columns].isnull()
        | (df[non_specified_columns] == 0)
        | (df[non_specified_columns] == "")
    ).all(axis=1)
    df["mark_value"] = "non_empty"
    df.loc[complete_empty_mask, "mark_value"] = "complete_empty"
    df.loc[partial_empty_mask & ~complete_empty_mask, "mark_value"] = "partial"

    new_column_order = specified_columns + [col for col in df.columns if col not in specified_columns]
    df = df[new_column_order]

    # Save to Excel
    # df.to_excel("marked_data.xlsx", index=False)


    return df["mark_value"]


def process_remove_duplicates(data):
    # data = data.drop(columns=["id"], axis=1)

    print(data.columns[4:])

    columns_to_convert = data.columns[4:]
    data[columns_to_convert] = data[columns_to_convert].apply(
        pd.to_numeric, errors="coerce"
    )

    data["mark_value"] = mark_rows_vectorized(data)

    # data.to_csv("data.csv", index=False)

    data = data[data["mark_value"] != "partial"]

    return data

@api_view(["POST"])
def old_del_trend(request):
    kpi = [
        "MV_RRC Setup Success Rate",
        "ERAB Setup Success Rate [CDBH]",
        "PS Drop Call Rate % [CBBH]",
        "MV_DL User Throughput_Kbps [CDBH]",
        "MV_UL User Throughput_Kbps [CDBH]",
        "PS handover success rate [LTE Intra System] [CDBH]",
        "PS handover success rate [LTE Inter System] [CDBH]",
        "MV_4G Data Volume_GB",
        "MV_CSFB Redirection Success Rate [CDBH]",
        "RRC Paging Discard Ratio",
        "MV_Average number of used DL PRBs [CDBH]",
        "VoLTE Call Setup Success rate [CBBH]",
        "VoLTE Drop Call Rate [CBBH]]",
        "MV_VoLTE Packet Loss DL [CBBH]",
        "MV_VoLTE Packet Loss UL [CBBH]",
        "VoLTE Intra HOSR [CBBH]",
        "VoLTE InterF HOSR Exec [CBBH]",
        "VoLTE SRVCC SR Exec [CBBH]",
        "VoLTE Traffic [CDBH]",
        "MV_VoLTE Traffic",
        "4G Data Volume [GB] [CDBH]",
        "UL RSSI [CBBH]",
        "MV_E-UTRAN Average CQI [CDBH]",
        "Max Connected User [CDBH]",
        "Max Connected User",
    ]

    gsm = [
        "Total Voice Traffic [BBH]",
        "SDCCH Drop Call Rate [BBH]",
        "SDCCH Drop Call Rate_Nom [BBH]",
        "SDCCH Blocking Rate [BBH]",
        "TCH Blocking Rate [BBH]",
        "Drop Call Rate [BBH]",
        "RX Quality [BBH]",
        "Handover Success Rate [BBH]",
        "Handover Success Rate_Nom [BBH]",
        "Handover Success Rate_Denom [BBH]",
        "TCH Assignment Success Rate [BBH]",
        "ICM Band4-5 [BBH]",
        "Number of Available TCH [BBH]",
        "TNDROP [BBH]",
    ]

    raw_kpi_4G = request.FILES["raw_kpi_4G"] if "raw_kpi_4G" in request.FILES else None

    df_raw_kpi_4G = pd.read_excel(raw_kpi_4G)

    print(df_raw_kpi_4G)

    sts, response = required_col_check(raw_kpi_4G, kpi)
    if sts:
        return Response(response)
    
    print("procced________________________")
    response, s_l = site_list_handler_4G(request)
    s_l = [str(site) for site in s_l]
    # print("_______________success_______________________", s_l)
    if s_l:
        site_list = s_l

        if response:
            return Response(response)

    raw_kpi_2G = request.FILES["raw_kpi_2G"] if "raw_kpi_2G" in request.FILES else None
    df_raw_kpi_2G = pd.read_excel(raw_kpi_2G)
    print("2G_:- ",df_raw_kpi_2G)

    sts, response = required_col_check(raw_kpi_2G, gsm)
    if sts:
        return Response(response)

    response, s_l = site_list_handler_2G(request)
    s_l = [str(site) for site in s_l]
    print("site_list_cate....", s_l)
    if s_l:
        df_2G_site = s_l
        # print("_________________________print", type(df_2G_site[0]))

        # print(df_2G_site, "______2G_ROW_KPI_________", s_l)
        if response:
            return Response(response)

    door_path = os.path.join(MEDIA_ROOT, "trends", "del", "del4G")
    if not os.path.exists(door_path):
        os.makedirs(door_path, exist_ok=True)

    for x in kpi:
        df_raw_kpi_4G[x] = df_raw_kpi_4G[x].replace(
            to_replace=".*", value=0, regex=True
        )

    df_raw_kpi_4G["Short name"].fillna(inplace=True, method="ffill")
    df_raw_kpi_4G.rename(columns={"Unnamed: 1": "date"}, inplace=True)
    df_raw_kpi_4G.rename(columns={"Date": "date"}, inplace=True)
    df_raw_kpi_4G=process_remove_duplicates(df_raw_kpi_4G)


    lis = list(df_raw_kpi_4G["Short name"])
    site_id_lis = []
    cell_id_lis = []
    for item in lis:
        if "_" in item:
            cell_id = item.split("_")[-2]
            ln = len(item.split("_")[-1])
            sit_id = item.split("_")[-2][:-ln]
        else:
            cell_id = item
            sit_id = item
        cell_id_lis.append(cell_id)
        site_id_lis.append(sit_id)

    df_raw_kpi_4G.insert(1, "SITE_ID", site_id_lis)
    df_raw_kpi_4G.insert(2, "CELL_ID", cell_id_lis)

    df_raw_kpi_4G.rename(columns={"Short name": "Shortname"}, inplace=True)
    df_raw_kpi_4G.fillna(value=0, inplace=True)

    print("final_4g df", df_raw_kpi_4G)

    enb = []
    for cell in df_raw_kpi_4G["SITE_ID"]:
        if pd.isnull(cell):
            pass
        else:
            enbid = cell[1:]
            enb.append(enbid)
    df_raw_kpi_4G.insert(3, "ENBID", enb)

    df_raw_kpi_4G["ENBID"] = df_raw_kpi_4G["ENBID"].astype(str)


    add = []
    for i in df_raw_kpi_4G["SITE_ID"]:
        ad = 9
        add.append(ad)
    df_raw_kpi_4G.insert(4, "add", add)

    df_raw_kpi_4G["SECTOR"] = df_raw_kpi_4G["add"].astype(str) + df_raw_kpi_4G["ENBID"]
    # df_raw_kpi_4G
    message_4G = site_comparision(site_id_lis, site_list)  # site_comparision_call

    PsOs_path1 = os.path.join(door_path, "process_outputs", "desired_input.xlsx")
    # df_raw_kpi_4G.to_excel(PsOs_path1)

    str_date = request.POST.get("offered_date")
    date1 = datetime.datetime.strptime(str_date, "%Y-%m-%d")
    dt1 = date1 - timedelta(1)
    dt2 = date1 - timedelta(2)
    dt3 = date1 - timedelta(3)
    dt4 = date1 - timedelta(4)
    dt5 = date1 - timedelta(5)
    ls = [dt1, dt2, dt3, dt4, dt5]

    print("final_4g df", df_raw_kpi_4G)

    def perticular_tech(tech, site_list):

        print(type(df_raw_kpi_4G['ENBID'][0]))
        print(df_raw_kpi_4G.columns)
        

        df_filtered = df_raw_kpi_4G[
            (df_raw_kpi_4G['ENBID'].isin(site_list))
            & (df_raw_kpi_4G['date'].isin(ls))
            & (df_raw_kpi_4G['Shortname'].str.contains("|".join(tech)))
        ]
        print("df_filtered:- ",df_filtered)
        
        print(df_filtered.columns)
        
        for col in df_filtered.columns:
            if col not in ['Shortname', 'SITE_ID', 'CELL_ID', 'ENBID', 'add', 'date', 'MV Freq_Band', 'SECTOR']:
             df_filtered[col] = df_filtered[col].astype(str).str.replace(",", "").str.strip()
             df_filtered[col] = pd.to_numeric(df_filtered[col], errors='coerce')
        

        PsOs_filter = os.path.join(
            door_path, "process_outputs", "last_filtered_input.xlsx"
        )

       # Better pivot_table call
        df_pivoted = df_filtered.pivot_table(
            index=["SITE_ID", "Shortname", "CELL_ID", "ENBID", "SECTOR"],
            columns="date",
            aggfunc='first'
        )
        df_raw_kpi_4G[x] = df_raw_kpi_4G[x].replace(to_replace="^-$|^na$|^N/A$", value=0, regex=True)

        print(df_pivoted)
        
        save_name = str(tech) + "_pivot.xlsx"
        PsOs_pivot = os.path.join(door_path, "process_outputs", save_name)

        return df_filtered, df_pivoted

    alpha = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    def num_hash(num):
        if num < 26:
            return alpha[num - 1]
        else:
            q, r = num // 26, num % 26
            if r == 0:
                if q == 1:
                    return alpha[r - 1]
                else:
                    return num_hash(q - 1) + alpha[r - 1]
            else:
                return num_hash(q) + alpha[r - 1]

    def titleToNumber(s):
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord("A") + 1
        return result

    def overwrite(df_pivoted, kpi_name, coln1, trend_ws):
        coln2 = num_hash(titleToNumber(coln1) + 1)
        coln3 = num_hash(titleToNumber(coln1) + 2)
        coln4 = num_hash(titleToNumber(coln1) + 3)
        coln5 = num_hash(titleToNumber(coln1) + 4)
        # #print(kpi_name)
        index_pivot = df_pivoted.index.to_list()
        # print("index ;###############################",index_pivot)
        # #print(len(index_pivot))
        # print("index of pivoted table: ",index_pivot)
        print("columns of dr table", df_pivoted)
        dr = df_pivoted[kpi_name]
        # print("columns of dr table",dr.columns)
        cl = dr.columns.to_list()
        # #print("column list",cl)

        # site_id=dr["SITE_ID"].to_list()
        # cell_id=dr["CELL_ID"].to_list()
        col1 = dr[str(cl[0])].to_list()
        col2 = dr[str(cl[1])].to_list()
        col3 = dr[str(cl[2])].to_list()
        col4 = dr[str(cl[3])].to_list()
        col5 = dr[str(cl[4])].to_list()

        trend_ws[coln1 + "2"].value = cl[0]
        trend_ws[coln2 + "2"].value = cl[1]
        trend_ws[coln3 + "2"].value = cl[2]
        trend_ws[coln4 + "2"].value = cl[3]
        trend_ws[coln5 + "2"].value = cl[4]

        # me=column_index_from_string(coln5)+1
        # me=get_column_letter(me)
        for i, value in enumerate(index_pivot):
            j = i + 3
            trend_ws["B" + str(j)].value = index_pivot[i][1]
            trend_ws["E" + str(j)].value = index_pivot[i][1]
            trend_ws["A" + str(j)].value = index_pivot[i][0]
            trend_ws["C" + str(j)].value = index_pivot[i][3]
            trend_ws["D" + str(j)].value = index_pivot[i][4]
            trend_ws["J" + str(j)].value = date1
            trend_ws["F" + str(j)].value = "DL"

            trend_ws[coln1 + str(j)].value = 0 if pd.isna(col1[i]) else col1[i]
            trend_ws[coln2 + str(j)].value = 0 if pd.isna(col2[i]) else col2[i]
            trend_ws[coln3 + str(j)].value = 0 if pd.isna(col3[i]) else col3[i]
            trend_ws[coln4 + str(j)].value = 0 if pd.isna(col4[i]) else col4[i]
            trend_ws[coln5 + str(j)].value = 0 if pd.isna(col5[i]) else col5[i]

            # trend_ws[me+str(j)].value='=COUNTIF(P5:T5,">=98.5")'

    # for fdd
    pivot_fdd = perticular_tech(["_F3_"], site_list)[1]
    PsOs_blnk_temp = os.path.join(
        door_path, "template", "DEL KPIs Submission_L1800.xlsx"
    )
    trend_wb_L1800 = xl.load_workbook(PsOs_blnk_temp)  # added xl.
    trend_ws = trend_wb_L1800["KPI"]
    for kpi_name in kpi:
        if kpi_name == "MV_RRC Setup Success Rate":
            overwrite(pivot_fdd, kpi_name, "N", trend_ws)

        if kpi_name == "ERAB Setup Success Rate [CDBH]":
            overwrite(pivot_fdd, kpi_name, "T", trend_ws)

        if kpi_name == "PS Drop Call Rate % [CBBH]":
            overwrite(pivot_fdd, kpi_name, "Z", trend_ws)

        if kpi_name == "MV_DL User Throughput_Kbps [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AF", trend_ws)

        if kpi_name == "MV_UL User Throughput_Kbps [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AL", trend_ws)

        if kpi_name == "PS handover success rate [LTE Intra System] [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AQ", trend_ws)

        if kpi_name == "PS handover success rate [LTE Inter System] [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AW", trend_ws)

        if kpi_name == "MV_4G Data Volume_GB":
            overwrite(pivot_fdd, kpi_name, "DW", trend_ws)

        if kpi_name == "MV_CSFB Redirection Success Rate [CDBH]":
            overwrite(pivot_fdd, kpi_name, "BH", trend_ws)

        if kpi_name == "RRC Paging Discard Ratio":
            overwrite(pivot_fdd, kpi_name, "BN", trend_ws)

        if kpi_name == "MV_Average number of used DL PRBs [CDBH]":
            overwrite(pivot_fdd, kpi_name, "BS", trend_ws)

        if kpi_name == "VoLTE Call Setup Success rate [CBBH]":
            overwrite(pivot_fdd, kpi_name, "BX", trend_ws)

        if kpi_name == "VoLTE Drop Call Rate [CBBH]]":
            overwrite(pivot_fdd, kpi_name, "CD", trend_ws)

        if kpi_name == "MV_VoLTE Packet Loss DL [CBBH]":
            overwrite(pivot_fdd, kpi_name, "CJ", trend_ws)

        if kpi_name == "MV_VoLTE Packet Loss UL [CBBH]":
            overwrite(pivot_fdd, kpi_name, "CP", trend_ws)

        if kpi_name == "VoLTE Intra HOSR [CBBH]":
            overwrite(pivot_fdd, kpi_name, "CV", trend_ws)

        if kpi_name == "VoLTE InterF HOSR Exec [CBBH]":
            overwrite(pivot_fdd, kpi_name, "DB", trend_ws)

        if kpi_name == "VoLTE SRVCC SR Exec [CBBH]":
            overwrite(pivot_fdd, kpi_name, "DH", trend_ws)

        if kpi_name == "VoLTE Traffic [CDBH]":
            overwrite(pivot_fdd, kpi_name, "DM", trend_ws)

        if kpi_name == "MV_VoLTE Traffic":
            overwrite(pivot_fdd, kpi_name, "DR", trend_ws)

        if kpi_name == "4G Data Volume [GB] [CDBH]":
            overwrite(pivot_fdd, kpi_name, "BC", trend_ws)

        if kpi_name == "UL RSSI [CBBH]":
            overwrite(pivot_fdd, kpi_name, "EB", trend_ws)

        if kpi_name == "MV_E-UTRAN Average CQI [CDBH]":
            overwrite(pivot_fdd, kpi_name, "EG", trend_ws)

    # for tdd
    pivot_fdd = perticular_tech(["_T1_", "_T2_"], site_list)[1]
    PsOs_blnk_temp = os.path.join(
        door_path, "template", "DEL KPIs Submission_L2300.xlsx"
    )
    path_of_blnk_temp = PsOs_blnk_temp
    trend_wb_L2300 = load_workbook(path_of_blnk_temp)
    # #print(trend_wb.sheetnames)
    trend_ws = trend_wb_L2300["KPI"]
    for kpi_name in kpi:
        if kpi_name == "MV_RRC Setup Success Rate":
            overwrite(pivot_fdd, kpi_name, "N", trend_ws)

        if kpi_name == "ERAB Setup Success Rate [CDBH]":
            overwrite(pivot_fdd, kpi_name, "T", trend_ws)

        if kpi_name == "PS Drop Call Rate % [CBBH]":
            overwrite(pivot_fdd, kpi_name, "Z", trend_ws)

        if kpi_name == "MV_DL User Throughput_Kbps [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AF", trend_ws)

        if kpi_name == "MV_UL User Throughput_Kbps [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AL", trend_ws)

        if kpi_name == "PS handover success rate [LTE Intra System] [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AQ", trend_ws)

        if kpi_name == "PS handover success rate [LTE Inter System] [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AW", trend_ws)

        if kpi_name == "MV_4G Data Volume_GB":
            overwrite(pivot_fdd, kpi_name, "DW", trend_ws)

        if kpi_name == "MV_CSFB Redirection Success Rate [CDBH]":
            overwrite(pivot_fdd, kpi_name, "BH", trend_ws)

        if kpi_name == "RRC Paging Discard Ratio":
            overwrite(pivot_fdd, kpi_name, "BN", trend_ws)

        if kpi_name == "MV_Average number of used DL PRBs [CDBH]":
            overwrite(pivot_fdd, kpi_name, "BS", trend_ws)

        if kpi_name == "VoLTE Call Setup Success rate [CBBH]":
            overwrite(pivot_fdd, kpi_name, "BX", trend_ws)

        if kpi_name == "VoLTE Drop Call Rate [CBBH]]":
            overwrite(pivot_fdd, kpi_name, "CD", trend_ws)

        if kpi_name == "MV_VoLTE Packet Loss DL [CBBH]":
            overwrite(pivot_fdd, kpi_name, "CJ", trend_ws)

        if kpi_name == "MV_VoLTE Packet Loss UL [CBBH]":
            overwrite(pivot_fdd, kpi_name, "CP", trend_ws)

        if kpi_name == "VoLTE Intra HOSR [CBBH]":
            overwrite(pivot_fdd, kpi_name, "CV", trend_ws)

        if kpi_name == "VoLTE InterF HOSR Exec [CBBH]":
            overwrite(pivot_fdd, kpi_name, "DB", trend_ws)

        if kpi_name == "VoLTE SRVCC SR Exec [CBBH]":
            overwrite(pivot_fdd, kpi_name, "DH", trend_ws)

        if kpi_name == "VoLTE Traffic [CDBH]":
            overwrite(pivot_fdd, kpi_name, "DM", trend_ws)

        if kpi_name == "MV_VoLTE Traffic":
            overwrite(pivot_fdd, kpi_name, "DR", trend_ws)

        if kpi_name == "4G Data Volume [GB] [CDBH]":
            overwrite(pivot_fdd, kpi_name, "BC", trend_ws)

        if kpi_name == "UL RSSI [CBBH]":
            overwrite(pivot_fdd, kpi_name, "EB", trend_ws)

        if kpi_name == "MV_E-UTRAN Average CQI [CDBH]":
            overwrite(pivot_fdd, kpi_name, "EG", trend_ws)

    pivot_fdd = perticular_tech(["_F8_"], site_list)[1]
    PsOs_blnk_temp = os.path.join(
        door_path, "template", "DEL KPIs Submission_L900.xlsx"
    )
    path_of_blnk_temp = PsOs_blnk_temp
    trend_wb_L900 = load_workbook(path_of_blnk_temp)

    trend_ws = trend_wb_L900["KPI"]
    for kpi_name in kpi:
        if kpi_name == "MV_RRC Setup Success Rate":
            overwrite(pivot_fdd, kpi_name, "N", trend_ws)

        if kpi_name == "ERAB Setup Success Rate [CDBH]":
            overwrite(pivot_fdd, kpi_name, "T", trend_ws)

        if kpi_name == "PS Drop Call Rate % [CBBH]":
            overwrite(pivot_fdd, kpi_name, "Z", trend_ws)

        if kpi_name == "MV_DL User Throughput_Kbps [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AF", trend_ws)

        if kpi_name == "MV_UL User Throughput_Kbps [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AL", trend_ws)

        if kpi_name == "PS handover success rate [LTE Intra System] [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AQ", trend_ws)

        if kpi_name == "PS handover success rate [LTE Inter System] [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AW", trend_ws)

        if kpi_name == "MV_4G Data Volume_GB":
            overwrite(pivot_fdd, kpi_name, "DW", trend_ws)

        if kpi_name == "MV_CSFB Redirection Success Rate [CDBH]":
            overwrite(pivot_fdd, kpi_name, "BH", trend_ws)

        if kpi_name == "RRC Paging Discard Ratio":
            overwrite(pivot_fdd, kpi_name, "BN", trend_ws)

        if kpi_name == "MV_Average number of used DL PRBs [CDBH]":
            overwrite(pivot_fdd, kpi_name, "BS", trend_ws)

        if kpi_name == "VoLTE Call Setup Success rate [CBBH]":
            overwrite(pivot_fdd, kpi_name, "BX", trend_ws)

        if kpi_name == "VoLTE Drop Call Rate [CBBH]]":
            overwrite(pivot_fdd, kpi_name, "CD", trend_ws)

        if kpi_name == "MV_VoLTE Packet Loss DL [CBBH]":
            overwrite(pivot_fdd, kpi_name, "CJ", trend_ws)

        if kpi_name == "MV_VoLTE Packet Loss UL [CBBH]":
            overwrite(pivot_fdd, kpi_name, "CP", trend_ws)

        if kpi_name == "VoLTE Intra HOSR [CBBH]":
            overwrite(pivot_fdd, kpi_name, "CV", trend_ws)

        if kpi_name == "VoLTE InterF HOSR Exec [CBBH]":
            overwrite(pivot_fdd, kpi_name, "DB", trend_ws)

        if kpi_name == "VoLTE SRVCC SR Exec [CBBH]":
            overwrite(pivot_fdd, kpi_name, "DH", trend_ws)

        if kpi_name == "VoLTE Traffic [CDBH]":
            overwrite(pivot_fdd, kpi_name, "DM", trend_ws)

        if kpi_name == "MV_VoLTE Traffic":
            overwrite(pivot_fdd, kpi_name, "DR", trend_ws)

        if kpi_name == "4G Data Volume [GB] [CDBH]":
            overwrite(pivot_fdd, kpi_name, "BC", trend_ws)

        if kpi_name == "UL RSSI [CBBH]":
            overwrite(pivot_fdd, kpi_name, "EB", trend_ws)

        if kpi_name == "MV_E-UTRAN Average CQI [CDBH]":
            overwrite(pivot_fdd, kpi_name, "EG", trend_ws)

    pivot_fdd = perticular_tech(["_F1_"], site_list)[1]
    PsOs_blnk_temp = os.path.join(
        door_path, "template", "DEL KPIs Submission_L2100.xlsx"
    )
    path_of_blnk_temp = PsOs_blnk_temp
    trend_wb_L2100 = load_workbook(path_of_blnk_temp)

    trend_ws = trend_wb_L2100["KPI"]
    for kpi_name in kpi:
        if kpi_name == "MV_RRC Setup Success Rate":
            overwrite(pivot_fdd, kpi_name, "N", trend_ws)

        if kpi_name == "ERAB Setup Success Rate [CDBH]":
            overwrite(pivot_fdd, kpi_name, "T", trend_ws)

        if kpi_name == "PS Drop Call Rate % [CBBH]":
            overwrite(pivot_fdd, kpi_name, "Z", trend_ws)

        if kpi_name == "MV_DL User Throughput_Kbps [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AF", trend_ws)

        if kpi_name == "MV_UL User Throughput_Kbps [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AL", trend_ws)

        if kpi_name == "PS handover success rate [LTE Intra System] [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AQ", trend_ws)

        if kpi_name == "PS handover success rate [LTE Inter System] [CDBH]":
            overwrite(pivot_fdd, kpi_name, "AW", trend_ws)

        if kpi_name == "MV_4G Data Volume_GB":
            overwrite(pivot_fdd, kpi_name, "DW", trend_ws)

        if kpi_name == "MV_CSFB Redirection Success Rate [CDBH]":
            overwrite(pivot_fdd, kpi_name, "BH", trend_ws)

        if kpi_name == "RRC Paging Discard Ratio":
            overwrite(pivot_fdd, kpi_name, "BN", trend_ws)

        if kpi_name == "MV_Average number of used DL PRBs [CDBH]":
            overwrite(pivot_fdd, kpi_name, "BS", trend_ws)

        if kpi_name == "VoLTE Call Setup Success rate [CBBH]":
            overwrite(pivot_fdd, kpi_name, "BX", trend_ws)

        if kpi_name == "VoLTE Drop Call Rate [CBBH]]":
            overwrite(pivot_fdd, kpi_name, "CD", trend_ws)

        if kpi_name == "MV_VoLTE Packet Loss DL [CBBH]":
            overwrite(pivot_fdd, kpi_name, "CJ", trend_ws)

        if kpi_name == "MV_VoLTE Packet Loss UL [CBBH]":
            overwrite(pivot_fdd, kpi_name, "CP", trend_ws)

        if kpi_name == "VoLTE Intra HOSR [CBBH]":
            overwrite(pivot_fdd, kpi_name, "CV", trend_ws)

        if kpi_name == "VoLTE InterF HOSR Exec [CBBH]":
            overwrite(pivot_fdd, kpi_name, "DB", trend_ws)

        if kpi_name == "VoLTE SRVCC SR Exec [CBBH]":
            overwrite(pivot_fdd, kpi_name, "DH", trend_ws)

        if kpi_name == "VoLTE Traffic [CDBH]":
            overwrite(pivot_fdd, kpi_name, "DM", trend_ws)

        if kpi_name == "MV_VoLTE Traffic":
            overwrite(pivot_fdd, kpi_name, "DR", trend_ws)

        if kpi_name == "4G Data Volume [GB] [CDBH]":
            overwrite(pivot_fdd, kpi_name, "BC", trend_ws)

        if kpi_name == "UL RSSI [CBBH]":
            overwrite(pivot_fdd, kpi_name, "EB", trend_ws)

        if kpi_name == "MV_E-UTRAN Average CQI [CDBH]":
            overwrite(pivot_fdd, kpi_name, "EG", trend_ws)

    save_wb_L2100 = os.path.join(MEDIA_ROOT,"trends","del","output","toBeZipped","out_DEL KPIs Submission_L2100.xlsx")


    # Ensure the directory exists before using the path
    os.makedirs(os.path.dirname(save_wb_L2100), exist_ok=True)
 
    trend_wb_L2100.save(save_wb_L2100)
    save_wb_L2300 = os.path.join(
        MEDIA_ROOT,
        "trends",
        "del",
        "output",
        "toBeZipped",
        "out_DEL KPIs Submission_L2300.xlsx",
    )
    if not os.path.exists(save_wb_L2300):
        os.makedirs(save_wb_L2300, exist_ok=True)
        
    trend_wb_L2300.save(save_wb_L2300)
    save_wb_L1800 = os.path.join(
        MEDIA_ROOT,
        "trends",
        "del",
        "output",
        "toBeZipped",
        "out_DEL KPIs Submission_L1800.xlsx",
    )
    if not os.path.exists(save_wb_L1800):
        os.makedirs(save_wb_L1800, exist_ok=True)
        
    trend_wb_L1800.save(save_wb_L1800)
    save_wb_L900 = os.path.join(
        MEDIA_ROOT,
        "trends",
        "del",
        "output",
        "toBeZipped",
        "out_DEL KPIs Submission_L900.xlsx",
    )
    if not os.path.exists(save_wb_L900):
        os.makedirs(save_wb_L900, exist_ok=True)
        
    trend_wb_L900.save(save_wb_L900)

    ##############################2G##################

    door_path1 = os.path.join(MEDIA_ROOT, "trends", "del", "del2G")
    if not os.path.exists(door_path1):
        os.makedirs(door_path1, exist_ok=True)

    df_raw_kpi_2G["Short name"] = df_raw_kpi_2G["Short name"].fillna(method=("ffill"))
    df_raw_kpi_2G.columns.values[1] = "DATE"
    a = []
    for cell in df_raw_kpi_2G["Short name"]:
        if '-' in cell:
            Site_ID = str(cell).split('-')[2][:-1]
        else:
            Site_ID = cell[:-1]
        a.append(Site_ID)
    df_raw_kpi_2G.insert(3, "Site_ID", a)
    df_raw_kpi_2G.fillna(value=0, inplace=True)
    df_raw_kpi_2G.rename(columns={"Short name": "Shortname"}, inplace=True)

    message_2G = site_comparision(a, df_2G_site) 

    PsOs_gsm = os.path.join(door_path1, "process_outputs", "fill_2g.xlsx")
    if not os.path.exists(PsOs_gsm):
        os.makedirs(PsOs_gsm, exist_ok=True)

    df_raw_kpi_2G.to_excel(PsOs_gsm, index=False)

    excel_1 = PsOs_gsm


    df1 = pd.read_excel(excel_1)


    print("df1\n\n", df1)
    print("df1\n\n", df1["Site_ID"])

    df1["Site_ID"] = df1["Site_ID"].apply(lambda x: str(x)[1:])
    G2_filter = df1[(df1.Site_ID.isin(df_2G_site))]
    print(G2_filter)


    G2Os_filter = os.path.join(door_path1, "process_outputs", "2Gfilter.xlsx") if os.path.exists(os.path.join(door_path1, "process_outputs", "2Gfilter.xlsx")) else os.makedirs(os.path.join(door_path1, "process_outputs", "2Gfilter.xlsx"),exist_ok=True)
    G2_filter.to_excel(G2Os_filter, index=False)

    df1 = pd.read_excel(G2Os_filter)
    
    print(df1)
    df1.fillna(value=0, inplace=True)
    G2_pivot = df1.pivot_table(
        values=gsm, columns="DATE", index=["Shortname", "Site_ID"]
    )

    G2Os_pivot = os.path.join(door_path1, "process_outputs", "G2_pivot.xlsx") if os.path.exists(os.path.join(door_path1, "process_outputs", "G2_pivot.xlsx")) else os.makedirs(os.path.join(door_path1, "process_outputs", "G2_pivot.xlsx"),exist_ok=True)
    
    G2_pivot.to_excel(G2Os_pivot)

    str_temp = os.path.join(door_path1, "template", "GSM KPI-OLD.xlsx")
    STR = str_temp
    wb = xl.load_workbook(STR)
    ws1 = wb.active

    str_date = request.POST.get("offered_date")
    date1 = datetime.datetime.strptime(str_date, "%Y-%m-%d")
    d1 = date1 - timedelta(1)
    d2 = date1 - timedelta(2)
    d3 = date1 - timedelta(3)
    d4 = date1 - timedelta(4)
    d5 = date1 - timedelta(5)
    cl = [d1, d2, d3, d4, d5]

    alpha = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    def num_hash(num):
        if num < 26:
            return alpha[num - 1]
        else:
            q, r = num // 26, num % 26
            if r == 0:
                if q == 1:
                    return alpha[r - 1]
                else:
                    return num_hash(q - 1) + alpha[r - 1]
            else:
                return num_hash(q) + alpha[r - 1]

    def titleToNumber(s):
        # This process is similar to binary-to-
        # decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord("A") + 1
        return result

    def overwrite(gsm_name, coln1, ws):
        coln2 = num_hash(titleToNumber(coln1) + 1)
        coln3 = num_hash(titleToNumber(coln1) + 2)
        coln4 = num_hash(titleToNumber(coln1) + 3)
        coln5 = num_hash(titleToNumber(coln1) + 4)

        index_GSM = G2_pivot.index
        dr = G2_pivot[gsm_name]
        li = dr.columns.to_list()

        col1 = dr[li[0]].to_list()
        col2 = dr[li[1]].to_list()
        col3 = dr[li[2]].to_list()
        col4 = dr[li[3]].to_list()
        col5 = dr[li[4]].to_list()

        ws[coln1 + "2"].value = cl[4]
        ws[coln2 + "2"].value = cl[3]
        ws[coln3 + "2"].value = cl[2]
        ws[coln4 + "2"].value = cl[1]
        ws[coln5 + "2"].value = cl[0]

        for i, value in enumerate(index_GSM):
            j = i + 3
            ws["A" + str(j)].value = "NT"

            ws["F" + str(j)].value = "DL"

            # ws['E'+str(j)].value=index_GSM[i][2]
            ws["C" + str(j)].value = index_GSM[i][0]
            ws["D" + str(j)].value = index_GSM[i][1]

            ws[coln1 + str(j)].value = col1[i]
            ws[coln2 + str(j)].value = col2[i]
            ws[coln3 + str(j)].value = col3[i]
            ws[coln4 + str(j)].value = col4[i]
            ws[coln5 + str(j)].value = col5[i]

    for gsm_name in gsm:
        if gsm_name == "Total Voice Traffic [BBH]":
            overwrite(gsm_name, "G", ws1)
        if gsm_name == "SDCCH Drop Call Rate [BBH]":
            overwrite(gsm_name, "L", ws1)
        if gsm_name == "SDCCH Drop Call Rate_Nom [BBH]":
            overwrite(gsm_name, "R", ws1)
        if gsm_name == "SDCCH Blocking Rate [BBH]":
            overwrite(gsm_name, "W", ws1)
        if gsm_name == "TCH Blocking Rate [BBH]":
            overwrite(gsm_name, "AC", ws1)
        if gsm_name == "Drop Call Rate [BBH]":
            overwrite(gsm_name, "AI", ws1)
        if gsm_name == "RX Quality [BBH]":
            overwrite(gsm_name, "AO", ws1)
        if gsm_name == "Handover Success Rate [BBH]":
            overwrite(gsm_name, "AU", ws1)
        if gsm_name == "Handover Success Rate_Nom [BBH]":
            overwrite(gsm_name, "BA", ws1)
        if gsm_name == "Handover Success Rate_Denom [BBH]":
            overwrite(gsm_name, "BF", ws1)
        if gsm_name == "TCH Assignment Success Rate [BBH]":
            overwrite(gsm_name, "BK", ws1)
        if gsm_name == "ICM Band4-5 [BBH]":
            overwrite(gsm_name, "BQ", ws1)
        if gsm_name == "Number of Available TCH [BBH]":
            overwrite(gsm_name, "BV", ws1)
        if gsm_name == "TNDROP [BBH]":
            overwrite(gsm_name, "CA", ws1)
    save_outputs = os.path.join(
        MEDIA_ROOT, "trends", "del", "output", "toBeZipped", "2G_Output_Trend.xlsx"
    )
    wb.save(save_outputs)

    def get_all_file_paths(directory):
        file_paths = []

        for root, directories, files in os.walk(directory):
            for filename in files:
                filepath = os.path.join(root, filename)
                file_paths.append(filepath)

        return file_paths

    def main1():
        directory = os.path.join(MEDIA_ROOT, "trends", "del", "output", "toBeZipped")
        file_paths = get_all_file_paths(directory)

        with ZipFile(
            os.path.join(
                MEDIA_ROOT, "trends", "del", "output", "output", "del_output.zip"
            ),
            "w",
        ) as zip:
            for file in file_paths:
                zip.write(file)

    main1()

    download_path = os.path.join(
        MEDIA_URL, "trends", "del", "output", "output", "del_output.zip"
    )
    
    print("The END--------------------------------------------------------------")
    return Response(
        {
            "status": True,
            "message": "successfully",
            "missing_sites_4g": message_4G,
            "missing_sites_2g": message_2G,
            "Download_url": download_path,
        }
    )