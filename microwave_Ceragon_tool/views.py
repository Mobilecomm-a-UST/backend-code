import re
import pandas as pd
import numpy as np
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from microwave_Ceragon_tool.models import *
from microwave_Ceragon_tool.serializers import MicrowaveparaSerializer
from microwave_Ceragon_tool.utils import *
from openpyxl import Workbook
import os
from pathlib import Path
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST
from django.shortcuts import get_object_or_404



main_folder = os.path.join(MEDIA_ROOT, 'MicroWave_Ceragone_Tool')
os.makedirs(main_folder, exist_ok=True)
final_folder=os.path.join(main_folder, 'MW_Final_Output')
os.makedirs(final_folder, exist_ok=True)


@api_view(['POST', 'GET' , 'DELETE'])
def upload_link_budget(request):
    try:
        Link_budget_path = os.path.join(main_folder, 'Link_budget_file')
        os.makedirs(Link_budget_path, exist_ok=True)
 
        if request.method == 'POST':
            files = request.FILES.getlist('link_buget_file')

            if not files:
                return Response(
                    {'error': 'Please upload link_buget_file'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Delete old files before uploading new ones
            if os.path.exists(Link_budget_path):
                for filename in os.listdir(Link_budget_path):
                    file_path = os.path.join(Link_budget_path, filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)

            # Save new files
            uploaded_files = []
            for f in files:
                file_path = os.path.join(Link_budget_path, f.name)
                with open(file_path, 'wb+') as destination:
                    for chunk in f.chunks():
                        destination.write(chunk)
                uploaded_files.append(f.name)

            return Response({
                'status': True,
                'message': 'Old files deleted and new files uploaded successfully.',
                'uploaded_files': uploaded_files
            }, status=status.HTTP_200_OK)
 
        elif request.method == 'GET':
            if not os.path.exists(Link_budget_path):
                return Response({'files': []}, status=status.HTTP_200_OK)
 
            files = os.listdir(Link_budget_path)
            return Response({
                'status': True,
                'message': f'{len(files)} Files found in Link_budget_file folder',
                'files': files,
                }, status=status.HTTP_200_OK)
           
        elif request.method == 'DELETE':
            if not os.path.exists(Link_budget_path):
                return Response({'error': 'Folder does not exist'}, status=status.HTTP_404_NOT_FOUND)
 
            deleted_files = []
            for filename in os.listdir(Link_budget_path):
                file_path = os.path.join(Link_budget_path, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    deleted_files.append(filename)
 
            return Response({
                'status': True,
                'message': f'{len(deleted_files)} Files deleted successfully',
                'deleted_files': deleted_files
            }, status=status.HTTP_200_OK)
           
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

    
@api_view(['POST', 'GET', 'DELETE'])
def upload_traffic_shifting(request):
    try:
        traffic_shifting_path = os.path.join(main_folder, "TRAFFIC_SHIFTING")
        os.makedirs(traffic_shifting_path, exist_ok=True)

        if request.method == "POST":
            files = request.FILES.getlist("ts_file")

            if not files:
                return Response(
                    {"error": "Please upload traffic_shifting_file"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Delete old files before uploading new ones
            for filename in os.listdir(traffic_shifting_path):
                file_path = os.path.join(traffic_shifting_path, filename)

                if os.path.isfile(file_path):
                    os.remove(file_path)

            # Save new files
            uploaded_files = []

            for f in files:
                file_path = os.path.join(traffic_shifting_path, f.name)

                with open(file_path, "wb+") as destination:
                    for chunk in f.chunks():
                        destination.write(chunk)

                uploaded_files.append(f.name)

            return Response(
                {
                    "status": True,
                    "message": "Old files deleted and new Traffic Shifting files uploaded successfully.",
                    "uploaded_files": uploaded_files,
                },
                status=status.HTTP_200_OK,
            )
        elif request.method == "GET":
            files = os.listdir(traffic_shifting_path)

            return Response(
                {
                    "status": True,
                    "message": f"{len(files)} Traffic Shifting file(s) found.",
                    "files": files,
                },
                status=status.HTTP_200_OK,
            )

        elif request.method == "DELETE":
            deleted_files = []

            for filename in os.listdir(traffic_shifting_path):
                file_path = os.path.join(traffic_shifting_path, filename)

                if os.path.isfile(file_path):
                    os.remove(file_path)
                    deleted_files.append(filename)

            return Response(
                {
                    "status": True,
                    "message": f"{len(deleted_files)} Traffic Shifting file(s) deleted successfully.",
                    "deleted_files": deleted_files,
                },
                status=status.HTTP_200_OK,
            )

    except Exception as e:
        return Response(
            {"status": False, "error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET", "POST", "PUT", "DELETE"])
def upload_server_ip(request, id=None):
    if request.method == "GET":

        if id:
            try:
                data = CircleServerIP.objects.values("id", "circle", "ip").get(id=id)
                return Response(data)

            except CircleServerIP.DoesNotExist:
                return Response(
                    {"error": "Record not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

        data = list(
            CircleServerIP.objects.all().values(
                "id",
                "circle",
                "ip"
            )
        )

        return Response(data)

    # ================= POST =================
    elif request.method == "POST":
        file = request.FILES.get("file")
        # ================= Excel Upload =================
        if file:

            try:
                df = pd.read_excel(file).fillna("")
            except Exception as e:
                return Response(
                    {"error": str(e)},
                    status=status.HTTP_400_BAD_REQUEST
                )

            inserted = 0
            skipped = 0

            for _, row in df.iterrows():

                circle = str(row.get("Circle", "")).strip().upper()
                ip = str(row.get("IP", "")).strip()

                if not circle or not ip:
                    continue

                if CircleServerIP.objects.filter(circle=circle, ip=ip).exists():
                    skipped += 1
                    continue

                CircleServerIP.objects.create(
                    circle=circle,
                    ip=ip
                )

                inserted += 1

            return Response({
                "message": "Upload completed successfully.",
                "inserted": inserted,
                "skipped": skipped
            })

        # ================= Single Record =================

        circle = request.data.get("circle")
        ip = request.data.get("ip")

        if not circle or not ip:
            return Response(
                {
                    "error": "Please provide circle and ip or upload Excel file."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        circle = circle.strip().upper()
        ip = ip.strip()

        if CircleServerIP.objects.filter(circle=circle, ip=ip).exists():
            return Response(
                {
                    "error": "Record already exists."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        obj = CircleServerIP.objects.create(
            circle=circle,
            ip=ip
        )

        return Response(
            {
                "message": "Added Successfully",
                "id": obj.id,
                "circle": obj.circle,
                "ip": obj.ip,
            },
            status=status.HTTP_201_CREATED
        )


    elif request.method == "PUT":

        if not id:
            return Response(
                {"error": "Please provide id in URL"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            obj = CircleServerIP.objects.get(id=id)
        except CircleServerIP.DoesNotExist:
            return Response(
                {"error": "Record not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        circle = request.data.get("circle")
        ip = request.data.get("ip")

        if circle is not None:
            obj.circle = circle.strip().upper()

        if ip is not None:
            obj.ip = ip.strip()

        obj.save()

        return Response({
            "message": "Updated Successfully"
        })

    elif request.method == "DELETE":

        if not id:
            return Response(
                {"error": "Please provide id in URL"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            obj = CircleServerIP.objects.get(id=id)
        except CircleServerIP.DoesNotExist:
            return Response(
                {"error": "Record not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        obj.delete()

        return Response({
            "message": "Deleted Successfully"
        })
        
@api_view(['GET', 'POST', 'PUT', 'DELETE'])
def microwave_para(request, pk=None):
    if request.method == "GET":
        idu = request.data.get('idu', '')
        queryset = Microwavepara.objects.all()
        if idu:
            queryset = queryset.filter(idu_model__iexact=idu)
        serializer = MicrowaveparaSerializer(queryset, many=True)

        return Response(serializer.data)

    elif request.method == "POST":
        data = request.data
          # Bulk upload
        if isinstance(data, list):
            serializer = MicrowaveparaSerializer(data=data, many=True)
        else:
            serializer = MicrowaveparaSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "status": True,
                    "data": serializer.data
                },
                status=status.HTTP_201_CREATED
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    elif request.method == "PUT":

        obj = get_object_or_404(Microwavepara, pk=pk)

        serializer = MicrowaveparaSerializer(
            obj,
            data=request.data,
            partial=True
        )

        if serializer.is_valid():
            serializer.save()

            return Response(
                {
                    "status": True,
                    "data": serializer.data
                }
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    elif request.method == "DELETE":

        obj = get_object_or_404(Microwavepara, pk=pk)
        obj.delete()

        return Response(
            {
                "status": True,
                "message": "Deleted successfully"
            }
        )

@api_view(["POST"])
def get_server_ip(request):

    plan_id = request.data.get("plan_id")

    if not plan_id:
        return Response({"message": "plan_id required"}, status=400)

    if isinstance(plan_id, str):
        plan_ids = [x.strip() for x in plan_id.split(",") if x.strip()]
    else:
        plan_ids = plan_id

    result = []

    for pid in plan_ids:

        try:
            circle = pid.split("-")[2].strip().upper()
        except IndexError:
            continue

        ips = list(
            CircleServerIP.objects.filter(
                circle=circle,
                is_active=True
            )
            .values_list("ip", flat=True)
            .distinct()
        )

        result.append({
            "plan_id": pid,
            "circle": circle,
            "ip_list": ips
        })

    return Response(result)



@api_view(['POST'])
def search_plan_id(request):
    search_text = request.data.get('plan_id', '').strip()
    if not search_text:
        return Response([])

    link_budget_folder = os.path.join(main_folder, "Link_budget_file")
    files = os.listdir(link_budget_folder)
    if not files:
        return Response([])

    file_path = os.path.join(link_budget_folder, files[0])
    if file_path.endswith(".xlsx"):
        df = pd.read_excel(file_path)
    else:
        df = pd.read_csv(file_path)
    df.columns = df.columns.str.strip()

    if "Plan Id" not in df.columns:
        return Response([])
    
    result = (
        df["Plan Id"].astype(str).str.strip().loc[lambda x: x.str.contains(search_text, case=False, na=False)]
        .drop_duplicates()
        .head(15)
        .tolist()
    )

    return Response(result)


import json

@api_view(['POST'])
def upload_cergon_dump(request):
    plan_id_str = request.data.get("plan_id")
    if not plan_id_str:
        return Response({"message": "plan_id not provided"}, status=400)

    plan_ids = [x.strip() for x in plan_id_str.split(",") if x.strip()]
    print(plan_ids)

  
    selected_ips = request.data.get("selected_ips", "{}")
    if isinstance(selected_ips, str):
        selected_ips = json.loads(selected_ips)
        print(selected_ips)

        print("Keys =", list(selected_ips.keys()))
     

    dump_files1 =  request.FILES.getlist("dump_file1")
    if not dump_files1:
        return Response({"message": "dump_file1 not provided"}, status=400)
    
    dump_files2 = request.FILES.getlist("dump_file2")
    if not dump_files2:
        return Response({"message": "dump_file2 not provided"}, status=400)

    link_budget_file = os.path.join(main_folder, 'Link_budget_file')
    if not os.path.exists(link_budget_file):
        return Response({"error": "Link_budget_file folder not found"}, status=400)

    link_budget_file_list = []
    for filename in os.listdir(link_budget_file):
        file_path = os.path.join(link_budget_file, filename)
        if filename.endswith('.xlsx'):
            df = pd.read_excel(file_path)
        elif filename.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            return Response(
                {"status": "ERROR", "message": f"Unsupported file type: {filename}"},
                status=HTTP_400_BAD_REQUEST
            )
        df.columns = df.columns.str.strip()
        link_budget_file_list.append(df)

    if link_budget_file_list:
        link_budget_df = pd.concat(link_budget_file_list, ignore_index=True)
    else:
        return Response(
            {"status": "ERROR", "message": "No valid files found"},
            status=HTTP_400_BAD_REQUEST
        )
    # print("----------------")
    # print(link_budget_df.columns.tolist())
    traffic_shifting_folder = os.path.join(main_folder, "TRAFFIC_SHIFTING")
    if not os.path.exists(traffic_shifting_folder):
        return Response({"error": "TRAFFIC_SHIFTING folder not found"},
            status=400
        )

    traffic_shifting_file_list = []
    cols = [
        "Site ID",
        "DCN Site A(HOP 1)",
        "DCN Site B(HOP 1)"
    ]

    for filename in os.listdir(traffic_shifting_folder):
        file_path = os.path.join(traffic_shifting_folder, filename)
        if filename.endswith(".xlsx"):
            df = pd.read_excel(
                file_path,
                engine="calamine",
                header=1,
                usecols=cols,
                dtype=str
            )

        elif filename.endswith(".csv"):
            df = pd.read_csv(
                file_path,
                header=1,
                usecols=cols,
                dtype=str
            )

        else:
            continue

        df.columns = df.columns.str.strip()
        traffic_shifting_file_list.append(df)

    if traffic_shifting_file_list:
        traffic_shifting_df = pd.concat(
            traffic_shifting_file_list,
            ignore_index=True
        )
    else:
        return Response({
            "status": False,
            "message": "No valid Traffic Shifting files found"
        }, status=400)

    # print(traffic_shifting_df.head())

    traffic_shifting_df["Site ID"] = (
        traffic_shifting_df["Site ID"]
        .str.strip()
        .str.upper()
    )

    traffic_shifting_df = traffic_shifting_df.drop_duplicates(
        subset="Site ID",
        keep="first"
    )

    ts_lookup = traffic_shifting_df.set_index("Site ID")[
        ["DCN Site A(HOP 1)", "DCN Site B(HOP 1)"]
    ].to_dict("index")

    import time
    start = time.time()
    print("Read Time:", time.time() - start)


#read dump file1 -----------
    dump_groups = [
        ("Report1", dump_files1),
        ("Report2", dump_files2)
    ]
    report_sheets = {}
    sl_no_map = {}
    sl_counter = 1
    Atcp_df = pd.DataFrame(columns=[
        "SL NO",
        "MW Plan Id",
        "Link ID (Site ID A - Site ID B and Site ID B - Site ID A)",
        "Site ID",
        "Site Name/System name/Unit Name",
        "Server/RDP -IP",
        "Type of Equipment",
        "ODU/IDU (New/Existing)",
        "MRMC Script ID/Bandw",
        "MRMC Profile/Bandw",
        "Link configuration ( 1+0 / 1+1 / 2+0/ XPIC/1G)",
        "IDU Model (Modular IDU-IP20N, IP-20G , High capacity packet radio IP20C/IP20S, IP-20GX, IP-20F)",
        "Link-Slot Number",
        "Link-Port Number",
        "Frequency Tx",
        "Frequency Rx",
        "Tx Power  (dBm)",
        "RSL(dBM) Main [<=3 dB Deviation permitted]",
        "Modulation",
        "ATPC (enabled/disable)",
        "ACM Enabled/Modulation mode (Fixed/adaptive)",
        "ODU IP Address",
        "IDU IP Address/Remarks",
        "ODU IP Ethernet slot port detail  /Other remarks",
        "GNOC Link ID (Site ID A - Site ID B and Site ID B - Site ID A)",
        "GNOC Site Name/System name/Unit Name",
        "Type of Equipment/IDU Model (Modular IDU-IP20N, IP-20G, High capacity packet radio IP20C, IP-20GX)",
        "HOP visible in NMS (Y/N) [NMS visibility is mandatory]",
        "Software Version IDU",
        "GNOC Frequency Tx",
        "GNOC Frequency Rx",
        "GNOC Tx Power (dBm)",
        "GNOC Current RSL (+/-3)",
        "GNOC MRMC Script ID/Bandw",
        "GNOC MRMC Profile/Bandw",
        "GNOC Modulation",
        "GNOC ACM Enabled/Modulation mode (Fixed/adaptive)",
        "GNOC ATPC (enabled/disable)",
        "GNOC High Low Violation (Y/N) [No violation accepted]",
        "GNOC QoS configured (Y/N) (As per planning guideline)",
        "GNOC Performance error in last 24 hour end",
        "GNOC Date and time settings correct (Y/N)",
        "GNOC MSTP Disable define status",
        "GNOC Undervoltage clear threshold =48",
        "GNOC Undervoltage raise threshold =46",
        "GNOC Critical alarm at New Node",
        "GNOC Ethernet Port speed status (Should be 1000)",
        "GNOC Final Remarks",
        "GNOC Final AT status (Accepted/Rejected)",
        "Done By",
        "Date"
    ])

    for sheet_name, files in dump_groups:
        final_report_df = pd.DataFrame()
        for idx, dump_file in enumerate(files):

            current_plan_id = ( plan_ids[idx]if idx < len(plan_ids)else "")
            lines = dump_file.read().decode("utf-8").splitlines()
            result = []
            # ---- Simple parameters ----
            patterns = {
                "Software Version IDU": r'-sw version\s*:\s*(.+)',
                "Site Name/System name/Unit Name": r'unit-name=(.*)',
             
                "SNMP Version": r'snmp-s-version=(.*)',
                "License Key": r'license-demo-admin=(.*)',
                "Network IP": r'-ip address\s*:\s*(.+)',
                "IDU Model": r'-system type\s*:\s*(.+)',
                "Type of Equipment": r'-system id\s*:\s*(.+)',
                "MTU Value": r'sw-ap-l2-if-global-mru=(.*)',
                "Tx Mute": r'rf-config-table\.0\.mute-tx=(.*)',
                "Server Location": r'unit-location=(.*)',
                "NodeLONG": r'unit-longitude=(.*)',
                "NodeLAT" :r'unit-latitude=(.*)',
                "Multicarrier ABC": r'.*workaround_manager_config\.0\.interface\(1\)\s*=\s*(.*)',
                "HTTPS": r'http-s-admin=(.*)',
                "SNMP": r'snmp-s-admin=(.*)',
                "Password": r'security-config-log-upload-configuration-table\.0\.Password=(.*)',
                "MRMC Script ID/Bandw": r'mrmc-script-config-table\.0\.mrmc-config-active-script-id=(.*)',
                "MRMC Profile/Bandw": r'mrmc-script-config-table\.0\.mrmc-config-max-profile=(.*)',
                "Ethernet Port Speed": r'radio-ethernet-config-table\.\d+\.threshold-capacity=(.*)',
                "Operational Mode": r'mrmc-script-config-table\.\d+\.mrmc-config-script-operational-mode=(.*)',
                "Frequency Tx (MHz)": r'rf-config-table\.\d+\.tx-frequency=(.*)',
                "Frequency Rx (MHz)": r'rf-config-table\.\d+\.rx-frequency=(.*)',
                "Tx Power (dBm)": r'rf-config-table\.\d+\.tx-level-config=(.*)',
                
            }
            utc_hours = "NA"
            utc_minutes = "NA"

            for line in lines:
                if line.startswith("time-services-config-table.0.time-services-utc-offset-hours="):
                    utc_hours = line.split("=", 1)[1].strip()

                elif line.startswith("time-services-config-table.0.time-services-utc-offset-minutes="):
                    utc_minutes = line.split("=", 1)[1].strip()

            datetime_value = f"offset-hours:{utc_hours}-offset-minutes:{utc_minutes}"
            result.append({
                "parameter": "Datetime",
                "value in dump": datetime_value
            })

            required_alarm_ids = ["11", "15", "915"]
            alarm_admin_status = {}

            current_alarm_id = None

            for line in lines:
                line = line.strip()

                m = re.match(r"alarm-services-configuration-table\.\d+\.alarm-id=(.+)", line)
                if m:
                    current_alarm_id = m.group(1).strip()
                    continue

                m = re.match(r"alarm-services-configuration-table\.\d+\.alarm-admin=(.+)", line)
                if m and current_alarm_id:
                    alarm_admin_status[current_alarm_id] = m.group(1).strip().capitalize()
                    current_alarm_id = None

            alarm_value = ", ".join(
                f"ID-{alarm_id}: {alarm_admin_status.get(alarm_id, 'not found')}"
                for alarm_id in required_alarm_ids
            )

            result.append({
                "parameter": "Alarm Not required",
                "value in dump": alarm_value
            })

            result.append({
                "parameter": "Datetime",
                "value in dump": datetime_value
            })

            snmp_values = {
                "username": "",
                "security_mode": "",
                "auth_algorithm": "",
                "access_mode": ""
            }

            for line in lines:
                line = line.strip()

                if line.startswith("snmp-v3-authentication-table.0.v3-username="):
                    snmp_values["username"] = line.split("=", 1)[1].strip()

                elif line.startswith("snmp-v3-authentication-table.0.v3-security-mode="):
                    snmp_values["security_mode"] = line.split("=", 1)[1].strip()

                elif line.startswith("snmp-v3-authentication-table.0.v3-auth-algorithm="):
                    snmp_values["auth_algorithm"] = line.split("=", 1)[1].strip()

                elif line.startswith("snmp-v3-authentication-table.0.v3-access-mode="):
                    snmp_values["access_mode"] = line.split("=", 1)[1].strip()

            snmp_config = (
                f"username:{snmp_values['username']}, "
                f"security_mode:{snmp_values['security_mode']}, "
                f"auth_algorithm:{snmp_values['auth_algorithm']}, "
                f"access_mode:{snmp_values['access_mode']}"
            )

            result.append({
                "parameter": "SNMP Configuration",
                "value in dump": snmp_config
            })

            utilization = {
                "threshold1": "",
                "threshold2": "",
                "threshold3": ""
            }

            for line in lines:
                line = line.strip()

                if line.startswith("radio-ethernet-config-table.1.threshold-utilization="):
                    utilization["threshold1"] = line.split("=", 1)[1].strip()

                elif line.startswith("radio-ethernet-config-table.1.threshold2-utilization="):
                    utilization["threshold2"] = line.split("=", 1)[1].strip()

                elif line.startswith("radio-ethernet-config-table.1.threshold3-utilization="):
                    utilization["threshold3"] = line.split("=", 1)[1].strip()

            utilization_config = ", ".join(
                                [f"threshold{i+1}:{utilization[f'threshold{i+1}']}" for i in range(3)]
                            )

            result.append({
                "parameter": "Utilization Configuration",
                "value in dump": utilization_config
            })

            for display_name, pattern in patterns.items():
                value = "NA"        # default value
                for line in lines:
                    match = re.search(pattern, line, re.IGNORECASE)
                    if match:
                        value = match.group(1).strip()

                        if value == "":
                            value = "NA"

                        break

                result.append({
                    "parameter": display_name,
                    "value in dump": value
                })

            rsl_value = "NA"
            header = None

            for i, line in enumerate(lines):
                if line.strip() == "rf-config-table":

                    # Header
                    header = lines[i + 1].strip().split("|")

                    # First data row
                    data = lines[i + 2].strip().split("|")

                    if "Reference RX Level (dBm)" in header:
                        idx = header.index("Reference RX Level (dBm)")
                        rsl_value = data[idx]

                    break

            result.append({
                "parameter": "Current RSL",
                "value in dump": rsl_value
            })

            amcc_group = "NA"

            for line in lines:
                line = line.strip()

                if "amcc-group-admin=" in line:          # IP-50
                    amcc_group = line.split("=", 1)[1].strip()
                    break

                elif "xpic_admin=" in line:              # IP-20
                    amcc_group = line.split("=", 1)[1].strip()
                    break

            result.append({
                "parameter": "AMCC Group",
                "value in dump": amcc_group
            })

            result.append({"parameter": "ACM Enabled/Modulation mode", "value in dump": "adaptive"})
            result.append({"parameter": "ATPC", "value in dump": "disable"})

            # ---- Voltage table parameters ----
            header = None
            values = None
            for i, line in enumerate(lines):
                if line.strip() == "unit-mgr-voltage-config-table":
                    for j in range(i + 1, len(lines)):
                        if lines[j].startswith("row|"):
                            header = [x.strip() for x in lines[j].split("|")]
                        elif lines[j].startswith("0|"):
                            values = [x.strip() for x in lines[j].split("|")]
                            break
                    break

            if header and values:
                required_columns = [
                    "Undervoltage clear threshold (V)",
                    "Undervoltage raise threshold (V)",
                ]
                for col in required_columns:
                    if col in header:
                        result.append({"parameter": col, "value in dump": values[header.index(col)]})

            # ---- Date and Time ----
            creation_date = ""
            timer = ""
            for line in lines:
                if line.startswith("unit-info-creation-date="):
                    creation_date = line.split("=", 1)[1].strip()
                elif line.startswith("software-mgt-s-timer="):
                    timer = line.split("=", 1)[1].strip()


            dscp_cos_list = []
            inside_table = False

            for line in lines:
                line = line.strip()
                if line == "sw-ap-global-ingress-mapping-dscp-to-cos-table":
                    inside_table = True
                    continue
                if inside_table:
                    # next section start ho gaya
                    if line.startswith("%%%"):
                        break
                    # data rows
                    if re.match(r'^\d+\|', line):
                        cols = line.split("|")
                        dscp = cols[1]
                        cos = cols[4]
                        dscp_cos_list.append({
                            "DSCP": dscp,
                            "CoS": cos
                        })

            # print(dscp_cos_list)  

            expected_set = {(x["DSCP"], x["CoS"]) for x in dscp_cos_mapping}
            actual_set = {(x["DSCP"], x["CoS"]) for x in dscp_cos_list}

            qos_status = "all cos value match" if expected_set == actual_set else "NOT match cos"       

            result.append({"parameter": "Date and Time", "value in dump": f"{creation_date} {timer}".strip()})
            result.append({
                    "parameter": "QoS configured (Y/N)",
                    "value in dump": qos_status
                })

            # convert dump result list -> dict for easy lookup
            dump_values = {item["parameter"]: item["value in dump"] for item in result}
            # print(dump_values)

            dump_values["Type of Equipment/IDU Model"] = (
                str(dump_values.get("Type of Equipment", "")).strip()
                + " | "
                + str(dump_values.get("IDU Model", "")).strip()
            )

            idu_model = (
                dump_values.get("Type of Equipment/IDU Model", "")
                .split("|")[-1]
                .strip()
                .upper()
            )

            queryset = Microwavepara.objects.filter(
                idu_model__iexact=idu_model
            ).values("parameter", "value")

            db_values = {
                row["parameter"].strip().lower(): str(row["value"]).strip()
                for row in queryset
            }

        
            site_id = dump_values.get("Site Name/System name/Unit Name", "").strip()

            mask_a = link_budget_df["Site ID-A"].astype(str).str.strip() == site_id
            mask_b = link_budget_df["Site ID -B"].astype(str).str.strip() == site_id

            row = None
            mapping = {}

            common_mapping = {
                "Link ID (Site ID A - Site ID B and Site ID B - Site ID A)":"Link ID",
                "Tx Power (dBm)": "BER10e6 Tx Power (dBm)",
                "MRMC Script ID/Bandw": "MRMC Script ID",
                "MRMC Profile/Bandw": "Tx Radio",
                "Operational Mode": "ACM Max QAM",
                "ACM Enabled/Modulation mode": "ACM Status",
                "ATPC": "ATPC Status",
                "Current RSL": "RSL",
                "Plan Id":"Plan Id",
                "Hop Type":"Hop Type",
                "Fiber POP Id":"Fiber POP Id"
            
                
            }

            site_a_mapping = {
                "Site Name/System name/Unit Name": "Site ID-A",
                "NodeLAT": "Site A Lat",
                "NodeLONG": "Site A Long",
                "Type of Equipment/IDU Model": "(MODEM(IF Card)/Ethernet Port) Site-A",
                "Frequency Tx (MHz)": "Tx Frequency (MHz)",
                "Frequency Rx (MHz)": "Rx Frequency (MHz)",

            }

            site_b_mapping = {
                "Site Name/System name/Unit Name": "Site ID -B",
                "NodeLAT": "Site B Lat",
                "NodeLONG": "Site B Long",
                "Type of Equipment/IDU Model": "(MODEM(IF Card)/Ethernet Port) Site-B",
                "Frequency Tx (MHz)": "Rx Frequency (MHz)",
                "Frequency Rx (MHz)": "Tx Frequency (MHz)"
            }

            if mask_a.any():
                row = link_budget_df.loc[mask_a].iloc[0]
                mapping = {**common_mapping, **site_a_mapping}

            elif mask_b.any():
                row = link_budget_df.loc[mask_b].iloc[0]
                mapping = {**common_mapping, **site_b_mapping}

            # ---------------- LB Values ----------------
            lb_values = {}

            if row is not None:
                for k, v in mapping.items():
                    if v in row.index:
                        lb_values[k] = row[v]

            lb_values["band"] = "D band" if "IP-50" in str(lb_values.get("Type of Equipment/IDU Model", "")).upper() else "C band"

            # clean NA
            for k in lb_values:
                if pd.isna(lb_values[k]) or str(lb_values[k]).strip() == "":
                    lb_values[k] = "NA"
                else:
                    lb_values[k] = str(lb_values[k]).strip()

            # print(lb_values)
            report = []
            plan_status = (
                "OK"
                if str(current_plan_id).strip() == str(row.get("Plan Id", "")).strip()
                else "NOT OK"
            )

            # report.append([
            #     "Plan-ID",
            #     current_plan_id,
            #     row.get("Plan Id", "NA"),
            #     plan_status
            # ])

            for param in master_parameters:
                param_key = param.strip().lower()
                if param == "Link ID (Site ID A - Site ID B and Site ID B - Site ID A)":
                    dump_value = lb_values.get(
                        "Link ID (Site ID A - Site ID B and Site ID B - Site ID A)",
                        "NA"
                    )

                    expected_value = dump_value

                    status = "OK" if dump_value not in ["NA", ""] else "NOT OK"

                else:
                    dump_value = dump_values.get(param, "NA")
                    if param == "DCN Synch Status":
                        dump_ip = str(dump_values.get("Network IP", "NA")).strip()

                        ts_row = ts_lookup.get(site_id)
                        # print(ts_row)

                        if ts_row:
                            site_a_ip = str(ts_row.get("DCN Site A(HOP 1)", "")).strip()
                            site_b_ip = str(ts_row.get("DCN Site B(HOP 1)", "")).strip()

                            dump_value = dump_ip
                            expected_value = f"A:{site_a_ip} / B:{site_b_ip}"

                            if dump_ip == site_a_ip:
                                status = "OK"

                            elif dump_ip == site_b_ip:
                                status = "OK"

                            elif site_a_ip == site_b_ip:
                                status = "PARTIALLY OK"

                            else:
                                status = "NOT OK"

                        else:
                            dump_value = dump_ip
                            expected_value = "NOT FOUND"
                            status = "NOT OK"

                        report.append([
                            param,
                            dump_value,
                            expected_value,
                            status
                        ])

                        continue
# ====================================================

                    if param == "High Low Violation (Y/N)":

                        dump_value = dump_values.get("Frequency Tx (MHz)", "NA")
                        expected_value = lb_values.get("Frequency Tx (MHz)", "NA")

                        status = check_status(
                            "Frequency Tx (MHz)",dump_value,expected_value,dump_values)

                        report.append([param, dump_value, expected_value, status])
                        continue

                    # ---------------- LB Parameters ----------------
                    if param in check_lb_params:

                        if param == "Current RSL":
                            idu = str(dump_values.get("IDU Model", "")).upper()
                            # expected_value = "+/-3" if "IP-50" in idu else "+/-3"
                            expected_value="+/-3"


                        elif param == "AMCC Group":
                            # print("INSIDE AMCC")
                            # print("Hop =", lb_values.get("Hop Type"))

                            hop_type = str(lb_values.get("Hop Type", "")).upper()

                            if "XPIC" in hop_type:
                                expected_value = "Enable"
                            elif "1+0" in hop_type:
                                expected_value = "Disable"
                            else:
                                expected_value = "NA"
    
                        else:
                            expected_value = lb_values.get(param, "NA")

                                 

                    elif param_key in db_values:
                        expected_value = db_values[param_key]

                    elif param in fix_values:
                        expected_value = fix_values[param]

                    elif param in check_in_manual:
                        expected_value = "Manual"

                    else:
                        expected_value = "NA"

                    status = check_status( param,dump_value,expected_value,dump_values)

                report.append([
                        param,
                        dump_value,
                        expected_value,
                        status
                    ])

            remove_params = [
                "MSTP",
                "Server Location",
                "NodeLONG",
                "NodeLAT",
                "Management Service",
                "Password",
                "Multicarrier ABC",
                "Date and Time",
            ]

            # Create report dataframe only once
            if final_report_df.empty:

                filtered_parameters = [
                    p for p in master_parameters
                    if p not in remove_params
                ]

                final_report_df = pd.DataFrame({
                    "Parameter": ["Plan-ID"] + filtered_parameters
                })

            site = dump_values.get("Site Name/System name/Unit Name", "NA")
            fiber_pop=row.get("Fiber Pop", "NA")

            idu_model_text = str(dump_values.get("Type of Equipment/IDU Model", "")).upper()

            if "IP-50" in idu_model_text:
                slot_number = 1
            elif "IP-20" in idu_model_text:
                slot_number = 2
            else:
                slot_number = "NA"

            hop_type = str(row.get("Hop Type", "")).upper()
            if "XPIC" in hop_type:
                port_number = "1,2"
            elif "1+0" in hop_type:
                port_number = "1"
            else:
                port_number = "NA"

            dump_col = [current_plan_id]
            expected_col = [row.get("Plan Id", "NA")]
            remark_col = [plan_status]

            for param, dump, expected, remark in report:

                if param in remove_params:
                    continue

                dump_col.append(dump)
                expected_col.append(expected)
                remark_col.append(remark)

            final_report_df[f"Dump Value-{site}"] = dump_col
            final_report_df[f"Expected Value-{site}"] = expected_col
            final_report_df[f"Remark-{site}"] = remark_col

         
    
            plan_key = str(row.get("Plan Id", "")).strip()
            if plan_key not in sl_no_map:
                sl_no_map[plan_key] = sl_counter
                sl_counter += 1

            sl_no = sl_no_map[plan_key]

            site_a = str(row.get("Site ID-A", "")).strip()
            site_b = str(row.get("Site ID -B", "")).strip()

            current_site = dump_values.get("Site Name/System name/Unit Name", "").strip()

            if current_site == site_a:
                link_id = f"{site_a}-{site_b}"
            elif current_site == site_b:
                link_id = f"{site_b}-{site_a}"
            else:
                link_id = row.get("Link ID", "NA")

            row_data = {
                "SL NO":sl_no,
                "MW Plan Id": current_plan_id,
                "Link ID (Site ID A - Site ID B and Site ID B - Site ID A)": link_id,
                "Site ID":dump_values.get("Site Name/System name/Unit Name", "NA"),
                "Site Name/System name/Unit Name":dump_values.get("Site Name/System name/Unit Name", "NA"),
                "Server/RDP -IP":"",
                "Type of Equipment":"Ceragone",
                "ODU/IDU (New/Existing)":"New",
                "MRMC Script ID/Bandw":dump_values.get("MRMC Script ID/Bandw", "NA"),
                "MRMC Profile/Bandw":dump_values.get("MRMC Profile/Bandw", "NA"),
                "Link configuration ( 1+0 / 1+1 / 2+0/ XPIC/1G)":row.get("Hop Type","NA"),
                "IDU Model (Modular IDU-IP20N, IP-20G , High capacity packet radio IP20C/IP20S, IP-20GX, IP-20F)":dump_values.get("Type of Equipment/IDU Model", "NA"),
                "Link-Slot Number":slot_number,
                "Link-Port Number":port_number,
                "Frequency Tx":dump_values.get("Frequency Tx (MHz)", "NA"),
                "Frequency Rx":dump_values.get("Frequency Rx (MHz)", "NA"),
                "Tx Power  (dBm)":row.get("BER10e6 Tx Power (dBm)", "NA"),
                "RSL(dBM) Main [<=3 dB Deviation permitted]":dump_values.get("Current RSL", "NA"),
                "Modulation":row.get("ACM Max QAM", "NA"),
                "ATPC (enabled/disable)":row.get("ATPC Status", "NA"),
                "ACM Enabled/Modulation mode (Fixed/adaptive)":row.get("ACM Status", "NA"),
                "ODU IP Address":dump_values.get("Network IP", "NA"),

                "IDU IP Address/Remarks": (
                    "Connected to Ciena"
                    if site == fiber_pop
                    else "Connected to BTS"
                ),
                "ODU IP Ethernet slot port detail  /Other remarks":"ethernet interfaces eth slot 1 port 1",
                "GNOC Link ID (Site ID A - Site ID B and Site ID B - Site ID A)": "",
                "GNOC Site Name/System name/Unit Name": "",
                "Type of Equipment/IDU Model (Modular IDU-IP20N, IP-20G, High capacity packet radio IP20C, IP-20GX)": "",
                "HOP visible in NMS (Y/N) [NMS visibility is mandatory]": "",
                "Software Version IDU": "",
                "GNOC Frequency Tx": "",
                "GNOC Frequency Rx": "",
                "GNOC Tx Power (dBm)": "",
                "GNOC Current RSL (+/-3)": "",
                "GNOC MRMC Script ID/Bandw": "",
                "GNOC MRMC Profile/Bandw": "",
                "GNOC Modulation": "",
                "GNOC ACM Enabled/Modulation mode (Fixed/adaptive)": "",
                "GNOC ATPC (enabled/disable)": "",
                "GNOC High Low Violation (Y/N) [No violation accepted]": "",
                "GNOC QoS configured (Y/N) (As per planning guideline)": "",
                "GNOC Performance error in last 24 hour end": "",
                "GNOC Date and time settings correct (Y/N)": "",
                "GNOC MSTP Disable define status": "",
                "GNOC Undervoltage clear threshold =48": "",
                "GNOC Undervoltage raise threshold =46": "",
                "GNOC Critical alarm at New Node": "",
                "GNOC Ethernet Port speed status (Should be 1000)": "",
                "GNOC Final Remarks": "",
                "GNOC Final AT status (Accepted/Rejected)": "",
                "Done By": "",
                "Date": "",
                "Fiber POP Id":row.get("Plan Id", "NA"),

            }
     
            Atcp_df = pd.concat(
                [Atcp_df, pd.DataFrame([row_data])],
                ignore_index=True
            )

           
                                

        report_sheets[sheet_name] = final_report_df
        Atcp_df["MW Plan Id"] = Atcp_df["MW Plan Id"].astype(str).str.strip()

        selected_ips = {
            str(k).strip(): str(v).strip()
            for k, v in selected_ips.items()
        }

        Atcp_df["Server/RDP -IP"] = (
            Atcp_df["MW Plan Id"]
            .map(selected_ips)
            .fillna("")
        )
        print("MW Plan IDs =", Atcp_df["MW Plan Id"].tolist())


    wb = Workbook()
    wb.remove(wb.active)

    # Report1, Report2 create karna
    for sheet_name, df in report_sheets.items():
        ws = wb.create_sheet(sheet_name)
        data = [df.columns.tolist()] + df.values.tolist()
        create_config_report(ws, data)

    def link_sort_key(link):
        if "-" in str(link):
            a, b = str(link).split("-", 1)
            return "-".join(sorted([a.strip(), b.strip()]))
        return str(link)

    # Same link ko ek group banana
    Atcp_df["SortKey"] = Atcp_df[
        "Link ID (Site ID A - Site ID B and Site ID B - Site ID A)"
    ].apply(link_sort_key)

    # Link ID ki direction identify karo
    Atcp_df["SiteOrder"] = Atcp_df.apply(
        lambda x: 0 if str(x["Link ID (Site ID A - Site ID B and Site ID B - Site ID A)"]).startswith(str(x["Site ID"])) else 1,
        axis=1
    )

    Atcp_df = Atcp_df.sort_values(
        by=["MW Plan Id", "SortKey", "SiteOrder"],
        ignore_index=True
    )

    Atcp_df.drop(columns=["SortKey", "SiteOrder"], inplace=True)

    plan_map = {}
    sl_no = []

    for plan in Atcp_df["MW Plan Id"]:
        if plan not in plan_map:
            plan_map[plan] = len(plan_map) + 1
        sl_no.append(plan_map[plan])

    Atcp_df["SL NO"] = sl_no
    # =========================


    # ATP Sheet
    if not Atcp_df.empty:
        ws_atp = wb.create_sheet("ATP")

        data_atp = [Atcp_df.columns.tolist()] + Atcp_df.values.tolist()

        create_config_report(ws_atp, data_atp)

    output_file = os.path.join(final_folder, "MW_Ceragone_Report.xlsx")
    wb.save(output_file)

    relative_path = os.path.relpath(output_file, MEDIA_ROOT).replace("\\", "/")
    download_url = request.build_absolute_uri(MEDIA_URL + relative_path)

    return Response({
        "status": True,
        "message": "Report generated successfully",
        "download_url": download_url
    })
    

