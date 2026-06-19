from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import re

def create_config_report(ws, data):

    # ---------- Styles ----------
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)

    ok_fill = PatternFill("solid", fgColor="00B050")
    notok_fill = PatternFill("solid", fgColor="FF0000")
    na_fill = PatternFill("solid", fgColor="FFD966")

    ok_font = Font(color="FFFFFF", bold=True)
    notok_font = Font(color="FFFFFF", bold=True)

    normal_font = Font(name="Calibri", size=10)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    num_cols = len(data[0])

    # ---------- HEADER ----------
    for col_idx, header in enumerate(data[0], 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 16

    # ---------- DATA ----------
    for row_idx, row in enumerate(data[1:], 2):
        for col_idx, value in enumerate(row, 1):

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            val = str(value).strip().upper() if value is not None else ""

            cell.alignment = center
            cell.border = border
            cell.font = normal_font

            # ---------- Status column ----------
            if col_idx == num_cols:

                if val == "OK":
                    cell.fill = ok_fill
                    cell.font = ok_font

                elif val in ["NOT OK", "NOT FOUND"]:
                    cell.fill = notok_fill
                    cell.font = notok_font

            # ---------- NA ----------sss

        ws.row_dimensions[row_idx].height = 20

    # ---------- FIX COLUMN WIDTH ----------
    ws.column_dimensions["A"].width = 45  # Parameter
    ws.column_dimensions["B"].width = 45  # Dump Value

    # ---------- AUTO WIDTH FOR OTHERS ----------
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        if col_letter not in ["A", "B"]:
            ws.column_dimensions[col_letter].width = max_len + 3

    for cell in ws[1]:
        if cell.value == "SL NO":
            col_letter = get_column_letter(cell.column)
            ws.column_dimensions[col_letter].width = 7
            break

import re

def check_status(parameter, actual, expected, dump_values):

    actual = "" if actual is None else str(actual).strip().lower()
    expected = "" if expected is None else str(expected).strip().lower()

    # Parameters where only presence in dump matters
    present_check_params = [
        "date and time",
        "license key",
        "network ip",
        "password"
    ]

    if parameter.strip().lower() in present_check_params:
        return "OK" if actual not in ["", "na", "none"] else "NOT OK"

    # ACM Enabled/Modulation mode
    if parameter.strip().lower() == "acm enabled/modulation mode":

        if actual in ["adaptive", "enable", "enabled"] and \
        expected in ["adaptive", "enable", "enabled"]:
            return "OK"

        if actual in ["fixed", "disable", "disabled"] and \
        expected in ["fixed", "disable", "disabled"]:
            return "OK"

        return "NOT OK"

    # Manual fields
    if expected == "manual":
        return ""

    if parameter.strip().lower() == "multicarrier abc":

        idu_model = str(
            dump_values.get("IDU Model", "")
        ).strip().upper()

        actual_upper = actual.upper()

        # IP-20 → Slot 2
        if "IP-20" in idu_model:
            return "OK" if "SLOT 2" in actual_upper else "NOT OK"

        # IP-50 → Slot 1
        if "IP-50" in idu_model:
            return "OK" if "SLOT 1" in actual_upper else "NOT OK"

        return "NOT OK"
    
    # Current RSL handling
    if parameter.strip().lower() == "current rsl":

        try:
            actual_rsl = float(actual)

            # LB RSL value
            lb_rsl = float(
                dump_values.get("RSL", dump_values.get("Current RSL LB", actual_rsl))
            )

            if expected == "+/-3":
                return "OK" if abs(actual_rsl - lb_rsl) <= 3 else "NOT OK"

            elif expected == "+/-4":
                return "OK" if abs(actual_rsl - lb_rsl) <= 4 else "NOT OK"

        except:
            return "NOT OK"

    # Numeric comparison
    try:
        a = float(actual)
        e = float(expected)

        # Frequency handling: 15033000 → 15033
        if a > 100000:
            a = a / 1000

        if abs(a - e) < 0.001:
            return "OK"

    except:
        pass

    # MRMC Profile/Bandw handling
    try:
        actual_num = str(int(float(actual)))

        last_part = expected.split("-")[-1].upper()

        xyz_map = {
            "X": "10",
            "Y": "11",
            "Z": "12"
        }

        expected_num = xyz_map.get(last_part, last_part)

        if actual_num == expected_num:
            return "OK"

    except:
        pass

    # IDU Model handling
    actual_ip = re.search(r'ip-\d+', actual)
    expected_ip = re.search(r'ip-\d+', expected)

    if actual_ip and expected_ip:
        if actual_ip.group() == expected_ip.group():
            return "OK"

    # Text comparison (case insensitive)
    if actual == expected:
        return "OK"

    return "NOT OK"


master_parameters = [
        "Link ID (Site ID A - Site ID B and Site ID B - Site ID A)",
        "Site Name/System name/Unit Name",
        "Type of Equipment/IDU Model",
        "HOP visible in NMS (Y/N)",
        "Software Version IDU",
        "Frequency Tx (MHz)",
        "Frequency Rx (MHz)",
        "Tx Power (dBm)",
        "Current RSL",
        "MRMC Script ID/Bandw",
        "MRMC Profile/Bandw",
        "Modulation",
        "ACM Enabled/Modulation mode",
        "ATPC",
        "High Low Violation (Y/N)",
        "QoS configured (Y/N)",
        "Performance error in last 24 hour end",
        "Date and Time",
        "MSTP",
        "Undervoltage clear threshold (V)",
        "Undervoltage raise threshold (V)",
        "Critical Alarm at New Node",
        "Ethernet Port Speed",
        "MTU Value",
        "Tx Mute",
        "Server Location",
        "NodeLONG",
        "NodeLAT",
        "XPI Value",
        "SNMP",
        "HTTPS",
        "SNMP Version",
        "License Key",
        "Management Service",
        "Multicarrier ABC",
        "Network IP",
        "Password"
    ]

check_in_manual = [
    "HOP visible in NMS (Y/N)",
    "Performance error in last 24 hour end",
    "Critical Alarm at New Node",
    "XPI Value",
]


check_lb_params = [
    "Link ID (Site ID A - Site ID B and Site ID B - Site ID A)",
    "Site Name/System name/Unit Name",
    "Frequency Tx (MHz)",
    "Frequency Rx (MHz)",
    "Tx Power (dBm)",
    "Current RSL",
    "MRMC Script ID/Bandw",
    "MRMC Profile/Bandw",
    "ACM Enabled/Modulation mode",
    "ATPC",
    "NodeLAT",
    "NodeLONG",
    "Type of Equipment/IDU Model",
    "High Low Violation (Y/N)",
]


fix_values = {
    "Ethernet Port Speed": "1000",
    "MTU Value": "9612",
    "Tx Mute": "OFF",
    "SNMP": "ENABLE",
    "HTTPS": "ENABLE",
    "Undervoltage clear threshold (V)":"48",
    "Undervoltage raise threshold (V)":"46",
    "QoS configured (Y/N)":"all cos value match"

}

dscp_cos_mapping = [
    {"DSCP": "46", "CoS": "7"},
    {"DSCP": "20", "CoS": "7"},
    {"DSCP": "34", "CoS": "6"},
    {"DSCP": "18", "CoS": "6"},
    {"DSCP": "42", "CoS": "6"},
    {"DSCP": "44", "CoS": "6"},
    {"DSCP": "16", "CoS": "6"},
    {"DSCP": "28", "CoS": "5"},
    {"DSCP": "4",  "CoS": "5"},
    {"DSCP": "40", "CoS": "4"},
    {"DSCP": "14", "CoS": "4"},
    {"DSCP": "12", "CoS": "4"},
    {"DSCP": "32", "CoS": "3"},
    {"DSCP": "30", "CoS": "3"},
    {"DSCP": "10", "CoS": "3"},
    {"DSCP": "8",  "CoS": "3"},
    {"DSCP": "6",  "CoS": "3"},
    {"DSCP": "26", "CoS": "2"},
    {"DSCP": "24", "CoS": "2"},
    {"DSCP": "22", "CoS": "2"},
    {"DSCP": "56", "CoS": "1"},
    {"DSCP": "54", "CoS": "1"},
    {"DSCP": "52", "CoS": "1"},
    {"DSCP": "51", "CoS": "1"},
    {"DSCP": "48", "CoS": "1"},
    {"DSCP": "38", "CoS": "1"},
    {"DSCP": "36", "CoS": "1"},
    {"DSCP": "0",  "CoS": "1"}
]














