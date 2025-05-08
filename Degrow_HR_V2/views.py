from django.shortcuts import render

# Create your views here.
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
# Create your views here.
from openpyxl import Workbook,load_workbook

import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
# Create your views here.
from openpyxl import Workbook,load_workbook
import zipfile

from django.shortcuts import render
from django.shortcuts import render,redirect
from django.http import HttpResponse
import sys
import pandas as pd
from django.contrib import messages
# import json
from django.contrib.auth.decorators import login_required
from mcom_website.settings import MEDIA_URL,BASE_DIR,ALLOWED_HOSTS
import numpy as np
import os
import datetime
from rest_framework.decorators import api_view,parser_classes
from rest_framework.response import Response
from django.http import JsonResponse

# from .serializer import *

from django.forms.models import model_to_dict

from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import authentication_classes,permission_classes
from django.core.files.storage import FileSystemStorage
from mcom_website.settings import MEDIA_ROOT
from .models import *

import openpyxl
from openpyxl.styles import PatternFill
# Create your views here.
from openpyxl import Workbook,load_workbook
from statistics import mean

from django.db.models import Sum


import pandas as pd
import matplotlib  as mpl
import numpy as np

from datetime import date, timedelta


def delete_files_in_folder(folder_path):
    try:
        # Iterate over all files in the folder
        for file_name in os.listdir(folder_path):
            # Construct the full file path
            file_path = os.path.join(folder_path, file_name)
            
            # Check if it's a file and not a directory
            if os.path.isfile(file_path):
                # Delete the file
                os.remove(file_path)
                print(f"Deleted file: {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")

def traverse_folder(folder_path):
    # Iterate over all the files and folders in the specified folder
    # #print("in fun")
        for root, dirs, files in os.walk(folder_path):
                file_list=files
            # for file in files:
            #     file_path = os.path.join(root, file)  # Get the absolute path of the file
            #     #print(file_path)  # Do whatever you want with the file path
            #     files

        return file_list

def extract_date(string):
        str_list=string.split("_")
        #print(str_list)
        date_str=str_list[-1].split(".")[0]
        #print(date_str)
        year=int(date_str[0:4])
        month=int(date_str[4:6])
        day=int(date_str[6:])

        #print(year,month,day)

        date=datetime.date(year,month,day)
        
        date = date - timedelta(1)
        #print(date)
        return date


alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
def num_hash(num):
    if num < 26:
        return alpha[num-1]
    else:
        q, r = num//26, num % 26
        if r == 0:
            if q == 1:
                return alpha[r-1]
            else:
                return num_hash(q-1) + alpha[r-1]
        else:
            return num_hash(q) + alpha[r-1]
        
def template_colouring(ws,no_pre_cols,no_post_cols):
    BLUE="1a90ff"
    PURPLE="db70ff"
    CYAN="1addff"
    ORANGE="ffd53d"
    ws["A1"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE
    ws["A2"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE
    ws["B1"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE
    ws["B2"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE
    ws["A3"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE
    ws["B3"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE
    
    play=3
    ws[str(num_hash(play))+"1"].fill=PatternFill(patternType='solid',fgColor=ORANGE) #PURPLE #pre

    ws[str(num_hash(play+no_pre_cols))+"1"].fill=PatternFill(patternType='solid',fgColor=BLUE) #PURPLE #post

    ws[str(num_hash(play+no_pre_cols + no_post_cols))+"1"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE #comparision

    for i in range(0,no_pre_cols):
        j=play+i
        ws[str(num_hash(j))+"2"].fill=PatternFill(patternType='solid',fgColor=PURPLE)

    for x in range(0,no_post_cols):
        K=play + no_pre_cols + x
        ws[str(num_hash(K))+"2"].fill=PatternFill(patternType='solid',fgColor=PURPLE)

    ws[str(num_hash(play+no_pre_cols + no_post_cols))+"2"].fill=PatternFill(patternType='solid',fgColor=ORANGE) #ORANGE
    ws[str(num_hash(play+no_pre_cols + no_post_cols+1))+"2"].fill=PatternFill(patternType='solid',fgColor=BLUE) #BLUE
    ws[str(num_hash(play+no_pre_cols + no_post_cols+2))+"2"].fill=PatternFill(patternType='solid',fgColor=CYAN) #CYAN
    ws[str(num_hash(play+no_pre_cols + no_post_cols+3))+"2"].fill=PatternFill(patternType='solid',fgColor=CYAN) #CYAN

    return ws
def pre_degrow(additional_inputs,pre_file_dict): 
        print("print pre degrow...")
        print(additional_inputs)
        print(type(additional_inputs))
       ############################################## code for additional inputs ######################################################
       
        additional_inputs_df=additional_inputs
        freq_band=int(additional_inputs.at[0,"Freq_Band"])
        site_id=additional_inputs_df.at[0,"Site_ID"]
        global SITE_ID
        SITE_ID=site_id
        site_list=[site_id,("F" + site_id)]
        sector=additional_inputs_df.at[0,"Sector"]

        global site_sector
        site_sector = str(SITE_ID) +"_"+ str(sector) + "_" + str(freq_band)
        # #print(additional_inputs_df)
        # #print("freq_band:--",freq_band,type(freq_band))
        # #print("Site List:--",site_list,)
        # #print("sector:--",sector,type(sector))
        # #print("nishant verma...")
        ######################################## ******************************** ################################################
       

       #################################################  main processing  #######################################################                  
       
        # file_list=traverse_folder(folder_path_pre)
        # #print(file_list)
        K=0
        data1={} # dictionary containing data for "4G Data Volume[GB]" from 24hrs file
        data2={}  # dictionary containing data for ALL Other KPis from BBH  file
        code_executed=False
        for date,df in pre_file_dict.items():
            
                i=date
            
            ############################################ code for 24hrs ###########################
            # if :
                # #print("------------------------------------------------------------------################## pre ####################----------------------------------------------------")
                # #print("------------------------------------------------------------------##################",i,"####################----------------------------------------------------")
                # #print(".......-------------------------------------------------------------------------inside 24hrs------------------------------------------------------------.........")
                # path=os.path.join(folder_path_pre,file)
                # df=pd.read_csv(file)
                
                df.rename(columns={"Site_ID" :"Site_Name" } ,inplace = True )
                df=df[(df.Site_Name.isin(site_list))]                              # filtering the df with sitename 
                df['Short name'] = df['Short name'].str.strip()
                
                ################## to get the frequency band ##########################
                # band_list=df['Freq_Band'].unique()
                # band_list=pd.Series(band_list).dropna().astype(int)
                
                band_list=[1800,2100,900,2300]
                #print("Band List:-",band_list)

                dif_tech_data={}
                # row_sum = df.sum(axis=0)              
                # VoLTE_Traffic=row_sum["MV_4G Data Volume_GB"]
                # dif_tech_data[("MV_4G Data Volume_GB","NW")]=VoLTE_Traffic

                for band in band_list:  
                    
                    #print(band,freq_band)
                    if band == freq_band: # to compare the de grow technology
                        # #print(band)

                        df_fil=df[df['MV Freq_Band'] == freq_band]
                        if band == 2300:
                            # if df_fil['Short name'].str.contains("_T1_"):__
                                if sector == "A":
                                    
                                    df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "B":
                                    
                                    df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "C":
                                    
                                    df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "D":
                                    
                                    df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                #print("df_fil_t1---------")
                                # #print(df_fil_t1)
                                row_sum_t1 = df_fil_t1.sum(axis=0)
                                
                                t=row_sum_t1["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","TDDC1")]=t
                            # if df_fil['Short name'].str.contains("_T2_"):
                                if sector == "A":
                                    
                                    df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "B":
                                    
                                    df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "C":
                                    
                                    df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                
                                if sector == "D":
                                    
                                    df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                #print("df_fil_t2---------")

                                # #print(df_fil_t2)
                                row_sum_t2 = df_fil_t2.sum(axis=0)
                                t=row_sum_t2["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","TDDC2")]=t
                    
                        else:
                            if sector == "A":  
                                df_fil_sec = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")]
                            if sector == "B":
                                df_fil_sec = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")]
                            if sector == "C":  
                                df_fil_sec = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")]
                            if sector == "D":
                                df_fil_sec = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")]
                            

                            row_sum = df_fil_sec.sum(axis=0)
                            if band==1800:
                               
                                t=row_sum["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD18")]=t
                            if band==2100:
                               
                                t=row_sum["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD21")]=t
                            if band==900:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD8")]=t
                            
                            # #print(df_fil_sec)

                            
                        
                        
                    else :
                        df_fil=df[df['MV Freq_Band'] == band]
                        # #print(df_fil)
                        # #print(band)

                        if band == 2300:
                            # if df_fil['Short name'].str.contains("_T1_"):__
                                if sector == "A":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "B":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "C":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "D":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]

                                # #print("df_fil_t1---------")
                                # #print(df_fil_t1)
                                row_sum_t1 = df_fil_t1.sum(axis=0)
                                
                                t=row_sum_t1["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","TDDC1")]=t
                            # if df_fil['Short name'].str.contains("_T2_"):
                                if sector == "A":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "B":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "C":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "D":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                #print("df_fil_t2---------")


                                # #print(df_fil_t2)
                                row_sum_t2 = df_fil_t2.sum(axis=0)
                                t=row_sum_t2["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","TDDC2")]=t


                        else:


                            if sector == "A":
                                    
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")]
                            if sector == "B":
                                
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")]
                            if sector == "C":
                                
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")]
                            if sector == "D":
                                
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")]
                            

                            row_sum = df_fil.sum(axis=0)
                            if band==1800:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                # data[i]={("MV_4G Data Volume_GB","FD18"):t}
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD18")]=t
                            if band==2100:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                # data[i]={("MV_4G Data Volume_GB","FD21"):t}
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD21")]=t
                            if band==900:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                # data[i]={("MV_4G Data Volume_GB","FD8"):t}
                
    
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD9")]=t
                
                #---------------- code for nettwork level data ------------------------#
                row_sum = df.sum(axis=0)              
                VoLTE_Traffic=row_sum["MV_4G Data Volume_GB"]
                dif_tech_data[("DATA TRAFFIC (24hrs)","NW")]=VoLTE_Traffic
                #------------------***************************-------------------------#

                data1[i]=dif_tech_data        
                
              
          ############################################################ code for bbh ###################################################
        for date,df in pre_file_dict.items():
                i=date
                # if not code_executed:
                #     K=1
                #     i="Day"+ str(K)
                #     code_executed=True
                

                # #print("------------------------------------------------------------------################## pre ####################----------------------------------------------------")
                # #print("------------------------------------------------------------------##################",i,"####################----------------------------------------------------")
                # #print("_________________--------------------------------------------------------------####  in bbh ####--------------------------------------------_____________________")
               
                # #print("yeas................")
                # path=os.path.join(folder_path_pre,file)
                # df=pd.read_csv(file)
                
                df.rename(columns={"Site_ID" :"Site_Name" } ,inplace = True )
                df=df[(df.Site_Name.isin(site_list))]
                df['Short name'] = df['Short name'].str.strip()
                
                ################## to get the frequency band ##########################
                # band_list=df['Freq_Band'].unique()
                # band_list=pd.Series(band_list).dropna().astype(int)
                
                KPI=["MV_VoLTE Traffic [CBBH]","MV_DL User Throughput_Kbps [CDBH]","MV_UL User Throughput_Kbps [CUBH]","DL PRB Utilization [CDBH]","MV_Avg Connected User [CDBH]","MV_RRC Setup Success Rate [CDBH]","MV_ERAB Setup Success Rate [CDBH]","MV_VoLTE DCR [CBBH]"]
                KPI_dic ={
                          "MV_VoLTE Traffic [CBBH]":"VOLTE TRAFFIC (CBBH)",
                          "MV_DL User Throughput_Kbps [CDBH]":"DL THPT (CDBH)",
                          "MV_UL User Throughput_Kbps [CUBH]":"UL THPT (CUBH)",
                          "DL PRB Utilization [CDBH]":"DL PRB Utilization (CDBH)",
                          "MV_Avg Connected User [CDBH]":"Avg RRC Connectd Users (CDBH)",
                          "MV_RRC Setup Success Rate [CDBH]":"RRC SR (CDBH)",
                          "MV_ERAB Setup Success Rate [CDBH]":"RAB SR (CDBH)",
                          "MV_VoLTE DCR [CBBH]":"VoLTE DCR (CBBH)",
                          }
                
                dif_tech_data={}
                for kpi in KPI:
                    if kpi =="MV_VoLTE Traffic [CBBH]":
                            band_list=[1800,2100,900,2300]
                            # #print("Band List:-",band_list)
                            
                            for band in band_list:  
                                
                                # #print(band,freq_band)

                                if band == freq_band: # to compare the de grow technology
                                    # #print(band)

                                    df_fil=df[df['MV Freq_Band'] == freq_band]
                                    if band == 2300:
                                        # if df_fil['Short name'].str.contains("_T1_"):__
                                            if sector == "A":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "B":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "C":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "D":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                           
                                            # #print("df_fil_t1---------")
                                            #print(df_fil_t1)
                                            row_sum_t1 = df_fil_t1.sum(axis=0)
                                            
                                            t=row_sum_t1[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC1")]=t
                                        # if df_fil['Short name'].str.contains("_T2_"):
                                            if sector == "A":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "B":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]

                                            if sector == "C":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                        
                                            if sector == "D":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                           
                                            # #print("df_fil_t2---------")
                                            # #print(df_fil_t2)
                                            row_sum_t2 = df_fil_t2.sum(axis=0)
                                            t=row_sum_t2[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC2")]=t
                                
                                    else:
                                        if sector == "A":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")]
                                        if sector == "B":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")]
                                        if sector == "C":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")]
                                        if sector == "D":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")]
                                        

                                        row_sum = df_fil.sum(axis=0)
                                        if band==1800:
                                        
                                            t=row_sum[kpi]
                                            dif_tech_data[(KPI_dic[kpi]
                                            ,"FD18")]=t
                                        if band==2100:
                                        
                                            t=row_sum[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"FD21")]=t
                                        if band==900:
                                            
                                            t=row_sum[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"FD8")]=t
                                        
                                        #print(df_fil)

                                        
                                    
                                    
                                else :
                                    df_fil=df[df['MV Freq_Band'] == band]
                                    # #print(df_fil)
                                    # #print(band)
                                    if band == 2300:
                                        # if df_fil['Short name'].str.contains("_T1_"):__
                                            if sector == "A":
                                                
                                                df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "B":
                                                
                                                df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "C":
                                                
                                                df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "D":
                                                
                                                df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                            
                                            # #print("df_fil_t1---------")
                                            # #print(df_fil_t1)
                                            row_sum_t1 = df_fil_t1.sum(axis=0)
                                            
                                            t=row_sum_t1[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC1")]=t
                                        # if df_fil['Short name'].str.contains("_T2_"):
                                            if sector == "A":
                                                
                                                df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "B":
                                                
                                                df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "C":
                                                
                                                df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "D":
                                                
                                                df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                            #print("df_fil_t2---------")

                                            # #print(df_fil_t2)
                                            row_sum_t2 = df_fil_t2.sum(axis=0)
                                            t=row_sum_t2[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC2")]=t


                                    else:

                                            if sector == "A":
                                                    
                                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")]
                                            if sector == "B":
                                                
                                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")]
                                            if sector == "C":
                                                
                                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")]
                                            if sector == "D":
                                                
                                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")]
                                            

                                            row_sum = df_fil.sum(axis=0)
                                            if band==1800:
                                                
                                                t=row_sum[kpi]
                                                # data[i]={("MV_4G Data Volume_GB","FD18"):t}
                                                dif_tech_data[(KPI_dic[kpi],"FD18")]=t
                                            if band==2100:
                                                
                                                t=row_sum[kpi]
                                                # data[i]={("MV_4G Data Volume_GB","FD21"):t}
                                                dif_tech_data[(KPI_dic[kpi],"FD21")]=t
                                            if band==900:
                                                
                                                t=row_sum[kpi]
                                                # data[i]={("MV_4G Data Volume_GB","FD8"):t}
                                                dif_tech_data[(KPI_dic[kpi],"FD9")]=t
                            row_sum = df.sum(axis=0)                
                            t=row_sum[kpi]
                            dif_tech_data[(KPI_dic[kpi],"NW")]=t
                    else:
                            band_list=[1800,2100,900,2300]
                            # #print("Band List:-",band_list)
                            
                            for band in band_list:  
                                
                                # #print(band,freq_band)
                                # #print("Day--",K,"    KPI:--",kpi,"   Band:---",band)

                                if band == freq_band: # to compare the de grow technology
                                    # #print(band)

                                    df_fil=df[df['MV Freq_Band'] == freq_band]
                                    if band == 2300:
                                        # if df_fil['Short name'].str.contains("_T1_"):__
                                            if sector == "A":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "B":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "C":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "D":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                            #print("df_fil_t1---------")
                                            #print(df_fil_t1)
                                            row_sum_t1 = df_fil_t1.mean(axis=0)
                                            
                                            t=row_sum_t1[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC1")]=t
                                        # if df_fil['Short name'].str.contains("_T2_"):
                                            if sector == "A":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "B":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "C":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "D":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                            ##print("df_fil_t2---------")

                                            ##print(df_fil_t2)
                                            row_sum_t2 = df_fil_t2.mean(axis=0)
                                            t=row_sum_t2[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC2")]=t
                                
                                    else:
                                        if sector == "A":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")]
                                        if sector == "B":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")]
                                        if sector == "C":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")]
                                        if sector == "D":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")]
                                        

                                        row_sum = df_fil.mean(axis=0)
                                        if band==1800:
                                        
                                            t=row_sum[kpi]
                                            dif_tech_data[(KPI_dic[kpi]
                                            ,"FD18")]=t
                                        if band==2100:
                                        
                                            t=row_sum[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"FD21")]=t
                                        if band==900:
                                            
                                            t=row_sum[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"FD8")]=t
                                        
                                        ##print(df_fil)

                                        
                                    
                                    
                                else :
                                    df_fil=df[df['MV Freq_Band'] == band]
                                    ##print(df_fil)
                                    ##print(band)
                                    if band == 2300:
                                        # if df_fil['Short name'].str.contains("_T1_"):__
                                            if sector == "A":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "B":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "C":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "D":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                            #print("df_fil_t1---------")
                                            #print(df_fil_t1)
                                            row_sum_t1 = df_fil_t1.mean(axis=0)
                                            
                                            t=row_sum_t1[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC1")]=t
                                        # if df_fil['Short name'].str.contains("_T2_"):
                                            if sector == "A":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "B":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "C":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "D":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                            #print("df_fil_t2---------")

                                            #print(df_fil_t2)
                                            row_sum_t2 = df_fil_t2.mean(axis=0)
                                            t=row_sum_t2[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC2")]=t


                                    else:

                                        if sector == "A":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")]
                                        if sector == "B":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")]
                                        if sector == "C":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")]
                                        if sector == "D":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")]
                                            

                                        row_sum = df_fil.mean(axis=0)
                                        if band==1800:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD18"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD18")]=t
                                        if band==2100:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD21"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD21")]=t
                                        if band==900:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD8"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD9")]=t
                            row_sum = df.mean(axis=0)                
                            t=row_sum[kpi]
                            dif_tech_data[(KPI_dic[kpi],"NW")]=t
                data2[i]=dif_tech_data 
                #print("DAY-------",i)  
                
                
        
        df1out=pd.DataFrame(data1)
        df2out=pd.DataFrame(data2)
        pre_df = pd.concat([df1out, df2out])
        
        print(pre_df)
        return pre_df
        ############################################## ***************************** ##############################################             

def post_degrow(additional_inputs,post_file_dict):
      # Example usage

       ############################################## code for additional inputs ######################################################
      
        additional_inputs_df=additional_inputs
        freq_band=int(additional_inputs_df.at[0,"Freq_Band"])
        
        site_id=additional_inputs_df.at[0,"Site_ID"]
        site_list=[site_id,("F" + site_id)]
        sector=additional_inputs_df.at[0,"Sector"]

        #print(additional_inputs_df)

        #print("freq_band:--",freq_band,type(freq_band))
        #print("sector:--",sector,type(sector))
        
        ######################################## ******************************** ################################################
       

       #################################################  main processing  #######################################################                  
       
        # file_list=traverse_folder(folder_path_post)
        # #print(file_list)
        K=5
        data1={}
        data2={}
        code_executed=False
        for date,df in post_file_dict.items():
        
            
                i=date
            
            ############################################ code for 24hrs ###########################
            # if "_24Hrs_" in file.name:
                #print("------------------------------------------------------------------################## post ####################----------------------------------------------------")
                #print("------------------------------------------------------------------##################",i,"####################----------------------------------------------------")
                #print(".......-------------------------------------------------------------------------inside 24hrs------------------------------------------------------------.........")
                #print("yeas................")
                # path=os.path.join(folder_path_post,file)
                # df=pd.read_csv(file)
                
                df.rename(columns={"Site_ID" :"Site_Name" } ,inplace = True )
                df=df[(df.Site_Name.isin(site_list))]
                df['Short name'] = df['Short name'].str.strip()
                
                ################## to get the frequency band ##########################
                # band_list=df['Freq_Band'].unique()
                # band_list=pd.Series(band_list).dropna().astype(int)

                
                band_list=[1800,2100,900,2300]
                #print("Band List:-",band_list)
                dif_tech_data={}
                # row_sum = df.sum(axis=0)              
                # VoLTE_Traffic=row_sum["MV_4G Data Volume_GB"]
                # dif_tech_data[("MV_4G Data Volume_GB","NW")]=VoLTE_Traffic

                for band in band_list:  
                    
                        #print(band,freq_band)
                   
                    
                        df_fil=df[df['MV Freq_Band'] == band]
                        #print(df_fil)
                        #print(band)

                        if band == 2300:
                            # if df_fil['Short name'].str.contains("_T1_"):__
                                if sector == "A":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "B":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "C":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "D":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                #print("df_fil_t1---------")
                                #print(df_fil_t1)
                                row_sum_t1 = df_fil_t1.sum(axis=0)
                                
                                t=row_sum_t1["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","TDDC1")]=t
                            # if df_fil['Short name'].str.contains("_T2_"):
                                if sector == "A":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "B":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "C":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "D":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                #print("df_fil_t2---------")

                                #print(df_fil_t2)
                                row_sum_t2 = df_fil_t2.sum(axis=0)
                                t=row_sum_t2["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","TDDC2")]=t


                        else:


                            if sector == "A":
                                    
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")]
                            if sector == "B":
                                
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")]
                            if sector == "C":
                                
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")]
                            if sector == "D":
                                
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")]
                            

                            row_sum = df_fil.sum(axis=0)
                            if band==1800:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                # data[i]={("MV_4G Data Volume_GB","FD18"):t}
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD18")]=t
                            if band==2100:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                # data[i]={("MV_4G Data Volume_GB","FD21"):t}
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD21")]=t
                            if band==900:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                # data[i]={("MV_4G Data Volume_GB","FD8"):t}
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD9")]=t
                row_sum = df.sum(axis=0)              
                VoLTE_Traffic=row_sum["MV_4G Data Volume_GB"]
                dif_tech_data[("DATA TRAFFIC (24hrs)","NW")]=VoLTE_Traffic
                data1[i]=dif_tech_data        
                
              
          ############################################################ code for bbh ###################################################
        for date,df in post_file_dict.items():
                i=date
                # if not code_executed:
                #     K=6
                #     i="Day"+ str(K)
                #     code_executed=True

                #print("------------------------------------------------------------------################## post ####################----------------------------------------------------")
                #print("------------------------------------------------------------------##################",i,"####################----------------------------------------------------")
                #print(".......-------------------------------------------------------------------------inside bbh ------------------------------------------------------------.........")
                #print("yeas................")
                # path=os.path.join(folder_path_post,file)
                # df=pd.read_csv(file)
                
                df.rename(columns={"Site Name" :"Site_Name" } ,inplace = True )
                df=df[(df.Site_Name.isin(site_list))]
                df['Short name'] = df['Short name'].str.strip()
                
                ################## to get the frequency band ##########################
                # band_list=df['Freq_Band'].unique()
                # band_list=pd.Series(band_list).dropna().astype(int)

                KPI=["MV_VoLTE Traffic [CBBH]","MV_DL User Throughput_Kbps [CDBH]","MV_UL User Throughput_Kbps [CUBH]","DL PRB Utilization [CDBH]","MV_Avg Connected User [CDBH]","MV_RRC Setup Success Rate [CDBH]","MV_ERAB Setup Success Rate [CDBH]","MV_VoLTE DCR [CBBH]"]
                KPI_dic ={
                          "MV_VoLTE Traffic [CBBH]":"VOLTE TRAFFIC (CBBH)",
                          "MV_DL User Throughput_Kbps [CDBH]":"DL THPT (CDBH)",
                          "MV_UL User Throughput_Kbps [CUBH]":"UL THPT (CUBH)",
                          "DL PRB Utilization [CDBH]":"DL PRB Utilization (CDBH)",
                          "MV_Avg Connected User [CDBH]":"Avg RRC Connectd Users (CDBH)",
                          "MV_RRC Setup Success Rate [CDBH]":"RRC SR (CDBH)",
                          "MV_ERAB Setup Success Rate [CDBH]":"RAB SR (CDBH)",
                          "MV_VoLTE DCR [CBBH]":"VoLTE DCR (CBBH)",
                          }
                dif_tech_data={}
                for kpi in KPI:
                    if kpi == "MV_VoLTE Traffic [CBBH]":
                            band_list=[1800,2100,900,2300]
                            #print("Band List:-",band_list)
                            
                            for band in band_list:  
                                
                                #print(band,freq_band)

                                
                                        
                                    
                                    
                                
                                df_fil=df[df['MV Freq_Band'] == band]
                                #print(df_fil)
                                #print(band)
                                if band == 2300:
                                    # if df_fil['Short name'].str.contains("_T1_"):__
                                        if sector == "A":
                                            df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                        if sector == "B":
                                    
                                            df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                        if sector == "C":
                                    
                                            df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                        if sector == "D":
                                    
                                            df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                        #print("df_fil_t1---------")
                                        #print(df_fil_t1)
                                        row_sum_t1 = df_fil_t1.sum(axis=0)
                                        
                                        t=row_sum_t1[kpi]
                                        dif_tech_data[(KPI_dic[kpi],"TDDC1")]=t
                                    # if df_fil['Short name'].str.contains("_T2_"):
                                        if sector == "A":
                                    
                                            df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                        if sector == "B":
                                    
                                            df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                        if sector == "C":
                                    
                                            df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                        if sector == "D":
                                    
                                            df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                        #print("df_fil_t2---------")

                                        #print(df_fil_t2)
                                        row_sum_t2 = df_fil_t2.sum(axis=0)
                                        t=row_sum_t2[kpi]
                                        dif_tech_data[(KPI_dic[kpi],"TDDC2")]=t


                                else:

                                        if sector == "A":
                                                
                                            df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")]
                                        if sector == "B":
                                            
                                            df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")]
                                        if sector == "C":
                                            
                                            df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")]
                                        if sector == "D":
                                            
                                            df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")]
                                        

                                        row_sum = df_fil.sum(axis=0)
                                        if band==1800:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD18"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD18")]=t
                                        if band==2100:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD21"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD21")]=t
                                        if band==900:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD8"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD9")]=t
                            row_sum = df.sum(axis=0)                
                            t=row_sum[kpi]
                            dif_tech_data[(KPI_dic[kpi],"NW")]=t
                    else:
                            band_list=[1800,2100,900,2300]
                            #print("Band List:-",band_list)
                            
                            for band in band_list:  
                                
                                #print(band,freq_band)

                                df_fil=df[df['MV Freq_Band'] == band]
                                #print(df_fil)
                                #print(band)
                                if band == 2300:
                                    # if df_fil['Short name'].str.contains("_T1_"):__
                                        if sector == "A":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                        if sector == "B":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                        if sector == "C":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                        if sector == "D":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                        #print("df_fil_t1---------")
                                        #print(df_fil_t1)
                                        row_sum_t1 = df_fil_t1.mean(axis=0)
                                        
                                        t=row_sum_t1[kpi]
                                        dif_tech_data[(KPI_dic[kpi],"TDDC1")]=t
                                    # if df_fil['Short name'].str.contains("_T2_"):
                                        if sector == "A":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                        if sector == "B":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                        if sector == "C":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                        if sector == "D":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                        #print("df_fil_t2---------")

                                        #print(df_fil_t2)
                                        row_sum_t2 = df_fil_t2.mean(axis=0)
                                        t=row_sum_t2[kpi]
                                        dif_tech_data[(KPI_dic[kpi],"TDDC2")]=t


                                else:

                                        if sector == "A":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")]
                                        if sector == "B":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")]
                                        if sector == "C":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")]
                                        if sector == "D":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")]
                                        

                                        row_sum = df_fil.mean(axis=0)
                                        if band==1800:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD18"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD18")]=t
                                        if band==2100:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD21"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD21")]=t
                                        if band==900:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD8"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD9")]=t
                            row_sum = df.mean(axis=0)                
                            t=row_sum[kpi]
                            dif_tech_data[(KPI_dic[kpi],"NW")]=t

                data2[i]=dif_tech_data        
                
        
        df1out=pd.DataFrame(data1)
        df2out=pd.DataFrame(data2)
        post_df = pd.concat([df1out, df2out])
        # #print(post_df)
        return post_df


def mid_process(additional_inputs,pre_file_dict,post_file_dict):
    pre_degrow_df=pre_degrow(additional_inputs,pre_file_dict)
    # print("------------------------------------- Pre degrow ---------------------------------------")
    # print(pre_degrow_df)
    
    # exit(0)
    # post_degrow_df =post_degrow(additional_input_file_path,folder_path_post)
    post_degrow_df =post_degrow(additional_inputs,post_file_dict)
    # # #print("------------------------------------- Post degrow ---------------------------------------")
    # # #print(post_degrow_df)
    
    degrow_output_df = pd.concat([pre_degrow_df, post_degrow_df], axis=1)
    #print(degrow_output_df)

    

    degrow_output_df['Pre avg'] = round(pre_degrow_df.mean(axis=1),2)
    degrow_output_df['Post avg'] = round(post_degrow_df.mean(axis=1),2)
    degrow_output_df['Delta'] = round(degrow_output_df['Post avg']-degrow_output_df['Pre avg'],2)

    degrow_output_df['% Change'] = round(((degrow_output_df['Delta'] /degrow_output_df['Pre avg']) *100),2)

    #print(degrow_output_df)
    
   
    ###################### for giving name to the index ########################

    degrow_output_df.reset_index(inplace=True)
    degrow_output_df.rename( columns={'level_0':'KPI'}, inplace=True )
    degrow_output_df.rename( columns={'level_1':'TECH'}, inplace=True)
    degrow_output_df.set_index(['KPI', 'TECH'], inplace=True)
    ######################### ************************* ########################
    tuples=[]
    cols_pre_df=pre_degrow_df.columns
    no_pre_cols=len(cols_pre_df)
    cols_post_df=post_degrow_df.columns
    no_post_cols = len(cols_post_df)

    for pre_col in cols_pre_df:
         tuples.append(("Pre",pre_col))
    for post_col in cols_post_df:
         tuples.append(("Post",post_col))
        
    comp=[("Comparision","Pre Avg"),
            ("Comparision","Post Avg"),
            ("Comparision","Delta"),
            ("Comparision","% Change")]
    tuples.extend(comp)
    #print(tuples)

    cols = pd.MultiIndex.from_tuples(tuples)
    degrow_output_df.columns=cols

    degrow_output_df=degrow_output_df.round(2)
    #print(degrow_output_df)

#################################################################################################
    xlname="v3degrow_output_" + str(site_sector)+ ".xlsx"
    saving_path=os.path.join(MEDIA_ROOT,'Degrow','output',xlname)
    #print(saving_path)
    degrow_output_df.to_excel(saving_path,sheet_name=SITE_ID)

    # wb = openpyxl.load_workbook(r"C:\Users\Lenovo\Desktop\general python\v3degrow_output.xlsx")
    wb = openpyxl.load_workbook(saving_path)
    ws = wb[SITE_ID]
    
    ws=template_colouring(ws,no_pre_cols, no_post_cols)
###########################################################################################################
    # return ws
    xlname="v3degrow_output_coloured_" + str(site_sector)+ ".xlsx"
    output_saving_path=os.path.join(MEDIA_ROOT,'Degrow','coloured_op',xlname)
    #print(output_saving_path)
    wb.save(output_saving_path)
    # trend_wb.save("opopopop.xlsx")
       
    # wb.save("v3degrow_output_coloured.xlsx")
    #print("excel operation done.........................")
    # download_path=os.path.join(MEDIA_URL,'Degrow',"v3degrow_output_coloured.xlsx")


#     #################################################################################################
#     saving_path=os.path.join(MEDIA_ROOT,'Degrow',"v3degrow_output.xlsx")
#     #print(saving_path)
#     degrow_output_df.to_excel(saving_path,sheet_name=SITE_ID)

#     # wb = openpyxl.load_workbook(r"C:\Users\Lenovo\Desktop\general python\v3degrow_output.xlsx")
#     wb = openpyxl.load_workbook(saving_path)
#     ws = wb[SITE_ID]
    
#     ws=template_colouring(ws,no_pre_cols, no_post_cols)
# ###########################################################################################################
    
#     output_saving_path=os.path.join(MEDIA_ROOT,'Degrow',"v3degrow_output_coloured.xlsx")
#     #print(output_saving_path)
#     wb.save(output_saving_path)
#     # trend_wb.save("opopopop.xlsx")
       
#     # wb.save("v3degrow_output_coloured.xlsx")
#     #print("excel operation done.........................")
#     download_path=os.path.join(MEDIA_URL,'Degrow',"v3degrow_output_coloured.xlsx")
def delete_file(file_path):
    try:
        # Check if the file exists
        if os.path.exists(file_path):
            # Delete the file
            os.remove(file_path)
            print(f"File deleted successfully: {file_path}")
        else:
            print(f"File does not exist: {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")

@api_view(["POST","GET"])
# @authentication_classes([TokenAuthentication])
# @permission_classes([IsAuthenticated])
def Degrow_trend(request):
    saved_folder_path=os.path.join(MEDIA_ROOT,'Degrow','output')
    delete_files_in_folder(saved_folder_path)

    saved_folder_path=os.path.join(MEDIA_ROOT,'Degrow','coloured_op')
    delete_files_in_folder(saved_folder_path)
    # exit()
    zip_folder_path=os.path.join(MEDIA_ROOT,'Degrow.zip')
    delete_file(zip_folder_path)
    # exit(0)
    pre_file = request.FILES["pre_file"] if 'pre_file' in request.FILES else None
    post_file = request.FILES["post_file"] if 'post_file' in request.FILES else None
    additional_inputs = request.FILES["additional_inputs"] if 'additional_inputs' in request.FILES else None

    df_pre_file=pd.read_excel(pre_file)
    df_post_file = pd.read_excel(post_file)
    df_additional_inputs = pd.read_excel(additional_inputs)


    df_pre_file["Site_ID"]=df_pre_file["Short name"].str.split("_").str[-2].str[:-1].str.strip()
    df_post_file["Site_ID"]=df_post_file["Short name"].str.split("_").str[-2].str[:-1].str.strip()
    df_pre_file
    bulk_input=[]
    for index,raw in df_additional_inputs.iterrows():
        # print(raw)
        # exit(0)
       
        additional_input_df =pd.DataFrame(raw)
        additional_input_df = additional_input_df.T
        additional_input_df = additional_input_df.reset_index(drop=True)
        # print("....................................................",additional_input_df)
        dismantle_date = raw["Dismantled date"]
        # print(dismantle_date)
        SITE_ID = raw["Site_ID"]
        sited_df_pre_file=df_pre_file[df_pre_file["Site_ID"]== SITE_ID]
    #      print(sited_df_pre_file)
        df_date=list(set(sited_df_pre_file['Date']))
        sorted_dates=sorted(df_date)
        # print(sorted_dates)
        for i,date in enumerate(sorted_dates):
                pre_dict={}
                if dismantle_date - timedelta(1) == date:
                    dt1=sorted_dates[i-4]
                    df_d1=sited_df_pre_file[sited_df_pre_file['Date'] ==dt1]
                    pre_dict[dt1]=df_d1
                    
                    dt2=sorted_dates[i-3]
                    df_d2=sited_df_pre_file[sited_df_pre_file['Date'] ==dt2]
                    pre_dict[dt2]=df_d2
                    
                    dt3=sorted_dates[i-2]
                    df_d3=sited_df_pre_file[sited_df_pre_file['Date'] ==dt3]
                    pre_dict[dt3]=df_d3
                    
                    dt4=sorted_dates[i-1]
                    df_d4=sited_df_pre_file[sited_df_pre_file['Date'] ==dt4]
                    pre_dict[dt4]=df_d4
                    
                    dt5=sorted_dates[i]
                    df_d5=sited_df_pre_file[sited_df_pre_file['Date'] == dt5]
                    pre_dict[dt5]=df_d5
                    
                    dt_list=[dt1,dt2,dt3,dt4,dt5]
                    break
                else:
                    dt_list=[]
        # print("............ pre dict.........",pre_dict)
        # exit(0)
        print("............... working for post..............")
        sited_df_post_file=df_post_file[df_post_file["Site_ID"] == SITE_ID]
        df_date=list(set(sited_df_post_file['Date']))
        sorted_dates=sorted(df_date)
        print(sorted_dates)
        # exit(0)
        max_date=sorted_dates[-1]
        g_index = sorted_dates.index(max_date)
        post_dict={}
        
        dt1=sorted_dates[g_index -4]
        df_d1=sited_df_post_file[sited_df_post_file['Date'] ==dt1]
        post_dict[dt1]=df_d1
        
        dt2=sorted_dates[g_index -3]
        df_d2=sited_df_post_file[sited_df_post_file['Date'] ==dt2]
        post_dict[dt2]=df_d2
        
        dt3=sorted_dates[g_index -2]
        df_d3=sited_df_post_file[sited_df_post_file['Date'] ==dt3]
        post_dict[dt3]=df_d3
        
        dt4=sorted_dates[g_index -1]
        df_d4=sited_df_post_file[sited_df_post_file['Date'] ==dt4]
        post_dict[dt4]=df_d4
        
        dt5=sorted_dates[g_index]
        df_d5=sited_df_post_file[sited_df_post_file['Date'] == dt5]
        post_dict[dt5]=df_d5
        
       
        # print("post dict................",post_dict)
        # exit(0)
        bulk_input.append([additional_input_df,pre_dict,post_dict])
         
    # print(bulk_input)
    
    for row in bulk_input:
        #  print(raw[2])
         mid_process(row[0],row[1],row[2])
         


         
    try:
        # Define the folder path
        folder_name='Degrow'
        folder_path = os.path.join(MEDIA_ROOT, folder_name)

        # Create a zip file
        zip_filename = f"{folder_name}.zip"
        zip_file_path = os.path.join(MEDIA_ROOT,zip_filename)
        with zipfile.ZipFile(zip_file_path, 'w') as zipf:
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    rel_path = os.path.relpath(file_path, folder_path)
                    zipf.write(file_path, rel_path)
        
        # Prepare the response
        # response = HttpResponse(open(zip_file_path, 'rb'), content_type='application/zip')
        # response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
        # return response
        download_path= os.path.join(MEDIA_URL,"Degrow.zip")
        return Response({"status":True,"message":'Sucesfully Executed',"Download_url":download_path})
    except Exception as e:
        return Response({'message': str(e)}, status=500) 
    
    


   
    print("..................................................download path fetched..........................................")
    return Response({"status":True,"message":"Succesfully uploaded","Download_url":download_path})