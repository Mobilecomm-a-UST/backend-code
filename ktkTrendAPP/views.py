from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.core.files.storage import FileSystemStorage
from datetime import date,timedelta
from openpyxl import workbook,load_workbook
import datetime
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import pandas as pd
from openpyxl.styles import Font,PatternFill,Alignment
from django.conf import settings
from rest_framework import status
import os


def G2_trend(df_2G_raw_kpi,df_site_list_2G,trend_wb,offered_date,df_mapping_file):
        print(df_2G_raw_kpi)
        ########################## merging fies ##############################################
        df_2G_raw_kpi=pd.merge(df_2G_raw_kpi,df_mapping_file,left_on='Site Name',right_on='Site ID',how='left')
        print(df_2G_raw_kpi.columns)
        print("dchbdhjcbdsjbjdsbbdvbjdbvdbvjdbjvbsv--3-kcn")
        ######################## ********************************* ###########################

        date1=offered_date
        # date1=date.today()
        dt1 = date1 - timedelta(1)
        dt2 = date1 - timedelta(2)
        dt3 = date1 - timedelta(3)
        dt4 = date1 - timedelta(4)
        dt5 = date1 - timedelta(5)
        ls=[dt1,dt2,dt3,dt4,dt5]
        
        alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        def num_hash(num):
                if num < 26:
                    return alpha[num-1]
                else:
                    q, r = num//26, num % 26
                    if r == 0:
                        if q == 1:
                            return alpha[r-1]
                        else:
                            return num_hash(q-1) + alpha[r-1]
                    else:
                        return num_hash(q) + alpha[r-1]
    
    # Driver code

    # printString(27906)

        def titleToNumber(s):
        # This process is similar to binary-to-
        # decimal conversion
            result = 0
            for B in range(len(s)):
                result *= 26
                result += ord(s[B]) - ord('A') + 1
            return result


        def overwrite(df_pivoted,kpi_name,coln1,trend_ws):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            print(kpi_name)
            index_pivot=df_pivoted.index.to_list()
            print("index ;###############################",index_pivot)
            print(len(index_pivot))
            print("index of pivoted table: ",index_pivot)
           
            dr=df_pivoted[kpi_name]
           

            print("columns of dr table",dr.columns)
            cl=dr.columns.to_list()
            print("column list",cl)
            
            # site_id=dr["SITE_ID"].to_list() 
            # cell_id=dr["CELL_ID"].to_list()    
            col1=dr[str(cl[0])].to_list()
            col2=dr[str(cl[1])].to_list()
            col3=dr[str(cl[2])].to_list()
            col4=dr[str(cl[3])].to_list()
            col5=dr[str(cl[4])].to_list()

            trend_ws[coln1+"3"].value=cl[0]
            trend_ws[coln2+"3"].value=cl[1]
            trend_ws[coln3+"3"].value=cl[2]
            trend_ws[coln4+"3"].value=cl[3]
            trend_ws[coln5+"3"].value=cl[4]

            # me=column_index_from_string(coln5)+1
            # me=get_column_letter(me)
            for i,value in enumerate(index_pivot):
                j=i+4
                
                trend_ws["H"+str(j)].value=date1
                trend_ws["I"+str(j)].value="Done"
                trend_ws["E"+str(j)].value="G900"
                trend_ws["F"+str(j)].value="Done"
                trend_ws["D"+str(j)].value="2G+FDD"



                trend_ws["A"+str(j)].value=index_pivot[i][1]
                trend_ws["B"+str(j)].value=index_pivot[i][0] 
                trend_ws["C"+str(j)].value=index_pivot[i][0] 
                trend_ws["G"+str(j)].value=index_pivot[i][2] 
               

                
                
                
                trend_ws[coln1+str(j)].value=col1[i]
                trend_ws[coln2+str(j)].value=col2[i]
                trend_ws[coln3+str(j)].value=col3[i]
                trend_ws[coln4+str(j)].value=col4[i]
                trend_ws[coln5+str(j)].value=col5[i]
                
                    
        
        G2_kpi1=[ "TCH Drop Call Rate KTK(%)",
                "HOSR(%)",
                "TCH Blocking_Mapa(%)",
                "SDCCH Blocking_Mapa(%)",
                "SDCCH Drop_Mapa(%)",
                "RxQual DL(%)",
                "K3014:Traffic Volume on TCH(Erl)",
                "TR373:Cell Availability(%)",
                
        ]

        site_list_2G=list(df_site_list_2G["Site ID"])
             
        # df_2G_raw_kpi.rename(columns={"Cell Name": "CellName"}, inplace=True)
               
        G2_filter_df = df_2G_raw_kpi[(df_2G_raw_kpi["Site Name_x"].isin(site_list_2G))]
        print("---------------------------- 2g df=-----------------------------------")
        G2_filter_df.fillna(value=0,inplace=True)
        print("G2_filter_df:",G2_filter_df)
        


        G2_filter_df.to_excel("filter.xlsx",index=False)

        df_pivoted_2G = G2_filter_df.pivot_table(index=["Site Name_x","Cell Name","Integration"],columns="Date")
        print("___________pivoted________",df_pivoted_2G)
        df_pivoted_2G.to_excel("Pivoted.xlsx")

        trend_ws=trend_wb["2G KPI"]
        if not df_pivoted_2G.empty:
            for kpi_name in G2_kpi1:
                    if(kpi_name=="TCH Drop Call Rate KTK(%)"):
                        overwrite(df_pivoted_2G,kpi_name,"J",trend_ws)
                    if(kpi_name=="HOSR(%)"):
                        overwrite(df_pivoted_2G,kpi_name,"P",trend_ws)
                    if(kpi_name=="TCH Blocking_Mapa(%)"):
                        overwrite(df_pivoted_2G,kpi_name,"V",trend_ws)
                    if(kpi_name=="SDCCH Blocking_Mapa(%)"):
                        overwrite(df_pivoted_2G,kpi_name,"AB",trend_ws)
                    if(kpi_name=="SDCCH Drop_Mapa(%)"):
                        overwrite(df_pivoted_2G,kpi_name,"AH",trend_ws)
                    if(kpi_name=="RxQual DL(%)"):
                        overwrite(df_pivoted_2G,kpi_name,"AN",trend_ws)
                    if(kpi_name=="K3014:Traffic Volume on TCH(Erl)"):
                        overwrite(df_pivoted_2G,kpi_name,"AT",trend_ws)
                    if(kpi_name=="TR373:Cell Availability(%)"):
                        overwrite(df_pivoted_2G,kpi_name,"AZ",trend_ws)                            

                    



      
    
    

    




@api_view(["POST"])
def old_ktk_trend(request):
        raw_kpi = request.FILES["4G_raw_kpi"] if '4G_raw_kpi' in request.FILES else None
        
        site_list = request.FILES["4G_site_list"] if '4G_site_list' in request.FILES else None
        offered_date = request.POST.get("offered_date")
        print(offered_date)
        offered_date=datetime.datetime.strptime(offered_date,"%Y-%m-%d").date()
        print("offered_date:",offered_date)

        location= MEDIA_ROOT + r"\trends\temporary_files"
        fs = FileSystemStorage(location=location)

       ####################### for converting raw kpi file to a dataframe ###########################
        raw_kpi = fs.save(raw_kpi.name, raw_kpi)
        # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
        filepath = fs.path(raw_kpi)
        print("file_path:-",filepath)
        df_raw_kpi=pd.read_excel(filepath)
        # df_raw_kpi=pd.read_csv(filepath)
        os.remove(path=filepath)
        print(filepath,"deleted...............")
        print(df_raw_kpi)

      ####################### for converting sitelist file to a dataframe ###########################
        
        site_list = fs.save(site_list.name, site_list)
        # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
        filepath = fs.path(site_list)
        print("file_path:-",filepath)
        df_site_list=pd.read_excel(filepath)
        os.remove(path=filepath)
        print(filepath,"deleted...............")
        print(df_site_list)

############################# 2G FILE###############################
        

       ####################### for converting 2G raw kpi file to a dataframe ###########################
        raw_kpi_2G = request.FILES["2G_raw_kpi"] if '2G_raw_kpi' in request.FILES else None
        raw_kpi_2G= fs.save(raw_kpi_2G.name, raw_kpi_2G)
        # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
        filepath = fs.path(raw_kpi_2G)
        print("file_path:-",filepath)
        df_2G_raw_kpi=pd.read_excel(filepath)
        # df_raw_kpi=pd.read_csv(filepath)
        os.remove(path=filepath)
        print(filepath,"deleted...............")
        print(df_2G_raw_kpi)

      ####################### for converting 2G sitelist file to a dataframe ###########################
        site_list_2G = request.FILES["2G_site_list"] if '2G_site_list' in request.FILES else None
        site_list_2G = fs.save(site_list_2G.name, site_list_2G)
        # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
        filepath = fs.path(site_list_2G)
        print("file_path:-",filepath)
        df_site_list_2G=pd.read_excel(filepath)
        os.remove(path=filepath)
        print(filepath,"deleted...............")
        print(df_site_list_2G)
#####################################################################################

    ###################### mapping file read and saving in dataframe ########################
        mapping_file = request.FILES["mapping_file"] if 'mapping_file' in request.FILES else None
        mapping_file = fs.save(mapping_file.name,mapping_file)
        # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
        filepath = fs.path(mapping_file)
        print("file_path:-",filepath)
        df_mapping_file=pd.read_excel(filepath)
        os.remove(path=filepath)
        print(filepath,"deleted...............")
        print(df_mapping_file)
    ################# *************************************************######################

        door_root= os.path.join(MEDIA_ROOT,'trends',"ktk")

        path_of_blnk_temp=os.path.join(door_root,"template","ktk_template.xlsx")
        trend_wb=load_workbook(path_of_blnk_temp)
        print(trend_wb.sheetnames)


        print("##################################################################################")

        ############### 2G trend process call ################
     
        G2_trend(df_2G_raw_kpi,df_site_list_2G,trend_wb,offered_date,df_mapping_file)


        ################## ************* #######################

        g2_site=[]
        f8_site=[]
        f3_site=[]
        t1t2_site=[]


        kpi=[ "RRC setup Success  ratio",
        "E-RAB Setup Success ratio",
        "LTE DCR",
        "DL user thput_H",
        "UL User throughput",
        "Radio Network Availability Rate(%)",
        "Average CQI",
        # "SINRrrrrrrrrrrrrrrrrrrrrrrrrrrrr",# to add kpi
        "Intra Freq HO Success Rate",
        "Inter Freq HO Success Rate",
        "CSFB SR",
        "Paging Discards",
        "PRB Util Downlink",
        "Total Traffic/Mbps",
        "eRAB Success Rate Volte_A",
        "VoLTE drop call rate_A",
        "Downlink Packet Loss Rate (VoLTE)",
        "Uplink Packet Loss Rate (VoLTE)",
        "1_SRVCC Success Rate (LTE to GSM)",
        "PUSCH SINR",
        "VoLTE HO Success Rate(IntraFreq)",
        "VoLTE HO Success Rate(InterFreq)",
        "VoLTE SRVCC IRAT Per Cal Rate",
        "VoLTE Call Setup Success Rate",  
        "UL BLER", ]

        kpiStrToZero=[  "RRC setup Success  ratio",
                "E-RAB Setup Success ratio",
                "LTE DCR",
                "DL user thput_H",
                "UL User throughput",
                "Radio Network Availability Rate(%)",
                "Average CQI",
                # "SINRrrrrrrrrrrrrrrrrrrrrrrrrrrrr",# to add kpi
                "Intra Freq HO Success Rate",
                "Inter Freq HO Success Rate",
                "CSFB SR",
                "Paging Discards",
                "PRB Util Downlink",
                "Total Traffic/Mbps",
                "eRAB Success Rate Volte_A",
                "VoLTE drop call rate_A",
                "Downlink Packet Loss Rate (VoLTE)",
                "Uplink Packet Loss Rate (VoLTE)",
                "1_SRVCC Success Rate (LTE to GSM)",
                "PUSCH SINR",
                "VoLTE HO Success Rate(IntraFreq)",
                "VoLTE HO Success Rate(InterFreq)",
                "VoLTE SRVCC IRAT Per Cal Rate",
                "VoLTE Call Setup Success Rate",
                "UL BLER",
    
    
                    ]


            
       
      
        
        ########################## the below code is to replace every string to zero from numeric columns ##################

        for x in kpiStrToZero:
             df_raw_kpi[x] = df_raw_kpi[x].replace(to_replace='.*', value=0, regex=True)
             print("______________________running____________________-----")
        ######################################  * * * * * * * * * * * * * * ################################################
       
        site_list=list(df_site_list["Site ID"])

 

        print("__________________raw KPI after converting str to zero_______________")
        print(df_raw_kpi)

        # df_raw_kpi[""].fillna( inplace=True, method="ffill")
        df_raw_kpi["Cell Name"] =df_raw_kpi["Cell Name"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..


        # df_raw_kpi.rename( columns={'Unnamed: 1':'date'}, inplace=True )

        # df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"]/1024)
        # df_raw_kpi.rename(columns={"MV_DL User Throughput_Kbps [CDBH]" :"MV_DL User Throughput_Mbps [CDBH]" } ,inplace = True )

        # df_raw_kpi["MV_UL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_UL User Throughput_Kbps [CDBH]"]/1024)
        # df_raw_kpi.rename(columns={"MV_UL User Throughput_Kbps [CDBH]" :"MV_UL User Throughput_Mbps [CDBH]" } ,inplace = True )

        print(df_raw_kpi)

        
        print(df_raw_kpi.columns)



        lis=list(df_raw_kpi["Cell Name"])
        sit_id_lis=[]
        cell_id_lis=[]
        for item in lis:
            if("_" in item):
                cell_id=item.split("_")[-2]
                ln=len(item.split("_")[-1])
                #print(ln)
                sit_id=item.split("_")[-2][:-ln]
            else:
                cell_id=item
                sit_id=item
            cell_id_lis.append(cell_id)
            sit_id_lis.append(sit_id)

        print(sit_id)
        print(cell_id_lis)

        df_raw_kpi.insert(1, "SITE_ID", sit_id_lis)
        df_raw_kpi.insert(2, "CELL_ID", cell_id_lis)
         



       ############################### tech extraction code by alpana mam ###############################

        lis1=list(df_raw_kpi["Cell Name"])
        techlist=[]
        for cell in lis1:
            if ('_F1_' in cell or '_F3_' in cell or '_F8_' in cell or '_T1' in cell or '_T2_' in cell):
                if('_F1_' in cell or '_F3_' in cell or '_F8_'):
                    tech="FDD"
                if("_T1_" in cell or "_T2_"in cell):
                    tech="TDD" 
                techlist.append(tech) 
            else:
                tech=cell
                techlist.append(tech)     
        # df_raw_kpi["tech"]=techlist
        df_raw_kpi.insert(3, "tech", techlist)

        ############################ **************************************** ############################
    



        df_raw_kpi.rename(columns={"Cell Name" :"CellName" } ,inplace = True )
        df_raw_kpi.fillna(value=0,inplace=True)
        

        ########################## merging fies ##############################################
        df_raw_kpi=pd.merge(df_raw_kpi,df_mapping_file,left_on='SITE_ID',right_on='Site ID',how='left')
        print(df_raw_kpi)
        ######################## ********************************* ###########################


        process_op_path=os.path.join(door_root,"process output")
        
        savepath=os.path.join(process_op_path,"desired input.xlsx")
        df_raw_kpi.to_excel(savepath)
    
        # date1=date(2023,2,27)
        date1=offered_date
        # date1=date.today()
        dt1 = date1 - timedelta(1)
        dt2 = date1 - timedelta(2)
        dt3 = date1 - timedelta(3)
        dt4 = date1 - timedelta(4)
        dt5 = date1 - timedelta(5)
        ls=[dt1,dt2,dt3,dt4,dt5]
        only_site_fil = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list))]
        only_site_fil.fillna(value=0,inplace=True)

        
        savepath=os.path.join(process_op_path,"only_site_date_filtered_input.xlsx")
        
       
        
       

        def perticular_tech(tech,site_list):
            # df_filtered = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.date.isin(ls))]
            # df_filtered = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.CellName.str.contains('|'.join(tech)))]
            df_filtered = only_site_fil[(only_site_fil.SITE_ID.isin(site_list)) & (only_site_fil.CellName.str.contains('|'.join(tech)))]
            
            print("__________filter___________-",df_filtered)
            if not df_filtered.empty:
                address="last_filtered_input"  + str(tech) + ".xlsx"
                savepath=os.path.join(process_op_path,address)
                df_filtered.to_excel(savepath)
                df_pivoted = df_filtered.pivot_table(index=["SITE_ID","CellName","CELL_ID","eNodeB Name","tech","LocalCell Id","Site Name","Integration"],columns="Date")
                df_pivoted.fillna(value=0,inplace=True)
                print("technology:",tech)
                print(df_pivoted)
                address_pivot="pivoted_input"  +str(tech)+ ".xlsx"
                savepath=os.path.join(process_op_path,address_pivot)
                df_pivoted.to_excel(savepath)
                print("________________pivoted______________",df_pivoted)
                return df_pivoted
                
            
            return df_filtered


      
        alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        def num_hash(num):
                if num < 26:
                    return alpha[num-1]
                else:
                    q, r = num//26, num % 26
                    if r == 0:
                        if q == 1:
                            return alpha[r-1]
                        else:
                            return num_hash(q-1) + alpha[r-1]
                    else:
                        return num_hash(q) + alpha[r-1]
    
    # Driver code

    # printString(27906)

        def titleToNumber(s):
        # This process is similar to binary-to-
        # decimal conversion
            result = 0
            for B in range(len(s)):
                result *= 26
                result += ord(s[B]) - ord('A') + 1
            return result


        def overwrite(df_pivoted,kpi_name,coln1,trend_ws):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            print(kpi_name)
            index_pivot=df_pivoted.index.to_list()
            print("index ;###############################",index_pivot)
            print(len(index_pivot))
            print("index of pivoted table: ",index_pivot)
           
            dr=df_pivoted[kpi_name]
           

            print("columns of dr table",dr.columns)
            cl=dr.columns.to_list()
            print("column list",cl)
            
            # site_id=dr["SITE_ID"].to_list() 
            # cell_id=dr["CELL_ID"].to_list()
            try:    
                col1=dr[str(cl[0])].to_list()
            except:
                col1=[0]*len(dr.index)
            try:    
                col2=dr[str(cl[1])].to_list()
            except:
                col2=[0]*len(dr.index)
            try:    
                col3=dr[str(cl[2])].to_list()
            except:
                col3=[0]*len(dr.index)
            try:    
                col4=dr[str(cl[3])].to_list()
            except:
                col4=[0]*len(dr.index)
            try:    
                col5=dr[str(cl[4])].to_list()
            except:
                col5=[0]*len(dr.index)
            # col2=dr[str(cl[1])].to_list()
            # col3=dr[str(cl[2])].to_list()
            # col4=dr[str(cl[3])].to_list()
            # col5=dr[str(cl[4])].to_list()
            try:
                trend_ws[coln1+"3"].value=cl[0]
            except:
                trend_ws[coln1+"3"].value='UNNAMMED0'
            try:
                trend_ws[coln2+"3"].value=cl[1]
            except:
                trend_ws[coln2+"3"].value='UNNAMMED1'
            try:
                trend_ws[coln3+"3"].value=cl[2]
            except:
                trend_ws[coln3+"3"].value='UNNAMMED2'
            try:
                trend_ws[coln4+"3"].value=cl[3]
            except:
                trend_ws[coln4+"3"].value='UNNAMMED3'
            try:
                trend_ws[coln5+"3"].value=cl[4]
            except:
                trend_ws[coln5+"3"].value='UNNAMMED4'
            # trend_ws[coln2+"3"].value=cl[1]
            # trend_ws[coln3+"3"].value=cl[2]
            # trend_ws[coln4+"3"].value=cl[3]
            # trend_ws[coln5+"3"].value=cl[4]

            # me=column_index_from_string(coln5)+1
            # me=get_column_letter(me)
            for i,value in enumerate(index_pivot):
                j=i+4
                trend_ws["A"+str(j)].value='KTK'
                trend_ws["L"+str(j)].value='Done'

                trend_ws["K"+str(j)].value=date1
                trend_ws["G"+str(j)].value=index_pivot[i][1]
                trend_ws["B"+str(j)].value=index_pivot[i][0] 
                trend_ws["E"+str(j)].value=index_pivot[i][2]
                trend_ws["F"+str(j)].value=index_pivot[i][3]
                trend_ws["I"+str(j)].value=index_pivot[i][4]
                trend_ws["C"+str(j)].value=index_pivot[i][5]
                trend_ws["H"+str(j)].value=index_pivot[i][6] 
                trend_ws["J"+str(j)].value=index_pivot[i][7] 


                # trend_ws["K"+str(j)].value=date1
                
                
                trend_ws[coln1+str(j)].value=col1[i]
                trend_ws[coln2+str(j)].value=col2[i]
                trend_ws[coln3+str(j)].value=col3[i]
                trend_ws[coln4+str(j)].value=col4[i]
                trend_ws[coln5+str(j)].value=col5[i]
                
                # trend_ws[me+str(j)].value='=COUNTIF(P5:T5,">=98.5")'

        # def overwrite_g2(df_pivoted,kpi_name,coln1,coln2,coln3,coln4,coln5,trend_ws):
        #     index_pivot=df_pivoted.index.to_list()
        #     print("index ;###############################",index_pivot)
        #     print(len(index_pivot))
        #     print("index of pivoted table: ",index_pivot)
        #     dr=df_pivoted[kpi_name]
        #     print("columns of dr table",dr.columns)
        #     cl=dr.columns.to_list()
        #     print("column list",cl)
            
        #     # site_id=dr["SITE_ID"].to_list() 
        #     # cell_id=dr["CELL_ID"].to_list()    
        #     col1=dr[str(cl[0])].to_list()
        #     col2=dr[str(cl[1])].to_list()
        #     col3=dr[str(cl[2])].to_list()
        #     col4=dr[str(cl[3])].to_list()
        #     col5=dr[str(cl[4])].to_list()

        #     trend_ws[coln1+"4"].value=cl[0]
        #     trend_ws[coln2+"4"].value=cl[1]
        #     trend_ws[coln3+"4"].value=cl[2]
        #     trend_ws[coln4+"4"].value=cl[3]
        #     trend_ws[coln5+"4"].value=cl[4]

        #     # me=column_index_from_string(coln5)+1
        #     # me=get_column_letter(me)
        #     for i,value in enumerate(index_pivot):
        #         j=i+5
        #         trend_ws["B"+str(j)].value=index_pivot[i][0]
        #         trend_ws["C"+str(j)].value=index_pivot[i][1]
        #         trend_ws["G"+str(j)].value=date1
                
                
                
        #         trend_ws[coln1+str(j)].value=col1[i]
        #         trend_ws[coln2+str(j)].value=col2[i]
        #         trend_ws[coln3+str(j)].value=col3[i]
        #         trend_ws[coln4+str(j)].value=col4[i]
        #         trend_ws[coln5+str(j)].value=col5[i]



        # for fdd
        pivot_fdd=perticular_tech(["_F3_"],site_list)
        trend_ws=trend_wb["FDD KPI"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="RRC setup Success  ratio"):
                    overwrite(pivot_fdd,kpi_name,"O",trend_ws)

                if(kpi_name=="E-RAB Setup Success ratio"):
                    overwrite(pivot_fdd,kpi_name,"V",trend_ws)

                if(kpi_name=="LTE DCR"):
                    overwrite(pivot_fdd,kpi_name,"AC",trend_ws)

                if(kpi_name=="DL user thput_H"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="UL User throughput"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="Radio Network Availability Rate(%)"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)
                
                if(kpi_name=="Average CQI"): 
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)
                
                # if(kpi_name=="SINR"):             # TO ADD
                #     overwrite(pivot_fdd,kpi_name,"BL",trend_ws)
                
                if(kpi_name=="Intra Freq HO Success Rate"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="Inter Freq HO Success Rate"):
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="CSFB SR"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="Paging Discards"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="PRB Util Downlink"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="Total Traffic/Mbps"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="eRAB Success Rate Volte_A"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)

                if(kpi_name=="VoLTE drop call rate_A"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)

                if(kpi_name=="Downlink Packet Loss Rate (VoLTE)"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)

                if(kpi_name=="Uplink Packet Loss Rate (VoLTE)"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)
                if(kpi_name=="1_SRVCC Success Rate (LTE to GSM)"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)
                if(kpi_name=="PUSCH SINR"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="VoLTE drop call rate_A"):    #####duplicate two drop present
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws) 

                if(kpi_name=="VoLTE HO Success Rate(IntraFreq)"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws) 

                if(kpi_name=="VoLTE HO Success Rate(InterFreq)"):
                    overwrite(pivot_fdd,kpi_name,"FM",trend_ws)  
                if(kpi_name=="VoLTE SRVCC IRAT Per Cal Rate"):
                    overwrite(pivot_fdd,kpi_name,"FT",trend_ws)  

                    
                print("___________________________running__________________")           

   

        # for tdd
        pivot_fdd=perticular_tech(["_T1_","_T2_"],site_list)
        trend_ws=trend_wb["TDD KPI"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="RRC setup Success  ratio"):
                    overwrite(pivot_fdd,kpi_name,"O",trend_ws)

                if(kpi_name=="E-RAB Setup Success ratio"):
                    overwrite(pivot_fdd,kpi_name,"V",trend_ws)

                if(kpi_name=="LTE DCR"):
                    overwrite(pivot_fdd,kpi_name,"AC",trend_ws)

                if(kpi_name=="DL user thput_H"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="UL User throughput"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="Radio Network Availability Rate(%)"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)
                
                if(kpi_name=="Average CQI"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)
                
                # if(kpi_name=="SINR"):#TO ADD
                #     overwrite(pivot_fdd,kpi_name,"BL",trend_ws)
                
                if(kpi_name=="Intra Freq HO Success Rate"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="Inter Freq HO Success Rate"):
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="CSFB SR"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="Paging Discards"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="PRB Util Downlink"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="Total Traffic/Mbps"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="eRAB Success Rate Volte_A"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)

                if(kpi_name=="VoLTE drop call rate_A"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)

                if(kpi_name=="Downlink Packet Loss Rate (VoLTE)"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)

                if(kpi_name=="Uplink Packet Loss Rate (VoLTE)"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)
                if(kpi_name=="1_SRVCC Success Rate (LTE to GSM)"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)
                if(kpi_name=="PUSCH SINR"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)
                if(kpi_name=="VoLTE drop call rate_A"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)    
                if(kpi_name=="VoLTE HO Success Rate(IntraFreq)"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)
                if(kpi_name=="VoLTE HO Success Rate(InterFreq)"):
                    overwrite(pivot_fdd,kpi_name,"FM",trend_ws)
                if(kpi_name=="VoLTE SRVCC IRAT Per Cal Rate"):
                    overwrite(pivot_fdd,kpi_name,"FT",trend_ws)            

             
            

        pivot_fdd=perticular_tech(["_F8_"],site_list)
        trend_ws=trend_wb["L900 KPI"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="RRC setup Success  ratio"):
                    overwrite(pivot_fdd,kpi_name,"O",trend_ws)

                if(kpi_name=="E-RAB Setup Success ratio"):
                    overwrite(pivot_fdd,kpi_name,"V",trend_ws)

                if(kpi_name=="LTE DCR"):
                    overwrite(pivot_fdd,kpi_name,"AC",trend_ws)

                if(kpi_name=="Intra Freq HO Success Rate"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="Inter Freq HO Success Rate"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="VoLTE Call Setup Success Rate"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)
                
                if(kpi_name=="VoLTE drop call rate_A"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)
                
                if(kpi_name=="DL user thput_H"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)
                
                if(kpi_name=="Downlink Packet Loss Rate (VoLTE)"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="Uplink Packet Loss Rate (VoLTE)"):
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="UL BLER"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="VoLTE HO Success Rate(IntraFreq)"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="PRB Util Downlink"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="Total Traffic/Mbps"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
               
               
                    




        pivot_fdd=perticular_tech(["_F1_"],site_list)
        trend_ws=trend_wb["L2100 KPI"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="RRC setup Success  ratio"):
                    overwrite(pivot_fdd,kpi_name,"P",trend_ws)

                if(kpi_name=="E-RAB Setup Success ratio"):
                    overwrite(pivot_fdd,kpi_name,"W",trend_ws)

                if(kpi_name=="LTE DCR"):
                    overwrite(pivot_fdd,kpi_name,"AD",trend_ws)

                if(kpi_name=="Intra Freq HO Success Rate"):
                    overwrite(pivot_fdd,kpi_name,"AK",trend_ws)

                if(kpi_name=="Inter Freq HO Success Rate"):
                    overwrite(pivot_fdd,kpi_name,"AR",trend_ws)

                if(kpi_name=="VoLTE Call Setup Success Rate"):
                    overwrite(pivot_fdd,kpi_name,"AY",trend_ws)
                
                if(kpi_name=="VoLTE drop call rate_A"):
                    overwrite(pivot_fdd,kpi_name,"BF",trend_ws)
                
                if(kpi_name=="DL user thput_H"):
                    overwrite(pivot_fdd,kpi_name,"BM",trend_ws)
                
                if(kpi_name=="Downlink Packet Loss Rate (VoLTE)"):
                    overwrite(pivot_fdd,kpi_name,"BT",trend_ws)
                
                if(kpi_name=="Uplink Packet Loss Rate (VoLTE)"):
                    overwrite(pivot_fdd,kpi_name,"CA",trend_ws)
                
                if(kpi_name=="UL BLER"):
                    overwrite(pivot_fdd,kpi_name,"CH",trend_ws)
                
                if(kpi_name=="VoLTE HO Success Rate(IntraFreq)"):
                    overwrite(pivot_fdd,kpi_name,"CO",trend_ws)
                
                if(kpi_name=="PRB Util Downlink"):
                    overwrite(pivot_fdd,kpi_name,"CV",trend_ws)
                
                if(kpi_name=="Total Traffic/Mbps"):
                    overwrite(pivot_fdd,kpi_name,"DC",trend_ws)
                
                


        # # print(g2_tech(g2_site)[1])
        # pivote_g2=g2_tech(g2_site)[1]
        # # pivote_g2.to_excel("g2_pivoted.xlsx")
        # trend_ws=trend_wb["2G"]



        # for kpi_name in g2_kpi:
        #     if(kpi_name=="SDCCH Blocking Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"L","M","N","O","P",trend_ws)
            
        #     if(kpi_name=="SDCCH Drop Call Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"S","T","U","V","W",trend_ws)
            
        #     if(kpi_name=="TCH Blocking Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"Z","AA","AB","AC","AD",trend_ws)

        #     if(kpi_name=="Drop Call Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"AG","AH","AI","AJ","AK",trend_ws)
            
        #     if(kpi_name=="Handover Success Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"AO","AP","AQ","AR","AS",trend_ws)
            

        #     if(kpi_name=="RX Quality"):
        #         overwrite_g2(pivote_g2,kpi_name,"AV","AW","AX","AY","AZ",trend_ws)
    
        #     if(kpi_name=="RACH Success Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"BC","BD","BE","BF","BG",trend_ws)

            
        #     if(kpi_name=="Total Voice Traffic [BBH]"):
        #         overwrite_g2(pivote_g2,kpi_name,"BJ","BK","BL","BM","BN",trend_ws)

            
        #     if(kpi_name=="TCH Assignment Success Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"BP","BQ","BR","BS","BT",trend_ws)

            




        output_path=os.path.join(door_root,"output","ktk_trend_output.xlsx")
        trend_wb.save(output_path)
        # load_excel_data()
        # directory = "Tnch_OUTPUT_trend"
    
        # # Parent Directory path
        # parent_dir = "C:/Users/dell7480/Desktop/" 
        # # Path
        # path = os.path.join(parent_dir, directory)
        # print(path)
        # #Create the directory
        # if(not (os.path.isdir(path))):
        #     os.makedirs(path)
        #     print("Directory '% s' created" % directory)
        # path= path + "/trend_output.xlsx"
        # trend_wb.save(path)
        download_path=os.path.join(MEDIA_URL,"trends","ktk","output","ktk_trend_output.xlsx")
        return Response({"status":True,"message":"Succesfully uploaded","Download_url":download_path})

@api_view(["POST"])
def pre_post_tech(request):
    try:
        raw_kpi_pre = request.FILES.get("raw_kpi_pre")
        raw_kpi_post = request.FILES.get("raw_kpi_post")
        site_list = request.FILES.get("site_list")

        if raw_kpi_pre and raw_kpi_post and site_list:
            output_dir = os.path.join(settings.MEDIA_ROOT,"KTK_PRE_POST", "Temprary_output")
            os.makedirs(output_dir, exist_ok=True)
#_____________________________________work for pre file__________________________________
            df_pre = pd.read_excel(raw_kpi_pre)
            df_pre['Short name'] = df_pre['Short name'].ffill()
            df_pre.fillna(value=0, inplace=True)
            df_pre['Layer'] = df_pre['Short name'].apply(lambda x: x.split('_')[2] if isinstance(x, str) and len(x.split('_')) > 2 else '')
            df_pre['Site_id'] = df_pre['Short name'].apply(lambda x: x.split('_')[4] if isinstance(x, str) and len(x.split('_')) > 4 else '')
            df_pre['Layer'] = df_pre['Layer'].replace({ 'F1': 'L21', 'F3': 'L18', 'F8': 'L9', 'T1': 'L23', 'T2': 'L23', 'F5': 'L85' })
            df_pre['Layer'] = df_pre['Layer'].ffill()
            df_pre.to_excel(os.path.join(output_dir, "desirePre.xlsx"), index=False)
#_____________filter site list and pre file____________________
            site_list = pd.read_excel(site_list)
            site_list.columns = site_list.columns.str.strip().str.replace(' ', '_')
            site_list['Integration'] = pd.to_datetime(site_list['Integration'], errors='coerce')
            df_pre.columns = df_pre.columns.str.strip().str.replace(' ', '_')
            df_pre['Date'] = pd.to_datetime(df_pre['Date'], errors='coerce')

        
            filtered_rows = []
            for row in site_list.itertuples(index=False):
                site_id = row.Cell_id
                integration_date = row.Integration

                start_date = integration_date - timedelta(days=3)
                end_date = integration_date - timedelta(days=1)

                mask = (
                    (df_pre['Site_id'] == site_id) &
                    (df_pre['Date'] >= start_date) &
                    (df_pre['Date'] <= end_date)
                )
                filtered = df_pre[mask].copy()
                if not filtered.empty:
                    filtered['Integration'] = integration_date
                    filtered_rows.append(filtered)

            filter_pre = pd.concat(filtered_rows, ignore_index=True)

    
            non_kpi_cols = ['Site_id', 'Layer', 'Date', 'Integration', 'Short_name']
            kpi_columns = [col for col in filter_pre.columns if col not in non_kpi_cols]

            for col in kpi_columns:
                filter_pre[col] = pd.to_numeric(filter_pre[col], errors='coerce')

 
            final_rows = []
            for _, row in filter_pre.iterrows():
                row_values = [
                    row['Site_id'],
                    row['Layer'],
                    row['Integration'],
                    row['Date'].strftime('%Y-%m-%d'),
                    row['Short_name'] 
                ] + [round(row[col], 2) if pd.notna(row[col]) else None for col in kpi_columns]
                final_rows.append(row_values)

            columns = ['Circle_Site_Sec', 'Layer', 'Integration_Date', 'Date','Short_name'] + kpi_columns
            expanded_pre_df = pd.DataFrame(final_rows, columns=columns)
            filtered_df = expanded_pre_df[expanded_pre_df['Short_name'].str.contains('T1|T2', na=False)]

            t1_t2_sum_pre = filtered_df.groupby(['Circle_Site_Sec', 'Date', 'Layer']).sum(numeric_only=True).reset_index()
            t1_t2_avg_pre =t1_t2_sum_pre.groupby(['Circle_Site_Sec', 'Layer']).mean(numeric_only=True).reset_index()


            l23_df = expanded_pre_df[expanded_pre_df['Layer'] == 'L23'].copy()

            l23_avg_merged = pd.merge(
                l23_df,
                t1_t2_avg_pre,  
                on=['Circle_Site_Sec', 'Layer'],
                suffixes=('', '_avg')
            )

            numeric_cols = t1_t2_avg_pre.select_dtypes(include='number').columns.tolist()

            for col in numeric_cols:
                l23_avg_merged[col] = l23_avg_merged[f"{col}_avg"]

            l23_avg_merged.drop(columns=[f"{col}_avg" for col in numeric_cols], inplace=True)

            non_l23_df = expanded_pre_df[expanded_pre_df['Layer'] != 'L23']
            updated_expanded_pre_df = pd.concat([non_l23_df, l23_avg_merged], ignore_index=True)
            updated_expanded_pre_df.to_excel(os.path.join(output_dir, "expanded_pre.xlsx"), index=False)
#_____________________________________pivot table for pre file__________________________________
            aggfunc = {
                'MV_4G_Data_Volume_GB': 'mean',
                'MV_DL_User_Throughput_Kbps_[CDBH]': 'mean',
                'MV_Average_number_of_used_DL_PRBs_[CDBH]': 'mean',
                'MV_E-UTRAN_Average_CQI_[CDBH]': 'mean',
                'MV_Average_UE_Distance_KM_[CDBH]': 'mean',
                'MV_Avg_Connected_User': 'mean'
            }

            pivot_pre =   updated_expanded_pre_df.pivot_table(
                columns=['Circle_Site_Sec', 'Layer'],
                values=list(aggfunc.keys()),
                aggfunc=aggfunc
            ).T.round(2)
            pivot_pre.columns = [f"{col}_Pre" for col in pivot_pre.columns]
            pivot_pre.to_excel(os.path.join(output_dir, "pivot_pre.xlsx"), index=True)
#_____________________________________work for post file__________________________________
            df_post = pd.read_excel(raw_kpi_post)
            df_post['Short name'] = df_post['Short name'].ffill()
            df_post.fillna(value=0, inplace=True)
            df_post['Layer'] = df_post['Short name'].apply(lambda x: x.split('_')[2] if isinstance(x, str) and len(x.split('_')) > 2 else '')
            df_post['Site_id'] = df_post['Short name'].apply(lambda x: x.split('_')[4] if isinstance(x, str) and len(x.split('_')) > 4 else '')
            df_post['Layer'] = df_post['Layer'].replace({ 'F1': 'L21', 'F3': 'L18', 'F8': 'L9', 'T1': 'L23', 'T2': 'L23', 'F5': 'L85' })
            df_post['Layer'] = df_post['Layer'].ffill()
            df_post.to_excel(os.path.join(output_dir, "desirePost.xlsx"), index=False)
#_____________filter site list and post file____________________
 
            site_list.columns = site_list.columns.str.strip().str.lower().str.replace(' ', '_')
            filter_pre = df_post[df_post['Site_id'].isin(site_list['cell_id'])]
            site_layer_group = filter_pre.groupby('Site_id')['Layer'].apply(lambda x: sorted(set(x))).reset_index()

            final_rows = []
            for _, row in site_layer_group.iterrows():
                site_id = row['Site_id']
                layers = row['Layer']
                for layer in layers:
                    final_rows.append([site_id, layer])

            site_layer_df = pd.DataFrame(final_rows, columns=['Circle_Site_Sec', 'Layer'])

            expanded_post_df = pd.merge(
                site_layer_df,
                filter_pre,
                left_on=['Circle_Site_Sec', 'Layer'],
                right_on=['Site_id', 'Layer'],
                how='left'
            )

            filtered_df = expanded_post_df[expanded_post_df['Short name'].str.contains('T1|T2', na=False)]
            t1_t2_sum_post = filtered_df.groupby(['Circle_Site_Sec', 'Date', 'Layer']).sum(numeric_only=True).reset_index()
            t1_t2_avg_post =t1_t2_sum_post.groupby(['Circle_Site_Sec', 'Layer']).mean(numeric_only=True).reset_index()
            l23_df = expanded_post_df[expanded_post_df['Layer'] == 'L23'].copy()
            l23_avg_merged = pd.merge(
                l23_df,
                t1_t2_avg_post,  
                on=['Circle_Site_Sec', 'Layer'],
                suffixes=('', '_avg')
            )
            numeric_cols = t1_t2_avg_post.select_dtypes(include='number').columns.tolist()

            for col in numeric_cols:
                l23_avg_merged[col] = l23_avg_merged[f"{col}_avg"]

            l23_avg_merged.drop(columns=[f"{col}_avg" for col in numeric_cols], inplace=True)
            non_l23_df = expanded_post_df[expanded_post_df['Layer'] != 'L23']
            updated_expanded_post_df = pd.concat([non_l23_df, l23_avg_merged], ignore_index=True)
            updated_expanded_post_df.to_excel(os.path.join(output_dir, "expanded_post.xlsx"), index=False)
#_____________________________________pivot table for post file__________________________________
            aggfunc_post = {
                'MV_4G Data Volume_GB': 'mean',
                'MV_DL User Throughput_Kbps [CDBH]': 'mean',
                'MV_Average number of used DL PRBs [CDBH]': 'mean',
                'MV_E-UTRAN Average CQI [CDBH]': 'mean',
                'MV_Average UE Distance KM [CDBH]': 'mean',
                'MV_Avg Connected User': 'mean'
            }

            pivot_post = updated_expanded_post_df.pivot_table(
                columns=['Circle_Site_Sec', 'Layer'],
                values=list(aggfunc_post.keys()),
                aggfunc=aggfunc_post
            ).T.round(2)
            pivot_post.columns = [f"{col}_Post" for col in pivot_post.columns]
            pivot_post.to_excel(os.path.join(output_dir, "pivot_post.xlsx"), index=True)
#_____________________________________work for final file__________________________________
            df_pre_post = pd.concat([pivot_pre, pivot_post], axis=1).reset_index()
            df_pre_post.fillna(value=0, inplace=True)
            df_pre_post["Circle_Site_Sec"] = df_pre_post["Circle_Site_Sec"].ffill()

            processed_data = []
            for site in df_pre_post["Circle_Site_Sec"].unique():
                site_df = df_pre_post[df_pre_post["Circle_Site_Sec"] == site]

                df_new = pd.DataFrame({
                    ("", "Circle_Site_Sec"): site_df["Circle_Site_Sec"],
                    ("", "Layer"): site_df["Layer"],

                    ("Payload 24 Hour", "Pre"): site_df["MV_4G_Data_Volume_GB_Pre"],
                    ("Payload 24 Hour", "Post"): site_df["MV_4G Data Volume_GB_Post"],
                    ("Payload 24 Hour", "Change %"): ((site_df["MV_4G Data Volume_GB_Post"] - site_df["MV_4G_Data_Volume_GB_Pre"]) / site_df["MV_4G_Data_Volume_GB_Pre"] * 100).round(2),

                    ("Throughput_CDBH", "Pre"): site_df["MV_DL_User_Throughput_Kbps_[CDBH]_Pre"],
                    ("Throughput_CDBH", "Post"): site_df["MV_DL User Throughput_Kbps [CDBH]_Post"],
                    ("Throughput_CDBH", "Change"): site_df["MV_DL User Throughput_Kbps [CDBH]_Post"] - site_df["MV_DL_User_Throughput_Kbps_[CDBH]_Pre"],

                    ("DL PRB CDBH", "Pre"): site_df["MV_Average_number_of_used_DL_PRBs_[CDBH]_Pre"],
                    ("DL PRB CDBH", "Post"): site_df["MV_Average number of used DL PRBs [CDBH]_Post"],
                    ("DL PRB CDBH", "Change"): site_df["MV_Average number of used DL PRBs [CDBH]_Post"] - site_df["MV_Average_number_of_used_DL_PRBs_[CDBH]_Pre"],

                    ("Avg CQI", "Pre"): site_df["MV_E-UTRAN_Average_CQI_[CDBH]_Pre"],
                    ("Avg CQI", "Post"): site_df["MV_E-UTRAN Average CQI [CDBH]_Post"],
                    ("Avg CQI", "Change"): site_df["MV_E-UTRAN Average CQI [CDBH]_Post"] - site_df["MV_E-UTRAN_Average_CQI_[CDBH]_Pre"],

                    ("Avg UE Distance", "Pre"): site_df["MV_Average_UE_Distance_KM_[CDBH]_Pre"],
                    ("Avg UE Distance", "Post"): site_df["MV_Average UE Distance KM [CDBH]_Post"],
                    ("Avg UE Distance", "Change"): site_df["MV_Average UE Distance KM [CDBH]_Post"] - site_df["MV_Average_UE_Distance_KM_[CDBH]_Pre"],

                    ("Avg Connected User", "Pre"): site_df["MV_Avg_Connected_User_Pre"],
                    ("Avg Connected User", "Post"): site_df["MV_Avg Connected User_Post"],
                    ("Avg Connected User", "Change"): site_df["MV_Avg Connected User_Post"] - site_df["MV_Avg_Connected_User_Pre"],
                })

                total_pre = df_new[("Payload 24 Hour", "Pre")].sum()
                total_post = df_new[("Payload 24 Hour", "Post")].sum()
                total_change = ((total_post - total_pre) / total_pre * 100).round(2)

                total_row = {
                    ("", "Circle_Site_Sec"): "",
                    ("", "Layer"): "Total",
                    ("Payload 24 Hour", "Pre"): total_pre,
                    ("Payload 24 Hour", "Post"): total_post,
                    ("Payload 24 Hour", "Change %"): total_change
                }

                df_total = pd.DataFrame([total_row])
                processed_data.append(df_new)
                processed_data.append(df_total)

            final_df = pd.concat(processed_data, ignore_index=True)

            output_dir_final = os.path.join(settings.MEDIA_ROOT,"KTK_PRE_POST" ,"Final_Output")
            os.makedirs(output_dir_final, exist_ok=True)
            output_path = os.path.join(output_dir_final, "Final_Pre_post.xlsx")
            final_df.to_excel(output_path)

            download_link = request.build_absolute_uri(settings.MEDIA_URL + "KTK_PRE_POST/Final_Output/Final_Pre_post.xlsx")
    #_____________________work with excel file_______________________
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            center_align = Alignment(horizontal='center', vertical='center')
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = center_align
            
            
            fill = PatternFill(start_color="90CCDC", end_color="90CCDC", fill_type="solid",)
            payload_columns = ["B","C"] 
            for col in payload_columns:
                for row in range(2,3):  
                        cell = ws[f"{col}{row}"]
                        cell.fill = fill
    
            fill1 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            target_header1= "Payload 24 Hour"
            for col in range(1, ws.max_column + 1):
                header1 = ws.cell(row=1, column=col).value  
                if header1 == target_header1:  
                    for row in range(1, 2): 
                        cell = ws.cell(row=row, column=col)
                        cell.fill = fill1
                    payload_columns = ["D", "E", "F"] 
                    for col in payload_columns:
                        for row in range(2,3):  
                            cell = ws[f"{col}{row}"]
                            cell.fill = fill1 
                            
            fill2 = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
            target_header2= "Throughput_CDBH"
            for col in range(1, ws.max_column + 1):
                header1 = ws.cell(row=1, column=col).value  
                if header1 == target_header2:  
                    for row in range(1, 2): 
                        cell = ws.cell(row=row, column=col)
                        cell.fill =fill2 
                    payload_columns = ["G", "H", "I"] 
                    for col in payload_columns:
                        for row in range(2,3):  
                            cell = ws[f"{col}{row}"]
                            cell.fill = fill2              

            fill3 = PatternFill(start_color="FFEAA7", end_color="FFEAA7", fill_type="solid")
            target_header3= "DL PRB CDBH"
            for col in range(1, ws.max_column + 1):
                header1 = ws.cell(row=1, column=col).value  
                if header1 == target_header3:  
                    for row in range(1, 2): 
                        cell = ws.cell(row=row, column=col)
                        cell.fill =fill3 
                    payload_columns = ["J", "K", "L"] 
                    for col in payload_columns:
                        for row in range(2,3):  
                            cell = ws[f"{col}{row}"]
                            cell.fill = fill3  
                            
            fill4 = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
            target_header4= "Avg CQI"
            for col in range(1, ws.max_column + 1):
                header1 = ws.cell(row=1, column=col).value  
                if header1 == target_header4:  
                    for row in range(1, 2):  
                        cell = ws.cell(row=row, column=col)
                        cell.fill =fill4
                    payload_columns = ["M", "N", "O"] 
                    for col in payload_columns:
                        for row in range(2,3):  
                            cell = ws[f"{col}{row}"]
                            cell.fill = fill4               
            fill5 = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
            target_header5= "Avg UE Distance"
            for col in range(1, ws.max_column + 1):
                header1 = ws.cell(row=1, column=col).value  
                if header1 == target_header5:  
                    for row in range(1, 2):
                        cell = ws.cell(row=row, column=col)
                        cell.fill =fill5 
                    payload_columns = ["P", "Q", "R"] 
                    for col in payload_columns:
                        for row in range(2,3):  
                            cell = ws[f"{col}{row}"]
                            cell.fill = fill5   
                            
            fill6= PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
            target_header6= "Avg Connected User"
            for col in range(1, ws.max_column + 1):
                header1 = ws.cell(row=1, column=col).value  
                if header1 == target_header6:  
                    for row in range(1, 2):  
                        cell = ws.cell(row=row, column=col)
                        cell.fill =fill6 
                    payload_columns = ["S", "T", "U"]
                    for col in payload_columns:
                        for row in range(2,3):  
                            cell = ws[f"{col}{row}"]
                            cell.fill = fill6       
                                                 
            circle_col = 2  
            current_value = None
            start_row = 2

            for row in range(2, ws.max_row + 2):  
                cell = ws.cell(row=row, column=circle_col)
                if cell.value != current_value:
                    if current_value is not None:
                        end_row = row - 1
                        if end_row > start_row:
                            col_letter = get_column_letter(circle_col)
                            ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                    current_value = cell.value
                    start_row = row


            for merged_range in ws.merged_cells.ranges:
                top_left_cell = ws[merged_range.coord.split(":")[0]]
                top_left_cell.font = Font(bold=True)


            for row in range(2, ws.max_row + 1):
                cell = ws[f"F{row}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00 "%"'

            for cell in ws["B"]:
                cell.font = cell.font.copy(bold=True)
            for cell in ws["C"]:
                cell.font = cell.font.copy(bold=True)

            wb.save(output_path)

            return Response({"message": "File uploaded successfully", "download_path": download_link, "status": True}, status=status.HTTP_200_OK)

        return Response({"error": "Please provide all required files: raw_kpi_pre, raw_kpi_post, site_list"}, status=status.HTTP_400_BAD_REQUEST)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)