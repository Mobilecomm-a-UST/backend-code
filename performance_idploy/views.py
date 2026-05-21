
import os
import pandas as pd
from datetime import date
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.http import FileResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
#from project_pat.settings import MEDIA_ROOT, MEDIA_URL
from django.shortcuts import render
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL

# ─────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────

main_folder = os.path.join(settings.MEDIA_ROOT, "performance_idploy")
input_path  = os.path.join(main_folder, "input")
output_path = os.path.join(main_folder, "output")

os.makedirs(input_path, exist_ok=True)
os.makedirs(output_path, exist_ok=True)

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────

OUTPUT_FILENAME     = "output_Offered Vs OA TAT.xlsx"
CATEGORIES          = ['<=12days', '13-21days', '22-30days', '>30days']
HEADER_BG           = '4472C4'
HEADER_FG           = 'FFFFFF'
TOTAL_BG            = 'D9E1F2'
ALT_ROW_BG          = 'F2F2F2'
BORDER_CLR          = '000000'

LAYERS_4G_EXCLUDE   = {"3500", "L3500", "N3500", "N2600"}
LAYERS_5G_INCLUDE   = {"3500", "L3500", "N3500"}
LAYERS_4G5G_EXCLUDE = {"N2600"}


# ─────────────────────────────────────────────
# API 1 — UPLOAD
# ─────────────────────────────────────────────

def _clear_old_files():
    """
    Delete all existing files in input and output folders.
    Guarantees both folders exist after clearing.
    """
    for folder in [input_path, output_path]:
        if os.path.exists(folder):
            for f in os.listdir(folder):
                os.remove(os.path.join(folder, f))
        os.makedirs(folder, exist_ok=True)


def _save_uploaded_file(uploaded):
    """Save uploaded file to input folder with its original name."""
    save_path = os.path.join(input_path, uploaded.name)
    with open(save_path, 'wb+') as destination:
        for chunk in uploaded.chunks():
            destination.write(chunk)
    return save_path


def _validate_file(filepath):
    """
    Read file and check all required columns exist.
    Required: Circle, On Air Date, Performance AT Offered Date, Offered Layer
    Called only once at upload time.
    Returns (df, error_message). If valid, error_message is None.
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls'):
        df = pd.read_excel(filepath)
    elif ext == '.csv':
        df = pd.read_csv(filepath)
    else:
        return None, f"Unsupported file type: {ext}"

    # strip whitespace from all column names
    df.rename(columns={c: c.strip() for c in df.columns}, inplace=True)

    lower_map = {c.lower(): c for c in df.columns}

    required = {
        'on air date':                 'On Air Date',
        'performance at offered date': 'Performance AT Offered Date',
        'performance at status date':  'Performance AT Status Date',
        'circle':                      'Circle',
        'offered layer':               'Offered Layer',
    }

    rename = {}
    for key, canonical in required.items():
        if canonical in df.columns:
            rename[canonical] = canonical
        elif key in lower_map:
            rename[lower_map[key]] = canonical
        else:
            match = next((c for c in df.columns if key in c.lower()), None)
            if match:
                rename[match] = canonical
            else:
                return None, (
                    f"Column '{canonical}' not found. "
                    f"Available columns: {list(df.columns)}"
                )

    df.rename(columns=rename, inplace=True)

    for col in ('On Air Date', 'Performance AT Offered Date', 'Performance AT Status Date'):
        df[col] = pd.to_datetime(df[col], errors='coerce').dt.date

    # drop rows where On Air Date is missing — cannot process without it
    df.dropna(subset=['On Air Date'], inplace=True)

    return df, None


@api_view(['POST', 'GET', 'DELETE'])
def upload_file(request):
    try:
        os.makedirs(input_path, exist_ok=True)
        os.makedirs(output_path, exist_ok=True)

        if request.method == 'POST':
            files = request.FILES.getlist('file')
            if not files:
                return Response(
                    {'error': 'No files uploaded'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            uploaded = files[0]
            ext      = os.path.splitext(uploaded.name)[1].lower()
            if ext not in ('.xlsx', '.xls', '.csv'):
                return Response(
                    {'error': f"Unsupported type '{ext}'. Use .xlsx, .xls, or .csv."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            _clear_old_files()

            save_path = _save_uploaded_file(uploaded)
            df, error = _validate_file(save_path)

            if error:
                os.remove(save_path)
                return Response({'error': error}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

            return Response({
                'status':    True,
                'message':   'File saved successfully',
                'filename':  uploaded.name,
                'rows_read': len(df),
            }, status=status.HTTP_200_OK)


        elif request.method == 'GET':
            # check input file
            input_file_info = None
            if os.path.exists(input_path):
                input_files = os.listdir(input_path)
                if input_files:
                    input_file_info = {
                        'filename': input_files[0],
                        'status':   'ready',
                    }

            # check output files
            output_files = []
            file_map = [
                ("output_Offered Vs OA TAT.xlsx",              'offered'),
                ("output_Performance vs OA TAT.xlsx",          'performance'),
                ("output_Offered Vs OA TAT_DateRange.xlsx",    'offered_range'),
                ("output_Performance vs OA TAT_DateRange.xlsx", 'perf_range'),
                ("output_performance_FTR.xlsx",                'ftr'),
                ("output_performance_SCFT FTR.xlsx",           'scft'),
            ]


           
            for filename, report_type in file_map:
                filepath = os.path.join(output_path, filename)
                if os.path.exists(filepath):
                    output_files.append({
                        'report_type':  report_type,
                        'filename':     filename,
                        'download_url': request.build_absolute_uri(
                            f"{settings.MEDIA_URL.rstrip('/')}/performance_idploy/output/{filename}"
                        ),
                    })

            return Response({
                'status':       True,
                'message':      'Files found' if input_file_info else 'No files found',
                'input_file':   input_file_info,
                'output_files': output_files,
            }, status=status.HTTP_200_OK)


        elif request.method == 'DELETE':
            if not os.path.exists(input_path) and not os.path.exists(output_path):
                return Response(
                    {'error': 'Folder does not exist'},
                    status=status.HTTP_404_NOT_FOUND,
                )

            deleted_files = _delete_all_files()

            return Response({
                'status':        True,
                'message':       'Files deleted successfully',
                'deleted_files': deleted_files,
            }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





# ─────────────────────────────────────────────
# API 2 — GET MONTHS
# ─────────────────────────────────────────────

def _get_input_df():
    """
    Read the already-validated file from input folder.
    Parse date columns back to date objects after reading.
    Returns (df, error_message).
    """
    if not os.path.exists(input_path):
        return None, "No input file found. Upload first via POST /idploy/upload/"

    for f in os.listdir(input_path):
        filepath = os.path.join(input_path, f)
        ext      = os.path.splitext(filepath)[1].lower()

        if ext in ('.xlsx', '.xls'):
            df = pd.read_excel(filepath)
        elif ext == '.csv':
            df = pd.read_csv(filepath)
        else:
            continue

       

        for col in ('On Air Date', 'Performance AT Offered Date', 'Performance AT Status Date'):
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.date

        return df, None

    return None, "No input file found. Upload first via POST /idploy/upload/"
# Added  start and end date validation here

def _validate_date_range(start_str, end_str):
    """
    Validate start_date and end_date strings.
    Returns (start_dt, end_dt, error_response) where error_response
    is None if valid, or a Response object if invalid.
    """
    if not start_str or not end_str:
        return None, None, Response(
            {'error': 'Provide both start_date and end_date. Format: YYYY-MM-DD'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        start_dt = date.fromisoformat(start_str)
        end_dt   = date.fromisoformat(end_str)
    except ValueError:
        return None, None, Response(
            {'error': f"Invalid date format. Use YYYY-MM-DD. Got: start='{start_str}' end='{end_str}'"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if start_dt > end_dt:
        return None, None, Response(
            {'error': f"start_date ({start_str}) must not be after end_date ({end_str})."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    return start_dt, end_dt, None

# get month

def _bucket_start(month_label):
    """'Dec 2025' -> date(2025, 11, 26) — 26th of previous month."""
    dt   = pd.to_datetime(month_label, format='%b %Y')
    prev = dt - relativedelta(months=1)
    return date(prev.year, prev.month, 26)


def _bucket_end(month_label):
    """'Dec 2025' -> date(2025, 12, 25) — 25th of current month."""
    dt = pd.to_datetime(month_label, format='%b %Y')
    return date(dt.year, dt.month, 25)


def _get_available_months(df):
    """
    Derive month bucket labels from On Air Date column.
    26th rollover rule:
      day >= 26 -> current month label
      day <  26 -> previous month label
    Returns sorted list e.g. ['Nov 2025', 'Dec 2025', 'Jan 2026']
    """
    def _to_label(d):
        if d.day >= 26:
            return d.strftime('%b %Y')
        prior = date(d.year, d.month, 1) - relativedelta(months=1)
        return prior.strftime('%b %Y')

    labels = df['On Air Date'].dropna().apply(_to_label)
    return sorted(set(labels), key=lambda x: pd.to_datetime(x, format='%b %Y'))


@api_view(['GET'])
def get_months(request):
    """
    GET /idploy/months/
    Returns available month bucket labels from the uploaded file.
    User picks one label and passes it to POST /idploy/generate-offered/
    """
    df, error = _get_input_df()
    if error:
        return Response({'error': error}, status=status.HTTP_404_NOT_FOUND)

    try:
        months = _get_available_months(df)
    except Exception as exc:
        return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    result = [{
        'label': m,
        'start': _bucket_start(m).strftime('%d-%b-%Y'),
        'end':   _bucket_end(m).strftime('%d-%b-%Y'),
    } for m in months]

    return Response({
        'months':    result,
        #'next_step': 'POST /idploy/generate-offered/ with body {"month": "<label>"}',
    }, status=status.HTTP_200_OK)

# Date range selection API

@api_view(['POST'])
def date_range_selection(request):
    """
    POST /idploy/date-range-selection/

    Validates and confirms a custom start/end date range.
    Use this to pick dates before calling generate-offered,
    generate-performance, generate-ftr, or generate-scft.

    Body:
    {
        "start_date": "2025-12-26",
        "end_date":   "2026-01-25"
    }

    Returns the confirmed date range + count of rows that fall
    within that range, so the frontend can show a preview
    before the user triggers report generation.
    """
    start_str = request.data.get('start_date', '').strip()
    end_str   = request.data.get('end_date',   '').strip()

    start_dt, end_dt, err = _validate_date_range(start_str, end_str)
    if err:
        return err

    df, error = _get_input_df()
    if error:
        return Response({'error': error}, status=status.HTTP_404_NOT_FOUND)

    # preview: count rows whose On Air Date falls in range
    mask        = (df['On Air Date'] >= start_dt) & (df['On Air Date'] <= end_dt)
    rows_in_range = int(mask.sum())

    return Response({
        'status':        True,
        'message':       'Date range confirmed. Use these dates in your report generation call.',
        'start_date':    start_dt.strftime('%d-%b-%Y'),
        'end_date':      end_dt.strftime('%d-%b-%Y'),
        'rows_in_range': rows_in_range,
        'next_steps': {
            'offered':     'POST /idploy/generate-offered/     with {"start_date": "...", "end_date": "..."}',
            'performance': 'POST /idploy/generate-performance/ with {"start_date": "...", "end_date": "..."}',
            'ftr':         'POST /idploy/generate-ftr/         with {"start_date": "...", "end_date": "..."}',
            'scft':        'POST /idploy/generate-scft/        with {"start_date": "...", "end_date": "..."}',
        },
    }, status=status.HTTP_200_OK)



# ─────────────────────────────────────────────
# API 3 — GENERATE + DOWNLOAD
# ─────────────────────────────────────────────

def _apply_layer_filter(df, case):
    """
    Filter DataFrame rows based on Offered Layer column.
    case '4G'    — exclude rows where Offered Layer is in LAYERS_4G_EXCLUDE
    case '5G'    — include only rows where Offered Layer is in LAYERS_5G_INCLUDE
    case '4G+5G' — exclude rows where Offered Layer is in LAYERS_4G5G_EXCLUDE
    """
    df      = df.copy()
    col     = 'Offered Layer'
    df[col] = df[col].astype(str).str.strip()

    if case == '4G':
        return df[~df[col].isin(LAYERS_4G_EXCLUDE)].copy()
    elif case == '5G':
        return df[df[col].isin(LAYERS_5G_INCLUDE)].copy()
    elif case == '4G+5G':
        return df[~df[col].isin(LAYERS_4G5G_EXCLUDE)].copy()
    else:
        raise ValueError(f"Unknown layer case: {case}")


def _classify(days):
    """
    Classify day difference into one of the 4 categories.
    0 and negatives also fall into <=12days.
    """
    if days <= 12:
        return '<=12days'
    elif days <= 21:
        return '13-21days'
    elif days <= 30:
        return '22-30days'
    else:
        return '>30days'


def _process_data(df, month_label, layer_case, date_col='Performance AT Offered Date'):
    """
    Step 1: Apply Offered Layer filter first.
    Step 2: Filter On Air Date within bucket (26th prev -> 25th current).
    Step 3: Split into pending (blank date_col) and filled.
    Step 4: Compute diff = date_col - On Air Date (days).
            Keep 0 and negatives — counted in <=12days.
            Only drop rows where diff could not be computed at all.
    Step 5: Classify diff into category.
    Step 6: Count per circle alphabetically + Grand Total.

    date_col: column to use for diff calculation.
              'Performance AT Offered Date' or 'Performance AT Status Date'
    """
    start = _bucket_start(month_label)
    end   = _bucket_end(month_label)
    return _process_data_by_date_range(df, start, end, layer_case, date_col, label=month_label)

def _process_data_by_date_range(df, start_dt, end_dt, layer_case, date_col='Performance AT Offered Date', label=None):
    # Step 1 — layer filter first
    df_layer = _apply_layer_filter(df, layer_case)

    # Step 2 — filter On Air Date within bucket
    mask_oa = (df_layer['On Air Date'] >= start_dt) & (df_layer['On Air Date'] <= end_dt)
    df_oa   = df_layer[mask_oa].copy()

    if df_oa.empty:
        available = _get_available_months(df)
        raise ValueError(
            f"No data found for '{label or f'{start_dt} to {end_dt}'}' with layer '{layer_case}'. "
            f"Available months: {available}"
        )

    # Step 3 — split into pending and filled using date_col ← FIXED
    mask_blank = df_oa[date_col].isna()
    df_pending = df_oa[mask_blank].copy()
    df_filled  = df_oa[~mask_blank].copy()

    # Step 4 — compute diff using date_col ← FIXED
    df_filled['_diff'] = (
        df_filled[date_col] - df_filled['On Air Date']
    ).apply(lambda x: x.days if hasattr(x, 'days') else None)

    df_filled = df_filled[df_filled['_diff'].notna()]

    # Step 5 — classify
    df_filled['_category'] = df_filled['_diff'].apply(_classify)

    # Step 6 — count per circle alphabetically
    all_circles = sorted(df_oa['Circle'].dropna().unique())

    result_circles = {}
    grand          = {c: 0 for c in CATEGORIES}
    grand_pending  = 0
    grand_total    = 0

    for circle in all_circles:
        df_c    = df_filled[df_filled['Circle'] == circle]
        counts  = df_c['_category'].value_counts()
        row     = {cat: int(counts.get(cat, 0)) for cat in CATEGORIES}

        pending = int((df_pending['Circle'] == circle).sum())
        total   = sum(row.values()) + pending

        pct_lt12  = round(
            row['<=12days'] / total * 100, 1
        ) if total > 0 else 0.0

        pct_lt21  = round(
            (row['<=12days'] + row['13-21days']) / total * 100, 1
        ) if total > 0 else 0.0

        pct_22_30 = round(
            (row['22-30days'] + row['<=12days'] + row['13-21days'])/ total * 100, 1
        ) if total > 0 else 0.0
        pct_gt_30 = round(
            (row['22-30days'] + row['<=12days'] + row['13-21days']+ row['>30days'])/ total * 100, 1
        ) if total > 0 else 0.0

        result_circles[circle] = {
            **row,
            'Pending':  pending,
            'Total':    total,
            '<12%':     pct_lt12,
            '<13-21%':     pct_lt21,
            '<22-30%':  pct_22_30,
            '>30days%':  pct_gt_30,
        }

        for cat in CATEGORIES:
            grand[cat] += row[cat]
        grand_pending += pending
        grand_total   += total

    grand_pct_lt12  = round(
        grand['<=12days'] / grand_total * 100, 1
    ) if grand_total > 0 else 0.0

    grand_pct_lt21  = round(
        (grand['<=12days'] + grand['13-21days']) / grand_total * 100, 1
    ) if grand_total > 0 else 0.0

    grand_pct_22_30 = round(
        (grand['<=12days'] + grand['13-21days']+grand['22-30days']) / grand_total * 100, 1
    ) if grand_total > 0 else 0.0
    grand_pct_gt_30 = round(
        (grand['<=12days'] + grand['13-21days']+grand['22-30days']+grand['>30days']) / grand_total * 100, 1
    ) if grand_total > 0 else 0.0


    period = label if label else f"{start_dt.strftime('%d-%b-%Y')} to {end_dt.strftime('%d-%b-%Y')}"

    return {
        'month':       period,
        'start':       start_dt.strftime('%d-%b-%Y'),
        'end':         end_dt.strftime('%d-%b-%Y'),
        'circles':     result_circles,
        'grand_total': {
            **grand,
            'Pending':  grand_pending,
            'Total':    grand_total,
            '<12%':     grand_pct_lt12,
            '<13-21%':     grand_pct_lt21,
            '<22-30%':  grand_pct_22_30,
            '>30days%':  grand_pct_gt_30,
        },
    }


def _write_sheet(ws, result, thin):
    """
    Write data into a single worksheet.
    Called once per layer case — 4G, 5G, 4G+5G.
    """
    ALL_COLS = CATEGORIES + ['Pending', 'Total','<12%', '<13-21%', '<22-30%','>30days%']

    def _hdr_cell(cell, value, bg=HEADER_BG, fg=HEADER_FG):
        cell.value     = value
        cell.font      = Font(name='Arial', bold=True, color=fg, size=10)
        cell.fill      = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin

    def _data_cell(cell, value, bold=False, bg=None):
        cell.value     = value
        cell.font      = Font(name='Arial', bold=bold, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin
        if bg:
            cell.fill = PatternFill('solid', start_color=bg)

    # Row 1 — month label spanning all columns A to I
    ws.merge_cells('A1:K1')
    _hdr_cell(ws['A1'], result['month'])

    # Row 2 — column headers
    _hdr_cell(ws['A2'], 'Circle')
    for col_idx, col_name in enumerate(ALL_COLS, start=2):
        _hdr_cell(ws.cell(row=2, column=col_idx), col_name)

    

    # Data rows
    row = 3
    for i, (circle, counts) in enumerate(result['circles'].items()):
        
        _data_cell(ws.cell(row=row, column=1), circle, bg='FFFFFF')
        for col_idx, col_name in enumerate(ALL_COLS, start=2):
            val = counts.get(col_name)
            # show blank instead of 0 for category counts and pending
            if col_name not in ('%<13-21', '%<22-30', '%>30days','Total') and val == 0:
                val = None
            _data_cell(ws.cell(row=row, column=col_idx), val, bg='FFFFFF')
        row += 1

    # Grand Total row
    grand = result['grand_total']
    _data_cell(ws.cell(row=row, column=1), 'Grand Total', bold=True, bg=TOTAL_BG)
    for col_idx, col_name in enumerate(ALL_COLS, start=2):
        val = grand.get(col_name)
        if col_name not in ('<12%','13-<21%', '<22-30%','>30days%', 'Total') and val == 0:
            val = None
        _data_cell(ws.cell(row=row, column=col_idx), val, bold=True, bg=TOTAL_BG)

    # Column widths
    col_widths = {
        'A': 10,   # Circle
        'B': 10,   # <=12days
        'C': 11,   # 13-21days
        'D': 11,   # 22-30days
        'E': 10,   # >30days
        'F': 10,   # Pending
        'G': 8,    # Total
        'H': 8,    # %<12 
        'I': 8,    # %<13-21
        'J': 10,   # %<22-30
        'K':10, # %>30days
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30


def _write_output_excel(results, out_filepath):
    """
    Write one Excel file with 3 sheets — 4G, 5G, 4G+5G.
    results is a dict: {'4G': result_dict, '5G': result_dict, '4G+5G': result_dict}
    """
    wb = Workbook()

    # remove default empty sheet
    wb.remove(wb.active)

    side = Side(style='thin', color=BORDER_CLR)
    thin = Border(left=side, right=side, top=side, bottom=side)

    for sheet_name in ['4G', '5G', '4G+5G']:
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, results[sheet_name], thin)

    wb.save(out_filepath)


def _delete_all_files():
    """Delete all input and output files."""
    deleted = []
    for folder in [input_path, output_path]:
        if os.path.exists(folder):
            for f in os.listdir(folder):
                os.remove(os.path.join(folder, f))
                deleted.append(f)
    return deleted


@api_view(['POST'])
def generate_offered(request):
    """
    POST /idploy/generate-offered/

    Option A — by month bucket:
    { "month": "Dec 2025" }

    Option B — by custom date range:
    { "start_date": "2025-12-26", "end_date": "2026-01-25" }

    Processes data for all 3 layer cases (4G, 5G, 4G+5G),
    writes one Excel with 3 sheets, returns JSON + download URL.
    """
    start_str   = request.data.get('start_date', '').strip()
    end_str     = request.data.get('end_date',   '').strip()
    month_label = request.data.get('month', '').strip()
    if not month_label and not (start_str or end_str):
        return Response(
            {'error': 'Provide either {"month": "Dec 2025"} or {"start_date": "...", "end_date": "..."}'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    df, error = _get_input_df()
    if error:
        return Response({'error': error}, status=status.HTTP_404_NOT_FOUND)

    

    # ── resolve date range ───────────────────────────────────────
    if month_label:
        try:
            out_filename = OUTPUT_FILENAME
            results = {
                '4G':    _process_data(df, month_label, '4G'),
                '5G':    _process_data(df, month_label, '5G'),
                '4G+5G': _process_data(df, month_label, '4G+5G'),
            }
        except ValueError as ve:
            return Response({'error': str(ve)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        start_dt, end_dt, err = _validate_date_range(start_str, end_str)
        if err:
            return err
        try:
            out_filename = "output_Offered Vs OA TAT_DateRange.xlsx"
            results = {
                case: _process_data_by_date_range(df, start_dt, end_dt, case, 'Performance AT Offered Date')
                for case in ['4G', '5G', '4G+5G']
            }
        except ValueError as ve:
            return Response({'error': str(ve)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    out_file = os.path.join(output_path, out_filename)
    _write_output_excel(results, out_file)

    download_url = request.build_absolute_uri(
        f"{settings.MEDIA_URL.rstrip('/')}/performance_idploy/output/{out_filename}"
    )

    return Response({
        'status':       True,
        'message':      'Offered report generated successfully.',
        'date_range':   f"{results['4G']['start']} to {results['4G']['end']}",
        'data': {
            '4G':    {'circles': results['4G']['circles'],    'grand_total': results['4G']['grand_total']},
            '5G':    {'circles': results['5G']['circles'],    'grand_total': results['5G']['grand_total']},
            '4G+5G': {'circles': results['4G+5G']['circles'], 'grand_total': results['4G+5G']['grand_total']},
        },
        'download_url': download_url,
    }, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────
# API 5 — GENERATE PERFORMANCE
# Supports both month label and start/end date range
# ─────────────────────────────────────────────

@api_view(['POST'])
def generate_performance(request):
    """
    POST /idploy/generate-performance/

    Option A — by month bucket:
    { "month": "Dec 2025" }

    Option B — by custom date range:
    { "start_date": "2025-12-26", "end_date": "2026-01-25" }
    """
    month_label = request.data.get('month', '').strip()
    start_str   = request.data.get('start_date', '').strip()
    end_str     = request.data.get('end_date',   '').strip()

    if not month_label and not (start_str or end_str):
        return Response(
            {'error': 'Provide either {"month": "Dec 2025"} or {"start_date": "...", "end_date": "..."}'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    df, error = _get_input_df()
    if error:
        return Response({'error': error}, status=status.HTTP_404_NOT_FOUND)
    
    if 'Performance AT Status' not in df.columns:
        return Response(
            {'error': "Column 'Performance AT Status' not found in uploaded file."},
            status=status.HTTP_422_UNPROCESSABLE_ENTITY,
        )

    df['Performance AT Status'] = df['Performance AT Status'].astype(str).str.strip()
    df = df[df['Performance AT Status'] == 'Accepted'].copy()

    if df.empty:
        return Response(
            {'error': "No rows found where Performance AT Status is 'Accepted'."},
            status=status.HTTP_404_NOT_FOUND,
        )

    # ── resolve date range ───────────────────────────────────────
    if month_label:
        try:
            out_filename = "output_Performance vs OA TAT.xlsx"
            results = {
                '4G':    _process_data(df, month_label, '4G',    'Performance AT Status Date'),
                '5G':    _process_data(df, month_label, '5G',    'Performance AT Status Date'),
                '4G+5G': _process_data(df, month_label, '4G+5G', 'Performance AT Status Date'),
            }
        except ValueError as ve:
            return Response({'error': str(ve)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        start_dt, end_dt, err = _validate_date_range(start_str, end_str)
        if err:
            return err
        try:
            out_filename = "output_Performance vs OA TAT_DateRange.xlsx"
            results = {
                case: _process_data_by_date_range(df, start_dt, end_dt, case, 'Performance AT Status Date')
                for case in ['4G', '5G', '4G+5G']
            }
        except ValueError as ve:
            return Response({'error': str(ve)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    out_file = os.path.join(output_path, out_filename)
    _write_output_excel(results, out_file)
    download_url = request.build_absolute_uri(
        f"{settings.MEDIA_URL.rstrip('/')}/performance_idploy/output/{out_filename}"
    )

    return Response({
        'status':       True,
        'message':      'Performance report generated successfully.',
        'date_range':   f"{results['4G']['start']} to {results['4G']['end']}",
        'data': {
            '4G':    {'circles': results['4G']['circles'],    'grand_total': results['4G']['grand_total']},
            '5G':    {'circles': results['5G']['circles'],    'grand_total': results['5G']['grand_total']},
            '4G+5G': {'circles': results['4G+5G']['circles'], 'grand_total': results['4G+5G']['grand_total']},
        },
        'download_url': download_url,
    }, status=status.HTTP_200_OK)




#added code new performance ftr
# ─────────────────────────────────────────────
# FTR CONSTANTS
# ─────────────────────────────────────────────

FTR_HEADER_FILL = PatternFill("solid", start_color="C55A11")
FTR_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
FTR_DATA_FONT   = Font(name="Arial", size=10)
FTR_TOTAL_FONT  = Font(name="Arial", bold=True, size=10)
FTR_TOTAL_FILL  = PatternFill("solid", start_color="F2F2F2")
FTR_BORDER_SIDE = Side(style="thin", color="BFBFBF")
FTR_BORDER      = Border(
    left=FTR_BORDER_SIDE, right=FTR_BORDER_SIDE,
    top=FTR_BORDER_SIDE,  bottom=FTR_BORDER_SIDE
)
FTR_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
FTR_LEFT        = Alignment(horizontal="left",   vertical="center")

OUTPUT_FTR_MONTH = "output_performance_FTR_Month.xlsx"
OUTPUT_FTR_WEEK  = "output_performance_FTR_Week.xlsx"


# ─────────────────────────────────────────────
# FTR HELPERS
# ─────────────────────────────────────────────

def _ftr_fill(ftr_value):
    """6-band FTR colour coding."""
    if ftr_value >= 95:
        return PatternFill("solid", start_color="375623")   # dark green
    elif ftr_value >= 85:
        return PatternFill("solid", start_color="70AD47")   # light green
    elif ftr_value >= 75:
        return PatternFill("solid", start_color="FFFF00")   # yellow
    elif ftr_value >= 65:
        return PatternFill("solid", start_color="FFC000")   # light orange
    elif ftr_value >= 35:
        return PatternFill("solid", start_color="F98D8D")   # light red
    else:
        return PatternFill("solid", start_color="C00000")   # dark red


def _ftr_font(ftr_value):
    """White bold for dark backgrounds, black for light backgrounds."""
    if ftr_value >= 95 or ftr_value < 35:
        return Font(name="Arial", size=10, color="FFFFFF", bold=True)
    return Font(name="Arial", size=10, color="000000")


def _calc_metrics(group):
    """
    Calculate FTR metrics for a group of rows.
    Total Site     = total rows in group
    Pending        = rows where Performance AT Status == 'Pending'
    Accepted 0     = Accepted + Rejection Counter == 0
    Acc Pending 0  = Acceptance Pending + Rejection Counter == 0
    FTR            = Accepted 0 / (Total - Pending - Acc Pending 0) * 100
    """
    total   = len(group)
    pending = (group["Performance AT Status"] == "Pending").sum()

    acc_zero = (
        (group["Performance AT Status"] == "Accepted") &
        (group["Performance AT Rejection Counter"] == 0)
    ).sum()

    acc_pend_zero = (
        (group["Performance AT Status"] == "Acceptance Pending") &
        (group["Performance AT Rejection Counter"] == 0)
    ).sum()

    denominator = total - pending - acc_pend_zero
    ftr = round((acc_zero / denominator) * 100, 2) if denominator > 0 else 0.0

    return pd.Series({
        "Total Site":                           total,
        "Pending":                              int(pending),
        "Accepted with 0 counter":              int(acc_zero),
        "Acceptance pending with 0 Counter":    int(acc_pend_zero),
        "FTR":                                  ftr,
    })


def _build_ftr_summary(df):
    """
    Circle-wise FTR summary sorted alphabetically.
    Grand Total row pinned at bottom.
    """
    if df.empty:
        return pd.DataFrame(columns=[
            "Circle", "Total Site", "Pending",
            "Accepted with 0 counter",
            "Acceptance pending with 0 Counter", "FTR"
        ])

    if "Circle" not in df.columns:
        metrics = _calc_metrics(df)
        summary = pd.DataFrame([metrics])
        summary.insert(0, "Circle", "ALL")
        return summary

    summary     = df.groupby("Circle", sort=True).apply(_calc_metrics).reset_index()
    summary     = summary.sort_values("Circle", ascending=True).reset_index(drop=True)
    total_row   = pd.DataFrame([_calc_metrics(df)])
    total_row.insert(0, "Circle", "Grand Total")
    return pd.concat([summary, total_row], ignore_index=True)


def _validate_ftr_columns(df):
    """
    Check that FTR-specific columns exist in the DataFrame.
    Returns (df, error_message).
    """
    required = [
        "Performance AT Status",
        "Performance AT Rejection Counter",
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        return None, f"Missing columns for FTR report: {missing}"

    # parse numeric
    df["Performance AT Rejection Counter"] = pd.to_numeric(
        df["Performance AT Rejection Counter"], errors="coerce"
    ).fillna(0)
    df["Performance AT Status"] = df["Performance AT Status"].astype(str).str.strip()

    return df, None

def _process_ftr(df, start_dt, end_dt):
    start = pd.Timestamp(start_dt)
    end   = pd.Timestamp(end_dt)

    results = {}
    for case in ['4G', '5G', '4G+5G']:
        df_layer = _apply_layer_filter(df, case)
        df_layer['On Air Date'] = pd.to_datetime(df_layer['On Air Date'], errors='coerce')
        mask     = (df_layer['On Air Date'] >= start) & (df_layer['On Air Date'] <= end)
        df_range = df_layer[mask].copy()
        results[case] = _build_ftr_summary(df_range)

    return results


def _write_ftr_sheet(ws, summary, period_label, sheet_title, metric_col = "FTR", report_title = "Performance_FTR"):
    """
    metric_col: column name to apply colour coding on.
                'FTR' for performance report, 'SCFT' for SCFT report.
    """
    if summary.empty:
        ws.cell(row=1, column=1, value="No data available for this filter.")
        return

    columns = list(summary.columns)

    col_widths = {
        "Circle":                               14,
        "Total Site":                           12,
        "Pending":                              10,
        "Accepted with 0 counter":              22,
        "Acceptance pending with 0 Counter":    30,
        "FTR":                                  8,
        "SCFT FTR":                             8,
    }

    # Row 1 — title bar
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell              = ws.cell(row=1, column=1,
                                      value=f"{report_title} |  {period_label}  |  {sheet_title}")
    title_cell.font         = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    title_cell.fill         = PatternFill("solid", start_color="1F4E79")
    title_cell.alignment    = FTR_CENTER

    # Row 2 — column headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell            = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font       = FTR_HEADER_FONT
        cell.fill       = FTR_HEADER_FILL
        cell.alignment  = FTR_CENTER
        cell.border     = FTR_BORDER

    # Data rows
    for row_idx, row in summary.iterrows():
        excel_row = row_idx + 3
        is_total  = row["Circle"] == "Grand Total"
        alt_fill  = PatternFill("solid", start_color="FFFFFF") if row_idx % 2 == 0 else None

        for col_idx, col_name in enumerate(columns, start=1):
            value       = row[col_name]
            cell        = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = FTR_BORDER

            if is_total:
                cell.font = FTR_TOTAL_FONT
                cell.fill = FTR_TOTAL_FILL
                cell.alignment = FTR_LEFT if col_idx == 1 else FTR_CENTER
                if col_name == metric_col:
                    cell.number_format = "0.0"
            else:
                if col_name == metric_col:
                    cell.fill = _ftr_fill(value)
                    cell.font = _ftr_font(value)
                    cell.number_format = "0.0"
                else:
                    cell.font = FTR_DATA_FONT
                    if alt_fill:
                        cell.fill = alt_fill

            cell.alignment = FTR_LEFT if col_idx == 1 else FTR_CENTER
            

    # column widths + row heights
    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 16)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30
    ws.freeze_panes             = "A3"


def _write_ftr_excel(summaries, period_label, out_filepath, metric_col = "FTR", report_title = "Performance_FTR"):
    """
    Write one Excel file with 3 sheets — 4G, 5G, 4G+5G.
    summaries = {'4G': df, '5G': df, '4G+5G': df}
    """
    wb = Workbook()
    wb.remove(wb.active)

    side = Side(style= "thin", color = "BFBFBF")
    thin = Border(left = side, right = side, top = side, bottom = side)

    for sheet_name in ['4G', '5G', '4G+5G']:
        ws = wb.create_sheet(title=sheet_name)
        _write_ftr_sheet(ws, summaries[sheet_name], period_label, sheet_name, metric_col= metric_col,
                         report_title = report_title)

    wb.save(out_filepath)







# ─────────────────────────────────────────────
# API — GENERATE FTR REPORT (date range only)
# ─────────────────────────────────────────────

@api_view(['POST'])
def generate_ftr(request):
    """
    POST /idploy/generate-ftr/

    Uses start/end date from date-range-selection.
    Body: { "start_date": "2025-12-26", "end_date": "2026-01-25" }
    """
    start_str = request.data.get('start_date', '').strip()
    end_str   = request.data.get('end_date',   '').strip()

    start_dt, end_dt, err = _validate_date_range(start_str, end_str)
    if err:
        return err

    df, error = _get_input_df()
    if error:
        return Response({'error': error}, status=status.HTTP_404_NOT_FOUND)

    df, col_error = _validate_ftr_columns(df)
    if col_error:
        return Response({'error': col_error}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

    period_label = f"{start_dt.strftime('%d-%b-%Y')} to {end_dt.strftime('%d-%b-%Y')}"
    out_filename = "output_performance_FTR.xlsx"
    out_file     = os.path.join(output_path, out_filename)

    try:
        summaries = _process_ftr(df, start_dt, end_dt)
        if all(s.empty for s in summaries.values()):
            return Response(
                {'error': f"No data found for date range: {period_label}"},
                status=status.HTTP_404_NOT_FOUND,
            )
        _write_ftr_excel(summaries, period_label, out_file)
    except Exception as exc:
        return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    download_url = request.build_absolute_uri(
        f"{settings.MEDIA_URL.rstrip('/')}/performance_idploy/output/{out_filename}"
    )

    return Response({
        'status':       True,
        'message':      'FTR report generated successfully.',
        'date_range':   period_label,
        'sheets':       ['4G', '5G', '4G+5G'],
        'data': {
            '4G':    summaries['4G'].to_dict(orient='records'),
            '5G':    summaries['5G'].to_dict(orient='records'),
            '4G+5G': summaries['4G+5G'].to_dict(orient='records'),
        },
        'download_url': download_url,
    }, status=status.HTTP_200_OK)
# ─────────────────────────────────────────────
# API 4 — CLEANUP
# ─────────────────────────────────────────────

# ADDED FOR SCFT
# ─────────────────────────────────────────────
# API — GENERATE SCFT REPORT
# ─────────────────────────────────────────────

def _validate_scft_columns(df):
    """
    Check that SCFT-specific columns exist in the DataFrame.
    Returns (df, error_message).
    """
    required = [
        "SCFT AT Status",
        "SCFT AT Rejection Counter",
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        return None, f"Missing columns for SCFT report: {missing}"

    df["SCFT AT Rejection Counter"] = pd.to_numeric(
        df["SCFT AT Rejection Counter"], errors="coerce"
    ).fillna(0)
    df["SCFT AT Status"] = df["SCFT AT Status"].astype(str).str.strip()

    return df, None


def _calc_scft_metrics(group):
    """
    Same FTR logic — just uses SCFT columns instead of Performance AT columns.
    Total Site     = total rows
    Pending        = SCFT AT Status == 'Pending'
    Accepted 0     = Accepted + SCFT AT Rejection Counter == 0
    Acc Pending 0  = Acceptance Pending + SCFT AT Rejection Counter == 0
    SCFT            = Accepted with zero counter/(Total no of sites-Acceptance pending with zero counter)* 100
    """
    total   = len(group)
    pending = (group["SCFT AT Status"] == "Pending").sum()

    acc_zero = (
        (group["SCFT AT Status"] == "Accepted") &
        (group["SCFT AT Rejection Counter"] == 0)
    ).sum()

    acc_pend_zero = (
        (group["SCFT AT Status"] == "Acceptance Pending") &
        (group["SCFT AT Rejection Counter"] == 0)
    ).sum()

    denominator = total - acc_pend_zero
    scft = round((acc_zero / denominator) * 100, 2) if denominator > 0 else 0.0

    return pd.Series({
        "Total Site":                           total,
        "Pending":                              int(pending),
        "Accepted with 0 counter":              int(acc_zero),
        "Acceptance pending with 0 Counter":    int(acc_pend_zero),
        "SCFT FTR":                             scft,
    })


def _build_scft_summary(df):
    """
    Circle-wise SCFT summary sorted alphabetically.
    Grand Total row pinned at bottom.
    Same as _build_ftr_summary — uses _calc_scft_metrics.
    """
    if df.empty:
        return pd.DataFrame(columns=[
            "Circle", "Total Site", "Pending",
            "Accepted with 0 counter",
            "Acceptance pending with 0 Counter", "SCFT FTR"
        ])

    if "Circle" not in df.columns:
        metrics = _calc_scft_metrics(df)
        summary = pd.DataFrame([metrics])
        summary.insert(0, "Circle", "ALL")
        return summary

    summary   = df.groupby("Circle", sort=True).apply(_calc_scft_metrics).reset_index()
    summary   = summary.sort_values("Circle", ascending=True).reset_index(drop=True)
    total_row = pd.DataFrame([_calc_scft_metrics(df)])
    total_row.insert(0, "Circle", "Grand Total")
    return pd.concat([summary, total_row], ignore_index=True)


def _process_scft(df, start_dt, end_dt):
    """
    Apply layer filters then date filter then build SCFT summary.
    Same flow as _process_ftr — uses _build_scft_summary.
    """
    start = pd.Timestamp(start_dt)
    end   = pd.Timestamp(end_dt)

    results = {}
    for case in ['4G', '5G', '4G+5G']:
        df_layer = _apply_layer_filter(df, case)

        df_layer['On Air Date'] = pd.to_datetime(df_layer['On Air Date'], errors='coerce')
        mask     = (df_layer['On Air Date'] >= start) & (df_layer['On Air Date'] <= end)
        df_range = df_layer[mask].copy()

        results[case] = _build_scft_summary(df_range)

    return results

# ─────────────────────────────────────────────
# API 7 — GENERATE SCFT (date range only)
# ─────────────────────────────────────────────

@api_view(['POST'])
def generate_scft(request):
    """
    POST /idploy/generate-scft/

    Uses start/end date from date-range-selection.
    Body: { "start_date": "2025-12-26", "end_date": "2026-01-25" }
    """
    start_str = request.data.get('start_date', '').strip()
    end_str   = request.data.get('end_date',   '').strip()

    start_dt, end_dt, err = _validate_date_range(start_str, end_str)
    if err:
        return err

    df, error = _get_input_df()
    if error:
        return Response({'error': error}, status=status.HTTP_404_NOT_FOUND)

    df, col_error = _validate_scft_columns(df)
    if col_error:
        return Response({'error': col_error}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

    period_label = f"{start_dt.strftime('%d-%b-%Y')} to {end_dt.strftime('%d-%b-%Y')}"
    out_filename = "output_performance_SCFT FTR.xlsx"
    out_file     = os.path.join(output_path, out_filename)

    try:
        summaries = _process_scft(df, start_dt, end_dt)
        if all(s.empty for s in summaries.values()):
            return Response(
                {'error': f"No data found for date range: {period_label}"},
                status=status.HTTP_404_NOT_FOUND,
            )
        _write_ftr_excel(summaries, period_label, out_file, metric_col='SCFT FTR', report_title='SCFT FTR')
    except Exception as exc:
        return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    download_url = request.build_absolute_uri(
        f"{settings.MEDIA_URL.rstrip('/')}/performance_idploy/output/{out_filename}"
    )

    return Response({
        'status':       True,
        'message':      'SCFT FTR report generated successfully.',
        'date_range':   period_label,
        'sheets':       ['4G', '5G', '4G+5G'],
        'data': {
            '4G':    summaries['4G'].to_dict(orient='records'),
            '5G':    summaries['5G'].to_dict(orient='records'),
            '4G+5G': summaries['4G+5G'].to_dict(orient='records'),
        },
        'download_url': download_url,
    }, status=status.HTTP_200_OK)

# New report at report

# ─────────────────────────────────────────────
# AT REPORT CONSTANTS
# ─────────────────────────────────────────────

AT_REPORT_FILENAME  = "Performance_AT_SR-WISE_Report.xlsx"

AT_HEADER_FILL      = PatternFill("solid", start_color="1F4E79")
AT_HEADER_FONT      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
AT_DATA_FONT        = Font(name="Arial", size=10)
AT_TOTAL_FONT       = Font(name="Arial", bold=True, size=10)
AT_PENDING_FILL     = PatternFill("solid", start_color="FFF2CC")   # light yellow for Pending rows
AT_ALT_FILL         = PatternFill("solid", start_color="F2F2F2")
AT_BORDER_SIDE      = Side(style="thin", color="BFBFBF")
AT_BORDER           = Border(
    left=AT_BORDER_SIDE, right=AT_BORDER_SIDE,
    top=AT_BORDER_SIDE,  bottom=AT_BORDER_SIDE
)
AT_CENTER           = Alignment(horizontal="center", vertical="center", wrap_text=True)
AT_LEFT             = Alignment(horizontal="left",   vertical="center")

AT_COL_WIDTHS = {
    "SR_Site ID": 26,
    "Site ID":    14,
    "Circle":     14,
    "PAT":        18,
    "PAT Date":   14,
    "SAT":        18,
    "SAT Date":   14,
    "KAT":        18,
    "KAT Date":   14,
    "SCFT":       18,
    "SCFT Date":  14,
}

# Maps output column names to input source columns
AT_STATUS_MAP = {
    "PAT":  "Physical AT Status",
    "SAT":  "Soft AT Status",
    "KAT":  "Performance AT Status",
    "SCFT": "SCFT AT Status",
}
AT_DATE_MAP = {
    "PAT Date":  "Physical AT Status Date",
    "SAT Date":  "Soft AT Status Date",
    "KAT Date":  "Performance AT Status Date",
    "SCFT Date": "SCFT AT Status Date",
}

AT_OUTPUT_COLS = [
    "SR_Site ID", "Site ID","Circle",
    "PAT", "PAT Date",
    "SAT", "SAT Date",
    "KAT", "KAT Date",
    "SCFT", "SCFT Date",
]


# ─────────────────────────────────────────────
# AT REPORT HELPERS
# ─────────────────────────────────────────────

def _validate_at_columns(df):
    """
    Check that all AT-report-specific columns are present.
    Normalises column names by stripping whitespace.
    Returns (df, error_message).
    """
    df.rename(columns={c: c.strip() for c in df.columns}, inplace=True)

    required = [
        "SR. No.", "Site ID", "Circle", "On Air Date",
        "Physical AT Status",   "Physical AT Status Date",
        "Soft AT Status",       "Soft AT Status Date",
        "Performance AT Status","Performance AT Status Date",
        "SCFT AT Status",       "SCFT AT Status Date",
        
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        return None, f"Missing columns for AT report: {missing}"

    # parse all date columns
    for col in (
        "On Air Date",
        "Physical AT Status Date", "Soft AT Status Date",
        "Performance AT Status Date", "SCFT AT Status Date",
    ):
        df[col] = pd.to_datetime(df[col], errors="coerce")

    # normalise status columns to stripped strings
    for col in (
        "Physical AT Status", "Soft AT Status",
        "Performance AT Status", "SCFT AT Status",
    ):
        df[col] = df[col].astype(str).str.strip()

    # SR. No. as string for concatenation
    df["SR. No."] = df["SR. No."].astype(str).str.strip()
    df["Circle"]  = df["Circle"].astype(str).str.strip()

    return df, None


def _resolve_status(group_col):
    """
    Priority rule across all rows in the group:
      1. ANY row == 'Pending'            → 'Pending'
      2. ANY row == 'Rejected'           → 'Pending'
      3. ANY row == 'Acceptance Pending' → 'Offered'
      4. All rows == 'Accepted'          → 'Accepted'

    Case-insensitive comparison. Ignores blank / NaN values.
    """
    # clean: drop nulls, strip whitespace, lowercase for comparison
    values = (
        group_col
        .dropna()
        .astype(str)
        .str.strip()
    )
    # remove empty strings and literal "nan" from unclean data
    values = values[~values.str.lower().isin(["", "nan", "none"])]

    if values.empty:
        return None

    lower_vals = values.str.lower()

    if (lower_vals == "pending").any():
        return "Pending"

    if (lower_vals == "rejected").any():
        return "Pending"

    if (lower_vals == "acceptance pending").any():
        return "Offered"

    return "Accepted"


def _resolve_date(date_col):
    """
    Return the most recent (max) date from the group.
    Returns None if all values are NaT.
    """
    valid = date_col.dropna()
    if valid.empty:
        return None
    return valid.max()


def _build_at_summary(df, start_dt, end_dt):
    """
    Build the AT report rows.

    Steps:
    1. Filter On Air Date within [start_dt, end_dt].
    2. Group by (SR. No., Site ID).
    3. For each group:
       - SR_Site ID  = SR_No + "_" + Site_ID
       - Status cols = pending-priority rule
       - Date cols   = most recent date
    4. Sort output by SR. No. ascending (numeric where possible).

    Returns a list of dicts, one per unique (SR. No., Site ID).
    """
    # ── Step 1: date filter on On Air Date ──────────────────────
    start_ts = pd.Timestamp(start_dt)
    end_ts   = pd.Timestamp(end_dt)
    mask     = (df["On Air Date"] >= start_ts) & (df["On Air Date"] <= end_ts)
    df_range = df[mask].copy()

    if df_range.empty:
        return []

    # ── Step 2: sort by date desc so first row = most recent ────
    # We sort on all four date cols; any one being recent is enough.
    # Primary sort key: On Air Date desc (secondary keys don't matter
    # for status resolution because _resolve_status scans all rows).
    df_range = df_range.sort_values("On Air Date", ascending=False)

    # ── Step 3: group and resolve ────────────────────────────────
    rows = []
    grouped = df_range.groupby(["SR. No.", "Site ID", "Circle"], sort=False)

    for (sr_no, site_id, circle), grp in grouped:
        row = {
            "SR_Site ID": f"{sr_no}_{site_id}",
            "Site ID":    site_id,
            "Circle": circle,
        }

        for out_col, src_col in AT_STATUS_MAP.items():
            row[out_col] = _resolve_status(grp[src_col])

        for out_col, src_col in AT_DATE_MAP.items():
            dt = _resolve_date(grp[src_col])
            row[out_col] = dt.strftime("%d-%b-%Y") if dt is not None else None

        rows.append(( circle,sr_no, row))

    # ── Step 4: sort by SR. No. numerically where possible ───────
    def _sort_key(item):
        circle,sr, = item[0]
        try:
            return (circle, 0, int(sr))
        except (ValueError, TypeError):
            return (circle, 1, str(sr))

    rows.sort(key=_sort_key)
    return [r for _, _, r in rows]


def _write_at_sheet(ws, rows, period_label):
    """
    Write AT report rows into a worksheet.
    Highlights rows where any status column is 'Pending'.
    """
    # ── Row 1: title bar ─────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(AT_OUTPUT_COLS))
    title_cell           = ws.cell(row=1, column=1,
                                   value=f"AT Report  |  {period_label}")
    title_cell.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    title_cell.fill      = AT_HEADER_FILL
    title_cell.alignment = AT_CENTER

    # ── Row 2: column headers ────────────────────────────────────
    for col_idx, col_name in enumerate(AT_OUTPUT_COLS, start=1):
        cell            = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font       = AT_HEADER_FONT
        cell.fill       = PatternFill("solid", start_color="2E75B6")
        cell.alignment  = AT_CENTER
        cell.border     = AT_BORDER

    if not rows:
        ws.cell(row=3, column=1, value="No data found for the selected date range.")
        return

    # ── Data rows ────────────────────────────────────────────────
    status_cols = list(AT_STATUS_MAP.keys())   # PAT, SAT, KAT, SCFT

    for row_idx, row_data in enumerate(rows, start=3):
        has_pending = any(
            str(row_data.get(sc, "")).strip() == "Pending"
            for sc in status_cols
        )
        row_fill = AT_PENDING_FILL if has_pending else (
            AT_ALT_FILL if row_idx % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        )

        for col_idx, col_name in enumerate(AT_OUTPUT_COLS, start=1):
            value        = row_data.get(col_name)
            cell         = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font    = AT_DATA_FONT
            cell.border  = AT_BORDER
            cell.fill    = row_fill
            cell.alignment = AT_LEFT if col_idx in (1, 2) else AT_CENTER

    # ── Column widths and row heights ────────────────────────────
    for col_idx, col_name in enumerate(AT_OUTPUT_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = AT_COL_WIDTHS.get(col_name, 16)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"


def _write_at_excel(rows, period_label, out_filepath):
    """
    Write a single-sheet Excel file for the AT report.
    One sheet named 'AT Report'.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "AT Report"

    _write_at_sheet(ws, rows, period_label)
    wb.save(out_filepath)


# ─────────────────────────────────────────────
# API — GENERATE AT REPORT
# ─────────────────────────────────────────────

@api_view(["POST"])
def performance_at_sr_wise_tracking(request):
    """
    POST /idploy/generate-at-report/

    Body:
    {
        "start_date": "2025-12-26",
        "end_date":   "2026-01-25"
    }

    Filters rows by On Air Date within [start_date, end_date].
    Groups by (SR. No., Site ID), applies pending-priority rule on
    status columns, picks most-recent date for date columns.

    Returns JSON data + download URL for the Excel file.
    """
    start_str = request.data.get("start_date", "").strip()
    end_str   = request.data.get("end_date",   "").strip()

    # ── validate date range ──────────────────────────────────────
    start_dt, end_dt, err = _validate_date_range(start_str, end_str)
    if err:
        return err

    # ── load input file ──────────────────────────────────────────
    df, error = _get_input_df()
    if error:
        return Response({"error": error}, status=status.HTTP_404_NOT_FOUND)

    # ── validate required columns ────────────────────────────────
    df, col_error = _validate_at_columns(df)
    if col_error:
        return Response({"error": col_error}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

    period_label = (
        f"{start_dt.strftime('%d-%b-%Y')} to {end_dt.strftime('%d-%b-%Y')}"
    )
    out_filename = AT_REPORT_FILENAME
    out_file     = os.path.join(output_path, out_filename)

    # ── process and write ────────────────────────────────────────
    try:
        rows = _build_at_summary(df, start_dt, end_dt)

        if not rows:
            return Response(
                {"error": f"No data found for On Air Date range: {period_label}"},
                status=status.HTTP_404_NOT_FOUND,
            )

        _write_at_excel(rows, period_label, out_file)

    except Exception as exc:
        return Response({"error": str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    download_url = request.build_absolute_uri(
        f"{settings.MEDIA_URL.rstrip('/')}/performance_idploy/output/{out_filename}"
    )

    return Response({
        "status":       True,
        "message":      "AT report generated successfully.",
        "date_range":   period_label,
        "total_sites":  len(rows),
        "data":         rows,
        "download_url": download_url,
    }, status=status.HTTP_200_OK)


    

@api_view(['DELETE'])
def cleanup(request):
    """
    DELETE /idploy/cleanup/
    Manually delete all input and output files without downloading.
    Use this to reset at any point in the flow.
    """
    deleted = _delete_all_files()
    return Response({
        'message':       'Cleanup complete.',
        'deleted_files': deleted,
    }, status=status.HTTP_200_OK)





