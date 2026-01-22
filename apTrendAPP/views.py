from django.shortcuts import render
from django.http import HttpResponse
from commom_utilities.utils import *
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from rest_framework.decorators import api_view
from rest_framework.response import Response
from tkinter import *
from django.core.files.storage import FileSystemStorage
from tkinter import filedialog, messagebox, ttk
import openpyxl
import pandas as pd
import numpy as np
from datetime import date, timedelta
import datetime
from datetime import date, timedelta
import re
from openpyxl import Workbook, load_workbook
import os
from zipfile import ZipFile
import os


@api_view(["POST"])
def old_ap4G_trend(request):
    raw_kpi_4G = request.FILES["raw_kpi_4G"] if "raw_kpi_4G" in request.FILES else None
    print(raw_kpi_4G)
    if raw_kpi_4G:
        location = MEDIA_ROOT + r"\trends\temporary_files"
        fs = FileSystemStorage(location=location)
        file = fs.save(raw_kpi_4G.name, raw_kpi_4G)
        file_path = fs.path(file)
        print("file_path: ", file_path)
        df_raw_kpi_4G = pd.read_excel(file_path)
        print(df_raw_kpi_4G)
        os.remove(path=file_path)

    required_cols = [
        "MV_4G Data Volume_GB",
        "RRC Setup Success Rate [CDBH]",
        "ERAB Setup Success Rate [CDBH]",
        "PS Drop Call Rate % [CDBH]",
        "DL User Throughput_Kbps [CDBH]",
        "MV_Average number of used DL PRBs [CDBH]",
        "MV_UL User Throughput_Kbps [CDBH]",
        "MV_PS handover success rate [LTE Intra System] [CDBH]",
        "PS handover success rate [LTE Inter System] [CDBH]",
        "MV_LTE_AVG_CQI",
        "VoLTE Call Setup Success rate [CBBH]",
        "MV_VoLTE DCR [CBBH]",
        "VoLTE IntraF HOSR Exec [CBBH]",
        "VoLTE InterF HOSR Exec [CBBH]",
        "MV_VoLTE Packet Loss DL [CBBH]",
        "MV_VoLTE Packet Loss UL [CBBH]",
        "Radio NW Availability",
        "UL RSSI [CDBH]",
        "TA Sample<500 M_Nom",
        "TA Sample<1 KM_Nom",
        "TA Sample >1.5 KM_Nom_Ericsson_1",
        "TA Sample >3.4 KM_Nom",
        "TA Sample >4.45 KM_Nom",
        "TA Sample >5.7 KM_Nom",
        "TA Sample >7.2 KM_Nom",
        "TA Sample >8.4 KM_Nom",
    ]
    sts, response = required_col_check(raw_kpi_4G, required_cols)
    if sts:
        return Response(response)

    site_list_4G = (
        request.FILES["site_list_4G"] if "site_list_4G" in request.FILES else None
    )
    df_site_list_4G = pd.DataFrame()

    if site_list_4G:
        location = MEDIA_ROOT + r"\trends\temporary_files"
        fs = FileSystemStorage(location=location)
        file = fs.save(site_list_4G.name, site_list_4G)
        file_site_path = fs.path(file)
        df_site_list_4G = pd.read_excel(file_site_path)
        print(df_site_list_4G)
        os.remove(path=file_site_path)
        # door_root= os.path.join(MEDIA_ROOT,'trends',"ap",'ap4G')
    response, site_list = site_list_handler(request)
    # if site_list:
    #     sts, response = required_col_check(raw_kpi_4G, required_cols)
    #     if sts:
    #         return Response(response)

    door_path = os.path.join(MEDIA_ROOT, "trends", "ap", "ap4G")

    kpii = [
        "TA Sample<500 M_Nom",
        "TA Sample<1 KM_Nom",
        "TA Sample >1.5 KM_Nom_Ericsson_1",
        "TA Sample >3.4 KM_Nom",
        "TA Sample >4.45 KM_Nom",
        "TA Sample >5.7 KM_Nom",
        "TA Sample >7.2 KM_Nom",
        "TA Sample >8.4 KM_Nom",
    ]

    kpi = [
        "RRC Setup Success Rate [CDBH]",
        "ERAB Setup Success Rate [CDBH]",
        "PS Drop Call Rate % [CDBH]",
        "DL User Throughput_Kbps [CDBH]",
        "MV_Average number of used DL PRBs [CDBH]",
        "MV_UL User Throughput_Kbps [CDBH]",
        "MV_PS handover success rate [LTE Intra System] [CDBH]",
        "PS handover success rate [LTE Inter System] [CDBH]",
        "MV_LTE_AVG_CQI",
        "VoLTE Call Setup Success rate [CBBH]",
        "MV_VoLTE DCR [CBBH]",
        "VoLTE IntraF HOSR Exec [CBBH]",
        "VoLTE InterF HOSR Exec [CBBH]",
        "MV_VoLTE Packet Loss DL [CBBH]",
        "MV_VoLTE Packet Loss UL [CBBH]",
        "Radio NW Availability",
        "UL RSSI [CDBH]",
        "MV_4G Data Volume_GB",
    ]

    # site_list = list(df_site_list_4G["2G ID"])

    df_raw_kpi_4G["Short name"].fillna(inplace=True, method="ffill")
    df_raw_kpi_4G.rename(columns={"Unnamed: 1": "date"}, inplace=True)

    unique_dates = pd.to_datetime(df_raw_kpi_4G["date"]).unique()

    # unique_dates = [np.datetime_as_string(val, unit="D") for val in unique_dates]

    print("", df_raw_kpi_4G)

    df_raw_kpi_4G["DL User Throughput_Kbps [CDBH]"] = (
        df_raw_kpi_4G["DL User Throughput_Kbps [CDBH]"] / 1024
    )
    df_raw_kpi_4G.rename(
        columns={"DL User Throughput_Kbps [CDBH]": "DL User Throughput_Mbps [CDBH]"},
        inplace=True,
    )

    df_raw_kpi_4G["MV_UL User Throughput_Kbps [CDBH]"] = (
        df_raw_kpi_4G["MV_UL User Throughput_Kbps [CDBH]"] / 1024
    )
    df_raw_kpi_4G.rename(
        columns={
            "MV_UL User Throughput_Kbps [CDBH]": "MV_UL User Throughput_Mbps [CDBH]"
        },
        inplace=True,
    )

    print(df_raw_kpi_4G)
    print(df_raw_kpi_4G.columns)

    for i in range(len(kpi)):
        if kpi[i] == "DL User Throughput_Kbps [CDBH]":
            kpi[i] = "DL User Throughput_Mbps [CDBH]"
        if kpi[i] == "MV_UL User Throughput_Kbps [CDBH]":
            kpi[i] = "MV_UL User Throughput_Mbps [CDBH]"

    lis = list(df_raw_kpi_4G["Short name"])
    techlist = []
    sit_id_lis = []
    # cell_id_lis=[]
    for item in lis:
        if "_" in item:

            sit_id = item.split("_")[-2][:-1]
            sit_id_lis.append(sit_id)

        else:
            # cell_id=item
            sit_id = item[1:-1]
            # cell_id_lis.append(cell_id)
            sit_id_lis.append(sit_id)

        # print(cell_id_lis)
        if "F3" in item or "F8" in item or "T1" in item or "T2" in item:
            if "F3" in item:
                tech = "L1800"
                # techlist.append(tech)

            if "F8" in item:
                tech = "L900"
                # techlist.append(tech)

            if "T2" in item or "T1" in item:
                tech = "TDD"
            techlist.append(tech)

        else:
            tech = item
            techlist.append(tech)

    df_raw_kpi_4G.insert(2, "SITE_ID", sit_id_lis)
    df_raw_kpi_4G.insert(3, "tech", techlist)

    df_raw_kpi_4G.rename(columns={"Short name": "Shortname"}, inplace=True)
    # df_all_tech_kpi.rename(columns={"site id" :"site_id" } ,inplace = True )
    df_raw_kpi_4G.fillna(value=0, inplace=True)
    message = site_comparision(
        list(df_raw_kpi_4G["SITE_ID"]), list(site_list)
    ) 
       ##getting F3, F1, F8, T1, T2 dfs seprately  ######################
       
    df_F3= df_raw_kpi_4G[df_raw_kpi_4G['Shortname'].str.contains("F3", na=False)]
    df_F1=df_raw_kpi_4G[df_raw_kpi_4G['Shortname'].str.contains("F1", na=False)]
    df_F8=df_raw_kpi_4G[df_raw_kpi_4G['Shortname'].str.contains("F8", na=False)]
    df_T1=df_raw_kpi_4G[df_raw_kpi_4G['Shortname'].str.contains("T1", na=False)]
    df_T2=df_raw_kpi_4G[df_raw_kpi_4G['Shortname'].str.contains("T2", na=False)]
 
    # PsOs_pathF3 = os.path.join(door_path, "process_output", "df_F3.xlsx")
    # df_F3.to_excel(PsOs_pathF3)
    # PsOs_pathF1 = os.path.join(door_path, "process_output", "df_F1.xlsx")
    # df_F1.to_excel(PsOs_pathF1)
    # PsOs_pathF8 = os.path.join(door_path, "process_output", "df_F8.xlsx")
    # df_F8.to_excel(PsOs_pathF8)
    # PsOs_pathT1 = os.path.join(door_path, "process_output", "df_T1.xlsx")
    # df_T1.to_excel(PsOs_pathT1)
    # PsOs_pathT2 = os.path.join(door_path, "process_output", "df_T2.xlsx")
    # df_T2.to_excel(PsOs_pathT2)
    
    print("df_f3",df_F3)
    print("df_f1",df_F1)
    print("df_f8",df_F8)
    print("df_t1",df_T1)
    print("df_t2",df_T2)
    
    
    ####################################################################
   
    
    
    # site_comparision_calling
    PsOs_path1 = os.path.join(door_path, "process_output", "desired_input.xlsx")
    df_raw_kpi_4G.to_excel(PsOs_path1)
    
 
   

    str_date = request.POST.get("offered_date")
    date1 = datetime.datetime.strptime(str_date, "%Y-%m-%d")
    # dt1 = date1 - timedelta(1)
    # dt2 = date1 - timedelta(2)
    # dt3 = date1 - timedelta(3)
    # dt4 = date1 - timedelta(4)
    # dt5 = date1 - timedelta(5)
    dt1 = unique_dates[0]
    dt2 = unique_dates[1]
    dt3 = unique_dates[2]
    dt4 = unique_dates[3]
    dt5 = unique_dates[4]
    ls = [dt1, dt2, dt3, dt4, dt5]

    def perticular_tech(tech, site_list):

        print("df_new_kpi;- ", df_raw_kpi_4G)
        print("site_list;- ", site_list)
        print(list(df_raw_kpi_4G["SITE_ID"]))
        df_filtered = df_raw_kpi_4G[
            (df_raw_kpi_4G.SITE_ID.isin(site_list))
            & (df_raw_kpi_4G.date.isin(ls))
            & (df_raw_kpi_4G.Shortname.str.contains("|".join(tech)))
        ]
        print("df_filtered: ",df_filtered)
        df_filtered.fillna(value=0, inplace=True)
        PsOs_filter = os.path.join(
            door_path, "process_output", "last_filtered_input.xlsx"
        )
        df_filtered.to_excel(PsOs_filter)

        print(df_filtered)
        
        df_filtered.fillna(value=0, inplace=True)

        df_pivoted = df_filtered.pivot_table(
            index=["SITE_ID", "Shortname", "tech"], columns="date",aggfunc=lambda x: x.mean() if pd.api.types.is_numeric_dtype(x) else ''
        )

        print(df_pivoted)
        save_name = str(tech) + "_pivot.xlsx"
        print("savename", tech)
        PsOs_pivot = os.path.join(door_path, "process_output", save_name)

        df_pivoted.to_excel(PsOs_pivot)
        return df_filtered, df_pivoted

    alpha = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    def num_hash(num):
        if num < 26:
            return alpha[num - 1]
        else:
            q, r = num // 26, num % 26
            if r == 0:
                if q == 1:
                    return alpha[r - 1]
                else:
                    return num_hash(q - 1) + alpha[r - 1]
            else:
                return num_hash(q) + alpha[r - 1]

    # Driver code

    # printString(27906)

    def titleToNumber(s):
        # This process is similar to binary-to-
        # decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord("A") + 1
        return result

    def overwrite(df_pivoted, kpi_name, coln1, trend_ws):
        coln2 = num_hash(titleToNumber(coln1) + 1)
        coln3 = num_hash(titleToNumber(coln1) + 2)
        coln4 = num_hash(titleToNumber(coln1) + 3)
        coln5 = num_hash(titleToNumber(coln1) + 4)
        print(kpi_name)
        index_pivot = df_pivoted.index.to_list()
        print("index ;###############################", index_pivot)
        print(len(index_pivot))
        print("index of pivoted table: ", index_pivot)
        dr = df_pivoted[kpi_name]
        print("columns of dr table", dr.columns)
        cl = dr.columns.to_list()
        print("column list", cl)

        # site_id=dr["SITE_ID"].to_list()
        # cell_id=dr["CELL_ID"].to_list()
        col1 = dr[str(cl[0])].to_list()
        col2 = dr[str(cl[1])].to_list()
        col3 = dr[str(cl[2])].to_list()
        col4 = dr[str(cl[3])].to_list()
        col5 = dr[str(cl[4])].to_list()

        trend_ws[coln1 + "2"].value = cl[0]
        trend_ws[coln2 + "2"].value = cl[1]
        trend_ws[coln3 + "2"].value = cl[2]
        trend_ws[coln4 + "2"].value = cl[3]
        trend_ws[coln5 + "2"].value = cl[4]

        # me=column_index_from_string(coln5)+1
        # me=get_column_letter(me)
        for i, value in enumerate(index_pivot):
            j = i + 3
            trend_ws["B" + str(j)].value = index_pivot[i][0]
            trend_ws["A" + str(j)].value = index_pivot[i][1]

            trend_ws[coln1 + str(j)].value = col1[i]
            trend_ws[coln2 + str(j)].value = col2[i]
            trend_ws[coln3 + str(j)].value = col3[i]
            trend_ws[coln4 + str(j)].value = col4[i]
            trend_ws[coln5 + str(j)].value = col5[i]

    def overwrite1(df_pivoted, kpi_name, coln1, trend_ws):
        print(kpi_name)
        index_pivot = df_pivoted.index.to_list()
        print("index of pivoted table: ", index_pivot)
        dr = df_pivoted[kpi_name]
        print("columns of dr table", dr.columns)
        cl = dr.columns.to_list()
        print("column list", cl)
        col1 = dr[str(cl[0])].to_list()
        trend_ws[coln1 + "2"].value = cl[0]

        for i, value in enumerate(index_pivot):
            j = i + 3
            trend_ws[coln1 + str(j)].value = col1[i]

    def overwrite2(df_pivoted, kpi_name, coln1, trend_ws):
        print(kpi_name)
        index_pivot = df_pivoted.index.to_list()
        print("index of pivoted table: ", index_pivot)
        dr = df_pivoted[kpi_name]
        print("columns of dr table", dr.columns)
        cl = dr.columns.to_list()
        print("column list", cl)
        col1 = dr[str(cl[1])].to_list()
        trend_ws[coln1 + "2"].value = cl[1]

        for i, value in enumerate(index_pivot):
            j = i + 3
            trend_ws[coln1 + str(j)].value = col1[i]

    def overwrite3(df_pivoted, kpi_name, coln1, trend_ws):
        print(kpi_name)
        index_pivot = df_pivoted.index.to_list()
        print("index of pivoted table: ", index_pivot)
        dr = df_pivoted[kpi_name]
        print("columns of dr table", dr.columns)
        cl = dr.columns.to_list()
        print("column list", cl)
        col1 = dr[str(cl[2])].to_list()
        trend_ws[coln1 + "2"].value = cl[2]

        for i, value in enumerate(index_pivot):
            j = i + 3
            trend_ws[coln1 + str(j)].value = col1[i]

    # for fdd
    if not df_F3.empty:
        print("Enters f3 category")
        pivot_fdd = perticular_tech(["F3"], site_list)[1]
        PsOs_blnk_temp = os.path.join(door_path, "templates", "L1800_KPI_Submission.xlsx")
        # trend_wb_L1800=load_workbook(path_of_blnk_temp)
        path_of_blnk_temp = PsOs_blnk_temp
        trend_wb_L1800 = openpyxl.load_workbook(path_of_blnk_temp)
        print(trend_wb_L1800.sheetnames)
    
        trend_ws = trend_wb_L1800["LTE-VOLTE KPI"]
        for ta in kpii:
            if ta == "TA Sample<500 M_Nom":
                overwrite1(pivot_fdd, ta, "DG", trend_ws)

            if ta == "TA Sample<1 KM_Nom":
                overwrite1(pivot_fdd, ta, "DH", trend_ws)

            if ta == "TA Sample >1.5 KM_Nom_Ericsson_1":
                overwrite1(pivot_fdd, ta, "DI", trend_ws)

            if ta == "TA Sample >3.4 KM_Nom":
                overwrite1(pivot_fdd, ta, "DJ", trend_ws)

            if ta == "TA Sample >4.45 KM_Nom":
                overwrite1(pivot_fdd, ta, "DK", trend_ws)
            if ta == "TA Sample >5.7 KM_Nom":
                overwrite1(pivot_fdd, ta, "DL", trend_ws)

            if ta == "TA Sample >7.2 KM_Nom":
                overwrite1(pivot_fdd, ta, "DM", trend_ws)

            if ta == "TA Sample >8.4 KM_Nom":
                overwrite1(pivot_fdd, ta, "DN", trend_ws)
        for ta in kpii:
            if ta == "TA Sample<500 M_Nom":
                overwrite2(pivot_fdd, ta, "DR", trend_ws)

            if ta == "TA Sample<1 KM_Nom":
                overwrite2(pivot_fdd, ta, "DS", trend_ws)

            if ta == "TA Sample >1.5 KM_Nom_Ericsson_1":
                overwrite2(pivot_fdd, ta, "DT", trend_ws)

            if ta == "TA Sample >3.4 KM_Nom":
                overwrite2(pivot_fdd, ta, "DU", trend_ws)

            if ta == "TA Sample >4.45 KM_Nom":
                overwrite2(pivot_fdd, ta, "DV", trend_ws)
            if ta == "TA Sample >5.7 KM_Nom":
                overwrite2(pivot_fdd, ta, "DW", trend_ws)

            if ta == "TA Sample >7.2 KM_Nom":
                overwrite2(pivot_fdd, ta, "DX", trend_ws)

            if ta == "TA Sample >8.4 KM_Nom":
                overwrite2(pivot_fdd, ta, "DY", trend_ws)
        for ta in kpii:
            if ta == "TA Sample<500 M_Nom":
                overwrite3(pivot_fdd, ta, "EC", trend_ws)

            if ta == "TA Sample<1 KM_Nom":
                overwrite3(pivot_fdd, ta, "ED", trend_ws)

            if ta == "TA Sample >1.5 KM_Nom_Ericsson_1":
                overwrite3(pivot_fdd, ta, "EE", trend_ws)

            if ta == "TA Sample >3.4 KM_Nom":
                overwrite3(pivot_fdd, ta, "EF", trend_ws)

            if ta == "TA Sample >4.45 KM_Nom":
                overwrite3(pivot_fdd, ta, "EG", trend_ws)
            if ta == "TA Sample >5.7 KM_Nom":
                overwrite3(pivot_fdd, ta, "EH", trend_ws)

            if ta == "TA Sample >7.2 KM_Nom":
                overwrite3(pivot_fdd, ta, "EI", trend_ws)

            if ta == "TA Sample >8.4 KM_Nom":
                overwrite3(pivot_fdd, ta, "EJ", trend_ws)
        for kpi_name in kpi:
            if kpi_name == "RRC Setup Success Rate [CDBH]":
                overwrite(pivot_fdd, kpi_name, "E", trend_ws)

            if kpi_name == "ERAB Setup Success Rate [CDBH]":
                overwrite(pivot_fdd, kpi_name, "K", trend_ws)

            if kpi_name == "PS Drop Call Rate % [CDBH]":
                overwrite(pivot_fdd, kpi_name, "Q", trend_ws)

            if kpi_name == "DL User Throughput_Mbps [CDBH]":
                overwrite(pivot_fdd, kpi_name, "W", trend_ws)

            if kpi_name == "MV_Average number of used DL PRBs [CDBH]":
                overwrite(pivot_fdd, kpi_name, "AC", trend_ws)

            if kpi_name == "MV_UL User Throughput_Mbps [CDBH]":
                overwrite(pivot_fdd, kpi_name, "AI", trend_ws)

            if kpi_name == "MV_PS handover success rate [LTE Intra System] [CDBH]":
                overwrite(pivot_fdd, kpi_name, "AN", trend_ws)

            if kpi_name == "PS handover success rate [LTE Inter System] [CDBH]":
                overwrite(pivot_fdd, kpi_name, "AT", trend_ws)

            if kpi_name == "MV_LTE_AVG_CQI":
                overwrite(pivot_fdd, kpi_name, "AZ", trend_ws)

            if kpi_name == "VoLTE Call Setup Success rate [CBBH]":
                overwrite(pivot_fdd, kpi_name, "BE", trend_ws)

            if kpi_name == "MV_VoLTE DCR [CBBH]":
                overwrite(pivot_fdd, kpi_name, "BK", trend_ws)

            if kpi_name == "VoLTE IntraF HOSR Exec [CBBH]":
                overwrite(pivot_fdd, kpi_name, "BQ", trend_ws)

            if kpi_name == "VoLTE InterF HOSR Exec [CBBH]":
                overwrite(pivot_fdd, kpi_name, "BW", trend_ws)

            if kpi_name == "MV_VoLTE Packet Loss DL [CBBH]":
                overwrite(pivot_fdd, kpi_name, "CC", trend_ws)

            if kpi_name == "MV_VoLTE Packet Loss UL [CBBH]":
                overwrite(pivot_fdd, kpi_name, "CI", trend_ws)

            if kpi_name == "Radio NW Availability":
                overwrite(pivot_fdd, kpi_name, "CO", trend_ws)

            if kpi_name == "UL RSSI [CDBH]":
                overwrite(pivot_fdd, kpi_name, "CU", trend_ws)

            if kpi_name == "MV_4G Data Volume_GB":
                overwrite(pivot_fdd, kpi_name, "DA", trend_ws)
        
        save_output_L18 = os.path.join(
            door_path, "output", "toBeZipped", "out_L1800_KPI_Submission.xlsx"
        )
        trend_wb_L1800.save(save_output_L18)

    # for fdd
    if not df_F8.empty:
        pivot_fdd = perticular_tech(["F8"], site_list)[1]
        print("enters the f8 category")
        PsOs_blnk_temp = os.path.join(door_path, "templates", "L900_KPI_Submission.xlsx")
        path_of_blnk_temp = PsOs_blnk_temp
        # trend_wb_L900=xl.load_workbook(path_of_blnk_temp)
        trend_wb_L900 = openpyxl.load_workbook(path_of_blnk_temp)  # added xl.
        trend_ws = trend_wb_L900["L900 KPI"]
        for ta in kpii:
            if ta == "TA Sample<500 M_Nom":
                overwrite1(pivot_fdd, ta, "DG", trend_ws)

            if ta == "TA Sample<1 KM_Nom":
                overwrite1(pivot_fdd, ta, "DH", trend_ws)

            if ta == "TA Sample >1.5 KM_Nom_Ericsson_1":
                overwrite1(pivot_fdd, ta, "DI", trend_ws)

            if ta == "TA Sample >3.4 KM_Nom":
                overwrite1(pivot_fdd, ta, "DJ", trend_ws)

            if ta == "TA Sample >4.45 KM_Nom":
                overwrite1(pivot_fdd, ta, "DK", trend_ws)
            if ta == "TA Sample >5.7 KM_Nom":
                overwrite1(pivot_fdd, ta, "DL", trend_ws)

            if ta == "TA Sample >7.2 KM_Nom":
                overwrite1(pivot_fdd, ta, "DM", trend_ws)

            if ta == "TA Sample >8.4 KM_Nom":
                overwrite1(pivot_fdd, ta, "DN", trend_ws)
        for ta in kpii:
            if ta == "TA Sample<500 M_Nom":
                overwrite2(pivot_fdd, ta, "DR", trend_ws)

            if ta == "TA Sample<1 KM_Nom":
                overwrite2(pivot_fdd, ta, "DS", trend_ws)

            if ta == "TA Sample >1.5 KM_Nom_Ericsson_1":
                overwrite2(pivot_fdd, ta, "DT", trend_ws)

            if ta == "TA Sample >3.4 KM_Nom":
                overwrite2(pivot_fdd, ta, "DU", trend_ws)

            if ta == "TA Sample >4.45 KM_Nom":
                overwrite2(pivot_fdd, ta, "DV", trend_ws)
            if ta == "TA Sample >5.7 KM_Nom":
                overwrite2(pivot_fdd, ta, "DW", trend_ws)

            if ta == "TA Sample >7.2 KM_Nom":
                overwrite2(pivot_fdd, ta, "DX", trend_ws)

            if ta == "TA Sample >8.4 KM_Nom":
                overwrite2(pivot_fdd, ta, "DY", trend_ws)
        for ta in kpii:
            if ta == "TA Sample<500 M_Nom":
                overwrite3(pivot_fdd, ta, "EC", trend_ws)

            if ta == "TA Sample<1 KM_Nom":
                overwrite3(pivot_fdd, ta, "ED", trend_ws)

            if ta == "TA Sample >1.5 KM_Nom_Ericsson_1":
                overwrite3(pivot_fdd, ta, "EE", trend_ws)

            if ta == "TA Sample >3.4 KM_Nom":
                overwrite3(pivot_fdd, ta, "EF", trend_ws)

            if ta == "TA Sample >4.45 KM_Nom":
                overwrite3(pivot_fdd, ta, "EG", trend_ws)
            if ta == "TA Sample >5.7 KM_Nom":
                overwrite3(pivot_fdd, ta, "EH", trend_ws)

            if ta == "TA Sample >7.2 KM_Nom":
                overwrite3(pivot_fdd, ta, "EI", trend_ws)

            if ta == "TA Sample >8.4 KM_Nom":
                overwrite3(pivot_fdd, ta, "EJ", trend_ws)
        for kpi_name in kpi:
            if kpi_name == "RRC Setup Success Rate [CDBH]":
                overwrite(pivot_fdd, kpi_name, "E", trend_ws)

            if kpi_name == "ERAB Setup Success Rate [CDBH]":
                overwrite(pivot_fdd, kpi_name, "K", trend_ws)

            if kpi_name == "PS Drop Call Rate % [CDBH]":
                overwrite(pivot_fdd, kpi_name, "Q", trend_ws)

            if kpi_name == "DL User Throughput_Mbps [CDBH]":
                overwrite(pivot_fdd, kpi_name, "W", trend_ws)

            if kpi_name == "MV_Average number of used DL PRBs [CDBH]":
                overwrite(pivot_fdd, kpi_name, "AC", trend_ws)

            if kpi_name == "MV_UL User Throughput_Mbps [CDBH]":
                overwrite(pivot_fdd, kpi_name, "AI", trend_ws)

            if kpi_name == "MV_PS handover success rate [LTE Intra System] [CDBH]":
                overwrite(pivot_fdd, kpi_name, "AN", trend_ws)

            if kpi_name == "PS handover success rate [LTE Inter System] [CDBH]":
                overwrite(pivot_fdd, kpi_name, "AT", trend_ws)

            if kpi_name == "MV_LTE_AVG_CQI":
                overwrite(pivot_fdd, kpi_name, "AZ", trend_ws)

            if kpi_name == "VoLTE Call Setup Success rate [CBBH]":
                overwrite(pivot_fdd, kpi_name, "BE", trend_ws)

            if kpi_name == "MV_VoLTE DCR [CBBH]":
                overwrite(pivot_fdd, kpi_name, "BK", trend_ws)

            if kpi_name == "VoLTE IntraF HOSR Exec [CBBH]":
                overwrite(pivot_fdd, kpi_name, "BQ", trend_ws)

            if kpi_name == "VoLTE InterF HOSR Exec [CBBH]":
                overwrite(pivot_fdd, kpi_name, "BW", trend_ws)

            if kpi_name == "MV_VoLTE Packet Loss DL [CBBH]":
                overwrite(pivot_fdd, kpi_name, "CC", trend_ws)

            if kpi_name == "MV_VoLTE Packet Loss UL [CBBH]":
                overwrite(pivot_fdd, kpi_name, "CI", trend_ws)

            if kpi_name == "Radio NW Availability":
                overwrite(pivot_fdd, kpi_name, "CO", trend_ws)

            if kpi_name == "UL RSSI [CDBH]":
                overwrite(pivot_fdd, kpi_name, "CU", trend_ws)

            if kpi_name == "MV_4G Data Volume_GB":
                overwrite(pivot_fdd, kpi_name, "DA", trend_ws)
                
        save_output_L900 = os.path.join(
            door_path, "output", "toBeZipped", "out_L900_KPI_Submission.xlsx"
        )

        trend_wb_L900.save(save_output_L900)

    # for tdd
    if not df_T1.empty or not df_T2.empty:
        print("enters T1 AND T2 category")
        pivot_fdd = perticular_tech(["T1", "T2"], site_list)[1]
        PsOs_blnk_temp = os.path.join(door_path, "templates", "TDD_KPI_Submission.xlsx")
        # trend_wb_TDD=load_workbook(path_of_blnk_temp)
        path_of_blnk_temp = PsOs_blnk_temp
        trend_wb_TDD = openpyxl.load_workbook(path_of_blnk_temp)  # added xl.
        # print(trend_wb.sheetnames)
        trend_ws = trend_wb_TDD["LTE-VOLTE KPI"]
        for ta in kpii:
            if ta == "TA Sample<500 M_Nom":
                overwrite1(pivot_fdd, ta, "DG", trend_ws)

            if ta == "TA Sample<1 KM_Nom":
                overwrite1(pivot_fdd, ta, "DH", trend_ws)

            if ta == "TA Sample >1.5 KM_Nom_Ericsson_1":
                overwrite1(pivot_fdd, ta, "DI", trend_ws)

            if ta == "TA Sample >3.4 KM_Nom":
                overwrite1(pivot_fdd, ta, "DJ", trend_ws)

            if ta == "TA Sample >4.45 KM_Nom":
                overwrite1(pivot_fdd, ta, "DK", trend_ws)
            if ta == "TA Sample >5.7 KM_Nom":
                overwrite1(pivot_fdd, ta, "DL", trend_ws)

            if ta == "TA Sample >7.2 KM_Nom":
                overwrite1(pivot_fdd, ta, "DM", trend_ws)

            if ta == "TA Sample >8.4 KM_Nom":
                overwrite1(pivot_fdd, ta, "DN", trend_ws)
        for ta in kpii:
            if ta == "TA Sample<500 M_Nom":
                overwrite2(pivot_fdd, ta, "DR", trend_ws)

            if ta == "TA Sample<1 KM_Nom":
                overwrite2(pivot_fdd, ta, "DS", trend_ws)

            if ta == "TA Sample >1.5 KM_Nom_Ericsson_1":
                overwrite2(pivot_fdd, ta, "DT", trend_ws)

            if ta == "TA Sample >3.4 KM_Nom":
                overwrite2(pivot_fdd, ta, "DU", trend_ws)

            if ta == "TA Sample >4.45 KM_Nom":
                overwrite2(pivot_fdd, ta, "DV", trend_ws)
            if ta == "TA Sample >5.7 KM_Nom":
                overwrite2(pivot_fdd, ta, "DW", trend_ws)

            if ta == "TA Sample >7.2 KM_Nom":
                overwrite2(pivot_fdd, ta, "DX", trend_ws)

            if ta == "TA Sample >8.4 KM_Nom":
                overwrite2(pivot_fdd, ta, "DY", trend_ws)
        for ta in kpii:
            if ta == "TA Sample<500 M_Nom":
                overwrite3(pivot_fdd, ta, "EC", trend_ws)

            if ta == "TA Sample<1 KM_Nom":
                overwrite3(pivot_fdd, ta, "ED", trend_ws)

            if ta == "TA Sample >1.5 KM_Nom_Ericsson_1":
                overwrite3(pivot_fdd, ta, "EE", trend_ws)

            if ta == "TA Sample >3.4 KM_Nom":
                overwrite3(pivot_fdd, ta, "EF", trend_ws)

            if ta == "TA Sample >4.45 KM_Nom":
                overwrite3(pivot_fdd, ta, "EG", trend_ws)
            if ta == "TA Sample >5.7 KM_Nom":
                overwrite3(pivot_fdd, ta, "EH", trend_ws)

            if ta == "TA Sample >7.2 KM_Nom":
                overwrite3(pivot_fdd, ta, "EI", trend_ws)

            if ta == "TA Sample >8.4 KM_Nom":
                overwrite3(pivot_fdd, ta, "EJ", trend_ws)
        for kpi_name in kpi:
            if kpi_name == "RRC Setup Success Rate [CDBH]":
                overwrite(pivot_fdd, kpi_name, "E", trend_ws)

            if kpi_name == "ERAB Setup Success Rate [CDBH]":
                overwrite(pivot_fdd, kpi_name, "K", trend_ws)

            if kpi_name == "PS Drop Call Rate % [CDBH]":
                overwrite(pivot_fdd, kpi_name, "Q", trend_ws)

            if kpi_name == "DL User Throughput_Mbps [CDBH]":
                overwrite(pivot_fdd, kpi_name, "W", trend_ws)

            if kpi_name == "MV_Average number of used DL PRBs [CDBH]":
                overwrite(pivot_fdd, kpi_name, "AC", trend_ws)

            if kpi_name == "MV_UL User Throughput_Mbps [CDBH]":
                overwrite(pivot_fdd, kpi_name, "AI", trend_ws)

            if kpi_name == "MV_PS handover success rate [LTE Intra System] [CDBH]":
                overwrite(pivot_fdd, kpi_name, "AN", trend_ws)

            if kpi_name == "PS handover success rate [LTE Inter System] [CDBH]":
                overwrite(pivot_fdd, kpi_name, "AT", trend_ws)

            if kpi_name == "MV_LTE_AVG_CQI":
                overwrite(pivot_fdd, kpi_name, "AZ", trend_ws)

            if kpi_name == "VoLTE Call Setup Success rate [CBBH]":
                overwrite(pivot_fdd, kpi_name, "BE", trend_ws)

            if kpi_name == "MV_VoLTE DCR [CBBH]":
                overwrite(pivot_fdd, kpi_name, "BK", trend_ws)

            if kpi_name == "VoLTE IntraF HOSR Exec [CBBH]":
                overwrite(pivot_fdd, kpi_name, "BQ", trend_ws)

            if kpi_name == "VoLTE InterF HOSR Exec [CBBH]":
                overwrite(pivot_fdd, kpi_name, "BW", trend_ws)

            if kpi_name == "MV_VoLTE Packet Loss DL [CBBH]":
                overwrite(pivot_fdd, kpi_name, "CC", trend_ws)

            if kpi_name == "MV_VoLTE Packet Loss UL [CBBH]":
                overwrite(pivot_fdd, kpi_name, "CI", trend_ws)

            if kpi_name == "Radio NW Availability":
                overwrite(pivot_fdd, kpi_name, "CO", trend_ws)

            if kpi_name == "UL RSSI [CDBH]":
                overwrite(pivot_fdd, kpi_name, "CU", trend_ws)

            if kpi_name == "MV_4G Data Volume_GB":
                overwrite(pivot_fdd, kpi_name, "DA", trend_ws)
        save_output_tdd = os.path.join(
            door_path, "output", "toBeZipped", "out_TDD_KPI_Submission.xlsx"
        )
        trend_wb_TDD.save(save_output_tdd)
        
        # save_output_L900 = os.path.join(
        #     door_path, "output", "toBeZipped", "out_L900_KPI_Submission.xlsx"
        # )

        # trend_wb_L900.save(save_output_L900)
        # save_output_L18 = os.path.join(
        #     door_path, "output", "toBeZipped", "out_L1800_KPI_Submission.xlsx"
        # )
        # trend_wb_L1800.save(save_output_L18)
        # save_output_tdd = os.path.join(
        #     door_path, "output", "toBeZipped", "out_TDD_KPI_Submission.xlsx"
        # )
        # trend_wb_TDD.save(save_output_tdd)
        print("successfully")

    def get_all_file_paths(directory):

        file_paths = []

        for root, directories, files in os.walk(directory):
            for filename in files:

                filepath = os.path.join(root, filename)
                file_paths.append(filepath)

        return file_paths

    def main1():
        directory = os.path.join(door_path, "output", "toBeZipped")
        file_paths = get_all_file_paths(directory)
        print("Following files will be zipped:")
        for file_name in file_paths:
            print(file_name)

        with ZipFile(
            os.path.join(door_path, "output", "output", "Ap_output.zip"), "w"
        ) as zip:
            # writing each file one by one
            for file in file_paths:
                zip.write(file)

        print("All files zipped successfully!")

    main1()
    # download_path1=os.path.join(MEDIA_URL,'trends','ap',"ap4G",'output','22april_L900_KPI_Submission_01Mar.xlsx')
    # download_path2=os.path.join(MEDIA_URL,'trends','ap',"ap4G",'output','22april_L1800_KPI_Submission_01Mar.xlsx')
    # download_path3=os.path.join(MEDIA_URL,'trends','ap',"ap4G",'output','22april_TDD_KPI_Submission_01Mar.xlsx')
    download_path = os.path.join(
        MEDIA_URL, "trends", "ap", "ap4G", "output", "output", "Ap_output.zip"
    )
    # return Response({'status':True,'Download_url':download_path1,'Download_url2':download_path2,'Download_url3':download_path3})
    return Response(
        {
            "status": True,
            "message": "successfully",
            "missing_sites": message,
            "Download_url": download_path,
        }
    )


#################################################### GSM ###########################################
@api_view(["POST"])
def old_ap2G_trend(request):
    raw_kpi_2G = request.FILES["raw_kpi_2G"] if "raw_kpi_2G" in request.FILES else None
    print(raw_kpi_2G)
    if raw_kpi_2G:
        location = MEDIA_ROOT + r"\trends\temporary_files"
        fs = FileSystemStorage(location=location)
        file_2G = fs.save(raw_kpi_2G.name, raw_kpi_2G)
        file_2G_path = fs.path(file_2G)
        df_raw_kpi_2G = pd.read_excel(file_2G_path)

        print(df_raw_kpi_2G)
        os.remove(path=file_2G_path)

    required_cols = [
        "SDCCH Drop Call Rate",
        "SDCCH Blocking Rate",
        "TCH Blocking Rate",
        "Drop Call Rate",
        "RX Quality [BBH]",
        "Handover Success Rate",
        "Handover Success Rate_Denom",
        "Handover Success Rate_Nom",
        "TCH Assignment Success Rate",
        "Network availability [RNA]",
        "Total Voice Traffic",
    ]

    sts, response = required_col_check(raw_kpi_2G, required_cols)
    if sts:
        return Response(response)

    site_list_2G = (
        request.FILES["site_list_2G"] if "site_list_2G" in request.FILES else None
    )
    df_2G_site = pd.DataFrame()
    if site_list_2G:
        location = MEDIA_ROOT + r"\trends\temporary_files"
        fs = FileSystemStorage(location=location)
        file_2G_sheet = fs.save(site_list_2G.name, site_list_2G)
        file_2G_path = fs.path(file_2G_sheet)
        df_2G_site = pd.read_excel(file_2G_path)
        print(df_2G_site)
        os.remove(path=file_2G_path)

    # if site_list_2G:
    #     sts, response = required_col_check(raw_kpi_2G, required_cols)
    #     if sts:
    #         return Response(response)

    response, site_list_2G = site_list_handler(request)

    door_path = os.path.join(MEDIA_ROOT, "trends", "ap", "ap2G")
    # save_path=os.path.join(door_path,'input','GSM_AP.xlsx')
    # df_raw_kpi_2G = pd.read_excel(save_path)
    STR = os.path.join(door_path, "templates", "G1800_KPI_01-Mar.xlsx")
    wb = openpyxl.load_workbook(STR)
    ws = wb["2G KPI"]
    print(df_raw_kpi_2G)

    df_raw_kpi_2G["Short name"] = df_raw_kpi_2G["Short name"].fillna(
        method="ffill"
    )  #######for (forward fill)ffill is used to copy  and fill

    df_raw_kpi_2G.fillna(value=0, inplace=True)

    print(df_raw_kpi_2G)
    df_raw_kpi_2G.rename(columns={"Unnamed: 0": "Date"}, inplace=True)
    df_raw_kpi_2G.columns.values[1] = "Date"  #####for empty colum

    A = []
    B = []

    unique_dates = pd.to_datetime(df_raw_kpi_2G["Date"]).unique()

    # unique_dates = [np.datetime_as_string(val, unit="D") for val in unique_dates]
    unique_dates = np.sort(unique_dates)
    unique_dates = [str(date)[:10] for date in unique_dates]
    
    print(unique_dates)
    
    # exit(0)    
    new_columns = df_raw_kpi_2G.columns[2:]
    for column in new_columns:
        df_raw_kpi_2G[column] = df_raw_kpi_2G[column].apply(pd.to_numeric, errors='coerce')
        
    df_raw_kpi_2G.fillna(value=0, inplace=True)

    for i, x in enumerate(df_raw_kpi_2G["Short name"]):
        if " " in str(x):
            site_id = x.split(" ")[0]
            A.append(site_id)
        elif "-" in str(x):
            site_id = x.split("-")[2][:-1]
            A.append(site_id)
        else:
            site_id = str(x)[:-1]
            A.append(site_id)
        if " " in str(x):
            cell_id = x.split(" ")[-1][:6]
            B.append(cell_id)
        elif "-" in str(x):
            cell_id = x.split("-")[-1]
            B.append(cell_id)
        else:
            cell_id = str(x)[:7]
            B.append(cell_id)
    df_raw_kpi_2G.insert(0, "SITE_ID", A)
    df_raw_kpi_2G.insert(4, "cell_id", B)
    print("after adding cell name and site id: \n", df_raw_kpi_2G)
    PsOsPath = os.path.join(door_path, "process_output", "final.xlsx")
    df_raw_kpi_2G.to_excel(PsOsPath, index=False)

    excel_file_1 = PsOsPath
    # PsOsPath2=os.path.join(door_path,'input','SITES2G.xlsx')
    # excel_file_2 = PsOsPath2

    df1 = pd.read_excel(excel_file_1)
    # df2 = pd.read_excel(excel_file_2)

    #     # df2 = pd.read_excel(excel_file_2)

    df1.rename(columns={"Site ID": "SITE_ID"}, inplace=True)
    df1 = df1.apply(lambda col: pd.to_datetime(col) if col.dtype == 'datetime64[ns]' else col)

    # df_2G_site.rename(columns={"Site ID": "SITE_ID"}, inplace=True)
    message = site_comparision(A, list(site_list_2G))  # site_co

    g2_kpi = [
        "SDCCH Drop Call Rate [BBH]",
        "SDCCH Blocking Rate [BBH]",
        "TTCH Blocking Rate [BBH]",
        "TCH Drop Call Rate [BBH]",
        "TCH Drop Call Rate_Denom [BBH]",
        "TCH Drop Call Rate_Nom [BBH]",
        "RX Quality [BBH]",
        "Handover Success Rate [BBH]",
        "Handover Success Rate_Denom [BBH]",
        "Handover Success Rate_Nom [BBH]",
        "TCH Assignment Success Rate [BBH]",
        "Network availability [RNA] [BBH]",
    ]

    filtered_df_1 = df1[(df1.SITE_ID.isin(site_list_2G))]

    print(filtered_df_1)
    PsOs_filtr = os.path.join(door_path, "process_output", "filtered_df_1.xlsx")
    filtered_df_1.to_excel(PsOs_filtr, index=False)
    df1 = pd.read_excel(PsOs_filtr)
    
    df1.fillna(value=0, inplace=True)
    df_pivot = df1.pivot_table(columns="Date", index=["SITE_ID", "cell_id"], aggfunc=lambda x: x.mean() if pd.api.types.is_numeric_dtype(x) else '')

    
    # df_pivot.fillna(value=0, inplace=True)
    print("df_pivot", df_pivot)

    PsOs_pivot = os.path.join(door_path, "process_output", "pivot.xlsx")
    df_pivot.to_excel(PsOs_pivot)

    alpha = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    def num_hash(num):
        if num < 26:
            return alpha[num - 1]
        else:
            q, r = num // 26, num % 26
            if r == 0:
                if q == 1:
                    return alpha[r - 1]
                else:
                    return num_hash(q - 1) + alpha[r - 1]
            else:
                return num_hash(q) + alpha[r - 1]

    def titleToNumber(s):
        # This process is similar to binary-to-
        # decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord("A") + 1
        return result

    str_date = request.POST.get("offered_date")
    date1 = datetime.datetime.strptime(str_date, "%Y-%m-%d")
    # d1 = date1 - timedelta(1)
    # d2 = date1 - timedelta(2)
    # d3 = date1 - timedelta(3)
    # d4 = date1 - timedelta(4)
    # d5 = date1 - timedelta(5)
    d1 = unique_dates[0]
    d2 = unique_dates[1]
    d3 = unique_dates[2]
    d4 = unique_dates[3]
    d5 = unique_dates[4]
    cl = [d1, d2, d3, d4, d5]
    # index=df_pivot.index

    def overwrite(kpi_name, coln1):
        coln2 = num_hash(titleToNumber(coln1) + 1)
        coln3 = num_hash(titleToNumber(coln1) + 2)
        coln4 = num_hash(titleToNumber(coln1) + 3)
        coln5 = num_hash(titleToNumber(coln1) + 4)
        print(kpi_name)
        index = df_pivot.index
        # print("index ;###############################",index_pivot)
        # print(len(index_pivot))
        print("doooooonnenn")

        # print(df_pivot)
        print(f"KPI_name - {kpi_name}:- \n", df_pivot[kpi_name])
        dr = df_pivot[kpi_name]
        # print(dr)
        print("columns of dr tab")
        li = dr.columns
        print("column list", li)

        col1 = dr[li[0]].to_list()
        col2 = dr[li[1]].to_list()
        col3 = dr[li[2]].to_list()
        col4 = dr[li[3]].to_list()
        col5 = dr[li[4]].to_list()

        ws[coln1 + "2"].value = cl[4]
        ws[coln2 + "2"].value = cl[3]
        ws[coln3 + "2"].value = cl[2]
        ws[coln4 + "2"].value = cl[1]
        ws[coln5 + "2"].value = cl[0]

        for i, value in enumerate(index):
            j = i + 6

            ws["A" + str(j)].value = index[i][0]
            ws["B" + str(j)].value = index[i][1]

            ws[coln1 + str(j)].value = col1[i]
            ws[coln2 + str(j)].value = col2[i]
            ws[coln3 + str(j)].value = col3[i]
            ws[coln4 + str(j)].value = col4[i]
            ws[coln5 + str(j)].value = col5[i]

    for kpi_name in g2_kpi:
        if kpi_name == "SDCCH Drop Call Rate [BBH]":
            overwrite(kpi_name, "E")

        if kpi_name == "SDCCH Blocking Rate [BBH]":
            overwrite(kpi_name, "K")

        if kpi_name == "TCH Blocking Rate [BBH]":
            overwrite(kpi_name, "Q")

        if kpi_name == "TCH Drop Call Rate [BBH]":
            overwrite(kpi_name, "W")

        if kpi_name == "TCH Drop Call Rate_Denom [BBH]":
            overwrite(kpi_name, "AC")

        if kpi_name == "TCH Drop Call Rate_Nom [BBH]":
            overwrite(kpi_name, "AI")

        if kpi_name == "RX Quality [BBH]":
            overwrite(kpi_name, "AM")

        if kpi_name == "Handover Success Rate [BBH]":
            overwrite(kpi_name, "AS")

        if kpi_name == "Handover Success Rate_Denom [BBH]":
            overwrite(kpi_name, "AY")

        if kpi_name == "Handover Success Rate_Nom [BBH]":
            overwrite(kpi_name, "BD")

        if kpi_name == "TCH Assignment Success Rate [BBH]":
            overwrite(kpi_name, "BI")

        if kpi_name == "Network availability [RNA] [BBH]":
            overwrite(kpi_name, "BO")


    SaveOutput = os.path.join(MEDIA_ROOT, "trends", "ap", "ap2G", "output", "gsm1.xlsx")
    wb.save(SaveOutput)

    download_path = os.path.join(
        MEDIA_URL, "trends", "ap", "ap2G", "output", "gsm1.xlsx"
    )
    return Response(
        {
            "status": True,
            "missing_sites": message,
            "Download_url": download_path,
            "message": "successfully uploaded",
        }
    )


##############--------------------------------------------------------------------##################################