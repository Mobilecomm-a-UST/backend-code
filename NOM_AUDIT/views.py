from rest_framework.decorators import api_view
from rest_framework.response import Response
from collections import Counter
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
import os
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from django.core.files.storage import FileSystemStorage
from rest_framework import status
import json
import shutil
from NOM_AUDIT.reorder import reorder_post_values
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import subprocess
import re
import os
import shutil
import subprocess
from datetime import datetime
from urllib.parse import urljoin
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage



def is_date(value):
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def is_time(value):
    try:
        datetime.strptime(value, "%H:%M:%S")
        return True
    except ValueError:
        return False


def combine_values_for_hget_group(data):
    result = []
    for entry in data:
        new_result = []
        i = 0
        while i < len(entry):
            if (
                (entry[i] == "1" or entry[i] == "0")
                and i + 1 < len(entry)
                and entry[i + 1].startswith("(")
            ):
                new_result.append(f"{entry[i]} {entry[i+1]}")
                i += 2
            else:
                new_result.append(entry[i])
                i += 1
            result.append(new_result)
    return result


def hget_EUtranCell(df):
    actual_column_names = {
        ("PRE", "Unnamed: 0_level_1"): ("PRE", "Index"),
        ("PRE", 0): ("PRE", "MO"),
        ("PRE", 1): ("PRE", "administrativeState"),
        ("PRE", 2): ("PRE", "cellBarred"),
        ("PRE", 3): ("PRE", "CellId"),
        ("PRE", 4): ("PRE", "dlPowerState"),
        ("PRE", 5): ("PRE", "earfcndl"),
        ("PRE", 6): ("PRE", "earfcnul"),
        ("PRE", 7): ("PRE", "operationalState"),
        ("PRE", 8): ("PRE", "primaryPlmnReserved"),
        ("PRE", 9): ("PRE", "sectorCarrierRef"),
        ("PRE", 10): ("PRE", "Date & Time(UTC)"),
        ("POST", 11): ("PRE", "IP ADDR"),
        ("POST", 12): ("POST", "MO"),
        ("POST", 13): ("POST", "administrativeState"),
        ("POST", 14): ("POST", "cellBarred"),
        ("POST", 15): ("POST", "CellId"),
        ("POST", 16): ("POST", "dlPowerState"),
        ("POST", 17): ("POST", "earfcndl"),
        ("POST", 18): ("POST", "earfcnul"),
        ("POST", 19): ("POST", "operationalState"),
        ("POST", 20): ("POST", "primaryPlmnReserved"),
        ("POST", 21): ("POST", "sectorCarrierRef"),
        ("POST", 22): ("POST", "Date & Time(UTC)"),
        ("POST", 23): ("POST", "IP ADDR"),
    }

    df.rename(columns=actual_column_names, inplace=True)

    def check_mo(row):
        try:
            if (
                str(row[("PRE", 0)]) == str(row[("POST", 12)])
                and "_" not in str(row[("PRE", 0)])
                and "_" not in str(row[("POST", 12)])
            ):
                return "OK"
            return (
                "OK"
                if (
                    str(row[("PRE", 0)]).split("=")[1][1:-1]
                    == str(row[("POST", 12)]).split("=")[1].split("_")[4][:-1]
                )
                else "NOT OK"
            )
        except (IndexError, ValueError):
            return "ERROR"

    df[("audit", "MO")] = df.apply(lambda row: check_mo(row), axis=1)

    df[("audit", "administrativeState")] = df.apply(
        lambda row: "OK" if (row[("PRE", 1)] == row[("POST", 13)]) else "NOT OK", axis=1
    )
    df[("audit", "cellBarred")] = df.apply(
        lambda row: "OK" if (row[("PRE", 2)] == row[("POST", 14)]) else "NOT OK", axis=1
    )
    df[("audit", "CellId")] = df.apply(
        lambda row: "OK" if (row[("PRE", 3)] == row[("POST", 15)]) else "NOT OK", axis=1
    )
    df[("audit", "dlPowerState")] = df.apply(
        lambda row: "OK" if (row[("PRE", 4)] == row[("POST", 16)]) else "NOT OK", axis=1
    )
    df[("audit", "operationalState")] = df.apply(
        lambda row: "OK" if (row[("PRE", 7)] == row[("POST", 19)]) else "NOT OK", axis=1
    )
    df[("audit", "sectorCarrierRef")] = df.apply(
        lambda row: "OK" if (row[("PRE", 9)] == row[("POST", 21)]) else "NOT OK", axis=1
    )

    df[("audit", "IP ADDR")] = df.apply(
        lambda row: "OK" if (row[("POST", 11)] == row[("POST", 23)]) else "NOT OK",
        axis=1,
    )

    df[("OverAll", "Status")] = df.apply(
        lambda row: (
            "OK"
            if (
                row[("audit", "MO")] == "OK"
                and row[("audit", "cellBarred")] == "OK"
                and row[("audit", "CellId")] == "OK"
                and row[("audit", "dlPowerState")] == "OK"
                and row[("audit", "operationalState")] == "OK"
                and row[("audit", "sectorCarrierRef")] == "OK"
                and row[("audit", "IP ADDR")] == "OK"
            )
            else "NOT OK"
        ),
        axis=1,
    )

    df.columns = pd.MultiIndex.from_tuples(
        list(actual_column_names.values())
        + [
            ("audit", "MO"),
            ("audit", "administrativeState"),
            ("audit", "cellBarred"),
            ("audit", "CellId"),
            ("audit", "dlPowerState"),
            ("audit", "operationalState"),
            ("audit", "sectorCarrierRef"),
            ("audit", "IP ADDR"),
            ("Overall", "Status"),
        ]
    )

    return df


def st_cell_log_audit(df):
    actual_column_names = {
        ("PRE", "Unnamed: 0_level_1"): ("PRE", "Index"),
        ("PRE", 0): ("PRE", "Proxy"),
        ("PRE", 1): ("PRE", "Adm"),
        ("PRE", 2): ("PRE", "State"),
        ("PRE", 3): ("PRE", "Op."),
        ("PRE", 4): ("PRE", "State.1"),
        ("PRE", 5): ("PRE", "MO"),
        ("PRE", 6): ("PRE", "Date & Time(UTC)"),
        ("PRE", 7): ("PRE", "IP ADDR"),
        ("POST", 8): ("POST", "Proxy"),
        ("POST", 9): ("POST", "Adm"),
        ("POST", 10): ("POST", "State"),
        ("POST", 11): ("POST", "Op."),
        ("POST", 12): ("POST", "State.1"),
        ("POST", 13): ("POST", "MO"),
        ("POST", 14): ("POST", "Date & Time(UTC)"),
        ("POST", 15): ("POST", "IP ADDR"),
    }

    df.rename(columns=actual_column_names, inplace=True)

    def check_mo(row):
        try:
            if (
                str(row[("PRE", 5)]) == str(row[("POST", 13)])
                and "_" in str(row[("PRE", 5)])
                and "_" in str(row[("POST", 13)])
            ):
                return "NOT OK"
            elif str(row[("PRE", 5)]) == str(row[("POST", 13)]):
                return "NOT OK"
            elif (
                str(row[("PRE", 5)]).split(",")[1].split("=")[1]
                != str(row[("POST", 13)]).split(",")[1].split("=")[1]
            ):
                if "F" in str(row[("POST", 13)]).split(",")[1].split("=")[1].split("_"):
                    return "ERROR"
            return (
                "OK"
                if (
                    str(row[("PRE", 5)]).split(",")[1].split("=")[1]
                    != str(row[("POST", 13)]).split(",")[1].split("=")[1]
                )
                else "NOT OK"
            )
        except (IndexError, ValueError):
            return "ERROR"

    df[("audit", "AdmState")] = df.apply(
        lambda row: (
            "OK"
            if row[("PRE", 1)] == row[("POST", 9)]
            and row[("PRE", 2)] == row[("POST", 10)]
            else "NOT OK"
        ),
        axis=1,
    )
    df[("audit", "OpState")] = df.apply(
        lambda row: (
            "OK"
            if row[("PRE", 3)] == row[("POST", 11)]
            and row[("PRE", 4)] == row[("POST", 12)]
            else "NOT OK"
        ),
        axis=1,
    )
    df[("audit", "MO")] = df.apply(
        check_mo,
        axis=1,
    )
    df[("audit", "IP ADDR")] = df.apply(
        lambda row: "OK" if row[("POST", 7)] == row[("POST", 15)] else "NOT OK", axis=1
    )

    df[("OverAll", "Status")] = df.apply(
        lambda row: (
            "OK"
            if (
                row[("audit", "AdmState")] == "OK"
                and row[("audit", "OpState")] == "OK"
                and row[("audit", "MO")] == "OK"
                and row[("audit", "IP ADDR")] == "OK"
            )
            else "NOT OK"
        ),
        axis=1,
    )

    df.columns = pd.MultiIndex.from_tuples(
        list(actual_column_names.values())
        + [
            ("audit", "AdmState"),
            ("audit", "OpState"),
            ("audit", "MO"),
            ("audit", "IP ADDR"),
            ("OverAll", "Status"),
        ]
    )

    return df


def get_EutrancellTDD_audit(df):
    if df.isnull().all().all():
        return pd.DataFrame(columns=df.columns)
    else:
        # print(f"get_EutrancellTDD_audit", df.columns)

        actual_column_names = {
            ("PRE", "Unnamed: 0_level_1"): ("PRE", "Index"),
            ("PRE", 0): ("PRE", "MO"),
            ("PRE", 1): ("PRE", "Attribute"),
            ("PRE", 2): ("PRE", "Value"),
            ("PRE", 3): ("PRE", "Date & Time(UTC)"),
            ("POST", 4): ("PRE", "IP ADDR"),
            ("POST", 5): ("POST", "MO"),
            ("POST", 6): ("POST", "Attribute"),
            ("POST", 7): ("POST", "Value"),
            ("POST", 8): ("POST", "Date & Time(UTC)"),
            ("POST", 9): ("POST", "IP ADDR"),
        }

        def check_mo(row):
            try:
                if (
                    str(row[("PRE", 0)]) == str(row[("POST", 5)])
                    and "_" not in row[("PRE", 0)]
                    and "_" not in row[("POST", 5)]
                ):
                    return "OK"
                return (
                    "OK"
                    if (
                        str(row[("PRE", 0)]).split("=")[1][1:-1]
                        == str(row[("POST", 5)]).split("=")[1].split("_")[4][:-1]
                    )
                    else "NOT OK"
                )
            except (IndexError, ValueError):
                return "ERROR"

        df[("audit", "MO")] = df.apply(check_mo, axis=1)
        df[("audit", "Value")] = df.apply(
            lambda row: "OK" if row[("PRE", 2)] == row[("POST", 7)] else "NOT OK",
            axis=1,
        )
        df[("audit", "IP ADDR")] = df.apply(
            lambda row: "OK" if row[("POST", 4)] == row[("POST", 9)] else "NOT OK",
            axis=1,
        )

        # Overall Status
        df[("OverAll", "Status")] = df.apply(
            lambda row: (
                "OK"
                if (
                    row[("audit", "MO")] == "OK"
                    and row[("audit", "Value")] == "OK"
                    and row[("audit", "IP ADDR")] == "OK"
                )
                else "NOT OK"
            ),
            axis=1,
        )

        df.rename(columns=actual_column_names, inplace=True)
        df.columns = pd.MultiIndex.from_tuples(
            list(actual_column_names.values())
            + [
                ("audit", "MO"),
                ("audit", "Value"),
                ("audit", "IP ADDR"),
                ("OverAll", "Status"),
            ]
        )

        # print(df.columns)

        return df


def get_EutrancellFDD_audit(df):
    if df.isnull().all().all():
        # Return an empty DataFrame with the same columns
        return pd.DataFrame(columns=df.columns)

    def check_mo(row):
        try:
            if (
                str(row[("PRE", 0)]) == str(row[("POST", 5)])
                and "_" not in row[("PRE", 0)]
                and "_" not in row[("POST", 5)]
            ):
                return "OK"
            return (
                "OK"
                if (
                    str(row[("PRE", 0)]).split("=")[1][1:-1]
                    == str(row[("POST", 5)]).split("=")[1].split("_")[4][:-1]
                )
                else "NOT OK"
            )
        except (IndexError, ValueError):
            return "ERROR"

    actual_column_names = {
        ("PRE", "Unnamed: 0_level_1"): ("PRE", "Index"),
        ("PRE", 0): ("PRE", "MO"),
        ("PRE", 1): ("PRE", "Attribute"),
        ("PRE", 2): ("PRE", "Value"),
        ("PRE", 3): ("PRE", "Date & Time(UTC)"),
        ("POST", 4): ("PRE", "IP ADDR"),
        ("POST", 5): ("POST", "MO"),
        ("POST", 6): ("POST", "Attribute"),
        ("POST", 7): ("POST", "Value"),
        ("POST", 8): ("POST", "Date & Time(UTC)"),
        ("POST", 9): ("POST", "IP ADDR"),
    }

    df[("audit", "MO")] = df.apply(check_mo, axis=1)
    df[("audit", "Value")] = df.apply(
        lambda row: "OK" if row[("PRE", 2)] == row[("POST", 7)] else "NOT OK", axis=1
    )
    df[("audit", "IP ADDR")] = df.apply(
        lambda row: "OK" if row[("POST", 4)] == row[("POST", 9)] else "NOT OK", axis=1
    )
    df[("OverAll", "Status")] = df.apply(
        lambda row: (
            "OK"
            if (
                row[("audit", "MO")] == "OK"
                and row[("audit", "Value")] == "OK"
                and row[("audit", "IP ADDR")] == "OK"
            )
            else "NOT OK"
        ),
        axis=1,
    )

    df.columns = pd.MultiIndex.from_tuples(
        list(actual_column_names.values())
        + [
            ("audit", "MO"),
            ("audit", "Value"),
            ("audit", "IP ADDR"),
            ("OverAll", "Status"),
        ]
    )

    return df


def getEutrancellTDD_audit(df):
    if df.isnull().all().all():
        # Return an empty DataFrame with the same columns
        return pd.DataFrame(columns=df.columns)

    actual_column_names = {
        ("PRE", "Unnamed: 0_level_1"): ("PRE", "Index"),
        ("PRE", 0): ("PRE", "MO"),
        ("PRE", 1): ("PRE", "Attribute"),
        ("PRE", 2): ("PRE", "Value"),
        ("PRE", 3): ("PRE", "Date & Time(UTC)"),
        ("POST", 4): ("PRE", "IP ADDR"),
        ("POST", 5): ("POST", "MO"),
        ("POST", 6): ("POST", "Attribute"),
        ("POST", 7): ("POST", "Value"),
        ("POST", 8): ("POST", "Date & Time(UTC)"),
        ("POST", 9): ("POST", "IP ADDR"),
    }

    def check_mo(row):
        try:
            if (
                str(row[("PRE", 0)]) == str(row[("POST", 5)])
                and "_" not in row[("PRE", 0)]
                and "_" not in row[("POST", 5)]
            ):
                return "OK"
            return (
                "OK"
                if (
                    str(row[("PRE", 0)]).split("=")[1][1:-1]
                    == str(row[("POST", 5)]).split("=")[1].split("_")[4][:-1]
                )
                else "NOT OK"
            )
        except (IndexError, ValueError):
            return "ERROR"

    df[("audit", "MO")] = df.apply(check_mo, axis=1)
    df[("audit", "Value")] = df.apply(
        lambda row: "OK" if row[("PRE", 2)] == row[("POST", 7)] else "NOT OK", axis=1
    )
    df[("audit", "IP ADDR")] = df.apply(
        lambda row: "OK" if row[("POST", 4)] == row[("POST", 9)] else "NOT OK", axis=1
    )
    df[("OverAll", "Status")] = df.apply(
        lambda row: (
            "OK"
            if (
                row[("audit", "MO")] == "OK"
                and row[("audit", "Value")] == "OK"
                and row[("audit", "IP ADDR")] == "OK"
            )
            else "NOT OK"
        ),
        axis=1,
    )

    df.columns = pd.MultiIndex.from_tuples(
        list(actual_column_names.values())
        + [
            ("audit", "MO"),
            ("audit", "Value"),
            ("audit", "IP ADDR"),
            ("OverAll", "Status"),
        ]
    )

    return df


def getEutrancellFDD_audit(df):
    if df.isnull().all().all():
        # Return an empty DataFrame with the same columns
        return pd.DataFrame(columns=df.columns)

    actual_column_names = {
        ("PRE", "Unnamed: 0_level_1"): ("PRE", "Index"),
        ("PRE", 0): ("PRE", "MO"),
        ("PRE", 1): ("PRE", "Attribute"),
        ("PRE", 2): ("PRE", "Value"),
        ("PRE", 3): ("PRE", "Date & Time(UTC)"),
        ("POST", 4): ("PRE", "IP ADDR"),
        ("POST", 5): ("POST", "MO"),
        ("POST", 6): ("POST", "Attribute"),
        ("POST", 7): ("POST", "Value"),
        ("POST", 8): ("POST", "Date & Time(UTC)"),
        ("POST", 9): ("POST", "IP ADDR"),
    }

    def check_mo(row):
        try:
            if (
                str(row[("PRE", 0)]) == str(row[("POST", 5)])
                and "_" not in row[("PRE", 0)]
                and "_" not in row[("POST", 5)]
            ):
                return "OK"
            return (
                "OK"
                if (
                    str(row[("PRE", 0)]).split("=")[1][1:-1]
                    == str(row[("POST", 5)]).split("=")[1].split("_")[4][:-1]
                )
                else "NOT OK"
            )
        except (IndexError, ValueError):
            return "ERROR"

    df[("audit", "MO")] = df.apply(check_mo, axis=1)
    df[("audit", "Value")] = df.apply(
        lambda row: "OK" if row[("PRE", 2)] == row[("POST", 7)] else "NOT OK", axis=1
    )
    df[("audit", "IP ADDR")] = df.apply(
        lambda row: "OK" if row[("POST", 4)] == row[("POST", 9)] else "NOT OK", axis=1
    )
    df[("OverAll", "Status")] = df.apply(
        lambda row: (
            "OK"
            if (
                row[("audit", "MO")] == "OK"
                and row[("audit", "Value")] == "OK"
                and row[("audit", "IP ADDR")] == "OK"
            )
            else "NOT OK"
        ),
        axis=1,
    )

    df.columns = pd.MultiIndex.from_tuples(
        list(actual_column_names.values())
        + [
            ("audit", "MO"),
            ("audit", "Value"),
            ("audit", "IP ADDR"),
            ("OverAll", "Status"),
        ]
    )

    return df


def get_prbpairsperframe(df):
    actual_column_names = {
        ("PRE", "Unnamed: 0_level_1"): ("PRE", "Index"),
        ("PRE", 0): ("PRE", "MO"),
        ("PRE", 1): ("PRE", "Attribute"),
        ("PRE", 2): ("PRE", "Value"),
        ("PRE", 3): ("PRE", "Date & Time(UTC)"),
        ("POST", 4): ("PRE", "IP ADDR"),
        ("POST", 5): ("POST", "MO"),
        ("POST", 6): ("POST", "Attribute"),
        ("POST", 7): ("POST", "Value"),
        ("POST", 8): ("POST", "Date & Time(UTC)"),
        ("POST", 9): ("POST", "IP ADDR"),
    }

    def check_mo(row):
        try:
            if (
                str(row[("PRE", 0)]) == str(row[("POST", 5)])
                and "_" not in row[("PRE", 0)]
                and "_" not in row[("POST", 5)]
            ):
                return "OK"
            return (
                "OK"
                if (
                    str(row[("PRE", 0)]).split("=")[1][1:-1]
                    == str(row[("POST", 5)]).split("=")[1].split("_")[4][:-1]
                )
                else "NOT OK"
            )
        except (IndexError, ValueError):
            return "ERROR"

    df[("audit", "MO")] = df.apply(check_mo, axis=1)

    df[("audit", "Value")] = df.apply(
        lambda row: "OK" if row[("PRE", 2)] == row[("POST", 7)] else "NOT OK", axis=1
    )

    df[("audit", "IP ADDR")] = df.apply(
        lambda row: "OK" if row[("POST", 4)] == row[("POST", 9)] else "NOT OK", axis=1
    )

    df[("OverAll", "Status")] = df.apply(
        lambda row: (
            "OK"
            if (
                row[("audit", "MO")] == "OK"
                and row[("audit", "Value")] == "OK"
                and row[("audit", "IP ADDR")] == "OK"
            )
            else "NOT OK"
        ),
        axis=1,
    )

    df.columns = pd.MultiIndex.from_tuples(
        list(actual_column_names.values())
        + [
            ("audit", "MO"),
            ("audit", "Value"),
            ("audit", "IP ADDR"),
            ("OverAll", "Status"),
        ]
    )

    return df


def pr_EUtranFreqRelation(df):

    actual_column_names = {
        ("PRE", "Unnamed: 0_level_1"): ("PRE", "Index"),
        ("PRE", 0): ("PRE", "Proxy"),
        ("PRE", 1): ("PRE", "MO"),
        ("PRE", 2): ("PRE", "Date & Time(UTC)"),
        ("POST", 3): ("PRE", "IP ADDR"),
        ("POST", 4): ("POST", "Proxy"),
        ("POST", 5): ("POST", "MO"),
        ("POST", 6): ("POST", "Date & Time(UTC)"),
        ("POST", 7): ("POST", "IP ADDR"),
    }

    def check_mo(row):
        try:
            if (
                str(row[("PRE", 1)]) == str(row[("POST", 5)])
                and "_" not in row[("PRE", 1)]
                and "_" not in row[("POST", 5)]
            ):
                return "OK"
            return (
                "OK"
                if (
                    str(row[("PRE", 1)]).split("=")[1][1:-1]
                    == str(row[("POST", 5)]).split("=")[1].split("_")[4][:-1]
                )
                else "NOT OK"
            )
        except (IndexError, ValueError):
            return "ERROR"

    df[("audit", "Proxy")] = df.apply(
        lambda row: "OK" if row[("PRE", 0)] == row[("POST", 4)] else "NOT OK", axis=1
    )
    df[("audit", "MO")] = df.apply(check_mo, axis=1)

    df[("audit", "IP ADDR")] = df.apply(
        lambda row: "OK" if row[("POST", 3)] == row[("POST", 7)] else "NOT OK", axis=1
    )
    df[("OverAll", "Status")] = df.apply(
        lambda row: (
            "OK"
            if (
                row[("audit", "MO")] == "OK"
                and row[("audit", "Proxy")] == "OK"
                and row[("audit", "IP ADDR")] == "OK"
            )
            else "NOT OK"
        ),
        axis=1,
    )

    df.columns = pd.MultiIndex.from_tuples(
        list(actual_column_names.values())
        + [
            ("audit", "MO"),
            ("audit", "Proxy"),
            ("audit", "IP ADDR"),
            ("OverAll", "Status"),
        ]
    )

    return df


def get_AntennaUnitGroup(df):
    actual_column_names = {
        ("PRE", "Unnamed: 0_level_1"): ("PRE", "Index"),
        ("PRE", 0): ("PRE", "MO"),
        ("PRE", 1): ("PRE", "Attribute"),
        ("PRE", 2): ("PRE", "Value"),
        ("PRE", 3): ("PRE", "Date & Time(UTC)"),
        ("POST", 4): ("PRE", "IP ADDR"),
        ("POST", 5): ("POST", "MO"),
        ("POST", 6): ("POST", "Attribute"),
        ("POST", 7): ("POST", "Value"),
        ("POST", 8): ("POST", "Date & Time(UTC)"),
        ("POST", 9): ("POST", "IP ADDR"),
    }

    def check_mo(row):
        try:
            if (
                str(row[("PRE", 0)]) == str(row[("POST", 5)])
                and "_" not in row[("PRE", 0)]
                and "_" not in row[("POST", 5)]
            ):
                return "OK"
            elif str(row[("PRE", 0)]) == str(row[("POST", 5)]):
                return "OK"
            return (
                "OK"
                if (
                    str(row[("PRE", 0)]).split("=")[1][1:-1]
                    == str(row[("POST", 5)]).split("=")[1].split("_")[4][:-1]
                )
                else "NOT OK"
            )
        except (IndexError, ValueError):
            return "ERROR"

    df[("audit", "MO")] = df.apply(check_mo, axis=1)
    df[("audit", "Value")] = df.apply(
        lambda row: "OK" if row[("PRE", 2)] == row[("POST", 7)] else "NOT OK", axis=1
    )
    df[("audit", "Attribute")] = df.apply(
        lambda row: "OK" if row[("PRE", 1)] == row[("POST", 6)] else "NOT OK", axis=1
    )
    df[("audit", "IP ADDR")] = df.apply(
        lambda row: "OK" if row[("POST", 4)] == row[("POST", 9)] else "NOT OK", axis=1
    )
    df[("OverAll", "Status")] = df.apply(
        lambda row: (
            "OK"
            if (
                row[("audit", "MO")] == "OK"
                and row[("audit", "Value")] == "OK"
                and row[("audit", "IP ADDR")] == "OK"
                and row[("audit", "Attribute")] == "OK"
            )
            else "NOT OK"
        ),
        axis=1,
    )
    df.columns = pd.MultiIndex.from_tuples(
        list(actual_column_names.values())
        + [
            ("audit", "MO"),
            ("audit", "Value"),
            ("audit", "Attribute"),
            ("audit", "IP ADDR"),
            ("OverAll", "Status"),
        ]
    )
    return df


def hget_tss(df):

    actual_column_names = {
        ("PRE", "Unnamed: 0_level_1"): ("PRE", "Index"),
        ("PRE", 0): ("PRE", "MO"),
        ("PRE", 1): ("PRE", "abisTsState"),
        # ('PRE', 2): ('PRE', 'ductIntSspStatus'),
        ("PRE", 2): ("PRE", "Date & Time(UTC)"),
        ("POST", 3): ("PRE", "IP ADDR"),
        ("POST", 4): ("POST", "MO"),
        ("POST", 5): ("POST", "abisTsState"),
        # ('POST', 7): ('POST', 'ductIntSspStatus'),
        ("POST", 6): ("POST", "Date & Time(UTC)"),
        ("POST", 7): ("POST", "IP ADDR"),
    }

    def check_mo(row):
        try:
            if (
                str(row[("PRE", 0)]) == str(row[("POST", 4)])
                and "_" not in row[("PRE", 0)]
                and "_" not in row[("POST", 5)]
            ):
                return "OK"
            return (
                "OK"
                if (
                    str(row[("PRE", 0)]).split("=")[1][1:-1]
                    == str(row[("POST", 4)]).split("=")[1].split("_")[4][:-1]
                )
                else "NOT OK"
            )
        except (IndexError, ValueError):
            return "ERROR"

    df[("audit", "MO")] = df.apply(check_mo, axis=1)
    if ("PRE", 1) in df.columns and ("POST", 5) in df.columns:
        df[("audit", "abisTsState")] = df.apply(
            lambda row: "OK" if row[("PRE", 1)] == row[("POST", 5)] else "NOT OK",
            axis=1,
        )
    if ("audit", "ductIntSspStatus") in df.columns:
        df[("audit", "ductIntSspStatus")] = df.apply(
            lambda row: "OK" if row[("PRE", 2)] == row[("POST", 6)] else "NOT OK",
            axis=1,
        )
    df[("audit", "IP ADDR")] = df.apply(
        lambda row: "OK" if row[("POST", 3)] == row[("POST", 7)] else "NOT OK", axis=1
    )
    df[("OverAll", "Status")] = df.apply(
        lambda row: (
            "OK"
            if (
                row[("audit", "MO")] == "OK"
                and row[("audit", "abisTsState")] == "OK"
                and row[("audit", "IP ADDR")] == "OK"
            )
            else "NOT OK"
        ),
        axis=1,
    )
    df.columns = pd.MultiIndex.from_tuples(
        list(actual_column_names.values())
        + [
            ("audit", "MO"),
            ("audit", "abisTsState"),
            ("audit", "IP ADDR"),
            ("OverAll", "Status"),
        ]
    )
    return df


def altk(df):
    # print("altk", df.columns)
    actual_column_names = {
        ("PRE", "Unnamed: 0_level_1"): ("PRE", "Index"),
        ("PRE", 0): ("PRE", "Date & Time (UTC)"),
        ("PRE", 1): ("PRE", "S"),
        ("PRE", 2): ("PRE", "Specific Problem"),
        ("PRE", 3): ("PRE", "MO (Cause/AdditionalInfo)"),
        ("PRE", 4): ("PRE", "Date & Time(UTC)"),
        ("POST", 5): ("PRE", "IP ADDR"),
        ("POST", 6): ("PRE", "Date & Time (UTC)"),
        ("POST", 7): ("PRE", "S"),
        ("POST", 8): ("POST", "Specific Problem"),
        ("POST", 9): ("POST", "MO (Cause/AdditionalInfo)"),
        ("POST", 10): ("POST", "Date & Time(UTC)"),
        ("POST", 11): ("POST", "IP ADDR"),
    }

    def check_mo(row):
        try:
            if (
                str(row[("PRE", 3)]) == str(row[("POST", 8)])
                and "_" not in row[("PRE", 3)]
                and "_" not in row[("POST", 8)]
            ):
                return "OK"
            return (
                "OK"
                if (
                    str(row[("PRE", 3)]).split("=")[1][1:-1]
                    == str(row[("POST", 8)]).split("=")[1].split("_")[4][:-1]
                )
                else "NOT OK"
            )
        except (IndexError, ValueError):
            return "ERROR"

    df[("audit", "Specific Problem")] = df.apply(
        lambda row: "OK" if row[("PRE", 2)] == row[("POST", 8)] else "NOT OK", axis=1
    )
    df[("audit", "MO (Cause/AdditionalInfo)")] = df.apply(
        lambda row: "OK" if row[("PRE", 3)] == row[("POST", 9)] else "NOT OK", axis=1
    )
    df[("audit", "IP ADDR")] = df.apply(
        lambda row: "OK" if row[("POST", 5)] == row[("POST", 11)] else "NOT OK", axis=1
    )
    df[("OverAll", "Status")] = df.apply(
        lambda row: (
            "OK"
            if (
                row[("audit", "Specific Problem")] == "OK"
                and row[("audit", "MO (Cause/AdditionalInfo)")] == "OK"
                and row[("audit", "IP ADDR")] == "OK"
            )
            else "NOT OK"
        ),
        axis=1,
    )
    df.columns = pd.MultiIndex.from_tuples(
        list(actual_column_names.values())
        + [
            ("audit", "Specific Problem"),
            ("audit", "MO (Cause/AdditionalInfo)"),
            ("audit", "IP ADDR"),
            ("OverAll", "Status"),
        ]
    )
    return df


def extract_tdd_fdd_data(tech: str, content: list[str], found: bool) -> pd.DataFrame:
    """
    Extracts TDD/FDD data from a given content.

    :param tech: The technology type (TDD/FDD)
    :param content: The content to extract data from
    :param found: Whether the data extraction has been started
    :return: A pandas DataFrame containing the extracted data
    """
    data = []
    for line in content:
        line = line.strip()
        if re.search(rf"\sget\sEutrancell{tech}=.*", line):
            found = True
            continue

        if found and not re.match(r"==+", line):
            if re.match(
                r"EUtranCell[A-Z][A-Z]{2}=\S+\s+(earfcn)?(earfcndl)?\s+\d+", line
            ):
                list_val = re.split(r"\s{2,}+", line)
                data.append(
                    {
                        "MO": list_val[0].split("=")[1],
                        "Attribute": list_val[1],
                        "Value": list_val[2],
                    }
                )
        if found and re.match(r"Total:\s\d+\sMOs", line) and not re.match(r"==+", line):
            found = False
            break
    df = pd.DataFrame(data, columns=["MO", "Attribute", "Value"])

    return df


def create_dataframe(headers, values, date_ip_info, start_point):
    if start_point == "cvls":
        for val in values:
            idx = 1
            date_id = -1
            time_id = -1
            try:
                for id, value in enumerate(val):
                    if is_date(value):
                        date_id = id
                    if is_time(value):
                        time_id = id
            except Exception as e:
                # print(f"Error processing cvls date/time detection: {e}")
                continue

            try:
                if idx == 1 and date_id != -1:
                    segment = val[idx:date_id]
                    combind = " ".join(segment)
                    val[idx:date_id] = [combind]
                if date_id != -1 and time_id != -1:
                    date_str = val[date_id]
                    time_str = val[time_id]
                    date_time_str = f"{date_str} {time_str}"
                    val[date_id] = date_time_str
                    del val[time_id]
            except Exception as e:
                # print(f"Error processing cvls value adjustment: {e}")
                continue

    if start_point == "altk":
        headers = [
            "Date & Time (UTC)",
            "S",
            "Specific Problem",
            "MO (Cause/AdditionalInfo)",
        ]
        for val in values:
            date_id = -1
            time_id = -1
            find_id_from_equal = -1
            try:
                for id, value in enumerate(val):
                    if is_date(value):
                        date_id = id
                    if is_time(value):
                        time_id = id
                    if "=" in value:
                        find_id_from_equal = id
            except Exception as e:
                # print(f"Error detecting altk attributes: {e}")
                continue

            try:
                if date_id != -1 and time_id != -1:
                    date_str = val[date_id]
                    time_str = val[time_id]
                    date_time_str = f"{date_str} {time_str}"
                    val[date_id] = date_time_str
                    del val[time_id]
                if find_id_from_equal != -1:
                    segment = val[find_id_from_equal - 1 :]
                    combind = " ".join(segment)
                    val[find_id_from_equal:] = [combind]
                segment_line = val[2:-1]
                combind_line = " ".join(segment_line)
                val[2:-1] = [combind_line]
            except Exception as e:
                # print(f"Error processing altk value adjustment: {e}")
                continue

    if start_point == "hget . tss":
        for value in values:
            try:
                segment = value[1:]
                combind = " ".join(segment)
                value[1:] = [combind]
            except Exception as e:
                # print(f"Error processing hget . tss segment: {e}")
                continue

    if (
        start_point
        == "hget ^EUtranCell*.* cellBARred|sectorCARrierRef|primARyPlmnReserved|state|eARfcndl|eARfcnUL|^cellId"
    ):
        for value in values:
            try:
                sector_carrier = -1
                for id, val in enumerate(value):
                    if val == "[1]":
                        sector_carrier = id
                if sector_carrier != -1:
                    segment = value[sector_carrier:]
                    combind = " ".join(segment)
                    value[sector_carrier:] = [combind]
            except Exception as e:
                print(f"Error processing hget sectorCarrier segment: {e}")
                continue
        try:
            values = combine_values_for_hget_group(values)
        except Exception as e:
            print(f"Error combining hget group values: {e}")

    if (
        start_point
        == "get AntennaUnitGroup=*.*,AntennaNearUnit=*.*,RetSubUnit=*.*   userLabel"
    ):
        for value in values:
            try:
                sector_carrier = -1
                for id, val in enumerate(value):
                    if val == "userLabel":
                        sector_carrier = id
                if sector_carrier != -1:
                    segment = value[sector_carrier + 1 :]
                    combind = " ".join(segment)
                    value[sector_carrier + 1 :] = [combind]
            except Exception as e:
                # print(f"Error processing AntennaUnitGroup userLabel: {e}")
                continue

    try:
        if date_ip_info and date_ip_info[0] is not None:
            try:
                date = datetime.strptime(date_ip_info[0].split("-")[0], "%y%m%d")
                date = datetime.strftime(date, "%Y-%m-%d")
                time = date_ip_info[0].split("-")[1]
                date_time = date + " " + time
            except ValueError as e:
                # Handle the case where the date format is incorrect
                print(f"Error parsing date: {e}")

            date = None  # or handle accordingly
        else:
            # print("date_ip_info is None or empty")
            date = None

    except (ValueError, IndexError) as e:
        # print(f"Error processing date and time: {e}")
        date_time = "Unknown Date & Time"
    try:
        df = pd.DataFrame(values, columns=headers)
    except Exception as e:
        # print(f"Error creating DataFrame: {e}")
        return None

    try:
        df[("Date & Time(UTC)")] = df[f"{headers[0]}"].apply(lambda x: date_time)
        df[("IP ADDR")] = df[f"{headers[0]}"].apply(
            lambda x: date_ip_info[1] + " " + date_ip_info[-1]
        )
    except Exception as e:
        print(f"Error processing DataFrame columns: {e}")

    return df


def extract_header_and_extract_values(
    header_start, file_content, start_point, end_point
):
    headers = []
    values = []
    command_info = []
    start_found = False
    header_found = False
    site_id = None
    for id, line in enumerate(file_content):
        if start_point in line:
            site_id = line.split(">")[0]
            start_found = True
            continue

        if start_found:
            if end_point in line:
                break
            if "MSRBS_NODE_MODEL" in line:
                command_info.extend(line.strip().split())

            else:
                if header_start in line:
                    header_found = True
                    headers = [val for val in line.strip().split() if val]
                    # print(headers)
                    continue
                if header_found:
                    values.append([val for val in line.strip().split() if val])

    return headers, values[1:], command_info, site_id


def delete_existing_files(upload_files):
    for file_name in os.listdir(upload_files):
        file_path = os.path.join(upload_files, file_name)
        if os.path.isfile(file_path):
            os.remove(file_path)


OUTPUT_DIR = os.path.join(settings.MEDIA_ROOT, "NOM_AUDIT", "scripts_outputs")

# Ensure output directory exists
def clean_output_directory():
    if os.path.isdir(OUTPUT_DIR):
        shutil.rmtree(OUTPUT_DIR)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

def find_tech(band):
    band_parts = band.split("_")
    band_part = band_parts[2] if len(band_parts) > 2 else ""
    
    if band_part in ["F1", "F3", "F8"]:
        return "FDD"
    elif band_part in ["T1", "T2"]:
        return "TDD"
    return ""





############################ Constants ################################################
OUTPUT_DIR = os.path.join(settings.MEDIA_ROOT, 'NOM_AUDIT', 'SCRIPTS_OUTPUT')
RAR_EXE_PATH = getattr(settings, "WINRAR_PATH", r"C:\Program Files\WinRAR\rar.exe")
#######################################################################################

def clean_output_directory():
    if os.path.exists(OUTPUT_DIR):
        shutil.rmtree(OUTPUT_DIR)
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def find_tech(cell_name):
    if "FDD" in cell_name:
        return "FDD"
    elif "TDD" in cell_name:
        return "TDD"
    return "UNKNOWN"


def find_band(cell_name):
    if "_F1_" in cell_name or "_F3_" in cell_name or "_F8_" in cell_name:
        return "FDD"
    elif "_T1_" in cell_name or "_T2_" in cell_name:
        return "TDD"
    return "UNKNOWN"


@api_view(['POST'])
def generate_site_scripts(request):
    if 'planned_site_file' not in request.FILES:
        return Response({"error": "No file provided"}, status=400)

    uploaded_file = request.FILES['planned_site_file']
    file_path = default_storage.save(
        os.path.join(settings.MEDIA_ROOT, "NOM_AUDIT", "temp_uploaded.xlsx"),
        ContentFile(uploaded_file.read())
    )
    file_path = default_storage.path(file_path)

    try:
        df_planed_sites = pd.read_excel(file_path, sheet_name="CELL DETAIL", engine='openpyxl')
    except Exception as e:
        return Response({"error": f"Error reading Excel file: {str(e)}"}, status=400)

    clean_output_directory()

    distinct_4g_sitenames = df_planed_sites["MO Name"].unique()
    total_files_count = 0

    df_planed_sites['Band_tech'] = df_planed_sites['Planned nomenclature'].apply(lambda x: find_band(str(x)))

    fdd_string = """lbl EUtranCellFDD={old_cel_name}
rset EUtranCellFDD={old_cel_name} EUtranCellFDDId {new_cel_name}
ldeb {new_cel_name}
"""

    tdd_string = """lbl EUtranCellTDD={old_cel_name}
rset EUtranCellTDD={old_cel_name} EUtranCellTDDId {new_cel_name}
ldeb {new_cel_name}
"""

    ending_string = """confbd-
$date = `date +%y%m%d_%H%M`
cvms post_cv_nomenclaturechange_$date
"""

    try:
        for site in distinct_4g_sitenames:
            op_string = """$date = `date +%y%m%d_%H%M`
cvms pre_cv_nomenclaturechange_$date

confbd+

"""
            df_planed_site = df_planed_sites[df_planed_sites["MO Name"] == site]

            for _, row in df_planed_site.iterrows():
                old_cel_name = row["MV Cell Name"]
                new_cel_name = row["Planned nomenclature"]
                tech = row['Band_tech']

                if tech == "FDD":
                    op_string += fdd_string.format(old_cel_name=old_cel_name, new_cel_name=new_cel_name)
                elif tech == "TDD":
                    op_string += tdd_string.format(old_cel_name=old_cel_name, new_cel_name=new_cel_name)

            op_string += ending_string

            script_name = f"{site}.txt"
            script_path = os.path.join(OUTPUT_DIR, script_name)

            try:
                with open(script_path, "w") as file:
                    file.write(op_string)
                total_files_count += 1
            except Exception as e:
                return Response({"error": f"Failed to write script for site {site}: {str(e)}"}, status=500)

        # Archive all scripts into a RAR file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        rar_filename = f"{'nomenclature_change_scripts'.upper()}_{timestamp}.rar"
        rar_file_path = os.path.join(settings.MEDIA_ROOT, 'NOM_AUDIT', rar_filename)

        if os.path.exists(rar_file_path):
            os.remove(rar_file_path)

        cmd = [RAR_EXE_PATH, "a", rar_file_path, os.path.join(OUTPUT_DIR, "*")]

        try:
            subprocess.run(cmd, check=True)
        except subprocess.CalledProcessError as e:
            return Response({"error": f"Error creating RAR file: {str(e)}"}, status=500)
        
        print("total file proccessed:- ", total_files_count)

        # Build the download URL safely
        download_url = urljoin(settings.MEDIA_URL, f"NOM_AUDIT/{rar_filename}")

        return Response({
            "message": "Scripts generated successfully",
            "download_url": download_url,
            "total_files": total_files_count
        })

    finally:
        # Cleanup temporary files
        if os.path.exists(file_path):
            os.remove(file_path)
        if os.path.exists(OUTPUT_DIR):
            shutil.rmtree(OUTPUT_DIR)





@api_view(["POST"])
def pre_post_audit_process(request):
    pre_files = request.FILES.getlist("pre_files")

    pre_folder = os.path.join(settings.MEDIA_ROOT, "NOM_AUDIT", "pre")
    UNPROCESSED_FILES = []
    if not os.path.exists(pre_folder):
        os.mkdir(pre_folder)

    post_files = request.FILES.getlist("post_files")

    post_folder = os.path.join(settings.MEDIA_ROOT, "NOM_AUDIT", "post")

    if not os.path.exists(post_folder):
        os.mkdir(post_folder)

    delete_existing_files(pre_folder)
    delete_existing_files(post_folder)

    for file in pre_files:
        with open(os.path.join(pre_folder, file.name), "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)

    for file in post_files:
        with open(os.path.join(post_folder, file.name), "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)

    circle = request.POST.get("circle")

    pre_file_names = set(file.name for file in pre_files)
    post_file_names = set(file.name for file in post_files)

    matched_files = sorted(list(set(pre_file_names).intersection(set(post_file_names))))

    matched_pairs = []
    for file_name in matched_files:
        matched_pairs.append(
            (os.path.join(pre_folder, file_name), os.path.join(post_folder, file_name))
        )

    new_pre_folder = os.path.join(settings.MEDIA_ROOT, "NOM_AUDIT", "new_pre")
    if not os.path.exists(new_pre_folder):
        os.mkdir(new_pre_folder)
    new_post_folder = os.path.join(settings.MEDIA_ROOT, "NOM_AUDIT", "new_post")
    if not os.path.exists(new_post_folder):
        os.mkdir(new_post_folder)

    delete_existing_files(new_pre_folder)
    delete_existing_files(new_post_folder)

    copied_pre_files = []
    copied_post_files = []
    for file_pair in matched_pairs:
        if os.path.exists(file_pair[0]):
            file_name = os.path.basename(file_pair[0])
            dest_path = os.path.join(new_pre_folder, file_name)
            shutil.copy(file_pair[0], dest_path)
            copied_pre_files.append(file_name)

        if os.path.exists(file_pair[1]):
            file_name = os.path.basename(file_pair[1])
            dest_path = os.path.join(new_post_folder, file_name)
            shutil.copy(file_pair[1], dest_path)
            copied_post_files.append(file_name)

    matched_files = [
        (pre, post)
        for pre, post in zip(copied_pre_files, copied_post_files)
        if pre == post
    ]

    # print(matched_files)

    def process_files(
        pre_file_content,
        post_file_content,
        list_commands,
        fdd_mapping_dict,
        tdd_mapping_dict,
    ):

        sheet_name_val = {}
        header = {
            "st cell": "Proxy",
            # "get AntennaUnitGroup=*.*,AntennaNearUnit=*.*,RetSubUnit=*.*   userLabel": "MO",
            # "cvls": "Id",
            "altk": "Date & Time (UTC)",
            # "hget . tss": "MO",
            # "pr EUtranFreqRelation=": "Proxy",
            # "get EutrancellTDD=.* Earfcn": "MO",
            # "get EutrancellFDD=.* Earfcn": "MO",
            # "get ^EutrancellTDD=.* ChannelBandwidth": "MO",
            # "get ^EutrancellFDD=.* ChannelBandwidth": "MO",
            # "get . prbpairsperframe": "MO",
            # "hget ^EUtranCell*.* cellBARred|sectorCARrierRef|primARyPlmnReserved|state|eARfcndl|eARfcnUL|^cellId": "MO",
        }

        for command in list_commands:
            header_start = header.get(command)
            if not header_start:
                continue

            start_point = command
            end_point = "Total"

            pre_header, pre_value, pre_date_ip_info, pre_site_id = (
                extract_header_and_extract_values(
                    header_start, pre_file_content, start_point, end_point
                )
            )

            pre_date_ip_info.append(pre_site_id)
            post_header, post_value, post_date_ip_info, post_site_id = (
                extract_header_and_extract_values(
                    header_start, post_file_content, start_point, end_point
                )
            )
            post_date_ip_info.append(post_site_id)
            pre_value = [val for val in pre_value if not "".join(val).startswith("=")]
            post_value = [val for val in post_value if not "".join(val).startswith("=")]
            try:
                pre_df = create_dataframe(
                    pre_header, pre_value, pre_date_ip_info, start_point
                )
                post_df = create_dataframe(
                    post_header, post_value, post_date_ip_info, start_point
                )

                print("pre_post", post_file, pre_file)

                # Handle DataFrame reordering based on start_point
                if start_point in ["st cell", "pr EUtranFreqRelation="]:
                    reorder_post_df = reorder_post_values(
                        post_df,
                        pre_df,
                        start_point,
                        circle,
                        fdd_mapping_dict,
                        tdd_mapping_dict,
                    )
                elif start_point in [
                    "hget ^EUtranCell*.* cellBARred|sectorCARrierRef|primARyPlmnReserved|state|eARfcndl|eARfcnUL|^cellId",
                    "get ^EutrancellFDD=.* ChannelBandwidth",
                    # "get EutrancellFDD=.* Earfcn",
                    # "get EutrancellTDD=.* Earfcn",
                    "get . prbpairsperframe",
                ]:
                    reorder_post_df = reorder_post_values(
                        post_df, pre_df, start_point, circle
                    )
                else:
                    reorder_post_df = post_df

                # Check if the DataFrames are empty
                if not pre_df.empty and not reorder_post_df.empty:
                    pre_df.columns = pd.MultiIndex.from_tuples(
                        [("PRE", val) for val in pre_df.columns]
                    )
                    reorder_post_df.columns = pd.MultiIndex.from_tuples(
                        [("POST", val) for val in reorder_post_df.columns]
                    )

                    print("pre_df \n", pre_df)
                    print("post_df \n", reorder_post_df)

                    # Concatenate the DataFrames
                    df = pd.concat([pre_df, reorder_post_df], ignore_index=True, axis=1)

                    sheet_name_val[command] = df

                else:
                    # Handle the case where DataFrames are empty
                    # print("Empty DataFrame encountered")
                    sheet_name_val[command] = (
                        pd.DataFrame()
                    )  # Set an empty DataFrame or handle it as needed
            except Exception as e:
                print(pre_date_ip_info, post_date_ip_info)
                UNPROCESSED_FILES.append((pre_file, post_file))
                print(f"Error processing DataFrame columns: {e}")
        return sheet_name_val

    df_dict = {
        "st cell": pd.DataFrame(),
        "cvls": pd.DataFrame(),
        "altk": pd.DataFrame(),
        "hget . tss": pd.DataFrame(),
        "pr EUtranFreqRelation=": pd.DataFrame(),
        "get EutrancellTDD=.* Earfcn": pd.DataFrame(),
        "get EutrancellFDD=.* Earfcn": pd.DataFrame(),
        "get ^EutrancellTDD=.* ChannelBandwidth": pd.DataFrame(),
        "get ^EutrancellFDD=.* ChannelBandwidth": pd.DataFrame(),
        "get . prbpairsperframe": pd.DataFrame(),
        "hget ^EUtranCell*.* cellBARred|sectorCARrierRef|primARyPlmnReserved|state|eARfcndl|eARfcnUL|^cellId": pd.DataFrame(),
    }

    all_sheet_data = {}
    for pre_file, post_file in matched_files:
        pre_file_path = os.path.join(new_pre_folder, pre_file)
        post_file_path = os.path.join(new_post_folder, post_file)

        with open(pre_file_path, "r") as file:
            pre_file_content = file.readlines()

        with open(post_file_path, "r") as file:
            post_file_content = file.readlines()

        with open(pre_file_path, "r") as file:
            file = file.readlines()
            tdd_df = extract_tdd_fdd_data("TDD", file, False)
            fdd_df = extract_tdd_fdd_data("FDD", file, False)

        tdd_mapping_dict = {
            last: freq for last, freq in zip(tdd_df["MO"], tdd_df["Value"])
        }
        fdd_df = fdd_df[fdd_df["Attribute"] == "earfcndl"]

        fdd_mapping_dict = {
            last: freq for last, freq in zip(fdd_df["MO"], fdd_df["Value"])
        }

        list_commands = [
            "st cell",
            "get AntennaUnitGroup=*.*,AntennaNearUnit=*.*,RetSubUnit=*.*   userLabel",
            "cvls",
            "altk",
            "hget . tss",
            "pr EUtranFreqRelation=",
            "get EutrancellTDD=.* Earfcn",
            "get EutrancellFDD=.* Earfcn",
            "get ^EutrancellTDD=.* ChannelBandwidth",
            "get ^EutrancellFDD=.* ChannelBandwidth",
            "get . prbpairsperframe",
            "hget ^EUtranCell*.* cellBARred|sectorCARrierRef|primARyPlmnReserved|state|eARfcndl|eARfcnUL|^cellId",
        ]
        tdd_df = pd.DataFrame()
        fdd_df = pd.DataFrame()
        sheet_name_val = process_files(
            pre_file_content,
            post_file_content,
            list_commands,
            fdd_mapping_dict,
            tdd_mapping_dict,
        )

        for key, df in sheet_name_val.items():
            key = key.split("*")[0] if "*" in key else key
            original_columns = df.columns
            if key in df_dict:
                df_dict[key] = pd.concat([df_dict[key], df], axis=0)
            else:
                df_dict[key] = df
    all_sheet_data = {key: df for key, df in df_dict.items() if not df.empty}

    output_folder = os.path.join(settings.MEDIA_ROOT, "NOM_AUDIT", "output")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    excel_path = os.path.join(output_folder, "Pre_Post_File_logs.xlsx")
    with pd.ExcelWriter(excel_path, engine="xlsxwriter") as writer:
        for key, df in all_sheet_data.items():
            sheet_name = key.split("*")[0] if "*" in key else key
            df.to_excel(writer, sheet_name=sheet_name)

    wb = load_workbook(excel_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        num_cols = ws.max_column

        ws.insert_rows(1)
        pre_row = ["PRE"] + [""] * (num_cols // 2 - 1)
        post_row = ["POST"] + [""] * (num_cols // 2 - 1)

        for col_num, value in enumerate(pre_row, 1):
            cell = ws.cell(row=1, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            # cell.fill = pre_fill

        for col_num, value in enumerate(post_row, len(pre_row) + 1):
            cell = ws.cell(row=1, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            # cell.fill = post_fill

    if len(pre_row) <= num_cols:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(pre_row))
    else:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

    if len(pre_row) + 1 <= num_cols:
        ws.merge_cells(
            start_row=1, start_column=len(pre_row) + 1, end_row=1, end_column=num_cols
        )

        wb.save(excel_path)

    audit_sheet_df = {}
    mapped_columns = {}
    for sheet_name in wb.sheetnames:
        data_df = pd.read_excel(excel_path, sheet_name=sheet_name, header=[0, 1])
        # print(data_df.columns)

        if sheet_name == "st cell":
            df = st_cell_log_audit(data_df)
            mapped_columns[sheet_name] = df.columns.tolist()
        # elif sheet_name == "get EutrancellTDD=.":
        #     df = get_EutrancellTDD_audit(data_df)
        #     mapped_columns[sheet_name] = df.columns.tolist()
        # elif sheet_name == "get EutrancellFDD=.":
        #     df = get_EutrancellFDD_audit(data_df)
        #     mapped_columns[sheet_name] = df.columns.tolist()
        # elif sheet_name == "get ^EutrancellTDD=.":
        #     df = getEutrancellTDD_audit(data_df)
        #     mapped_columns[sheet_name] = df.columns.tolist()
        # elif sheet_name == "get ^EutrancellFDD=.":
        #     df = getEutrancellFDD_audit(data_df)
        #     mapped_columns[sheet_name] = df.columns.tolist()
        # elif sheet_name == "get . prbpairsperframe":
        #     df = get_prbpairsperframe(data_df)
        #     mapped_columns[sheet_name] = df.columns.tolist()
        # elif sheet_name == "pr EUtranFreqRelation=":
        #     df = pr_EUtranFreqRelation(data_df)
        #     mapped_columns[sheet_name] = df.columns.tolist()
        # elif sheet_name == "get AntennaUnitGroup=":
        #     df = get_AntennaUnitGroup(data_df)
        #     mapped_columns[sheet_name] = df.columns.tolist()
        # elif sheet_name == "hget . tss":
        #     df = hget_tss(data_df)
        #     mapped_columns[sheet_name] = df.columns.tolist()
        elif sheet_name == "altk":
            df = altk(data_df)
            mapped_columns[sheet_name] = df.columns.tolist()
        # elif sheet_name == "hget ^EUtranCell":
        #     df = hget_EUtranCell(data_df)
        #     mapped_columns[sheet_name] = df.columns.tolist()
        else:
            continue
        audit_sheet_df[sheet_name] = df
    audit_file_name = "audit_pre_post.xlsx"
    audit_file_path = os.path.join(output_folder, audit_file_name)
    all_data_df = {}

    def make_unique_columns(columns):
        seen = {}
        new_columns = []
        for col in columns:
            if col in seen:
                seen[col] += 1
                new_columns.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                new_columns.append(col)
        return new_columns

    with pd.ExcelWriter(audit_file_path, engine="xlsxwriter") as writer:
        for key, df in audit_sheet_df.items():
            sheet_name = key.split("*")[0] if "*" in key else key
            df.to_excel(writer, sheet_name=sheet_name)
            new_cols = []
            for val in df.columns:
                col_name = f"{val[0]}_{val[1]}"
                if " " in col_name:
                    col_name = f"{col_name.replace(' ', '_')}"

                if "&" in col_name:
                    col_name = f"{col_name.replace('&', '')}"

                if "(" in col_name:
                    col_name = f"{col_name.replace('(', '_')}"

                if ")" in col_name:
                    col_name = f"{col_name.replace(')', '')}"

                if "." in col_name:
                    col_name = f"{col_name.replace('.', '_')}"

                new_cols.append(col_name)

            df.columns = new_cols

            name = "_".join(key.split(" ")) if " " in key else key
            df.columns = make_unique_columns(df.columns)
            all_data_df[name] = json.loads(df.to_json(orient="records"))

    wb = load_workbook(audit_file_path)
    pre_fill = PatternFill(fgColor="538DD5", fill_type="solid")
    post_fill = PatternFill(fgColor="66CCFF", fill_type="solid")
    audit_fill = PatternFill(fgColor="DA9694", fill_type="solid")
    pre_columns_fill = PatternFill(fgColor="CCECFF", fill_type="solid")
    post_columns_fill = PatternFill(fgColor="C7DBEF", fill_type="solid")
    audit_columns_fill = PatternFill(fgColor="E6B8B7", fill_type="solid")
    max_col = ws.max_column

    # print(max_col)

    # exit(0)

    def find_range_pre_post_audit(ws):
        pre_start_col, pre_end_col = None, None
        post_start_col, post_end_col = None, None
        audit_start_col, audit_end_col = None, None

        # Iterate through the first row to find column headers
        for row in ws.iter_rows(min_row=1, max_row=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Detect the PRE range
                    if "PRE" in cell.value and pre_start_col is None:
                        pre_start_col = cell.column
                        for col in range(pre_start_col, ws.max_column + 1):
                            if ws.cell(row=cell.row, column=col).value == "POST":
                                pre_end_col = col - 1
                                break

                    # Detect the POST range
                    elif "POST" in cell.value and post_start_col is None:
                        post_start_col = cell.column
                        for col in range(post_start_col, ws.max_column + 1):
                            if ws.cell(row=cell.row, column=col).value == "audit":
                                post_end_col = col - 1
                                break

                    # Detect the AUDIT range
                    elif "audit" in cell.value and audit_start_col is None:
                        audit_start_col = cell.column
                        for col in range(audit_start_col, ws.max_column + 1):
                            if ws.cell(row=cell.row, column=col).value == "OverAll":
                                audit_end_col = col - 1
                                break

        return (
            (pre_start_col, pre_end_col),
            (post_start_col, post_end_col),
            (audit_start_col, audit_end_col),
        )

    mid_col = max_col // 2

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find the column ranges dynamically
        pre_range = find_range_pre_post_audit(ws)[0]
        post_range = find_range_pre_post_audit(ws)[1]
        audit_range = find_range_pre_post_audit(ws)[2]

        pre_start_col, pre_end_col = pre_range
        post_start_col, post_end_col = post_range
        audit_start_col, audit_end_col = audit_range

        print(
            pre_start_col,
            pre_end_col,
            post_start_col,
            post_end_col,
            audit_start_col,
            audit_end_col,
        )

        # Remove the third row (if needed)
        ws.delete_rows(3)

        pre_col = pre_start_col  # No need for chr(ord()), these are already integers
        cell = ws.cell(row=1, column=pre_col)
        cell.fill = pre_fill

        post_col = post_start_col
        cell = ws.cell(row=1, column=post_col)
        cell.fill = post_fill

        audit_col = audit_start_col
        cell = ws.cell(row=1, column=audit_col)
        cell.fill = audit_fill

        pre_range_str = (
            f"{get_column_letter(pre_start_col)}2:{get_column_letter(pre_end_col)}2"
        )
        for row in ws[pre_range_str]:
            for cell in row:
                cell.fill = pre_columns_fill

        post_range_str = (
            f"{get_column_letter(post_start_col)}2:{get_column_letter(post_end_col)}2"
        )
        for row in ws[post_range_str]:
            for cell in row:
                cell.fill = post_columns_fill

        audit_range_str = (
            f"{get_column_letter(audit_start_col)}2:{get_column_letter(audit_end_col)}2"
        )
        for row in ws[audit_range_str]:
            for cell in row:
                cell.fill = audit_columns_fill

        for row in ws.iter_rows(min_row=3):
            for cell in row:
                if cell.value == "OK":
                    cell.fill = PatternFill(fgColor="5ce65c", fill_type="solid")
                elif cell.value == "NOT OK":
                    cell.fill = PatternFill(fgColor="cd1c18", fill_type="solid")
                elif cell.value == "ERROR":
                    cell.fill = PatternFill(fgColor="ED4337", fill_type="solid")

    wb.save(audit_file_path)
    print(UNPROCESSED_FILES)
    audit_file_url = os.path.join(
        settings.MEDIA_URL, "NOM_AUDIT", "output", audit_file_name
    )

    return Response(
        {
            "Status": True,
            "message": "File converted successfully",
            "pre_post_json": all_data_df,
            "Download_url": audit_file_url,
        },
        status=status.HTTP_200_OK,
    )
