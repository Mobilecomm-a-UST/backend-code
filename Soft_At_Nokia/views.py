
import os
from .models import MOData  # Adjust the path if the model is in a different app
import pandas as pd
from .models import ExpectedParameter, SummaryData
import xml.etree.ElementTree as ET
from django.conf import settings
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from rest_framework.response import Response
from rest_framework import status
# from rest_framework import status
from .models import ExpectedParameter ,SummaryData , UserCounter , ExpectedParameter_5G, SummaryData_5G
from  mcom_website.settings import MEDIA_ROOT
from .serializers import ExpectedParameterSerializer,SummaryDataSerializer ,UserCounterSerializer, ExpectedParameterSerializer_5G, SummaryDataSerializer_5G
import json
from datetime import datetime
import re
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST
import math
import gzip
# from openpyxl import workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone

def format_excel(writer, sheet_name, df):

    worksheet = writer.sheets[sheet_name]

    # -------- COLORS --------
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    thick_side = Side(style='thick')
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    # ---------------- HEADER STYLE ----------------
    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=1, column=col)

        if cell.value:
            cell.font = Font(bold=True, color="000000")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center',
                                        vertical='center',
                                        wrap_text=True)
            cell.border = thin_border

    # Increase header row height
    worksheet.row_dimensions[1].height = 30

    # ---------------- DATA STYLE ----------------
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)

            cell.alignment = Alignment(horizontal='center',
                                    vertical='center')
            cell.border = thin_border

    # ---------------- AUTO COLUMN WIDTH ----------------
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        max_length = 0

        for row in range(1, max_row + 1):
            value = worksheet.cell(row=row, column=col).value
            if value:
                max_length = max(max_length, len(str(value)))

        worksheet.column_dimensions[column_letter].width = max_length + 4

    # ---------------- THICK OUTER BORDER ----------------
    for col in range(1, max_col + 1):
        # Top border
        worksheet.cell(row=1, column=col).border = Border(
            top=thick_side,
            left=thin_border.left,
            right=thin_border.right,
            bottom=thin_border.bottom
        )

        # Bottom border
        worksheet.cell(row=max_row, column=col).border = Border(
            bottom=thick_side,
            left=thin_border.left,
            right=thin_border.right,
            top=thin_border.top
        )

    for row in range(1, max_row + 1):
        # Left border
        worksheet.cell(row=row, column=1).border = Border(
            left=thick_side,
            top=thin_border.top,
            bottom=thin_border.bottom,
            right=thin_border.right
        )

        # Right border
        worksheet.cell(row=row, column=max_col).border = Border(
            right=thick_side,
            top=thin_border.top,
            bottom=thin_border.bottom,
            left=thin_border.left
        )

    worksheet.freeze_panes = "A2"
# -------------------- Utility Functions --------------------

def extract_site_id(dist_name):

    """Extract site id from distName attribute."""

    try:
        parts = dist_name.split('/')
        for part in parts:
            if part.startswith("MRBTS-"):
                return part.split('-')[1]
    except Exception:
        return None
    return None

def get_mrbts_site_map(root, ns):
    mrbts_site_map = {}

    # namespace-safe tags
    mo_path = ".//ns:managedObject[@class='MRBTS']" if ns else ".//managedObject"
    p_tag = "ns:p" if ns else "p"

    for mrbts in root.findall(mo_path, ns):
        dist_name = mrbts.attrib.get("distName", "")
        mrbts_id = extract_site_id(dist_name)

        bts_name = ""
        for p in mrbts.findall(p_tag, ns):
            if p.attrib.get("name", "").lower() == "btsname":
                bts_name = (p.text or "").strip()
                break

        # extract WBJ1503 from WB_ZQ_F_WBJ1503
        site_id = bts_name.split("_")[-1] if bts_name else ""

        if mrbts_id and site_id:
            mrbts_site_map[mrbts_id] = site_id

    return mrbts_site_map



def normalize_path(path):

    """Normalize MO path to lowercase and remove 'managedobject=' prefix if present."""

    if isinstance(path, str):
        path = path.strip().lower()
        if path.startswith("managedObject ="):
            path = path[len("managedObject ="):]
        return path
    return ''


def get_default_namespace(element):

    """Extract the default namespace URI from the root element tag."""
    if element.tag.startswith("{"):
        return element.tag[1:].split("}")[0]
    return ''

def format_excel_sheet(writer, sheet_name, df, startrow=0, startcol=0):
    import pandas as pd

    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # =========================
    # 🎨 FORMATS
    # =========================
    header_format = workbook.add_format({
        "bold": True, "bg_color": "#236A6D", "border": 2,
        "font_color": "#ffffff", "align": "center", "valign": "vcenter"
    })

    center_format = workbook.add_format({
        "align": "center", "valign": "center", "border": 1, "bold": True
    })

    ok_format = workbook.add_format({
        "bg_color": "#90EE90", "align": "center", "valign": "center"
    })

    not_ok_format = workbook.add_format({
        "bg_color": "#FF0000", "font_color": "#FFFFFF",
        "align": "center", "valign": "center"
    })

    green_fmt = workbook.add_format({"bg_color": "#B4E6BE", "border": 1})
    red_fmt   = workbook.add_format({"bg_color": "#E9AAB1", "border": 1})
    blank_fmt = workbook.add_format({"bg_color": "#FCF259", "font_color": "#222831",
                                    "align": "center", "valign": "center", "bold": True, "border": 1})

    worksheet.set_row(startrow, 23)

    # =========================
    # 🧾 HEADER + WIDTH
    # =========================
    for col_num, col_name in enumerate(df.columns):
        worksheet.write(startrow, startcol + col_num, str(col_name), header_format)
        max_length = max(
            df[col_name].astype(str).str.len().max(skipna=True) or 0,
            len(str(col_name)),
        )
        worksheet.set_column(startcol + col_num, startcol + col_num, min(max_length, 255) + 5)

    # =========================
    # 📊 NORMAL FORMATTING
    # =========================
    for row_num in range(len(df)):
        worksheet.set_row(startrow + row_num + 1, 15)
        for col_num in range(len(df.columns)):
            cell_value = df.iloc[row_num, col_num]
            fmt = center_format

            # Only highlight blank key columns
            key_cols = []
            if sheet_name == "RET Counter":
                key_cols = ["Mech. angle", "Min. angle", "Max. angle", "Angle"]
            elif sheet_name == "VSWR":
                key_cols = ["vswrMajorThreshold", "vswrMinorThreshold"]
            elif sheet_name == "Nomenclature 4G":
                key_cols = ["TAC"]
            elif sheet_name == "Nomenclature 2G":
                key_cols = ["LAC"]

            if df.columns[col_num] in key_cols and (pd.isna(cell_value) or str(cell_value).strip() == ""):
                fmt = blank_fmt
            elif str(cell_value) == "OK":
                fmt = ok_format
            elif str(cell_value) in ["NOT OK", "NOK"]:
                fmt = not_ok_format
            elif str(cell_value) in ["Missing", "Missing in Post"]:
                fmt = workbook.add_format({
                    "bg_color": "#FF6347", "font_color": "#FFFFFF",
                    "align": "center", "valign": "center", "bold": True, "border": 1
                })
            elif "|" in str(cell_value):
                fmt = workbook.add_format({
                    "font_color": "#FF0000", "align": "center",
                    "valign": "center", "bold": True, "border": 1
                })

            worksheet.write(row_num + 1, col_num, cell_value, fmt)

    # =========================
    # 🔥 RET LOGIC
    # =========================
    if sheet_name == "RET Counter":
        check_cols = ["Mech. angle", "Min. angle", "Max. angle", "Angle"]
        if all(col in df.columns for col in check_cols):
            col_index = {col: df.columns.get_loc(col) for col in check_cols}
            for site, group in df.groupby("Site_id"):
                rows = group.index.tolist()
                for col in check_cols:
                    values = df.loc[rows, col]
                    fmt = green_fmt if values.nunique() == 1 else red_fmt
                    col_idx = col_index[col]
                    for r in rows:
                        worksheet.write(r + 1, col_idx, df.iloc[r, col_idx], fmt)

    # =========================
    # 🔥 VSWR LOGIC
    # =========================
    if sheet_name == "VSWR":
        check_cols = ["vswrMajorThreshold", "vswrMinorThreshold"]
        if all(col in df.columns for col in check_cols):
            col_index = {col: df.columns.get_loc(col) for col in check_cols}
            for site, group in df.groupby("Site_id"):
                rows = group.index.tolist()
                for col in check_cols:
                    values = df.loc[rows, col]
                    fmt = green_fmt if values.nunique() == 1 else red_fmt
                    col_idx = col_index[col]
                    for r in rows:
                        worksheet.write(r + 1, col_idx, df.iloc[r, col_idx], fmt)

    # =========================
    # 🔥 TAC LOGIC (Nomenclature)
    # =========================
    if sheet_name == "Nomenclature 4G" and "TAC" in df.columns and "Site_id" in df.columns:
        tac_col_idx = df.columns.get_loc("TAC")
        for site, group in df.groupby("Site_id"):
            rows = group.index.tolist()
            valid_tac = group["TAC"].dropna()
            valid_tac = valid_tac[valid_tac.astype(str).str.strip() != ""]
            unique_tac = valid_tac.unique()
            majority_tac = valid_tac.mode()[0] if len(valid_tac) > 0 else None
            for r in rows:
                tac_value = df.iloc[r]["TAC"]
                if pd.isna(tac_value) or str(tac_value).strip() == "":
                    worksheet.write(r + 1, tac_col_idx, tac_value, red_fmt)
                elif len(unique_tac) > 1 and tac_value != majority_tac:
                    worksheet.write(r + 1, tac_col_idx, tac_value, red_fmt)

                

# --- ADDED: match_value() for flexible comparison ---
def match_value(actual, expected):
    actual = actual.strip().lower()
    expected = expected.strip().lower()
    if ';' in expected:
        options = [opt.split('(')[0].strip().lower() for opt in expected.split(';')]
        return any(actual == opt for opt in options)
    if "should be" in expected or "if" in expected:
        return expected.split()[0] in actual
    return actual == expected


def normalize_value(val):
    """Clean value: uppercase, strip spaces, remove special chars"""
    if not val:
        return ''
    
    val = val.strip().upper()
    digits = re.sub(r'\D', '', val)

    # For values like '30m' — return only digits if present
    if digits and val.lower().endswith('m'):
        return digits
    
    return re.sub(r'\W+', '', val)




def match_value(actual, expected):
    """Check if normalized actual is in any normalized expected values"""
    actual_norm = normalize_value(actual)
    expected_values = [normalize_value(val) for val in expected.split("or")]
    return actual_norm in expected_values

def matches_path(expected_path, dist_name):
    expected_parts = [p.split('-')[0].lower() for p in expected_path.split('/')]
    dist_parts = [p.split('-')[0].lower() for p in dist_name.split('/')]
    idx = 0
    for part in dist_parts:
        if idx < len(expected_parts) and part == expected_parts[idx]:
            idx += 1
    return idx == len(expected_parts)


# ------------------- Main API VieW ####################################################
def deduplicate_alarms(alarms):
    seen = set()
    unique_alarms = []
    for alarm in alarms:
        key = (alarm['file'], alarm['MRBTS'], alarm['mo_path'], alarm['parameter'])
        if key not in seen:
            seen.add(key)
            unique_alarms.append(alarm)
    return unique_alarms


# def api_usage_all(userId, api):
#     print(userId)
#     qs = UserCounter.objects.filter(user_name=userId, api_name=api).first()

#     if qs:
#         qs.count += 1
#         qs.save()
    
#     else:
#         qs = UserCounter.objects.create(user_name=userId, api_name=api, count=1)



# def api_usage_all(userId, api, site_id):
#     today = timezone.now().date()

#     obj, created = UserCounter.objects.get_or_create(
#         user_name=userId,
#         api_name=api,
#         site_id=site_id,
#         Date=today,
#         defaults={'count': 1}
#     )

#     if not created:
#         obj.count += 1
#         obj.save()




# @api_view(['GET', 'POST'])
# def user_count(request):
#     if request.method == 'GET':
#         qs = UserCounter.objects.all()
#         serializer = UserCounterSerializer(qs, many=True)
#         return Response(serializer.data)

#     elif request.method == 'POST':
#         serializer = UserCounterSerializer(data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def api_usage_all(userId, api, site_id):
    today = timezone.now().date()

    print("USER:", userId)
    # print("SITE:", site_id)
    print("Successfully Data received for API usage tracking........")   # 🔥 DEBUG

    qs = UserCounter.objects.filter(
        user_name=userId,
        api_name=api,
        Date=today   # ✅ same date condition
    ).first()

    if qs:
        qs.count += 1

        # existing site_ids split
        existing_sites = []
        if qs.site_id:
            existing_sites = [s.strip() for s in qs.site_id.split(",")]

        # ✅ add only if not duplicate
        if site_id and site_id not in existing_sites:
            existing_sites.append(site_id)
            qs.site_id = ",".join(existing_sites)

        qs.save()

    else:
        # new row (new date / new user / new api)
        UserCounter.objects.create(
            user_name=userId,
            api_name=api,
            site_id=site_id if site_id else "",
            count=1,
            Date=today
        )


@api_view(['GET', 'POST'])
def user_count(request):

    if request.method == 'GET':
        data = request.query_params
    else:
        data = request.data

    print("REQUEST DATA:", data)

    user = data.get('user')
    api_name = data.get('api')
    site_id = data.get('site_id')
    date_str = data.get('date')
    start_date = data.get('start_date')
    end_date = data.get('end_date')

    qs = UserCounter.objects.all()

    # -------- USER FILTER --------
    if user:
        qs = qs.filter(user_name=user)

    # -------- API FILTER --------
    if api_name:
        qs = qs.filter(api_name=api_name)

    # -------- SITE FILTER --------
    if site_id:
        qs = qs.filter(site_id__icontains=site_id)

    # -------- SINGLE DATE --------
    if date_str:
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
            qs = qs.filter(Date=date_obj)   # ⚠️ important
        except ValueError:
            return Response({"error": "Invalid date format"}, status=400)

    # -------- DATE RANGE --------
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
            qs = qs.filter(Date__range=[start, end])
        except ValueError:
            return Response({"error": "Invalid date range"}, status=400)

    print("FILTERED COUNT:", qs.count())   # 🔥 DEBUG

    serializer = UserCounterSerializer(qs, many=True)

    return Response({
        "count": qs.count(),
        "data": serializer.data
    })

@api_view(['GET', 'PUT', 'POST'])
def upload_and_compare_xml(request):
    userId = request.data.get('userId')
    api = "Checkpoints Data"
    if request.method == 'GET':
        expected_parameters = ExpectedParameter.objects.all()
        serializer = ExpectedParameterSerializer(expected_parameters, many=True)
        print(serializer.data)
        # api_usage_all(userId, api)
        return Response(serializer.data)
    
    elif request.method == 'POST':
            serializer= ExpectedParameterSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                api_usage_all(userId, api,"")
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'PUT':
        data = request.data
        serializer = ExpectedParameterSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            # api_usage_all(userId, api)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



# -------------------------expeted parameter------------

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_excel(request):
    userId = request.data.get('userId')
    api = "checkpoints Upload"
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return Response({"error": "Excel file is required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        df = pd.read_excel(excel_file)
        # df = df.dropna(subset=['MO Class', 'Parameter', 'Value'])

        ExpectedParameter.objects.all().delete()  


        errors = []

        for _, row in df.iterrows():
            data = {
            "path": str(row["path"]).strip(),
            "parameter_name": str(row["parameter_name"]).strip(),
            "expected_value": str(row["expected_value"]).strip(),
                            }
            serializer = ExpectedParameterSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
            else:
                errors.append(serializer.errors)

        if errors:
            return Response({"error": "Some rows failed validation", "details": errors}, status=status.HTTP_400_BAD_REQUEST)
        api_usage_all(userId, api,"")
        return Response({"status": True,"message": "Excel uploaded and saved successfully"},status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": f"Failed to process Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

# Upload XMLs and compare with database values

FIXED_PARAMETERS = [
    {"parameter": "actLBpowersaving", "mo_class": "MRBTS/LNBTS"},
    {"parameter": "gnssAntennaLatitude", "mo_class": "MRBTS/EQM_R-1/APEQM_R-1/CABINET_R-1/SMOD_R-1/GNSSE_R"},
    {"parameter": "gnssAntennaLongitude", "mo_class": "MRBTS/EQM_R-1/APEQM_R-1/CABINET_R-1/SMOD_R-1/GNSSE_R"},
    {"parameter": "activeSWReleaseVersion", "mo_class": "MNL_R"},
    {"parameter": "speedAndDuplex", "mo_class": "MRBTS/TNLSVC-1/TNL-1/ETHSVC-1/ETHLK-1"},
    {"parameter": "detectedSpeedAndDuplex", "mo_class": "MRBTS/TNLSVC-1/TNL-1/ETHSVC-1/ETHLK-1/ETHLK_R-1"},
    {"parameter": "holdOverModeUsed", "mo_class": "MRBTS/MNL/MNLENT/SYNC-1/CLOCK-1"},
    {"parameter": "syncInputType", "mo_class": "MRBTS/MNL/MNLENT/SYNC-1/CLOCK-1"},
    {"parameter": "transmissionrate", "mo_class": "MRBTS/EQM_R-1/APEQM_R-1/RMOD_R-1/SFP_R"},
    {"parameter": "autoDetectTotalDelayFromSHM", "mo_class": "MRBTS/MNL-1/MNLENT-1/SYNC-1/CLOCK_R-1"},
    {"parameter": "productName", "mo_class": "CABINET_R"}

]


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_and_compare_xml_files(request):
    userId = request.data.get('userId')
    api = "checkpoints Report Downloaded"

    xml_files = request.FILES.getlist('xml_files')
    circle_name = request.data.get('circle_name', 'default_circle').strip()
    project_type = request.data.get('project_type', 'project_type').strip()

    if not xml_files:
        return Response({"error": "At least one XML file is required"}, status=status.HTTP_400_BAD_REQUEST)

    # Load ExpectedParameter DB
    expected_params = ExpectedParameter.objects.all()
    serializer = ExpectedParameterSerializer(expected_params, many=True)
    expected_params_json = serializer.data

    # Normalize expected values
    expected_values = {}
    for param in expected_params_json:
        norm_path = normalize_path(param['path'])
        if norm_path not in expected_values:
            expected_values[norm_path] = {}
        expected_values[norm_path][param['parameter_name'].lower()] = param['expected_value']

    # Containers
    results = []
    all_alarms = []
    summary_rows = []
    ipmtu_rows = []
    nomenclature_4G = []
    ret_counter = []
    VSWR = []
    nomenclature_2G = []


    # ------------------------
    # Helper function
    # ------------------------
    def get_param_value(mo, param_name, ns_uri):
        p_tag = f"{{{ns_uri}}}p" if ns_uri else "p"
        list_tag = f"{{{ns_uri}}}list" if ns_uri else "list"
        item_tag = f"{{{ns_uri}}}item" if ns_uri else "item"

        # Search direct <p>
        for p in mo.findall(p_tag):
            if p.attrib.get("name", "").lower() == param_name:
                return (p.text or "").strip()

        # Search inside list/item/p
        for lst in mo.findall(list_tag):
            for item in lst.findall(item_tag):
                for p in item.findall(p_tag):
                    if p.attrib.get("name", "").lower() == param_name:
                        return (p.text or "").strip()

        return ""

    # ------------------------
    # MAIN XML LOOP
    # ------------------------
    for xml_file in xml_files:
        xml_file.seek(0)
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
        except ET.ParseError:
            results.append({
                "file": xml_file.name,
                "MRBTS": '',
                "site_id": '',
                "mo_path": '',
                "parameter": '',
                "actual_value": '',
                "expected_value": '',
                "status": "Failed to parse XML"
            })
            continue

    # ----------------------to map site id to MRBTS---------------------

        ns_uri = get_default_namespace(root)
        ns = {'ns': ns_uri} if ns_uri else {}

        mo_elements = root.findall('.//ns:managedObject', ns) if ns else root.findall('.//managedObject')

        mrbts_site_map = get_mrbts_site_map(root, ns)
        # print("MRBTS → SITE MAP:", mrbts_site_map)

    #--------------------------------------------------------------------------------------
        # 1. IPMTU VALIDATION
        REQUIRED_IPIF_DISTNAMES = [
            "MRBTS/TNLSVC-1/TNL-1/IPNO-1/IPIF-1",
            "MRBTS/TNLSVC-1/TNL-1/IPNO-1/IPIF-2",
            "MRBTS/TNLSVC-1/TNL-1/IPNO-1/IPIF-3",
        ]
        # First loop only for debug (optional)
        for mo in mo_elements:
            dist_name = mo.attrib.get("distName", "")
            site_id = extract_site_id(dist_name)
            site = mrbts_site_map.get(site_id, "")
            # print("Site_id of Fixed parameters:", site_id)


        # Now, for each site, iterate over each parameter rule
        for site_id, site in mrbts_site_map.items():

            for rule in FIXED_PARAMETERS:
                param = rule["parameter"].lower().strip()
                mo_class = rule["mo_class"]

                value_found = "Missing"

                # Find the MO that matches both site_id and mo_class
                for mo in mo_elements:
                    dist_name = mo.attrib.get("distName", "")

                    if (
                        extract_site_id(dist_name) == site_id
                        and matches_path(mo_class, dist_name)
                    ):
                        value_found = get_param_value(mo, param, ns_uri)
                        break  # Take first match

                summary_rows.append({
                    "file": xml_file.name,
                    "site_id": site,        # actual site name/value
                    "MRBTS": site_id,       # site ID
                    "mo_class": mo_class,
                    "parameter": param,
                    "value": value_found
                })
        for mo in mo_elements:
            if mo.attrib.get("class", "") != "IPIF":
                continue

            dist_name = mo.attrib.get("distName", "")
            site_id = extract_site_id(dist_name)
            site = mrbts_site_map.get(site_id, "")
            
            ipmtu_val = get_param_value(mo, "ipmtu", ns_uri)
            status_ip = "OK" if any(matches_path(req, dist_name) for req in REQUIRED_IPIF_DISTNAMES) else "Extra"

            ipmtu_rows.append({
                "file": xml_file.name,
                "Site_id":site,
                "MRBTS": site_id,
                "dist_name": dist_name,
                "ipMtu": ipmtu_val,
                "status": status_ip
            })


        alarms = []
        if project_type == 'uls':
            max_alarms_per_site = 4
        elif project_type == 'nt relocation':
            max_alarms_per_site = 12
        else:
            max_alarms_per_site = None

        site_alarm_counts = {}

        for mo in mo_elements:
            dist_name = mo.attrib.get('distName', '')
            site_id = extract_site_id(dist_name)
            site = mrbts_site_map.get(site_id, "")

            port_id = get_param_value(mo, "portid", ns_uri)
            descr = get_param_value(mo, "descr", ns_uri)
            polarity = get_param_value(mo, "polarity", ns_uri)

            if not (port_id and descr and polarity):
                continue

            if site_id not in site_alarm_counts:
                site_alarm_counts[site_id] = 0

            if max_alarms_per_site is None or site_alarm_counts[site_id] + 2 <= max_alarms_per_site:
                alarms.append({
                    "file": xml_file.name,
                    "Site_id":site,
                    "MRBTS": site_id,
                    "mo_path": dist_name,
                    "parameter": "descr",
                    "Status": descr
                })
                alarms.append({
                    "file": xml_file.name,
                    "Site_id":site,
                    "MRBTS": site_id,
                    "mo_path": dist_name,
                    "parameter": "polarity",
                    "Status": polarity
                })
                site_alarm_counts[site_id] += 2

        all_alarms.extend(alarms)

        # ---------------------------------------
        # 4. EXPECTED PARAMETER COMPARISON
        # ---------------------------------------
        for mo in mo_elements:
            dist_name = mo.attrib.get("distName", "")
            site_id = extract_site_id(dist_name)
            site = mrbts_site_map.get(site_id, "")

            for path, params in expected_values.items():
                if not matches_path(path, dist_name):
                    continue

                for param_name, expected_value in params.items():
                    param_name = param_name.lower().strip()

                    # Get all values for this parameter in the current MO
                    def get_all_param_values(mo, param_name, ns_uri):
                        values = []
                        p_tag = f"{{{ns_uri}}}p" if ns_uri else "p"
                        for p in mo.findall(p_tag):
                            if p.attrib.get("name", "").lower() == param_name:
                                values.append(p.text.strip() if p.text else "")
                        return values

                    actual_values = get_all_param_values(mo, param_name, ns_uri)

                    # Skip if parameter does not exist in MO
                    if not actual_values:
                        continue

                    # Pick the value that matches DB (expected_value), if any
                    matching_value = next((v for v in actual_values if match_value(v, expected_value)), None)

                    if matching_value:
                        results.append({
                            "file": xml_file.name,
                            "site_id": site,
                            "MRBTS": site_id,
                            "mo_path": path,
                            "parameter": param_name,
                            "actual_value": matching_value,
                            "expected_value": expected_value,
                            "status": "OK"
                        })


        # ---------------------------------------
        # 5. Nomenclature Validation (FIXED)
        # ---------------------------------------
        for mo in mo_elements:

            mo_class = mo.attrib.get("class")

            dist_name = mo.attrib.get("distName", "")
            site_id = extract_site_id(dist_name)
            site = mrbts_site_map.get(site_id, "")

            bts_name = ""
            cell_name = ""
            itemname = ""
            tac = ""

            p_tag = f"{{{ns_uri}}}p" if ns_uri else "p"

            # 👉 MRBTS → BTS NAME
            if mo_class == "MRBTS":
                for p in mo.findall(p_tag):
                    if p.attrib.get("name", "").lower() == "btsname":
                        bts_name = (p.text or "").strip()
                        break

            # 👉 LNCEL → CELL NAME + TAC
            elif mo_class == "LNCEL":
                for p in mo.findall(p_tag):
                    pname = p.attrib.get("name", "").lower()

                    if pname == "name":
                        cell_name = (p.text or "").strip()

                    elif pname == "tac":
                        tac = (p.text or "").strip()

            # 👉 LNBTS → siteTemplateName
            # elif mo_class == "LNBTS":
            #     for p in mo.findall(p_tag):
            #         if p.attrib.get("name", "").lower() == "sitetemplatename":
            #             itemname = (p.text or "").strip()
            #             break
            
            elif mo_class == "LNBTS":
                for p in mo.findall(p_tag):
                    if p.attrib.get("name", "").lower() == "sitetemplatename":
                        itemname = (p.text or "").strip()

                        if len(itemname) != 15:
                            itemname = f"{itemname[:15]} (Invalid Length)"

                        break

            if bts_name or cell_name or itemname or tac:
                nomenclature_4G.append({
                    "Site_id": site,
                    "MRBTS": site_id,
                    "BTS_NAME": bts_name,
                    "CELL_NAME": cell_name,
                    "SiteTemplateName": itemname,
                    "TAC": tac,
                })
        # ---------------------------------------
        # 2G Nomenclature
        # ---------------------------------------
        for mo in mo_elements:

            mo_class = mo.attrib.get("class")

            dist_name = mo.attrib.get("distName", "")
            site_id = extract_site_id(dist_name)
            site = mrbts_site_map.get(site_id, "")

            sector_name = ""
            bcf_name = ""
            mrbts_name = ""
            site_name = ""
            itemname = ""
            lac = ""

            p_tag = f"{{{ns_uri}}}p" if ns_uri else "p"
            if mo_class == "BTS":
                for p in mo.findall(p_tag):
                    pname = p.attrib.get("name", "").lower()

                    if pname == "name":
                        sector_name = (p.text or "").strip()
                        print("SECTOR_NAME:", sector_name)

                    # elif pname == "sitetemplatename":
                    #     itemname = (p.text or "").strip()
                    #     print("TEMPLATE:", itemname)
                    
                    elif pname == "sitetemplatename":
                        itemname = (p.text or "").strip()

                        # check length BEFORE trimming
                        is_invalid = len(itemname) > 15
                        # trim for display only
                        itemname = itemname[:15]
                        # store only value (no flag)
                    
                    elif pname == "locationareaidlac":
                        lac = (p.text or "").strip()
                        print("LAC:", lac)

            elif mo_class == "BCF":
                bcf_name = dist_name.split("BCF-")[-1]
                for p in mo.findall(p_tag):
                    pname = p.attrib.get("name", "").lower()
                    if pname == "sbtsid":
                        mrbts_name = (p.text or "").strip()
                        print("MRBTS_NAME:", mrbts_name)

                    elif pname == "name":
                        site_name = (p.text or "").strip()
                        print("SITE_NAME:", site_name)

            if bcf_name or sector_name or itemname or lac or mrbts_name or site_name:
                nomenclature_2G.append({
                    "SITE_NAME": site_name,
                    "MRBTS_NAME": mrbts_name,
                    "BCF": bcf_name,
                    "SECTOR_NAME": sector_name,
                    "SiteTemplateName": itemname,
                    "LAC": lac,
                })



        for mo in mo_elements:

            if mo.attrib.get("class") != "RETU_R":
                continue

            dist_name = mo.attrib.get("distName", "")
            site_id = extract_site_id(dist_name)
            site = mrbts_site_map.get(site_id, "")

            # -------- Extract ALD and RETU --------
            ald_match = re.search(r"(ALD_R-\d+)", dist_name)
            retu_match = re.search(r"(RETU_R-\d+)", dist_name)

            ALD = ald_match.group(1) if ald_match else ""
            RET_Unit_ID = retu_match.group(1) if retu_match else ""

            # -------- Extract parameters --------
            values = {
                "angle": "",
                "maxangle": "",
                "minangle": "",
                "mechanicalangle": ""
            }

            p_tag = f"{{{ns_uri}}}p" if ns_uri else "p"

            for p in mo.findall(p_tag):
                pname = p.attrib.get("name", "").lower()
                if pname in values:
                    values[pname] = (p.text or "").strip()

            # -------- Append row --------
            ret_counter.append({
                "Site_id": site,
                "MRBTS": site_id,
                "ALD": ALD,
                "RET Unit ID": RET_Unit_ID,
                "Mech. angle": values["mechanicalangle"],
                "Min. angle": values["minangle"],
                "Max. angle": values["maxangle"],
                "Angle": values["angle"],
            })

        # #-------------------------------------
        # # 6. vswr
        # #-------------------------------------
        for mo in mo_elements:

            mo_class = mo.attrib.get("class")
            if mo_class != "ANTL":
                continue

            dist_name = mo.attrib.get("distName", "")
            site_id = extract_site_id(dist_name)
            site = mrbts_site_map.get(site_id, "")

            # -------- Extract ANT and RMOD --------
            rmod = re.search(r"RMOD-\d+", dist_name)
            antl = re.search(r"ANTL-\d+", dist_name)

            RMOD = rmod.group(0) if rmod else ""
            ANT = antl.group(0) if antl else ""

            # print("Processing:", ANT, RMOD, "Site:", site)


            vswrMajor = ""
            vswrMinor = ""

            for p in mo.findall(p_tag):
                pname = p.attrib.get("name", "").lower()

                if pname == "vswrmajorthreshold":
                    vswrMajor = (p.text or "").strip()

                elif pname == "vswrminorthreshold":
                    vswrMinor = (p.text or "").strip()

            if vswrMajor or vswrMinor:
                VSWR.append({
                    "Site_id": site,
                    "MRBTS": site_id,
                    "RMOD": RMOD,
                    "ANT": ANT,
                    "vswrMajorThreshold": vswrMajor,
                    "vswrMinorThreshold": vswrMinor,
                })
    # ---------------------------------------
    # Deduplicate final outputs
    # ---------------------------------------
    all_alarms = deduplicate_alarms(all_alarms)

    df_results = pd.DataFrame(results).drop_duplicates(
        subset=["file","MRBTS", "mo_path", "parameter", "actual_value", "expected_value"],
        keep="first"
    )

    df_alarms = pd.DataFrame(all_alarms).drop_duplicates(
        subset=["file", "MRBTS", "mo_path", "parameter", "Status"],
        keep="first"
    )

    df_summary = pd.DataFrame(summary_rows).drop_duplicates()

    df_ipmtu = pd.DataFrame(ipmtu_rows).drop_duplicates(
        subset=["file", "MRBTS","dist_name", "ipMtu", "status"],
        keep="first"
    )
    df_nomenclature_4G = pd.DataFrame(nomenclature_4G).drop_duplicates()
    # ---------------- Fix Nomenclature 4G format ----------------
    if not df_nomenclature_4G.empty:

        fill_columns = [
            "Site_id",
            "MRBTS",
            "BTS_NAME",
            "SiteTemplateName"
        ]

        # upar wali values neeche fill karo
        df_nomenclature_4G[fill_columns] = (
            df_nomenclature_4G[fill_columns].replace("", pd.NA).ffill()
        )

        # sirf CELL_NAME wali rows rakho
        df_nomenclature_4G = df_nomenclature_4G[
            df_nomenclature_4G["CELL_NAME"].notna()
            & (df_nomenclature_4G["CELL_NAME"] != "")
        ]

        df_nomenclature_4G.reset_index(drop=True, inplace=True)

    df_nomenclature_2G = pd.DataFrame(nomenclature_2G).drop_duplicates()
    print("df_nomenclature_2G", df_nomenclature_2G  )
    
    # ---------------- Fix Nomenclature 2G format ----------------
    if not df_nomenclature_2G.empty:

        fill_columns = [
            "SITE_NAME",
            "MRBTS_NAME",
            "BCF",
            "SiteTemplateName",
            "LAC"
        ]

        # blank values ko NA bana ke forward fill
        df_nomenclature_2G[fill_columns] = (
            df_nomenclature_2G[fill_columns]
            .replace("", pd.NA)
            .ffill()
        )

        # sirf sector rows rakho
        df_nomenclature_2G = df_nomenclature_2G[
            df_nomenclature_2G["SECTOR_NAME"].notna()
            & (df_nomenclature_2G["SECTOR_NAME"] != "")
        ]

        df_nomenclature_2G.reset_index(drop=True, inplace=True)

    df_vswr = pd.DataFrame(VSWR)

    df_ret = pd.DataFrame(ret_counter).drop_duplicates()

    # ---------------------------------------
    # SAVE EXCEL REPORT
    # ---------------------------------------
    report_folder = os.path.join(settings.MEDIA_ROOT, "reports")
    os.makedirs(report_folder, exist_ok=True)

    filename = f"report_{circle_name}.xlsx"
    filepath = os.path.join(report_folder, filename)

    with pd.ExcelWriter(filepath, engine='xlsxwriter') as writer:
        df_results.to_excel(writer, index=False, sheet_name="Comparison")
        format_excel_sheet(writer, "Comparison", df_results)

        if not df_alarms.empty:
            df_alarms["sort_key"] = df_alarms["mo_path"].str.extract(r"EAC_IN-(\d+)$").astype(int)
            df_alarms["param_order"] = df_alarms["parameter"].map({"descr": 1, "polarity": 2})

            df_alarms = df_alarms.sort_values(
                by=["MRBTS", "sort_key", "param_order"],
                ascending=[True, True, True]
            ).drop(columns=["sort_key", "param_order"])


            # ---- Flattened Lists ----
            list1 = [
                'MAINS FAIL', 'Normally_open',
                'RECTIFIER FAIL', 'Normally_open',
                'SITE ON BATTERY', 'Normally_open',
                'AC FAIL or CANOPY FAN FAIL', 'Normally_open',
                'FIRE AND SMOKE', 'Normally_open',
                'DG ON LOAD', 'Normally_open',
                'DG FAILED TO START', 'Normally_open',
                'DG FAILED TO STOP', 'Normally_open',
                'HIGH TEMPERATURE', 'Normally_open',
                'DOOR OPEN', 'Normally_open',
                'LOW BATTERY VOLTAGE', 'Normally_open',
                'DG FUEL LEVEL LOW', 'Normally_open'
            ]

            list2 = [
                'ACOC DOOR OPEN', 'Normally_closed',
                'ACOC EXT FAN FAIL', 'Normally_closed',
                'ACOC INT FAN FAIL', 'Normally_closed',
                'ACOC OVER TEMPERATURE', 'Normally_open',
                'MAINS FAIL', 'Normally_open',
                'SITE ON BATTERY', 'Normally_open',
                'LOW BATTERY VOLTAGE', 'Normally_open',
                'DG ON LOAD', 'Normally_open',
                'DG FUEL LEVEL LOW', 'Normally_open',
                'DG FAILED TO START', 'Normally_open',
                'HIGH TEMPERATURE', 'Normally_open',
                'FIRE AND SMOKE', 'Normally_open'
            ]

            valid_orders = {
                "list1": list1,
                "list2": list2,
                "list1_list2": list1 + list2,
                "list2_list1": list2 + list1,
            }

            # ---- Detect order ----
            def detect_order(status_values, valid_orders):
                scores = {}
                for name, order in valid_orders.items():
                    score = 0
                    for s, o in zip(status_values, order):
                        if s == o:
                            score += 1
                        else:
                            break
                    scores[name] = score
                return max(scores, key=scores.get)

            status_list = df_alarms["Status"].astype(str).tolist()
            detected_order = detect_order(status_list, valid_orders)
            expected_sequence = valid_orders[detected_order]

            status_ok = []
            seq_len = len(expected_sequence)

            for idx, actual in enumerate(status_list):

                # 🔹 normal range → PURANA behaviour
                if idx < seq_len:
                    status_ok.append(actual == expected_sequence[idx])

                # 🔁 repeat start → cyclic allow
                else:
                    expected = expected_sequence[idx % seq_len]
                    status_ok.append(actual == expected)


            df_alarms["Status_OK"] = status_ok

            # ---- Write alarms sheet ----
            sheet_name = "Alarms"
            df_alarms.to_excel(writer, index=False, sheet_name=sheet_name)
            format_excel_sheet(writer, sheet_name, df_alarms)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            green_fmt = workbook.add_format({
                "bg_color": "#4DC765",
                "border": 1
            })
            red_fmt = workbook.add_format({
                "bg_color": "#DB6976",
                "border": 1
            })

            status_col_idx = df_alarms.columns.get_loc("Status")
            ok_col_idx = df_alarms.columns.get_loc("Status_OK")

            # ---- Apply coloring ----
            for row_idx in range(len(df_alarms)):
                fmt = green_fmt if df_alarms.iloc[row_idx, ok_col_idx] else red_fmt
                worksheet.write(
                    row_idx + 1,
                    status_col_idx,
                    df_alarms.iloc[row_idx, status_col_idx],
                    fmt
                )

            # ---- Hide helper column ----
            worksheet.set_column(ok_col_idx, ok_col_idx, None, None, {"hidden": True})


        if not df_summary.empty:
            df_summary.to_excel(writer, index=False, sheet_name="FixedParameters")
            format_excel_sheet(writer, "FixedParameters", df_summary)

        if not df_ipmtu.empty:
            df_ipmtu.to_excel(writer, index=False, sheet_name="ipMtu Status")
            format_excel_sheet(writer, "ipMtu Status", df_ipmtu)
        
        if not df_nomenclature_4G.empty:
            df_nomenclature_4G.to_excel(writer, index=False, sheet_name="Nomenclature 4G")
            format_excel_sheet(writer, "Nomenclature 4G", df_nomenclature_4G)
            workbook = writer.book
            worksheet = writer.sheets["Nomenclature 4G"]

            red_fmt = workbook.add_format({
                "bg_color": "#DB6976",
                "border": 1
            })

            template_col = df_nomenclature_4G.columns.get_loc("SiteTemplateName")

            for row_idx in range(len(df_nomenclature_4G)):
                value = str(df_nomenclature_4G.iloc[row_idx, template_col])

                # highlight invalid length
                if len(value) != 15:
                    worksheet.write(row_idx + 1, template_col, value, red_fmt)
                    
                    
        if not df_nomenclature_2G.empty:
            df_nomenclature_2G.to_excel(writer, index=False, sheet_name="Nomenclature 2G")
            format_excel_sheet(writer, "Nomenclature 2G", df_nomenclature_2G)

            workbook = writer.book
            worksheet = writer.sheets["Nomenclature 2G"]

            red_fmt = workbook.add_format({
                "bg_color": "#DB6976",
                "border": 1
            })

            template_col = df_nomenclature_2G.columns.get_loc("SiteTemplateName")

            for row_idx in range(len(df_nomenclature_2G)):
                value = str(df_nomenclature_2G.iloc[row_idx, template_col])

                # highlight invalid length
                if len(value) != 15:
                    worksheet.write(row_idx + 1, template_col, value, red_fmt)
                    
        
            
            
        if not df_ret.empty:
            df_ret = df_ret[
                [
                    "Site_id",
                    "MRBTS",
                    "ALD",
                    "RET Unit ID",
                    "Mech. angle",
                    "Min. angle",
                    "Max. angle",
                    "Angle"
                ]
            ]
            df_ret.to_excel(writer, index=False, sheet_name="RET Counter")
            format_excel_sheet(writer, "RET Counter", df_ret)
        
        if not df_vswr.empty:
            df_vswr.to_excel(writer, index=False, sheet_name="VSWR")
            format_excel_sheet(writer, "VSWR", df_vswr)
        # print("==========ret=======",df_ret)


    all_sites = []
    # from ipmtu
    for row in ipmtu_rows:
        if row.get("Site_id"):
            all_sites.append(row["Site_id"])

    # from summary
    for row in summary_rows:
        if row.get("site_id"):
            all_sites.append(row["site_id"])

    # from nomenclature_4G
    for row in nomenclature_4G:
        if row.get("Site_id"):
            all_sites.append(row["Site_id"])
        
    for row in nomenclature_2G:
        if row.get("Site_id"):
            all_sites.append(row["Site_id"])

    # from alarms
    for row in all_alarms:
        if row.get("Site_id"):
            all_sites.append(row["Site_id"])

    unique_sites = list(set(all_sites))
    site_id_str = ",".join(unique_sites)
    # print("FINAL SITE LIST:", site_id_str)

    api_usage_all(userId, api, site_id_str)

    return Response({
        "message": f"Report generated successfully for circle: {circle_name}",
        "download_link": request.build_absolute_uri(settings.MEDIA_URL + 'reports/' + filename),
        "status": True
    })


##################################################### summary ##############################
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_Summary_excel(request):
    userId = request.data.get('userId')
    api = "Summary Upload"

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return Response({"error": "Excel file is required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        df = pd.read_excel(excel_file)
        df = df.dropna(subset=['MO Class', 'Parameter'])
        
        SummaryData.objects.all().delete()
        errors = []

        for _, row in df.iterrows():
            data = {
                "MO_Class": str(row['MO Class']).strip(),
                "Parameter": str(row['Parameter']).strip(),
            
            }
            serializer = SummaryDataSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
            else:
                errors.append(serializer.errors)

        if errors:
            return Response({"error": "Some rows failed validation", "details": errors}, status=status.HTTP_400_BAD_REQUEST)
        api_usage_all(userId, api, "")
        return Response({"status": True,"message": "Excel uploaded and saved successfully"},status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": f"Failed to process Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
    


    
@api_view(['GET','POST','PUT'])
def get_summary_data(request):
    userId = request.data.get('userId')
    api = "Summary Data"
    if request.method == 'GET':
        summary_data = SummaryData.objects.all()
        serializer = SummaryDataSerializer(summary_data, many=True)
        # api_usage_all(userId, api)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = SummaryDataSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            api_usage_all(userId, api)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'PUT':
        try:
            param_id = request.data.get('id')
            if not param_id:
                return Response({'error': 'ID is required for update.'}, status=status.HTTP_400_BAD_REQUEST)

            summary_data = SummaryData.objects.get(id=param_id)
        except SummaryData.DoesNotExist:
            return Response({'error': 'SummaryData not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = SummaryDataSerializer(summary_data, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            # api_usage_all(userId, api)
            return Response(serializer.data)
        return Response(serializer.errors,status=status.HTTP_400_BAD_REQUEST)
    

    
####################################### summary report##########################################################################
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_summary_xml_files(request):
    userId = request.data.get('userId')
    api = "Summary Report Downloaded"

    xml_files = request.FILES.getlist('xml_files')
    circle_name = request.data.get('circle_name', 'default_circle').strip()
    project_type = request.data.get('project_type', 'project_type').strip()

    if not xml_files:
        return Response({"error": "At least one XML file is required"}, status=status.HTTP_400_BAD_REQUEST)

    # Paths
    base_media_url = os.path.join(MEDIA_ROOT, "Soft_AT_Nokia")
    template_file_path = os.path.join(base_media_url, "template_file")
    output_dir = os.path.join(base_media_url, 'output_files')

    os.makedirs(template_file_path, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    template_path = os.path.join(template_file_path, "Sample Site_Soft AT.xlsx")
    output_path = os.path.join(output_dir, "AT_Summary_Filled.xlsx")

    # Load template ONCE (important)
    base_template_df = pd.read_excel(template_path, engine='openpyxl', sheet_name='Sheet1')

    all_dataframes = []

    # -----------------------------
    #       XML Extractor
    # -----------------------------
    def extract_data_from_xml(xml_file):
        tree = ET.parse(xml_file)
        root = tree.getroot()

        # Namespace
        ns = {}
        if '}' in root.tag:
            ns = {'ns': root.tag.split('}')[0].strip('{')}

        def find(path): return root.find(path, ns)
        def findall(path): return root.findall(path, ns)

        all_sites = []

        for mrbts in findall(".//ns:managedObject[@class='MRBTS']"):
            data = {}

            # Dictionary for MRBTS
            p_dict = {
                p.attrib.get("name"): (p.text.strip() if p.text else "")
                for p in mrbts.findall("ns:p", ns)
            }

            dist_name = mrbts.attrib.get("distName", "")
            mrbts_id = dist_name.split("MRBTS-")[-1] if "MRBTS-" in dist_name else ""

            data["MRBTS_ID"] = mrbts_id
            data["MRBTS_Name"] = p_dict.get("btsName", "")

            # Site ID from name
            site_name = p_dict.get("btsName", "")
            data["Site_ID"] = site_name.split('_')[-1] if site_name else ""

            data["latitude"] = p_dict.get("latitude", "")
            data["longitude"] = p_dict.get("longitude", "")

            # ----------------- Latitude & Longitude -----------------
            # data["latitude"] = ""
            # for latitude in findall(".//ns:managedObject[@class='GNSSE_R']"):
            #     if mrbts_id in latitude.attrib.get("distName", ""):
            #         for p in latitude.findall("ns:p", ns):
            #             if p.attrib.get("name") == "gnssAntennaLatitude":
            #                 data["latitude"] = p.text.strip()
            #                 break
                    

            # data["longitude"] = ""
            # for longitude in findall(".//ns:managedObject[@class='GNSSE_R']"):
            #     if mrbts_id in longitude.attrib.get("distName", ""):
            #         for p in longitude.findall("ns:p", ns):
            #             if p.attrib.get("name") == "gnssAntennaLongitude":
            #                 data["longitude"] = p.text.strip()
            #                 break


            # ----------------- BCF -----------------
            bcfid = ""
            for gnbcf in findall(".//ns:managedObject[@class='GNBCF']"):
                if "distName" in gnbcf.attrib and mrbts_id in gnbcf.attrib["distName"]:
                    for p in gnbcf.findall("ns:p", namespaces=ns):
                        if p.attrib.get("name") == "bcfId":
                            bcfid = p.text.strip() if p.text else ""
                            break
            data["bcfid"] = bcfid

            # ----------------- SW Version -----------------
            data["swVersion"] = ""
            for hw in findall(".//ns:managedObject[@class='MNL_R']"):
                if mrbts_id in hw.attrib.get("distName", ""):
                    for p in hw.findall("ns:p", ns):
                        if p.attrib.get("name") == "activeSWReleaseVersion":
                            data["swVersion"] = p.text.strip()
                            break

            # ----------------- MPlane IP -----------------
            data["Mplane"] = ""

            for ip in root.findall(".//ns:managedObject", ns):
                dist = ip.attrib.get("distName", "")

                # Check for correct class AND IPIF-3 in distName
                if "IPIF-3" in dist and mrbts_id in dist:
                    for p in ip.findall("ns:p", ns):
                        if p.attrib.get("name") == "localIpAddrAllocated":
                            data["Mplane"] = p.text.strip()
                            break


            # ----------------- Cells -----------------
            # data["cells"] = []
            # for cell in findall(".//ns:managedObject[@class='NRCELLGRP']"):
            #     if mrbts_id in cell.attrib.get("distName", ""):
            #         # find the nrCellList list
            #         nr_list = cell.find("ns:list[@name='nrCellList']", ns)
            #         if nr_list is not None:
            #             # each entry is a <p> element containing the cell ID
            #             for item in nr_list.findall("ns:p", ns):
            #                 data["cells"].append({"LNCEL_ID": item.text.strip()})

            # ----------------- RET Count -----------------
            data["RET Count"] = len([
                r for r in findall(".//ns:managedObject[@class='RET']")
                if mrbts_id in r.attrib.get("distName", "")
            ])

            # ----------------- TAC -----------------
            data["TAC"] = ""
            for tracking in findall(".//ns:managedObject[@class='LNCEL']"):
                if mrbts_id in tracking.attrib.get("distName", ""):
                    for p in tracking.findall("ns:p", ns):
                        if p.attrib.get("name") == "tac":
                            data["TAC"] = p.text.strip()
                            break
            # tracking = find(".//ns:managedObject[@class='LNCEL']")
            # if tracking:
            #     for p in tracking.findall("ns:p", ns):
            #         if p.attrib.get("name") == "tac":
            #             data["TAC"] = p.text.strip()
            #             break

            # ----------------- LAC -----------------
            # data["LAC"] = ""
            # lnbts = find(".//ns:managedObject[@class='LNRELG']")
            # if lnbts:
            #     for p in lnbts.findall("ns:p", ns):
            #         if p.attrib.get("name") == "lac":
            #             data["LAC"] = p.text.strip()
            #             break

            # ----------------- Sync Status -----------------
            data["Sync Status"] = "Not Available"
            for clock in findall(".//ns:managedObject[@class='CLOCK']"):
                if mrbts_id in clock.attrib.get("distName", ""):
                    sync_list = clock.find("ns:list[@name='syncInputList']", ns)
                    if sync_list is not None:
                        for item in sync_list.findall("ns:item", ns):
                            for p in item.findall("ns:p", ns):
                                # We only need syncInputType
                                if p.attrib.get("name") == "syncInputType" and p.text:
                                    value = p.text.strip()
                                    if "GNSS" in value:
                                        data["Sync Status"] = "GPS"
                                    elif "Sync Hub Master" in value:
                                        data["Sync Status"] = "HUB"
                                    else:
                                        data["Sync Status"] = value
                                    break

            # ----------------- DPR Cell Name -----------------
            data["DPR_Cell_Name"] = ""
            for lncel in findall(".//ns:managedObject[@class='LNCEL']"):
                dist = lncel.attrib.get("distName", "")
                if mrbts_id in dist:
                    for p in lncel.findall("ns:p", ns):
                        if p.attrib.get("name") == "name":
                            data["DPR_Cell_Name"] = p.text.strip()
                            break

            # ----------------- Bands -----------------
            band_mapping = {
                "T1":"L2300",
                "T2": "L2300",
                "F1": "L2100",
                "F3": "L1800",
                "F8": "L900",
                "G1": "G900",
                "G3": "G1800"
            }

            band_set = set()
            for lncel in findall(".//ns:managedObject[@class='LNCEL']"):
                if mrbts_id in lncel.attrib.get("distName", ""):
                    for p in lncel.findall("ns:p", ns):
                        if p.attrib.get("name") == "name" and p.text:
                            cell_name = p.text.strip()
                            for code, band in band_mapping.items():
                                if code in cell_name:
                                    band_set.add(band)

            data["Band"] = "_".join(sorted(band_set)) if band_set else ""

            # ----------------- LCR_ID (LNCEL IDs) -----------------
            lncel_ids = []
            for lncel in findall(".//ns:managedObject[@class='LNCEL']"):
                dist = lncel.attrib.get("distName", "")
                if mrbts_id in dist:
                    match = re.search(r"LNCEL-(\d+)", dist)
                    if match:
                        lncel_ids.append(match.group(1))

            data["LCR_ID"] = "&".join(lncel_ids)

            # ----------Determine Activity Type based on project_type---------------------------------
            Activity_Type = []
            if project_type.lower() == 'uls':
                Activity_Type.append('NEW_TOWER_ULS')
            elif project_type.lower() == 'nt':
                Activity_Type.append('NEW_TOWER')
            elif project_type.lower() == 'relocation':
                Activity_Type.append('NT_Relocation')
            elif project_type.lower() == 'upgrade':
                Activity_Type.append('L2100_UPGRADE')

            data["Activity_Type"] = ",".join(Activity_Type) if Activity_Type else ""


            # Add to site list
            all_sites.append(data)

        return all_sites

    # -----------------------------
    #       Process each XML
    # -----------------------------
    for xml_file in xml_files:
        sites_data = extract_data_from_xml(xml_file)

        for data in sites_data:
            template_df = base_template_df.copy()

            # Static fields
            template_df.loc[0, 'OEM'] = 'Nokia'
            template_df.loc[0, 'AoP'] = '2025-26'
            template_df.loc[0, 'AT_Type'] = 'Soft_AT'
            template_df.loc[0, 'TSP'] = 'Mobilecomm'
            template_df.loc[0, 'Offer/Reoffer'] = 'OFFER'
            template_df.loc[0, 'Project Remarks'] = '4G GPL parameters are as per the guidelines is OK.'
            template_df.loc[0, 'SMP_ID'] = 'NA'
            template_df.loc[0, 'Media '] = 'Microwave'
            template_df.loc[0, 'Processed_By'] = 'Mobilecomm'
            template_df.loc[0, 'Tech_info'] = 'NT'
            template_df.loc[0, 'Tech'] = 'NT'
            template_df.loc[0, 'Activity Type'] = data.get("Activity_Type", '')
            template_df.loc[0, 'MRBTS_Name'] = data.get('MRBTS_Name', '')
            template_df.loc[0, 'Site_ID'] = data.get('Site_ID', '')
            template_df.loc[0, 'BCF'] = data.get('bcfid', '')
            template_df.loc[0, 'Latitude (N)'] = data.get('latitude', '')
            template_df.loc[0, 'Longitude (E)'] = data.get('longitude', '')
            template_df.loc[0, 'Profile'] = data.get("swVersion",'')
            template_df.loc[0, 'TAC'] = data.get('TAC', '')
            # template_df.loc[0, 'LAC'] = data.get('LAC', '')
            template_df.loc[0, 'RET Count'] = data.get('RET Count', '')
            template_df.loc[0, 'Sync Status'] = data.get('Sync Status', '')
            # template_df.loc[0, 'LNCEL_ID'] = data['cells'][0]['LNCEL_ID'] if data.get("cells") else ''
            template_df.loc[0, 'DPR_Cell_Name'] = data.get('DPR_Cell_Name', '')
            template_df.loc[0, 'Band'] = data.get('Band', '')
 
            

            # Dynamic (from XML)
            for key, value in data.items():
                if key in template_df.columns:
                    template_df.loc[0, key] = value

            template_df.loc[0, 'Circle'] = circle_name

            all_dataframes.append(template_df)

    # -----------------------------
    # Final Excel Output
    # -----------------------------
    final_df = pd.concat(all_dataframes, ignore_index=True)
    final_df.drop_duplicates(inplace=True)
    final_df = final_df.fillna("")

    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        final_df.to_excel(writer, index=False, sheet_name='Sheet1')
        format_excel_sheet(writer, 'Sheet1', final_df)

    relative_output_path = os.path.relpath(output_path, MEDIA_ROOT)
    download_link = request.build_absolute_uri(os.path.join(MEDIA_URL, relative_output_path).replace('\\', '/'))
    
    # Extract unique site_ids from final_df
    site_ids = list(set(final_df['Site_ID'].tolist()))
    site_id_str = ",".join([str(s) for s in site_ids if s])

    # Call counter API
    api_usage_all(userId, api, site_id_str)

    return Response({
        "message": f"Report generated successfully for circle: {circle_name}",
        "download_link": download_link,
        "status": True
    }, status=status.HTTP_200_OK)


#----------------------------------------------------------
#------------------------------------------ 5G API's --------------------------------------


#-----------SUMMARY REPORT 5G------------------------------
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_summary_xml_files_5G(request):
    userId = request.data.get('userId')
    api = "Summary Report Downloaded 5G"
    # site_id = request.data.get('Site_ID')
    xml_files = request.FILES.getlist('xml_files')
    circle_name = request.data.get('circle_name', 'default_circle').strip()
    project_type = request.data.get('project_type', 'project_type').strip()

    if not xml_files:
        return Response({"error": "At least one XML file is required"}, status=status.HTTP_400_BAD_REQUEST)

    # Paths
    base_media_url = os.path.join(MEDIA_ROOT, "Soft_AT_Nokia")
    template_file_path = os.path.join(base_media_url, "template_file")
    output_dir = os.path.join(base_media_url, 'output_files')

    os.makedirs(template_file_path, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    template_path = os.path.join(template_file_path, "5G Soft AT.xlsx")
    output_path = os.path.join(output_dir, "AT_Summary_5G.xlsx")

    # Load template ONCE (important)
    # base_template_df = pd.read_excel(template_path, engine='openpyxl', sheet_name='Sheet1')
    template_sheets = pd.read_excel(template_path, engine='openpyxl', sheet_name=None)

    # Extract all 5 sheets
    sheet1_df = template_sheets.get("Sheet1")
    sheet2_df = template_sheets.get("Sheet2")
    pci_df = template_sheets.get("PCI Count")
    nrrel_df = template_sheets.get("NRREL OK")
    stencil_df = template_sheets.get("Stencil X2 Data")

    all_dataframes = []
    

    # -----------------------------
    #       XML Extractor
    # -----------------------------
    def extract_data_from_xml(xml_file):
        tree = ET.parse(xml_file)
        root = tree.getroot()

        # Namespace
        ns = {}
        if '}' in root.tag:
            ns = {'ns': root.tag.split('}')[0].strip('{')}

        def find(path): return root.find(path, ns)
        def findall(path): return root.findall(path, ns)

        all_sites = []

        for mrbts in findall(".//ns:managedObject[@class='MRBTS']"):
            data = {}

            # Dictionary for MRBTS
            p_dict = {
                p.attrib.get("name"): (p.text.strip() if p.text else "")
                for p in mrbts.findall("ns:p", ns)
            }

            dist_name = mrbts.attrib.get("distName", "")
            mrbts_id = dist_name.split("MRBTS-")[-1] if "MRBTS-" in dist_name else ""

            data["MRBTS_ID"] = mrbts_id
            data["MRBTS_Name"] = p_dict.get("btsName", "")

            # Site ID from name
            site_name = p_dict.get("btsName", "")
            data["Site_ID"] = site_name.split('_')[-1] if site_name else ""

            data["latitude"] = p_dict.get("latitude", "")
            data["longitude"] = p_dict.get("longitude", "")

            # ----------------- BCF -----------------
            gnbid = ""

            for gnbcf in root.findall(".//ns:managedObject[@class='LNADJGNB']", ns):
                dist_name = gnbcf.attrib.get("distName", "")

                if dist_name.endswith("LNADJGNB-0") and mrbts_id in dist_name:
                    for p in gnbcf.findall("ns:p", ns):
                        if p.attrib.get("name") == "adjGnbId":
                            gnbid = p.text.strip() if p.text else ""
                            break
                    break

            data["gnbid"] = gnbid


            # ----------------- SW Version -----------------
            data["swVersion"] = ""
            for hw in findall(".//ns:managedObject[@class='MNL_R']"):
                if mrbts_id in hw.attrib.get("distName", ""):
                    for p in hw.findall("ns:p", ns):
                        if p.attrib.get("name") == "activeSWReleaseVersion":
                            data["swVersion"] = p.text.strip()
                            break

            # ----------------- MPlane IP -----------------
            data["Mplane"] = ""
            # data["5G Node IP"]=""
            for ip in root.findall(".//ns:managedObject", ns):
                dist = ip.attrib.get("distName", "")

                # Check for correct class AND IPIF-3 in distName
                if "IPIF-3" in dist and mrbts_id in dist:
                    for p in ip.findall("ns:p", ns):
                        if p.attrib.get("name") == "localIpAddrAllocated":
                            data["Mplane"] = p.text.strip()
                            # data["5G Node IP"] = p.text.strip()
                            break


            # ----------------- Cells -----------------
            cells = []

            for cell in root.findall(".//ns:managedObject[@class='NRCELLGRP']", ns):
                if mrbts_id in cell.attrib.get("distName", ""):
                    nr_list = cell.find("ns:list[@name='nrCellList']", ns)
                    if nr_list is not None:
                        cells.extend(
                            p.text.strip() for p in nr_list.findall("ns:p", ns) if p.text
                        )

            data["cells"] = " & ".join(cells)

            # ----------------- RET Count -----------------
            data["RET Count"] = len([
                r for r in findall(".//ns:managedObject[@class='RET']")
                if mrbts_id in r.attrib.get("distName", "")
            ])

            # ----------------- TAC -----------------
            data["TAC"] = ""
            for tracking in findall(".//ns:managedObject[@class='LNCEL']"):
                if mrbts_id in tracking.attrib.get("distName", ""):
                    for p in tracking.findall("ns:p", ns):
                        if p.attrib.get("name") == "tac":
                            data["TAC"] = p.text.strip()
                            break

            # ----------------- Sync Status -----------------
            data["Sync Status"] = "Not Available"
            for clock in findall(".//ns:managedObject[@class='CLOCK']"):
                if mrbts_id in clock.attrib.get("distName", ""):
                    sync_list = clock.find("ns:list[@name='syncInputList']", ns)
                    if sync_list is not None:
                        for item in sync_list.findall("ns:item", ns):
                            for p in item.findall("ns:p", ns):
                                # We only need syncInputType
                                if p.attrib.get("name") == "syncInputType" and p.text:
                                    value = p.text.strip()
                                    if "GNSS" in value:
                                        data["Sync Status"] = "GPS"
                                    elif "Sync Hub Master" in value:
                                        data["Sync Status"] = "HUB"
                                    else:
                                        data["Sync Status"] = value
                                    break

            # ----------------- DPR Cell Name -----------------
            data["DPR_Cell_Name"] = ""
            for lncel in findall(".//ns:managedObject[@class='LNCEL']"):
                dist = lncel.attrib.get("distName", "")
                if mrbts_id in dist:
                    for p in lncel.findall("ns:p", ns):
                        if p.attrib.get("name") == "name":
                            data["DPR_Cell_Name"] = p.text.strip()
                            break

            # ----------------- LCR_ID (LNCEL IDs) -----------------
            lncel_ids = []
            for lncel in findall(".//ns:managedObject[@class='LNCEL']"):
                dist = lncel.attrib.get("distName", "")
                if mrbts_id in dist:
                    match = re.search(r"LNCEL-(\d+)", dist)
                    if match:
                        lncel_ids.append(match.group(1))

            data["LCR_ID"] = "&".join(lncel_ids)

            # ----------Determine Activity Type based on project_type---------------------------------
            Activity_Type = []
            if project_type.lower() == 'uls':
                Activity_Type.append('NEW_TOWER_ULS')
            elif project_type.lower() == 'nt':
                Activity_Type.append('NEW_TOWER')
            elif project_type.lower() == 'relocation':
                Activity_Type.append('NT_Relocation')
            elif project_type.lower() == 'upgrade':
                Activity_Type.append('L2100_UPGRADE')

            data["Activity_Type"] = ",".join(Activity_Type) if Activity_Type else ""


            # Add to site list
            all_sites.append(data)

        return all_sites
    
    circle_sgw_map = {
    "WB": [
        "2401:4900:40:200::1:48c",
        "2401:4900:44:2600::102",
        "2401:4900:0040:0200:0:0:0001:1500",
        "2401:4900:0040:0200:0:0:0001:1501",
        "2401:4900:0040:0200:0:0:0001:1502",
        "2401:4900:0040:0200:0:0:0001:1503",
        "2401:4900:0040:0200:0:0:0001:1504",
        "2401:4900:0040:0200:0:0:0001:1505",
        "2401:4900:40:3::1:c00",
        "2401:4900:0040:0003:0::0001:0942",
    ],
    "AP": [
        "2401:4900:50:200::1:48c",
        "2401:4900:50:3::1:c01"
    ],
}


    # -----------------------------
    #       Process each XML
    # -----------------------------
    for xml_file in xml_files:
        sites_data = extract_data_from_xml(xml_file)

        for data in sites_data:
            sheet1_df = template_sheets.get("Sheet1").copy()
            sheet2_df = template_sheets.get("Sheet2").copy()

            # Static fields
            sheet1_df.loc[0, 'OEM'] = 'Nokia'
            sheet2_df.loc[0, 'OEM'] = 'Nokia'
            sheet1_df.loc[0, 'AoP'] = '2025-26'
            sheet1_df.loc[0, 'AT_Type'] = 'Soft_AT'
            sheet2_df.loc[0, 'AT Type'] = 'Soft AT'
            sheet1_df.loc[0, 'TSP'] = 'Mobilecomm'
            sheet2_df.loc[0, 'TSP'] = 'Nokia'
            sheet1_df.loc[0, 'Offer/Reoffer'] = 'OFFER'
            sheet2_df.loc[0, 'Offer/Re-offer'] = 'OFFER'
            sheet1_df.loc[0, 'Project_Remarks'] = ' OK '
            sheet1_df.loc[0, 'Band'] = 'N3500'
            sheet2_df.loc[0, 'Band'] = 'N78'
            sheet1_df.loc[0, 'SMP_ID'] = 'NA'
            sheet1_df.loc[0, 'BSC'] = 'NA'
            sheet1_df.loc[0, 'Offered_Date'] = datetime.now().date()
            sheet2_df.loc[0, 'Offer Date'] = datetime.now().date()
            sheet1_df.loc[0, 'Media '] = 'Microwave'
            sheet1_df.loc[0, 'Ckt_Id'] = 'Microwave'
            sheet1_df.loc[0, 'Processed_By'] = 'Mobilecomm'
            sheet1_df.loc[0, 'Tech_info'] = '5G'
            sheet1_df.loc[0, 'Tech'] = '5G'
            sheet2_df.loc[0, 'Technology'] = 'Standalone'
            sheet1_df.loc[0, 'Cell_ID'] = 'NA'
            sheet1_df.loc[0, 'Activity'] ='5G Combo_Relocation'
            sheet1_df.loc[0, 'Activity'] ='5G-Standalone'
            sheet1_df.loc[0, 'MRBTS_Name'] = data.get('MRBTS_Name', '')
            sheet1_df.loc[0, 'Site ID (5G/)/MRBTS ID'] = data.get('MRBTS_Name', '')
            sheet1_df.loc[0, 'Site_ID'] = data.get('Site_ID', '')
            sheet2_df.loc[0, 'Site ID (2G)'] = data.get('Site_ID', '')
            sheet1_df.loc[0, 'Site ID (4G)'] = data.get('Site_ID', '')
            sheet1_df.loc[0, 'Reference_Id'] = f"{circle_name}_{data.get('Site_ID', '')}_5G_NA"
            sheet1_df.loc[0, 'BCF'] = 'NA'
            sheet1_df.loc[0, 'FDD_MRBTS_ID'] = 'NA'
            sheet1_df.loc[0, 'FDD_Mplane_IP'] = 'NA'
            sheet1_df.loc[0, 'FDD_Mplane_IP'] = 'NA'
            sheet1_df.loc[0, 'Nominal_Type'] = 'NA'
            sheet1_df.loc[0, 'Latitude (N)'] = data.get('latitude', '')
            sheet1_df.loc[0, 'Longitude (E)'] = data.get('longitude', '')
            sheet1_df.loc[0, 'Profile/SW Version'] = data.get("swVersion",'')
            sheet1_df.loc[0, 'SW Version'] = data.get("swVersion",'')
            sheet1_df.loc[0, 'TAC'] = data.get('TAC', '') 
            sheet1_df.loc[0, 'gNB_ID'] = data.get('gnbid', '') 
            sheet1_df.loc[0, 'Circle'] = circle_name
            sheet1_df.loc[0, "SGW IP"] = " ; ".join(circle_sgw_map.get(circle_name, []))
            sheet1_df.loc[0, 'Other Hardware Related Additional Information'] = '1-ABIO/2-AQQF'
            sheet1_df.loc[0, 'All Approved features & Compliance implemented'] = ' OK '
            sheet2_df.loc[0, 'Xn interface Configuration check'] = ' OK '
            sheet2_df.loc[0, 'Site should be on Latest Software release'] = ' OK '
            sheet2_df.loc[0, 'SCTP Association should be UP'] = ' OK '
            sheet2_df.loc[0, 'Ping test towards AMF/AMF enable snapshot'] = ' OK '
            sheet2_df.loc[0, 'Ping test towards UPF'] = ' Required-under discussion '
            sheet2_df.loc[0, '5G should be properly configured and should be enabled and unlocked in core nodes'] = ' OK '
            sheet2_df.loc[0, 'QIA Alarms : No active alarm ( except external alarm)/No false alarm in iOMS or OSS Monitor'] = ' OK '
            sheet1_df.loc[0, 'RET_Count'] = data.get('RET Count', '')
            sheet1_df.loc[0, 'Sync_Status'] = data.get('Sync Status', '')
            sheet1_df.loc[0, 'NRCEL_ID'] = data['cells'] if data.get("cells") else ''
            sheet1_df.loc[0, 'DPR_Cell_Name'] = data.get('DPR_Cell_Name', '')


            # Dynamic (from XML)
            for key, value in data.items():
                if key in sheet1_df.columns:
                    sheet1_df.loc[0, key] = value

            sheet1_df.loc[0, 'Circle'] = circle_name
            sheet2_df.loc[0, 'Circle'] = circle_name
        

            # all_dataframes.append(sheet1_df)
            all_dataframes.append(("Sheet1", sheet1_df))
            all_dataframes.append(("Sheet2", sheet2_df))

    # -----------------------------
    # Final Excel Output
    # -----------------------------
    # Combine all Sheet1 data
    sheet1_list = [df for name, df in all_dataframes if name == "Sheet1"]
    sheet2_list = [df for name, df in all_dataframes if name == "Sheet2"]

    final_sheet1_df = pd.concat(sheet1_list, ignore_index=True).fillna("")
    final_sheet2_df = pd.concat(sheet2_list, ignore_index=True).fillna("")

    # final_sheet1_df = pd.concat(all_dataframes, ignore_index=True)
    # final_sheet1_df.drop_duplicates(inplace=True)
    # final_sheet1_df = final_sheet1_df.fillna("")



    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        final_sheet1_df.to_excel(writer, sheet_name='Sheet1', index=False)
        final_sheet2_df.to_excel(writer, sheet_name='Sheet2', index=False)

        # Write other template sheets as-is
        # template_sheets.get("Sheet2").to_excel(writer, sheet_name='Sheet2', index=False)
        template_sheets.get("PCI Count").to_excel(writer, sheet_name='PCI Count', index=False)
        template_sheets.get("NRREL OK").to_excel(writer, sheet_name='NRREL OK', index=False)
        template_sheets.get("Stencil X2 Data").to_excel(writer, sheet_name='Stencil X2 Data', index=False)

        # 3️⃣ Apply formatting to ALL sheets
        format_excel(writer, 'Sheet1', final_sheet1_df)
        format_excel(writer, 'Sheet2', final_sheet2_df)
        format_excel(writer, 'PCI Count', template_sheets.get("PCI Count"))
        format_excel(writer, 'NRREL OK', template_sheets.get("NRREL OK"))
        format_excel(writer, 'Stencil X2 Data', template_sheets.get("Stencil X2 Data"))

    relative_output_path = os.path.relpath(output_path, MEDIA_ROOT)
    download_link = request.build_absolute_uri(os.path.join(MEDIA_URL, relative_output_path).replace('\\', '/'))
    api_usage_all(userId, api)
    return Response({
        "message": f"Report generated successfully for circle: {circle_name}",
        "download_link": download_link,
        "status": True
    }, status=status.HTTP_200_OK)



#-----------------------------------Checklist 5G APIs---------------------------------------

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_checklist_xml_files_5G(request):

    xml_file = request.FILES.get("xml_file")

    if not xml_file:
        return Response({"error": "XML file required"}, status=status.HTTP_400_BAD_REQUEST)

    # Expected external values
    external_values = {
        "actSliceSwitchToDefault": "1",
        "sd": "2",
        "sst": "0",
        "userLabel": "Premium"
    }

    rows = []

    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Namespace handling
    ns_uri = get_default_namespace(root)
    ns = {'ns': ns_uri} if ns_uri else {}

    mo_elements = root.findall('.//ns:managedObject', ns) if ns else root.findall('.//managedObject')

    mrbts_site_map = get_mrbts_site_map(root, ns)

    # Only these parameters required
    allowed_parameters = {
        "NRBTS": ["actSliceSwitchToDefault","x2LinkSupervisionTmr","actEnhancedX2LinkSupervision"],
        "SNSSAI": ["sd", "sst", "userLabel"]
    } 

    for mo in mo_elements:

        mo_class = mo.attrib.get("class", "")
        dist_name = mo.attrib.get("distName", "")

        if mo_class not in allowed_parameters:
            continue

        site_id = extract_site_id(dist_name)
        site = mrbts_site_map.get(site_id, "")

        snssai_id = ""

        # Extract SNSSAI ID
        if mo_class == "SNSSAI":
            if "SNSSAI-" in dist_name:
                snssai_id = dist_name.split("SNSSAI-")[-1]

            # Only SNSSAI ID = 2 required
            if snssai_id != "2":
                continue

        p_elements = mo.findall("ns:p", ns) if ns else mo.findall("p")

        for p in p_elements:

            param = p.attrib.get("name", "")

            if param not in allowed_parameters.get(mo_class, []):
                continue

            internal_value = (p.text or "").strip()
            external_value = external_values.get(param, "")

            # Custom logic for actSliceSwitchToDefault
            if param == "actSliceSwitchToDefault":
                if internal_value.lower() in ["true", "1"]:
                    status_result = "OK"
                else:
                    status_result = "NOK"
            else:
                status_result = "OK" if internal_value == external_value else "NOK"

            rows.append({
                "Site_ID": site,
                "MRBTS": site_id,
                "MO": mo_class,
                "ID": snssai_id,
                "Parameter": param,
                "Value(Internal)": internal_value,
                "Value(External)": external_value,
                "Status": status_result
            })

    df = pd.DataFrame(rows)

    # Save Excel report
    report_folder = os.path.join(settings.MEDIA_ROOT, "reports")
    os.makedirs(report_folder, exist_ok=True)

    filepath = os.path.join(report_folder, "5G-AT_Checklist.xlsx")

    with pd.ExcelWriter(filepath, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Comparison")
        format_excel_sheet(writer, "Comparison", df)

    return Response({
        "status": True,
        "message": "XML parsed successfully",
        "rows_generated": len(df),
        "file_path": filepath
    })
