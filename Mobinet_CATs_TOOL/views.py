import os
from django.conf import settings
from rest_framework.decorators import api_view , parser_classes
from rest_framework.response import Response
import pandas as pd
import warnings
from pathlib import Path
import stat
import shutil
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
import datetime
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
# from .models import MobinetDump
from rest_framework.parsers import MultiPartParser
from rest_framework import status
from pyxlsb import open_workbook


def write_df_in_chunks(df, writer, sheet_name, chunk_size=100000):
    """
    Write a large DataFrame to Excel in chunks to avoid memory/time issues.
    """
    row_start = 0
    for start in range(0, len(df), chunk_size):
        end = start + chunk_size
        chunk = df.iloc[start:end]

        # Only write header for the first chunk
        header = (start == 0)

        chunk.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=header,
            startrow=row_start
        )

        row_start += len(chunk) + (1 if header else 0)


def safe_write(df, writer, sheet_name, threshold=50000, chunk_size=100000):
    """
    Automatically decide whether to write normally or in chunks.
    """
    if len(df) > threshold:
        write_df_in_chunks(df, writer, sheet_name, chunk_size=chunk_size)
    else:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
 
 
 
 
# Excel  formate  function.........  
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
# def final_excel(df, writer, sheet_name='Sheet1', fast_mode=False):
#     df.to_excel(writer, index=False, sheet_name=sheet_name)
#     ws = writer.sheets[sheet_name]
 
#     if fast_mode:
#         # Only style the header for large files
#         header_fill = PatternFill(start_color='FFFF00', fill_type='solid')
#         bold = Font(bold=True)
#         align = Alignment(horizontal='center', vertical='center')
#         border = Border(*[Side(style='thin')] * 4)
 
#         for col_idx, col in enumerate(df.columns, start=1):
#             cell = ws.cell(row=1, column=col_idx)
#             cell.fill = header_fill
#             cell.font = bold
#             cell.alignment = align
#             cell.border = border
#         return  # Skip data cell styling
#     else:
#         # Full formatting (original logic)
#         header_fill = PatternFill(start_color='FFFF00', fill_type='solid')
#         bold = Font(bold=True)
#         align = Alignment(horizontal='center', vertical='center')
#         border = Border(*[Side(style='thin')] * 4)
 
#         for col_idx, col in enumerate(df.columns, start=1):
#             cell = ws.cell(row=1, column=col_idx)
#             cell.fill = header_fill
#             cell.font = bold
#             cell.alignment = align
#             cell.border = border
#             ws.row_dimensions[1].height = 25
 
#         for col_idx, col in enumerate(df.columns, 1):
#             max_len = len(col)
#             for row in range(2, ws.max_row + 1):
#                 cell = ws.cell(row=row, column=col_idx)
#                 cell.alignment = align
#                 cell.border = border
#                 if cell.value:
#                     max_len = max(max_len, len(str(cell.value)))
#             ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
 
       
# Delete the existing file function_______________
def delete_existing_files(folder_path):
    if os.path.exists(folder_path):
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print(f"Error deleting {file_path}: {e}")
               
# Read the both  excel and csv file function___________
def read_file(file):
    filename = file.name.lower()
    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(file)
        elif filename.endswith(('.xls', '.xlsx')):
            df = pd.read_excel(file, engine='openpyxl')
        else:
            raise ValueError("supported only csv and excel file")
        df.columns = df.columns.str.strip()
        return df
    except Exception as e:
        raise ValueError(f"Error reading file '{file.name}': {str(e)}")          
   
   
#--main folder in media folder--------------------
main_folder = os.path.join(MEDIA_ROOT, 'Mobinet_CATs_TOOL')
os.makedirs(main_folder, exist_ok=True)
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
 
# API Delete singal file from current folder ---------
@api_view(['DELETE'])
def delete_single_mobinet_file(request):
    try:
        mobinet_data_path = os.path.join(main_folder, request.data.get('foldername'))
        filename = request.data.get('filename')  # expecting JSON body: { "filename": "example.csv" }

        if not filename:
            return Response({'error': 'Filename not provided'}, status=status.HTTP_400_BAD_REQUEST)

        file_path = os.path.join(mobinet_data_path, filename)

        if not os.path.exists(file_path):
            return Response({'error': f'File "{filename}" not found'}, status=status.HTTP_404_NOT_FOUND)

        os.remove(file_path)

        return Response({
            'status': True,
            'message': f'File "{filename}" deleted successfully'
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
# API to upload and manage Mobinet dumps--------
@api_view(['POST', 'GET' , 'DELETE'])
def upload_mobinet_dumps(request):
    try:
        mobinet_data_path = os.path.join(main_folder, 'mobinet_dump_data')
        os.makedirs(mobinet_data_path, exist_ok=True)
 
        if request.method == 'POST':
            files = request.FILES.getlist('files')
            # files = request.FILES.getlist('mobinet_dumps')
            if not files:
                return Response({'error': 'No files uploaded'}, status=status.HTTP_400_BAD_REQUEST)
 
            for f in files:
                file_path = os.path.join(mobinet_data_path, f.name)
                with open(file_path, 'wb+') as destination:
                    for chunk in f.chunks():
                        destination.write(chunk)
 
            return Response({'status': True, 'message': 'Files uploaded and saved successfully'}, status=status.HTTP_200_OK)
 
        elif request.method == 'GET':
            if not os.path.exists(mobinet_data_path):
                return Response({'files': []}, status=status.HTTP_200_OK)
 
            files = os.listdir(mobinet_data_path)
            return Response({
                'status': True,
                'message': f'{len(files)} Files found in mobinet_dump_data folder',
                'files': files,
                }, status=status.HTTP_200_OK)
           
        elif request.method == 'DELETE':
            if not os.path.exists(mobinet_data_path):
                return Response({'error': 'Folder does not exist'}, status=status.HTTP_404_NOT_FOUND)
 
            deleted_files = []
            for filename in os.listdir(mobinet_data_path):
                file_path = os.path.join(mobinet_data_path, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    deleted_files.append(filename)
 
            return Response({
                'status': True,
                'message': f'{len(deleted_files)} Files deleted successfully',
                'deleted_files': deleted_files
            }, status=status.HTTP_200_OK)
           
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
    #Api to upload and manage RFS(CATS) ends here-----------


@api_view(['POST', 'GET', 'DELETE'])
def upload_rfs_data(request):
    try:
        rfs_data_path = os.path.join(main_folder, 'rfs_data')
        os.makedirs(rfs_data_path, exist_ok=True)
 
        if request.method == 'POST':
            files = request.FILES.getlist('files')
            if not files:
                return Response({'error': 'No files uploaded'}, status=status.HTTP_400_BAD_REQUEST)
 
            for f in files:
                file_path = os.path.join(rfs_data_path, f.name)
                with open(file_path, 'wb+') as destination:
                    for chunk in f.chunks():
                        destination.write(chunk)
 
            return Response({'status': True, 'message': 'Files saved successfully'}, status=status.HTTP_200_OK)
       
        elif request.method == 'GET':
            if not os.path.exists(rfs_data_path):
                return Response({'files': []}, status=status.HTTP_200_OK)
            files = os.listdir(rfs_data_path)
            return Response({
                'status': True,
                'message': 'Files found in rfs_data folder',
                'files': files,
            }, status=status.HTTP_200_OK)
           
        elif request.method == 'DELETE':
            if not os.path.exists(rfs_data_path):
                return Response({'error': 'Folder does not exist'}, status=status.HTTP_404_NOT_FOUND)
            deleted_files = []
            for filename in os.listdir(rfs_data_path):
                file_path = os.path.join(rfs_data_path, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    deleted_files.append(filename)
 
            return Response({
                'status': True,
                'message': 'Files deleted successfully',
                'deleted_files': deleted_files  }, status=status.HTTP_200_OK)
       
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
# api to upload and manage MSMF ends here-----------  
@api_view(['POST', 'GET', 'DELETE'])  
def upload_msmf_data(request):
    try:
        msmf_folder_path = os.path.join(main_folder, 'msmf_data')
        os.makedirs(msmf_folder_path, exist_ok=True)
        if request.method == 'POST':
            files = request.FILES.getlist('files')
            if not files:
                return Response({'error': 'No files uploaded'}, status=status.HTTP_400_BAD_REQUEST)
 
            for f in files:
                file_path = os.path.join(msmf_folder_path, f.name)
                with open(file_path, 'wb+') as destination:
                    for chunk in f.chunks():
                        destination.write(chunk)
 
            return Response({'status': True, 'message': 'Files saved successfully'}, status=status.HTTP_200_OK)
        elif request.method == 'GET':
            if not os.path.exists(msmf_folder_path):
                return Response({'files': []}, status=status.HTTP_200_OK)
            files = os.listdir(msmf_folder_path)
            return Response({
                'status': True,
                'message': 'Files found in msmf_data folder',
                'files': files,
            }, status=status.HTTP_200_OK)
        elif request.method == 'DELETE':
            if not os.path.exists(msmf_folder_path):
                return Response({'error': 'Folder does not exist'}, status=status.HTTP_404_NOT_FOUND)
            deleted_files = []
            for filename in os.listdir(msmf_folder_path):
                file_path = os.path.join(msmf_folder_path, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    deleted_files.append(filename)
 
            return Response({
                'status': True,
                'message': 'Files deleted successfully',
                'deleted_files': deleted_files
            }, status=status.HTTP_200_OK)
           
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)  
   
    #api for locator file  data ends here-----------


@api_view(['POST', 'GET', 'DELETE'])
def upload_locator_data(request):
    try:
        locator_folder_path = os.path.join(main_folder, 'locator_data')
        os.makedirs(locator_folder_path, exist_ok=True)
 
        if request.method == 'POST':
            files = request.FILES.getlist('files')
            if not files:
                return Response({'error': 'No files uploaded'}, status=status.HTTP_400_BAD_REQUEST)
 
            for f in files:
                file_path = os.path.join(locator_folder_path, f.name)
                with open(file_path, 'wb+') as destination:
                    for chunk in f.chunks():
                        destination.write(chunk)
 
            return Response({'status': True, 'message': 'Files saved successfully'}, status=status.HTTP_200_OK)
 
        elif request.method == 'GET':
            if not os.path.exists(locator_folder_path):
                return Response({'files': []}, status=status.HTTP_200_OK)
            files = os.listdir(locator_folder_path)
            return Response({
                'status': True,
                'message': 'Files found in locator_data folder',
                'files': files,
            }, status=status.HTTP_200_OK)
 
        elif request.method == 'DELETE':
            if not os.path.exists(locator_folder_path):
                return Response({'error': 'Folder does not exist'}, status=status.HTTP_404_NOT_FOUND)
            deleted_files = []
            for filename in os.listdir(locator_folder_path):
                file_path = os.path.join(locator_folder_path, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    deleted_files.append(filename)
 
            return Response({
                'status': True,
                'message': 'Files deleted successfully',
                'deleted_files': deleted_files
            }, status=status.HTTP_200_OK)
 
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
#api to upload and manage stock report data ends here-----------
@api_view(['POST', 'GET', 'DELETE'])
def upload_stock_report_data(request):
    stock_report_folder_path = os.path.join(main_folder, 'stock_report_data')
    os.makedirs(stock_report_folder_path, exist_ok=True)
    try:
        if request.method == 'POST':
            files = request.FILES.getlist('files')
            if not files:
                return Response({'error': 'No files uploaded'}, status=status.HTTP_400_BAD_REQUEST)
 
            for f in files:
                file_path = os.path.join(stock_report_folder_path, f.name)
                with open(file_path, 'wb+') as destination:
                    for chunk in f.chunks():
                        destination.write(chunk)
 
            return Response({'status': True, 'message': 'Files saved successfully'}, status=status.HTTP_200_OK)
        elif request.method == 'GET':
            if not os.path.exists(stock_report_folder_path):
                return Response({'files': []}, status=status.HTTP_200_OK)
            files = os.listdir(stock_report_folder_path)
            return Response({
                'status': True,
                'message': 'Files found in stock_report_data folder',
                'files': files,
            }, status=status.HTTP_200_OK)    
           
        elif request.method == 'DELETE':
            if not os.path.exists(stock_report_folder_path):
                return Response({'error': 'Folder does not exist'}, status=status.HTTP_404_NOT_FOUND)
            deleted_files = []
            for filename in os.listdir(stock_report_folder_path):
                file_path = os.path.join(stock_report_folder_path, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    deleted_files.append(filename)
 
            return Response({
                'status': True,
                'message': 'Files deleted successfully',
                'deleted_files': deleted_files
            }, status=status.HTTP_200_OK)
           
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)        
         
 
 
 
       
# Frist Api for mobinet_dump site_match and hw_files-------------
@api_view(['POST'])
def mobinet_dump(request):
    print("Start Process_________")
    try:
        # Read files from mobinet_dump_data folder-------------
        mobinet_log_folder = os.path.join(main_folder, 'mobinet_dump_data')
 
        if not os.path.exists(mobinet_log_folder):
            return Response({"error": "mobinet_dump_data folder not found"}, status=400)
 
        log_df_list = []
        for file in os.listdir(mobinet_log_folder):
            file_path = os.path.join(mobinet_log_folder, file)
 
            if not os.path.isfile(file_path):
                continue  # skip subfolders
 
            try:
                if file_path.endswith('.csv'):
                   df = pd.read_csv(file_path)
                elif file_path.endswith(('.xls', '.xlsx')):
                    df = pd.read_excel(file_path, engine='openpyxl')
                else:
                    return Response({"error": f"Unsupported file format: {file}"}, status=400)
                log_df_list.append(df)
            except Exception as e:
                return Response({"error": f"Error in file {file}: {str(e)}"}, status=500)
 
        if not log_df_list:
            return Response({"error": "No valid log files found"}, status=400)
 
        log_df = pd.concat(log_df_list, ignore_index=True)
        print("Concatenated log_df columns:", log_df.columns)
       
       
       
 
        # Read site list file (CSV or Excel)
        site_list_file = request.FILES.get("site_list")
        if not site_list_file:
            return Response({"error": "site_list file not provided"}, status=400)
 
        try:
            site_df = read_file(site_list_file)
            site_df.columns = site_df.columns.str.strip()
            print(site_df)
        except ValueError as e:
            return Response({"error": str(e)}, status=500)
 
        # Read multiple hw_files
        # hw_files = request.FILES.getlist("hw_file")
        # if not hw_files:
        #     return Response({"error": "hw_files not provided"}, status=400)
 
        # hw_df_list = []
        # for hw_file in hw_files:
        #     try:
        #         df = read_file(hw_file)
        #         df.columns = df.columns.str.strip()
        #         hw_df_list.append(df)
        #     except Exception as e:
        #         return Response({"error": f"Error reading one of the hw_files: {str(e)}"}, status=500)
 
        # hw_df = pd.concat(hw_df_list, ignore_index=True)
 
        # Strip and match values
        log_df['Parent Site'] = log_df['Parent Site'].astype(str).str.strip()
        site_df['Unique ID'] = site_df['Unique ID'].astype(str).str.strip()
 
        # Merge logs with site list
        matched_df = pd.merge(
            log_df,
            site_df[['Unique ID', 'Dismantled date']],
            how='inner',
            left_on='Parent Site',
            right_on='Unique ID'
        )
        print("Matched DF:")
        print(matched_df)
 
        # Unmatched sites
        unmatched_df = site_df.loc[
            ~site_df['Unique ID'].isin(log_df['Parent Site']), ['Unique ID']
        ].copy()
        unmatched_df.rename(columns={'Unique ID': 'Unmatched Unique ID'}, inplace=True)
        print("Unmatched Unique IDs:")
        print(unmatched_df)
 
        # Merge with HW file
        #   matched_df = pd.merge(
        #     matched_df,
        #     hw_df[['Module Name', 'ITEMCODE']],
        #     how='outer',
        #     left_on='Board Model',
        #     right_on='Module Name'
        # )
        #   matched_df.drop(columns='Module Name', inplace=True)
 
        # Remove blank and '-' data in serial number
        matched_df = matched_df[
              matched_df['Serial Number'].notna() &
            (  matched_df['Serial Number'].astype(str).str.strip() != '') &
            (  matched_df['Serial Number'].astype(str).str.strip() != '-')
        ]
 
        matched_df['Site+Module'] =   matched_df['Parent Site'].astype(str).str.strip() + "_" +   matched_df['Board Model'].astype(str).str.strip()
        matched_df['Site+Serial'] =   matched_df['Parent Site'].astype(str).str.strip() + "_" +   matched_df['Serial Number'].astype(str).str.strip()
        matched_df.drop_duplicates(subset='Serial Number', inplace=True)
 
        # Count serial numbers per Site+Module
        data_count =   matched_df.copy()
        summary_df = (
            data_count.groupby("Site+Module", as_index=False)
            .agg({"Serial Number": "count"})
            .rename(columns={"Serial Number": "Serial Number_count"})
        )
        print("Summary DF:")
        print(summary_df)
 
        # Export Excel
        output_dir = os.path.join(main_folder, 'Final_Mobinet_Summary_output')
        os.makedirs(output_dir, exist_ok=True)
        delete_existing_files(output_dir)
 
        filename = f"mobinet_summary_data_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, filename)
 
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            safe_write(summary_df, writer, sheet_name='Mobinet_Summary')
            safe_write(data_count, writer, sheet_name='Mobinet_Backup_Data')
            safe_write(unmatched_df, writer, sheet_name='Unmatched_Site_IDs')
 
        # Generate download URL
        relative_path = os.path.relpath(output_path, MEDIA_ROOT)
        download_url = request.build_absolute_uri(os.path.join(MEDIA_URL, relative_path).replace("\\", "/"))
        print("End Process_________")
 
        return Response({
            "status": True,
            "message": "File saved successfully",
            "download_url": download_url
        })
 
    except Exception as e:
        return Response({"error": f"find an error: {str(e)}"}, status=500)

circle_map_code = {
        151: 'AP', 132: 'BR',152: 'CN',111: 'DL',113: 'HR', 114: 'JK',
        153: 'KK', 174: 'MU', 172: 'MP',  134: 'OR', 115: 'PB', 176: 'RJ',
        116: 'UE', 117: 'UW', 135: 'WB', 136: 'KO', 155: 'TN',133: 'NE',131: 'AS',
        173:"MH"
    }        
  
# Second Api for CATs(rfs) and Locator files and hw files-------------
@api_view(['POST'])
def rfs_dump(request):
    print("Start Process_________")
    try:
        # RFS File
        rfs_folder = os.path.join(main_folder, 'rfs_data')
        if not os.path.exists(rfs_folder):
            return Response({"error": "rfs_data folder not found"}, status=400)
       
        for file in os.listdir(rfs_folder):
            rfs_file_path = os.path.join(rfs_folder, file)
            if not os.path.isfile(rfs_file_path):
                continue
            try:
                if rfs_file_path.endswith('.csv'):
                    rfs_df = pd.read_csv(rfs_file_path)
                elif rfs_file_path.endswith(('.xls', '.xlsx')):
                    rfs_df = pd.read_excel(rfs_file_path, engine='openpyxl')    
                elif rfs_file_path.endswith('.xlsb'):
                    with open_workbook(rfs_file_path) as wb:
                        with wb.get_sheet(wb.sheets[0]) as sheet:
                            data = []
                            for row in sheet.rows():
                                data.append([item.v for item in row])
                            rfs_df = pd.DataFrame(data[1:], columns=data[0])  
                else:
                    return Response({"error": f"Unsupported file format: {file}"}, status=400)  
                rfs_df.columns = rfs_df.columns.str.strip()
                # print("RFS Data columns:", rfs_df.columns)
            except Exception as e:
                return Response({"error": f"error reading rfs file: {str(e)}"}, status=500)
           
    
    
 
       # Multiple HW Files --------------------------------------------------
        hw_files = request.FILES.getlist("hw_file")

        if not hw_files:
            return Response({"error": "hw_file(s) not provided"}, status=400)

        hw_df_list = []
        no_need_list = []

        for hw_file in hw_files:
            try:
                if hw_file.name.endswith('.csv'):

                    df = pd.read_csv(hw_file)
                    df.columns = df.columns.str.strip()

                    hw_df_list.append(df[['ITEMCODE', 'Module Name']])

                elif hw_file.name.endswith(('.xls', '.xlsx')):

                    # main sheet
                    df = pd.read_excel(hw_file, engine='openpyxl')
                    df.columns = df.columns.str.strip()

                    hw_df_list.append(df[['ITEMCODE', 'Module Name']])

                    # optional No module list sheet
                    try:
                        module_df = pd.read_excel( hw_file,engine='openpyxl',sheet_name="No module list")
                        module_df.columns = module_df.columns.str.strip()

                        if 'No Module' in module_df.columns:
                            no_need_list.append(module_df[['No Module']])
                    except:
                        pass
                else:
                    return Response(
                        {"error": "Unsupported file format for hw_file"},
                        status=400
                    )

            except Exception as e:
                return Response(
                    {"error": f"error reading a hw_file: {str(e)}"},
                    status=500
                )

        hw_df = pd.concat(hw_df_list, ignore_index=True).drop_duplicates()

        no_need_module = (
            pd.concat(no_need_list, ignore_index=True).drop_duplicates()
            if no_need_list
            else pd.DataFrame(columns=['No Module'])
        )
 

        #  Site List File-----------------------------------------------------------
        site_list_file = request.FILES.get("site_list_file")
        if not site_list_file:
            return Response({"error": "site_list_file not provided"}, status=400)
        try:
            if site_list_file.name.endswith('.csv'):
                site_list_df = pd.read_csv(site_list_file)
            elif site_list_file.name.endswith(('.xls', '.xlsx')):
              site_list_df = pd.read_excel(site_list_file, engine='openpyxl', usecols=['Unique ID', 'Dismantled date'])
            site_list_df.columns = site_list_df.columns.str.strip()
            # print(site_list_df.columns)
        except Exception as e:
            return Response({"error": f"error reading site_list_file: {str(e)}"}, status=500)
 
        # Multiple Locator Files -------------------------------------------------
        locator_folder = os.path.join(main_folder, 'locator_data')
        if not os.path.exists(locator_folder):
            return Response({"error": "locator_data folder not found"}, status=400)
       
        locator_df_list = []
        for file in os.listdir(locator_folder):
            locator_file_path = os.path.join(locator_folder, file)
            if not os.path.isfile(locator_file_path):
                continue
            try:
                if locator_file_path.endswith('.csv'):
                    df = pd.read_csv(locator_file_path, usecols=['Location Name', 'Siteid'])    
                elif locator_file_path.endswith(('.xls', '.xlsx')):
                    df = pd.read_excel(locator_file_path, engine='openpyxl', usecols=['Location Name', 'Siteid'])  
                elif locator_file_path.endswith('.xlsb'):
                    with open_workbook(locator_file_path) as wb:
                        with wb.get_sheet(wb.sheets[0]) as sheet:
                            data = []
                            for row in sheet.rows():
                                data.append([item.v for item in row])
                            df = pd.DataFrame(data[1:], columns=data[0])
                else:
                    return Response({"error": f"Unsupported file format: {file}"}, status=400)
                df.columns = df.columns.str.strip()
                locator_df_list.append(df)
            except Exception as e:
                return Response({"error": f"error reading locator file: {str(e)}"}, status=500)
        locator_df = pd.concat(locator_df_list, ignore_index=True).drop_duplicates()
        # print(locator_df.columns)
           
     
    #    # Unmatched Sites...
    #     unmatched_sites = request.POST.getlist('unmatched_sites')
    #     if not unmatched_sites:
    #         return Response({"error": "unmatched_sites not provided"}, status=400)
    #     if len(unmatched_sites) == 1 and "\n" in unmatched_sites[0]:
    #      unmatched_sites = unmatched_sites[0].strip().split("\n")
    #     print("Aarry of unmatched_sites:",unmatched_sites)
    #     unmatched_sites_df = pd.DataFrame(unmatched_sites, columns=["Unmatched Sites"])
    #     print(unmatched_sites_df.head())
       
        # Mobinet Dump File--------------------------------------------
        mobinet_dump_file = request.FILES.get("mobinet_dump_file")
        if not mobinet_dump_file:
            return Response({"error": "mobinet_dump_file not provided"}, status=400)
        try:
            mobinet_output_df = pd.read_excel(mobinet_dump_file, engine='openpyxl', sheet_name='Mobinet_Summary')
            mobinet_output_df.columns = mobinet_output_df.columns.str.strip()
            mobinet_dump_df = pd.read_excel(mobinet_dump_file, engine='openpyxl',sheet_name='Mobinet_Backup_Data')
            mobinet_dump_df.columns = mobinet_dump_df.columns.str.strip()
            # print(mobinet_dump_df.columns)
        except Exception as e:
            return Response({"error": f"error reading mobinet_dump_file: {str(e)}"}, status=500)
       
       
        #MS-MF file-----------------------------------------------
        msmf_folder = os.path.join(main_folder, 'msmf_data')
        if not os.path.exists(msmf_folder):
            return Response({"error": "msmf_data folder not found"}, status=400)
       
        for file in os.listdir(msmf_folder):
            msmf_file_path = os.path.join(msmf_folder, file)
            if not os.path.isfile(msmf_file_path):
                continue
            try:
                if msmf_file_path.endswith('.csv'):
                    msmf_df = pd.read_csv(msmf_file_path,engine='python', on_bad_lines='skip')
                elif msmf_file_path.endswith(('.xls', '.xlsx')):
                    msmf_df = pd.read_excel(msmf_file_path, engine='openpyxl')
                elif msmf_file_path.endswith('.xlsb'):
                    with open_workbook(msmf_file_path) as wb:
                        with wb.get_sheet(wb.sheets[0]) as sheet:
                            data = []
                            for row in sheet.rows():
                                data.append([item.v for item in row])
                            msmf_df = pd.DataFrame(data[1:], columns=data[0])
                else:
                    return Response({"error": f"Unsupported file format: {file}"}, status=400)
                   
                msmf_df.columns = msmf_df.columns.str.strip()
                # print("MS-MF Data columns:", msmf_df.columns)
            except Exception as e:
                return Response({"error": f"error reading ms-mf file: {str(e)}"}, status=500)
           
       
       
           
       
        # stock-report_file -----------------------------------------------
        stock_report_folder=os.path.join(main_folder, 'stock_report_data')
        stock_report_df = None
        if not os.path.exists(stock_report_folder):
            return Response({"error": "stock_report_data folder not found"}, status=400)
       
        for file in os.listdir(stock_report_folder):
            stock_report_file_path = os.path.join(stock_report_folder, file)
            if not os.path.isfile(stock_report_file_path):
                continue
            try:
                if stock_report_file_path.endswith('.csv'):
                    stock_report_df = pd.read_csv(stock_report_file_path)
                elif stock_report_file_path.endswith(('.xls', '.xlsx')):
                    stock_report_df = pd.read_excel(stock_report_file_path, engine='openpyxl')
                elif stock_report_file_path.endswith('.xlsb'):
                    with open_workbook(stock_report_file_path) as wb:
                        with wb.get_sheet(wb.sheets[0]) as sheet:  # First sheet
                            data = []
                            for row in sheet.rows():
                                data.append([item.v for item in row])
                            stock_report_df = pd.DataFrame(data[1:], columns=data[0])
                stock_report_df.columns = stock_report_df.columns.str.strip()
                # print("Stock Report Data columns:", stock_report_df.columns)
            except Exception as e:
                return Response({"error": f"error reading stock report file: {str(e)}"}, status=500)
       
 
   
        print("#step1------")
########for RFS file#######################################
        
      
        rfs_df["Partners"] = "Other TSP"

        mobile_mask = rfs_df["Partner Name"].str.contains(
            "Mobilecom|Mobilecomm",
            case=False,
            na=False
        )
        # Step 2: if Partner Name blank then se check Creatordepartment ---
        creator_mask = (
            rfs_df["Partner Name"].isna() |
            (rfs_df["Partner Name"].str.strip() == "")
        ) & rfs_df["Creatordepartment"].str.contains(
            "Mobilecom|Mobilecomm",
            case=False,
            na=False
        )

        rfs_df.loc[mobile_mask | creator_mask, "Partners"] = "Mobliecomm"
        rfs_df.drop(["Creatordepartment", "Partner Name"], axis=1, inplace=True)
 
        # Merge with locator----
        merged_df = pd.merge(
            rfs_df,
            locator_df[['Location Name', 'Siteid']],
            how='inner',
            left_on='Fromlocation',
            right_on='Location Name',
            sort=False
        )
        # print('marge with locator',merged_df.head())
       
        # Circle Mapping ----
        merged_df['Circle_Map'] = merged_df['Fromlocation'].astype(str).str[:3].astype(int)
        merged_df['Circle'] = merged_df['Circle_Map'].map(circle_map_code)
        merged_df['Unique Site ID'] = merged_df['Siteid'].astype(str) + "_" + merged_df['Circle'].astype(str)
        merged_df.drop(columns=['Location Name', 'Circle_Map'], inplace=True)
        print("#step2------")
     
 
        # Merge with HW File -----
        # merged_hw = pd.merge(
        #     merged_df,
        #     hw_df,
        #     how='left',
        #     left_on="Itemcode",
        #     right_on='ITEMCODE',
        #     sort=False
        # )
        # print('merge with Hardware',merged_hw.head())
        #add filter filter and remove no neeed module----
        print(no_need_module['No Module'])

        merged_hw = pd.merge(
            merged_df,
            hw_df,
            how='left',
            left_on="Itemcode",
            right_on='ITEMCODE'
        )

        # remove no need modules
        if not no_need_module.empty:
            print("remove module---")
            no_modules = (
                no_need_module['No Module']
                .dropna()
                .astype(str)
                .str.strip()
                .unique()
            )

            merged_hw = merged_hw[
                ~merged_hw['Module Name']
                .astype(str)
                .str.strip()
                .isin(no_modules)
            ]
       


 
        #  Merge with site list
        merged_site_df = pd.merge(
            merged_hw,
            site_list_df,
            how='inner',
            left_on='Unique Site ID',
            right_on='Unique ID',
            sort=False
        )
        merged_site_df.drop(columns=['Unique ID'], inplace=True)
       
        # print('merge with site list file',merged_site_df.head())
        # print(site_list_df['Unique ID'])
     
       
        #  #data for unmatched site.
        # unmatched_data= merged_site_df[merged_site_df['Unique Site ID'].isin(unmatched_sites_df["Unmatched Sites"])]
        # unmatched_data["Remark"]="Unmatched Site"
        # print(unmatched_data.head())
       
 
 
        merged_site_df["Dismantled date"] = pd.to_datetime(merged_site_df["Dismantled date"], errors='coerce')
        merged_site_df["Rfs Created Date"] = pd.to_datetime(merged_site_df["Rfs Created Date"], errors='coerce')
 
        # merged_site_df["Aging"] = (merged_site_df["Dismantled date"] - merged_site_df["Rfs Created Date"]).dt.days
        # merged_site_df = merged_site_df[merged_site_df["Aging"]<=-45]
 
       
        merged_site_df["Site+Module"] = merged_site_df["Unique Site ID"].astype(str) + "_" + merged_site_df["Module Name"].astype(str)
        merged_site_df.drop_duplicates(inplace=True)
        
        # merged_site_df.to_excel("mereged_site_data.xlsx")
        print("#step3------")
       
 
       
       
        #work on serial number----------------------------------------------------------
        rfs_matched_sn =merged_site_df.copy()
        rfs_matched_sn["Site+Serial"]= rfs_matched_sn["Unique Site ID"].astype(str) + "_" + rfs_matched_sn["Serialnumber"].astype(str)
       
        #remove S from serial number
        rfs_matched_sn_2= rfs_matched_sn.copy()
        rfs_matched_sn_2["Serialnumber"]=rfs_matched_sn_2["Serialnumber"].str.replace(r'^[Ss]', '', regex=True)
        rfs_matched_sn_2["Site+Serial"]= rfs_matched_sn_2["Unique Site ID"].astype(str) + "_" + rfs_matched_sn_2["Serialnumber"].astype(str)
       
        #merge with serial number from mobinet dump file
        rfs_matched_sn=pd.merge(
            rfs_matched_sn,
            mobinet_dump_df[['Site+Serial']],
            how='inner',
            on='Site+Serial',
        )
        rfs_matched_sn["Remark"] = "Matched by Serial Number"
       
        #merge without S serial number from mobinet dump file
        rfs_matched_sn_2=pd.merge(
            rfs_matched_sn_2,
            mobinet_dump_df[['Site+Serial']],
            how='inner',
            on='Site+Serial',
        )
        rfs_matched_sn_2["Remark"] = "Matched by Serial Number"
       
        rfs_matched_sn_merege=pd.concat([rfs_matched_sn, rfs_matched_sn_2], ignore_index=True)
        rfs_matched_sn_merege.drop_duplicates(inplace=True)
        
       
         
 # Merge with stock report file if provided-----------------------------------
        if stock_report_df is not None and not stock_report_df.empty:
            rfs_finally =merged_site_df.copy()
            rfs_with_stock_report = pd.merge(
            rfs_finally,
            stock_report_df[['LOT_MRR_NUMBER', 'CURRENTLOCATION', 'CURRENTSTATUS', 'LOCATION_TYPE', 'SUBINVENTORY', 'ORACLE_STOCK_LOCATOR']],
            left_on='ORACLE_LOT_NO',
            right_on='LOT_MRR_NUMBER',
            how='left',
            )  
            rfs_with_stock_report.drop(columns=['LOT_MRR_NUMBER'], inplace=True)
            rfs_with_stock_report.drop_duplicates(inplace=True)
            print("rfs_with_stock_report", rfs_with_stock_report)
            
        else:
            rfs_with_stock_report= merged_site_df.copy()   
        
        print("#phase5-------------")
        # print(rfs_with_stock_report)
        # rfs_with_stock_report.to_excel("mereged_site_data.xlsx")
       


    
       
       
# summary_df for RFS file------
        summary_df = (
            merged_site_df.groupby("Site+Module", as_index=False)
            .agg({
                "Received Qty": "count",
                "Rfs Qty": "count",
                "Srncamreqstatus": lambda x: ", ".join(set(x.dropna().astype(str)))
            })
            .rename(columns={
                "Received Qty": "RECEIVED_QTY_COUNT(RFS_FILE)",
                "Rfs Qty": "RFS_QTY_COUNT(RFS_FILE)",
                "Srncamreqstatus":"SRNCAMREQSTATUS"
            })
        )
       
        # print(summary_df)
    #######################################################################################



   
   
    #######for MS-MF file####################
        msmf_df["Partners"] = "Other TSP"

        msmf_df.loc[
            msmf_df["Partner"].astype(str).str.contains(
                "Mobilecom|Mobilecomm",
                case=False,
                na=False
            ),
            "Partners"
        ] = "Mobliecomm"

        msmf_df.drop(["Partner"], axis=1,
            inplace=True,
            errors="ignore"
        )

        # Merge with locator----
        merged_df_ms_mf = pd.merge(
            msmf_df,
            locator_df[['Location Name', 'Siteid']],
            how='inner',
            left_on='Ms Fromlocation',
            right_on='Location Name',
            sort=False
        )
        # print('marge with locator',merged_df_ms_mf.head())
       
        # Circle Mapping ----
        merged_df_ms_mf['Circle_Map'] = merged_df_ms_mf['Ms Fromlocation'].astype(str).str[:3].astype(int)
        merged_df_ms_mf['Circle'] = merged_df_ms_mf['Circle_Map'].map(circle_map_code)
        merged_df_ms_mf['Unique Site ID'] = merged_df_ms_mf['Siteid'].astype(str) + "_" + merged_df_ms_mf['Circle'].astype(str)
        merged_df_ms_mf.drop(columns=['Location Name', 'Circle_Map'], inplace=True)
       
        print("#step6------")
 
        # Merge with HW File -----
        merged_hw_ms_mf = pd.merge(
            merged_df_ms_mf,
            hw_df,
            how='left',
            left_on='Partcode',
            right_on='ITEMCODE',
            sort=False
        )
        merged_hw_ms_mf.drop(columns=['ITEMCODE'], inplace=True)

        print(no_need_module['No Module'])

        # remove no need modules
        if not no_need_module.empty:
            print("remove module---")
            no_modules = (
                no_need_module['No Module']
                .dropna()
                .astype(str)
                .str.strip()
                .unique()
            )


            merged_hw_ms_mf =  merged_hw_ms_mf[
                ~ merged_hw_ms_mf['Module Name']
                .astype(str)
                .str.strip()
                .isin(no_modules)
            ]
    
        # print('merge with Hardware',merged_hw_ms_mf.head())
 
        #  Merge with site list
        merged_site_df_ms_mf = pd.merge(
            merged_hw_ms_mf,
            site_list_df,
            how='inner',
            left_on='Unique Site ID',
            right_on='Unique ID',
            sort=False
        )
        merged_site_df_ms_mf.drop(columns=['Unique ID'], inplace=True)
        # print('merge with site list file',merged_site_df_ms_mf.head())
       
        # #data for unmatched site.
        # unmatched_data_ms_mf= merged_site_df_ms_mf[merged_site_df_ms_mf['Unique Site ID'].isin(unmatched_sites_df["Unmatched Sites"])]
        # unmatched_data_ms_mf["Remark"]="Unmatched Site"
        # print(unmatched_data_ms_mf.head())
       
     
       
        merged_site_df_ms_mf["Dismantled date"] = pd.to_datetime(merged_site_df_ms_mf["Dismantled date"], errors='coerce')
        merged_site_df_ms_mf["Move Start Date"] = pd.to_datetime(merged_site_df_ms_mf["Move Start Date"], errors='coerce')
       
        # merged_site_df_ms_mf["Aging"] = (merged_site_df_ms_mf["Dismantled date"] - merged_site_df_ms_mf["Move Start Date"]).dt.days
        # merged_site_df_ms_mf =merged_site_df_ms_mf[merged_site_df_ms_mf["Aging"] <= -45]
 
       
        merged_site_df_ms_mf["Site+Module"] = merged_site_df_ms_mf["Unique Site ID"].astype(str) + "_" + merged_site_df_ms_mf["Module Name"].astype(str)
        merged_site_df_ms_mf.drop_duplicates(inplace=True)
       
        #work with serial number----------------------------------------
        ms_mf_matched_sn =merged_site_df_ms_mf.copy()
        ms_mf_matched_sn["Site+Serial"]= ms_mf_matched_sn["Unique Site ID"].astype(str) + "_" + ms_mf_matched_sn["Serial Number"].astype(str)
         #remove S from serial number
        ms_mf_matched_sn_2= ms_mf_matched_sn.copy()
        ms_mf_matched_sn_2["Serial Number"]=ms_mf_matched_sn_2["Serial Number"].str.replace(r'^[Ss]', '', regex=True)
        ms_mf_matched_sn_2["Site+Serial"]= ms_mf_matched_sn_2["Unique Site ID"].astype(str) + "_" + ms_mf_matched_sn_2["Serial Number"].astype(str)
       
         
        #merge with serial number from mobinet dump file
        ms_mf_matched_sn=pd.merge(
            ms_mf_matched_sn,
            mobinet_dump_df[['Site+Serial']],
            how='inner',
            on='Site+Serial',
        )
       
        #merge without S serial number from mobinet dump file
        ms_mf_matched_sn_2=pd.merge(
            ms_mf_matched_sn_2,
            mobinet_dump_df[['Site+Serial']],
            how='inner',
            on='Site+Serial',
        )
       
        ms_mf_matched_sn_2["Remark"]="Matched by Serial Number"
        ms_mf_matched_sn["Remark"]="Matched by Serial Number"
       
        ms_mf_matched_merged=pd.concat([ms_mf_matched_sn, ms_mf_matched_sn_2], ignore_index=True)
        ms_mf_matched_merged.drop_duplicates(inplace=True)
        # print(ms_mf_matched_merged)
        print("#step7------")
       
       
   
       
       # summary of MS-MF file------
        ms_mf_data_cout =merged_site_df_ms_mf.copy()
       
        ms_mf_summary_df = (
            merged_site_df_ms_mf.groupby("Site+Module", as_index=False)
            .agg({
                "Received Qty": "count",
                "Qty": "count"
            })
            .rename(columns={
                "Received Qty": "RECEIVED_QTY_COUNT(MS-MF FILE)",
                "Qty": "RFS_QTY_COUNT(MS-MF FILE)"
            })
        )
       
        # print("ms_mf_summary_df",ms_mf_summary_df)
     
    #merge summary_df and ms_mf_summary_df
        final_summary_df=pd.merge(
            summary_df,
            ms_mf_summary_df,
            how='outer',
            on='Site+Module',)
        # print(final_summary_df)
       
        ## final Api work here-------------------------
        #site wise summary and circle wise summary data site wise  data count
        merged_find_summary = pd.merge(
            left=mobinet_output_df,
            right=final_summary_df[['Site+Module', 'RECEIVED_QTY_COUNT(RFS_FILE)' ,'RFS_QTY_COUNT(RFS_FILE)','RECEIVED_QTY_COUNT(MS-MF FILE)','RFS_QTY_COUNT(MS-MF FILE)','SRNCAMREQSTATUS']].copy(),  # include ITEMCODE here
            how='outer',
            on='Site+Module',
           
        )
        merged_find_summary['Unique ID'] = merged_find_summary['Site+Module'].str.split('_', expand=True)[0] + "_" + merged_find_summary['Site+Module'].str.split('_', expand=True)[1]
     #site wise summay----------------------------------------------
        site_wise_summary = (
        merged_find_summary
        .groupby('Unique ID', as_index=False)
        .agg({
            'Serial Number_count': 'sum',
            
            'RECEIVED_QTY_COUNT(RFS_FILE)': 'sum',
            'RECEIVED_QTY_COUNT(MS-MF FILE)': 'sum',
            'RFS_QTY_COUNT(RFS_FILE)': 'sum',
            'RFS_QTY_COUNT(MS-MF FILE)': 'sum',
          'SRNCAMREQSTATUS': 'first'
 
        })
        .assign(
            Total_In_CATS =lambda x:  x['RFS_QTY_COUNT(RFS_FILE)'] + x['RFS_QTY_COUNT(MS-MF FILE)'],
            Total_Receipt_CATS=lambda x: x['RECEIVED_QTY_COUNT(RFS_FILE)'] + x['RECEIVED_QTY_COUNT(MS-MF FILE)'],
            Dismantle_vs_SREQ=lambda x: x['Total_In_CATS'] - x['Serial Number_count'],
            SREQ_vs_WH_Receipt=lambda x:x['Total_Receipt_CATS'] - x['Total_In_CATS'],
           
            )
       
        .rename(columns={
            'Unique ID': 'Site ID',
            'Serial Number_count': 'Mobinet_Count(Serial Number)',
        })
    )
        # print(site_wise_summary)
       
        #Circle wise summary----------------------------------------------
        merged_find_summary['Circle'] = merged_find_summary['Unique ID'].str.split('_', expand=True)[1]
       
        circle_wise_summary = (
            merged_find_summary
                .groupby('Circle', as_index=False)
                .agg({
                    'Serial Number_count': 'sum',
                    'RECEIVED_QTY_COUNT(RFS_FILE)': 'sum',
                    'RECEIVED_QTY_COUNT(MS-MF FILE)': 'sum',
                    'RFS_QTY_COUNT(RFS_FILE)': 'sum',
                    'RFS_QTY_COUNT(MS-MF FILE)': 'sum',
                    'SRNCAMREQSTATUS': 'first'
 
                })
               
            .assign(
                    Total_In_CATS =lambda x:  x['RFS_QTY_COUNT(RFS_FILE)'] + x['RFS_QTY_COUNT(MS-MF FILE)'],
                    Total_Receipt_CATS=lambda x: x['RECEIVED_QTY_COUNT(RFS_FILE)'] + x['RECEIVED_QTY_COUNT(MS-MF FILE)'],
                    Dismantle_vs_SREQ=lambda x: x['Total_In_CATS'] - x['Serial Number_count'],
                    SREQ_vs_WH_Receipt=lambda x:x['Total_Receipt_CATS'] - x['Total_In_CATS'],
                   
                    )
                .rename(columns={
                    'Serial Number_count': 'Mobinet_Count(Serial Number)',
               
                })
            )
        # print(circle_wise_summary)
       
        #count of matched sites in mobinet and rfs file ms-mf------------------------
       
        site_count_mobinet = mobinet_dump_df['Parent Site'].value_counts()
        site_count_rfs =merged_site_df['Unique Site ID'].value_counts()
        site_count_msmf = ms_mf_data_cout['Unique Site ID'].value_counts()
 
        new_df = site_list_df[['Unique ID']].copy()
        new_df.rename(columns={'Unique ID': 'Site ID'}, inplace=True)
        new_df['SiteID in Mobinet'] = site_list_df['Unique ID'].map(site_count_mobinet).fillna(0).astype(int)
        new_df['SiteID in RFS'] = site_list_df['Unique ID'].map(site_count_rfs).fillna(0).astype(int)
        new_df['SiteID in MS-MF'] = site_list_df['Unique ID'].map(site_count_msmf).fillna(0).astype(int)
 
        # print("new_df", new_df)
        print("End all files Process. Saving the data in excel file")
 
       
        output_dir = os.path.join(main_folder,'Finall_CATs_MSMF_Summary_Output')
        os.makedirs(output_dir, exist_ok=True)
        delete_existing_files(output_dir)
        filename = f"CATs_Summary_data{timestamp}.xlsx"
        output_path = os.path.join(output_dir, filename)
        with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
            safe_write(site_wise_summary, writer, "Site Wise Summary")
            safe_write(circle_wise_summary, writer, "Circle Wise Summary")
            safe_write(final_summary_df, writer, "CATs_MS-MF_Summary")
            safe_write(summary_df, writer, "CATs_Summary")
            safe_write(ms_mf_summary_df, writer, "MS-MF_Summary")
            safe_write(rfs_with_stock_report, writer, "CATs_Backup Data")
            safe_write(rfs_matched_sn_merege, writer, "CATs_Matched_by_SN")
            safe_write(ms_mf_data_cout, writer, "MS-MF_Backup Data")
            safe_write(ms_mf_matched_merged, writer, "MS-MF_Matched_by_SN")
            safe_write(new_df, writer, "SiteID Count Summary")

         
        relative_path = os.path.relpath(output_path, MEDIA_ROOT)
        download_url = request.build_absolute_uri(os.path.join(MEDIA_URL, relative_path).replace("\\", "/"))
        print("End Process_________")
       
        return Response({
            "status": True,
            "message": "File saved successfully",
            "download_url": download_url
        })
    except Exception as e:
        return Response({"error": f"find an error: {str(e)}"}, status=500)
    
    
#api for  Rfs mapping tool----------------------------------------
@api_view(["POST"])
def rfs_site_mapping(request):
    print("Start Process rfsmapping tool_________")
    try:
        # ----------------- Helpers -----------------
        def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
            """Clean and normalize column names"""
            df.columns = (
                df.columns.astype(str)
                .str.strip()
                .str.replace(r"\n+", "", regex=True)
                .str.replace(r"\s+", " ", regex=True)
            )
            return df

        def move_after(cols: list, col_to_move: str, after_col: str) -> list:
            """Reorder columns by moving col_to_move after after_col"""
            if col_to_move in cols and after_col in cols:
                cols.insert(cols.index(after_col) + 1, cols.pop(cols.index(col_to_move)))
            return cols

        # ----------------- RFS File -----------------
        rfs_folder = os.path.join(main_folder, "rfs_data")
        if not os.path.exists(rfs_folder):
            return Response({"error": "rfs_data folder not found"}, status=400)

        rfs_df = None
        for file in os.listdir(rfs_folder):
            rfs_file_path = os.path.join(rfs_folder, file)
            if not os.path.isfile(rfs_file_path):
                continue
            if rfs_file_path.endswith(".csv"):
                rfs_df = pd.read_csv(rfs_file_path, chunksize=50000)  # <-- chunked
            elif rfs_file_path.endswith((".xls", ".xlsx")):
                rfs_df = pd.read_excel(rfs_file_path, engine="openpyxl")
            else:
                return Response({"error": f"Unsupported file format: {file}"}, status=400)
        if rfs_df is None:
            return Response({"error": "No RFS file found"}, status=400)

        # ----------------- Locator Files -----------------
        locator_folder = os.path.join(main_folder, "locator_data")
        if not os.path.exists(locator_folder):
            return Response({"error": "locator_data folder not found"}, status=400)

        locator_df_list = []
        for file in os.listdir(locator_folder):
            locator_file_path = os.path.join(locator_folder, file)
            if not os.path.isfile(locator_file_path):
                continue
            if locator_file_path.endswith(".csv"):
                df = pd.read_csv(locator_file_path, usecols=["LOCATION_NAME", "SITEID"])
            elif locator_file_path.endswith((".xls", ".xlsx")):
                df = pd.read_excel(locator_file_path, engine="openpyxl", usecols=["LOCATION_NAME", "SITEID"])
            else:
                return Response({"error": f"Unsupported file format: {file}"}, status=400)
            clean_columns(df)
            locator_df_list.append(df)
        locater_df = pd.concat(locator_df_list, ignore_index=True).drop_duplicates()

        # ----------------- OLM ID File -----------------
        olm_id_file = request.FILES.get("olm")
        if not olm_id_file:
            return Response({"error": "olm_id not provided"}, status=400)

        if olm_id_file.name.endswith(".csv"):
            olm_df = pd.read_csv(olm_id_file)
        elif olm_id_file.name.endswith((".xls", ".xlsx")):
            olm_df = pd.read_excel(olm_id_file, engine="openpyxl")
        else:
            return Response({"error": "Invalid file format for olm_id_file"}, status=400)
        clean_columns(olm_df)

        # ----------------- HW Files -----------------
        hw_files = request.FILES.getlist("hw")
        if not hw_files:
            return Response({"error": "hw_file(s) not provided"}, status=400)

        hw_df_list = []
        for hw_file in hw_files:
            if hw_file.name.endswith(".csv"):
                df = pd.read_csv(hw_file, usecols=["ITEMCODE", "Module Name", "Module type", "Tech"])
            elif hw_file.name.endswith((".xls", ".xlsx")):
                df = pd.read_excel(hw_file, engine="openpyxl", usecols=["ITEMCODE", "Module Name", "Module type", "Tech"])
            else:
                return Response({"error": "Unsupported file format for hw_file"}, status=400)
            clean_columns(df)
            hw_df_list.append(df)
        hw_df = pd.concat(hw_df_list, ignore_index=True).drop_duplicates()

        # ----------------- Process RFS in Chunks -----------------
        export_dir = os.path.join(main_folder, "RFS_Mapping_Data")
        os.makedirs(export_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(export_dir, f"RFS_Mapping_Output{timestamp}.csv")

        total_rows = 0
        if isinstance(rfs_df, pd.io.parsers.TextFileReader):  # chunked CSV
            for i, rfs_chunk in enumerate(rfs_df):
                clean_columns(rfs_chunk)
                rfs_chunk["FROMLOCATION"] = rfs_chunk["FROMLOCATION"].astype(str).str.strip()
                rfs_chunk["ITEMCODE"] = rfs_chunk["ITEMCODE"].astype(str).str.strip()

                merged_rfs = (
                    rfs_chunk
                    .merge(olm_df[["OLM ID", "Partner"]], left_on="RFS_CREATED_BY", right_on="OLM ID", how="left")
                    .drop(columns=["OLM ID"])
                    .merge(locater_df, left_on="FROMLOCATION", right_on="LOCATION_NAME", how="left")
                    .drop(columns=["LOCATION_NAME"])
                    .merge(hw_df, on="ITEMCODE", how="left")
                )

                merged_rfs["Partner"] = merged_rfs["Partner"].replace("", None).fillna("Other TSP")

                # Column reordering
                cols = merged_rfs.columns.tolist()
                cols = move_after(cols, "SITEID", "FROMLOCATION")
                for col in ["Module Name", "Module type", "Tech"]:
                    cols = move_after(cols, col, "ITEM_DESCRIPTION")
                cols = move_after(cols, "Partner", "RFS_CREATED_BY")
                merged_rfs = merged_rfs[cols]

                total_rows += len(merged_rfs)
                merged_rfs.to_csv(output_path, mode="a", header=(i == 0), index=False, encoding="utf-8-sig")
        else:
            # Excel file (no chunks)
            clean_columns(rfs_df)
            merged_rfs = (
                rfs_df
                .merge(olm_df[["OLM ID", "Partner"]], left_on="RFS_CREATED_BY", right_on="OLM ID", how="left")
                .drop(columns=["OLM ID"])
                .merge(locater_df, left_on="FROMLOCATION", right_on="LOCATION_NAME", how="left")
                .drop(columns=["LOCATION_NAME"])
                .merge(hw_df, on="ITEMCODE", how="left")
            )
            merged_rfs["Partner"] = merged_rfs["Partner"].replace("", None).fillna("Other TSP")
            merged_rfs.to_csv(output_path, index=False, encoding="utf-8-sig")
            total_rows = len(merged_rfs)

        # ----------------- API Response -----------------
        relative_path = os.path.relpath(output_path, MEDIA_ROOT)
        download_url = request.build_absolute_uri(os.path.join(MEDIA_URL, relative_path).replace("\\", "/"))

        print("End Process_________")
        return Response({
            "status": True,
            "rows_processed": total_rows,
            "message": "File saved successfully",
            "download_url": download_url
        })

    except Exception as e:
        return Response({"error": f"find an error: {str(e)}"}, status=500)
    


# API for MS-MF Mapping Tool ----------------------------------------
@api_view(["POST"])
def ms_mf_site_mapping(request):
    print("Start Process for msmfmapping tool_________")
    try:
        # ----------------- Helpers -----------------
        def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
            """Clean and normalize column names"""
            df.columns = (
                df.columns.astype(str)
                .str.strip()
                .str.replace(r"\n+", "", regex=True)
                .str.replace(r"\s+", " ", regex=True)
            )
            return df

        def move_after(cols: list, col_to_move: str, after_col: str) -> list:
            """Reorder columns by moving col_to_move after after_col"""
            if col_to_move in cols and after_col in cols:
                cols.insert(cols.index(after_col) + 1, cols.pop(cols.index(col_to_move)))
            return cols

        # ----------------- MS-MF File -----------------
        msmf_folder = os.path.join(main_folder, "msmf_data")
        if not os.path.exists(msmf_folder):
            return Response({"error": "msmf_data folder not found"}, status=400)

        msmf_df = None
        for file in os.listdir(msmf_folder):
            msmf_file_path = os.path.join(msmf_folder, file)
            if not os.path.isfile(msmf_file_path):
                continue
            if msmf_file_path.endswith(".csv"):
                msmf_df = pd.read_csv(msmf_file_path, chunksize=50000)  # chunked
            elif msmf_file_path.endswith((".xls", ".xlsx")):
                msmf_df = pd.read_excel(msmf_file_path, engine="openpyxl")
            else:
                return Response({"error": f"Unsupported file format: {file}"}, status=400)
        if msmf_df is None:
            return Response({"error": "No MS-MF file found"}, status=400)

        # ----------------- Locator Files -----------------
        locator_folder = os.path.join(main_folder, "locator_data")
        if not os.path.exists(locator_folder):
            return Response({"error": "locator_data folder not found"}, status=400)

        locator_df_list = []
        for file in os.listdir(locator_folder):
            locator_file_path = os.path.join(locator_folder, file)
            if not os.path.isfile(locator_file_path):
                continue
            if locator_file_path.endswith(".csv"):
                df = pd.read_csv(locator_file_path, usecols=["LOCATION_NAME", "SITEID"])
            elif locator_file_path.endswith((".xls", ".xlsx")):
                df = pd.read_excel(locator_file_path, engine="openpyxl", usecols=["LOCATION_NAME", "SITEID"])
            else:
                return Response({"error": f"Unsupported file format: {file}"}, status=400)
            clean_columns(df)
            locator_df_list.append(df)
        locator_df = pd.concat(locator_df_list, ignore_index=True).drop_duplicates()

        # ----------------- OLM ID File -----------------
        olm_id_file = request.FILES.get("olm")
        if not olm_id_file:
            return Response({"error": "olm_id not provided"}, status=400)

        if olm_id_file.name.endswith(".csv"):
            olm_df = pd.read_csv(olm_id_file)
        elif olm_id_file.name.endswith((".xls", ".xlsx")):
            olm_df = pd.read_excel(olm_id_file, engine="openpyxl")
        else:
            return Response({"error": "Invalid file format for olm_id_file"}, status=400)
        clean_columns(olm_df)

        # ----------------- HW Files -----------------
        hw_files = request.FILES.getlist("hw")
        if not hw_files:
            return Response({"error": "hw_file(s) not provided"}, status=400)

        hw_df_list = []
        for hw_file in hw_files:
            if hw_file.name.endswith(".csv"):
                df = pd.read_csv(hw_file, usecols=["ITEMCODE", "Module Name", "Module type", "Tech"])
            elif hw_file.name.endswith((".xls", ".xlsx")):
                df = pd.read_excel(hw_file, engine="openpyxl", usecols=["ITEMCODE", "Module Name", "Module type", "Tech"])
            else:
                return Response({"error": "Unsupported file format for hw_file"}, status=400)
            clean_columns(df)
            hw_df_list.append(df)
        hw_df = pd.concat(hw_df_list, ignore_index=True).drop_duplicates()

        # ----------------- Process MSMF in Chunks -----------------
        export_dir = os.path.join(main_folder, "MSMF_Mapping_Data")
        os.makedirs(export_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(export_dir, f"MSMF_Mapping_Output{timestamp}.csv")

        total_rows = 0
        if isinstance(msmf_df, pd.io.parsers.TextFileReader):  # chunked CSV
            for i, msmf_chunk in enumerate(msmf_df):
                clean_columns(msmf_chunk)
                msmf_chunk["MS_FROMLOCATION"] = msmf_chunk["MS_FROMLOCATION"].astype(str).str.strip()
                msmf_chunk["PARTCODE"] = msmf_chunk["PARTCODE"].astype(str).str.strip()

                merged_msmf = (
                    msmf_chunk
                    .merge(olm_df[["OLM ID", "Partner"]], left_on="MS_CREATED_BY", right_on="OLM ID", how="left")
                    .drop(columns=["OLM ID"])
                    .merge(locator_df, left_on="MS_FROMLOCATION", right_on="LOCATION_NAME", how="left")
                    .drop(columns=["LOCATION_NAME"])
                    .merge(hw_df, left_on="PARTCODE", right_on="ITEMCODE", how="left")
                )

                merged_msmf["Partner"] = merged_msmf["Partner"].replace("", None).fillna("Other TSP")

                # Column reordering
                cols = merged_msmf.columns.tolist()
                cols = move_after(cols, "SITEID", "MS_FROMLOCATION")
                for col in ["Module Name", "Module type", "Tech"]:
                    cols = move_after(cols, col, "PART_DESCRIPTION")
                cols = move_after(cols, "Partner", "MS_CREATED_BY")
                merged_msmf = merged_msmf[cols]

                total_rows += len(merged_msmf)
                merged_msmf.to_csv(output_path, mode="a", header=(i == 0), index=False, encoding="utf-8-sig")
        else:
            # Excel file (no chunks)
            clean_columns(msmf_df)
            merged_msmf = (
                msmf_df
                .merge(olm_df[["OLM ID", "Partner"]], left_on="MS_CREATED_BY", right_on="OLM ID", how="left")
                .drop(columns=["OLM ID"])
                .merge(locator_df, left_on="MS_FROMLOCATION", right_on="LOCATION_NAME", how="left")
                .drop(columns=["LOCATION_NAME"])
                .merge(hw_df, left_on="PARTCODE", right_on="ITEMCODE", how="left")
            )
            merged_msmf["Partner"] = merged_msmf["Partner"].replace("", None).fillna("Other TSP")
            merged_msmf.to_csv(output_path, index=False, encoding="utf-8-sig")
            total_rows = len(merged_msmf)

        # ----------------- API Response -----------------
        relative_path = os.path.relpath(output_path, MEDIA_ROOT)
        download_url = request.build_absolute_uri(os.path.join(MEDIA_URL, relative_path).replace("\\", "/"))

        print("End Process_________")
        return Response({
            "status": True,
            "rows_processed": total_rows,
            "message": "File saved successfully",
            "download_url": download_url
        })

    except Exception as e:
        return Response({"error": f"find an error: {str(e)}"}, status=500)

    print("Start Process for msmfmapping  tool_________")
    try:
         #MS-MF file-----------------------------------------------
        msmf_folder = os.path.join(main_folder, 'msmf_data')
        if not os.path.exists(msmf_folder):
            return Response({"error": "msmf_data folder not found"}, status=400)
       
        for file in os.listdir(msmf_folder):
            msmf_file_path = os.path.join(msmf_folder, file)
            if not os.path.isfile(msmf_file_path):
                continue
            try:
                if msmf_file_path.endswith('.csv'):
                    msmf_df = pd.read_csv(msmf_file_path)
                elif msmf_file_path.endswith(('.xls', '.xlsx')):
                    msmf_df = pd.read_excel(msmf_file_path, engine='openpyxl')
                else:
                    return Response({"error": f"Unsupported file format: {file}"}, status=400)  
                msmf_df.columns = msmf_df.columns.str.strip()
                print("MS-MF Data -------------:", msmf_df.head())
            except Exception as e:
             return Response({"error": f"error reading olm_id_file: {str(e)}"}, status=500)
 
        # Locator Files -------------------------------------------------
        locator_folder = os.path.join(main_folder, 'locator_data')
        if not os.path.exists(locator_folder):
            return Response({"error": "locator_data folder not found"}, status=400)
       
        locator_df_list = []
        for file in os.listdir(locator_folder):
            locator_file_path = os.path.join(locator_folder, file)
            if not os.path.isfile(locator_file_path):
                continue
            try:
                if locator_file_path.endswith('.csv'):
                    df = pd.read_csv(locator_file_path, usecols=['LOCATION_NAME', 'SITEID'])    
                elif locator_file_path.endswith(('.xls', '.xlsx')):
                    df = pd.read_excel(locator_file_path, engine='openpyxl', usecols=['LOCATION_NAME', 'SITEID'])
                else:
                    return Response({"error": f"Unsupported file format: {file}"}, status=400)
                df.columns = df.columns.str.strip()
                locator_df_list.append(df)
            except Exception as e:
                return Response({"error": f"error reading locator file: {str(e)}"}, status=500)
        locator_df = pd.concat(locator_df_list, ignore_index=True).drop_duplicates()
        print("locator_df:",locator_df.head())
       
           
        #OLM ID File------------------------------------------------------------``
        olm_id_file = request.FILES.get("olm_id")
        if not olm_id_file:
            return Response({"error": "olm_id not provided"}, status=400)
        try:
            if olm_id_file.name.endswith('.csv'):
                olm_id_df = pd.read_csv(olm_id_file)
            elif olm_id_file.name.endswith(('.xls', '.xlsx')):
                olm_id_df = pd.read_excel(olm_id_file, engine='openpyxl')
            else:
                return Response({"error": "Invalid file format for olm_id_file"}, status=400)
            olm_id_df.columns = olm_id_df.columns.str.strip()
            print(olm_id_df.columns)
        except Exception as e:
            return Response({"error": f"error reading olm_id_file: {str(e)}"}, status=500)
 
        #  Multiple HW Files--------------------------------------------------
        hw_files = request.FILES.getlist("hw_file")
        if not hw_files:
            return Response({"error": "hw_file(s) not provided"}, status=400)
 
        hw_df_list = []
        for hw_file in hw_files:
            try:
                if hw_file.name.endswith('.csv'):
                    df = pd.read_csv(hw_file, usecols=["ITEMCODE", "Module Name", "Module type", "Tech"])
                elif hw_file.name.endswith(('.xls', '.xlsx')):
                 df = pd.read_excel(hw_file, engine='openpyxl', usecols=["ITEMCODE", "Module Name", "Module type", "Tech"])
                else:
                    return Response({"error": "Unsupported file format for hw_file"}, status=400)
                df.columns = df.columns.str.strip()
                hw_df_list.append(df)
            except Exception as e:
                return Response({"error": f"error reading a hw_file: {str(e)}"}, status=500)
        hw_df = pd.concat(hw_df_list, ignore_index=True).drop_duplicates()
        print(hw_df.columns)
       
       
        merged_data = (
                msmf_df
                .merge(olm_id_df[["OLM ID", "Partner"]], left_on="MS_CREATED_BY", right_on="OLM ID", how="left")
                .drop(columns=["OLM ID"])
                .merge(locator_df, left_on="MS_FROMLOCATION", right_on="LOCATION_NAME", how="left")
                .drop(columns=["LOCATION_NAME"])
                .merge(hw_df, left_on="PARTCODE", right_on="ITEMCODE", how="left")
            )
 
        if "Partner" in merged_data.columns:
            merged_data["Partner"] = merged_data["Partner"].replace("", None).fillna("Other TSP")
               
        # ----------------- Save Output -----------------
            export_dir = os.path.join(main_folder, "MSMF_Mapping_Data")
            os.makedirs(export_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(export_dir, f"MSMF_Mapping_Output{timestamp}.xlsx")
 
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                merged_data.to_excel(writer, sheet_name="Msmf_Mapping", index=False)
                final_excel(merged_data, sheet_name="Msmf_Mapping", fast_mode=True)
 
   
            relative_path = os.path.relpath(output_path, MEDIA_ROOT)
            download_url = request.build_absolute_uri(
                os.path.join(MEDIA_URL, relative_path).replace("\\", "/")
            ).replace("\\", "/")
 
            print("End Process_________")
 
            return Response({
                "status": True,
                "message": "File saved successfully",
                "download_url": download_url
            })      
    except Exception as e:
     return Response({"error": f"find an error: {str(e)}"}, status=500)            




    #new api------ site+searial match
    #here the logic is to match the serial number from mobinet dump file with rfs and ms-mf file------


@api_view(['GET','POST'])
def mobinet_sitecircle_match(request):
    mobinet_log_folder = os.path.join(main_folder, 'mobinet_dump_data')
    if not os.path.exists(mobinet_log_folder):
        return Response({"error": "mobinet_dump_data folder not found"}, status=400)

    log_df_list = []
    for file in os.listdir(mobinet_log_folder):
        file_path = os.path.join(mobinet_log_folder, file)
        if not os.path.isfile(file_path):
            continue
        try:
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            elif file_path.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file_path, engine='openpyxl')
            else:
                continue
            log_df_list.append(df)
        except Exception as e:
            return Response({"error": f"Error reading file {file}: {str(e)}"}, status=500)

    if not log_df_list:
        return Response({"error": "No valid Mobinet log files found"}, status=400)

    log_df = pd.concat(log_df_list, ignore_index=True)

    # ------------------- RFS DATA -------------------
    rfs_df = pd.DataFrame()  # default blank
    rfs_folder = os.path.join(main_folder, 'rfs_data')
    if os.path.exists(rfs_folder):
        for file in os.listdir(rfs_folder):
            rfs_file_path = os.path.join(rfs_folder, file)
            if not os.path.isfile(rfs_file_path):
                continue
            try:
                if rfs_file_path.endswith('.csv'):
                    rfs_df = pd.read_csv(rfs_file_path)
                elif rfs_file_path.endswith(('.xls', '.xlsx')):
                    rfs_df = pd.read_excel(rfs_file_path, engine='openpyxl')
                else:
                    continue
                rfs_df.columns = rfs_df.columns.str.strip()
                # print("RFS columns:", rfs_df.columns)
                break  # take only first valid file
            except Exception as e:
                print(f"Error reading RFS file: {e}")
    else:
        print("rfs_data folder not found — continuing with blank DataFrame.")

    # ------------------- MSMF DATA -------------------
    msmf_df = pd.DataFrame()  # default blank
    msmf_folder = os.path.join(main_folder, 'msmf_data')
    if os.path.exists(msmf_folder):
        for file in os.listdir(msmf_folder):
            msmf_file_path = os.path.join(msmf_folder, file)
            if not os.path.isfile(msmf_file_path):
                continue
            try:
                if msmf_file_path.endswith('.csv'):
                    msmf_df = pd.read_csv(msmf_file_path)
                elif msmf_file_path.endswith(('.xls', '.xlsx')):
                    msmf_df = pd.read_excel(msmf_file_path, engine='openpyxl')
                else:
                    continue
                msmf_df.columns = msmf_df.columns.str.strip()
                # print("MS-MF columns:", msmf_df.columns)
                break
            except Exception as e:
                print(f"Error reading MS-MF file: {e}")
    else:
        print("msmf_data folder not found — continuing with blank DataFrame.")

    # ------------------- LOCATOR DATA -------------------
    locator_folder = os.path.join(main_folder, 'locator_data')
    if not os.path.exists(locator_folder):
        return Response({"error": "locator_data folder not found"}, status=400)

    locator_df_list = []
    for file in os.listdir(locator_folder):
        locator_file_path = os.path.join(locator_folder, file)
        if not os.path.isfile(locator_file_path):
            continue
        try:
            if locator_file_path.endswith('.csv'):
                df = pd.read_csv(locator_file_path, usecols=['LOCATION_NAME', 'SITEID'])
            elif locator_file_path.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(locator_file_path, engine='openpyxl', usecols=['LOCATION_NAME', 'SITEID'])
            else:
                continue
            df.columns = df.columns.str.strip()
            locator_df_list.append(df)
        except Exception as e:
            print(f"Error reading locator file {file}: {e}")

    locator_df = pd.concat(locator_df_list, ignore_index=True).drop_duplicates()

    # ------------------- FILTER MOBINET SERIALS -------------------
    log_df = log_df[
        log_df['Serial Number'].notna() &
        (log_df['Serial Number'].astype(str).str.strip() != '') &
        (log_df['Serial Number'].astype(str).str.strip() != '-')
    ]
    log_df['SiteID+SN'] = log_df['Parent Site'].astype(str) + "_" + log_df['Serial Number'].astype(str)

    # ------------------- Circle Mapping -------------------
    circle_map_code = {
        151: 'AP', 132: 'BR', 152: 'CN', 111: 'DL', 113: 'HR', 114: 'JK',
        153: 'KK', 174: 'MU', 172: 'MP', 134: 'OR', 115: 'PB', 176: 'RJ',
        116: 'UE', 117: 'UW', 135: 'WB', 136: 'KO', 155: 'TN', 133: 'NE', 131: 'AS'
    }

    # ------------------- RFS MERGE -------------------
    if not rfs_df.empty:
        merged_rfs_df = pd.merge(
            rfs_df, locator_df[['LOCATION_NAME', 'SITEID']],
            how='inner', left_on='FROMLOCATION', right_on='LOCATION_NAME', sort=False
        )
        merged_rfs_df['Circle_Map'] = merged_rfs_df['FROMLOCATION'].astype(str).str[:3].astype(int)
        merged_rfs_df['Circle'] = merged_rfs_df['Circle_Map'].map(circle_map_code)
        merged_rfs_df['Unique Site ID'] = merged_rfs_df['SITEID'].astype(str) + "_" + merged_rfs_df['Circle'].astype(str)
        merged_rfs_df.drop(columns=['LOCATION_NAME', 'Circle_Map'], inplace=True)
        merged_rfs_df["SiteID+SN"] = merged_rfs_df["Unique Site ID"].astype(str) + "_" + merged_rfs_df["SERIALNUMBER"].astype(str)

        matched_rfs_sn = pd.merge(
            merged_rfs_df, 
            log_df[['SiteID+SN', 'Board Model']],
            how='inner', 
            on='SiteID+SN', 
            sort=False
        )
    else:
        matched_rfs_sn = pd.DataFrame(columns=['No RFS Data Found'])

    # ------------------- MSMF MERGE -------------------
    if not msmf_df.empty:
        merged_msmf_df = pd.merge(
            msmf_df, locator_df[['LOCATION_NAME', 'SITEID']],
            how='inner', left_on='MS_FROMLOCATION', right_on='LOCATION_NAME', sort=False
        )
        merged_msmf_df['Circle_Map'] = merged_msmf_df['MS_FROMLOCATION'].astype(str).str[:3].astype(int)
        merged_msmf_df['Circle'] = merged_msmf_df['Circle_Map'].map(circle_map_code)
        merged_msmf_df['Unique Site ID'] = merged_msmf_df['SITEID'].astype(str) + "_" + merged_msmf_df['Circle'].astype(str)
        merged_msmf_df.drop(columns=['LOCATION_NAME', 'Circle_Map'], inplace=True)
        merged_msmf_df["SiteID+SN"] = merged_msmf_df["Unique Site ID"].astype(str) + "_" + merged_msmf_df["SERIAL_NUMBER"].astype(str)

        matched_msmf_sn = pd.merge(
            merged_msmf_df, 
            log_df[['SiteID+SN', 'Board Model']],
            how='inner',
            on='SiteID+SN', 
            sort=False
        )
    else:
        matched_msmf_sn = pd.DataFrame(columns=['No MS-MF Data Found'])
        
   

    # ------------------ OUTPUT FILE --------------------
    output_dir = os.path.join(main_folder, 'SN_Match_Output')
    os.makedirs(output_dir, exist_ok=True)
    delete_existing_files(output_dir)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"SN_Match_Output_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)

    with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
        # safe_write(log_df, writer, "Mobinet_Summary")
        safe_write(matched_rfs_sn, writer, "RFS_Summary")
        safe_write(matched_msmf_sn, writer, "MS-MF_Summary")
        
    print("End Of Process_________")
    relative_path = os.path.relpath(output_path, MEDIA_ROOT)
    download_url = request.build_absolute_uri(
        os.path.join(MEDIA_URL, relative_path).replace("\\", "/")
    ).replace("\\", "/")

    return Response({
        "status": True,
        "message": "All Files processed successfully",
        "download_url": download_url
    })








@api_view(["GET", "POST", "DELETE"])
def dismental_template(request):
    print("Start Process_________")
    folder = os.path.join(main_folder, "dismental_tem")
    os.makedirs(folder, exist_ok=True)

    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)

        path = os.path.join(folder, file.name)

        with open(path, 'wb+') as f:
            for chunk in file.chunks():
                f.write(chunk)

        return Response({'status': True, 'file': file.name})

    if request.method == 'GET':
        return Response({'files': os.listdir(folder)})

    if request.method == 'DELETE':
        for f in os.listdir(folder):
            os.remove(os.path.join(folder, f))
        return Response({'status': True, 'message': 'All files deleted'})


# new api for degrow dismental deshboard-------------------------------
from .utils import *
@api_view(['POST'])
def dismental_desh(request):
    print("Start Process_________")
 
    #read_template----
    temp_folder=os.path.join(main_folder, "dismental_tem")
    if not os.path.exists(temp_folder):
        return Response({"error": "temp folder not found"}, status=400)
 
    tem_df = None
    for file in os.listdir(temp_folder):
        file_path = os.path.join(temp_folder, file)
        if not os.path.isfile(file_path):
            continue
        if file_path.endswith(".csv"):
            tem_df = pd.read_csv(file_path)
        elif file_path.endswith((".xls", ".xlsx")):
            tem_df = pd.read_excel(file_path, engine="openpyxl")
        break
    # print(tem_df.columns)
 
 
    mobinet_log_folder = os.path.join(main_folder, "mobinet_dump_data")
    if not os.path.exists(mobinet_log_folder):
        return Response({"error": "mobinet_dump_data folder not found"}, status=400)
 
    log_df_list = []
    for file in os.listdir(mobinet_log_folder):
        file_path = os.path.join(mobinet_log_folder, file)
        if not os.path.isfile(file_path):
            continue
        if file_path.endswith(".csv"):
            df = pd.read_csv(file_path)
        elif file_path.endswith((".xls", ".xlsx")):
            df = pd.read_excel(file_path, engine="openpyxl")
        else:
            continue
        log_df_list.append(df)
 
    if not log_df_list:
        return Response({"error": "No valid log files found"}, status=400)
 
    mob_df = pd.concat(log_df_list, ignore_index=True)
    print("Concatenated all mobinates:")
    # print(mob_df.columns)
 
 
    #Read site list file
    site_list_file = request.FILES.get("site_list")
    if not site_list_file:
        return Response({"error": "site_list file not provided"}, status=400)
    site_df = read_file(site_list_file)
    site_df.columns = site_df.columns.str.strip()
    # print(site_df)
   
#rfs-----
    rfs_files = request.FILES.getlist("rfs_file")  
    rfs_file_list = []
    for file in rfs_files:
        if file.name.endswith(".xlsx"):
            rfs_df = pd.read_excel(file)
        elif file.name.endswith(".csv"):
            rfs_df = pd.read_csv(file)
        elif file.name.endswith(".xlsb"):
            rfs_df = pd.read_excel(file, engine="pyxlsb")
        else:
            continue
        rfs_df.columns = rfs_df.columns.str.strip()
        rfs_file_list.append(rfs_df)
 
    rfs_df = (pd.concat(rfs_file_list, ignore_index=True)
        if rfs_file_list
        else pd.DataFrame()
    )
 
#msmf-----
    msmf_files = request.FILES.getlist("msmf_file")  
    msmf_file_list = []
    for file in msmf_files:
        if file.name.endswith(".xlsx"):
            msmf_df = pd.read_excel(file)
        elif file.name.endswith(".csv"):
            msmf_df = pd.read_csv(file)
        elif file.name.endswith(".xlsb"):
            msmf_df = pd.read_excel(file, engine="pyxlsb")
        else:
            continue
        msmf_df.columns = msmf_df.columns.str.strip()
        msmf_file_list.append(msmf_df)
 
    msmf_df = (pd.concat(msmf_file_list, ignore_index=True)
        if msmf_file_list
        else pd.DataFrame())
   
 
#upload by hardware file-----
    hardware_file = request.FILES.getlist("hw_file")
    if not hardware_file:
        return Response(
            {"status": "ERROR", "message": "hardware_file not uploaded"},
            status=HTTP_400_BAD_REQUEST
        )
    hardware_file_list = []
    for file in hardware_file:
        if file.name.endswith('.xlsx'):
            hardware_df = pd.read_excel(file)
        elif file.name.endswith('.csv'):
            hardware_df = pd.read_csv(file)
        elif file.name.endswith('.xlsb'):
            hardware_df = pd.read_excel(file, engine="pyxlsb")
        else:
            return Response(
                {"status": "ERROR", "message": "Unsupported file type in hardware_file"},
                status=HTTP_400_BAD_REQUEST
            )
        hardware_df.columns = hardware_df.columns.str.strip()
        hardware_file_list.append(hardware_df)
    hardware_df = pd.concat(hardware_file_list, ignore_index=True)  
 
   
 
 
    #start concept-----
    # phase 1 -> work with mobinate file
    mob_df["Parent Site"] = mob_df["Parent Site"].astype(str).str.strip()
    site_df["Unique ID"] = site_df["Unique ID"].astype(str).str.strip()
    # Merge logs with site list
    matched_df = pd.merge(
        mob_df,
        site_df,
        how="inner",
        left_on="Parent Site",
        right_on="Unique ID"
    )
    matched_df = matched_df.drop_duplicates(
        subset=["Board Serial Number"],
        keep="first"
    )
    # print(matched_df.columns)
   
    # fill mobinate data in template---
    tem_df['Circle']=matched_df['Zone']
    tem_df['Degrow type']=matched_df['Degrow type']
    tem_df['Plan Released-DD/MM/YY']=matched_df['Plan Released-DD/MM/YY']
    tem_df['Cir_Site']=matched_df['Parent Site']
    tem_df['Site ID']=matched_df['Parent Site'].str.split('_').str[0]
    tem_df['Module Qty']=1
    tem_df['Modules Name As per Mobinet']=matched_df['Board Model']
    tem_df['Module-Serial No']=matched_df['Board Serial Number']
    tem_df['Survey Date-DD/MM/YY']=matched_df['Survey Date-DD/MM/YY']
    tem_df['Dismantling Date-DD/MM/YY']=matched_df['Dismantled date']
   
    tem_df["matched_by_sn"] =(tem_df['Cir_Site'].astype(str).str.strip()+ "_"+tem_df['Module-Serial No'].astype(str).str.strip())
    tem_df["matched_by_module"] =(tem_df['Cir_Site'].astype(str).str.strip()+ "_"+tem_df['Modules Name As per Mobinet'].astype(str).str.strip())
 
    print("Phase 1-> mobinate part run----")
 
    #work with rfs file-----
    print(rfs_df.columns)
    rfs_df = rfs_df[~rfs_df["Srncamreqstatus"].astype(str).str.strip().str.upper()
    .isin(["SHORT RECEIVED", "CANCELLED", "FAILED"])]
    rfs_df["Partners"] = "Other TSP"
    mobile_mask = rfs_df["Partner Name"].astype(str).str.contains(
        "Mobilecom|Mobilecomm|MCOM",
        case=False,
        na=False
    )
    creator_mask = (
        rfs_df["Partner Name"].isna()
        |
        (rfs_df["Partner Name"].astype(str).str.strip() == "")
    ) & (
        rfs_df["Creatordepartment"].astype(str).str.contains(
            "Mobilecom|Mobilecomm|MCOM",
            case=False,
            na=False
        )
    )
    rfs_df.loc[
        mobile_mask | creator_mask,
        "Partners"
    ] = "Mobilecomm"
 
 
    final_rfs = pd.merge(
        rfs_df,
        hardware_df,
        how="left",
        left_on="Itemcode",
        right_on="ITEMCODE"
    )
 
    final_rfs["matched_by_sn"] =(final_rfs["Site ID"].astype(str).str.strip()+ "_"+ final_rfs["Circle"].astype(str).str.strip()+ "_"+ final_rfs["Serialnumber"].astype(str).str.strip())
    final_rfs["matched_by_module"] =( final_rfs["Site ID"].astype(str).str.strip()+ "_"+ final_rfs["Circle"].astype(str).str.strip()+ "_"+ final_rfs["Module Name"].astype(str).str.strip())
    # Match by SN
    merged_exact = pd.merge(
        tem_df,
        final_rfs,
        how="left",
        on="matched_by_sn",
        suffixes=("", "_rfs"),
        indicator=True
    )
 
    merged_exact["Remark"] = np.where(
        merged_exact["_merge"] == "both",
        "Matched by SN Number",
        ""
    )
 
    matched_mask = merged_exact["_merge"] == "both"
 
    # Not matched by SN
    unmatched_temp = merged_exact[
        ~matched_mask
    ][tem_df.columns].copy()
 
    # Match by Module
    merged_module = pd.merge(
        unmatched_temp,
        final_rfs,
        how="left",
        on="matched_by_module",
        suffixes=("", "_rfs"),
        indicator=True
    )
 
    merged_module["Remark"] = np.where(
        merged_module["_merge"] == "both",
        "Matched by Module",
        "Not found In RFS"
    )

    final_output_rfs = pd.concat(
    [
        merged_exact[matched_mask].drop(columns=["_merge"]),
        merged_module.drop(columns=["_merge"])
    ],
    ignore_index=True

)    
    
    final_output_rfs['Locator ID'] = final_output_rfs['Fromlocation']
    final_output_rfs['Partner Name'] = final_output_rfs['Partners']
    final_output_rfs['Module Name as per Dictionary'] = final_output_rfs['Module Name']
    final_output_rfs['Survey-Modules Name'] = final_output_rfs['Module Name']
    final_output_rfs['Survey-Module Qty'] = final_output_rfs['Rfs Qty']
    final_output_rfs['Survey-Module Serial No'] = final_output_rfs['Serialnumber']
    final_output_rfs['SREQ Number'] = final_output_rfs['Order ID']
    final_output_rfs['SREQ Date-DD/MM/YY'] = final_output_rfs['Rfs Created Date']
    final_output_rfs['Dismantled-Modules Name'] = final_output_rfs['Module Name']
    final_output_rfs['Dismantled-Module Qty'] = final_output_rfs['Rfs Qty']
    final_output_rfs['Dismantled-Module-Serial No'] = final_output_rfs['Serialnumber']
    final_output_rfs['DC NUMBER'] = final_output_rfs['Oracle Dc No']
    final_output_rfs['DC Request Date'] = final_output_rfs['Oracle Dc Date']
    final_output_rfs['Need By Date-DD/MM/YY'] = final_output_rfs['Needbydate']
 
    final_output_rfs["Pick Date"] = final_output_rfs['Picked Date']
    final_output_rfs["Pick/Ship Qty QTY"] = final_output_rfs['Picked Qty']
 
    final_output_rfs["PICK/SHIP Status"] = np.where(
        final_output_rfs["Pick/Ship Qty QTY"].isna(),
        "",
        np.where(final_output_rfs["Pick/Ship Qty QTY"] == 0, "Pending", "Done")
    )
 
    final_output_rfs["WH Submission Date-DD/MM/YY"] = final_output_rfs['Receipt Date']
    final_output_rfs["WH Submission Qty"] = final_output_rfs['Received Qty']
    final_output_rfs["OCI Date-DD/MM/YY"] = final_output_rfs['Receipt Date']
    final_output_rfs["OCI QTY"] = final_output_rfs['Received Qty']
 
    final_output_rfs["Usable/Non Usable"] = ""
    final_output_rfs["Final Status (Completed/Pending/WIP)"] = ""
    final_output_rfs["All SREQ Status"] = final_output_rfs['Srncamreqstatus']
    final_output_rfs["Module Type"] = final_output_rfs['Module type']
    final_output_rfs["SRN Lot Number"] = final_output_rfs['Oracle Lot No']
    final_output_rfs["Technology"] = final_output_rfs['Tech']
    final_output_rfs["Item Code"] = final_output_rfs['ITEMCODE']
    final_output_rfs["Item Description"] = final_output_rfs['ITEM_DESCRIPTION']
    final_output_rfs['Condtion Applied'] = final_output_rfs['Remark']
    # final_output_rfs['Remarks'] = final_output_rfs['Status_remark']
 

 
    # ==================================================
    # EXTRA SN LOGIC START
    # ==================================================
 
    valid_sites = tem_df["Site ID"].unique()
 
    final_same_site = final_rfs[
        final_rfs["Site ID"].isin(valid_sites)
    ].copy()
 
    matched_serials = set(
        final_output_rfs["Serialnumber"]
        .dropna()
        .astype(str)
        .str.strip()
    )
 
    extra_final = final_same_site[
        (~final_same_site["Serialnumber"]
            .astype(str)
            .str.strip()
            .isin(matched_serials))
    ].copy()
 
    extra_final["Remark"] = "Extra SN in RFS"
    extra_final_df = pd.DataFrame(columns=final_output_rfs.columns)
    extra_final_df["Site ID"]=extra_final["Site ID"]
    extra_final_df["Circle"]=extra_final["Circle"]
    extra_final_df["Locator ID"] = extra_final["Fromlocation"]
    extra_final_df["Partner Name"] = extra_final["Partners"]
    extra_final_df["Module Name as per Dictionary"] = extra_final["Module Name"]
    extra_final_df["Survey-Modules Name"] = extra_final["Module Name"]
    extra_final_df["Survey-Module Qty"] = extra_final["Rfs Qty"]
    extra_final_df["Survey-Module Serial No"] = extra_final["Serialnumber"]
    extra_final_df["SREQ Number"] = extra_final["Order ID"]
    extra_final_df["SREQ Date-DD/MM/YY"] = extra_final["Rfs Created Date"]
    extra_final_df["Dismantled-Modules Name"] = extra_final["Module Name"]
    extra_final_df["Dismantled-Module Qty"] = extra_final["Rfs Qty"]
    extra_final_df["Dismantled-Module-Serial No"] = extra_final["Serialnumber"]
    extra_final_df["DC NUMBER"] = extra_final["Oracle Dc No"]
    extra_final_df["DC Request Date"] = extra_final["Oracle Dc Date"]
    extra_final_df["Need By Date-DD/MM/YY"] = extra_final["Needbydate"]
    extra_final_df["Pick Date"] = extra_final["Picked Date"]
    extra_final_df["Pick/Ship Qty QTY"] = extra_final["Picked Qty"]
    extra_final_df["Ship Date"] = extra_final["Shipped Date"]
    extra_final_df["WH Submission Date-DD/MM/YY"] = extra_final["Receipt Date"]
    extra_final_df["WH Submission Qty"] = extra_final["Received Qty"]
    extra_final_df["OCI Date-DD/MM/YY"] = extra_final["Receipt Date"]
    extra_final_df["OCI QTY"] = extra_final["Received Qty"]
    extra_final_df["Usable/Non Usable"] = ""
    extra_final_df["Final Status (Completed/Pending/WIP)"] = ""
    extra_final_df["All SREQ Status"] = extra_final["Srncamreqstatus"]
    extra_final_df["Module Type"] = extra_final["Module type"]
    extra_final_df["SRN Lot Number"] = extra_final["Oracle Lot No"]
    extra_final_df["Technology"] = extra_final["Tech"]
    extra_final_df["Item Code"] = extra_final["ITEMCODE"]
    extra_final_df["Item Description"] = extra_final["ITEM_DESCRIPTION"]
    extra_final_df["Condtion Applied"] = "Extra SN in RFS"
    extra_final_df["Remarks"] = "Found in RFS"
    extra_final_df["Remark"] = "Extra SN in RFS"


    #wwork with msmf----
    rfs_found = final_output_rfs[
    final_output_rfs["Remark"] != "Not found In RFS"
    ].copy()

    rfs_not_found = final_output_rfs[
        final_output_rfs["Remark"] == "Not found In RFS"
    ].copy()


    msmf_df["Partners"] = "Other TSP"
    msmf_df.loc[
        msmf_df["Partner"].astype(str).str.contains(
            "Mobilecom|Mobilecomm",
            case=False,
            na=False
        ),
        "Partners"
    ] = "Mobliecomm"
    

    msmf_final = pd.merge(
        msmf_df,
        hardware_df,
        how="left",
        left_on="Partcode",
        right_on="ITEMCODE"
    )

    msmf_final["Circle"] = msmf_final["Circle Name"]
    msmf_final["Site ID"] = msmf_final["Ms Siteid"]

    msmf_final["matched_by_sn"] =(msmf_final["Site ID"].astype(str).str.strip()+ "_" + msmf_final["Circle"].astype(str).str.strip() + "_"+ msmf_final["Serial Number"].astype(str).str.strip())
    msmf_final["matched_by_module"] =(msmf_final["Site ID"].astype(str).str.strip()+ "_"+ msmf_final["Circle"].astype(str).str.strip()+ "_"+ msmf_final["Module Name"].astype(str).str.strip())
    msmf_final = msmf_final[
        [   "Partners",
            "matched_by_sn",
            "matched_by_module",
            "Ms Fromlocation",
            "Partner",
            "Module Name",
            "Module type",
            "Tech",
            "ITEMCODE",
            "ITEM_DESCRIPTION",
            "Serial Number",
            "Qty",
            "Dc Number",
            "Move Start Date",
            "Received Date",
            "Received Qty",
            "Site ID",
            "Circle"
        ]
    ].copy()
    merged_exact_msmf = pd.merge(
        rfs_not_found,
        msmf_final,
        how="left",
        on="matched_by_sn",
        suffixes=("", "_msmf"),
        indicator=True
    )

    matched_mask_msmf = merged_exact_msmf["_merge"] == "both"
    merged_exact_msmf["Remark"] = np.where(
        matched_mask_msmf,
        "Matched by SN in MSMF",
        ""
    )

    unmatched_msmf = merged_exact_msmf[
        ~matched_mask_msmf
    ].drop(columns=["_merge"], errors="ignore").copy()
    
    merged_module_msmf = pd.merge(
        unmatched_msmf,
        msmf_final,
        how="left",
        on="matched_by_module",
        indicator=True,
        suffixes=("", "_msmf")
    )
    merged_module_msmf["Remark"] = np.where(
        merged_module_msmf["_merge"] == "both",
        "Matched by Module in MSMF",
        "Not Found in MSMF"
    )
    merged_exact_msmf = merged_exact_msmf.loc[:, ~merged_exact_msmf.columns.duplicated()]
    merged_module_msmf = merged_module_msmf.loc[:, ~merged_module_msmf.columns.duplicated()]
    final_output_msmf = pd.concat(
    [
        merged_exact_msmf[matched_mask_msmf].drop(columns=["_merge"]),
        merged_module_msmf.drop(columns=["_merge"])
    ],
    ignore_index=True
    )
    print(final_output_msmf.columns)

        
    final_output_msmf['Locator ID'] = final_output_msmf['Ms Fromlocation']
    final_output_msmf['Partner Name'] = final_output_msmf['Partners']
    final_output_msmf['Module Name as per Dictionary'] = final_output_msmf['Module Name']
    final_output_msmf['Survey-Modules Name'] = final_output_msmf['Module Name']
    final_output_msmf['Survey-Module Qty'] = final_output_msmf['Qty']
    final_output_msmf['Survey-Module Serial No'] = final_output_msmf['Serial Number']
    final_output_msmf['SREQ Number'] = final_output_msmf['Dc Number']
    final_output_msmf['SREQ Date-DD/MM/YY'] = final_output_msmf['Move Start Date']
    final_output_msmf['Dismantled-Modules Name'] = final_output_msmf['Module Name']
    final_output_msmf['Dismantled-Module Qty'] = final_output_msmf['Qty']
    final_output_msmf['Dismantled-Module-Serial No'] = final_output_msmf['Serial Number']
    final_output_msmf['DC NUMBER'] = final_output_msmf['Dc Number']
    final_output_msmf['DC Request Date'] = ""
    final_output_msmf['Need By Date-DD/MM/YY'] = ""
 
    final_output_msmf["Pick Date"] =""
    final_output_msmf["Pick/Ship Qty QTY"] =""
 
    final_output_msmf["PICK/SHIP Status"] =""

    final_output_msmf["WH Submission Date-DD/MM/YY"] = final_output_msmf['Received Date']
    final_output_msmf["WH Submission Qty"] = final_output_msmf['Received Qty']
    final_output_msmf["OCI Date-DD/MM/YY"] = final_output_msmf['Received Date']
    final_output_msmf["OCI QTY"] = final_output_msmf['Received Qty']
 
    final_output_msmf["Usable/Non Usable"] = ""
    final_output_msmf["Final Status (Completed/Pending/WIP)"] = ""
    final_output_msmf["All SREQ Status"] =""
    final_output_msmf["Module Type"] = final_output_msmf['Module type']
    final_output_msmf["SRN Lot Number"] = ""
    final_output_msmf["Technology"] = final_output_msmf['Tech']
    final_output_msmf["Item Code"] = final_output_msmf['ITEMCODE']
    final_output_msmf["Item Description"] = final_output_msmf['ITEM_DESCRIPTION']
    final_output_msmf['Condtion Applied'] = final_output_msmf['Remark']


    valid_sites = tem_df["Site ID"].unique()
 
    final_same_site = msmf_final[
        msmf_final["Site ID"].isin(valid_sites)
    ].copy()
    
    matched_serials = set(
        final_output_msmf["Serial Number"]
        .dropna()
        .astype(str)
        .str.strip()
    )
 
    extra_final = final_same_site[
        (~final_same_site["Serial Number"]
            .astype(str)
            .str.strip()
            .isin(matched_serials))
    ].copy()
 
    extra_final["Remark"] = "Extra SN Msmf"
    extra_final_df_msmf = pd.DataFrame(columns=final_output_msmf.columns)
    extra_final_df_msmf["Site ID"]=extra_final["Site ID"]
    extra_final_df_msmf["Circle"]=extra_final["Circle"]
    extra_final_df_msmf["Locator ID"] = extra_final["Ms Fromlocation"]
    extra_final_df_msmf["Partner Name"] = extra_final["Partners"]
    extra_final_df_msmf["Module Name as per Dictionary"] = extra_final["Module Name"]
    extra_final_df_msmf["Survey-Modules Name"] = extra_final["Module Name"]
    extra_final_df_msmf["Survey-Module Qty"] = extra_final["Qty"]
    extra_final_df_msmf["Survey-Module Serial No"] = extra_final['Serial Number']
    extra_final_df_msmf["SREQ Number"] =""
    extra_final_df_msmf["SREQ Date-DD/MM/YY"] = extra_final["Move Start Date"]
    extra_final_df_msmf["Dismantled-Modules Name"] = extra_final["Module Name"]
    extra_final_df_msmf["Dismantled-Module Qty"] = extra_final["Qty"]
    extra_final_df_msmf["Dismantled-Module-Serial No"] = extra_final['Serial Number']
    extra_final_df_msmf["DC NUMBER"] = extra_final['Dc Number']
    extra_final_df_msmf["DC Request Date"] = ""
    extra_final_df_msmf["Need By Date-DD/MM/YY"] = ""
    extra_final_df_msmf["Pick Date"] = ""
    extra_final_df_msmf["Pick/Ship Qty QTY"] = ""
    extra_final_df_msmf["Ship Date"] = ""
    extra_final_df_msmf["WH Submission Date-DD/MM/YY"] = extra_final["Received Date"]
    extra_final_df_msmf["WH Submission Qty"] = extra_final["Received Qty"]
    extra_final_df_msmf["OCI Date-DD/MM/YY"] = extra_final["Received Date"]
    extra_final_df_msmf["OCI QTY"] = extra_final["Received Qty"]
    extra_final_df_msmf["Usable/Non Usable"] = ""
    extra_final_df_msmf["Final Status (Completed/Pending/WIP)"] = ""
    extra_final_df_msmf["All SREQ Status"] = ""
    extra_final_df_msmf["Module Type"] = extra_final["Module type"]
    extra_final_df_msmf["SRN Lot Number"] = ""
    extra_final_df_msmf["Technology"] = extra_final["Tech"]
    extra_final_df_msmf["Item Code"] = extra_final["ITEMCODE"]
    extra_final_df_msmf["Item Description"] = extra_final["ITEM_DESCRIPTION"]
    extra_final_df_msmf["Condtion Applied"] = "Extra SN in Msmf"
    extra_final_df_msmf["Remarks"] = "Found in Msmf"
    extra_final_df_msmf["Remark"] = "Extra SN in Msmf"
    


    # Final Output
    tem_df = pd.concat(
    [
        final_output_rfs,
        final_output_msmf,
        extra_final_df,
        extra_final_df_msmf
    ],
    ignore_index=True,
    sort=False
    )
    

 
    print(tem_df.columns)
    tem_df = tem_df.drop(columns=columns_list_rfs, errors="ignore")
    output_path = os.path.join(main_folder, "Dismtental_Deshboard")
    os.makedirs(output_path, exist_ok=True)
 
    output_file = os.path.join(output_path, "recodata.xlsx")
 
    tem_df.to_excel(output_file, index=False)
    format_excel(output_file)
    relative_path = os.path.relpath(output_file,MEDIA_ROOT)
    download_url=request.build_absolute_uri(
            MEDIA_URL + relative_path.replace("\\", "/"))
      

    return Response({
        "status": True,
        "message": "Report generated successfully",
        "download_url":download_url
    })
     