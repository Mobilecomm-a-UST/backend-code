from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side,PatternFill
from openpyxl.utils import get_column_letter
import os

def format_excel(file_path):
    wb = load_workbook(file_path)
    ws =wb.active
    header_fill = PatternFill(start_color="C4D79B",end_color="C4D79B",fill_type="solid")

    header_font = Font(name="Aptos", bold=True, size=9)
    center_align = Alignment(
        vertical="center",
        horizontal="center",
        wrap_text=True
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Header row height
    ws.row_dimensions[1].height =24

    # ---------- Data formatting ----------
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Aptos", size=9)
            cell.alignment = Alignment(vertical="center",horizontal="center", wrap_text=True)
            cell.border = border
            
    # ---------- Auto column width ----------
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        header = str(col[0].value).strip().upper() if col[0].value else ""

        # ITEM_DESCRIPTION column width fixed
        if header == "ITEM DESCRIPTION":
            ws.column_dimensions[col_letter].width = 75
            continue

        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 3, 45)
    wb.save(file_path)



def delete_existing_files(folder_path):
    if os.path.exists(folder_path):
        for file in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except Exception as e:
                print(f"Error deleting {file_path}: {e}")


columns_list_rfs = [
    "Partners",
    "matched_by_sn",
    "matched_by_module",
    "Order ID",
    "Circle_rfs",
    "Operatingunit",
    "Organization",
    "Book Type Code",
    "Fromlocation",
    "Site ID_rfs",
    "From Location Active",
    "From Location Locked",
    "Fromstatus",
    "Tolocation",
    "To Location Active",
    "To Location Locked",
    "Itemcode",
    "Gbv",
    "Nbv",
    "Item Description_rfs",
    "Spare Flag",
    "Assetcode",
    "Serialnumber",
    "E2e Faid",
    "Oracle Faid",
    "Partner Name_rfs",
    "Date Of Survey",
    "Items Found During Survey",
    "Date Of Dismantling",
    "No Of Items Dismantled",
    "Reason",
    "Movement Type",
    "Item Stocked",
    "Stock Source",
    "Stock Reason",
    "Rejection Reason",
    "Stockstatus",
    "Stockin",
    "Stock Approved By",
    "Stock Approved User Org",
    "Manufacturer_x",
    "Mfgpartnumber",
    "Serialized Flag",
    "Rfs Qty",
    "Needbydate",
    "Rfs Created By",
    "Rfs Created User Org",
    "Rfs Created Date",
    "Srncamreqnum",
    "Srncamreqstatus",
    "Srncam Creation Reprocessed By",
    "Creation Reprocessed User Org",
    "Recordid",
    "Rfs Process Flag",
    "Oracle Srn Cam No",
    "Reqnumber",
    "Reqstatus",
    "Req Child Level Status",
    "Site Receiving Hdr Status",
    "Transfernumber",
    "Trstatus",
    "Picked Qty",
    "Picked Date",
    "Picked By",
    "Picked User Org",
    "Shipped Qty",
    "Shipped Date",
    "Shipped By",
    "Shipped User Org",
    "Receipt Date",
    "Receipt By",
    "Receipt User Org",
    "Received Qty",
    "Oracle Lot No",
    "Oracle Dc No",
    "Oracle Dc Date",
    "Oraclestockloc Code",
    "Shortage Reason",
    "Error Message",
    "Notes",
    "Rfs Device Latitude",
    "Rfs Device Longitude",
    "Fault Code",
    "Fault Description",
    "Fault Detected Date",
    "Transfertype",
    "Gps Latitude",
    "Gps Longitude",
    "Geo Tag",
    "Creatorname",
    "Creatordepartment",
    "Receivername",
    "Receiverdepartment",
    "Pickedusername",
    "Pickeduserdept",
    "Shippedusername",
    "Shippeduserdept",
    "ITEMCODE",
    "ITEM_DESCRIPTION",
    "Module Name",
    "Module type",
    "Tech",
    "Manufacturer_y",
    "matched_by_module_rfs",
    "Remark",
    "matched_by_sn_rfs",
    "Status_remark",
     "Partners_msmf",
    "matched_by_module_msmf",
    "Ms Fromlocation",
    "Partner",
    "Module Name_msmf",
    "Module type_msmf",
    "Tech_msmf",
    "ITEMCODE_msmf",
    "ITEM_DESCRIPTION_msmf",
    "Serial Number",
    "Qty",
    "Dc Number",
    "Move Start Date",
    "Received Date",
    "Received Qty_msmf",
    "Site ID_msmf",
    "Circle_msmf",
    "matched_by_sn_msmf",
    "Ms Fromlocation_msmf",
    "Partner_msmf",
    "Serial Number_msmf",
    "Qty_msmf",
    "Dc Number_msmf",
    "Move Start Date_msmf",
    "Received Date_msmf"
]


