from django.shortcuts import render
from django.db import connection
# Create your views here.
from rest_framework.parsers import FileUploadParser
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework import status
from openpyxl import load_workbook
import pandas as pd
from .models import IntegrationDataVI, Document
import json
from rest_framework.exceptions import ValidationError
from asgiref.sync import sync_to_async
import os
from django.db import transaction
from django.conf import settings

from .serializers import IntegrationDataSerializer

from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from SOFT_AT_VINAY.models import Soft_At_Table
from datetime import datetime, timedelta
import pandas as pd
from django.http import JsonResponse
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser

from rest_framework.response import Response
from django.db.models import Count
from .models import IntegrationDataVI
from datetime import datetime, timedelta
from .parser import process_log

from .models import *




def generate_date_list(start_date):
    # Parse the input date (assuming it's a string in 'YYYY-MM-DD' format)
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    # Get the current date
    current_date = datetime.now().date()
    
    # Initialize an empty list to store the dates
    date_list = []
    
    # Generate dates from start_date to current_date
    delta = timedelta(days=1)  # Step by 1 day
    while start_date <= current_date:
        date_list.append(start_date)
        start_date += delta
    
    return date_list
@api_view(['GET'])
def get_excel_temp_link(request):
    #mcom123 temp id.
    # Path to the Excel file
    file_path = os.path.join(settings.MEDIA_ROOT, 'IntegrationTrackerVI', 'Integration_Tracker_VI_Template_V1.1.1.xlsm')

    # Check if the file exists
    if os.path.exists(file_path):
        # Construct the URL to access the file
        file_url = os.path.join(settings.MEDIA_URL ,'IntegrationTrackerVI' , 'Integration_Tracker_VI_Template_V1.1.1.xlsm')              
        return Response({'file_url': file_url,'template_version':'v1.1.1'}, status=status.HTTP_200_OK)
    else:
        # Return a 404 response if the file does not exist
        return Response({'error': 'Excel file not found'}, status=status.HTTP_404_NOT_FOUND)
    


import openpyxl
from io import BytesIO

def read_excel_cell(file_contents, sheet_name, cell):
    # Load the Excel file from memory
    wb = openpyxl.load_workbook(filename=BytesIO(file_contents))
    
    # Select the specific sheet by name
    sheet = wb[sheet_name]
    
    # Get the value of the specified cell
    cell_value = sheet[cell].value
    
    return cell_value


@api_view(['POST'])
def upload_excel(request):
    print("User:", request.user.username)

    file = request.data.get('file')
    if not file:
        return Response(
            {'error': 'No file uploaded'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # ✅ Validate template version
    try:
        value_in_cell = read_excel_cell(
            file.read(),
            sheet_name='Sheet2',
            cell='BE37'
        )
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    if value_in_cell != "mcom_v1.1.1":
        return Response(
            {'error': 'Only tool template can be uploaded. Please check template version'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Reset file pointer
    file.seek(0)

    # ✅ Read Excel
    try:
        df = pd.read_excel(
            file,
            sheet_name="Tracker",
            skiprows=1,
            keep_default_na=False
        )
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    # ✅ Clean column names
    df.columns = [col.strip() for col in df.columns]

    # Ensure string columns
    df["Cell ID"] = df["Cell ID"].astype(str)

    # ✅ Convert date columns safely
    date_columns = [
        'Integration Date',
        'FR Date',
        '4G HOTO OFFERED DATE',
        '4G HOTO ACCEPTED DATE',
        '2G HOTO OFFERED DATE',
        '2G HOTO ACCEPTED DATE',
    ]

    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')

    # ✅ DB Transaction (ALL or NOTHING)
    with transaction.atomic():

        for _, row in df.iterrows():

            # --- Safe date extraction ---
            integration_date = (
                row['Integration Date'].date()
                if not pd.isna(row['Integration Date'])
                else None
            )

            fr_date = row['FR Date'].date() if not pd.isna(row.get('FR Date')) else None
            hoto_offered_4g = row['4G HOTO OFFERED DATE'].date() if not pd.isna(row.get('4G HOTO OFFERED DATE')) else None
            hoto_accepted_4g = row['4G HOTO ACCEPTED DATE'].date() if not pd.isna(row.get('4G HOTO ACCEPTED DATE')) else None
            hoto_offered_2g = row['2G HOTO OFFERED DATE'].date() if not pd.isna(row.get('2G HOTO OFFERED DATE')) else None
            hoto_accepted_2g = row['2G HOTO ACCEPTED DATE'].date() if not pd.isna(row.get('2G HOTO ACCEPTED DATE')) else None

            # --- Build UNIQUE KEY (NEW LOGIC) ---
            circle = str(row['CIRCLE']).strip().upper()
            site_id = str(row['Site ID']).strip()
            technology = str(row['Technology (SIWA)']).strip().upper()

            unique_key = f"{circle}_{site_id}_{technology}"

            # --- UPSERT ---
            IntegrationDataVI.objects.update_or_create(
                unique_key=unique_key,  # 🔑 MAIN IDENTIFIER
                defaults={
                    'Integration_Date': integration_date,
                    'Activity_Name': str(row['Activity Name']).upper(),
                    'Site_ID': site_id,
                    'Technology_SIWA': technology,
                    'CIRCLE': circle,
                    'Cell_ID': row['Cell ID'],
                    'LNBTS_ID': row['LNBTS ID'],
                    'OEM': str(row['OEM']).upper(),

                    'MO_NAME': str(row['MO NAME']).upper(),
                    'OSS_Details': row['OSS Details'],
                    'CELL_COUNT': row['CELL COUNT'],
                    'TRX_Count': row['TRX Count'],
                    'PRE_ALARM': row['PRE-ALARM'],
                    'GPS_IP_CLK': row['GPS/IP CLK'],
                    'RET': row['RET'],
                    'POST_VSWR': row['POST-VSWR'],
                    'POST_Alarms': row['POST Alarms'],
                    'CELL_STATUS': row['CELL STATUS'],
                    'CTR_STATUS': row['CTR STATUS'],
                    'Integration_Remark': row['Integration Remark'],
                    'T2T4R': row['2T2R/4T4R'],
                    'BBU_TYPE': row['BBU TYPE'],
                    'BB_CARD': row['BB CARD'],
                    'RRU_Type': row['RRU Type'],
                    'Media_Status': row['Media Status'],
                    'Mplane_IP': row['Mplane IP'],
                    'SCF_PREPARED_BY': row['SCF PREPARED BY'],
                    'SITE_INTEGRATE_BY': row['SITE INTEGRATE BY(Integrator Name)'],
                    'Site_Status': row['Site Status'],
                    'External_Alarm_Confirmation': row['External Alarm Confirmation'],
                    'SOFT_AT_STATUS': row['SOFT AT STATUS'],
                    'LICENCE_Status': row['LICENCE Status'],
                    'ESN_NO': row['ESN NO'],
                    'Responsibility_for_alarm_clearance': row['Responsibility for alarm clearance'],
                    'TAC': row['TAC'],
                    'Activity_Type_SIWA': row['Activity Type (SIWA)'],
                    'Activity_Mode': row['Activity Mode (SA/NSA)'],
                    'Band_SIWA': row['Band (SIWA)'],
                    'FR_Date': fr_date,
                    'HOTO_Offered_Date_4g': hoto_offered_4g,
                    'HOTO_Accepted_Date_4g': hoto_accepted_4g,
                    'HOTO_Offered_Date_2g': hoto_offered_2g,
                    'HOTO_Accepted_Date_2g': hoto_accepted_2g,
                    'uploaded_by': request.user.username
                }
            )

        # ✅ Save uploaded document record
        Document.objects.create(
            uploaded_file=file,
            uploaded_by=request.user,
            uploaded_by_username=request.user.username
        )

    return Response(
        {'message': 'Data uploaded successfully'},
        status=status.HTTP_201_CREATED
    )



# ########### DATE WISE INTEGRATION DATA new ###########
# @api_view(['GET','POST'])
# def datewise_integration_data(request):
#     print("User: ", request.user.username)
#     date_str = request.POST.get('date')
    
#     # Parse dates
#     if date_str:
#         try:
#             print("user defined date")
#             date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
#             date1 = date_obj
#             date2 = date_obj - timedelta(days=1)
#             date3 = date_obj - timedelta(days=2)
#             print(date1, date2, date3)
#             year = date_obj.year
#         except ValueError:
#             return JsonResponse({'error': 'Invalid date format'}, status=400)
#     else:
#         latest_date = IntegrationDataVI.objects.latest('Integration_Date').Integration_Date
#         date1 = latest_date
#         date2 = latest_date - timedelta(days=1)
#         date3 = latest_date - timedelta(days=2)
#         print(date1, date2, date3)
#         year = latest_date.year

#     fixed_circles = [
#         'AP','BIH','CHN','DEL','HP','HRY','JK','JRK','KK','KOL',
#         'MAH','MP','MUM','NESA','ORI','PUN','RAJ','ROTN','UPE','UPW','WB'
#     ]

#     with connection.cursor() as cursor:
#         activity_type = [
#             '5G SECTOR ADDITION','5G RELOCATION','HT INCREMENT','FEMTO','DE-GROW',
#             'IBS','IDSC','MACRO','ODSC','5G BW UPGRADE','RRU UPGRADE','OPERATIONS',
#             'OTHERS','RECTIFICATION','RELOCATION','5G RRU SWAP','RET',
#             'TRAFFIC SHIFTING','ULS_HPSC','UPGRADE','RRU SWAP'
#         ]
        
#         activity_columns = [a.replace(' ', '_').replace('-', '_') for a in activity_type]
#         dates = [date1, date2, date3]

#         join_clauses = []
#         alias_list = []

#         for index, date in enumerate(dates):
#             alias = f"d{index+1}"
#             alias_list.append(alias)
#             date_tag = f"D{index+1}"
#             columns_def = ",".join([f'"{date_tag}_{col}" INTEGER' for col in activity_columns])

#             select_cte = f"""
#                 SELECT * FROM crosstab($$
#                     SELECT 
#                         c.cir AS "CIRCLE",
#                         a."Activity_Name",
#                         COALESCE(r.cnt, 0)::INTEGER AS cnt
#                     FROM 
#                         (SELECT unnest(ARRAY{fixed_circles}) AS cir) c
#                     CROSS JOIN
#                         (SELECT unnest(ARRAY{activity_type}) AS "Activity_Name") a
#                     LEFT JOIN (
#                         SELECT UPPER("CIRCLE") AS "CIRCLE", "Activity_Name", COUNT("id") AS cnt
#                         FROM public."IX_Tracker_VI_integrationdatavi"
#                         WHERE "Integration_Date" = '{date}'
#                         GROUP BY "CIRCLE", "Activity_Name"
#                     ) r 
#                     ON UPPER(c.cir) = r."CIRCLE" 
#                     AND a."Activity_Name" = r."Activity_Name"
#                     ORDER BY c.cir, array_position(ARRAY{activity_type}, a."Activity_Name")
#                 $$) AS ct(cir text, {columns_def})
#             """
#             join_clauses.append(f"({select_cte}) AS {alias}")

#         # Build the FULL OUTER JOIN dynamically
#         final_query = join_clauses[0]
#         for i in range(1, len(join_clauses)):
#             final_query = f"""
#                 {final_query}
#                 FULL OUTER JOIN
#                 {join_clauses[i]}
#                 USING (cir)
#             """

#         query = f"SELECT * FROM {final_query};"
#         cursor.execute(query)
#         results = cursor.fetchall()

#         results_as_strings = [[str(element) for element in row] for row in results]
#         rows_as_dict = [dict(zip([column[0] for column in cursor.description], row)) for row in results_as_strings]

#     # Serialize download data
#     objs = IntegrationDataVI.objects.filter(Integration_Date=date1)
#     serializer = IntegrationDataSerializer(objs, many=True)

#     data = {
#         "table_data": json.dumps(rows_as_dict),
#         "latest_dates": [date1, date2, date3],
#         "download_data": serializer.data
#     }
#     print('integration data in datewise',data)
#     return Response(data)

@api_view(['GET', 'POST'])
def datewise_integration_data(request):
    print("User:", request.user.username)
    date_str = request.POST.get('date')

    # ----------------------------
    # DATE HANDLING
    # ----------------------------
    if date_str:
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            date1 = date_obj
            date2 = date_obj - timedelta(days=1)
            date3 = date_obj - timedelta(days=2)
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)
    else:
        latest_obj = (
                IntegrationDataVI.objects
                .filter(Integration_Date__isnull=False)
                .order_by('-Integration_Date')
                .first()
            )
        latest_date = latest_obj.Integration_Date
        # latest_date = IntegrationDataVI.objects.latest('Integration_Date').Integration_Date
        date1 = latest_date
        date2 = latest_date - timedelta(days=1)
        date3 = latest_date - timedelta(days=2)

    dates = [date1, date2, date3]

    fixed_circles = [
        'AP','BIH','CHN','DEL','HP','HRY','JK','JRK','KK','KOL',
        'MAH','MP','MUM','NESA','ORI','PUN','RAJ','ROTN','UPE','UPW','WB'
    ]

    activity_type = [
        '5G SECTOR ADDITION','5G RELOCATION','HT INCREMENT','FEMTO','DE-GROW',
        'IBS','IDSC','MACRO','ODSC','5G BW UPGRADE','RRU UPGRADE','OPERATIONS',
        'OTHERS','RECTIFICATION','RELOCATION','5G RRU SWAP','RET',
        'TRAFFIC SHIFTING','ULS_HPSC','UPGRADE','RRU SWAP',
    ]

    activity_columns = [a.replace(' ', '_').replace('-', '_') for a in activity_type]

    join_clauses = []

    # ----------------------------
    # MAIN QUERY BUILD
    # ----------------------------
    with connection.cursor() as cursor:

        for index, date in enumerate(dates):
            alias = f"d{index + 1}"
            date_tag = f"D{index + 1}"

            # ---- Activity columns
            columns_def = ",".join(
                [f'"{date_tag}_{col}" INTEGER' for col in activity_columns]
            )

            # ---- Activity Crosstab
            select_cte = f"""
                SELECT * FROM crosstab($$
                    SELECT 
                        c.cir AS "CIRCLE",
                        a."Activity_Name",
                        COALESCE(r.cnt, 0)::INTEGER
                    FROM 
                        (SELECT unnest(ARRAY{fixed_circles}) AS cir) c
                    CROSS JOIN
                        (SELECT unnest(ARRAY{activity_type}) AS "Activity_Name") a
                    LEFT JOIN (
                        SELECT 
                            UPPER("CIRCLE") AS "CIRCLE",
                            "Activity_Name",
                            COUNT("id") AS cnt
                        FROM public."IX_Tracker_VI_integrationdatavi"
                        WHERE "Integration_Date" = '{date}'
                        GROUP BY UPPER("CIRCLE"), "Activity_Name"
                    ) r 
                    ON UPPER(c.cir) = r."CIRCLE"
                    AND a."Activity_Name" = r."Activity_Name"
                    ORDER BY c.cir, array_position(ARRAY{activity_type}, a."Activity_Name")
                $$) AS ct(cir text, {columns_def})
            """

            # ---- Milestone Counts
            extra_cte = f"""
                SELECT
                    UPPER("CIRCLE") AS cir,
                    COUNT("FR_Date") FILTER (WHERE "FR_Date" IS NOT NULL) AS "{date_tag}_FR_Date",
                    COUNT("HOTO_Offered_Date_4g") FILTER (WHERE "HOTO_Offered_Date_4g" IS NOT NULL) AS "{date_tag}_HOTO_Offered_4g",
                    COUNT("HOTO_Accepted_Date_4g") FILTER (WHERE "HOTO_Accepted_Date_4g" IS NOT NULL) AS "{date_tag}_HOTO_Accepted_4g",
                    COUNT("HOTO_Offered_Date_2g") FILTER (WHERE "HOTO_Offered_Date_2g" IS NOT NULL) AS "{date_tag}_HOTO_Offered_2g",
                    COUNT("HOTO_Accepted_Date_2g") FILTER (WHERE "HOTO_Accepted_Date_2g" IS NOT NULL) AS "{date_tag}_HOTO_Accepted_2g"
                FROM public."IX_Tracker_VI_integrationdatavi"
                WHERE "Integration_Date" = '{date}'
                GROUP BY UPPER("CIRCLE")
            """

            join_clauses.append(f"""
                (
                    {select_cte}
                    LEFT JOIN ({extra_cte}) ex_{alias}
                    USING (cir)
                ) AS {alias}
            """)

        # ---- FULL OUTER JOIN ALL DAYS
        final_query = join_clauses[0]
        for i in range(1, len(join_clauses)):
            final_query = f"""
                {final_query}
                FULL OUTER JOIN
                {join_clauses[i]}
                USING (cir)
            """

        query = f"SELECT * FROM {final_query};"
        cursor.execute(query)

        results = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

        # 🔥 NONE → 0 FIX (IMPORTANT)
        rows_as_dict = [
            {col: (0 if val is None else val) for col, val in zip(columns, row)}
            for row in results
        ]

    # ----------------------------
    # DOWNLOAD DATA
    # ----------------------------
    objs = IntegrationDataVI.objects.filter(Integration_Date=date1)
    serializer = IntegrationDataSerializer(objs, many=True)

    return Response({
        "table_data": json.dumps(rows_as_dict),
        "latest_dates": [date1, date2, date3],
        "download_data": serializer.data
    })
################################


# @api_view(["POST"])
# def hyperlink_datewise_integration_data(request):
#     print("User: ", request.user.username)
#     date=request.POST.get("date")
#     circle=request.POST.get("circle")
#     activity_name=request.POST.get("Activity_Name")
#      # Activity list jinka combined data chahiye
#     special_activities = [
#         "FR_Date",
#         "HOTO_Offered_2g",
#         "HOTO_Accepted_2g",
#         "HOTO_Offered_4g",
#         "HOTO_Accepted_4g"
#     ]

#     if activity_name in special_activities:
#         # Multiple activities ka data
#         objs = IntegrationDataVI.objects.filter(
#             Integration_Date=date,
#             CIRCLE=circle,
#             Activity_Name__in=special_activities
#         )
#     else:
#         # Normal single activity case
#         objs = IntegrationDataVI.objects.filter(
#             Integration_Date=date,
#             CIRCLE=circle,
#             Activity_Name=activity_name
#         )

#      # Filter IntegrationData objects based on provided parameters
#     # objs = IntegrationDataVI.objects.filter(Integration_Date=date, CIRCLE=circle, Activity_Name=activity_name)
    
#     # Serialize queryset using serializer
#     serializer = IntegrationDataSerializer(objs, many=True)
    
#     # Return serialized data in Response
#     return Response({"data":serializer.data})
    
@api_view(["POST"])
def hyperlink_datewise_integration_data(request):
    print("User: ", request.user.username)

    date = request.POST.get("date")
    circle = request.POST.get("circle")
    activity_name = request.POST.get("Activity_Name")

    # Mapping activity name to field that must NOT be null
    activity_field_map = {
        "FR_Date": "FR_Date",
        "HOTO_Offered_2g": "HOTO_Offered_Date_2g",
        "HOTO_Accepted_2g": "HOTO_Accepted_Date_2g",
        "HOTO_Offered_4g": "HOTO_Offered_Date_4g",
        "HOTO_Accepted_4g": "HOTO_Accepted_Date_4g",
    }

    # Base filter
    filters = {
        "Integration_Date": date,
        "CIRCLE": circle,
        "Activity_Name":activity_name
    }

    # If activity exists in mapping, add NOT NULL condition
    if activity_name in activity_field_map:
        filters[f"{activity_field_map[activity_name]}__isnull"] = False

    # Final queryset
    objs = IntegrationDataVI.objects.filter(**filters).order_by('Integration_Date')

    serializer = IntegrationDataSerializer(objs, many=True)
    return Response({"data": serializer.data})



# -------------------------- DATE RANGE WISE INTEGRATION DATA --------------------------
from datetime import datetime, timedelta

def get_dates():
    # Get the current date
    today = datetime.today()
    
    # Get the 25th of the current month
    # current_month_25 = today.replace(day=25).date()
    current_month_25 = datetime.now().date()
    
    # Calculate the first day of the current month
    first_day_of_current_month = today.replace(day=1)
    
    # Calculate the last day of the previous month
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
    
    # Get the 26th of the previous month
    previous_month_26 = last_day_of_previous_month.replace(day=26).date()
    # previous_month_26 = datetime.now().date()
    
    return previous_month_26, current_month_25

import asyncio



# ############# DATE RANGE WISE INTEGRATION DATA END New ################

# @api_view(["POST"])
# def date_range_wise_integration_data(request):
#     print("User:", request.user.username)

#     from_date = request.data.get("from_date")
#     to_date = request.data.get("to_date")

#     try:
#         if from_date:
#             from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
#         else:
#             from_date = (datetime.today() - timedelta(days=7)).date()  # default: last 7 days

#         if to_date:
#             to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
#         else:
#             to_date = datetime.today().date()  # default: today
#     except ValueError:
#         raise ValidationError("Dates must be in YYYY-MM-DD format.")

#     # ✅ Activity list (unique + sorted)
#     raw_activities = [
#         "5G SECTOR ADDITION",
#         "5G RELOCATION",
#         "HT INCREMENT",
#         "FEMTO",
#         "DE-GROW",
#         "IBS",
#         "IDSC",
#         "MACRO",
#         "ODSC",
#         "OPERATIONS",
#         "RRU UPGRADE",
#         "5G BW UPGRADE",
#         "OTHERS",
#         "5G RRU SWAP",
#         "RELOCATION",
#         "RECTIFICATION",
#         "RET",
#         "TRAFFIC SHIFTING",
#         "ULS_HPSC",
#         "UPGRADE",
#         "5G AIR SWAP",
#         "RRU SWAP",
#     ]
#     activity_type = sorted(set(raw_activities))

#     # ✅ Create SQL-safe column aliases
#     def sql_safe(name):
#         return f"D1_{name.upper().replace(' ', '_').replace('-', '_')}"

#     activity_columns = [f'"{sql_safe(a)}" INTEGER' for a in activity_type]

#     print("activity columns →", activity_columns)

#     # ✅ Crosstab fetcher
#     def fetch_crosstab_data(activity_type: list, activity_columns: list) -> tuple:
#         with connection.cursor() as cursor:
#             # Escape activity array for SQL
#             activity_array = ",".join([f"'{a}'" for a in activity_type])
#             activity_columns_def = ",\n".join(activity_columns)

#             query = f"""
#                 SELECT * FROM crosstab($$
#                     SELECT 
#                         UPPER(c."CIRCLE") AS "CIRCLE",
#                         a."Activity_Name",
#                         COALESCE(r.cnt, 0)::INTEGER AS cnt
#                     FROM 
#                         (SELECT DISTINCT UPPER("CIRCLE") AS "CIRCLE" 
#                          FROM public."IX_Tracker_VI_integrationdatavi") c
#                     CROSS JOIN
#                         (SELECT unnest(ARRAY[{activity_array}]) AS "Activity_Name") a
#                     LEFT JOIN (
#                         SELECT UPPER("CIRCLE") AS "CIRCLE", 
#                                UPPER("Activity_Name") AS "Activity_Name", 
#                                COUNT(id) AS cnt
#                         FROM public."IX_Tracker_VI_integrationdatavi"
#                         WHERE "Integration_Date" BETWEEN '{from_date}' AND '{to_date}'
#                         GROUP BY "CIRCLE", "Activity_Name"
#                     ) r ON c."CIRCLE" = r."CIRCLE" AND a."Activity_Name" = r."Activity_Name"
#                     ORDER BY 1, array_position(ARRAY[{activity_array}], a."Activity_Name")
#                 $$) AS ct (
#                     cir TEXT,
#                     {activity_columns_def}
#                 )
#             """
#             cursor.execute(query)
#             results = cursor.fetchall()
#             columns = [col[0] for col in cursor.description]
#             return results, columns

#     # ✅ Async wrapper
#     async def get_data_async():
#         results, columns = await sync_to_async(fetch_crosstab_data)(activity_type, activity_columns)
#         rows_as_dict = [dict(zip(columns, row)) for row in results]
#         jsonResult = json.dumps(rows_as_dict)

#         objs = await sync_to_async(
#             lambda: list(IntegrationDataVI.objects.filter(Integration_Date__range=(from_date, to_date)))
#         )()

#         serializer = IntegrationDataSerializer(objs, many=True)

#         return {
#             "table_data": jsonResult,
#             "date_range": [from_date.isoformat(), to_date.isoformat()],
#             "download_data": serializer.data,
#         }

#     data = asyncio.run(get_data_async())
#     return Response(data)

@api_view(["POST"])
def date_range_wise_integration_data(request):
    print("User:", request.user.username)
    dashboard_type = request.data.get("dashboard_type")  # future use
    from_date = request.data.get("from_date")
    to_date = request.data.get("to_date")
 
    # ----------------------------
    # DATE PARSING
    # ----------------------------
    try:
        if from_date:
            from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
        else:
            from_date = (datetime.today() - timedelta(days=7)).date()
 
        if to_date:
            to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
        else:
            to_date = datetime.today().date()
    except ValueError:
        raise ValidationError("Dates must be in YYYY-MM-DD format.")
 
    # ----------------------------
    # ACTIVITY LIST
    # ----------------------------
    raw_activities = [
        "5G SECTOR ADDITION", "5G RELOCATION", "HT INCREMENT", "FEMTO",
        "DE-GROW", "IBS", "IDSC", "MACRO", "ODSC", "OPERATIONS",
        "RRU UPGRADE", "5G BW UPGRADE", "OTHERS", "5G RRU SWAP",
        "RELOCATION", "RECTIFICATION", "RET", "TRAFFIC SHIFTING",
        "ULS_HPSC", "UPGRADE", "5G AIR SWAP", "RRU SWAP",
    ]
    activity_type = sorted(set(raw_activities))
 
    # ----------------------------
    # SQL SAFE COLUMN NAMES
    # ----------------------------
    def sql_safe(name):
        return f"D1_{name.upper().replace(' ', '_').replace('-', '_')}"
 
    activity_columns = [f'"{sql_safe(a)}" INTEGER' for a in activity_type]
 
    # ----------------------------
    # FETCH DATA
    # ----------------------------
    def fetch_data():
        with connection.cursor() as cursor:
            activity_array = ",".join([f"'{a}'" for a in activity_type])
            activity_columns_def = ",\n".join(activity_columns)
 
            query = f"""
                SELECT
                    ct.*,
                    ex."FR_Date_Count",
                    ex."HOTO_Offered_4g_Count",
                    ex."HOTO_Accepted_4g_Count",
                    ex."HOTO_Offered_2g_Count",
                    ex."HOTO_Accepted_2g_Count"
                FROM crosstab($$
                    SELECT
                        UPPER(c."CIRCLE") AS "CIRCLE",
                        a."Activity_Name",
                        COALESCE(r.cnt, 0)::INTEGER
                    FROM
                        (SELECT DISTINCT UPPER("CIRCLE") AS "CIRCLE"
                         FROM public."IX_Tracker_VI_integrationdatavi") c
                    CROSS JOIN
                        (SELECT unnest(ARRAY[{activity_array}]) AS "Activity_Name") a
                    LEFT JOIN (
                        SELECT
                            UPPER("CIRCLE") AS "CIRCLE",
                            UPPER("Activity_Name") AS "Activity_Name",
                            COUNT(id) AS cnt
                        FROM public."IX_Tracker_VI_integrationdatavi"
                        WHERE "{dashboard_type}"
                              BETWEEN '{from_date}' AND '{to_date}'
                        GROUP BY "CIRCLE", "Activity_Name"
                    ) r
                    ON c."CIRCLE" = r."CIRCLE"
                    AND a."Activity_Name" = r."Activity_Name"
                    ORDER BY 1,
                        array_position(ARRAY[{activity_array}], a."Activity_Name")
                $$) AS ct (
                    cir TEXT,
                    {activity_columns_def}
                )
                LEFT JOIN (
                    SELECT
                        UPPER("CIRCLE") AS cir,
                        COUNT("FR_Date") FILTER (WHERE "FR_Date" IS NOT NULL) AS "FR_Date_Count",
                        COUNT("HOTO_Offered_Date_4g") FILTER (WHERE "HOTO_Offered_Date_4g" IS NOT NULL) AS "HOTO_Offered_4g_Count",
                        COUNT("HOTO_Accepted_Date_4g") FILTER (WHERE "HOTO_Accepted_Date_4g" IS NOT NULL) AS "HOTO_Accepted_4g_Count",
                        COUNT("HOTO_Offered_Date_2g") FILTER (WHERE "HOTO_Offered_Date_2g" IS NOT NULL) AS "HOTO_Offered_2g_Count",
                        COUNT("HOTO_Accepted_Date_2g") FILTER (WHERE "HOTO_Accepted_Date_2g" IS NOT NULL) AS "HOTO_Accepted_2g_Count"
                    FROM public."IX_Tracker_VI_integrationdatavi"
                    WHERE "{dashboard_type}"
                          BETWEEN '{from_date}' AND '{to_date}'
                    GROUP BY UPPER("CIRCLE")
                ) ex
                USING (cir);
            """
 
            cursor.execute(query)
            results = cursor.fetchall()
            columns = [col[0] for col in cursor.description]
            return results, columns
 
    # ----------------------------
    # ASYNC WRAPPER
    # ----------------------------
    async def get_data_async():
        results, columns = await sync_to_async(fetch_data)()
 
        # 🔥 NONE → 0 FIX (IMPORTANT)
        rows_as_dict = [
            {col: (0 if val is None else val) for col, val in zip(columns, row)}
            for row in results
        ]
 
        objs = await sync_to_async(
            lambda: list(
                IntegrationDataVI.objects.filter(
                    Integration_Date__range=(from_date, to_date)
                )
            )
        )()
 
        serializer = IntegrationDataSerializer(objs, many=True)
 
        return {
            "table_data": json.dumps(rows_as_dict),
            "date_range": [from_date.isoformat(), to_date.isoformat()],
            "download_data": serializer.data,
        }
 
    data = asyncio.run(get_data_async())
    return Response(data)

################################


# @api_view(["POST"])
# def hyperlink_date_range_integration_data(request):
#     print("User: ", request.user.username)
#     from_date=request.POST.get("from_date")
#     to_date=request.POST.get("to_date")

#     circle=request.POST.get("circle")
#     activity_name=request.POST.get("Activity_Name")
#      # Filter IntegrationData objects based on provided parameters
#     objs = IntegrationDataVI.objects.filter(Integration_Date__range=[from_date, to_date], CIRCLE=circle, Activity_Name=activity_name)
    
#     # Serialize queryset using serializer
#     serializer = IntegrationDataSerializer(objs, many=True)
    
#     # Return serialized data in Response
#     return Response({"data":serializer.data})

@api_view(["POST"])
def hyperlink_date_range_integration_data(request):
    print("User: ", request.user.username)

    from_date = request.POST.get("from_date")
    to_date = request.POST.get("to_date")
    circle = request.POST.get("circle")
    activity_name = request.POST.get("Activity_Name")

    # Mapping activity name to field that must NOT be null
    activity_field_map = {
        "FR_Date": "FR_Date",
        "HOTO_Offered_2g": "HOTO_Offered_Date_2g",
        "HOTO_Accepted_2g": "HOTO_Accepted_Date_2g",
        "HOTO_Offered_4g": "HOTO_Offered_Date_4g",
        "HOTO_Accepted_4g": "HOTO_Accepted_Date_4g",
    }

    # Base filters
    filters = {
        "Integration_Date__range": [from_date, to_date],
        "CIRCLE": circle
    }

    # Add NOT NULL condition if activity is in mapping
    if activity_name in activity_field_map:
        filters[f"{activity_field_map[activity_name]}__isnull"] = False
    else:
        # Normal activity case
        filters["Activity_Name"] = activity_name

    # Final queryset
    objs = IntegrationDataVI.objects.filter(**filters).order_by('Integration_Date')

    serializer = IntegrationDataSerializer(objs, many=True)
    return Response({"data": serializer.data})

# -----------------------------*********************************************************-----------------------------



from dateutil.relativedelta import relativedelta
# def calculate_previous_months(given_month,given_year):
#     print("Month1:", given_month)
#     print("Year1:", given_year)
#     given_date = datetime(given_year, given_month, 1)

#     # Calculate previous month
#     month2_date = given_date - relativedelta(months=1)
#     month3_date = given_date - relativedelta(months=2)
#     month4_date = given_date - relativedelta(months=3)
#     month5_date = given_date - relativedelta(months=4)
#     month6_date = given_date - relativedelta(months=5)

#     # Extract the month and year from the previous month date
#     month2 =month2_date.month
#     year2 = month2_date.year

#     month3 =month3_date.month
#     year3 = month3_date.year

#     month4 =month4_date.month
#     year4 = month4_date.year

#     month5 =month5_date.month
#     year5 = month5_date.year

#     month6 =month6_date.month
#     year6 = month6_date.year

#     print("Previous Month2:", month2)
#     print("Previous Year2:", year2)

#     print("Previous Month3:", month3)
#     print("Previous Year3:", year3)

#     return {"Month2":month2,"Month3":month3,"Year2":year2,"Year3":year3,"Month4":month4,"Year4":year4,"Month5":month5,"Year5":year5,"Month6":month6,"Year6":year6}

def calculate_previous_months(given_month, given_year):
    given_date = datetime(given_year, given_month, 1)
    result = {}

    for i in range(2, 7):
        prev = given_date - relativedelta(months=i - 1)
        result[f"Month{i}"] = prev.month
        result[f"Year{i}"] = prev.year

    return result


######### Monthwise Integration Data New API ########
# @api_view(['GET', 'POST'])
# def monthwise_integration_data(request):
#     try:
#         print("User: ", request.user.username)

#         # ✅ Use request.data instead of request.POST
#         month = request.data.get("month")
#         year = request.data.get("year")

#         if month and year:
#             print("User-defined month/year")
#             base_month = int(month)
#             base_year = int(year)
#         else:
#             print("Using default latest month/year")
#             latest_date = IntegrationDataVI.objects.latest('Integration_Date').Integration_Date
#             base_month = latest_date.month
#             base_year = latest_date.year

#         date_dict = calculate_previous_months(base_month, base_year)

#         # ✅ Prepare month-year list
#         month_year_list = [(base_month, base_year, 1)]
#         for i in range(2, 7):
#             month_year_list.append((date_dict[f"Month{i}"], date_dict[f"Year{i}"], i))

#         # ✅ Activity list (correct indentation)
#         activity_list = [
#             '5G BW UPGRADE', '5G RELOCATION', '5G RRU SWAP', '5G SECTOR ADDITION',
#             'DE-GROW', 'FEMTO', 'HT INCREMENT', 'IBS', 'IDSC', 'MACRO', 'ODSC',
#             'OPERATIONS', 'OTHERS', 'RECTIFICATION', 'RELOCATION', 'RET',
#             'RRU UPGRADE', 'TRAFFIC SHIFTING', 'ULS_HPSC', 'UPGRADE', 'RRU SWAP'
#         ]

#         def build_crosstab(month, year, index):
#             label = f"M{index}"
#             safe_activities = [a.upper() for a in activity_list]
#             columns = ",".join([f'"{label}_{a.replace(" ", "_").replace("-", "_")}" INTEGER' for a in safe_activities])
#             array_activities = ",".join([f"'{a}'" for a in safe_activities])

#             return f"""
#                 (
#                     SELECT * FROM crosstab($$
#                         SELECT 
#                             UPPER("CIRCLE") AS "CIRCLE",
#                             UPPER("Activity_Name") AS "Activity_Name",
#                             COALESCE(cnt, 0)::INTEGER AS cnt
#                         FROM 
#                             (SELECT DISTINCT UPPER("CIRCLE") AS "CIRCLE" FROM public."IX_Tracker_VI_integrationdatavi") AS "C"
#                         CROSS JOIN
#                             (SELECT unnest(ARRAY[{array_activities}]) AS "Activity_Name") AS a
#                         LEFT JOIN (
#                             SELECT UPPER("CIRCLE") AS "CIRCLE", UPPER("Activity_Name") AS "Activity_Name", COUNT("id") AS cnt
#                             FROM public."IX_Tracker_VI_integrationdatavi"
#                             WHERE EXTRACT(MONTH FROM "Integration_Date") = {month}
#                               AND EXTRACT(YEAR FROM "Integration_Date") = {year}
#                             GROUP BY "CIRCLE", "Activity_Name"
#                         ) AS r USING ("CIRCLE", "Activity_Name")
#                         ORDER BY 1, 2
#                     $$, $$ SELECT unnest(ARRAY[{array_activities}]) $$) 
#                     AS ct(cir TEXT, {columns})
#                 ) AS d{index}
#             """

#         # ✅ Build final SQL
#         join_parts = []
#         for i, (month, year, index) in enumerate(month_year_list):
#             part = build_crosstab(month, year, index)
#             if i == 0:
#                 join_parts.append(part)
#             else:
#                 join_parts.append(f"FULL OUTER JOIN {part} USING (cir)")

#         full_sql = f"SELECT * FROM {' '.join(join_parts)};"

#         # ✅ Execute query
#         with connection.cursor() as cursor:
#             cursor.execute(full_sql)
#             results = cursor.fetchall()
#             columns = [col[0] for col in cursor.description]

#         results_as_dicts = [dict(zip(columns, row)) for row in results]
#         jsonResult = json.dumps(results_as_dicts)

#         # ✅ Latest month/year for filtering
#         latest_month, latest_year, _ = month_year_list[0]

#         objs = IntegrationDataVI.objects.filter(
#             Integration_Date__year=latest_year,
#             Integration_Date__month=latest_month
#         )
#         serializer = IntegrationDataSerializer(objs, many=True)

#         data = {
#             "table_data": jsonResult,
#             "latest_months": [month for month, _, _ in month_year_list],
#             "latest_year": [year for _, year, _ in month_year_list],
#             "download_data": serializer.data
#         }
#         return Response(data)

#     except Exception as e:
#         # ✅ Always return a Response even on error
#         return Response({"error": str(e)}, status=400)


@api_view(['GET', 'POST'])
def monthwise_integration_data(request):
    try:
        print("User:", request.user.username)

        month = request.data.get("month")
        year = request.data.get("year")

        # ---------------------------------------
        # BASE MONTH / YEAR
        # ---------------------------------------
        if month and year:
            base_month = int(month)
            base_year = int(year)
        else:
            latest_obj = (
                IntegrationDataVI.objects
                .filter(Integration_Date__isnull=False)
                .order_by('-Integration_Date')
                .first()
            )

            if not latest_obj or not latest_obj.Integration_Date:
                return Response(
                    {"error": "No valid Integration_Date found in database"},
                    status=400
                )

            latest_date = latest_obj.Integration_Date

            # If DateTimeField → convert to date
            if hasattr(latest_date, "date"):
                latest_date = latest_date.date()

            base_month = latest_date.month
            base_year = latest_date.year

        date_dict = calculate_previous_months(base_month, base_year)

        # ---------------------------------------
        # MONTH LIST (M1–M6)
        # ---------------------------------------
        month_year_list = [(base_month, base_year, 1)]
        for i in range(2, 7):
            month_year_list.append((date_dict[f"Month{i}"], date_dict[f"Year{i}"], i))

        # ---------------------------------------
        # ACTIVITY LIST
        # ---------------------------------------
        activity_list = [
            '5G BW UPGRADE', '5G RELOCATION', '5G RRU SWAP', '5G SECTOR ADDITION',
            'DE-GROW', 'FEMTO', 'HT INCREMENT', 'IBS', 'IDSC', 'MACRO', 'ODSC',
            'OPERATIONS', 'OTHERS', 'RECTIFICATION', 'RELOCATION', 'RET',
            'RRU UPGRADE', 'TRAFFIC SHIFTING', 'ULS_HPSC', 'UPGRADE', 'RRU SWAP'
        ]

        # ---------------------------------------
        # BUILD MONTH BLOCK
        # ---------------------------------------
        def build_month_block(month, year, index):
            label = f"M{index}"
            activities = [a.upper() for a in activity_list]

            columns = ",".join(
                [f'"{label}_{a.replace(" ", "_").replace("-", "_")}" INTEGER' for a in activities]
            )
            array_activities = ",".join([f"'{a}'" for a in activities])

            return f"""
                (
                    SELECT
                        ct.*,
                        ex."{label}_FR_Date",
                        ex."{label}_HOTO_Offered_4g",
                        ex."{label}_HOTO_Accepted_4g",
                        ex."{label}_HOTO_Offered_2g",
                        ex."{label}_HOTO_Accepted_2g"
                    FROM crosstab($$
                        SELECT
                            UPPER("CIRCLE") AS "CIRCLE",
                            UPPER("Activity_Name") AS "Activity_Name",
                            COUNT(id)::INTEGER
                        FROM public."IX_Tracker_VI_integrationdatavi"
                        WHERE EXTRACT(MONTH FROM "Integration_Date") = {month}
                          AND EXTRACT(YEAR FROM "Integration_Date") = {year}
                        GROUP BY "CIRCLE", "Activity_Name"
                        ORDER BY 1, 2
                    $$, $$ SELECT unnest(ARRAY[{array_activities}]) $$)
                    AS ct(cir TEXT, {columns})
                    LEFT JOIN (
                        SELECT
                            UPPER("CIRCLE") AS cir,
                            COUNT("FR_Date") FILTER (WHERE "FR_Date" IS NOT NULL) AS "{label}_FR_Date",
                            COUNT("HOTO_Offered_Date_4g") FILTER (WHERE "HOTO_Offered_Date_4g" IS NOT NULL) AS "{label}_HOTO_Offered_4g",
                            COUNT("HOTO_Accepted_Date_4g") FILTER (WHERE "HOTO_Accepted_Date_4g" IS NOT NULL) AS "{label}_HOTO_Accepted_4g",
                            COUNT("HOTO_Offered_Date_2g") FILTER (WHERE "HOTO_Offered_Date_2g" IS NOT NULL) AS "{label}_HOTO_Offered_2g",
                            COUNT("HOTO_Accepted_Date_2g") FILTER (WHERE "HOTO_Accepted_Date_2g" IS NOT NULL) AS "{label}_HOTO_Accepted_2g"
                        FROM public."IX_Tracker_VI_integrationdatavi"
                        WHERE EXTRACT(MONTH FROM "Integration_Date") = {month}
                          AND EXTRACT(YEAR FROM "Integration_Date") = {year}
                        GROUP BY UPPER("CIRCLE")
                    ) ex USING (cir)
                ) AS d{index}
            """

        # ---------------------------------------
        # BUILD FINAL SQL
        # ---------------------------------------
        join_parts = []
        for i, (m, y, idx) in enumerate(month_year_list):
            block = build_month_block(m, y, idx)
            join_parts.append(block if i == 0 else f"FULL OUTER JOIN {block} USING (cir)")

        final_sql = f"SELECT * FROM {' '.join(join_parts)};"

        # ---------------------------------------
        # EXECUTE QUERY
        # ---------------------------------------
        with connection.cursor() as cursor:
            cursor.execute(final_sql)
            results = cursor.fetchall()
            columns = [col[0] for col in cursor.description]

        # ---------------------------------------
        # 🔥 NULL → 0 FIX HERE
        # ---------------------------------------
        rows_as_dict = [
            {col: (0 if val is None else val) for col, val in zip(columns, row)}
            for row in results
        ]

        # ---------------------------------------
        # DOWNLOAD DATA (LATEST MONTH)
        # ---------------------------------------
        latest_month, latest_year, _ = month_year_list[0]

        objs = IntegrationDataVI.objects.filter(
            Integration_Date__month=latest_month,
            Integration_Date__year=latest_year
        )
        serializer = IntegrationDataSerializer(objs, many=True)

        return Response({
            "table_data": json.dumps(rows_as_dict),
            "latest_months": [m for m, _, _ in month_year_list],
            "latest_years": [y for _, y, _ in month_year_list],
            "download_data": serializer.data
        })

    except Exception as e:
        return Response({"error": str(e)}, status=400)
####################################################

# @api_view(["POST"])
# def hyperlink_monthwise_integration_data(request):
#     print("User: ", request.user.username)
#     month=request.POST.get("month")
#     year=request.POST.get("year")
#     circle=request.POST.get("circle")
#     activity_name=request.POST.get("Activity_Name")
#      # Filter IntegrationData objects based on provided parameters
#     objs = IntegrationDataVI.objects.filter(Integration_Date__year=year,Integration_Date__month=month, CIRCLE=circle, Activity_Name=activity_name)
    
#     # Serialize queryset using serializer
#     serializer = IntegrationDataSerializer(objs, many=True)
    
#     # Return serialized data in Response
#     return Response(serializer.data)

@api_view(["POST"])
def hyperlink_monthwise_integration_data(request):
    print("User: ", request.user.username)

    month = request.POST.get("month")
    year = request.POST.get("year")
    circle = request.POST.get("circle")
    activity_name = request.POST.get("Activity_Name")

    # Mapping activity name to field that must NOT be null
    activity_field_map = {
        "FR_Date": "FR_Date",
        "HOTO_Offered_2g": "HOTO_Offered_Date_2g",
        "HOTO_Accepted_2g": "HOTO_Accepted_Date_2g",
        "HOTO_Offered_4g": "HOTO_Offered_Date_4g",
        "HOTO_Accepted_4g": "HOTO_Accepted_Date_4g",
    }

    # Base filters
    filters = {
        "Integration_Date__year": year,
        "Integration_Date__month": month,
        "CIRCLE": circle
    }

    # Add NOT NULL condition for special activities
    if activity_name in activity_field_map:
        filters[f"{activity_field_map[activity_name]}__isnull"] = False
    else:
        # Normal activity case
        filters["Activity_Name"] = activity_name

    # Final queryset
    objs = IntegrationDataVI.objects.filter(**filters).order_by('Integration_Date')

    serializer = IntegrationDataSerializer(objs, many=True)
    return Response({"data": serializer.data})


@api_view(['POST'])
def monthly_oemwise_integration_data(request):
    print("User: ", request.user.username) 
    month=request.POST.get("month")
    year=request.POST.get("year")
    if month and year:
        print("user defined......")
        month=int(month)
        year=int(year)

    else:
        print("default months......")

        latest_obj = (
            IntegrationDataVI.objects
            .filter(Integration_Date__isnull=False)
            .order_by('-Integration_Date')
            .first()
        )

        if not latest_obj or not latest_obj.Integration_Date:
            return Response(
                {"error": "No valid Integration_Date found in database"},
                status=400
            )

        latest_date = latest_obj.Integration_Date

        # If DateTimeField → convert to date
        if hasattr(latest_date, "date"):
            latest_date = latest_date.date()

        month = latest_date.month
        year = latest_date.year
        

    with connection.cursor() as cursor:

        query = f""" 
        
        select * from crosstab($$
	    SELECT  "CIRCLE","OEM",COALESCE("cnt", 0)::INTEGER as cnt -- Replace 0 with appropriate default value
     FROM 
        (SELECT DISTINCT UPPER("CIRCLE") as "CIRCLE" FROM public."IX_Tracker_VI_integrationdatavi") AS "CIRCLE"
     CROSS JOIN
           (
            SELECT unnest(ARRAY['SAMSUNG', 'NOKIA', 'ERICSSON','HUAWEI','ZTE']) AS "OEM"
        ) AS activities
     LEFT JOIN
        (select "CIRCLE","OEM", count("id") as cnt from 
 						(select "id", "CIRCLE", UPPER("OEM") as "OEM" 
						 from public."IX_Tracker_VI_integrationdatavi" WHERE EXTRACT(MONTH FROM "Integration_Date") = {month} and EXTRACT(YEAR FROM "Integration_Date") = {year}) in_0
    group by "CIRCLE","OEM") as first_t
        USING ("CIRCLE", "OEM") order by 1,2 $$) as 	 
    ct(cir text, "ERICSSON" INTEGER,"HUAWEI" INTEGER,"NOKIA" INTEGER,"SAMSUNG" INTEGER, "ZTE" INTEGER)
         """
        cursor.execute(query)
        results = cursor.fetchall()
        columns = cursor.description
        # print(results)        
    results_as_strings = []
    for row in results:
        # Convert each element in the row to a string 
        row_as_string = [str(element) for element in row]
        # Append the row as a string to the list
        results_as_strings.append(row_as_string)
    
    rows_as_dict = [
    dict(zip([column[0] for column in columns], row))
    for row in results_as_strings
]
    jsonResult =  json.dumps(rows_as_dict)

    # print(jsonResult)
    data={"table_data":jsonResult,'month':month,'year':year}
    return Response(data)
        



@api_view(['POST'])
def hyperlink_monthly_oemwise_integration_data(request):
    print("User: ", request.user.username)
    month=request.POST.get("month")
    year=request.POST.get("year")
    circle=request.POST.get("circle")
    oem=request.POST.get("oem")
    with connection.cursor() as cursor:

        query = f"""
SELECT
    "CIRCLE",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = '5G BW UPGRADE')        AS "5G BW UPGRADE",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = '5G RELOCATION')       AS "5G RELOCATION",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'RRU SWAP')            AS "RRU SWAP",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = '5G RRU SWAP')         AS "5G RRU SWAP",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = '5G SECTOR ADDITION')  AS "5G SECTOR ADDITION",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'DE-GROW')             AS "DE-GROW",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'FEMTO')               AS "FEMTO",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'HT INCREMENT')        AS "HT INCREMENT",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'IBS')                 AS "IBS",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'IDSC')                AS "IDSC",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'MACRO')               AS "MACRO",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'ODSC')                AS "ODSC",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'OPERATIONS')          AS "OPERATIONS",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'OTHERS')              AS "OTHERS",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'RECTIFICATION')       AS "RECTIFICATION",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'RELOCATION')          AS "RELOCATION",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'RET')                 AS "RET",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'RRU UPGRADE')         AS "RRU UPGRADE",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'TRAFFIC SHIFTING')    AS "TRAFFIC SHIFTING",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'ULS_HPSC')            AS "ULS_HPSC",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = 'UPGRADE')             AS "UPGRADE",
    COUNT(*) FILTER (WHERE UPPER("Activity_Name") = '5G AIR SWAP')         AS "5G AIR SWAP"
FROM public."IX_Tracker_VI_integrationdatavi"
WHERE
    EXTRACT(MONTH FROM "Integration_Date") = {month}
    AND EXTRACT(YEAR FROM "Integration_Date") = {year}
    AND "OEM" = '{oem}'
    AND "CIRCLE" = '{circle}'
GROUP BY "CIRCLE";
"""
        cursor.execute(query)
        results = cursor.fetchall()
        columns = cursor.description        
    results_as_strings = []
    for row in results:
        # Convert each element in the row to a string
        row_as_string = [str(element) for element in row]
        # Append the row as a string to the list
        results_as_strings.append(row_as_string)
    
    rows_as_dict = [
    dict(zip([column[0] for column in columns], row))
    for row in results_as_strings
]
    jsonResult =  json.dumps(rows_as_dict)

    # print(jsonResult)
    data={"table_data":jsonResult}
    return Response(data)
    

@api_view(["POST"])
def hyperlink_hyperlink_monthly_oemwise_integration_data(request):
    print("User: ", request.user.username) 
    oem=request.POST.get("oem")
    month=request.POST.get("month")
    year=request.POST.get("year")
    circle=request.POST.get("circle")
    activity_name=request.POST.get("Activity_Name")
     # Filter IntegrationData objects based on provided parameters
    objs = IntegrationDataVI.objects.filter(Integration_Date__year=year,Integration_Date__month=month, CIRCLE=circle, Activity_Name=activity_name,OEM=oem)
    
    # Serialize queryset using serializer
    serializer = IntegrationDataSerializer(objs, many=True)
    
    # Return serialized data in Response
    return Response(serializer.data)




@api_view(['GET'])
def overall_record_summary(request):
    print("User: ", request.user.username)
  
    with connection.cursor() as cursor:

        query = f"""
        SELECT 
            "OEM",
            MIN("Integration_Date") AS from_integration_date,
            MAX("Integration_Date") AS to_integration_date,
            COUNT(*) AS record_count
        FROM 
            PUBLIC."IX_Tracker_VI_integrationdatavi"
        GROUP BY 
            "OEM";

        """
        cursor.execute(query)
        results = cursor.fetchall()
        columns=cursor.description
        # print(results)        
    results_as_strings = []
    for row in results:
        # Convert each element in the row to a string
        row_as_string = [str(element) for element in row]
        # Append the row as a string to the list
        results_as_strings.append(row_as_string)
        
    
    rows_as_dict = [
    dict(zip([column[0] for column in columns], row))
    for row in results_as_strings
]
    jsonResult =  json.dumps(rows_as_dict)

    # print(jsonResult)
    data={"table_data":jsonResult}
    return Response(data)

@api_view(["Delete"])
def delete_integration_record(request, pk):
    print("User: ", request.user.username)
    user=request.user.username
    print("pk: ", pk)
    nokia_spocks=['chandan.kumar@mcpsinc.com']
    zte_spocks=['aashish.s@mcpsinc.com','chandan.kumar@mcpsinc.com']
    huawei_spocks=['rahul.dahiya@mcpsinc.com','chandan.kumar@mcpsinc.com']
    samsung_spocks=['rahul.dahiya@mcpsinc.com','chandan.kumar@mcpsinc.com']
    ericsson_spocks=['aashish.s@mcpsinc.com','chandan.kumar@mcpsinc.com']
    print("here")
    try:
        obj = IntegrationDataVI.objects.get(pk=pk)
    except IntegrationDataVI.DoesNotExist:
        return Response({"status":True,'message': 'Record not found'}, status=status.HTTP_404_NOT_FOUND)
    print(obj)

    if obj.OEM == "NOKIA" and user in nokia_spocks:
        obj.delete()
        print("deleted")
        return Response({"status":True,'message': 'Record deleted successfully'}, status=status.HTTP_200_OK)
    elif obj.OEM == "ZTE" and user in zte_spocks:
        obj.delete()
        print("deleted")
        return Response({"status":True,'message': 'Record deleted successfully'}, status=status.HTTP_200_OK)
    elif obj.OEM == "HUAWEI" and user in huawei_spocks:
        obj.delete()
        print("deleted")
        return Response({"status":True,'message': 'Record deleted successfully'}, status=status.HTTP_200_OK)
    elif obj.OEM == "SAMSUNG" and user in samsung_spocks:
        obj.delete()
        print("deleted")
        return Response({"status":True,'message': 'Record deleted successfully'}, status=status.HTTP_200_OK)
    elif obj.OEM == "ERICSSON" and user in ericsson_spocks:
        obj.delete()
        print("deleted")
        return Response({"status":True,'message': 'Record deleted successfully'}, status=status.HTTP_200_OK)
    else:
        print("exception..")
        return Response({'status':False,'message': 'You are not authorized to delete this record'}, status=status.HTTP_403_FORBIDDEN)

        
@api_view(['PUT'])
def integration_table_update(request, id=None):
    user = request.user.username
    print("username: ", user)
    nokia_spocks=['chandan.kumar@mcpsinc.com','Prerna.PramodKumar@ust.com','girraj.singh@mcpsinc.in','mohit.batra@mcpsinc.com','abhishek.gupta','Abhinav.Verma@ust.com']
    zte_spocks=['chandan.kumar@mcpsinc.com','aashish.s@mcpsinc.com','mohit.batra@mcpsinc.com','abhishek.gupta','Prerna.PramodKumar@ust.com','Abhinav.Verma@ust.com']
    huawei_spocks=['chandan.kumar@mcpsinc.com','rahul.dahiya@mcpsinc.com','mohit.batra@mcpsinc.com','harish.singh@ust.com','abhishek.gupta', 'Prerna.PramodKumar@ust.com','Abhinav.Verma@ust.com']
    samsung_spocks=['chandan.kumar@mcpsinc.com','rahul.dahiya@mcpsinc.com','mohit.batra@mcpsinc.com','harish.singh@ust.com', 'abhishek.gupta', 'Prerna.PramodKumar@ust.com','Abhinav.Verma@ust.com']
    ericsson_spocks=['chandan.kumar@mcpsinc.com','aashish.s@mcpsinc.com','mohit.batra@mcpsinc.com','abhishek.gupta', 'Prerna.PramodKumar@ust.com','Abhinav.Verma@ust.com']
  
    if request.method == 'PUT':
        try:
            integration_record = IntegrationDataVI.objects.get(pk=id)
        except IntegrationDataVI.DoesNotExist:
            return Response({'message': 'Record not found'}, status=status.HTTP_404_NOT_FOUND)
    print('OEM: ',integration_record.OEM)
    if integration_record.OEM == "NOKIA" and user in nokia_spocks :
        serializer = IntegrationDataSerializer(integration_record, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"data":serializer.data,"status":True,'message': 'Record deleted successfully'}, status=status.HTTP_200_OK)
        else:
            print(serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
    elif integration_record.OEM == "ZTE" and user in zte_spocks:
            serializer = IntegrationDataSerializer(integration_record, data=request.data)
            if serializer.is_valid():
             serializer.save()
             return Response({"data":serializer.data,"status":True,'message': 'Record edited successfully'}, status=status.HTTP_200_OK)
    elif integration_record.OEM == "HUAWEI" and user in huawei_spocks:
            serializer = IntegrationDataSerializer(integration_record, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"status":True,'message': 'Record edited successfully'}, status=status.HTTP_200_OK)
    
    elif integration_record.OEM == "SAMSUNG" and user in samsung_spocks:
           serializer = IntegrationDataSerializer(integration_record, data=request.data)
           if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"status":True,'message': 'Record edited successfully'}, status=status.HTTP_200_OK)
        
    elif integration_record.OEM == "ERICSSON" and user in ericsson_spocks:
           serializer = IntegrationDataSerializer(integration_record, data=request.data)
           if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"status":True,'message': 'Record edited successfully'}, status=status.HTTP_200_OK)   
    else:
        print("exception..")
        return Response({'status':False,'message': 'You are not authorized to edit this record'}, status=status.HTTP_403_FORBIDDEN)


@api_view(["POST"])
def overall_integration_for_perticular_oem(request):
    print("User: ", request.user.username)
    oem=request.POST.get("oem")
    print(oem)
    objs = IntegrationDataVI.objects.filter(OEM=oem)
    serializer = IntegrationDataSerializer(objs, many=True)
    return Response({"table_data":serializer.data})




 
 
@api_view(["POST"])
def upload_log(request):

    uploaded_file = request.FILES.get("log_file")

    if not uploaded_file:
        return Response(
            {"error": "Please upload a file."},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Create logs directory inside MEDIA_ROOT
    logs_dir = os.path.join(settings.MEDIA_ROOT, "logs")
    os.makedirs(logs_dir, exist_ok=True)

    # Full file path
    temp_path = os.path.join(logs_dir, uploaded_file.name)

    # Save uploaded file
    with open(temp_path, "wb+") as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)

    try:
        report_path = process_log(temp_path)

        # Remove uploaded log file after processing
        if os.path.exists(temp_path):
            os.remove(temp_path)

        # Convert absolute path to relative path
        relative_path = os.path.relpath(report_path, settings.MEDIA_ROOT)

        return Response({
            "status": True,
            "message": "Report generated successfully",
            "download_url": request.build_absolute_uri(
                settings.MEDIA_URL + relative_path.replace("\\", "/")
            )
        })

    except Exception as e:
        return Response(
            {
                "status": False,
                "error": str(e)
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

import os
from django.conf import settings
from rest_framework.decorators import api_view
from rest_framework.response import Response
from openpyxl import load_workbook
from .models import IntegrationDataVI

VI_FIELDS = [
    "unique_key", "OEM", "Integration_Date", "CIRCLE", "Activity_Name", "Site_ID",
    "MO_NAME", "LNBTS_ID", "Technology_SIWA", "OSS_Details", "Cell_ID", "CELL_COUNT",
    "BSC_NAME", "BCF", "TRX_Count", "PRE_ALARM", "GPS_IP_CLK", "RET", "POST_VSWR",
    "POST_Alarms", "Activity_Mode", "Activity_Type_SIWA", "Band_SIWA", "CELL_STATUS",
    "CTR_STATUS", "Integration_Remark", "T2T4R", "BBU_TYPE", "BB_CARD", "RRU_Type",
    "Media_Status", "Mplane_IP", "SCF_PREPARED_BY", "SITE_INTEGRATE_BY", "Site_Status",
    "External_Alarm_Confirmation", "SOFT_AT_STATUS", "LICENCE_Status", "ESN_NO",
    "Responsibility_for_alarm_clearance", "TAC", "PCI_TDD_20", "PCI_TDD_10_20",
    "PCI_FDD_2100", "PCI_FDD_1800", "PCI_L900", "PCI_5G", "RSI_TDD_20", "RSI_TDD_10_20",
    "RSI_FDD_2100", "RSI_FDD_1800", "RSI_L900", "RSI_5G", "GPL", "Pre_Post_Check",
    "CRQ", "Customer_Approval", "FR_Date", "HOTO_Offered_Date_4g", "HOTO_Accepted_Date_4g",
    "HOTO_Offered_Date_2g","HOTO_Accepted_Date_2g",
    None, None, None, None, None,  # Overall HOTO Status, Pending Bucket, Pending Remarks, Responsibility, TAT
]

# Header occupies row 1 (version/category) + row 2 (column names) -> data starts row 3
HEADER_ROWS = 2
DATA_START_ROW = HEADER_ROWS + 1  # = 3
@api_view(['GET'])
def get_vi_temp_link(request):
    template_name = 'VI_Tracker_HOTO_Template_V1.1.1.xlsm'
    sub_dir = 'IntegrationTrackerVI'
    template_path = os.path.join(settings.MEDIA_ROOT, sub_dir, template_name)

    if not os.path.exists(template_path):
        return Response({'error': 'Excel template not found'}, status=404)

    wb = load_workbook(template_path, keep_vba=True)
    ws = wb.active

    # --- Clear any leftover/sample data rows below header before writing fresh data ---
    if ws.max_row >= DATA_START_ROW:
        for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, max_col=len(VI_FIELDS)):
            for cell in row:
                cell.value = None

    # queryset = IntegrationDataVI.objects.all().order_by('id')
    queryset = IntegrationDataVI.objects.filter(
            Activity_Name__in=[
                "MACRO",
                "UPGRADE",
                "ULS_HPSC",
                "RELOCATION"
            ]
        ).order_by('id')

    row_idx = DATA_START_ROW
    for obj in queryset:
        for col_idx, field_name in enumerate(VI_FIELDS, start=1):
            if field_name is None:
                value = ''
            else:
                value = getattr(obj, field_name, '')
                if value is None:
                    value = ''
                elif hasattr(value, 'isoformat'):
                    value = value.isoformat()
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1

    output_sub_dir = os.path.join(sub_dir, 'generated')
    output_dir = os.path.join(settings.MEDIA_ROOT, output_sub_dir)
    os.makedirs(output_dir, exist_ok=True)

    output_name = 'VI_Tracker_HOTO_Filled.xlsm'
    output_path = os.path.join(output_dir, output_name)
    wb.save(output_path)

    file_url = request.build_absolute_uri(
        f"{settings.MEDIA_URL.rstrip('/')}/{output_sub_dir}/{output_name}".replace('\\', '/')
    )

    return Response({
        'file_url': file_url,
        'template_version': 'v1.8',
        'total_records': queryset.count(),
    }, status=200)