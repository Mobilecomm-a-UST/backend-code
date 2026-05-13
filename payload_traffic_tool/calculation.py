import os
import sys
from datetime import datetime, timedelta
from django.db import connection
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL


main_folder=os.path.join(MEDIA_ROOT, "Payload_traffic")
output_path = os.path.join(main_folder, "Report")
os.makedirs(output_path, exist_ok=True)

def generate_dates(on_air_date_str):
    for fmt in (
        "%d-%m-%Y","%Y-%m-%d", "%m-%d-%Y",
        "%d/%m/%Y", "%Y/%m/%d","%d-%b-%Y",
        "%d %b %Y", "%d %B %Y",
    ):
         
         try:
            on_air = datetime.strptime(on_air_date_str.strip(), fmt)
            break
         except ValueError:
              continue
    else:
        raise ValueError(f"Cannot parse on-air date: {on_air_date_str}")
    prev_dates = []
    for i in range(1,11):
         prev_dates.append(on_air - timedelta(days=i))

    fwd_dates = []
    for i in range(1,11):
        fwd_dates.append(on_air + timedelta(days=i))
    return on_air, prev_dates, fwd_dates
    
def get_traffic_for_date(cursor, table_name, site_id, traffic_date):
    query = (
        "SELECT COALESCE(SUM(traffic_value), 0) FROM " + table_name +
        " WHERE site_id = %s AND traffic_date = %s"
    )

    print("Executing SQL:", query, (site_id, traffic_date))

    cursor.execute(query, [site_id, traffic_date])
    result = cursor.fetchone()
    return round(result[0], 6) if result else 0.0
    
def calculate_site_traffic(site_id, on_air_date_str):
    on_air, prev_dates, fwd_dates = generate_dates(on_air_date_str)
    all_dates = prev_dates + fwd_dates
    date_strings = [d.strftime("%Y-%m-%d") for d in all_dates]

    cursor = connection.cursor()

    placeholders =  ",".join(["%s" for _ in date_strings])

    # ✅ FIXED TABLE NAME
    cursor.execute(f"""
        SELECT traffic_date, SUM(traffic_value)
        FROM payload_traffic_4g
        WHERE site_id = %s
        AND traffic_date IN ({placeholders})
        GROUP BY traffic_date
    """, [site_id] + date_strings)

    rows_4g = cursor.fetchall()
    map_4g = {str(row[0]): round(row[1], 6) for row in rows_4g}

    # ✅ FIXED TABLE NAME
    # FIXED — all %s
    cursor.execute(f"""
    SELECT traffic_date, SUM(traffic_value)
    FROM payload_traffic_4g
    WHERE site_id = %s
    AND traffic_date IN ({placeholders})
    GROUP BY traffic_date""", [site_id] + date_strings)

    rows_5g = cursor.fetchall()
    map_5g = {str(row[0]): round(row[1], 6) for row in rows_5g}

    # ── Build results from maps ───────────────────────────────
    def build_traffic_row(dates):
        date_labels, traffic_4g, traffic_5g, traffic_total = [], [], [], []
        for date in dates:
            date_str = date.strftime("%Y-%m-%d")
            date_labels.append(date_str)
            val_4g = map_4g.get(date_str, 0.0)
            val_5g = map_5g.get(date_str, 0.0)
            total  = round(val_4g + val_5g, 6)
            traffic_4g.append(val_4g)
            traffic_5g.append(val_5g)
            traffic_total.append(total)
        return date_labels, traffic_4g, traffic_5g, traffic_total

    prev_labels, prev_4g, prev_5g, prev_total = build_traffic_row(prev_dates)
    fwd_labels,  fwd_4g,  fwd_5g,  fwd_total  = build_traffic_row(fwd_dates)

    def avg(values):
        non_zero = [v for v in values if v > 0]
        return round(sum(non_zero) / len(non_zero), 6) if non_zero else 0.0

    return {
        "site_id":      site_id,
        "on_air_date":  on_air.strftime("%d-%m-%Y"),
        "prev_labels":  prev_labels,
        "prev_4g":      prev_4g,
        "prev_5g":      prev_5g,
        "prev_total":   prev_total,
        "prev_avg_4g":  avg(prev_4g),
        "prev_avg_5g":  avg(prev_5g),
        "prev_avg_tot": avg(prev_total),
        "fwd_labels":   fwd_labels,
        "fwd_4g":       fwd_4g,
        "fwd_5g":       fwd_5g,
        "fwd_total":    fwd_total,
        "fwd_avg_4g":   avg(fwd_4g),
        "fwd_avg_5g":   avg(fwd_5g),
        "fwd_avg_tot":  avg(fwd_total),
    }


def get_styles():
    thin = Side(style= "thin")
    border = Border(left = thin, right= thin, top = thin, bottom=thin)
    center = Alignment(horizontal="center", vertical = "center", wrap_text=True)
    left = Alignment(horizontal= "left", vertical = "center")

    header_fill = PatternFill("solid",fgColor= "1F4E79")
    header_font = Font(color="FFFFFF", bold= True, size=10)

    fill_4g = PatternFill("solid", fgColor= "DEEAF1")
    fill_5g = PatternFill("solid", fgColor= "E2EFDA")
    fill_total = PatternFill("solid", fgColor="FCE4D6")

    avg_fill = PatternFill("solid", fgColor ="FFF2CC")
    avg_font = Font(bold=True, size=10)

    return {
        "border" : border,
        "center": center,
        "left": left,
        "header_fill": header_fill,
        "header_font": header_font,
        "fill_4g": fill_4g,
        "fill_5g": fill_5g,
        "fill_total": fill_total,
        "avg_fill": avg_fill,
        "avg_font": avg_font
    }
def write_site_block(ws, result, start_row, date_labels,
                     traffic_4g, traffic_5g, traffic_total,
                     avg_4g, avg_5g, avg_total, styles):
    s= styles
    on_air = result["on_air_date"]
    site_id = result["site_id"]

    headers= ["On_air_date", "Site_ID", "Values"] + date_labels + ["Average"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column = col_idx)
        cell.value = header
        cell.fill = s["header_fill"]
        cell.font = s["header_font"]
        cell.alignment = s["center"]
        cell.border = s["border"] 
    rows = [
        ("4G Traffic", traffic_4g, avg_4g, s["fill_4g"]),
        ("5G Traffic", traffic_5g, avg_5g, s["fill_5g"]),
        ("Total Traffic", traffic_total, avg_total, s["fill_total"]),
    ]
    
    for row_offset, (label, values, avg, fill) in enumerate(rows, start=1):
        row = start_row + row_offset

        c = ws.cell(row=row, column = 1)
        c.value = on_air
        c.fill = fill
        c.alignment = s["center"]
        c.border = s["border"]

        c = ws.cell(row=row, column = 2)
        c.value = site_id
        c.fill = fill
        c.alignment = s["center"]
        c.border = s["border"]

        c = ws.cell(row=row, column = 3)
        c.value = label
        c.fill = fill
        c.font = Font(bold= True, size= 10)
        c.alignment = s["left"]
        c.border = s["border"]

        for col_offset, val in enumerate(values, start=4):
            c = ws.cell(row=row, column = col_offset)
            c.value = val
            c.fill = fill    
            c.alignment = s["center"]
            c.border = s["border"]

        avg_col = 4 + len(values)
        c = ws.cell(row=row, column= avg_col)
        c.value = avg 
        c.fill = s["avg_fill"]
        c.font = s["avg_font"]
        c.alignment = s["center"]
        c.border = s["border"]

    return start_row + 3 + 2 + 1

def set_column_widths(ws, num_dates):
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14

    for i in range(4,4 + num_dates):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = 13

    avg_col = openpyxl.utils.get_column_letter(4 + num_dates)
    ws.column_dimensions[avg_col].width =12

def export_to_excel(all_results, output_path):
    # today = datetime.now().strftime("%d-%m-%Y")
    filename = f"Traffic_Report_File.xlsx"
    filepath = os.path.join(output_path, filename)

    wb = openpyxl.Workbook()
    ws_prev = wb.active
    ws_prev.title = "Previous 10 Days"
    ws_fwd = wb.create_sheet(title = "Future 10 Days")

    styles = get_styles()

    prev_row = 1
    fwd_row = 1

    for result in all_results:
        prev_row = write_site_block(
            ws = ws_prev,
            result = result,
            start_row = prev_row,
            date_labels = result["prev_labels"],
            traffic_4g = result["prev_4g"],
            traffic_5g = result["prev_5g"],
            traffic_total = result["prev_total"],
            avg_4g = result["prev_avg_4g"],
            avg_5g = result["prev_avg_5g"],
            avg_total = result["prev_avg_tot"],
            styles = styles)
        
        fwd_row = write_site_block(
            ws = ws_fwd,
            result = result,
            start_row = fwd_row,
            date_labels = result["fwd_labels"],
            traffic_4g = result["fwd_4g"],
            traffic_5g = result["fwd_5g"],
            traffic_total = result["fwd_total"],
            avg_4g = result["fwd_avg_4g"],
            avg_5g = result["fwd_avg_5g"],
            avg_total = result["fwd_avg_tot"],
            styles = styles)

    set_column_widths(ws_prev, 10)
    set_column_widths(ws_fwd, 10)

    wb.save(filepath)
    print(f"\n Excel saved  {filepath}")
    return filepath

def calculate_traffic(site_ids_str, on_air_date_str):
    site_ids = [s.strip() for s in site_ids_str.split(",")]
    on_air, prev_dates, fwd_dates = generate_dates(on_air_date_str)
    all_dates = prev_dates + fwd_dates
    date_strings = [d.strftime("%Y-%m-%d") for d in all_dates]

    # ── ONE query for ALL sites and ALL dates ─────────────────
    cursor = connection.cursor()

    site_placeholders = ",".join(["%s" for _ in site_ids])
    date_placeholders = ",".join(["%s" for _ in date_strings])
    

    cursor.execute(f"""
        SELECT site_id, traffic_date, SUM(traffic_value)
        FROM payload_traffic_4g
        WHERE site_id IN ({site_placeholders})
        AND traffic_date IN ({date_placeholders})
        GROUP BY site_id, traffic_date
    """, site_ids + date_strings)

    all_4g = {}
    for site_id, date, val in cursor.fetchall():
        all_4g[(site_id, str(date))] = round(val, 6)

   
    cursor.execute(f"""
        SELECT site_id, traffic_date, SUM(traffic_value)
        FROM payload_traffic_5g
        WHERE site_id IN ({site_placeholders})
        AND traffic_date IN ({date_placeholders})
        GROUP BY site_id, traffic_date
    """, site_ids + date_strings)

    all_5g = {}
    for site_id, date, val in cursor.fetchall():
        all_5g[(site_id, str(date))] = round(val, 6)

    # ── Build results for each site ───────────────────────────
    def avg(values):
        non_zero = [v for v in values if v > 0]
        return round(sum(non_zero) / len(non_zero), 6) if non_zero else 0.0

    def build_row(site_id, dates, map_4g, map_5g):
        date_labels, traffic_4g, traffic_5g, traffic_total = [], [], [], []
        for date in dates:
            date_str = date.strftime("%Y-%m-%d")
            date_labels.append(date_str)
            val_4g = map_4g.get((site_id, date_str), 0.0)
            val_5g = map_5g.get((site_id, date_str), 0.0)
            traffic_4g.append(val_4g)
            traffic_5g.append(val_5g)
            traffic_total.append(round(val_4g + val_5g, 6))
        return date_labels, traffic_4g, traffic_5g, traffic_total

    all_results = []
    for site_id in site_ids:
        print(f" Calculating {site_id}...")
        prev_labels, prev_4g, prev_5g, prev_total = build_row(
            site_id, prev_dates, all_4g, all_5g)
        fwd_labels, fwd_4g, fwd_5g, fwd_total = build_row(
            site_id, fwd_dates, all_4g, all_5g)

        all_results.append({
            "site_id":      site_id,
            "on_air_date":  on_air.strftime("%d-%m-%Y"),
            "prev_labels":  prev_labels,
            "prev_4g":      prev_4g,
            "prev_5g":      prev_5g,
            "prev_total":   prev_total,
            "prev_avg_4g":  avg(prev_4g),
            "prev_avg_5g":  avg(prev_5g),
            "prev_avg_tot": avg(prev_total),
            "fwd_labels":   fwd_labels,
            "fwd_4g":       fwd_4g,
            "fwd_5g":       fwd_5g,
            "fwd_total":    fwd_total,
            "fwd_avg_4g":   avg(fwd_4g),
            "fwd_avg_5g":   avg(fwd_5g),
            "fwd_avg_tot":  avg(fwd_total),
        })

    # Print summary
    for result in all_results:
        print(f"\n{'='*80}")
        print(f" Site: {result['site_id']}")
        print(f" On_air_date: {result['on_air_date']}")
        print(f" PREVIOUS 10 DAYS:")
        print(f" 4G: {result['prev_4g']}  Avg: {result['prev_avg_4g']}")
        print(f" 5G: {result['prev_5g']}  Avg: {result['prev_avg_5g']}")
        print(f" Total: {result['prev_total']}  Avg: {result['prev_avg_tot']}")
        print(f" FUTURE 10 DAYS:")
        print(f" 4G: {result['fwd_4g']}  Avg: {result['fwd_avg_4g']}")
        print(f" 5G: {result['fwd_5g']}  Avg: {result['prev_avg_5g']}")
        print(f" Total: {result['fwd_total']}  Avg: {result['fwd_avg_tot']}")
        print(f"{'='*80}")

    filepath = export_to_excel(all_results, output_path)
    return all_results, filepath

if __name__ == "__main__":
    import sys


    if len(sys.argv) < 3:
        print("\nUsage: python calculation.py <site_ids> <on_air_dates>")
        print("Example: python calculation.py STD958 26-12-2025")
        print("Example: python calculation.py 'STD958, BKL722' 26-12-2025\n")
    

    else:
        site_ids_str = sys.argv[1]
        on_air_date_str = sys.argv[2]
        calculate_traffic(site_ids_str, on_air_date_str)

            








