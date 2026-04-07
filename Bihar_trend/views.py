from django.shortcuts import render
from django.http import HttpResponse
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from rest_framework.decorators import api_view
from rest_framework.response import Response
from tkinter import *
from django.core.files.storage import FileSystemStorage
from tkinter import filedialog, messagebox, ttk
import openpyxl
import pandas as pd
import numpy as np
from datetime import date, timedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
import datetime
import os
import glob
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment 
from openpyxl.utils import get_column_letter

from commom_utilities.utils import *
from rest_framework import status 
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST
from openpyxl.utils import column_index_from_string


@api_view(["POST"])
def old_bih_trend_smallcell(request):
    kpi=["RRC Setup Success Rate [CDBH]",
        "ERAB Setup Success Rate [CDBH]",
        "PS Drop Call Rate % [CDBH]",
        "DL User Throughput_Mbps [CDBH]",
        "UL User Throughput_Kbps [CDBH]",
        "PS handover success rate [LTE Intra System] [CDBH]",
        "PS handover success rate [LTE Inter System] [CDBH]",
        "Overall CSFB Success rate",
        "VoLTE ERAB Setup Success Rate [CBBH]",
        "VoLTE DCR [CBBH]",
        "VoLTE Packet Loss DL [CBBH]",
        "VoLTE Packet Loss UL [CBBH]",
        "VoLTE SRVCC SR",
        "VoLTE IntraF HOSR Exec",
        "VoLTE InterF HOSR [CBBH]",
        "E-UTRAN Average CQI [CDBH]",
        "UL NI [RSSI-SINR] [CDBH]",
        "4G Data Volume [GB]",
        "VoLTE Traffic_24Hrs",
        "Average number of used DL PRBs [CDBH]",
        "Radio NW Availability"]
    

    raw_kpi_4G=request.FILES['raw_kpi'] if "raw_kpi" in request.FILES else None
    df_raw_kpi=pd.read_excel( raw_kpi_4G)
    required_cols=["RRC Setup Success Rate [CDBH]",
        "ERAB Setup Success Rate [CDBH]",
        "PS Drop Call Rate % [CDBH]",
        "DL User Throughput_Kbps [CDBH]",
        "UL User Throughput_Kbps [CDBH]",
        "PS handover success rate [LTE Intra System] [CDBH]",
        "PS handover success rate [LTE Inter System] [CDBH]",
        "Overall CSFB Success rate",
        "VoLTE ERAB Setup Success Rate [CBBH]",
        "VoLTE DCR [CBBH]",
        "VoLTE Packet Loss DL [CBBH]",
        "VoLTE Packet Loss UL [CBBH]",
        "VoLTE SRVCC SR",
        "VoLTE IntraF HOSR Exec",
        "VoLTE InterF HOSR [CBBH]",
        "E-UTRAN Average CQI [CDBH]",
        "UL NI [RSSI-SINR] [CDBH]",
        "4G Data Volume [GB]",
        "VoLTE Traffic_24Hrs",
        "Average number of used DL PRBs [CDBH]",
        "Radio NW Availability"]
    sts,response=required_col_check(raw_kpi_4G,required_cols)
    if sts:
        return Response(response) 


    # site_list_4G=request.FILES["site_list"] if "site_list" in request.FILES else None
    # df_site_list=pd.read_excel(site_list_4G)

    # if site_list_4G:
    #     sts,response=required_col_check(raw_kpi_4G,required_cols)
    #     if sts:
    #         return Response(response)
    response,s_l= site_list_handler(request)
    if s_l:
        df_site_list=s_l
        site_list_str = [str(x) for x in df_site_list]
        print("_________________________--print",type(site_list_str[0]))

        print(df_site_list,"_______________")
        if response:
            return Response(response)


    door_path=os.path.join(MEDIA_ROOT,'trends','Bihar')
   
    df_raw_kpi['Short name']=df_raw_kpi['Short name'].fillna(method='ffill')
    print('_________date_________',df_raw_kpi)
    df_raw_kpi.columns.values[1]='Date'
    
    df_raw_kpi['DL User Throughput_Kbps [CDBH]']=(df_raw_kpi['DL User Throughput_Kbps [CDBH]']/1024)
    df_raw_kpi.rename(columns={"DL User Throughput_Kbps [CDBH]":"DL User Throughput_Mbps [CDBH]"},inplace=True)

    df_raw_kpi['site_id']=[site.split('_')[-2][:-1] if("_") in site else site for site in  df_raw_kpi['Short name']]
    df_raw_kpi["ecgi"]=[cell.split('-')[-1] if('-') in str(cell) else cell for cell in  df_raw_kpi["4G_ECGI"]]
    df_raw_kpi["ecgiid"]=[cell.split('-')[-2] if('-') in str(cell) else cell for cell in  df_raw_kpi["4G_ECGI"]]
    # df_raw_kpi["cell_id"]=df_raw_kpi["ecgiid"]+df_raw_kpi["ecgi"]
    df_raw_kpi["cell_id"]=(df_raw_kpi["ecgiid"]+df_raw_kpi["ecgi"]).tolist()
    cell_id1=df_raw_kpi["cell_id"]
    df_raw_kpi.fillna(value=0,inplace=True) 
   
    df_unique = df_raw_kpi.drop_duplicates(subset=['Short name',"Date",'ecgi'],keep='last')

    message=site_comparision(cell_id1,site_list_str) # site_comparision_call 
    print("message____",message)

    PsOsPath1=os.path.join(door_path,'process output','desired.xlsx')
    
    df_unique.to_excel(PsOsPath1,index=False)
    
    excelfile_1=PsOsPath1

    df1=pd.read_excel(excelfile_1)

   

    filtered_df_1 = df1[(df1.cell_id.isin(list(df_site_list)))]
    
    print(filtered_df_1)
    PsOsPath=os.path.join(door_path,'process output','filtered1.xlsx')
    filtered_df_1.to_excel(PsOsPath,index=False)       
    # print(df)
    
    df1 = pd.read_excel(PsOsPath)
    df_pivot = df1.pivot_table(values=kpi, columns='Date', index=['Short name','site_id','cell_id'])

    # df.fillna(value=0,inplace=True)
    PsOsPathpivot=os.path.join(door_path,'process output','pivot.xlsx')
    df_pivot.to_excel(PsOsPathpivot)
    

    STR=os.path.join(door_path,'templates','smallcell(uls).xlsx')
    wb=openpyxl.load_workbook(STR)
    ws=wb.active
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
        result= 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    str_date=request.POST.get("offered_date")
    date1 =datetime.datetime.strptime(str_date,'%Y-%m-%d')
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        print('donnnnnnnnnnnnnnnnne')
        print(len(index))
        dr=df_pivot[kpi_name]
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()

        ws[coln1+"2"].value=cl[4]
        ws[coln2+"2"].value=cl[3]
        ws[coln3+"2"].value=cl[2]
        ws[coln4+"2"].value=cl[1]
        ws[coln5+"2"].value=cl[0]

        for i,value in enumerate(index):
                j=i+5
                # ws['K'+str(j)].value=date1
                ws['A'+str(j)].value='BH'
                ws['D'+str(j)].value='NOKIA'
                
                ws['B'+str(j)].value=index[i][2]
                ws['C'+str(j)].value=index[i][1]
        
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
                
    for kpi_name in kpi:
        if(kpi_name=='RRC Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'E') 
        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'K') 
        if(kpi_name=='PS Drop Call Rate % [CDBH]'):
            overwrite(kpi_name,'Q') 
        if(kpi_name=='DL User Throughput_Mbps [CDBH]'):
            overwrite(kpi_name,'W') 
        if(kpi_name=='UL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name,'AC') 
        if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'AI') 
        if(kpi_name=='PS handover success rate [LTE Inter System] [CDBH]'):
            overwrite(kpi_name,'AO') 
        if(kpi_name=='Overall CSFB Success rate'):
            overwrite(kpi_name,'AU') 
        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'BA') 
        if(kpi_name=='VoLTE DCR [CBBH]'):
            overwrite(kpi_name,'BG') 
        if(kpi_name=='VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name,'BM') 
        if(kpi_name=='VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name,'BS') 
        if(kpi_name=='VoLTE SRVCC SR'):
            overwrite(kpi_name,'BY') 
        if(kpi_name=='VoLTE IntraF HOSR Exec'):
            overwrite(kpi_name,'CE') 
        if(kpi_name=='VoLTE InterF HOSR [CBBH]'):
            overwrite(kpi_name,'CK') 
        if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name,'CQ') 
        if(kpi_name=='UL NI [RSSI-SINR] [CDBH] '):
            overwrite(kpi_name,'CY') 
        if(kpi_name=='4G Data Volume [GB]'):
            overwrite(kpi_name,'DE') 
        if(kpi_name=='VoLTE Traffic_24Hrs'):
            overwrite(kpi_name,'DM') 
        if(kpi_name=='Average number of used DL PRBs [CDBH]'):
            overwrite(kpi_name,'DS')   
        if(kpi_name=='Radio NW Availability'):
            overwrite(kpi_name,'DY')  
    SaveOutput=os.path.join(door_path,'output','biharoutput.xlsx')        
    wb.save(SaveOutput) 
    download_path=os.path.join(MEDIA_URL,"trends","Bihar","output","biharoutput.xlsx")
    return Response({"status":True,"message":"successfully","missing_sites":message,"Download_url":download_path}) 
 
############################################################################################
@api_view(["POST"])
def old_bih_trend_macro(request):
    kpi=["Radio NW Availability",
         "MV_4G Data Volume_GB",
        "Data Volume DL - Total [MB] [CDBH]",
        "VoLTE Traffic_24Hrs",
        "DL User Throughput_Mbps [CDBH]",
        "UL User Throughput_Kbps [CDBH]",
        "Average number of used DL PRBs [CDBH]",
        "ERAB Setup Success Rate [CDBH]",
        "ERAB Setup Successful [CDBH]",
        "ERAB Setup Attempts [CDBH]",
        "RRC Setup Success Rate [CDBH]",
        "RRC Attempts [CDBH]",
        "RRC Success [CDBH]",
        "VoLTE Packet Loss DL [CBBH]",
        "VoLTE Packet Loss UL [CBBH]",
        "VoLTE ERAB Setup Success Rate [CBBH]",
        "VoLTE ERAB Setup Success Rate_Nom [CBBH]",
        "VoLTE ERAB Setup Success Rate_Denom [CBBH]",
        "PS Drop Call Rate % [CDBH]",
        "PS Drop Call Rate NOM [CDBH]",
        "PS Drop Call Rate DENOM [CDBH]",
        "VoLTE DCR [CBBH]",
        "VoLTE DCR_Nom [CBBH]",
        "VoLTE DCR_Denom [CBBH]",
        "PS handover success rate [LTE Intra System] [CDBH]",
        "PS handover success rate [LTE Inter System] [CDBH]",
        "PS handover success rate [LTE Intra System]_NOM [CDBH]",
        "PS handover success rate [LTE Intra System]_DENOM [CDBH]",
        "VoLTE InterF HOSR Exec",
        "VoLTE InterF HOSR Exec_Nom [CBBH]",
        "VoLTE InterF HOSR Exec_Denom [CBBH]",
        "VoLTE IntraF HOSR Exec",
        "VoLTE IntraF HOSR Exec_Nom [CBBH]",
        "VoLTE IntraF HOSR Exec_Denom [CBBH]",
        "E-UTRAN Average CQI [CDBH]",
        "UL PUCCH SINR [CDBH]",
        "UL PUSCH SINR [CDBH]",
        "Average UE Distance_KM [CDBH]",
        "VoLTE SRVCC SR",
        "VoLTE SRVCC SR [CBBH]",
        "VoLTE SRVCC SR_Nom [CBBH]",
        "VoLTE SRVCC SR_Denom [CBBH]"]
    
    raw_kpi_4G=request.FILES['raw_kpi'] if "raw_kpi" in request.FILES else None
    df_raw_kpi=pd.read_excel( raw_kpi_4G)
    required_cols=["Radio NW Availability",
         "MV_4G Data Volume_GB",
        "Data Volume DL - Total [MB] [CDBH]",
        "VoLTE Traffic_24Hrs",
        "DL User Throughput_Kbps [CDBH]",
        "UL User Throughput_Kbps [CDBH]",
        "Average number of used DL PRBs [CDBH]",
        "ERAB Setup Success Rate [CDBH]",
        "ERAB Setup Successful [CDBH]",
        "ERAB Setup Attempts [CDBH]",
        "RRC Setup Success Rate [CDBH]",
        "RRC Attempts [CDBH]",
        "RRC Success [CDBH]",
        "VoLTE Packet Loss DL [CBBH]",
        "VoLTE Packet Loss UL [CBBH]",
        "VoLTE ERAB Setup Success Rate [CBBH]",
        "VoLTE ERAB Setup Success Rate_Nom [CBBH]",
        "VoLTE ERAB Setup Success Rate_Denom [CBBH]",
        "PS Drop Call Rate % [CDBH]",
        "PS Drop Call Rate NOM [CDBH]",
        "PS Drop Call Rate DENOM [CDBH]",
        "VoLTE DCR [CBBH]",
        "VoLTE DCR_Nom [CBBH]",
        "VoLTE DCR_Denom [CBBH]",
        "PS handover success rate [LTE Intra System] [CDBH]",
        "PS handover success rate [LTE Inter System] [CDBH]",
        "PS handover success rate [LTE Intra System]_NOM [CDBH]",
        "PS handover success rate [LTE Intra System]_DENOM [CDBH]",
        "VoLTE InterF HOSR Exec",
        "VoLTE InterF HOSR Exec_Nom [CBBH]",
        "VoLTE InterF HOSR Exec_Denom [CBBH]",
        "VoLTE IntraF HOSR Exec",
        "VoLTE IntraF HOSR Exec_Nom [CBBH]",
        "VoLTE IntraF HOSR Exec_Denom [CBBH]",
        "E-UTRAN Average CQI [CDBH]",
        "UL PUCCH SINR [CDBH]",
        "UL PUSCH SINR [CDBH]",
        "Average UE Distance_KM [CDBH]",
        "VoLTE SRVCC SR",
        "VoLTE SRVCC SR [CBBH]",
        "VoLTE SRVCC SR_Nom [CBBH]",
        "VoLTE SRVCC SR_Denom [CBBH]"]
    sts,response=required_col_check(raw_kpi_4G,required_cols)
    if sts:
        return Response(response)

    # site_list_4G=request.FILES["site_list"] if "site_list" in request.FILES else None
    # df_site_list=pd.read_excel(site_list_4G)
    # if site_list_4G:
    #     sts,response=required_col_check(raw_kpi_4G,required_cols)
    #     if sts:
    #         return Response(response)
    response,s_l= site_list_handler(request)
    if s_l:
        df_site_list=s_l
        
        print("_________________________--print",type(df_site_list[0]))

        print(df_site_list,"_______________")
        if response:
            return Response(response)

    door_path=os.path.join(MEDIA_ROOT,'trends','Bihar')
   
    df_raw_kpi['Short name']=df_raw_kpi['Short name'].fillna(method='ffill')
    df_raw_kpi["DL User Throughput_Kbps [CDBH]"]=(df_raw_kpi["DL User Throughput_Kbps [CDBH]"]/1024)
    df_raw_kpi.rename(columns={"DL User Throughput_Kbps [CDBH]":"DL User Throughput_Mbps [CDBH]"},inplace=True)
    df_raw_kpi.columns.values[1]='Date'
    df_raw_kpi.fillna(value=0,inplace=True) 
    df_raw_kpi["SITE_ID"]=[site.split('_')[-2][:-1] if'_' in site else site for site in df_raw_kpi["Short name"]]
    df_raw_kpi["SITE"]=[site.split('_')[-2] if'_' in site else site for site in df_raw_kpi["Short name"]]
    df_raw_kpi["lnbts_id"]=[cell.split('-')[-2] if '-' in str(cell) else cell for cell in df_raw_kpi["4G_ECGI"]]
    df_raw_kpi["L1800 LCR"]=[cell1.split('-')[-1] if'-' in str(cell1) else cell1 for cell1 in df_raw_kpi["4G_ECGI"]]
    df_raw_kpi['sec']=[str(cell2)[-1:] for cell2 in df_raw_kpi["L1800 LCR"]]
    df_raw_kpi["SECTOR"]=df_raw_kpi["lnbts_id"].astype(str) + '_' + df_raw_kpi['sec'].astype(str)
    print(df_raw_kpi)
    df_unique = df_raw_kpi.drop_duplicates(subset=['Short name',"Date",'4G_ECGI'],keep='last')
    message=site_comparision(df_raw_kpi["ecgiid"],df_site_list)
    PsOsPath1=os.path.join(door_path,'process output','desired.xlsx')
    
    df_unique.to_excel(PsOsPath1,index=False)
    
    excelfile_1=PsOsPath1

    df1=pd.read_excel(excelfile_1)

    
    
    kpi1=["MV_4G Data Volume_GB"]
    filtered_df_1 = df1[(df1.lnbts_id.isin(df_site_list))]
    
    print(filtered_df_1)
    PsOsPath=os.path.join(door_path,'process output','filtered1.xlsx')
    filtered_df_1.to_excel(PsOsPath,index=False)       
    # print(df)
    
    df1 = pd.read_excel(PsOsPath)
    df_pivot = df1.pivot_table(values=kpi, columns='Date',index=['Short name','4G_ECGI','SITE_ID','lnbts_id',"SECTOR",'sec','L1800 LCR','SITE'])

    df_pivot_vol = df1.pivot_table(values=kpi1,columns='Date',index=['SITE'],aggfunc='sum')
    # df.fillna(value=0,inplace=True)
    PsOsPathpivot=os.path.join(door_path,'process output','pivot.xlsx')
    PsOsPathpivot_vol=os.path.join(door_path,'process output','pivot_vol.xlsx')

    df_pivot.to_excel(PsOsPathpivot)
    df_pivot_vol.to_excel(PsOsPathpivot_vol)

    

    STR=os.path.join(door_path,'templates','macro(uls).xlsx')
    wb=openpyxl.load_workbook(STR)
    ws=wb["Sheet1"]
    ws.insert_rows(312)
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
        result= 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    str_date=request.POST.get("offered_date")
    date1 =datetime.datetime.strptime(str_date,'%Y-%m-%d')
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        print('donnnnnnnnnnnnnnnnne')
        print(len(index))
        dr=df_pivot[kpi_name]
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()

        ws[coln1+"2"].value=cl[4]
        ws[coln2+"2"].value=cl[3]
        ws[coln3+"2"].value=cl[2]
        ws[coln4+"2"].value=cl[1]
        ws[coln5+"2"].value=cl[0]

        for i,value in enumerate(index):
                j=i+3
                # ws['K'+str(j)].value=date1
                ws['A'+str(j)].value=index[i][1]
                ws['N'+str(j)].value=index[i][1]
                # ws['D'+str(j)].value='NOKIA'
                
                ws['B'+str(j)].value=index[i][2]
                ws['D'+str(j)].value=index[i][3]
                ws['M'+str(j)].value=index[i][3]
                ws['E'+str(j)].value=index[i][4]
                ws['I'+str(j)].value=index[i][0]
                ws['K'+str(j)].value=index[i][5]
                ws['L'+str(j)].value=index[i][6]
                ws['KZ'+str(j)].value=index[i][7]
        
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
                
    for kpi_name in kpi:
        if(kpi_name=='Radio NW Availability'):
            overwrite(kpi_name,'AB') 
        if(kpi_name=='Data Volume DL - Total [MB] [CDBH]'):
            overwrite(kpi_name,'BR')    
        if(kpi_name=='VoLTE Traffic_24Hrs'):
            overwrite(kpi_name,'BX') 
        if(kpi_name=='DL User Throughput_Mbps [CDBH]'):
            overwrite(kpi_name,'CD') 
        if(kpi_name=='UL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name,'CJ') 
        if(kpi_name=='Average number of used DL PRBs [CDBH]'):
            overwrite(kpi_name,'CP') 
        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'CV') 
        if(kpi_name=='ERAB Setup Successful [CDBH]'):
            overwrite(kpi_name,'DB') 
        if(kpi_name=='ERAB Setup Attempts [CDBH]'):
            overwrite(kpi_name,'DH') 
        if(kpi_name=='RRC Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'DN') 
        if(kpi_name=='RRC Attempts [CDBH]'):
            overwrite(kpi_name,'DT') 
        if(kpi_name=='RRC Success [CDBH]'):
            overwrite(kpi_name,'DZ') 
        if(kpi_name=='VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name,'EF') 
        if(kpi_name=='VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name,'EL') 
        if(kpi_name=='VoLTE ERAB Setup Success Rate [CBBH]'):
            overwrite(kpi_name,'ER') 
        if(kpi_name=='VoLTE ERAB Setup Success Rate_Nom [CBBH]'):
            overwrite(kpi_name,'EX') 
        if(kpi_name=='VoLTE ERAB Setup Success Rate_Denom [CBBH]'):
            overwrite(kpi_name,'FD') 
        if(kpi_name=='PS Drop Call Rate % [CDBH]'):
            overwrite(kpi_name,'FJ') 
        if(kpi_name=='PS Drop Call Rate NOM [CDBH]'):
            overwrite(kpi_name,'FP') 
        if(kpi_name=='PS Drop Call Rate DENOM [CDBH]'):
            overwrite(kpi_name,'FV') 
        if(kpi_name=='VoLTE DCR [CBBH]'):
            overwrite(kpi_name,'GB')   
        if(kpi_name=='VoLTE DCR_Nom [CBBH]'):
            overwrite(kpi_name,'GH')  
        if(kpi_name=='VoLTE DCR_Denom [CBBH]'):
            overwrite(kpi_name,'GN')  
        if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'GT')  
        if(kpi_name=='PS handover success rate [LTE Inter System] [CDBH]'):
            overwrite(kpi_name,'GZ')  
        if(kpi_name=='PS handover success rate [LTE Intra System]_NOM [CDBH]'):
            overwrite(kpi_name,'HF')  
        if(kpi_name=='PS handover success rate [LTE Intra System]_DENOM [CDBH]'):
            overwrite(kpi_name,'HL')  
        if(kpi_name=='VoLTE InterF HOSR Exec'):
            overwrite(kpi_name,'HR')  
        if(kpi_name=='VoLTE InterF HOSR Exec_Nom [CBBH]'):
            overwrite(kpi_name,'HX')  
        if(kpi_name=='VoLTE InterF HOSR Exec_Denom [CBBH]'):
            overwrite(kpi_name,'ID')  
        if(kpi_name=='VoLTE IntraF HOSR Exec'):
            overwrite(kpi_name,'IJ')  
        if(kpi_name=='VoLTE IntraF HOSR Exec_Nom [CBBH]'):
            overwrite(kpi_name,'IP')  
        if(kpi_name=='VoLTE IntraF HOSR Exec_Denom [CBBH]'):
            overwrite(kpi_name,'IV')  
        if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name,'JB')  
        if(kpi_name=='UL PUCCH SINR [CDBH]'):
            overwrite(kpi_name,'JH')  
        if(kpi_name=='UL PUSCH SINR [CDBH]'):
            overwrite(kpi_name,'JN')  
        if(kpi_name=='Average UE Distance_KM [CDBH]'):
            overwrite(kpi_name,'JT')  
        if(kpi_name=='VoLTE SRVCC SR'):
            overwrite(kpi_name,'JZ')  
        if(kpi_name=='VoLTE SRVCC SR [CBBH]'):
            overwrite(kpi_name,'KF')  
        if(kpi_name=='VoLTE SRVCC SR_Nom [CBBH]'):
            overwrite(kpi_name,'KL')  
        if(kpi_name=='VoLTE SRVCC SR_Denom [CBBH]'):
            overwrite(kpi_name,'KR') 
############################ MV_4G_DATAVOLUME ######################################
    # for col_idx, date_value in enumerate(cl, start=0):

    #     target_column = chr(ord("U") + col_idx)
    

    #     cell = ws[target_column + '2']
    #     cell.value = date_value

    #     print("__________cell",cell.value)
    # for i in range(1,1000):
    #     for site in df_pivot_vol.index:     
    #         if ws['KZ'+str(i)].value==site:
    #             print(ws['KZ'+str(i)].value==site,"DONE")        
    #             a2=df_pivot_vol[df_pivot_vol.index ==site ]["MV_4G Data Volume_GB"].values[0]
     
    #             for j, value in enumerate(a2, start=0):
    
    #                 target_column = chr(ord("U") + len(cl) + j)
    
    #                 cell = ws[chr(ord('U') + j) + str(i)]
    #                 cell.value = value




    SaveOutput=os.path.join(door_path,'output','biharoutput_macro.xlsx')        
    wb.save(SaveOutput) 
    download_path=os.path.join(MEDIA_URL,"trends","Bihar","output","biharoutput_macro.xlsx")
    return Response({"status":True,"message":"successfully","missing_sites":message,"Download_url":download_path}) 

    ############################################## DEGROW #############################
# @api_view(["POST"])

# def old_bih_trend_degrow(request):
#     # required_cols=['MV_4G Data Volume_GB',"VoLTE Traffic",'MV_DL User Throughput_Kbps',
#     #                 'RRC Setup Success Rate','VoLTE DCR','VoLTE ERAB Setup Success Rate',
#     #                 'ERAB Setup Success Rate [CDBH]','MV_UL User Throughput_Kbps']
#     raw_kpi_pre=request.FILES['raw_kpi_pre'] if "raw_kpi_pre" in request.FILES else None
#     df_raw_kpi_pre=pd.read_csv(raw_kpi_pre)



#     raw_kpi_post=request.FILES['raw_kpi_post'] if "raw_kpi_post" in request.FILES else None
#     df_raw_kpi_post=pd.read_csv(raw_kpi_post)


#     site_list=request.FILES["site_list"] if "site_list" in request.FILES else None
#     degrow_site_list=pd.read_excel(site_list)


#     door_path=os.path.join(MEDIA_ROOT,'trends','Bihar')


#     # df_pre=pd.read_csv("actual input/Pre Post/Pre Post Part 9/Pre 6 JUL TO 10 JUL.csv")

#     df_raw_kpi_pre['Short name']=df_raw_kpi_pre['Short name'].fillna(method='ffill')
#     df_raw_kpi_pre = df_raw_kpi_pre[
#     ~df_raw_kpi_pre['Short name'].str.startswith('@Nokia', na=False)
# ]


#     print(df_raw_kpi_pre['Short name'].dtype)
#     df_raw_kpi_pre["DL User Throughput_Kbps"]=(df_raw_kpi_pre["DL User Throughput_Kbps"]/1024)
#     df_raw_kpi_pre.rename(columns={"DL User Throughput_Kbps":"DL User Throughput_Mbps"},inplace=True)
#     #df.fillna(method='ffill',inplace=true)
#     df_raw_kpi_pre.fillna(value=0,inplace=True)                     
#     df_raw_kpi_pre.columns.values[1]='Date'
#     df_raw_kpi_pre['band2']=[ band.split('_')[2] for band in df_raw_kpi_pre['Short name']]
#     df_raw_kpi_pre['SITE_ID']=[site.split('_')[-2][:-1] for site in df_raw_kpi_pre['Short name']]
#     df_raw_kpi_pre['CELL_ID']=[site.split('_')[-2] if '_' in str(site) else str(site) for site in df_raw_kpi_pre['Short name']]


#     df_raw_kpi_pre['band2']=df_raw_kpi_pre['band2'].replace(['F1','F3','F8'],['L2100','L1800','L900'])
#     # df_pre['band2']=df_pre['band2'].fillna(method='ffill')
#     # print(df_pre)
#     tech=[]
#     for band1 in df_raw_kpi_pre['band2']:
#         if('T1' in band1 or 'T2' in band1):
#             if('T1' in band1):
#                 band='TDDC2'
#             if('T2' in band1):
#                 band='TDDC1'
#             tech.append(band) 
#         else:
#             band=band1
#             tech.append(band)
#     df_raw_kpi_pre.insert(1,'tech',tech) 

#     df_raw_kpi_pre=df_raw_kpi_pre.drop('band2',axis=1)############### drop ######
#     pre_os_path=os.path.join(door_path,'process output','desired_pre.xlsx')
#     df_raw_kpi_pre.to_excel(pre_os_path,index=False)
#     # site_list="project file/pre_post_site.xlsx"
#     # site_list=pd.read_excel(site_list)

#     filter_pre=df_raw_kpi_pre[(df_raw_kpi_pre.SITE_ID.isin(list(degrow_site_list['Site_ID'])))]
#     pre_os_filter_path=os.path.join(door_path,'process output','filter.xlsx')
#     filter_pre.to_excel(pre_os_filter_path,index=False)

#     # pivot_pre=filter_pre.pivot_table(index="Date",columns=['Short name','tech','SITE_ID','CELL_ID','4G_ECGI'],aggfunc={'MV_4G Data Volume_GB':'sum',"VoLTE Traffic [CBBH]":'sum','DL User Throughput_Mbps [CDBH]':'mean',
#     #                                                                             'RRC Setup Success Rate [CDBH]':'mean','VoLTE DCR [CBBH]':'mean','ERAB Setup Success Rate [CDBH]':'mean',
#     #                                                                             "MV_Avg Connected User [CDBH]":"mean",'Average number of used DL PRBs [CDBH]':'mean','UL User Throughput_Kbps [CUBH]':'mean'})
    
#     pivot_pre=filter_pre.pivot_table(index="Date",columns=['Short name','tech','SITE_ID','CELL_ID','4G_ECGI'],aggfunc={'MV_4G Data Volume_GB':'sum',"VoLTE Traffic":'sum','DL User Throughput_Mbps':'mean',
#                                                                                 'RRC Setup Success Rate':'mean','VoLTE DCR':'mean','VoLTE ERAB Setup Success Rate':'mean',
#                                                                             'ERAB Setup Success Rate [CDBH]':'mean','UL User Throughput_Kbps':'mean'})
#     pivot_pre=pivot_pre.T
#     # rounded_df_post['Pre Avg']=round(rounded_df_pre.mean(axis=1),2)   
#     pre_os_path_pivot=os.path.join(door_path,'process output','pivoted.xlsx')
#     pivot_pre.to_excel(pre_os_path_pivot)

#     rounded_df_pre=pivot_pre.round(2)
# ###################################### POST #####################################

#     df_raw_kpi_post['Short name']=df_raw_kpi_post['Short name'].fillna(method='ffill')
#     df_raw_kpi_post = df_raw_kpi_post[
#     ~df_raw_kpi_post['Short name'].str.startswith('@Nokia', na=False)
# ]

#     df_raw_kpi_post["DL User Throughput_Kbps"]=(df_raw_kpi_post["DL User Throughput_Kbps"]/1024)
#     df_raw_kpi_post.rename(columns={"DL User Throughput_Kbps":"DL User Throughput_Mbps"},inplace=True)
#     df_raw_kpi_post.fillna(value=0,inplace=True)                     
#     #df.fillna(method='ffill',inplace=true)
#     df_raw_kpi_post.columns.values[1]='Date'
#     df_raw_kpi_post['band2']=[band.split('_')[2] if '_' in str(band) else str(band) for band in df_raw_kpi_post['Short name']]
#     df_raw_kpi_post['SITE_ID']=[site.split('_')[-2][:-1] if '_' in str(site) else str(site)[:-1] for site in df_raw_kpi_post['Short name']]
#     df_raw_kpi_post['band2']=df_raw_kpi_post['band2'].replace(['F1','F3','F8'],['L2100','L1800','L900'])
#     df_raw_kpi_post['CELL_ID']=[site.split('_')[-2] if '_' in str(site) else str(site) for site in df_raw_kpi_post['Short name']]

#     tech=[]
#     for band1 in df_raw_kpi_post['band2']:
#         if('T1' in band1 or 'T2' in band1):
#             if('T1' in band1):
#                 band='TDDC2'
#             if('T2' in band1):
#                 band='TDDC1'
#             tech.append(band) 
#         else:
#             band=band1
#             tech.append(band)
#     df_raw_kpi_post.insert(1,'tech',tech) 

#     df_raw_kpi_post=df_raw_kpi_post.drop('band2',axis=1)############### drop ######
#     post_os_path=os.path.join(door_path,"process output","desired_post.xlsx")
#     df_raw_kpi_post.to_excel(post_os_path,index=False)

# #     ######################## filter ##################################


#     filter_post=df_raw_kpi_post[(df_raw_kpi_post.SITE_ID.isin(list(degrow_site_list['Site_ID'])))]
#     post_os_filter_path=os.path.join(door_path,'process output','filter_post.xlsx')
#     filter_post.to_excel(post_os_filter_path,index=False)
#     # print("_________________filter_______",filter_post)

# #     ###################################################################################################################

#     # pivot_post_vol=filter_post.pivot_table(index="Date",columns=['Short name','tech','SITE_ID','CELL_ID','4G_ECGI'],aggfunc={'MV_4G Data Volume_GB':'sum',"VoLTE Traffic":'sum','MV_DL User Throughput_Mbps':'mean',
#     #                                                                             'RRC Setup Success Rate':'mean','VoLTE DCR':'mean','VoLTE ERAB Setup Success Rate':'mean',
#     #                                                                             "MV_Avg Connected User":"mean",'ERAB Setup Success Rate [CDBH]':'mean','MV_UL User Throughput_Kbps':'mean'})
#     pivot_post_vol=filter_post.pivot_table(index="Date",columns=['Short name','tech','SITE_ID','CELL_ID','4G_ECGI'],aggfunc={'MV_4G Data Volume_GB':'sum',"VoLTE Traffic":'sum','DL User Throughput_Mbps':'mean',
#                                                                                 'RRC Setup Success Rate':'mean','VoLTE DCR':'mean','VoLTE ERAB Setup Success Rate':'mean',
#                                                                             'ERAB Setup Success Rate [CDBH]':'mean','UL User Throughput_Kbps':'mean'})
    
#     pivot_post=pivot_post_vol.T
#     rounded_df_post=pivot_post.round(2)
#     # rounded_df_post['Pre Avg']=round(rounded_df_pre.mean(axis=1),2)   
#     post_os_pivot_path=os.path.join(door_path,"process output","post_pivoted.xlsx")
#     rounded_df_post.to_excel(post_os_pivot_path)
   

# ######################################### concat ##########################
#     concat_pre_post = pd.concat([rounded_df_pre, rounded_df_post],axis=1) 
#     concat_pre_post.fillna(value=0,inplace=True) 
#     concat_pre_post["Pre Avg"]=round(concat_pre_post.iloc[:,[0,1,2,3,4]].mean(axis=1),2)
#     concat_pre_post['Post Avg']=round(concat_pre_post.iloc[:,[5,6,7,8,9]].mean(axis=1),2)############
#     concat_pre_post['Delta']=round(concat_pre_post['Post Avg']-concat_pre_post['Pre Avg'],2)
#     concat_pre_post['Change%']=round(concat_pre_post['Delta']/concat_pre_post['Pre Avg'],2)
    
#     concat_pre_post.replace([np.inf, -np.inf], 0, inplace=True)
#     concat_pre_post['Change%'].fillna(0,inplace=True)

#     concat_pre_post_os_path=os.path.join(door_path,"process output","concat_pre_post.xlsx")

#     concat_pre_post.to_excel( concat_pre_post_os_path)

#     df=pd.read_excel(concat_pre_post_os_path)
# #######################################################################
#     df.rename(columns={'Unnamed: 0':'KPI'}, inplace=True)
#     df.rename(columns={'Short name':'Row Labels'},inplace=True)
#     df.rename(columns={'tech':'technology'},inplace=True)
#     df.set_index(['KPI','Row Labels'])
#     df_os_ws_path=os.path.join(door_path,"process output","concat_pre_post2.xlsx")
#     df.to_excel(df_os_ws_path,index=False,header=True,startrow=1)####### TO CHIFT A HEADER USE START ROW ###
#     df['KPI'].ffill()

#     df1=df_os_ws_path

#     wb=openpyxl.load_workbook(df1)
#     ws=wb.active
# ############################### font size################################
#     font=Font(size=9)
#     alignment = Alignment(horizontal='center', vertical='center')
#     for col_cells in ws.columns:
#         for cell in col_cells:
#             cell.font=font
#             cell.alignment=alignment 
# ###################################### merge cell ##############################
#     ws.merge_cells('G1:K1')
#     ws.merge_cells('L1:P1')  
# ############################# AFTER MERGING PUT COLUMNS NAME ################
#     ws['G1']='PRE'  
#     ws['L1']='POST'
# ################################ IN CENTER HEADER ###########################
#     merged_cell=ws['G1']
#     merged_cell1=ws['L1']  
#     merged_cell.alignment = Alignment(horizontal='center', vertical='center')
#     merged_cell1.alignment = Alignment(horizontal='center', vertical='center')
# # ############################ to change row wise color change ####################################
#     row_number=2
#     # fillcolor='FFFFFF'

#     for cell in ws[row_number]:
#         cell.font = Font(bold=True,size=9)
# ############### FOR COLORING ###############
#     YELLOW="FFE699"
#     LIGHT_GREEN="C6E0B4"
#     LIGHT_GREY="D9E1F2"
#     PURPLE="7030A0"
#     CYAN="00B0F0"
#     ORANGE="FFC000"
#     light_blue='BDD7EE'

#     ws["G1"].fill=PatternFill(patternType='solid',fgColor=YELLOW)
#     ws['L1'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREEN)
#     ws['A2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
#     ws['F2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
#     ws['G2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
#     ws['H2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
#     ws['I2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
#     ws['J2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
#     ws['K2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
#     ws['L2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
#     ws['M2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
#     ws['N2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
#     ws['O2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
#     ws['B2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
#     ws['P2'].fill=PatternFill(patternType='solid',fgColor=ORANGE)
#     ws['Q2'].fill=PatternFill(patternType='solid',fgColor=CYAN)
#     ws['R2'].fill=PatternFill(patternType='solid',fgColor=PURPLE)
#     ws['S2'].fill=PatternFill(patternType='solid',fgColor=light_blue)


    


# #  ############################################# till your row value it will be border thin  not to fix #####################
#     thin_border = Border(left=Side(style='thin'),
#                     right=Side(style='thin'),
#                     top=Side(style='thin'),
#                     bottom=Side(style='thin'))
#     for row in ws.iter_rows():
#         for cell in row:
#             if cell.value is not None:
#                 cell.border = thin_border 

# ############################## column changing ###########################################
#     # source_column_index = 5  # Replace with the source column index
#     # target_column_index = 0  # Replace with the target column index

#     # source_column_values = []

#     # for row in ws.iter_rows(min_row=6, values_only=True):
#     #     source_value = row[source_column_index]
#     #     source_column_values.append(source_value)
#     # for index, row in enumerate(ws.iter_rows(min_row=6), start=6):
#     #     target_cell = row[target_column_index]
#     #     target_cell.value = source_column_values[index - 6]     

#     # ws.delete_cols(source_column_index + 1)
#     save_output=os.path.join(door_path,'output','BIH_DEGROW_OUTPUT.xlsx')

#     wb.save(save_output)  

#     download_path=os.path.join(MEDIA_URL,"trends","Bihar","output","BIH_DEGROW_OUTPUT.xlsx")
#     return Response({"status":True,"message":"uploaded successfully","Download_url":download_path}) 

# ---- Added helper ----
def keep_only_5days_per_cell(df, date_col="Date", site_col="SITE_ID", mode="pre"):

    """
    mode = 'pre'  -> last 5 days BEFORE event
    mode = 'post' -> last 5 available days AFTER filtering
    """

    df = df.copy()
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")

    final_parts = []

    # GROUP BY SITE
    for site, g in df.groupby(site_col):

        unique_dates = sorted(g[date_col].dropna().unique())

        if len(unique_dates) == 0:
            continue

        # both pre & post need latest 5 KPI days
        selected_dates = unique_dates[-5:]

        g = g[g[date_col].isin(selected_dates)]

        final_parts.append(g)

    if final_parts:
        return pd.concat(final_parts, ignore_index=True)
    else:
        return pd.DataFrame(columns=df.columns)

    
def force_only_5_dates_after_pivot(df,mode="pre"):

    date_cols = [c for c in df.columns if isinstance(c, (pd.Timestamp, datetime.date)) or str(c)[:2].isdigit()]

    if not date_cols:
        return df

    date_cols = sorted(date_cols)

    # keep latest 5 dates
    keep = date_cols[-5:]

    non_date_cols = [c for c in df.columns if c not in date_cols]

    return df[non_date_cols + keep]




def get_pre_post_5days(df, site, dismantle_date):
    df = df.copy()
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df[df['SITE_ID'] == site]

    before = df[df['Date'] < dismantle_date]
    after  = df[df['Date'] >= dismantle_date]

    return before, after





@api_view(["POST"])
def old_bih_trend_degrow(request):
    raw_kpi_pre=request.FILES['raw_kpi_pre'] if "raw_kpi_pre" in request.FILES else None
    df_raw_kpi_pre=pd.read_csv(raw_kpi_pre)

    raw_kpi_post=request.FILES['raw_kpi_post'] if "raw_kpi_post" in request.FILES else None
    df_raw_kpi_post=pd.read_csv(raw_kpi_post)

    site_list=request.FILES["site_list"] if "site_list" in request.FILES else None
    degrow_site_list=pd.read_excel(site_list)

    door_path=os.path.join(MEDIA_ROOT,'trends','Bihar')

    # ---------- PRE PROCESS ----------
    df_raw_kpi_pre['Short name']=df_raw_kpi_pre['Short name'].ffill()
    df_raw_kpi_pre = df_raw_kpi_pre[~df_raw_kpi_pre['Short name'].str.startswith('@Nokia', na=False)]

    df_raw_kpi_pre["DL User Throughput_Kbps"]=(df_raw_kpi_pre["DL User Throughput_Kbps"]/1024)
    df_raw_kpi_pre.rename(columns={"DL User Throughput_Kbps":"DL User Throughput_Mbps"},inplace=True)
    df_raw_kpi_pre.fillna(value=0,inplace=True)
    df_raw_kpi_pre.columns.values[1]='Date'

    df_raw_kpi_pre['band2']=[band.split('_')[2] if '_' in str(band) else str(band) for band in df_raw_kpi_pre['Short name']]
    df_raw_kpi_pre['SITE_ID']=[site.split('_')[-2][:-1] if '_' in str(site) else str(site)[:-1] for site in df_raw_kpi_pre['Short name']]
    df_raw_kpi_pre['band2']=df_raw_kpi_pre['band2'].replace(['F1','F3','F8'],['L2100','L1800','L900'])
    df_raw_kpi_pre['CELL_ID']=[site.split('_')[-2] if '_' in str(site) else str(site) for site in df_raw_kpi_pre['Short name']]


    tech=[]
    for band1 in df_raw_kpi_pre['band2']:
        if('T1' in band1 or 'T2' in band1):
            if('T1' in band1): band='TDDC2'
            if('T2' in band1): band='TDDC1'
            tech.append(band)
        else:
            tech.append(band1)
    df_raw_kpi_pre.insert(1,'tech',tech)
    df_raw_kpi_pre=df_raw_kpi_pre.drop('band2',axis=1)

    # ---------- DISMANTLE BASED FILTER PRE ----------
    degrow_site_list["DismantleDate"] = pd.to_datetime(degrow_site_list["DismantleDate"], errors="coerce")

    pre_collect = []
    for _, r in degrow_site_list.iterrows():
        site = r["Site_ID"]
        d_date = r["DismantleDate"]

        pre5, _ = get_pre_post_5days(df_raw_kpi_pre, site, d_date)

        if not pre5.empty:
            pre_collect.append(pre5)

    filter_pre = pd.concat(pre_collect, ignore_index=True) if pre_collect else pd.DataFrame()


    filter_pre = keep_only_5days_per_cell(filter_pre, mode="pre")
    pivot_pre = filter_pre.pivot_table(
        index="Date",
        columns=['Short name','tech','SITE_ID','CELL_ID','4G_ECGI'],
        aggfunc={
            'MV_4G Data Volume_GB':'sum',
            "VoLTE Traffic":'sum',
            'DL User Throughput_Mbps':'mean',
            'RRC Setup Success Rate':'mean',
            'VoLTE DCR':'mean',
            'VoLTE ERAB Setup Success Rate':'mean',
            'ERAB Setup Success Rate [CDBH]':'mean',
            'UL User Throughput_Kbps':'mean'
        }
    )

    rounded_df_pre = pivot_pre.T.round(2)
    rounded_df_pre = force_only_5_dates_after_pivot(rounded_df_pre, "pre")

    # ---------- POST PROCESS ----------
    df_raw_kpi_post['Short name']=df_raw_kpi_post['Short name'].fillna(method='ffill')
    df_raw_kpi_post = df_raw_kpi_post[
    ~df_raw_kpi_post['Short name'].str.startswith('@Nokia', na=False)
    ]

    df_raw_kpi_post["DL User Throughput_Kbps"]=(df_raw_kpi_post["DL User Throughput_Kbps"]/1024)
    df_raw_kpi_post.rename(columns={"DL User Throughput_Kbps":"DL User Throughput_Mbps"},inplace=True)
    df_raw_kpi_post.fillna(value=0,inplace=True)                     
    #df.fillna(method='ffill',inplace=true)
    df_raw_kpi_post.columns.values[1]='Date'
    df_raw_kpi_post['band2']=[band.split('_')[2] if '_' in str(band) else str(band) for band in df_raw_kpi_post['Short name']]
    df_raw_kpi_post['SITE_ID']=[site.split('_')[-2][:-1] if '_' in str(site) else str(site)[:-1] for site in df_raw_kpi_post['Short name']]
    df_raw_kpi_post['band2']=df_raw_kpi_post['band2'].replace(['F1','F3','F8'],['L2100','L1800','L900'])
    df_raw_kpi_post['CELL_ID']=[site.split('_')[-2] if '_' in str(site) else str(site) for site in df_raw_kpi_post['Short name']]

    tech=[]
    for band1 in df_raw_kpi_post['band2']:
        if('T1' in band1 or 'T2' in band1):
            if('T1' in band1):
                band='TDDC2'
            if('T2' in band1):
                band='TDDC1'
            tech.append(band) 
        else:
            band=band1
            tech.append(band)
    df_raw_kpi_post.insert(1,'tech',tech) 

    df_raw_kpi_post=df_raw_kpi_post.drop('band2',axis=1)############### drop ######
    post_os_path=os.path.join(door_path,"process output","desired_post.xlsx")
    df_raw_kpi_post.to_excel(post_os_path,index=False)

    #     ######################## filter ##################################


    filter_post=df_raw_kpi_post[(df_raw_kpi_post.SITE_ID.isin(list(degrow_site_list['Site_ID'])))]
    post_os_filter_path=os.path.join(door_path,'process output','filter_post.xlsx')
    filter_post.to_excel(post_os_filter_path,index=False)
    # print("_________________filter_______",filter_post)

    #     ###################################################################################################################

    # pivot_post_vol=filter_post.pivot_table(index="Date",columns=['Short name','tech','SITE_ID','CELL_ID','4G_ECGI'],aggfunc={'MV_4G Data Volume_GB':'sum',"VoLTE Traffic":'sum','MV_DL User Throughput_Mbps':'mean',
    #                                                                             'RRC Setup Success Rate':'mean','VoLTE DCR':'mean','VoLTE ERAB Setup Success Rate':'mean',
    #                                                                             "MV_Avg Connected User":"mean",'ERAB Setup Success Rate [CDBH]':'mean','MV_UL User Throughput_Kbps':'mean'})
    pivot_post_vol=filter_post.pivot_table(index="Date",columns=['Short name','tech','SITE_ID','CELL_ID','4G_ECGI'],aggfunc={'MV_4G Data Volume_GB':'sum',"VoLTE Traffic":'sum','DL User Throughput_Mbps':'mean',
                                                                                'RRC Setup Success Rate':'mean','VoLTE DCR':'mean','VoLTE ERAB Setup Success Rate':'mean',
                                                                            'ERAB Setup Success Rate [CDBH]':'mean','UL User Throughput_Kbps':'mean'})

    pivot_post=pivot_post_vol.T
    rounded_df_post=pivot_post.round(2)
    # rounded_df_post['Pre Avg']=round(rounded_df_pre.mean(axis=1),2)   
    post_os_pivot_path=os.path.join(door_path,"process output","post_pivoted.xlsx")
    rounded_df_post.to_excel(post_os_pivot_path)
    rounded_df_post=force_only_5_dates_after_pivot(rounded_df_post, "post")


    concat_pre_post = pd.concat([rounded_df_pre, rounded_df_post],axis=1)
    concat_pre_post.fillna(value=0,inplace=True)

    pre_cols  = rounded_df_pre.columns
    post_cols = rounded_df_post.columns

    concat_pre_post["Pre Avg"]  = round(concat_pre_post[pre_cols].mean(axis=1),2)
    concat_pre_post["Post Avg"] = round(concat_pre_post[post_cols].mean(axis=1),2)
    concat_pre_post['Delta']=round(concat_pre_post['Post Avg']-concat_pre_post['Pre Avg'],2)
    concat_pre_post['Change%']=round(concat_pre_post['Delta']/concat_pre_post['Pre Avg'],2)

    concat_pre_post.replace([np.inf, -np.inf], 0, inplace=True)
    concat_pre_post['Change%'].fillna(0,inplace=True)

    concat_pre_post_os_path=os.path.join(door_path,"process output","concat_pre_post.xlsx")

    concat_pre_post.to_excel( concat_pre_post_os_path)

    df=pd.read_excel(concat_pre_post_os_path)
#######################################################################
    df.rename(columns={'Unnamed: 0':'KPI'}, inplace=True)
    df.rename(columns={'Short name':'Row Labels'},inplace=True)
    df.rename(columns={'tech':'technology'},inplace=True)
    df['KPI']=df['KPI'].ffill()
    df.set_index(['KPI','Row Labels'])
    df_os_ws_path=os.path.join(door_path,"process output","concat_pre_post2.xlsx")
    df.to_excel(df_os_ws_path,index=False,header=True,startrow=1)####### TO CHIFT A HEADER USE START ROW ###


    df1=df_os_ws_path

    wb=openpyxl.load_workbook(df1)
    ws=wb.active
############################### font size################################
    font=Font(size=9)
    alignment = Alignment(horizontal='center', vertical='center')
    for col_cells in ws.columns:
        for cell in col_cells:
            cell.font=font
            cell.alignment=alignment 
###################################### merge cell ##############################
    ws.merge_cells('G1:K1')
    ws.merge_cells('L1:P1')  
############################# AFTER MERGING PUT COLUMNS NAME ################
    ws['G1']='PRE'  
    ws['L1']='POST'
################################ IN CENTER HEADER ###########################
    merged_cell=ws['G1']
    merged_cell1=ws['L1']  
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell1.alignment = Alignment(horizontal='center', vertical='center')
# ############################ to change row wise color change ####################################
    row_number=2
    # fillcolor='FFFFFF'

    for cell in ws[row_number]:
        cell.font = Font(bold=True,size=9)
############### FOR COLORING ###############
    YELLOW="FFE699"
    LIGHT_GREEN="C6E0B4"
    LIGHT_GREY="D9E1F2"
    PURPLE="7030A0"
    CYAN="00B0F0"
    ORANGE="FFC000"
    light_blue='BDD7EE'

    ws["G1"].fill=PatternFill(patternType='solid',fgColor=YELLOW)
    ws['L1'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREEN)
    ws['A2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['F2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['G2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['H2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['I2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['J2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['K2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['L2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['M2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['N2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['O2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['B2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['P2'].fill=PatternFill(patternType='solid',fgColor=ORANGE)
    ws['Q2'].fill=PatternFill(patternType='solid',fgColor=CYAN)
    ws['R2'].fill=PatternFill(patternType='solid',fgColor=PURPLE)
    ws['S2'].fill=PatternFill(patternType='solid',fgColor=light_blue)


    


#  ############################################# till your row value it will be border thin  not to fix #####################
    thin_border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border 

############################## column changing ###########################################
    # source_column_index = 5  # Replace with the source column index
    # target_column_index = 0  # Replace with the target column index

    # source_column_values = []

    # for row in ws.iter_rows(min_row=6, values_only=True):
    #     source_value = row[source_column_index]
    #     source_column_values.append(source_value)
    # for index, row in enumerate(ws.iter_rows(min_row=6), start=6):
    #     target_cell = row[target_column_index]
    #     target_cell.value = source_column_values[index - 6]     

    # ws.delete_cols(source_column_index + 1)
    save_output=os.path.join(door_path,'output','BIH_DEGROW_OUTPUT.xlsx')

    wb.save(save_output)  

    download_path=os.path.join(MEDIA_URL,"trends","Bihar","output","BIH_DEGROW_OUTPUT.xlsx")
    return Response({"status":True,"message":"uploaded successfully","Download_url":download_path}) 






   



#Bihar Degrow new-----------------------
@api_view(["POST"])
def Bihar_degrow(request):
    pre_file= request.FILES.get("pre_file")
    if not pre_file:
        return Response({"message": "pre file not uploaded"}, status=HTTP_400_BAD_REQUEST)
    post_file=request.FILES.get("post_file")
    if not post_file:
        return Response({"message": "post file not uploaded"}, status=HTTP_400_BAD_REQUEST)
    
    site_list=request.FILES.get("site_list")
    if not site_list:
        return Response({"message": "sitelist not uploaded"}, status=HTTP_400_BAD_REQUEST)
    

    base_media_url = os.path.join(MEDIA_ROOT, "Bihar_Degrow")
    output_path = os.path.join(base_media_url, "Final_Output")
    input_folder = os.path.join(base_media_url, "Process_output")


    os.makedirs(input_folder, exist_ok=True)
    os.makedirs(output_path, exist_ok=True)
  
    

#__________________________
    if pre_file.name.lower().endswith(".csv"):
        df_pre = pd.read_csv(pre_file)
    else:
        df_pre = pd.read_excel(pre_file)

    df_pre.columns = df_pre.columns.str.strip()

    


    if post_file.name.lower().endswith(".csv"):
        df_post= pd.read_csv(post_file)
    else:
        df_post= pd.read_excel(post_file)
    df_post.columns = df_post.columns.str.strip()    


    site_df = pd.read_excel(site_list)
    site_df.columns = site_df.columns.str.strip()
   
  #processs for pre file---------  
    df_pre['Short name']=df_pre['Short name'].ffill()
    df_pre.rename(columns={"Unnamed: 1": "Date"}, inplace=True)
    df_pre['Date'] = pd.to_datetime(df_pre['Date'])
    df_pre = df_pre[df_pre['Date'].notna()]
    df_pre = df_pre.sort_values(["Short name", "Date"])
    df_pre = df_pre.groupby("Short name").tail(5)
    # df_pre.to_excel("date5_pre.xlsx")

    df_pre["DL User Throughput_Kbps [CDBH]"]=(df_pre["DL User Throughput_Kbps [CDBH]"]/1024)
    df_pre.rename(columns={"DL User Throughput_Kbps [CDBH]":"DL User Throughput_Mbps [CDBH]"},inplace=True)
    for col in df_pre.columns:
        if df_pre[col].dtype in ['object', 'string']:
            df_pre[col] = df_pre[col].fillna("")
        else:
            df_pre[col] = df_pre[col].fillna(0)  

  
    df_pre['band2']=[ band.split('_')[2] if '_' in band else band for band in df_pre['Short name']]
    df_pre['SITE_ID']=[site.split('_')[-2][:-1] if '_' in site else site for site in df_pre['Short name']]
    df_pre['CELL_ID']=[site.split('_')[-2] if '_' in str(site) else str(site) for site in df_pre['Short name']]
    df_pre['band2']=df_pre['band2'].replace(['F1','F3','F8'],['L2100','L1800','L900'])

    df_pre['tech'] = df_pre['band2'].astype(str).str.upper().apply(
    lambda x: 'TDDC2' if 'T1' in x else ('TDDC1' if 'T2' in x else x)
)
    df_pre=df_pre.drop('band2',axis=1)
# pre filter with site id---- and make pivot
    filter_pre = df_pre[df_pre["SITE_ID"].isin(site_df["Site_ID"])]
    # print(filter_pre)
    pivot_pre=filter_pre.pivot_table(index="Date",columns=['Short name','tech','SITE_ID','CELL_ID','4G_ECGI'],aggfunc={'MV_4G Data Volume_GB':'sum',"VoLTE Traffic_24Hrs":'sum','DL User Throughput_Mbps [CDBH]':'mean',
                                                                                'RRC Setup Success Rate [CDBH]':'mean','VoLTE DCR [CBBH]':'mean','ERAB Setup Success Rate [CDBH]':'mean',
                                                                            'DL PRB Utilization [CDBH]':'mean','UL User Throughput_Kbps [CUBH]':'mean'})
    pivot_pre=pivot_pre.T
    rounded_df_pre=pivot_pre.round(2)
    pre_file_name = "pre_pivoted.xlsx"
    pre_file_path = os.path.join(input_folder,pre_file_name)
    rounded_df_pre.to_excel(pre_file_path)
 

 #process for post data----------
    df_post['Short name']=df_post['Short name'].ffill()
    df_post.rename(columns={"Unnamed: 1": "Date"}, inplace=True)
    df_post['Date'] = pd.to_datetime(df_post['Date'])
    df_post = df_post[df_post['Date'].notna()]
    df_post = df_post.sort_values(["Short name", "Date"])
    df_post = df_post.groupby("Short name").tail(5)

    df_post["DL User Throughput_Kbps [CDBH]"]=(df_post["DL User Throughput_Kbps [CDBH]"]/1024)
    df_post.rename(columns={"DL User Throughput_Kbps [CDBH]":"DL User Throughput_Mbps [CDBH]"},inplace=True)
    for col in df_post.columns:
     if df_post[col].dtype in ['object', 'string']:
        df_post[col] = df_post[col].fillna("")
     else:
        df_post[col] = df_post[col].fillna(0) 

    df_post['band2']=[band.split('_')[2] if '_' in str(band) else str(band) for band in df_post['Short name']]
    df_post['SITE_ID']=[site.split('_')[-2][:-1] if '_' in str(site) else str(site)[:-1] for site in df_post['Short name']]
    df_post['band2']=df_post['band2'].replace(['F1','F3','F8'],['L2100','L1800','L900'])
    df_post['CELL_ID']=[site.split('_')[-2] if '_' in str(site) else str(site) for site in df_post['Short name']]
    df_post['tech'] = df_post['band2'].astype(str).str.upper().apply(
     lambda x: 'TDDC2' if 'T1' in x else ('TDDC1' if 'T2' in x else x))
    df_post=df_post.drop('band2',axis=1)
   
# pre filter with site id---- and make pivot
    filter_post =df_post[df_post["SITE_ID"].isin(site_df["Site_ID"])]
    pivot_post_vol=filter_post.pivot_table(index="Date",columns=['Short name','tech','SITE_ID','CELL_ID','4G_ECGI'],aggfunc={'MV_4G Data Volume_GB':'sum',"VoLTE Traffic_24Hrs":'sum','DL User Throughput_Mbps [CDBH]':'mean',
                                                                                'RRC Setup Success Rate [CDBH]':'mean','VoLTE DCR [CBBH]':'mean','ERAB Setup Success Rate [CDBH]':'mean',
                                                                               'DL PRB Utilization [CDBH]':'mean','UL User Throughput_Kbps [CUBH]':'mean'})
    
    pivot_post=pivot_post_vol.T
    rounded_df_post=pivot_post.round(2)
    post_file_name = "post_pivoted.xlsx"
    post_file_path = os.path.join(input_folder,post_file_name)
    rounded_df_post.to_excel(post_file_path)
    

    #concat to pre-post
    concat_pre_post = pd.concat([rounded_df_pre, rounded_df_post], axis=1)
    concat_pre_post.fillna(0, inplace=True)

    pre_cols = concat_pre_post.columns[:5]
    post_cols = concat_pre_post.columns[5:10]

    concat_pre_post["Pre Avg"] = concat_pre_post[pre_cols].mean(axis=1).round(2)
    concat_pre_post["Post Avg"] = concat_pre_post[post_cols].mean(axis=1).round(2)

    concat_pre_post["Delta"] = (concat_pre_post["Post Avg"] - concat_pre_post["Pre Avg"]).round(2)

    concat_pre_post["Change%"] = np.where(
        concat_pre_post["Pre Avg"] == 0,
        0,
        (concat_pre_post["Delta"] / concat_pre_post["Pre Avg"]).round(2)
    )

    concat_pre_post.replace([np.inf, -np.inf], 0, inplace=True)
    concat_pre_post["Change%"].fillna(0, inplace=True)
    concat_pre_post.reset_index(inplace=True)
    concat_pre_post.rename(columns={
    'level_0': 'KPI',
    'Short name': 'Row Labels',
    'tech': 'Technology'
    }, inplace=True)
    concat_pre_post.set_index(['KPI', 'Row Labels'], inplace=True)
    file_name = "priovt_concate_pre_post.xlsx"
    file_path = os.path.join(input_folder, file_name)
    concat_pre_post.to_excel(file_path)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    font = Font(size=9,bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    merged_ranges = list(ws.merged_cells.ranges)

    for m in merged_ranges:
        ws.unmerge_cells(str(m))
    ws.insert_rows(1)


    for m in merged_ranges:
        min_col, min_row, max_col, max_row = m.bounds
        ws.merge_cells(
            start_row=min_row + 1,
            start_column=min_col,
            end_row=max_row + 1,
            end_column=max_col
        )

    ws.merge_cells('G1:K1')
    ws['G1'] = "PRE"

    ws.merge_cells('L1:P1')
    ws['L1'] = "POST"


    for cell in ['G1', 'L1']:
        ws[cell].font = Font(bold=True, size=11)
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')

 
    pre_fill = PatternFill(start_color="C4BD97", end_color="C4BD97", fill_type="solid")
    post_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")

    for col in range(7, 12):   # G-K
        ws.cell(row=1, column=col).fill = pre_fill

    for col in range(12, 17):  # L-P
        ws.cell(row=1, column=col).fill = post_fill


    header_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")

    for col in range(1, 7):  # KPI to ECGI
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill


    date_fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")

    for col in range(7, 17):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = date_fill
 

    back_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    for col in range(17,21):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True, size=9) 
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = back_fill    



    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(cell.column)

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_len + 2


    kpi_col = column_index_from_string("A")
    merged_ranges = list(ws.merged_cells.ranges)

    for m in merged_ranges:
        min_col, min_row, max_col, max_row = m.bounds
        if min_col == kpi_col and max_col == kpi_col:
            kpi_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(m))
            for row in range(min_row, max_row + 1):
                cell = ws.cell(row=row, column=kpi_col)
                cell.value = kpi_value
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')


    file_name = "BIH_DEGROW_OUTPUT.xlsx"
    final_file_path = os.path.join(output_path, file_name)

    # Save workbook
    wb.save(final_file_path)
    

  
    download_url = request.build_absolute_uri(
    os.path.join(MEDIA_URL, "Bihar_Degrow", "Final_Output", file_name).replace("\\", "/")
)
   


    print("End of Process---")
    return Response({
        "status":True,
        "message": "Files processed successfully",
        "download_url":download_url,
        }, status=HTTP_200_OK)












