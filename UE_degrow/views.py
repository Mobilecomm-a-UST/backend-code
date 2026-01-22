from openpyxl.styles import Alignment
from tkinter import filedialog, messagebox, ttk
from openpyxl import workbook,load_workbook

import pandas as pd
import numpy as np
from datetime import date, timedelta
from datetime import datetime
from openpyxl import Workbook
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.core.files.storage import FileSystemStorage
from datetime import date
import glob
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
import os

#################################################################################################

########################################## NOTE MISSING F1 TECHNOLOGY ###########################################
@api_view(['POST'])
def UE_degrow(request):
    pre_kpi=request.FILES['PRE'] if 'PRE' in request.FILES else None
    print("______________pre kpi______--",pre_kpi)
    if pre_kpi:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        print("___________location________",location)
        fs=FileSystemStorage(location=location)
        file=fs.save(pre_kpi.name,pre_kpi)
        file_path=fs.path(file)
        df_pre=pd.read_csv(file_path)
        print(df_pre)
        os.remove(path=file_path)
    post_kpi=request.FILES['POST'] if 'POST' in request.FILES else None
    print("______________post kpi______",post_kpi)
    if post_kpi:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        print("___________location________",location)
        fs=FileSystemStorage(location=location)
        file=fs.save(post_kpi.name,post_kpi)
        file_path=fs.path(file)
        df_post=pd.read_csv(file_path)
        print(df_post)
        os.remove(path=file_path)
    site_list=request.FILES['site'] if 'site' in request.FILES else None
    print("______________site______--",site_list)
    if site_list:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        print("___________location________",location)
        fs=FileSystemStorage(location=location)
        file_sheet=fs.save(site_list.name,site_list)
        file_path=fs.path(file_sheet)
        site_list=pd.read_excel(file_path)
        print(site_list)
        os.remove(path=file_path)


    # df_pre=pd.read_csv("actual input/Pre Post/Pre Post Part 9/Pre 6 JUL TO 10 JUL.csv")
    door_path=os.path.join(MEDIA_ROOT,'trends','UEdegrow')

    df_pre['Short name']=df_pre['Short name'].fillna(method='ffill')
    print(df_pre['Short name'].dtype,"--------hhhhhh------------------------------------")
    df_pre["DL User Throughput_Kbps [CDBH]"]=(df_pre["DL User Throughput_Kbps [CDBH]"]/1024)
    df_pre.rename(columns={"DL User Throughput_Kbps [CDBH]":"DL User Throughput_Mbps [CDBH]"},inplace=True)
    #df.fillna(method='ffill',inplace=true)
    df_pre.fillna(value=0,inplace=True)                     
    # df_pre.columns.values[1]='Date'
    df_pre['band2']=[ band.split('_')[2] for band in df_pre['Short name']]
    df_pre['SITE_ID']=[site.split('_')[-2][:-1] for site in df_pre['Short name']]
    df_pre['CELL_ID']=[site.split('_')[-2] if '_' in str(site) else str(site) for site in df_pre['Short name']]


    df_pre['band2']=df_pre['band2'].replace(['F1','F3','F8'],['L2100','L1800','L900'])
    # df_pre['band2']=df_pre['band2'].fillna(method='ffill')

    tech=[]
    for band1 in df_pre['band2']:
        if('T1' in band1 or 'T2' in band1):
            if('T1' in band1):
                band='TDDC2'
            if('T2' in band1):
                band='TDDC1'
            tech.append(band) 
        else:
            band=band1
            tech.append(band)
    df_pre.insert(1,'tech',tech) 

    df_pre=df_pre.drop('band2',axis=1)############### drop ######
    
    Pre_path1=os.path.join(door_path,'process output','desired_pre.xlsx')
    df_pre.to_excel(Pre_path1,index=False)
    

    filter_pre=df_pre[(df_pre.CELL_ID.isin(list(site_list['2G ID'])))]
    Pre_path2=os.path.join(door_path,'process output','filter_pre.xlsx')
    filter_pre.to_excel(Pre_path2,index=False)
   

    pivot_pre=filter_pre.pivot_table(index="Date",columns=['Short name','tech','SITE_ID','CELL_ID'],aggfunc={'MV_4G Data Volume_GB':'sum',"VoLTE Traffic [CBBH]":'sum','DL User Throughput_Mbps [CDBH]':'mean',
                                                                                'RRC Setup Success Rate [CDBH]':'mean','ERAB Setup Success Rate [CDBH]':'mean',
                                                                                "Avg_RRC_Connected User":"mean",'DL PRB Utilization [CDBH]':'mean','UL User Throughput_Kbps [CUBH]':'mean','E-UTRAN Average CQI [CDBH]':'sum','Average UE Distance_KM':'sum'})
    pivot_pre=pivot_pre.T
    pre_path3=os.path.join(door_path,'process output','pivoted_pre.xlsx')
    pivot_pre.to_excel(pre_path3)

    rounded_df_pre=pivot_pre.round(2)
###################################### POST #####################################

    df_post['Short name']=df_post['Short name'].fillna(method='ffill')
    df_post["DL User Throughput_Kbps [CDBH]"]=(df_post["DL User Throughput_Kbps [CDBH]"]/1024)
    df_post.rename(columns={"DL User Throughput_Kbps [CDBH]":"DL User Throughput_Mbps [CDBH]"},inplace=True)
    df_post.fillna(value=0,inplace=True)                     
    #df.fillna(method='ffill',inplace=true)
    # df_post.columns.values[1]='Date'
    df_post['band2']=[band.split('_')[2] if '_' in str(band) else str(band) for band in df_post['Short name']]
    df_post['SITE_ID']=[site.split('_')[-2][:-1] if '_' in str(site) else str(site)[:-1] for site in df_post['Short name']]
    df_post['band2']=df_post['band2'].replace(['F1','F3','F8'],['L2100','L1800','L900'])
    df_post['CELL_ID']=[site.split('_')[-2] if '_' in str(site) else str(site) for site in df_post['Short name']]

    tech=[]
    for band1 in df_post['band2']:
        if('T1' in band1 or 'T2' in band1):
            if('T1' in band1):
                band='TDDC2'
            if('T2' in band1):
                band='TDDC1'
            tech.append(band) 
        else:
            band=band1
            tech.append(band)
    df_post.insert(1,'tech',tech) 

    df_post=df_post.drop('band2',axis=1)############### drop ######

    post_path1=os.path.join(door_path,'process output','desired_post.xlsx')
    df_post.to_excel(post_path1,index=False)

#     ######################## filter ##################################
    filter_post=df_post[(df_post.CELL_ID.isin(list(site_list['2G ID'])))]
    post_path2=os.path.join(door_path,'process output','filter_post.xlsx')
    filter_post.to_excel(post_path2,index=False)

#     ###################################################################################################################

    pivot_post_vol=filter_post.pivot_table(index="Date",columns=['Short name','tech','SITE_ID','CELL_ID'],aggfunc={'MV_4G Data Volume_GB':'sum',"VoLTE Traffic [CBBH]":'sum','DL User Throughput_Mbps [CDBH]':'mean',
                                                                                'RRC Setup Success Rate [CDBH]':'mean','ERAB Setup Success Rate [CDBH]':'mean',
                                                                                "Avg_RRC_Connected User":"mean",'DL PRB Utilization [CDBH]':'mean','UL User Throughput_Kbps [CUBH]':'mean','E-UTRAN Average CQI [CDBH]':'sum','Average UE Distance_KM':'sum'})
    pivot_post=pivot_post_vol.T
    rounded_df_post=pivot_post.round(2)
    # rounded_df_post['Pre Avg']=round(rounded_df_pre.mean(axis=1),2)   
    post_path3=os.path.join(door_path,'process output','pivoted_post.xlsx')

    print("________done__________________",post_path3)
    rounded_df_post.to_excel(post_path3)
    

######################################### concat ##########################
    concat_pre_post = pd.concat([rounded_df_pre, rounded_df_post],axis=1) 
    concat_pre_post.fillna(value=0,inplace=True) 
    concat_pre_post["Pre Avg"]=round(concat_pre_post.iloc[:,[0,1,2,3,4]].mean(axis=1),2)
    concat_pre_post['Post Avg']=round(concat_pre_post.iloc[:,[5,6,7,8,9]].mean(axis=1),2)############
    concat_pre_post['Delta']=round(concat_pre_post['Post Avg']-concat_pre_post['Pre Avg'],2)
    concat_pre_post['Change%']=round(concat_pre_post['Delta']/concat_pre_post['Pre Avg'],2)
     
    # concat_pre_post.replace([np.inf, -np.inf], 'NA', inplace=True)
    concat_pre_post['Change%'].fillna(0,inplace=True)




    concat_path=os.path.join(door_path,'process output','concat_pre_post.xlsx')
    concat_pre_post.to_excel(concat_path)

    df=pd.read_excel(concat_path)
#######################################################################
    df.rename(columns={'Unnamed: 0':'KPI'}, inplace=True)
    df.rename(columns={'Short name':'Row Labels'},inplace=True)
    df.rename(columns={'tech':'technology'},inplace=True)
    df.set_index(['KPI','Row Labels'])
    
    concat_path2=os.path.join(door_path,'process output','concat_pre_post2.xlsx')
    df.to_excel(concat_path2,index=False,header=True,startrow=1)####### TO CHIFT A HEADER USE START ROW ###

    df1=concat_path2

    wb=openpyxl.load_workbook(df1)
    ws=wb.active
############################### font size ################################
    font=Font(size=9)
    alignment = Alignment(horizontal='center', vertical='center')
    for col_cells in ws.columns:
        for cell in col_cells:
            cell.font=font
            cell.alignment=alignment 
###################################### merge cell ##############################
    ws.merge_cells('F1:J1')
    ws.merge_cells('K1:O1')  
############################# AFTER MERGING PUT COLUMNS NAME ################
    ws['F1']='PRE'  
    ws['K1']='POST'
################################ IN CENTER HEADER ###########################
    merged_cell=ws['F1']
    merged_cell1=ws['K1']  
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell1.alignment = Alignment(horizontal='center', vertical='center')
# ############################ to change row wise color change ####################################
    row_number=2
    # fillcolor='FFFFFF'

    for cell in ws[row_number]:
        cell.font = Font(bold=True,size=9)
############### FOR COLORING ###############
    YELLOW="FFE699"
    LIGHT_GREEN="C6E0B4"
    LIGHT_GREY="D9E1F2"
    PURPLE="7030A0"
    CYAN="00B0F0"
    ORANGE="FFC000"
    light_blue='BDD7EE'

    ws["F1"].fill=PatternFill(patternType='solid',fgColor=YELLOW)
    ws['K1'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREEN)
    ws['A2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['F2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['G2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['H2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['I2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['J2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['K2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['L2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['M2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['N2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['O2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['B2'].fill=PatternFill(patternType='solid',fgColor=LIGHT_GREY)
    ws['P2'].fill=PatternFill(patternType='solid',fgColor=ORANGE)
    ws['Q2'].fill=PatternFill(patternType='solid',fgColor=CYAN)
    ws['R2'].fill=PatternFill(patternType='solid',fgColor=PURPLE)
    ws['S2'].fill=PatternFill(patternType='solid',fgColor=light_blue)


    


#  ############################################# till your row value it will be border thin  not to fix #####################
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border 

############################## column changing ###########################################

    saveoutput=os.path.join(door_path,'output','Degrow_output.xlsx')
    wb.save(saveoutput) 
    download_path=os.path.join(MEDIA_URL,'trends','UEdegrow','output','Degrow_output.xlsx')
    return Response({"Status":True,"Message":"Successfully",'download_url':download_path}) 




