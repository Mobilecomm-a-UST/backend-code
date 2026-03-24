import pandas as pd
import os
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
import xml.etree.ElementTree as ET
import re
import math
import gzip
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import Counter
from datetime import datetime


def convert_int(val):
    s = str(val).strip()
    if re.fullmatch(r"-?\d+", s):
        return int(s)
    else:
        return val


def in_excel(df, writer, sheet_name):
    """
    Save DataFrame to Excel with chunking if rows exceed Excel limit.
    """
    max_rows = 1_048_000
    total_rows = len(df)
    num_sheets = math.ceil(total_rows / max_rows)

    start = 0
    for i in range(num_sheets):
        end = start + max_rows
        chunk_df = df.iloc[start:end]

        if num_sheets == 1:
            chunk_df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            chunk_df.to_excel(writer, sheet_name=f"{sheet_name}_{i+1}", index=False)

        start = end

main_folder = os.path.join(MEDIA_ROOT, "Xml_parsar_script")

@api_view(["POST"])
def xml_bulk_to_excel(request):
    try:

        xml_files = request.FILES.getlist("xml_files")

        if not xml_files:
            return Response({"error": "Please upload XML files"}, status=HTTP_400_BAD_REQUEST)

        
        output_folder = os.path.join(main_folder, "Parsed_Output")
        os.makedirs(output_folder, exist_ok=True)

        parse_data = []
        xnlink_data = []

        for file in xml_files:

            try:

                file_name = file.name.lower()

                if file_name.endswith(".gz"):
                    xml_bytes = gzip.decompress(file.read())
                else:
                    xml_bytes = file.read()

                root = ET.fromstring(xml_bytes)

                m = re.match(r"\{(.*)\}", root.tag)
                ns_url = m.group(1) if m else root.attrib.get("xmlns", "")
                ns = {"ns": ns_url} if ns_url else {}

                managed_objects = root.findall(".//ns:managedObject", ns) if ns else []

                if not managed_objects:
                    managed_objects = root.findall(".//managedObject")

                for mo in managed_objects:

                    record = {
                        "File_Name": file.name,
                        "SW_Version": mo.get("version"),
                        "distName": mo.get("distName", ""),
                        "id": mo.get("id", "")
                    }

                    mo_class = mo.get("class", "")
                    distname = mo.get("distName", "")

                    p_tags = mo.findall("ns:p", ns) if ns else []

                    if not p_tags:
                        p_tags = mo.findall("p")

                    for p in p_tags:
                        key = p.get("name")
                        value = (p.text or "").strip()
                        record[key] = value

                    # ---- XNLINK condition ----
                    if mo_class == "XNLINK" or "XNLINK-" in distname:
                        xnlink_data.append(record)
                    else:
                        parse_data.append(record)

            except Exception as e:

                parse_data.append({
                    "File_Name": file.name,
                    "Error": f"Error parsing XML: {str(e)}"
                })

        # ---------------- LNADJGNB DATA ---------------- #

        dump_df = pd.DataFrame(parse_data)

        if not dump_df.empty:

            dump_df[["MRBTS", "LNBTS", "LNADJGNB"]] = dump_df["distName"].str.extract(
                r"MRBTS-*(\d+).*LNBTS-*(\d+).*LNADJGNB-*(\d+)", expand=True
            )

            priority_cols = ["MRBTS", "LNBTS", "LNADJGNB", "id"]
            remaining_cols = [c for c in dump_df.columns if c not in priority_cols]

            dump_df = dump_df[priority_cols + remaining_cols]
            cols_to_int = ['id','adjGnbId','adjGnbIdLength','administrativeState',                                
                       'cPlaneIpAddrCtrl','mcc','mnc','mncLength',                                
                       'x2ToGnbLinkStatus','lbpsblockedNrAct','MRBTS','LNBTS','LNADJGNB'
                       ]      
        for col in cols_to_int:            
            dump_df[col] = dump_df[col].apply(convert_int)
           
     
        # print (dump_df.columns)
       
        excel_columns = [
            "MRBTS", "LNBTS", "LNADJGNB", "id", "File_Name", "SW_Version", "distName",
            "adjGnbId", "adjGnbIdLength", "administrativeState", "cPlaneIpAddr",
            "cPlaneIpAddrCtrl", "mcc", "mnc", "mncLength", "x2ToGnbLinkStatus",
            "lbpsblockedNrAct","xnLinkStatus"
        ]
       
        for col in excel_columns:
            if col not in dump_df.columns:
                dump_df[col] = ""
 
     
        dump_df = dump_df[excel_columns]
        print(dump_df.tail())

        # ---------------- XNLINK DATA ---------------- #

        xnlink_df = pd.DataFrame(xnlink_data)

        if not xnlink_df.empty:

            xnlink_df[["MRBTS", "NRBTS", "XNLINK"]] = xnlink_df["distName"].str.extract(
                r"MRBTS-(\d+).*NRBTS-(\d+).*XNLINK-(\d+)", expand=True
            )

            priority_cols = ["MRBTS", "NRBTS", "XNLINK", "id"]
            remaining_cols = [c for c in xnlink_df.columns if c not in priority_cols]

            xnlink_df = xnlink_df[priority_cols + remaining_cols]

        output_path = os.path.join(output_folder, "Dump_Parsed_Output.xlsx")

        with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:

            if not dump_df.empty:
                in_excel(dump_df, writer, "LNADJGNB")

            if not xnlink_df.empty:
                in_excel(xnlink_df, writer, "XNLINK")

      

        relative_path = os.path.relpath(output_path, MEDIA_ROOT)

        download_url = request.build_absolute_uri(
            os.path.join(MEDIA_URL, relative_path).replace("\\", "/")
        )

        return Response({
            "status": True,
            "message": "All XML files successfully parsed",
            "download_url": download_url
        }, status=HTTP_200_OK)

    except Exception as e:

        return Response({
            "find an error": str(e)
        }, status=HTTP_400_BAD_REQUEST)
    



    # Nokia Alarm NREEL-----



# Function to extract PCI & Count pairs
def extract_pci_count(text):
    if pd.isna(text):
        return []
    
    pattern = r"NR-PCI\s*=\s*(\d+)\s*COUNT\s*=\s*(\d+)"
    matches = re.findall(pattern, text)
    
    return [(int(pci), int(cnt)) for pci, cnt in matches]

def format_and_autofit_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Header formatting
    header_fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    ws.row_dimensions[1].height =20  

    # Data cell formatting
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # bold_font = Font(bold=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            # cell.font = bold_font
            if str(cell.value).strip().upper() == "NA":
                cell.fill = yellow_fill

    # Autofit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # 1-based
        column_letter = get_column_letter(column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(file_path)


@api_view(["POST"])
def nrrel_parse_status(request):
    parse_file=request.FILES.get("file")
    base_name = os.path.splitext(parse_file.name)[0]
    if not parse_file:
        return Response({"error": "Please upload XML files"}, status=HTTP_400_BAD_REQUEST)
    
    nokia_output_folder = os.path.join(main_folder, "Nokia_Alarm_parsed")
    os.makedirs(nokia_output_folder, exist_ok=True)

     
    file_name = parse_file.name.lower()
    if file_name.endswith(".csv"):
        alarm_df = pd.read_csv(parse_file)
    elif file_name.endswith((".xlsx", ".xls")):
        alarm_df = pd.read_excel(parse_file)
    else:
        return Response({"error": "Unsupported file format"}, status=HTTP_400_BAD_REQUEST)

    

    #-------------------------------
    alarm_df= alarm_df[alarm_df["Diagnostic Info"].str.contains("8150 supplAlarmInfo", na=False)]
    if alarm_df.empty:
        return Response({"message": "No 8150 alarm found"}, status=200)
    
    
    alarm_df=alarm_df[["Distinguished Name", "Diagnostic Info"]].copy()
    alarm_df["pci_list"] = alarm_df["Diagnostic Info"].apply(extract_pci_count)
    alarm_df = alarm_df.explode("pci_list")

    alarm_df[["Target_PCI_", "Count"]] = pd.DataFrame(
        alarm_df["pci_list"].tolist(), index=alarm_df.index
    )
    alarm_df = alarm_df[["Distinguished Name", "Target_PCI_", "Count"]].dropna()
    alarm_df=alarm_df.reset_index().rename(columns={"index":"","Target_PCI_":"Target_PCI"})
    print(alarm_df.head())


   
    file_name = f"{base_name}_Output.xlsx"
    output_path = os.path.join(nokia_output_folder, file_name)
    alarm_df.to_excel(output_path, index=False, engine="openpyxl")
    format_and_autofit_excel(output_path)

    relative_path = os.path.relpath(output_path, MEDIA_ROOT).replace("\\", "/")
    download_url = request.build_absolute_uri(MEDIA_URL + relative_path)
    print("End the Process-")
   

    return Response(
        {
            "status": True,
            "message": "Files processed successfully",
            "download_url": download_url,
        }
      
    )


   




    

