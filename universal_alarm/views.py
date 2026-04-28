from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Font, Alignment, Side
from rest_framework.response import Response
from rest_framework.decorators import api_view
import os
import re
import pandas as pd
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from rest_framework import status
import shutil
from datetime import datetime
from .models import *


#function for clean alarm data
Remove_alarm = {"external", "security", "xnc"}
Piu_eternet = re.compile(r"ethernetport=tn_[a-z]_piu", re.IGNORECASE)

def specific_problem(val, mo_val):
    if pd.isna(val):
        return val

    parts = [p.strip() for p in val.split("/")]

    if (
        any(p.lower() == "ethernet" for p in parts)
        and isinstance(mo_val, str)
        and Piu_eternet.search(mo_val)
    ):
        return "No Alarm"

    if all(p.lower() in Remove_alarm for p in parts):
        return "No Alarm"
    cleaned = [p for p in parts if p.lower() not in Remove_alarm]

    if not cleaned:
        return "No Alarm"

    return "/".join(cleaned)

 
def format_excel_sheet(writer, sheet_name, df, startrow=0, startcol=0):
    """Apply formatting to an Excel sheet with adjustable start positions."""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
 
    header_format = workbook.add_format(
        {
            "bold": True,
            "bg_color": "#000957",
            "border": 2,
            "font_color": "#ffffff",
            "align": "center",
            "valign": "vcenter",
        }
    )
    center_format = workbook.add_format(
        {
            "align": "center",
            "valign": "center",
            "border": 0,
            "border_color": "#000000",
        }
    )
    ok_format = workbook.add_format(
        {
            "bg_color": "#90EE90",
            "font_color": "#000000",
            "align": "center",
            "valign": "center",
        }
    )
    not_ok_format = workbook.add_format(
        {
            "bg_color": "#FF0000",
            "font_color": "#FFFFFF",
            "align": "center",
            "valign": "center",
        }
    )

 
    worksheet.set_row(startrow, 23)
 
    for col_num, col_name in enumerate(df.columns):
        worksheet.write(startrow, startcol + col_num, str(col_name), header_format)
 
        column_series = df[col_name]
        if isinstance(column_series, pd.DataFrame):
            column_series = column_series.iloc[:, 0]
 
        max_length = max(
            column_series.fillna("").astype(str).str.len().max(skipna=True) or 0,
            len(str(col_name)),
        )
        max_length = min(max_length, 255)
        worksheet.set_column(startcol + col_num, startcol + col_num, max_length + 5)
 
    for row_num in range(len(df)):
        worksheet.set_row(startrow + row_num + 1, 15)
 
        for col_num in range(len(df.columns)):
            cell_value = str(df.iloc[row_num, col_num])
            format_style = center_format
            if cell_value == "OK":
                format_style = ok_format
            elif cell_value == "NOT OK":
                format_style = not_ok_format
            elif cell_value == "Missing":
                format_style = workbook.add_format(
                    {
                        "bg_color": "#FF6347",
                        "font_color": "#FFFFFF",
                        "align": "center",
                        "valign": "center",
                    }
                )
 
            worksheet.write(
                startrow + row_num + 1, startcol + col_num, cell_value, format_style
            )
 
            writer._save()
 
 
def delete_existing_files(upload_files):
    for file_name in os.listdir(upload_files):
        file_path = os.path.join(upload_files, file_name)
        if os.path.isfile(file_path):
            os.remove(file_path)


@api_view(['GET'])
def get_sites(request):
    data = list(UniversalAlarm.objects.values())
    OldvsNew_data = list(OldvsNew.objects.values())
    return Response({"universal_alarms": data, "old_vs_new": OldvsNew_data})

@api_view(['DELETE'])
def delete_site(request):
    try:
        obj = UniversalAlarm.objects.all()
        obj.delete()
        OldvsNew.objects.all().delete()

        return Response({"message": "Deleted"})
    except UniversalAlarm.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

@api_view(['POST'])
def upload_site_list(request):
    file = request.FILES.get('site_file')

    if not file:
        return Response({"error": "File required"}, status=400)

    if not file.name.endswith('.xlsx'):
        return Response({"error": "Only Excel file allowed"}, status=400)

    try:
        UniversalAlarm.objects.all().delete()
        OldvsNew.objects.all().delete()
        # Read all sheets
        sheets = pd.read_excel(file, sheet_name=None)

        # =========================
        # Sheet1 = UniversalAlarm
        # =========================
        alarm_df = sheets.get("Sheet 1")

        if alarm_df is not None:
            alarm_df.columns = alarm_df.columns.str.strip().str.lower()

            for _, row in alarm_df.iterrows():

                UniversalAlarm.objects.update_or_create(
                    site_id=str(row.get('site id')).strip(),
                    defaults={
                        "mplane_ip": str(row.get('mplane ip')).strip(),
                        "status": str(row.get('status')).strip()
                    }
                )

        # ======================
        # Sheet2 = OldvsNew
        # ======================
        map_df = sheets.get("Sheet 2")

        if map_df is not None:
            map_df.columns = map_df.columns.str.strip().str.lower()

            for _, row in map_df.iterrows():

                OldvsNew.objects.update_or_create(
                    new_site=str(row.get('newsite')).strip(),
                    defaults={
                        "old_site": str(row.get('oldsite')).strip()
                    }
                )

        return Response({
            "status": True,
            "message": "Sheet1 + Sheet2 uploaded successfully"
        })

    except Exception as e:
        return Response({
            "status": False,
            "error": str(e)
        }, status=500)
    
# @api_view(['POST'])
# def upload_site_list(request):
#     file = request.FILES.get('site_file')

#     if not file:
#         return Response({"error": "File required"}, status=400)

#     if file.name.endswith('.csv'):
#         df = pd.read_csv(file)

#     elif file.name.endswith('.xlsx'):
#         df = pd.read_excel(file)

#     else:
#         return Response({"error": "Only CSV/Excel allowed"}, status=400)

#     df.columns = df.columns.str.strip().str.lower()

#     for _, row in df.iterrows():
#         UniversalAlarm.objects.create(
#             site_id=str(row.get('site id')).strip(),
#             mplane_ip=str(row.get('mplane ip')).strip(),
#             status=str(row.get('status')).strip()
#         )


#     return Response({"message": "Saved", "status": True})
    
@api_view(['POST'])
def upload_5g_log_file(request):
    if request.method != 'POST':
        return Response({"error": "Invalid request method"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
 
    log_files = request.FILES.getlist('log_files')
   
    if not log_files:
        return Response({"error": "5G log file must not be empty"}, status=status.HTTP_400_BAD_REQUEST)
 
    circle = request.POST.get("circle")
   
    if not circle:
        return Response({"error": "circle not found in input"}, status=status.HTTP_400_BAD_REQUEST)
   
    base_folder = os.path.join(MEDIA_ROOT, "Universal_alarm")
    if not os.path.exists(base_folder):
        os.makedirs(base_folder, exist_ok=True)
 
    output_folder = os.path.join(base_folder, "daily_status_alarm")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder, exist_ok=True)
 
    log_folder = os.path.join(base_folder, "log_files_5g")
    os.makedirs(log_folder, exist_ok=True)
    delete_existing_files(log_folder)
################################################# empty output with circle folder ####################################################################################
    today=datetime.now().strftime("%Y-%m-%d")
    circle_based_output_path = os.path.join(output_folder, today, circle)

#######################################################################################################################################################################
 
    for file in log_files:
        file_path = os.path.join(log_folder, file.name)
        with open(file_path, "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)
 
    # def extract_sync_status(content):
    #     sync_ok = False
    #     sync_lines = []
    #     for line in content.splitlines():
    #         if "TimeSyncIO=1" in line and "1 (ENABLED)" in line:
    #             sync_ok = True
    #             sync_lines.append(line.strip())
    #     return "OK" if sync_ok else "NOT OK"


 
    def extract_time_phase_status(content):
        match = re.search(r'timeAndPhaseSynchAlignment\s+(true|false)', content, re.IGNORECASE)

        if match and match.group(1).lower() == "true":
            return "OK"
        else:
            return "NOT OK"
    
    def extract_amf_status(content):
        pattern = re.compile(
            r'\d+\s+\d+\s+\((UNLOCKED|LOCKED)\)\s+\d+\s+\((ENABLED|DISABLED)\)\s+GNBCUCPFunction=\d+,TermPointToAmf=\d+',
            re.IGNORECASE
        )

        matches = pattern.findall(content)

        if not matches:
            return "NOT OK"

        for adm, op in matches:
            if adm.upper() != "UNLOCKED" or op.upper() != "ENABLED":
                return "NOT OK"

        return "OK"
    

    def extract_snssai_status(content):

        # Extract NRCellCU / NRCellDU blocks containing sNSSAIList
        blocks = re.findall(
            r'(NRCellC[UD][\s\S]*?sNSSAIList[\s\S]*?)(?=\nNRCellC[UD]|$)',
            content
        )

        if not blocks:
            return "NOT OK"

        required_sd = {'1', '2', '3', '4'}
        required_sst = {'1'}

        for block in blocks:

            # extract sd values
            sd_values = re.findall(r'\.\s*sd\s*=\s*(\d+)', block)

            # extract sst values
            sst_values = re.findall(r'\.\s*sst\s*=\s*(\d+)', block)

            sd_set = set(sd_values)
            sst_set = set(sst_values)

            if not required_sd.issubset(sd_set):
                return "NOT OK"

            if not required_sst.issubset(sst_set):
                return "NOT OK"

        return "OK"
    
    def extract_site_id_from_cell_name(cell_name):
        match = re.search(r'_([A-Za-z0-9]{5,12})[A-Z]_', cell_name)
        return match.group(1) if match else ""
 
    def extract_circle_from_cell_name(cell_name):
        match = re.match(r'([A-Z]+)_', cell_name)
        return match.group(1) if match else "Unknown"
    

    def parse_cell_block(content):
        pattern = re.compile(
            r'\d+\s+\d+ \((?P<adm_state>UNLOCKED|LOCKED)\)\s+'
            r'\d+ \((?P<op_state>ENABLED|DISABLED)\)\s+'
            r'(ENodeBFunction|GNBDUFunction)=\d+,(?P<cell_type>EUtranCellFDD|EUtranCellTDD|NRCellDU)=(?P<cell_name>[\w\-]+)'
        )
        cells = []
        for match in pattern.finditer(content):
            adm_state = match.group("adm_state")
            op_state = match.group("op_state")
            cell_name = match.group("cell_name")
            cell_type_str = match.group("cell_type")
 
            site_id = extract_site_id_from_cell_name(cell_name)
            circle = extract_circle_from_cell_name(cell_name)
            cell_type = "5G" if "NRCellDU" in cell_type_str else "4G"
 
            cells.append({
                'Adm State': adm_state,
                'Op. State': op_state,
                'Cell Name': cell_name,
                'Site ID': site_id,
                'Circle': circle,
                'Cell Type': cell_type
            })
        return cells
 
    def site_down_5g(content, site_id="Unknown"):
        site_down = []

        ip_match = re.search(
            r'Logging to file .*?/([\w\.]+)\.log',
            content
        )

        # ip_address = ip_match.group(1) if ip_match else "IP Not Found"
        # if ip_address.count(".") > 3:
        #     ip_address = ip_address.replace(".", ":")

        # if "Checking ip contact...Not OK" in content:
        #     remark = "Unable to connect"
        # elif "Checking ip contact...OK" in content:
        #     remark = "OK"
        # else:
        #     remark = "NOT OK"
        # data = UniversalAlarm.objects.filter(mplane_ip=ip_address).first()
        

        # site_down.append({
        # "Site ID": data.site_id if data else "Unknown",
        # "Mplane IP": ip_address,
        # "Remark": remark,
        # "Status": data.status if data else ""
        # })
        ip_address = ip_match.group(1).strip() if ip_match else ""
        # Sirf IPv6 dot format convert karo
        if ip_address and ip_address.count(".") > 3:
            ip_address = ip_address.replace(".", ":")
        
        ip_address = ip_address.lower()

        if "Checking ip contact...Not OK" in content:
            remark = "Unable to connect"
        elif "Checking ip contact...OK" in content:
            remark = "OK"
        else:
            remark = "NOT OK"

        # IP mila tabhi DB query
        data = UniversalAlarm.objects.filter(mplane_ip__iexact=ip_address).first() if ip_address else None

        site_down.append({
            "Site ID": data.site_id if data else "Unknown",
            "Mplane IP": ip_address if ip_address else "IP Not Found",
            "Remark": remark,
            "Status": data.status if data else ""
        })
        return site_down
    
    
    def extract_alarms_from_alt(content, site_id="Unknown"):
        alarms = []
        node_id_match = re.search(r'(\S+)>[\s]*alt', content)
        node_id = node_id_match.group(1).strip() if node_id_match else "Unknown"
 
        ip_match = re.search(r'\d{6}-\d{2}:\d{2}:\d{2}[+|-]\d{4}\s+([a-fA-F0-9:.]+)', content)
        # ip = ip_match.group(1) if ip_match else "No IP found"
        ip = ip_match.group(1).strip() if ip_match else ""
        # Sirf IPv6 dot format convert karo
        if ip and ip.count(".") > 3:
            ip = ip.replace(".", ":")
        
        # ip = ip.lower()
 
        alarm_pattern = re.compile(r'''
            ^\s*
            (?P<dt>\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})
            \s+(?P<sev>[A-Za-z])
            \s+(?P<prob>.+?)
            \s+(?P<mo>[^()]+?)
            (?:\s*\(\s*(?P<add>.*?)\))?
            \s*$
        ''', re.MULTILINE | re.VERBOSE)
        # print("🔍 Alarm pattern compiled successfully", alarm_pattern)

        no_alarm_keywords = {"external", "security", "ethernet"}

        noc_mapping = {
            "Security": ("No Alarm", "NSA", "Soft Alarm", "Noc Team"),
            "Ethernet": ("Link Failure", "NSA", "HW ALARM", "Circle Team"),
            "External": ("HW Faulty", "SA", "HW ALARM", "Circle Team"),
            "Inter": ("Configuration Issue", "NSA", "Soft Alarm", "Noc Team"),
            "License": ("LKF Alarm", "SA", "LKF Alarm", "Noc Team"),
            "SFP": ("SFP Not Connected", "NSA", "HW ALARM", "Circle Team"),
            "Configuration": ("RET Down", "SA", "HW ALARM", "Circle Team"),
            "SW": ("RET Down", "SA", "HW ALARM", "Circle Team"),
            "VSWR": ("VSWR High", "SA", "HW ALARM", "Circle Team"),
            "Degraded": ("Ri link Down", "SA", "HW ALARM", "Circle Team"),
            "RET": ("RET Down", "SA", "HW ALARM", "Circle Team"),
            "No": ("RET Down", "SA", "HW ALARM", "Circle Team"),
            "SUP": ("Sau Down", "NSA", "HW ALARM", "Circle Team"),
            "SAU": ("Sau Down", "NSA", "HW ALARM", "Circle Team"),
            "Hatch": ("Hatch Open", "NSA", "HW ALARM", "Circle Team"),
            "Resource": ("GSM Sector Down", "SA", "HW ALARM", "Circle Team"),
            "Link": ("Ri link Down", "SA", "HW ALARM", "Circle Team"),
            "TimeSyncIO": ("Configuration Issue", "NSA", "Soft Alarm", "Noc Team"),
            # "Inconsistent": ("RET Down", "NSA", "HW ALARM", "Circle Team"),
            "BSC": ("2G Down", "SA", "Soft Alarm", "Noc Team"),
            "HW": ("HW Faulty", "SA", "HW ALARM", "Circle Team"),
            "Fault": ("HW Faulty", "SA", "HW ALARM", "Circle Team"),
            "Current": ("RRU Down", "SA", "HW ALARM", "Circle Team"),
            "MO": ("Configuration Issue", "NSA", "Soft Alarm", "Noc Team"),
            "Fan": ("Fan Faulty", "SA", "HW ALARM", "Circle Team"),
            "XnC": ("X2 alarm", "NSA", "Soft Alarm", "Noc Team"),
            "Sync": ("GPS Alarm", "SA", "HW ALARM", "Circle Team"),
            "Clock": ("GPS Alarm", "SA", "HW ALARM", "Circle Team"),
            "Service": ("Service Alarm", "SA", "Soft Alarm", "Noc Team"),
            "Inconsistent": ("Configuration Issue", "NSA", "Soft Alarm", "Noc Team"),
            "Input":("Input Power alarm", "SA", "Power Alarm", "Circle Team"),
            "Service": ("VSWR High", "SA", "HW Alarm", "Circle Team"),
            "EC":("Configuration Issue", "NSA", "HW Alarm", "Noc Team"),
            "Feature":("Configuration Issue", "NSA", "HW Alarm", "Circle Team"),

            # Special
            "RX_BRANCH": ("Rx Branch Imbalance Fault", "SA", "HW ALARM", "Circle Team"),
            "RX_DIVERSITY": ("Rx Diversity lost", "SA", "HW ALARM", "Circle Team"),
            "VSWR_REFLECTED": ("VSWR High", "SA", "HW ALARM", "Circle Team"),
        }

        # ✅ FINAL MERGED LOGIC
        def map_noc_fields(prob, mo):
            parts = [p.strip().lower() for p in prob.split("/")]

            noc_list, bucket_list, resp_list, sa_list = [], [], [], []

            for part in parts:
                combined = part + " " + mo.lower()
                # 🔥 RX FIX
                if "rx" in part:
                    if "branch imbalance" in combined:
                        vals = noc_mapping["RX_BRANCH"]
                    elif "diversity lost" in combined:
                        vals = noc_mapping["RX_DIVERSITY"]
                    else:
                        vals = None

                    if vals:
                        noc_list.append(vals[0]); sa_list.append(vals[1])
                        bucket_list.append(vals[2]); resp_list.append(vals[3])
                        continue

                if "reflected power" in part:
                    vals = noc_mapping["VSWR_REFLECTED"]
                    noc_list.append(vals[0]); sa_list.append(vals[1])
                    bucket_list.append(vals[2]); resp_list.append(vals[3])
                    continue

                if "sup-1" in mo.lower():
                    vals = noc_mapping["SUP"]
                    noc_list.append(vals[0]); sa_list.append(vals[1])
                    bucket_list.append(vals[2]); resp_list.append(vals[3])
                    continue

                if "sau-1" in mo.lower():
                    vals = noc_mapping["SAU"]
                    noc_list.append(vals[0]); sa_list.append(vals[1])
                    bucket_list.append(vals[2]); resp_list.append(vals[3])
                    continue

                if "service" in part:
                    if "eutrancelltdd" in  mo.lower():
                        vals = ("VSWR High", "SA", "HW Alarm", "Circle Team")
                    elif "eutrancellfdd" in  mo.lower():
                        vals = ("Service Alarm", "SA", "HW ALARM", "Circle Team")
                    else:
                        vals = noc_mapping.get("Service")

                    noc_list.append(vals[0]); sa_list.append(vals[1])
                    bucket_list.append(vals[2]); resp_list.append(vals[3])
                    continue

                if "inconsistent" in part:
                    if "ret" in  mo.lower() or "antenna" in  mo.lower():
                        vals = ("RET Down", "NSA", "HW ALARM", "Circle Team")
                    else:
                        vals = noc_mapping.get("Inconsistent")

                    noc_list.append(vals[0]); sa_list.append(vals[1])
                    bucket_list.append(vals[2]); resp_list.append(vals[3])
                    continue

                # 🔥 Normal mapping
                for key, vals in noc_mapping.items():
                    if key in ["RX_BRANCH", "RX_DIVERSITY", "VSWR_REFLECTED"]:
                        continue
                    if key.lower() in part:
                        noc_list.append(vals[0])
                        sa_list.append(vals[1])
                        bucket_list.append(vals[2])
                        resp_list.append(vals[3])

            if not noc_list:
                return ("Unknown", "NO Alarm", "Unknown", "Unknown")

            noc_remark = " / ".join(noc_list)
            alarm_bucket = " / ".join(bucket_list)
            responsibility = " / ".join(resp_list)

            # ✅ SA/NSA FIX
            if "SA" in sa_list:
                sa_nsa = "SA"
            elif "NSA" in sa_list:
                sa_nsa = "NSA"
            else:
                sa_nsa = "NO Alarm"

            return (noc_remark, sa_nsa, alarm_bucket, responsibility)

        # ---- Extract alarms ----
        for m in alarm_pattern.finditer(content):
            prob = m.group("prob").strip()
            mo = m.group("mo").strip()

            if any(k in prob.lower() for k in no_alarm_keywords):
                alarm_status = "No Alarm"
                noc_remark, sa_nsa, alarm_bucket, responsibility = ("", "NO Alarm", "", "")
            else:
                alarm_status = "Alarm"
                noc_remark, sa_nsa, alarm_bucket, responsibility = map_noc_fields(prob, mo)

            data = UniversalAlarm.objects.filter(mplane_ip__iexact=ip).first() if ip else None
            status_flag = data.status if data else ""
            alarms.append({
                "IP": ip,
                "Site ID": site_id,
                "Node ID": node_id,
                "Date & Time": m.group("dt"),
                "Severity": m.group("sev"),
                "Specific Problem": prob,
                "MO": prob + mo,
                "Alarm/No Alarm": alarm_status,
                "Noc Remark": noc_remark,
                "SA/NSA": sa_nsa,
                "Alarm Bucket": alarm_bucket,
                "Responsibility": responsibility,
                "Alarm Status": status_flag
            })

        # ✅ Filter only alarms
        if any(a["Alarm/No Alarm"] == "Alarm" for a in alarms):
            alarms = [a for a in alarms if a["Alarm/No Alarm"] == "Alarm"]

        # ✅ FINAL SA/NSA
        sa_values = [a["SA/NSA"] for a in alarms]

        if "SA" in sa_values:
            final_sa = "SA"
        elif "NSA" in sa_values:
            final_sa = "NSA"
        else:
            final_sa = "NO Alarm"

        for a in alarms:
            a["SA/NSA"] = final_sa

        # print(f"✅ Extracted {len(alarms)} alarms for IP: {ip}")
        # print("1.Alarm file processed........")

        return alarms
    
        # for m in alarm_pattern.finditer(content):
        #     alarms.append({
        #         "IP": ip,
        #         "Site ID": site_id,
        #         "Node ID": node_id,
        #         "Date & Time": m.group("dt"),
        #         "Severity": m.group("sev"),
        #         "Specific Problem": m.group("prob").strip(),
        #         "MO": m.group("mo").strip(),
        #         "Additional": (m.group("add") or "").strip()
        #     })
        # return alarms
 
    all_rows = []
    all_alarms = []
    site_down = []
 
    for log_file in os.listdir(log_folder):
        log_file_path = os.path.join(log_folder, log_file)
        with open(log_file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content_5g = file.read()
 
        ip_match = re.search(r'([\w:.]+)(?=>\s*lt all)', content_5g)
        # ip_5g = ip_match.group(1) if ip_match else "No IP found"

        ip_5g = ip_match.group(1).strip() if ip_match else ""
        # Sirf IPv6 dot format convert karo
        if ip_5g and ip_5g.count(".") > 3:
            ip_5g = ip_5g.replace(".", ":")
        
        ip_5g = ip_5g.lower()

        node_match = re.search(r'([\w:-]+)(?=>\s*alt)', content_5g)
        node_id = node_match.group(1) if node_match else "No ID found"
        
        def extract_plmnr_cellwise(content_5g):
            results = {}

            # 4G cells
            pattern_4g = re.compile(
                r"(EUtranCellFDD=\S+|EUtranCellTDD=\S+)\s+(?:additionalPlmnReservedList.*?=\s*(.+?)\s|primaryPlmnReserved\s+(true|false))",
                re.IGNORECASE
            )
            for match in pattern_4g.finditer(content_5g):
                mo = match.group(1).strip()
                values = []
                if match.group(2):
                    values.extend(match.group(2).split())
                if match.group(3):
                    values.append(match.group(3))
                results[mo] = "true" if any(v.lower() == "true" for v in values) else "false"

            # 5G cells (NRCellCU)
            pattern_5g = re.compile(
                r"NRCellCU=(\S+?),NRFreqRelation=.*?plmnRestriction\s+(true|false)",
                re.IGNORECASE
            )
            for match in pattern_5g.finditer(content_5g):
                mo = f"NRCellCU={match.group(1).strip()}"
                val = match.group(2).lower()
                results[mo] = "true" if val == "true" else "false"

            return results




        
        
        plmnr_status = extract_plmnr_cellwise(content_5g)
        sync_status = extract_time_phase_status(content_5g)
        amf_status = extract_amf_status(content_5g)
        snssai_status = extract_snssai_status(content_5g)
        cells_5g = parse_cell_block(content_5g)

        # ---------------- DB Match (ONCE per IP) ----------------
        data = UniversalAlarm.objects.filter(mplane_ip__iexact=ip_5g).first() if ip_5g else None
        status_flag = data.status if data else ""

 
        for cell in cells_5g:
            cell_name = cell['Cell Name']
            site_id = cell['Site ID']
            if site_id.startswith("NL"):
                site_id = site_id[2:]

            possible_keys = [
                f"EUtranCellFDD={cell_name}",
                f"EUtranCellTDD={cell_name}",
                f"NRCellCU={cell_name}"
            ]

            # Default false, update if any match found
            plmnr_value = "NA"
            for key in possible_keys:
                if key in plmnr_status:
                    plmnr_value = plmnr_status[key]
                    break
        
            # print(f"cell_name:{cell_name}---------plmnr_value:{plmnr_value}")
          
            row = {
                ("Circle", ""): cell['Circle'],
                ("2G Site ID", ""): cell['Site ID'],
                ("5G Site ID", ""): site_id,
                ("5G Node IP", ""): ip_5g,
                ("5G Node ID", ""): node_id,
                ("5G Cell Status", "Adm State"): cell['Adm State'],
                ("5G Cell Status", "Op. State"): cell['Op. State'],
                ("Cells", ""): cell['Cell Name'],
                ("SYNC", "Status"): sync_status,
                ("AMF", "Status"): amf_status,
                ("sNSSAI", "Status"): snssai_status,
                ("Cell Type", ""): cell["Cell Type"],
                ("PLMNR_Status", ""): plmnr_value,
                ("Alarm Status", ""): status_flag,
            }
            all_rows.append(row)
            # site_id = cell['Site ID']
            # alarms = extract_alarms_from_alt(content_5g, site_id)
            # all_alarms.extend(alarms)
        # alarms cell wise
        for cell in cells_5g:
            site_id = cell['Site ID']
            if site_id.startswith("NL"):
                site_id = site_id[2:]
            alarms = extract_alarms_from_alt(content_5g, site_id)
            all_alarms.extend(alarms)
        # site down always check
        site_down.extend(site_down_5g(content_5g))
        
    columns = [
        ("Circle", ""), ("2G Site ID", ""), ("5G Site ID", ""), ("5G Node IP", ""),
        ("5G Node ID", ""), ("5G Cell Status", "Adm State"), ("5G Cell Status", "Op. State"),
        ("Cells", ""),
        ("SYNC", "Status"), 
        ("AMF", "Status"),
        ("sNSSAI", "Status"),
        ("Cell Type", ""),("PLMNR_Status", ""),
        ("Alarm Status", "")
    ]
    multi_index = pd.MultiIndex.from_tuples(columns)
    if all_rows:
        df_5g = pd.DataFrame(all_rows)
    else:
        df_5g = pd.DataFrame(columns=multi_index) 

    # df_5g = pd.DataFrame(all_rows, columns=multi_index)
    df_5g.columns = [' - '.join(filter(None, col)).strip() for col in df_5g.columns.values]
    
     #Alarm sheet
    if all_alarms:
        alarm_df = pd.DataFrame(all_alarms)
 
        agg_alarms = alarm_df.groupby(["IP", "Site ID", "Node ID"]).agg({
            "Date & Time": lambda x: " / ".join(x.dropna().sort_values().astype(str).unique()),
            "Severity": lambda x: " / ".join(sorted(set(x))),
            "Specific Problem": lambda x: " / ".join(sorted(set(x))),
            "MO": lambda x: " / ".join(sorted(set(x))),
 
            # ✅ Include all computed columns
            "Alarm/No Alarm": lambda x: " / ".join(sorted(set(x))),
            "Noc Remark": lambda x: " / ".join(sorted(set(filter(None, x)))),
            "SA/NSA": lambda x: " / ".join(sorted(set(filter(None, x)))),
            "Alarm Bucket": lambda x: " / ".join(sorted(set(filter(None, x)))),
            "Responsibility": lambda x: " / ".join(sorted(set(filter(None, x)))),
            "Alarm Status": lambda x: " / ".join(sorted(set(filter(None, x))))

        }).reset_index()
 
        df_alarms = agg_alarms
 
    else:
        df_alarms = pd.DataFrame(columns=[
            "IP", "Site ID", "Node ID", "Date & Time", "Severity",
            "Specific Problem", "MO",
            "Alarm/No Alarm", "Noc Remark", "SA/NSA",
            "Alarm Bucket", "Responsibility" ,"Alarm Status"
        ])


    # df_alarms = pd.DataFrame(all_alarms)
 
    # if not df_alarms.empty:
    #     grouped_rows = []
    #     grouped = df_alarms.groupby(["IP", "Site ID", "Node ID"])
 
    #     for (ip, site_id, node_id), group in grouped:
    #         combined = {
    #             "IP": ip,
    #             "Site ID": site_id,
    #             "Node ID": node_id,
    #             "Date & Time": '/ '.join(group["Date & Time"].dropna().unique()),
    #             "Severity": '/ '.join(group["Severity"].dropna().unique()),
    #             "Specific Problem": '/ '.join(group["Specific Problem"].dropna().unique()),
    #             "MO": '/ '.join(group["MO"].dropna().unique()),
    #             "Additional": '/ '.join(group["Additional"].dropna().unique()),
    #         }
    #         grouped_rows.append(combined)
 
    #     df_alarms = pd.DataFrame(grouped_rows)
    # else:
    #     df_alarms = pd.DataFrame(columns=[
    #         "IP", "Site ID", "Node ID", "Date & Time", "Severity",
    #         "Specific Problem", "MO",
    #         "Alarm/No Alarm", "Noc Remark", "SA/NSA",
    #         "Alarm Bucket", "Responsibility"
    #     ])

    if site_down:
        site_down_df = pd.DataFrame(site_down)
    else:
        site_down_df = pd.DataFrame(columns=["Site ID","Mplane IP","Remark", "Status"])
    
    

    circles = sorted(set(
        row[('Circle', '')].strip().upper().replace(" ", "_")
        for row in all_rows if row.get(('Circle', ''))
    ))

    if not circles:
        circle_name = circle.upper().replace(" ", "_")
    elif len(circles) > 1:
        circle_name = "".join(circles)
    else:
        circle_name = circles[0]    

    output_filename = f"5G_Alarm_Logs_{circle}.xlsx"
    output_path = os.path.join(circle_based_output_path, output_filename)
 
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df_5g["Cells"] = df_5g["Cells"].astype(str)
        df_5g = df_5g[df_5g["Cells"].str.contains("_5_", na=False)]
        df_5g["5G Site ID"] =df_5g["5G Node ID"].astype(str).str.split("-").str[1]
        df_5g["5G Site ID"] = df_5g["5G Site ID"].apply(lambda x: x[2:] if x.startswith("NL") else x)
        df_5g.drop(columns=["2G Site ID"], inplace=True)
        df_alarms.to_excel(writer, index=False, sheet_name="5G Alarms")
        site_down_df.to_excel(writer, index=False, sheet_name="5G Site Down")
        df_5g.to_excel(writer, index=False, sheet_name="5G Cells")

        format_excel_sheet(writer, '5G Site Down', site_down_df)
        format_excel_sheet(writer, '5G Cells', df_5g)
        format_excel_sheet(writer, '5G Alarms', df_alarms)
    # After saving Excel:
    relative_path = os.path.relpath(output_path, MEDIA_ROOT)
    download_link = request.build_absolute_uri("/media/" + relative_path.replace("\\", "/"))

 
    return Response({
        "message": f"Report generated successfully for circle: {circle_name}",
        "download_link": download_link,
        "status": True
    })


############################################################################################################################

def delete_existing_files(upload_files):  # noqa: F811
    for file_name in os.listdir(upload_files):
        file_path = os.path.join(upload_files, file_name)
        if os.path.isfile(file_path):
            os.remove(file_path)



@api_view(['POST'])
def upload_4g_log_file(request):
    if request.method != 'POST':
        return Response({"error": "Invalid request method"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
 
    log_files = request.FILES.getlist('log_files')
   
    if not log_files:
        return Response({"error": "4G log file must not be empty"}, status=status.HTTP_400_BAD_REQUEST)
 
    circle = request.POST.get("circle")
   
    if not circle:
        return Response({"error": "circle not found in input"}, status=status.HTTP_400_BAD_REQUEST)
   
    base_folder = os.path.join(MEDIA_ROOT, "Universal_alarm")
    if not os.path.exists(base_folder):
        os.makedirs(base_folder, exist_ok=True)
 
    output_folder = os.path.join(base_folder, "daily_status_alarm")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder, exist_ok=True)
 
    log_folder = os.path.join(base_folder, "log_files_4g")
    os.makedirs(log_folder, exist_ok=True)
    delete_existing_files(log_folder)
################################################# empty output with circle folder ####################################################################################
    today=datetime.now().strftime("%Y-%m-%d")
    circle_based_output_path = os.path.join(output_folder,today, circle)
    if os.path.exists(circle_based_output_path):
        shutil.rmtree(circle_based_output_path)
        # print("All olde files deleted successfully")
    else:
        os.makedirs(circle_based_output_path, exist_ok=True)
 
    for file in log_files:
        file_path = os.path.join(log_folder, file.name)
        with open(file_path, "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)
 
   
 
    def get_band(cell_name):
        band_map = {
            "F8": "L900",
            "F1": "L2100",
            "F3": "L1800",
            "T1": "L2300",
            "T2": "L2300",
        }
 
        # Use regex to extract the band code from cell name (e.g., F8, F1, etc.)
        match = re.search(r'_([FT]\d)_', cell_name)
        if match:
            band_code = match.group(1)
            return band_map.get(band_code, "")
        return ""
    # def extract_4g_sync_status(content_4g):
    #     sync_ok = False
    #     sync_lines = []
 
    #     for line in content_4g.splitlines():
    #         if "TimeSyncIO=1" in line and "1 (ENABLED)" in line:
    #             sync_ok = True
    #             sync_lines.append(line.strip())
 
    #     print("🧪 Sync lines found:", sync_lines)
 
    #     return "OK" if sync_ok else "NOT OK"

    def extract_time_phase_status(content_4g):
        match = re.search(r'timeAndPhaseSynchAlignment\s+(true|false)', content_4g, re.IGNORECASE)

        if match and match.group(1).lower() == "true":
            return "OK"
        else:
            return "NOT OK"
    
    def extract_4g_site_id_from_cell_name(cell_name):
   
        match = re.search(r'_([A-Za-z0-9]{5,12})[A-Z]_', cell_name)
        return match.group(1) if match else ""
   
   
    def extract_4g_circle_from_cell_name(cell_name):
        match = re.match(r'([A-Z]+)_', cell_name)
        return match.group(1) if match else "Unknown"
 

 
 
    # def site_down_4g(content_4g, site_id="Unknown"):
       
    # #   def site_down_4g(content_4g, site_id="Unknown"):
    #     site_down=[]
    #     ip_match = re.search(
    #     # r'\d{6}-\d{2}:\d{2}:\d{2}[+|-]\d{4}\s+([a-fA-F0-9:.]+)',
    #      r'Logging to file .+?/([a-fA-F0-9:.]+)\.log',
       
    #     content_4g)
    #     ip_address = ip_match.group(1) if ip_match else "IP Not Found"
    #     status = "OK" if "Checking ip contact...OK" in content_4g else "NOT OK"
       
    #     site_down.append(
    #         {
    #             "Status": status,
    #             "IP Address": ip_address,
        
    #         }
    #     )
    #     return site_down

    def site_down_4g(content_4g, site_id="Unknown"):
        site_down = []

        ip_match = re.search(
            r'Logging to file .*?/([\w\.]+)\.log',
            content_4g
        )

        # ip_address = ip_match.group(1) if ip_match else "IP Not Found"
        ip_address = ip_match.group(1).strip() if ip_match else ""
        # Sirf IPv6 dot format convert karo
        if ip_address and ip_address.count(".") > 3:
            ip_address = ip_address.replace(".", ":")
        
        ip_address = ip_address.lower()

        if "Checking ip contact...Not OK" in content_4g:
            remark = "Unable to connect"
        elif "Checking ip contact...OK" in content_4g:
            remark = "OK"
        else:
            remark = "NOT OK"
        data = UniversalAlarm.objects.filter(mplane_ip__iexact=ip_address).first() if ip_address else None

        site_down.append({
            "Site ID": data.site_id if data else "Unknown",
            "Mplane IP": ip_address if ip_address else "IP Not Found",
            "Remark": remark,
            "Status": data.status if data else ""
        })
        return site_down

    def extract_4galarms_from_alt(content, site_id="Unknown"):
        alarms = []

        node_id_match = re.search(r'(\S+)>[\s]*alt', content)
        node_id = node_id_match.group(1).strip() if node_id_match else "Unknown"

        ip_match = re.search(
            r'\d{6}-\d{2}:\d{2}:\d{2}[+|-]\d{4}\s+([a-fA-F0-9:.]+)',
            content
        )
        # ip = ip_match.group(1) if ip_match else "No IP found"
        ip = ip_match.group(1).strip() if ip_match else ""
        # Sirf IPv6 dot format convert karo
        if ip and ip.count(".") > 3:
            ip = ip.replace(".", ":")

        # ip = ip.lower()

        alarm_pattern = re.compile(r'''
            (?P<dt>\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+
            (?P<sev>[A-Za-z])\s+
            (?P<prob>.+?)\s+
            (?P<mo>[^()\n]+)
        ''', re.MULTILINE | re.VERBOSE)

        

        no_alarm_keywords = {"external", "security", "ethernet"}

        noc_mapping = {
            "Security": ("No Alarm", "NSA", "Soft Alarm", "Noc Team"),
            "Ethernet": ("Link Failure", "NSA", "HW ALARM", "Circle Team"),
            "External": ("HW Faulty", "SA", "HW ALARM", "Circle Team"),
            "Inter": ("Configuration Issue", "NSA", "Soft Alarm", "Noc Team"),
            "License": ("LKF Alarm", "SA", "LKF Alarm", "Noc Team"),
            "SFP": ("SFP Not Connected", "NSA", "HW ALARM", "Circle Team"),
            "Configuration": ("RET Down", "SA", "HW ALARM", "Circle Team"),
            "SW": ("RET Down", "SA", "HW ALARM", "Circle Team"),
            "VSWR": ("VSWR High", "SA", "HW ALARM", "Circle Team"),
            "Degraded": ("Ri link Down", "SA", "HW ALARM", "Circle Team"),
            "RET": ("RET Down", "SA", "HW ALARM", "Circle Team"),
            "No": ("RET Down", "SA", "HW ALARM", "Circle Team"),
            "SUP": ("Sau Down", "NSA", "HW ALARM", "Circle Team"),
            "SAU": ("Sau Down", "NSA", "HW ALARM", "Circle Team"),
            "Hatch": ("Hatch Open", "NSA", "HW ALARM", "Circle Team"),
            "Resource": ("GSM Sector Down", "SA", "HW ALARM", "Circle Team"),
            "Link": ("Ri link Down", "SA", "HW ALARM", "Circle Team"),
            "TimeSyncIO": ("Configuration Issue", "NSA", "Soft Alarm", "Noc Team"),
            # "Inconsistent": ("RET Down", "NSA", "HW ALARM", "Circle Team"),
            "BSC": ("2G Down", "SA", "Soft Alarm", "Noc Team"),
            "HW": ("HW Faulty", "SA", "HW ALARM", "Circle Team"),
            "Fault": ("HW Faulty", "SA", "HW ALARM", "Circle Team"),
            "Current": ("RRU Down", "SA", "HW ALARM", "Circle Team"),
            "MO": ("Configuration Issue", "NSA", "Soft Alarm", "Noc Team"),
            "Fan": ("Fan Faulty", "SA", "HW ALARM", "Circle Team"),
            "XnC": ("X2 alarm", "NSA", "Soft Alarm", "Noc Team"),
            "Sync": ("GPS Alarm", "SA", "HW ALARM", "Circle Team"),
            "Clock": ("GPS Alarm", "SA", "HW ALARM", "Circle Team"),
            "Service": ("Service Alarm", "SA", "Soft Alarm", "Noc Team"),
            "Inconsistent": ("Configuration Issue", "NSA", "Soft Alarm", "Noc Team"),
            "Input":("Input Power alarm", "SA", "Power Alarm", "Circle Team"),
            "Service": ("VSWR High", "SA", "HW Alarm", "Circle Team"),
            "EC":("Configuration Issue", "NSA", "HW Alarm", "Noc Team"),
            "Feature":("Configuration Issue", "NSA", "HW Alarm", "Circle Team"),


            # Special
            "RX_BRANCH": ("Rx Branch Imbalance Fault", "SA", "HW ALARM", "Circle Team"),
            "RX_DIVERSITY": ("Rx Diversity lost", "SA", "HW ALARM", "Circle Team"),
            "VSWR_REFLECTED": ("VSWR High", "SA", "HW ALARM", "Circle Team"),
        }

        # ✅ FINAL MERGED LOGIC
        def map_noc_fields(prob, mo):
            parts = [p.strip().lower() for p in prob.split("/")]

            noc_list, bucket_list, resp_list, sa_list = [], [], [], []

            for part in parts:
                combined = part + " " + mo.lower()
                # 🔥 RX FIX
                if "rx" in part:
                    if "branch imbalance" in combined:
                        vals = noc_mapping["RX_BRANCH"]
                    elif "diversity lost" in combined:
                        vals = noc_mapping["RX_DIVERSITY"]
                    else:
                        vals = None

                    if vals:
                        noc_list.append(vals[0]); sa_list.append(vals[1])
                        bucket_list.append(vals[2]); resp_list.append(vals[3])
                        continue

                if "reflected power" in part:
                    vals = noc_mapping["VSWR_REFLECTED"]
                    noc_list.append(vals[0]); sa_list.append(vals[1])
                    bucket_list.append(vals[2]); resp_list.append(vals[3])
                    continue

                if "sup-1" in mo.lower():
                    vals = noc_mapping["SUP"]
                    noc_list.append(vals[0]); sa_list.append(vals[1])
                    bucket_list.append(vals[2]); resp_list.append(vals[3])
                    continue

                if "sau-1" in mo.lower():
                    vals = noc_mapping["SAU"]
                    noc_list.append(vals[0]); sa_list.append(vals[1])
                    bucket_list.append(vals[2]); resp_list.append(vals[3])
                    continue
                
                if "service" in part:
                    if "eutrancelltdd" in  mo.lower():
                        vals = ("VSWR High", "SA", "HW Alarm", "Circle Team")
                    elif "eutrancellfdd" in  mo.lower():
                        vals = ("Service Alarm", "SA", "HW ALARM", "Circle Team")
                    else:
                        vals = noc_mapping.get("Service")

                    noc_list.append(vals[0]); sa_list.append(vals[1])
                    bucket_list.append(vals[2]); resp_list.append(vals[3])
                    continue

                if "inconsistent" in part:
                    if "ret" in  mo.lower() or "antenna" in  mo.lower():
                        vals = ("RET Down", "NSA", "HW ALARM", "Circle Team")
                    else:
                        vals = noc_mapping.get("Inconsistent")

                    noc_list.append(vals[0]); sa_list.append(vals[1])
                    bucket_list.append(vals[2]); resp_list.append(vals[3])
                    continue

                
                for key, vals in noc_mapping.items():
                    if key in ["RX_BRANCH", "RX_DIVERSITY", "VSWR_REFLECTED"]:
                        continue
                    if key.lower() in part:
                        noc_list.append(vals[0])
                        sa_list.append(vals[1])
                        bucket_list.append(vals[2])
                        resp_list.append(vals[3])

            if not noc_list:
                return ("Unknown", "NO Alarm", "Unknown", "Unknown")

            noc_remark = " / ".join(noc_list)
            alarm_bucket = " / ".join(bucket_list)
            responsibility = " / ".join(resp_list)

            # ✅ SA/NSA FIX
            if "SA" in sa_list:
                sa_nsa = "SA"
            elif "NSA" in sa_list:
                sa_nsa = "NSA"
            else:
                sa_nsa = "NO Alarm"

            return (noc_remark, sa_nsa, alarm_bucket, responsibility)

        # ---- Extract alarms ----
        for m in alarm_pattern.finditer(content):
            prob = m.group("prob").strip()
            mo = m.group("mo").strip()
            
            complete_mo = f"{m.group('dt')[5:]} {m.group('sev')} {prob} {mo}"
 

            if any(k in prob.lower() for k in no_alarm_keywords):
                alarm_status = "No Alarm"
                noc_remark, sa_nsa, alarm_bucket, responsibility = ("", "NO Alarm", "", "")
            else:
                alarm_status = "Alarm"
                noc_remark, sa_nsa, alarm_bucket, responsibility = map_noc_fields(prob, mo)
            data = UniversalAlarm.objects.filter(mplane_ip__iexact=ip).first() if ip else None
            status_flag = data.status if data else ""
            
            alarms.append({
                "IP": ip,
                "Site ID": site_id,
                "Node ID": node_id,
                "Date & Time": m.group("dt"),
                "Severity": m.group("sev"),
                "Specific Problem": prob,
                "MO": mo,
                "MO Alarm": complete_mo,
                "Alarm/No Alarm": alarm_status,
                "Noc Remark": noc_remark,
                "SA/NSA": sa_nsa,
                "Alarm Bucket": alarm_bucket,
                "Responsibility": responsibility,
                "Alarm Status": status_flag
            })

        # ✅ Filter only alarms
        if any(a["Alarm/No Alarm"] == "Alarm" for a in alarms):
            alarms = [a for a in alarms if a["Alarm/No Alarm"] == "Alarm"]

        # ✅ FINAL SA/NSA
        sa_values = [a["SA/NSA"] for a in alarms]

        if "SA" in sa_values:
            final_sa = "SA"
        elif "NSA" in sa_values:
            final_sa = "NSA"
        else:
            final_sa = "NO Alarm"

        for a in alarms:
            a["SA/NSA"] = final_sa

        # print(f"✅ Extracted {len(alarms)} alarms for IP: {ip}")
        # print("1.Alarm file processed........")

        return alarms
    
    def extract_all_2g_trx_status(content):

        pattern = re.compile(
            r'GsmSector=\d+,Trx=([A-Z0-9-]+)\s+abisTsState\s+i\[\d+\]\s*=\s*[\d\s]+\(([A-Z\s]+)\)',
            re.IGNORECASE
        )

        trx_status_dict = {}

        for match in pattern.finditer(content):
            trx_name = match.group(1)
            raw_statuses = match.group(2).split()

            unique_statuses = {s.upper() for s in raw_statuses}

            if "ENABLED" in unique_statuses:
                final_status = "UNLOCKED"
            elif "DISABLED" in unique_statuses:
                final_status = "LOCKED"
            elif "RESET" in unique_statuses:
                final_status = "DOWN"
            else:
                final_status = "UNKNOWN"

            trx_status_dict[trx_name] = final_status

        # 🔹 If all TRX are UNLOCKED
        if trx_status_dict and all(v == "UNLOCKED" for v in trx_status_dict.values()):
            return [("", "UNLOCKED")]
        
        if trx_status_dict and all(v == "LOCKED" for v in trx_status_dict.values()):
            return [("", "LOCKED")]

        
        # 🔹 Return normal list
        return [(trx, status) for trx, status in trx_status_dict.items()]
        

    def parse_4g_st_cell_output(content_4g):
        pattern = re.compile(
            r'\d+\s+\d+ \((?P<adm_state>UNLOCKED|LOCKED)\)\s+'
            r'\d+ \((?P<op_state>ENABLED|DISABLED)\)\s+'
            r'(ENodeBFunction|GNBDUFunction)=\d+,(?:EUtranCellFDD|EUtranCellTDD)=(?P<cell_name>[\w\-]+)'
        )
 
 
 
 
        cells= []
        for match in pattern.finditer(content_4g):
            cell_name = match.group("cell_name")
            adm_state = match.group("adm_state")
            op_state = match.group("op_state")
            circle =  extract_4g_circle_from_cell_name(cell_name)
            site_id = extract_4g_site_id_from_cell_name(cell_name)
 
            band = get_band(cell_name)
            # print("2. Status file processed........")
            cells.append({
                ("4G Cell Status", "Adm State"): adm_state,
                ("4G Cell Status", "Op. State"): op_state,
                'Cell Name': cell_name,
                'Site ID': site_id,
                'Circle': circle,
                'Band': band
            })
        return cells
    all_rows = []
    all_alarms = []
    site_down = []
    for log_file in os.listdir(log_folder):
        log_file_path = os.path.join(log_folder, log_file)
        with open(log_file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content_4g = file.read()
        ip_match = re.search(r'([\w:.]+)(?=>\s*lt all)', content_4g)
        # ip_4g = ip_match.group(1) if ip_match else "No IP found"

        ip_4g = ip_match.group(1).strip() if ip_match else ""
        # Sirf IPv6 dot format convert karo
        if ip_4g and ip_4g.count(".") > 3:
            ip_4g = ip_4g.replace(".", ":")

        # ip_4g = ip_4g.lower()
        

        node_match = re.search(r'([\w:-]+)(?=>\s*alt)', content_4g)
        node_id = node_match.group(1) if node_match else "No ID found"
        
  
        # --- Extract PLMNR cell-wise ---
        def extract_plmnr_cellwise(content_4g):
            """
            Extracts PLMNR status for each cell (EUtranCellFDD, EUtranCellTDD, NRCellCU).
            Returns dict {full_mo: "true"/"false"}
            """
            results = {}

            # Pattern to capture MO and value
            pattern = re.compile(
                r"(EUtranCellFDD=\S+|EUtranCellTDD=\S+|NRCellCU=\S+).*?\s(true|false)",
                re.IGNORECASE
            )

            for match in pattern.finditer(content_4g):
                mo = match.group(1).strip()
                value = match.group(2).lower()

                if mo not in results:
                    results[mo] = []
                results[mo].append(value)

            # Finalize → if any "true" then true else false
            for mo, vals in results.items():
                results[mo] = "true" if "true" in vals else "false"

            return results


        plmm_status = extract_plmnr_cellwise(content_4g)
        sync_status = extract_time_phase_status(content_4g)
        cells_4g = parse_4g_st_cell_output(content_4g)
      
        trx_status_list = extract_all_2g_trx_status(content_4g)
        trx_status_str = "; ".join([f"{trx} - {status}" for trx, status in trx_status_list])

        data = UniversalAlarm.objects.filter(mplane_ip__iexact=ip_4g).first() if ip_4g else None
        status_flag = data.status if data else ""

        print(data)



        for cell in cells_4g:
            cell_name = cell['Cell Name']

            # Possible MO keys for this cell
            possible_keys = [
                f"EUtranCellFDD={cell_name}",
                f"EUtranCellTDD={cell_name}",
                f"NRCellCU={cell_name}"
            ]

            # Default false, update if any match found
            plmm_value = "NA"
            for key in possible_keys:
                if key in plmm_status:
                    plmm_value = plmm_status[key]
                    break
                
            row = {
                ('Circle', ''): cell['Circle'],
                ('2G Site ID', ''): cell['Site ID'],
                ('4G Site ID', ''): cell['Site ID'],  # Assuming naming rule
                ('4G Node IP', ''): ip_4g,
                ('4G Node ID', ''): node_id,
                ('4G Cell Status', 'Adm State'): cell[('4G Cell Status', 'Adm State')],
                ('4G Cell Status', 'Op. State'): cell[('4G Cell Status', 'Op. State')],
                ('Cells', ''): cell['Cell Name'],
                ('PLMNR Status',''): plmm_value,
                # ('2G Cell(TRX) Status', ''): [extract_all_2g_trx_status(content_4g)] ,
                ("2G Cell(TRX) Status", "Status"): trx_status_str,
                # ('2G Cell(TRX) Status', ''): "; ".join([f"{trx} - {status}" for trx, status in extract_all_2g_trx_status(content_4g)]),
                ('SYNC', 'Status'): sync_status,
                ('Band', ''): cell['Band'],
                ('Alarm Status', ''): status_flag
            
            }
            all_rows.append(row)

        # alarms cell wise
        for cell in cells_4g:
            site_id = cell['Site ID']
            alarms = extract_4galarms_from_alt(content_4g, site_id)
            all_alarms.extend(alarms)

        # site down always check
        site_down.extend(site_down_4g(content_4g))
        # for cell in cells_4g:
        #     site_id = cell['Site ID']
        #     alarms = extract_4galarms_from_alt(content_4g, site_id)
        #     all_alarms.extend(alarms)
        #     site_down.extend(site_down_4g(content_4g, site_id))
 
    columns = [
        ("Circle", ""),
        ("2G Site ID", ""),
        ("4G Site ID", ""),
        ("4G Node IP", ""),
        ("4G Node ID", ""),
        ("4G Cell Status", "Adm State"),
        ("4G Cell Status", "Op. State"),
        ("Cells", ""),
        ("2G Cell(TRX)", "Status"),
        ("SYNC", "Status"),
        ("Band", ""),
        ("Plmmr Status", ""),
        ("Alarm Status", "")

        
    ]
    multi_index = pd.MultiIndex.from_tuples(columns)
    if all_rows:
        df_4g = pd.DataFrame(all_rows)
    else:
        df_4g = pd.DataFrame(columns=multi_index) 

# ❗ Flatten before saving to Excel
    df_4g.columns = [' - '.join(filter(None, col)).strip() for col in df_4g.columns.values]
    
    
# Alarm sheet
    if all_alarms:
        alarm_df = pd.DataFrame(all_alarms)
 
        agg_alarms = alarm_df.groupby(["IP", "Site ID", "Node ID"]).agg({
            "Date & Time": lambda x: " / ".join(x.dropna().sort_values().astype(str).unique()),
            "Severity": lambda x: " / ".join(sorted(set(x))),
            "Specific Problem": lambda x: " / ".join(sorted(set(x))),
            "MO": lambda x: " / ".join(sorted(set(x))),
            "MO Alarm": lambda x: " / ".join(sorted(set(x))),
 
            # ✅ Include all computed columns
            "Alarm/No Alarm": lambda x: " / ".join(sorted(set(x))),
            "Noc Remark": lambda x: " / ".join(sorted(set(filter(None, x)))),
            "SA/NSA": lambda x: " / ".join(sorted(set(filter(None, x)))),
            "Alarm Bucket": lambda x: " / ".join(sorted(set(filter(None, x)))),
            "Responsibility": lambda x: " / ".join(sorted(set(filter(None, x)))),
            "Alarm Status": lambda x: " / ".join(sorted(set(filter(None, x))))
        }).reset_index()
 
        df_alarms = agg_alarms
 
    else:
        df_alarms = pd.DataFrame(columns=[
            "IP", "Site ID", "Node ID", "Date & Time", "Severity",
            "Specific Problem", "MO",
            "Alarm/No Alarm", "Noc Remark", "SA/NSA",
            "Alarm Bucket", "Responsibility","Alarm Status"
        ])


    if site_down:
        site_down_df = pd.DataFrame(site_down).drop_duplicates()
    else:
        site_down_df = pd.DataFrame(columns=["Site ID","Mplane IP","Remark", "Status"])
   

    circles = sorted(set(
    row[('Circle', '')].strip().upper().replace(" ", "_")
    for row in all_rows if row.get(('Circle', ''))
    ))
 
        # Join all circle names (like TNCH or TN_ORI_CH)
 
    if not circles:
        circle_name = circle.upper().replace(" ", "_")
    elif len(circles) > 1:
        circle_name = "".join(circles)
    else:
        circle_name = circles[0] 

    df_alarms["MO Alarm"] = df_alarms["MO Alarm"].apply(lambda val: " / ".join([
        re.sub(r"^\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}\s+[A-Za-z]\s+", "", i.strip())
        for i in val.split(" / ")
    ]))
    
    df_alarms.drop(columns=["MO"], inplace=True)
    
    output_filename = f"4G_Alarm_Logs_{circle}.xlsx"
    output_path = os.path.join(circle_based_output_path, output_filename)
    ##################################
    # with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df_4g.to_excel(writer, index=False, sheet_name="4G Cells")
        df_alarms.to_excel(writer, index=False, sheet_name="4G Alarms")
        site_down_df.to_excel(writer, sheet_name="4G Site Down", index=False)
        format_excel_sheet(writer, '4G Alarms', df_alarms ,)
 
    # After saving Excel:
    relative_path = os.path.relpath(output_path, MEDIA_ROOT)
    download_link = request.build_absolute_uri("/media/" + relative_path.replace("\\", "/"))

 
    wb = load_workbook(output_path)
    ws = wb.active
    for sheet_name in wb.sheetnames:
        ws = wb["4G Cells"]
 
        header_fill = PatternFill(start_color="259aa8", end_color="4682B4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal="center", vertical="center")
 
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        prev_value = None
        start_col = 1
        for col in range(1, ws.max_column + 1):
            top_cell = ws.cell(row=1, column=col)
            sub_cell = ws.cell(row=2, column=col)
            top_val = top_cell.value
            sub_val = sub_cell.value  # noqa: F841
            # Detect merged header range
            if top_val != prev_value:
                if col > start_col:
                    if prev_value not in (None, ""):
                        ws.merge_cells(
                            start_row=1, start_column=start_col,
                            end_row=1, end_column=col - 1
                        )
                start_col = col
                prev_value = top_val
        # Final merge for last group
        if prev_value not in (None, "") and start_col <= ws.max_column:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=ws.max_column
            )
        # Styling headers
        for row in [1, 1]:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = thin_border
        # Style entire data range
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        wb.save(output_path)
    return Response({
        # "message": "4G log file processed successfully",
        "message": f"Report generated successfully for circle: {circle_name}",
        "download_link": download_link,
        "status": True
    }, status=200)





