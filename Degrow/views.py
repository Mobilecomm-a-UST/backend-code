import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
# Create your views here.
from openpyxl import Workbook,load_workbook

import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
# Create your views here.
from openpyxl import Workbook,load_workbook


from django.shortcuts import render
from django.shortcuts import render,redirect
from django.http import HttpResponse
import sys
import pandas as pd
from django.contrib import messages
# import json
from django.contrib.auth.decorators import login_required
from mcom_website.settings import MEDIA_URL,BASE_DIR,ALLOWED_HOSTS
import numpy as np
import os
import datetime
from rest_framework.decorators import api_view,parser_classes
from rest_framework.response import Response
from django.http import JsonResponse

# from .serializer import *

from django.forms.models import model_to_dict

from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import authentication_classes,permission_classes
from django.core.files.storage import FileSystemStorage
from mcom_website.settings import MEDIA_ROOT
from .models import *

import openpyxl
from openpyxl.styles import PatternFill
# Create your views here.
from openpyxl import Workbook,load_workbook
from statistics import mean

from django.db.models import Sum


import pandas as pd
import matplotlib  as mpl
import numpy as np

from datetime import date, timedelta

def traverse_folder(folder_path):
    # Iterate over all the files and folders in the specified folder
    # #print("in fun")
        for root, dirs, files in os.walk(folder_path):
                file_list=files
            # for file in files:
            #     file_path = os.path.join(root, file)  # Get the absolute path of the file
            #     #print(file_path)  # Do whatever you want with the file path
            #     files

        return file_list

def extract_date(string):
        str_list=string.split("_")
        #print(str_list)
        date_str=str_list[-1].split(".")[0]
        #print(date_str)
        year=int(date_str[0:4])
        month=int(date_str[4:6])
        day=int(date_str[6:])

        #print(year,month,day)

        date=datetime.date(year,month,day)
        
        date = date - timedelta(1)
        #print(date)
        return date


alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
def num_hash(num):
    if num < 26:
        return alpha[num-1]
    else:
        q, r = num//26, num % 26
        if r == 0:
            if q == 1:
                return alpha[r-1]
            else:
                return num_hash(q-1) + alpha[r-1]
        else:
            return num_hash(q) + alpha[r-1]
        
def template_colouring(ws,no_pre_cols,no_post_cols):
    BLUE="1a90ff"
    PURPLE="db70ff"
    CYAN="1addff"
    ORANGE="ffd53d"
    ws["A1"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE
    ws["A2"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE
    ws["B1"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE
    ws["B2"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE
    ws["A3"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE
    ws["B3"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE
    
    play=3
    ws[str(num_hash(play))+"1"].fill=PatternFill(patternType='solid',fgColor=ORANGE) #PURPLE #pre

    ws[str(num_hash(play+no_pre_cols))+"1"].fill=PatternFill(patternType='solid',fgColor=BLUE) #PURPLE #post

    ws[str(num_hash(play+no_pre_cols + no_post_cols))+"1"].fill=PatternFill(patternType='solid',fgColor=PURPLE) #PURPLE #comparision

    for i in range(0,no_pre_cols):
        j=play+i
        ws[str(num_hash(j))+"2"].fill=PatternFill(patternType='solid',fgColor=PURPLE)

    for x in range(0,no_post_cols):
        K=play + no_pre_cols + x
        ws[str(num_hash(K))+"2"].fill=PatternFill(patternType='solid',fgColor=PURPLE)

    ws[str(num_hash(play+no_pre_cols + no_post_cols))+"2"].fill=PatternFill(patternType='solid',fgColor=ORANGE) #ORANGE
    ws[str(num_hash(play+no_pre_cols + no_post_cols+1))+"2"].fill=PatternFill(patternType='solid',fgColor=BLUE) #BLUE
    ws[str(num_hash(play+no_pre_cols + no_post_cols+2))+"2"].fill=PatternFill(patternType='solid',fgColor=CYAN) #CYAN
    ws[str(num_hash(play+no_pre_cols + no_post_cols+3))+"2"].fill=PatternFill(patternType='solid',fgColor=CYAN) #CYAN

    return ws
def pre_degrow(additional_inputs,pre_file_dict): 
       # Example usage

       ############################################## code for additional inputs ######################################################
       
        additional_inputs_df=pd.read_excel(additional_inputs)
        freq_band=int(additional_inputs_df.at[0,"Freq_Band"])
        site_id=additional_inputs_df.at[0,"Site_ID"]
        global SITE_ID
        SITE_ID=site_id
        site_list=[site_id,("F" + site_id)]
        sector=additional_inputs_df.at[0,"Sector"]

        
        # #print(additional_inputs_df)
        # #print("freq_band:--",freq_band,type(freq_band))
        # #print("Site List:--",site_list,)
        # #print("sector:--",sector,type(sector))
        # #print("nishant verma...")
        ######################################## ******************************** ################################################
       

       #################################################  main processing  #######################################################                  
       
        # file_list=traverse_folder(folder_path_pre)
        # #print(file_list)
        K=0
        data1={} # dictionary containing data for "4G Data Volume[GB]" from 24hrs file
        data2={}  # dictionary containing data for ALL Other KPis from BBH  file
        code_executed=False
        for file,date in pre_file_dict.items():
            
            i=date
            
            ############################################ code for 24hrs ###########################
            if "_24Hrs_" in file.name:
                # #print("------------------------------------------------------------------################## pre ####################----------------------------------------------------")
                # #print("------------------------------------------------------------------##################",i,"####################----------------------------------------------------")
                # #print(".......-------------------------------------------------------------------------inside 24hrs------------------------------------------------------------.........")
                # path=os.path.join(folder_path_pre,file)
                df=pd.read_csv(file)
                
                df.rename(columns={"Site Name" :"Site_Name" } ,inplace = True )
                df=df[(df.Site_Name.isin(site_list))]                              # filtering the df with sitename 
                df['Short name'] = df['Short name'].str.strip()
                
                ################## to get the frequency band ##########################
                # band_list=df['Freq_Band'].unique()
                # band_list=pd.Series(band_list).dropna().astype(int)
                
                band_list=[1800,2100,900,2300]
                #print("Band List:-",band_list)

                dif_tech_data={}
                # row_sum = df.sum(axis=0)              
                # VoLTE_Traffic=row_sum["MV_4G Data Volume_GB"]
                # dif_tech_data[("MV_4G Data Volume_GB","NW")]=VoLTE_Traffic

                for band in band_list:  
                    
                    #print(band,freq_band)
                    if band == freq_band: # to compare the de grow technology
                        # #print(band)

                        df_fil=df[df.Freq_Band == freq_band]
                        if band == 2300:
                            # if df_fil['Short name'].str.contains("_T1_"):__
                                if sector == "A":
                                    
                                    df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "B":
                                    
                                    df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "C":
                                    
                                    df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "D":
                                    
                                    df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                #print("df_fil_t1---------")
                                # #print(df_fil_t1)
                                row_sum_t1 = df_fil_t1.sum(axis=0)
                                
                                t=row_sum_t1["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","TDDC1")]=t
                            # if df_fil['Short name'].str.contains("_T2_"):
                                if sector == "A":
                                    
                                    df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "B":
                                    
                                    df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "C":
                                    
                                    df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                
                                if sector == "D":
                                    
                                    df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                #print("df_fil_t2---------")

                                # #print(df_fil_t2)
                                row_sum_t2 = df_fil_t2.sum(axis=0)
                                t=row_sum_t2["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","TDDC2")]=t
                    
                        else:
                            if sector == "A":  
                                df_fil_sec = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")]
                            if sector == "B":
                                df_fil_sec = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")]
                            if sector == "C":  
                                df_fil_sec = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")]
                            if sector == "D":
                                df_fil_sec = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")]
                            

                            row_sum = df_fil_sec.sum(axis=0)
                            if band==1800:
                               
                                t=row_sum["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD18")]=t
                            if band==2100:
                               
                                t=row_sum["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD21")]=t
                            if band==900:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD8")]=t
                            
                            # #print(df_fil_sec)

                            
                        
                        
                    else :
                        df_fil=df[df.Freq_Band == band]
                        # #print(df_fil)
                        # #print(band)

                        if band == 2300:
                            # if df_fil['Short name'].str.contains("_T1_"):__
                                if sector == "A":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "B":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "C":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "D":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]

                                # #print("df_fil_t1---------")
                                # #print(df_fil_t1)
                                row_sum_t1 = df_fil_t1.sum(axis=0)
                                
                                t=row_sum_t1["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","TDDC1")]=t
                            # if df_fil['Short name'].str.contains("_T2_"):
                                if sector == "A":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "B":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "C":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "D":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                #print("df_fil_t2---------")


                                # #print(df_fil_t2)
                                row_sum_t2 = df_fil_t2.sum(axis=0)
                                t=row_sum_t2["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","TDDC2")]=t


                        else:


                            if sector == "A":
                                    
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")]
                            if sector == "B":
                                
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")]
                            if sector == "C":
                                
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")]
                            if sector == "D":
                                
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")]
                            

                            row_sum = df_fil.sum(axis=0)
                            if band==1800:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                # data[i]={("MV_4G Data Volume_GB","FD18"):t}
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD18")]=t
                            if band==2100:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                # data[i]={("MV_4G Data Volume_GB","FD21"):t}
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD21")]=t
                            if band==900:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                # data[i]={("MV_4G Data Volume_GB","FD8"):t}
                
                    
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD9")]=t
                
                #---------------- code for nettwork level data ------------------------#
                row_sum = df.sum(axis=0)              
                VoLTE_Traffic=row_sum["MV_4G Data Volume_GB"]
                dif_tech_data[("DATA TRAFFIC (24hrs)","NW")]=VoLTE_Traffic
                #------------------***************************-------------------------#

                data1[i]=dif_tech_data        
                
              
          ############################################################ code for bbh ###################################################
            else:
                
                # if not code_executed:
                #     K=1
                #     i="Day"+ str(K)
                #     code_executed=True
                

                # #print("------------------------------------------------------------------################## pre ####################----------------------------------------------------")
                # #print("------------------------------------------------------------------##################",i,"####################----------------------------------------------------")
                # #print("_________________--------------------------------------------------------------####  in bbh ####--------------------------------------------_____________________")
               
                # #print("yeas................")
                # path=os.path.join(folder_path_pre,file)
                df=pd.read_csv(file)
                
                df.rename(columns={"Site Name" :"Site_Name" } ,inplace = True )
                df=df[(df.Site_Name.isin(site_list))]
                df['Short name'] = df['Short name'].str.strip()
                
                ################## to get the frequency band ##########################
                # band_list=df['Freq_Band'].unique()
                # band_list=pd.Series(band_list).dropna().astype(int)
                
                KPI=["VoLTE Traffic [CBBH]","DL User Throughput_Kbps [CDBH]","UL User Throughput_Kbps [CDBH]","Average number of used DL PRBs [CDBH]","Avg Connected User [CDBH]","RRC Setup Success Rate [CDBH]","ERAB Setup Success Rate [CDBH]","VoLTE DCR [CBBH]"]
                KPI_dic ={
                          "VoLTE Traffic [CBBH]":"VOLTE TRAFFIC (CBBH)",
                          "DL User Throughput_Kbps [CDBH]":"DL THPT (CDBH)",
                          "UL User Throughput_Kbps [CDBH]":"UL THPT (CUBH)",
                          "Average number of used DL PRBs [CDBH]":"DL PRB Utilization (CDBH)",
                          "Avg Connected User [CDBH]":"Avg RRC Connectd Users (CDBH)",
                          "RRC Setup Success Rate [CDBH]":"RRC SR (CDBH)",
                          "ERAB Setup Success Rate [CDBH]":"RAB SR (CDBH)",
                          "VoLTE DCR [CBBH]":"VoLTE DCR (CBBH)",
                          }
                dif_tech_data={}
                for kpi in KPI:
                    if kpi =="VoLTE Traffic [CBBH]":
                            band_list=[1800,2100,900,2300]
                            # #print("Band List:-",band_list)
                            
                            for band in band_list:  
                                
                                # #print(band,freq_band)

                                if band == freq_band: # to compare the de grow technology
                                    # #print(band)

                                    df_fil=df[df.Freq_Band == freq_band]
                                    if band == 2300:
                                        # if df_fil['Short name'].str.contains("_T1_"):__
                                            if sector == "A":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "B":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "C":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "D":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                           
                                            # #print("df_fil_t1---------")
                                            #print(df_fil_t1)
                                            row_sum_t1 = df_fil_t1.sum(axis=0)
                                            
                                            t=row_sum_t1[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC1")]=t
                                        # if df_fil['Short name'].str.contains("_T2_"):
                                            if sector == "A":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "B":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]

                                            if sector == "C":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                        
                                            if sector == "D":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                           
                                            # #print("df_fil_t2---------")
                                            # #print(df_fil_t2)
                                            row_sum_t2 = df_fil_t2.sum(axis=0)
                                            t=row_sum_t2[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC2")]=t
                                
                                    else:
                                        if sector == "A":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")]
                                        if sector == "B":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")]
                                        if sector == "C":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")]
                                        if sector == "D":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")]
                                        

                                        row_sum = df_fil.sum(axis=0)
                                        if band==1800:
                                        
                                            t=row_sum[kpi]
                                            dif_tech_data[(KPI_dic[kpi]
                                            ,"FD18")]=t
                                        if band==2100:
                                        
                                            t=row_sum[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"FD21")]=t
                                        if band==900:
                                            
                                            t=row_sum[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"FD8")]=t
                                        
                                        #print(df_fil)

                                        
                                    
                                    
                                else :
                                    df_fil=df[df.Freq_Band == band]
                                    # #print(df_fil)
                                    # #print(band)
                                    if band == 2300:
                                        # if df_fil['Short name'].str.contains("_T1_"):__
                                            if sector == "A":
                                                
                                                df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "B":
                                                
                                                df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "C":
                                                
                                                df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "D":
                                                
                                                df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                            
                                            # #print("df_fil_t1---------")
                                            # #print(df_fil_t1)
                                            row_sum_t1 = df_fil_t1.sum(axis=0)
                                            
                                            t=row_sum_t1[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC1")]=t
                                        # if df_fil['Short name'].str.contains("_T2_"):
                                            if sector == "A":
                                                
                                                df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "B":
                                                
                                                df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "C":
                                                
                                                df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "D":
                                                
                                                df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                            #print("df_fil_t2---------")

                                            # #print(df_fil_t2)
                                            row_sum_t2 = df_fil_t2.sum(axis=0)
                                            t=row_sum_t2[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC2")]=t


                                    else:

                                            if sector == "A":
                                                    
                                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")]
                                            if sector == "B":
                                                
                                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")]
                                            if sector == "C":
                                                
                                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")]
                                            if sector == "D":
                                                
                                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")]
                                            

                                            row_sum = df_fil.sum(axis=0)
                                            if band==1800:
                                                
                                                t=row_sum[kpi]
                                                # data[i]={("MV_4G Data Volume_GB","FD18"):t}
                                                dif_tech_data[(KPI_dic[kpi],"FD18")]=t
                                            if band==2100:
                                                
                                                t=row_sum[kpi]
                                                # data[i]={("MV_4G Data Volume_GB","FD21"):t}
                                                dif_tech_data[(KPI_dic[kpi],"FD21")]=t
                                            if band==900:
                                                
                                                t=row_sum[kpi]
                                                # data[i]={("MV_4G Data Volume_GB","FD8"):t}
                                                dif_tech_data[(KPI_dic[kpi],"FD9")]=t
                            row_sum = df.sum(axis=0)                
                            t=row_sum[kpi]
                            dif_tech_data[(KPI_dic[kpi],"NW")]=t
                    else:
                            band_list=[1800,2100,900,2300]
                            # #print("Band List:-",band_list)
                            
                            for band in band_list:  
                                
                                # #print(band,freq_band)
                                # #print("Day--",K,"    KPI:--",kpi,"   Band:---",band)

                                if band == freq_band: # to compare the de grow technology
                                    # #print(band)

                                    df_fil=df[df.Freq_Band == freq_band]
                                    if band == 2300:
                                        # if df_fil['Short name'].str.contains("_T1_"):__
                                            if sector == "A":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "B":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "C":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "D":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                            #print("df_fil_t1---------")
                                            #print(df_fil_t1)
                                            row_sum_t1 = df_fil_t1.mean(axis=0)
                                            
                                            t=row_sum_t1[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC1")]=t
                                        # if df_fil['Short name'].str.contains("_T2_"):
                                            if sector == "A":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "B":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "C":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "D":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                            ##print("df_fil_t2---------")

                                            ##print(df_fil_t2)
                                            row_sum_t2 = df_fil_t2.mean(axis=0)
                                            t=row_sum_t2[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC2")]=t
                                
                                    else:
                                        if sector == "A":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")]
                                        if sector == "B":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")]
                                        if sector == "C":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")]
                                        if sector == "D":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")]
                                        

                                        row_sum = df_fil.mean(axis=0)
                                        if band==1800:
                                        
                                            t=row_sum[kpi]
                                            dif_tech_data[(KPI_dic[kpi]
                                            ,"FD18")]=t
                                        if band==2100:
                                        
                                            t=row_sum[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"FD21")]=t
                                        if band==900:
                                            
                                            t=row_sum[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"FD8")]=t
                                        
                                        ##print(df_fil)

                                        
                                    
                                    
                                else :
                                    df_fil=df[df.Freq_Band == band]
                                    ##print(df_fil)
                                    ##print(band)
                                    if band == 2300:
                                        # if df_fil['Short name'].str.contains("_T1_"):__
                                            if sector == "A":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "B":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "C":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                            if sector == "D":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                            #print("df_fil_t1---------")
                                            #print(df_fil_t1)
                                            row_sum_t1 = df_fil_t1.mean(axis=0)
                                            
                                            t=row_sum_t1[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC1")]=t
                                        # if df_fil['Short name'].str.contains("_T2_"):
                                            if sector == "A":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "B":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "C":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                            if sector == "D":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                            #print("df_fil_t2---------")

                                            #print(df_fil_t2)
                                            row_sum_t2 = df_fil_t2.mean(axis=0)
                                            t=row_sum_t2[kpi]
                                            dif_tech_data[(KPI_dic[kpi],"TDDC2")]=t


                                    else:

                                        if sector == "A":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")]
                                        if sector == "B":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")]
                                        if sector == "C":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")]
                                        if sector == "D":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")]
                                            

                                        row_sum = df_fil.mean(axis=0)
                                        if band==1800:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD18"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD18")]=t
                                        if band==2100:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD21"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD21")]=t
                                        if band==900:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD8"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD9")]=t
                            row_sum = df.mean(axis=0)                
                            t=row_sum[kpi]
                            dif_tech_data[(KPI_dic[kpi],"NW")]=t
                data2[i]=dif_tech_data 
                #print("DAY-------",i)  
                
                
        
        df1out=pd.DataFrame(data1)
        df2out=pd.DataFrame(data2)
        pre_df = pd.concat([df1out, df2out])
        
        # #print(pre_df)
        return pre_df
        ############################################## ***************************** ##############################################             

def post_degrow(additional_inputs,post_file_dict):
      # Example usage

       ############################################## code for additional inputs ######################################################
      
        additional_inputs_df=pd.read_excel(additional_inputs)
        freq_band=int(additional_inputs_df.at[0,"Freq_Band"])
        
        site_id=additional_inputs_df.at[0,"Site_ID"]
        site_list=[site_id,("F" + site_id)]
        sector=additional_inputs_df.at[0,"Sector"]

        #print(additional_inputs_df)

        #print("freq_band:--",freq_band,type(freq_band))
        #print("sector:--",sector,type(sector))
        
        ######################################## ******************************** ################################################
       

       #################################################  main processing  #######################################################                  
       
        # file_list=traverse_folder(folder_path_post)
        # #print(file_list)
        K=5
        data1={}
        data2={}
        code_executed=False
        for file,date in post_file_dict.items():
        
            
            i=date
            
            ############################################ code for 24hrs ###########################
            if "_24Hrs_" in file.name:
                #print("------------------------------------------------------------------################## post ####################----------------------------------------------------")
                #print("------------------------------------------------------------------##################",i,"####################----------------------------------------------------")
                #print(".......-------------------------------------------------------------------------inside 24hrs------------------------------------------------------------.........")
                #print("yeas................")
                # path=os.path.join(folder_path_post,file)
                df=pd.read_csv(file)
                
                df.rename(columns={"Site Name" :"Site_Name" } ,inplace = True )
                df=df[(df.Site_Name.isin(site_list))]
                df['Short name'] = df['Short name'].str.strip()
                
                ################## to get the frequency band ##########################
                # band_list=df['Freq_Band'].unique()
                # band_list=pd.Series(band_list).dropna().astype(int)

                
                band_list=[1800,2100,900,2300]
                #print("Band List:-",band_list)
                dif_tech_data={}
                # row_sum = df.sum(axis=0)              
                # VoLTE_Traffic=row_sum["MV_4G Data Volume_GB"]
                # dif_tech_data[("MV_4G Data Volume_GB","NW")]=VoLTE_Traffic

                for band in band_list:  
                    
                        #print(band,freq_band)
                   
                    
                        df_fil=df[df.Freq_Band == band]
                        #print(df_fil)
                        #print(band)

                        if band == 2300:
                            # if df_fil['Short name'].str.contains("_T1_"):__
                                if sector == "A":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "B":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "C":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                if sector == "D":
                                    
                                    df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                #print("df_fil_t1---------")
                                #print(df_fil_t1)
                                row_sum_t1 = df_fil_t1.sum(axis=0)
                                
                                t=row_sum_t1["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","TDDC1")]=t
                            # if df_fil['Short name'].str.contains("_T2_"):
                                if sector == "A":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "B":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "C":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                if sector == "D":
                                    
                                    df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                #print("df_fil_t2---------")

                                #print(df_fil_t2)
                                row_sum_t2 = df_fil_t2.sum(axis=0)
                                t=row_sum_t2["MV_4G Data Volume_GB"]
                                dif_tech_data[("DATA TRAFFIC (24hrs)","TDDC2")]=t


                        else:


                            if sector == "A":
                                    
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")]
                            if sector == "B":
                                
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")]
                            if sector == "C":
                                
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")]
                            if sector == "D":
                                
                                df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")]
                            

                            row_sum = df_fil.sum(axis=0)
                            if band==1800:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                # data[i]={("MV_4G Data Volume_GB","FD18"):t}
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD18")]=t
                            if band==2100:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                # data[i]={("MV_4G Data Volume_GB","FD21"):t}
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD21")]=t
                            if band==900:
                                
                                t=row_sum["MV_4G Data Volume_GB"]
                                # data[i]={("MV_4G Data Volume_GB","FD8"):t}
                                dif_tech_data[("DATA TRAFFIC (24hrs)","FD9")]=t
                row_sum = df.sum(axis=0)              
                VoLTE_Traffic=row_sum["MV_4G Data Volume_GB"]
                dif_tech_data[("DATA TRAFFIC (24hrs)","NW")]=VoLTE_Traffic
                data1[i]=dif_tech_data        
                
              
          ############################################################ code for bbh ###################################################
            else:
                # if not code_executed:
                #     K=6
                #     i="Day"+ str(K)
                #     code_executed=True

                #print("------------------------------------------------------------------################## post ####################----------------------------------------------------")
                #print("------------------------------------------------------------------##################",i,"####################----------------------------------------------------")
                #print(".......-------------------------------------------------------------------------inside bbh ------------------------------------------------------------.........")
                #print("yeas................")
                # path=os.path.join(folder_path_post,file)
                df=pd.read_csv(file)
                
                df.rename(columns={"Site Name" :"Site_Name" } ,inplace = True )
                df=df[(df.Site_Name.isin(site_list))]
                df['Short name'] = df['Short name'].str.strip()
                
                ################## to get the frequency band ##########################
                # band_list=df['Freq_Band'].unique()
                # band_list=pd.Series(band_list).dropna().astype(int)

                KPI=["VoLTE Traffic [CBBH]","DL User Throughput_Kbps [CDBH]","UL User Throughput_Kbps [CDBH]","Average number of used DL PRBs [CDBH]","Avg Connected User [CDBH]","RRC Setup Success Rate [CDBH]","ERAB Setup Success Rate [CDBH]","VoLTE DCR [CBBH]"]
                KPI_dic ={
                        "VoLTE Traffic [CBBH]":"VOLTE TRAFFIC (CBBH)",
                        "DL User Throughput_Kbps [CDBH]":"DL THPT (CDBH)",
                        "UL User Throughput_Kbps [CDBH]":"UL THPT (CUBH)",
                        "Average number of used DL PRBs [CDBH]":"DL PRB Utilization (CDBH)",
                        "Avg Connected User [CDBH]":"Avg RRC Connectd Users (CDBH)",
                        "RRC Setup Success Rate [CDBH]":"RRC SR (CDBH)",
                        "ERAB Setup Success Rate [CDBH]":"RAB SR (CDBH)",
                        "VoLTE DCR [CBBH]":"VoLTE DCR (CBBH)",

                        }
                dif_tech_data={}
                for kpi in KPI:
                    if kpi == "VoLTE Traffic [CBBH]":
                            band_list=[1800,2100,900,2300]
                            #print("Band List:-",band_list)
                            
                            for band in band_list:  
                                
                                #print(band,freq_band)

                                
                                        
                                    
                                    
                                
                                df_fil=df[df.Freq_Band == band]
                                #print(df_fil)
                                #print(band)
                                if band == 2300:
                                    # if df_fil['Short name'].str.contains("_T1_"):__
                                        if sector == "A":
                                            df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                        if sector == "B":
                                    
                                            df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                        if sector == "C":
                                    
                                            df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                        if sector == "D":
                                    
                                            df_fil_t1 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                        #print("df_fil_t1---------")
                                        #print(df_fil_t1)
                                        row_sum_t1 = df_fil_t1.sum(axis=0)
                                        
                                        t=row_sum_t1[kpi]
                                        dif_tech_data[(KPI_dic[kpi],"TDDC1")]=t
                                    # if df_fil['Short name'].str.contains("_T2_"):
                                        if sector == "A":
                                    
                                            df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                        if sector == "B":
                                    
                                            df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                        if sector == "C":
                                    
                                            df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                        if sector == "D":
                                    
                                            df_fil_t2 = df_fil[((df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                        #print("df_fil_t2---------")

                                        #print(df_fil_t2)
                                        row_sum_t2 = df_fil_t2.sum(axis=0)
                                        t=row_sum_t2[kpi]
                                        dif_tech_data[(KPI_dic[kpi],"TDDC2")]=t


                                else:

                                        if sector == "A":
                                                
                                            df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("L")]
                                        if sector == "B":
                                            
                                            df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("M")]
                                        if sector == "C":
                                            
                                            df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("N")]
                                        if sector == "D":
                                            
                                            df_fil = df_fil[(df_fil['Short name'].str.endswith(sector)) | df_fil['Short name'].str.endswith("O")]
                                        

                                        row_sum = df_fil.sum(axis=0)
                                        if band==1800:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD18"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD18")]=t
                                        if band==2100:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD21"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD21")]=t
                                        if band==900:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD8"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD9")]=t
                            row_sum = df.sum(axis=0)                
                            t=row_sum[kpi]
                            dif_tech_data[(KPI_dic[kpi],"NW")]=t
                    else:
                            band_list=[1800,2100,900,2300]
                            #print("Band List:-",band_list)
                            
                            for band in band_list:  
                                
                                #print(band,freq_band)

                                df_fil=df[df.Freq_Band == band]
                                #print(df_fil)
                                #print(band)
                                if band == 2300:
                                    # if df_fil['Short name'].str.contains("_T1_"):__
                                        if sector == "A":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T1_")]
                                        if sector == "B":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T1_")]
                                        if sector == "C":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T1_")]
                                        if sector == "D":
                                                
                                                df_fil_t1 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T1_")]
                                        #print("df_fil_t1---------")
                                        #print(df_fil_t1)
                                        row_sum_t1 = df_fil_t1.mean(axis=0)
                                        
                                        t=row_sum_t1[kpi]
                                        dif_tech_data[(KPI_dic[kpi],"TDDC1")]=t
                                    # if df_fil['Short name'].str.contains("_T2_"):
                                        if sector == "A":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")) & df_fil['Short name'].str.contains("_T2_")]
                                        if sector == "B":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")) & df_fil['Short name'].str.contains("_T2_")]
                                        if sector == "C":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")) & df_fil['Short name'].str.contains("_T2_")]
                                        if sector == "D":
                                                
                                                df_fil_t2 = df_fil[(df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")) & df_fil['Short name'].str.contains("_T2_")]
                                        #print("df_fil_t2---------")

                                        #print(df_fil_t2)
                                        row_sum_t2 = df_fil_t2.mean(axis=0)
                                        t=row_sum_t2[kpi]
                                        dif_tech_data[(KPI_dic[kpi],"TDDC2")]=t


                                else:

                                        if sector == "A":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("L")]
                                        if sector == "B":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("M")]
                                        if sector == "C":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("N")]
                                        if sector == "D":
                                            
                                            df_fil = df_fil[df_fil['Short name'].str.endswith(sector) | df_fil['Short name'].str.endswith("O")]
                                        

                                        row_sum = df_fil.mean(axis=0)
                                        if band==1800:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD18"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD18")]=t
                                        if band==2100:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD21"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD21")]=t
                                        if band==900:
                                            
                                            t=row_sum[kpi]
                                            # data[i]={("MV_4G Data Volume_GB","FD8"):t}
                                            dif_tech_data[(KPI_dic[kpi],"FD9")]=t
                            row_sum = df.mean(axis=0)                
                            t=row_sum[kpi]
                            dif_tech_data[(KPI_dic[kpi],"NW")]=t

                data2[i]=dif_tech_data        
                
        
        df1out=pd.DataFrame(data1)
        df2out=pd.DataFrame(data2)
        post_df = pd.concat([df1out, df2out])
        # #print(post_df)
        return post_df



@api_view(["POST","GET"])
# @authentication_classes([TokenAuthentication])
# @permission_classes([IsAuthenticated])
def Degrow_trend(request):
     
    # additional_input_file_path=r"E:\Mcom_Projects_files\Pre-Post degrow\referance files2\additional requirement.xlsx" # mhra02
    # additional_input_file_path=r"E:\Mcom_Projects_files\Pre-Post degrow\referance files1\additional requirement_ABLA16.xlsx"  #for abla16
    # additional_input_file_path=r"E:\Mcom_Projects_files\Pre-Post degrow\referance files1\additional requirement_ASDH16.xlsx"  #for asdh16


    # folder_path_pre = r"C:\Users\Lenovo\Desktop\general python\test\pre" # Replace with the actual folder path # mhra02
    # folder_path_pre = r"E:\Mcom_Projects_files\Pre-Post degrow\referance files1\inputs\PRE" # Replace with the actual folder path abla16
    # folder_path_pre = r"E:\Mcom_Projects_files\Pre-Post degrow\referance files1\inputs\PRE" # Replace with the actual folder path asdh16


    # folder_path_post = r"C:\Users\Lenovo\Desktop\general python\test\post" # Replace with the actual folder path # mhra02
    # folder_path_post = r"E:\Mcom_Projects_files\Pre-Post degrow\referance files1\inputs\POST" # Replace with the actual folder path for abla16
    # folder_path_post = r"E:\Mcom_Projects_files\Pre-Post degrow\referance files1\inputs\POST" # Replace with the actual folder path for asdh16
    #print(request.FILES)

    additional_inputs = request.FILES["additional_inputs"] if 'additional_inputs' in request.FILES else None
    ######################################  dev pre files  #################################################
    # pre_d1_24 = request.FILES["pre_d1_24"] if 'pre_d1_24' in request.FILES else None
    # pre_d2_24 = request.FILES["pre_d2_24"] if 'pre_d2_24' in request.FILES else None
    # pre_d3_24 = request.FILES["pre_d3_24"] if 'pre_d3_24' in request.FILES else None
    # pre_d4_24 = request.FILES["pre_d4_24"] if 'pre_d4_24' in request.FILES else None
    # pre_d5_24 = request.FILES["pre_d5_24"] if 'pre_d5_24' in request.FILES else None

    # pre_d1_bbh = request.FILES["pre_d1_bbh"] if 'pre_d1_bbh' in request.FILES else None
    # pre_d2_bbh = request.FILES["pre_d2_bbh"] if 'pre_d2_bbh' in request.FILES else None
    # pre_d3_bbh = request.FILES["pre_d3_bbh"] if 'pre_d3_bbh' in request.FILES else None
    # pre_d4_bbh = request.FILES["pre_d4_bbh"] if 'pre_d4_bbh' in request.FILES else None
    # pre_d5_bbh = request.FILES["pre_d5_bbh"] if 'pre_d5_bbh' in request.FILES else None

    pre_d1_24 = request.FILES["pre_0"] if 'pre_0' in request.FILES else None
    pre_d2_24 = request.FILES["pre_1"] if 'pre_1' in request.FILES else None
    pre_d3_24 = request.FILES["pre_2"] if 'pre_2' in request.FILES else None
    pre_d4_24 = request.FILES["pre_3"] if 'pre_3' in request.FILES else None
    pre_d5_24 = request.FILES["pre_4"] if 'pre_4' in request.FILES else None

    pre_d1_bbh = request.FILES["pre_5"] if 'pre_5' in request.FILES else None
    pre_d2_bbh = request.FILES["pre_6"] if 'pre_6' in request.FILES else None
    pre_d3_bbh = request.FILES["pre_7"] if 'pre_7' in request.FILES else None
    pre_d4_bbh = request.FILES["pre_8"] if 'pre_8' in request.FILES else None
    pre_d5_bbh = request.FILES["pre_9"] if 'pre_9' in request.FILES else None
    
    #print(".............................................................pre files name....................................................................")
    #print(pre_d1_24)
    #print(pre_d2_24)
    #print(pre_d3_24)
    #print(pre_d4_24)
    #print(pre_d5_24)

    #print( pre_d1_bbh )
    #print( pre_d2_bbh )
    #print( pre_d3_bbh )
    #print( pre_d4_bbh )
    #print( pre_d5_bbh )

    pre_file_list=[pre_d1_24, pre_d2_24,pre_d3_24, pre_d4_24, pre_d5_24, pre_d1_bbh, pre_d2_bbh, pre_d3_bbh, pre_d4_bbh, pre_d5_bbh]
    
    pre_file_dict={}
    for f_obj in pre_file_list:
         if f_obj:
            pre_file_dict[f_obj]= extract_date(f_obj.name)
    pre_file_dict = dict(sorted(pre_file_dict.items(), key=lambda x: x[1]))
    #print(pre_file_dict)
    for key, value in pre_file_dict.items():
            #print(key, value)
        pass
   
   
    # exit(0)
         


         
        
    ################################ ******************** #######################################


    ######################################  dev post files  #################################################
    post_d1_24 = request.FILES["post_0"] if 'post_0' in request.FILES else None
    post_d2_24 = request.FILES["post_1"] if 'post_1' in request.FILES else None
    post_d3_24 = request.FILES["post_2"] if 'post_2' in request.FILES else None
    post_d4_24 = request.FILES["post_3"] if 'post_3' in request.FILES else None
    post_d5_24 = request.FILES["post_4"] if 'post_4' in request.FILES else None

    post_d1_bbh = request.FILES["post_5"] if 'post_5' in request.FILES else None
    post_d2_bbh = request.FILES["post_6"] if 'post_6' in request.FILES else None
    post_d3_bbh = request.FILES["post_7"] if 'post_7' in request.FILES else None
    post_d4_bbh = request.FILES["post_8"] if 'post_8' in request.FILES else None
    post_d5_bbh = request.FILES["post_9"] if 'post_9' in request.FILES else None
   
    
        
          
    #print(".............................................................post files name....................................................................")
    #print(post_d1_24)
    #print(post_d2_24)
    #print(post_d3_24)
    #print(post_d4_24)
    #print(post_d5_24)

    #print( post_d1_bbh )
    #print( post_d2_bbh )
    #print( post_d3_bbh )
    #print( post_d4_bbh )
    #print( post_d5_bbh )

    post_file_list=[post_d1_24,post_d2_24,post_d3_24,post_d4_24,post_d5_24,post_d1_bbh, post_d2_bbh, post_d3_bbh, post_d4_bbh, post_d5_bbh]


    post_file_dict={}
    for f_obj in post_file_list:
         if f_obj:
            post_file_dict[f_obj]= extract_date(f_obj.name)
    post_file_dict = dict(sorted(post_file_dict.items(), key=lambda x: x[1]))
    #print(post_file_dict)\
    
    for key, value in post_file_dict.items():
            #print(key, value)
         pass
    ################################ ******************** #######################################
    # exit(0)
    
    ################################## dev post files ############################################
    
    ##############################################################################################
    
    # pre_degrow_df=pre_degrow(additional_input_file_path,folder_path_pre)
    pre_degrow_df=pre_degrow(additional_inputs,pre_file_dict)
    # #print("------------------------------------- Pre degrow ---------------------------------------")
    # #print(pre_degrow_df)
    
    # exit(0)
    # post_degrow_df =post_degrow(additional_input_file_path,folder_path_post)
    post_degrow_df =post_degrow(additional_inputs,post_file_dict)
    # #print("------------------------------------- Post degrow ---------------------------------------")
    # #print(post_degrow_df)
    
    degrow_output_df = pd.concat([pre_degrow_df, post_degrow_df], axis=1)
    #print(degrow_output_df)
    
    
    
    degrow_output_df['Pre avg'] = round(pre_degrow_df.mean(axis=1),2)
    degrow_output_df['Post avg'] = round(post_degrow_df.mean(axis=1),2)
    degrow_output_df['Delta'] = round(degrow_output_df['Post avg']-degrow_output_df['Pre avg'],2)

    degrow_output_df['% Change'] = round(((degrow_output_df['Delta'] /degrow_output_df['Pre avg']) *100),2)

    #print(degrow_output_df)
    
   
    ###################### for giving name to the index ########################

    degrow_output_df.reset_index(inplace=True)
    degrow_output_df.rename( columns={'level_0':'KPI'}, inplace=True )
    degrow_output_df.rename( columns={'level_1':'TECH'}, inplace=True)
    degrow_output_df.set_index(['KPI', 'TECH'], inplace=True)
    ######################### ************************* ########################
    tuples=[]
    cols_pre_df=pre_degrow_df.columns
    no_pre_cols=len(cols_pre_df)
    cols_post_df=post_degrow_df.columns
    no_post_cols = len(cols_post_df)

    for pre_col in cols_pre_df:
         tuples.append(("Pre",pre_col))
    for post_col in cols_post_df:
         tuples.append(("Post",post_col))
        
    comp=[("Comparision","Pre Avg"),
            ("Comparision","Post Avg"),
            ("Comparision","Delta"),
            ("Comparision","% Change")]
    tuples.extend(comp)
    #print(tuples)

    cols = pd.MultiIndex.from_tuples(tuples)
    degrow_output_df.columns=cols

    degrow_output_df=degrow_output_df.round(2)
    #print(degrow_output_df)

#################################################################################################
    saving_path=os.path.join(MEDIA_ROOT,'Degrow',"v3degrow_output.xlsx")
    #print(saving_path)
    degrow_output_df.to_excel(saving_path,sheet_name=SITE_ID)

    # wb = openpyxl.load_workbook(r"C:\Users\Lenovo\Desktop\general python\v3degrow_output.xlsx")
    wb = openpyxl.load_workbook(saving_path)
    ws = wb[SITE_ID]
    
    ws=template_colouring(ws,no_pre_cols, no_post_cols)
###########################################################################################################
    
    output_saving_path=os.path.join(MEDIA_ROOT,'Degrow',"v3degrow_output_coloured.xlsx")
    #print(output_saving_path)
    wb.save(output_saving_path)
    # trend_wb.save("opopopop.xlsx")
       
    # wb.save("v3degrow_output_coloured.xlsx")
    #print("excel operation done.........................")
    download_path=os.path.join(MEDIA_URL,'Degrow',"v3degrow_output_coloured.xlsx")
    print("..................................................download path fetched..........................................")
    return Response({"status":True,"message":"Succesfully uploaded","Download_url":download_path})