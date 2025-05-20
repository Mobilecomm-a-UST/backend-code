import logging
from django.shortcuts import render

# Create your views here.from django.shortcuts import render
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
import os
from django.shortcuts import render

from django.shortcuts import render
from django.shortcuts import render,redirect
from django.http import HttpResponse

import sys
import pandas as pd
from django.contrib import messages
# import json
from django.contrib.auth.decorators import login_required
# from mcom_websites.settings import MEDIA_URL,BASE_DIR,ALLOWED_HOSTS
import numpy as np
import os
import datetime
from rest_framework.decorators import api_view,parser_classes
from rest_framework.response import Response
from django.http import JsonResponse
# from .serializer import *
from django.forms.models import model_to_dict

from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import authentication_classes,permission_classes
from django.core.files.storage import FileSystemStorage
# from .settings import MEDIA_ROOT
from .models import *

import openpyxl
from openpyxl.styles import PatternFill
# Create your views here.
from openpyxl import Workbook,load_workbook
from statistics import mean

from django.db.models import Sum


import pandas as pd
import matplotlib  as mpl
import numpy as np

from datetime import date, timedelta

@api_view(["POST"])
def old_or_trend(request):
    try:
        raw_kpi = request.FILES["raw_kpi"] if 'raw_kpi' in request.FILES else None
        if raw_kpi:
             location=MEDIA_ROOT+r'\trends\temporary_files'
             fs=FileSystemStorage(location=location)
             file=fs.save(raw_kpi.name,raw_kpi)
             file_path=fs.path(file)
             df_raw_kpi=pd.read_excel(file_path)
             print(df_raw_kpi)
            #  os.remove(path=file_path)
        site_list=request.FILES['site_list'] if 'site_list' in request.FILES else None
        if site_list:
             location=MEDIA_ROOT+r'\trends\temporary_files'
             fs=FileSystemStorage(location=location)
             file=fs.save(site_list.name,site_list)
             file_path=fs.path(file)  
             df_site_list=pd.read_excel(file_path)  
             print(df_site_list)
             os.remove(path=file_path) 

        door_root= os.path.join(MEDIA_ROOT,'trends',"or","OR_4G")
        
        
        
        df_raw_kpi["Short name"] = df_raw_kpi["Short name"].fillna(method="ffill")  #######for (forward fill)ffill is used to copy  and fill
        df_raw_kpi.columns.values[1] = 'Date'  #####for empty colum

        df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"] / 1024)  #####for change kbps to mbps
        df_raw_kpi.rename(columns={"MV_DL User Throughput_Kbps [CDBH]" :"MV_DL User Throughput_Mbps [CDBH]" } , inplace = True )
       
        df_raw_kpi.fillna(value=0,inplace=True)
        # df_raw_kpi.columns.values[23] ='MV_DL User Throughput_Mbps [CDBH]'
        
        # A = []
        # for x in df_raw_kpi['4G_ECGI']:
        #     if("-" in x):
        #         cell_id = x.split('-')[-2][:6]
        #         A.append(cell_id)
        #     else:
        #         cell_id = x[:-3]
        #         A.append(cell_id)
        # df_raw_kpi.insert(0,'cell_id',A)

        # # A = []
        # B = []
        # c = []
        # for x in df_raw_kpi['Short name']:
        #     if('_' in x):
        #         id=x.split('_')[-2][:-1]
        #         c.append(id)
        #     elif('-' in x):
        #         id=x.split('-')[-3]
        #         c.append(id) 
        #     else:
        #         id=x[-3]   
        #     if('-' in x):
        #         mrbts_name = x.split(' ')[0][:-3]
        #         B.append(mrbts_name)
        #     else:
        #         mrbts_name = x[:-3]
        #         B.append(mrbts_name) 
        A = []
        B = []
        for x in df_raw_kpi['Short name']:
            if("-" in str(x)):
                mrbts_name = str(x).split(' ')[0][:-3]
                A.append(mrbts_name)
            else:
                mrbts_name = str(x)[:-3]
                A.append(mrbts_name)

            if("_" in str(x)):
                id=x.split("_")[-2][:-1]
                B.append(id)
            elif("_" in str(x)):
                id=x.split("_")[-3]
                B.append(id) 
            else:
                id=str(x)[-3]  
                B.append(id)

        A = []
        for x in df_raw_kpi['4G_ECGI']:
            if("-" in str(x)):
                cell_id = str(x).split('-')[-2][:6]
                A.append(cell_id)
            else:
                cell_id = str(x)[:-3]
                A.append(cell_id)
        df_raw_kpi.insert(3,'cell_id',A)
        print(df_raw_kpi)
        # df.to_excel('process_output/smalcell.xlsx')

      


        df_raw_kpi.insert(1, "mrbts_name",A)
        df_raw_kpi.insert(2,'Site ID',B)
        process_op_path=os.path.join(door_root,"process_output",'desired input.xlsx')
        
        
        df_raw_kpi.to_excel(process_op_path, index=False)

        excel_file_1 = process_op_path
        df1 = pd.read_excel(excel_file_1) 

        df1.rename(columns={"Site ID": "SITE_ID"}, inplace=True)
 
        kpi = ['RRC Setup Success Rate [CDBH]','ERAB Setup Success Rate [CDBH]','PS Drop Call Rate % [CDBH]',
            'MV_DL User Throughput_Mbps [CDBH]','MV_UL User Throughput_Kbps [CDBH]','UL RSSI [CDBH]','Average number of used DL PRBs [CDBH]'
            ,'MV_PS handover success rate [LTE Intra System] [CDBH]','MV_PS handover success rate [LTE Inter System] [CDBH]','MV_CSFB Redirection Success Rate [CDBH]',
            'MV_VoLTE ERAB Setup Success Rate [CBBH]','MV_VoLTE DCR [CBBH]','VoLTE Packet Loss DL [CBBH]','VoLTE Packet Loss UL [CBBH]','VoLTE SRVCC Per Call Rate [CBBH]',
            'VoLTE Intra-LTE Handover Success Ratio [CBBH]','VoLTE Inter-Frequency Handover Success Ratio [CBBH]','4G Data Volume [GB]','VoLTE Traffic','E-UTRAN Average CQI [CDBH]']

        # filtered_df_1 = df1[(df1.Site_ID.isin(list(df_site_list['SITE_ID'])))]

        # # print(filtered_df_1)
        # or_path=os.path.join(door_root,'process_output','filtered_df_1.xlsx')
        # filtered_df_1.to_excel(or_path,index=False)
        # # df1 = pd.read_excel(or_filtr)
        # df1 = pd.read_excel(or_path)


        filtered_df_1 = df1[(df1.SITE_ID.isin(list(df_site_list['SITE_ID'])))]

        print(filtered_df_1)
        mp_filtr=os.path.join(door_root,'process_output','filtered_df_1.xlsx')
        filtered_df_1.to_excel(mp_filtr,index=False)
        df1 = pd.read_excel(mp_filtr)

        df_pivot = df1.pivot_table( values=kpi,columns='Date', index=['Short name','SITE_ID','mrbts_name','cell_id'])
        orPathpivot=os.path.join(door_root,'process_output','pivot.xlsx')
        df_pivot.to_excel(orPathpivot)
        
        path_of_blnk_temp=os.path.join(door_root,"template","Small_cell_KPI_trend.xlsx")
        wb=openpyxl.load_workbook(path_of_blnk_temp)
        ws=wb['OK']


        alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
        def titleToNumber(s):
            result= 0
            for B in range(len(s)):
                result *= 26
                result += ord(s[B]) - ord('A') + 1
            return result  
            # result += ord(s[B]) - ord('A') + 1
            # return result

        str_date=request.POST.get("offered_date")
        date1=datetime.datetime.strptime(str_date,'%Y-%m-%d')
        d1=date1-timedelta(1)
        d2=date1-timedelta(2)
        d3=date1-timedelta(3)
        d4=date1-timedelta(4)
        d5=date1-timedelta(5)
        cl=[d1,d2,d3,d4,d5]

        
        def overwrite(kpi_name,coln1):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            print(kpi_name)
            index=df_pivot.index
            # print("index ;###############################",index_pivot)
            # print(len(index_pivot))
            print("doooooonnenn")
            dr=df_pivot[kpi_name]
            print("columns of dr tab")
            li=dr.columns
            print("column list")

            col1=dr[li[0]].to_list()
            col2=dr[li[1]].to_list()
            col3=dr[li[2]].to_list()
            col4=dr[li[3]].to_list()
            col5=dr[li[4]].to_list()

            ws[coln1+"3"].value=cl[4]
            ws[coln2+"3"].value=cl[3]
            ws[coln3+"3"].value=cl[2]
            ws[coln4+"3"].value=cl[1]
            ws[coln5+"3"].value=cl[0]

            for i,value in enumerate(index):
                j=i+4
                ws['N'+str(j)].value='LPSC Small Cell'
                ws['M'+str(j)].value='NOKIA'
                ws['J'+str(j)].value='TDD'

            # ws['L'+str(j)].value=date1
        
                # ws['B'+str(j)].value=index[i][1]
                # ws['C'+str(j)].value=index[i][3]
                # ws['D'+str(j)].value=index[i][3]
                # ws['F'+str(j)].value=index[i][2]
                # ws['G'+str(j)].value=index[i][2]
                # ws['H'+str(j)].value=index[i][0]
                ws['B'+str(j)].value=index[i][0]
                ws['C'+str(j)].value=index[i][2]
                ws['D'+str(j)].value=index[i][2]
                ws['F'+str(j)].value=index[i][3]
                ws['G'+str(j)].value=index[i][3]
                ws['H'+str(j)].value=index[i][1]

                ws[coln1 + str(j)].value = col1[i]
                ws[coln2 + str(j)].value = col2[i]
                ws[coln3 + str(j)].value = col3[i]
                ws[coln4 + str(j)].value = col4[i]
                ws[coln5 + str(j)].value = col5[i]

   
                       
        # for kpi_name in kpi:
        #         print("_______________________________________done_________________________________")
        #         if(kpi_name=='RRC Setup Success Rate [CBBH]'):
        #           overwrite(kpi_name,'P')

        #         if(kpi_name=='VoLTE ERAB Setup Success Rate [CBBH]'):
        #                     overwrite(kpi_name,'V') 

        #         if(kpi_name=='PS Drop Call Rate % [CDBH]'):
        #             overwrite(kpi_name,'AB')

        #         if(kpi_name=='MV_DL User Throughput_Mbps [CDBH]'):
                    
        #             overwrite(kpi_name,'AH')
        #         print('____________________________________dl___________________________')

        #         if(kpi_name=='MV_UL User Throughput_Kbps [CDBH]'):
        #             overwrite(kpi_name,'AN') 
                    
        #         if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
        #             overwrite(kpi_name,'AT') 

        #         if(kpi_name=='PS handover success rate [LTE Inter System] [CBBH]'):
        #             overwrite(kpi_name,'AZ')
                        
        #         if(kpi_name=='MV_CSFB Redirection Success Rate [CDBH]'):
        #             overwrite(kpi_name,'BF') 

        #         if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
        #             overwrite(kpi_name,'BL') 

        #         if(kpi_name=='VoLTE Packet Loss DL [CBBH]'):
        #             overwrite(kpi_name,'BX')    

        #         if(kpi_name=='VoLTE DCR [CBBH]'):
        #             overwrite(kpi_name,'BR')    

        #         if(kpi_name=='VoLTE Packet Loss UL [CBBH]'):
        #             overwrite(kpi_name,'CD')  

        #         if(kpi_name=='VoLTE SRVCC SR'):
        #             overwrite(kpi_name,'CJ')   

        #         if(kpi_name=='VoLTE Intra-LTE Handover Success Ratio [CBBH]'):
        #             overwrite(kpi_name,'CP')  

        #         if(kpi_name=='VoLTE Inter-Frequency Handover Success Ratio [CBBH]'):
        #             overwrite(kpi_name,'CV')  

        #         if(kpi_name=='UL RSSI [CDBH]'):
        #             overwrite(kpi_name,'DH')   

        #         if(kpi_name=='4G Data Volume [GB]'):
        #             overwrite(kpi_name,'DN')  

        #         if(kpi_name=='VoLTE Traffic'):
        #             overwrite(kpi_name,'DV')  

        #         if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
        #             overwrite(kpi_name,'DB')    

        #         if(kpi_name=='Average number of used DL PRBs'):
        #             overwrite(kpi_name,'EB') 
        for kpi_name in kpi:
            if(kpi_name=='RRC Setup Success Rate [CDBH]'):
                overwrite(kpi_name,'P')

            if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
                overwrite(kpi_name,'W') 

            if(kpi_name=='PS Drop Call Rate % [CDBH]'):
                overwrite(kpi_name,'AD')

            if(kpi_name=='MV_DL User Throughput_Mbps [CDBH]'):
                overwrite(kpi_name,'AK')

            if(kpi_name=='MV_UL User Throughput_Kbps [CDBH]'):
                overwrite(kpi_name,'AR') 
                
            if(kpi_name=='MV_PS handover success rate [LTE Intra System] [CDBH]'):
                overwrite(kpi_name,'AY') 

            if(kpi_name=='MV_PS handover success rate [LTE Inter System] [CDBH]'):
                overwrite(kpi_name,'BF')
                    
            if(kpi_name=='MV_CSFB Redirection Success Rate [CDBH]'):
                overwrite(kpi_name,'BM') 

            if(kpi_name=='MV_VoLTE ERAB Setup Success Rate [CBBH]'):
                overwrite(kpi_name,'BT') 

            if(kpi_name=='MV_VoLTE DCR [CBBH]'):
                overwrite(kpi_name,'CA')    

            if(kpi_name=='VoLTE Packet Loss DL [CBBH]'):
                overwrite(kpi_name,'CH')    


            if(kpi_name=='VoLTE Packet Loss UL [CBBH]'):
                overwrite(kpi_name,'CO')  

            if(kpi_name=='VoLTE SRVCC Per Call Rate [CBBH]'):
                overwrite(kpi_name,'CV')   

            if(kpi_name=='VoLTE Intra-LTE Handover Success Ratio [CBBH]'):
                overwrite(kpi_name,'DC')  

            if(kpi_name=='VoLTE Inter-Frequency Handover Success Ratio [CBBH]'):
                overwrite(kpi_name,'DJ')  


            if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
                overwrite(kpi_name,'DQ')   

            if(kpi_name=='UL RSSI [CDBH]'):
                overwrite(kpi_name,'DX')   

            if(kpi_name=='4G Data Volume [GB]'):
                overwrite(kpi_name,'EE')  

            if(kpi_name=='VoLTE Traffic'):
                overwrite(kpi_name,'EN')  

            
            if(kpi_name=='Average number of used DL PRBs [CDBH]'):
                overwrite(kpi_name,'EU') 
            


    




         
        SaveOutput=os.path.join(door_root, 'output','ortrendoutput.xlsx')        
        wb.save(SaveOutput) 
            # print("_______________________________---save_____________________")
        download_path=os.path.join(MEDIA_URL,'trends','or','OR_4G','output','ortrendoutput.xlsx')
        return Response({"status":True,"message":"Succesfully uploaded","Download_url":download_path})
    
    except ValueError as ve:
        logging.error(f"ValueError occurred: {ve}")
        return Response({'status': False, 'message': 'Invalid value provided.'})
    except TypeError as te:
        logging.error(f"TypeError occurred: {te}")
        return Response({'status': False, 'message': 'Invalid type provided.'})
    except Exception as e:
        logging.error(f"Unexpected error occurred: {e}")
        return Response({'status': False, 'message': 'An unexpected error occurred.'})

#############################################----------------------------------------------------------------########################
      
@api_view(["POST"])
def old_or_Trend_rna(request):
    try:
        raw_kpi = request.FILES["raw_kpi"] if 'raw_kpi' in request.FILES else None
        if raw_kpi:
             location=MEDIA_ROOT+r'\trends\temporary_files'
             fs=FileSystemStorage(location=location)
             file=fs.save(raw_kpi.name,raw_kpi)
             file_path=fs.path(file)
             df_raw_kpi=pd.read_excel(file_path)
             print(df_raw_kpi)
            #  os.remove(path=file_path)
        site_list=request.FILES['site_list'] if 'site_list' in request.FILES else None
        if site_list:
             location=MEDIA_ROOT+r'\trends\temporary_files'
             fs=FileSystemStorage(location=location)
             file=fs.save(site_list.name,site_list)
             file_path=fs.path(file)  
             df_site_list=pd.read_excel(file_path)  
             print(df_site_list)
             os.remove(path=file_path) 


        door_root= os.path.join(MEDIA_ROOT,'trends',"or","RNA")
        path_of_blnk_temp=os.path.join(door_root,"templates","Relocation.xlsx")
        trend_wb=load_workbook(path_of_blnk_temp)
        print(trend_wb.sheetnames)
        ##################

        # path_of_blnk_temp=os.path.join(door_root,"template","template2.xlsx")
        # trend_wb=load_workbook(path_of_blnk_temp)
        # print(trend_wb.sheetnames)

        # trend_ws=trend_wb["L1800"]
        # trend_


####################
        df_raw_kpi["Short name"] = df_raw_kpi["Short name"].fillna(method="ffill")  #######for (forward fill)ffill is used to copy  and fill
        df_raw_kpi.columns.values[1] = 'Date'  #####for empty colum
        
        
        df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"] / 1024)  #####for change kbps to mbps
        df_raw_kpi.rename(columns={"MV_DL User Throughput_Kbps [CDBH]" :" MV_DL User Throughput_Mbps [CDBH]" } , inplace = True )

        df_raw_kpi.columns.values[3] = 'MV_DL User Throughput_Mbps [CDBH]'


        split=[]
        for cell in df_raw_kpi['Short name']:
            if("_" in cell):
                sit_id = cell.split('_')[-2][:7]
                # print(sit_id)
                split.append(sit_id)
            else:
                sit_id=cell[:7]
                split.append(sit_id)
        df_raw_kpi.insert(2,'SITE_ID',split)

        ecgi=[]
        add=[]
        for cell in df_raw_kpi['4G_ECGI']:
            if("-" in cell):
                eNode_ID = cell.split('-')[-2][:6]

                ecgi.append(eNode_ID)
        
            else:
                eNode_ID=cell[:6]
                ecgi.append(eNode_ID)

            if("-" in cell):
                lnbts_id = cell.split("-")[-1][-1]
                add.append(lnbts_id)
                    
            else:
                lnbts_id=cell[:-1]
                add.append(lnbts_id)
        df_raw_kpi.insert(3,'lnbts id',add)
        df_raw_kpi.insert(1,'eNode ID',ecgi)
        print(df_raw_kpi)
        df_raw_kpi['comb']=  df_raw_kpi['eNode ID'].astype(str) + '_' +df_raw_kpi['lnbts id']

        process_op_path=os.path.join(door_root,"process_output",'desired input.xlsx')
        
        
        df_raw_kpi.to_excel(process_op_path, index=False)
        excel_file_1 = process_op_path
        df1 = pd.read_excel(excel_file_1) 

        df1.rename(columns={"SITE_ID": "SITE_ID"}, inplace=True)

        kpi = ['MV_4G Data Volume_GB','Radio NW Availability' ,'MV_DL User Throughput_Mbps [CDBH]','MV_UL User Throughput_Kbps [CDBH]','MV_Average number of used DL PRBs [CDBH]',
              'MV_E-UTRAN Average CQI [CDBH]','VoLTE DCR [CDBH]',]



        filtered_df_1 = df1[(df1.SITE_ID.isin(list(df_site_list['SITE_ID'])))]

        print(filtered_df_1)
        mp_filtr=os.path.join(door_root,'process_output','filtered_df_1.xlsx')
        filtered_df_1.to_excel(mp_filtr,index=False)
        df1 = pd.read_excel(mp_filtr)

        df_pivot = df1.pivot_table( values=kpi,columns='Date', index=['Short name','SITE_ID','eNode ID','comb','lnbts id'])
        orPathpivot=os.path.join(door_root,'process_output','pivot.xlsx')
        df_pivot.to_excel(orPathpivot)

        alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
        def titleToNumber(s):
            result= 0
            for B in range(len(s)):
                result *= 26
                result += ord(s[B]) - ord('A') + 1
            return result  
            # result += ord(s[B]) - ord('A') + 1
            # return result

        str_date=request.POST.get("offered_date")
        date1=datetime.datetime.strptime(str_date,'%Y-%m-%d')
        d1=date1-timedelta(1)
        d2=date1-timedelta(2)
        d3=date1-timedelta(3)
        d4=date1-timedelta(4)
        d5=date1-timedelta(5)
        cl=[d1,d2,d3,d4,d5]

        
        def overwrite(kpi_name,coln1):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            print(kpi_name)
            index=df_pivot.index
            # print("index ;###############################",index_pivot)
            # print(len(index_pivot))
            print("doooooonnenn")
            dr=df_pivot[kpi_name]
            print("columns of dr tab")
            li=dr.columns
            print("column list")

            col1=dr[li[0]].to_list()
            col2=dr[li[1]].to_list()
            col3=dr[li[2]].to_list()
            col4=dr[li[3]].to_list()
            col5=dr[li[4]].to_list()

            trend_ws[coln1 + "2"].value = cl[4]
            trend_ws[coln2 + "2"].value = cl[3]
            trend_ws[coln3 + "2"].value = cl[2]
            trend_ws[coln4 + "2"].value = cl[1]
            trend_ws[coln5 + "2"].value = cl[0]

            for i,value in enumerate(index):
                j=i+3
          
            # ws['D'+str(j)].value='TDD'
                trend_ws['B'+str(j)].value=index[i][2]
                trend_ws['C'+str(j)].value=index[i][1]
            

                trend_ws[coln1 + str(j)].value = col1[i]
                trend_ws[coln2 + str(j)].value = col2[i]
                trend_ws[coln3 + str(j)].value = col3[i]
                trend_ws[coln4 + str(j)].value = col4[i]
                trend_ws[coln5 + str(j)].value = col5[i]
        trend_ws=trend_wb['High Power small cell payload']
        for kpi_name in kpi:
            if(kpi_name =='MV_4G Data Volume_GB'):
                overwrite(kpi_name,'O')

            if(kpi_name=='Radio NW Availability'):
                overwrite(kpi_name,'W') 
#######################################--------------------------------------------------------------################################
   
        def overwrite(kpi_name,coln1):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            print(kpi_name)
            index=df_pivot.index
            # print("index ;###############################",index_pivot)
            # print(len(index_pivot))
            print("doooooonnenn")
            dr=df_pivot[kpi_name]
            print("columns of dr tab")
            li=dr.columns
            print("column list")

            col1=dr[li[0]].to_list()
            col2=dr[li[1]].to_list()
            col3=dr[li[2]].to_list()
            col4=dr[li[3]].to_list()
            col5=dr[li[4]].to_list()

            trend_ws[coln1 + "2"].value = cl[4]
            trend_ws[coln2 + "2"].value = cl[3]
            trend_ws[coln3 + "2"].value = cl[2]
            trend_ws[coln4 + "2"].value = cl[1]
            trend_ws[coln5 + "2"].value = cl[0]

            for i,value in enumerate(index):
                j=i+3
          
            # ws['D'+str(j)].value='TDD'
                trend_ws['B'+str(j)].value=index[i][1]
                trend_ws['D'+str(j)].value=index[i][2]
            

                trend_ws[coln1 + str(j)].value = col1[i]
                trend_ws[coln2 + str(j)].value = col2[i]
                trend_ws[coln3 + str(j)].value = col3[i]
                trend_ws[coln4 + str(j)].value = col4[i]
                trend_ws[coln5 + str(j)].value = col5[i]
        trend_ws=trend_wb['Low Power small cell Payload']
        for kpi_name in kpi:
            if(kpi_name =='MV_4G Data Volume_GB'):
                overwrite(kpi_name,'L')

            if(kpi_name=='Radio NW Availability'):
                overwrite(kpi_name,'T') 
        #############################-------------------------------------------################################3

        def overwrite(kpi_name,coln1):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            print(kpi_name)
            index=df_pivot.index
            # print("index ;###############################",index_pivot)
            # print(len(index_pivot))
            print("doooooonnenn")
            dr=df_pivot[kpi_name]
            print("columns of dr tab")
            li=dr.columns
            print("column list")

            col1=dr[li[0]].to_list()
            col2=dr[li[1]].to_list()
            col3=dr[li[2]].to_list()
            col4=dr[li[3]].to_list()
            col5=dr[li[4]].to_list()

            trend_ws[coln1 + "2"].value = cl[4]
            trend_ws[coln2 + "2"].value = cl[3]
            trend_ws[coln3 + "2"].value = cl[2]
            trend_ws[coln4 + "2"].value = cl[1]
            trend_ws[coln5 + "2"].value = cl[0]

            for i,value in enumerate(index):
                j=i+3
          
            # ws['D'+str(j)].value='TDD'
                trend_ws['C'+str(j)].value=index[i][2]
                trend_ws['D'+str(j)].value=index[i][1]
            

                trend_ws[coln1 + str(j)].value = col1[i]
                trend_ws[coln2 + str(j)].value = col2[i]
                trend_ws[coln3 + str(j)].value = col3[i]
                trend_ws[coln4 + str(j)].value = col4[i]
                trend_ws[coln5 + str(j)].value = col5[i]
        trend_ws=trend_wb['Relocation & Sector Addition']
        for kpi_name in kpi:
            if(kpi_name =='MV_4G Data Volume_GB'):
                overwrite(kpi_name,'R')

            if(kpi_name=='Radio NW Availability'):
                overwrite(kpi_name,'W') 

        ##########################################------------------------------------------------------hpsc cell wise-####################################
        def overwrite(kpi_name,coln1):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            print(kpi_name)
            index=df_pivot.index
            # print("index ;###############################",index_pivot)
            # print(len(index_pivot))
            print("doooooonnenn")
            dr=df_pivot[kpi_name]
            print("columns of dr tab")
            li=dr.columns
            print("column list")

            col1=dr[li[0]].to_list()
            col2=dr[li[1]].to_list()
            col3=dr[li[2]].to_list()
            col4=dr[li[3]].to_list()
            col5=dr[li[4]].to_list()

            trend_ws[coln1 + "2"].value = cl[4]
            trend_ws[coln2 + "2"].value = cl[3]
            trend_ws[coln3 + "2"].value = cl[2]
            trend_ws[coln4 + "2"].value = cl[1]
            trend_ws[coln5 + "2"].value = cl[0]

            for i,value in enumerate(index):
                j=i+3
          
            # ws['D'+str(j)].value='TDD'

                # trend_ws['C'+str(j)].value=index[i][2]
                trend_ws['c'+str(j)].value=date1
                trend_ws['B'+str(j)].value=index[i][1]
                trend_ws['D'+str(j)].value=index[i][0]
                trend_ws['E'+str(j)].value=index[i][4]
                trend_ws['F'+str(j)].value=index[i][3]
                trend_ws['H'+str(j)].value=index[i][2]
            

                trend_ws[coln1 + str(j)].value = col1[i]
                trend_ws[coln2 + str(j)].value = col2[i]
                trend_ws[coln3 + str(j)].value = col3[i]
                trend_ws[coln4 + str(j)].value = col4[i]
                trend_ws[coln5 + str(j)].value = col5[i]


        trend_ws=trend_wb['HPSC Cell Wise']
        for kpi_name in kpi:
            if(kpi_name =='MV_4G Data Volume_GB'):
                overwrite(kpi_name,'K')

            if(kpi_name=='MV_DL User Throughput_Mbps [CDBH]'):
                overwrite(kpi_name,'P') 

            if(kpi_name=='MV_UL User Throughput_Kbps [CDBH]'):
                overwrite(kpi_name,'U')

            if(kpi_name=='MV_Average number of used DL PRBs [CDBH]'):
                overwrite(kpi_name,'Z')

            if(kpi_name=='MV_E-UTRAN Average CQI [CDBH]'):
                overwrite(kpi_name,'AE') 
                
            if(kpi_name=='VoLTE DCR [CDBH]'):
                overwrite(kpi_name,'AJ') 



        ###########################################--------------------------------------###########################################3
        SaveOutput=os.path.join(door_root, 'output','rna.xlsx')        
        trend_wb.save(SaveOutput) 
            # print("_______________________________---save_____________________")
        download_path=os.path.join(MEDIA_URL,'trends','or','RNA','output','rna.xlsx')
        return Response({"status":True,"message":"Succesfully uploaded","Download_url":download_path})
        
        # return Response({"status":True,"message":"Succesfully uploaded"})
        
    except ValueError as ve:
        logging.error(f"ValueError occurred: {ve}")
        return Response({'status': False, 'message': 'Invalid value provided.'})
    except TypeError as te:
        logging.error(f"TypeError occurred: {te}")
        return Response({'status': False, 'message': 'Invalid type provided.'})
    except Exception as e:
        logging.error(f"Unexpected error occurred: {e}")
        return Response({'status': False, 'message': 'An unexpected error occurred.'})


