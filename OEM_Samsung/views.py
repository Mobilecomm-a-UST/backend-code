from django.shortcuts import render
from django.db.models import *
from rest_framework.decorators import api_view
from rest_framework.response import Response
import pandas as pd
import os
from django.conf import settings
from rest_framework import status
import numpy as np
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from rest_framework.status import HTTP_200_OK
from .models import SamsungAlarm
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font,Border,Side
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin
import xml.etree.ElementTree as ET
import re
import zipfile
import shutil
from .models import *


def compare_lock_status_5g(row):

    old_cell = str(
        row.get("Cell Name_old", "")
    ).strip()

    new_cell = str(
        row.get("Cell Name_new", "")
    ).strip()

    old_state = str(
        row.get(
                "MRBTS.NRBTS.NRCELL.administrativeState_old",
                ""
        )
    ).strip().lower()

    new_state = str(
        row.get(
                "MRBTS.NRBTS.NRCELL.administrativeState_new",
                ""
        )
    ).strip().lower()

    # split cell name
    old_parts = old_cell.split("_")
    new_parts = new_cell.split("_")

    # prefix match
    old_prefix = "_".join(old_parts[:4])
    new_prefix = "_".join(new_parts[:4])

    # suffix match (A_A / B_B)
    old_suffix = "_".join(old_parts[-2:])
    new_suffix = "_".join(new_parts[-2:])

    # pattern mismatch
    if (
        old_prefix != new_prefix
        or
        old_suffix != new_suffix
    ):
        return " "

    # both locked
    if (
        old_state == "locked"
        and new_state == "locked"
    ):
        return "Both Locked"

    # both unlocked
    elif (
        old_state == "unlocked"
        and new_state == "unlocked"
    ):
        return "Both Unlocked"

    return " "


def compare_lock_status(row):

    old_cell = str(
        row.get("MV Cell Name_old", "")
    ).strip()

    new_cell = str(
        row.get("MV Cell Name_new", "")
    ).strip()

    old_state = str(
        row.get("administrativeState_old", "")
    ).strip().lower()

    new_state = str(
        row.get("administrativeState_new", "")
    ).strip().lower()

    # ---------------- PREFIX ----------------
    old_parts = old_cell.split("_")
    new_parts = new_cell.split("_")

    # MH_E_F1_OM
    old_prefix = "_".join(old_parts[:4])
    new_prefix = "_".join(new_parts[:4])

    # ---------------- SUFFIX ----------------
    old_suffix_match = re.search(
        r'([A-Z]_[A-Z])$',
        old_cell
    )

    new_suffix_match = re.search(
        r'([A-Z]_[A-Z])$',
        new_cell
    )

    old_suffix = (
        old_suffix_match.group(1)
        if old_suffix_match
        else ""
    )

    new_suffix = (
        new_suffix_match.group(1)
        if new_suffix_match
        else ""
    )

    # print(
    #       "OLD:", old_cell,
    #       "| NEW:", new_cell
    # )
    # print(
    #       old_prefix,
    #       old_suffix,
    #       "----",
    #       new_prefix,
    #       new_suffix
    # )

    # ---------------- MATCH ----------------
    if (
        old_prefix != new_prefix
        or
        old_suffix != new_suffix
    ):
        return ""

    # ---------------- LOCK STATUS ----------------
    if (
        old_state == "locked"
        and new_state == "locked"
    ):
        return "Both Locked"

    elif (
        old_state == "unlocked"
        and new_state == "unlocked"
    ):
        return "Both Unlocked"

    return ""

def format_of_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # ---------- Styles ----------
    header_fill = PatternFill(start_color="E87529", end_color="E87529", fill_type="solid")
    alarm_header_fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")

    header_font = Font(color="000000", bold=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ---------- Header Formatting ----------
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.row_dimensions[1].height = 40  # visible header

    header_list = [cell.value for cell in ws[1]]

    lb_cols = [
        "Site Id","MS2 Status","Circle","Partner Name","Old/New",
        "OEM","On-air Date","On-air month","MS2 status",
        "ip","LNBTS ID","NRBTS ID","Alarm Status (Yes/No)",
    ]

    for col_name in lb_cols:
        if col_name in header_list:
                col_idx = header_list.index(col_name) + 1
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = alarm_header_fill
                cell.font = header_font

    # ---------- Data Borders ----------
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
                cell.border = thin_border

    # ---------- UNIFORM COLUMN WIDTH ----------
    DEFAULT_WIDTH = 25

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = DEFAULT_WIDTH

    wb.save(file_path)


@api_view(['GET'])
def get_sites(request):
    try:
        data = list(Old_New.objects.values())

        return Response({
                "status": True,
                "data": data
        })

    except Exception as e:
        return Response({
                "status": False,
                "error": str(e)
        }, status=500)


@api_view(['DELETE'])
def delete_sites(request):
    try:
        Old_New.objects.all().delete()

        return Response({
                "status": True,
                "message": "Deleted successfully"
        })
    except Exception as e:
        return Response({
                "status": False,
                "error": str(e)
        }, status=500)


@api_view(['POST'])
def upload_site_list(request):
    file = request.FILES.get('site_file')
    if not file:
        return Response({
                "status": False,
                "error": "File required"
        }, status=400)
    if not file.name.endswith('.xlsx'):
        return Response({
                "status": False,
                "error": "Only Excel file allowed"
        }, status=400)

    try:
        Old_New.objects.all().delete()
        df = pd.read_excel(file, sheet_name="Sheet 1")

        if df is not None:

                df.columns = df.columns.str.strip()

                for _, row in df.iterrows():

                    Old_New.objects.update_or_create(

                            new_site=str(row.get('New SiteId', '')).strip(),
                            old_site=str(row.get('Old SiteId', '')).strip(),
                            new_4g_mrbts=str(row.get('New LNBTS ID', '')).strip(),
                            old_4g_mrbts=str(row.get('OLD LNBTS ID', '')).strip(),
                            new_5g_nrbts=str(row.get('New NRBTS ID', '')).strip(),
                            old_5g_nrbts=str(row.get('OLD NRBTS ID', '')).strip(),

                    )

        return Response({
        "status": True,
        "message": "Data uploaded successfully"
        })

    except Exception as e:
        return Response({
                "status": False,
                "error": str(e)
        }, status=500)

@api_view(['GET', 'POST'])
def fileupload(request):

    # ---------------- GET ----------------
    if request.method == 'GET':
        alarms = SamsungAlarm.objects.all().values(
            'oem',
            'alarm_name',
            'alarm_type',
            'sa_nsa',
            'noc_circle',
        )
        return Response(list(alarms), status=status.HTTP_200_OK)

    # ---------------- POST ----------------
    elif request.method == 'POST':

        file = request.FILES.get('file')

        if not file:
            return Response(
                {'error': 'File is required!'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Required columns (normalized uppercase)
        required_cols = {
            "OEM",
            "ALARM NAME",
            "SA/NSA",
            "NOC/CIRCLE",
            "ALARM TYPE"
        }

        try:
            valid_dfs = []

            # ---------- CSV ----------
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)

                df = df.dropna(how="all")
                df.columns = df.columns.str.strip().str.upper()

                if required_cols.issubset(set(df.columns)):
                    valid_dfs.append(df)

            # ---------- EXCEL ----------
            else:
                excel_data = pd.read_excel(file, sheet_name=None)

                for sheet_name, df in excel_data.items():

                    df = df.dropna(how="all")

                    # Normalize columns
                    df.columns = df.columns.str.strip().str.upper()

                    if required_cols.issubset(set(df.columns)):
                        valid_dfs.append(df)

            # ❌ No valid sheet found
            if not valid_dfs:
                return Response(
                    {'error': 'No sheet contains required columns!'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Clear old data
            SamsungAlarm.objects.all().delete()

        except Exception as e:
            return Response(
                {'error': f'Error reading file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ---------- SAVE DATA ----------
        objs = []
        count = 0

        for df in valid_dfs:
            for _, row in df.iterrows():

                objs.append(
                    SamsungAlarm(
                        oem=str(row.get("OEM", "")).strip(),
                        alarm_name=str(row.get("ALARM NAME", "")).strip(),
                        sa_nsa=str(row.get("SA/NSA", "")).strip(),
                        noc_circle=str(row.get("NOC/CIRCLE", "")).strip(),
                        # remark=str(row.get("REMARK", "")).strip(),
                        alarm_type=str(row.get("ALARM TYPE", "")).strip()
                    )
                )
                count += 1

        # Bulk insert (fast)
        SamsungAlarm.objects.bulk_create(objs)

        return Response(
            {
                "status": True,
                "message": f"{count} alarms saved successfully!"
            },
            status=status.HTTP_201_CREATED
        )
@api_view(['POST'])
def alarmfileUpload(request):

    alarm_files = request.FILES.getlist('alarm_file')
    mapping_file = request.FILES.get('mapping_file')

    if not alarm_files or not mapping_file:
        return Response(
                {'error': 'Both alarm_file and mapping_file are required!'},
                status=status.HTTP_400_BAD_REQUEST
        )
    
    # ---------------- Clean Old Output ----------------
    output_root = os.path.join(
        MEDIA_ROOT,
        "SAMSUNG_OUTPUT"
    )

    final_output_folder = os.path.join(
        output_root,
        "OUTPUT"
    )

    zip_path = os.path.join(
        output_root,
        "OUTPUT.zip"
    )

    # delete old output folder
    if os.path.exists(final_output_folder):
        shutil.rmtree(final_output_folder)

    # delete old zip
    if os.path.exists(zip_path):
        os.remove(zip_path)

    # ---------------- Read Alarm Files ----------------
    alarm_dfs = []

    # for alarm_file in alarm_files:
    #     try:
    #             filename = alarm_file.name.lower()

    #             if filename.endswith('.csv'):
    #                 df = pd.read_csv(alarm_file)

    #             elif filename.endswith(('.xlsx', '.xls')):
    #                 df = pd.read_excel(alarm_file)

    #             else:
    #                 continue
    #             df.columns = df.columns.str.strip()
    #             alarm_dfs.append(df)

    #     except Exception as e:
    #             return Response(
    #             {'error': f'Error reading alarm file {alarm_file.name}: {str(e)}'},
    #             status=status.HTTP_400_BAD_REQUEST
    #             )
    for alarm_file in alarm_files:
        try:
            filename = alarm_file.name.lower()

            # ---------------- CSV ----------------
            if filename.endswith('.csv'):
                df = pd.read_csv(alarm_file)

            # ---------------- EXCEL ----------------
            elif filename.endswith(('.xlsx', '.xls')):

                excel_sheets = pd.read_excel(alarm_file, sheet_name=None)

                df_list = []

                for sheet_name, temp_df in excel_sheets.items():
                    print("SHEET NAME:", sheet_name)
                    temp_df.columns = temp_df.columns.astype(str).str.strip()
                    temp_df = temp_df.loc[:, ~temp_df.columns.str.contains("Unnamed")]

                    # skip empty sheet
                    if temp_df.empty:
                        continue

                    df_list.append(temp_df)

                df = pd.concat(df_list, ignore_index=True)

            else:
                continue

            df.columns = df.columns.str.strip()
            alarm_dfs.append(df)
            print("FILE NAME:", filename)
            print("SHAPE:", df.shape)
            print("COLUMNS:", df.columns.tolist())
        except Exception as e:
            return Response(
                {'error': f'Error reading alarm file {alarm_file.name}: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    if not alarm_dfs:
        return Response(
                {'error': 'No valid alarm files found'},
                status=status.HTTP_400_BAD_REQUEST
        )
    df_alarm = pd.concat(alarm_dfs, ignore_index=True)
    
    # ---------------- Separate 4G / 5G Status Files ----------------
    df_4g = pd.DataFrame()
    df_5g = pd.DataFrame()

    for df in alarm_dfs:

        cols = [str(c).strip() for c in df.columns]

        
        # ---------- 4G ----------
        if "NE ID" in cols and df["NE ID"].astype(str).str.startswith("eNB_").any():

            temp_4g = df.copy()

            temp_4g.columns = temp_4g.columns.astype(str).str.strip()

            required_cols = [
                "NE ID",
                "NE Name",
                "administrative-state",
                "user-label",
            ]

            temp_4g = temp_4g[
                [c for c in required_cols if c in temp_4g.columns]
            ]

            # -------- LNBTS ID --------
            temp_4g["LNBTS ID"] = (
                temp_4g["NE ID"]
                .astype(str)
                .str.extract(r'eNB_(\d+)')[0]
            )

            temp_4g = temp_4g[
                temp_4g["LNBTS ID"].notna()
            ]

            temp_4g = temp_4g[
                temp_4g["LNBTS ID"].str.match(r'^\d+$', na=False)
            ]

            # -------- SITE ID (PB_E_F3XXXXXX_LUDH45 → LUDH45) --------
            temp_4g["Site Id"] = (
                temp_4g["NE Name"]
                .astype(str)
                .str.strip()
                .str.rsplit("_", n=1)
                .str[-1]
            )

            # -------- FINAL 4G OUTPUT --------
            temp_4g = temp_4g[
                ["LNBTS ID", "Site Id", "administrative-state","user-label"]
            ].drop_duplicates()

            df_4g = pd.concat(
                [df_4g, temp_4g],
                ignore_index=True
            )


        # ---------- 5G ----------
        elif "NE ID" in cols and df["NE ID"].astype(str).str.contains(r'\d').any():

            temp_5g = df.copy()

            temp_5g.columns = temp_5g.columns.astype(str).str.strip()

            required_cols = [
                "NE ID",
                "NE Name",
                "administrative-state",
                "user-label"
            ]

            temp_5g = temp_5g[
                [c for c in required_cols if c in temp_5g.columns]
            ]

            # -------- NRBTS ID --------
            temp_5g["NRBTS ID"] = (
                temp_5g["NE ID"]
                .astype(str)
                .str.extract(r'(\d+)$')[0]
            )

            temp_5g = temp_5g[temp_5g["NRBTS ID"].str.len() > 6]
            
            temp_5g = temp_5g[
                temp_5g["NRBTS ID"].notna()
            ]

            temp_5g = temp_5g[
                temp_5g["NRBTS ID"].str.match(r'^\d+$', na=False)
            ]
            if "NE Name" not in temp_5g.columns:
                continue

            # -------- SITE ID --------
            temp_5g["Site Id"] = (
                temp_5g["NE Name"]
                .astype(str)
                .str.strip()
                .str.rsplit("_", n=1)
                .str[-1]
            )
            if "administrative-state" not in temp_5g.columns:
                continue
            
            # -------- FINAL 5G OUTPUT --------
            temp_5g = temp_5g[
                ["NRBTS ID", "Site Id", "administrative-state","user-label"]
            ].drop_duplicates()

            df_5g = pd.concat(
                [df_5g, temp_5g],
                ignore_index=True
            )
                        
        # print("-----5g data-----",df_5g)
        
    
    # ---------------- Read Mapping File ----------------
    try:
        mapping_filename = mapping_file.name.lower()

        if mapping_filename.endswith('.csv'):
            df_map = pd.read_csv(mapping_file)

        # elif mapping_filename.endswith(('.xlsx', '.xls')):
        #     df_map = pd.read_excel(mapping_file)
        elif mapping_filename.endswith(('.xlsx', '.xls')):

            excel_sheets = pd.read_excel(
                mapping_file,
                sheet_name=None
            )

            df_map = None

            for sheet_name, temp_df in excel_sheets.items():

                temp_df = temp_df.dropna(how="all")
                temp_df.columns = temp_df.columns.str.strip()

                if "SiteId" in temp_df.columns:
                    df_map = temp_df
                    print(f"Using mapping sheet: {sheet_name}")
                    break

            if df_map is None:
                return Response(
                    {
                        "error":
                        "No Mapping sheet found containing SiteId column"
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

        else:
            return Response(
                {'error': 'Unsupported mapping file format'},
                status=status.HTTP_400_BAD_REQUEST
            )

        df_map.columns = df_map.columns.str.strip()
        # print("mapping file data:",df_map)
    except Exception as e:
        return Response(
                {'error': f'Error reading mapping file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
        )
    # ---------------- Ensure MRBTS ----------------

    if "LNBTS ID" not in df_alarm.columns:
        if "NE ID" in df_alarm.columns:
                df_alarm["LNBTS ID"] = (
                df_alarm["NE ID"]
                .astype(str)
                .where(
                    df_alarm["NE ID"].astype(str).str.len() <= 6,
                    ""
                )
                )
        else:
                df_alarm["LNBTS ID"] = ""

    # ---------------- Ensure NRBTS ----------------

    if "NRBTS ID" not in df_alarm.columns:
        if "NE ID" in df_alarm.columns:
                df_alarm["NRBTS ID"] = (
                df_alarm["NE ID"]
                .astype(str)
                .where(
                    df_alarm["NE ID"].astype(str).str.len() > 6,
                    ""
                )
                )
        else:
                df_alarm["NRBTS ID"] = ""
    
    # needed_cols = [
    #       "MRBTS",
    #       "Supplementary Information",
    #       # "Origin Alarm Time",
    #       # "Origin Alarm Update Time",
    # ]
    
    # # ---------------- Extract 5G Alarms ----------------
    # if (
    #     "Distinguished Name" in df_alarm.columns
    #     and "Supplementary Information" in df_alarm.columns
    # ):

    #     df_alarm["5G Alarms"] = np.where(
    #             df_alarm["Distinguished Name"]
    #             .astype(str)
    #             .str.contains("NRBTS-", na=False),

    #             df_alarm["Supplementary Information"]
    #             .astype(str),

    #             ""
    #     )

    # else:
    #     df_alarm["5G Alarms"] = ""


    needed_cols = [
        "LNBTS ID",
        "NRBTS ID",
        "Specific Problem",
    ]
    df_alarm = df_alarm[[c for c in needed_cols if c in df_alarm.columns]]

    # ---------------- Load SA / NSA from DB ----------------
    # sa = list(SamsungAlarm.objects.exclude(SA__isnull=True).values_list('SA', flat=True))
    # nsa = list(SamsungAlarm.objects.exclude(NSA__isnull=True).values_list('NSA', flat=True))
    
    # ---------------- Load SA / NSA from DB ----------------
    alarm_master = SamsungAlarm.objects.all()
    
    alarm_lookup = {
        str(obj.alarm_name).strip().lower(): {
            "sa_nsa": obj.sa_nsa,
            "alarm_name": obj.alarm_name,
            "alarm_type": obj.alarm_type,
            "noc_circle": obj.noc_circle
        }
        for obj in alarm_master
        if obj.alarm_name
    }
    
    # Keep old variables to avoid breaking existing code
    sa = [
    k for k, v in alarm_lookup.items()
    if str(v.get("sa_nsa")).strip().upper() == "SA"
    ]

    nsa = [
    k for k, v in alarm_lookup.items()
    if str(v.get("sa_nsa")).strip().upper() == "NSA"
    ]
    # combined_alarms = set(sa + nsa)

    # ---------------- Timeout Regex ----------------
    # timeout_regex = re.compile(r"Timeout\s+connecting\s+to|No\s+route\s+to\s+host|CommunicationTimeout|CommunicationTimeout", re.IGNORECASE)

    # ---------------- Classify Alarms ----------------
    def classify_alarms(info):

        if pd.isna(info):
            return pd.Series([
                "No",
                "No Alarms",
                "No Alarms",
                "",
                ""
            ])

        alarm = str(info).strip().lower()

        matched_sa_nsa = []
        matched_alarm_names = []
        matched_alarm_types = []
        matched_noc_circles = []

        if alarm in alarm_lookup:

            details = alarm_lookup[alarm]

            if str(details["sa_nsa"]).upper() == "SA":
                matched_sa_nsa.append(
                    "SA"
                )

            elif str(details["sa_nsa"]).upper() == "NSA":
                matched_sa_nsa.append(
                    "NSA"
                )

            matched_alarm_names.append(
                details["alarm_name"]
            )

            matched_alarm_types.append(
                details["alarm_type"]
            )
            
            matched_noc_circles.append(
                details["noc_circle"]
            )


        return pd.Series([
            "Yes" if matched_alarm_names else "No",

            ", ".join(set(matched_sa_nsa))
            if matched_sa_nsa
            else "No Alarms",

            ", ".join(set(matched_alarm_names))
            if matched_alarm_names
            else "No Alarms",

            ", ".join(set(matched_alarm_types))
            if matched_alarm_types
            else "",
            ", ".join(set(matched_noc_circles))
            if matched_noc_circles
            else "",
            str(info) if pd.notna(info) else "No Alarm"
        ])
    df_alarm[
    [
        "Alarm Status (Yes/No)",
        "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms",
        "Service Affecting Alarms 1",
        "Alarm Type",
        "NOC Circle",
        "Raw Alarm"
    ]
    ] = (
    df_alarm["Specific Problem"]
    .apply(classify_alarms)
    )
    # ---------------- NRBTS Alarm Lookup ----------------

    nrbts_alarm_lookup = (
        df_alarm[
            df_alarm["NRBTS ID"].astype(str).str.strip().ne("")
        ]
        .groupby("NRBTS ID")["Specific Problem"]
        .agg(lambda x: ", ".join(sorted(set(
            str(i).strip()
            for i in x
            if pd.notna(i) and str(i).strip()
        ))))
        .to_dict()
    )
    
# ---------------- Group & Deduplicate ----------------
    # df_alarm = (
    #       df_alarm.drop_duplicates()
    #       .groupby("MRBTS", as_index=False)
    #       .agg(lambda x: ', '.join(
    #             sorted(set(str(i) for i in x if pd.notna(i) and str(i).strip()))
    #       ))
    # )
    
    # ---------------- Clean + Group ----------------

    def clean_and_merge(values, column_name=""):

        # remove nan / empty
        vals = [
                str(v).strip()
                for v in values
                if pd.notna(v)
                and str(v).strip()
                and str(v).strip().lower() != "nan"
        ]

        vals_set = set(vals)

        # ---------------- Alarm Status ----------------
        if column_name == "Alarm Status (Yes/No)":

                # if both Yes and No exist -> keep Yes
                if "Yes" in vals_set:
                    return "Yes"

                return "No"

        # ---------------- SA / NSA ----------------
        if column_name == "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms":

                vals_upper = {
                    str(v).strip().upper()
                    for v in vals_set
                }

                # Priority:
                # SITE DOWN > SA > NSA > No Alarms
                if "SITE DOWN" in vals_upper:
                    return "SITE DOWN"

                if "SA" in vals_upper:
                    return "SA"

                if "NSA" in vals_upper:
                    return "NSA"

                return "No Alarms"

        # ---------------- Other Columns ----------------
        cleaned = sorted(set(vals))

        return ", ".join(cleaned)

    nrbts_details_lookup = (
        df_alarm[
            df_alarm["NRBTS ID"].astype(str).str.strip().ne("")
        ]
        .groupby("NRBTS ID")
        .agg({
            "Alarm Status (Yes/No)":
                lambda x: clean_and_merge(x, "Alarm Status (Yes/No)"),

            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms":
                lambda x: clean_and_merge(
                    x,
                    "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms"
                ),

            "Alarm Type":
                lambda x: ", ".join(sorted(set(
                    str(i).strip()
                    for i in x
                    if pd.notna(i) and str(i).strip()
                ))),

            "NOC Circle":
                lambda x: ", ".join(sorted(set(
                    str(i).strip()
                    for i in x
                    if pd.notna(i) and str(i).strip()
                )))
        })
        .to_dict("index")
    )
    df_alarm = (
        df_alarm.drop_duplicates()
        .groupby(["LNBTS ID", "NRBTS ID"], as_index=False)
        .agg(lambda x: clean_and_merge(x, x.name))
    )
    df_alarm["Service Affecting Alarms 1"] = (
        df_alarm["Service Affecting Alarms 1"]
        .astype(str)
        .apply(
            lambda x: ", ".join(
                sorted(
                    {
                        i.strip()
                        for i in x.split(",")
                        if i.strip() and i.strip() != "No Alarms"
                    }
                )
            ) or "No Alarms"
        )
    )
    # ---------------- Normalize LNBTS ID ----------------
    df_alarm["LNBTS ID"] = df_alarm["LNBTS ID"].astype(str).str.replace('.0', '', regex=False)
    df_alarm["NRBTS ID"] = df_alarm["NRBTS ID"].astype(str).str.replace('.0', '', regex=False)
    # df_map["MRBTS"] = df_map["MRBTS"].astype(str).str.replace('.0', '', regex=False)
    df_map["LNBTS ID"] = (
        df_map["LNBTS ID"]
        .astype(str)
        .str.replace(r'\.0', '', regex=True)
        .str.strip()
        )

    # df_map.rename(
    #     columns={"Service Affecting Alarms 1": "Service Affecting Alarms"},
    #     inplace=True
    # )
    for df in [df_map, df_alarm]:
        df["LNBTS ID"] = (
            df["LNBTS ID"]
            .astype(str)
            .str.replace(r"\.0", "", regex=True)
            .str.strip()
        )

        df["NRBTS ID"] = (
            df["NRBTS ID"]
            .astype(str)
            .str.replace(r"\.0", "", regex=True)
            .str.strip()
        )

    # ---------------- Merge ----------------
    df_merged = pd.merge(
        df_map,
        df_alarm,
        on="LNBTS ID",
        how="left",
        suffixes=("", "_alarm")
    )
    cols_to_update = [
        "Alarm Status (Yes/No)",
        "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms",
        "Service Affecting Alarms 1",
        "Alarm Type",
        "NOC Circle"
    ]

    for col in cols_to_update:

        alarm_col = f"{col}_alarm"

        if alarm_col in df_merged.columns:

            df_merged[col] = df_merged[col].mask(
                df_merged[col].isna() |
                (df_merged[col].astype(str).str.strip() == ""),
                df_merged[alarm_col]
            )

            df_merged.drop(columns=[alarm_col], inplace=True)
            


    # ---------------- NRBTS Alarm Column ----------------

    df_merged["NRBTS ID_alarm"] = (
        df_merged["NRBTS ID"]
        .astype(str)
        .str.replace(r"\.0", "", regex=True)
        .str.strip()
        .map(nrbts_alarm_lookup)
        .fillna("")
    )
    df_merged["5G Alarm Status (Yes/No)"] = (
        df_merged["NRBTS ID"]
        .astype(str)
        .str.strip()
        .map(lambda x: nrbts_details_lookup.get(x, {}).get(
            "Alarm Status (Yes/No)", ""
        ))
    )

    df_merged["5G No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms"] = (
        df_merged["NRBTS ID"]
        .astype(str)
        .str.strip()
        .map(lambda x: nrbts_details_lookup.get(x, {}).get(
            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms", ""
        ))
    )

    df_merged["5G Alarm Type"] = (
        df_merged["NRBTS ID"]
        .astype(str)
        .str.strip()
        .map(lambda x: nrbts_details_lookup.get(x, {}).get(
            "Alarm Type", ""
        ))
    )

    df_merged["5G NOC Circle"] = (
        df_merged["NRBTS ID"]
        .astype(str)
        .str.strip()
        .map(lambda x: nrbts_details_lookup.get(x, {}).get(
            "NOC Circle", ""
        ))
    )
    # ---------------- Fill Blank SA/NSA ----------------
    cols_fill = [
        "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms",
        "5G No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms",
        "Service Affecting Alarms 1"
        
    ]

    df_merged[cols_fill] = df_merged[cols_fill].replace(
        ["", "nan", None],
        "No Alarms"
    )

    # ---------------- 4G / 5G Alarm Status ----------------
    old_new_df = pd.DataFrame(
        list(
                Old_New.objects.values(
                    "new_4g_mrbts",
                    "old_4g_mrbts",
                    "new_5g_nrbts",
                    "old_5g_nrbts"
                )
        )
    )

    # clean values
    if not old_new_df.empty:
        for col in old_new_df.columns:
            old_new_df[col] = (
                old_new_df[col]
                .astype(str)
                .str.replace(r"\.0", "", regex=True)
                .str.strip()
            )
    # create lookup set
    new_4g_set = set(old_new_df["new_4g_mrbts"])
    old_4g_set = set(old_new_df["old_4g_mrbts"])

    new_5g_set = set(old_new_df["new_5g_nrbts"])
    old_5g_set = set(old_new_df["old_5g_nrbts"])

    # clean merged cols
    df_merged["LNBTS ID"] = (
        df_merged["LNBTS ID"]
        .astype(str)
        .str.replace(r"\.0", "", regex=True)
        .str.strip()
    )

    df_merged["NRBTS ID"] = (
        df_merged["NRBTS ID"]
        .astype(str)
        .str.replace(r"\.0", "", regex=True)
        .str.strip()
    )
    # print("=======df_merged=======",df_merged["NRBTS"])


    # 4G Status
    def get_4g_status(LNBTS_ID):

        LNBTS_ID = str(LNBTS_ID).strip()

        if LNBTS_ID in new_4g_set:
                return "New Site"

        elif LNBTS_ID in old_4g_set:
                return "OLD Site"

        return ""


    # 5G Status
    def get_5g_status(NRBTS_ID):

        # null check
        if pd.isna(NRBTS_ID):
                return ""

        # clean nrbts
        NRBTS_ID = (
                str(NRBTS_ID)
                .replace(".0", "")
                .strip()
        )

        # ignore nan string
        if not NRBTS_ID or NRBTS_ID.lower() == "nan":
                return ""

        # print("Checking NRBTS:", nrbts)

        # match new
        if NRBTS_ID in new_5g_set:
                return "New Site"

        # match old
        elif NRBTS_ID in old_5g_set:
                return "OLD Site"

        return ""


    # create columns
    df_merged["4G Alarm Status"] = (
        df_merged["LNBTS ID"]
        .apply(get_4g_status)
    )

    df_merged["5G Alarm Status"] = (
        df_merged["NRBTS ID"]
        .apply(get_5g_status)
    )
    # ---------------- Save Circle Wise Output ----------------
    # output_root = os.path.join(MEDIA_ROOT, "SAMSUNG_OUTPUT")
    # os.makedirs(output_root, exist_ok=True)
    # saved_files = []
    # ---------------- Save Circle Wise Output ----------------
    output_root = os.path.join(
        MEDIA_ROOT,
        "SAMSUNG_OUTPUT"
    )

    os.makedirs(output_root, exist_ok=True)

    # OUTPUT folder
    final_output_folder = os.path.join(
        output_root,
        "OUTPUT"
    )

    # remove old OUTPUT folder
    if os.path.exists(final_output_folder):
        shutil.rmtree(final_output_folder)

    os.makedirs(final_output_folder, exist_ok=True)

    saved_files = []
    # Ensure Circle column exists
    if "Circle" not in df_merged.columns:
        return Response(
                {"error": "Circle column not found in mapping file"},
                status=status.HTTP_400_BAD_REQUEST
        )

    # Clean circle values
    df_merged["Circle"] = (
        df_merged["Circle"]
        .astype(str)
        .str.strip()
    )
    # keep only rows having actual alarms
    # df_merged = df_merged[
    #       df_merged["Total Alarms"].notna()
    # ]
    # ---------------- Split Circle Wise ----------------
    for circle, circle_df in df_merged.groupby("Circle"):
        # skip fully blank circles
        # if circle_df["Total Alarms"].isna().all():
        #         continue
        # if not circle or circle.lower() == "nan":
        #         continue
        
        # Circle folder inside OUTPUT
        circle_folder = os.path.join(
                final_output_folder,
                circle
        )

        os.makedirs(circle_folder, exist_ok=True)

        # File name
        output_filename = (
                f"Samsung_SA_NSA_OUTPUT_{circle}.xlsx"
        )

        full_output_path = os.path.join(
                circle_folder,
                output_filename
        )

        # Save excel
        # circle_df.to_excel(
        #       full_output_path,
        #       index=False
        # )
        # ---------------- FORCE REQUIRED FORMAT ----------------
        required_columns = [
            "Circle",
            "SiteId",
            "Countif",
            "Partner Name",
            "OEM",
            "On-air Date",
            "MS2 status",
            "Alarm Status (Yes/No)",
            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms",
            "Service Affecting Alarms 1"
        ]

        # ensure columns exist
        for col in required_columns:
            if col not in circle_df.columns:
                circle_df[col] = ""

        # reorder columns (keep extra also)
        circle_df = circle_df[required_columns + [c for c in circle_df.columns if c not in required_columns]]
        
        with pd.ExcelWriter(full_output_path,engine="openpyxl") as writer:
                # Main Alarm Sheet
                circle_df.to_excel(writer,sheet_name="Alarm_Output",index=False)

                # ---------------- 4G Status ----------------
                if not df_4g.empty:

                    circle_4g = df_4g[
                        df_4g["LNBTS ID"]
                        .astype(str)
                        .isin(
                            circle_df["LNBTS ID"]
                            .astype(str)
                        )
                    ].copy()

                    if not circle_4g.empty:
                        # ---------------- 4G Alarm Status ----------------
                        circle_4g["LNBTS ID"] = (
                            circle_4g["LNBTS ID"]
                            .astype(str)
                            .str.replace(r"\.0", "", regex=True)
                            .str.strip()
                        )

                        circle_4g["4G Alarm Status"] = (
                            circle_4g["LNBTS ID"]
                            .apply(get_4g_status)
                        )

                        circle_4g.to_excel(
                            writer,
                            sheet_name="4G_Status",
                            index=False
                        )

                # ---------------- 5G Status ----------------
                if not df_5g.empty:

                    circle_5g = df_5g[
                            df_5g["NRBTS ID"]
                            .astype(str)
                            .isin(
                                circle_df["NRBTS ID"]
                                .astype(str)
                            )
                    ].copy()

                    if not circle_5g.empty:

                        # ---------------- 5G Alarm Status ----------------
                        circle_5g["NRBTS ID"] = (
                            circle_5g["NRBTS ID"]
                            .astype(str)
                            .str.replace(r"\.0", "", regex=True)
                            .str.strip()
                        )

                        circle_5g["5G Alarm Status"] = (
                            circle_5g["NRBTS ID"]
                            .apply(get_5g_status)
                        )
                        circle_5g.to_excel(
                            writer,
                            sheet_name="5G_Status",
                            index=False
                        )
                        
                # ---------------- 4G OLD VS NEW SHEET ----------------
                if not df_4g.empty:

                    # clean MRBTS
                    temp_4g_compare = df_4g.copy()

                    temp_4g_compare["LNBTS ID"] = (
                            temp_4g_compare["LNBTS ID"]
                            .astype(str)
                            .str.replace(r"\.0", "", regex=True)
                            .str.strip()
                    )

                    compare_4g_rows = []

                    for _, row in old_new_df.iterrows():
                            

                            old_mrbts = str(
                                row.get("old_4g_mrbts", "")
                            ).replace(".0", "").strip()

                            new_mrbts = str(
                                row.get("new_4g_mrbts", "")
                            ).replace(".0", "").strip()
                            # ---------------- Current Circle Filter ----------------
                            circle_mrbts = set(
                                circle_df["LNBTS ID"]
                                .astype(str)
                                .str.replace(".0", "", regex=False)
                                .str.strip()
                            )

                            # skip other circle data
                            if (
                                old_mrbts not in circle_mrbts
                                and
                                new_mrbts not in circle_mrbts
                            ):
                                continue
                            
                            # find old row
                            old_row = temp_4g_compare[
                                temp_4g_compare["LNBTS ID"] == old_mrbts
                            ]

                            # find new row
                            new_row = temp_4g_compare[
                                temp_4g_compare["LNBTS ID"] == new_mrbts
                            ]

                            if old_row.empty and new_row.empty:
                                continue

                            old_dict = {}
                            new_dict = {}

                            # old data
                            if not old_row.empty:
                                old_dict = {
                                        f"{col}_old": val
                                        for col, val in old_row.iloc[0].to_dict().items()
                                }

                            # new data
                            if not new_row.empty:
                                new_dict = {
                                        f"{col}_new": val
                                        for col, val in new_row.iloc[0].to_dict().items()
                                }

                            compare_4g_rows.append({
                                **old_dict,
                                **new_dict
                            })

                    if compare_4g_rows:

                            compare_4g_df = pd.DataFrame(
                                compare_4g_rows
                            )
                            compare_4g_df["Lock Status"] = (
                                compare_4g_df.apply(
                                        compare_lock_status,
                                        axis=1
                                )
                            )
                            compare_4g_df.to_excel(
                                writer,
                                sheet_name="4G Old vs New",
                                index=False
                            )
                            print("old vs new :",compare_4g_df)


                # ---------------- 5G OLD VS NEW SHEET ----------------
                if not df_5g.empty:

                    temp_5g_compare = df_5g.copy()
                    temp_5g_compare["NRBTS ID"] = (
                            temp_5g_compare["NRBTS ID"]
                            .astype(str)
                            .str.replace(r"\.0", "", regex=True)
                            .str.strip()
                    )

                    compare_5g_rows = []

                    for _, row in old_new_df.iterrows():
                            

                            old_nrbts = str(
                                row.get("old_5g_nrbts", "")
                            ).replace(".0", "").strip()

                            new_nrbts = str(
                                row.get("new_5g_nrbts", "")
                            ).replace(".0", "").strip()

                            # ---------------- Current Circle Filter ----------------
                            circle_nrbts = set(
                                circle_df["NRBTS ID"]
                                .astype(str)
                                .str.replace(".0", "", regex=False)
                                .str.strip()
                            )

                            # skip other circle data
                            if (
                                old_nrbts not in circle_nrbts
                                and
                                new_nrbts not in circle_nrbts
                            ):
                                continue
                            # find old row
                            old_row = temp_5g_compare[
                                temp_5g_compare["NRBTS ID"] == old_nrbts
                            ]

                            # find new row
                            new_row = temp_5g_compare[
                                temp_5g_compare["NRBTS ID"] == new_nrbts
                            ]

                            if old_row.empty and new_row.empty:
                                continue

                            old_dict = {}
                            new_dict = {}

                            # old data
                            if not old_row.empty:
                                old_dict = {
                                        f"{col}_old": val
                                        for col, val in old_row.iloc[0].to_dict().items()
                                }

                            # new data
                            if not new_row.empty:
                                new_dict = {
                                        f"{col}_new": val
                                        for col, val in new_row.iloc[0].to_dict().items()
                                }

                            compare_5g_rows.append({
                                **old_dict,
                                **new_dict
                            })

                    if compare_5g_rows:

                            compare_5g_df = pd.DataFrame(
                                compare_5g_rows
                            )
                            compare_5g_df["Lock Status"] = (
                                compare_5g_df.apply(
                                        compare_lock_status_5g,
                                        axis=1
                                )
                            )
                            compare_5g_df.to_excel(
                                writer,
                                sheet_name="5G Old vs New",
                                index=False
                            )
        # Format excel
        # format_of_excel(full_output_path)
        try:
                format_of_excel(full_output_path)
        except Exception as e:
                print(f"Formatting failed: {e}")

        saved_files.append({
                "circle": circle,
                "path": full_output_path
        })

    print("Saved Circle Wise Files:", saved_files)
    # # ---------------- Save Output ----------------
    # output_path = os.path.join(MEDIA_ROOT, "SAMSUNG_OUTPUT")
    # os.makedirs(output_path, exist_ok=True)
    # output_filename = "Nokia_SA_NSA_OUTPUT.xlsx"
    # full_output_path = os.path.join(output_path, output_filename)
    # df_merged.to_excel(full_output_path, index=False)
    # format_of_excel(full_output_path)

    # # Build absolute download link robustly
    # relative_url = f"{output_filename}"
    # download_link = request.build_absolute_uri(urljoin(MEDIA_URL, relative_url))

    # ---------------- Create ZIP of OUTPUT Folder ----------------
    zip_filename = "OUTPUT.zip"

    zip_path = os.path.join(
        output_root,
        zip_filename
    )

    # delete old zip
    if os.path.exists(zip_path):
        os.remove(zip_path)

    # create zip from OUTPUT folder
    shutil.make_archive(
        base_name=os.path.splitext(zip_path)[0],
        format="zip",
        root_dir=final_output_folder
    )

    print("ZIP Created:", zip_path)

    # ---------------- Download Link ----------------
    relative_url = (
        f"SAMSUNG_OUTPUT/{zip_filename}"
    )

    download_link = request.build_absolute_uri(
        urljoin(
                MEDIA_URL,
                relative_url
        )
    )

    # ---------------- Response ----------------
    return Response(
        {
                "status": True,
                "message": "alarm files processed successfully",
                "files_saved": len(saved_files),
                "circles": [
                    x["circle"]
                    for x in saved_files
                ],
                "download_url": download_link
        },
        status=status.HTTP_200_OK
    )
    
    
    
