from django.http import JsonResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from nt_site_tracker.models import *  # noqa: F403
from alok_tracker.models import RelocationUser  # noqa: F403
import pandas as pd
from django.db import transaction
import os
from django.conf import settings
from datetime import datetime as dtime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.utils import timezone
import json
from datetime import datetime, timedelta, date
import shutil
from django.db.models import Q
from django.forms.models import model_to_dict
import numpy as np
from rest_framework import status
from openpyxl.utils import get_column_letter
from .utils import format_excel

ALL_COLUMNS = [
    "Circle",
    "Site ID",
    "Project (NT/IBS/ULS)",
    "TUR Breakup (Urban/Rural/Metro)",
    "TOCO Name",
    "SR Number",
    "RAN OEM",
    "Media Type",
    "MW OEM",
    "MW Installation Partner",
    "Site Band",
    "RFAI Date",
    "Allocation Date",
    "RFAI Survey Date",
    "WRFAI (YES/NO)",
    "WRFAI Date",
    "MO Punch Date",
    "Material Dispatch Date",
    "Material Delivered Date",
    "Material Type - HW(Fresh/SRN)",
    "Material Type - IM(Fresh/SRN)",
    "Installation Start Date",
    "Installation End Date",
    "Integration Date",
    "Sacfa Appied date",
    "WPC No",
    "WPC Date",
    "EMF Submission Date",
    "NEP ID",
    "RAN LKF Status",
    "Alarm Status",
    "Alarm Rectification Done Date",
    "SCFT Done Date",
    "SCFT Offered Date",
    "RAN PAT Offer Date",
    "RAN SAT Offer Date",
    "MW Plan ID",
    "MW PAT Offer Date",
    "RSL Value Status",
    "ENM Status",
    "MW LKF",
    "MW SAT Offer Date",
    "MW MS1 MIDS Date",
    "MW Sacfa - A End",
    "MW WPC - A End",
    "MW Sacfa - B End",
    "MW WPC - B End",
    "Site ONAIR Date",
    "I-Deploy ONAIR Date",
    "Current Status",
    "Ideploy Status",
    "Detailed Remarks",
    "RFAI Rejected Date",
    "Ideploy PRI Taging",
    "Re RFAI Date",
    "PRI Count",
    "PRI Issue Ageing",
    "Other UST Issue Ageing",
    "Other Airtel Issue Ageing",
    "Total Issue Ageing",
    "Clear RFAI to MS1 Ageing",
    "RFAI to MS1 Ageing",
    "RAN PAT Accepted Date",
    "RAN SAT Accepted Date",
    "MW PAT Accepted Date",
    "MW SAT Accepted Date",
    "SCFT Accepted Date",
    "KPI AT offer Date",
    "KPI AT Accepted Date",
    "4G MS2 Date",
    "SSID",
    "PMIS Month",
    "Airtel Sign off( Yes/No)",
    "Last Updated Date",
    "Last Updated By"
]
# chnage model and coulm list

COLUMN_MAP = {
    "Circle": "circle",
    "Site ID": "site_id",

    "Project (NT/IBS/ULS)": "project_nt",
    "TUR Breakup (Urban/Rural/Metro)": "tur_breakup",

    "TOCO Name": "new_toco_name",
    "SR Number": "sr_number",

    "RAN OEM": "ran_oem",
    "Media Type": "media_type",
    "MW OEM": "mw_oem",
    "MW Installation Partner": "mw_nstallation_partner",

    "Site Band": "site_band",
    "RFAI Date": "rfai_date",
    "Allocation Date": "allocation_date",
    "RFAI Survey Date": "rfai_survey_date",

    "WRFAI (YES/NO)": "wrfai",
    "WRFAI Date":"wrfai_date",

    "MO Punch Date": "mo_punch_date",
    "Material Dispatch Date": "material_dispatch_date",
    "Material Delivered Date": "material_delivered_date",

    "Material Type - HW(Fresh/SRN)": "material_type_hw",
    "Material Type - IM(Fresh/SRN)": "material_type_im",

    "Installation Start Date": "installation_start_date",
    "Installation End Date": "installation_end_date",
    "Integration Date": "integration_date",

    "Sacfa Appied date": "Sacfa_Appied_date",
    "WPC No": "wpc_no",
    "WPC Date": "wpc_date",

    "EMF Submission Date": "emf_submission_date",
    "NEP ID": "nep_id",

    "RAN LKF Status": "ran_lkf_status",
    "Alarm Status": "alarm_status",
    "Alarm Rectification Done Date": "alarm_rectification_done_date",

    "SCFT Done Date": "scft_done_date",
    "SCFT Offered Date": "scft_offered_date",

    "RAN PAT Offer Date": "ran_pat_offer_date",
    "RAN SAT Offer Date": "ran_sat_offer_date",

    "MW Plan ID": "mw_plan_id",
    "MW PAT Offer Date": "mw_pat_offer_date",

    "RSL Value Status": "rsl_value_status",
    "ENM Status": "enm_status",
    "MW LKF": "mw_lkf",

    "MW SAT Offer Date": "mw_sat_offer_date",
    "MW MS1 MIDS Date": "mw_ms1_mids_date",

    "MW Sacfa - A End": "mw_sacfa_a_end",
    "MW WPC - A End": "mw_wpc_a_end",
    "MW Sacfa - B End": "mw_sacfa_b_end",
    "MW WPC - B End": "mw_wpc_b_end",

    "Site ONAIR Date": "site_onair_date",
    "I-Deploy ONAIR Date": "i_deploy_onair_date",

    "Current Status": "current_status",
    "Ideploy Status": "ideploy_status",
    "Detailed Remarks": "detailed_remarks",

    "RFAI Rejected Date": "rfai_rejected_date",
    "Ideploy PRI Taging": "ideploy_pri_taging",
    "Re RFAI Date": "re_rfai_date",

    "PRI Count": "pri_count",
    "PRI Issue Ageing": "pri_issue_ageing",
    "Other UST Issue Ageing": "other_ust_issue_ageing",
    "Other Airtel Issue Ageing": "other_airtel_issue_ageing",
    "Total Issue Ageing": "total_issue_ageing",

    "Clear RFAI to MS1 Ageing": "clear_rfai_ms1_ageing",
    "RFAI to MS1 Ageing": "rfai_to_ms1_ageing",

    "RAN PAT Accepted Date": "ran_pat_accepted_date",
    "RAN SAT Accepted Date": "ran_sat_accepted_date",
    "MW PAT Accepted Date": "mw_pat_accepted_date",
    "MW SAT Accepted Date": "mw_sat_accepted_date",
    "SCFT Accepted Date": "scft_accepted_date",

    "KPI AT offer Date": "kpi_at_offer_date",
    "KPI AT Accepted Date": "kpi_at_accepted_date",

    "4G MS2 Date": "four_g_ms2_date",
    "SSID": "ssid",
    "PMIS Month": "pmis_month",

    "Airtel Sign off( Yes/No)": "airtel_sign_off",

    "Last Updated Date": "last_updated_date",
    "Last Updated By": "last_updated_by",
}

# Create your views here.


def compute_union_ageing(rows):

    if rows.empty:
        return 0

    today = datetime.today()

    def to_datetime(x):

        if isinstance(x, datetime):
            return x

        if isinstance(x, date):
            return datetime.combine(x, datetime.min.time())

        if isinstance(x, str) and x not in ["", "-", "nan", None]:
            return datetime.strptime(x, "%d-%b-%y")

        return None

    ranges = []

    for _, row in rows.iterrows():

        start = to_datetime(row["Start Date"])
        close = row["Close Date"]
        end = today if close in ["", "-", "nan", None, "undefined"] else close

        end = to_datetime(end)

        if start and end:
            ranges.append((start, end))

    if not ranges:
        return 0

    ranges.sort(key=lambda x: x[0])

    merged = []

    current_start, current_end = ranges[0]

    for start, end in ranges[1:]:

        if start <= current_end:
            current_end = max(current_end, end)

        else:
            merged.append((current_start, current_end))
            current_start, current_end = start, end

    merged.append((current_start, current_end))

    return sum((end - start).days for start, end in merged)


def update_ageing_new(circle, site_id):

    data_obj = NTSiteTracker.objects.filter(
        circle=circle,
        site_id=site_id
    ).first()

    if not data_obj:
        return

    issue_df = pd.DataFrame(
        NTIssue.objects.filter(
            circle=circle,
            site_id=site_id
        ).values()
    )

    if issue_df.empty:

        pri_ageing = 0
        other_ust_ageing = 0
        other_airtel_ageing = 0
        total_issue_ageing = 0
        pri_count = 0

    else:

        rename_map = {
            "start_date": "Start Date",
            "close_date": "Close Date",
            "issue_name": "Issue Name",
            "issue_owner": "Issue Owner"
        }

        issue_df = issue_df.rename(columns=rename_map)

        pri_rows = issue_df[
            issue_df["Issue Name"] == "PRI"
        ]

        other_ust_rows = issue_df[
            (issue_df["Issue Name"] != "PRI") &
            (issue_df["Issue Owner"] == "Mobilecomm")
        ]

        other_airtel_rows = issue_df[
            (issue_df["Issue Name"] != "PRI") &
            (issue_df["Issue Owner"] == "Airtel")
        ]

        pri_ageing = compute_union_ageing(pri_rows)

        other_ust_ageing = compute_union_ageing(
            other_ust_rows
        )

        other_airtel_ageing = compute_union_ageing(
            other_airtel_rows
        )

        total_issue_ageing = compute_union_ageing(
            issue_df
        )

        pri_count = len(pri_rows)

    data_obj.pri_issue_ageing = pri_ageing
    data_obj.other_ust_issue_ageing = other_ust_ageing
    data_obj.other_airtel_issue_ageing = other_airtel_ageing
    data_obj.total_issue_ageing = total_issue_ageing
    data_obj.pri_count = pri_count

    site_onair = pd.to_datetime(
        data_obj.site_onair_date,
        errors="coerce"
    )

    re_rfai = pd.to_datetime(
        data_obj.re_rfai_date,
        errors="coerce"
    )

    rfai = pd.to_datetime(
        data_obj.rfai_date,
        errors="coerce"
    )

    rfai_final = re_rfai if pd.notna(re_rfai) else rfai

    if pd.isna(rfai_final):

        data_obj.rfai_to_ms1_ageing = "-"

    else:

        end_date = (
            datetime.today()
            if pd.isna(site_onair)
            else site_onair
        )

        data_obj.rfai_to_ms1_ageing = (
            (end_date - rfai_final).days
            - total_issue_ageing
        )

    data_obj.save()

    return


############################################################  UPLOAD DATA #####################################################################################

@api_view(["POST"])
def upload_tracker_data_view(request):

    user_id = request.data.get("userId")

    if not user_id:
        return Response({"error": "userId required"}, status=400)

    user = RelocationUser.objects.filter(email__iexact=user_id).first()

    if not user:
        return Response({"error": "User not found"}, status=404)

    if user.right == "Read":
        return Response({"error": "ACCESS DENIED"}, status=403)

    file = request.data.get("file")

    if not file:
        return Response({"error": "No file uploaded"}, status=400)

    try:
        # ================= READ FILE =================
        if file.name.endswith(".csv"):
            df = pd.read_csv(file, header=1)
        else:
            df = pd.read_excel(file, header=1)

        # ================= CLEAN COLUMN NAMES =================
        df.columns = df.columns.str.strip()

        # ================= REPLACE EMPTY VALUES =================
        df = df.replace(
            ["NaN", "nan", "", "NA", "N/A", "-", "NA_IBS"],
            None
        )

        # ================= CIRCLE FILTER =================
        circles = user.circles
        if circles and "CENTRAL" not in circles:
            df = df[df["Circle"].isin(circles)]

        # ================= CHECK REQUIRED COLUMNS =================
        missing = [
            c for c in COLUMN_MAP.keys()
            if c not in df.columns
        ]

        if missing:
            return Response(
                {
                    "status": False,
                    "message": f"Missing columns: {', '.join(missing)}"
                }
            )

        # ================= RENAME COLUMNS =================
        df = df.rename(columns=COLUMN_MAP)
        df = df[list(COLUMN_MAP.values())]

        # ================= DATE CLEANING =================
        INVALID_DATE_STRINGS = {
            "need to check",
            "pending",
            "na",
            "na_ibs",
            "tbd",
            "n/a",
            "-"
        }

        def safe_date(val):
            # Handle None / NaT
            if val is None or pd.isna(val):
                return None

            # pandas Timestamp
            if isinstance(val, pd.Timestamp):
                return val.date()

            # python datetime
            if isinstance(val, datetime):
                return val.date()

            # python date
            if isinstance(val, date):
                return val

            # string dates
            if isinstance(val, str):
                v = val.strip().lower()

                if v in INVALID_DATE_STRINGS:
                    return None

                parsed = pd.to_datetime(
                    val,
                    format="%d-%b-%y",
                    errors="coerce"
                )

                if pd.isna(parsed):
                    return None

                return parsed.date()

            return None

        # Apply date conversion
        for col in df.columns:
            if col.endswith("_date"):
                df[col] = df[col].apply(safe_date)

        # ================= INTEGER FIELD CLEANING =================
        INT_FIELDS = [
            "pri_count",
            "pri_issue_ageing",
            "other_ust_issue_ageing",
            "other_airtel_issue_ageing",
            "total_issue_ageing",
            "clear_rfai_ms1_ageing",
            "rfai_to_ms1_ageing"
        ]

        for col in INT_FIELDS:
            if col in df.columns:
                df[col] = pd.to_numeric(
                    df[col],
                    errors="coerce"
                )

                # convert NaN to None
                df[col] = df[col].where(
                    pd.notna(df[col]),
                    None
                )

        # ================= FINAL CLEANUP =================
        # Convert all pandas types to python objects
        df = df.astype(object)

        # Convert NaT / NaN / <NA> → None
        df = df.where(pd.notna(df), None)

        # ================= DEBUG (REMOVE LATER) =================
        print("EMF CHECK:")
        print(df[["emf_submission_date"]].head(20))

        # ================= RECORDS =================
        records = df.to_dict("records")

        # ================= EXISTING DATA =================
        existing = {
            (
                str(x.circle).strip().lower() if x.circle else "",
                str(x.site_id).strip().lower() if x.site_id else ""
            ): x
            for x in NTSiteTracker.objects.only(
                "circle",
                "site_id"
            )
        }

        to_create = []
        to_update = []

        # ================= SAVE =================
        with transaction.atomic():

            for row in records:

                circle = row.get("circle")
                site_id = row.get("site_id")

                if not circle or not site_id:
                    continue

                circle_key = str(circle).strip().lower()
                site_id_key = str(site_id).strip().lower()

                key = (circle_key, site_id_key)

                row["last_updated_by"] = user.email
                row["last_updated_date"] = timezone.now()

                if key in existing:
                    obj = existing[key]

                    for field, value in row.items():
                        setattr(obj, field, value)

                    to_update.append(obj)

                else:
                    to_create.append(
                        NTSiteTracker(**row)
                    )

            if to_create:
                NTSiteTracker.objects.bulk_create(
                    to_create,
                    batch_size=500
                )

            if to_update:
                NTSiteTracker.objects.bulk_update(
                    to_update,
                    fields=list(COLUMN_MAP.values()) + [
                        "last_updated_by",
                        "last_updated_date"
                    ],
                    batch_size=500
                )

        return Response(
            {
                "status": True,
                "message": "NT Tracker Upload Successfully"
            }
        )

    except Exception as e:
        return Response(
            {
                "status": False,
                "message": str(e)
            },
            status=500
        )
    
############################################################ DOWNLOAD DATA ###################################################################

@api_view(['POST'])
def download_tracker_data_view(request):
    userId = request.data.get('userId')
    year = request.data.get('year')


    user = RelocationUser.objects.filter(email__iexact=userId).first()
    if not user:
        return Response({"error": "User not found"}, status=404)
    
    sync_nt_site_status(request._request)

    # ================= Financial Year Logic =================

    final_filter = Q()
    if year:
        year = int(year)

        fy_start = pd.Timestamp(year=year,
            month=3,
            day=26).date()

        fy_end = pd.Timestamp(year=year + 1,
            month=3,
            day=25).date()

        today = dtime.today().date()

        fy_filter = Q(
            site_onair_date__range=(fy_start,fy_end)
        )

        null_filter = Q(
            site_onair_date__isnull=True
        )

        if fy_start <= today <= fy_end:
            final_filter = fy_filter | null_filter

        else:
            final_filter = fy_filter
   
    circles = user.circles

    

    try:
        # ================= Fetch Data =================
        if 'CENTRAL' in circles:
            if year:
                qs = NTSiteTracker.objects.filter(final_filter)
            else:
                qs = NTSiteTracker.objects.all()

            issue_qs = NTIssue.objects.all()
            site_status_qs = SiteStatus.objects.all()

        else:
            if year:
                qs = NTSiteTracker.objects.filter(
                    final_filter,
                    circle__in=circles
                )
            else:
                qs = NTSiteTracker.objects.filter(
                    circle__in=circles
                )
            issue_qs = NTIssue.objects.filter(
                circle__in=circles
            )

            site_status_qs = SiteStatus.objects.filter(
                circle__in=circles
            )

        df = pd.DataFrame(qs.values())
        issue_df = pd.DataFrame(issue_qs.values())
        site_status_df = pd.DataFrame(site_status_qs.values())

        if df.empty:
            return Response({
                "status":False,
                "message": "No data available"})
        
        if issue_df.empty:
            issue_df = pd.DataFrame(columns=[
                "circle","site_id","issue_owner","milestone",
                "issue_name","start_date","close_date","status",
                "duration","remarks","updated_by","updated_at","created_by","created_at"
            ])

        if site_status_df.empty:
            site_status_df = pd.DataFrame(columns=[
                "site_id",
                "circle",
                "status",
                "date"
            ])    

        # ================= Date Formatting =================
        # Main tracker date formatting
        for col in df.columns:
            if col.endswith("_date"):
                df[col] = pd.to_datetime(df[col],errors='coerce').dt.strftime('%d-%b-%y')

        # Issue tracker date formatting
        for col in issue_df.columns:
            if col.endswith("_date"):
                issue_df[col] = pd.to_datetime(issue_df[col],errors='coerce').dt.strftime('%d-%b-%y')


        if "last_updated_date" in df.columns:
            df["last_updated_date"] = pd.to_datetime(df["last_updated_date"], errors='coerce').dt.strftime('%d-%b-%y %H:%M:%S')

        if "updated_at" in issue_df.columns:
            issue_df["updated_at"] = pd.to_datetime(issue_df["updated_at"], errors='coerce').dt.strftime('%d-%b-%y %H:%M:%S')

        if "created_at" in issue_df.columns:
            issue_df["created_at"] = pd.to_datetime(issue_df["created_at"], errors='coerce').dt.strftime('%d-%b-%y %H:%M:%S')

        issue_df = issue_df.where(pd.notna(issue_df), "")    

        # ================= Unique ID =================
        df.insert(0, "Unique ID", range(1, len(df) + 1))
        df.drop(columns=["id"], inplace=True)

        # ================= Order Columns =================
        df = df[[
            "Unique ID",

            "circle",
            "site_id",
            "project_nt",
            "tur_breakup",
            "new_toco_name",
            "sr_number",

            "ran_oem",
            "media_type",
            "mw_oem",
            "mw_nstallation_partner",

            "site_band",

            "rfai_date",
            "allocation_date",
            "rfai_survey_date",
            "wrfai",
            "wrfai_date",


            "mo_punch_date",
            "material_dispatch_date",
            "material_delivered_date",

            "material_type_hw",
            "material_type_im",

            "installation_start_date",
            "installation_end_date",
            "integration_date",

            "Sacfa_Appied_date",
            "wpc_no",
            "wpc_date",

            "emf_submission_date",
            "nep_id",

            "ran_lkf_status",
            "alarm_status",
            "alarm_rectification_done_date",

            "scft_done_date",
            "scft_offered_date",

            "ran_pat_offer_date",
            "ran_sat_offer_date",

            "mw_plan_id",
            "mw_pat_offer_date",

            "rsl_value_status",
            "enm_status",
            "mw_lkf",
            "mw_sat_offer_date",
            "mw_ms1_mids_date",

            "mw_sacfa_a_end",
            "mw_wpc_a_end",
            "mw_sacfa_b_end",
            "mw_wpc_b_end",

            "site_onair_date",
            "i_deploy_onair_date",

            "current_status",
            "ideploy_status",
            "detailed_remarks",

            "rfai_rejected_date",
            "ideploy_pri_taging",
            "re_rfai_date",

            "pri_count",
            "pri_issue_ageing",
            "other_ust_issue_ageing",
            "other_airtel_issue_ageing",
            "total_issue_ageing",
            "clear_rfai_ms1_ageing",
            "rfai_to_ms1_ageing",

            "ran_pat_accepted_date",
            "ran_sat_accepted_date",
            "mw_pat_accepted_date",
            "mw_sat_accepted_date",
            "scft_accepted_date",

            "kpi_at_offer_date",
            "kpi_at_accepted_date",

            "four_g_ms2_date",
            "ssid",
            "pmis_month",
            "airtel_sign_off",

            "last_updated_date",
            "last_updated_by"
        ]]
        
        df = df.where(pd.notna(df), "")
        df = df.replace(["nan", "NaN", "NAN", "None", "NONE", "null", "NULL"], "")
        df = df.where(pd.notna(df), "")

        # ================= Write Excel =================
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE, exist_ok=True)

        folder = os.path.join(BASE, f"generated_files_{current_date}")
        shutil.rmtree(folder, ignore_errors=True)
        os.makedirs(folder, exist_ok=True)

        file_path = os.path.join(folder, f"NT_TRACKER_{current_date}_{current_time}.xlsx")

        template = os.path.join(BASE, "template", "nt_template.xlsx")
        wb = load_workbook(template)
        ws = wb["Main Tracker"]
        issue_ws = wb["Issue Tracker"]
        status_ws = wb["Site Status"]
      
        status_ws["A1"] = "Circle"
        status_ws["B1"] = "Site ID"
        # ================= Fill Issue Tracker Sheet =================
        issue_start_row = 2

        # remove id column
        if "id" in issue_df.columns:
            issue_df.drop(columns=["id"], inplace=True)

        # column order
        issue_df = issue_df[
            [
                "circle",
                "site_id",
                "issue_owner",
                "milestone",
                "issue_name",
                "start_date",
                "close_date",
                "status",
                "duration",
                "remarks",
                "updated_by",
                "updated_at",
                "created_by",
                "created_at",
            ]
        ]

        for r_idx, row in enumerate(
            dataframe_to_rows(issue_df, index=False, header=False),
            start=issue_start_row
        ):
            for c_idx, value in enumerate(row, start=1):
                cell = issue_ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if issue_df.columns[c_idx - 1] in ["start_date", "close_date"]:
                    cell.number_format = "dd-mmm-yy"

    #site sttus--------------------------------------
        if not site_status_df.empty:

            # remove id if exists
            if "id" in site_status_df.columns:
                site_status_df.drop(columns=["id"], inplace=True)

            # keep required columns
            site_status_df = site_status_df[
                [
                    "circle",
                    "site_id",
                    "date",
                    "status"
                ]
            ]

            # date format
            site_status_df["date"] = pd.to_datetime(
                site_status_df["date"],
                errors="coerce"
            ).dt.strftime("%d-%b-%y")

            # pivot
            pivot_status = site_status_df.pivot_table(
                index=["circle", "site_id"],
                columns="date",
                values="status",
                aggfunc="first",
                fill_value=""
            ).reset_index()

            # header rename
            pivot_status = pivot_status.rename(
                columns={
                    "circle": "Circle",
                    "site_id": "Site ID"
                }
            )

            # clear old sheet
            status_ws.delete_rows(1, status_ws.max_row)
            # write pivot data
            for r_idx, row in enumerate(
                dataframe_to_rows(
                    pivot_status,
                    index=False,
                    header=True
                ),
                start=1
            ):
                for c_idx, value in enumerate(
                    row,
                    start=1
                ):
                    cell = status_ws.cell(
                        row=r_idx,
                        column=c_idx
                    )

                    cell.value = value

                    # header color
                    if r_idx == 1:
                        from openpyxl.styles import PatternFill, Font

                        cell.fill = PatternFill(
                            fill_type="solid",
                            start_color="800080",
                            end_color="800080"
                        )

                        cell.font = Font(
                            color="FFFFFF",
                            bold=True
                        )      
       #--------------------------------         
        date_columns = [col for col in df.columns if col.endswith("_date") and col != 'last_updated_date']

        for col in date_columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        start_row = 3
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                if df.columns[c_idx - 1].endswith("_date") and value:
                    cell.number_format = "dd-mmm-yy"

        for sheet in [ws, issue_ws, status_ws]:
            for column_cells in sheet.columns:
                max_length = 0
                column = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(
                                max_length,
                                len(str(cell.value))
                            )
                    except:
                        pass

                adjusted_width = max_length + 2
                sheet.column_dimensions[
                    column
                ].width = adjusted_width            

        wb.save(file_path)

        relative = file_path.replace(settings.MEDIA_ROOT, "").lstrip(os.sep)
        download_url = request.build_absolute_uri(os.path.join(settings.MEDIA_URL, relative))

        return Response({"download_link":download_url})

    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['DELETE'])
def delete_tracker_data_view(request):
    sr_number = request.data.get('sr_number')   
    try:
        if sr_number:
            qs = NTSiteTracker.objects.filter(sr_number=sr_number)
            
            if not qs.exists():
                return Response(
                    {"message": "No record found with this sr number"},
                    status=status.HTTP_404_NOT_FOUND
                )        
            qs.delete()        
            return Response({"message": "1 site deleted successfully"})
        qs = NTSiteTracker.objects.all()
        
        qs.delete()
        
        return Response({"message": " all sites deleted successfully"})

    except Exception as e:
        return Response({"error": str(e)}, status=500)

        


@api_view(['GET', 'POST'])
def daily_dashboard_view(request):
    circle = request.data.get('circle', [])
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    start_date = request.data.get('from_date')
    end_date = request.data.get('to_date')
    view = request.data.get('view')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None

    all_unique_circles = list(
        NTSiteTracker.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    # all_unique_site_tagging = list(
    #     NTSiteTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        NTSiteTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        NTSiteTracker.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )

    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]

    start_date = dtime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    end_date = dtime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

    try:
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name

        obj = NTSiteTracker.objects.filter(**filters)
        df = pd.DataFrame(obj.values())

        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        # df['rfai_date'] = df['re_rfai_date'].where(df["re_rfai_date"].notna(), df['rfai_date'])

        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")

        today = dtime.today().date()

        if (not month_start or not month_end) and (not start_date or not end_date):
            if today.day >= 26:
                start_date = today.replace(day=26)
                if today.month == 12:
                    end_date = today.replace(year=today.year + 1, month=1, day=25)
                else:
                    end_date = today.replace(month=today.month + 1, day=25)
            else:
                if today.month == 1:
                    start_date = today.replace(year=today.year - 1, month=12, day=26)
                else:
                    start_date = today.replace(month=today.month - 1, day=26)
                end_date = today.replace(day=25)
                
        if month_start and month_end:
            date_range = pd.date_range(start=month_start, end=min(month_end, today)).date
        else:
            date_range = pd.date_range(start=start_date, end=min(end_date, today)).date

        formatted_dates = [d.strftime("%d-%b-%y") for d in date_range]

        if month_start and month_end:
            result_columns = ["Milestone Track/Site Count", "AOP", "CF"] + formatted_dates
        else:
            result_columns = ["Milestone Track/Site Count", "AOP"] + formatted_dates
        result = pd.DataFrame(columns=result_columns)

        milestones = [
            "Allocation Date",
            "RFAI Date",
            "WRFAI",
            "RFAI Survey Date",
            "MO Punch Date",
            "Material Dispatch Date",
            "Material Delivered Date",
            "Installation End Date",
            "Integration Date",
            "EMF Submission Date",
            "Alarm Rectification Done Date",
            # "SCFT I-Deploy Offered Date",
            "SCFT Offered Date",
            "RAN PAT Offer Date",
            "RAN SAT Offer Date",
            "MW PAT Offer Date",
            "MW SAT Offer Date",
            "Site ONAIR Date",
            "I-Deploy ONAIR Date",
        ]

        unique_data.update(**{"Milestone": milestones})

        for milestone in milestones:
            milestone_df_format = (
                milestone.lower()
                .replace(" ", "_")
                .replace("-", "_")
                .replace("(", "")
                .replace(")", "")
            )

            # WRFAI special fix
            if milestone == "WRFAI":
                milestone_df_format = "wrfai_date"

            if milestone_df_format not in df.columns:
                continue

            df[milestone_df_format] = pd.to_datetime(df[milestone_df_format], errors="coerce").dt.date
            if month_start and month_end:
                mask = df['site_onair_date'].isna()
                valid_dates_cf = df.loc[mask, milestone_df_format].dropna()
            # else:
            #     valid_dates = df[milestone_df_format].dropna()
            
            valid_dates_aop = df[milestone_df_format].dropna()

            if valid_dates_aop.empty:
                if month_start and month_end:
                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": "-",
                        "CF": "-",
                        **{d.strftime("%d-%b-%y"): "-" for d in date_range}
                    }
                else:
                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": "-",
                        **{d.strftime("%d-%b-%y"): "-" for d in date_range}
                    }
                
                result.loc[len(result)] = row
                continue
            
            if month_start and month_end:
                cf_count = (valid_dates_cf < month_start).sum()
                aop_count = (valid_dates_aop < month_start).sum()
            else:
            #     cf_count = (valid_dates < start_date).sum()
                aop_count = (valid_dates_aop < start_date).sum()
            
            if month_start and month_end:
                cumulative = cf_count
                row = {"Milestone Track/Site Count": milestone,"AOP": aop_count, "CF": cf_count}
            else:
                cumulative = aop_count
                row = {"Milestone Track/Site Count": milestone,"AOP": aop_count}


            for d in date_range:
                count = (valid_dates_aop == d).sum()
                cumulative += count

                if view == "Cumulative":
                    row[d.strftime("%d-%b-%y")] = cumulative
                else:
                    row[d.strftime("%d-%b-%y")] = count

            result.loc[len(result)] = row

        result.columns = [
            col.strftime("%d-%b-%y") if isinstance(col, (dtime,)) or hasattr(col, "strftime") else col
            for col in result.columns
        ]
        if month_start and month_end:
            result = result[["Milestone Track/Site Count", "AOP" , "CF"] + formatted_dates]
        else:
            result = result[["Milestone Track/Site Count", "AOP"] + formatted_dates]
 
        last_col = formatted_dates[-1]
        dash_mask = result[last_col] == '-'

        result[last_col] = pd.to_numeric(result[last_col], errors='coerce')
        result['Gap'] = -result[last_col].diff()

        nan_mask = result['Gap'].isna()

        result[last_col] = result[last_col].fillna(0).astype(int)
        result['Gap'] = result['Gap'].fillna(0).astype(int)

        result.loc[dash_mask, last_col] = '-'
        result.loc[nan_mask, 'Gap'] = '-'

        result['Gap'] = result['Gap'].astype(str)
        result = result.astype(str).reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(
            lambda col: col.replace(" Date", "") if " Date" in col else col
        )

        new_result = result.copy()

        for i, d in enumerate(formatted_dates, start=1):
            new_result.rename(columns={d: f'date_{i}'}, inplace=True)

        result_json = new_result.to_dict(orient="records")
        json_data = json.dumps(result_json)

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE_URL = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE_URL, exist_ok=True)

        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)

        dashboard_file_path = os.path.join(
            output_folder,
            f"NT_TRACKING_DAILY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx"
        )

        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Daily Waterfall MS1')

        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)

        return Response({
            'message': 'Dashboard created successfully !!!',
            "download_link": download_link,
            "data": json_data,
            "dates": formatted_dates,
            "unique_data": unique_data
        }, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
    


@api_view(['GET', 'POST'])
def weekly_monthly_dashboard_view(request):
    circle = request.data.get('circle', [])
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    view = request.data.get('view')
 
    # Default to 'ALL' if not provided
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None
   
    all_unique_circles = list(
        NTSiteTracker.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    # all_unique_site_tagging = list(
    #     NTSiteTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        NTSiteTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        NTSiteTracker.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )

    # ✅ Add "ALL" at the top of each list
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    try:
        # 🔹 Dynamic filters
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name
 
        # 🔹 Fetch data
        obj = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        # df['rfai_date'] = df['re_rfai_date'].where(df["re_rfai_date"].notna(), df['rfai_date'])
       
        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")
               
        today = dtime.today().date()
 
        if today.month >= 4:
            fy_start = dtime(today.year, 3, 26).date()   # 26-Apr current year start
            fy_end = dtime(today.year + 1, 3, 25).date() # 25-Mar next year end
        else:
            fy_start = dtime(today.year - 1, 3, 26).date()
            fy_end = dtime(today.year, 3, 25).date()
 
        # 🧩 2. Determine current cycle (26th prev → 25th curr)
        if today.day >= 26:
            if today.month == 12:
                current_cycle_start = dtime(today.year, 12, 26).date()
                current_cycle_end = dtime(today.year + 1, 1, 25).date()
            else:
                current_cycle_start = dtime(today.year, today.month, 26).date()
                current_cycle_end = dtime(today.year, today.month + 1, 25).date()
        else:
            if today.month == 1:
                current_cycle_start = dtime(today.year - 1, 12, 26).date()
            else:
                current_cycle_start = dtime(today.year, today.month - 1, 26).date()
            current_cycle_end = dtime(today.year, today.month, 25).date()
 
        #  3. Create monthly periods (Apr → month before current)
        months = []
        start = fy_start
        while start <= current_cycle_start:
            if start.month == 12:
                end = dtime(start.year + 1, 1, 25).date()
            else:
                end = dtime(start.year, start.month + 1, 25).date()
            months.append((start, end))
            start = end + timedelta(days=1)
 
        #  4. Create weekly periods for current month
        weeks = []
        
        if month_start and month_end:
            current_cycle_start = month_start
            current_cycle_end = month_end
        
        week_start = current_cycle_start
 
        while week_start <= current_cycle_end:
            week_end = week_start + timedelta(days=6)
            if week_end > current_cycle_end:
                week_end = current_cycle_end
            weeks.append((week_start, week_end))
            week_start = week_end + timedelta(days=1)
 
 
        result = pd.DataFrame()
       
       
        # result.set_index("Milestone Track/Site Count", inplace=True)
 
        milestones = [
        "Allocation Date",
        "RFAI Date",
        "RFAI Survey Date",
        "WRFAI",
        # "RFAI Survey Done Date",
        "MO Punch Date",
        "Material Dispatch Date",
        "Material Delivered Date",
        # "Installation Start Date",
        "Installation End Date",
        "Integration Date",
        "EMF Submission Date",
        "Alarm Rectification Done Date",
        # "SCFT Done Date",
        # "SCFT I-Deploy Offered Date",
        "SCFT Offered Date",
        "RAN PAT Offer Date",
        "RAN SAT Offer Date",
        "MW PAT Offer Date",
        "MW SAT Offer Date",
        # "MW MS1 Date (MIDS)",
        "Site ONAIR Date",
        "I-Deploy ONAIR Date",
    ]
 
        data = []
        
        for milestone in milestones:
            milestone_df_format = (
                milestone.lower()
                .replace(" ", "_")
                .replace("-", "_")
                .replace("(", "")
                .replace(")", "")
            )

            if milestone == "WRFAI":
                milestone_df_format = "wrfai_date"

            if milestone_df_format not in df.columns:
                continue
            
            milestone_data = pd.to_datetime(df[milestone_df_format], errors='coerce').dt.date
            row = {"Milestone Track/Site Count": milestone}
            if milestone_data.dropna().empty:
                row["CF"] = "-"
                # Set all month and week columns to "-"
                for _, end in months:
                    month_name = end.strftime("%b-%y")
                    row[month_name] = "-"
                for i, _ in enumerate(weeks, 1):
                    week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                    row[week_name] = "-"
                data.append(row)
                continue
 
 
            # CF → sites before financial year start
            row["CF"] = (milestone_data < fy_start).sum()
 
            cumulative_month = row["CF"]
            
            for start, end in months:
                month_name = end.strftime("%b-%y")
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_month += count
                if view == 'Cumulative':
                    row[month_name] = cumulative_month
                else:
                    row[month_name] = count
 
            cumulative_week = 0
            for i, (start, end) in enumerate(weeks, 1):
                week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                if(start >= today):
                    row[week_name] = 0
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_week += count
                if view == 'Cumulative':
                    row[week_name] = cumulative_week
                else:
                    row[week_name] = count
 
            data.append(row)
 

        result = pd.DataFrame(data)
 
        for col in result.columns:
            if col != "Milestone Track/Site Count":
                result[col] = result[col].astype(str)
 
        result = result.reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(lambda col: col.replace(" Date", "") if " Date" in col else col)
 
        new_data = result.copy()
        months_columns = [col for col in new_data.columns.to_list()[2:] if " W" not in col]
        months_data = new_data[new_data.columns.to_list()[:2] + months_columns].copy()
        
        weeks_columns = [col for col in new_data.columns.to_list()[2:] if " W" in col]
        
        week_data = new_data[new_data.columns.to_list()[:1] + weeks_columns].copy()
        
        months_data.rename(columns={
            col : f"Month-{i}" for i, col in enumerate(months_columns, start=1)
        }, inplace=True)
        
        week_data.rename(columns={
            col : f"Month_Week-{i}" for i, col in enumerate(weeks_columns, start=1)
        }, inplace=True)
        
        unique_data.update(**{"month_columns": months_columns, "week_columns" : weeks_columns})
        

        month_dict_data = months_data.to_dict(orient="records")
        month_json_data = json.dumps(month_dict_data)
        week_dict_data = week_data.to_dict(orient="records")
        week_json_data = json.dumps(week_dict_data)
        
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"NT_TRACKING_WEEKLY_MONTHY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx")
       
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Monthly MS1')
       
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
   
        return Response({'message': 'Weekly and Monthly Dashboard created successfully !!!', "download_link": download_link, "unique_data": unique_data, "months_data": month_json_data, "week_data": week_json_data}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
    

@api_view(['GET', 'POST'])
def gap_view(request):
    userId = request.data.get('userId')
    
    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    
    circles = user.circles
    
    circle = request.data.get('circle', [])
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    last_date = request.data.get('last_date')
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    gap = request.data.get('gap')
    
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]

    last_date = dtime.strptime(last_date, "%Y-%m-%d").date() if last_date else None
    milestone1 = (milestone1 or "").strip().lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")
    milestone2 = (milestone2 or "").strip().lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")

    print("milestone1:", repr(milestone1))
    print("milestone2:", repr(milestone2))

    # if not milestone1:
    #     return Response({"error": "milestone1 missing"}, status=400)

    # if not milestone2:
    #     return Response({"error": "milestone2 missing"}, status=400)

    try:
        
        today = dtime.today().date()
 
        if not last_date:
            if today.day >= 26:
                if today.month == 12:
                    last_date = today.replace(year=today.year + 1, month=1, day=25)
                else:
                    last_date = today.replace(month=today.month + 1, day=25)
            else:
                last_date = today.replace(day=25)
                
            last_date = min(today - timedelta(days=1), last_date)
                
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name
        
        if int(gap)>0 :
            filters[f"{milestone1}__lte"] = last_date
        else:
            filters[f"{milestone2}__lte"] = last_date

        # if f"{milestone1}__lte" in filters:
        #     del filters[f"{milestone1}__lte"]
 
        # if f"{milestone2}__lte" in filters:
        #     del filters[f"{milestone2}__lte"]
 
        # 🔹 Fetch data
        obj1 = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        df1 = pd.DataFrame(obj1.values())
        
        if int(gap)>0 :
            filters[f"{milestone2}__lte"] = last_date
        else:
            filters[f"{milestone1}__lte"] = last_date
        
        obj2 = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        df2 = pd.DataFrame(obj2.values())
        
        df1['key'] = df1['circle'].astype(str) + "_" + df1['site_id'].astype(str)
        df2['key'] = df2['circle'].astype(str) + "_" + df2['site_id'].astype(str)

        # df = df1[~df1['key'].isin(df2['key'])].drop(columns=['key'])
        # if len(df1) > len(df2):  
        df = df1[~df1['key'].isin(df2['key'])].drop(columns=['key'])
        # else:
        #     df = df2[~df2['key'].isin(df1['key'])].drop(columns=['key'])
        
        sites = df['site_id'].dropna().unique().tolist()

        
        # filters = {
        #     "site_id__in" : sites
        # }
        
        # issue_obj = RelocationIssue.objects.filter(site_id__in=sites)
        # issue_df = pd.DataFrame(issue_obj.values())

        # rename_map = {
        #     "circle": "Circle",
        #     "site_id": "Site ID",
        #     "issue_owner": "Issue Owner",
        #     "milestone": "Milestone",
        #     "issue_name": "Issue Name",
        #     "start_date": "Start Date",
        #     "close_date": "Close Date",
        #     "status": "Status",
        #     "duration": "Duration",
        #     "updated_by": "Updated_by",
        #     "updated_at": "Updated_at",
        #     "created_by": "Created_by",
        #     "created_at": "Created_at"
        # }

        # print(issue_df)
        
        # issue_df = issue_df.rename(columns=rename_map)

        # required_cols = ["Circle", "Site ID", "Issue Owner", "Milestone", "Issue Name", "Start Date", "Close Date", "Status", "Duration", "Updated_by", "Updated_at", "Created_by", "Created_at"]
        # for col in required_cols:
        #     if col not in issue_df.columns:
        #         issue_df[col] = None

        # # Convert date columns safely
        # for col in ["Start Date", "Close Date"]:
        #     if col in issue_df.columns:
        #         issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
        
        # df = update_ageing(df, issue_df)
 
        
        for col in df.columns:
            if col != 'last_updated_date':
                converted = pd.to_datetime(df[col], errors='coerce')
 
                if converted.notna().sum() > 0:
                    df[col] = converted.dt.strftime('%d-%b-%y')
            else:
                converted = pd.to_datetime(df[col], errors='coerce')
                if converted.notna().sum() > 0:
                    df[col] = converted.dt.strftime('%d-%b-%y %H:%M:%S')
 
 
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        tracker_file_path = os.path.join(output_folder, f"NT_TRACKER_GAP_FILE_{milestone1}_{milestone2}_{circle}_{current_status}_{new_toco_name}_{current_date}_{current_time}.xlsx")
        
        df.insert(0, "Unique ID", range(1, len(df) + 1))
        
       
        df.drop(columns=['id'], inplace=True)

        df = df[[
            
            "circle",
            "site_id",
            "project_nt",
            "tur_breakup",
            "new_toco_name",
            "sr_number",

            "ran_oem",
            "media_type",
            "mw_oem",
            "mw_nstallation_partner",

            "site_band",

            "rfai_date",
            "allocation_date",
            "rfai_survey_date",
            "wrfai",

            "mo_punch_date",
            "material_dispatch_date",
            "material_delivered_date",

            "material_type_hw",
            "material_type_im",

            "installation_start_date",
            "installation_end_date",
            "integration_date",

            "Sacfa_Appied_date",
            "wpc_no",
            "wpc_date",

            "nep_id",
            "emf_submission_date",

            "ran_lkf_status",
            "alarm_status",
            "alarm_rectification_done_date",

            "scft_done_date",
            "scft_offered_date",

            "ran_pat_offer_date",
            "ran_sat_offer_date",

            "mw_plan_id",
            "mw_pat_offer_date",

            "rsl_value_status",
            "enm_status",
            "mw_lkf",
            "mw_sat_offer_date",
            "mw_ms1_mids_date",

            "mw_sacfa_a_end",
            "mw_wpc_a_end",
            "mw_sacfa_b_end",
            "mw_wpc_b_end",

            "site_onair_date",
            "i_deploy_onair_date",

            "current_status",
            "ideploy_status",
            "detailed_remarks",

            "rfai_rejected_date",
            "ideploy_pri_taging",
            "re_rfai_date",

            "pri_count",
            "pri_issue_ageing",
            "other_ust_issue_ageing",
            "other_airtel_issue_ageing",
            "total_issue_ageing",
            "clear_rfai_ms1_ageing",
            "rfai_to_ms1_ageing",

            "ran_pat_accepted_date",
            "ran_sat_accepted_date",
            "mw_pat_accepted_date",
            "mw_sat_accepted_date",
            "scft_accepted_date",

            "kpi_at_offer_date",
            "kpi_at_accepted_date",

            "four_g_ms2_date",
            "ssid",
            "pmis_month",
            "airtel_sign_off",

            "last_updated_date",
            "last_updated_by"
        ]]
        

 
 
        json_dict_data = df.to_dict(orient="records")
        json_val = json.dumps(json_dict_data)
        
        
        template_path = os.path.join(BASE_URL, "template", "nt_template.xlsx")
        wb = load_workbook(template_path)
        ws = wb.active  
        date_columns = [col for col in df.columns if col.endswith("_date") and col != 'last_updated_date']

        for col in date_columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        start_row = 3
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                # apply date format for *_date columns
                if df.columns[c_idx - 1].endswith("_date") and hasattr(value, "strftime") and df.columns[c_idx - 1] != 'last_updated_date':
                    cell.number_format = "dd-mmm-yy"
 

        # ws2 = wb.create_sheet(title="Issues Tracker")

        # # Clean issue_df (remove id if present)
        # if "id" in issue_df.columns:
        #     issue_df = issue_df.drop(columns=["id"])
            
        # # print(issue_df['updated_by'])

        # def remove_tz_safe(x):
        #     if isinstance(x, datetime) and x.tzinfo is not None:
        #         return x.replace(tzinfo=None)
        #     return x

        # # apply to every cell in issue_df
        # issue_df = issue_df.applymap(remove_tz_safe)
        
        # # print(issue_df['updated_by'])


        # # Convert date columns
        # issue_date_columns = [col for col in issue_df.columns if "Date" in col]
        # for col in issue_date_columns:
        #     issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
            
        # # print(issue_df['updated_by'])

        # # ---- Write header ----
        # for col_idx, column_name in enumerate(issue_df.columns, start=1):
        #     ws2.cell(row=1, column=col_idx, value=column_name)
            
        # # print(issue_df['updated_by'])

        # # ---- Write data ----
        # for r_idx, row in enumerate(dataframe_to_rows(issue_df, index=False, header=False), start=2):
        #     for c_idx, value in enumerate(row, start=1):
        #         cell = ws2.cell(row=r_idx, column=c_idx)
        #         cell.value = value

        #         # Apply date formatting
        #         if issue_df.columns[c_idx - 1] in issue_date_columns and hasattr(value, "strftime"):
        #             cell.number_format = "dd-mmm-yy"

        wb.save(tracker_file_path)
        
        relative_path = tracker_file_path.replace(settings.MEDIA_ROOT, '').lstrip(os.sep)
        download_url = request.build_absolute_uri(
            os.path.join(settings.MEDIA_URL, relative_path).replace('\\', '/')
        )
        
        
        
        
        return Response({'message': 'request processed successfully !!!', "data": json_val, "download_link": download_url}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500) 






@api_view(['GET', 'POST'])
def frontend_nt_display_view(request):
    userId = request.data.get('userId')

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    
    # circles = user.circles
    circles ="CENTRAL"

    type = request.data.get('day_type')
    milestone = request.data.get('milestone')
    col_name = request.data.get('col_name')
    view = request.data.get('view')
    circle = request.data.get('circle', [])
    current_status = request.data.get('current_status', [])
    toco_name = request.data.get('toco_name', [])
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None

    # circle = [ACCESS_RIGHTS[userId]["Circle"]] if ACCESS_RIGHTS[userId]["Circle"] != 'CENTRAL' else [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    # if 'ALL' in circle:
    #     circle = circles
    # elif 'ALL' not in circles:
    #     circle = [c for c in circle if c in circles]

    circle = [c.strip() for c in circle.split(',')] if circle else ["CENTRAL"]
    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    toco_name = [n.strip() for n in toco_name.split(',')] if toco_name else ["ALL"]
    
    if 'CENTRAL' in circle:
        circle = circles
    elif 'CENTRAL' not in circles:
        circle = [c for c in circle if c in circles]
    
    if not type or not milestone or not col_name or not view:
        return Response({"error": "Missing required parameters"}, status=400)
    
    milestone = milestone.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
    
    try:
        def get_week_range(col_name):
            month_str, week_part = col_name.split()
            month_abbr, year_suffix = month_str.split('-')
            week_num = int(week_part.replace('W', ''))

            year = 2000 + int(year_suffix)  # '25' -> 2025
            month = datetime.strptime(month_abbr, "%b").month

            # Define cycle range: 26th of previous month → 25th of current month
            prev_month = month - 1 if month > 1 else 12
            prev_year = year if month > 1 else year - 1

            cycle_start = date(prev_year, prev_month, 26)
            cycle_end = date(year, month, 25)

            # Build week ranges: every 7 days from cycle_start to cycle_end
            weeks = []
            start = cycle_start
            while start <= cycle_end:
                end = start + timedelta(days=6)
                if end > cycle_end:
                    end = cycle_end
                weeks.append((start, end))
                start = end + timedelta(days=1)

            # Get requested week
            week_start, week_end = weeks[week_num - 1]
            return week_start, week_end, cycle_start
        
        filters = {}
        if type == 'daily':
            col_name = dtime.strptime(col_name, "%d-%b-%y").date()
            if view == 'Cumulative':
                filters[f"{milestone}__lte"] = col_name
            else:
                filters[milestone] = col_name
            
        elif type == 'weekly':
            week_start, week_end, cycle_start = get_week_range(col_name)

            if view == 'Cumulative':
                filters[f"{milestone}__range"] = (cycle_start, week_end)
            else:
                filters[f"{milestone}__range"] = (week_start, week_end)

        else:
            month_abbr, year_suffix = col_name.split('-')
            year = 2000 + int(year_suffix)
            month = datetime.strptime(month_abbr, "%b").month

            # Compute cycle range
            prev_month = month - 1 if month > 1 else 12
            prev_year = year if month > 1 else year - 1

            cycle_start = datetime(prev_year, prev_month, 26)
            cycle_end = datetime(year, month, 25)

            if view == 'Cumulative':
                filters[f"{milestone}__lte"] = cycle_end
            else:
                filters[f"{milestone}__range"] = (cycle_start, cycle_end)


        if "CENTRAL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in toco_name:
            filters["new_toco_name__in"] = toco_name


        obj = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        if df.empty :
            json_data = []
            json_val = json.dumps(json_data)
            return Response({"message" : "No data found", "data": json_val},status = 200)
        
        if month_start and month_end:
            df = df[df['site_onair_date'].isna() | (df['site_onair_date'] >= month_start)]
        
        sites = df['site_id'].dropna().unique().tolist()
        
        issue_obj = NTIssue.objects.filter(site_id__in=sites)
        issue_df = pd.DataFrame(issue_obj.values())

        rename_map = {
            "circle": "Circle",
            "site_id": "Site ID",
            "issue_owner": "Issue Owner",
            "milestone": "Milestone",
            "issue_name": "Issue Name",
            "start_date": "Start Date",
            "close_date": "Close Date",
            "status": "Status",
            "duration": "Duration",
            "remarks": "Remarks",
            "updated_by": "Updated_by",
            "updated_at": "Updated_at",
            "created_by": "Created_by",
            "created_at": "Created_at"
        }

        print(issue_df)
        
        issue_df = issue_df.rename(columns=rename_map)

        required_cols = ["Circle", "Site ID", "Issue Owner", "Milestone", "Issue Name", "Start Date", "Close Date", "Status", "Duration", "Remarks", "Updated_by", "Updated_at", "Created_by", "Created_at"]
        for col in required_cols:
            if col not in issue_df.columns:
                issue_df[col] = None

        # # Convert date columns safely
        # for col in ["Start Date", "Close Date"]:
        #     if col in issue_df.columns:
        #         issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
        
        # df = update_ageing(df, issue_df)

        for col in df.columns:
            if 'date' in col:
                if col != 'last_updated_date' :
                    converted = pd.to_datetime(df[col], errors='coerce')
    
                    if converted.notna().sum() > 0:
                        df[col] = converted.dt.strftime('%d-%b-%y')
                else:
                    converted = pd.to_datetime(df[col], errors='coerce')
                    if converted.notna().sum() > 0:
                        df[col] = converted.dt.strftime('%d-%b-%y %H:%M:%S')
 
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        tracker_file_path = os.path.join(output_folder, f"FRONTEND_NT_TRACKER_FILE_{current_date}_{current_time}.xlsx")
        
        df.insert(0, "Unique ID", range(1, len(df) + 1))
        df.drop(columns=['id'], inplace=True)

        df = df[[
            'Unique ID','circle','site_id',
            'project_nt',
            'tur_breakup',
            'new_toco_name',
            'sr_number',

            'ran_oem',
            'media_type',
            'mw_oem',
            'mw_nstallation_partner',

            'site_band',

            'rfai_date',
            'allocation_date',
            'rfai_survey_date',
            'wrfai',
            'wrfai_date',

            'mo_punch_date',
            'material_dispatch_date',
            'material_delivered_date',

            'material_type_hw',
            'material_type_im',

            'installation_start_date',
            'installation_end_date',
            'integration_date',

            'Sacfa_Appied_date',
            'wpc_no',
            'wpc_date',

            'emf_submission_date',
            'nep_id',

            'ran_lkf_status',
            'alarm_status',
            'alarm_rectification_done_date',

            'scft_done_date',
            'scft_offered_date',

            'ran_pat_offer_date',
            'ran_sat_offer_date',

            'mw_plan_id',
            'mw_pat_offer_date',

            'rsl_value_status',
            'enm_status',
            'mw_lkf',
            'mw_sat_offer_date',
            'mw_ms1_mids_date',

            'mw_sacfa_a_end',
            'mw_wpc_a_end',
            'mw_sacfa_b_end',
            'mw_wpc_b_end',

            'site_onair_date',
            'i_deploy_onair_date',

            'current_status',
            'ideploy_status',
            'detailed_remarks',

            'rfai_rejected_date',
            'ideploy_pri_taging',
            're_rfai_date',

            'pri_count',
            'pri_issue_ageing',
            'other_ust_issue_ageing',
            'other_airtel_issue_ageing',
            'total_issue_ageing',
            'clear_rfai_ms1_ageing',
            'rfai_to_ms1_ageing',

            'ran_pat_accepted_date',
            'ran_sat_accepted_date',
            'mw_pat_accepted_date',
            'mw_sat_accepted_date',
            'scft_accepted_date',

            'kpi_at_offer_date',
            'kpi_at_accepted_date',

            'four_g_ms2_date',
            'ssid',
            'pmis_month',
            'airtel_sign_off',

            'last_updated_date',
            'last_updated_by'
        ]]
        
        
        template_path = os.path.join(BASE_URL, "template", "nt_template.xlsx")
        wb = load_workbook(template_path)
        ws = wb.active

        # JSON response data
        df_copy = df.copy()
        df_copy = df_copy.fillna('').astype(str)

        json_dict_data = df_copy.to_dict(orient="records")
        json_val = json.dumps(json_dict_data)

        # Convert date columns
        date_columns = [
            col for col in df.columns
            if col.endswith("_date") and col != "last_updated_date"
        ]

        for col in date_columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        # Write main tracker data
        start_row = 3
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False),
            start=start_row
        ):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if (
                    df.columns[c_idx - 1].endswith("_date")
                    and hasattr(value, "strftime")
                    and df.columns[c_idx - 1] != "last_updated_date"
                ):
                    cell.number_format = "dd-mmm-yy"


        # ================= Issues Tracker Sheet =================
        ws2 = wb.create_sheet(title="Issues Tracker")

        # Remove id column
        if "id" in issue_df.columns:
            issue_df = issue_df.drop(columns=["id"])


        def remove_tz_safe(x):
            if isinstance(x, datetime) and x.tzinfo is not None:
                return x.replace(tzinfo=None)
            return x


        # Remove timezone
        issue_df = issue_df.applymap(remove_tz_safe)

        # Convert issue dates
        issue_date_columns = [
            col for col in issue_df.columns
            if "Date" in col
        ]

        for col in issue_date_columns:
            issue_df[col] = pd.to_datetime(
                issue_df[col],
                errors="coerce"
            )

        # Write issue header
        for col_idx, column_name in enumerate(
            issue_df.columns,
            start=1
        ):
            ws2.cell(
                row=1,
                column=col_idx,
                value=column_name
            )

        # Write issue data
        for r_idx, row in enumerate(
            dataframe_to_rows(
                issue_df,
                index=False,
                header=False
            ),
            start=2
        ):
            for c_idx, value in enumerate(row, start=1):
                cell = ws2.cell(
                    row=r_idx,
                    column=c_idx
                )
                cell.value = value

                if (
                    issue_df.columns[c_idx - 1]
                    in issue_date_columns
                    and hasattr(value, "strftime")
                ):
                    cell.number_format = "dd-mmm-yy"


        # Save file
        wb.save(tracker_file_path)

        relative_path = tracker_file_path.replace(
            settings.MEDIA_ROOT,
            ""
        ).lstrip(os.sep)

        download_url = request.build_absolute_uri(
            os.path.join(
                settings.MEDIA_URL,
                relative_path
            ).replace("\\", "/")
        )

        return Response(
            {
                "message": "request processed successfully !!!",
                "data": json_val,
                "download_link": download_url
            },
            status=200
        )

    except Exception as e:
        return Response(
                {"error": str(e)},
                status=500
            )


@api_view(['POST'])
def frontend_nt_update_view(request):
    userId = request.data.get('userId')
    data = request.data.get('data')

    if not userId or not data:
        return Response({'error': 'userId and data are required.'},status=400
        )

    try:
        user = RelocationUser.objects.filter(
            email=userId.lower()
        ).first()

        if not user:
            return Response({"error": "User not found"},status=404
            )

        if user.right == 'Read':
            return Response(
                {'error': 'ACCESS DENIED'},
                status=403
            )

        # circles = user.circles
        circles ="CENTRAL"
        # user="Admin"

        # ================= Parse data =================
        if isinstance(data, str):
            data = json.loads(data)
        df = pd.DataFrame([data])

        if df.empty:
            return Response(
                {'error': 'No data provided for update.'},
                status=400
            )

        record = df.iloc[0].to_dict()

        print(record)

        # ================= Safe converters =================
        INVALID_DATE_STRINGS = {
            "need to check",
            "pending",
            "na",
            "tbd",
            "n/a",
            "nan",
            "-"
        }

        def safe_datetime(value):
            if pd.isna(value) or value is pd.NaT:
                return None

            if isinstance(value, (pd.Timestamp, datetime)):
                return value

            if isinstance(value, str):
                v = value.strip().lower()

                if not v or v in INVALID_DATE_STRINGS:
                    return None

                try:
                    val = pd.to_datetime(
                        value,
                        errors="coerce"
                    )

                    if pd.isna(val):
                        return None

                    return val.to_pydatetime()

                except:
                    return None

            return None

        def safe_int(value):
            try:
                return int(value)
            except:
                return None

        # ================= Get identifiers =================
        circle_val = (
            str(record.get("circle") or record.get("Circle")).strip()
            if record.get("circle") or record.get("Circle")
            else None
        )

        site_id_val = (
            str(record.get("site_id") or record.get("Site ID")).strip()
            if record.get("site_id") or record.get("Site ID")
            else None
        )

        if not circle_val or not site_id_val:
            return Response(
                {'error': 'Circle and Site ID are required.'},
                status=400
            )

        # ================= Circle restriction =================
        if 'CENTRAL' not in circles and circle_val not in circles:
            return Response(
                {
                    'error':
                    f'You are not allowed to edit {circle_val} circle data.'
                },
                status=403
            )

        # ================= Validate editable columns =================
        required_columns = [
            col.lower()
            .strip()
            .replace(" ", "_")
            .replace("-", "_")
            .replace("/", "_")
            .replace("4", "four_")
            .replace("5", "five_")
            for col in user.columns
        ]

        allowed_data = {}

        for col in record:
            field_name = (
                col.lower()
                .strip()
                .replace(" ", "_")
                .replace("-", "_")
                .replace("/", "_")
                .replace("4", "four_")
                .replace("5", "five_")
            )

            if field_name not in required_columns:
                continue

            val = record[col]

            if "date" in field_name:
                val = safe_datetime(val)

            elif "ageing" in field_name or "count" in field_name:
                val = safe_int(val)

            allowed_data[field_name] = val

        # ================= System fields =================
        allowed_data["last_updated_by"] = userId
        allowed_data["last_updated_date"] = timezone.now()

        # ================= Fetch record =================
        try:
            obj = NTSiteTracker.objects.get(
                circle__iexact=circle_val,
                site_id__iexact=site_id_val
            )

        except NTSiteTracker.DoesNotExist:
            return Response(
                {'error': 'Record not found in database.'},
                status=404
            )

        # ================= Update =================
        for field, value in allowed_data.items():
            setattr(obj, field, value)

        obj.save(
            update_fields=list(allowed_data.keys())
        )

        return Response(
            {
                'status': True,
                'message': 'Record updated successfully.'
            },
            status=200
        )

    except Exception as e:
        return Response(
            {'error': str(e)},
            status=500
        )




@api_view(['GET', 'POST'])
def ms1_ageing_dashboard_table1(request):
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

  

    all_unique_current_status = list(
        NTSiteTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Allocation",
        "RFAI",
        "WRFAI",
        "RFAI Survey",
        "MO Punch",
        "Material Dispatch",
        "Material Delivered",
        "Installation End",
        "Integration",
        "EMF Submission",
        "Alarm Rectification Done",
        "SCFT Offered",
        "RAN PAT Offer",
        "RAN SAT Offer",
        "MW PAT Offer",
        "MW SAT Offer",
        "Site ONAIR",
        "I-Deploy ONAIR",
    ]
    
    unique_data = {
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones
    }

    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:
        month_filtered = int(month_filtered) if str(month_filtered).strip() else None
        year_filtered = int(year_filtered) if str(year_filtered).strip() else None

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        # if site_tagging and "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging

        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status

        # if month_start and month_end:
        #     milestone1_col = "wrfai_date" if milestone1 == "WRFAI" else milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + "_date"
        #     filters[f"{milestone1_col}__range"] = (month_start, month_end)
            
        obj = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        date_cols = [
            "wrfai_date" if col == "WRFAI"else col.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + "_date" for col in milestones
        ]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']
        
        
        def generate_table_summary(df, milestone1, milestone2):


            milestone_cols = ["wrfai_date" if col == "WRFAI"else col.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
                for col in milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]
            ]

            temp = df.copy()

            circles = ["ALL"] + temp["circle"].unique().tolist()

            final_df = pd.DataFrame()

            milestone1_col = "wrfai_date" if milestone1 == "WRFAI" else milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
            milestone2_col = "wrfai_date" if milestone2 == "WRFAI" else milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
            
            for circle in circles:
                if circle != "ALL":
                    temp  = df[df["circle"] == circle]

                done_row = {}
                pending_row = {}

                for col, label in zip(
                    milestone_cols,
                    milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]
                ):

                    done_count = (
                        temp[col].notna() &
                        temp[milestone1_col].notna()
                    ).sum()

                    pending_count = (
                        temp[col].isna() &
                        temp[milestone1_col].notna()
                    ).sum()

                    # normal columns
                    done_row[label] = done_count
                    pending_row[label] = pending_count

                    # only for RFAI
                    if label == "RFAI":
                        if month_start and month_end:
                            current_rfai = (
                                temp["rfai_date"].notna() &
                                (temp["rfai_date"] >= pd.Timestamp(month_start)) &
                                (temp["rfai_date"] <= pd.Timestamp(month_end))
                            ).sum()

                            cf_rfai = (
                                temp["rfai_date"].notna() &
                                (temp["rfai_date"] < pd.Timestamp(month_start)) &
                                temp["site_onair_date"].isna()
                            ).sum()

                            total_rfai = current_rfai + cf_rfai
                        else:
                            current_rfai = temp["rfai_date"].notna().sum()
                            cf_rfai = 0
                            total_rfai = current_rfai

                        done_row["RFAI"] = int(current_rfai)
                        done_row["CF RFAI"] = int(cf_rfai)
                        done_row["Total RFAI"] = int(total_rfai)

                        pending_row["RFAI"] = 0
                        pending_row["CF RFAI"] = 0
                        pending_row["Total RFAI"] = 0

                        continue


                rows_to_add = pd.DataFrame([
                    {"Circle": circle, "Site Status": "Done", **done_row},
                    {"Circle": circle, "Site Status": "Pending", **pending_row},
                ])

                final_df = pd.concat([final_df, rows_to_add], ignore_index=True)

            for c in final_df.columns:
                if c != "Site Status" and c != "Circle":
                    final_df[c] = (
                        pd.to_numeric(final_df[c], errors="coerce")
                        .fillna(0)
                        .astype(int)
                    )

            return final_df

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        table_summary = generate_table_summary(df, start_label, end_label)
 
        table_summary = table_summary.applymap(lambda x: str(x) if pd.notna(x) else "")      


        result_json = table_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "table_summary": json_data1
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
  


  
@api_view(['GET', 'POST'])
def ms1_ageing_dashboard_table2(request):
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    breakpoint1 = request.data.get('breakpoint1')
    breakpoint2 = request.data.get('breakpoint2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    type = request.data.get('type')

    # all_unique_site_tagging = list(
    #     NTSiteTracker.objects.exclude(site_tagging__isnull=True)

    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        NTSiteTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Allocation",
        "RFAI",
        "WRFAI",
        "RFAI Survey",
        "MO Punch",
        "Material Dispatch",
        "Material Delivered",
        "Installation End",
        "Integration",
        "EMF Submission",
        "Alarm Rectification Done",
        "SCFT Offered",
        "RAN PAT Offer",
        "RAN SAT Offer",
        "MW PAT Offer",
        "MW SAT Offer",
        "Site ONAIR",
        "I-Deploy ONAIR",
    ]
    
    unique_data = {
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones" : milestones
    }
    

    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    
    try:

        milestone1_col = "wrfai_date" if milestone1 == "WRFAI" else (milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date")
        milestone2_col = "wrfai_date" if milestone1 == "WRFAI" else (milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date")

        breakpoint1 = int(breakpoint1)
        breakpoint2 = int(breakpoint2)
        
        temp = breakpoint1
        breakpoint1 = min(breakpoint1, breakpoint2)
        breakpoint2 = max(temp, breakpoint2)

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None

        milestones = [
        "Allocation",
        "RFAI",
        "WRFAI",
        "RFAI Survey",
        "MO Punch",
        "Material Dispatch",
        "Material Delivered",
        "Installation End",
        "Integration",
        "EMF Submission",
        "Alarm Rectification Done",
        "SCFT Offered",
        "RAN PAT Offer",
        "RAN SAT Offer",
        "MW PAT Offer",
        "MW SAT Offer",
        "Site ONAIR",
        "I-Deploy ONAIR",
    ]
        

        filters = {}
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
            
        if type == "type2":
            if month_start and month_end:
                filters[f"{milestone1_col}__range"] = (month_start, month_end)
                

        obj = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        
        if type == "type1":
            if month_start and month_end:
                obj = obj.filter(
                    Q(**{f"{milestone1_col}__range": (month_start, month_end)}) |
                    Q(**{f"{milestone2_col}__range": (month_start, month_end)})
                )
        

        df = pd.DataFrame(obj.values())

        
        date_cols = ["wrfai_date" if col == "WRFAI" else col.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_") .replace("5", "five_") + '_date'for col in milestones]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']

        if month_start is not None:
            month_start = pd.to_datetime(month_start)

        if month_end is not None:
            month_end = pd.to_datetime(month_end)

        def generate_done_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            temp = df.copy()
            
            temp["days_diff"] = (temp[end_col] - temp[start_col]).dt.days
            
            if type == "type2":
        
                summary = (
                    temp.groupby("Circle").apply(lambda g: pd.Series({
                        f"{start_label} Done Count": g[start_col].notna().sum(),
                        f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].notna(), "days_diff"].mean()
                    }))
                    .reset_index()
                )
                
                total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                total_row["Circle"] = "Total"

                overall_valid = temp.loc[
                    temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].notna(),
                    "days_diff"
                ]
                total_row["Average Days"] = overall_valid.mean()

            else:
                if month_start is not None and month_end is not None:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": (g[start_col].notna() & (g[start_col] >= month_start) & (g[start_col] <= month_end)).sum(),
                            f"{end_label} Done Count": (g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"<= {breakpoint1} days": (g["days_diff"].notna() & (g["days_diff"] <= breakpoint1) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": (g["days_diff"].notna() & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f">= {breakpoint2} days": (g["days_diff"].notna() & (g["days_diff"] >= breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna() & (temp[end_col] >= month_start) & (temp[end_col] <= month_end),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                else:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": g[start_col].notna().sum(),
                            f"{end_label} Done Count": g[end_col].notna().sum(),
                            f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & g[end_col].notna()).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna()).sum(),
                            f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & g[end_col].notna()).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna(), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna(),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                    
            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]

            return summary


        done_summary = generate_done_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        done_summary = done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        def generate_pending_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            
            today = pd.Timestamp.today()
            
            temp = df.copy()
            
            temp["days_diff"] = (today - temp[start_col]).dt.days

            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                    f"<= {breakpoint1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[start_col].notna())).sum(),
                    f"{breakpoint1+1}-{breakpoint2-1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[start_col].notna())).sum(),
                    f">= {breakpoint2} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[start_col].notna())).sum(),
                    "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].isna(), "days_diff"].mean()
                }))
                .reset_index()
            )

            total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
            total_row["Circle"] = "Total"

            overall_valid = temp.loc[
                temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].isna(),
                "days_diff"
            ]
            total_row["Average Days"] = overall_valid.mean()

            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Pending Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]
            
            return summary
        

        pending_summary = generate_pending_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        pending_summary = pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"AGEING_DASHBOARD_FILE_{current_date}_{current_time}.xlsx")

        result_json = pending_summary.to_dict(orient="records")
        json_data2 = json.dumps(result_json)

        result_json = done_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "done_summary": json_data1,
            "pending_summary": json_data2
        }
        
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            pending_summary.to_excel(writer, index=False, sheet_name='PENDING SUMMARY')
            done_summary.to_excel(writer, index=False, sheet_name='DONE SUMMARY')
        
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
        
        return Response({'message': 'request processed successfully !!!', "download_link": download_link, "json_data": json_data, "unique_data": unique_data, "breakpoint1" : breakpoint1, "breakpoint2" : breakpoint2}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)


@api_view(['GET', 'POST'])
def lifecycle_display(request):
    userId = request.data.get('userId').lower()
    siteId = request.data.get('siteId')
    circle = request.data.get('circle')

    if not userId or not siteId or not circle:
        return Response({"error": "input not provided"}, status = 400)
    
    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    circles = user.circles
    # circles='CENTRAL'
    
    if 'CENTRAL' not in circles and circle not in circles:
        return Response({"error" : f"Access to {circle} sites denied"}, status = 403)
    
    try:
        filter = {
            "circle": circle,
            "site_id": siteId
        }

        obj = NTSiteTracker.objects.filter(**filter).first()
        
        if obj is None:
            return Response({"error": "Site not found!"}, status=500)
        else:
            df = pd.DataFrame([model_to_dict(obj)])

        filters = {
            "circle" : circle,
            "site_id" : siteId
        }
        
        issue_obj = NTIssue.objects.filter(**filters)
        issue_df = pd.DataFrame(issue_obj.values())

        required_cols = ["milestone", "status"]
        for col in required_cols:
            if col not in issue_df.columns:
                issue_df[col] = None
        
        data_cols = [
        "Allocation",
        "RFAI",
        "WRFAI",
        "RFAI Survey",
        "MO Punch",
        "Material Dispatch",
        "Material Delivered",
        "Installation End",
        "Integration",
        "EMF Submission",
        "Alarm Rectification Done",
        "SCFT Offered",
        "RAN PAT Offer",
        "RAN SAT Offer",
        "MW PAT Offer",
        "MW SAT Offer",
        "Site ONAIR",
        "I-Deploy ONAIR",
    ]
        
        data_cols_1 = [col.lower().replace("-", "_").replace(" ", "_").replace("/", "_") + "_date" for col in data_cols] + ['circle']
        
        df_data = df[data_cols_1]
        
        issue_counts = {}
        
        for col in data_cols:
            counts = {}
            counts['total'] = (issue_df['milestone'] == col).sum()
            counts['open'] = ((issue_df['milestone'] == col) & (issue_df['status'] == 'Open')).sum()
            counts['closed'] = ((issue_df['milestone'] == col) & (issue_df['status'] == 'Closed')).sum()
            issue_counts[col] = counts
        
        print("C")
        
        if not df_data.empty:
            df_data = df_data.applymap(
                lambda x: x.strftime("%d-%b-%y") if hasattr(x, 'strftime') else x
            )
            
        
        result_json = df_data.to_dict(orient="records")
        json_data = json.dumps(result_json)
        
        return Response({'message': 'request processed successfully !!!', "json_data" : json_data, "issue_counts" : issue_counts }, status = 200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
    

@api_view(['POST'])
def issue_timeline_display(request):
    userId = request.data.get('userId')
    if not userId:
        return Response({"error": "userId required"}, status=400)

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    
    # circles = "CENTRAL"
    circles = user.circles
    
    siteId = request.data.get('siteId')
    circle = request.data.get('circle')
    owner = request.data.get('owner')
    milestone = request.data.get('milestone')

    if not all([siteId, circle]):
        return Response({"message": "Circle/Site-ID not provided"}, status=400)

    if 'CENTRAL' not in circles and circle not in circles:
        return Response({"error": f"Access to {circle} sites denied!"}, status=403)

    try:
        filters = {
            "circle": circle,
            "site_id": siteId
        }
        if owner:
            filters['issue_owner'] = owner
        if milestone:
            filters['milestone'] = milestone

        obj = NTIssue.objects.filter(**filters)

        df = pd.DataFrame(obj.values())

        if not obj.exists():
            # required_cols = ["Circle", "Site ID", "Issue Owner", "Milestone", "Issue Name", "Start Date", "Close Date", "Status", "Duration", "Updated_by", "Updated_at", "Created_by", "Created_at"]
            # for col in required_cols:
            #     if col not in df.columns:
            #         df[col] = None
            json_data = df.to_dict(orient="records")
            json_data = json.dumps(json_data)
            return Response({"error": "Data not found!", "json_data": json_data}, status=200)

        df["start_date"] = df["start_date"].apply(
            lambda x: x.strftime("%d-%b-%y") if pd.notnull(x) else "-"
        )

        df["close_date"] = df["close_date"].apply(
            lambda x: x.strftime("%d-%b-%y") if pd.notnull(x) else "-"
        )
        
        rename_map = {
            "circle": "Circle",
            "site_id": "Site ID",
            "issue_owner": "Issue Owner",
            "milestone": "Milestone",
            "issue_name": "Issue Name",
            "start_date": "Start Date",
            "close_date": "Close Date",
            "status": "Status",
            "duration": "Duration",
            "remarks": "Remarks",
            "updated_by": "Updated_by",
            "updated_at": "Updated_at",
            "created_by": "Created_by",
            "created_at": "Created_at"
        }
        
        df = df.rename(columns=rename_map)

        json_data = df.to_dict(orient="records")

        return Response({"message": "Request processed successfully!", "json_data": json_data}, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
 
@api_view(['POST'])
def issue_timeline_add(request):
    userId = request.data.get('userId')
    if not userId:
        return Response({"error": "userId required"}, status=400)

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    if user.right == 'Read':
        return Response({"error": "Access denied!"}, status=status.HTTP_403_FORBIDDEN)

    siteId = request.data.get('siteId')
    circle = request.data.get('circle')
    owner = request.data.get('owner')
    milestone = request.data.get('milestone')
    issue = request.data.get('issue')
    start_date_str = request.data.get('start_date')   
    close_date_str = request.data.get('close_date')  
    remarks = request.data.get('remarks') 

    required_fields = [siteId, circle, owner, milestone, issue, start_date_str]
    if not all(required_fields):
        return Response({"message": "Required data not provided"}, status=status.HTTP_400_BAD_REQUEST)
    
        

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()

        close_date = None
        if close_date_str not in ["", None, "-", "nan", "undefined"]:
            close_date = datetime.strptime(close_date_str, "%Y-%m-%d").date()

    except Exception:
        return Response({"error": "Invalid date format. Expected yyyy-mm-dd"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        filters = {
            "circle": circle,
            "site_id": siteId,
            "issue_owner": owner,
            "milestone": milestone,
            "issue_name": issue,
            "status": "Open"
        }

        if NTIssue.objects.filter(**filters).exists():
            return Response({"error": "Open issue already exists"}, status=409)

        today = date.today()

        if close_date:
            duration = (close_date - start_date).days
        else:
            duration = (today - start_date).days

        new_issue = NTIssue.objects.create(
            circle=circle,
            site_id=siteId,
            issue_owner=owner,
            milestone=milestone,
            issue_name=issue,
            start_date=start_date,
            close_date=close_date,
            status="Closed" if close_date else "Open",
            duration=duration,
            remarks=remarks,
            updated_by=userId,
            created_by=userId,
        )
        
        update_ageing_new(circle, siteId)

        return Response(
            {
                "message": "Issue added successfully!",
                "json_data":{"id": new_issue.id,
                "circle": circle,
                "site": siteId,
                "milestone": milestone,
                "issue": issue,
                "start_date": start_date.strftime("%d-%b-%y"),
                "close_date": close_date.strftime("%d-%b-%y") if close_date else "-",
                "status": new_issue.status,
                "duration_days": duration,
                "remarks": remarks}
            },
            status=200
        )

    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
def issue_timeline_update(request):
    userId = request.data.get('userId')
    issue_id = request.data.get('id') 
    circle = request.data.get('circle')

    if not userId or not issue_id or not circle:
        return Response({"error": "Required data not provided"}, status=400)

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    if user.right == 'Read':
        return Response({"error": "Access denied!"}, status=status.HTTP_403_FORBIDDEN)

    # user_circle = ACCESS_RIGHTS[userId]["Circle"]
    # if user_circle != "CENTRAL" and user_circle != circle:
    #     return Response({"error": "Access denied!"}, status=403)

    try:
        issue_obj = NTIssue.objects.get(id=issue_id)
    except NTIssue.DoesNotExist:
        return Response({"error": "Issue not found"}, status=404)
    
    if issue_obj.status == 'Closed':
        return Response({"error": "Issue already closed"}, status=404)

    start_date_str = request.data.get("start_date")
    close_date_str = request.data.get("close_date")
    new_owner = request.data.get("owner")
    new_milestone = request.data.get("milestone")
    new_issue_name = request.data.get("issue")
    remarks = request.data.get('remarks')

    try:
        if start_date_str:
            issue_obj.start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()

        if close_date_str in ["", None, "-", "nan", "undefined"]:
            issue_obj.close_date = None
        else:
            issue_obj.close_date = datetime.strptime(close_date_str, "%Y-%m-%d").date()

    except Exception:
        return Response({"error": "Invalid date format. Expected yyyy-mm-dd"}, status=400)

    try:
        
        if new_owner:
            issue_obj.issue_owner = new_owner

        if new_milestone:
            issue_obj.milestone = new_milestone

        if new_issue_name:
            issue_obj.issue_name = new_issue_name
            
        if remarks:
            issue_obj.remarks = remarks

        if issue_obj.close_date:
            issue_obj.status = "Closed"
        else:
            issue_obj.status = "Open"

        today = date.today()

        if issue_obj.close_date:
            issue_obj.duration = (issue_obj.close_date - issue_obj.start_date).days
        else:
            issue_obj.duration = (today - issue_obj.start_date).days

        issue_obj.updated_by = userId

        issue_obj.save()

        out_start = issue_obj.start_date.strftime("%d-%b-%y")
        out_close = issue_obj.close_date.strftime("%d-%b-%y") if issue_obj.close_date else "-"

        response_data = {
            "id": issue_obj.id,
            "circle": issue_obj.circle,
            "site_id": issue_obj.site_id,
            "issue_owner": issue_obj.issue_owner,
            "milestone": issue_obj.milestone,
            "issue_name": issue_obj.issue_name,
            "start_date": out_start,
            "close_date": out_close,
            "status": issue_obj.status,
            "duration": issue_obj.duration,
            "remarks": remarks,
            "updated_by": userId
        }
        
        update_ageing_new(circle, issue_obj.site_id)

        return Response(
            {"message": "Issue updated successfully!", "json_data": response_data},
            status=200
        )
        
    except Exception as e:
        return Response({"Error: " : f"{str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def issue_timeline_delete(request):
    userId = request.data.get('userId')
    issue_id = request.data.get('id')
    circle = request.data.get('circle')

    if not userId or not issue_id or not circle:
        return Response(
            {"error": "userId, id and circle are required"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    if user.right == 'Read':
        return Response({"error": "Access denied!"}, status=status.HTTP_403_FORBIDDEN)

    # Access control
    # user_circle = ACCESS_RIGHTS[userId]["Circle"]
    # if user_circle != "CENTRAL" and user_circle != circle:
    #     return Response({"error": "Access denied!"}, status=status.HTTP_403_FORBIDDEN)

    try:
        issue_obj = NTIssue.objects.get(id=issue_id)
    except NTIssue.DoesNotExist:
        return Response({"error": "Issue not found"}, status=status.HTTP_404_NOT_FOUND)

    # Optional safety check (recommended)
    # if issue_obj.status == "Closed":
    #     return Response(
    #         {"error": "Closed issues cannot be deleted"},
    #         status=status.HTTP_400_BAD_REQUEST
    #     )
    
    update_ageing_new(circle, issue_obj.site_id)

    issue_obj.delete()

    return Response(
        {
            "message": "Issue deleted successfully!",
            "id": issue_id
        },
        status=status.HTTP_200_OK
    )


@api_view(['GET', 'POST'])
def graphs_view(request):
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    # all_unique_site_tagging = list(
    #     NTSiteTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        NTSiteTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Allocation",
        "RFAI",
        "WRFAI",
        "RFAI Survey",
        "MO Punch",
        "Material Dispatch",
        "Material Delivered",
        "Installation End",
        "Integration",
        "EMF Submission",
        "Alarm Rectification Done",
        "SCFT Offered",
        "RAN PAT Offer",
        "RAN SAT Offer",
        "MW PAT Offer",
        "MW SAT Offer",
        "Site ONAIR",
        "I-Deploy ONAIR",
    ]
    
    unique_data = {
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]
    }

    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        # if site_tagging and "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if month_start and month_end:
            filter_col = milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '') + '_date'
            
            if milestone1 == "WRFAI":
                filter_col = "wrfai_date"

            filters[f"{filter_col}__range"] = (month_start, month_end)

        obj = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        print(df)
        
        df['Circle'] = df['circle']

        def generate_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"{end_label} Pending Count"]
            ]

            return summary

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        start_col = start_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_") + "_date"
        end_col = end_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_") + "_date"

        if start_label == "WRFAI":
            start_col = "wrfai_date"

        if end_label == "WRFAI":
            end_col = "wrfai_date"

            
        graph_summary = generate_summary(df, start_label, end_label, start_col, end_col)

        graph_summary = graph_summary.applymap(lambda x: str(x) if pd.notna(x) else "")   

        result_json = graph_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "graph_summary": json_data1,
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
 

@api_view(['GET', 'POST'])
def monthly_graph(request):
    circle = request.data.get('circle', [])
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    view = request.data.get('view')
    type = request.data.get('type')
    year_filtered = request.data.get('years')
    
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    milestone1_col = (
        "wrfai_date"
        if milestone1 == "WRFAI"
        else milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"
    )

    milestone2_col = (
        "wrfai_date"
        if milestone2 == "WRFAI"
        else milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"
    )
    
    all_unique_circles = list(
        NTSiteTracker.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    # all_unique_site_tagging = list(
    #     NTSiteTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        NTSiteTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        NTSiteTracker.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )
    
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }
    
    try:
        year_filtered = int(year_filtered)

        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name
        
            
            
        obj = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())

        
        # print(df)
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        df = df[[milestone1_col, milestone2_col]]
       
        for col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        
        today = dtime.today().date()
        
        df["month_num1"] = (
            df[milestone1_col].dt.month + (df[milestone1_col].dt.day >= 26)
        ).replace({13: 1}).astype("Int64")

        df["year1"] = (
            df[milestone1_col].dt.year - (
                (df[milestone1_col].dt.month < 3) |
                ((df[milestone1_col].dt.month == 3) & (df[milestone1_col].dt.day < 26))
            )
        ).astype("Int64")

        # if year_filtered:
        #     df = df[df["year1"] == year_filtered]


        # 3️⃣ Display year
        df["display_year"] = df["year1"] + (df["month_num1"] <= 3)

        # 4️⃣ Month label
        df["month_name1"] = pd.to_datetime(
            df["month_num1"].astype(str) + "-" + df["display_year"].astype(str),
            format="%m-%Y",
            errors="coerce"
        ).dt.strftime("%b-%y")

        # 5️⃣ Current FY
        current_fy = year_filtered

# CF
        df.loc[
            df["year1"].notna() & (df["year1"] < current_fy),
            "month_name1"
        ] = "CF"

        # Future
        df.loc[
            df["year1"].notna() & (df["year1"] > current_fy),
            "month_name1"
        ] = "Future"


        df['month_num2'] = df[milestone2_col].dt.month + (df[milestone2_col].dt.day >= 26)
        df['month_num2'] = df['month_num2'].replace({13: 1})

        # 2️⃣ Financial year (derived ONLY from original date)
        df['year2'] = df[milestone2_col].dt.year - (
            (df[milestone2_col].dt.month < 3) |
            ((df[milestone2_col].dt.month == 3) & (df[milestone2_col].dt.day < 26))
        )

        # 3️⃣ Safe numeric conversion (IMPORTANT)
        df['month_num2'] = pd.to_numeric(df['month_num2'], errors='coerce')
        df['year2'] = pd.to_numeric(df['year2'], errors='coerce')

        # 4️⃣ Display year
        df['display_year2'] = df['year2'] + (df['month_num2'] <= 3)

        # 5️⃣ Month label (NaT-safe)
        df['month_name2'] = pd.to_datetime(
            df['month_num2'].astype('Int64').astype(str) + '-' +
            df['display_year2'].astype('Int64').astype(str),
            format='%m-%Y',
            errors='coerce'
        ).dt.strftime('%b-%y')

        # 6️⃣ Current financial year
        current_fy = today.year if today >= date(today.year, 3, 26) else today.year - 1

        # 7️⃣ Carry Forward (only where year2 exists)
        df.loc[df['year2'].notna() & (df['year2'] < current_fy), 'month_name2'] = 'CF'
        
        print(df['month_name1'].dropna().unique().tolist())

        
        print(df['year1'])
        
        print(df['month_name2'])
        
        print(df['year2'])

        def sort_financial_year(summary):
            # CF aur Future hata ke sirf actual month values lo
            months = [
                m for m in summary['month_name'].unique()
                if m not in ["CF", "Future"] and pd.notna(m)
            ]

            # Sirf real months sort honge
            months_sorted = sorted(
                months,
                key=lambda x: pd.to_datetime(x, format='%b-%y')
            )

            ordered = []

            # CF hamesha first
            if "CF" in summary['month_name'].values:
                ordered.append("CF")

            # Actual months
            ordered += months_sorted

            # Future hamesha last
            if "Future" in summary['month_name'].values:
                ordered.append("Future")

            summary['month_name'] = pd.Categorical(
                summary['month_name'],
                categories=ordered,
                ordered=True
            )

            return summary.sort_values('month_name')
        
        def generate_type1_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()

            start_counts = (
                temp[temp[start_col].notna()]
                .groupby("month_name1")
                .size()
                .reset_index(name=f"{start_label} Done Count")
                .rename(columns={"month_name1": "month_name"})
            )

            end_counts = (
                temp[temp[end_col].notna()]
                .groupby("month_name2")
                .size()
                .reset_index(name=f"{end_label} Done Count")
                .rename(columns={"month_name2": "month_name"})
            )

            summary = pd.merge(start_counts, end_counts, on="month_name", how="outer")

            # Convert to Int64 nullable type
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")
            summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].astype("Int64")

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            else:
                # Only for RFAI → add CF into last visible month
                if start_label == "RFAI" and "CF" in summary["month_name"].values:

                    cf_rfai = summary.loc[
                        summary["month_name"] == "CF",
                        f"{start_label} Done Count"
                    ].values[0]

                    # last month (May)
                    last_idx = summary[summary["month_name"] != "CF"].index.max()

                    summary.loc[
                        last_idx,
                        f"{start_label} Done Count"
                    ] += cf_rfai

            return summary.reset_index(drop=True)
        
        def generate_type2_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("month_name1").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    # f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
                .rename(columns={"month_name1": "month_name"})
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")

            summary = summary[
                ["month_name", f"{start_label} Done Count", f"{end_label} Done Count"]
            ]

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            else:
                if "CF" in summary["month_name"].values:

                    cf_row = summary[summary["month_name"] == "CF"]

                    # CF value nikal lo
                    cf_rfai = 0
                    if start_label == "RFAI":
                        cf_rfai = cf_row[f"{start_label} Done Count"].iloc[0]

                    # Pehle CF row hatao
                    summary = summary[summary["month_name"] != "CF"].copy()

                    # Fir last visible month me add karo
                    if start_label == "RFAI" and len(summary) > 0:
                        summary.iloc[-1, summary.columns.get_loc(f"{start_label} Done Count")] += cf_rfai

            return summary.reset_index(drop=True)
        
        summary = generate_type1_summary(df, milestone1, milestone2, milestone1_col, milestone2_col) if type == 'type1' else generate_type2_summary(df, milestone1, milestone2, milestone1_col, milestone2_col)

        numeric_cols = [col for col in summary.columns if "Done Count" in col]

        summary[numeric_cols] = summary[numeric_cols].apply(pd.to_numeric, errors="coerce")
        summary[numeric_cols] = summary[numeric_cols].fillna(0).astype(int)

        
        result_json = summary.to_dict(orient = "records")
        json_data = json.dumps(result_json)
        
        return Response({"message" : "request processed successfully ! ", "json_data": json_data, "unique_data": unique_data}, status=200)
    
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
    


@api_view(['GET', 'POST'])
def ms2_daily_waterfall(request):
    circle = request.data.get('circle', [])
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    start_date = request.data.get('from_date')
    end_date = request.data.get('to_date')
    view = request.data.get('view')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None

    all_unique_circles = list(
        NTSiteTracker.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    # all_unique_site_tagging = list(
    #     NTSiteTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        NTSiteTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        NTSiteTracker.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )

    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]

    start_date = dtime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    end_date = dtime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

    try:
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name

        obj = NTSiteTracker.objects.filter(**filters)
        df = pd.DataFrame(obj.values())

        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        # df['rfai_date'] = df['re_rfai_date'].where(df["re_rfai_date"].notna(), df['rfai_date'])

        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")

        today = dtime.today().date()

        if (not month_start or not month_end) and (not start_date or not end_date):
            if today.day >= 26:
                start_date = today.replace(day=26)
                if today.month == 12:
                    end_date = today.replace(year=today.year + 1, month=1, day=25)
                else:
                    end_date = today.replace(month=today.month + 1, day=25)
            else:
                if today.month == 1:
                    start_date = today.replace(year=today.year - 1, month=12, day=26)
                else:
                    start_date = today.replace(month=today.month - 1, day=26)
                end_date = today.replace(day=25)
                
        if month_start and month_end:
            date_range = pd.date_range(start=month_start, end=min(month_end, today)).date
        else:
            date_range = pd.date_range(start=start_date, end=min(end_date, today)).date

        formatted_dates = [d.strftime("%d-%b-%y") for d in date_range]

        if month_start and month_end:
            result_columns = ["Milestone Track/Site Count", "AOP", "CF"] + formatted_dates
        else:
            result_columns = ["Milestone Track/Site Count", "AOP"] + formatted_dates
        result = pd.DataFrame(columns=result_columns)

        milestones = [
            "Site ONAIR Date",
            "RAN PAT Accepted Date",
            "RAN SAT Accepted Date",
            "MW PAT Accepted Date",
            "MW SAT Accepted Date",
            "SCFT Accepted Date",

            "KPI AT offer Date",
            "KPI AT Accepted Date",

            "4G MS2 Date",
            # "5G MS2 Date",
            # "Final MS2 Date",
        ]

        unique_data.update(**{"Milestone": milestones})

        for milestone in milestones:
            milestone_df_format = (
                milestone.lower()
                .replace(" ", "_")
                .replace("-", "_")
                .replace("(", "")
                .replace(")", "")
            )

            if milestone_df_format not in df.columns:
                continue

            df[milestone_df_format] = pd.to_datetime(df[milestone_df_format], errors="coerce").dt.date
            if month_start and month_end:
                mask = df['site_onair_date'].isna()
                valid_dates_cf = df.loc[mask, milestone_df_format].dropna()
            # else:
            #     valid_dates = df[milestone_df_format].dropna()
            
            valid_dates_aop = df[milestone_df_format].dropna()

            if valid_dates_aop.empty:
                if month_start and month_end:
                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": "-",
                        "CF": "-",
                        **{d.strftime("%d-%b-%y"): "-" for d in date_range}
                    }
                else:
                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": "-",
                        **{d.strftime("%d-%b-%y"): "-" for d in date_range}
                    }
                
                result.loc[len(result)] = row
                continue
            
            if month_start and month_end:
                cf_count = (valid_dates_cf < month_start).sum()
                aop_count = (valid_dates_aop < month_start).sum()
            else:
            #     cf_count = (valid_dates < start_date).sum()
                aop_count = (valid_dates_aop < start_date).sum()
            
            if month_start and month_end:
                cumulative = cf_count
                row = {"Milestone Track/Site Count": milestone,"AOP": aop_count, "CF": cf_count}
            else:
                cumulative = aop_count
                row = {"Milestone Track/Site Count": milestone,"AOP": aop_count}


            for d in date_range:
                count = (valid_dates_aop == d).sum()
                cumulative += count

                if view == "Cumulative":
                    row[d.strftime("%d-%b-%y")] = cumulative
                else:
                    row[d.strftime("%d-%b-%y")] = count

            result.loc[len(result)] = row

        result.columns = [
            col.strftime("%d-%b-%y") if isinstance(col, (dtime,)) or hasattr(col, "strftime") else col
            for col in result.columns
        ]
        if month_start and month_end:
            result = result[["Milestone Track/Site Count", "AOP" , "CF"] + formatted_dates]
        else:
            result = result[["Milestone Track/Site Count", "AOP"] + formatted_dates]
 
        last_col = formatted_dates[-1]
        dash_mask = result[last_col] == '-'

        result[last_col] = pd.to_numeric(result[last_col], errors='coerce')
        result['Gap'] = -result[last_col].diff()

        nan_mask = result['Gap'].isna()

        result[last_col] = result[last_col].fillna(0).astype(int)
        result['Gap'] = result['Gap'].fillna(0).astype(int)

        result.loc[dash_mask, last_col] = '-'
        result.loc[nan_mask, 'Gap'] = '-'

        result['Gap'] = result['Gap'].astype(str)
        result = result.astype(str).reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(
            lambda col: col.replace(" Date", "") if " Date" in col else col
        )

        new_result = result.copy()

        for i, d in enumerate(formatted_dates, start=1):
            new_result.rename(columns={d: f'date_{i}'}, inplace=True)

        result_json = new_result.to_dict(orient="records")
        json_data = json.dumps(result_json)

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE_URL = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE_URL, exist_ok=True)

        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)

        dashboard_file_path = os.path.join(
            output_folder,
            f"NT_TRACKING_DAILY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx"
        )

        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Daily Waterfall MS1')

        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)

        return Response({
            'message': 'Dashboard created successfully !!!',
            "download_link": download_link,
            "data": json_data,
            "dates": formatted_dates,
            "unique_data": unique_data
        }, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)


@api_view(['GET', 'POST'])
def ms2_weekly_monthly_waterfall(request):
    circle = request.data.get('circle', [])
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    view = request.data.get('view')
 
    # Default to 'ALL' if not provided
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None
   
    all_unique_circles = list(
        NTSiteTracker.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    # all_unique_site_tagging = list(
    #     NTSiteTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        NTSiteTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        NTSiteTracker.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )

    # ✅ Add "ALL" at the top of each list
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    try:
        # 🔹 Dynamic filters
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name
 
        # 🔹 Fetch data
        obj = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        # df['rfai_date'] = df['re_rfai_date'].where(df["re_rfai_date"].notna(), df['rfai_date'])
       
        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")
               
        today = dtime.today().date()
 
        if today.month >= 4:
            fy_start = dtime(today.year, 3, 26).date()   # 26-Apr current year start
            fy_end = dtime(today.year + 1, 3, 25).date() # 25-Mar next year end
        else:
            fy_start = dtime(today.year - 1, 3, 26).date()
            fy_end = dtime(today.year, 3, 25).date()
 
        # 🧩 2. Determine current cycle (26th prev → 25th curr)
        if today.day >= 26:
            if today.month == 12:
                current_cycle_start = dtime(today.year, 12, 26).date()
                current_cycle_end = dtime(today.year + 1, 1, 25).date()
            else:
                current_cycle_start = dtime(today.year, today.month, 26).date()
                current_cycle_end = dtime(today.year, today.month + 1, 25).date()
        else:
            if today.month == 1:
                current_cycle_start = dtime(today.year - 1, 12, 26).date()
            else:
                current_cycle_start = dtime(today.year, today.month - 1, 26).date()
            current_cycle_end = dtime(today.year, today.month, 25).date()
 
        # 🗓️ 3. Create monthly periods (Apr → month before current)
        months = []
        start = fy_start
        while start <= current_cycle_start:
            if start.month == 12:
                end = dtime(start.year + 1, 1, 25).date()
            else:
                end = dtime(start.year, start.month + 1, 25).date()
            months.append((start, end))
            start = end + timedelta(days=1)
 
        # 🗓️ 4. Create weekly periods for current month
        weeks = []
        
        if month_start and month_end:
            current_cycle_start = month_start
            current_cycle_end = month_end
        
        week_start = current_cycle_start
 
        while week_start <= current_cycle_end:
            week_end = week_start + timedelta(days=6)
            if week_end > current_cycle_end:
                week_end = current_cycle_end
            weeks.append((week_start, week_end))
            week_start = week_end + timedelta(days=1)
 
 
        # 📊 5. Prepare result DataFrame
        result = pd.DataFrame()
       
       
        # result.set_index("Milestone Track/Site Count", inplace=True)
 
        milestones = [
            "Site ONAIR Date",
            "RAN PAT Accepted Date",
            "RAN SAT Accepted Date",
            "MW PAT Accepted Date",
            "MW SAT Accepted Date",
            "SCFT Accepted Date",

            "KPI AT offer Date",
            "KPI AT Accepted Date",

            "4G MS2 Date",
            # "5G MS2 Date",
            # "Final MS2 Date",
        ]
 
        data = []
        
        for milestone in milestones:
            
            milestone_df_format = milestone.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")

            if milestone_df_format not in df.columns:
                continue
            
            milestone_data = pd.to_datetime(df[milestone_df_format], errors='coerce').dt.date
            row = {"Milestone Track/Site Count": milestone}
            if milestone_data.dropna().empty:
                row["CF"] = "-"
                # Set all month and week columns to "-"
                for _, end in months:
                    month_name = end.strftime("%b-%y")
                    row[month_name] = "-"
                for i, _ in enumerate(weeks, 1):
                    week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                    row[week_name] = "-"
                data.append(row)
                continue
 
 
            # CF → sites before financial year start
            row["CF"] = (milestone_data < fy_start).sum()
 
            cumulative_month = row["CF"]
            
            for start, end in months:
                month_name = end.strftime("%b-%y")
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_month += count
                if view == 'Cumulative':
                    row[month_name] = cumulative_month
                else:
                    row[month_name] = count
 
            cumulative_week = 0
            for i, (start, end) in enumerate(weeks, 1):
                week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                if(start >= today):
                    row[week_name] = 0
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_week += count
                if view == 'Cumulative':
                    row[week_name] = cumulative_week
                else:
                    row[week_name] = count
 
            data.append(row)
 

        result = pd.DataFrame(data)
 
        for col in result.columns:
            if col != "Milestone Track/Site Count":
                result[col] = result[col].astype(str)
 
        result = result.reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(lambda col: col.replace(" Date", "") if " Date" in col else col)
 
        new_data = result.copy()
        months_columns = [col for col in new_data.columns.to_list()[2:] if " W" not in col]
        months_data = new_data[new_data.columns.to_list()[:2] + months_columns].copy()
        
        weeks_columns = [col for col in new_data.columns.to_list()[2:] if " W" in col]
        
        week_data = new_data[new_data.columns.to_list()[:1] + weeks_columns].copy()
        
        months_data.rename(columns={
            col : f"Month-{i}" for i, col in enumerate(months_columns, start=1)
        }, inplace=True)
        
        week_data.rename(columns={
            col : f"Month_Week-{i}" for i, col in enumerate(weeks_columns, start=1)
        }, inplace=True)
        
        unique_data.update(**{"month_columns": months_columns, "week_columns" : weeks_columns})
        

        month_dict_data = months_data.to_dict(orient="records")
        month_json_data = json.dumps(month_dict_data)
        week_dict_data = week_data.to_dict(orient="records")
        week_json_data = json.dumps(week_dict_data)
        
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"NT_TRACKING_WEEKLY_MONTHY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx")
       
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Monthly MS1')
       
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
   
        return Response({'message': 'Weekly and Monthly Dashboard created successfully !!!', "download_link": download_link, "unique_data": unique_data, "months_data": month_json_data, "week_data": week_json_data}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
   


@api_view(['GET', 'POST'])
def ms2_ageing_dashboard_table1(request):
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    # all_unique_site_tagging = list(
    #     NTSiteTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        NTSiteTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "RAN PAT Accepted",
        "RAN SAT Accepted",
        "MW PAT Accepted",
        "MW SAT Accepted",
        "SCFT Accepted",

        "KPI AT offer",
        "KPI AT Accepted",

        "4G MS2",
        # "5G MS2",
        # "Final MS2",
    ]
    
    unique_data = {
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones
    }

    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        # if site_tagging and "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if month_start and month_end:
            filters[f"{milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + '_date'}__range"] = (month_start, month_end)

        obj = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        date_cols = [
            col.lower()
            .replace(' ', '_')
            .replace('-', '_')
            .replace('(', '')
            .replace(')', '')
            .replace('/', '_')
            .replace("4", "four_")
            .replace("5", "five_") + '_date'
                    for col in milestones
                ]

        # Convert only existing columns, create empty if missing
        for col in date_cols:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")
            else:
                df[col] = pd.NaT

        df['Circle'] = df['circle']
        
        
        def generate_table_summary(df, milestone1, milestone2):


            milestone_cols = [(col.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date") for col in milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]]

            temp = df.copy()

            circles = ["ALL"] + temp["circle"].unique().tolist()

            final_df = pd.DataFrame()

            milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
            milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"

            for circle in circles:
                if circle != "ALL":
                    temp  = df[df["circle"] == circle]

                done_row = {}
                pending_row = {}

                for col, label in zip(milestone_cols, milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]):

                    done_row[label] = (temp[col].notna() & temp[milestone1_col].notna() ).sum()

                    pending_row[label] = (temp[col].isna() & temp[milestone1_col].notna() ).sum()

                rows_to_add = pd.DataFrame([
                    {"Circle": circle, "Site Status": "Done", **done_row},
                    {"Circle": circle, "Site Status": "Pending", **pending_row},
                ])

                final_df = pd.concat([final_df, rows_to_add], ignore_index=True)

            for c in final_df.columns:
                if c != "Site Status" and c != "Circle":
                    final_df[c] = final_df[c].astype("Int64")

            return final_df

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        table_summary = generate_table_summary(df, start_label, end_label)
 
        table_summary = table_summary.applymap(lambda x: str(x) if pd.notna(x) else "")      


        result_json = table_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "table_summary": json_data1
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
  
  
@api_view(['GET', 'POST'])
def ms2_ageing_dashboard_table2(request):
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    breakpoint1 = request.data.get('breakpoint1')
    breakpoint2 = request.data.get('breakpoint2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    type = request.data.get('type')

    # all_unique_site_tagging = list(
    #     NTSiteTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        NTSiteTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "RAN PAT Accepted",
        "RAN SAT Accepted",
        "MW PAT Accepted",
        "MW SAT Accepted",
        "SCFT Accepted",

        "KPI AT offer",
        "KPI AT Accepted",

        "4G MS2",
        # "5G MS2",
        # "Final MS2",
    ]
    
    unique_data = {
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones" : milestones
    }
    

    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    
    try:

        milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
        milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"

        breakpoint1 = int(breakpoint1)
        breakpoint2 = int(breakpoint2)
        
        temp = breakpoint1
        breakpoint1 = min(breakpoint1, breakpoint2)
        breakpoint2 = max(temp, breakpoint2)

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None

        milestones = [
            "Site ONAIR",
            "RAN PAT Accepted",
            "RAN SAT Accepted",
            "MW PAT Accepted",
            "MW SAT Accepted",
            "SCFT Accepted",

            "KPI AT offer",
            "KPI AT Accepted",

            "4G MS2",
            # "5G MS2",
            # "Final MS2",
        ]
        

        filters = {}
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
            
        if type == "type2":
            if month_start and month_end:
                filters[f"{milestone1_col}__range"] = (month_start, month_end)
                

        obj = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        
        if type == "type1":
            if month_start and month_end:
                obj = obj.filter(
                    Q(**{f"{milestone1_col}__range": (month_start, month_end)}) |
                    Q(**{f"{milestone2_col}__range": (month_start, month_end)})
                )
        

        df = pd.DataFrame(obj.values())

        
        date_cols = [
            col.lower()
            .replace(' ', '_')
            .replace('-', '_')
            .replace('(', '')
            .replace(')', '')
            .replace('/', '_')
            .replace("4", "four_")
            .replace("5", "five_") + '_date'
            for col in milestones
        ]

        # Convert only existing columns, create empty if missing
        for col in date_cols:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")
            else:
                df[col] = pd.NaT

        df['Circle'] = df['circle']

        if month_start is not None:
            month_start = pd.to_datetime(month_start)

        if month_end is not None:
            month_end = pd.to_datetime(month_end)

        def generate_done_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            temp = df.copy()
            
            temp["days_diff"] = (temp[end_col] - temp[start_col]).dt.days
            
            if type == "type2":
        
                summary = (
                    temp.groupby("Circle").apply(lambda g: pd.Series({
                        f"{start_label} Done Count": g[start_col].notna().sum(),
                        f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].notna(), "days_diff"].mean()
                    }))
                    .reset_index()
                )
                
                total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                total_row["Circle"] = "Total"

                overall_valid = temp.loc[
                    temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].notna(),
                    "days_diff"
                ]
                total_row["Average Days"] = overall_valid.mean()

            else:
                if month_start is not None and month_end is not None:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": (g[start_col].notna() & (g[start_col] >= month_start) & (g[start_col] <= month_end)).sum(),
                            f"{end_label} Done Count": (g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"<= {breakpoint1} days": (g["days_diff"].notna() & (g["days_diff"] <= breakpoint1) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": (g["days_diff"].notna() & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f">= {breakpoint2} days": (g["days_diff"].notna() & (g["days_diff"] >= breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna() & (temp[end_col] >= month_start) & (temp[end_col] <= month_end),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                else:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": g[start_col].notna().sum(),
                            f"{end_label} Done Count": g[end_col].notna().sum(),
                            f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & g[end_col].notna()).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna()).sum(),
                            f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & g[end_col].notna()).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna(), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna(),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                    
            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]

            return summary


        done_summary = generate_done_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        done_summary = done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        def generate_pending_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            
            today = pd.Timestamp.today()
            
            temp = df.copy()
            
            temp["days_diff"] = (today - temp[start_col]).dt.days

            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                    f"<= {breakpoint1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[start_col].notna())).sum(),
                    f"{breakpoint1+1}-{breakpoint2-1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[start_col].notna())).sum(),
                    f">= {breakpoint2} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[start_col].notna())).sum(),
                    "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].isna(), "days_diff"].mean()
                }))
                .reset_index()
            )

            total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
            total_row["Circle"] = "Total"

            overall_valid = temp.loc[
                temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].isna(),
                "days_diff"
            ]
            total_row["Average Days"] = overall_valid.mean()

            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Pending Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]
            
            return summary
        

        pending_summary = generate_pending_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        pending_summary = pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"AGEING_DASHBOARD_FILE_{current_date}_{current_time}.xlsx")

        result_json = pending_summary.to_dict(orient="records")
        json_data2 = json.dumps(result_json)

        result_json = done_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "done_summary": json_data1,
            "pending_summary": json_data2
        }
        
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            pending_summary.to_excel(writer, index=False, sheet_name='PENDING SUMMARY')
            done_summary.to_excel(writer, index=False, sheet_name='DONE SUMMARY')
        
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
        
        return Response({'message': 'request processed successfully !!!', "download_link": download_link, "json_data": json_data, "unique_data": unique_data, "breakpoint1" : breakpoint1, "breakpoint2" : breakpoint2}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
    


@api_view(['GET', 'POST'])
def ms2_graphs_view(request):
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    # all_unique_site_tagging = list(
    #     NTSiteTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        NTSiteTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "RAN PAT Accepted",
        "RAN SAT Accepted",
        "MW PAT Accepted",
        "MW SAT Accepted",
        "SCFT Accepted",

        "KPI AT offer",
        "KPI AT Accepted",

        "4G MS2",
        # "5G MS2",
        # "Final MS2",
    ]
    
    unique_data = {
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones
    }

    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        # if site_tagging and "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if month_start and month_end:
            filters[f"{milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_")+ '_date'}__range"] = (month_start, month_end)

        obj = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        df['Circle'] = df['circle']

        def generate_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"{end_label} Pending Count"]
            ]

            return summary

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        start_col = start_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_").replace("4", "four_") + "_date"
        end_col = end_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_").replace("4", "four_") + "_date"
        
        graph_summary = generate_summary(df, start_label, end_label, start_col, end_col)

        graph_summary = graph_summary.applymap(lambda x: str(x) if pd.notna(x) else "")   

        result_json = graph_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "graph_summary": json_data1,
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
 

@api_view(['GET', 'POST'])
def ms2_monthly_graph(request):
    circle = request.data.get('circle', [])
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    view = request.data.get('view')
    type = request.data.get('type')
    year_filtered = request.data.get('years')
   
    
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_") + "_date"
    milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_") + "_date"
    
    all_unique_circles = list(
        NTSiteTracker.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    # all_unique_site_tagging = list(
    #     NTSiteTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        NTSiteTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        NTSiteTracker.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "RAN PAT Accepted",
        "RAN SAT Accepted",
        "MW PAT Accepted",
        "MW SAT Accepted",
        "SCFT Accepted",

        "KPI AT offer",
        "KPI AT Accepted",

        "4G MS2",
    #     "5G MS2",
    #     "Final MS2",
    ]
    
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
        "milestones": milestones
    }
    
    try:
        year_filtered = int(year_filtered)

        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name
        
            
            
        obj = NTSiteTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        # print(df)
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        df = df[[milestone1_col, milestone2_col]]
       
        for col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        
        today = dtime.today().date()
        
        df["month_num1"] = (
            df[milestone1_col].dt.month + (df[milestone1_col].dt.day >= 26)
        ).replace({13: 1}).astype("Int64")

        df["year1"] = (
            df[milestone1_col].dt.year - (
                (df[milestone1_col].dt.month < 3) |
                ((df[milestone1_col].dt.month == 3) & (df[milestone1_col].dt.day < 26))
            )
        ).astype("Int64")

        # 3️⃣ Display year
        df["display_year"] = df["year1"] + (df["month_num1"] <= 3)

        # 4️⃣ Month label
        df["month_name1"] = pd.to_datetime(
            df["month_num1"].astype(str) + "-" + df["display_year"].astype(str),
            format="%m-%Y",
            errors="coerce"
        ).dt.strftime("%b-%y")

        # 5️⃣ Current FY
        current_fy = year_filtered

        # 6️⃣ CF logic
        df.loc[
            df["year1"].notna() & (df["year1"] < current_fy),
            "month_name1"
        ] = "CF"

        # Future
        df.loc[
            df["year1"].notna() & (df["year1"] > current_fy),
            "month_name1"
        ] = "Future"

        df['month_num2'] = df[milestone2_col].dt.month + (df[milestone2_col].dt.day >= 26)
        df['month_num2'] = df['month_num2'].replace({13: 1})

        # 2️⃣ Financial year (derived ONLY from original date)
        df['year2'] = df[milestone2_col].dt.year - (
            (df[milestone2_col].dt.month < 3) |
            ((df[milestone2_col].dt.month == 3) & (df[milestone2_col].dt.day < 26))
        )

        # 3️⃣ Safe numeric conversion (IMPORTANT)
        df['month_num2'] = pd.to_numeric(df['month_num2'], errors='coerce')
        df['year2'] = pd.to_numeric(df['year2'], errors='coerce')

        # 4️⃣ Display year
        df['display_year2'] = df['year2'] + (df['month_num2'] <= 3)

        # 5️⃣ Month label (NaT-safe)
        df['month_name2'] = pd.to_datetime(
            df['month_num2'].astype('Int64').astype(str) + '-' +
            df['display_year2'].astype('Int64').astype(str),
            format='%m-%Y',
            errors='coerce'
        ).dt.strftime('%b-%y')

        # 6️⃣ Current financial year
        # 6️⃣ Selected financial year
        current_fy = year_filtered

        # 7️⃣ Carry Forward
        df.loc[
            df['year2'].notna() & (df['year2'] < current_fy),
            'month_name2'
        ] = 'CF'

        # 8️⃣ Future
        df.loc[
            df['year2'].notna() & (df['year2'] > current_fy),
            'month_name2'
        ] = 'Future'
        
        print(df['month_name1'].dropna().unique().tolist())

        
        print(df['year1'])
        
        print(df['month_name2'])
        
        print(df['year2'])

        def sort_financial_year(summary):
            # CF aur Future hata ke sirf actual month values lo
            months = [
                m for m in summary['month_name'].unique()
                if m not in ["CF", "Future"] and pd.notna(m)
            ]

            # Sirf real months sort honge
            months_sorted = sorted(
                months,
                key=lambda x: pd.to_datetime(x, format='%b-%y')
            )

            ordered = []

            # CF hamesha first
            if "CF" in summary['month_name'].values:
                ordered.append("CF")

            # Actual months
            ordered += months_sorted

            # Future hamesha last
            if "Future" in summary['month_name'].values:
                ordered.append("Future")

            summary['month_name'] = pd.Categorical(
                summary['month_name'],
                categories=ordered,
                ordered=True
            )

            return summary.sort_values('month_name')
        
        def generate_type1_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()

            # Start counts by month_name
            start_counts = (
                temp[temp[start_col].notna()]
                .groupby("month_name1")
                .size()
                .reset_index(name=f"{start_label} Done Count")
                .rename(columns={"month_name1": "month_name"})
            )

            # End counts by month_name2
            end_counts = (
                temp[temp[end_col].notna()]
                .groupby("month_name2")
                .size()
                .reset_index(name=f"{end_label} Done Count")
                .rename(columns={"month_name2": "month_name"})
            )

            # Merge both results
            summary = pd.merge(start_counts, end_counts, on="month_name", how="outer")

            # Convert to Int64 nullable type
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")
            summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].astype("Int64")

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            return summary.reset_index(drop=True)
        
        def generate_type2_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("month_name1").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    # f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
                .rename(columns={"month_name1": "month_name"})
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")

            summary = summary[
                ["month_name", f"{start_label} Done Count", f"{end_label} Done Count"]
            ]

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            return summary.reset_index(drop=True)
        
        summary = generate_type1_summary(df, milestone1, milestone2, milestone1_col, milestone2_col) if type == 'type1' else generate_type2_summary(df, milestone1, milestone2, milestone1_col, milestone2_col)

        numeric_cols = [col for col in summary.columns if "Done Count" in col]

        summary[numeric_cols] = summary[numeric_cols].apply(pd.to_numeric, errors="coerce")
        summary[numeric_cols] = summary[numeric_cols].fillna(0).astype(int)

        
        result_json = summary.to_dict(orient = "records")
        json_data = json.dumps(result_json)
        
        return Response({"message" : "request processed successfully ! ", "json_data": json_data, "unique_data": unique_data}, status=200)
    
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)

@api_view(["POST"])
def nt_issue_summary(request):
    try:
        status = request.data.get('status', 'ALL')
        milestone = request.data.get('milestone', '')
        owner = request.data.get('owner', '')
        duration_start = request.data.get('duration_start')
        duration_end = request.data.get('duration_end')
        site_on_air = request.data.get('is_on_air')

        duration_start = int(duration_start) if duration_start not in [None, ""] else None
        duration_end = int(duration_end) if duration_end not in [None, ""] else None


        qs = NTIssue.objects.all().values(
            "circle",
            "issue_name",
            "issue_owner",
            "milestone",
            "site_id",
            "status",
            "duration",   
            "start_date",
            "close_date"
        )

        df = pd.DataFrame(list(qs))

        if df.empty:
            return Response({"message": "No data found"}, status=200)

        df["circle"] = df["circle"].astype(str).str.strip()
        df["issue_name"] = df["issue_name"].astype(str).str.strip()
        df["site_id"] = df["site_id"].astype(str).str.strip()
        df["milestone"]=df["milestone"].astype(str).str.strip()

  
        if site_on_air in ["Yes", "No"]:
            qs1 = NTSiteTracker.objects.all().values(
                "circle", "site_id", "site_onair_date"
            )

            df1 = pd.DataFrame(list(qs1))

            if not df1.empty:
                df1["circle"] = df1["circle"].fillna("").astype(str).str.strip()
                df1["site_id"] = df1["site_id"].fillna("").astype(str).str.strip()

                df1 = df1.rename(columns={"site_id": "site_id"})
                df1["is_on_air"] = df1["site_onair_date"].notna()

           
                df = df.merge(
                    df1[["circle", "site_id", "is_on_air"]],
                    on=["circle", "site_id"],
                    how="left"
                )

             
                df["is_on_air"] = df["is_on_air"].fillna(False)

                if site_on_air == "No":
                    df = df[df["is_on_air"] == False]

                elif site_on_air == "Yes":
                    df = df[df["is_on_air"] == True]

                
                df = df.drop(columns=["is_on_air"])

                if df.empty:
                    return Response({"message": "No data found"}, status=200)
                
        
        milestone = [c.strip() for c in milestone.split(',')] if milestone else ["ALL"]
        owner = [c.strip() for c in owner.split(',')] if owner else ["ALL"]
        
        if status != 'ALL':
            df = df[df['status'] == status]

        if duration_start is not None:
            df = df[df['duration'] >= duration_start]

        if duration_end is not None:
            df = df[df['duration'] <= duration_end]

        if 'ALL' not in milestone:
            df = df[df['milestone'].isin(milestone)]

        if 'ALL' not in owner:
            df = df[df['issue_owner'].isin(owner)]

        df = df.drop_duplicates(subset=["circle", "issue_name", "site_id", 'issue_owner', 'milestone'])
        
        all_unique_milestones = list(
            NTIssue.objects.exclude(milestone__isnull=True)
            .distinct("milestone")
            .values_list("milestone", flat=True)
        )
        
        all_unique_owners = list(
             NTIssue.objects.exclude(issue_owner__isnull=True)
            .distinct("issue_owner")
            .values_list("issue_owner", flat=True)
        )
        
        unique_data = {
            "unique_milestone": sorted(all_unique_milestones),
            "unique_owners": sorted(all_unique_owners),
        }
        

        pivot_df = pd.pivot_table(
            df,
            index="issue_name",
            columns="circle",
            values="site_id",
            aggfunc="size",
            fill_value=0
        )

    
       
        pivot_df["Total"] = pivot_df.sum(axis=1)
        circle_unique_sites = (
            df.groupby("circle")["site_id"]
            .nunique()
        )

        total_row = circle_unique_sites.to_dict() 
        total_row["Total"] = df["site_id"].nunique()
        total_row_df = pd.DataFrame([total_row], index=["Total"])
        pivot_df = pd.concat([pivot_df, total_row_df])
        circle_order = [
            "AP","ASM","BIH","CHN","DEL","HRY","JK","JRK",
            "KK","KOL","MAH","MP","MUM","NE","ORI","PUN",
            "RAJ","ROTN","UPE","UPW","WB"
        ]

        existing_cols = [c for c in circle_order if c in pivot_df.columns]
        pivot_df = pivot_df.reindex(columns=existing_cols + ["Total"], fill_value=0)
        pivot_df = pivot_df.sort_values(by="Total", ascending=False)
        pivot_df = pivot_df.reset_index().rename(columns={"index": "Issue"})

        return Response({
            "status": True,
            "message": "request processed successfully!",
            "data": pivot_df.to_dict(orient="records"),
            "unique_data": unique_data
        }, status=200)

    except Exception as e:
        return Response({
            "status": False,
            "message": str(e),}, status=500)




ISSUE_COLUMN_MAP = {
    "Circle": "circle",
    "Site ID": "site_id",
    "Issue Owner": "issue_owner",
    "Milestone": "milestone",
    "Issue Name": "issue_name",
    "Start Date": "start_date",
    "Close Date": "close_date",
    "Status": "status",
    "Duration (Days)": "duration",
    "Remarks": "remarks",
}


@api_view(["POST"])
def upload_nt_issue_view(request):
   
    user_id =request.user.username
    if not user_id:
        return Response(
            {"status": False, "message": "userId required"},
            status=400
        )
    user_email = user_id.strip().lower()
  
    file = request.data.get("file")
    if not file:
        return Response(
            {"status": False, "message": "No file uploaded"},
            status=400
        )
    try:
        df = pd.read_excel(file,sheet_name="Issue Tracker")
        df.columns = df.columns.str.strip()
        df = df.replace(
            ["NaN", "nan", "", "NA", "N/A", "-", "null"],
            None
        )
        missing = [
            c for c in ISSUE_COLUMN_MAP.keys()
            if c not in df.columns
        ]

        if missing:
            return Response(
                {
                    "status": False,
                    "message": f"Missing columns: {', '.join(missing)}"
                }
            )
        df = df.rename(columns=ISSUE_COLUMN_MAP)
        df = df[list(ISSUE_COLUMN_MAP.values())]

        INVALID_DATE_STRINGS = {
            "need to check", "pending",
            "na","n/a","-","null","none"
        }

        def safe_date(val):
            try:
                if val is None or pd.isna(val):
                    return None

                if isinstance(val, pd.Timestamp):
                    return val.date()

                if isinstance(val, datetime):
                    return val.date()

                if isinstance(val, date):
                    return val

                val = str(val).strip()

                if val.lower() in INVALID_DATE_STRINGS:
                    return None

                parsed = pd.to_datetime(
                    val,
                    errors="coerce",
                    dayfirst=True
                )

                if pd.isna(parsed):
                    return None

                return parsed.date()

            except:
                return None

        df["start_date"] = df["start_date"].apply(safe_date)
        df["close_date"] = df["close_date"].apply(safe_date)


        df["duration"] = pd.to_numeric(
            df["duration"],
            errors="coerce"
        )

        df["duration"] = df["duration"].where(
            pd.notna(df["duration"]),
            None
        )

        df = df.astype(object)
        df = df.where(pd.notna(df), None)

        records = df.to_dict("records")
        existing = {
            (
                str(x.circle).strip().lower(),
                str(x.site_id).strip().lower(),
                str(x.issue_name).strip().lower(),
                str(x.issue_owner).strip().lower(),
                str(x.milestone).strip().lower(),
            ): x
            for x in NTIssue.objects.all()
        }

        to_create = []
        to_update = []
        with transaction.atomic():
            for row in records:
                if not row.get("circle") or not row.get("site_id"):
                    continue
                key = (
                    str(row["circle"]).strip().lower(),
                    str(row["site_id"]).strip().lower(),
                    str(row["issue_name"]).strip().lower(),
                    str(row["issue_owner"]).strip().lower(),
                    str(row["milestone"]).strip().lower(),
                )
                row["updated_by"] = user_email

                if key in existing:
                    obj = existing[key]

                    for field, value in row.items():
                        setattr(obj, field, value)

                    to_update.append(obj)

                else:
                    row["created_by"] = user_email
                    row["updated_by"] = user_email

                    to_create.append(
                        NTIssue(**row)
                    )
            if to_create:
                NTIssue.objects.bulk_create(
                    to_create,
                    batch_size=500
                )

            if to_update:
                NTIssue.objects.bulk_update(
                    to_update,
                    fields=[
                        "start_date",
                        "close_date",
                        "status",
                        "duration",
                        "remarks",
                        "updated_by",
                    ],
                    batch_size=500
                )

        return Response(
            {
                "status": True,
                "message": "NT Issue Upload Successfully"
            }
        )

    except Exception as e:
        return Response(
            {
                "status": False,
                "message": str(e)
            },
            status=500
        )
    
#site Status Tracker
@api_view(["POST"])
def sync_nt_site_status(request):
    today = date.today()-timedelta(days=1)

    qs = NTSiteTracker.objects.values(
        "site_id",
        "circle",
        "current_status"
        
    )
    if not qs.exists():
        return Response(
            {"error": "No data found in NT_Tracker"},
            status=400
        )

    records = []
    seen = set()

    for row in qs:
        site_id = row["site_id"]
        circle = row["circle"]
        status = row["current_status"]

        if not site_id or not status:
            continue

        site_id = str(site_id).strip()
        circle = str(circle).strip()
        status = str(status).strip()

        key = (site_id, today)
        if key in seen:
            continue

        seen.add(key)

        records.append(
            SiteStatus(
                site_id=site_id,
                circle=circle,
                status=status,
                date=today
            )
        )

    if not records:
        return Response(
            {"error": "No valid records to insert"},
            status=400
        )

    # Optional overwrite logic for today
    # SiteStatus.objects.filter(date=today).delete()

    SiteStatus.objects.bulk_create(records)

    return Response({
        "message": " NTD data synced successfully",
        "records_created": len(records),
        "date": today
    })






















@api_view(['GET', 'POST', 'DELETE'])
def updtae_nt_template(request):
    BASE = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
    template = os.path.join(BASE, "template", "nt_template.xlsx")

    if request.method == "GET":
        if not os.path.exists(template):
            return Response(
                {"error": "Template not found"},
                status=404
            )

        return Response(
            {"status": True,
              "file": template},status=200
        )

    elif request.method == "POST":
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file uploaded"},
                status=400
            )
        os.makedirs(
            os.path.dirname(template),
            exist_ok=True
        )

        with open(template, "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        return Response(
            {"status": True,
              "message": "Template updated successfully"},status=200
        )

    elif request.method == "DELETE":
        if os.path.exists(template):
            os.remove(template)

            return Response(
                {"status": True,
                  "message": "Template deleted successfully"},status=200
            )
        return Response(
            {"error": "Template not found"},
            status=404
        )


@api_view(["POST"])
def nt_issue_summary_frontend(request):

    try:
        issue_name = request.data.get("issue_name")
        circle = request.data.get("circle")
        status = request.data.get("status", "ALL")
        owner = request.data.get("owner", "")
        milestone = request.data.get("milestone", "")
        duration_start = request.data.get("duration_start")
        duration_end = request.data.get("duration_end")
        site_on_air = request.data.get("is_on_air")

        duration_start = int(duration_start) if duration_start not in [None, ""] else None
        duration_end = int(duration_end) if duration_end not in [None, ""] else None

        qs = NTIssue.objects.all().values(
            "circle",
            "site_id",
            "issue_name",
            "issue_owner",
            "milestone",
            "status",
            "duration",
            "start_date",
            "close_date",
            "remarks"
        )

        df = pd.DataFrame(list(qs))

        if df.empty:
            return Response({
                "status": False,
                "message": "No data found"
            }, status=200)

        if issue_name:
            df = df[df["issue_name"] == issue_name]

        if circle and circle != "Total":
            df = df[df["circle"] == circle]

        if status != "ALL":
            df = df[df["status"] == status]

        milestone = [x.strip() for x in milestone.split(",")] if milestone else ["ALL"]
        owner = [x.strip() for x in owner.split(",")] if owner else ["ALL"]

        if "ALL" not in milestone:
            df = df[df["milestone"].isin(milestone)]

        if "ALL" not in owner:
            df = df[df["issue_owner"].isin(owner)]

        if duration_start is not None:
            df = df[df["duration"] >= duration_start]

        if duration_end is not None:
            df = df[df["duration"] <= duration_end]

        if site_on_air in ["Yes", "No"]:

            tracker_df = pd.DataFrame(
                NTSiteTracker.objects.all().values(
                    "circle",
                    "site_id",
                    "site_onair_date"
                )
            )

            if not tracker_df.empty:

                tracker_df["is_on_air"] = tracker_df["site_onair_date"].notna()

                df = df.merge(
                    tracker_df[["circle", "site_id", "is_on_air"]],
                    on=["circle", "site_id"],
                    how="left"
                )

                df["is_on_air"] = df["is_on_air"].fillna(False)

                if site_on_air == "Yes":
                    df = df[df["is_on_air"] == True]

                else:
                    df = df[df["is_on_air"] == False]

                df.drop(columns=["is_on_air"], inplace=True)

        if df.empty:
            return Response({
                "status": False,
                "message": "No data found"
            }, status=200)

        df["start_date"] = pd.to_datetime(
            df["start_date"],
            errors="coerce"
        ).dt.strftime("%d-%b-%y")

        df["close_date"] = pd.to_datetime(
            df["close_date"],
            errors="coerce"
        ).dt.strftime("%d-%b-%y")

        df = df.fillna("")
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE_URL = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE_URL, exist_ok=True)

        output_folder = os.path.join(
            BASE_URL,
            f"generated_files_{current_date}"
        )

        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)

        file_path = os.path.join(
            output_folder,
            f"NT_ISSUE_SUMMARY_DETAIL_{current_date}_{current_time}.xlsx"
        )

        download_df = df.copy()

        download_df.columns = [
            "Circle",
            "Site ID",
            "Issue Name",
            "Issue Owner",
            "Milestone",
            "Status",
            "Duration",
            "Start Date",
            "Close Date",
            "Remarks"
        ]

        with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
            download_df.to_excel( writer, index=False, sheet_name="Issue Summary"
            )
        format_excel(file_path)


        relative_path = file_path.replace(
            settings.MEDIA_ROOT,
            ""
        ).lstrip(os.sep)

        download_link = request.build_absolute_uri(
            os.path.join(
                settings.MEDIA_URL,
                relative_path
            ).replace("\\", "/")
        )

        return Response({
            "status": True,
            "message": "request processed successfully!",
            "data": df.to_dict(orient="records"),
            "download_url":download_link
        }, status=200)

    except Exception as e:

        return Response({
            "status": False,
            "message": str(e)
        }, status=500)
    


@api_view(["POST"])
def ageing_dashboard_hyperlink(request):
    try:
        circle = request.data.get("circle")
        site_status = request.data.get("site_status")   # Done / Pending
        milestone = request.data.get("milestone")

        current_status = request.data.get("current_status", [])
        milestone1 = request.data.get("milestone1")
        milestone2 = request.data.get("milestone2")
        month_filtered = request.data.get("month")
        year_filtered = request.data.get("year")

        current_status = [x.strip() for x in current_status.split(",")] if current_status else ["ALL"]

        month_filtered = int(month_filtered) if str(month_filtered).strip() else None
        year_filtered = int(year_filtered) if str(year_filtered).strip() else None

        if month_filtered and year_filtered:
            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None

        filters = {}

        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status

        if circle and circle != "ALL":
            filters["circle"] = circle

        obj = NTSiteTracker.objects.filter(**filters)
        df = pd.DataFrame(obj.values())

        if df.empty:
            return Response(
                {"message": "No data found", "data": []},
                status=200
            )

        milestone_col = (
            "wrfai_date"
            if milestone == "WRFAI"
            else milestone.lower()
                .replace(" ", "_")
                .replace("-", "_")
                .replace("/", "_")
                .replace("4", "four_")
                .replace("5", "five_") + "_date"
        )

        milestone1_col = (
            "wrfai_date"
            if milestone1 == "WRFAI"
            else milestone1.lower()
                .replace(" ", "_")
                .replace("-", "_")
                .replace("/", "_")
                .replace("4", "four_")
                .replace("5", "five_") + "_date"
        )

        df[milestone_col] = pd.to_datetime(df[milestone_col], errors="coerce")
        df[milestone1_col] = pd.to_datetime(df[milestone1_col], errors="coerce")

        # ================= FILTER LOGIC =================

        if milestone == "RFAI":

            if site_status == "Done":

                if month_start and month_end:
                    df = df[
                        (
                            (df["rfai_date"] >= pd.Timestamp(month_start)) &
                            (df["rfai_date"] <= pd.Timestamp(month_end))
                        )
                        |
                        (
                            (df["rfai_date"] < pd.Timestamp(month_start)) &
                            (df["site_onair_date"].isna())
                        )
                    ]
                else:
                    df = df[df["rfai_date"].notna()]

            else:
                df = df.iloc[0:0]

        else:

            if site_status == "Done":
                df = df[
                    df[milestone_col].notna() &
                    df[milestone1_col].notna()
                ]

            else:
                df = df[
                    df[milestone_col].isna() &
                    df[milestone1_col].notna()
                ]

        if df.empty:
            return Response(
                {"message": "No data found", "data": []},
                status=200
            )

        # ================= DATE FORMAT =================

        for col in df.columns:
            if "date" in col:
                converted = pd.to_datetime(df[col], errors="coerce")

                if col != "last_updated_date":
                    df[col] = converted.dt.strftime("%d-%b-%y")
                else:
                    df[col] = converted.dt.strftime("%d-%b-%y %H:%M:%S")

        # ================= DOWNLOAD FILE =================

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE_URL = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE_URL, exist_ok=True)

        output_folder = os.path.join(
            BASE_URL,
            f"generated_files_{current_date}"
        )

        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)

        tracker_file_path = os.path.join(
            output_folder,
            f"MS1_AGEING_HYPERLINK_{current_date}_{current_time}.xlsx"
        )

        df.insert(0, "Unique ID", range(1, len(df) + 1))
        drop_cols = ["id", "Unique ID"]

        df = df.drop(
            columns=[c for c in drop_cols if c in df.columns],
            errors="ignore"
        )

        reverse_column_map = {
            v: k for k, v in COLUMN_MAP.items()
        }

        df.rename(
            columns=reverse_column_map,
            inplace=True
        )

        with pd.ExcelWriter(tracker_file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="AGEING_DETAIL")
           
        format_excel(tracker_file_path)    

        relative_path = tracker_file_path.replace(
            settings.MEDIA_ROOT,
            ""
        ).lstrip(os.sep)

        download_url = request.build_absolute_uri(
            os.path.join(
                settings.MEDIA_URL,
                relative_path
            ).replace("\\", "/")
        )

        json_data = json.dumps(
            df.fillna("").astype(str).to_dict(orient="records")
        )

        return Response(
            {
                "message": "request processed successfully !!!",
                "data": json_data,
                "download_link": download_url
            },
            status=200
        )

    except Exception as e:
        return Response(
            {"error": str(e)},
            status=500
        )
    
@api_view(['DELETE'])
def delete_nt_tracker(request):
    try:
        deleted_count, _ = NTSiteTracker.objects.all().delete()

        return Response(
            {
                "status": True,
                "message": f"{deleted_count} records deleted successfully"
            },
            status=200
        )

    except Exception as e:
        return Response(
            {"status": False,
              "message": str(e)},
          
        )
    

@api_view(['DELETE'])
def delete_ntissue(request):
    try:
        deleted_count, _ = NTIssue.objects.all().delete()

        return Response(
            {
                "status": True,
                "message": f"{deleted_count} records deleted successfully"
            },
            status=200
        )

    except Exception as e:
        return Response(
            {"status": False,
              "message": str(e)},
          
        )
    


