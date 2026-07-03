from django.shortcuts import render
from django.db.models import *
from rest_framework.decorators import api_view
from rest_framework.response import Response
import pandas as pd
import os
from django.conf import settings
from rest_framework import status
import numpy as np
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from rest_framework.status import HTTP_200_OK
from .models import NokiaAlarm
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font,Border,Side
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin
import xml.etree.ElementTree as ET
import re
import zipfile
import shutil
from .models import *

# def parse_nokia_xml(xml_file):
#       tree = ET.parse(xml_file)
#       root = tree.getroot()

#       data = []

#       for mo in root.iter():

#             if "managedObject" not in mo.tag:
#                   continue

#             row = {}

#             # ---------------- distName ----------------
#             dist_name = mo.attrib.get("distName", "")
#             row["distName"] = dist_name

#             # ---------------- MRBTS / LNBTS ----------------
#             if dist_name:
#                   for part in dist_name.split("/"):
#                         if part.startswith("MRBTS-"):
#                               row["MRBTS"] = part.split("-")[1]
#                         if part.startswith("LNBTS-"):
#                               row["LNBTS"] = part.split("-")[1]

#             # ---------------- <p name=""> parsing ----------------
#             for p in mo:
#                   if p.tag.endswith("p"):
#                         key = p.attrib.get("name")
#                         if key and p.text is not None:
#                               row[key] = p.text

#             data.append(row)
      
#       print("Parsed XML Data:", data)

#       return pd.DataFrame(data)
def compare_lock_status_5g(row):

      old_cell = str(
            row.get("Cell Name_old", "")
      ).strip()

      new_cell = str(
            row.get("Cell Name_new", "")
      ).strip()

      old_state = str(
            row.get(
                  "MRBTS.NRBTS.NRCELL.administrativeState_old",
                  ""
            )
      ).strip().lower()

      new_state = str(
            row.get(
                  "MRBTS.NRBTS.NRCELL.administrativeState_new",
                  ""
            )
      ).strip().lower()

      # split cell name
      old_parts = old_cell.split("_")
      new_parts = new_cell.split("_")

      # prefix match
      old_prefix = "_".join(old_parts[:4])
      new_prefix = "_".join(new_parts[:4])

      # suffix match (A_A / B_B)
      old_suffix = "_".join(old_parts[-2:])
      new_suffix = "_".join(new_parts[-2:])

      # pattern mismatch
      if (
            old_prefix != new_prefix
            or
            old_suffix != new_suffix
      ):
            return " "

      # both locked
      if (
            old_state == "locked"
            and new_state == "locked"
      ):
            return "Both Locked"

      # both unlocked
      elif (
            old_state == "unlocked"
            and new_state == "unlocked"
      ):
            return "Both Unlocked"

      return " "


def compare_lock_status(row):

      old_cell = str(
            row.get("MV Cell Name_old", "")
      ).strip()

      new_cell = str(
            row.get("MV Cell Name_new", "")
      ).strip()

      old_state = str(
            row.get("administrativeState_old", "")
      ).strip().lower()

      new_state = str(
            row.get("administrativeState_new", "")
      ).strip().lower()

      # ---------------- PREFIX ----------------
      old_parts = old_cell.split("_")
      new_parts = new_cell.split("_")

      # MH_E_F1_OM
      old_prefix = "_".join(old_parts[:4])
      new_prefix = "_".join(new_parts[:4])

      # ---------------- SUFFIX ----------------
      old_suffix_match = re.search(
            r'([A-Z]_[A-Z])$',
            old_cell
      )

      new_suffix_match = re.search(
            r'([A-Z]_[A-Z])$',
            new_cell
      )

      old_suffix = (
            old_suffix_match.group(1)
            if old_suffix_match
            else ""
      )

      new_suffix = (
            new_suffix_match.group(1)
            if new_suffix_match
            else ""
      )

      # print(
      #       "OLD:", old_cell,
      #       "| NEW:", new_cell
      # )
      # print(
      #       old_prefix,
      #       old_suffix,
      #       "----",
      #       new_prefix,
      #       new_suffix
      # )

      # ---------------- MATCH ----------------
      if (
            old_prefix != new_prefix
            or
            old_suffix != new_suffix
      ):
            return ""

      # ---------------- LOCK STATUS ----------------
      if (
            old_state == "locked"
            and new_state == "locked"
      ):
            return "Both Locked"

      elif (
            old_state == "unlocked"
            and new_state == "unlocked"
      ):
            return "Both Unlocked"

      return ""

# def parse_nokia_xml(xml_file):

#       tree = ET.parse(xml_file)
#       root = tree.getroot()

#       data = []

#       for mo in root.iter():

#             # only managedObject
#             if "managedObject" not in mo.tag:
#                   continue

#             row = {}

#             # ---------------- Extract IDs ----------------
#             dist_name = mo.attrib.get("distName", "")

#             if dist_name:

#                   for part in dist_name.split("/"):

#                         if part.startswith("MRBTS-"):
#                               row["MRBTS"] = (
#                                     part.split("-")[1]
#                               )

#                         elif part.startswith("LNBTS-"):
#                               row["LNBTS"] = (
#                                     part.split("-")[1]
#                               )

#                         elif part.startswith("NRBTS-"):
#                               row["NRBTS"] = (
#                                     part.split("-")[1]
#                               )

#             # ---------------- Required p tags only ----------------
#             for p in mo:

#                   if not p.tag.endswith("p"):
#                         continue

#                   key = p.attrib.get("name")

#                   if key == "name":
#                         row["name"] = (
#                               p.text.strip()
#                               if p.text
#                               else ""
#                         )

#                   elif key == "administrativeState":
#                         row["administrativeState"] = (
#                               p.text.strip()
#                               if p.text
#                               else ""
#                         )

#             # only append useful rows
#             if (
#                   row.get("name")
#                   or row.get("administrativeState")
#                   or row.get("MRBTS")
#             ):
#                   data.append(row)

#       print("Parsed XML Data is processing........")

#       return pd.DataFrame(data)



def format_of_excel(file_path):
      wb = load_workbook(file_path)
      ws = wb.active

      # ---------- Styles ----------
      header_fill = PatternFill(start_color="E87529", end_color="E87529", fill_type="solid")
      alarm_header_fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")

      header_font = Font(color="000000", bold=True)

      thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
      )

      # ---------- Header Formatting ----------
      for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

      ws.row_dimensions[1].height = 40  # visible header

      header_list = [cell.value for cell in ws[1]]

      lb_cols = [
            "Site Id","MS2 Status","Circle","Partner Name","Old/New",
            "OEM","On-air Date","On-air month","MS2 status",
            "ip","MRBTS"
      ]

      for col_name in lb_cols:
            if col_name in header_list:
                  col_idx = header_list.index(col_name) + 1
                  cell = ws.cell(row=1, column=col_idx)
                  cell.fill = alarm_header_fill
                  cell.font = header_font

      # ---------- Data Borders ----------
      for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                  cell.border = thin_border

      # ---------- UNIFORM COLUMN WIDTH ----------
      DEFAULT_WIDTH = 25

      for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = DEFAULT_WIDTH

      wb.save(file_path)


@api_view(['GET'])
def get_sites(request):
      try:
            data = list(Old_New.objects.values())

            return Response({
                  "status": True,
                  "data": data
            })

      except Exception as e:
            return Response({
                  "status": False,
                  "error": str(e)
            }, status=500)


@api_view(['DELETE'])
def delete_sites(request):
      try:
            Old_New.objects.all().delete()

            return Response({
                  "status": True,
                  "message": "Deleted successfully"
            })
      except Exception as e:
            return Response({
                  "status": False,
                  "error": str(e)
            }, status=500)


@api_view(['POST'])
def upload_site_list(request):
      file = request.FILES.get('site_file')
      if not file:
            return Response({
                  "status": False,
                  "error": "File required"
            }, status=400)
      if not file.name.endswith('.xlsx'):
            return Response({
                  "status": False,
                  "error": "Only Excel file allowed"
            }, status=400)

      try:
            Old_New.objects.all().delete()
            df = pd.read_excel(file, sheet_name="Sheet 1")

            if df is not None:

                  df.columns = df.columns.str.strip()

                  for _, row in df.iterrows():

                        Old_New.objects.update_or_create(

                              new_site=str(row.get('New SiteId', '')).strip(),
                              old_site=str(row.get('Old SiteId', '')).strip(),
                              new_4g_mrbts=str(row.get('New 4G MRBTS', '')).strip(),
                              old_4g_mrbts=str(row.get('OLD 4G MRBTS', '')).strip(),
                              new_5g_nrbts=str(row.get('New 5G NRBTS', '')).strip(),
                              old_5g_nrbts=str(row.get('OLD 5G NRBTS', '')).strip(),

                        )

            return Response({
            "status": True,
            "message": "Data uploaded successfully"
            })

      except Exception as e:
            return Response({
                  "status": False,
                  "error": str(e)
            }, status=500)

@api_view(['GET', 'POST'])
def fileupload(request):
      if request.method == 'GET':
            # alarms = NokiaAlarm.objects.all().values('id', 'SA', 'NSA')
            alarms = NokiaAlarm.objects.all().values(
                  'id',
                  'alarm_name',
                  'alarm_status',
                  'sa_nsa',
                  'alarm_bucket',
                  'responsibility'
                  )
            return Response(list(alarms), status=status.HTTP_200_OK)

      elif request.method == 'POST':
            file = request.FILES.get('file')

            if not file:
                  return Response({'error': 'SA/NSA file is required!'}, status=status.HTTP_400_BAD_REQUEST)

            # Read Excel or CSV file
            try:
                  if file.name.endswith('.csv'):
                        df = pd.read_csv(file)
                  else:
                        df = pd.read_excel(file)

                        df = df.dropna(how="all")

                  NokiaAlarm.objects.all().delete() 
            except Exception as e:
                  return Response({'error': f'Error reading file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

            # Expecting at least one of these columns
            # if not {'SA', 'NSA'}.intersection(df.columns):
            #       return Response({'error': 'File must contain SA or NSA column.'}, status=status.HTTP_400_BAD_REQUEST)
            required_cols = {
                  'ALARMS',
                  'Alarm/No Alarm',
                  'SA/NSA',
                  'Alarm Bucket',
                  'Responsibility'
            }

            missing_cols = required_cols - set(df.columns)

            if missing_cols:
                  return Response(
                        {
                              'error': f'Missing columns: {list(missing_cols)}'
                        },
                        status=status.HTTP_400_BAD_REQUEST
                  )
            # Save each row to database
            count = 0
            # for _, row in df.iterrows():
            #       NokiaAlarm.objects.create(
            #       SA=row.get('SA'),
            #       NSA=row.get('NSA')
            #       )
            for _, row in df.iterrows():
                  NokiaAlarm.objects.create(
                        alarm_name=row.get("ALARMS"),
                        alarm_status=row.get("Alarm/No Alarm"),
                        sa_nsa=row.get("SA/NSA"),
                        alarm_bucket=row.get("Alarm Bucket"),
                        responsibility=row.get("Responsibility")
                  )
                  count += 1

            return Response({"status": True,'message': f'{count} alarms saved successfully!'}, status=status.HTTP_201_CREATED)

@api_view(['POST'])
def alarmfileUpload(request):

      alarm_files = request.FILES.getlist('alarm_file')
      mapping_file = request.FILES.get('mapping_file')

      if not alarm_files or not mapping_file:
            return Response(
                  {'error': 'Both alarm_file and mapping_file are required!'},
                  status=status.HTTP_400_BAD_REQUEST
            )
      
      # ---------------- Clean Old Output ----------------
      output_root = os.path.join(
            MEDIA_ROOT,
            "NOKIA_OUTPUT"
      )

      final_output_folder = os.path.join(
            output_root,
            "OUTPUT"
      )

      zip_path = os.path.join(
            output_root,
            "OUTPUT.zip"
      )

      # delete old output folder
      if os.path.exists(final_output_folder):
            shutil.rmtree(final_output_folder)

      # delete old zip
      if os.path.exists(zip_path):
            os.remove(zip_path)

      # ---------------- Read Alarm Files ----------------
      alarm_dfs = []

      for alarm_file in alarm_files:
            try:
                  filename = alarm_file.name.lower()

                  if filename.endswith('.csv'):
                        df = pd.read_csv(alarm_file)

                  elif filename.endswith(('.xlsx', '.xls')):
                        df = pd.read_excel(alarm_file)

                  else:
                        continue
                  df.columns = df.columns.str.strip()
                  alarm_dfs.append(df)

            except Exception as e:
                  return Response(
                  {'error': f'Error reading alarm file {alarm_file.name}: {str(e)}'},
                  status=status.HTTP_400_BAD_REQUEST
                  )

      if not alarm_dfs:
            return Response(
                  {'error': 'No valid alarm files found'},
                  status=status.HTTP_400_BAD_REQUEST
            )
      df_alarm = pd.concat(alarm_dfs, ignore_index=True)
      
      # ---------------- Separate 4G / 5G Status Files ----------------
      df_4g = pd.DataFrame()
      df_5g = pd.DataFrame()

      for df in alarm_dfs:

            cols = [str(c).strip() for c in df.columns]

            
            # ---------- 4G ----------
            if "ECGI" in cols:

                  temp_4g = df.copy()

                  # ---------------- CLEAN COLUMNS ----------------
                  temp_4g.columns = temp_4g.columns.astype(str).str.strip()

                  required_cols = [
                        "ECGI",
                        "MV Site Name",
                        "MV Cell Name",
                        "administrativeState",
                        "operationalState",
                        "NE_STATE"
                  ]

                  temp_4g = temp_4g[
                        [c for c in required_cols if c in temp_4g.columns]
                  ]
                  # Keep only ACTIVE NE_STATE rows
                  if "NE_STATE" in temp_4g.columns:

                        temp_4g["NE_STATE"] = (
                              temp_4g["NE_STATE"]
                              .astype(str)
                              .str.strip()
                              .str.upper()
                        )

                        temp_4g = temp_4g[
                              temp_4g["NE_STATE"] == "ACTIVE"
                        ]
                  # ---------------- EXTRACT MRBTS FROM ECGI ----------------
                  # Example: 405-51-38163-91 → 38163
                  temp_4g["MRBTS"] = (
                        temp_4g["ECGI"]
                        .astype(str)
                        .str.extract(r'^\d+-\d+-(\d+)-')[0]
                  )

                  # ---------------- CLEAN MRBTS ----------------
                  temp_4g["MRBTS"] = (
                        temp_4g["MRBTS"]
                        .astype(str)
                        .str.replace(r'\.0', '', regex=True)
                        .str.strip()
                  )

                  # ---------------- REMOVE BAD VALUES ----------------
                  temp_4g = temp_4g[
                        temp_4g["MRBTS"].notna()
                  ]

                  temp_4g = temp_4g[
                        temp_4g["MRBTS"].str.match(r'^\d+$', na=False)
                  ]

                  # ---------------- SITE ID EXTRACTION ----------------
                  temp_4g["Site Id"] = (
                        temp_4g["MV Site Name"]
                        .astype(str)
                        .str.strip()
                        .str.extract(r'([^_]+$)')[0]
                  )

                  # ---------------- APPEND ----------------
                  df_4g = pd.concat(
                        [df_4g, temp_4g],
                        ignore_index=True
                  )

            # ---------- 5G ----------
            elif "NCGI" in cols:

                  temp_5g = df.copy()
                  temp_5g.columns = temp_5g.columns.astype(str).str.strip()
                  required_cols = [
                        "NCGI",
                        "Site Name",
                        "Cell Name",
                        "MRBTS.NRBTS.NRCELL.administrativeState",
                        "MRBTS.NRBTS.NRCELL.operationalState",
                        "NE_STATE",
                  ]

                  temp_5g = temp_5g[
                        [c for c in required_cols if c in temp_5g.columns]
                  ]
                  # Keep only ACTIVE NE_STATE rows
                  if "NE_STATE" in temp_5g.columns:
                        temp_5g["NE_STATE"] = (
                              temp_5g["NE_STATE"]
                              .astype(str)
                              .str.strip()
                              .str.upper()
                        )

                        temp_5g = temp_5g[
                              temp_5g["NE_STATE"] == "ACTIVE"
                        ]
                  # NCGI → NRBTS
                  temp_5g["NRBTS"] = (
                        temp_5g["NCGI"]
                        .astype(str)
                        .str.extract(r'^\d+-\d+-(\d+)-')[0]
                  )
                  # ---------------- CLEAN MRBTS ----------------
                  temp_5g["NRBTS"] = (
                        temp_5g["NRBTS"]
                        .astype(str)
                        .str.replace(r'\.0', '', regex=True)
                        .str.strip()
                  )
                  
                  # ---------------- REMOVE BAD VALUES ----------------
                  temp_5g = temp_5g[
                        temp_5g["NRBTS"].notna()
                  ]

                  temp_5g = temp_5g[
                        temp_5g["NRBTS"].str.match(r'^\d+$', na=False)
                  ]

                  df_5g = pd.concat(
                        [df_5g, temp_5g],
                        ignore_index=True
                  )
                  
            # print("-----5g data-----",df_5g)
            
      
      
      # ---------------- Read Mapping File ----------------
      try:
            mapping_filename = mapping_file.name.lower()

            if mapping_filename.endswith('.csv'):
                  df_map = pd.read_csv(mapping_file)

            elif mapping_filename.endswith(('.xlsx', '.xls')):
                  df_map = pd.read_excel(mapping_file)

            else:
                  return Response(
                  {'error': 'Unsupported mapping file format'},
                  status=status.HTTP_400_BAD_REQUEST
                  )

            df_map.columns = df_map.columns.str.strip()

      except Exception as e:
            return Response(
                  {'error': f'Error reading mapping file: {str(e)}'},
                  status=status.HTTP_400_BAD_REQUEST
            )

      # ---------------- Ensure MRBTS ----------------
      
      if "MRBTS" not in df_alarm.columns:
            if "Distinguished Name" in df_alarm.columns:
                  df_alarm["MRBTS"] = (
                  df_alarm["Distinguished Name"]
                  .astype(str)
                  .str.extract(r"MRBTS-(\d+)")
                  )
            else:
                  df_alarm["MRBTS"] = ""
                  
      # ---------------- Extract NRBTS ----------------
      if "Distinguished Name" in df_alarm.columns:

            df_alarm["NRBTS"] = (
                  df_alarm["Distinguished Name"]
                  .astype(str)
                  .str.extract(r"NRBTS-(\d+)")
            )

      else:
            df_alarm["NRBTS"] = ""
      
      # ---------------- Extract 5G Alarms ----------------
      if (
            "Distinguished Name" in df_alarm.columns
            and "Supplementary Information" in df_alarm.columns
      ):

            df_alarm["5G Alarms"] = np.where(
                  df_alarm["Distinguished Name"]
                  .astype(str)
                  .str.contains("NRBTS-", na=False),

                  df_alarm["Supplementary Information"]
                  .astype(str),

                  ""
            )

      else:
            df_alarm["5G Alarms"] = ""


      needed_cols = [
            "MRBTS",
            "NRBTS",
            "5G Alarms",
            "Supplementary Information",
      ]
      df_alarm = df_alarm[[c for c in needed_cols if c in df_alarm.columns]]

      # ---------------- Load SA / NSA from DB ----------------
      # sa = list(NokiaAlarm.objects.exclude(SA__isnull=True).values_list('SA', flat=True))
      # nsa = list(NokiaAlarm.objects.exclude(NSA__isnull=True).values_list('NSA', flat=True))
      
      alarm_master = NokiaAlarm.objects.all()

      alarm_lookup = {
      obj.alarm_name.strip(): {
            "sa_nsa": obj.sa_nsa,
            "alarm_bucket": obj.alarm_bucket,
            "responsibility": obj.responsibility
      }
      for obj in alarm_master
      if obj.alarm_name
      }
      # Keep old variables to avoid breaking existing code
      sa = [
      k for k, v in alarm_lookup.items()
      if str(v.get("sa_nsa")).strip().upper() == "SA"
      ]

      nsa = [
      k for k, v in alarm_lookup.items()
      if str(v.get("sa_nsa")).strip().upper() == "NSA"
      ]
      # combined_alarms = set(sa + nsa)

      # ---------------- Timeout Regex ----------------
      timeout_regex = re.compile(r"Timeout\s+connecting\s+to|No\s+route\s+to\s+host|CommunicationTimeout|CommunicationTimeout", re.IGNORECASE)

      # ---------------- Classify Alarms ----------------
      def classify_alarms(info):

            if pd.isna(info):
                  return pd.Series([
                        "No",
                        "No Alarms",
                        "No Alarms",
                        "",
                        ""
                  ])

            alarms = [
                  a.strip()
                  for a in str(info).split(',')
                  if a.strip()
            ]

            matched_sa_nsa = []
            matched_bucket = []
            matched_responsibility = []

            for alarm in alarms:

                  # timeout → site down
                  if timeout_regex.search(alarm):

                        matched_sa_nsa.append("SITE DOWN")
                        matched_bucket.append("SITE DOWN")
                        matched_responsibility.append(
                        "Circle Team"
                        )

                        continue

                  if alarm in alarm_lookup:

                        details = alarm_lookup[alarm]

                        matched_sa_nsa.append(
                        details["sa_nsa"]
                        )

                        matched_bucket.append(
                        details["alarm_bucket"]
                        )

                        matched_responsibility.append(
                        details["responsibility"]
                        )

            return pd.Series([
                  "Yes" if matched_sa_nsa else "No",

                  ", ".join(set(matched_sa_nsa))
                  if matched_sa_nsa
                  else "No Alarms",

                  ", ".join(set(matched_bucket))
                  if matched_bucket else "",

                  ", ".join(set(matched_responsibility))
                  if matched_responsibility else "",

                  info if info else "No Alarms"
            ])
      df_alarm[
      [
            "Alarm Status (Yes/No)",
            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms",
            "Alarm Bucket",
            "Responsibility",
            "Service Affecting Alarms"
      ]
      ] = (
      df_alarm["Supplementary Information"]
      .apply(classify_alarms)
      )
      
    # ---------------- Group & Deduplicate ----------------
      # df_alarm = (
      #       df_alarm.drop_duplicates()
      #       .groupby("MRBTS", as_index=False)
      #       .agg(lambda x: ', '.join(
      #             sorted(set(str(i) for i in x if pd.notna(i) and str(i).strip()))
      #       ))
      # )
      
      # ---------------- Clean + Group ----------------

      def clean_and_merge(values, column_name=""):

            # remove nan / empty
            vals = [
                  str(v).strip()
                  for v in values
                  if pd.notna(v)
                  and str(v).strip()
                  and str(v).strip().lower() != "nan"
            ]

            vals_set = set(vals)

            # ---------------- Alarm Status ----------------
            if column_name == "Alarm Status (Yes/No)":

                  # if both Yes and No exist -> keep Yes
                  if "Yes" in vals_set:
                        return "Yes"

                  return "No"

            # ---------------- SA / NSA ----------------
            if column_name == "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms":

                  vals_upper = {
                        str(v).strip().upper()
                        for v in vals_set
                  }

                  # Priority:
                  # SITE DOWN > SA > NSA > No Alarms
                  if "SITE DOWN" in vals_upper:
                        return "SITE DOWN"

                  if "SA" in vals_upper:
                        return "SA"

                  if "NSA" in vals_upper:
                        return "NSA"

                  return "No Alarms"

            # ---------------- Other Columns ----------------
            cleaned = sorted(set(vals))

            return ", ".join(cleaned)


      df_alarm = (
            df_alarm.drop_duplicates()
            .groupby("MRBTS", as_index=False)
            .agg(lambda x: clean_and_merge(x, x.name))
      )

      # ---------------- Normalize MRBTS ----------------
      df_alarm["MRBTS"] = df_alarm["MRBTS"].astype(str).str.replace('.0', '', regex=False)
      df_alarm["NRBTS"] = df_alarm["NRBTS"].astype(str).str.replace('.0', '', regex=False)
      # df_map["MRBTS"] = df_map["MRBTS"].astype(str).str.replace('.0', '', regex=False)
      df_map["MRBTS"] = (
            df_map["MRBTS"]
            .astype(str)
            .str.replace(r'\.0', '', regex=True)
            .str.strip()
            )

      df_map.rename(
            columns={"Service Affecting Alarms 1": "Service Affecting Alarms"},
            inplace=True
      )
      # ---------------- Merge ----------------
      df_merged = pd.merge(
            df_map,
            df_alarm,
            on="MRBTS",
            how="left",
            suffixes=('', '_alarm')
      )

      # ---------------- Update Existing Columns Only ----------------
      cols_to_update = [
            "Alarm Status (Yes/No)",
            "Service Affecting Alarms",
            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms"
      ]

      for col in cols_to_update:
            alarm_col = f"{col}_alarm"
            if alarm_col in df_merged.columns:
                  df_merged[col] = df_merged[col].mask(
                  df_merged[col].isna() | (df_merged[col].astype(str).str.strip() == ""),
                  df_merged[alarm_col]
                  )
                  df_merged.drop(columns=[alarm_col], inplace=True)
                  
      # ---------------- Fill Blank SA/NSA ----------------
      df_merged[
            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms"
      ] = df_merged[
            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms"
      ].replace(
            ["", "nan", None],
            "No Alarms"
      )

      # ---------------- Add Extra Columns If Missing ----------------
      extra_cols = [
            # "Effective Alarms",
            "Supplementary Information",
            # "Origin Alarm Time",
            # "Origin Alarm Update Time"
      ]

      for col in extra_cols:
            alarm_col = f"{col}_alarm"
            if alarm_col in df_merged.columns:
                  if col not in df_merged.columns:
                        df_merged[col] = df_merged[alarm_col]
                  df_merged.drop(columns=[alarm_col], inplace=True)
      
      if "Effective Alarms" in df_merged.columns:
            df_merged.drop(columns=["Effective Alarms"], inplace=True)
            
      # Rename column
      df_merged.rename(
            columns={
                  "Supplementary Information": "Total Alarms"
            },
            inplace=True
      )
      # ---------------- 4G / 5G Alarm Status ----------------
      old_new_df = pd.DataFrame(
            list(
                  Old_New.objects.values(
                        "new_4g_mrbts",
                        "old_4g_mrbts",
                        "new_5g_nrbts",
                        "old_5g_nrbts"
                  )
            )
      )

      # clean values
      for col in old_new_df.columns:
            old_new_df[col] = (
                  old_new_df[col]
                  .astype(str)
                  .str.replace(r"\.0", "", regex=True)
                  .str.strip()
            )

      # create lookup set
      new_4g_set = set(old_new_df["new_4g_mrbts"])
      old_4g_set = set(old_new_df["old_4g_mrbts"])

      new_5g_set = set(old_new_df["new_5g_nrbts"])
      old_5g_set = set(old_new_df["old_5g_nrbts"])

      # clean merged cols
      df_merged["MRBTS"] = (
            df_merged["MRBTS"]
            .astype(str)
            .str.replace(r"\.0", "", regex=True)
            .str.strip()
      )

      df_merged["NRBTS"] = (
            df_merged["NRBTS"]
            .astype(str)
            .str.replace(r"\.0", "", regex=True)
            .str.strip()
      )
      # print("=======df_merged=======",df_merged["NRBTS"])


      # 4G Status
      def get_4g_status(mrbts):

            mrbts = str(mrbts).strip()

            if mrbts in new_4g_set:
                  return "New Site"

            elif mrbts in old_4g_set:
                  return "OLD Site"

            return ""


      # 5G Status
      def get_5g_status(nrbts):

            # null check
            if pd.isna(nrbts):
                  return ""

            # clean nrbts
            nrbts = (
                  str(nrbts)
                  .replace(".0", "")
                  .strip()
            )

            # ignore nan string
            if not nrbts or nrbts.lower() == "nan":
                  return ""

            # print("Checking NRBTS:", nrbts)

            # match new
            if nrbts in new_5g_set:
                  return "New Site"

            # match old
            elif nrbts in old_5g_set:
                  return "OLD Site"

            return ""


      # create columns
      df_merged["4G Alarm Status"] = (
            df_merged["MRBTS"]
            .apply(get_4g_status)
      )

      df_merged["5G Alarm Status"] = (
            df_merged["NRBTS"]
            .apply(get_5g_status)
      )
      # ---------------- Save Circle Wise Output ----------------
      # output_root = os.path.join(MEDIA_ROOT, "NOKIA_OUTPUT")
      # os.makedirs(output_root, exist_ok=True)
      # saved_files = []
      # ---------------- Save Circle Wise Output ----------------
      output_root = os.path.join(
            MEDIA_ROOT,
            "NOKIA_OUTPUT"
      )

      os.makedirs(output_root, exist_ok=True)

      # OUTPUT folder
      final_output_folder = os.path.join(
            output_root,
            "OUTPUT"
      )

      # remove old OUTPUT folder
      if os.path.exists(final_output_folder):
            shutil.rmtree(final_output_folder)

      os.makedirs(final_output_folder, exist_ok=True)

      saved_files = []
      # Ensure Circle column exists
      if "Circle" not in df_merged.columns:
            return Response(
                  {"error": "Circle column not found in mapping file"},
                  status=status.HTTP_400_BAD_REQUEST
            )

      # Clean circle values
      df_merged["Circle"] = (
            df_merged["Circle"]
            .astype(str)
            .str.strip()
      )
      # keep only rows having actual alarms
      # df_merged = df_merged[
      #       df_merged["Total Alarms"].notna()
      # ]
      # ---------------- Split Circle Wise ----------------
      for circle, circle_df in df_merged.groupby("Circle"):
            # skip fully blank circles
            if circle_df["Total Alarms"].isna().all():
                  continue
            if not circle or circle.lower() == "nan":
                  continue
            
            # Circle folder inside OUTPUT
            circle_folder = os.path.join(
                  final_output_folder,
                  circle
            )

            os.makedirs(circle_folder, exist_ok=True)

            # File name
            output_filename = (
                  f"Nokia_SA_NSA_OUTPUT_{circle}.xlsx"
            )

            full_output_path = os.path.join(
                  circle_folder,
                  output_filename
            )

            # Save excel
            # circle_df.to_excel(
            #       full_output_path,
            #       index=False
            # )
            with pd.ExcelWriter(
                  full_output_path,
                  engine="openpyxl"
            ) as writer:

                  # Main Alarm Sheet
                  circle_df.to_excel(
                        writer,
                        sheet_name="Alarm_Output",
                        index=False
                  )

                  # ---------------- 4G Status ----------------
                  if not df_4g.empty:

                        circle_4g = df_4g[
                              df_4g["MRBTS"]
                              .astype(str)
                              .isin(
                                    circle_df["MRBTS"]
                                    .astype(str)
                              )
                        ]

                        if not circle_4g.empty:

                              # 4G Lock / Unlock Mapping
                              circle_4g["administrativeState"] = (
                                    pd.to_numeric(
                                          circle_4g["administrativeState"],
                                          errors="coerce"
                                    )
                                    .map({
                                          1: "Unlocked",
                                          3: "Locked",
                                          2: "lock/Shutter Down"
                                    })
                                    .fillna(
                                          circle_4g["administrativeState"]
                                    )
                              )

                              # 4G Enable / Disable Mapping
                              circle_4g["operationalState"] = (
                                    pd.to_numeric(
                                          circle_4g["operationalState"],
                                          errors="coerce"
                                    )
                                    .map({
                                          0: "Disabled",
                                          1: "Enabled"
                                    })
                                    .fillna(
                                          circle_4g["operationalState"]
                                    )
                              )
                              # ---------------- 4G Alarm Status ----------------
                              circle_4g["MRBTS"] = (
                                    circle_4g["MRBTS"]
                                    .astype(str)
                                    .str.replace(r"\.0", "", regex=True)
                                    .str.strip()
                              )

                              circle_4g["4G Alarm Status"] = (
                                    circle_4g["MRBTS"]
                                    .apply(get_4g_status)
                              )

                              circle_4g.to_excel(
                                    writer,
                                    sheet_name="4G_Status",
                                    index=False
                              )

                  # ---------------- 5G Status ----------------
                  if not df_5g.empty:

                        circle_5g = df_5g[
                              df_5g["NRBTS"]
                              .astype(str)
                              .isin(
                                    circle_df["NRBTS"]
                                    .astype(str)
                              )
                        ]

                        if not circle_5g.empty:

                              circle_5g = circle_5g[
                                    [
                                          "NCGI",
                                          "NRBTS",
                                          "Site Name",
                                          "Cell Name",
                                          "MRBTS.NRBTS.NRCELL.administrativeState",
                                          "MRBTS.NRBTS.NRCELL.operationalState"
                                    ]
                              ]
                              # 5G Lock / Unlock Mapping
                              circle_5g[
                                    "MRBTS.NRBTS.NRCELL.administrativeState"
                              ] = (
                                    pd.to_numeric(
                                          circle_5g[
                                                "MRBTS.NRBTS.NRCELL.administrativeState"
                                          ],
                                          errors="coerce"
                                    )
                                    .map({
                                          0: "Locked",
                                          2: "Unlocked",
                                          1: "lock/Shutter Down"
                                    })
                                    .fillna(
                                          circle_5g[
                                                "MRBTS.NRBTS.NRCELL.administrativeState"
                                          ]
                                    )
                              )

                              # 5G Enable / Disable Mapping
                              circle_5g[
                                    "MRBTS.NRBTS.NRCELL.operationalState"
                              ] = (
                                    pd.to_numeric(
                                          circle_5g[
                                                "MRBTS.NRBTS.NRCELL.operationalState"
                                          ],
                                          errors="coerce"
                                    )
                                    .map({
                                          0: "Disabled",
                                          1: "Enabled"
                                    })
                                    .fillna(
                                          circle_5g[
                                                "MRBTS.NRBTS.NRCELL.operationalState"
                                          ]
                                    )
                              )
                              # ---------------- 5G Alarm Status ----------------
                              circle_5g["NRBTS"] = (
                                    circle_5g["NRBTS"]
                                    .astype(str)
                                    .str.replace(r"\.0", "", regex=True)
                                    .str.strip()
                              )

                              circle_5g["5G Alarm Status"] = (
                                    circle_5g["NRBTS"]
                                    .apply(get_5g_status)
                              )
                              circle_5g.to_excel(
                                    writer,
                                    sheet_name="5G_Status",
                                    index=False
                              )
                              
                              # ---------------- 4G OLD VS NEW SHEET ----------------
                              if not df_4g.empty:

                                    # clean MRBTS
                                    temp_4g_compare = df_4g.copy()
                                    # ---------------- 4G Lock / Unlock Mapping ----------------
                                    if "administrativeState" in temp_4g_compare.columns:

                                          temp_4g_compare["administrativeState"] = (
                                                pd.to_numeric(
                                                      temp_4g_compare["administrativeState"],
                                                      errors="coerce"
                                                )
                                                .map({
                                                      1: "Unlocked",
                                                      3: "Locked",
                                                      2: "lock/Shutter Down"
                                                })
                                                .fillna(
                                                      temp_4g_compare["administrativeState"]
                                                )
                                          )

                                    # ---------------- 4G Enable / Disable Mapping ----------------
                                    if "operationalState" in temp_4g_compare.columns:

                                          temp_4g_compare["operationalState"] = (
                                                pd.to_numeric(
                                                      temp_4g_compare["operationalState"],
                                                      errors="coerce"
                                                )
                                                .map({
                                                      0: "Disabled",
                                                      1: "Enabled"
                                                })
                                                .fillna(
                                                      temp_4g_compare["operationalState"]
                                                )
                                          )

                                    temp_4g_compare["MRBTS"] = (
                                          temp_4g_compare["MRBTS"]
                                          .astype(str)
                                          .str.replace(r"\.0", "", regex=True)
                                          .str.strip()
                                    )

                                    compare_4g_rows = []

                                    for _, row in old_new_df.iterrows():
                                          

                                          old_mrbts = str(
                                                row.get("old_4g_mrbts", "")
                                          ).replace(".0", "").strip()

                                          new_mrbts = str(
                                                row.get("new_4g_mrbts", "")
                                          ).replace(".0", "").strip()
                                          # ---------------- Current Circle Filter ----------------
                                          circle_mrbts = set(
                                                circle_df["MRBTS"]
                                                .astype(str)
                                                .str.replace(".0", "", regex=False)
                                                .str.strip()
                                          )

                                          # skip other circle data
                                          if (
                                                old_mrbts not in circle_mrbts
                                                and
                                                new_mrbts not in circle_mrbts
                                          ):
                                                continue
                                          
                                          # find old row
                                          old_row = temp_4g_compare[
                                                temp_4g_compare["MRBTS"] == old_mrbts
                                          ]

                                          # find new row
                                          new_row = temp_4g_compare[
                                                temp_4g_compare["MRBTS"] == new_mrbts
                                          ]

                                          if old_row.empty and new_row.empty:
                                                continue

                                          old_dict = {}
                                          new_dict = {}

                                          # old data
                                          if not old_row.empty:
                                                old_dict = {
                                                      f"{col}_old": val
                                                      for col, val in old_row.iloc[0].to_dict().items()
                                                }

                                          # new data
                                          if not new_row.empty:
                                                new_dict = {
                                                      f"{col}_new": val
                                                      for col, val in new_row.iloc[0].to_dict().items()
                                                }

                                          compare_4g_rows.append({
                                                **old_dict,
                                                **new_dict
                                          })

                                    if compare_4g_rows:

                                          compare_4g_df = pd.DataFrame(
                                                compare_4g_rows
                                          )
                                          compare_4g_df["Lock Status"] = (
                                                compare_4g_df.apply(
                                                      compare_lock_status,
                                                      axis=1
                                                )
                                          )
                                          compare_4g_df.to_excel(
                                                writer,
                                                sheet_name="4G Old vs New",
                                                index=False
                                          )


                              # ---------------- 5G OLD VS NEW SHEET ----------------
                              if not df_5g.empty:

                                    temp_5g_compare = df_5g.copy()
                                    # ---------------- 5G Lock / Unlock Mapping ----------------
                                    if "MRBTS.NRBTS.NRCELL.administrativeState" in temp_5g_compare.columns:

                                          temp_5g_compare[
                                                "MRBTS.NRBTS.NRCELL.administrativeState"
                                          ] = (
                                                pd.to_numeric(
                                                      temp_5g_compare[
                                                            "MRBTS.NRBTS.NRCELL.administrativeState"
                                                      ],
                                                      errors="coerce"
                                                )
                                                .map({
                                                      0: "Locked",
                                                      2: "Unlocked",
                                                      1: "lock/Shutter Down"
                                                })
                                                .fillna(
                                                      temp_5g_compare[
                                                            "MRBTS.NRBTS.NRCELL.administrativeState"
                                                      ]
                                                )
                                          )

                                    # ---------------- 5G Enable / Disable Mapping ----------------
                                    if "MRBTS.NRBTS.NRCELL.operationalState" in temp_5g_compare.columns:

                                          temp_5g_compare[
                                                "MRBTS.NRBTS.NRCELL.operationalState"
                                          ] = (
                                                pd.to_numeric(
                                                      temp_5g_compare[
                                                            "MRBTS.NRBTS.NRCELL.operationalState"
                                                      ],
                                                      errors="coerce"
                                                )
                                                .map({
                                                      0: "Disabled",
                                                      1: "Enabled"
                                                })
                                                .fillna(
                                                      temp_5g_compare[
                                                            "MRBTS.NRBTS.NRCELL.operationalState"
                                                      ]
                                                )
                                          )

                                    temp_5g_compare["NRBTS"] = (
                                          temp_5g_compare["NRBTS"]
                                          .astype(str)
                                          .str.replace(r"\.0", "", regex=True)
                                          .str.strip()
                                    )

                                    compare_5g_rows = []

                                    for _, row in old_new_df.iterrows():
                                          

                                          old_nrbts = str(
                                                row.get("old_5g_nrbts", "")
                                          ).replace(".0", "").strip()

                                          new_nrbts = str(
                                                row.get("new_5g_nrbts", "")
                                          ).replace(".0", "").strip()

                                          # ---------------- Current Circle Filter ----------------
                                          circle_nrbts = set(
                                                circle_df["NRBTS"]
                                                .astype(str)
                                                .str.replace(".0", "", regex=False)
                                                .str.strip()
                                          )

                                          # skip other circle data
                                          if (
                                                old_nrbts not in circle_nrbts
                                                and
                                                new_nrbts not in circle_nrbts
                                          ):
                                                continue
                                          # find old row
                                          old_row = temp_5g_compare[
                                                temp_5g_compare["NRBTS"] == old_nrbts
                                          ]

                                          # find new row
                                          new_row = temp_5g_compare[
                                                temp_5g_compare["NRBTS"] == new_nrbts
                                          ]

                                          if old_row.empty and new_row.empty:
                                                continue

                                          old_dict = {}
                                          new_dict = {}

                                          # old data
                                          if not old_row.empty:
                                                old_dict = {
                                                      f"{col}_old": val
                                                      for col, val in old_row.iloc[0].to_dict().items()
                                                }

                                          # new data
                                          if not new_row.empty:
                                                new_dict = {
                                                      f"{col}_new": val
                                                      for col, val in new_row.iloc[0].to_dict().items()
                                                }

                                          compare_5g_rows.append({
                                                **old_dict,
                                                **new_dict
                                          })

                                    if compare_5g_rows:

                                          compare_5g_df = pd.DataFrame(
                                                compare_5g_rows
                                          )
                                          compare_5g_df["Lock Status"] = (
                                                compare_5g_df.apply(
                                                      compare_lock_status_5g,
                                                      axis=1
                                                )
                                          )
                                          compare_5g_df.to_excel(
                                                writer,
                                                sheet_name="5G Old vs New",
                                                index=False
                                          )
                  # Format excel
            # format_of_excel(full_output_path)
            try:
                  format_of_excel(full_output_path)
            except Exception as e:
                  print(f"Formatting failed: {e}")

            saved_files.append({
                  "circle": circle,
                  "path": full_output_path
            })

      print("Saved Circle Wise Files:", saved_files)
      # ---------------- Create ZIP of OUTPUT Folder ----------------
      zip_filename = "OUTPUT.zip"

      zip_path = os.path.join(
            output_root,
            zip_filename
      )

      # delete old zip
      if os.path.exists(zip_path):
            os.remove(zip_path)

      # create zip from OUTPUT folder
      shutil.make_archive(
            base_name=os.path.splitext(zip_path)[0],
            format="zip",
            root_dir=final_output_folder
      )

      print("ZIP Created:", zip_path)

      # ---------------- Download Link ----------------
      relative_url = (
            f"NOKIA_OUTPUT/{zip_filename}"
      )

      download_link = request.build_absolute_uri(
            urljoin(
                  MEDIA_URL,
                  relative_url
            )
      )

      # ---------------- Response ----------------
      return Response(
            {
                  "status": True,
                  "message": "alarm files processed successfully",
                  "files_saved": len(saved_files),
                  "circles": [
                        x["circle"]
                        for x in saved_files
                  ],
                  "download_url": download_link
            },
            status=status.HTTP_200_OK
      )
      
      
      
