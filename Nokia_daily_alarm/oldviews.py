from django.shortcuts import render
from django.db.models import *
from rest_framework.decorators import api_view
from rest_framework.response import Response
import pandas as pd
import os
from django.conf import settings
from rest_framework import status
import numpy as np
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from rest_framework.status import HTTP_200_OK
from .models import NokiaAlarm
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font,Border,Side
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin
import re




def format_of_excel(file_path):
      wb = load_workbook(file_path)
      ws = wb.active

      # ---------- Styles ----------
      header_fill = PatternFill(start_color="E87529", end_color="E87529", fill_type="solid")
      alarm_header_fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")

      header_font = Font(color="000000", bold=True)

      thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
      )

      # ---------- Header Formatting ----------
      for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

      ws.row_dimensions[1].height = 40  # visible header

      header_list = [cell.value for cell in ws[1]]

      lb_cols = [
            "Site Id","MS2 Status","Circle","Partner Name","Old/New",
            "OEM","On-air Date","On-air month","MS2 status",
            "ip","MRBTS"
      ]

      for col_name in lb_cols:
            if col_name in header_list:
                  col_idx = header_list.index(col_name) + 1
                  cell = ws.cell(row=1, column=col_idx)
                  cell.fill = alarm_header_fill
                  cell.font = header_font

      # ---------- Data Borders ----------
      for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                  cell.border = thin_border

      # ---------- UNIFORM COLUMN WIDTH ----------
      DEFAULT_WIDTH = 25

      for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = DEFAULT_WIDTH

      wb.save(file_path)



@api_view(['GET', 'POST'])
def fileupload(request):
      if request.method == 'GET':
            alarms = NokiaAlarm.objects.all().values('id', 'SA', 'NSA')
            return Response(list(alarms), status=status.HTTP_200_OK)

      elif request.method == 'POST':
            file = request.FILES.get('file')

            if not file:
                  return Response({'error': 'SA/NSA file is required!'}, status=status.HTTP_400_BAD_REQUEST)

            # Read Excel or CSV file
            try:
                  if file.name.endswith('.csv'):
                        df = pd.read_csv(file)
                  else:
                        df = pd.read_excel(file)

                        df = df.dropna(how="all")

                  NokiaAlarm.objects.all().delete() 
            except Exception as e:
                  return Response({'error': f'Error reading file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

            # Expecting at least one of these columns
            if not {'SA', 'NSA'}.intersection(df.columns):
                  return Response({'error': 'File must contain SA or NSA column.'}, status=status.HTTP_400_BAD_REQUEST)

            # Save each row to database
            count = 0
            for _, row in df.iterrows():
                  NokiaAlarm.objects.create(
                  SA=row.get('SA'),
                  NSA=row.get('NSA')
                  )
                  count += 1

            return Response({"status": True,'message': f'{count} alarms saved successfully!'}, status=status.HTTP_201_CREATED)
      
@api_view(['POST'])
def alarmfileUpload(request):
      alarm_files = request.FILES.getlist('alarm_file')
      mapping_file = request.FILES.get('mapping_file')

      if not alarm_files or not mapping_file:
            return Response(
                  {'error': 'Both alarm_file and mapping_file are required!'},
                  status=status.HTTP_400_BAD_REQUEST
            )

      # ---------------- Read Alarm Files ----------------
      alarm_dfs = []
      for alarm_file in alarm_files:
            try:
                  if alarm_file.name.endswith('.csv'):
                        df = pd.read_csv(alarm_file)
                  elif alarm_file.name.endswith(('.xlsx', '.xls')):
                        df = pd.read_excel(alarm_file)
                  else:
                        continue

                  df.columns = df.columns.str.strip()
                  alarm_dfs.append(df)

            except Exception as e:
                  return Response(
                  {'error': f'Error reading alarm file {alarm_file.name}: {str(e)}'},
                  status=status.HTTP_400_BAD_REQUEST
                  )

      if not alarm_dfs:
            return Response(
                  {'error': 'No valid alarm files found'},
                  status=status.HTTP_400_BAD_REQUEST
            )

      df_alarm = pd.concat(alarm_dfs, ignore_index=True)

    # ---------------- Read Mapping File ----------------
      try:
            if mapping_file.name.endswith('.csv'):
                  df_map = pd.read_csv(mapping_file)
            else:
                  df_map = pd.read_excel(mapping_file)
            df_map.columns = df_map.columns.str.strip()
      except Exception as e:
            return Response(
                  {'error': f'Error reading mapping file: {str(e)}'},
                  status=status.HTTP_400_BAD_REQUEST
            )

    # ---------------- Ensure MRBTS ----------------
      if "MRBTS" not in df_alarm.columns:
            if "Distinguished Name" in df_alarm.columns:
                  df_alarm["MRBTS"] = (
                  df_alarm["Distinguished Name"]
                  .astype(str)
                  .str.extract(r"MRBTS-(\d+)")
                  )
            else:
                  df_alarm["MRBTS"] = ""

      needed_cols = [
            "MRBTS",
            "Supplementary Information",
            "Origin Alarm Time",
            "Origin Alarm Update Time",
      ]
      df_alarm = df_alarm[[c for c in needed_cols if c in df_alarm.columns]]

      # ---------------- Load SA / NSA from DB ----------------
      sa = list(NokiaAlarm.objects.exclude(SA__isnull=True).values_list('SA', flat=True))
      nsa = list(NokiaAlarm.objects.exclude(NSA__isnull=True).values_list('NSA', flat=True))
      combined_alarms = set(sa + nsa)

      # ---------------- Timeout Regex ----------------
      timeout_regex = re.compile(r"Timeout\s+connecting\s+to|No\s+route\s+to\s+host|CommunicationTimeout|CommunicationTimeout", re.IGNORECASE)

    # ---------------- Classify Alarms ----------------
      def classify_alarms(info):
            if pd.isna(info):
                  return pd.Series(["", ""])

            alarms = [a.strip() for a in str(info).split(',') if a.strip()]

            service = []
            effective = []

            for alarm in alarms:
            # ✅ Timeout check
                  if timeout_regex.search(alarm):
                        service.append(alarm)  # Change to effective.append(alarm) if needed in NSA
                  elif alarm in combined_alarms:
                        service.append(alarm)
                  else:
                        effective.append(alarm)

                  return pd.Series([
                  ', '.join(service),
                  ', '.join(effective)
                  ])

      df_alarm[["Service Affecting Alarms", "Effective Alarms"]] = (
            df_alarm["Supplementary Information"].apply(classify_alarms)
      )

    # ---------------- Group & Deduplicate ----------------
      df_alarm = (
            df_alarm.drop_duplicates()
            .groupby("MRBTS", as_index=False)
            .agg(lambda x: ', '.join(
                  sorted(set(str(i) for i in x if pd.notna(i) and str(i).strip()))
            ))
      )


      # df_alarm[["Service Affecting Alarms", "Effective Alarms"]] = (
      #       df_alarm[["Service Affecting Alarms", "Effective Alarms"]]
      #       .replace('', 'No Alarms')
      # )

      df_alarm["Service Affecting Alarms"] = df_alarm["Service Affecting Alarms"].apply(
      lambda x: "Site Down"
      if timeout_regex.search(str(x))
      else ("No Alarms" if pd.isna(x) or str(x).strip() == "" else x)
      )



      df_alarm["Alarm Status (Yes/No)"] = df_alarm["Service Affecting Alarms"].apply(
            lambda x: "Yes" if x != "No Alarms" else "No"
      )

      sa_set = set(sa)
      nsa_set = set(nsa)


      df_alarm["No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms"] = (
      df_alarm["Service Affecting Alarms"].apply(
            lambda x: (
                  "SA" if str(x).strip().lower() == "site down"
                  else "SA" if any(a.strip() in sa_set for a in str(x).split(','))
                  else "NSA" if any(a.strip() in nsa_set for a in str(x).split(','))
                  else "No Alarms"
            )
      )
      )


    # ---------------- Normalize MRBTS ----------------
      df_alarm["MRBTS"] = df_alarm["MRBTS"].astype(str).str.replace('.0', '', regex=False)
      df_map["MRBTS"] = df_map["MRBTS"].astype(str).str.replace('.0', '', regex=False)

      df_map.rename(
            columns={"Service Affecting Alarms 1": "Service Affecting Alarms"},
            inplace=True
      )

      # ---------------- Merge ----------------
      df_merged = pd.merge(
            df_map,
            df_alarm,
            on="MRBTS",
            how="left",
            suffixes=('', '_alarm')
      )

      # ---------------- Update Existing Columns Only ----------------
      cols_to_update = [
            "Alarm Status (Yes/No)",
            "Service Affecting Alarms",
            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms"
      ]

      for col in cols_to_update:
            alarm_col = f"{col}_alarm"
            if alarm_col in df_merged.columns:
                  df_merged[col] = df_merged[col].mask(
                  df_merged[col].isna() | (df_merged[col].astype(str).str.strip() == ""),
                  df_merged[alarm_col]
                  )
                  df_merged.drop(columns=[alarm_col], inplace=True)
      
      

    # ---------------- Add Extra Columns If Missing ----------------
      extra_cols = [
            # "Effective Alarms",
            "Supplementary Information",
            "Origin Alarm Time",
            "Origin Alarm Update Time"
      ]

      for col in extra_cols:
            alarm_col = f"{col}_alarm"
            if alarm_col in df_merged.columns:
                  if col not in df_merged.columns:
                        df_merged[col] = df_merged[alarm_col]
                  df_merged.drop(columns=[alarm_col], inplace=True)
      
      if "Effective Alarms" in df_merged.columns:
            df_merged.drop(columns=["Effective Alarms"], inplace=True)


      # ---------------- Save Output ----------------
      output_path = os.path.join(MEDIA_ROOT, "NOKIA_OUTPUT")
      os.makedirs(output_path, exist_ok=True)
      output_filename = "Nokia_SA_NSA_OUTPUT.xlsx"
      full_output_path = os.path.join(output_path, output_filename)
      df_merged.to_excel(full_output_path, index=False)
      format_of_excel(full_output_path)

      # Build absolute download link robustly
      relative_url = f"NOKIA_OUTPUT/{output_filename}"
      download_link = request.build_absolute_uri(urljoin(MEDIA_URL, relative_url))

      return Response(
            {
                  "status": True,
                  "message": "alarm files processed successfully",
                  "download_url": download_link
            },
            status=status.HTTP_200_OK
      )