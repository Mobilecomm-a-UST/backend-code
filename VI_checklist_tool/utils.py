import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter
import re

def create_config_report(ws, data):

    # Header
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
        name="Calibri",
        size=10
    )

    # Alternate rows
    stripe_fill = PatternFill(
        start_color="F3F6F8",
        end_color="F3F6F8",
        fill_type="solid"
    )

    # NA cells
    yellow_fill = PatternFill(
        start_color="FFD966",
        end_color="FFD966",
        fill_type="solid"
    )

    # OK
    ok_fill = PatternFill(
        start_color="00B050",
        end_color="00B050",
        fill_type="solid"
    )

    ok_font = Font(
        color="000000",
        name="Calibri",
        bold=True,
        size=10
    )

    # NOT OK
    notok_fill = PatternFill(
        start_color="E43C3C",
        end_color="E43C3C",
        fill_type="solid"
    )

    notok_font = Font(
        color="000000",
        name="Calibri",
        bold=True,
        size=10
    )

    normal_font = Font(
        name="Calibri",
        size=10
    )

    first_col_font = Font(
        name="Calibri",
        size=10
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    thin = Side(style="thin", color="D9D9D9")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    num_cols = len(data[0])

    # Header row
    for col_idx, header in enumerate(data[0], 1):

        cell = ws.cell(
            row=1,
            column=col_idx,
            value=header
        )

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 25

    # Freeze header
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row_data in enumerate(data[1:], 2):

        for col_idx, value in enumerate(row_data, 1):

            cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value=value
            )

            cell.alignment = center
            cell.border = border

            # Alternate row color
            if row_idx % 2 == 0:
                cell.fill = stripe_fill

            if col_idx == 1:
                cell.font = first_col_font
            else:
                cell.font = normal_font

            # Status column
            if col_idx == num_cols:

                status = str(value).strip().upper()

                if status == "OK":
                    cell.fill = ok_fill
                    cell.font = ok_font

                elif status in ["NOT OK", "NOT FOUND"]:
                    cell.fill = notok_fill
                    cell.font = notok_font

            # NA highlight
            if str(value).strip().upper() == "NA":
                cell.fill = yellow_fill

        ws.row_dimensions[row_idx].height = 20

    # Auto-fit columns
    for col in ws.columns:

        max_length = 0
        column_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 4


def normalize_value(v):
    if pd.isna(v):
        return ""

    v = str(v).strip().lower()

    if v in ["true", "1"]:
        return "1"

    if v in ["false", "0"]:
        return "0"

    return v


def check_status(actual, expected):

    if pd.isna(actual):
        actual = ""

    if pd.isna(expected):
        expected = ""

    # -------- Actual values --------
    actual_values = set()

    for val in re.split(r"[;,/]", str(actual)):   # <-- / added
        val = re.sub(r"\s+", "", val.strip().lower())

        if not val:
            continue

        if val in ["true", "1"]:
            actual_values.update(["true", "1"])

        elif val in ["false", "0"]:
            actual_values.update(["false", "0"])

        else:
            actual_values.add(val)

    # -------- Expected values --------
    expected_values = set()

    for item in re.split(r"[,;/]", str(expected)):   # <-- / added

        item = item.strip()

        if not item:
            continue

        # Remove TDD: / FDD:
        if ":" in item:
            val = item.split(":", 1)[1].strip()
        else:
            val = item

        val = re.sub(r"\s+", "", val.lower())

        if val in ["true", "1"]:
            expected_values.update(["true", "1"])

        elif val in ["false", "0"]:
            expected_values.update(["false", "0"])

        else:
            expected_values.add(val)

    if expected_values in ({"0db"}, {"3db"}):
        if actual_values == expected_values:
            return "OK"
        else:
            return "NOT OK"   

    if len(actual_values) > 1 and len(expected_values) == 1:
        if actual_values != expected_values:
            return "NOT OK"    
             
    # Normal matching for all other parameters
    if actual_values.intersection(expected_values):
        return "OK"
    return "NOT OK"

parameter_map = {
    "actCoMp": "fixedULCoMp",
    "actDLCAggr": "TRUE",
    "actFlexScellSelect": "TRUE",
    "actTtiBundling": "TDD:0, FDD:1",
    "actDlMuMimo": "1",
    "actFastMimoSwitch": "1",
    "actMMimo": "1",
    "actUlMuMimo": "1",
    "actUlpcMethod": "TDD:PuschCLSrsPucchCL, FDD:PuschIAwPucchCL",
    "actVoipCovBoostEnh": "TDD:off, FDD:on_ttib_entry_PHR_based",
    "dlCaMinPcellCqiQci1": "0",
    "harqMaxTrUlTtiBundling": "20",
    "ttibOperMode": "sinrBasedNotWithMeasGap",
    "ulsSchedMethod": "TDD:channelaware, FDD:interference aware",
    "b2Threshold2RssiGERANQci1": "15",
    "qrxlevmin": "-124",
    "dlRsBoost-FDD":"0dB",
    "dlRsBoost-TDD":"3dB",
    "actLBPowerSaving": "1",
    "lbpsLastCellMinLoad": "40",
    "lbpsLastCellRTXMinLoad": "40",
    "lbpsLastCellSOEnabled": "0",
    "lbpsMaxLoad": "60",
    "lbpsMinLoad": "40",
    "lbpsPdcchLoadOffset": "20",
    "lbpsRTXCellCapOffset": "50",
    "lbpsRTXMaxLoad": "60",
    "lbpsRTXMinLoad": "40",
    "mdtxPdcchSymb": "1",
    "lbpsDayOfWeek": "Sun;Mon;Tue;Wed;Thu;Fri;Sat",
    "lbpsDuration": "360;360;360;360;360;360;360",
    "lbpsStartTimeHour": "0;0;0;0;0;0;0",
    "lbpsStartTimeMinute": "0;0;0;0;0;0;0",
    "lbpsSuspended": "notSuspended",
    "actRadioFanFailureEarlyNotif": "true",
    "rrcGuardTimer":"60",
    "validate_ntp": "The NTP IP should be present inside the reference sheet and it should match with the XML data.",
    "ntpStatus": "Available",
    "mdtxAggressiveness":"8",
    "actAutoPucchAllo":"1",
    "actMicroDtx":"1",
    "actAutoPucchAlloc":"FDD:1,TDD:0",
    "allowTrafficConcentration":"1",
    "alVoltHighThreshold":"58/107",
    "alVoltLowThreshold":"40.5/-100",
    "alVoltUnstableThreshold":"5",
    "ttibSinrThresholdIn":"3",
    "scellFastSchedulingSelect":"fast",
    "userLabel-IPIF-1": "LTE_UP",
    "userLabel-IPIF-2": "LTE_CP",
    "userLabel-IPIF-3": "LTE_OM",
    "userLabel-IPIF-5": "GSM_ABIS",
    "userLabel(SR)-IPRT-1":	"LTE_UP",
    "userLabel(SR)-IPRT-2":"LTE_CP",
    "userLabel(SR)-IPRT-3":"LTE_OM",
    "userLabel(SR)-IPRT-5":"GSM_ABIS",
    "userLabel(SR)-IPRTV6-1":"LTE_UP",	
   "userLabel(SR)-IPRTV6-2":"LTE_CP",
   "userLabel(SR)-IPRTV6-3":"LTE_OM",
   "userLabel(SR)-IPRTV6-5":"GSM_ABIS",
    "qciTab6schedulWeight":"40",
    "qciTab7schedulWeight":"40",
    "qciTab8schedulWeight":"40",
    "qciTab9schedulWeight":"40",
    "qciTab6schedulPrio":"10",
    "qciTab7schedulPrio":"10",
    "qciTab8schedulPrio":"10",
    "qciTab9schedulPrio":"10",

    "n310Qci1":	"n4",
    "t301":	"400ms",
    "t310Qci1":	"500ms",
    "t311":	"10000ms",
    "actNwReqUeCapa":"TRUE"



}

master_parameters = [

    "actAutoPucchAlloc",
    "actCoMp",
    "actDLCAggr",
    "actDlMuMimo",
    "actFastMimoSwitch",
    "actFlexScellSelect",
    "actLBPowerSaving",
    "actMMimo",
    "actMicroDtx",
    "actRadioFanFailureEarlyNotif",
    "actTtiBundling",
    "actUlpcMethod",
    "actVoipCovBoostEnh",
    "allowTrafficConcentration",
    "b2Threshold2RssiGERANQci1",
    "dlCaMinPcellCqiQci1",
    "dlRsBoost-FDD",
    "dlRsBoost-TDD",
    "harqMaxTrUlTtiBundling",
    "lbpsDayOfWeek",
    "lbpsDuration",
    "lbpsLastCellMinLoad",
    "lbpsLastCellRTXMinLoad",
    "lbpsLastCellSOEnabled",
    "lbpsMaxLoad",
    "lbpsMinLoad",
    "lbpsPdcchLoadOffset",
    "lbpsRTXCellCapOffset",
    "lbpsRTXMaxLoad",
    "lbpsRTXMinLoad",
    "lbpsStartTimeHour",
    "lbpsStartTimeMinute",
    "lbpsSuspended",
    "mdtxAggressiveness",
    "mdtxPdcchSymb",
    "qrxlevmin",
    "rrcGuardTimer",
    "ttibOperMode",
    "ulsSchedMethod",
    "alVoltHighThreshold",
    "alVoltLowThreshold",
    "alVoltUnstableThreshold",
    "ttibSinrThresholdIn",
    "scellFastSchedulingSelect",
     "userLabel-IPIF-1",
    "userLabel-IPIF-2",
    "userLabel-IPIF-3",
    "userLabel-IPIF-5", 
    "userLabel(SR)-IPRT-3",
    "userLabel(SR)-IPRT-5",
    "userLabel(SR)-IPRTV6-1",	
    "userLabel(SR)-IPRTV6-2",
    "qciTab6schedulWeight",
    "qciTab7schedulWeight",
    "qciTab8schedulWeight",
    "qciTab9schedulWeight",
    "qciTab6schedulPrio",
    "qciTab7schedulPrio",
    "qciTab8schedulPrio",
    "qciTab9schedulPrio",
    "n310Qci1",
    "t301",
    "t310Qci1",
    "t311",
    "actNwReqUeCapa",

	

]