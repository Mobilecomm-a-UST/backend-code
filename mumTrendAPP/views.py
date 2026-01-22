from django.shortcuts import render
from django.http import HttpResponse
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from rest_framework.decorators import api_view
from rest_framework.response import Response
from tkinter import *
from django.core.files.storage import FileSystemStorage
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl import workbook,load_workbook
import pandas as pd
import numpy as np
from datetime import date, timedelta
import datetime
import pandas as pd
import os
import threading
# from Merged_APP.views import merge_App


def merge_dataframe(dict_list):
    list = ["4G_raw_kpi","4G_site_list"]

    for x in list:
        if x in dict_list.keys():
            dict_list.pop(x)


    dfs = []
    i=0
    for name,file in dict_list.items():
        print("doning...",i)
        i=i+1
        df = pd.read_csv(file)
        dfs.append(df)
    merge_df = pd.concat(dfs, ignore_index=True)

    return merge_df

@api_view(['POST'])
def old_mum_trend(request):
    
    # raw_kpi = request.FILES["4G_raw_kpi"] if '4G_raw_kpi' in request.FILES else None
        
    # site_list = request.FILES["4G_site_list"] if '4G_site_list' in request.FILES else None
    offered_date = request.POST.get("offered_date")
    print(offered_date)
    offered_date=datetime.datetime.strptime(offered_date,"%Y-%m-%d").date()
    print("offered_date:",offered_date)

    location= MEDIA_ROOT + r"\trends\temporary_files"
    fs = FileSystemStorage(location=location)

    # raw_kpi = fs.save(raw_kpi.name, raw_kpi)
    #     # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
    # filepath = fs.path(raw_kpi)
    # print("file_path:-",filepath)
    # df_raw_kpi=pd.read_excel(filepath)
    # # df_raw_kpi=pd.read_csv(filepath)
    # os.remove(path=filepath)
    # print(filepath,"deleted...............")
    # print(df_raw_kpi)

    file_dict=request.FILES
    print(file_dict)
    df_raw_kpi=merge_dataframe(file_dict)



    # site_list = fs.save(site_list.name, site_list)
    #     # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
    # filepath = fs.path(site_list)
    # print("file_path:-",filepath)
    # df_site_list=pd.read_excel(filepath)
    # os.remove(path=filepath)
    # print(filepath,"deleted...............")
    # print(df_site_list)


    door_root= os.path.join(MEDIA_ROOT,'trends',"mum")

    

    kpi=["Radio NW Availability",
        "4G Data Volume [GB]",
        "RRC Setup Success Rate [CDBH]",
        "RRC Fails [CDBH]",
        "RRC Attempts [CDBH]",
        "ERAB Setup Success Rate [CDBH]",
        "ERAB Fails [CDBH]",
        "ERAB Setup Attempts [CDBH]",
        "PS Drop Call Rate % [CDBH]",
        "PS Drop Call Rate NOM [CDBH]",
        "DL User Throughput_Kbps [CDBH]",
        "DL User Throughput_Kbps",
        "UL User Throughput_Kbps [CDBH]",
        "Average number of used DL PRBs [CDBH]",
        "Average number of used UL PRBs [CDBH]",
        "Avg Connected User [CDBH]",
        "Max Connected User [CDBH]",
        "E-UTRAN Average CQI [CDBH]",
        "E-UTRAN Average CQI ",
        "UL PUCCH SINR [CBBH]",
        "Average UE Distance_KM [CDBH]",
        "TA Sample <500M [CDBH]",
        "RSRP Samples<-116 dBm [CDBH]",
        "PS handover success rate [LTE Inter System] [CDBH]",
        "PS handover success rate [LTE Intra System] [CDBH]",
        "PS handover success rate [LTE Inter System]_DENOM [CDBH]",
        "PS handover success rate [LTE Inter System]_NOM [CDBH]",
        "UL RSSI [CDBH]",
        "UL NI [RSSI-SINR] [CDBH]",
        "VoLTE Traffic",
        "VoLTE DCR [CBBH]",
        "VoLTE DCR_Nom [CBBH]",
        "VoLTE ERAB Setup Success Rate [CBBH]",
        "VoLTE CSSR Fail Nom [CBBH]",
        "VoLTE InterF HOSR Exec [CBBH]",
        "VoLTE IntraF HOSR Exec [CBBH]",
        "VoLTE InterF HOSR Exec_Denom [CBBH]",
        "VoLTE InterF HOSR Exec_Nom [CBBH]",
        "VoLTE SRVCC SR [CBBH]",
        "VoLTE Packet Loss DL [CBBH]",
        "VoLTE Packet Loss UL [CBBH]",
        "VoLTE Bler DL [CBBH]",
        "VoLTE Bler UL [CBBH]",
        "PS Inter System handover success Fail Nom [CDBH]",
        "PS Intra System handover success Fail Nom [CDBH]",
        "VoLTE InterF HOSR Exec Fail Nom [CBBH]",
        "VoLTE IntraF HOSR Exec Fail Nom [CBBH]",
]


# kpi_path=r"E:\Mcom_Projects_files\Trend\Mum Trend New\MU 4G TOOL KPI.xlsx"
# df_raw_kpi=pd.read_excel(kpi_path)

    # df_raw_kpi

    df_raw_kpi["Short name"]=df_raw_kpi["Short name"].fillna(method='ffill')
    df_raw_kpi.columns.values[1]='Date'
    df_raw_kpi

    df_raw_kpi["Site ID"]=df_raw_kpi["4G_ECGI"].str.split("-").str[2]
    df_raw_kpi["LNBTS ID"]=df_raw_kpi["4G_ECGI"].str.split("-").str[-1]
    df_raw_kpi["CELL ID"]=df_raw_kpi["Site ID"] + "-" +df_raw_kpi["LNBTS ID"]
    df_raw_kpi.fillna(value=0,inplace=True)
    df_raw_kpi["PS Inter System handover success Fail Nom [CDBH]"]=df_raw_kpi["PS handover success rate [LTE Inter System]_DENOM [CDBH]"]-df_raw_kpi["PS handover success rate [LTE Inter System]_NOM [CDBH]"]
    df_raw_kpi["PS Intra System handover success Fail Nom [CDBH]"]=df_raw_kpi["PS handover success rate [LTE Intra System]_DENOM [CDBH]"]-df_raw_kpi["PS handover success rate [LTE Intra System]_NOM [CDBH]"]
    df_raw_kpi["VoLTE InterF HOSR Exec Fail Nom [CBBH]"]=df_raw_kpi["VoLTE InterF HOSR Exec_Denom [CBBH]"]-df_raw_kpi["VoLTE InterF HOSR Exec_Nom [CBBH]"]
    df_raw_kpi["VoLTE IntraF HOSR Exec Fail Nom [CBBH]"]=df_raw_kpi["VoLTE IntraF HOSR Exec_Denom [CBBH]"]-df_raw_kpi["VoLTE IntraF HOSR Exec_Nom [CBBH]"]
    df_raw_kpi["VoLTE CSSR Fail Nom [CBBH]"] = df_raw_kpi["VoLTE CSSR_Denom [CDBH]"] - df_raw_kpi["VoLTE CSSR_Nom [CBBH]"]

    # df_raw_kpi.set_index(["Short name","Date","Freq_Band"],inplace=True)
    # print(df_raw_kpi)
    # df_raw_kpi.to_csv("tttttttttttttttttt.csv")
    # exit(1)




    
    
    # df_site_path=r"E:\Mcom_Projects_files\Trend\Mum Trend New\site_list.xlsx"
    # df_site_list=pd.read_excel(df_site_path)

    ################## code for site list ###########################
    # df_site_list['Site Id']=df_site_list["Site Id"].astype(str)
    # site_list=list(df_site_list["Site Id"])
    # site_list


    # df_raw_kpi_filter= df_raw_kpi[(df_raw_kpi["Site ID"].isin(site_list))]
    
    ###################### ******************* ##############################
    ########################## the below code is to replace every string to zero from numeric columns ##################
                  ############### without filtering ###############
    for x in kpi:
            df_raw_kpi[x] = df_raw_kpi[x].replace(to_replace='.*', value=0, regex=True)
            print("....................running.....................")
              
                ################ after filtering #################
    # for x in kpi:
    #         df_raw_kpi_filter[x] = df_raw_kpi_filter[x].replace(to_replace='.*', value=0, regex=True)
    #         print("..................running.....................")
    #####################################  * * * * * * * * * * * * * * ################################################

    df_raw_kpi_filter=df_raw_kpi

    # print("filtered................",df_raw_kpi_filter)
    # exit(1)
    print("piovting...")
    df_pivote=df_raw_kpi_filter.pivot_table(index=["Site ID","LNBTS ID","CELL ID","Site Name","Short name","4G_ECGI","Freq_Band","Freq_Bandwidth"],columns="Date",values=kpi)
    print("finished pivoting...")
    # savepath=os.path.join(door_root,'process output','mum_pivot.xlsx')
    # df_pivote.to_excel(savepath)

    path_of_blnk_temp=os.path.join(door_root,"template","KPI TREND SAMPLE MUMBAI.xlsx")
    trend_wb=load_workbook(path_of_blnk_temp)
    trend_ws=trend_wb.active
    print(trend_ws)

    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]

    def titleToNumber(s):
        # This process is similar to binary-to-
        # decimal conversion
            result = 0
            for B in range(len(s)):
                result *= 26
                result += ord(s[B]) - ord('A') + 1
            return result
    
    date1=offered_date
    dt1 = date1 - timedelta(1)
    dt2 = date1 - timedelta(2)
    dt3 = date1 - timedelta(3)
    dt4 = date1 - timedelta(4)
    dt5 = date1 - timedelta(5)
    cl=[dt1,dt2,dt3,dt4,dt5]
    
    def overwrite(df_pivote,kpi_name,coln1,trend_ws):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            # print(kpi_name)
            index_pivot=df_pivote.index.to_list()
            # print("index ;###############################",index_pivot)
            # print(len(index_pivot))
            # print("index of pivoted table: ",index_pivot)
           
            dr=df_pivote[kpi_name]
           

            # print("columns of dr table",dr.columns)
            cl=dr.columns.to_list()
            # print("column list",cl)
            
            # site_id=dr["SITE_ID"].to_list() 
            # cell_id=dr["CELL_ID"].to_list()    
            col1=dr[str(cl[0])].to_list()
            col2=dr[str(cl[1])].to_list()
            col3=dr[str(cl[2])].to_list()
            col4=dr[str(cl[3])].to_list()
            col5=dr[str(cl[4])].to_list()

            trend_ws[coln1+"2"].value=cl[0]
            trend_ws[coln2+"2"].value=cl[1]
            trend_ws[coln3+"2"].value=cl[2]
            trend_ws[coln4+"2"].value=cl[3]
            trend_ws[coln5+"2"].value=cl[4]

            # me=column_index_from_string(coln5)+1
            # me=get_column_letter(me)
            for i,value in enumerate(index_pivot):
                j=i+3
                trend_ws["A"+str(j)].value=index_pivot[i][0] 
                trend_ws["B"+str(j)].value=index_pivot[i][1]
                trend_ws["C"+str(j)].value=index_pivot[i][2] 
                trend_ws["D"+str(j)].value=index_pivot[i][3]
                trend_ws["E"+str(j)].value=index_pivot[i][4]
                trend_ws["F"+str(j)].value=index_pivot[i][5]
                trend_ws["G"+str(j)].value=index_pivot[i][6]
                trend_ws["H"+str(j)].value=index_pivot[i][7]

 
                trend_ws[coln1+str(j)].value=col1[i]
                trend_ws[coln2+str(j)].value=col2[i]
                trend_ws[coln3+str(j)].value=col3[i]
                trend_ws[coln4+str(j)].value=col4[i]
                trend_ws[coln5+str(j)].value=col5[i]
    print("processing start...")
    # for kpi_name in kpi:
        
    #     if(kpi_name=='Radio NW Availability'):
    #         overwrite(df_pivote,kpi_name,'I',trend_ws)
    #     if(kpi_name=='4G Data Volume [GB]'):
    #         overwrite(df_pivote,kpi_name,'O',trend_ws)   

    #     if(kpi_name=='RRC Setup Success Rate [CDBH]'):
    #         overwrite(df_pivote,kpi_name,'T',trend_ws) 

    #     if(kpi_name=='RRC Fails [CDBH]'):
    #         overwrite(df_pivote,kpi_name,'Z',trend_ws)
    #     if(kpi_name=='RRC Attempts [CDBH]'):
    #         overwrite(df_pivote,kpi_name,'AE',trend_ws)
        # if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'AJ',trend_ws)
        # if(kpi_name=='ERAB Fails [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'AP',trend_ws)
        # if(kpi_name=='ERAB Setup Attempts [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'AU',trend_ws)
        # if(kpi_name=='PS Drop Call Rate % [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'AZ',trend_ws)
        # if(kpi_name=='PS Drop Call Rate NOM [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'BF',trend_ws)
        # if(kpi_name=='DL User Throughput_Kbps [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'BK',trend_ws)
        # if(kpi_name=='DL User Throughput_Kbps'):
        #     overwrite(df_pivote,kpi_name,'BQ',trend_ws)
        # if(kpi_name=='UL User Throughput_Kbps [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'BV',trend_ws)
        # if(kpi_name=='Average number of used DL PRBs [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'CB',trend_ws)
        # if(kpi_name=='Average number of used UL PRBs [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'CG',trend_ws)
        # if(kpi_name=='Avg Connected User [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'CL',trend_ws)
        # if(kpi_name=='Max Connected User [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'CQ',trend_ws)
        # if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'CV',trend_ws)
        # if(kpi_name=='E-UTRAN Average CQI '):
        #     overwrite(df_pivote,kpi_name,'DB',trend_ws)
        # if(kpi_name=='UL PUCCH SINR [CBBH]'):
        #     overwrite(df_pivote,kpi_name,'DG',trend_ws)
        # if(kpi_name=='Average UE Distance_KM [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'DM',trend_ws)
        # if(kpi_name=='TA Sample <500M [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'DR',trend_ws)
        # if(kpi_name=='RSRP Samples<-116 dBm [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'DW',trend_ws)
        # if(kpi_name=='PS handover success rate [LTE Inter System] [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'EB',trend_ws)
        # if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'EH',trend_ws)
        # if(kpi_name=='PS Inter System handover success Fail Nom [CDBH]'):  #new
        #     overwrite(df_pivote,kpi_name,'EN',trend_ws)
        # if(kpi_name=='PS Intra System handover success Fail Nom [CDBH]'):  #new
        #     overwrite(df_pivote,kpi_name,'ES',trend_ws)
        # if(kpi_name=='UL RSSI [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'EX',trend_ws)
        # if(kpi_name=='UL NI [RSSI-SINR] [CDBH]'):
        #     overwrite(df_pivote,kpi_name,'FD',trend_ws)
        # if(kpi_name=='VoLTE Traffic'):
        #     overwrite(df_pivote,kpi_name,'FI',trend_ws)
        # if(kpi_name=='VoLTE DCR [CBBH]'):
        #     overwrite(df_pivote,kpi_name,'FN',trend_ws)
        # if(kpi_name=='VoLTE DCR_Nom [CBBH]'):
        #     overwrite(df_pivote,kpi_name,'FT',trend_ws)
        # if(kpi_name=='VoLTE ERAB Setup Success Rate [CBBH]'):
        #     overwrite(df_pivote,kpi_name,'FY',trend_ws)
        # if(kpi_name=='VoLTE CSSR Fail Nom [CBBH]'):
        #     overwrite(df_pivote,kpi_name,'GE',trend_ws)
        # if(kpi_name=='VoLTE InterF HOSR Exec [CBBH]'):
        #     overwrite(df_pivote,kpi_name,'GJ',trend_ws)
        # if(kpi_name=='VoLTE IntraF HOSR Exec [CBBH]'):
        #     overwrite(df_pivote,kpi_name,'GP',trend_ws)
        # if(kpi_name=='VoLTE InterF HOSR Exec Fail Nom [CBBH]'): #new
        #     overwrite(df_pivote,kpi_name,'GV',trend_ws)
        # if(kpi_name=='VoLTE IntraF HOSR Exec Fail Nom [CBBH]'): #new
        #     overwrite(df_pivote,kpi_name,'HA',trend_ws)
        # if(kpi_name=='VoLTE SRVCC SR [CBBH]'):
        #     overwrite(df_pivote,kpi_name,'HF',trend_ws)
        # if(kpi_name=='VoLTE Packet Loss DL [CBBH]'):
        #     overwrite(df_pivote,kpi_name,'HK',trend_ws) 
        # if(kpi_name=='VoLTE Packet Loss UL [CBBH]'):
        #     overwrite(df_pivote,kpi_name,'HQ',trend_ws)  
        # if(kpi_name=='VoLTE Bler DL [CBBH]'):
        #     overwrite(df_pivote,kpi_name,'HW',trend_ws)  
        # if(kpi_name=='VoLTE Bler UL [CBBH]'):
        #     overwrite(df_pivote,kpi_name,'IB',trend_ws)     
    
    
    t1=threading.Thread(target=overwrite,args=(df_pivote,"Radio NW Availability",'I',trend_ws))
    t2=threading.Thread(target=overwrite,args=(df_pivote,'4G Data Volume [GB]','O',trend_ws))
    t3=threading.Thread(target=overwrite,args=(df_pivote,'RRC Setup Success Rate [CDBH]','T',trend_ws))
    t4=threading.Thread(target=overwrite,args=(df_pivote,'RRC Fails [CDBH]','Z',trend_ws))
    
    t5=threading.Thread(target=overwrite,args=(df_pivote,'RRC Attempts [CDBH]','AE',trend_ws))
    t6=threading.Thread(target=overwrite,args=(df_pivote,'ERAB Setup Success Rate [CDBH]','AJ',trend_ws))
    t7=threading.Thread(target=overwrite,args=(df_pivote,'ERAB Fails [CDBH]','AP',trend_ws))
    t8=threading.Thread(target=overwrite,args=(df_pivote,'ERAB Setup Attempts [CDBH]','AU',trend_ws))

    t9=threading.Thread(target=overwrite,args=(df_pivote,'PS Drop Call Rate % [CDBH]','AZ',trend_ws))
    t10=threading.Thread(target=overwrite,args=(df_pivote,'PS Drop Call Rate NOM [CDBH]','BF',trend_ws))
    t11=threading.Thread(target=overwrite,args=(df_pivote,'DL User Throughput_Kbps [CDBH]','BK',trend_ws))
    t12=threading.Thread(target=overwrite,args=(df_pivote,'DL User Throughput_Kbps','BQ',trend_ws))
    t13=threading.Thread(target=overwrite,args=(df_pivote,'UL User Throughput_Kbps [CDBH]','BV',trend_ws))
    t14=threading.Thread(target=overwrite,args=(df_pivote,'Average number of used DL PRBs [CDBH]','CB',trend_ws))
    t15=threading.Thread(target=overwrite,args=(df_pivote,'Average number of used UL PRBs [CDBH]','CG',trend_ws))
    t16=threading.Thread(target=overwrite,args=(df_pivote,'Avg Connected User [CDBH]','CL',trend_ws))
    t17=threading.Thread(target=overwrite,args=(df_pivote,'Max Connected User [CDBH]','CQ',trend_ws))
    t18=threading.Thread(target=overwrite,args=(df_pivote,'E-UTRAN Average CQI [CDBH]','CV',trend_ws))
    t19=threading.Thread(target=overwrite,args=(df_pivote,'E-UTRAN Average CQI ','DB',trend_ws))
    t20=threading.Thread(target=overwrite,args=(df_pivote,'UL PUCCH SINR [CBBH]','DG',trend_ws))
    t21=threading.Thread(target=overwrite,args=(df_pivote,'Average UE Distance_KM [CDBH]','DM',trend_ws))
    t22=threading.Thread(target=overwrite,args=(df_pivote,'TA Sample <500M [CDBH]','DR',trend_ws))
    t23=threading.Thread(target=overwrite,args=(df_pivote,'RSRP Samples<-116 dBm [CDBH]','DW',trend_ws))
    t24=threading.Thread(target=overwrite,args=(df_pivote,'PS handover success rate [LTE Inter System] [CDBH]','EB',trend_ws))
    t25=threading.Thread(target=overwrite,args=(df_pivote,'PS handover success rate [LTE Intra System] [CDBH]','EH',trend_ws))
    t26=threading.Thread(target=overwrite,args=(df_pivote,'PS Inter System handover success Fail Nom [CDBH]','EN',trend_ws))
    t27=threading.Thread(target=overwrite,args=(df_pivote,'PS Intra System handover success Fail Nom [CDBH]','ES',trend_ws))
    t28=threading.Thread(target=overwrite,args=(df_pivote,'UL RSSI [CDBH]','EX',trend_ws))
    t29=threading.Thread(target=overwrite,args=(df_pivote,'UL NI [RSSI-SINR] [CDBH]','FD',trend_ws))
    t30=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE Traffic','FI',trend_ws))
    t31=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE DCR [CBBH]','FN',trend_ws))
    t32=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE DCR_Nom [CBBH]','FT',trend_ws))
    t33=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE ERAB Setup Success Rate [CBBH]','FY',trend_ws))
    t34=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE CSSR Fail Nom [CBBH]','GE',trend_ws))
    t35=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE InterF HOSR Exec [CBBH]','GJ',trend_ws))

    t36=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE IntraF HOSR Exec [CBBH]','GP',trend_ws))
    t37=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE InterF HOSR Exec Fail Nom [CBBH]','GV',trend_ws))
    t38=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE IntraF HOSR Exec Fail Nom [CBBH]','HA',trend_ws))
    t39=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE SRVCC SR [CBBH]','HF',trend_ws))
    t40=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE Packet Loss DL [CBBH]','HK',trend_ws))
    t41=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE Packet Loss UL [CBBH]','HQ',trend_ws))
    t42=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE Bler DL [CBBH]','HW',trend_ws))
    t43=threading.Thread(target=overwrite,args=(df_pivote,'VoLTE Bler UL [CBBH]','IB',trend_ws))
    
    t1.start()
    t2.start()
    t3.start()
    t4.start()
    t5.start()
    t6.start()
    t7.start()
    t8.start()
    t9.start()
    t10.start()
    t11.start()
    t12.start()
    t13.start()
    t14.start()
    t15.start()
    t16.start()
    t17.start()
    t18.start()
    t19.start()
    t20.start()
    t21.start()
    t22.start()
    t23.start()
    t24.start()
    t25.start()
    t26.start()
    t27.start()
    t28.start()
    t29.start()
    t30.start()
    t31.start()
    t32.start()
    t33.start()
    t34.start()
    t35.start()
    t36.start()
    t37.start()
    t38.start()
    t39.start()
    t40.start()
    t41.start()
    t42.start()
    t43.start()

    t1.join()
    t2.join()
    t3.join()
    t4.join()
    t5.join()
    t6.join()
    t7.join()
    t8.join()
    t9.join()
    t10.join()
    t11.join()
    t12.join()
    t13.join()
    t14.join()
    t15.join()
    t16.join()
    t17.join()
    t18.join()
    t19.join()
    t20.join()
    t21.join()
    t22.join()
    t23.join()
    t24.join()
    t25.join()
    t26.join()
    t27.join()
    t28.join()
    t29.join()
    t30.join()
    t31.join()
    t32.join()
    t33.join()
    t34.join()
    t35.join()
    t36.join()
    t37.join()
    t38.join()
    t39.join()
    t40.join()
    t41.join()
    t42.join()
    t43.join()
    
    print("processing finished...")
    output_path=os.path.join(door_root,"output","mum_trend_output.xlsx")
    print("saving file.....")
    trend_wb.save(output_path)
    print("file saved..")
    download_path=os.path.join(MEDIA_URL,"trends","mum","output","mum_trend_output.xlsx")
    return Response({"status":True,"message":"Succesfully uploaded","Download_url":download_path})                                                                                                                                                   
                


