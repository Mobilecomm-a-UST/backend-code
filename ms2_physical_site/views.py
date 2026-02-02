from django.shortcuts import render
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST
from pathlib import Path
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
import os
import pandas as pd
import re
import stat
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import Counter
from datetime import datetime
 
#function to format and autofit excel-------------
def format_and_autofit_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
 
    # Header formatting
    header_fill = PatternFill(start_color="215967", end_color="215967", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
 
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
 
    # Data cell formatting
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
 
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.font = bold_font
            if str(cell.value).strip().upper() == "NA":
                cell.fill = yellow_fill
 
    # Autofit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # 1-based
        column_letter = get_column_letter(column)
 
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
 
        ws.column_dimensions[column_letter].width = max_length + 2
 
    wb.save(file_path)
 
 
#function to extract data from log files----------------
def explode_data_from_log(command, start_pattern, row_pattern, end_pattern, file_content):
    command_found = False
    header_found = False
    header_values = []
    data_rows = []
    node_id = None
    ip_addr = None
    timestamp = None
 
    # Compile patterns----------------
    command_regex = re.compile(command)
    start_regex = re.compile(start_pattern)
    row_regex = re.compile(row_pattern)
    end_regex = re.compile(end_pattern)
    log_info_regex = re.compile(
        r'(\d{6}-\d{2}:\d{2}:\d{2}\+\d{4})\s+([\da-fA-F:.]+)\s+(\d+\.\d+[a-zA-Z]*)\s+([\w.-]+)\s+(stopfile=[/\w\d]+)'
    )
 
    for line in file_content:
        line = line.strip()
 
        # Detect command
        if not command_found and command_regex.match(line):
            command_found = True
            node_id = line.split('>')[0].strip()
            continue
 
        # New prompt before header or end
        if command_found and not header_found and re.match(r'^[A-Z0-9_-]+>', line):
            break
 
        # Optional log info extraction
        log_match = log_info_regex.match(line)
        if log_match:
            timestamp, ip_addr, *_ = log_match.groups()
 
        # Detect header
        if command_found and not header_found:
            header_match = start_regex.match(line)
            if header_match:
                header_found = True
                header_values = list(header_match.groups()) or header_match.group(0).split()
                continue
 
        # Detect end of block
        if command_found and header_found and end_regex.match(line):
            break
 
        # Collect row data
        if command_found and header_found and not line.startswith('==='):
            row_match = row_regex.search(line)
            if row_match:
                data_rows.append(list(row_match.groups()))
 
    if data_rows and header_values:
        df = pd.DataFrame(data_rows, columns=header_values)
        df['Node_ID'] = node_id
        df['IP_Addr'] = ip_addr
        return df
   
    if not data_rows and (node_id or ip_addr):
        return pd.DataFrame([{
            "Proxy": "",
            "Adm State": "",
            "Op. State": "",
            "MO": "",
            "Node_ID": node_id,
            "IP_Addr": ip_addr
        }])
 
    return pd.DataFrame()
 
 
def delete_existing_files(folder_path):
    if os.path.exists(folder_path):
        for file in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except Exception as e:
                print(f"Error deleting {file_path}: {e}")

#post api to upload and process files----------------
@api_view(["POST"])
def MS2_upload(request):
    final_output_df_list = []

    files = request.FILES.getlist("files")
    if not files:
        return Response({"status": "ERROR", "message": "No files uploaded"}, status=HTTP_400_BAD_REQUEST)

    base_media_url = os.path.join(MEDIA_ROOT, "MS2_physical_site")
    output_path = os.path.join(base_media_url, "MS2_Output")
    log_folder = os.path.join(base_media_url, "logs_files")
    log_excel_folder = os.path.join(base_media_url, "logs_excel")
    os.makedirs(log_folder, exist_ok=True)
    os.makedirs(output_path, exist_ok=True)
    os.makedirs(log_excel_folder, exist_ok=True)

    delete_existing_files(log_folder)
    delete_existing_files(output_path)
    delete_existing_files(log_excel_folder)

    saved_files = []
    for file in files:
        file_path = os.path.join(log_folder, file.name)
        with open(file_path, "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        saved_files.append(file_path)
 
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
 
    # Process each uploaded file
    for uploaded_file in saved_files:
        node_name = os.path.splitext(os.path.basename(uploaded_file))[0]
        excel_filename = os.path.join(log_excel_folder, f"{node_name}_{timestamp}.xlsx")
 
        with open(uploaded_file, "r") as file:
            file_content = file.readlines()
 
        # Extract DataFrame

        get0_df = explode_data_from_log(
                r'[A-Z0-9_-]+>\sget\s0',
                r'^(\d+)\s+([A-Za-z0-9=_\-]+)$',
                r'^\s*([a-zA-Z0-9]+)\s+(.+?)\s*$',
                r'^Total:\s\d+',
                file_content
        )

        # Save as Excel
        with pd.ExcelWriter(excel_filename, engine="xlsxwriter") as writer:
            if not get0_df.empty:
                get0_df.to_excel(writer, sheet_name='get0', index=False)
 
    # --- Generate final output ---
    for file in os.listdir(log_excel_folder):
        file_path = os.path.join(log_excel_folder, file)
        xls = pd.ExcelFile(file_path, engine="openpyxl")

        for sheet_name in xls.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
            if df.empty:
                continue

            if sheet_name == "get0":

                # Site ID extraction
                site_id_val = df["Node_ID"].dropna().iloc[0] if "Node_ID" in df.columns else "NA"

                # UserLabel extraction
                if "0" in df.columns:
                    userlabel_series = df.loc[df["0"] == "userLabel"]
                    if not userlabel_series.empty:
                        userlabel_val = userlabel_series.iloc[0, 1]
                    else:
                        userlabel_val = "NA"
                else:
                    userlabel_val = "NA"

                final_output_df_list.append({
                    "Site ID": site_id_val,
                    "Userlabel": userlabel_val
                })

    # Final output DataFrame
    final_output_df = pd.DataFrame(final_output_df_list)
    final_output_df.drop_duplicates(inplace=True)

    # Save final Excel File
    output_file_name = "MS2_Status_Output.xlsx"
    final_output_path = os.path.join(output_path, output_file_name)
    final_output_df.to_excel(final_output_path, index=False, engine="openpyxl")

    # Apply formatting
    format_and_autofit_excel(final_output_path)

    # Generate download URL
    relative_path = os.path.relpath(final_output_path, MEDIA_ROOT).replace("\\", "/")
    download_url = request.build_absolute_uri(MEDIA_URL + relative_path)

    return Response(
        {
            "status": True,
            "message": "Files processed successfully",
            "download_url": download_url,
        }
    )
