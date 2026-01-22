from celery import shared_task
from celery.result import AsyncResult
from .models import Task
import time
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from mailapp.tasks import send_email
from celery import shared_task
@shared_task
def print_message():
    print("Hello, Celery!")


@shared_task
def add_numbers(x, y):

    result = (x + y )
    print(f"The sum of {x} and {y} is {result}")
    time.sleep(15)
    return result

import pandas as pd
import math
import os
from openpyxl import load_workbook
def read_excel_file(file_path):
    """
    Read an Excel file and store each sheet as a DataFrame in a dictionary.
    
    Parameters:
    - file_path: str, path to the Excel file
    
    Returns:
    - sheet_data: dict, containing sheet names as keys and DataFrames as values
    """
    sheet_data = {}

    # Read the Excel file
    xls = pd.ExcelFile(file_path)

    # Iterate over each sheet in the Excel file
    for sheet_name in xls.sheet_names:
        # if sheet_name in ['GBssCSS','GHandoverControl']:
        # if sheet_name in ['GPriorityResel']:
            df = pd.read_excel(file_path, sheet_name=sheet_name,dtype=str, header=0, skiprows=lambda X : X in [1,2,3,4])
            print(sheet_name)
            sheet_data[sheet_name] = df

    return sheet_data


@shared_task
def find_2G_audit_KOL( DUMP_2G , BSC_SITE ,GPL_2G,output_file_save_name): 
    sheet_data = read_excel_file(DUMP_2G)
    GPL_data = pd.read_excel(GPL_2G, header=0,dtype=str)
    BSC_SITE_list_df = pd.read_excel(BSC_SITE, header=0,dtype=str)

    unique_mo_list=list(set(GPL_data["MO Name"]))
    #print(len(unique_mo_list))
    #print("------------------------")

    unique_mo_list = [x for x in unique_mo_list if not (isinstance(x, float) and math.isnan(x))]
    #print(len(unique_mo_list))
    print(unique_mo_list)
    # exit(0)
    output_mo_dict={}
    for mo_name in unique_mo_list:
        print(mo_name)
        try:
            mo_df=sheet_data[mo_name]
            mo_df.columns = mo_df.columns.str.lower()
        #print("----------------------------------------- -------------------------- -----------------------")
        except:
            print('mo not fount in dump')
            continue
        
        #######################     #########################
        if "GBtsSiteManager".lower() in mo_df.columns:
            print('site wise mo')
            mo_df = mo_df[mo_df.set_index(['GBssFunction'.lower(), 'GBtsSiteManager'.lower()]).index.isin(BSC_SITE_list_df.set_index(['BSC ID', 'SITE ID']).index)]
            print(mo_df)
        else:
            print('bsc wise mo')
            mo_df = mo_df[mo_df['GBssFunction'.lower()].isin(BSC_SITE_list_df['BSC ID'])]
            print(mo_df)
        #####################################################    
            
        mo_gpl_data = GPL_data[GPL_data["MO Name"]==mo_name]
        
        

        for i,raw in mo_gpl_data.iterrows():
            MO_name=raw["MO Name"]
            parameter_name = raw["parameter name"].lower()
            Suggested_Value =raw['Suggested Value']
            # #print("-----------------------------------------=====++++++++++++++++++++++++++++++++++++",additional_conditions)
            print(MO_name,parameter_name,Suggested_Value)
            mo_df[parameter_name]= Suggested_Value
            mo_df["MODIND".lower()]= "M"
            
        print("output......................",mo_df)
        output_mo_dict[mo_name]=mo_df

    excel_file = os.path.join(MEDIA_ROOT, "Audit_ZTE_HR", "2G_AUDIT", "template","2G_dump_template_kolkata.xlsx")
    
    wb = load_workbook(excel_file)
    for mo_name, mo_df in output_mo_dict.items():
        ws = wb[mo_name]
        start_row = 6
        # Convert DataFrame to Excel-compatible data (2D list)
        data = mo_df.values.tolist()

        # Paste data to Excel starting from the fourth row
        for row_index, row_data in enumerate(data, start=start_row):
            for col_index, cell_data in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index, value=cell_data)

    wb.save(os.path.join(MEDIA_ROOT, "Audit_ZTE_HR", "2G_AUDIT","KOL_outputs", output_file_save_name+".xlsx"))
    os.remove(DUMP_2G)
    os.remove(BSC_SITE)
    os.remove(GPL_2G)
        
        
    # download_path = os.path.join(MEDIA_URL, "Audit_ZTE_HR", "2G_AUDIT", "output.xlsx")
    # return Response({"status":True,"message":"Succesfully uploaded","Download_url":download_path})

@shared_task
def find_2G_audit_HRY( DUMP_2G , BSC_SITE ,GPL_2G,output_file_save_name): 
    sheet_data = read_excel_file(DUMP_2G)
    GPL_data = pd.read_excel(GPL_2G, header=0,dtype=str)
    BSC_SITE_list_df = pd.read_excel(BSC_SITE, header=0,dtype=str)

    unique_mo_list=list(set(GPL_data["MO Name"]))
    #print(len(unique_mo_list))
    #print("------------------------")

    unique_mo_list = [x for x in unique_mo_list if not (isinstance(x, float) and math.isnan(x))]
    #print(len(unique_mo_list))
    print(unique_mo_list)
    # exit(0)
    output_mo_dict={}
    for mo_name in unique_mo_list:
        print(mo_name)
        try:
            mo_df=sheet_data[mo_name]
            mo_df.columns = mo_df.columns.str.lower()
        #print("----------------------------------------- -------------------------- -----------------------")
        except:
            print('mo not fount in dump')
            continue
        
        #######################     #########################
        if "GBtsSiteManager".lower() in mo_df.columns:
            print('site wise mo')
            mo_df = mo_df[mo_df.set_index(['GBssFunction'.lower(), 'GBtsSiteManager'.lower()]).index.isin(BSC_SITE_list_df.set_index(['BSC ID', 'SITE ID']).index)]
            print(mo_df)
        else:
            print('bsc wise mo')
            mo_df = mo_df[mo_df['GBssFunction'.lower()].isin(BSC_SITE_list_df['BSC ID'])]
            print(mo_df)
        #####################################################    
            
        mo_gpl_data = GPL_data[GPL_data["MO Name"]==mo_name]
        
        

        for i,raw in mo_gpl_data.iterrows():
            MO_name=raw["MO Name"]
            parameter_name = raw["parameter name"].lower()
            Suggested_Value =raw['Suggested Value']
            # #print("-----------------------------------------=====++++++++++++++++++++++++++++++++++++",additional_conditions)
            print(MO_name,parameter_name,Suggested_Value)
            mo_df[parameter_name]= Suggested_Value
            mo_df["MODIND".lower()]= "M"
            
        print("output......................",mo_df)
        output_mo_dict[mo_name]=mo_df

    excel_file = os.path.join(MEDIA_ROOT, "Audit_ZTE_HR", "2G_AUDIT", "template","2G_dump_template.xlsx")

    
    wb = load_workbook(excel_file)
    for mo_name, mo_df in output_mo_dict.items():
        ws = wb[mo_name]
        start_row = 6
        # Convert DataFrame to Excel-compatible data (2D list)
        data = mo_df.values.tolist()

        # Paste data to Excel starting from the fourth row
        for row_index, row_data in enumerate(data, start=start_row):
            for col_index, cell_data in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index, value=cell_data)

    wb.save(os.path.join(MEDIA_ROOT, "Audit_ZTE_HR", "2G_AUDIT","HRY_outputs", output_file_save_name+".xlsx"))
    os.remove(DUMP_2G)
    os.remove(BSC_SITE)
    os.remove(GPL_2G)


@shared_task
def find_2G_audit_PNB( DUMP_2G , BSC_SITE ,GPL_2G,output_file_save_name): 
    sheet_data = read_excel_file(DUMP_2G)
    GPL_data = pd.read_excel(GPL_2G, header=0,dtype=str)
    BSC_SITE_list_df = pd.read_excel(BSC_SITE, header=0,dtype=str)

    unique_mo_list=list(set(GPL_data["MO Name"]))
    #print(len(unique_mo_list))
    #print("------------------------")

    unique_mo_list = [x for x in unique_mo_list if not (isinstance(x, float) and math.isnan(x))]
    #print(len(unique_mo_list))
    print(unique_mo_list)
    # exit(0)
    output_mo_dict={}
    for mo_name in unique_mo_list:
        print(mo_name)
        try:
            mo_df=sheet_data[mo_name]
            mo_df.columns = mo_df.columns.str.lower()
            mo_df['GBtsSiteManager'.lower()] = mo_df['ldn'].str.extract(r'GBtsSiteManager=(\d+)')
        #print("----------------------------------------- -------------------------- -----------------------")
        except:
            print('mo not fount in dump')
            continue
        
        #######################     #########################
        if "GBtsSiteManager".lower() in mo_df.columns:
            print('site wise mo')
            mo_df = mo_df[mo_df.set_index(['NE_Name'.lower(), 'GBtsSiteManager'.lower()]).index.isin(BSC_SITE_list_df.set_index(['BSC ID', 'SITE ID']).index)]
            print(mo_df)
        else:
            print('bsc wise mo')
            mo_df = mo_df[mo_df['NE_Name'.lower()].isin(BSC_SITE_list_df['BSC ID'])]
            print(mo_df)
        #####################################################    
            
        mo_gpl_data = GPL_data[GPL_data["MO Name"]==mo_name]
        
        

        for i,raw in mo_gpl_data.iterrows():
            MO_name=raw["MO Name"]
            parameter_name = raw["parameter name"].lower()
            Suggested_Value =raw['Suggested Value']
            # #print("-----------------------------------------=====++++++++++++++++++++++++++++++++++++",additional_conditions)
            print(MO_name,parameter_name,Suggested_Value)
            mo_df[parameter_name]= Suggested_Value
            mo_df["MODIND".lower()]= "M"
            
        print("output......................",mo_df)

        mo_df = mo_df.drop('GBtsSiteManager'.lower(), axis=1)
        output_mo_dict[mo_name]=mo_df

    excel_file = os.path.join(MEDIA_ROOT, "Audit_ZTE_HR", "2G_AUDIT", "template","2G_dump_template_pb.xlsx")

    
    wb = load_workbook(excel_file)
    for mo_name, mo_df in output_mo_dict.items():
        ws = wb[mo_name]
        start_row = 6
        # Convert DataFrame to Excel-compatible data (2D list)
        data = mo_df.values.tolist()

        # Paste data to Excel starting from the fourth row
        for row_index, row_data in enumerate(data, start=start_row):
            for col_index, cell_data in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index, value=cell_data)

    wb.save(os.path.join(MEDIA_ROOT, "Audit_ZTE_HR", "2G_AUDIT","PB_outputs", output_file_save_name+".xlsx"))
    os.remove(DUMP_2G)
    os.remove(BSC_SITE)
    os.remove(GPL_2G)


@shared_task
def monitor_task_status(user_name,task_id, task_pk,output_file_save_name,circle):
    # Get the Celery task result object
    result = AsyncResult(task_id)

    # Update the task status in the database
    task = Task.objects.get(pk=task_pk)
    task.status = result.status  # Update status based on Celery task status
    task.save()

    # Check if the task has been completed
    if result.ready():
        # Additional logic to handle completion of the task, if needed
        if circle == "HRY":
            file_link=os.path.join(MEDIA_URL, "Audit_ZTE_HR", "2G_AUDIT","HRY_outputs", output_file_save_name+".xlsx")
            subject = "2G Audit Report/ circle: HRY"
            
        if circle == "KOL":
             file_link=os.path.join(MEDIA_URL, "Audit_ZTE_HR", "2G_AUDIT","KOL_outputs", output_file_save_name+".xlsx")
             subject = "2G Audit Report / circle: KOL"
        if circle == "PNB":
            file_link=os.path.join(MEDIA_URL, "Audit_ZTE_HR", "2G_AUDIT","PB_outputs", output_file_save_name+".xlsx")
            subject = "2G Audit Report / circle: PB"
        task.file_link = file_link
        task.circle = circle
        task.save()
        to_address = user_name
        cc_mails = "Manoj.Kumar@ust.com;Mohit.Batra@ust.com"
        # subject = "2G Audit Report"
        body = "Your 2G Audit Report is ready. Please download the report from the below link. \n\n"+ str("http://122.176.141.197:8000") + str(file_link)
        send_email.delay(to_address, cc_mails ,subject, body)
        
        
    else:
        # If the task is not yet completed, schedule this task to run again after a certain interval
        monitor_task_status.apply_async((task_id, task_pk), countdown=10)  # Retry after 10 seconds
