from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST
from pathlib import Path
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
import os
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

base_media_url = os.path.join(MEDIA_ROOT, "Nokia2G_Script")
output_path = os.path.join(base_media_url, "Output")
os.makedirs(output_path, exist_ok=True)


def excel_formate(file_path):
    wb = load_workbook(file_path)

    header_fill = PatternFill(start_color="215967", end_color="215967", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    strip_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

 
    command_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    command_font = Font(color="333333", bold=True)

    thin_border = Side(border_style="thin", color="808080")
    full_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    MAX_WIDTH = 60

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

     
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = full_border
            cell.font=Font(size=9,bold=True,color="FFFFFF")

        ws.row_dimensions[1].height =35

        col_letter = None
        col_index = None
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == "Command":
                col_letter = col[0].column_letter
                col_index = col[0].column 
                break

    
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
             
                if col_index and cell.column == col_index:
                    cell.fill = command_fill
                    cell.font = command_font

                
                elif cell.row % 2 == 0:
                    cell.fill = strip_fill

                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = full_border

       
        for column_cells in ws.columns:
            column = column_cells[0].column_letter

            if column == col_letter:
                ws.column_dimensions[column].width = 170
                continue

            max_length = 0
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column].width = max_length + 2

     
       

    wb.save(file_path)

@api_view(['POST'])
def nokia_2g(request):
    circle = request.data.get("circle")
    ref_file = request.FILES.get("file")
    if not ref_file:
        return Response({"status": False, "message": "ref_file not uploaded"}, status=400)
    
    # work for STCP TRX 2 for two command-------
    ref_sctp_df=pd.read_excel(ref_file, sheet_name="STCP TRX 2")
    ref_sctp_df.columns = ref_sctp_df.columns.str.strip()


    ref_sctp_df["Command"] = ref_sctp_df.apply(
        lambda row: (
            f"ZOYX:{row['SCTP association name']}:{row['SCTP user']}:{row['role']}:"
            f"{row['unit']},{row['index(BCSU)']}:AFAST{row['SCTP parameter set']}:;\n"
            
            f"ZOYP:{row['SCTP user']}:{row['SCTP association name']}:"
            f"\"{row['SOURCE ADDRESS']}\",,{row['SOURCE PORT']}:"
            f"\"{row['DESTINATION ADDRESS']}\",{row['Net Mask Length']},,,{row['DESTINATION PORT']}:;\n"
            
            # f"ZOYS:{row['SCTP user']}:{row['SCTP association name']}:ACT;"
        ),
        axis=1
    )

    #work for LAPD------------------
    lapd_df=pd.read_excel(ref_file,sheet_name="LAPD")
    lapd_df["Command"]=lapd_df.apply(
        lambda raw :(
            f"ZDWP:{raw['LAPD_ID']}:{raw['unit']},{raw['index(BCSU)']}:0,{raw['SCTP parameter set']}:{raw['SCTP association name']}"
        ),axis=1
    )

    # work for BCF for command-------

    bcf_df=pd.read_excel(ref_file,sheet_name="BCF")
    bcf_df.columns = bcf_df.columns.str.strip()
    bcf_df["Command"]= bcf_df.apply(
        lambda row:( 
        f"ZEFC:{row['BCF_ID']},{row['BTS_TYPE']}:{row['DNAME']},REF={row['REF BCF'] if pd.notna(row['REF BCF']) else ''}"
        f",::VLANID={row['VLAN ID']},"
        f"BCUIP={row['BCUIP']},SMCUP={row['SMCUP']},"
        f"BMIP={row['BMIP']},SMMP={row['SMMP']},"
        f"ETPGID={row['Used ETME ID']},::;\n"

        f"ZEFM:{row['BCF_ID']}::::PACC={row['PACC']},PL1={row['PL1']},PL2={row['PL2']},DLCIR={row['DLCIR']},"
        f"ULTS={row['ULTS']},ULCIR={row['ULCIR']},ULCBS={row['ULCBS']},MBMWT={row['MBMWT']},MEMWT={row['MEMWT']},MMPS={row['MMPS']},;\n"

        f"ZEXA:LMUA={row['LMUA']}:TRE={row['TRE']},BCF={row['BCF_ID']},:"
        f"FNO={row['FNO']},LTO={row['LTO']},FMU={row['FMU']},LTD={row['LTD']},:;\n"

        f"ZEFM:{row['BCF_ID']}:CS={row['CS']},SENA={row['SENA']};\n"

        f"ZEFM:{row['BCF_ID']}::BU1={row['BU1']},BU2={row['BU2']};\n"

        f"ZEFM:{row['BCF_ID']}:CS={row['CS']};\n"

        f"ZEEI:BCF={row['BCF_ID']}:;"

        ),axis=1
    )
   
    # work for BTS for command-------
    bts_df=pd.read_excel(ref_file,sheet_name="BTS")
    bts_df.columns = bts_df.columns.str.strip()
    bts_df["Command"] = bts_df.apply(
    lambda row: "\n".join([
        f"ZEQC:BCF={row['BCF']},BTS={row['BTS_ID']},SEG={row['SEG_ID']},REF={row['MR ID']},NAME={row['NW_NAME']},SEGNAME={row['SEG_NAME']}:CI={row['CI']},BAND={row['BAND']}:NCC={row['NCC']},BCC={row['BCC']}:MCC={row['MCC']},MNC={row['MNC']},LAC={row['LAC']}::GENA={row['GENA']},RAC={row['RAC']};",

        f"ZEQA:BTS={row['BTS_ID']}:MAL={row['MAL']},MO={row['MO']},MS={row['MS']};",

        f"ZEQE:BTS={row['BTS_ID']}:HOP={row['HOP']},HSN1={row['HSN1']};",

        f"ZEUC:SEG={row['SEG_ID']},RSEG={row['MR ID']};",
        f"ZEHC:SEG={row['SEG_ID']},RSEG={row['MR ID']};",
        f"ZEHC:SEG={row['SEG_ID']},RSEG={row['MR ID']};",

        f"ZEQV:SEG={row['SEG_ID']}:GENA={row['GENA']},:PCU={row['PCU']};",
        f"ZEQV:BTS={row['BTS_ID']}:EGENA={row['EGENA']};",
        f"ZEQV:BTS={row['BTS_ID']}:CDED={row['CDED']},CDEF={row['CDEF']};",

        f"ZEFM:{row['BCF']}::T200S={row['T200S']},T200F={row['T200F']};",

        f"ZEQM:BTS={row['BTS_ID']}:RDIV={row['RXDIV']};",
        f"ZEQV:SEG={row['SEG_ID']}:BFG={row['BFG']};",

        f"ZEQM:SEG={row['SEG_ID']}:PMAX1={row['PMAX1']},PMAX2={row['PMAX2']},FRL={row['FRL']},FRU={row['FRU']},AFRL={row['AFRL']},AFRU={row['AFRU']},BMA={row['BMA']};",

        f"ZEQF:SEG={row['SEG_ID']}:PLMN={row['PLMN']},;",
        f"ZEQY:SEG={row['SEG_ID']}:AHRLT={row['AHRLT']},ARLT={row['ARLT']},;",
        f"ZEQG:SEG={row['SEG_ID']}:RLT={row['RLT']};",

        f"ZEQM:SEG={row['SEG_ID']}::::QSRI={row['QSRI']},QSRP={row['QSRP']},:::::;",
        f"ZEQM:BTS={row['BTS_ID']}:RDIV={row['RXDIV']},;",
        f"ZEQM:SEG={row['SEG_ID']}::::FDD={row['FDD']},FDM={row['FDM']},;",

        f"ZEQF:SEG={row['SEG_ID']}:FRLTE={row['FRLTE']};",
        f"ZEQM:SEG={row['SEG_ID']}:SLO={row['SLO']},CB={row['CB']},BLT={row['BLT']},NECI={row['NECI']},CABE={row['CABE']};",
        f"ZEQB:SEG={row['SEG_ID']}:IDLE={row['BAL']},ACT={row['ACT']},;",

        f"ZEQF:SEG={row['SEG_ID']}:RE={row['RE']},;",
        f"ZEQM:SEG={row['SEG_ID']}:TRP={row['TRP']};",
        f"ZEQE:BTS={row['BTS_ID']}:AHOP={row['AHOP']};",
        f"ZEQF:SEG={row['SEG_ID']}:DR={row['DR']};",
        f"ZEQJ:SEG={row['SEG_ID']}:PER={row['PER']};",
        f"ZEQV:SEG={row['SEG_ID']}:GENA={row['GENA']};"
    ]),
    axis=1
)
   
    
    trx_df=pd.read_excel(ref_file,sheet_name="TRX")
    trx_df.columns = trx_df.columns.str.strip()
    trx_df["Command"] = trx_df.apply(
    lambda row: (
        f"ZDWP:{row['LAPD_NAME']}:"
        f"{'BCSU' if 'BCSU' in row and pd.notna(row['BCSU']) else 'BCXU'},"
        f"{row['BCSU'] if 'BCSU' in row and pd.notna(row['BCSU']) else row['BCXU']}"
        f":0,1:{row['SCTP Name']},:;\n"

        f"ZERC:BTS={row['BTS_ID']},TRX={row['TRX_ID']}:"
        f"PREF={row['PREF']},GTRX={row['GTRX']},:"
        f"FREQ={row['FREQ']},TSC={row['TSC']}:"
        f"DNAME={row['LAPD_NAME']}:"
        f"CH0={row['CH_0']},CH1={row['CH_1']},CH2={row['CH_2']},CH3={row['CH_3']},"
        f"CH4={row['CH_4']},CH5={row['CH_5']},CH6={row['CH_6']},CH7={row['CH_7']},:;\n"
    ),
    axis=1
)
    
    lbs_df=pd.read_excel(ref_file,sheet_name="LBS Data")
    lbs_df.columns = lbs_df.columns.str.strip()
    lbs_df["Command"]=lbs_df.apply(
    lambda row:(
        f"ZEXC:LAC={row['LAC']},CI={row['CI']}:"
        f"FREQ={row['FREQ']},NCC={row['NCC']},BCC={row['BCC']}:"
        f"LADS={row['LADS']},LAD={row['LAD']},LAM={row['LAM']},LAS={row['LAS']},LAF={row['LAF']},"
        f"LODS={row['LODS']},LOD={row['LOD']},LOM={row['LOM']},LOS={row['LOS']},LOF={row['LOF']},AL={row['AL']}:"
        f"ABE={row['ABE']},CIT={row['CIT']},AHE={row['AHE']},AHB={row['AHB']},MRP={row['MRP']},"
        f"FSR={row['FSR']},BSR={row['BSR']};"
    ),
    axis=1
    )

    adce_df=pd.read_excel(ref_file,sheet_name="ADCE")
    adce_df.columns = adce_df.columns.str.strip()
    adce_df["Command"]=adce_df.apply(
    lambda row:(
    f"ZEAC:SEG={row['S_BTS_ID']}::"
    f"MCC={row['MCC']},MNC={row['MNC']},LAC={row['A_LAC']},CI={row['A_CI']}:"
    f"NCC={row['A_NCC']},BCC={row['A_BCC']},FREQ={row['A_FREQ']}:"
    f"SYNC={row['SYNC']},DRT={row['DRT']},SL={row['SL']},"
    f"PMRG={row['PMRG']},LMRG={row['LMRG']},QMRG={row['QMRG']},RAC={row['RAC']};"

    ),axis=1   
    )
    
    ref_sctp_df2=pd.read_excel(ref_file, sheet_name="STCP TRX 2")
    ref_sctp_df2.columns = ref_sctp_df2.columns.str.strip()
    ref_sctp_df2["Command"] = ref_sctp_df2.apply(
        lambda row: (
            f"ZOYS:{row['SCTP user']}:{row['SCTP association name']}:ACT;"
        ),
        axis=1
    )
    

    trx_df2=pd.read_excel(ref_file,sheet_name="TRX")
    trx_df2.columns = trx_df2.columns.str.strip()
    trx_df2["Command"] = trx_df2.apply(
    lambda row: (
        f"ZERS:BTS={row['BTS_ID']},TRX={row['TRX_ID']}:U;\n"
        f"ZEQS:BTS={row['BTS_ID']}:U;"

    ),
    axis=1
    )

    dmal_df = bts_df.copy()
    dmal_df["Command"] = dmal_df.apply(
        lambda row: "\n".join([
            f"ZEQA:BTS={row['BTS_ID']}:OPE=A,DMAL=41&42&43,DUMAL=100;",
            f"ZEQM:BTS={row['BTS_ID']}:::::DMOD=1;",
            f"ZEQM:BTS={row['BTS_ID']}:::::DMOD=2;",
            f"ZEQE:BTS={row['BTS_ID']}:HOP=RF;"
        ]),
        axis=1
    ) 


    trx_modification_df=pd.read_excel(ref_file,sheet_name="TRX")
    trx_modification_df.columns=trx_modification_df.columns.str.strip()
    trx_modification_df["DFCA"] = trx_modification_df["CH_0"].apply(
        lambda x: "Y" if x == "TCHD" else "N"
    )

    trx_modification_df["Command"] =trx_modification_df.apply(
    lambda row: (
           f"ZERM:BTS={row['BTS_ID']},TRX={row['TRX_ID']}:DFCA={row['DFCA']};"
    ),
    axis=1
)
    
    dfca_df=bts_df.copy()
    dfca_df["Command"] = dfca_df.apply(
    lambda row: "\n".join([
        f"ZEUB:BTS={row['BTS_ID']}:UURH=2,UURF=2,UDRH=2,UDRF=2,LURH=3,LURF=3,LDRH=3,LDRF=3;",
        f"ZEHB:BTS={row['BTS_ID']}:IHRF=2,IHRH=4,QDRH=5,QURH=4,QDRF=5,QURF=4;",
        f"ZEQF:BTS={row['BTS_ID']}:DRM=1,MADR=12,MIDR=7;",
        f"ZEQH:BTS={row['BTS_ID']}:MQL=100,MPU=Y,QPC=1,QPH=2,QPN=10;",
        f"ZEQM:SEG={row['BTS_ID']}:BMA=2;",
        f"ZEQM:BTS={row['BTS_ID']}:::::FAHT=14,FHR=5,FHT=0,FHH=6;",
        f"ZEQM:BTS={row['BTS_ID']}:FRL=99,FRU=100;",
        f"ZEHA:SEG={row['BTS_ID']}:QDW=2,QUW=2,LDW=2,LUW=2;",
        f"ZEHG:SEG={row['BTS_ID']}:MIH=3;",
        f"ZEHQ:SEG={row['BTS_ID']}:QDN=4,QDP=3,QDR=5,QUN=6,QUP=4,QUR=4;",
        f"ZEHS:SEG={row['BTS_ID']}:LDR=-110,LUR=-110;",
        f"ZEUG:SEG={row['BTS_ID']}:INT=1,RED=4;",
        f"ZEUQ:SEG={row['BTS_ID']}:LDN=1,LDP=1,LDR=3,LUN=1,LUP=1,LUR=4,UDN=1,UDP=1,UDR=2,UUN=1,UUP=1,UUR=2;",
        f"ZEUS:SEG={row['BTS_ID']}:LDR=-90,LUR=-95,UDR=-80,UUR=-47;"
    ]),
    axis=1
)  
      
    gpl_df = bts_df.copy()

    gpl_df["Command"] = gpl_df.apply(
        lambda row: "\n".join([
            f"ZEQF:SEG={row['BTS_ID']}:BAR=N,EC=N,RE=Y,DR=Y,MIDR=2,MADR=9,PLMN=0&&7,FRLTE=1;",
            f"ZEQV:BTS={row['BTS_ID']}:ALA=Y,MCA=9,MCU=9,MBG=6,MBP=6,DLPC=Y,CS34=Y;",
            f"ZEQM:SEG={row['BTS_ID']}:::::::::GPRIO=1,WPRIO=2,TIMEH=0;",
            f"ZEQY:SEG={row['BTS_ID']}:URIS=10,AURIS=10,AHURIS=10,ARLT=36,AHRLT=36;",
            f"ZEHG:SEG={row['BTS_ID']}:ESD=N,EFA=Y,EFP=Y,EFH=Y,EUM=Y,EMS=Y,HPU=6;",
            f"ZEUG:SEG={row['BTS_ID']}:TPR=2,ALIM=6,PENA=Y;",
            f"ZEQV:SEG={row['BTS_ID']}::::DENA=Y;",
            f"ZEQM:BTS={row['BTS_ID']}:ISIC=0,FRL=99,FRU=100,AFRL=99,AFRU=100,BMA=2,DTX=1,RDIV=Y,STIRC=Y;",
            f"ZEQM:SEG={row['BTS_ID']}:ESI=Y,RXLT=-102,RET=1,TRP=1,DMAX=63,PMAX2=30,FRL=99,FRU=100,AFRL=99,AFRU=100,SLO=32;",
            f"ZEQJ:SEG={row['BTS_ID']}:PER=1.5,AG=2,MFR=4;",
            f"ZEHB:BTS={row['BTS_ID']}:IHRF=2,IHRH=7;",
            f"ZEHN:SEG={row['BTS_ID']}:QSRC=7;",
            f"ZEQG:SEG={row['BTS_ID']}:RLT=32;",
            f"ZEQV:SEG={row['BTS_ID']}:GENA=Y;",
            f"ZEQV:BTS={row['BTS_ID']}:EGENA=Y;",
            f"ZEQE:BTS={row['BTS_ID']}:HOP=N,AHOP=Y;",
            f"ZEUM:BTS={row['BTS_ID']}:PCPOW=2,PCWS=4;",
            f"ZEQM:BTS={row['BTS_ID']}::::QSRI=7,QSRP=7,FDD=N,FDM=-14;",
            f"ZEHY:SEG={row['BTS_ID']}:EFHO=DIS;",
            f"ZEQV:SEG={row['BTS_ID']}:::::QPEU=7;",
            f"ZEQV:BTS={row['BTS_ID']}:CMAX=100;"
        ]),
        axis=1
    )
    
    gpl_df2 = bts_df.copy()

    gpl_df2["Command"] = gpl_df2.apply(
        lambda row: "\n".join([
            f"ZEQF:SEG={row['BTS_ID']}:BAR=N,EC=N,RE=Y,DR=Y,MIDR=2,MADR=9,PLMN=0&&7,FRLTE=1;",
            f"ZEQV:BTS={row['BTS_ID']}:ALA=Y,MCA=9,MCU=9,MBG=6,MBP=6,DLPC=Y,CS34=Y;",
            f"ZEQM:SEG={row['BTS_ID']}:::::::::GPRIO=1,WPRIO=2,TIMEH=0;",
            f"ZEQY:SEG={row['BTS_ID']}:URIS=10,AURIS=10,AHURIS=10,ARLT=36;",
            f"ZEHG:SEG={row['BTS_ID']}:ESD=N,EFA=Y,EFP=Y,EFH=Y,EUM=Y,EMS=Y,HPU=6;",
            f"ZEUG:SEG={row['BTS_ID']}:TPR=2,ALIM=6,PENA=Y;",
            f"ZEQV:SEG={row['BTS_ID']}::::DENA=Y;",
            f"ZEQM:BTS={row['BTS_ID']}:ISIC=0,FRL=99,FRU=100,AFRL=99,AFRU=100,BMA=2,DTX=1,RDIV=Y,STIRC=Y;",
            f"ZEQM:SEG={row['BTS_ID']}:ESI=Y,RXLT=-102,RET=1,TRP=1,DMAX=63,PMAX2=30,FRL=99,FRU=100,AFRL=99,AFRU=100,SLO=32;",
            f"ZEQJ:SEG={row['BTS_ID']}:PER=1.5,AG=2,MFR=4;",
            f"ZEHB:BTS={row['BTS_ID']}:IHRF=2,IHRH=7;",
            f"ZEHN:SEG={row['BTS_ID']}:QSRC=7;",
            f"ZEQG:SEG={row['BTS_ID']}:RLT=32;",
            f"ZEQV:SEG={row['BTS_ID']}:GENA=Y;",
            f"ZEQV:BTS={row['BTS_ID']}:EGENA=Y;",
            f"ZEQE:BTS={row['BTS_ID']}:HOP=N,AHOP=Y;",
            f"ZEUM:BTS={row['BTS_ID']}:PCPOW=2,PCWS=4;",
            f"ZEQM:BTS={row['BTS_ID']}::::QSRI=7,QSRP=7,FDD=N,FDM=-14;",
            f"ZEHY:SEG={row['BTS_ID']}:EFHO=DIS;",
            f"ZEQV:SEG={row['BTS_ID']}:::::QPEU=7;",
            f"ZEQV:BTS={row['BTS_ID']}:CMAX=100;"
        ]),
        axis=1
    )

    
    all_commands = []
    all_commands += [("STCP TRX 2", cmd) for cmd in ref_sctp_df["Command"].dropna()]
    all_commands += [("LAPD", cmd) for cmd in lapd_df["Command"].dropna()]
    all_commands += [("BCF", cmd) for cmd in bcf_df["Command"].dropna()]
    all_commands += [("BTS", cmd) for cmd in bts_df["Command"].dropna()]
    all_commands += [("TRX", cmd) for cmd in trx_df["Command"].dropna()]
    all_commands += [("LBS Data", cmd) for cmd in lbs_df["Command"].dropna()]
    all_commands += [("ADCE", cmd) for cmd in adce_df["Command"].dropna()]
    all_commands += [("DMAL DUMAL ATTACH", cmd) for cmd in dmal_df["Command"].dropna()]
    all_commands += [("TRX MODIFICATION", cmd) for cmd in trx_modification_df["Command"].dropna()]
    all_commands +=[("DFCA Implementation on BCF BTS", cmd) for cmd in dfca_df["Command"].dropna()]
    all_commands +=[("GPL", cmd) for cmd in gpl_df["Command"].dropna()]
    all_commands +=[("BSC174BHU GPL", cmd) for cmd in gpl_df2["Command"].dropna()]
    all_commands += [("STCP TRX 2", cmd) for cmd in ref_sctp_df2["Command"].dropna()]
    all_commands += [("TRX", cmd) for cmd in trx_df2["Command"].dropna()]
    final_df = pd.DataFrame(all_commands, columns=["Sheet Name","Command"])
    final_df["Command"] = final_df["Command"].str.replace(r"=(nan|None)\b", "=", regex=True)
    

    file_name = f"2G_Script_{circle}.xlsx"
    final_output_path = os.path.join(output_path, file_name)

    with pd.ExcelWriter(final_output_path, engine='openpyxl') as writer:
        # ref_sctp_df.to_excel(writer, sheet_name="STCP TRX 2", index=False)
        # lapd_df.to_excel(writer,sheet_name="LAPD",index=False)
        # bcf_df.to_excel(writer, sheet_name="BCF", index=False)
        # bts_df.to_excel(writer,sheet_name="BTS", index=False)
        # trx_df.to_excel(writer,sheet_name="TRX",index=False)
        # lbs_df.to_excel(writer,sheet_name="LBS Data",index=False)
        # adce_df.to_excel(writer,sheet_name="ADCE",index=False)
        final_df.to_excel(writer, sheet_name="Command", index=False)
    excel_formate(final_output_path)

    relative_path = os.path.relpath(final_output_path, MEDIA_ROOT).replace("\\", "/")
    download_url = request.build_absolute_uri(f"{MEDIA_URL}{relative_path}")

    return Response({
        "status": True,
        "message": "2G Script Generated Successfully",
        "download_url": download_url
    })
 
   

