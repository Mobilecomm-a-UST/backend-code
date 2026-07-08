from django.shortcuts import render

# Create your views here.
import os
import pandas as pd
from django.conf import settings
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
import re
from openpyxl.styles import Alignment
import operator
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL


main_folder = os.path.join(settings.MEDIA_ROOT, "kpi_monitoring")
input_path  = os.path.join(main_folder, "input")
output_path = os.path.join(main_folder, "output")
os.makedirs(input_path, exist_ok=True)
os.makedirs(output_path, exist_ok=True)

VENDOR_FIELDS = ['ericsson', 'nokia', 'samsung']


def _clear_files(return_deleted=False):
    """Remove all files from input and output folders."""
    deleted = []
    for folder in (input_path, output_path):
        if os.path.exists(folder):
            for fname in os.listdir(folder):
                fpath = os.path.join(folder, fname)
                if os.path.isfile(fpath):
                    if return_deleted:
                        deleted.append(fname)
                    os.remove(fpath)
    return deleted


def _save_uploaded_file(vendor, uploaded_file):
    """Save uploaded file with vendor name as filename so it's identifiable later."""
    ext = os.path.splitext(uploaded_file.name)[1].lower()
    save_path = os.path.join(input_path, f"{vendor}{ext}")
    with open(save_path, 'wb+') as dest:
        for chunk in uploaded_file.chunks():
            dest.write(chunk)
    return save_path


def _read_file_to_df(file_path):
    """Read an uploaded file (xlsx/xls/csv) into a DataFrame."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.csv':
        return pd.read_csv(file_path)
    return pd.read_excel(file_path)


def _get_input_files():
    """Return dict of vendor -> filename currently present in input folder."""
    found = {}
    if os.path.exists(input_path):
        for fname in os.listdir(input_path):
            name_no_ext = os.path.splitext(fname)[0].lower()
            if name_no_ext in VENDOR_FIELDS:
                found[name_no_ext] = fname
    return found


@api_view(['POST', 'GET', 'DELETE'])
def upload_file(request):
    """
    POST   → upload Ericsson, Nokia, Samsung files (xlsx/xls/csv) in one call
    GET    → check what input/output files currently exist
    DELETE → clear all input and output files
    """
    try:
        # ── POST: upload ───────────────────────────────────────────────────
        if request.method == 'POST':
            missing = [v for v in VENDOR_FIELDS if v not in request.FILES]
            if missing:
                return Response(
                    {'status': False, 'error': f"Missing file(s) for: {', '.join(missing)}."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            _clear_files()

            saved_info = {}
            for vendor in VENDOR_FIELDS:
                uploaded = request.FILES[vendor]
                ext = os.path.splitext(uploaded.name)[1].lower()
                if ext not in ('.xlsx', '.xls', '.csv'):
                    return Response(
                        {'status': False, 'error': f"Unsupported type '{ext}' for {vendor}. Use .xlsx, .xls, or .csv."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                save_path = _save_uploaded_file(vendor, uploaded)
                try:
                    df = _read_file_to_df(save_path)
                except Exception as read_err:
                    os.remove(save_path)
                    return Response(
                        {'status': False, 'error': f"Could not read {vendor} file: {str(read_err)}"},
                        status=status.HTTP_422_UNPROCESSABLE_ENTITY
                    )
                saved_info[vendor] = {
                    'filename':  uploaded.name,
                    'rows_read': len(df),
                }

            return Response({
                'status':  True,
                'message': 'Ericsson, Nokia, and Samsung files uploaded and read successfully.',
                'files':   saved_info,
            }, status=status.HTTP_200_OK)

        # ── GET: check files ───────────────────────────────────────────────
        elif request.method == 'GET':
            input_files = _get_input_files()
            input_status = {
                vendor: ({'filename': input_files[vendor], 'status': 'ready'} if vendor in input_files else None)
                for vendor in VENDOR_FIELDS
            }

            output_files = []
            if os.path.exists(output_path):
                for filename in os.listdir(output_path):
                    if filename.endswith('.xlsx'):
                        output_files.append({
                            'filename':     filename,
                            'download_url': request.build_absolute_uri(
                                f"{settings.MEDIA_URL.rstrip('/')}/kpi_monitoring/output/{filename}"
                            ),
                        })

            return Response({
                'status':       True,
                'message':      'Files found.' if input_files else 'No input files found.',
                'input_files':  input_status,
                'output_files': output_files,
            }, status=status.HTTP_200_OK)

        # ── DELETE: clear files ────────────────────────────────────────────
        elif request.method == 'DELETE':
            deleted_files = _clear_files(return_deleted=True)
            return Response({
                'status':        True,
                'message':       'Files deleted successfully.',
                'deleted_files': deleted_files,
            }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




    


CELL_NAME_COLUMN = 'Cell Name'
DATE_COLUMN_INDEX = 1
MAIN_REPORT_FILENAME = "kpi_monitoring_report.xlsx"


COMMON_METRIC_COLUMNS = [
    '5G Data Volume [GB]',
    'Session PS Drop Rate % [CDBH]',
    'SgNb addition Success Rate % [CDBH]',
    'Average CQI [CDBH]',
    'Average DL Latency [ms] [CDBH]',
    'Average DL UE Throughput_Mbps [CDBH]',
    'Average UL UE Throughput_Mbps [CDBH]',
    'Session PS Drop Rate_Denom [CDBH]',
    'Session PS Drop Rate_Nom [CDBH]',
    'SgNb addition Success Rate_Denom [CDBH]',
    'SgNb addition Success Rate_Nom [CDBH]',
    'Radio Network Availability %',
    'PDSCH Slot Utilization % [CDBH]',
    'PUSCH Slot Utilization % [CDBH]',
    'Average UE Distance [Mtr] [CDBH]',



]

# Metrics only present in specific vendors

VENDOR_METRIC_COLUMNS = {
    'ericsson': [
        ('Inter sgNodeB PSCell Change Success Rate % [CDBH]', 'Inter sgNodeB PSCell Change Success Rate % [CDBH]'),
        ('Intra sgNodeB PSCell Change Success Rate % [CDBH]','Intra sgNodeB PSCell Change Success Rate % [CDBH]'),
        ('DL BLER [CDBH]','DL BLER [CDBH]'),
        ('UL BLER [CDBH]','UL BLER [CDBH]'),
        ('DL Packet Loss [CDBH]','DL Packet Loss [CDBH]'),
        ('RACH Msg3 Success Rate % [CDBH]','RACH Msg3 Success Rate % [CDBH]'),
        ('UL Packet Loss [CDBH]','UL Packet Loss [CDBH]'),
        ('UL RSSI PUSCH [dBm] [CDBH]','UL RSSI PUSCH [dBm] [CDBH]'),
        ('Inter sgNodeB PSCell Change Success Rate_Denom [CDBH]','Inter sgNodeB PSCell Change Success Rate_Denom [CDBH]'),
        ('Inter sgNodeB PSCell Change Success Rate_Nom [CDBH]','Inter sgNodeB PSCell Change Success Rate_Nom [CDBH]'),
        ('Intra sgNodeB PSCell Change Success Rate_Denom [CDBH]','Intra sgNodeB PSCell Change Success Rate_Denom [CDBH]'),
        ('Intra sgNodeB PSCell Change Success Rate_Nom [CDBH]','Intra sgNodeB PSCell Change Success Rate_Nom [CDBH]'),
        ('Average TA [CDBH]','Average TA [CDBH]'),
        ('SA Data Volume DL SD1 [GB]','SD1 Traffic'),
    ],
    'nokia': [
        ('Inter gNodeB PSCell Change Success Rate % [CDBH]', 'Inter sgNodeB PSCell Change Success Rate % [CDBH]'),
        ('Intra gNodeB PSCell Change Success Rate % [CDBH]','Intra sgNodeB PSCell Change Success Rate % [CDBH]'),
        ('Average DL Rank [CDBH]','DL BLER [CDBH]'),
        ('Initial BLER in PUSCH % [CDBH]','UL BLER [CDBH]'),
        ('Packet Loss Rate DL % [CDBH]','DL Packet Loss [CDBH]'),
        ('Random Access Success Rate % [CDBH]','RACH Msg3 Success Rate % [CDBH]'),
        ('Packet Loss Rate UL % [CDBH]','UL Packet Loss [CDBH]'),
        ('UL RSSI [CDBH]','UL RSSI PUSCH [dBm] [CDBH]'),
        ('Inter gNodeB PSCell Change Success Rate_Denom [CDBH]','Inter sgNodeB PSCell Change Success Rate_Denom [CDBH]'),
        ('Inter gNodeB PSCell Change Success Rate_Nom [CDBH]','Inter sgNodeB PSCell Change Success Rate_Nom [CDBH]'),
        ('Intra gNodeB PSCell Change Success Rate_Denom [CDBH]','Intra sgNodeB PSCell Change Success Rate_Denom [CDBH]'),
        ('Intra gNodeB PSCell Change Success Rate_Nom [CDBH]','Intra sgNodeB PSCell Change Success Rate_Nom [CDBH]'),
        ('Average UE Distance_Denom','Average TA [CDBH]'),
        ('SA Data Volume DL SD1 [GB]','SD1 Traffic'),
        
    ],
    'samsung': [
        ('Inter sgNodeB PSCell Change Success Rate % [CDBH]', 'Inter sgNodeB PSCell Change Success Rate % [CDBH]'),
        ('Intra sgNodeB PSCell Change Success Rate % [CDBH]','Intra sgNodeB PSCell Change Success Rate % [CDBH]'),
        ('DL Initial BLER [CDBH]','DL BLER [CDBH]'),
        ('UL Initial BLER % [CDBH]','UL BLER [CDBH]'),
        ('Packet Loss Rate DL [CDBH]','DL Packet Loss [CDBH]'),
        ('Random Access Success Rate % [CDBH]','RACH Msg3 Success Rate % [CDBH]'),
        ('Packet Loss Rate UL [CDBH]','UL Packet Loss [CDBH]'),
        ('UL RSSI PUSCH [dBm] [CDBH]','UL RSSI PUSCH [dBm] [CDBH]'),
        ('Inter sgNodeB PSCell Change Success Rate_Denom [CDBH]','Inter sgNodeB PSCell Change Success Rate_Denom [CDBH]'),
        ('Inter sgNodeB PSCell Change Success Rate_Nom [CDBH]','Inter sgNodeB PSCell Change Success Rate_Nom [CDBH]'),
        ('Intra sgNodeB PSCell Change Success Rate_Denom [CDBH]','Intra sgNodeB PSCell Change Success Rate_Denom [CDBH]'),
        ('Intra sgNodeB PSCell Change Success Rate_Nom [CDBH]','Intra sgNodeB PSCell Change Success Rate_Nom [CDBH]'),
        ('Average TA [CDBH]','Average TA [CDBH]'),
        ('SA Data Volume DL SD1 [GB]','SD1 Traffic'),
        
    
    ],
}

COMPUTED_COLUMNS = [
    {
        'name':           'SgNb',
        'source_metric':  'SgNb addition Success Rate % [CDBH]',
        'condition':      lambda v: pd.notna(v) and v >= 99,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '>=99%',   # shown in the merged row-1 header above this column
    },

    {
        'name':           'Drop Rate',
        'source_metric':  'Session PS Drop Rate % [CDBH]',
        'condition':      lambda v: pd.notna(v) and v <=1.3,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '<=1.3',   # shown in the merged row-1 header above this column
    },

    {
        'name':           'Inter_PSCell',
        'source_metric':  'Inter sgNodeB PSCell Change Success Rate % [CDBH]',
        'condition':      lambda v: pd.notna(v) and v >=95,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '>=95%',   # shown in the merged row-1 header above this column
    },

    {
        'name':           'Intra_PSCell',
        'source_metric':  'Intra sgNodeB PSCell Change Success Rate % [CDBH]',
        'condition':      lambda v: pd.notna(v) and v >=98,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '>=98%',   # shown in the merged row-1 header above this column
    },

    {
        'name':           'Average CQI',
        'source_metric':  'Average CQI [CDBH]',
        'condition':      lambda v: pd.notna(v) and v >=10,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '>=10',   # shown in the merged row-1 header above this column
    },

    {
        'name':           'UL RSSI PUSCH',
        'source_metric':  'UL RSSI PUSCH [dBm] [CDBH]',
        'condition':      lambda v: pd.notna(v) and v <=-110,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '<=-110',   # shown in the merged row-1 header above this column
    },

    {
        'name':           'UL BLER',
        'source_metric':  'UL BLER [CDBH]',
        'condition':      lambda v: pd.notna(v) and v <=10,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '<=10',   # shown in the merged row-1 header above this column
    },
    {
        'name':           'DL BLER',
        'source_metric':  'DL BLER [CDBH]',
        'condition':      lambda v: pd.notna(v) and v <=14,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '<=14',   # shown in the merged row-1 header above this column
    },

    {
        'name':           'RACH',
        'source_metric':  'RACH Msg3 Success Rate % [CDBH]',
        'condition':      lambda v: pd.notna(v) and v >=60,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '>=60',   # shown in the merged row-1 header above this column
    },

    {
        'name':           'DL Latency',
        'source_metric':  'Average DL Latency [ms] [CDBH]',
        'condition':      lambda v: pd.notna(v) and v <=15,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '<=15',   # shown in the merged row-1 header above this column
    },

    {
        'name':           'DL UE THPT',
        'source_metric':  'Average DL UE Throughput_Mbps [CDBH]',
        'condition':      lambda v: pd.notna(v) and v >=100,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '>=100',   # shown in the merged row-1 header above this column
    },

    {
        'name':           'DL Packet Loss',
        'source_metric':  'DL Packet Loss [CDBH]',
        'condition':      lambda v: pd.notna(v) and v <0.5,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '<0.5',   # shown in the merged row-1 header above this column
    },

    {
        'name':           'UL Packet Loss',
        'source_metric':  'UL Packet Loss [CDBH]',
        'condition':      lambda v: pd.notna(v) and v <0.5,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '<0.5',   # shown in the merged row-1 header above this column
    },

    {
        'name':           'SD1',
        'source_metric':  'SD1 Traffic',
        'condition':      lambda v: pd.notna(v) and v >0,
        'window_days':    5,          # most recent N date columns, row-wise
        'insert_before':  '5G Data Volume [GB]',
        'header_label':  '>0',   # shown in the merged row-1 header above this column
    },
]

OP_MAP = {
    '>=': operator.ge,
    '<=': operator.le,
    '>':  operator.gt,
    '<':  operator.lt,
    '==': operator.eq,
}
DAILY_FORMULA_CHECKS = [
    {
        'name_prefix': 'Day',
        'window_days': 5,
        'conditions': [
            ('Inter sgNodeB PSCell Change Success Rate % [CDBH]', '>=', 96),
            ('Intra sgNodeB PSCell Change Success Rate % [CDBH]', '>=', 98),
            ('Session PS Drop Rate % [CDBH]',                     '<=', 1.3),
            ('SgNb addition Success Rate % [CDBH]',               '>=', 99),
        ],
        'true_label':  'Meeting',
        'false_label': 'Not Meeting',
    },
    # Add more multi-metric daily checks here in the same shape.
]
# All unique standardised output metric names (drives pivot + Excel headers)
ALL_METRIC_COLUMNS = list(dict.fromkeys(
    [m for m in COMMON_METRIC_COLUMNS] +
    [output_col for vendor_metrics in VENDOR_METRIC_COLUMNS.values() for _, output_col in vendor_metrics]
))

def _extract_circle(cell_name):
    """Circle is the first underscore-separated token of Cell Name (e.g. 'AP')."""
    if not isinstance(cell_name, str):
        return None
    return cell_name.split('_')[0]


def _extract_site_id(cell_name):
    """Site ID = 2 uppercase letters followed by 3+ digits (e.g. 'HY9229')."""
    if not isinstance(cell_name, str):
        return None
    
    match = re.search(r'x+([A-Za-z0-9-]+)_[ABC]$', cell_name)
    return match.group(1) if match else None


def _extract_short_name(cell_name):
    """Short Name is the raw Cell Name value as-is."""
    return cell_name

def _compute_extra_column(raw_pivots, spec):
    """
    Count, per row, how many of the last `window_days` date columns of
    `source_metric` satisfy `condition`. Returns a Series aligned to the
    same (Circle, Short Name, Site ID) index as the source pivot.
    """
    source_pivot = raw_pivots.get(spec['source_metric'])
    if source_pivot is None or source_pivot.shape[1] == 0:
        return pd.Series(dtype='int64', name=spec['name'])

    window_cols = source_pivot.columns[-spec['window_days']:]
    windowed = source_pivot[window_cols]
    counts = windowed.map(spec['condition']).sum(axis=1)
    counts.name = spec['name']
    return counts


def _compute_daily_formula_columns(raw_pivots, spec):
    """
    For the last `window_days` dates, evaluate an AND of multiple
    (metric, operator, threshold) conditions row-wise, producing one
    'Meeting'/'Not Meeting' column per day: oldest -> Day1 ... newest -> DayN.
    """
    metric_names = [cond[0] for cond in spec['conditions']]
    missing = [m for m in metric_names if m not in raw_pivots or raw_pivots[m].shape[1] == 0]
    if missing:
        return []

    # Dates should align across metrics (same source rows/Date column);
    # intersect to be safe against any per-metric gaps.
    common_dates = raw_pivots[metric_names[0]].columns
    for m in metric_names[1:]:
        common_dates = common_dates.intersection(raw_pivots[m].columns)
    common_dates = sorted(common_dates)[-spec['window_days']:]

    day_columns = []
    for i, date in enumerate(common_dates, start=1):
        result = None
        for metric, op_symbol, threshold in spec['conditions']:
            op_func = OP_MAP[op_symbol]
            col_values = raw_pivots[metric][date]
            condition_result = col_values.apply(
                lambda v: op_func(v, threshold) if pd.notna(v) else False
            )
            result = condition_result if result is None else (result & condition_result)

        labels = result.map({True: spec['true_label'], False: spec['false_label']})
        labels.name = f"{spec['name_prefix']}{i}"
        day_columns.append(labels)

    return day_columns

def _compute_meeting_count(day_columns, true_label, column_name):
    """
    Count, per row, how many of the given day-check Series equal true_label.
    """
    if not day_columns:
        return None
    day_df = pd.concat(day_columns, axis=1)
    count = (day_df == true_label).sum(axis=1)
    count.name = column_name
    return count


CELL_NAME_EXTRACTORS = {
    'Circle':     _extract_circle,
    'Site ID':    _extract_site_id,
    'Short Name': _extract_short_name,
    
}

WINDOW_DAYS = 5


PER_KPI_THRESHOLDS = [
    ('SgNb addition Success Rate % [CDBH]',               '>=', 99,    'SgNb'),
    ('Session PS Drop Rate % [CDBH]',                       '<=', 1.3,  'Drop Rate'),
    ('Inter sgNodeB PSCell Change Success Rate % [CDBH]',  '>=', 95,    'Inter PS Cell'),
    ('Intra sgNodeB PSCell Change Success Rate % [CDBH]',  '>=', 98,    'Intra PS Cell'),
    ('Average CQI [CDBH]',                                  '>=', 10,   'Average CQI'),
    ('UL RSSI PUSCH [dBm] [CDBH]',                           '<=', -110,'UL RSSI PUSCH'),
    ('UL BLER [CDBH]',                                       '<=', 10,  'UL BLER'),
    ('DL BLER [CDBH]',                                       '<=', 14,  'DL BLER'),
    ('RACH Msg3 Success Rate % [CDBH]',                     '>=', 60,   'RACH'),
    ('Average DL Latency [ms] [CDBH]',                       '<=', 15,  'DL Latency'),
    ('Average DL UE Throughput_Mbps [CDBH]',                 '>=', 100, 'DL UE Throughput'),
    ('DL Packet Loss [CDBH]',                                 '<', 0.5, 'DL Packet Loss'),
    ('UL Packet Loss [CDBH]',                                 '<', 0.5, 'UL Packet Loss'),
]

# Red-zone rules for metric date-column blocks. Each condition returns True
# if the cell should be marked red. Metrics not listed here get no coloring.
CELL_COLOR_RULES = {
    'Inter sgNodeB PSCell Change Success Rate % [CDBH]': lambda v: pd.notna(v) and v < 96,
    'Intra sgNodeB PSCell Change Success Rate % [CDBH]': lambda v: pd.notna(v) and v == 0,
    'Session PS Drop Rate % [CDBH]':                     lambda v: pd.notna(v) and v > 1.16,
    'SgNb addition Success Rate % [CDBH]':               lambda v: pd.notna(v) and v < 99,
    'Average CQI [CDBH]':                                 lambda v: pd.notna(v) and v < 10,
    'Average DL Latency [ms] [CDBH]':                     lambda v: pd.notna(v) and v > 15,
    'Average DL UE Throughput_Mbps [CDBH]':               lambda v: pd.notna(v) and v < 100,
    'DL BLER [CDBH]':                                     lambda v: pd.notna(v) and v > 14,
    'UL BLER [CDBH]':                                     lambda v: pd.notna(v) and v > 10,
    'RACH Msg3 Success Rate % [CDBH]':                    lambda v: pd.notna(v) and v < 55,
    'UL Packet Loss [CDBH]':                               lambda v: pd.notna(v) and v > 0.55,
    'UL RSSI PUSCH [dBm] [CDBH]':                         lambda v: pd.notna(v) and v < -110,
    'Radio Network Availability %':                        lambda v: pd.notna(v) and v < 99,
}

# Computed columns (SgNb ... SD1): red if value < 3, green otherwise.
COMPUTED_COLUMN_RED_THRESHOLD = 3

MEETING_KPIS_CONFIG = {
    'column_name':     'Meeting KPIs',
    'count_threshold': 3,   # a KPI counts toward 'Meeting KPIs' if its own 5-day count >= this
    'insert_before':   '5G Data Volume [GB]',
}


def _build_per_kpi_count_df(raw_pivots):
    """
    Build a DataFrame where each column is one KPI's display name, valued
    0-WINDOW_DAYS — how many of the last WINDOW_DAYS days that KPI met its
    own threshold. Shared by Meeting KPIs and Failed KPI logic.
    """
    columns = {}
    for metric, operator, threshold, display_name in PER_KPI_THRESHOLDS:
        counts = _compute_per_kpi_day_count(raw_pivots, metric, operator, threshold, WINDOW_DAYS)
        if counts is not None:
            columns[display_name] = counts
    if not columns:
        return None
    return pd.DataFrame(columns)


def _compute_per_kpi_day_count(raw_pivots, metric, operator, threshold, window_days):
    """
    Count, per row, how many of the last `window_days` date columns for
    `metric` satisfy `operator threshold`.
    """
    pivot = raw_pivots.get(metric)
    if pivot is None or pivot.shape[1] == 0:
        return None
    op_func = OP_MAP[operator]
    window_cols = sorted(pivot.columns)[-window_days:]
    windowed = pivot[window_cols]
    return windowed.map(lambda v: 1 if (pd.notna(v) and op_func(v, threshold)) else 0).sum(axis=1)


def _compute_meeting_kpis_column(raw_pivots):
    """
    Two-level calculation:
    1. For each KPI in PER_KPI_THRESHOLDS, count how many of the last
       WINDOW_DAYS days met that KPI's own threshold (0-WINDOW_DAYS).
    2. 'Meeting KPIs' = count of how many of those per-KPI counts are
       >= MEETING_KPIS_CONFIG['count_threshold'].
    """
    per_kpi_df = _build_per_kpi_count_df(raw_pivots)
    if per_kpi_df is None:
        return None
    meeting_kpis = (per_kpi_df >= MEETING_KPIS_CONFIG['count_threshold']).sum(axis=1)
    meeting_kpis.name = MEETING_KPIS_CONFIG['column_name']
    return meeting_kpis

FAILED_KPI_CONFIG = {
    'column_name':      'Failed KPI',
    'pass_threshold':   2,   # KPI counts as failing if its value is <= this
    'all_failed_label': 'All KPIs failed to meet the target',
    'attention_prefix': '',
    'insert_after':     'Short Name',
}


def _compute_failed_kpi_column(raw_pivots):
    """
    Per row:
    - If every KPI value is <= pass_threshold -> all_failed_label.
    - Else (at least one KPI > pass_threshold) -> list the names of KPIs
      whose value is <= pass_threshold, as 'KPIs requiring attention: ...'.
    - If none are failing (all KPIs > pass_threshold) -> '' (nothing to flag).
    """
    per_kpi_df = _build_per_kpi_count_df(raw_pivots)
    if per_kpi_df is None:
        return None

    threshold = FAILED_KPI_CONFIG['pass_threshold']
    failing_mask = per_kpi_df <= threshold

    def _row_label(row):
        failing_names = row.index[row].tolist()
        if len(failing_names) == len(row):
            return FAILED_KPI_CONFIG['all_failed_label']
        if not failing_names:
            return ''
        return FAILED_KPI_CONFIG['attention_prefix'] + ', '.join(failing_names)

    labels = failing_mask.apply(_row_label, axis=1)
    labels.name = FAILED_KPI_CONFIG['column_name']
    return labels

OFFERING_REMARKS_CONFIG = {
    'column_name':         'Offering Remarks',
    'meeting_kpis_equals': len(PER_KPI_THRESHOLDS),  # all 13 KPIs must meet
    'criteria_threshold':  3,
    'true_label':          'Meeting',
    'false_label':         'Non-Meeting',
    'insert_before':       'Meeting KPIs',
}


def _compute_offering_remarks(meeting_kpis_column, criteria_count_column, config):
    """
    'Offering Remarks' = true_label if Meeting KPIs equals the full KPI count
    AND 2+1 Criteria >= criteria_threshold, else false_label.
    """
    if meeting_kpis_column is None or criteria_count_column is None:
        return None
    condition = (
        (meeting_kpis_column == config['meeting_kpis_equals']) &
        (criteria_count_column >= config['criteria_threshold'])
    )
    remarks = condition.map({True: config['true_label'], False: config['false_label']})
    remarks.name = config['column_name']
    return remarks





def _build_output_df(df, vendor):
    if CELL_NAME_COLUMN not in df.columns:
        raise ValueError(f"'{CELL_NAME_COLUMN}' column not found in input file.")
    if df.shape[1] <= DATE_COLUMN_INDEX:
        raise ValueError(f"Expected a date column at position {DATE_COLUMN_INDEX + 1}, but file only has {df.shape[1]} column(s).")

    missing_common = [m for m in COMMON_METRIC_COLUMNS if m not in df.columns]
    if missing_common:
        raise ValueError(f"Missing common metric column(s): {', '.join(missing_common)}")

    date_series = pd.to_datetime(df.iloc[:, DATE_COLUMN_INDEX], errors='coerce')
    index_cols  = list(CELL_NAME_EXTRACTORS.keys())

    base_df = pd.DataFrame({
        col_name: df[CELL_NAME_COLUMN].apply(extractor)
        for col_name, extractor in CELL_NAME_EXTRACTORS.items()
    })
    base_df['Date'] = date_series

    for metric in COMMON_METRIC_COLUMNS:
        base_df[metric] = df[metric]

    for source_col, output_col in VENDOR_METRIC_COLUMNS.get(vendor, []):
        if source_col not in df.columns:
            raise ValueError(f"Missing vendor-specific column '{source_col}' in {vendor} file.")
        base_df[output_col] = df[source_col]

    for metric in ALL_METRIC_COLUMNS:
        if metric not in base_df.columns:
            base_df[metric] = None

    base_df = base_df.drop_duplicates(subset=index_cols + ['Date'])

    def _pivot(metric):
        pivot = base_df.pivot_table(
            index=index_cols,
            columns='Date',
            values=metric,
            aggfunc='first'
        )
        return pivot.reindex(sorted(pivot.columns), axis=1)

  
    raw_pivots = {metric: _pivot(metric) for metric in ALL_METRIC_COLUMNS}

    # --- Computed columns (all use raw_pivots, must come after the line above) ---
    extra_series = {
        spec['name']: _compute_extra_column(raw_pivots, spec)
        for spec in COMPUTED_COLUMNS
    }

    meeting_kpis_column = _compute_meeting_kpis_column(raw_pivots)
    failed_kpi_column    = _compute_failed_kpi_column(raw_pivots)

    daily_blocks = []
    criteria_count_column = None
    for spec in DAILY_FORMULA_CHECKS:
        day_series_list = _compute_daily_formula_columns(raw_pivots, spec)
        daily_blocks.extend(s.to_frame() for s in day_series_list)
        criteria_count = _compute_meeting_count(day_series_list, spec['true_label'], '2+1 Criteria')
        if criteria_count is not None:
            daily_blocks.append(criteria_count.to_frame())
            criteria_count_column = criteria_count

    offering_remarks_column = _compute_offering_remarks(
        meeting_kpis_column, criteria_count_column, OFFERING_REMARKS_CONFIG
    )

    # Stringify + prefix date headers for the final layout.
    named_pivots = {}
    for metric, pivot in raw_pivots.items():
        renamed = pivot.copy()
        renamed.columns = [f"{metric}||{col.strftime('%Y-%m-%d')}" for col in pivot.columns]
        named_pivots[metric] = renamed

    # --- Assemble final column order ---
    ordered_blocks = []
    inserted_computed = set()
    meeting_kpis_inserted = False
    for metric in ALL_METRIC_COLUMNS:
        if not meeting_kpis_inserted and meeting_kpis_column is not None and MEETING_KPIS_CONFIG['insert_before'] == metric:
            if offering_remarks_column is not None:
                ordered_blocks.append(offering_remarks_column.to_frame())
            ordered_blocks.append(meeting_kpis_column.to_frame())
            meeting_kpis_inserted = True
        for spec in COMPUTED_COLUMNS:
            if spec['insert_before'] == metric and spec['name'] not in inserted_computed:
                ordered_blocks.append(extra_series[spec['name']].to_frame())
                inserted_computed.add(spec['name'])
        ordered_blocks.append(named_pivots[metric])

    if not meeting_kpis_inserted:
        if offering_remarks_column is not None:
            ordered_blocks.append(offering_remarks_column.to_frame())
        if meeting_kpis_column is not None:
            ordered_blocks.append(meeting_kpis_column.to_frame())
    for spec in COMPUTED_COLUMNS:
        if spec['name'] not in inserted_computed:
            ordered_blocks.append(extra_series[spec['name']].to_frame())

    extra_blocks = []
    if failed_kpi_column is not None:
        extra_blocks.append(failed_kpi_column.to_frame())

    merged = pd.concat(ordered_blocks + extra_blocks + daily_blocks, axis=1).reset_index()
    merged.columns.name = None

    if 'Failed KPI' in merged.columns:
        cols = list(merged.columns)
        cols.remove('Failed KPI')
        insert_pos = cols.index('Short Name') + 1
        cols.insert(insert_pos, 'Failed KPI')
        merged = merged[cols]

    return merged




def _apply_report_styling(ws, columns_list, data_start_row=3):
    """
    Apply the full KPI report styling to a worksheet that already has its
    header rows (1 and 2) and data written:
    - Row 1: alternating block-color labels (metric blocks + computed columns)
    - Row 2: light blue header row
    - Computed columns (SgNb...SD1): red if <3, green otherwise, thin borders
    - Metric date-columns: red fill if value breaches CELL_COLOR_RULES
    Recomputes block positions from columns_list, so it works on both the
    full report and any filtered subset with the same column structure.
    """
    metric_block_positions = {}
    for metric in ALL_METRIC_COLUMNS:
        cols = [
            i + 1 for i, c in enumerate(columns_list)
            if '||' in c and c.split('||', 1)[0] == metric
        ]
        if cols:
            metric_block_positions[metric] = (min(cols), max(cols))

    computed_col_positions = {
        spec['name']: (columns_list.index(spec['name']) + 1, spec['header_label'])
        for spec in COMPUTED_COLUMNS
        if spec['name'] in columns_list and spec.get('header_label')
    }
    computed_color_cols = {idx for idx, _label in computed_col_positions.values()}

    # Row 1 / Row 2 colors
    ROW1_COLOR_A = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ROW1_COLOR_B = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ROW2_FILL    = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ROW1_FONT    = Font(color="FFFFFF", bold=True)
    ROW2_FONT    = Font(bold=True)

    all_row1_blocks = sorted(
        list(metric_block_positions.items()) +
        [(name, (idx, idx)) for name, (idx, _label) in computed_col_positions.items()],
        key=lambda item: item[1][0]
    )
    for block_idx, (label, (start_col, end_col)) in enumerate(all_row1_blocks):
        fill = ROW1_COLOR_A if block_idx % 2 == 0 else ROW1_COLOR_B
        for col in range(start_col, end_col + 1):
            cell      = ws.cell(row=1, column=col)
            cell.fill = fill
            cell.font = ROW1_FONT

    covered_cols = set()
    for _, (s, e) in all_row1_blocks:
        covered_cols.update(range(s, e + 1))
    for col in range(1, len(columns_list) + 1):
        if col not in covered_cols:
            ws.cell(row=1, column=col).fill = ROW1_COLOR_A

    for col in range(1, len(columns_list) + 1):
        cell      = ws.cell(row=2, column=col)
        cell.fill = ROW2_FILL
        cell.font = ROW2_FONT

    # Cell-level color coding (metric date-columns + computed columns)
    RED_FILL   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    metric_color_lookup = {}
    for metric, condition in CELL_COLOR_RULES.items():
        for col_idx, col_name in enumerate(columns_list, start=1):
            if '||' in col_name and col_name.split('||', 1)[0] == metric:
                metric_color_lookup[col_idx] = condition

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    max_row = ws.max_row
    for r in range(data_start_row, max_row + 1):
        for c in range(1, len(columns_list) + 1):
            cell  = ws.cell(row=r, column=c)
            value = cell.value

            if c in metric_color_lookup:
                if metric_color_lookup[c](value):
                    cell.fill = RED_FILL
            elif c in computed_color_cols:
                if pd.notna(value) and value < COMPUTED_COLUMN_RED_THRESHOLD:
                    cell.fill = RED_FILL
                elif pd.notna(value):
                    cell.fill = GREEN_FILL

    for col in computed_color_cols:
        for r in range(1, max_row + 1):
            ws.cell(row=r, column=col).border = thin_border


@api_view(['POST'])
def generate_kpi_monitoring_report(request):
    """
    Reads each uploaded vendor input file, builds the full KPI report, and
    writes one merged Excel output with a two-row colored header:
        Row 1 — metric block labels (merged, alternating colors) + computed
                column threshold labels
        Row 2 — actual column headers (light blue), dates stripped of metric prefix
        Row 3+ — data
    """
    try:
        input_files = _get_input_files()
        missing_vendors = [v for v in VENDOR_FIELDS if v not in input_files]
        if missing_vendors:
            return Response(
                {
                    'status': False,
                    'error':  f"Missing input file(s) for: {', '.join(missing_vendors)}.",
                    'found':  list(input_files.keys()),
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        combined_frames = []
        for vendor in VENDOR_FIELDS:
            if vendor not in input_files:
                continue
            filename = input_files[vendor]
        # for vendor, filename in input_files.items():
            file_path = os.path.join(input_path, filename)
            df = _read_file_to_df(file_path)
            try:
                output_df = _build_output_df(df, vendor)
            except ValueError as ve:
                return Response(
                    {'status': False, 'error': f"{vendor}: {str(ve)}"},
                    status=status.HTTP_422_UNPROCESSABLE_ENTITY
                )
            output_df.insert(0, 'Vendor', vendor.capitalize())
            combined_frames.append(output_df)

        merged_df = pd.concat(combined_frames, ignore_index=True)

       

        output_filename  = MAIN_REPORT_FILENAME
        output_file_path = os.path.join(output_path, output_filename)
        columns_list      = list(merged_df.columns)

        # Exact-match column positions — avoids prefix-overlap MergedCell errors
        metric_block_positions = {}
        for metric in ALL_METRIC_COLUMNS:
            cols = [
                i + 1 for i, c in enumerate(columns_list)
                if '||' in c and c.split('||', 1)[0] == metric
            ]
            if cols:
                metric_block_positions[metric] = (min(cols), max(cols))

        computed_col_positions = {
            spec['name']: (columns_list.index(spec['name']) + 1, spec['header_label'])
            for spec in COMPUTED_COLUMNS
            if spec['name'] in columns_list and spec.get('header_label')
        }

        wb = Workbook()
        ws = wb.active
        ws.title = 'KPI Report'

        # Row 1: metric block labels (merged)
        for metric, (start_col, end_col) in metric_block_positions.items():
            if end_col > start_col:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            ws.cell(row=1, column=start_col).value     = metric
            ws.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')

        # Row 1: computed column threshold labels (e.g. '>=99%')
        for col_name, (col_idx, label) in computed_col_positions.items():
            ws.cell(row=1, column=col_idx).value     = label
            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')

        # Row 2: column headers (strip "MetricName||" prefix from date columns)
        for col_idx, col_name in enumerate(columns_list, start=1):
            header = col_name.split('||', 1)[1] if '||' in col_name else col_name
            ws.cell(row=2, column=col_idx).value = header

        
      
        # Row 3+: data
        for r, row in enumerate(merged_df.itertuples(index=False), start=3):
            for c, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c)
                cell.value = value
        _apply_report_styling(ws, columns_list, data_start_row=3)

               

       

        wb.save(output_file_path)

        download_url = request.build_absolute_uri(
            f"{settings.MEDIA_URL.rstrip('/')}/kpi_monitoring/output/{output_filename}"
        )

        return Response({
            'status':       True,
            'message':      'KPI monitoring report generated successfully.',
            'rows_written': len(merged_df),
            'download_url': download_url,
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



    
@api_view(['POST'])
def filter_kpi_report(request):
    """
    POST body (JSON):
    {
      "circle": "AP",
      "site_id": ["HY9229", "DL1234"],
      "short_name": ["AP_5_EE_...", "..."]
    }
    Rules:
    - 'circle' is an AND filter (optional), supports single value or list.
    - 'site_id' and 'short_name' are OR filters, at least one required.
    - All fields accept single string or comma-separated string or list.
    """
    try:
        def _to_list(value):
            if value is None:
                return []
            if isinstance(value, list):
                return [str(v).strip().lower() for v in value if str(v).strip()]
            if isinstance(value, str):
                return [v.strip().lower() for v in value.split(',') if v.strip()]
            return [str(value).strip().lower()]

        circle      = _to_list(request.data.get('circle'))
        site_ids    = _to_list(request.data.get('site_id'))
        short_names = _to_list(request.data.get('short_name'))

        if not site_ids and not short_names:
            return Response(
                {'status': False, 'error': "Provide at least one of 'site_id' or 'short_name'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        source_path = os.path.join(output_path, MAIN_REPORT_FILENAME)
        if not os.path.exists(source_path):
            return Response(
                {'status': False, 'error': "No report found. Generate the KPI monitoring report first."},
                status=status.HTTP_400_BAD_REQUEST
            )

        src_wb = load_workbook(source_path)
        src_ws = src_wb.active

        header_row = [cell.value for cell in src_ws[2]]
        try:
            circle_col_idx     = header_row.index('Circle')
            site_id_col_idx    = header_row.index('Site ID')
            short_name_col_idx = header_row.index('Short Name')
        except ValueError as ve:
            return Response(
                {'status': False, 'error': f"Could not locate filter column in report: {str(ve)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        filtered_rows = []
        for row in src_ws.iter_rows(min_row=3, values_only=True):
            row_circle     = str(row[circle_col_idx]).strip().lower()     if row[circle_col_idx]     is not None else ''
            row_site_id    = str(row[site_id_col_idx]).strip().lower()    if row[site_id_col_idx]    is not None else ''
            row_short_name = str(row[short_name_col_idx]).strip().lower() if row[short_name_col_idx] is not None else ''

            if circle and row_circle not in circle:
                continue

            # matches_site_id    = bool(site_ids)    and row_site_id    in site_ids
            # matches_short_name = bool(short_names) and row_short_name in short_names
            matches_site_id    = bool(site_ids)    and any(s in row_site_id    for s in site_ids)
            matches_short_name = bool(short_names) and any(s in row_short_name for s in short_names)
            if not (matches_site_id or matches_short_name):
                continue

            filtered_rows.append(row)

        if not filtered_rows:
            return Response(
                {'status': False, 'error': 'No matching rows found for the given filter.'},
                status=status.HTTP_404_NOT_FOUND
            )

        max_col = src_ws.max_column

        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = 'Filtered KPI Report'

        # Copy row 1 and row 2 values
        for col in range(1, max_col + 1):
            new_ws.cell(row=1, column=col).value = src_ws.cell(row=1, column=col).value
            new_ws.cell(row=2, column=col).value = src_ws.cell(row=2, column=col).value

        # Copy merged cells from row 1
        for merged_range in src_ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                new_ws.merge_cells(
                    start_row=1, start_column=merged_range.min_col,
                    end_row=1,   end_column=merged_range.max_col
                )

        for col in range(1, max_col + 1):
            cell = new_ws.cell(row=1, column=col)
            if cell.value is not None:
                cell.alignment = Alignment(horizontal='center')

        # Write filtered data rows
        for r, row in enumerate(filtered_rows, start=3):
            for c, value in enumerate(row, start=1):
                new_ws.cell(row=r, column=c).value = value

        # Reconstruct metric||date prefixed columns_list that
        # _apply_report_styling expects. Computed columns kept as plain names
        # so _apply_report_styling finds them via computed_color_cols.
        computed_names = {spec['name'] for spec in COMPUTED_COLUMNS}

        row1_values = [''] * max_col
        for merged_range in src_ws.merged_cells.ranges:
            if merged_range.min_row == 1:
                val = src_ws.cell(row=1, column=merged_range.min_col).value or ''
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    row1_values[col - 1] = val
        for col in range(1, max_col + 1):
            if not row1_values[col - 1]:
                row1_values[col - 1] = src_ws.cell(row=1, column=col).value or ''

        prefixed_columns_list = []
        for col_idx in range(1, max_col + 1):
            row1_val = str(row1_values[col_idx - 1]).strip()
            row2_val = str(src_ws.cell(row=2, column=col_idx).value or '').strip()

            if row2_val in computed_names:
                # Computed column — plain name so _apply_report_styling
                # finds it in computed_color_cols by spec['name']
                prefixed_columns_list.append(row2_val)
            elif row1_val and row2_val and row1_val != row2_val:
                # Metric date column — prefix with metric name
                prefixed_columns_list.append(f"{row1_val}||{row2_val}")
            else:
                # Static column (Vendor, Circle, Site ID, Short Name, etc.)
                prefixed_columns_list.append(row2_val)

        _apply_report_styling(new_ws, prefixed_columns_list, data_start_row=3)

        filter_label     = '_'.join(circle) if circle else 'filtered'
        output_filename  = f"kpi_report_{filter_label}_{len(filtered_rows)}rows.xlsx"
        output_file_path = os.path.join(output_path, output_filename)
        new_wb.save(output_file_path)

        download_url = request.build_absolute_uri(
            f"{settings.MEDIA_URL.rstrip('/')}/kpi_monitoring/output/{output_filename}"
        )

        return Response({
            'status':       True,
            'message':      'Filtered report generated successfully.',
            'rows_matched': len(filtered_rows),
            'download_url': download_url,
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)