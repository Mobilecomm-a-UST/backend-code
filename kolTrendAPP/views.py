from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.core.files.storage import FileSystemStorage
from datetime import date,timedelta
import datetime
import openpyxl
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
import pandas as pd
import os
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill
from commom_utilities.utils import *



@api_view(['POST'])
def old_kol_trend_uls(request):
    kpi=['RRC Setup Success Rate [CBBH]','VoLTE ERAB Setup Success Rate [CBBH]','PS Drop Call Rate % [CDBH]_Old','MV_DL User Throughput_Mbps [CDBH]',
    'MV_UL User Throughput_Kbps [CDBH]','PS handover success rate [LTE Intra System] [CDBH]','PS handover success rate [LTE Inter System] [CBBH]',
    'MV_CSFB Redirection Success Rate [CDBH]','ERAB Setup Success Rate [CDBH]','VoLTE DCR [CBBH]','VoLTE Packet Loss DL [CBBH]',
    'VoLTE Packet Loss UL [CBBH]','VoLTE Intra-LTE Handover Success Ratio [CBBH]','VoLTE Inter-Frequency Handover Success Ratio [CBBH]',
    'E-UTRAN Average CQI [CDBH]','UL RSSI [CDBH]','4G Data Volume [GB]','VoLTE Traffic','Average number of used DL PRBs','VoLTE SRVCC SR']
    
    raw_kpi_4G=request.FILES['raw_kpi'] if "raw_kpi" in request.FILES else None
    df_raw_kpi=pd.read_excel( raw_kpi_4G)

    required_cols=['RRC Setup Success Rate [CBBH]','VoLTE ERAB Setup Success Rate [CBBH]','PS Drop Call Rate % [CDBH]_Old','MV_DL User Throughput_Kbps [CDBH]',
    'MV_UL User Throughput_Kbps [CDBH]','PS handover success rate [LTE Intra System] [CDBH]','PS handover success rate [LTE Inter System] [CBBH]',
    'MV_CSFB Redirection Success Rate [CDBH]','ERAB Setup Success Rate [CDBH]','VoLTE DCR [CBBH]','VoLTE Packet Loss DL [CBBH]',
    'VoLTE Packet Loss UL [CBBH]','VoLTE Intra-LTE Handover Success Ratio [CBBH]','VoLTE Inter-Frequency Handover Success Ratio [CBBH]',
    'E-UTRAN Average CQI [CDBH]','UL RSSI [CDBH]','4G Data Volume [GB]','VoLTE Traffic','Average number of used DL PRBs','VoLTE SRVCC SR']
    
   
    sts,response=required_col_check(raw_kpi_4G,required_cols)
    if sts:
        return Response(response)

    response,s_l= site_list_handler(request)
    if s_l:
        # site_list = list(map(lambda a: str(a), s_l))  ##################### converting in str
        site_list=s_l
      
        site_list_str = [str(x) for x in site_list]
        
        print("Converted site_list to string:", site_list_str)
        print("__________site_list______", type(site_list_str[2]))

        print(site_list_str, "_______________", s_l)
        print("Checkpoint 1")
        if response:
            return Response(response)
        print("Checkpoint 2")
#####################################################################################
    door_path=os.path.join(MEDIA_ROOT,'trends','kol')
    df_raw_kpi['Short name'].fillna(inplace=True, method="ffill")
    df_raw_kpi["Short name"] =df_raw_kpi["Short name"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..
 
    df_raw_kpi.rename( columns={'Unnamed: 1':'date'}, inplace=True )
    df_raw_kpi['MV_DL User Throughput_Kbps [CDBH]']=(df_raw_kpi['MV_DL User Throughput_Kbps [CDBH]']/1024)
    df_raw_kpi.rename(columns={"MV_DL User Throughput_Kbps [CDBH]" :"MV_DL User Throughput_Mbps [CDBH]" } ,inplace = True )
    
    print(df_raw_kpi)

    split=[]
    tech=[]
    for cell in df_raw_kpi['Short name']:
        if('_' in cell):
            site_id = cell.split("_")[-2][:-1]
            split.append(site_id)
            
        else:
            site_id=cell[:-1]
            split.append(site_id) 

            
        if('KO' in cell or 'WB' in cell ):  
            if('KO' in cell):
                circle='KOL' 
                tech.append(circle)
            if('WB' in cell ):
                circle='ROB'
                tech.append(circle)                   
        else:
            circle=site_id
            tech.append(circle)     
            
    df_raw_kpi.insert(0,'Site ID', split)
    df_raw_kpi.insert(6,'circle',tech )  
    
    print( df_raw_kpi)

    ecgi=[]
    cgi=[]
    for cell1 in  df_raw_kpi['4G_ECGI']:
        if('-' in str(cell1)):
            lnbts_id = cell1.split("-")[-2]
            ecgi.append(lnbts_id)   
        else:
            
            lnbts_id=cell1
            ecgi.append(lnbts_id)
            

        if('-' in str(cell1)):
            lnbts_Name=cell1.split('-')[-1]
            cgi.append(lnbts_Name) 
        else:
            lnbts_Name=cell1
            cgi.append(lnbts_Name) 
                   
    df_raw_kpi.insert(2,'ENODEB_ID', ecgi)
    df_raw_kpi.insert(3,'ci',cgi) 
    # df["split_short"]=[for cell2 in df['Short name']]
    print( df_raw_kpi) 
    df_raw_kpi.fillna(value=0,inplace=True)  

    message=site_comparision(ecgi,site_list_str) # site_comparision_call 
    print("message",message)
    

    PsOsPath1=os.path.join(door_path,'process output','desird_output.xlsx')    
    df_raw_kpi.to_excel(PsOsPath1,index=False)

    fill_excelfile_1=PsOsPath1


    df1=pd.read_excel(fill_excelfile_1)


    df1.rename(columns={"ENODEB ID": "ENODEB_ID"}, inplace=True)



    filtered_df_1 = df1[(df1.ENODEB_ID.isin(site_list))]

    print(filtered_df_1)
    PsOs_Filter=os.path.join(door_path,'process output','filtered.xlsx')
    filtered_df_1.to_excel(PsOs_Filter, index=False)

    df1 = pd.read_excel(PsOs_Filter)
    df_pivot = df1.pivot_table(values=kpi,columns='date', index=['Short name', 'Site ID','circle','ENODEB_ID','ci','4G_ECGI'])
    
    PsOs_Pivot=os.path.join(door_path,"process output','pivot.xlsx")
    df_pivot.to_excel(PsOs_Pivot) 
    
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
	# This process is similar to binary-to-
	# decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    str_date=request.POST.get("offered_date")
    date1=datetime.datetime.strptime(str_date,"%Y-%m-%d")
    print("strdate--------------------------------------",date1)

    # date1 =date(2023,1,27)
    # date1=cal.get_date()
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
    # index=df_pivot.index
    
    STR=os.path.join(door_path,'template','new up.xlsx')
    wb=openpyxl.load_workbook(STR)
    ws=wb.active
    
    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        print('donnnnnnnnnnnnnnnnne')
        print(len(index))
        dr=df_pivot[kpi_name]
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()

        ws[coln1+"4"].value=cl[4]
        ws[coln2+"4"].value=cl[3]
        ws[coln3+"4"].value=cl[2]
        ws[coln4+"4"].value=cl[1]
        ws[coln5+"4"].value=cl[0]

        for i,value in enumerate(index):
                j=i+5
                ws['A'+str(j)].value=index[i][2]
                ws['D'+str(j)].value=index[i][1]
                ws['E'+str(j)].value=index[i][3]
                ws['F'+str(j)].value=index[i][0]
                ws['G'+str(j)].value=index[i][4]
                ws['H'+str(j)].value=index[i][5]
            
                
                
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
                
    for kpi_name in kpi:
        
        if(kpi_name=='RRC Setup Success Rate [CBBH]'):
            overwrite(kpi_name,'I') 
            
        if(kpi_name=='VoLTE ERAB Setup Success Rate [CBBH]'):
            overwrite(kpi_name,'P')    
            
        if(kpi_name=='PS Drop Call Rate % [CDBH]_Old'):
            overwrite(kpi_name,'W')   
            
        if(kpi_name=='MV_DL User Throughput_Mbps [CDBH]'):
            overwrite(kpi_name,'AD')
            
        if(kpi_name=='MV_UL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name,'AK')
            
        if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'AR')
            
        if(kpi_name=='PS handover success rate [LTE Inter System] [CBBH]'):
            overwrite(kpi_name,'AY')  
            
        if(kpi_name=='MV_CSFB Redirection Success Rate [CDBH]'):
            overwrite(kpi_name,'BF')
            
        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'BM') 
            
        if(kpi_name=='VoLTE DCR [CBBH]'):
            overwrite(kpi_name,'BT') 
            
        if(kpi_name=='VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name,'CA')
            
        if(kpi_name=='VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name,'CH')
            
        if(kpi_name=='VoLTE SRVCC SR'):
            overwrite(kpi_name,'CO')
            
        if(kpi_name=='VoLTE Intra-LTE Handover Success Ratio [CBBH]'):
            overwrite(kpi_name,'CV')
            
        if(kpi_name=='VoLTE Inter-Frequency Handover Success Ratio [CBBH]'):
            overwrite(kpi_name,'DC')
            
        if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name,'DJ')
            
        if(kpi_name=='UL RSSI [CDBH]'):
            overwrite(kpi_name,'DQ')
            
        if(kpi_name=='4G Data Volume [GB]'):
            overwrite(kpi_name,'DX')
            
        if(kpi_name=='VoLTE Traffic'):
            overwrite(kpi_name,'EE') 
            
        if(kpi_name=='Average number of used DL PRBs'):
            overwrite(kpi_name,'EL')
    save_output=os.path.join(door_path,'output','KOL_KPI_TREND_OUTPUT.xlsx')    
    wb.save(save_output)
    print('successfully') 
    download_path=os.path.join(MEDIA_URL,"trends",'kol','output','KOL_KPI_TREND_OUTPUT.xlsx')
    return Response({"status":True,"message":"uploaded sucessfully","missing_sites":message,'Download_url':download_path}) 
  
##################################### samsung ############################
@api_view(['POST'])
def old_kol_trend_sam(request):
    kpi=['RRC Setup Success Rate [CDBH]',
         'VoLTE ERAB Setup Success Rate [CBBH]',
         'PS Drop Call Rate % [CDBH]',
         'DL User Throughput_Mbps [CDBH]',
         'UL User Throughput_Kbps [CDBH]','PS IntraF HOSR [CDBH]',
         'PS InterF HOSR [CDBH]',
         'CSFB Redirection Success Rate [CDBH]','ERAB Setup Success Rate [CDBH]','VoLTE DCR [CBBH]',
         'VoLTE Packet Loss DL [CBBH]',
         'VoLTE Packet Loss UL [CBBH]','VoLTE IntraF HOSR [CBBH]',
         'VoLTE InterF HOSR [CBBH]',
         'E-UTRAN Average CQI [CDBH]','Avg RSSI [dBm] [CDBH]','4G Data Volume [GB]',
         'VoLTE Traffic','Average number of used DL PRBs [CDBH]','VoLTE SRVCC Per Call Rate [CBBH]',]
      
    raw_kpi_4G=request.FILES['raw_kpi'] if "raw_kpi" in request.FILES else None
    df_raw_kpi=pd.read_excel( raw_kpi_4G)

    required_col=['RRC Setup Success Rate [CDBH]',
         'VoLTE ERAB Setup Success Rate [CBBH]',
         'PS Drop Call Rate % [CDBH]',
         'DL User Throughput_Kbps [CDBH]',
         'UL User Throughput_Kbps [CDBH]','PS IntraF HOSR [CDBH]',
         'PS InterF HOSR [CDBH]',
         'CSFB Redirection Success Rate [CDBH]','ERAB Setup Success Rate [CDBH]','VoLTE DCR [CBBH]',
         'VoLTE Packet Loss DL [CBBH]',
         'VoLTE Packet Loss UL [CBBH]','VoLTE IntraF HOSR [CBBH]',
         'VoLTE InterF HOSR [CBBH]',
         'E-UTRAN Average CQI [CDBH]','Avg RSSI [dBm] [CDBH]','4G Data Volume [GB]',
         'VoLTE Traffic','Average number of used DL PRBs [CDBH]','VoLTE SRVCC Per Call Rate [CBBH]',]   
    sts,response=required_col_check(raw_kpi_4G,required_col)
    if sts:
        return Response(response)

    response,s_l= site_list_handler(request)
    if s_l:
        # site_list = list(map(lambda a: str(a), s_l))  ##################### converting in str
        site_list=s_l
      
        site_list_str = [str(x) for x in site_list]
        
        print("Converted site_list to string:", site_list_str)
        print("__________site_list______", type(site_list_str[2]))

        print(site_list_str, "_______________", s_l)
        print("Checkpoint 1")
        if response:
            return Response(response)
        print("Checkpoint 2")
       
#####################################################################################
    door_path=os.path.join(MEDIA_ROOT,'trends','kol')
    df_raw_kpi['Short name'].fillna(inplace=True, method="ffill")
    df_raw_kpi["Short name"] =df_raw_kpi["Short name"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..
 
    df_raw_kpi.rename( columns={'Unnamed: 1':'date'}, inplace=True )
    df_raw_kpi['DL User Throughput_Kbps [CDBH]']=(df_raw_kpi['DL User Throughput_Kbps [CDBH]']/1024)
    df_raw_kpi.rename(columns={"DL User Throughput_Kbps [CDBH]" :"DL User Throughput_Mbps [CDBH]"} ,inplace = True )
    
    print(df_raw_kpi)

    
    ecgi=[]
    cgi=[]
    for cell1 in  df_raw_kpi['ECGI']:
        if('-' in str(cell1)):
            lnbts_id = cell1.split("-")[-2]
            ecgi.append(lnbts_id)   
        else:
            
            lnbts_id=cell1
            ecgi.append(lnbts_id)
            

        if('-' in str(cell1)):
            lnbts_Name=cell1.split('-')[-1]
            cgi.append(lnbts_Name) 
        else:
            lnbts_Name=cell1
            cgi.append(lnbts_Name) 
                   
    df_raw_kpi.insert(2,'ENODEB_ID', ecgi)
    df_raw_kpi.insert(3,'ci',cgi) 

    print( df_raw_kpi) 
    df_raw_kpi.fillna(value=0,inplace=True)      

    message=site_comparision(ecgi,site_list_str) # site_comparision_call 
    print(message)
    
    PsOsPath1=os.path.join(door_path,'process output','desird_output.xlsx')    
    df_raw_kpi.to_excel(PsOsPath1,index=False)

    fill_excelfile_1=PsOsPath1


    df1=pd.read_excel(fill_excelfile_1)


    df1.rename(columns={"ENODEB ID": "ENODEB_ID"}, inplace=True)


    filtered_df_1 = df1[(df1.ENODEB_ID.isin(site_list))]

    print(filtered_df_1)
    PsOs_Filter=os.path.join(door_path,'process output','filtered.xlsx')
    filtered_df_1.to_excel(PsOs_Filter, index=False)

    df1 = pd.read_excel(PsOs_Filter)
    df_pivot = df1.pivot_table(values=kpi,columns='date', index=['Short name', 'ENODEB_ID','ci','ECGI'])
    
    PsOs_Pivot=os.path.join(door_path,"process output','pivot.xlsx")
    df_pivot.to_excel(PsOs_Pivot) 
    
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
	# This process is similar to binary-to-
	# decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    str_date=request.POST.get("offered_date")
    date1=datetime.datetime.strptime(str_date,"%Y-%m-%d")
    print("strdate--------------------------------------",date1)

    # date1 =date(2023,1,27)
    # date1=cal.get_date()
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
    # index=df_pivot.index
    
    STR=os.path.join(door_path,'template','new up.xlsx')
    wb=openpyxl.load_workbook(STR)
    ws=wb.active
    
    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        print('done')
        print(len(index))
        dr=df_pivot[kpi_name]
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()

        ws[coln1+"4"].value=cl[4]
        ws[coln2+"4"].value=cl[3]
        ws[coln3+"4"].value=cl[2]
        ws[coln4+"4"].value=cl[1]
        ws[coln5+"4"].value=cl[0]

        for i,value in enumerate(index):
                j=i+5
                ws['B'+str(j)].value='ULS'
                ws['A'+str(j)].value="SUMSUNG"
                # ws['A'+str(j)].value=index[i][2]
                ws['D'+str(j)].value=index[i][0]
                ws['E'+str(j)].value=index[i][1]
                ws['F'+str(j)].value=index[i][0]
                ws['G'+str(j)].value=index[i][2]
                ws['H'+str(j)].value=index[i][3]
            
                
                
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
                
    for kpi_name in kpi:
        
        if(kpi_name=='RRC Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'I') 
            
        if(kpi_name=='VoLTE ERAB Setup Success Rate [CBBH]'):
            overwrite(kpi_name,'P')    
            
        if(kpi_name=='PS Drop Call Rate % [CDBH]'):
            overwrite(kpi_name,'W')   
            
        if(kpi_name=='DL User Throughput_Mbps [CDBH]'):
            overwrite(kpi_name,'AD')
            
        if(kpi_name=='UL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name,'AK')
            
        if(kpi_name=='PS IntraF HOSR [CDBH]'):
            overwrite(kpi_name,'AR')
            
        if(kpi_name=='PS InterF HOSR [CDBH]'):
            overwrite(kpi_name,'AY')  
            
        if(kpi_name=='CSFB Redirection Success Rate [CDBH]'):
            overwrite(kpi_name,'BF')
            
        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'BM') 
            
        if(kpi_name=='VoLTE DCR [CBBH]'):
            overwrite(kpi_name,'BT') 
            
        if(kpi_name=='VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name,'CA')
            
        if(kpi_name=='VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name,'CH')
            
        if(kpi_name=='VoLTE SRVCC Per Call Rate [CBBH]'):
            overwrite(kpi_name,'CO')
            
        if(kpi_name=='VoLTE IntraF HOSR [CBBH]'):
            overwrite(kpi_name,'CV')
            
        if(kpi_name=='VoLTE InterF HOSR [CBBH]'):
            overwrite(kpi_name,'DC')
            
        if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name,'DJ')
            
        if(kpi_name=='Avg RSSI [dBm] [CDBH]'):
            overwrite(kpi_name,'DQ')
            
        if(kpi_name=='4G Data Volume [GB]'):
            overwrite(kpi_name,'DX')
            
        if(kpi_name=='VoLTE Traffic'):
            overwrite(kpi_name,'EE') 
            
        if(kpi_name=='Average number of used DL PRBs [CDBH]'):
            overwrite(kpi_name,'EL')
    save_output=os.path.join(door_path,'output','KOL_KPI_SAM_TREND_OUTPUT.xlsx')    
    wb.save(save_output)
    download_path=os.path.join(MEDIA_URL,"trends",'kol','output','KOL_KPI_SAM_TREND_OUTPUT.xlsx')
    return Response({"status":True,"message":"uploaded sucessfully","missing_sites":message,'Download_url':download_path})

################################### RELOCATION ###################################### 
@api_view(['POST'])
def old_kol_trend_relocation(request):
    kpi=['RRC Setup Success Rate [CBBH]',
    'ERAB Setup Success Rate [CDBH]',
    'PS Drop Call Rate % [CDBH]',
    'MV_DL User Throughput_Mbps [CDBH]',
    'MV_UL User Throughput_Kbps [CDBH]',
    '4G Data Volume [GB]',
    'E-UTRAN Average CQI [CDBH]','UL PUCCH SINR [CDBH]',
    'PS handover success rate [LTE Intra System] [CDBH]',
    'PS handover success rate [LTE Inter System] [CBBH]',
    'MV_CSFB Redirection Success Rate [CDBH]','Paging record discarded At eNodeB [CDBH]',
    'Average number of used DL PRBs','VoLTE Call Setup Success rate [CBBH]',
    'VoLTE DCR [CBBH]',
    'VoLTE Packet Loss DL [CBBH]',
    'VoLTE Packet Loss UL [CBBH]','VoLTE BLER UL [CDBH]',
    'VoLTE ERAB Setup Success Rate [CBBH]',
    'No of cells with UL PUSCH SINR >10',
    'VoLTE Inter-Frequency Handover Success Ratio [CBBH]',
    'VoLTE SRVCC SR',]
    raw_kpi_4G=request.FILES['raw_kpi'] if "raw_kpi" in request.FILES else None
    df_raw_kpi=pd.read_excel( raw_kpi_4G)

    required_cols=['RRC Setup Success Rate [CBBH]',
    'ERAB Setup Success Rate [CDBH]',
    'PS Drop Call Rate % [CDBH]',
    'MV_DL User Throughput_Kbps [CDBH]',
    'MV_UL User Throughput_Kbps [CDBH]',
    '4G Data Volume [GB]',
    'E-UTRAN Average CQI [CDBH]','UL PUCCH SINR [CDBH]',
    'PS handover success rate [LTE Intra System] [CDBH]',
    'PS handover success rate [LTE Inter System] [CBBH]',
    'MV_CSFB Redirection Success Rate [CDBH]','Paging record discarded At eNodeB [CDBH]',
    'Average number of used DL PRBs','VoLTE Call Setup Success rate [CBBH]',
    'VoLTE DCR [CBBH]',
    'VoLTE Packet Loss DL [CBBH]',
    'VoLTE Packet Loss UL [CBBH]','VoLTE BLER UL [CDBH]',
    'VoLTE ERAB Setup Success Rate [CBBH]',
    'No of cells with UL PUSCH SINR >10',
    'VoLTE Inter-Frequency Handover Success Ratio [CBBH]',
    'VoLTE SRVCC SR',]
    sts,response=required_col_check(raw_kpi_4G,required_cols)
    if sts:
        return Response(response)


    site_list_4G=request.FILES["site_list"] if "site_list" in request.FILES else None
    df_site_list=pd.read_excel(site_list_4G)
    if site_list_4G:
        sts,response=required_col_check(raw_kpi_4G,required_cols)
        if sts:
            return Response(response)
       
#####################################################################################
    door_path=os.path.join(MEDIA_ROOT,'trends','kol')
    df_raw_kpi['Short name'].fillna(inplace=True, method="ffill")
    df_raw_kpi["Short name"] =df_raw_kpi["Short name"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..
 
    df_raw_kpi.rename( columns={'Unnamed: 1':'date'}, inplace=True )
    df_raw_kpi['MV_DL User Throughput_Kbps [CDBH]']=( df_raw_kpi['MV_DL User Throughput_Kbps [CDBH]']/1024)
    df_raw_kpi.rename(columns={'MV_DL User Throughput_Kbps [CDBH]':'MV_DL User Throughput_Mbps [CDBH]'} ,inplace = True )
    
    print(df_raw_kpi)

    
    split=[]
    tech=[]
    for cell in df_raw_kpi['Short name']:
        if('_' in cell):
            site_id = cell.split("_")[-2][:-1]
            split.append(site_id)
            
        else:
            site_id=cell[:-1]
            split.append(site_id) 
    df_raw_kpi.insert(0,'Site ID', split)       

    ecgi=[]
    # cgi=[]
    for cell1 in df_raw_kpi['4G_ECGI']:
        if('-' in str(cell1)):
            lnbts_id = cell1.split("-")[-2]
            ecgi.append(lnbts_id)   
        else:
            
            lnbts_id=cell1
            ecgi.append(lnbts_id)
    df_raw_kpi.insert(2,'Lnbts_id', ecgi)

    df_raw_kpi["LNBTS NAME"]=[cell2[:-3] if '_' in cell2 else cell2  for cell2 in df_raw_kpi['Short name']]
    df_raw_kpi["tech"]=[cell3.split('_')[2] if '_' in cell3 else cell3 for cell3 in df_raw_kpi["Short name"]]
    tech1=[]
    for tech in df_raw_kpi["tech"]: 
        if ('F1' in tech or 'F3' in tech or 'F8' in tech or 'T1' in tech or 'T2' in tech):
            if("F1" in tech or "F3" in tech or "F8" in tech):
                tech='FDD' 
            if("T1" in tech or'T2' in tech):
                tech="TDD"
            tech1.append(tech)    
        else:
            tech=tech
            tech1.append(tech)
    df_raw_kpi.insert(2,"band",tech1) 

    df_raw_kpi.fillna(value=0,inplace=True)    
    message=site_comparision(ecgi,list(df_site_list['2G ID'])) # site_comparision_call 


    PsOsPath1=os.path.join(door_path,'process output','desird_output.xlsx')    
    df_raw_kpi.to_excel(PsOsPath1,index=False)

    fill_excelfile_1=PsOsPath1


    df1=pd.read_excel(fill_excelfile_1)


    filtered_df_1 = df1[(df1.Lnbts_id.isin(list(df_site_list['2G ID'])))]

    print(filtered_df_1)
    PsOs_Filter=os.path.join(door_path,'process output','filtered.xlsx')
    filtered_df_1.to_excel(PsOs_Filter, index=False)

    df1 = pd.read_excel(PsOs_Filter)
    df_pivot = df1.pivot_table(values=kpi,columns='date', index=['Short name', 'Site ID','Lnbts_id',"LNBTS NAME","band"])
    
    PsOs_Pivot=os.path.join(door_path,"process output','pivot.xlsx")
    df_pivot.to_excel(PsOs_Pivot) 
    
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
	# This process is similar to binary-to-
	# decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    str_date=request.POST.get("offered_date")
    date1=datetime.datetime.strptime(str_date,"%Y-%m-%d")
    print("strdate--------------------------------------",date1)

    # date1 =date(2023,1,27)
    # date1=cal.get_date()
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
    # index=df_pivot.index
    
    STR=os.path.join(door_path,'template','KOL_NEW_FORMAT.xlsx')
    wb=openpyxl.load_workbook(STR)
    ws=wb.active
    
    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        print('donnnnnnnnnnnnnnnnne')
        print(len(index))
        dr=df_pivot[kpi_name]
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()

        ws[coln1+"6"].value=cl[4]
        ws[coln2+"6"].value=cl[3]
        ws[coln3+"6"].value=cl[2]
        ws[coln4+"6"].value=cl[1]
        ws[coln5+"6"].value=cl[0]

        for i,value in enumerate(index):
                j=i+7
                ws['A'+str(j)].value='WB'
                ws['H'+str(j)].value=date1
                # ws['A'+str(j)].value=index[i][2]
                ws['B'+str(j)].value=index[i][1]
                ws['C'+str(j)].value=index[i][2]
                ws['D'+str(j)].value=index[i][3]
                ws['E'+str(j)].value=index[i][0]
                ws['F'+str(j)].value=index[i][4]
            
                
                
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
                
    for kpi_name in kpi:
        
        if(kpi_name=='RRC Setup Success Rate [CBBH]'):
            overwrite(kpi_name,'I') 
            
        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'P')    
            
        if(kpi_name=='PS Drop Call Rate % [CDBH]'):
            overwrite(kpi_name,'W')   
            
        if(kpi_name=='MV_DL User Throughput_Mbps [CDBH]'):
            overwrite(kpi_name,'AD')
            
        if(kpi_name=='MV_UL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name,'AL')
            
        if(kpi_name=='4G Data Volume [GB]'):
            overwrite(kpi_name,'AS')
            
        if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name,'AZ')  
            
        if(kpi_name=='UL PUCCH SINR [CDBH]'):
            overwrite(kpi_name,'BG')
            
        if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'BN') 
            
        if(kpi_name=='PS handover success rate [LTE Inter System] [CBBH]'):
            overwrite(kpi_name,'BU') 
            
        if(kpi_name=='MV_CSFB Redirection Success Rate [CDBH]'):
            overwrite(kpi_name,'CB')
            
        if(kpi_name=='Paging record discarded At eNodeB [CDBH]'):
            overwrite(kpi_name,'CI')
            
        if(kpi_name=='Average number of used DL PRBs'):
            overwrite(kpi_name,'CP')
            
        if(kpi_name=='VoLTE Call Setup Success rate [CBBH]'):
            overwrite(kpi_name,'CW')
            
        if(kpi_name=='VoLTE DCR [CBBH]'):
            overwrite(kpi_name,'DD')
            
        if(kpi_name=='VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name,'DK')
            
        if(kpi_name=='VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name,'DR')
            
        if(kpi_name=='VoLTE BLER UL [CDBH]'):
            overwrite(kpi_name,'DY')
            
        if(kpi_name=='VoLTE ERAB Setup Success Rate [CBBH]'):
            overwrite(kpi_name,'EF') 
            
        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'EM')

        if(kpi_name=='No of cells with UL PUSCH SINR >10'):
            overwrite(kpi_name,'ET')

        if(kpi_name=='VoLTE Inter-Frequency Handover Success Ratio [CBBH]'):
            overwrite(kpi_name,'FB')

        if(kpi_name=='VoLTE SRVCC SR'):
            overwrite(kpi_name,'FI') 

    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border  
    font=Font(size=9)
    alignment = Alignment(horizontal='center', vertical='center')
    for col_cells in ws.columns:
        for cell in col_cells:
            cell.font=font
            cell.alignment=alignment 

    row_index=4
    for cell in ws[row_index]:
        cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')    

    save_output=os.path.join(door_path,'output','KOL_KPI_SAM_TREND_OUTPUT.xlsx')    
    wb.save(save_output)
    download_path=os.path.join(MEDIA_URL,"trends",'kol','output','KOL_KPI_RELOCATION_TREND_OUTPUT.xlsx')
    return Response({"status":True,"message":"uploaded sucessfully","missing_sites":message,'Download_url':download_path}) 
###################################### 2G #######################
@api_view(['POST'])
def old_kol_trend_gsm_2g(request):
    kpi=['SDCCH Blocking Rate [BBH]','SDCCH Drop Call Rate [BBH]','TCH Blocking Rate [BBH]','Drop Call Rate [BBH]',
        'Handover Success Rate [BBH]','RX Quality [BBH]','TCH Drop Call Rate[BBH]','Total Voice Traffic [BBH]']
    raw_kpi_2G=request.FILES['raw_kpi'] if "raw_kpi" in request.FILES else None
    gsm_raw_kpi=pd.read_excel(raw_kpi_2G)

    required_cols=['SDCCH Blocking Rate [BBH]','SDCCH Drop Call Rate [BBH]','TCH Blocking Rate [BBH]','Drop Call Rate [BBH]',
        'Handover Success Rate [BBH]','RX Quality [BBH]','TCH Drop Call Rate[BBH]','Total Voice Traffic [BBH]']
    
    sts,response=required_col_check(raw_kpi_2G,required_cols)
    if sts:
        return Response(response)
    
    site_list_2G=request.FILES["site_list"] if "site_list" in request.FILES else None
    gsm_site_list=pd.read_excel(site_list_2G)

    if  site_list_2G:
        sts,response=required_col_check(raw_kpi_2G,required_cols)
        if sts:
            return Response(response)

    door_path=os.path.join(MEDIA_ROOT,'trends','kol')

    gsm_raw_kpi['Short name']=gsm_raw_kpi['Short name'].fillna(method='ffill')
    gsm_raw_kpi.columns.values[1]='Date'



    split=[]
    for cell in gsm_raw_kpi['Short name']:
        # if('-' in cell):
            site_id = cell[:-1]
            split.append(site_id)
                      
    gsm_raw_kpi.insert(0,'Site ID', split)
    

    gsm_raw_kpi.fillna(value=0,inplace=True)   
    gsm_path=os.path.join(door_path,'process output','gsm_desired.xlsx')   
    gsm_raw_kpi.to_excel(gsm_path,index=False)

    fill_excelfile_1= gsm_path 

    df1=pd.read_excel(fill_excelfile_1)

    df1.rename(columns={"Site ID": "Site_ID"}, inplace=True)

    gsm_filtered = df1[(df1.Site_ID.isin(list(gsm_site_list['Site ID'])))]
    message=site_comparision(split,list(gsm_site_list['Site ID']))######## site comparision #####

    gsm_filter_path=os.path.join(door_path,'process output','2Gfiltered.xlsx')
    gsm_filtered.to_excel(gsm_filter_path, index=False)

    df1 = pd.read_excel(gsm_filter_path)
    gsm_pivot = df1.pivot_table(columns='Date', index=['Short name', 'Site_ID'])

    gsm_pivot_path=os.path.join(door_path,'process output','2Gpivot.xlsx')
    gsm_pivot.to_excel(gsm_pivot_path) 



    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
    # This process is similar to binary-to-
    # decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result

    str_date=request.POST.get("offered_date")
    date1 =datetime.datetime.strptime(str_date,'%Y-%m-%d')
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]

    
    STR=os.path.join(door_path,'template','2gkol.xlsx')
    wb=openpyxl.load_workbook(STR)
    ws=wb.active

    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        index=gsm_pivot.index
        print(len(index))
        dr=gsm_pivot[kpi_name]
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()

        ws[coln1+"4"].value=cl[4]
        ws[coln2+"4"].value=cl[3]
        ws[coln3+"4"].value=cl[2]
        ws[coln4+"4"].value=cl[1]
        ws[coln5+"4"].value=cl[0]

        for i,value in enumerate(index):
                j=i+5
                ws['A'+str(j)].value='WB'
                ws['E'+str(j)].value='2G'
                ws['F'+str(j)].value=date1
                ws['B'+str(j)].value=index[i][1]
                
            
                
                
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
                
    for kpi_name in kpi:
        if(kpi_name=='SDCCH Blocking Rate [BBH]'):
            overwrite(kpi_name,'L') 
            
        if(kpi_name=='SDCCH Drop Call Rate [BBH]'):
            overwrite(kpi_name,'S')    
            
        if(kpi_name=='TCH Blocking Rate [BBH]'):
            overwrite(kpi_name,'Z')   
            
        if(kpi_name=='Drop Call Rate [BBH]'):
            overwrite(kpi_name,'AG')
            
        if(kpi_name=='Handover Success Rate [BBH]'):
            overwrite(kpi_name,'AO')
            
        if(kpi_name=='RX Quality [BBH]'):
            overwrite(kpi_name,'AV')
            
        if(kpi_name=='TCH Drop Call Rate[BBH]'):
            overwrite(kpi_name,'BC')  
            
        if(kpi_name=='Total Voice Traffic [BBH]'):
            overwrite(kpi_name,'BJ')

    save_output=os.path.join(door_path,'output','2G KOL_KPI_TREND_OUTPUT.xlsx')                
    wb.save(save_output)
  
    download_path=os.path.join(MEDIA_URL,"trends",'kol','output','2G KOL_KPI_TREND_OUTPUT.xlsx')
    return Response({"status":True,"message":"uploaded sucessfully","missing_sites":message,'Download_url':download_path}) 