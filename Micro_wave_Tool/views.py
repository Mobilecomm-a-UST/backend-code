from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST
from pathlib import Path
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
import os
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font,Border,Side
from openpyxl.utils import get_column_letter
import re
from .models import MicrowaveAviat
from django.db import transaction
from zipfile import BadZipFile
import numpy as np
from rest_framework import status
import html

 
def get_col(df, col, default=None):
    if col not in df.columns:
        return pd.Series([default]*len(df), index=df.index)
    s = df[col]
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:, 0]
    return s

def to_bool(s):
    return (
        s.astype(str)
        .str.strip()
        .str.lower()
        .map({
            "true": True, "1": True, "yes": True, "ok": True, "pass": True,
            "false": False, "0": False, "no": False, "fail": False,
            "nan": False, "none": False, "": False
        })
        .fillna(False)
    )

def num(s):
    return pd.to_numeric(s, errors="coerce")


def to_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None
    

#function for given color formate in output-----
def format_of_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    header_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    radio_header_fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")

    header_font = Font(color="000000", bold=True)
    normal_font = Font(bold=False)
    center_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    ws.row_dimensions[1].height = 50

    header_list = [cell.value for cell in ws[1]]

    lb_cols = [
        "Circle", "Reference-Key", "Site-ID", "Equipment Make", "Plan Id",
        "Polarization", "Site ID-A", "Tx Frequency (MHz)",
        "BER10e6 Rx Level (dBm)", "Site ID -B", "Rx Frequency (MHz)",
        "Bandwidth (MHz)", "ACM Status", "ACM Min QAM",
        "ACM Max QAM", "ATPC Status", "ATPC MIN", "ATPC MAX",
    ]

    for col_name in lb_cols:
        if col_name in header_list:
            col_idx = header_list.index(col_name) + 1
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = radio_header_fill
            cell.font = header_font
            cell.alignment = center_alignment

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            col_name = header_list[cell.column - 1]

            cell.alignment = center_alignment
  
            if col_name not in lb_cols:
                cell.border = thin_border


    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)


#function to add circle wise data in excel---

def add_to_excel(df, file_path, sheet_name="Sheet1"):
    key_col = "Reference-Key"
    try:
        if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
            existing_df = pd.read_excel(file_path, sheet_name=sheet_name)

            existing_df = existing_df.loc[:, ~existing_df.columns.duplicated()]
            df = df.loc[:, ~df.columns.duplicated()]
            existing_df = existing_df.loc[:, ~existing_df.columns.str.startswith("Soft-AT-Output.")]

            combined_df = pd.concat([existing_df, df], ignore_index=True)
            combined_df = combined_df.loc[:, ~combined_df.columns.duplicated()]
            combined_df = combined_df.loc[:, ~combined_df.columns.str.startswith("Soft-AT-Output.")]

        else:
            combined_df = df.loc[:, ~df.columns.duplicated()]
            combined_df = combined_df.loc[:, ~combined_df.columns.str.startswith("Soft-AT-Output.")]

    except (BadZipFile, FileNotFoundError):
        combined_df = df.loc[:, ~df.columns.duplicated()]
        combined_df = combined_df.loc[:, ~combined_df.columns.str.startswith("Soft-AT-Output.")]

    combined_df.drop_duplicates(subset=[key_col], keep="first", inplace=True)

    combined_df.to_excel(file_path, sheet_name=sheet_name, index=False)

    add_to_db(df)   

    



# funtion for coloring in value
def apply_color(cell, condition):
    green_font = Font(color="00B050")
    red_font = Font(color="FF0000")
    cell.font = green_font if condition else red_font

#function to clean----
def extract_qam(val):
    if pd.isna(val):
        return None
    match = re.search(r'\d+', str(val))
    return int(match.group()) if match else None

#function to mapping reverse-- key
def reverse_key(key):
    if pd.isna(key):
        return None
    key = str(key)
    if "-" not in key:
        return key
    a, b = key.split("-", 1)
    return f"{b}-{a}"




#function for make Polarization column in Radio budget file
def get_polarization(interface):
    interface = str(interface)

    if "Carrier1/1" in interface:
        return "V"
    elif "Carrier1/2" in interface:
        return "H"
    return "Unknown"


#Define folder to save the data_________________               
main_folder = os.path.join(MEDIA_ROOT, 'MicroWave_Tool_Data')
os.makedirs(main_folder, exist_ok=True)
dumy_folder=os.path.join(main_folder, 'MW_Dumy_Output')
os.makedirs(dumy_folder, exist_ok=True)
final_folder=os.path.join(main_folder, 'MW_Final_Output')
os.makedirs(final_folder, exist_ok=True)




@api_view(['POST', 'GET' , 'DELETE'])
def upload_link_budget(request):
    try:
        Link_budget_path = os.path.join(main_folder, 'Link_budget_file')
        os.makedirs(Link_budget_path, exist_ok=True)
 
        if request.method == 'POST':
            files = request.FILES.getlist('link_buget_file')
            if not files:
                return Response({'error': 'link_buget_file files uploaded'}, status=status.HTTP_400_BAD_REQUEST)
 
            for f in files:
                file_path = os.path.join(Link_budget_path, f.name)
                with open(file_path, 'wb+') as destination:
                    for chunk in f.chunks():
                        destination.write(chunk)
 
            return Response({'status': True, 'message': 'Files uploaded and saved successfully'}, status=status.HTTP_200_OK)
 
        elif request.method == 'GET':
            if not os.path.exists(Link_budget_path):
                return Response({'files': []}, status=status.HTTP_200_OK)
 
            files = os.listdir(Link_budget_path)
            return Response({
                'status': True,
                'message': f'{len(files)} Files found in Link_budget_file folder',
                'files': files,
                }, status=status.HTTP_200_OK)
           
        elif request.method == 'DELETE':
            if not os.path.exists(Link_budget_path):
                return Response({'error': 'Folder does not exist'}, status=status.HTTP_404_NOT_FOUND)
 
            deleted_files = []
            for filename in os.listdir(Link_budget_path):
                file_path = os.path.join(Link_budget_path, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    deleted_files.append(filename)
 
            return Response({
                'status': True,
                'message': f'{len(deleted_files)} Files deleted successfully',
                'deleted_files': deleted_files
            }, status=status.HTTP_200_OK)
           
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




@api_view(["POST"])
def microwave_upload(request):
    username = request.user.username
    print(f"Start the Process or Microwave by {username}")

# Upload a Link budget files and read data as a dataframe....
    link_budget_file = os.path.join(main_folder, 'Link_budget_file')
    if not os.path.exists(link_budget_file):
        return Response({"error": "Link_budget_file folder not found"}, status=400)

    link_budget_file_list = []
    for filename in os.listdir(link_budget_file):
        file_path = os.path.join(link_budget_file, filename)
        if filename.endswith('.xlsx'):
            df = pd.read_excel(file_path)
        elif filename.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            return Response(
                {"status": "ERROR", "message": f"Unsupported file type: {filename}"},
                status=HTTP_400_BAD_REQUEST
            )
        df.columns = df.columns.str.strip()
        link_budget_file_list.append(df)

    if link_budget_file_list:
        link_budget_df = pd.concat(link_budget_file_list, ignore_index=True)
    else:
        return Response(
            {"status": "ERROR", "message": "No valid files found"},
            status=HTTP_400_BAD_REQUEST
        ) 

  

# upload Radio-Report files and read data as a dataframe....

    # ------------------- Upload Radio Report Files -------------------
    radio_report_files = request.FILES.getlist("radio_report_file")

    if not radio_report_files:
        return Response(
            {"status": "ERROR", "message": "radio_report_file not uploaded"},
            status=HTTP_400_BAD_REQUEST
        )

    radio_report_file_list = []

    for file in radio_report_files:

        if file.name.endswith(".xlsx"):
            df = pd.read_excel(file, header=2)

        elif file.name.endswith(".csv"):
            df = pd.read_csv(file, header=2)

        else:
            return Response(
                {"status": "ERROR", "message": f"Unsupported file type : {file.name}"},
                status=HTTP_400_BAD_REQUEST
            )

        # Strip column names
        df.columns = df.columns.astype(str).str.strip()

        # Remove duplicate columns (keep first occurrence)
        if df.columns.duplicated().any():
            duplicate_cols = df.columns[df.columns.duplicated()].tolist()
            print(f"Warning: {file.name} contains duplicate columns: {duplicate_cols}")

            df = df.loc[:, ~df.columns.duplicated()]

        # Reset index
        df.reset_index(drop=True, inplace=True)

        radio_report_file_list.append(df)

    # Merge all uploaded radio reports
    radio_report_df = pd.concat(radio_report_file_list, ignore_index=True)

    print("Radio Report Shape :", radio_report_df.shape)
    print(radio_report_df.head())


    # ------------------- Upload Link Report Files -------------------

    link_report_files = request.FILES.getlist("link_report_file")

    if not link_report_files:
        return Response(
            {"status": "ERROR", "message": "link_report_file not uploaded"},
            status=HTTP_400_BAD_REQUEST
        )

    link_report_file_list = []

    for file in link_report_files:

        if file.name.endswith(".xlsx"):
            df = pd.read_excel(file, header=2)

        elif file.name.endswith(".csv"):
            df = pd.read_csv(file, header=2)

        else:
            return Response(
                {"status": "ERROR", "message": f"Unsupported file type : {file.name}"},
                status=HTTP_400_BAD_REQUEST
            )

        # Strip column names
        df.columns = df.columns.astype(str).str.strip()

        # Remove duplicate columns (keep first occurrence)
        if df.columns.duplicated().any():
            duplicate_cols = df.columns[df.columns.duplicated()].tolist()
            print(f"Warning: {file.name} contains duplicate columns: {duplicate_cols}")

            df = df.loc[:, ~df.columns.duplicated()]

        # Reset index
        df.reset_index(drop=True, inplace=True)

        link_report_file_list.append(df)

    # Merge all uploaded link reports
    link_report_df = pd.concat(link_report_file_list, ignore_index=True)

    print("Link Report Shape :", link_report_df.shape)
    print(link_report_df.head())

    


# make Logic and get requried cloumn in MW budget file--------
    # make Logic and get requried cloumn in MW budget file--------
    link_budget_df = link_budget_df[
        link_budget_df["Equipment Make"].astype(str).str.upper() == "AVIAT"
    ]

    link_budget_df = link_budget_df[
        link_budget_df["Plan Stage"].astype(str).str.upper() != "CANCELLED"
    ]


    circle_map = {
        "DEL": "DL",
        "ROB": "WB",
        "MUM": "MB",
        "MAH": "MH",
        "ASM": "AS",
        "UPE": "UE"
    }

    link_budget_df["Circle"] = link_budget_df["Circle"].replace(circle_map)
    link_budget_df["Reference-Key"]=link_budget_df["Circle"]+"MW"+link_budget_df["Site ID-A"]+"-"+link_budget_df["Circle"]+"MW"+link_budget_df["Site ID -B"]
    link_budget_df = link_budget_df[[   
        "Circle",
        "Plan Id",
        "Polarization",
        "Equipment Make",
        "Site ID-A",
        "Tx Frequency (MHz)",
        "BER10e6 Rx Level (dBm)",
        "Site ID -B",
        "Rx Frequency (MHz)",
        "Bandwidth (MHz)",
        "ACM Status",
        "ACM Min QAM",
        "ACM Max QAM",
        "ATPC Status",
        "ATPC MIN",
        "ATPC MAX",
        "Reference-Key"
    ]]


    link_budget_df["Site-ID"] = link_budget_df["Site ID-A"].astype(str) + "-" + link_budget_df["Site ID -B"].astype(str)
    cols = ["Circle", "Reference-Key", "Site-ID","Equipment Make"] + [
        col for col in link_budget_df.columns 
        if col not in ["Circle", "Reference-Key", "Site-ID","Equipment Make"]
    ]
    link_budget_df = link_budget_df[cols]
    output_path = os.path.join(dumy_folder, "link_budget_dumy.xlsx")
    link_budget_df.to_excel(output_path, index=False)
    format_of_excel(output_path)
    # print(link_budget_df.head())


# make Logic and get requried cloumn in the Radio Report file----
    radio_report_df["Radio-Reference-key"]=radio_report_df["Interface"].str.extract(r"^(\S+)")
    rename_map = {
        "RSL Min (dB)": "RSL Min (dBm)",
        "RSL Max (dB)": "RSL Max (dBm)",
        "Tx Power Max (dB)": "Tx Power Max (dBm)",
        "XPD Min (dB)": "XPD Min (dBm)",
        "XPD Max (dB)": "XPD Max (dBm)",
        "SNR Min (dBm)": "SNR Min (dB)"
    }

    radio_report_df.rename(
        columns={k: v for k, v in rename_map.items() if k in radio_report_df.columns},
        inplace=True
    )
    radio_report_df["Polarization(Radio)"] = radio_report_df["Interface"].apply(get_polarization)
    radio_report_df=radio_report_df[[
        "Interface",
        "RSL Min (dBm)",
        "RSL Max (dBm)",
        "Tx Power Max (dBm)",
        "SNR Min (dB)",
        "XPD Min (dBm)",
        "XPD Max (dBm)",
        "Radio-Reference-key",
        "Polarization(Radio)"
        ]]
    radio_report_df= radio_report_df[~radio_report_df["Interface"].str.contains(r"\bRadio1\b", case=False, na=False)]
    radio_report_df["Radio-Reference-key"] = (
    radio_report_df["Radio-Reference-key"].apply(lambda x: html.unescape(x) if isinstance(x, str) else x)
    .str.replace("&", "", regex=False)
)
    output_path = os.path.join(dumy_folder, "radio_report_dumy.xlsx")
    radio_report_df.to_excel(output_path, index=False)
    format_of_excel(output_path)
    # print(radio_report_df.head())


# Process polarization-cloumn, aggregate radio KPIs, and merge results by Radio-Reference-key---
    polarization_df = (
        radio_report_df.groupby("Radio-Reference-key")["Polarization(Radio)"]
        .unique()
        .reset_index()
    )

    def combine_polarization(pols):
        pols = list(pols)
        if set(pols) == {"H", "V"}:
            return "H+V"
        return pols[0]

    polarization_df["Polarization(Radio)"] = polarization_df["Polarization(Radio)"].apply(combine_polarization)


    num_cols_rd = [
        "RSL Min (dBm)", "RSL Max (dBm)", "Tx Power Max (dBm)",
        "SNR Min (dB)", "XPD Min (dBm)", "XPD Max (dBm)"
    ]

    radio_report_df[num_cols_rd] = radio_report_df[num_cols_rd].apply(
        pd.to_numeric, errors="coerce"
    )

    agg_value = {
        "RSL Min (dBm)": "min",
        "RSL Max (dBm)": "max",
        "Tx Power Max (dBm)": "max",
        "SNR Min (dB)": "min",
        "XPD Min (dBm)": "min",
        "XPD Max (dBm)": "max"
    }

    df_radio_group = (
        radio_report_df
        .groupby("Radio-Reference-key", as_index=False)
        .agg(agg_value)  )

    df_radio_group = df_radio_group.merge(
        polarization_df,
        on="Radio-Reference-key",
        how="inner"
    )

# merge {radio file + budget file} based on Reference-Key--- normal merge
    mormal_merge_radio = pd.merge(
        link_budget_df,
        df_radio_group,
        left_on="Reference-Key",
        right_on="Radio-Reference-key",
        how="inner"
    )

# Unmatched  rows (not matched in normal merge)
    unmatch_budget_file = link_budget_df[
        ~link_budget_df["Reference-Key"].isin(
            mormal_merge_radio["Reference-Key"])]
    
    unmatch_radio_file = df_radio_group[
        ~df_radio_group["Radio-Reference-key"].isin(
            mormal_merge_radio["Radio-Reference-key"])].copy()
    

# Reverse only RADIO key (Budget key unchanged)
    unmatch_radio_file["un-Reference-key"] = (
        unmatch_radio_file["Radio-Reference-key"].apply(reverse_key)
    )

# REVERSE MERGE---------------
    reverse_merge_radio = pd.merge(
        unmatch_budget_file,
        unmatch_radio_file,
        left_on="Reference-Key",
        right_on="un-Reference-key",
        how="inner"
    )

#Combine merge NORMAL + REVERSE results---------

    final_merge_radio_budget = pd.concat( [mormal_merge_radio, reverse_merge_radio], ignore_index=True)
    final_merge_radio_budget.drop(columns=["Radio-Reference-key","un-Reference-key"], inplace=True,errors="ignore")
  
    



#make logic and get requried cloumn in the Link Report file----

    link_report_df=link_report_df[[
    "Link Name",
    "Site A Tx Freq (MHz)",
    "Site Z Tx Freq (MHz)",
    "Site A Current RSL",
    "Site Z Current RSL",
    "Site A Modulation Mode",
    "Site Z Modulation Mode",
    "Site A Min Modulation Last 24h",
    "Site Z Min Modulation Last 24h",
    "Site A Max Modulation Last 24h",
    "Site Z Max Modulation Last 24h",
    "Site A Min Configured Modulation",
    "Site Z Min Configured Modulation",
    "Site A Max Configured Modulation",
    "Site Z Max Configured Modulation",
    "ATPC Status",
    ]]

    link_report_df.rename(columns={
        "Site A Tx Freq (MHz)":"FREQ TX",
        "Site Z Tx Freq (MHz)":"FREQ RX",
        "ATPC Status":"ATPC Status (Link)"},
        inplace=True)
    

    num_cols_lk=[ "Site A Current RSL","Site Z Current RSL",]
    link_report_df[num_cols_lk]=link_report_df[num_cols_lk].apply(pd.to_numeric, errors="coerce")
    agg_value = {
        "Site A Current RSL": "max",
        "Site Z Current RSL": "max",
        "FREQ TX": "first",
        "FREQ RX": "first",
        "Site A Modulation Mode": "first",
        "Site Z Modulation Mode": "first",
        "Site A Min Modulation Last 24h": "first",
        "Site Z Min Modulation Last 24h": "first",
        "Site A Max Modulation Last 24h": "first",
        "Site Z Max Modulation Last 24h": "first",
        "Site A Min Configured Modulation": "first",
        "Site Z Min Configured Modulation": "first",
        "Site A Max Configured Modulation": "first",
        "Site Z Max Configured Modulation": "first",
        "ATPC Status (Link)": "first",
        }
    df_link_group=link_report_df.groupby("Link Name", as_index=False).agg(agg_value)
    
    output_path=os.path.join(dumy_folder, "link_report_dumy.xlsx")
    df_link_group.to_excel(output_path, index=False)
    format_of_excel(output_path)

    

#merge to mergerd MW budget+radio and link budget--- finallmerge->(Mwbudget+radio)+linkbudget normarl merge 
    df_link_group = df_link_group.copy()
    df_link_group["reverse_key"] = df_link_group["Link Name"].apply(reverse_key)

    final_df = pd.merge( 
        final_merge_radio_budget, 
        df_link_group, 
        left_on="Reference-Key", 
        right_on="Link Name", 
        how="left" 
        ) 
    
    mask = final_df["Link Name"].isna()
    
    # reverse merge (append to same final) 
    reverse_matched = pd.merge( 
        final_merge_radio_budget.loc[mask],
        df_link_group, 
        left_on="Reference-Key", 
        right_on="reverse_key", 
        how="left" 
        )


    final_df.loc[mask, reverse_matched.columns] = reverse_matched.values
    final_df.drop(columns=["Link Name","reverse_key"], errors="ignore", inplace=True) 
    # final_df.to_excel("abhinav.xlsx")


    # all condtion to make remark -----------------------------------------

    final_df.columns = final_df.columns.astype(str).str.strip()
    final_df = final_df.loc[:, ~final_df.columns.duplicated()]

    final_df["Tx Frequency (MHz)"] = num(get_col(final_df,"Tx Frequency (MHz)")).round().astype("Int64")
    final_df["Rx Frequency (MHz)"] = num(get_col(final_df,"Rx Frequency (MHz)")).round().astype("Int64")
    final_df["FREQ TX"] = num(get_col(final_df,"FREQ TX")).round().astype("Int64")
    final_df["FREQ RX"] = num(get_col(final_df,"FREQ RX")).round().astype("Int64")

    tx_lb = final_df["Tx Frequency (MHz)"]
    rx_lb = final_df["Rx Frequency (MHz)"]

    tx_dump = final_df["FREQ TX"]
    rx_dump = final_df["FREQ RX"]

    freq_status = (
        ((tx_dump == tx_lb) & (rx_dump == rx_lb)) |
        ((tx_dump == rx_lb) & (rx_dump == tx_lb))
    )

    final_df["feq_TX"] = freq_status
    final_df["feq_RX"] = freq_status


   
    rx = num(get_col(final_df,"BER10e6 Rx Level (dBm)"))

    rsl_min = num(get_col(final_df,"RSL Min (dBm)"))
    rsl_max = num(get_col(final_df,"RSL Max (dBm)"))

    final_df["RSL_Status"] = (rx >= rsl_min-3) & (rx <= rsl_max+3)

    siteA = num(get_col(final_df,"Site A Current RSL"))
    siteZ = num(get_col(final_df,"Site Z Current RSL"))

    final_df["curret_rsl_status_A"] = (rx >= siteA-3) & (rx <= siteA+3)
    final_df["curret_rsl_status_Z"] = (rx >= siteZ-3) & (rx <= siteZ+3)



    xpd_min = num(get_col(final_df,"XPD Min (dBm)"))
    xpd_max = num(get_col(final_df,"XPD Max (dBm)"))

    final_df["Xpd_satus"] = xpd_min.between(22,36) & xpd_max.between(22,36)



    atpc = num(get_col(final_df,"ATPC MAX"))
    txmax = num(get_col(final_df,"Tx Power Max (dBm)"))

    final_df["Tx_power_status"] = atpc.notna() & txmax.notna() & (np.floor(atpc)==np.floor(txmax))


  
    final_df["SNR_Status"] = num(get_col(final_df,"SNR Min (dB)")) > 40
    final_df["polarization_status"] = get_col(final_df,"Polarization(Radio)").astype(str) == get_col(final_df,"Polarization").astype(str)



    mod_cols = [
        "Site A Min Modulation Last 24h","Site Z Min Modulation Last 24h",
        "Site A Max Modulation Last 24h","Site Z Max Modulation Last 24h",
        "Site A Max Configured Modulation","Site Z Max Configured Modulation"
    ]

    def mod_status(row):
        acm = re.sub(r"\D", "", str(row.get("ACM Max QAM", "")))

        for c in mod_cols:
            v = row.get(c, None)
            if isinstance(v, pd.Series):
                v = v.iloc[0]

            if acm != re.sub(r"\D", "", str(v)):
                return False

        return True

    final_df["Modulation_status"] = final_df.apply(mod_status,axis=1)


    mod_config_cols=["Site A Min Configured Modulation","Site Z Min Configured Modulation"]

    def mod_config(row):
        acm=str(row.get("ACM Min QAM","")).lower()
        for c in mod_config_cols:
            v=row.get(c,None)
            if isinstance(v,pd.Series): v=v.iloc[0]
            if pd.notna(v) and acm in str(v).lower(): return True
        return False

    final_df["Modulation_config_status"]=final_df.apply(mod_config,axis=1)


    mod_mode_cols=["Site A Modulation Mode","Site Z Modulation Mode"]
    def mod_mode(row):
        atpc=str(row.get("ATPC Status","")).lower()
        for c in mod_mode_cols:
            v=str(row.get(c,"")).lower()
            if (atpc=="enable" and v=="adaptive") or (atpc=="disable" and v=="fixed"):
                return True
        return False

    final_df["Modulation_moud_status"]=final_df.apply(mod_mode,axis=1)



    far_cols=["feq_RX","feq_TX","RSL_Status",
              "curret_rsl_status_A",
              "curret_rsl_status_Z",
              "Xpd_satus"]
    far_df=pd.DataFrame({c:to_bool(get_col(final_df,c)) for c in far_cols})
    final_df["FAR_Remark"]=np.where(far_df.all(axis=1),pd.NA,"Field Action Required")


    bar_cols=["Tx_power_status","SNR_Status",
              "polarization_status",
              "Modulation_status","Modulation_config_status",
              "Modulation_moud_status"]
    bar_df=pd.DataFrame({c:to_bool(get_col(final_df,c)) for c in bar_cols})
    final_df["BAR_Remark"]=np.where(bar_df.all(axis=1),pd.NA,"Backend Action Required")


   
    equip=get_col(final_df,"Equipment Make").astype(str).str.upper()
    final_df["oem_remark"]=np.where(equip=="AVIAT",pd.NA,"OEM Mismatch")


 
    def join_remark(row):
        vals=[str(v).strip() for v in row if pd.notna(v) and str(v).strip()!=""]
        return ", ".join(vals) if vals else "Ready to offer"

    final_df["Remark"]=final_df[["BAR_Remark","FAR_Remark","oem_remark"]].apply(join_remark,axis=1)

    final_df["Soft-AT-Output"] = np.where(
    final_df[["BAR_Remark", "FAR_Remark", "oem_remark"]].isna().all(axis=1),
    "Ready to offer",
    "Action Required"
)
            
    
    final_df.to_excel("Avait.xlsx",index=False)
    final_df.drop(columns=["feq_RX", "feq_TX", "RSL_Status", "curret_rsl_status_A",
    "curret_rsl_status_Z", "Xpd_satus", "Tx_power_status", "SNR_Status",
    "polarization_status", "Modulation_status", "Modulation_config_status",
    "Modulation_moud_status", "FAR_Remark", "BAR_Remark", "oem_remark","Remark","Soft-AT-Output"], inplace=True, errors="ignore")

    final_df.drop_duplicates(inplace=True)
    print(final_df.head())
    output_path=os.path.join(final_folder, "Microwave_final(AVIAT).xlsx")
    add_to_excel(final_df,output_path)
    format_of_excel(output_path)
    
        


 # work of color formate of red and green as per condition--

    wb = load_workbook(output_path)
    ws = wb.active

    cols = {c.value: i+1 for i, c in enumerate(ws[1])}
    for r in range(2, ws.max_row + 1):

        Polarization = (
            ws.cell(r, cols["Polarization"]).value == ws.cell(r, cols["Polarization(Radio)"]).value)
        apply_color(ws.cell(r, cols["Polarization(Radio)"]),Polarization)

        tx = (
        ws.cell(r, cols["Tx Frequency (MHz)"]).value == ws.cell(r, cols["FREQ TX"]).value
        or ws.cell(r, cols["Tx Frequency (MHz)"]).value == ws.cell(r, cols["FREQ RX"]).value
        )
        apply_color(ws.cell(r, cols["FREQ TX"]), tx)

        rx = (
            ws.cell(r, cols["Rx Frequency (MHz)"]).value == ws.cell(r, cols["FREQ RX"]).value
            or ws.cell(r, cols["Tx Frequency (MHz)"]).value == ws.cell(r, cols["FREQ RX"]).value
        )
        apply_color(ws.cell(r, cols["FREQ RX"]), rx)



        atpc = ws.cell(r, cols["ATPC MAX"]).value
        tx_max = ws.cell(r, cols["Tx Power Max (dBm)"]).value
        tx_power = (
            isinstance(atpc, (int, float)) and
            isinstance(tx_max, (int, float)) and
            int(atpc) == int(tx_max)
        )
        apply_color(ws.cell(r, cols["Tx Power Max (dBm)"]), tx_power)
 


    
        
       
        cell = ws.cell(r, cols["ATPC Status (Link)"])
        value = cell.value
        if value in (1, "1", True):
                cell.value = True
        elif value in (0, "0", False):
                cell.value = False


        status = ws.cell(r, cols["ATPC Status (Link)"]).value
        if status is True:
            apply_color(ws.cell(r, cols["ATPC Status (Link)"]), True)
        elif status is False:
            apply_color(ws.cell(r, cols["ATPC Status (Link)"]), False)





      
        snr = ws.cell(r, cols["SNR Min (dB)"]).value
        cond_snr = isinstance(snr, (int, float)) and snr > 40
        apply_color(ws.cell(r, cols["SNR Min (dB)"]), cond_snr)

      



        for col_name in ["XPD Min (dBm)", "XPD Max (dBm)"]:
            cell = ws.cell(r, cols[col_name])
            val = cell.value
            if not isinstance(val, (int, float)):
                continue
            if 22 <= val <= 35.9:
                apply_color(cell, True)

            elif (20 <= val <= 22) or (35.9 <= val <= 38):
                apply_color(cell, False)

            else:
                cell.font = Font(color="C00000", bold=True)



        ber = to_float(ws.cell(r, cols["BER10e6 Rx Level (dBm)"]).value)
        rsl_min = to_float(ws.cell(r, cols["RSL Min (dBm)"]).value)
        rsl_max = to_float(ws.cell(r, cols["RSL Max (dBm)"]).value)
        rsl_a = to_float(ws.cell(r, cols["Site A Current RSL"]).value)
        rsl_z = to_float(ws.cell(r, cols["Site Z Current RSL"]).value)


        condition = (
            all(x is not None for x in [ber, rsl_a, rsl_z]) and
            (rsl_a - 3) <= ber <= (rsl_a + 3) and
            (rsl_z - 3) <= ber <= (rsl_z + 3)
        )

        apply_color(ws.cell(r, cols["Site A Current RSL"]), condition)
        apply_color(ws.cell(r, cols["Site Z Current RSL"]), condition)


        # RSL Min (Deviation = RSL - BER)
        min_condition = False
        if ber is not None and rsl_min is not None:
            deviation = rsl_min - ber
            min_condition = -3.999 <= deviation <= 3.999

        apply_color(ws.cell(r, cols["RSL Min (dBm)"]), min_condition)


        # RSL Max (Deviation = RSL - BER)
        max_condition = False
        if ber is not None and rsl_max is not None:
            deviation = rsl_max - ber
            max_condition = -3.999 <= deviation <= 3.999

        apply_color(ws.cell(r, cols["RSL Max (dBm)"]), max_condition)



        acm_max = extract_qam(ws.cell(r, cols["ACM Max QAM"]).value)
        for col_name in [
            "Site A Max Modulation Last 24h","Site Z Max Modulation Last 24h",
            "Site A Min Modulation Last 24h","Site Z Min Modulation Last 24h",
            "Site A Max Configured Modulation","Site Z Max Configured Modulation"]:
      

            max_mod = extract_qam(ws.cell(r, cols[col_name]).value)

            condition = (
                max_mod is not None and
                acm_max is not None and
                max_mod == acm_max
            )

            apply_color(ws.cell(r, cols[col_name]), condition)


        acm_min = ws.cell(r, cols["ACM Min QAM"]).value
        if isinstance(acm_min, str):
            acm_min = acm_min.strip().lower()

        for col_name in ["Site A Min Configured Modulation", "Site Z Min Configured Modulation"]:
            min_mod = ws.cell(r, cols[col_name]).value
            if isinstance(min_mod, str):
                min_mod = min_mod.strip().lower()

            condition = (
                min_mod is not None and
                acm_min is not None and
                min_mod == acm_min   # case-insensitive comparison
            )

            apply_color(ws.cell(r, cols[col_name]), condition)


        acm = ws.cell(r, cols["ACM Status"]).value

        if isinstance(acm, str):
            acm = acm.strip().lower()

        for col_name in [
            "Site A Modulation Mode",
            "Site Z Modulation Mode"]:
            mod = ws.cell(r, cols[col_name]).value
            if isinstance(mod, str):
                mod = mod.strip().lower()

            condition = (
                (acm == "enable" and mod == "adaptive") or
                (acm == "disable" and mod == "fixed")
            )
            apply_color(ws.cell(r, cols[col_name]), condition)


    remarks_col = ws.max_column + 1
    headers = {cell.value: cell.column for cell in ws[1]}
    if "Soft-AT-Output" in headers:
        remarks_col = headers["Soft-AT-Output"]
    else:
        remarks_col = ws.max_column + 1
        ws.cell(row=1, column=remarks_col).value = "Soft-AT-Output"

    RED_COLORS = {
        "FF0000",      # RGB Red
        "FFFF0000",    # ARGB Red
        "C00000",      # Dark Red
        "FFC00000"     # ARGB Dark Red
    }

    for r in range(2, ws.max_row + 1):
        remark = "Ready to offer"   # Default

        for c in range(1, remarks_col):
            cell = ws.cell(r, c)

            if cell.font.color is None:
                continue

            if cell.font.color.type == "rgb":
                color = cell.font.color.rgb

                if color:
                    color = color.upper()

                    # Match all supported red colors
                    if (
                        color in RED_COLORS
                        or color.endswith("FF0000")
                        or color.endswith("C00000")
                    ):
                        remark = "Action Required"
                        break

        ws.cell(row=r, column=remarks_col).value = remark

    wb.save(output_path)
    print("End The Process..?")  




    
    relative_path = os.path.relpath(output_path, MEDIA_ROOT).replace("\\", "/")
    download_url = request.build_absolute_uri(MEDIA_URL + relative_path)

    return Response({
        "status": True, 
        "message": "Files uploaded successfully",
        "download_url":download_url 

    }) 


#function for add data in database.......
@transaction.atomic
def add_to_db(df):
    for _, row in df.iterrows():
        MicrowaveAviat.objects.update_or_create(
            reference_key=row.get("Reference-Key"),
            circle=row.get("Circle"), 
            defaults={
                "site_id": row.get("Site-ID"),
                "equipment_make": row.get("Equipment Make"),
                "plan_id": row.get("Plan Id"),
                "polarization": row.get("Polarization"),
                "site_id_a": row.get("Site ID-A"),
                "tx_frequency_mhz": row.get("Tx Frequency (MHz)"),
                "ber10e6_rx_level_dbm": row.get("BER10e6 Rx Level (dBm)"),
                "site_id_b": row.get("Site ID -B"),
                "rx_frequency_mhz": row.get("Rx Frequency (MHz)"),
                "bandwidth_mhz": row.get("Bandwidth (MHz)"),
                "acm_status": row.get("ACM Status"),
                "acm_min_qam": row.get("ACM Min QAM"),
                "acm_max_qam": row.get("ACM Max QAM"),
                "atpc_status": row.get("ATPC Status"),
                "atpc_min": row.get("ATPC MIN"),
                "atpc_max": row.get("ATPC MAX"),
                "rsl_min_dbm": row.get("RSL Min (dBm)"),
                "rsl_max_dbm": row.get("RSL Max (dBm)"),
                "tx_power_max_dbm": row.get("Tx Power Max (dBm)"),
                "snr_min_db": row.get("SNR Min (dB)"),
                "xpd_min_dbm": row.get("XPD Min (dBm)"),
                "xpd_max_dbm": row.get("XPD Max (dBm)"),
                "polarization_radio": row.get("Polarization(Radio)"),
                "site_a_current_rsl": row.get("Site A Current RSL"),
                "site_z_current_rsl": row.get("Site Z Current RSL"),
                "freq_tx": row.get("FREQ TX"),
                "freq_rx": row.get("FREQ RX"),
                "site_a_modulation_mode": row.get("Site A Modulation Mode"),
                "site_z_modulation_mode": row.get("Site Z Modulation Mode"),
                "site_a_min_mod_last_24h": row.get("Site A Min Modulation Last 24h"),
                "site_z_min_mod_last_24h": row.get("Site Z Min Modulation Last 24h"),
                "site_a_max_mod_last_24h": row.get("Site A Max Modulation Last 24h"),
                "site_z_max_mod_last_24h": row.get("Site Z Max Modulation Last 24h"),
                "site_a_min_configured_mod": row.get("Site A Min Configured Modulation"),
                "site_z_min_configured_mod": row.get("Site Z Min Configured Modulation"),
                "site_a_max_configured_mod": row.get("Site A Max Configured Modulation"),
                "site_z_max_configured_mod": row.get("Site Z Max Configured Modulation"),
                "atpc_status_link": row.get("ATPC Status (Link)"),
                "remark":row.get("Soft-AT-Output")
            }
        )  

# # api to get and delete data in databse...
# @api_view(['GET'])
# def final_excel_indb(request):
#     queryset = MicrowaveAviat.objects.all()
#     site_id = request.data.get("site_id")
#     circle = request.data.get("circle")
#     equipment = request.data.get("equipment_make")

#     if site_id:
#         queryset = queryset.filter(site_id=site_id)

#     if circle:
#         queryset = queryset.filter(circle=circle)

#     if equipment:
#         queryset = queryset.filter(equipment_make=equipment)

#     data = queryset.values()

#     return Response({
#         "status": True,
#         "count": queryset.count(),
#         "data": data
#     }, status=200)


from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@api_view(['GET', 'POST'])
def final_excel_indb(request):
    queryset = MicrowaveAviat.objects.all()

   
    if request.method == 'POST':
        site_id = request.data.get("site_id")
        circle = request.data.get("circle")
        equipment = request.data.get("equipment_make")

    # (Optional) GET support bhi rehne do
    else:
        site_id = request.GET.get("site_id")
        circle = request.GET.get("circle")
        equipment = request.GET.get("equipment_make")

    if site_id:
        queryset = queryset.filter(site_id__icontains=site_id)

    if circle:
        queryset = queryset.filter(circle=circle)

    if equipment:
        queryset = queryset.filter(equipment_make__icontains=equipment)

    data = queryset.values()    

    return Response({
        "status": True,
        "count": queryset.count(),
        "data": data
    }, status=200)



#---------

@api_view(["GET", "DELETE"])
def get_delete_file(request):
    folder_path = os.path.join(main_folder, "MW_Final_Output")

    files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]
    if not files:
        return Response(
            {
                "status": False,
                "error": "No file found in folder"
            },
            status=status.HTTP_404_NOT_FOUND
        )

    file_name = files[0]
    file_path = os.path.join(folder_path, file_name)
    if request.method == "GET":
        relative_path = os.path.relpath(file_path, MEDIA_ROOT).replace("\\", "/")
        download_file = request.build_absolute_uri(MEDIA_URL + relative_path)

        return Response(
            {
                "status": True,
                "file_url": download_file
            },
            status=status.HTTP_200_OK
        )

    elif request.method == "DELETE":
        username = request.user.username
        print("user:", username)

        allowed_users = ["Abhinav", "Abhinav.Verma@ust.com","289581@ust.com"]

        if username not in allowed_users:
            return Response(
                {
                    "status": False,
                    "error": "You are not authorized for this action"
                },
                status=status.HTTP_403_FORBIDDEN
            )

        with transaction.atomic():
            if os.path.exists(file_path):
                os.remove(file_path)

            deleted_count, _ = MicrowaveAviat.objects.all().delete()

        return Response(
            {
                "status": True,
                "message": "File and database records deleted successfully",
                "excel_deleted": file_name,
                "db_rows_deleted": deleted_count
            },
            status=status.HTTP_200_OK
        )
    

#auto-----------------------------
@api_view(["POST"])
def auto_delete_reports(request):

    folder_path = os.path.join(main_folder, "MW_Final_Output")

    files = []
    if os.path.exists(folder_path):
        files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]

    with transaction.atomic():

        for file_name in files:
            file_path = os.path.join(folder_path, file_name)

            if os.path.exists(file_path):
                os.remove(file_path)

        deleted_count, _ = MicrowaveAviat.objects.all().delete()

    return Response(
        {
            "status": True,
            "excel_files_deleted": len(files),
            "db_rows_deleted": deleted_count,
            "message": "Reports deleted successfully"
        },
        status=status.HTTP_200_OK
    )    