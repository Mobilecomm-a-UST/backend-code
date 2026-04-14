from django.http import JsonResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from upgrade_tracker.models import *  # noqa: F403
from alok_tracker.models import RelocationUser  # noqa: F403
import pandas as pd
from django.db import transaction
import os
from django.conf import settings
from datetime import datetime as dtime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.utils import timezone
import json
from datetime import datetime, timedelta, date
import shutil
from django.db.models import Q
from django.forms.models import model_to_dict
import numpy as np
from rest_framework import status

ALL_COLUMNS = [
    "Circle",
    "Upgrade Technology",
    "Site id",
    "Cell id",
    "TOCO Name",
    "SR Number",
    "RAN OEM",
    "Media Type",
    "Band",
    "RFAI Date",
    "Allocation Date",
    "RFAI Survey Date",
    "Survey (Accepted/Rejected)",
    "WRFAI (YES/NO)",
    "NWRFAI Reason(Material, Media, PRI)",
    "Remarks For NWRFAI",
    "NWRFAI To WRFAI Date",
    "MO Punch Date",
    "Material Dispatch Date",
    "Material Delivered Date",
    "Installation Start Date",
    "Installation End Date",
    "Integration Date",
    "NEP ID",
    "EMF Submission Date",
    "RAN LKF Status",
    "Alarm Status",
    "Alarm Rectification Done Date",
    "SCFT Done Date",
    "SCFT I-Deploy Offered Date",
    "RAN PAT Offer Date",
    "RAN SAT Offer Date",
    "Site ONAIR Date",
    "I-Deploy ONAIR Date",
    "Current Status",
    "Detailed Remarks",
    "Manual History",
    "RFAI Rejected Date",
    "Re-RFAI Date",
    "PRI Count",
    "PRI Issue Ageing",
    "Other Issue Ageing",
    "Total Issue Ageing",
    "RFAI to MS1 Ageing",
    "RAN PAT Accepted Date",
    "RAN SAT Accepted Date",
    "SCFT Accepted Date",
    "KPI AT offer Date",
    "KPI AT Accepted Date",
    "4G MS2 Date",
    "5G MS2 Date",
    "Final MS2 Date",
    "Last Updated Date",
    "Last Updated By",
]


COLUMN_MAP = {
    "Circle": "circle",
    "Upgrade Technology": "upgrade_technology",
    "Site id": "site_id",
    "Cell id": "cell_id",
    "TOCO Name": "toco_name",
    "SR Number": "sr_number",
    "RAN OEM": "ran_oem",
    "Media Type": "media_type",
    "Band": "band",
    "RFAI Date": "rfai_date",
    "Allocation Date": "allocation_date",
    "RFAI Survey Date": "rfai_survey_date",
    "Survey (Accepted/Rejected)": "survey_status",
    "WRFAI (YES/NO)": "wrfai",
    "NWRFAI Reason(Material, Media, PRI)": "nwrfai_reason",
    "Remarks For NWRFAI": "remarks_for_nwrfai",
    "NWRFAI To WRFAI Date": "nwrfai_to_wrfai_date",
    "MO Punch Date": "mo_punch_date",
    "Material Dispatch Date": "material_dispatch_date",
    "Material Delivered Date": "material_delivered_date",
    "Installation Start Date": "installation_start_date",
    "Installation End Date": "installation_end_date",
    "Integration Date": "integration_date",
    "NEP ID": "nep_id",
    "EMF Submission Date": "emf_submission_date",
    "RAN LKF Status": "ran_lkf_status",
    "Alarm Status": "alarm_status",
    "Alarm Rectification Done Date": "alarm_rectification_done_date",
    "SCFT Done Date": "scft_done_date",
    "SCFT I-Deploy Offered Date": "scft_i_deploy_offered_date",
    "RAN PAT Offer Date": "ran_pat_offer_date",
    "RAN SAT Offer Date": "ran_sat_offer_date",
    "Site ONAIR Date": "site_onair_date",
    "I-Deploy ONAIR Date": "i_deploy_onair_date",
    "Current Status": "current_status",
    "Detailed Remarks": "detailed_remarks",
    "Manual History": "manual_history",
    "RFAI Rejected Date": "rfai_rejected_date",
    "Re-RFAI Date": "re_rfai_date",
    "PRI Count": "pri_count",
    "PRI Issue Ageing": "pri_issue_ageing",
    "Other Issue Ageing": "other_issue_ageing",
    "Total Issue Ageing": "total_issue_ageing",
    "RFAI to MS1 Ageing": "rfai_to_ms1_ageing",
    "RAN PAT Accepted Date": "ran_pat_accepted_date",
    "RAN SAT Accepted Date": "ran_sat_accepted_date",
    "SCFT Accepted Date": "scft_accepted_date",
    "KPI AT offer Date": "kpi_at_offer_date",
    "KPI AT Accepted Date": "kpi_at_accepted_date",
    "4G MS2 Date": "four_g_ms2_date",
    "5G MS2 Date": "five_g_ms2_date",
    "Final MS2 Date": "final_ms2_date",
    "Last Updated Date": "last_updated_date",
    "Last Updated By": "last_updated_by",
}



# Create your views here.


def compute_union_ageing(rows):
    if rows.empty:
        return 0

    today = datetime.today()

    def to_datetime(x):
        if isinstance(x, datetime):
            return x
        if isinstance(x, date):
            return datetime.combine(x, datetime.min.time())
        if isinstance(x, str) and x not in ["", "-", "nan", None]:
            return datetime.strptime(x, "%d-%b-%y")
        return None

    ranges = []

    for _, row in rows.iterrows():
        start = to_datetime(row["Start Date"])
        close = row["Close Date"]
        # close = to_datetime(close)

        end = today if close in ["", "-", "nan", None, "undefined"] else close
        end = to_datetime(end)
        ranges.append((start, end))

    # Sort
    ranges.sort(key=lambda x: x[0])

    # Merge intervals
    merged = []
    current_start, current_end = ranges[0]

    for start, end in ranges[1:]:
        if start <= current_end:
            current_end = max(current_end, end)
        else:
            merged.append((current_start, current_end))
            current_start, current_end = start, end

    merged.append((current_start, current_end))

    return sum((end - start).days for start, end in merged)


def update_ageing_new(circle, site_id, sr_number):
    # Fetch tracker row
    data_obj = UpgradeTracker.objects.filter(
        circle=circle,
        cell_id=site_id,
        sr_number=sr_number
    ).first()

    # if not data_obj:
    #     return None

    # Fetch issues
    issue_obj = UpgradeIssue.objects.filter(
        circle=circle,
        site_id=site_id,
        sr_number=sr_number
    )

    issue_df = pd.DataFrame(issue_obj.values())
    
    rename_map = {
        "circle": "Circle",
        "site_id": "Site ID",
        "sr_number": "SR Number",
        "issue_owner": "Issue Owner",
        "milestone": "Milestone",
        "issue_name": "Issue Name",
        "start_date": "Start Date",
        "close_date": "Close Date",
        "status": "Status",
        "duration": "Duration",
        "updated_by": "Updated_by",
        "updated_at": "Updated_at",
        "created_by": "Created_by",
        "created_at": "Created_at"
    }

    print(issue_df)
    
    issue_df = issue_df.rename(columns=rename_map)

    if issue_df.empty:
        pri_ageing = 0
        other_issue_ageing = 0
        total_issue_ageing = 0
        pri_count = 0
    else:
        pri_rows = issue_df[issue_df["Issue Name"] == "PRI"]
        other_rows = issue_df[issue_df["Issue Name"] != "PRI"]

        pri_ageing = compute_union_ageing(pri_rows)
        other_issue_ageing = compute_union_ageing(other_rows)
        total_issue_ageing = compute_union_ageing(issue_df)
        pri_count = len(pri_rows)

    # Assign ageing values
    data_obj.pri_issue_ageing = pri_ageing
    data_obj.other_issue_ageing = other_issue_ageing
    data_obj.total_issue_ageing = total_issue_ageing
    data_obj.pri_count = pri_count

    # Convert dates
    site_onair = pd.to_datetime(data_obj.site_onair_date, errors="coerce")
    re_rfai = pd.to_datetime(data_obj.re_rfai_date, errors="coerce")
    rfai = pd.to_datetime(data_obj.rfai_date, errors="coerce")

    rfai_final = re_rfai if pd.notna(re_rfai) else rfai

    # Calculate rfai_to_ms1_ageing
    if pd.isna(rfai_final):
        data_obj.rfai_to_ms1_ageing = "-"
    else:
        issue_ageing = total_issue_ageing or 0

        if pd.isna(site_onair):
            today = datetime.today()
            data_obj.rfai_to_ms1_ageing = (today - rfai_final).days - issue_ageing
        else:
            data_obj.rfai_to_ms1_ageing = (site_onair - rfai_final).days - issue_ageing

    # Save changes
    data_obj.save()

    return
 


############################################################  UPLOAD DATA #####################################################################################

@api_view(["POST"])
def upload_tracker_data_view(request):

    user_id = request.data.get("userId")
    if not user_id:
        return Response({"error": "userId required"}, status=400)

    user = RelocationUser.objects.filter(email__iexact=user_id).first()
    if not user:
        return Response({"error": "User not found"}, status=404)

    if user.right == "Read":
        return Response({"error": "ACCESS DENIED"}, status=403)

    file = request.data.get("file")
    if not file:
        return Response({"error": "No file uploaded"}, status=400)

    try:
        # ================= Read File =================
        df = (
            pd.read_csv(file, header=1)
            if file.name.endswith(".csv")
            else pd.read_excel(file, header=1)
        )

        df.columns = df.columns.str.strip()
        df = df.replace(["NaN", "nan", "", "NA", "N/A", "-"], None)

        # ================= Circle Filtering =================
        circles = user.circles
        if "CENTRAL" not in circles:
            df = df[df["Circle"].isin(circles)]

        # ================= Validate Columns =================
        missing = [c for c in ALL_COLUMNS if c not in df.columns]
        if missing:
            return Response(
                {"error": f"Missing columns: {', '.join(missing)}"},
                status=400
            )

        # ================= Rename Excel → Model =================
        df = df.rename(columns=COLUMN_MAP)
        df = df[list(COLUMN_MAP.values())]

        # ================= Safe Date Parsing =================
        INVALID_DATE_STRINGS = {"need to check", "pending", "na", "tbd", "n/a", "-"}

        def safe_date(val):
            if val is None:
                return None
            if isinstance(val, pd.Timestamp):
                return val.date()
            if isinstance(val, str):
                v = val.strip().lower()
                if not v or v in INVALID_DATE_STRINGS:
                    return None
                parsed = pd.to_datetime(val, errors="coerce")
                return None if pd.isna(parsed) else parsed.date()
            return None

        for col in df.columns:
            if col.endswith("_date"):
                df[col] = df[col].apply(safe_date)

        # ================= Integers =================
        INT_FIELDS = [
            "pri_count",
            "pri_issue_ageing",
            "other_issue_ageing",
            "total_issue_ageing",
            "rfai_to_ms1_ageing",
        ]

        for col in INT_FIELDS:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")

        # ================= Convert to Records =================
        records = df.to_dict("records")

        # ================= Existing Records (SR NUMBER UNIQUE) =================
        
        def normalize_sr(val):
            """
            Normalizes SR Number for safe comparison.
            Handles str, int, float, NaN, None.
            """
            if val is None:
                return None

            # pandas NaN
            if isinstance(val, float) and pd.isna(val):
                return None

            # 12345.0 -> "12345"
            if isinstance(val, float):
                if val.is_integer():
                    return str(int(val)).strip().lower()
                return str(val).strip().lower()

            return str(val).strip().lower()
        
        existing = {
            normalize_sr(x.sr_number): x
            for x in UpgradeTracker.objects.only("sr_number")
            if normalize_sr(x.sr_number)
        }

        to_create, to_update = [], []

        with transaction.atomic():
            for row in records:
                sr_number = row.get("sr_number")

                if not sr_number:
                    continue

                row["last_updated_by"] = user.email
                row["last_updated_date"] = timezone.now()

                sr_number = row.get("sr_number")
                key = normalize_sr(sr_number)

                if not key:
                    continue

                if key in existing:
                    obj = existing[key]
                    for field, value in row.items():
                        setattr(obj, field, value)
                    to_update.append(obj)
                else:
                    to_create.append(UpgradeTracker(**row))

            if to_create:
                UpgradeTracker.objects.bulk_create(
                    to_create,
                    batch_size=500,
                    ignore_conflicts=True
                )

            if to_update:
                UpgradeTracker.objects.bulk_update(
                    to_update,
                    fields=list(COLUMN_MAP.values()) + [
                        "last_updated_by",
                        "last_updated_date",
                    ],
                    batch_size=500,
                )

        return Response(
            {
                "status": True,
                "message": "Upload successful",
                "created": len(to_create),
                "updated": len(to_update),
            }
        )

    except Exception as e:
        return Response(
            {"status": False, "error": str(e)},
            status=500
        )
    
############################################################ DOWNLOAD DATA ###################################################################

@api_view(['POST'])
def download_tracker_data_view(request):

    userId = request.data.get('userId')

    user = RelocationUser.objects.filter(email__iexact=userId).first()
    if not user:
        return Response({"error": "User not found"}, status=404)

    circles = user.circles

    try:
        # ================= Fetch Data =================
        if 'CENTRAL' in circles:
            qs = UpgradeTracker.objects.all()
        else:
            qs = UpgradeTracker.objects.filter(circle__in=circles)

        df = pd.DataFrame(qs.values())

        if df.empty:
            return Response({"error": "No data available"}, status=404)

        # ================= Date Formatting =================
        for col in df.columns:
            if col.endswith("_date"):
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d-%b-%y')

        if "last_updated_date" in df.columns:
            df["last_updated_date"] = pd.to_datetime(
                df["last_updated_date"], errors='coerce'
            ).dt.strftime('%d-%b-%y %H:%M:%S')

        # ================= Unique ID =================
        df.insert(0, "Unique ID", range(1, len(df) + 1))
        df.drop(columns=["id"], inplace=True)

        # ================= Order Columns =================
        df = df[[
            "Unique ID",
            "circle",
            "upgrade_technology",
            "site_id",
            "cell_id",
            "toco_name",
            "sr_number",
            "ran_oem",
            "media_type",
            "band",
            "rfai_date",
            "allocation_date",
            "rfai_survey_date",
            "survey_status",
            "wrfai",
            "nwrfai_reason",
            "remarks_for_nwrfai",
            "nwrfai_to_wrfai_date",
            "mo_punch_date",
            "material_dispatch_date",
            "material_delivered_date",
            "installation_start_date",
            "installation_end_date",
            "integration_date",
            "nep_id",
            "emf_submission_date",
            "ran_lkf_status",
            "alarm_status",
            "alarm_rectification_done_date",
            "scft_done_date",
            "scft_i_deploy_offered_date",
            "ran_pat_offer_date",
            "ran_sat_offer_date",
            "site_onair_date",
            "i_deploy_onair_date",
            "current_status",
            "detailed_remarks",
            "manual_history",
            "rfai_rejected_date",
            "re_rfai_date",
            "pri_count",
            "pri_issue_ageing",
            "other_issue_ageing",
            "total_issue_ageing",
            "rfai_to_ms1_ageing",
            "ran_pat_accepted_date",
            "ran_sat_accepted_date",
            "scft_accepted_date",
            "kpi_at_offer_date",
            "kpi_at_accepted_date",
            "four_g_ms2_date",
            "five_g_ms2_date",
            "final_ms2_date",
            "last_updated_date",
            "last_updated_by",
        ]]
        
        df = df.where(pd.notna(df), "")
        df = df.replace(["nan", "NaN", "NAN", "None", "NONE", "null", "NULL"], "")
        df = df.where(pd.notna(df), "")

        # ================= Write Excel =================
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE = os.path.join(settings.MEDIA_ROOT, "upgrade_tracking")
        os.makedirs(BASE, exist_ok=True)

        folder = os.path.join(BASE, f"generated_files_{current_date}")
        shutil.rmtree(folder, ignore_errors=True)
        os.makedirs(folder, exist_ok=True)

        file_path = os.path.join(folder, f"UPGRADE_TRACKER_{current_date}_{current_time}.xlsx")

        template = os.path.join(BASE, "template", "upgrade_template.xlsx")
        wb = load_workbook(template)
        ws = wb["Main Tracker"]
        
        date_columns = [col for col in df.columns if col.endswith("_date") and col != 'last_updated_date']

        for col in date_columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        start_row = 3
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                if df.columns[c_idx - 1].endswith("_date") and value:
                    cell.number_format = "dd-mmm-yy"

        wb.save(file_path)

        relative = file_path.replace(settings.MEDIA_ROOT, "").lstrip(os.sep)
        url = request.build_absolute_uri(os.path.join(settings.MEDIA_URL, relative))

        return Response({"download_link": url})

    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['DELETE'])
def delete_tracker_data_view(request):
    try:
        qs = UpgradeTracker.objects.all()
        
        qs.delete()
        
        return Response({"message": "deleted successfully"})

    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['GET', 'POST'])
def daily_dashboard_view(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    start_date = request.data.get('from_date')
    end_date = request.data.get('to_date')
    view = request.data.get('view')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None

    all_unique_circles = list(
        UpgradeTracker.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    # all_unique_site_tagging = list(
    #     UpgradeTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        UpgradeTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        UpgradeTracker.objects.exclude(toco_name__isnull=True)
        .distinct("toco_name")
        .values_list("toco_name", flat=True)
    )

    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]

    start_date = dtime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    end_date = dtime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

    try:
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["toco_name__in"] = new_toco_name

        obj = UpgradeTracker.objects.filter(**filters)
        df = pd.DataFrame(obj.values())

        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        # df['rfai_date'] = df['re_rfai_date'].where(df["re_rfai_date"].notna(), df['rfai_date'])

        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")

        today = dtime.today().date()

        if (not month_start or not month_end) and (not start_date or not end_date):
            if today.day >= 26:
                start_date = today.replace(day=26)
                if today.month == 12:
                    end_date = today.replace(year=today.year + 1, month=1, day=25)
                else:
                    end_date = today.replace(month=today.month + 1, day=25)
            else:
                if today.month == 1:
                    start_date = today.replace(year=today.year - 1, month=12, day=26)
                else:
                    start_date = today.replace(month=today.month - 1, day=26)
                end_date = today.replace(day=25)
                
        if month_start and month_end:
            date_range = pd.date_range(start=month_start, end=min(month_end, today)).date
        else:
            date_range = pd.date_range(start=start_date, end=min(end_date, today)).date

        formatted_dates = [d.strftime("%d-%b-%y") for d in date_range]

        if month_start and month_end:
            result_columns = ["Milestone Track/Site Count", "AOP", "CF"] + formatted_dates
        else:
            result_columns = ["Milestone Track/Site Count", "AOP"] + formatted_dates
        result = pd.DataFrame(columns=result_columns)

        milestones = [
            "Allocation Date",
            "RFAI Date",
            "RFAI Survey Date",
            "NWRFAI To WRFAI Date",
            "MO Punch Date",
            "Material Dispatch Date",
            "Material Delivered Date",
            "Installation End Date",
            "Integration Date",
            "EMF Submission Date",
            "Alarm Rectification Done Date",
            "SCFT I-Deploy Offered Date",
            "RAN PAT Offer Date",
            "RAN SAT Offer Date",
            "Site ONAIR Date",
            "I-Deploy ONAIR Date",
        ]

        unique_data.update(**{"Milestone": milestones})

        for milestone in milestones:
            milestone_df_format = (
                milestone.lower()
                .replace(" ", "_")
                .replace("-", "_")
                .replace("(", "")
                .replace(")", "")
            )

            if milestone_df_format not in df.columns:
                continue

            df[milestone_df_format] = pd.to_datetime(df[milestone_df_format], errors="coerce").dt.date
            if month_start and month_end:
                mask = df['site_onair_date'].isna()
                valid_dates_cf = df.loc[mask, milestone_df_format].dropna()
            # else:
            #     valid_dates = df[milestone_df_format].dropna()
            
            valid_dates_aop = df[milestone_df_format].dropna()

            if valid_dates_aop.empty:
                if month_start and month_end:
                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": "-",
                        "CF": "-",
                        **{d.strftime("%d-%b-%y"): "-" for d in date_range}
                    }
                else:
                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": "-",
                        **{d.strftime("%d-%b-%y"): "-" for d in date_range}
                    }
                
                result.loc[len(result)] = row
                continue
            
            if month_start and month_end:
                cf_count = (valid_dates_cf < month_start).sum()
                aop_count = (valid_dates_aop < month_start).sum()
            else:
            #     cf_count = (valid_dates < start_date).sum()
                aop_count = (valid_dates_aop < start_date).sum()
            
            if month_start and month_end:
                cumulative = cf_count
                row = {"Milestone Track/Site Count": milestone,"AOP": aop_count, "CF": cf_count}
            else:
                cumulative = aop_count
                row = {"Milestone Track/Site Count": milestone,"AOP": aop_count}


            for d in date_range:
                count = (valid_dates_aop == d).sum()
                cumulative += count

                if view == "Cumulative":
                    row[d.strftime("%d-%b-%y")] = cumulative
                else:
                    row[d.strftime("%d-%b-%y")] = count

            result.loc[len(result)] = row

        result.columns = [
            col.strftime("%d-%b-%y") if isinstance(col, (dtime,)) or hasattr(col, "strftime") else col
            for col in result.columns
        ]
        if month_start and month_end:
            result = result[["Milestone Track/Site Count", "AOP" , "CF"] + formatted_dates]
        else:
            result = result[["Milestone Track/Site Count", "AOP"] + formatted_dates]
 
        last_col = formatted_dates[-1]
        dash_mask = result[last_col] == '-'

        result[last_col] = pd.to_numeric(result[last_col], errors='coerce')
        result['Gap'] = -result[last_col].diff()

        nan_mask = result['Gap'].isna()

        result[last_col] = result[last_col].fillna(0).astype(int)
        result['Gap'] = result['Gap'].fillna(0).astype(int)

        result.loc[dash_mask, last_col] = '-'
        result.loc[nan_mask, 'Gap'] = '-'

        result['Gap'] = result['Gap'].astype(str)
        result = result.astype(str).reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(
            lambda col: col.replace(" Date", "") if " Date" in col else col
        )

        new_result = result.copy()

        for i, d in enumerate(formatted_dates, start=1):
            new_result.rename(columns={d: f'date_{i}'}, inplace=True)

        result_json = new_result.to_dict(orient="records")
        json_data = json.dumps(result_json)

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE_URL = os.path.join(settings.MEDIA_ROOT, "upgrade_tracking")
        os.makedirs(BASE_URL, exist_ok=True)

        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)

        dashboard_file_path = os.path.join(
            output_folder,
            f"UPGRADE_TRACKING_DAILY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx"
        )

        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Daily Waterfall MS1')

        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)

        return Response({
            'message': 'Dashboard created successfully !!!',
            "download_link": download_link,
            "data": json_data,
            "dates": formatted_dates,
            "unique_data": unique_data
        }, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)


@api_view(['GET', 'POST'])
def weekly_monthly_dashboard_view(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    view = request.data.get('view')
 
    # Default to 'ALL' if not provided
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None
   
    all_unique_circles = list(
        UpgradeTracker.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    # all_unique_site_tagging = list(
    #     UpgradeTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        UpgradeTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        UpgradeTracker.objects.exclude(toco_name__isnull=True)
        .distinct("toco_name")
        .values_list("toco_name", flat=True)
    )

    # ✅ Add "ALL" at the top of each list
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    try:
        # 🔹 Dynamic filters
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["toco_name__in"] = new_toco_name
 
        # 🔹 Fetch data
        obj = UpgradeTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        # df['rfai_date'] = df['re_rfai_date'].where(df["re_rfai_date"].notna(), df['rfai_date'])
       
        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")
               
        today = dtime.today().date()
 
        if today.month >= 4:
            fy_start = dtime(today.year, 3, 26).date()   # 26-Apr current year start
            fy_end = dtime(today.year + 1, 3, 25).date() # 25-Mar next year end
        else:
            fy_start = dtime(today.year - 1, 3, 26).date()
            fy_end = dtime(today.year, 3, 25).date()
 
        # 🧩 2. Determine current cycle (26th prev → 25th curr)
        if today.day >= 26:
            if today.month == 12:
                current_cycle_start = dtime(today.year, 12, 26).date()
                current_cycle_end = dtime(today.year + 1, 1, 25).date()
            else:
                current_cycle_start = dtime(today.year, today.month, 26).date()
                current_cycle_end = dtime(today.year, today.month + 1, 25).date()
        else:
            if today.month == 1:
                current_cycle_start = dtime(today.year - 1, 12, 26).date()
            else:
                current_cycle_start = dtime(today.year, today.month - 1, 26).date()
            current_cycle_end = dtime(today.year, today.month, 25).date()
 
        # 🗓️ 3. Create monthly periods (Apr → month before current)
        months = []
        start = fy_start
        while start <= current_cycle_start:
            if start.month == 12:
                end = dtime(start.year + 1, 1, 25).date()
            else:
                end = dtime(start.year, start.month + 1, 25).date()
            months.append((start, end))
            start = end + timedelta(days=1)
 
        # 🗓️ 4. Create weekly periods for current month
        weeks = []
        
        if month_start and month_end:
            current_cycle_start = month_start
            current_cycle_end = month_end
        
        week_start = current_cycle_start
 
        while week_start <= current_cycle_end:
            week_end = week_start + timedelta(days=6)
            if week_end > current_cycle_end:
                week_end = current_cycle_end
            weeks.append((week_start, week_end))
            week_start = week_end + timedelta(days=1)
 
 
        # 📊 5. Prepare result DataFrame
        result = pd.DataFrame()
       
       
        # result.set_index("Milestone Track/Site Count", inplace=True)
 
        milestones = [
            "Allocation Date",
            "RFAI Date",
            "RFAI Survey Date",
            "NWRFAI To WRFAI Date",
            "MO Punch Date",
            "Material Dispatch Date",
            "Material Delivered Date",
            "Installation End Date",
            "Integration Date",
            "EMF Submission Date",
            "Alarm Rectification Done Date",
            "SCFT I-Deploy Offered Date",
            "RAN PAT Offer Date",
            "RAN SAT Offer Date",
            "Site ONAIR Date",
            "I-Deploy ONAIR Date",
        ]
 
        data = []
        
        for milestone in milestones:
            
            milestone_df_format = milestone.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")

            if milestone_df_format not in df.columns:
                continue
            
            milestone_data = pd.to_datetime(df[milestone_df_format], errors='coerce').dt.date
            row = {"Milestone Track/Site Count": milestone}
            if milestone_data.dropna().empty:
                row["CF"] = "-"
                # Set all month and week columns to "-"
                for _, end in months:
                    month_name = end.strftime("%b-%y")
                    row[month_name] = "-"
                for i, _ in enumerate(weeks, 1):
                    week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                    row[week_name] = "-"
                data.append(row)
                continue
 
 
            # CF → sites before financial year start
            row["CF"] = (milestone_data < fy_start).sum()
 
            cumulative_month = row["CF"]
            
            for start, end in months:
                month_name = end.strftime("%b-%y")
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_month += count
                if view == 'Cumulative':
                    row[month_name] = cumulative_month
                else:
                    row[month_name] = count
 
            cumulative_week = 0
            for i, (start, end) in enumerate(weeks, 1):
                week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                if(start >= today):
                    row[week_name] = 0
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_week += count
                if view == 'Cumulative':
                    row[week_name] = cumulative_week
                else:
                    row[week_name] = count
 
            data.append(row)
 

        result = pd.DataFrame(data)
 
        for col in result.columns:
            if col != "Milestone Track/Site Count":
                result[col] = result[col].astype(str)
 
        result = result.reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(lambda col: col.replace(" Date", "") if " Date" in col else col)
 
        new_data = result.copy()
        months_columns = [col for col in new_data.columns.to_list()[2:] if " W" not in col]
        months_data = new_data[new_data.columns.to_list()[:2] + months_columns].copy()
        
        weeks_columns = [col for col in new_data.columns.to_list()[2:] if " W" in col]
        
        week_data = new_data[new_data.columns.to_list()[:1] + weeks_columns].copy()
        
        months_data.rename(columns={
            col : f"Month-{i}" for i, col in enumerate(months_columns, start=1)
        }, inplace=True)
        
        week_data.rename(columns={
            col : f"Month_Week-{i}" for i, col in enumerate(weeks_columns, start=1)
        }, inplace=True)
        
        unique_data.update(**{"month_columns": months_columns, "week_columns" : weeks_columns})
        

        month_dict_data = months_data.to_dict(orient="records")
        month_json_data = json.dumps(month_dict_data)
        week_dict_data = week_data.to_dict(orient="records")
        week_json_data = json.dumps(week_dict_data)
        
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "upgrade_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"WEEKLY_MONTHY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx")
       
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Monthly MS1')
       
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
   
        return Response({'message': 'Weekly and Monthly Dashboard created successfully !!!', "download_link": download_link, "unique_data": unique_data, "months_data": month_json_data, "week_data": week_json_data}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
    

@api_view(['GET', 'POST'])
def gap_view(request):

    user_id = request.data.get('userId')
    if not user_id:
        return Response({"error": "userId required"}, status=400)

    user = RelocationUser.objects.filter(email__iexact=user_id).first()
    if not user:
        return Response({"error": "User not found"}, status=404)

    circles = user.circles or []

    # ================= Filters =================
    circle = request.data.get('circle', [])
    current_status = request.data.get('current_status', [])
    toco_name = request.data.get('new_toco_name', [])

    last_date = request.data.get('last_date')
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    gap = int(request.data.get('gap', 0))

    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    current_status = [c.strip() for c in current_status.split(',')] if current_status else ["ALL"]
    toco_name = [t.strip() for t in toco_name.split(',')] if toco_name else ["ALL"]

    last_date = dtime.strptime(last_date, "%Y-%m-%d").date() if last_date else None

    milestone1 = milestone1.lower().replace(" ", "_").replace("-", "_")
    milestone2 = milestone2.lower().replace(" ", "_").replace("-", "_")

    try:
        # ================= Default Last Date =================
        today = dtime.today().date()
        if not last_date:
            last_date = today - timedelta(days=1)

        # ================= Build Filters =================
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in toco_name:
            filters["toco_name__in"] = toco_name

        # milestone condition
        if gap > 0:
            filters[f"{milestone1}__lte"] = last_date
        else:
            filters[f"{milestone2}__lte"] = last_date

        # ================= First Dataset =================
        qs1 = UpgradeTracker.objects.filter(**filters)
        df1 = pd.DataFrame(qs1.values())

        # ================= Second Dataset =================
        if gap > 0:
            filters[f"{milestone2}__lte"] = last_date
        else:
            filters[f"{milestone1}__lte"] = last_date

        qs2 = UpgradeTracker.objects.filter(**filters)
        df2 = pd.DataFrame(qs2.values())

        if df1.empty:
            return Response({"message": "No data found"}, status=200)

        # ================= Key on SR Number =================
        
        def normalize_sr(val):
            if val is None:
                return None
            if isinstance(val, float):
                if pd.isna(val):
                    return None
                if val.is_integer():
                    return str(int(val))
            return str(val).strip()

        
        df1["key"] = df1["sr_number"].apply(normalize_sr)
        df2["key"] = df2["sr_number"].apply(normalize_sr)

        df = df1[~df1["key"].isin(df2["key"])].drop(columns=["key"])

        # ================= Date Formatting =================
        for col in df.columns:
            converted = pd.to_datetime(df[col], errors="coerce")
            if converted.notna().sum() > 0:
                if col == "last_updated_date":
                    df[col] = converted.dt.strftime("%d-%b-%y %H:%M:%S")
                else:
                    df[col] = converted.dt.strftime("%d-%b-%y")

        # ================= File Creation =================
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE_URL = os.path.join(settings.MEDIA_ROOT, "upgrade_tracking")
        os.makedirs(BASE_URL, exist_ok=True)

        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)

        file_path = os.path.join(
            output_folder,
            f"UPGRADE_GAP_{milestone1}_{milestone2}_{current_date}_{current_time}.xlsx"
        )

        # ================= Prepare Output =================
        df.insert(0, "Unique ID", range(1, len(df) + 1))
        if "id" in df.columns:
            df.drop(columns=["id"], inplace=True)

        json_val = json.dumps(df.to_dict(orient="records"))

        # ================= Excel Write =================
        template_path = os.path.join(BASE_URL, "template", "upgrade_template.xlsx")
        wb = load_workbook(template_path)
        ws = wb.active

        start_row = 3
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                if isinstance(value, datetime):
                    cell.number_format = "dd-mmm-yy"

        wb.save(file_path)

        relative_path = file_path.replace(settings.MEDIA_ROOT, "").lstrip(os.sep)
        download_url = request.build_absolute_uri(
            os.path.join(settings.MEDIA_URL, relative_path).replace("\\", "/")
        )

        return Response(
            {
                "message": "Request processed successfully",
                "data": json_val,
                "download_link": download_url,
            },
            status=200,
        )

    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['GET', 'POST'])
def ms1_ageing_dashboard_table1(request):
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    # all_unique_site_tagging = list(
    #     UpgradeTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        UpgradeTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Allocation",
        "RFAI",
        "RFAI Survey",
        "NWRFAI To WRFAI",
        "MO Punch",
        "Material Dispatch",
        "Material Delivered",
        "Installation End",
        "Integration",
        "EMF Submission",
        "Alarm Rectification Done",
        "SCFT I-Deploy Offered",
        "RAN PAT Offer",
        "RAN SAT Offer",
        "Site ONAIR",
        "I-Deploy ONAIR",
    ]
    
    unique_data = {
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones
    }

    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        # if site_tagging and "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if month_start and month_end:
            filters[f"{milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + '_date'}__range"] = (month_start, month_end)

        obj = UpgradeTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        date_cols = [col.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + '_date' for col in milestones]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']
        
        
        def generate_table_summary(df, milestone1, milestone2):


            milestone_cols = [(col.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date") for col in milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]]

            temp = df.copy()

            circles = ["ALL"] + temp["circle"].unique().tolist()

            final_df = pd.DataFrame()

            milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
            milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"

            for circle in circles:
                if circle != "ALL":
                    temp  = df[df["circle"] == circle]

                done_row = {}
                pending_row = {}

                for col, label in zip(milestone_cols, milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]):

                    done_row[label] = (temp[col].notna() & temp[milestone1_col].notna() ).sum()

                    pending_row[label] = (temp[col].isna() & temp[milestone1_col].notna() ).sum()

                rows_to_add = pd.DataFrame([
                    {"Circle": circle, "Site Status": "Done", **done_row},
                    {"Circle": circle, "Site Status": "Pending", **pending_row},
                ])

                final_df = pd.concat([final_df, rows_to_add], ignore_index=True)

            for c in final_df.columns:
                if c != "Site Status" and c != "Circle":
                    final_df[c] = final_df[c].astype("Int64")

            return final_df

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        table_summary = generate_table_summary(df, start_label, end_label)
 
        table_summary = table_summary.applymap(lambda x: str(x) if pd.notna(x) else "")      


        result_json = table_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "table_summary": json_data1
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
  
  
@api_view(['GET', 'POST'])
def ms1_ageing_dashboard_table2(request):
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    breakpoint1 = request.data.get('breakpoint1')
    breakpoint2 = request.data.get('breakpoint2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    type = request.data.get('type')

    # all_unique_site_tagging = list(
    #     UpgradeTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        UpgradeTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Allocation",
        "RFAI",
        "RFAI Survey",
        "NWRFAI To WRFAI",
        "MO Punch",
        "Material Dispatch",
        "Material Delivered",
        "Installation End",
        "Integration",
        "EMF Submission",
        "Alarm Rectification Done",
        "SCFT I-Deploy Offered",
        "RAN PAT Offer",
        "RAN SAT Offer",
        "Site ONAIR",
        "I-Deploy ONAIR",
    ]
    
    unique_data = {
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones" : milestones
    }
    

    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    
    try:

        milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
        milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"

        breakpoint1 = int(breakpoint1)
        breakpoint2 = int(breakpoint2)
        
        temp = breakpoint1
        breakpoint1 = min(breakpoint1, breakpoint2)
        breakpoint2 = max(temp, breakpoint2)

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None

        milestones = [
            "Allocation",
            "RFAI",
            "RFAI Survey",
            "NWRFAI To WRFAI",
            "MO Punch",
            "Material Dispatch",
            "Material Delivered",
            "Installation End",
            "Integration",
            "EMF Submission",
            "Alarm Rectification Done",
            "SCFT I-Deploy Offered",
            "RAN PAT Offer",
            "RAN SAT Offer",
            "Site ONAIR",
            "I-Deploy ONAIR",
        ]
        

        filters = {}
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
            
        if type == "type2":
            if month_start and month_end:
                filters[f"{milestone1_col}__range"] = (month_start, month_end)
                

        obj = UpgradeTracker.objects.filter(**filters)  # noqa: F405
        
        if type == "type1":
            if month_start and month_end:
                obj = obj.filter(
                    Q(**{f"{milestone1_col}__range": (month_start, month_end)}) |
                    Q(**{f"{milestone2_col}__range": (month_start, month_end)})
                )
        

        df = pd.DataFrame(obj.values())

        
        date_cols = [col.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + '_date' for col in milestones]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']

        if month_start is not None:
            month_start = pd.to_datetime(month_start)

        if month_end is not None:
            month_end = pd.to_datetime(month_end)

        def generate_done_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            temp = df.copy()
            
            temp["days_diff"] = (temp[end_col] - temp[start_col]).dt.days
            
            if type == "type2":
        
                summary = (
                    temp.groupby("Circle").apply(lambda g: pd.Series({
                        f"{start_label} Done Count": g[start_col].notna().sum(),
                        f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].notna(), "days_diff"].mean()
                    }))
                    .reset_index()
                )
                
                total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                total_row["Circle"] = "Total"

                overall_valid = temp.loc[
                    temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].notna(),
                    "days_diff"
                ]
                total_row["Average Days"] = overall_valid.mean()

            else:
                if month_start is not None and month_end is not None:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": (g[start_col].notna() & (g[start_col] >= month_start) & (g[start_col] <= month_end)).sum(),
                            f"{end_label} Done Count": (g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"<= {breakpoint1} days": (g["days_diff"].notna() & (g["days_diff"] <= breakpoint1) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": (g["days_diff"].notna() & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f">= {breakpoint2} days": (g["days_diff"].notna() & (g["days_diff"] >= breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna() & (temp[end_col] >= month_start) & (temp[end_col] <= month_end),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                else:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": g[start_col].notna().sum(),
                            f"{end_label} Done Count": g[end_col].notna().sum(),
                            f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & g[end_col].notna()).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna()).sum(),
                            f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & g[end_col].notna()).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna(), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna(),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                    
            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]

            return summary


        done_summary = generate_done_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        done_summary = done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        def generate_pending_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            
            today = pd.Timestamp.today()
            
            temp = df.copy()
            
            temp["days_diff"] = (today - temp[start_col]).dt.days

            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                    f"<= {breakpoint1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[start_col].notna())).sum(),
                    f"{breakpoint1+1}-{breakpoint2-1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[start_col].notna())).sum(),
                    f">= {breakpoint2} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[start_col].notna())).sum(),
                    "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].isna(), "days_diff"].mean()
                }))
                .reset_index()
            )

            total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
            total_row["Circle"] = "Total"

            overall_valid = temp.loc[
                temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].isna(),
                "days_diff"
            ]
            total_row["Average Days"] = overall_valid.mean()

            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Pending Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]
            
            return summary
        

        pending_summary = generate_pending_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        pending_summary = pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"AGEING_DASHBOARD_FILE_{site_tagging}_{current_date}_{current_time}.xlsx")

        result_json = pending_summary.to_dict(orient="records")
        json_data2 = json.dumps(result_json)

        result_json = done_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "done_summary": json_data1,
            "pending_summary": json_data2
        }
        
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            pending_summary.to_excel(writer, index=False, sheet_name='PENDING SUMMARY')
            done_summary.to_excel(writer, index=False, sheet_name='DONE SUMMARY')
        
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
        
        return Response({'message': 'request processed successfully !!!', "download_link": download_link, "json_data": json_data, "unique_data": unique_data, "breakpoint1" : breakpoint1, "breakpoint2" : breakpoint2}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)


@api_view(['GET', 'POST'])
def lifecycle_display(request):
    userId = request.data.get('userId').lower()
    siteId = request.data.get('siteId')
    sr_number = request.data.get('sr_number')
    circle = request.data.get('circle')

    if not userId or not siteId or not circle or not sr_number:
        return Response({"error": "input not provided"}, status = 400)
    
    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    circles = user.circles
    
    if 'CENTRAL' not in circles and circle not in circles:
        return Response({"error" : f"Access to {circle} sites denied"}, status = 403)
    
    try:
        filter = {
            "circle": circle,
            "cell_id": siteId,
            "sr_number": sr_number
        }

        obj = UpgradeTracker.objects.filter(**filter).first()
        
        if obj is None:
            return Response({"error": "Site not found!"}, status=500)
        else:
            df = pd.DataFrame([model_to_dict(obj)])

        filters = {
            "circle" : circle,
            "site_id" : siteId,
            "sr_number" : sr_number
        }
        
        issue_obj = UpgradeIssue.objects.filter(**filters)
        issue_df = pd.DataFrame(issue_obj.values())

        required_cols = ["milestone", "status"]
        for col in required_cols:
            if col not in issue_df.columns:
                issue_df[col] = None
        
        data_cols = [
            "Allocation",
            "RFAI",
            "RFAI Survey",
            "NWRFAI To WRFAI",
            "MO Punch",
            "Material Dispatch",
            "Material Delivered",
            "Installation End",
            "Integration",
            "EMF Submission",
            "Alarm Rectification Done",
            "SCFT I-Deploy Offered",
            "RAN PAT Offer",
            "RAN SAT Offer",
            "Site ONAIR",
            "I-Deploy ONAIR",
        ]
        
        data_cols_1 = [col.lower().replace("-", "_").replace(" ", "_").replace("/", "_") + "_date" for col in data_cols] + ['circle']
        
        df_data = df[data_cols_1]
        
        issue_counts = {}
        
        for col in data_cols:
            counts = {}
            counts['total'] = (issue_df['milestone'] == col).sum()
            counts['open'] = ((issue_df['milestone'] == col) & (issue_df['status'] == 'Open')).sum()
            counts['closed'] = ((issue_df['milestone'] == col) & (issue_df['status'] == 'Closed')).sum()
            issue_counts[col] = counts
        
        print("C")
        
        if not df_data.empty:
            df_data = df_data.applymap(
                lambda x: x.strftime("%d-%b-%y") if hasattr(x, 'strftime') else x
            )
            
        
        result_json = df_data.to_dict(orient="records")
        json_data = json.dumps(result_json)
        
        return Response({'message': 'request processed successfully !!!', "json_data" : json_data, "issue_counts" : issue_counts }, status = 200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
    

@api_view(['POST'])
def issue_timeline_display(request):
    userId = request.data.get('userId')
    if not userId:
        return Response({"error": "userId required"}, status=400)

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    
    circles = user.circles
    
    siteId = request.data.get('siteId')
    sr_number = request.data.get('sr_number')
    circle = request.data.get('circle')
    owner = request.data.get('owner')
    milestone = request.data.get('milestone')

    if not all([siteId, circle, sr_number]):
        return Response({"message": "Circle/Site-ID not provided"}, status=400)

    if 'CENTRAL' not in circles and circle not in circles:
        return Response({"error": f"Access to {circle} sites denied!"}, status=403)

    try:
        filters = {
            "circle": circle,
            "site_id": siteId,
            "sr_number": sr_number
        }
        if owner:
            filters['issue_owner'] = owner
        if milestone:
            filters['milestone'] = milestone

        obj = UpgradeIssue.objects.filter(**filters)

        df = pd.DataFrame(obj.values())

        if not obj.exists():
            # required_cols = ["Circle", "Site ID", "Issue Owner", "Milestone", "Issue Name", "Start Date", "Close Date", "Status", "Duration", "Updated_by", "Updated_at", "Created_by", "Created_at"]
            # for col in required_cols:
            #     if col not in df.columns:
            #         df[col] = None
            json_data = df.to_dict(orient="records")
            json_data = json.dumps(json_data)
            return Response({"error": "Data not found!", "json_data": json_data}, status=200)

        df["start_date"] = df["start_date"].apply(
            lambda x: x.strftime("%d-%b-%y") if pd.notnull(x) else "-"
        )

        df["close_date"] = df["close_date"].apply(
            lambda x: x.strftime("%d-%b-%y") if pd.notnull(x) else "-"
        )
        
        rename_map = {
            "circle": "Circle",
            "site_id": "Site ID",
            "sr_number": "SR Number",
            "issue_owner": "Issue Owner",
            "milestone": "Milestone",
            "issue_name": "Issue Name",
            "start_date": "Start Date",
            "close_date": "Close Date",
            "status": "Status",
            "duration": "Duration",
            "remarks": "Remarks",
            "updated_by": "Updated_by",
            "updated_at": "Updated_at",
            "created_by": "Created_by",
            "created_at": "Created_at"
        }
        
        df = df.rename(columns=rename_map)

        json_data = df.to_dict(orient="records")

        return Response({"message": "Request processed successfully!", "json_data": json_data}, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
 
@api_view(['POST'])
def issue_timeline_add(request):
    userId = request.data.get('userId')
    if not userId:
        return Response({"error": "userId required"}, status=400)

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    if user.right == 'Read':
        return Response({"error": "Access denied!"}, status=status.HTTP_403_FORBIDDEN)

    siteId = request.data.get('siteId')
    circle = request.data.get('circle')
    sr_number = request.data.get('sr_number')
    owner = request.data.get('owner')
    milestone = request.data.get('milestone')
    issue = request.data.get('issue')
    start_date_str = request.data.get('start_date')   
    close_date_str = request.data.get('close_date')  
    remarks = request.data.get('remarks') 

    required_fields = [siteId, circle, owner, milestone, issue, start_date_str]
    if not all(required_fields):
        return Response({"message": "Required data not provided"}, status=status.HTTP_400_BAD_REQUEST)
    
        

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()

        close_date = None
        if close_date_str not in ["", None, "-", "nan", "undefined"]:
            close_date = datetime.strptime(close_date_str, "%Y-%m-%d").date()

    except Exception:
        return Response({"error": "Invalid date format. Expected yyyy-mm-dd"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        filters = {
            "circle": circle,
            "site_id": siteId,
            "sr_number": sr_number,
            "issue_owner": owner,
            "milestone": milestone,
            "issue_name": issue,
            "status": "Open"
        }

        if UpgradeIssue.objects.filter(**filters).exists():
            return Response({"error": "Open issue already exists"}, status=409)

        today = date.today()

        if close_date:
            duration = (close_date - start_date).days
        else:
            duration = (today - start_date).days

        new_issue = UpgradeIssue.objects.create(
            circle=circle,
            site_id=siteId,
            sr_number=sr_number,
            issue_owner=owner,
            milestone=milestone,
            issue_name=issue,
            start_date=start_date,
            close_date=close_date,
            status="Closed" if close_date else "Open",
            duration=duration,
            remarks=remarks,
            updated_by=userId,
            created_by=userId,
        )
        
        update_ageing_new(circle, siteId, sr_number)

        return Response(
            {
                "message": "Issue added successfully!",
                "json_data":{"id": new_issue.id,
                "circle": circle,
                "site": siteId,
                "sr_number": sr_number,
                "milestone": milestone,
                "issue": issue,
                "start_date": start_date.strftime("%d-%b-%y"),
                "close_date": close_date.strftime("%d-%b-%y") if close_date else "-",
                "status": new_issue.status,
                "duration_days": duration,
                "remarks": remarks}
            },
            status=200
        )

    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
def issue_timeline_update(request):
    userId = request.data.get('userId')
    issue_id = request.data.get('id') 
    circle = request.data.get('circle')

    if not userId or not issue_id or not circle:
        return Response({"error": "Required data not provided"}, status=400)

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    if user.right == 'Read':
        return Response({"error": "Access denied!"}, status=status.HTTP_403_FORBIDDEN)

    # user_circle = ACCESS_RIGHTS[userId]["Circle"]
    # if user_circle != "CENTRAL" and user_circle != circle:
    #     return Response({"error": "Access denied!"}, status=403)

    try:
        issue_obj = UpgradeIssue.objects.get(id=issue_id)
    except UpgradeIssue.DoesNotExist:
        return Response({"error": "Issue not found"}, status=404)
    
    if issue_obj.status == 'Closed':
        return Response({"error": "Issue already closed"}, status=404)

    start_date_str = request.data.get("start_date")
    close_date_str = request.data.get("close_date")
    new_owner = request.data.get("owner")
    new_milestone = request.data.get("milestone")
    new_issue_name = request.data.get("issue")
    remarks = request.data.get('remarks')

    try:
        if start_date_str:
            issue_obj.start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()

        if close_date_str in ["", None, "-", "nan", "undefined"]:
            issue_obj.close_date = None
        else:
            issue_obj.close_date = datetime.strptime(close_date_str, "%Y-%m-%d").date()

    except Exception:
        return Response({"error": "Invalid date format. Expected yyyy-mm-dd"}, status=400)

    try:
        
        if new_owner:
            issue_obj.issue_owner = new_owner

        if new_milestone:
            issue_obj.milestone = new_milestone

        if new_issue_name:
            issue_obj.issue_name = new_issue_name
            
        if remarks:
            issue_obj.remarks = remarks

        if issue_obj.close_date:
            issue_obj.status = "Closed"
        else:
            issue_obj.status = "Open"

        today = date.today()

        if issue_obj.close_date:
            issue_obj.duration = (issue_obj.close_date - issue_obj.start_date).days
        else:
            issue_obj.duration = (today - issue_obj.start_date).days

        issue_obj.updated_by = userId

        issue_obj.save()

        out_start = issue_obj.start_date.strftime("%d-%b-%y")
        out_close = issue_obj.close_date.strftime("%d-%b-%y") if issue_obj.close_date else "-"

        response_data = {
            "id": issue_obj.id,
            "circle": issue_obj.circle,
            "site_id": issue_obj.site_id,
            "sr_number": issue_obj.sr_number,
            "issue_owner": issue_obj.issue_owner,
            "milestone": issue_obj.milestone,
            "issue_name": issue_obj.issue_name,
            "start_date": out_start,
            "close_date": out_close,
            "status": issue_obj.status,
            "duration": issue_obj.duration,
            "remarks": remarks,
            "updated_by": userId
        }
        
        update_ageing_new(circle, issue_obj.site_id, issue_obj.sr_number)

        return Response(
            {"message": "Issue updated successfully!", "json_data": response_data},
            status=200
        )
        
    except Exception as e:
        return Response({"Error: " : f"{str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def issue_timeline_delete(request):
    userId = request.data.get('userId')
    issue_id = request.data.get('id')
    circle = request.data.get('circle')

    if not userId or not issue_id or not circle:
        return Response(
            {"error": "userId, id and circle are required"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    if user.right == 'Read':
        return Response({"error": "Access denied!"}, status=status.HTTP_403_FORBIDDEN)

    # Access control
    # user_circle = ACCESS_RIGHTS[userId]["Circle"]
    # if user_circle != "CENTRAL" and user_circle != circle:
    #     return Response({"error": "Access denied!"}, status=status.HTTP_403_FORBIDDEN)

    try:
        issue_obj = UpgradeIssue.objects.get(id=issue_id)
    except UpgradeIssue.DoesNotExist:
        return Response({"error": "Issue not found"}, status=status.HTTP_404_NOT_FOUND)

    # Optional safety check (recommended)
    # if issue_obj.status == "Closed":
    #     return Response(
    #         {"error": "Closed issues cannot be deleted"},
    #         status=status.HTTP_400_BAD_REQUEST
    #     )
    
    update_ageing_new(circle, issue_obj.site_id, issue_obj.sr_number)

    issue_obj.delete()

    return Response(
        {
            "message": "Issue deleted successfully!",
            "id": issue_id
        },
        status=status.HTTP_200_OK
    )


@api_view(['GET', 'POST'])
def graphs_view(request):
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    # all_unique_site_tagging = list(
    #     UpgradeTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        UpgradeTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Allocation",
        "RFAI",
        "RFAI Survey",
        "NWRFAI To WRFAI",
        "MO Punch",
        "Material Dispatch",
        "Material Delivered",
        "Installation End",
        "Integration",
        "EMF Submission",
        "Alarm Rectification Done",
        "SCFT I-Deploy Offered",
        "RAN PAT Offer",
        "RAN SAT Offer",
        "Site ONAIR",
        "I-Deploy ONAIR",
    ]
    
    unique_data = {
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]
    }

    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        # if site_tagging and "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if month_start and month_end:
            filters[f"{milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '') + '_date'}__range"] = (month_start, month_end)

        obj = UpgradeTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        df['Circle'] = df['circle']

        def generate_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"{end_label} Pending Count"]
            ]

            return summary

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        start_col = start_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_") + "_date"
        end_col = end_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_") + "_date"
        
        graph_summary = generate_summary(df, start_label, end_label, start_col, end_col)

        graph_summary = graph_summary.applymap(lambda x: str(x) if pd.notna(x) else "")   

        result_json = graph_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "graph_summary": json_data1,
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)


@api_view(['GET', 'POST'])
def monthly_graph(request):
    circle = request.data.get('circle', [])
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    view = request.data.get('view')
    type = request.data.get('type')
    
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"
    milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"
    
    all_unique_circles = list(
        UpgradeTracker.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    # all_unique_site_tagging = list(
    #     UpgradeTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        UpgradeTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        UpgradeTracker.objects.exclude(toco_name__isnull=True)
        .distinct("toco_name")
        .values_list("toco_name", flat=True)
    )
    
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }
    
    try:
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["toco_name__in"] = new_toco_name
        
            
            
        obj = UpgradeTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        # print(df)
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        df = df[[milestone1_col, milestone2_col]]
       
        for col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        
        today = dtime.today().date()
        
        df["month_num1"] = (
            df[milestone1_col].dt.month + (df[milestone1_col].dt.day >= 26)
        ).replace({13: 1}).astype("Int64")

        df["year1"] = (
            df[milestone1_col].dt.year - (
                (df[milestone1_col].dt.month < 3) |
                ((df[milestone1_col].dt.month == 3) & (df[milestone1_col].dt.day < 26))
            )
        ).astype("Int64")

        # 3️⃣ Display year
        df["display_year"] = df["year1"] + (df["month_num1"] <= 3)

        # 4️⃣ Month label
        df["month_name1"] = pd.to_datetime(
            df["month_num1"].astype(str) + "-" + df["display_year"].astype(str),
            format="%m-%Y",
            errors="coerce"
        ).dt.strftime("%b-%y")

        # 5️⃣ Current FY
        current_fy = today.year if today >= date(today.year, 3, 26) else today.year - 1

        # 6️⃣ CF logic
        df.loc[df["year1"] < current_fy, "month_name1"] = "CF"



        df['month_num2'] = df[milestone2_col].dt.month + (df[milestone2_col].dt.day >= 26)
        df['month_num2'] = df['month_num2'].replace({13: 1})

        # 2️⃣ Financial year (derived ONLY from original date)
        df['year2'] = df[milestone2_col].dt.year - (
            (df[milestone2_col].dt.month < 3) |
            ((df[milestone2_col].dt.month == 3) & (df[milestone2_col].dt.day < 26))
        )

        # 3️⃣ Safe numeric conversion (IMPORTANT)
        df['month_num2'] = pd.to_numeric(df['month_num2'], errors='coerce')
        df['year2'] = pd.to_numeric(df['year2'], errors='coerce')

        # 4️⃣ Display year
        df['display_year2'] = df['year2'] + (df['month_num2'] <= 3)

        # 5️⃣ Month label (NaT-safe)
        df['month_name2'] = pd.to_datetime(
            df['month_num2'].astype('Int64').astype(str) + '-' +
            df['display_year2'].astype('Int64').astype(str),
            format='%m-%Y',
            errors='coerce'
        ).dt.strftime('%b-%y')

        # 6️⃣ Current financial year
        current_fy = today.year if today >= date(today.year, 3, 26) else today.year - 1

        # 7️⃣ Carry Forward (only where year2 exists)
        df.loc[df['year2'].notna() & (df['year2'] < current_fy), 'month_name2'] = 'CF'
        
        print(df['month_name1'].dropna().unique().tolist())

        
        print(df['year1'])
        
        print(df['month_name2'])
        
        print(df['year2'])

        def sort_financial_year(summary):
            # Extract sorted unique Month-Year (excluding CF)
            months = [m for m in summary['month_name'].unique() if m != "CF"]

            # Convert to datetime to sort correctly
            months_sorted = sorted(
                months,
                key=lambda x: pd.to_datetime(x, format='%b-%y')
            )

            # CF should always be first
            ordered = ["CF"] + months_sorted if "CF" in summary['month_name'].values else months_sorted

            # Convert into ordered categorical for final sort
            summary['month_name'] = pd.Categorical(summary['month_name'], ordered, ordered=True)

            return summary.sort_values('month_name')
        
        def generate_type1_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()

            # Start counts by month_name
            start_counts = (
                temp[temp[start_col].notna()]
                .groupby("month_name1")
                .size()
                .reset_index(name=f"{start_label} Done Count")
                .rename(columns={"month_name1": "month_name"})
            )

            # End counts by month_name2
            end_counts = (
                temp[temp[end_col].notna()]
                .groupby("month_name2")
                .size()
                .reset_index(name=f"{end_label} Done Count")
                .rename(columns={"month_name2": "month_name"})
            )

            # Merge both results
            summary = pd.merge(start_counts, end_counts, on="month_name", how="outer")

            # Convert to Int64 nullable type
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")
            summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].astype("Int64")

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            return summary.reset_index(drop=True)
        
        def generate_type2_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("month_name1").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    # f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
                .rename(columns={"month_name1": "month_name"})
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")

            summary = summary[
                ["month_name", f"{start_label} Done Count", f"{end_label} Done Count"]
            ]

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            return summary.reset_index(drop=True)
        
        summary = generate_type1_summary(df, milestone1, milestone2, milestone1_col, milestone2_col) if type == 'type1' else generate_type2_summary(df, milestone1, milestone2, milestone1_col, milestone2_col)

        numeric_cols = [col for col in summary.columns if "Done Count" in col]

        summary[numeric_cols] = summary[numeric_cols].apply(pd.to_numeric, errors="coerce")
        summary[numeric_cols] = summary[numeric_cols].fillna(0).astype(int)

        
        result_json = summary.to_dict(orient = "records")
        json_data = json.dumps(result_json)
        
        return Response({"message" : "request processed successfully ! ", "json_data": json_data, "unique_data": unique_data}, status=200)
    
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
 



@api_view(['GET', 'POST'])
def ms2_daily_waterfall(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    start_date = request.data.get('from_date')
    end_date = request.data.get('to_date')
    view = request.data.get('view')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None

    all_unique_circles = list(
        UpgradeTracker.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    # all_unique_site_tagging = list(
    #     UpgradeTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        UpgradeTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        UpgradeTracker.objects.exclude(toco_name__isnull=True)
        .distinct("toco_name")
        .values_list("toco_name", flat=True)
    )

    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]

    start_date = dtime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    end_date = dtime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

    try:
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["toco_name__in"] = new_toco_name

        obj = UpgradeTracker.objects.filter(**filters)
        df = pd.DataFrame(obj.values())

        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        # df['rfai_date'] = df['re_rfai_date'].where(df["re_rfai_date"].notna(), df['rfai_date'])

        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")

        today = dtime.today().date()

        if (not month_start or not month_end) and (not start_date or not end_date):
            if today.day >= 26:
                start_date = today.replace(day=26)
                if today.month == 12:
                    end_date = today.replace(year=today.year + 1, month=1, day=25)
                else:
                    end_date = today.replace(month=today.month + 1, day=25)
            else:
                if today.month == 1:
                    start_date = today.replace(year=today.year - 1, month=12, day=26)
                else:
                    start_date = today.replace(month=today.month - 1, day=26)
                end_date = today.replace(day=25)
                
        if month_start and month_end:
            date_range = pd.date_range(start=month_start, end=min(month_end, today)).date
        else:
            date_range = pd.date_range(start=start_date, end=min(end_date, today)).date

        formatted_dates = [d.strftime("%d-%b-%y") for d in date_range]

        if month_start and month_end:
            result_columns = ["Milestone Track/Site Count", "AOP", "CF"] + formatted_dates
        else:
            result_columns = ["Milestone Track/Site Count", "AOP"] + formatted_dates
        result = pd.DataFrame(columns=result_columns)

        milestones = [
            "Site ONAIR Date",
            "RAN PAT Accepted Date",
            "RAN SAT Accepted Date",
            "SCFT Accepted Date",
            "KPI AT offer Date",
            "KPI AT Accepted Date",
            "4G MS2 Date",
            "5G MS2 Date",
            "Final MS2 Date"
        ]

        unique_data.update(**{"Milestone": milestones})

        for milestone in milestones:
            milestone_df_format = (
                milestone.lower()
                .replace(" ", "_")
                .replace("-", "_")
                .replace("(", "")
                .replace(")", "")
            )

            if milestone_df_format not in df.columns:
                continue

            df[milestone_df_format] = pd.to_datetime(df[milestone_df_format], errors="coerce").dt.date
            if month_start and month_end:
                mask = df['site_onair_date'].isna()
                valid_dates_cf = df.loc[mask, milestone_df_format].dropna()
            # else:
            #     valid_dates = df[milestone_df_format].dropna()
            
            valid_dates_aop = df[milestone_df_format].dropna()

            if valid_dates_aop.empty:
                if month_start and month_end:
                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": "-",
                        "CF": "-",
                        **{d.strftime("%d-%b-%y"): "-" for d in date_range}
                    }
                else:
                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": "-",
                        **{d.strftime("%d-%b-%y"): "-" for d in date_range}
                    }
                
                result.loc[len(result)] = row
                continue
            
            if month_start and month_end:
                cf_count = (valid_dates_cf < month_start).sum()
                aop_count = (valid_dates_aop < month_start).sum()
            else:
            #     cf_count = (valid_dates < start_date).sum()
                aop_count = (valid_dates_aop < start_date).sum()
            
            if month_start and month_end:
                cumulative = cf_count
                row = {"Milestone Track/Site Count": milestone,"AOP": aop_count, "CF": cf_count}
            else:
                cumulative = aop_count
                row = {"Milestone Track/Site Count": milestone,"AOP": aop_count}


            for d in date_range:
                count = (valid_dates_aop == d).sum()
                cumulative += count

                if view == "Cumulative":
                    row[d.strftime("%d-%b-%y")] = cumulative
                else:
                    row[d.strftime("%d-%b-%y")] = count

            result.loc[len(result)] = row

        result.columns = [
            col.strftime("%d-%b-%y") if isinstance(col, (dtime,)) or hasattr(col, "strftime") else col
            for col in result.columns
        ]
        if month_start and month_end:
            result = result[["Milestone Track/Site Count", "AOP" , "CF"] + formatted_dates]
        else:
            result = result[["Milestone Track/Site Count", "AOP"] + formatted_dates]
 
        last_col = formatted_dates[-1]
        dash_mask = result[last_col] == '-'

        result[last_col] = pd.to_numeric(result[last_col], errors='coerce')
        result['Gap'] = -result[last_col].diff()

        nan_mask = result['Gap'].isna()

        result[last_col] = result[last_col].fillna(0).astype(int)
        result['Gap'] = result['Gap'].fillna(0).astype(int)

        result.loc[dash_mask, last_col] = '-'
        result.loc[nan_mask, 'Gap'] = '-'

        result['Gap'] = result['Gap'].astype(str)
        result = result.astype(str).reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(
            lambda col: col.replace(" Date", "") if " Date" in col else col
        )

        new_result = result.copy()

        for i, d in enumerate(formatted_dates, start=1):
            new_result.rename(columns={d: f'date_{i}'}, inplace=True)

        result_json = new_result.to_dict(orient="records")
        json_data = json.dumps(result_json)

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE_URL = os.path.join(settings.MEDIA_ROOT, "upgrade_tracking")
        os.makedirs(BASE_URL, exist_ok=True)

        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)

        dashboard_file_path = os.path.join(
            output_folder,
            f"UPGRADE_TRACKING_MS2_DAILY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx"
        )

        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Daily Waterfall MS2')

        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)

        return Response({
            'message': 'Dashboard created successfully !!!',
            "download_link": download_link,
            "data": json_data,
            "dates": formatted_dates,
            "unique_data": unique_data
        }, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)


@api_view(['GET', 'POST'])
def ms2_weekly_monthly_waterfall(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    view = request.data.get('view')
 
    # Default to 'ALL' if not provided
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None
   
    all_unique_circles = list(
        UpgradeTracker.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    # all_unique_site_tagging = list(
    #     UpgradeTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        UpgradeTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        UpgradeTracker.objects.exclude(toco_name__isnull=True)
        .distinct("toco_name")
        .values_list("toco_name", flat=True)
    )

    # ✅ Add "ALL" at the top of each list
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    try:
        # 🔹 Dynamic filters
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["toco_name__in"] = new_toco_name
 
        # 🔹 Fetch data
        obj = UpgradeTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        # df['rfai_date'] = df['re_rfai_date'].where(df["re_rfai_date"].notna(), df['rfai_date'])
       
        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")
               
        today = dtime.today().date()
 
        if today.month >= 4:
            fy_start = dtime(today.year, 3, 26).date()   # 26-Apr current year start
            fy_end = dtime(today.year + 1, 3, 25).date() # 25-Mar next year end
        else:
            fy_start = dtime(today.year - 1, 3, 26).date()
            fy_end = dtime(today.year, 3, 25).date()
 
        # 🧩 2. Determine current cycle (26th prev → 25th curr)
        if today.day >= 26:
            if today.month == 12:
                current_cycle_start = dtime(today.year, 12, 26).date()
                current_cycle_end = dtime(today.year + 1, 1, 25).date()
            else:
                current_cycle_start = dtime(today.year, today.month, 26).date()
                current_cycle_end = dtime(today.year, today.month + 1, 25).date()
        else:
            if today.month == 1:
                current_cycle_start = dtime(today.year - 1, 12, 26).date()
            else:
                current_cycle_start = dtime(today.year, today.month - 1, 26).date()
            current_cycle_end = dtime(today.year, today.month, 25).date()
 
        # 🗓️ 3. Create monthly periods (Apr → month before current)
        months = []
        start = fy_start
        while start <= current_cycle_start:
            if start.month == 12:
                end = dtime(start.year + 1, 1, 25).date()
            else:
                end = dtime(start.year, start.month + 1, 25).date()
            months.append((start, end))
            start = end + timedelta(days=1)
 
        # 🗓️ 4. Create weekly periods for current month
        weeks = []
        
        if month_start and month_end:
            current_cycle_start = month_start
            current_cycle_end = month_end
        
        week_start = current_cycle_start
 
        while week_start <= current_cycle_end:
            week_end = week_start + timedelta(days=6)
            if week_end > current_cycle_end:
                week_end = current_cycle_end
            weeks.append((week_start, week_end))
            week_start = week_end + timedelta(days=1)
 
 
        # 📊 5. Prepare result DataFrame
        result = pd.DataFrame()
       
       
        # result.set_index("Milestone Track/Site Count", inplace=True)
 
        milestones = [
            "Site ONAIR Date",
            "RAN PAT Accepted Date",
            "RAN SAT Accepted Date",
            "SCFT Accepted Date",
            "KPI AT offer Date",
            "KPI AT Accepted Date",
            "4G MS2 Date",
            "5G MS2 Date",
            "Final MS2 Date"
        ]
 
        data = []
        
        for milestone in milestones:
            
            milestone_df_format = milestone.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")

            if milestone_df_format not in df.columns:
                continue
            
            milestone_data = pd.to_datetime(df[milestone_df_format], errors='coerce').dt.date
            row = {"Milestone Track/Site Count": milestone}
            if milestone_data.dropna().empty:
                row["CF"] = "-"
                # Set all month and week columns to "-"
                for _, end in months:
                    month_name = end.strftime("%b-%y")
                    row[month_name] = "-"
                for i, _ in enumerate(weeks, 1):
                    week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                    row[week_name] = "-"
                data.append(row)
                continue
 
 
            # CF → sites before financial year start
            row["CF"] = (milestone_data < fy_start).sum()
 
            cumulative_month = row["CF"]
            
            for start, end in months:
                month_name = end.strftime("%b-%y")
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_month += count
                if view == 'Cumulative':
                    row[month_name] = cumulative_month
                else:
                    row[month_name] = count
 
            cumulative_week = 0
            for i, (start, end) in enumerate(weeks, 1):
                week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                if(start >= today):
                    row[week_name] = 0
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_week += count
                if view == 'Cumulative':
                    row[week_name] = cumulative_week
                else:
                    row[week_name] = count
 
            data.append(row)
 

        result = pd.DataFrame(data)
 
        for col in result.columns:
            if col != "Milestone Track/Site Count":
                result[col] = result[col].astype(str)
 
        result = result.reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(lambda col: col.replace(" Date", "") if " Date" in col else col)
 
        new_data = result.copy()
        months_columns = [col for col in new_data.columns.to_list()[2:] if " W" not in col]
        months_data = new_data[new_data.columns.to_list()[:2] + months_columns].copy()
        
        weeks_columns = [col for col in new_data.columns.to_list()[2:] if " W" in col]
        
        week_data = new_data[new_data.columns.to_list()[:1] + weeks_columns].copy()
        
        months_data.rename(columns={
            col : f"Month-{i}" for i, col in enumerate(months_columns, start=1)
        }, inplace=True)
        
        week_data.rename(columns={
            col : f"Month_Week-{i}" for i, col in enumerate(weeks_columns, start=1)
        }, inplace=True)
        
        unique_data.update(**{"month_columns": months_columns, "week_columns" : weeks_columns})
        

        month_dict_data = months_data.to_dict(orient="records")
        month_json_data = json.dumps(month_dict_data)
        week_dict_data = week_data.to_dict(orient="records")
        week_json_data = json.dumps(week_dict_data)
        
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "upgrade_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"WEEKLY_MONTHY_MS2_DASHBOARD_FILE_{current_date}_{current_time}.xlsx")
       
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Monthly MS2')
       
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
   
        return Response({'message': 'Weekly and Monthly Dashboard created successfully !!!', "download_link": download_link, "unique_data": unique_data, "months_data": month_json_data, "week_data": week_json_data}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
   
   
   

@api_view(['GET', 'POST'])
def ms2_ageing_dashboard_table1(request):
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    # all_unique_site_tagging = list(
    #     UpgradeTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        UpgradeTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "RAN PAT Accepted",
        "RAN SAT Accepted",
        "SCFT Accepted",
        "KPI AT offer",
        "KPI AT Accepted",
        "4G MS2",
        "5G MS2",
        "Final MS2"
    ]
    
    
    
    unique_data = {
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones
    }

    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        # if site_tagging and "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if month_start and month_end:
            filters[f"{milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + '_date'}__range"] = (month_start, month_end)

        obj = UpgradeTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        date_cols = [col.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + '_date' for col in milestones]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']
        
        
        def generate_table_summary(df, milestone1, milestone2):


            milestone_cols = [(col.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date") for col in milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]]

            temp = df.copy()

            circles = ["ALL"] + temp["circle"].unique().tolist()

            final_df = pd.DataFrame()

            milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
            milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"

            for circle in circles:
                if circle != "ALL":
                    temp  = df[df["circle"] == circle]

                done_row = {}
                pending_row = {}

                for col, label in zip(milestone_cols, milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]):

                    done_row[label] = (temp[col].notna() & temp[milestone1_col].notna() ).sum()

                    pending_row[label] = (temp[col].isna() & temp[milestone1_col].notna() ).sum()

                rows_to_add = pd.DataFrame([
                    {"Circle": circle, "Site Status": "Done", **done_row},
                    {"Circle": circle, "Site Status": "Pending", **pending_row},
                ])

                final_df = pd.concat([final_df, rows_to_add], ignore_index=True)

            for c in final_df.columns:
                if c != "Site Status" and c != "Circle":
                    final_df[c] = final_df[c].astype("Int64")

            return final_df

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        table_summary = generate_table_summary(df, start_label, end_label)
 
        table_summary = table_summary.applymap(lambda x: str(x) if pd.notna(x) else "")      


        result_json = table_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "table_summary": json_data1
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
  
  
@api_view(['GET', 'POST'])
def ms2_ageing_dashboard_table2(request):
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    breakpoint1 = request.data.get('breakpoint1')
    breakpoint2 = request.data.get('breakpoint2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    type = request.data.get('type')

    # all_unique_site_tagging = list(
    #     UpgradeTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        UpgradeTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "RAN PAT Accepted",
        "RAN SAT Accepted",
        "SCFT Accepted",
        "KPI AT offer",
        "KPI AT Accepted",
        "4G MS2",
        "5G MS2",
        "Final MS2"
    ]
    
    unique_data = {
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones" : milestones
    }
    

    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    
    try:

        milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
        milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"

        breakpoint1 = int(breakpoint1)
        breakpoint2 = int(breakpoint2)
        
        temp = breakpoint1
        breakpoint1 = min(breakpoint1, breakpoint2)
        breakpoint2 = max(temp, breakpoint2)

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None

        milestones = [
            "Site ONAIR",
            "RAN PAT Accepted",
            "RAN SAT Accepted",
            "SCFT Accepted",
            "KPI AT offer",
            "KPI AT Accepted",
            "4G MS2",
            "5G MS2",
            "Final MS2"
        ]
        

        filters = {}
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
            
        if type == "type2":
            if month_start and month_end:
                filters[f"{milestone1_col}__range"] = (month_start, month_end)
                

        obj = UpgradeTracker.objects.filter(**filters)  # noqa: F405
        
        if type == "type1":
            if month_start and month_end:
                obj = obj.filter(
                    Q(**{f"{milestone1_col}__range": (month_start, month_end)}) |
                    Q(**{f"{milestone2_col}__range": (month_start, month_end)})
                )
        

        df = pd.DataFrame(obj.values())

        
        date_cols = [col.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + '_date' for col in milestones]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']

        if month_start is not None:
            month_start = pd.to_datetime(month_start)

        if month_end is not None:
            month_end = pd.to_datetime(month_end)

        def generate_done_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            temp = df.copy()
            
            temp["days_diff"] = (temp[end_col] - temp[start_col]).dt.days
            
            if type == "type2":
        
                summary = (
                    temp.groupby("Circle").apply(lambda g: pd.Series({
                        f"{start_label} Done Count": g[start_col].notna().sum(),
                        f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].notna(), "days_diff"].mean()
                    }))
                    .reset_index()
                )
                
                total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                total_row["Circle"] = "Total"

                overall_valid = temp.loc[
                    temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].notna(),
                    "days_diff"
                ]
                total_row["Average Days"] = overall_valid.mean()

            else:
                if month_start is not None and month_end is not None:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": (g[start_col].notna() & (g[start_col] >= month_start) & (g[start_col] <= month_end)).sum(),
                            f"{end_label} Done Count": (g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"<= {breakpoint1} days": (g["days_diff"].notna() & (g["days_diff"] <= breakpoint1) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": (g["days_diff"].notna() & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f">= {breakpoint2} days": (g["days_diff"].notna() & (g["days_diff"] >= breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna() & (temp[end_col] >= month_start) & (temp[end_col] <= month_end),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                else:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": g[start_col].notna().sum(),
                            f"{end_label} Done Count": g[end_col].notna().sum(),
                            f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & g[end_col].notna()).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna()).sum(),
                            f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & g[end_col].notna()).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna(), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna(),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                    
            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]

            return summary


        done_summary = generate_done_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        done_summary = done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        def generate_pending_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            
            today = pd.Timestamp.today()
            
            temp = df.copy()
            
            temp["days_diff"] = (today - temp[start_col]).dt.days

            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                    f"<= {breakpoint1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[start_col].notna())).sum(),
                    f"{breakpoint1+1}-{breakpoint2-1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[start_col].notna())).sum(),
                    f">= {breakpoint2} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[start_col].notna())).sum(),
                    "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].isna(), "days_diff"].mean()
                }))
                .reset_index()
            )

            total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
            total_row["Circle"] = "Total"

            overall_valid = temp.loc[
                temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].isna(),
                "days_diff"
            ]
            total_row["Average Days"] = overall_valid.mean()

            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Pending Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]
            
            return summary
        

        pending_summary = generate_pending_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        pending_summary = pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "nt_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"AGEING_DASHBOARD_FILE_{current_date}_{current_time}.xlsx")

        result_json = pending_summary.to_dict(orient="records")
        json_data2 = json.dumps(result_json)

        result_json = done_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "done_summary": json_data1,
            "pending_summary": json_data2
        }
        
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            pending_summary.to_excel(writer, index=False, sheet_name='PENDING SUMMARY')
            done_summary.to_excel(writer, index=False, sheet_name='DONE SUMMARY')
        
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
        
        return Response({'message': 'request processed successfully !!!', "download_link": download_link, "json_data": json_data, "unique_data": unique_data, "breakpoint1" : breakpoint1, "breakpoint2" : breakpoint2}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)


@api_view(['GET', 'POST'])
def ms2_graphs_view(request):
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    # all_unique_site_tagging = list(
    #     UpgradeTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        UpgradeTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "RAN PAT Accepted",
        "RAN SAT Accepted",
        "SCFT Accepted",
        "KPI AT offer",
        "KPI AT Accepted",
        "4G MS2",
        "5G MS2",
        "Final MS2"
    ]
    
    unique_data = {
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones
    }

    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        # if site_tagging and "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if month_start and month_end:
            filters[f"{milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '') + '_date'}__range"] = (month_start, month_end)

        obj = UpgradeTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        df['Circle'] = df['circle']

        def generate_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"{end_label} Pending Count"]
            ]

            return summary

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        start_col = start_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_") + "_date"
        end_col = end_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_") + "_date"
        
        graph_summary = generate_summary(df, start_label, end_label, start_col, end_col)

        graph_summary = graph_summary.applymap(lambda x: str(x) if pd.notna(x) else "")   

        result_json = graph_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "graph_summary": json_data1,
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)


@api_view(['GET', 'POST'])
def ms2_monthly_graph(request):
    circle = request.data.get('circle', [])
    # site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    view = request.data.get('view')
    type = request.data.get('type')
    
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    # site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"
    milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"
    
    all_unique_circles = list(
        UpgradeTracker.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    # all_unique_site_tagging = list(
    #     UpgradeTracker.objects.exclude(site_tagging__isnull=True)
    #     .distinct("site_tagging")
    #     .values_list("site_tagging", flat=True)
    # )

    all_unique_current_status = list(
        UpgradeTracker.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        UpgradeTracker.objects.exclude(toco_name__isnull=True)
        .distinct("toco_name")
        .values_list("toco_name", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "RAN PAT Accepted",
        "RAN SAT Accepted",
        "SCFT Accepted",
        "KPI AT offer",
        "KPI AT Accepted",
        "4G MS2",
        "5G MS2",
        "Final MS2"
    ]
    
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        # "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
        "milestones": milestones
    }
    
    try:
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        # if "ALL" not in site_tagging:
        #     filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["toco_name__in"] = new_toco_name
        
            
            
        obj = UpgradeTracker.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        # print(df)
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        df = df[[milestone1_col, milestone2_col]]
       
        for col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        
        today = dtime.today().date()
        
        df["month_num1"] = (
            df[milestone1_col].dt.month + (df[milestone1_col].dt.day >= 26)
        ).replace({13: 1}).astype("Int64")

        df["year1"] = (
            df[milestone1_col].dt.year - (
                (df[milestone1_col].dt.month < 3) |
                ((df[milestone1_col].dt.month == 3) & (df[milestone1_col].dt.day < 26))
            )
        ).astype("Int64")

        # 3️⃣ Display year
        df["display_year"] = df["year1"] + (df["month_num1"] <= 3)

        # 4️⃣ Month label
        df["month_name1"] = pd.to_datetime(
            df["month_num1"].astype(str) + "-" + df["display_year"].astype(str),
            format="%m-%Y",
            errors="coerce"
        ).dt.strftime("%b-%y")

        # 5️⃣ Current FY
        current_fy = today.year if today >= date(today.year, 3, 26) else today.year - 1

        # 6️⃣ CF logic
        df.loc[df["year1"] < current_fy, "month_name1"] = "CF"



        df['month_num2'] = df[milestone2_col].dt.month + (df[milestone2_col].dt.day >= 26)
        df['month_num2'] = df['month_num2'].replace({13: 1})

        # 2️⃣ Financial year (derived ONLY from original date)
        df['year2'] = df[milestone2_col].dt.year - (
            (df[milestone2_col].dt.month < 3) |
            ((df[milestone2_col].dt.month == 3) & (df[milestone2_col].dt.day < 26))
        )

        # 3️⃣ Safe numeric conversion (IMPORTANT)
        df['month_num2'] = pd.to_numeric(df['month_num2'], errors='coerce')
        df['year2'] = pd.to_numeric(df['year2'], errors='coerce')

        # 4️⃣ Display year
        df['display_year2'] = df['year2'] + (df['month_num2'] <= 3)

        # 5️⃣ Month label (NaT-safe)
        df['month_name2'] = pd.to_datetime(
            df['month_num2'].astype('Int64').astype(str) + '-' +
            df['display_year2'].astype('Int64').astype(str),
            format='%m-%Y',
            errors='coerce'
        ).dt.strftime('%b-%y')

        # 6️⃣ Current financial year
        current_fy = today.year if today >= date(today.year, 3, 26) else today.year - 1

        # 7️⃣ Carry Forward (only where year2 exists)
        df.loc[df['year2'].notna() & (df['year2'] < current_fy), 'month_name2'] = 'CF'
        
        print(df['month_name1'].dropna().unique().tolist())

        
        print(df['year1'])
        
        print(df['month_name2'])
        
        print(df['year2'])

        def sort_financial_year(summary):
            # Extract sorted unique Month-Year (excluding CF)
            months = [m for m in summary['month_name'].unique() if m != "CF"]

            # Convert to datetime to sort correctly
            months_sorted = sorted(
                months,
                key=lambda x: pd.to_datetime(x, format='%b-%y')
            )

            # CF should always be first
            ordered = ["CF"] + months_sorted if "CF" in summary['month_name'].values else months_sorted

            # Convert into ordered categorical for final sort
            summary['month_name'] = pd.Categorical(summary['month_name'], ordered, ordered=True)

            return summary.sort_values('month_name')
        
        def generate_type1_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()

            # Start counts by month_name
            start_counts = (
                temp[temp[start_col].notna()]
                .groupby("month_name1")
                .size()
                .reset_index(name=f"{start_label} Done Count")
                .rename(columns={"month_name1": "month_name"})
            )

            # End counts by month_name2
            end_counts = (
                temp[temp[end_col].notna()]
                .groupby("month_name2")
                .size()
                .reset_index(name=f"{end_label} Done Count")
                .rename(columns={"month_name2": "month_name"})
            )

            # Merge both results
            summary = pd.merge(start_counts, end_counts, on="month_name", how="outer")

            # Convert to Int64 nullable type
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")
            summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].astype("Int64")

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            return summary.reset_index(drop=True)
        
        def generate_type2_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("month_name1").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    # f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
                .rename(columns={"month_name1": "month_name"})
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")

            summary = summary[
                ["month_name", f"{start_label} Done Count", f"{end_label} Done Count"]
            ]

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            return summary.reset_index(drop=True)
        
        summary = generate_type1_summary(df, milestone1, milestone2, milestone1_col, milestone2_col) if type == 'type1' else generate_type2_summary(df, milestone1, milestone2, milestone1_col, milestone2_col)

        numeric_cols = [col for col in summary.columns if "Done Count" in col]

        summary[numeric_cols] = summary[numeric_cols].apply(pd.to_numeric, errors="coerce")
        summary[numeric_cols] = summary[numeric_cols].fillna(0).astype(int)

        
        result_json = summary.to_dict(orient = "records")
        json_data = json.dumps(result_json)
        
        return Response({"message" : "request processed successfully ! ", "json_data": json_data, "unique_data": unique_data}, status=200)
    
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
 