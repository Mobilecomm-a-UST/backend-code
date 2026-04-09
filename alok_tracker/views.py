from django.http import JsonResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from alok_tracker.models import *  # noqa: F403
import pandas as pd
from django.db import transaction
import os
from django.conf import settings
from datetime import datetime as dtime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.utils import timezone
import json
from datetime import datetime, timedelta, date
import shutil
from django.db.models import Q
from django.forms.models import model_to_dict
import numpy as np
from rest_framework import status


CENTRAL_COLUMNS = [
    'Integration Date',    
    'EMF Submission Date',  
    'Alarm Status',    
    'Alarm Rectification Done Date',    
    'SCFT Done Date',  
    'SCFT Offered Date',    
    'RAN PAT Offer Date',
    'RAN SAT Offer Date' ,
    'MW PAT Offer Date',
    'MW SAT Offer Date',    
    'MW MS1 MIDS Date',  
    'Site ONAIR Date',  
    'I-Deploy ONAIR Date',  
    '5G ONAIR Date',    
    'RAN PAT Accepted Date',    
    'RAN SAT Accepted Date',    
    'MW PAT Accepted Date',
    'MW SAT Accepted Date',
    'SCFT Accepted Date',
    'KPI AT offer Date',
    'KPI AT Accepted Date',
    '4G MS2 Date',
    '5G MS2 Date',
    'Final MS2 Date',
    'Milestone Status'
]

CIRCLE_COLUMNS = [
    'Circle',
    'Site Tagging',
    'Old TOCO Name',
    'Old Site Id',
    'New Site ID',
    'New TOCO Name',
    'SR Number',
    'RAN OEM',
    'Media Type',
    'MW OEM',
    'Relocation Method',
    'Relocation Type',
    'OLD Site Band',
    'New Site Band',
    'RFAI Date' ,
    'Allocation Date',
    'RFAI Survey Date',
    'RFAI Survey Done Date',
    'MO Punch Date',
    'Material Dispatch Date',
    'Material Delivered Date',
    'Installation Start Date',
    'Installation End Date',
    'RAN LKF Status',
    'MW Plan ID',
    'RSL Value Status',
    'ENM Status',
    'MW LKF'
]

ALL_COLUMNS = [
    # 'Unique ID',
    'Circle',
    'Site Tagging',
    'Old TOCO Name',
    'Old Site Id',
    'New Site ID',
    'New TOCO Name',
    'SR Number',
    'RAN OEM',
    'Media Type',
    'MW OEM',
    'Relocation Method',
    'Relocation Type',
    'OLD Site Band',
    'New Site Band',
    'RFAI Date',
    'Allocation Date',
    'RFAI Survey Date',
    'MO Punch Date',
    'Material Dispatch Date',
    'Material Delivered Date',
    'Installation Start Date',
    'Installation End Date',
    'Integration Date',
    'EMF Submission Date',
    'RAN LKF Status',
    'Alarm Status',
    'Alarm Rectification Done Date',
    'SCFT Done Date',
    'SCFT I-Deploy Offered Date',
    'RAN PAT Offer Date',
    'RAN SAT Offer Date',
    'MW Plan ID',
    'MW PAT Offer Date',
    'RSL Value Status',
    'ENM Status',
    'MW LKF',
    'MW SAT Offer Date',
    'MW MS1 MIDS Date',
    'Site ONAIR Date',
    'I-Deploy ONAIR Date',
    'Current Status',
    'Detailed Remarks',
    'RFAI Rejected Date',
    'Clear RFAI Date',
    'PRI Count',
    'PRI Issue Ageing',
    'Other UST Issue Ageing',
    'Other Airtel Issue Ageing',
    'Total Issue Ageing',
    'Clear RFAI to MS1 Ageing',
    'RFAI to MS1 Ageing',
    'RAN PAT Accepted Date',
    'RAN SAT Accepted Date',
    'MW PAT Accepted Date',
    'MW SAT Accepted Date',
    'SCFT Accepted Date',
    'KPI AT offer Date',
    'KPI AT Accepted Date',
    '4G MS2 Date',
    '5G MS2 Date',
    'Final MS2 Date',
    "Dismantling Survey Date",
    "SREQ/CREQ Raised Date",
    "Dismantle Date",
    "Material Pickup Date",
    "Material Submission Date",
    "OCI Done Date",
    "Sign-off Date",
    "Dismantling Status"
    # "TOCO/Owner Issue",
    # "EXIT Notice Issue",
    # "Commercial Issue",
    # "Workable Sites"
]

DUMP_COLUMNS = [
    "Circle",
    "Site ID",
    "On Air Date",
    "Band",
    "Physical AT Status",
    "Physical AT Rejection Counter",
    "Performance AT Status",
    "Performance AT Rejection Counter",
    "Soft AT Status",
    "Soft AT Rejection Counter",
    "SCFT AT Status",
    "SCFT AT Rejection Counter"
]

######################################################### IMPORTANT FUNCTIONS ##############################################################

def compute_union_ageing(rows):
    if rows.empty:
        return 0

    today = datetime.today()

    def to_datetime(x):
        if isinstance(x, datetime):
            return x
        if isinstance(x, date):
            return datetime.combine(x, datetime.min.time())
        if isinstance(x, str) and x not in ["", "-", "nan", None]:
            return datetime.strptime(x, "%d-%b-%y")
        return None

    ranges = []

    for _, row in rows.iterrows():
        start = to_datetime(row["Start Date"])
        close = row["Close Date"]
        # close = to_datetime(close)

        end = today if close in ["", "-", "nan", None, "undefined"] else close
        end = to_datetime(end)
        ranges.append((start, end))

    # Sort
    ranges.sort(key=lambda x: x[0])

    # Merge intervals
    merged = []
    current_start, current_end = ranges[0]

    for start, end in ranges[1:]:
        if start <= current_end:
            current_end = max(current_end, end)
        else:
            merged.append((current_start, current_end))
            current_start, current_end = start, end

    merged.append((current_start, current_end))

    return sum((end - start).days for start, end in merged)

def update_ageing(df, issue_df):
    pri_ageing_list = []
    other_airtel_issue_ageing_list = []
    other_ust_issue_ageing_list = []
    total_issue_ageing_list = []
    pri_count_list = []

    # Ensure datetime
    issue_df["Start Date"] = pd.to_datetime(issue_df["Start Date"], errors="coerce")
    issue_df["Close Date"] = pd.to_datetime(issue_df["Close Date"], errors="coerce")

    for site in df["new_site_id"]:

        site_rows = issue_df[issue_df["Site ID"] == site]

        pri_rows = site_rows[site_rows["Issue Name"] == "PRI"]

        other_rows_airtel = site_rows[
            (site_rows["Issue Name"] != "PRI") &
            (site_rows["Issue Owner"] == 'Airtel')
        ]

        other_rows_ust = site_rows[
            (site_rows["Issue Name"] != "PRI") &
            (site_rows["Issue Owner"] == 'UST')
        ]

        all_rows = site_rows

        pri_ageing = compute_union_ageing(pri_rows)
        other_airtel_issue_ageing = compute_union_ageing(other_rows_airtel)
        other_ust_issue_ageing = compute_union_ageing(other_rows_ust)
        total_issue_ageing = compute_union_ageing(all_rows)

        pri_count = len(pri_rows)

        pri_ageing_list.append(pri_ageing)
        other_airtel_issue_ageing_list.append(other_airtel_issue_ageing)
        other_ust_issue_ageing_list.append(other_ust_issue_ageing)
        total_issue_ageing_list.append(total_issue_ageing)
        pri_count_list.append(pri_count)

    df["pri_issue_ageing"] = pri_ageing_list
    df["other_airtel_issue_ageing"] = other_airtel_issue_ageing_list
    df["other_ust_issue_ageing"] = other_ust_issue_ageing_list
    df["total_issue_ageing"] = total_issue_ageing_list
    df["pri_count"] = pri_count_list

    # Convert dates
    df['site_onair_date'] = pd.to_datetime(df['site_onair_date'], errors='coerce')
    df['clear_rfai_date'] = pd.to_datetime(df['clear_rfai_date'], errors='coerce')
    df['rfai_date'] = pd.to_datetime(df['rfai_date'], errors='coerce')

    # 🔥 NEW LOGIC
    def calculate_adjusted_issue_ageing(site, clear_rfai):
        if pd.isna(clear_rfai):
            return None

        site_rows = issue_df[issue_df["Site ID"] == site]

        total_days = 0

        for _, issue in site_rows.iterrows():
            start = issue["Start Date"]
            end = issue["Close Date"]

            if pd.isna(start):
                continue

            if pd.isna(end):
                end = datetime.today()

            # Case 1: starts after clear_rfai
            if start >= clear_rfai:
                total_days += (end - start).days

            # Case 2: overlaps clear_rfai
            elif start < clear_rfai and end > clear_rfai:
                total_days += (end - clear_rfai).days

        return total_days

    def calculate_clear_rfai_to_ms1(row):
        site = row["new_site_id"]
        site_onair = row['site_onair_date']
        rfai = row['clear_rfai_date'] if pd.notna(row['clear_rfai_date']) else row['rfai_date']

        if pd.isna(rfai):
            return "-"

        # 🔥 use adjusted ageing if clear_rfai present
        if pd.notna(row['clear_rfai_date']):
            issue_ageing = calculate_adjusted_issue_ageing(site, row['clear_rfai_date'])
        else:
            issue_ageing = row['total_issue_ageing'] if pd.notna(row['total_issue_ageing']) else 0

        if issue_ageing is None:
            issue_ageing = 0

        if pd.isna(site_onair):
            today = datetime.today()
            return (today - rfai).days - issue_ageing

        return (site_onair - rfai).days - issue_ageing

    df['clear_rfai_to_ms1_ageing'] = df.apply(calculate_clear_rfai_to_ms1, axis=1)
    
    def calculate_rfai_to_ms1(row):
        site = row["new_site_id"]
        site_onair = row['site_onair_date']
        rfai = row['rfai_date']

        if pd.isna(rfai):
            return "-"

        if pd.isna(site_onair):
            today = datetime.today()
            return (today - rfai).days

        return (site_onair - rfai).days
    
    df['rfai_to_ms1_ageing'] = df.apply(calculate_rfai_to_ms1, axis=1)

    return df

def update_ageing_new(circle, site_id):
    # Fetch tracker row
    data_obj = AlokTrackerModel.objects.filter(
        circle=circle,
        new_site_id=site_id
    ).first()

    # if not data_obj:
    #     return None

    # Fetch issues
    issue_obj = RelocationIssue.objects.filter(
        circle=circle,
        site_id=site_id
    )

    issue_df = pd.DataFrame(issue_obj.values())
    
    rename_map = {
        "circle": "Circle",
        "site_id": "Site ID",
        "issue_owner": "Issue Owner",
        "milestone": "Milestone",
        "issue_name": "Issue Name",
        "start_date": "Start Date",
        "close_date": "Close Date",
        "status": "Status",
        "duration": "Duration",
        "updated_by": "Updated_by",
        "updated_at": "Updated_at",
        "created_by": "Created_by",
        "created_at": "Created_at"
    }
    
    issue_df = issue_df.rename(columns=rename_map)

    if issue_df.empty:
        pri_ageing = 0
        other_issue_ageing = 0
        total_issue_ageing = 0
        pri_count = 0
    else:
        pri_rows = issue_df[issue_df["Issue Name"] == "PRI"]
        other_rows = issue_df[issue_df["Issue Name"] != "PRI"]

        pri_ageing = compute_union_ageing(pri_rows)
        other_issue_ageing = compute_union_ageing(other_rows)
        total_issue_ageing = compute_union_ageing(issue_df)
        pri_count = len(pri_rows)

    # Assign ageing values
    data_obj.pri_issue_ageing = pri_ageing
    data_obj.other_issue_ageing = other_issue_ageing
    data_obj.total_issue_ageing = total_issue_ageing
    data_obj.pri_count = pri_count

    # Convert dates
    site_onair = pd.to_datetime(data_obj.site_onair_date, errors="coerce")
    clear_rfai = pd.to_datetime(data_obj.clear_rfai_date, errors="coerce")
    rfai = pd.to_datetime(data_obj.rfai_date, errors="coerce")

    rfai_final = clear_rfai if pd.notna(clear_rfai) else rfai

    # Calculate clear_rfai_to_ms1_ageing
    if pd.isna(rfai_final):
        data_obj.clear_rfai_to_ms1_ageing = "-"
    else:
        issue_ageing = total_issue_ageing or 0

        if pd.isna(site_onair):
            today = datetime.today()
            data_obj.clear_rfai_to_ms1_ageing = (today - rfai_final).days - issue_ageing
        else:
            data_obj.clear_rfai_to_ms1_ageing = (site_onair - rfai_final).days - issue_ageing

    # Save changes
    data_obj.save()

    return


def get_ftr_month(date):
    if pd.isna(date):
        return None

    day = date.day
    month = date.month
    year = date.year

    # April special case: 29 Mar – 25 Apr
    if (month == 3 and day >= 29) or (month == 4 and day <= 25):
        return f"Apr'{str(year)[-2:]}"

    # Normal months: 26th prev → 25th current
    if day >= 26:
        next_month = month + 1 if month < 12 else 1
        next_year = year if month < 12 else year + 1
        return pd.to_datetime(f"{next_year}-{next_month}-01").strftime("%b'%y")
    else:
        return date.strftime("%b'%y")

def ftr_count(df, status_col, reject_col):
    return (
        df[
            (df[status_col] == 'Accepted') &
            (df[reject_col] == 0)
        ]
        .groupby('FTR_Month')['site_id']
        .count()
    )

MONTH_ORDER = {
    "Apr": 1, "May": 2, "Jun": 3, "Jul": 4,
    "Aug": 5, "Sep": 6, "Oct": 7, "Nov": 8,
    "Dec": 9, "Jan": 10, "Feb": 11, "Mar": 12
}

def month_sort_key(label):
    # label like "Apr'25"
    month = label.split("'")[0]
    year = int(label.split("'")[1])

    # Adjust year for FY (Jan–Mar should belong to previous year)
    if month in ["Jan", "Feb", "Mar"]:
        year -= 1

    return (year, MONTH_ORDER[month])

def build_circlewise_ftr(df, status_col, reject_col):
    # Total rows per circle + month
    total = (
        df.groupby(['circle', 'FTR_Month'])
        .size()
        .rename('Total')
    )

    # FTR rows
    ftr = (
        df[
            (df[status_col] == 'Accepted') &
            (df[reject_col] == 0)
        ]
        .groupby(['circle', 'FTR_Month'])
        .size()
        .rename('FTR')
    )

    cm = pd.concat([total, ftr], axis=1).fillna(0)

    cm['FTR %'] = (
        cm['FTR'] / cm['Total'] * 100
    ).round().astype(int)

    # Pivot (circle × month)
    table = cm['FTR %'].unstack('FTR_Month')

    # Sort months
    table = table[sorted(table.columns, key=month_sort_key)]

    # Overall per circle (IMPORTANT)
    overall = (
        cm
        .groupby('circle')[['FTR', 'Total']]
        .sum()
    )

    overall_pct = (
        overall['FTR']
        .div(overall['Total'])
        .mul(100)
        .round()
    )


    table['Overall'] = overall_pct

    # Total row (all circles combined)
    total_row = (
        cm.groupby('FTR_Month')
        .apply(
            lambda x: round(
                x['FTR'].sum() / x['Total'].sum() * 100
            ) if x['Total'].sum() else 0
        )
    )

    total_row = total_row.reindex(table.columns[:-1])
    total_row['Overall'] = round(
        cm['FTR'].sum() / cm['Total'].sum() * 100
    )

    table.loc['Total'] = total_row
    

    # Move Total row to top
    new_index = ['Total'] + [i for i in table.index if i != 'Total']
    table = table.reindex(new_index)

    
    def format_ftr_percentage_table(table):
        def format_cell(x):
            if pd.isna(x):
                return '-'
            return f"{int(round(x))}%"

        return table.applymap(format_cell)
    
    table = format_ftr_percentage_table(table)

    return table

########################################################### USER VIEWS ###################################################################

@api_view(['POST'])
def users_display(request):
    adminId = request.data.get('adminId').lower()
    
    if not adminId:
        return Response({"error": "adminId is required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        admin_user = RelocationUser.objects.get(email=adminId)
    except RelocationUser.DoesNotExist:
        return Response({"error": "Admin_user not found"}, status=status.HTTP_404_NOT_FOUND)

    if admin_user.right != 'Admin':
        return Response({"message": "ACCESS TO VIEW USER LIST DENIED!"}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        
        if 'CENTRAL' in admin_user.circles:
            obj = RelocationUser.objects.all()
        else:
            q = Q()
            for circle in admin_user.circles:
                q |= Q(circles__contains=[circle])

            obj = RelocationUser.objects.filter(q)

        df = pd.DataFrame(obj.values())
        
        df['created_at'] = pd.to_datetime(df['created_at']).dt.strftime('%d-%b-%y')
        df['updated_at'] = pd.to_datetime(df['updated_at']).dt.strftime('%d-%b-%y')

        
        json_data = df.to_dict(orient="records")
        
        return Response({"message": "request processed successfully", "json_data" : json_data, "columns": ALL_COLUMNS}, status=status.HTTP_200_OK)
    
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def user_create(request):
    adminId = request.data.get('adminId').lower()   # Admin who is creating user
    name = request.data.get('name').lower()
    email = request.data.get('email').lower()
    circles = request.data.get('circles')
    columns = request.data.get('columns')
    right = request.data.get('right')

    # Check required fields
    required_fields = [adminId, name, email, circles, columns, right]
    if not all(required_fields):
        return Response({"error": "All fields are required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        admin_user = RelocationUser.objects.get(email=adminId)
    except RelocationUser.DoesNotExist:
        return Response({"error": "Admin user not found"}, status=status.HTTP_404_NOT_FOUND)

    if admin_user.right != 'Admin':
        return Response({"message": "ACCESS TO CREATE USER DENIED!"}, status=status.HTTP_403_FORBIDDEN)

    if RelocationUser.objects.filter(email=email).exists():
        return Response({"error": "A user with this email already exists"}, status=status.HTTP_409_CONFLICT)

    try:
        circles = [c.strip() for c in circles.split(",")]
        columns = [c.strip() for c in columns.split(",")]
        
        new_user = RelocationUser.objects.create(
            name=name,
            email=email.lower(),
            circles=circles,
            columns=columns,
            right=right,
            created_by=adminId,
            updated_by=adminId
        )

        return Response({
            "message": "New user created successfully",
            "user": {
                "id": new_user.id,
                "name": new_user.name,
                "email": new_user.email,
                "right": new_user.right,
                "circles": circles,
                "columns": columns
            }
        }, status=status.HTTP_201_CREATED)
    
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST'])
def user_delete(request):
    adminId = request.data.get('adminId').lower()
    email = request.data.get('email').lower()  

    if not adminId or not email:
        return Response({"error": "adminId and user email are required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        admin_user = RelocationUser.objects.get(email=adminId)
    except RelocationUser.DoesNotExist:
        return Response({"error": "Admin user not found"}, status=status.HTTP_404_NOT_FOUND)

    # Permission check
    if admin_user.right != 'Admin':
        return Response({"message": "ACCESS TO DELETE USER DENIED!"}, status=status.HTTP_403_FORBIDDEN)

    # Check if target user exists
    try:
        target_user = RelocationUser.objects.get(email=email)
    except RelocationUser.DoesNotExist:
        return Response({"error": "User with this email does not exist"}, status=status.HTTP_404_NOT_FOUND)

    # Prevent admin from deleting themselves
    if admin_user.email == target_user.email:
        return Response({"error": "You cannot delete yourself"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        target_user.delete()
        
        return Response({"message": "User deleted successfully"}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST'])
def user_update(request):
    adminId = request.data.get('adminId').lower()
    email = request.data.get('email').lower()

    circles = request.data.get('circles')
    columns = request.data.get('columns')
    right = request.data.get('right')

    if not adminId or not email:
        return Response({"error": "adminId and user email are required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        admin_user = RelocationUser.objects.get(email=adminId)
    except RelocationUser.DoesNotExist:
        return Response({"error": "Admin user not found"}, status=status.HTTP_404_NOT_FOUND)

    if admin_user.right != 'Admin':
        return Response({"message": "ACCESS TO UPDATE USER DENIED!"}, status=status.HTTP_403_FORBIDDEN)

    try:
        target_user = RelocationUser.objects.get(email=email)
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

    try:
        if circles:
            target_user.circles = [c.strip() for c in circles.split(",")]

        if columns:
            target_user.columns = [c.strip() for c in columns.split(",")]

        if right:
            target_user.right = right

        target_user.updated_by = adminId 
        target_user.save()

        return Response({
            "message": "User updated successfully",
            "updated_user": {
                "name": target_user.name,
                "email": target_user.email,
                "circles": target_user.circles,
                "columns": target_user.columns,
                "right": target_user.right
            }
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


############################################################  UPLOAD DATA #####################################################################################

@api_view(["POST"])
def upload_tracker_data_view(request):
    userId = request.data.get('userId')
    
    if not userId:
        return Response({'error': 'userId required.'}, status=400)
    
    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    print(user)
    
    if user.right == 'Read':
        return Response({"message": "ACCESS DENIED!"}, status=status.HTTP_403_FORBIDDEN)
    
    file = request.FILES.get("tracker_file")

    # if not userId or not circle:
    #     return Response({'error': 'userId and circle are required.'}, status=400)

    if not file:
        return Response({"status": False, "message": "No file provided."}, status=400)

    # if not ACCESS_RIGHTS[userId]["can_upload"]:
    #     return Response({'error': 'Upload Rights Not Granted'}, status=400)

    try:
        circles = user.circles
        df = pd.read_csv(file, header=1) if file.name.endswith('.csv') else pd.read_excel(file, header=1)
        df.columns = [col.strip() for col in df.columns]
        df = df.where(pd.notnull(df), None)
        df.fillna("", inplace=True)
        df = df.replace("NaN", None)
        df = df.replace("", None)
        
        
        
        if 'CENTRAL' not in circles:
            df = df[df['Circle'].notna() & df['Circle'].isin(circles)]
        

        INVALID_DATE_STRINGS = {"need to check", "pending", "na", "tbd", "n/a", "nan", "-"}

        def safe_datetime(value):
            import pandas as pd
            if pd.isna(value) or value is pd.NaT:
                return None
            if isinstance(value, (pd.Timestamp, dtime)):
                return value
            if isinstance(value, str):
                v = value.strip().lower()
                if v in INVALID_DATE_STRINGS or not v:
                    return None
                try:
                    val = pd.to_datetime(value, errors="coerce")
                    if pd.isna(val):
                        return None
                    return val.to_pydatetime()
                except Exception:
                    return None
            return None

        for col in df.columns:
            if "date" in col:
                df[col] = df[col].apply(safe_datetime)

        def safe_int(val):
            try:
                return int(val)
            except (TypeError, ValueError):
                return None

        required_columns = ALL_COLUMNS if 'ALL' in user.columns else user.columns
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            return Response(
                {"status": False, "error": f"Missing columns: {', '.join(missing_columns)}"},
                status=400
            )

        df = df[required_columns]

        df['New Site ID'] = df['New Site ID'].astype(str)
        records = df.to_dict(orient='records')

        existing_records = {
            (rec.circle.strip().lower(), str(rec.new_site_id).strip().lower()): rec
            for rec in AlokTrackerModel.objects.only("circle", "new_site_id")
        }

        objects_to_create, objects_to_update = [], []

        with transaction.atomic():
            for row in records:
                circle_val = str(row.get("Circle")).strip() if row.get("Circle") else None
                new_site_id_val = str(row.get("New Site ID")).strip() if row.get("New Site ID") else None

                if not circle_val or not new_site_id_val:
                    continue

                allowed_data = {}
                for col in required_columns:
                    val = row.get(col)

                    if "date" in col.lower():
                        val = safe_datetime(val)
                        if val and val.date() > timezone.localdate():
                            return Response(
                                {
                                    "status": False,
                                    "error": f"Future date found in column '{col}'",
                                    "details": {
                                        "circle": circle_val,
                                        "new_site_id": new_site_id_val,
                                        "column": col,
                                        "value": val
                                    }
                                },
                                
                                status=400
                            )
                    elif "ageing" in col.lower() or "count" in col.lower():
                        val = safe_int(val)

                    field_name = (
                        col.lower()
                        .strip()
                        .replace(" ", "_")
                        .replace("-", "_")
                        .replace("4", "four_")
                        .replace("5", "five_")
                        .replace("/", "_")
                    )
                    allowed_data[field_name] = val

                allowed_data.update({
                    "circle": circle_val,
                    "new_site_id": new_site_id_val,
                    "last_updated_by": userId,
                    "last_updated_date": timezone.now(),
                })

                key = (circle_val.lower(), new_site_id_val.lower())

                if key in existing_records:
                    obj = existing_records[key]
                    for field, value in allowed_data.items():
                        setattr(obj, field, value)

                    obj.last_updated_date = timezone.now()
                    obj.last_updated_by = userId
                    objects_to_update.append(obj)

                else:
                    objects_to_create.append(AlokTrackerModel(**allowed_data))

            if objects_to_create and (userId.lower() == 'manish.kumar4@ust.com' or userId.lower() == 'mohsin.khan@ust.com' or userId.lower() == 'girraj.singh@mcpsinc.in' or userId.lower() == 'devansh.jain@ust.com'):
                AlokTrackerModel.objects.bulk_create(objects_to_create, batch_size=500)

            if objects_to_update:
                AlokTrackerModel.objects.bulk_update(
                    objects_to_update,
                    fields=list({
                        col.lower()
                        .strip()
                        .replace(" ", "_")
                        .replace("-", "_")
                        .replace("4", "four_")
                        .replace("5", "five_")
                        .replace("/", "_")
                        for col in required_columns
                    }) + ['last_updated_date', 'last_updated_by'],
                    batch_size=500,
                )

        return Response({"status": True, "message": "Data inserted successfully."})

    except Exception as e:
        return Response({"status": False, "error": str(e)}, status=500)


@api_view(["POST"])
def upload_issues_data_view(request):
    file = request.data.get("tracker_file")

    try:
        df = pd.read_csv(file) if file.name.endswith('.csv') else pd.read_excel(file)
        
        # Clean dataframe safely
        # df.columns = df.columns.astype(str).str.strip()

        # Convert NaN properly
        df = df.replace({pd.NA: None, "NaN": None})

        # DO NOT convert everything to ""
        # This breaks datetime columns
        # print(df.columns.tolist())
        # Rename columns
        df.columns = [
            str(col).strip().lower().replace(" ", "_")
            for col in df.columns
        ]

        df = df[
            [
                "circle",
                "site_id",
                "issue_owner",
                "milestone",
                "issue_name",
                "start_date",
                "close_date",
                "remarks",
            ]
        ]

        user_email = "devansh.jain@ust.com"
        today = date.today()

        issues_to_create = []
        
        def safe_strip(val):
            if val is None:
                return None
            return str(val).strip()

        for _, row in df.iterrows():
            try:
                # Required fields check
                if not all([
                    row["circle"],
                    row["site_id"],
                    row["issue_owner"],
                    row["milestone"],
                    row["issue_name"],
                    row["start_date"]
                ]):
                    continue  # skip invalid row

                # Parse dates
                try:
                    start_date = pd.to_datetime(row["start_date"]).date()
                except:
                    continue

                close_date = None
                if row["close_date"] not in [None, "-", "nan", "undefined"]:
                    try:
                        close_date = pd.to_datetime(row["close_date"]).date()
                    except:
                        close_date = None

                # Status + duration
                if close_date:
                    status_val = "Closed"
                    duration = (close_date - start_date).days
                else:
                    status_val = "Open"
                    duration = (today - start_date).days

                # Duplicate check (same as your API)
                filters = {
                    "circle": row["circle"],
                    "site_id": row["site_id"],
                    "issue_owner": row["issue_owner"],
                    "milestone": row["milestone"],
                    "issue_name": row["issue_name"],
                    "status": "Open"
                }

                if RelocationIssue.objects.filter(**filters).exists():
                    continue  # skip duplicate open issue

                issues_to_create.append(
                    RelocationIssue(
                        circle = safe_strip(row["circle"]),
                        site_id = safe_strip(row["site_id"]),
                        issue_owner = safe_strip(row["issue_owner"]),
                        milestone = safe_strip(row["milestone"]),
                        issue_name = safe_strip(row["issue_name"]),
                        start_date=start_date,
                        close_date=close_date,
                        status=status_val,
                        duration=duration,
                        remarks= safe_strip(row["remarks"]),
                        created_by=user_email,
                        updated_by=user_email,
                    )
                )

            except Exception:
                print("EXCEPTION----------------------------------------------------------------------------------------")
                continue  # skip bad rows safely

        # ✅ Bulk insert (important for performance)
        RelocationIssue.objects.bulk_create(issues_to_create)

        # # Optional: update ageing per unique site
        # unique_sites = df[["circle", "site_id"]].drop_duplicates()
        for _, row in df.iterrows():
            update_ageing_new(row["circle"], row["site_id"])

        return Response({
            "message": f"{len(issues_to_create)} issues uploaded successfully!"
        }, status=200)

    except Exception as e:
        return Response({"status": False, "error": str(e)}, status=500)

@api_view(["POST"])
def fetch_sites(request):
    circle = request.data.get("circle")
    siteId = request.data.get("siteId")

    try:
        queryset = AlokTrackerModel.objects.all()

        # Filter by circle if provided
        if circle:
            queryset = queryset.filter(circle__iexact=circle)

        # Filter by siteId (partial match) if provided
        if siteId:
            queryset = queryset.filter(new_site_id__icontains=siteId)

        # Get top 10 results (latest or just first 10)
        sites = queryset.values_list("new_site_id", flat=True).distinct()[:10]

        return Response({
            "status": True,
            "data": list(sites),
            "message": "Sites fetched"
        }, status=200)

    except Exception as e:
        return Response({
            "status": False,
            "error": str(e)
        }, status=500)
   

@api_view(["POST"])
def issue_summary(request):
    try:
        # ---------------- 0️⃣ Input Handling ----------------
        status = request.data.get('status', 'ALL')
        milestone = request.data.get('milestone', '')
        owner = request.data.get('owner', '')
        duration_start = request.data.get('duration_start')
        duration_end = request.data.get('duration_end')
        site_on_air = request.data.get('is_on_air')

        # # Ensure list type
        # if isinstance(milestone, str):
        #     milestone = [milestone]
        # if isinstance(owner, str):
        #     owner = [owner]

        # Convert duration safely
        duration_start = int(duration_start) if duration_start not in [None, ""] else None
        duration_end = int(duration_end) if duration_end not in [None, ""] else None

        # ---------------- 1️⃣ Fetch Data ----------------
        qs = RelocationIssue.objects.all().values(
            "circle",
            "issue_name",
            "issue_owner",
            "milestone",
            "site_id",
            "status",
            "duration",   # ✅ added
            "start_date",
            "close_date"
        )

        df = pd.DataFrame(list(qs))

        if df.empty:
            return Response({"message": "No data found"}, status=200)

        # ---------------- 2️⃣ Clean Data ----------------
        df["circle"] = df["circle"].astype(str).str.strip()
        df["issue_name"] = df["issue_name"].astype(str).str.strip()
        df["site_id"] = df["site_id"].astype(str).str.strip()
        
        # ---------------- Site ON AIR Filtering ----------------
        if site_on_air in ["Yes", "No"]:
            qs1 = AlokTrackerModel.objects.all().values(
                "circle", "new_site_id", "site_onair_date"
            )

            df1 = pd.DataFrame(list(qs1))

            if not df1.empty:
                # Clean
                df1["circle"] = df1["circle"].fillna("").astype(str).str.strip()
                df1["new_site_id"] = df1["new_site_id"].fillna("").astype(str).str.strip()

                # Rename for merge
                df1 = df1.rename(columns={"new_site_id": "site_id"})

                # 🔥 Mark ON AIR sites
                df1["is_on_air"] = df1["site_onair_date"].notna()

                # Merge
                df = df.merge(
                    df1[["circle", "site_id", "is_on_air"]],
                    on=["circle", "site_id"],
                    how="left"
                )

                # If site not found in df1 → treat as NOT ON AIR
                df["is_on_air"] = df["is_on_air"].fillna(False)

                # 🎯 Apply filter
                if site_on_air == "No":
                    # Keep only NOT ON AIR
                    df = df[df["is_on_air"] == False]

                elif site_on_air == "Yes":
                    # Keep only ON AIR
                    df = df[df["is_on_air"] == True]

                # Cleanup
                df = df.drop(columns=["is_on_air"])

                if df.empty:
                    return Response({"message": "No data found"}, status=200)
                

        # ---------------- 3️⃣ Filters ----------------
        
        milestone = [c.strip() for c in milestone.split(',')] if milestone else ["ALL"]
        owner = [c.strip() for c in owner.split(',')] if owner else ["ALL"]
        
        if status != 'ALL':
            df = df[df['status'] == status]

        if duration_start is not None:
            df = df[df['duration'] >= duration_start]

        if duration_end is not None:
            df = df[df['duration'] <= duration_end]

        if 'ALL' not in milestone:
            df = df[df['milestone'].isin(milestone)]

        if 'ALL' not in owner:
            df = df[df['issue_owner'].isin(owner)]

        # ---------------- 4️⃣ Deduplicate (IMPORTANT) ----------------
        df = df.drop_duplicates(subset=["circle", "issue_name", "site_id", 'issue_owner', 'milestone'])
        
        all_unique_milestones = list(
            RelocationIssue.objects.exclude(milestone__isnull=True)
            .distinct("milestone")
            .values_list("milestone", flat=True)
        )
        
        all_unique_owners = list(
            RelocationIssue.objects.exclude(issue_owner__isnull=True)
            .distinct("issue_owner")
            .values_list("issue_owner", flat=True)
        )
        
        unique_data = {
            "unique_milestone": sorted(all_unique_milestones),
            "unique_owners": sorted(all_unique_owners),
        }
        

        # ---------------- 5️⃣ Pivot ----------------
        pivot_df = pd.pivot_table(
            df,
            index="issue_name",
            columns="circle",
            values="site_id",
            aggfunc="size",
            fill_value=0
        )

        # ---------------- 6️⃣ Totals ----------------
        # Existing row-wise total (keep this)
        pivot_df["Total"] = pivot_df.sum(axis=1)

        # 🔥 NEW: Unique site count per circle
        circle_unique_sites = (
            df.groupby("circle")["site_id"]
            .nunique()
        )

        # Convert to row format
        total_row = circle_unique_sites.to_dict()

        # Add overall total (unique sites across all circles)
        total_row["Total"] = df["site_id"].nunique()

        # Convert to DataFrame row
        total_row_df = pd.DataFrame([total_row], index=["Total"])

        # Append
        pivot_df = pd.concat([pivot_df, total_row_df])

        # ---------------- 7️⃣ Column Order ----------------
        circle_order = [
            "AP","ASM","BIH","CHN","DEL","HRY","JK","JRK",
            "KK","KOL","MAH","MP","MUM","NE","ORI","PUN",
            "RAJ","ROTN","UPE","UPW","WB"
        ]

        existing_cols = [c for c in circle_order if c in pivot_df.columns]

        pivot_df = pivot_df.reindex(columns=existing_cols + ["Total"], fill_value=0)

        # ---------------- 8️⃣ Sort ----------------
        pivot_df = pivot_df.sort_values(by="Total", ascending=False)

        # ---------------- 9️⃣ Final Format ----------------
        pivot_df = pivot_df.reset_index().rename(columns={"index": "Issue"})

        return Response({
            "message": "Success",
            "data": pivot_df.to_dict(orient="records"),
            "unique_data": unique_data
        }, status=200)

    except Exception as e:
        return Response({
            "message": "Error",
            "error": str(e)
        }, status=500)


############################################################ DOWNLOAD DATA ###################################################################

@api_view(['POST'])
def download_tracker_data_view(request):
    userId = request.data.get('userId')
    year = request.data.get('year')
    
    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    
    circles = user.circles

    # if not userId or not circle:
    #     return Response({'error': 'userId and circle are required.'}, status=400)

    try:
        if year:
            year = int(year)
        
            fy_start = pd.Timestamp(year=year, month=3, day=26).date()
            fy_end   = pd.Timestamp(year=year + 1, month=3, day=25).date()
        
        
            today = dtime.today().date()

            # condition for FY range
            fy_filter = Q(site_onair_date__range=(fy_start, fy_end))
            null_filter = Q(site_onair_date__isnull=True)

            if fy_start <= today <= fy_end:
                final_filter = fy_filter | null_filter
            else:
                final_filter = fy_filter
        
        if 'CENTRAL' in circles:
            if year:
                obj = AlokTrackerModel.objects.filter(final_filter)
            else:
                obj = AlokTrackerModel.objects.all()
            issue_obj = RelocationIssue.objects.all()
        else:
            if year:
                obj = AlokTrackerModel.objects.filter(
                    final_filter,
                    circle__in=circles
                )
            else:
                obj = AlokTrackerModel.objects.filter(
                    circle__in=circles
                )
            issue_obj = RelocationIssue.objects.filter(circle__in=circles)

        df = pd.DataFrame(obj.values())
        
        if df.empty:
            return Response({"message": "No Data Found"}, status=404)
        
        issue_df = pd.DataFrame(issue_obj.values())
        
        rename_map = {
            "circle": "Circle",
            "site_id": "Site ID",
            "issue_owner": "Issue Owner",
            "milestone": "Milestone",
            "issue_name": "Issue Name",
            "start_date": "Start Date",
            "close_date": "Close Date",
            "status": "Status",
            "duration": "Duration",
            "remarks": "Remarks",
            "updated_by": "Updated_by",
            "updated_at": "Updated_at",
            "created_by": "Created_by",
            "created_at": "Created_at"
        }
        
        issue_df = issue_df.rename(columns=rename_map)

        print("1")
        
        required_cols = ["Circle", "Site ID", "Issue Owner", "Milestone", "Issue Name", "Start Date", "Close Date", "Status", "Duration", "Remarks", "Updated_by", "Updated_at", "Created_by", "Created_at"]
        for col in required_cols:
            if col not in issue_df.columns:
                issue_df[col] = None

        # Convert date columns safely
        for col in ["Start Date", "Close Date"]:
            if col in issue_df.columns:
                issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
        
        print("1.1")
        df = update_ageing(df, issue_df)

        print("2")

        for col in df.columns:
            if 'date' in col:
                if col != 'last_updated_date':
                    converted = pd.to_datetime(df[col], errors='coerce')
                    if converted.notna().sum() > 0:
                        df[col] = converted.dt.strftime('%d-%b-%y')
                else:
                    converted = pd.to_datetime(df[col], errors='coerce')
                    if converted.notna().sum() > 0:
                        df[col] = converted.dt.strftime('%d-%b-%y %H:%M:%S')

        print("3")
        
        qs = SiteStatus.objects.all().values(
            "site_id", "date", "status"
        )

        # Create DataFrame
        status_df = pd.DataFrame(qs)
        
        print(status_df)
        
        required_cols = ["site_id", "date", "status"]
        for col in required_cols:
            if col not in status_df.columns:
                status_df[col] = None
        
        print(status_df['date'].dropna().unique().tolist())
        

        

        print("4")
        # Ensure datetime
        status_df["date"] = pd.to_datetime(status_df["date"])
        
        print("5")

        # Pivot to required structure
        status_df = status_df.pivot_table(
            index="site_id",
            columns="date",
            values="status",
            aggfunc="last"
        )

        # Sort columns chronologically
        status_df = status_df.sort_index(axis=1)

        # Format date headers
        status_df.columns = status_df.columns.strftime("%d-%b-%y")

        # Rename index for display
        status_df.index.name = "Site ID"

        # 🔥 REMOVE JUNK ROW
        status_df = status_df[status_df.index != "Site ID"]
        
        print("Index values:", list(status_df.index))

        
        print(status_df)


        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)

        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)

        tracker_file_path = os.path.join(output_folder, f"TRACKER_FILE_{current_date}_{current_time}.xlsx")
        df.insert(0, "Unique ID", range(1, len(df) + 1))

        df.drop(columns=['id'], inplace=True)

        df = df[['Unique ID', 'circle', 'site_tagging', 'old_toco_name', 'old_site_id', 'new_site_id', 'new_toco_name',
                 'sr_number', 'ran_oem', 'media_type', 'mw_oem', 'relocation_method', 'relocation_type', 'old_site_band',
                 'new_site_band', 'rfai_date', 'allocation_date', 'rfai_survey_date', 'mo_punch_date',
                 'material_dispatch_date', 'material_delivered_date', 'installation_start_date', 'installation_end_date',
                 'integration_date', 'emf_submission_date', 'ran_lkf_status', 'alarm_status', 'alarm_rectification_done_date',
                 'scft_done_date', 'scft_i_deploy_offered_date', 'ran_pat_offer_date', 'ran_sat_offer_date', 'mw_plan_id',
                 'mw_pat_offer_date', 'rsl_value_status', 'enm_status', 'mw_lkf', 'mw_sat_offer_date', 'mw_ms1_mids_date',
                 'site_onair_date', 'i_deploy_onair_date', 'current_status', 'detailed_remarks', 'rfai_rejected_date', 
                 'clear_rfai_date', 'pri_count', 'pri_issue_ageing', 'other_ust_issue_ageing', 'other_airtel_issue_ageing', 'total_issue_ageing', 
                 'clear_rfai_to_ms1_ageing', 'rfai_to_ms1_ageing', 'ran_pat_accepted_date', 'ran_sat_accepted_date', 
                 'mw_pat_accepted_date', 'mw_sat_accepted_date', 'scft_accepted_date', 'kpi_at_offer_date', 'kpi_at_accepted_date',
                 'four_g_ms2_date', 'five_g_ms2_date', 'final_ms2_date', "dismantling_survey_date", "sreq_creq_raised_date",
                 "dismantle_date", "material_pickup_date", "material_submission_date", "oci_done_date", "sign_off_date", "dismantling_status", 
                 'last_updated_date', 'last_updated_by']]

        template_path = os.path.join(BASE_URL, "template", "templateAlok_v.1.xlsx")
        wb = load_workbook(template_path)
        ws = wb["Sheet1"]   # or wb.create_sheet("Tracker")
        ws.title = "Main Tracker"

        date_columns = [col for col in df.columns if col.endswith("_date") and col != 'last_updated_date']

        for col in date_columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        start_row = 3
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                # apply formatting for date columns
                if df.columns[c_idx - 1].endswith("_date") and hasattr(value, "strftime") and df.columns[c_idx - 1] != 'last_updated_date':
                    cell.number_format = "dd-mmm-yy"



        print("6")
        
        ws2 = wb.create_sheet(title="Issues Tracker")

        # Clean issue_df (remove id if present)
        if "id" in issue_df.columns:
            issue_df = issue_df.drop(columns=["id"])
            
        # print(issue_df['updated_by'])

        def remove_tz_safe(x):
            if isinstance(x, datetime) and x.tzinfo is not None:
                return x.replace(tzinfo=None)
            return x

        # apply to every cell in issue_df
        issue_df = issue_df.applymap(remove_tz_safe)
        
        # print(issue_df['updated_by'])


        # Convert date columns
        issue_date_columns = [col for col in issue_df.columns if "Date" in col]
        for col in issue_date_columns:
            issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
            
        # print(issue_df['updated_by'])

        # ---- Write header ----
        for col_idx, column_name in enumerate(issue_df.columns, start=1):
            ws2.cell(row=1, column=col_idx, value=column_name)
            
        # print(issue_df['updated_by'])

        # ---- Write data ----
        for r_idx, row in enumerate(dataframe_to_rows(issue_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws2.cell(row=r_idx, column=c_idx)
                cell.value = value

                # Apply date formatting
                if issue_df.columns[c_idx - 1] in issue_date_columns and hasattr(value, "strftime"):
                    cell.number_format = "dd-mmm-yy"
                    
        # print(issue_df['updated_by'])

        ws3 = wb.create_sheet(title="Site Status Tracker")

        # Remove index name to avoid Excel phantom row
        status_df.index.name = None

        # Write header
        headers = ["Site ID"] + list(status_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            ws3.cell(row=1, column=col_idx, value=header)

        # Write data
        for r_idx, row in enumerate(
            dataframe_to_rows(status_df, index=True, header=False),
            start=2
        ):
            for c_idx, value in enumerate(row, start=1):
                if value=='' and c_idx==0:
                    print("----------------------------------------------------------------------------------")
                ws3.cell(row=r_idx, column=c_idx, value=value)

        wb.save(tracker_file_path)

        relative_path = tracker_file_path.replace(settings.MEDIA_ROOT, '').lstrip(os.sep)
        download_url = request.build_absolute_uri(
            os.path.join(settings.MEDIA_URL, relative_path).replace('\\', '/')
        )

        return Response({'message': f'Welcome {userId}! Credentials verified.', "download_link": download_url}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)


@api_view(["POST"])
def upload_site_status_table(request):
    """
    Upload Excel file with pivoted Site Status table
    and populate SiteStatus model.
    """

    file = request.data.get("file")
    if not file:
        return Response({"error": "No file uploaded"}, status=400)

    df = pd.read_excel(file)

    if df.empty:
        return Response({"error": "Uploaded file is empty"}, status=400)

    # 🔹 Clean column names but keep datetimes intact
    df.columns = [
        col.strip() if isinstance(col, str) else col
        for col in df.columns
    ]

    site_col = df.columns[0]  # first column = Site ID

    records = []
    seen = set()  # 🔥 prevents duplicates (site_id, date)

    def parse_date(col):
        """
        Robust date parser for Excel headers
        """
        if isinstance(col, pd.Timestamp):
            return col.normalize().date()

        try:
            return pd.to_datetime(col).date()
        except Exception:
            return None

    for _, row in df.iterrows():
        site_id = row[site_col]

        if pd.isna(site_id):
            continue

        site_id = str(site_id).strip()

        for col in df.columns[1:]:
            status = row[col]

            if pd.isna(status):
                continue

            date_obj = parse_date(col)
            if not date_obj:
                continue

            print(date_obj)

            records.append(
                SiteStatus(
                    site_id=site_id,
                    date=date_obj,
                    status=str(status).strip()
                )
            )

    if not records:
        return Response(
            {"error": "No valid records found in file"},
            status=400
        )
        
    print(records)

    # Optional overwrite logic (safe)
    # SiteStatus.objects.filter(
    #     site_id__in=[r.site_id for r in records],
    #     date__in=[r.date for r in records]
    # ).delete()

    SiteStatus.objects.bulk_create(records)

    return Response({
        "message": "Site status uploaded successfully",
        "records_created": len(records)
    })


@api_view(['DELETE'])
def delete_dropped_sites_status(request):

    qs = SiteStatus.objects.all()

    if not qs.exists():
        return Response(
            {"message": "No records found"},
            status=status.HTTP_404_NOT_FOUND
        )

    deleted_count, _ = qs.delete()

    return Response(
        {"message": f"{deleted_count} records deleted successfully"},
        status=status.HTTP_200_OK
    )
    

@api_view(["POST"])
def sync_alok_tracker_to_site_status(request):
    """
    Fetch data from AlokTrackerModel and
    populate SiteStatus with today's date.
    """

    today = date.today()-timedelta(days=1)

    qs = AlokTrackerModel.objects.values(
        "new_site_id",
        "current_status"
    )

    if not qs.exists():
        return Response(
            {"error": "No data found in AlokTrackerModel"},
            status=400
        )

    records = []
    seen = set()  # prevent duplicate site_id for today

    for row in qs:
        site_id = row["new_site_id"]
        status = row["current_status"]

        if not site_id or not status:
            continue

        site_id = str(site_id).strip()
        status = str(status).strip()

        key = (site_id, today)
        if key in seen:
            continue

        seen.add(key)

        records.append(
            SiteStatus(
                site_id=site_id,
                status=status,
                date=today
            )
        )

    if not records:
        return Response(
            {"error": "No valid records to insert"},
            status=400
        )

    # Optional overwrite logic for today
    # SiteStatus.objects.filter(date=today).delete()

    SiteStatus.objects.bulk_create(records)

    return Response({
        "message": "AlokTracker data synced successfully",
        "records_created": len(records),
        "date": today
    })


############################################################# DASHBOARD ##########################################################################

# @api_view(['GET', 'POST'])
# def daily_dashboard_view(request):
#     circle = request.data.get('circle', [])
#     site_tagging = request.data.get('site_tagging', [])
#     current_status = request.data.get('relocation_method', [])
#     new_toco_name = request.data.get('new_toco_name', [])
#     start_date = request.data.get('from_date')
#     end_date = request.data.get('to_date')
#     view = request.data.get('view')

#     all_unique_circles = list(
#         AlokTrackerModel.objects.exclude(circle__isnull=True)
#         .distinct("circle")
#         .values_list("circle", flat=True)
#     )

#     all_unique_site_tagging = list(
#         AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
#         .distinct("site_tagging")
#         .values_list("site_tagging", flat=True)
#     )

#     all_unique_current_status = list(
#         AlokTrackerModel.objects.exclude(current_status__isnull=True)
#         .distinct("current_status")
#         .values_list("current_status", flat=True)
#     )

#     all_unique_new_toco_name = list(
#         AlokTrackerModel.objects.exclude(new_toco_name__isnull=True)
#         .distinct("new_toco_name")
#         .values_list("new_toco_name", flat=True)
#     )

#     unique_data = {
#         "unique_circle": sorted(all_unique_circles),
#         "unique_site_tagging": sorted(all_unique_site_tagging),
#         "unique_relocation_method": sorted(all_unique_current_status),
#         "unique_new_toco_name": sorted(all_unique_new_toco_name),
#     }

#     circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
#     site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
#     current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
#     new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]

#     start_date = dtime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
#     end_date = dtime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

#     try:
#         filters = {}
#         if "ALL" not in circle:
#             filters["circle__in"] = circle
#         if "ALL" not in site_tagging:
#             filters["site_tagging__in"] = site_tagging
#         if "ALL" not in current_status:
#             filters["current_status__in"] = current_status
#         if "ALL" not in new_toco_name:
#             filters["new_toco_name__in"] = new_toco_name

#         obj = AlokTrackerModel.objects.filter(**filters)
#         df = pd.DataFrame(obj.values())

#         if df.empty:
#             return Response({'error': 'No data found for given filters'}, status=404)
        
#         # df['rfai_date'] = df['clear_rfai_date'].where(df["clear_rfai_date"].notna(), df['rfai_date'])

#         for col in df.columns:
#             if "Date" in col:
#                 df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")

#         today = dtime.today().date()

#         if not start_date or not end_date:
#             if today.day >= 26:
#                 start_date = today.replace(day=26)
#                 if today.month == 12:
#                     end_date = today.replace(year=today.year + 1, month=1, day=25)
#                 else:
#                     end_date = today.replace(month=today.month + 1, day=25)
#             else:
#                 if today.month == 1:
#                     start_date = today.replace(year=today.year - 1, month=12, day=26)
#                 else:
#                     start_date = today.replace(month=today.month - 1, day=26)
#                 end_date = today.replace(day=25)

#         date_range = pd.date_range(start=start_date, end=min(end_date, today)).date
#         formatted_dates = [d.strftime("%d-%b-%y") for d in date_range]

#         result_columns = ["Milestone Track/Site Count", "CF"] + formatted_dates
#         result = pd.DataFrame(columns=result_columns)

#         milestones = [
#             "Allocation Date",
#             "RFAI Date",
#             "RFAI Survey Date",
#             "MO Punch Date",
#             "Material Dispatch Date",
#             "Material Delivered Date",
#             "Installation End Date",
#             "Integration Date",
#             "EMF Submission Date",
#             "Alarm Rectification Done Date",
#             "SCFT I-Deploy Offered Date",
#             "RAN PAT Offer Date",
#             "RAN SAT Offer Date",
#             "MW PAT Offer Date",
#             "MW SAT Offer Date",
#             "Site ONAIR Date",
#             "I-Deploy ONAIR Date",
#         ]

#         unique_data.update(**{"Milestone": milestones})

#         for milestone in milestones:
#             milestone_df_format = (
#                 milestone.lower()
#                 .replace(" ", "_")
#                 .replace("-", "_")
#                 .replace("(", "")
#                 .replace(")", "")
#                 .replace("/", "_")
#             )

#             if milestone_df_format not in df.columns:
#                 continue

#             df[milestone_df_format] = pd.to_datetime(df[milestone_df_format], errors="coerce").dt.date
#             valid_dates = df[milestone_df_format].dropna()

#             if valid_dates.empty:
#                 row = {
#                     "Milestone Track/Site Count": milestone,
#                     "CF": "-",
#                     **{d.strftime("%d-%b-%y"): "-" for d in date_range}
#                 }
#                 result.loc[len(result)] = row
#                 continue

#             cf_count = (valid_dates < start_date).sum()
#             cumulative = cf_count
#             row = {"Milestone Track/Site Count": milestone, "CF": cf_count}

#             for date in date_range:
#                 count = (valid_dates == date).sum()
#                 cumulative += count

#                 if view == "Cumulative":
#                     row[date.strftime("%d-%b-%y")] = cumulative
#                 else:
#                     row[date.strftime("%d-%b-%y")] = count

#             result.loc[len(result)] = row

#         result.columns = [
#             col.strftime("%d-%b-%y") if isinstance(col, (dtime,)) or hasattr(col, "strftime") else col
#             for col in result.columns
#         ]

#         result = result[["Milestone Track/Site Count", "CF"] + formatted_dates]

#         last_col = formatted_dates[-1]
#         dash_mask = result[last_col] == '-'

#         result[last_col] = pd.to_numeric(result[last_col], errors='coerce')
#         result['Gap'] = -result[last_col].diff()

#         nan_mask = result['Gap'].isna()

#         result[last_col] = result[last_col].fillna(0).astype(int)
#         result['Gap'] = result['Gap'].fillna(0).astype(int)

#         result.loc[dash_mask, last_col] = '-'
#         result.loc[nan_mask, 'Gap'] = '-'

#         result['Gap'] = result['Gap'].astype(str)
#         result = result.astype(str).reset_index(drop=True)

#         result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(
#             lambda col: col.replace(" Date", "") if " Date" in col else col
#         )

#         new_result = result.copy()

#         for i, date in enumerate(formatted_dates, start=1):
#             new_result.rename(columns={date: f'date_{i}'}, inplace=True)

#         result_json = new_result.to_dict(orient="records")
#         json_data = json.dumps(result_json)

#         current_date = dtime.now().strftime("%Y-%m-%d")
#         current_time = dtime.now().strftime("%H-%M-%S")

#         BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
#         os.makedirs(BASE_URL, exist_ok=True)

#         output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
#         shutil.rmtree(output_folder, ignore_errors=True)
#         os.makedirs(output_folder, exist_ok=True)

#         dashboard_file_path = os.path.join(
#             output_folder,
#             f"DAILY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx"
#         )

#         with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
#             result.to_excel(writer, index=False, sheet_name='Daily Waterfall MS1')

#         dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
#         download_link = request.build_absolute_uri(dashboard_file_path)

#         return Response({
#             'message': 'Dashboard created successfully !!!',
#             "download_link": download_link,
#             "data": json_data,
#             "dates": formatted_dates,
#             "unique_data": unique_data
#         }, status=200)

#     except Exception as e:
#         return Response({"error": f"{str(e)}"}, status=500)


@api_view(['GET', 'POST'])
def daily_dashboard_view(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    start_date = request.data.get('from_date')
    end_date = request.data.get('to_date')
    view = request.data.get('view')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None

    all_unique_circles = list(
        AlokTrackerModel.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        AlokTrackerModel.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )

    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]

    start_date = dtime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    end_date = dtime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

    try:
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name

        obj = AlokTrackerModel.objects.filter(**filters)
        df = pd.DataFrame(obj.values())

        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        # df['rfai_date'] = df['re_rfai_date'].where(df["re_rfai_date"].notna(), df['rfai_date'])

        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")

        today = dtime.today().date()

        if (not month_start or not month_end) and (not start_date or not end_date):
            if today.day >= 26:
                start_date = today.replace(day=26)
                if today.month == 12:
                    end_date = today.replace(year=today.year + 1, month=1, day=25)
                else:
                    end_date = today.replace(month=today.month + 1, day=25)
            else:
                if today.month == 1:
                    start_date = today.replace(year=today.year - 1, month=12, day=26)
                else:
                    start_date = today.replace(month=today.month - 1, day=26)
                end_date = today.replace(day=25)
                
        if month_start and month_end:
            date_range = pd.date_range(start=month_start, end=min(month_end, today)).date
        else:
            date_range = pd.date_range(start=start_date, end=min(end_date, today)).date

        formatted_dates = [d.strftime("%d-%b-%y") for d in date_range]

        if month_start and month_end:
            result_columns = ["Milestone Track/Site Count", "AOP", "CF"] + formatted_dates
        else:
            result_columns = ["Milestone Track/Site Count", "AOP"] + formatted_dates
        result = pd.DataFrame(columns=result_columns)

        milestones = [
            "Allocation Date",
            "RFAI Date",
            "RFAI Survey Date",
            "Workable Sites",
            "MO Punch Date",
            "Material Dispatch Date",
            "Material Delivered Date",
            "Installation End Date",
            "Integration Date",
            "EMF Submission Date",
            "Alarm Rectification Done Date",
            "SCFT I-Deploy Offered Date",
            "RAN PAT Offer Date",
            "RAN SAT Offer Date",
            "MW PAT Offer Date",
            "MW SAT Offer Date",
            "Site ONAIR Date",
            "I-Deploy ONAIR Date",
        ]

        unique_data.update(**{"Milestone": milestones})
        
        included_statuses = [
            'MO WIP',
            'On Air Done',
            'Onair WIP'
        ]

        for milestone in milestones:
            if milestone == 'Workable Sites':
                if month_start and month_end:
                    # CF → before month_start
                    aop_count = SiteStatus.objects.filter(date__lt=month_start) \
                        .filter(status__in=included_statuses) \
                        .values('site_id').distinct().count()

                    # AOP → before month_start (same logic here)
                    cf_count = SiteStatus.objects.filter(date__lt=month_start) \
                        .filter(status__in=['Onair WIP', 'MO WIP']) \
                        .values('site_id').distinct().count()
                    
                    prev_date = month_start - timedelta(days=1)
                    
                    
                    
                    prevCount = SiteStatus.objects.filter(date=prev_date) \
                                .filter(status__in=included_statuses) \
                                .values('site_id').distinct().count()
                            

                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": aop_count,
                        "CF": cf_count
                    }
                else:
                    aop_count = SiteStatus.objects.filter(date__lt=start_date) \
                        .filter(status__in=included_statuses) \
                        .values('site_id').distinct().count()
                        
                    prev_date = start_date - timedelta(days=1)
                        
                    prevCount = SiteStatus.objects.filter(date=prev_date) \
                                .filter(status__in=included_statuses) \
                                .values('site_id').distinct().count()

                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": aop_count
                    }

                for d in date_range:
                    if d == today:
                        row[d.strftime("%d-%b-%y")] = '-'
                        continue
                    count = SiteStatus.objects.filter(date=d) \
                        .filter(status__in=included_statuses) \
                        .values('site_id').distinct().count()

                    if view == "Cumulative":
                        row[d.strftime("%d-%b-%y")] = count
                    else:
                        row[d.strftime("%d-%b-%y")] = count-prevCount
                    
                    prevCount = count

                result.loc[len(result)] = row
                continue
                
            milestone_df_format = (
                milestone.lower()
                .replace(" ", "_")
                .replace("-", "_")
                .replace("(", "")
                .replace(")", "")
            )

            if milestone_df_format not in df.columns:
                continue

            df[milestone_df_format] = pd.to_datetime(df[milestone_df_format], errors="coerce").dt.date
            if month_start and month_end:
                mask = df['site_onair_date'].isna()
                valid_dates_cf = df.loc[mask, milestone_df_format].dropna()
            # else:
            #     valid_dates = df[milestone_df_format].dropna()
            
            valid_dates_aop = df[milestone_df_format].dropna()

            if valid_dates_aop.empty:
                if month_start and month_end:
                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": "-",
                        "CF": "-",
                        **{d.strftime("%d-%b-%y"): "-" for d in date_range}
                    }
                else:
                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": "-",
                        **{d.strftime("%d-%b-%y"): "-" for d in date_range}
                    }
                
                result.loc[len(result)] = row
                continue
            
            if month_start and month_end:
                cf_count = (valid_dates_cf < month_start).sum()
                aop_count = (valid_dates_aop < month_start).sum()
            else:
            #     cf_count = (valid_dates < start_date).sum()
                aop_count = (valid_dates_aop < start_date).sum()
            
            if month_start and month_end:
                cumulative = cf_count
                row = {"Milestone Track/Site Count": milestone,"AOP": aop_count, "CF": cf_count}
            else:
                cumulative = aop_count
                row = {"Milestone Track/Site Count": milestone,"AOP": aop_count}


            for d in date_range:
                count = (valid_dates_aop == d).sum()
                cumulative += count

                if view == "Cumulative":
                    row[d.strftime("%d-%b-%y")] = cumulative
                else:
                    row[d.strftime("%d-%b-%y")] = count

            result.loc[len(result)] = row

        result.columns = [
            col.strftime("%d-%b-%y") if isinstance(col, (dtime,)) or hasattr(col, "strftime") else col
            for col in result.columns
        ]
        if month_start and month_end:
            result = result[["Milestone Track/Site Count", "AOP" , "CF"] + formatted_dates]
        else:
            result = result[["Milestone Track/Site Count", "AOP"] + formatted_dates]
 
        last_col = formatted_dates[-1]
        dash_mask = result[last_col] == '-'

        result[last_col] = pd.to_numeric(result[last_col], errors='coerce')
        result['Gap'] = -result[last_col].diff()

        nan_mask = result['Gap'].isna()

        result[last_col] = result[last_col].fillna(0).astype(int)
        result['Gap'] = result['Gap'].fillna(0).astype(int)

        result.loc[dash_mask, last_col] = '-'
        result.loc[nan_mask, 'Gap'] = '-'

        result['Gap'] = result['Gap'].astype(str)
        result = result.astype(str).reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(
            lambda col: col.replace(" Date", "") if " Date" in col else col
        )

        new_result = result.copy()

        for i, d in enumerate(formatted_dates, start=1):
            new_result.rename(columns={d: f'date_{i}'}, inplace=True)

        result_json = new_result.to_dict(orient="records")
        json_data = json.dumps(result_json)

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)

        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)

        dashboard_file_path = os.path.join(
            output_folder,
            f"RELOCATION_TRACKING_MS1_DAILY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx"
        )

        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Daily Waterfall MS1')

        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)

        return Response({
            'message': 'Dashboard created successfully !!!',
            "download_link": download_link,
            "data": json_data,
            "dates": formatted_dates,
            "unique_data": unique_data
        }, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)


@api_view(['GET', 'POST'])
def weekly_monthly_dashboard_view(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    view = request.data.get('view')
    year2 = request.data.get('year2')
 
    # Default to 'ALL' if not provided
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    
    if year2:
        year2 = int(year2)
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None
   
    all_unique_circles = list(
        AlokTrackerModel.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        AlokTrackerModel.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )

    # ✅ Add "ALL" at the top of each list
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    try:
        # 🔹 Dynamic filters
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name
            
        # 🔹 Fetch data
        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        # df['rfai_date'] = df['re_rfai_date'].where(df["re_rfai_date"].notna(), df['rfai_date'])
       
        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")
                
               
        today = dtime.today().date()
        
        if year2:
            fy_start = dtime(year2, 3, 26).date()   # 26-Apr current year start
            fy_end = dtime(year2 + 1, 3, 25).date() # 25-Mar next year end
        else:
            if today.month >= 4:
                fy_start = dtime(today.year, 3, 26).date()   # 26-Apr current year start
                fy_end = dtime(today.year + 1, 3, 25).date() # 25-Mar next year end
            else:
                fy_start = dtime(today.year - 1, 3, 26).date()
                fy_end = dtime(today.year, 3, 25).date()
                
        print("2.5-------------------------------------------------------------")
 
        # 🧩 2. Determine current cycle (26th prev → 25th curr)
        if today.day >= 26:
            if today.month == 12:
                current_cycle_start = dtime(today.year, 12, 26).date()
                current_cycle_end = dtime(today.year + 1, 1, 25).date()
            else:
                current_cycle_start = dtime(today.year, today.month, 26).date()
                current_cycle_end = dtime(today.year, today.month + 1, 25).date()
        else:
            if today.month == 1:
                current_cycle_start = dtime(today.year - 1, 12, 26).date()
            else:
                current_cycle_start = dtime(today.year, today.month - 1, 26).date()
            current_cycle_end = dtime(today.year, today.month, 25).date()
            
        if year2:
            current_cycle_start = min(dtime(year2+1, 2, 26).date(), current_cycle_start)
            current_cycle_end = min(dtime(year2+1, 3, 25).date(), current_cycle_end)
            
        print(fy_start)
        print(fy_end)    
        print(current_cycle_start)
        print(current_cycle_end)
 
        # 🗓️ 3. Create monthly periods (Apr → month before current)
        months = []
        start = fy_start
        while start <= current_cycle_start:
            if start.month == 12:
                end = dtime(start.year + 1, 1, 25).date()
            else:
                end = dtime(start.year, start.month + 1, 25).date()
            months.append((start, end))
            start = end + timedelta(days=1)
            
        print(months)
 
        # 🗓️ 4. Create weekly periods for current month
        weeks = []
        
        if month_start and month_end:
            current_cycle_start = month_start
            current_cycle_end = month_end
        
        week_start = current_cycle_start
 
        while week_start <= current_cycle_end:
            week_end = week_start + timedelta(days=6)
            if week_end > current_cycle_end:
                week_end = current_cycle_end
            weeks.append((week_start, week_end))
            week_start = week_end + timedelta(days=1)
 
 
        # 📊 5. Prepare result DataFrame
        result = pd.DataFrame()
       
       
        # result.set_index("Milestone Track/Site Count", inplace=True)
 
        milestones = [
            "Allocation Date",
            "RFAI Date",
            "RFAI Survey Date",
            # "RFAI Survey Done Date",
            "MO Punch Date",
            "Material Dispatch Date",
            "Material Delivered Date",
            # "Installation Start Date",
            "Installation End Date",
            "Integration Date",
            "EMF Submission Date",
            "Alarm Rectification Done Date",
            # "SCFT Done Date",
            "SCFT I-Deploy Offered Date",
            "RAN PAT Offer Date",
            "RAN SAT Offer Date",
            "MW PAT Offer Date",
            "MW SAT Offer Date",
            # "MW MS1 Date (MIDS)",
            "Site ONAIR Date",
            "I-Deploy ONAIR Date",
        ]
 
        data = []
        
        print("5-----------------------------------")
        
        for milestone in milestones:
            
            milestone_df_format = milestone.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_")

            if milestone_df_format not in df.columns:
                continue
            
            milestone_data = pd.to_datetime(df[milestone_df_format], errors='coerce').dt.date
            row = {"Milestone Track/Site Count": milestone}
            if milestone_data.dropna().empty:
                row["CF"] = "-"
                # Set all month and week columns to "-"
                for _, end in months:
                    month_name = end.strftime("%b-%y")
                    row[month_name] = "-"
                for i, _ in enumerate(weeks, 1):
                    week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                    row[week_name] = "-"
                data.append(row)
                continue
 
 
            # CF → sites before financial year start
            row["CF"] = (milestone_data < fy_start).sum()
 
            cumulative_month = row["CF"]
            
            for start, end in months:
                month_name = end.strftime("%b-%y")
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_month += count
                if view == 'Cumulative':
                    row[month_name] = cumulative_month
                else:
                    row[month_name] = count
 
            cumulative_week = 0
            for i, (start, end) in enumerate(weeks, 1):
                week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                if(start >= today):
                    row[week_name] = 0
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_week += count
                if view == 'Cumulative':
                    row[week_name] = cumulative_week
                else:
                    row[week_name] = count
 
            data.append(row)
 

        result = pd.DataFrame(data)
 
        for col in result.columns:
            if col != "Milestone Track/Site Count":
                result[col] = result[col].astype(str)
 
        result = result.reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(lambda col: col.replace(" Date", "") if " Date" in col else col)
 
        new_data = result.copy()
        months_columns = [col for col in new_data.columns.to_list()[2:] if " W" not in col]
        months_data = new_data[new_data.columns.to_list()[:2] + months_columns].copy()
        
        weeks_columns = [col for col in new_data.columns.to_list()[2:] if " W" in col]
        
        week_data = new_data[new_data.columns.to_list()[:1] + weeks_columns].copy()
        
        months_data.rename(columns={
            col : f"Month-{i}" for i, col in enumerate(months_columns, start=1)
        }, inplace=True)
        
        week_data.rename(columns={
            col : f"Month_Week-{i}" for i, col in enumerate(weeks_columns, start=1)
        }, inplace=True)
        
        unique_data.update(**{"month_columns": months_columns, "week_columns" : weeks_columns})
        

        month_dict_data = months_data.to_dict(orient="records")
        month_json_data = json.dumps(month_dict_data)
        week_dict_data = week_data.to_dict(orient="records")
        week_json_data = json.dumps(week_dict_data)
        
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"WEEKLY_MONTHY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx")
       
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Monthly MS1')
       
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
   
        return Response({'message': 'Weekly and Monthly Dashboard created successfully !!!', "download_link": download_link, "unique_data": unique_data, "months_data": month_json_data, "week_data": week_json_data}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
 

@api_view(['GET', 'POST'])
def gap_view(request):
    
    userId = request.data.get('userId')
    
    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    
    circles = user.circles
    
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    last_date = request.data.get('last_date')
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    gap = request.data.get('gap')
    view = request.data.get('view')
    
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None
    
    
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]

    last_date = dtime.strptime(last_date, "%Y-%m-%d").date() if last_date else None
    milestone1 = milestone1.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_")
    milestone2 = milestone2.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_")

    try:
        
        today = dtime.today().date()
 
        if not last_date:
            if today.day >= 26:
                if today.month == 12:
                    last_date = today.replace(year=today.year + 1, month=1, day=25)
                else:
                    last_date = today.replace(month=today.month + 1, day=25)
            else:
                last_date = today.replace(day=25)
                
            last_date = min(today - timedelta(days=1), last_date)
                
        if 'CENTRAL' not in circles:
            circle = [c for c in circle if c in circles]
                
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name
        
        if int(gap)>0 :
            if view == 'Cumulative':
                filters[f"{milestone1}__lte"] = last_date
            else:
                filters[milestone1] = last_date
        else:
            if view == 'Cumulative':
                filters[f"{milestone2}__lte"] = last_date
            else:
                filters[milestone2] = last_date

        # if f"{milestone1}__lte" in filters:
        #     del filters[f"{milestone1}__lte"]
 
        # if f"{milestone2}__lte" in filters:
        #     del filters[f"{milestone2}__lte"]
 
        # 🔹 Fetch data
        obj1 = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df1 = pd.DataFrame(obj1.values())
        
        if month_start and month_end:
            df1 = df1[df1['site_onair_date'].isna() | (df1['site_onair_date'] >= month_start)]
        
        if int(gap)>0 :
            if view == 'Cumulative':
                filters[f"{milestone2}__lte"] = last_date
            else:
                filters[milestone2] = last_date
        else:
            if view == 'Cumulative':
                filters[f"{milestone1}__lte"] = last_date
            else:
                filters[milestone1] = last_date
        
        obj2 = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df2 = pd.DataFrame(obj2.values())
        
        if month_start and month_end:
            df2 = df2[df2['site_onair_date'].isna() | (df2['site_onair_date'] >= month_start)]
        
        df1['key'] = df1['circle'].astype(str) + "_" + df1['new_site_id'].astype(str)
        df2['key'] = df2['circle'].astype(str) + "_" + df2['new_site_id'].astype(str)

        # df = df1[~df1['key'].isin(df2['key'])].drop(columns=['key'])
        # if len(df1) > len(df2):  
        df = df1[~df1['key'].isin(df2['key'])].drop(columns=['key'])
        # else:
        #     df = df2[~df2['key'].isin(df1['key'])].drop(columns=['key'])
        
        sites = df['new_site_id'].dropna().unique().tolist()

        
        # filters = {
        #     "site_id__in" : sites
        # }
        
        issue_obj = RelocationIssue.objects.filter(site_id__in=sites)
        issue_df = pd.DataFrame(issue_obj.values())

        rename_map = {
            "circle": "Circle",
            "site_id": "Site ID",
            "issue_owner": "Issue Owner",
            "milestone": "Milestone",
            "issue_name": "Issue Name",
            "start_date": "Start Date",
            "close_date": "Close Date",
            "status": "Status",
            "duration": "Duration",
            "remarks": "Remarks",
            "updated_by": "Updated_by",
            "updated_at": "Updated_at",
            "created_by": "Created_by",
            "created_at": "Created_at"
        }

        print(issue_df)
        
        issue_df = issue_df.rename(columns=rename_map)

        required_cols = ["Circle", "Site ID", "Issue Owner", "Milestone", "Issue Name", "Start Date", "Close Date", "Status", "Duration", "Remarks", "Updated_by", "Updated_at", "Created_by", "Created_at"]
        for col in required_cols:
            if col not in issue_df.columns:
                issue_df[col] = None

        # # Convert date columns safely
        for col in ["Start Date", "Close Date"]:
            if col in issue_df.columns:
                issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
        
        df = update_ageing(df, issue_df)
 
        
        for col in df.columns:
            if 'date' in col:
                if col != 'last_updated_date':
                    converted = pd.to_datetime(df[col], errors='coerce')
    
                    if converted.notna().sum() > 0:
                        df[col] = converted.dt.strftime('%d-%b-%y')
                else:
                    converted = pd.to_datetime(df[col], errors='coerce')
                    if converted.notna().sum() > 0:
                        df[col] = converted.dt.strftime('%d-%b-%y %H:%M:%S')
 
 
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        tracker_file_path = os.path.join(output_folder, f"TRACKER_GAP_FILE_{milestone1}_{milestone2}_{circle}_{site_tagging}_{current_status}_{new_toco_name}_{current_date}_{current_time}.xlsx")
        
        df.insert(0, "Unique ID", range(1, len(df) + 1))
        
       
        df.drop(columns=['id'], inplace=True)

        df = df[['Unique ID', 'circle', 'site_tagging', 'old_toco_name', 'old_site_id', 'new_site_id', 'new_toco_name',
                 'sr_number', 'ran_oem', 'media_type', 'mw_oem', 'relocation_method', 'relocation_type', 'old_site_band',
                 'new_site_band', 'rfai_date', 'allocation_date', 'rfai_survey_date', 'mo_punch_date',
                 'material_dispatch_date', 'material_delivered_date', 'installation_start_date', 'installation_end_date',
                 'integration_date', 'emf_submission_date', 'ran_lkf_status', 'alarm_status', 'alarm_rectification_done_date',
                 'scft_done_date', 'scft_i_deploy_offered_date', 'ran_pat_offer_date', 'ran_sat_offer_date', 'mw_plan_id',
                 'mw_pat_offer_date', 'rsl_value_status', 'enm_status', 'mw_lkf', 'mw_sat_offer_date', 'mw_ms1_mids_date',
                 'site_onair_date', 'i_deploy_onair_date', 'current_status', 'detailed_remarks', 'rfai_rejected_date', 
                 'clear_rfai_date', 'pri_count', 'pri_issue_ageing', 'other_ust_issue_ageing', 'other_airtel_issue_ageing', 'total_issue_ageing', 
                 'clear_rfai_to_ms1_ageing', 'rfai_to_ms1_ageing', 'ran_pat_accepted_date', 'ran_sat_accepted_date', 
                 'mw_pat_accepted_date', 'mw_sat_accepted_date', 'scft_accepted_date', 'kpi_at_offer_date', 'kpi_at_accepted_date',
                 'four_g_ms2_date', 'five_g_ms2_date', 'final_ms2_date', "dismantling_survey_date", "sreq_creq_raised_date",
                 "dismantle_date", "material_pickup_date", "material_submission_date", "oci_done_date", "sign_off_date", "dismantling_status", 
                 'last_updated_date', 'last_updated_by']]
 
 
        # json_dict_data = df.to_dict(orient="records")
        # json_val = json.dumps(json_dict_data)
        
        template_path = os.path.join(BASE_URL, "template", "templateAlok_v.1.xlsx")
        wb = load_workbook(template_path)
        ws = wb.active  
        date_columns = [col for col in df.columns if col.endswith("_date") and col != 'last_updated_date']

        for col in date_columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        start_row = 3
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                # apply date format for *_date columns
                if df.columns[c_idx - 1].endswith("_date") and hasattr(value, "strftime") and df.columns[c_idx - 1] != 'last_updated_date':
                    cell.number_format = "dd-mmm-yy"
 

        ws2 = wb.create_sheet(title="Issues Tracker")

        # Clean issue_df (remove id if present)
        if "id" in issue_df.columns:
            issue_df = issue_df.drop(columns=["id"])
            
        # print(issue_df['updated_by'])

        def remove_tz_safe(x):
            if isinstance(x, datetime) and x.tzinfo is not None:
                return x.replace(tzinfo=None)
            return x

        # apply to every cell in issue_df
        issue_df = issue_df.applymap(remove_tz_safe)
        
        # print(issue_df['updated_by'])


        # Convert date columns
        issue_date_columns = [col for col in issue_df.columns if "Date" in col]
        for col in issue_date_columns:
            issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
            
        # print(issue_df['updated_by'])

        # ---- Write header ----
        for col_idx, column_name in enumerate(issue_df.columns, start=1):
            ws2.cell(row=1, column=col_idx, value=column_name)
            
        # print(issue_df['updated_by'])

        # ---- Write data ----
        for r_idx, row in enumerate(dataframe_to_rows(issue_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws2.cell(row=r_idx, column=c_idx)
                cell.value = value

                # Apply date formatting
                if issue_df.columns[c_idx - 1] in issue_date_columns and hasattr(value, "strftime"):
                    cell.number_format = "dd-mmm-yy"

        wb.save(tracker_file_path)
        
        relative_path = tracker_file_path.replace(settings.MEDIA_ROOT, '').lstrip(os.sep)
        download_url = request.build_absolute_uri(
            os.path.join(settings.MEDIA_URL, relative_path).replace('\\', '/')
        )
    
        
        
        return Response({'message': 'request processed successfully !!!', "download_link": download_url}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)   


@api_view(['GET', 'POST'])
def ageing_dashboard_view(request):
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    breakpoint1 = request.data.get('breakpoint1')
    breakpoint2 = request.data.get('breakpoint2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    type = request.data.get('type')
    # issue = request.data.get('issue')

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    unique_data = {
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
    }
    

    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    
    try:

        milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"
        milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"

        breakpoint1 = int(breakpoint1)
        breakpoint2 = int(breakpoint2)
        
        temp = breakpoint1
        breakpoint1 = min(breakpoint1, breakpoint2)
        breakpoint2 = max(temp, breakpoint2)

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None

        milestones = [
            "Allocation",
            "RFAI",
            "RFAI Survey",
            "MO Punch",
            "Material Dispatch",
            "Material Delivered",
            "Installation End",
            "Integration",
            "EMF Submission",
            "Alarm Rectification Done",
            "SCFT I-Deploy Offered",
            "RAN PAT Offer",
            "RAN SAT Offer",
            "MW PAT Offer",
            "MW SAT Offer",
            "Site ONAIR",
            "I-Deploy ONAIR",
        ]
        



        filters = {}
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
            
        if type == "type2":
            if month_start and month_end:
                filters[f"{milestone1_col}__range"] = (month_start, month_end)
                

 
        # 🔹 Fetch data
        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        
        if type == "type1":
            if month_start and month_end:
                obj = obj.filter(
                    Q(**{f"{milestone1_col}__range": (month_start, month_end)}) |
                    Q(**{f"{milestone2_col}__range": (month_start, month_end)})
                )
        

        df = pd.DataFrame(obj.values())

        sites = df['new_site_id'].dropna().unique().tolist()

        
        # filters = {
        #     "site_id__in" : sites
        # }
        
        issue_obj = RelocationIssue.objects.filter(site_id__in=sites)
        issue_df = pd.DataFrame(issue_obj.values())
        
        rename_map = {
            "circle": "Circle",
            "site_id": "Site ID",
            "issue_owner": "Issue Owner",
            "milestone": "Milestone",
            "issue_name": "Issue Name",
            "start_date": "Start Date",
            "close_date": "Close Date",
            "status": "Status",
            "duration": "Duration",
            "updated_by": "Updated_by",
            "updated_at": "Updated_at",
            "created_by": "Created_by",
            "created_at": "Created_at"
        }

        print(issue_df)
        
        issue_df = issue_df.rename(columns=rename_map)

        required_cols = ["Circle", "Site ID", "Issue Owner", "Milestone", "Issue Name", "Start Date", "Close Date", "Status", "Duration", "Updated_by", "Updated_at", "Created_by", "Created_at"]
        for col in required_cols:
            if col not in issue_df.columns:
                issue_df[col] = None

        # Convert date columns safely
        for col in ["Start Date", "Close Date"]:
            if col in issue_df.columns:
                issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
        
        # df = update_ageing(df, issue_df)
        
        date_cols = [col.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '') + '_date' for col in milestones]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']

        if month_start is not None:
            month_start = pd.to_datetime(month_start)

        if month_end is not None:
            month_end = pd.to_datetime(month_end)

        def generate_done_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            temp = df.copy()
            # if start_col == 'rfai_date' and end_col == 'site_onair_date':
            #     temp['rfai_date'] = temp['rfai_date'].where(temp['clear_rfai_date'].isna(), temp['clear_rfai_date'])
            # if issue == 'considered':
            #     temp['total_issue_ageing'] = temp['total_issue_ageing'].fillna(0)
            #     temp['days_diff'] = (temp[end_col] - temp[start_col]).dt.days - temp['total_issue_ageing']
            # else:
            temp["days_diff"] = (temp[end_col] - temp[start_col]).dt.days
            
            if type == "type2":
        
                summary = (
                    temp.groupby("Circle").apply(lambda g: pd.Series({
                        f"{start_label} Done Count": g[start_col].notna().sum(),
                        f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        # "<= 7 days": ((g["days_diff"].notna()) & (g["days_diff"] <= 7)).sum(),
                        # "8-15 days": ((g["days_diff"].notna()) & (g["days_diff"] > 7) & (g["days_diff"] <= 15)).sum(),
                        # ">= 16 days": ((g["days_diff"].notna()) & (g["days_diff"] > 15)).sum(),
                        # "<= 14 days": ((g["days_diff"].notna()) & (g["days_diff"] <= 14)).sum(),
                        # "15-30 days": ((g["days_diff"].notna()) & (g["days_diff"] > 14) & (g["days_diff"] <= 30)).sum(),
                        # ">= 31 days": ((g["days_diff"].notna()) & (g["days_diff"] > 30)).sum(),
                        "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].notna(), "days_diff"].mean()
                    }))
                    .reset_index()
                )
                # ✅ Add total row at the top
                total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                total_row["Circle"] = "Total"

                overall_valid = temp.loc[
                    temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].notna(),
                    "days_diff"
                ]
                total_row["Average Days"] = overall_valid.mean()

            else:
                if month_start is not None and month_end is not None:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": (g[start_col].notna() & (g[start_col] >= month_start) & (g[start_col] <= month_end)).sum(),
                            f"{end_label} Done Count": (g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"<= {breakpoint1} days": (g["days_diff"].notna() & (g["days_diff"] <= breakpoint1) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": (g["days_diff"].notna() & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f">= {breakpoint2} days": (g["days_diff"].notna() & (g["days_diff"] >= breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    # ✅ Add total row at the top
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna() & (temp[end_col] >= month_start) & (temp[end_col] <= month_end),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                else:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": g[start_col].notna().sum(),
                            f"{end_label} Done Count": g[end_col].notna().sum(),
                            f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & g[end_col].notna()).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna()).sum(),
                            f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & g[end_col].notna()).sum(),
                            # "<= 7 days": ((g["days_diff"].notna()) & (g["days_diff"] <= 7)).sum(),
                            # "8-15 days": ((g["days_diff"].notna()) & (g["days_diff"] > 7) & (g["days_diff"] <= 15)).sum(),
                            # ">= 16 days": ((g["days_diff"].notna()) & (g["days_diff"] > 15)).sum(),
                            # "<= 14 days": ((g["days_diff"].notna()) & (g["days_diff"] <= 14)).sum(),
                            # "15-30 days": ((g["days_diff"].notna()) & (g["days_diff"] > 14) & (g["days_diff"] <= 30)).sum(),
                            # ">= 31 days": ((g["days_diff"].notna()) & (g["days_diff"] > 30)).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna(), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    # ✅ Add total row at the top
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna(),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                    

            # # ✅ Add total row at the top
            # total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
            # total_row["Circle"] = "Total"

            # overall_valid = temp.loc[
            #     temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].notna(),
            #     "days_diff"
            # ]
            # total_row["Average Days"] = overall_valid.mean()

            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            # ✅ Column order
            # if end_label == "MOS":
            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]
            # elif end_label == "Integration":
            #     summary = summary[
            #         ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", "<= 7 days", "8-15 days", ">= 16 days", "Average Days"]
            #     ]
            # else:
            #     summary = summary[
            #         ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", "<= 14 days", "15-30 days", ">= 31 days", "Average Days"]
            #     ]

            return summary


        # ✅ Generate all three summaries
        # rfai_to_mos_done_summary = generate_done_summary(df, "RFAI", "MOS", "rfai_date", "material_delivered_date")
        # rfai_to_integration_done_summary = generate_done_summary(df, "RFAI", "Integration", "rfai_date", "integration_date")
        # rfai_to_onair_done_summary = generate_done_summary(df, "RFAI", "MS1", "rfai_date", "site_onair_date")


        # rfai_to_mos_done_summary = rfai_to_mos_done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # rfai_to_integration_done_summary = rfai_to_integration_done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # rfai_to_onair_done_summary = rfai_to_onair_done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        done_summary = generate_done_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        done_summary = done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        def generate_pending_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            
            today = pd.Timestamp.today()
            
            temp = df.copy()
            
            # if start_col == 'rfai_date' and end_col == 'site_onair_date':
            #     temp['rfai_date'] = temp['rfai_date'].where(temp['clear_rfai_date'].isna(), temp['clear_rfai_date'])
            
            # if issue == 'considered':
            #     temp['total_issue_ageing'] = temp['total_issue_ageing'].fillna(0)
            #     temp['days_diff'] = (today - temp[start_col]).dt.days - temp['total_issue_ageing']
            # else:
            temp["days_diff"] = (today - temp[start_col]).dt.days

            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                    f"<= {breakpoint1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[start_col].notna())).sum(),
                    f"{breakpoint1+1}-{breakpoint2-1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[start_col].notna())).sum(),
                    f">= {breakpoint2} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[start_col].notna())).sum(),
                    # "<= 7 days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] <= 7) & (g[start_col].notna())).sum(),
                    # "8-15 days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > 7) & (g["days_diff"] <= 15) & (g[start_col].notna())).sum(),
                    # ">= 16 days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > 15) & (g[start_col].notna())).sum(),
                    # "<= 14 days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] <= 14) & (g[start_col].notna())).sum(),
                    # "15-30 days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > 14) & (g["days_diff"] <= 30) & (g[start_col].notna())).sum(),
                    # ">= 31 days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > 30) & (g[start_col].notna())).sum(),
                    "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].isna(), "days_diff"].mean()
                }))
                .reset_index()
            )

            # ✅ Add total row at the top
            total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
            total_row["Circle"] = "Total"

            overall_valid = temp.loc[
                temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].isna(),
                "days_diff"
            ]
            total_row["Average Days"] = overall_valid.mean()

            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            # ✅ Column order
            # if end_label == "MOS":
            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Pending Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]
            # elif end_label == "Integration":
            #     summary = summary[
            #         ["Circle", f"{start_label} Done Count", f"{end_label} Pending Count", "<= 7 days", "8-15 days", ">= 16 days", "Average Days"]
            #     ]
            # else:
            #     summary = summary[
            #         ["Circle", f"{start_label} Done Count", f"{end_label} Pending Count", "<= 14 days", "15-30 days", ">= 31 days", "Average Days"]
            #     ]

            return summary
        
        # rfai_to_mos_pending_summary = generate_pending_summary(df, "RFAI", "MOS", "rfai_date", "material_delivered_date")
        # rfai_to_integration_pending_summary = generate_pending_summary(df, "RFAI", "Integration", "rfai_date", "integration_date")
        # rfai_to_onair_pending_summary = generate_pending_summary(df, "RFAI", "MS1", "rfai_date", "site_onair_date")

        # rfai_to_mos_pending_summary = rfai_to_mos_pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # rfai_to_integration_pending_summary = rfai_to_integration_pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # rfai_to_onair_pending_summary = rfai_to_onair_pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        pending_summary = generate_pending_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        pending_summary = pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"AGEING_DASHBOARD_FILE_{site_tagging}_{current_date}_{current_time}.xlsx")

        # result_json = rfai_to_mos_done_summary.to_dict(orient="records")
        # json_data1 = json.dumps(result_json)
        
        # result_json = rfai_to_mos_pending_summary.to_dict(orient="records")
        # json_data2 = json.dumps(result_json)
        
        # result_json = rfai_to_integration_done_summary.to_dict(orient="records")
        # json_data3 = json.dumps(result_json)
        
        # result_json = rfai_to_integration_pending_summary.to_dict(orient="records")
        # json_data4 = json.dumps(result_json)
        
        # result_json = rfai_to_onair_done_summary.to_dict(orient="records")
        # json_data5 = json.dumps(result_json)
        
        # result_json = rfai_to_onair_pending_summary.to_dict(orient="records")
        # json_data6 = json.dumps(result_json)

        result_json = pending_summary.to_dict(orient="records")
        json_data2 = json.dumps(result_json)

        result_json = done_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            # "mos_done": json_data1,
            # "mos_pending": json_data2,
            # "integration_done": json_data3,
            # "integration_pending": json_data4,
            # "ms1_done": json_data5,
            # "ms1_pending": json_data6
            "done_summary": json_data1,
            "pending_summary": json_data2
        }
        
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            # rfai_to_mos_done_summary.to_excel(writer, index=False, sheet_name='RFAI TO MOS DONE')
            # rfai_to_integration_done_summary.to_excel(writer, index=False, sheet_name='RFAI TO INTEGRATION DONE')
            # rfai_to_onair_done_summary.to_excel(writer, index=False, sheet_name='RFAI TO MS1 DONE')
            # rfai_to_mos_pending_summary.to_excel(writer, index=False, sheet_name='RFAI TO MOS PENDING')
            # rfai_to_integration_pending_summary.to_excel(writer, index=False, sheet_name='RFAI TO INTEGRATION PENDING')
            # rfai_to_onair_pending_summary.to_excel(writer, index=False, sheet_name='RFAI TO MS1 PENDING')
            pending_summary.to_excel(writer, index=False, sheet_name='PENDING SUMMARY')
            done_summary.to_excel(writer, index=False, sheet_name='DONE SUMMARY')
        
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
        
        return Response({'message': 'request processed successfully !!!', "download_link": download_link, "json_data": json_data, "unique_data": unique_data, "breakpoint1" : breakpoint1, "breakpoint2" : breakpoint2}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)


@api_view(['GET', 'POST'])
def ageing_dashboard_view_issues(request):
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = 'RFAI'
    milestone2 = 'Site ONAIR'
    breakpoint1 = request.data.get('breakpoint1')
    breakpoint2 = request.data.get('breakpoint2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    type = request.data.get('type')
    issue = request.data.get('issue')

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    unique_data = {
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
    }
    

    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    
    try:

        milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"
        milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"

        breakpoint1 = int(breakpoint1)
        breakpoint2 = int(breakpoint2)
        
        temp = breakpoint1
        breakpoint1 = min(breakpoint1, breakpoint2)
        breakpoint2 = max(temp, breakpoint2)

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None

        milestones = [
            "Allocation",
            "RFAI",
            "RFAI Survey",
            "MO Punch",
            "Material Dispatch",
            "Material Delivered",
            "Installation End",
            "Integration",
            "EMF Submission",
            "Alarm Rectification Done",
            "SCFT I-Deploy Offered",
            "RAN PAT Offer",
            "RAN SAT Offer",
            "MW PAT Offer",
            "MW SAT Offer",
            "Site ONAIR",
            "I-Deploy ONAIR",
        ]
        



        filters = {}
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
            
        if type == "type2":
            if month_start and month_end:
                filters[f"{milestone1_col}__range"] = (month_start, month_end)
                

 
        # 🔹 Fetch data
        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        
        if type == "type1":
            if month_start and month_end:
                obj = obj.filter(
                    Q(**{f"{milestone1_col}__range": (month_start, month_end)}) |
                    Q(**{f"{milestone2_col}__range": (month_start, month_end)})
                )
        

        df = pd.DataFrame(obj.values())

        sites = df['new_site_id'].dropna().unique().tolist()

        
        # filters = {
        #     "site_id__in" : sites
        # }
        
        issue_obj = RelocationIssue.objects.filter(site_id__in=sites)
        issue_df = pd.DataFrame(issue_obj.values())
        
        rename_map = {
            "circle": "Circle",
            "site_id": "Site ID",
            "issue_owner": "Issue Owner",
            "milestone": "Milestone",
            "issue_name": "Issue Name",
            "start_date": "Start Date",
            "close_date": "Close Date",
            "status": "Status",
            "duration": "Duration",
            "updated_by": "Updated_by",
            "updated_at": "Updated_at",
            "created_by": "Created_by",
            "created_at": "Created_at"
        }

        print(issue_df)
        
        issue_df = issue_df.rename(columns=rename_map)

        required_cols = ["Circle", "Site ID", "Issue Owner", "Milestone", "Issue Name", "Start Date", "Close Date", "Status", "Duration", "Updated_by", "Updated_at", "Created_by", "Created_at"]
        for col in required_cols:
            if col not in issue_df.columns:
                issue_df[col] = None

        # Convert date columns safely
        for col in ["Start Date", "Close Date"]:
            if col in issue_df.columns:
                issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
        
        # df = update_ageing(df, issue_df)
        
        date_cols = [col.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '') + '_date' for col in milestones]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']

        if month_start is not None:
            month_start = pd.to_datetime(month_start)

        if month_end is not None:
            month_end = pd.to_datetime(month_end)

        def generate_done_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            temp = df.copy()
            if start_col == 'rfai_date' and end_col == 'site_onair_date' and issue == 'considered':
                temp['rfai_date'] = temp['rfai_date'].where(temp['clear_rfai_date'].isna(), temp['clear_rfai_date'])
            if issue == 'considered':
                # temp['total_issue_ageing'] = temp['total_issue_ageing'].fillna(0)
                # temp['days_diff'] = (temp[end_col] - temp[start_col]).dt.days - temp['total_issue_ageing']
                temp['days_diff'] = temp['clear_rfai_to_ms1_ageing']
            else:
                temp["days_diff"] = (temp[end_col] - temp[start_col]).dt.days
            
            if type == "type2":
        
                summary = (
                    temp.groupby("Circle").apply(lambda g: pd.Series({
                        f"{start_label} Done Count": g[start_col].notna().sum(),
                        f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        # "<= 7 days": ((g["days_diff"].notna()) & (g["days_diff"] <= 7)).sum(),
                        # "8-15 days": ((g["days_diff"].notna()) & (g["days_diff"] > 7) & (g["days_diff"] <= 15)).sum(),
                        # ">= 16 days": ((g["days_diff"].notna()) & (g["days_diff"] > 15)).sum(),
                        # "<= 14 days": ((g["days_diff"].notna()) & (g["days_diff"] <= 14)).sum(),
                        # "15-30 days": ((g["days_diff"].notna()) & (g["days_diff"] > 14) & (g["days_diff"] <= 30)).sum(),
                        # ">= 31 days": ((g["days_diff"].notna()) & (g["days_diff"] > 30)).sum(),
                        "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].notna(), "days_diff"].mean()
                    }))
                    .reset_index()
                )
                # ✅ Add total row at the top
                total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                total_row["Circle"] = "Total"

                overall_valid = temp.loc[
                    temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].notna(),
                    "days_diff"
                ]
                total_row["Average Days"] = overall_valid.mean()

            else:
                if month_start is not None and month_end is not None:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": (g[start_col].notna() & (g[start_col] >= month_start) & (g[start_col] <= month_end)).sum(),
                            f"{end_label} Done Count": (g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"<= {breakpoint1} days": (g["days_diff"].notna() & (g["days_diff"] <= breakpoint1) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": (g["days_diff"].notna() & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f">= {breakpoint2} days": (g["days_diff"].notna() & (g["days_diff"] >= breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    # ✅ Add total row at the top
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna() & (temp[end_col] >= month_start) & (temp[end_col] <= month_end),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                else:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": g[start_col].notna().sum(),
                            f"{end_label} Done Count": g[end_col].notna().sum(),
                            f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & g[end_col].notna()).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna()).sum(),
                            f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & g[end_col].notna()).sum(),
                            # "<= 7 days": ((g["days_diff"].notna()) & (g["days_diff"] <= 7)).sum(),
                            # "8-15 days": ((g["days_diff"].notna()) & (g["days_diff"] > 7) & (g["days_diff"] <= 15)).sum(),
                            # ">= 16 days": ((g["days_diff"].notna()) & (g["days_diff"] > 15)).sum(),
                            # "<= 14 days": ((g["days_diff"].notna()) & (g["days_diff"] <= 14)).sum(),
                            # "15-30 days": ((g["days_diff"].notna()) & (g["days_diff"] > 14) & (g["days_diff"] <= 30)).sum(),
                            # ">= 31 days": ((g["days_diff"].notna()) & (g["days_diff"] > 30)).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna(), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    # ✅ Add total row at the top
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna(),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                    

            # # ✅ Add total row at the top
            # total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
            # total_row["Circle"] = "Total"

            # overall_valid = temp.loc[
            #     temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].notna(),
            #     "days_diff"
            # ]
            # total_row["Average Days"] = overall_valid.mean()

            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            # ✅ Column order
            # if end_label == "MOS":
            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]
            # elif end_label == "Integration":
            #     summary = summary[
            #         ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", "<= 7 days", "8-15 days", ">= 16 days", "Average Days"]
            #     ]
            # else:
            #     summary = summary[
            #         ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", "<= 14 days", "15-30 days", ">= 31 days", "Average Days"]
            #     ]

            return summary


        # ✅ Generate all three summaries
        # rfai_to_mos_done_summary = generate_done_summary(df, "RFAI", "MOS", "rfai_date", "material_delivered_date")
        # rfai_to_integration_done_summary = generate_done_summary(df, "RFAI", "Integration", "rfai_date", "integration_date")
        # rfai_to_onair_done_summary = generate_done_summary(df, "RFAI", "MS1", "rfai_date", "site_onair_date")


        # rfai_to_mos_done_summary = rfai_to_mos_done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # rfai_to_integration_done_summary = rfai_to_integration_done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # rfai_to_onair_done_summary = rfai_to_onair_done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        done_summary = generate_done_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        done_summary = done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        def generate_pending_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            
            today = pd.Timestamp.today()
            
            temp = df.copy()
            
            if start_col == 'rfai_date' and end_col == 'site_onair_date' and issue == 'considered':
                temp['rfai_date'] = temp['rfai_date'].where(temp['clear_rfai_date'].isna(), temp['clear_rfai_date'])
            
            if issue == 'considered':
                # temp['total_issue_ageing'] = temp['total_issue_ageing'].fillna(0)
                # temp['days_diff'] = (today - temp[start_col]).dt.days - temp['total_issue_ageing']
                temp['days_diff'] = temp['clear_rfai_to_ms1_ageing']
            else:
                temp["days_diff"] = (today - temp[start_col]).dt.days

            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                    f"<= {breakpoint1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[start_col].notna())).sum(),
                    f"{breakpoint1+1}-{breakpoint2-1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[start_col].notna())).sum(),
                    f">= {breakpoint2} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[start_col].notna())).sum(),
                    # "<= 7 days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] <= 7) & (g[start_col].notna())).sum(),
                    # "8-15 days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > 7) & (g["days_diff"] <= 15) & (g[start_col].notna())).sum(),
                    # ">= 16 days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > 15) & (g[start_col].notna())).sum(),
                    # "<= 14 days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] <= 14) & (g[start_col].notna())).sum(),
                    # "15-30 days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > 14) & (g["days_diff"] <= 30) & (g[start_col].notna())).sum(),
                    # ">= 31 days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > 30) & (g[start_col].notna())).sum(),
                    "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].isna(), "days_diff"].mean()
                }))
                .reset_index()
            )

            # ✅ Add total row at the top
            total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
            total_row["Circle"] = "Total"

            overall_valid = temp.loc[
                temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].isna(),
                "days_diff"
            ]
            total_row["Average Days"] = overall_valid.mean()

            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            # ✅ Column order
            # if end_label == "MOS":
            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Pending Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]
            # elif end_label == "Integration":
            #     summary = summary[
            #         ["Circle", f"{start_label} Done Count", f"{end_label} Pending Count", "<= 7 days", "8-15 days", ">= 16 days", "Average Days"]
            #     ]
            # else:
            #     summary = summary[
            #         ["Circle", f"{start_label} Done Count", f"{end_label} Pending Count", "<= 14 days", "15-30 days", ">= 31 days", "Average Days"]
            #     ]

            return summary
        
        # rfai_to_mos_pending_summary = generate_pending_summary(df, "RFAI", "MOS", "rfai_date", "material_delivered_date")
        # rfai_to_integration_pending_summary = generate_pending_summary(df, "RFAI", "Integration", "rfai_date", "integration_date")
        # rfai_to_onair_pending_summary = generate_pending_summary(df, "RFAI", "MS1", "rfai_date", "site_onair_date")

        # rfai_to_mos_pending_summary = rfai_to_mos_pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # rfai_to_integration_pending_summary = rfai_to_integration_pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # rfai_to_onair_pending_summary = rfai_to_onair_pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        pending_summary = generate_pending_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        pending_summary = pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"AGEING_DASHBOARD_FILE_{site_tagging}_{current_date}_{current_time}.xlsx")

        # result_json = rfai_to_mos_done_summary.to_dict(orient="records")
        # json_data1 = json.dumps(result_json)
        
        # result_json = rfai_to_mos_pending_summary.to_dict(orient="records")
        # json_data2 = json.dumps(result_json)
        
        # result_json = rfai_to_integration_done_summary.to_dict(orient="records")
        # json_data3 = json.dumps(result_json)
        
        # result_json = rfai_to_integration_pending_summary.to_dict(orient="records")
        # json_data4 = json.dumps(result_json)
        
        # result_json = rfai_to_onair_done_summary.to_dict(orient="records")
        # json_data5 = json.dumps(result_json)
        
        # result_json = rfai_to_onair_pending_summary.to_dict(orient="records")
        # json_data6 = json.dumps(result_json)

        result_json = pending_summary.to_dict(orient="records")
        json_data2 = json.dumps(result_json)

        result_json = done_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            # "mos_done": json_data1,
            # "mos_pending": json_data2,
            # "integration_done": json_data3,
            # "integration_pending": json_data4,
            # "ms1_done": json_data5,
            # "ms1_pending": json_data6
            "done_summary": json_data1,
            "pending_summary": json_data2
        }
        
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            # rfai_to_mos_done_summary.to_excel(writer, index=False, sheet_name='RFAI TO MOS DONE')
            # rfai_to_integration_done_summary.to_excel(writer, index=False, sheet_name='RFAI TO INTEGRATION DONE')
            # rfai_to_onair_done_summary.to_excel(writer, index=False, sheet_name='RFAI TO MS1 DONE')
            # rfai_to_mos_pending_summary.to_excel(writer, index=False, sheet_name='RFAI TO MOS PENDING')
            # rfai_to_integration_pending_summary.to_excel(writer, index=False, sheet_name='RFAI TO INTEGRATION PENDING')
            # rfai_to_onair_pending_summary.to_excel(writer, index=False, sheet_name='RFAI TO MS1 PENDING')
            pending_summary.to_excel(writer, index=False, sheet_name='PENDING SUMMARY')
            done_summary.to_excel(writer, index=False, sheet_name='DONE SUMMARY')
        
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
        
        return Response({'message': 'request processed successfully !!!', "download_link": download_link, "json_data": json_data, "unique_data": unique_data, "breakpoint1" : breakpoint1, "breakpoint2" : breakpoint2}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
   

@api_view(['GET', 'POST'])
def frontend_editing_display_view(request):
    userId = request.data.get('userId')

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    
    circles = user.circles

    type = request.data.get('day_type')
    milestone = request.data.get('milestone')
    col_name = request.data.get('col_name')
    view = request.data.get('view')
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    toco_name = request.data.get('toco_name', [])
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None

    # circle = [ACCESS_RIGHTS[userId]["Circle"]] if ACCESS_RIGHTS[userId]["Circle"] != 'CENTRAL' else [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    # if 'ALL' in circle:
    #     circle = circles
    # elif 'ALL' not in circles:
    #     circle = [c for c in circle if c in circles]

    circle = [c.strip() for c in circle.split(',')] if circle else ["CENTRAL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    toco_name = [n.strip() for n in toco_name.split(',')] if toco_name else ["ALL"]
    
    if 'CENTRAL' in circle:
        circle = circles
    elif 'CENTRAL' not in circles:
        circle = [c for c in circle if c in circles]
    
    if not type or not milestone or not col_name or not view:
        return Response({"error": "Missing required parameters"}, status=400)
    
    milestone = milestone.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
    
    try:
        def get_week_range(col_name):
            month_str, week_part = col_name.split()
            month_abbr, year_suffix = month_str.split('-')
            week_num = int(week_part.replace('W', ''))

            year = 2000 + int(year_suffix)  # '25' -> 2025
            month = datetime.strptime(month_abbr, "%b").month

            # Define cycle range: 26th of previous month → 25th of current month
            prev_month = month - 1 if month > 1 else 12
            prev_year = year if month > 1 else year - 1

            cycle_start = date(prev_year, prev_month, 26)
            cycle_end = date(year, month, 25)

            # Build week ranges: every 7 days from cycle_start to cycle_end
            weeks = []
            start = cycle_start
            while start <= cycle_end:
                end = start + timedelta(days=6)
                if end > cycle_end:
                    end = cycle_end
                weeks.append((start, end))
                start = end + timedelta(days=1)

            # Get requested week
            week_start, week_end = weeks[week_num - 1]
            return week_start, week_end, cycle_start
        
        filters = {}
        if type == 'daily':
            col_name = dtime.strptime(col_name, "%d-%b-%y").date()
            if view == 'Cumulative':
                filters[f"{milestone}__lte"] = col_name
            else:
                filters[milestone] = col_name
            
        elif type == 'weekly':
            week_start, week_end, cycle_start = get_week_range(col_name)

            if view == 'Cumulative':
                filters[f"{milestone}__range"] = (cycle_start, week_end)
            else:
                filters[f"{milestone}__range"] = (week_start, week_end)

        else:
            month_abbr, year_suffix = col_name.split('-')
            year = 2000 + int(year_suffix)
            month = datetime.strptime(month_abbr, "%b").month

            # Compute cycle range
            prev_month = month - 1 if month > 1 else 12
            prev_year = year if month > 1 else year - 1

            cycle_start = datetime(prev_year, prev_month, 26)
            cycle_end = datetime(year, month, 25)

            if view == 'Cumulative':
                filters[f"{milestone}__lte"] = cycle_end
            else:
                filters[f"{milestone}__range"] = (cycle_start, cycle_end)


        if "CENTRAL" not in circle:
            filters["circle__in"] = circle
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in toco_name:
            filters["new_toco_name__in"] = toco_name


        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        if df.empty :
            json_data = []
            json_val = json.dumps(json_data)
            return Response({"message" : "No data found", "data": json_val},status = 200)
        
        if month_start and month_end:
            df = df[df['site_onair_date'].isna() | (df['site_onair_date'] >= month_start)]
        
        sites = df['new_site_id'].dropna().unique().tolist()
        
        issue_obj = RelocationIssue.objects.filter(site_id__in=sites)
        issue_df = pd.DataFrame(issue_obj.values())

        rename_map = {
            "circle": "Circle",
            "site_id": "Site ID",
            "issue_owner": "Issue Owner",
            "milestone": "Milestone",
            "issue_name": "Issue Name",
            "start_date": "Start Date",
            "close_date": "Close Date",
            "status": "Status",
            "duration": "Duration",
            "remarks": "Remarks",
            "updated_by": "Updated_by",
            "updated_at": "Updated_at",
            "created_by": "Created_by",
            "created_at": "Created_at"
        }

        print(issue_df)
        
        issue_df = issue_df.rename(columns=rename_map)

        required_cols = ["Circle", "Site ID", "Issue Owner", "Milestone", "Issue Name", "Start Date", "Close Date", "Status", "Duration", "Remarks", "Updated_by", "Updated_at", "Created_by", "Created_at"]
        for col in required_cols:
            if col not in issue_df.columns:
                issue_df[col] = None

        # # Convert date columns safely
        # for col in ["Start Date", "Close Date"]:
        #     if col in issue_df.columns:
        #         issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
        
        # df = update_ageing(df, issue_df)

        for col in df.columns:
            if 'date' in col:
                if col != 'last_updated_date' :
                    converted = pd.to_datetime(df[col], errors='coerce')
    
                    if converted.notna().sum() > 0:
                        df[col] = converted.dt.strftime('%d-%b-%y')
                else:
                    converted = pd.to_datetime(df[col], errors='coerce')
                    if converted.notna().sum() > 0:
                        df[col] = converted.dt.strftime('%d-%b-%y %H:%M:%S')
 
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        tracker_file_path = os.path.join(output_folder, f"FRONTEND_EDITING_TRACKER_FILE_{current_date}_{current_time}.xlsx")
        
        df.insert(0, "Unique ID", range(1, len(df) + 1))
       
        df.drop(columns=['id'], inplace=True)

        df = df[['Unique ID', 'circle', 'site_tagging', 'old_toco_name', 'old_site_id', 'new_site_id', 'new_toco_name',
                 'sr_number', 'ran_oem', 'media_type', 'mw_oem', 'relocation_method', 'relocation_type', 'old_site_band',
                 'new_site_band', 'rfai_date', 'allocation_date', 'rfai_survey_date', 'mo_punch_date',
                 'material_dispatch_date', 'material_delivered_date', 'installation_start_date', 'installation_end_date',
                 'integration_date', 'emf_submission_date', 'ran_lkf_status', 'alarm_status', 'alarm_rectification_done_date',
                 'scft_done_date', 'scft_i_deploy_offered_date', 'ran_pat_offer_date', 'ran_sat_offer_date', 'mw_plan_id',
                 'mw_pat_offer_date', 'rsl_value_status', 'enm_status', 'mw_lkf', 'mw_sat_offer_date', 'mw_ms1_mids_date',
                 'site_onair_date', 'i_deploy_onair_date', 'current_status', 'detailed_remarks', 'rfai_rejected_date', 
                 'clear_rfai_date', 'pri_count', 'pri_issue_ageing', 'other_ust_issue_ageing', 'other_airtel_issue_ageing', 'total_issue_ageing', 
                 'clear_rfai_to_ms1_ageing', 'rfai_to_ms1_ageing', 'ran_pat_accepted_date', 'ran_sat_accepted_date', 
                 'mw_pat_accepted_date', 'mw_sat_accepted_date', 'scft_accepted_date', 'kpi_at_offer_date', 'kpi_at_accepted_date',
                 'four_g_ms2_date', 'five_g_ms2_date', 'final_ms2_date', "dismantling_survey_date", "sreq_creq_raised_date",
                 "dismantle_date", "material_pickup_date", "material_submission_date", "oci_done_date", "sign_off_date", "dismantling_status", 
                 'last_updated_date', 'last_updated_by']]
        
        
        template_path = os.path.join(BASE_URL, "template", "templateAlok_v.1.xlsx")
        wb = load_workbook(template_path)
        ws = wb.active  
        
        df_copy = df.copy()
        df_copy = df_copy.fillna('').astype(str)
        
        json_dict_data = df_copy.to_dict(orient="records")
        json_val = json.dumps(json_dict_data)
        
        date_columns = [col for col in df.columns if col.endswith("_date") and col != 'last_updated_date']

        for col in date_columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        start_row = 3
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                # apply date format for *_date columns
                if df.columns[c_idx - 1].endswith("_date") and hasattr(value, "strftime") and df.columns[c_idx - 1] != 'last_updated_date':
                    cell.number_format = "dd-mmm-yy"
                    
        ws2 = wb.create_sheet(title="Issues Tracker")

        # Clean issue_df (remove id if present)
        if "id" in issue_df.columns:
            issue_df = issue_df.drop(columns=["id"])
            
        # print(issue_df['updated_by'])

        def remove_tz_safe(x):
            if isinstance(x, datetime) and x.tzinfo is not None:
                return x.replace(tzinfo=None)
            return x

        # apply to every cell in issue_df
        issue_df = issue_df.applymap(remove_tz_safe)
        
        # print(issue_df['updated_by'])


        # Convert date columns
        issue_date_columns = [col for col in issue_df.columns if "Date" in col]
        for col in issue_date_columns:
            issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
            
        # print(issue_df['updated_by'])

        # ---- Write header ----
        for col_idx, column_name in enumerate(issue_df.columns, start=1):
            ws2.cell(row=1, column=col_idx, value=column_name)
            
        # print(issue_df['updated_by'])

        # ---- Write data ----
        for r_idx, row in enumerate(dataframe_to_rows(issue_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws2.cell(row=r_idx, column=c_idx)
                cell.value = value

                # Apply date formatting
                if issue_df.columns[c_idx - 1] in issue_date_columns and hasattr(value, "strftime"):
                    cell.number_format = "dd-mmm-yy"
 
        wb.save(tracker_file_path)
        relative_path = tracker_file_path.replace(settings.MEDIA_ROOT, '').lstrip(os.sep)
        download_url = request.build_absolute_uri(
            os.path.join(settings.MEDIA_URL, relative_path).replace('\\', '/')
        )
        
        
        return Response({'message': 'request processed successfully !!!', "data": json_val, "download_link": download_url}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
    
    
@api_view(['POST'])
def frontend_editing_update_view(request):
    userId = request.data.get('userId')
    data = request.data.get('data')

    if not userId or not data:
        return Response({'error': 'userId and data are required.'}, status=400)
    
    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    if user.right == 'Read':
        return Response({'error': 'ACCESS DENIED'}, status=403)
    
    
    circles = user.circles

    try:
        data = json.loads(data)
        df = pd.DataFrame([data])

        if df.empty:
            return Response({'error': 'No data provided for update.'}, status=400)
        
        print(df)
        
 ####################################### UPDATING PRI HISTORY ################################################

        # def update_pri_history(df):
        #     start_date_val = df.at[0, "pri_start_date"]
        #     close_date_val = df.at[0, "pri_close_date"]
        #     history_val = df.at[0, "pri_history"]
        #     count_val = df.at[0, "pri_count"]

        #     # Convert to proper types and handle NaN
        #     start_date = "" if pd.isna(start_date_val) else str(start_date_val).strip()
        #     close_date = "" if pd.isna(close_date_val) else str(close_date_val).strip()
        #     history = "" if pd.isna(history_val) else str(history_val).strip()
        #     print("1.1")
        #     count = 0 if not str(count_val).strip() else int(float(count_val))
        #     print("1.2")
        #     if history.lower() in ["nan", "none", ""]:
        #         history = ""
                
        #     print(history)

        #     # Parse existing history entries into list
        #     entries = [x.strip() for x in history.split(";") if x.strip()] if history else []
        #     last_entry = entries[-1] if entries else None
            
        #     print("ENTRIES" , entries)
            
        #     print("LAST ENTRY",last_entry,"end")

        #     # Check if last entry has open "C" part
        #     has_open_start = last_entry and last_entry.endswith("-")
            
        #     print("HAS OPEN START : ", has_open_start)
            
        #     if (not start_date or start_date.lower() == 'nan') and (not close_date or close_date.lower() == 'nan'):
        #         raise ValueError("start_date or close_date not provided !")

        #     # Start date update
        #     if start_date and start_date.lower() != "nan":
        #         if has_open_start:
        #             raise ValueError("Cannot start new PRI — previous PRI is not closed.")
                
        #         start_date = normalize_date_format(start_date)

        #         count += 1
        #         entries.append(f"S{count}: {start_date}, C{count}: -")

        #     last_entry = entries[-1] if entries else None
            
        #     print("ENTRIES" , entries)
            
        #     print("LAST ENTRY",last_entry,"end")

        #     # Check if last entry has open "C" part
        #     has_open_start = last_entry and last_entry.endswith("-")

        #     # Close date update
        #     if close_date and close_date.lower() != "nan":
        #         if not has_open_start:
        #             raise ValueError("Cannot close PRI — no open PRI found.")
                
        #         close_date = normalize_date_format(close_date)

        #         s_part = last_entry.split(",")[0]  # S part only
        #         entries[-1] = f"{s_part}, C{count}: {close_date}"

        #     # Update dataframe
        #     df.at[0, "pri_history"] = "; ".join(entries)
        #     df.at[0, "pri_count"] = count

        #     # Clear input date fields after update
        #     df.at[0, "pri_start_date"] = ""
        #     df.at[0, "pri_close_date"] = ""

        #     print(df.at[0,'pri_history'])
            
        #     return df
        
        # def update_other_issue_history(df):
        #     """
        #     Updates:
        #     - issue_history (per-issue timelines)
        #     - other_issue_ageing (sum of all issues)
        #     - total_issue_ageing (pri_issue_ageing + other_issue_ageing)

        #     Columns used:
        #     'issue', 'issue_start_date', 'issue_close_date',
        #     'issue_history', 'other_issue_ageing', 'total_issue_ageing', 'pri_issue_ageing'
        #     """

        #     # ---- Read raw values from the single row ----
        #     issue_val = df.at[0, "issue"] if "issue" in df.columns else None
        #     start_date_val = df.at[0, "issue_start_date"]
        #     close_date_val = df.at[0, "issue_close_date"]
        #     history_val = df.at[0, "issue_history"] if "issue_history" in df.columns else ""
        #     pri_age_val = df.at[0, "pri_issue_ageing"] if "pri_issue_ageing" in df.columns else 0

        #     # ---- Normalise values ----
        #     issue = "" if pd.isna(issue_val) else str(issue_val).strip()
        #     start_date = "" if pd.isna(start_date_val) else str(start_date_val).strip()
        #     close_date = "" if pd.isna(close_date_val) else str(close_date_val).strip()
        #     history = "" if pd.isna(history_val) else str(history_val).strip()
        #     pri_issue_ageing = 0 if not str(pri_age_val).strip() else int(float(pri_age_val))


        #     if history.lower() in ["nan", "none", ""]:
        #         history = ""

        #     # No issue name but dates provided → treat as error
        #     print("ISSUE", issue)
        #     if not issue:
        #         raise ValueError("Issue name not provided !")
            
        #     if (not start_date or start_date.lower() == 'nan') and (not close_date or close_date.lower() == 'nan'):
        #         raise ValueError("start_date or close_date not provided !")

        #     # ---- Parse existing history into entries ----
        #     entries = [x.strip() for x in history.split(";") if x.strip()] if history else []

        #     # Find the last entry for THIS issue (per-issue logic)
        #     last_issue_entry = None
        #     last_issue_index = None

        #     for idx in range(len(entries) - 1, -1, -1):
        #         e = entries[idx]
        #         if e.startswith(f"{issue}:"):
        #             last_issue_entry = e
        #             last_issue_index = idx
        #             break

        #     # Determine if the last entry for this issue is still open
        #     # (we only care about this issue's last segment, not others)
        #     has_open_for_issue = bool(
        #         last_issue_entry and last_issue_entry.strip().endswith("C: -")
        #     )

        #     # ---- Handle Start ----
        #     if start_date and start_date.lower() != "nan":
        #         if not issue:
        #             raise ValueError("Issue name cannot be empty when starting an issue.")

        #         if has_open_for_issue:
        #             raise ValueError(
        #                 f"Cannot start new cycle for issue '{issue}' — previous cycle is still open."
        #             )

        #         start_date = normalize_date_format(start_date)

        #         # New open entry for this issue
        #         new_entry = f"{issue}: S: {start_date}, C: -"
        #         entries.append(new_entry)

        #     for idx in range(len(entries) - 1, -1, -1):
        #         e = entries[idx]
        #         if e.startswith(f"{issue}:"):
        #             last_issue_entry = e
        #             last_issue_index = idx
        #             break

        #     # Determine if the last entry for this issue is still open
        #     # (we only care about this issue's last segment, not others)
        #     has_open_for_issue = bool(
        #         last_issue_entry and last_issue_entry.strip().endswith("C: -")
        #     )

        #     # ---- Handle Close ----
        #     if close_date and close_date.lower() != "nan":
        #         if not issue:
        #             raise ValueError("Issue name cannot be empty when closing an issue.")

        #         if not has_open_for_issue:
        #             raise ValueError(
        #                 f"Cannot close issue '{issue}' — no open cycle found for this issue."
        #             )

        #         # Update the last open entry for this issue
        #         # last_issue_entry = "{issue}: S: {start_date_old}, C: -"

        #         close_date = normalize_date_format(close_date)

        #         parts = [p.strip() for p in last_issue_entry.split(",")]
        #         s_part = parts[0]  # "{issue}: S: {start_date_old}"
        #         updated_entry = f"{s_part}, C: {close_date}"
        #         entries[last_issue_index] = updated_entry

        #     # ---- Write back updated issue_history ----
        #     final_issue_history = "; ".join(entries)
        #     df.at[0, "issue_history"] = final_issue_history

        #     # ---- Recalculate other_issue_ageing ----
        #     other_ageing = calculate_other_issue_ageing(final_issue_history)
        #     df.at[0, "other_issue_ageing"] = other_ageing

        #     # ---- Update total_issue_ageing (PRI + Other) ----
        #     df.at[0, "total_issue_ageing"] = int(pri_issue_ageing) + int(other_ageing)

        #     # ---- Optional: clear input fields after processing ----
        #     df.at[0, "issue_start_date"] = ""
        #     df.at[0, "issue_close_date"] = ""
        #     df.at[0, "issue"] = ""
        #     # you may or may not want to clear 'issue'; keeping it can be helpful for UI
        #     # df.at[0, "issue"] = ""

        #     return df
        
        # if (df.at[0, 'pri_start_date'] and pd.notna(df.at[0, 'pri_start_date']) and df.at[0, 'pri_start_date'].lower() != 'nan') or (df.at[0, 'pri_close_date'] and pd.notna(df.at[0, 'pri_close_date']) and df.at[0, 'pri_close_date'].lower() != 'nan'):
        #     print("1")
        #     df = update_pri_history(df)

        # df.at[0, "pri_issue_ageing"] = calculate_pri_ageing(df.at[0, "pri_history"])

        # print(df)
        
        # if df.at[0, 'issue'] and pd.notna(df.at[0, 'issue']) and df.at[0, 'issue'].lower() != 'nan' and ((df.at[0, 'issue_start_date'] and pd.notna(df.at[0, 'issue_start_date']) and df.at[0, 'issue_start_date'].lower() != 'nan') or (df.at[0, 'issue_close_date'] and pd.notna(df.at[0, 'issue_close_date']) and df.at[0, 'issue_close_date'].lower() != 'nan')):
        #     df = update_other_issue_history(df)
        # elif df.at[0, 'issue'] and pd.notna(df.at[0, 'issue']) and df.at[0, 'issue'].lower() != 'nan':
        #     df.at[0, 'issue'] = ""
        # elif (df.at[0, 'issue_start_date'] and pd.notna(df.at[0, 'issue_start_date']) and df.at[0, 'issue_start_date'].lower() != 'nan') or (df.at[0, 'issue_close_date'] and pd.notna(df.at[0, 'issue_close_date']) and df.at[0, 'issue_close_date'].lower() != 'nan'):
        #     df.at[0, 'issue_start_date'] = ""
        #     df.at[0, 'issue_close_date'] = ""

        # # df.at[0, "other_issue_ageing"] = calculate_other_issue_ageing(df.at[0, "issue_history"]) 

        # # # ---- Update total_issue_ageing (PRI + Other) ----
        # # df.at[0, "total_issue_ageing"] = calculate_total_issue_ageing( df.at[0, "pri_history"], df.at[0, "issue_history"])

        # df = update_ageing(df)

        # ✅ Only one record expected
        record = df.iloc[0].to_dict()
        
        print(record)

        # ✅ Safe converters
        INVALID_DATE_STRINGS = {"need to check", "pending", "na", "tbd", "n/a", "nan", "-"}

        def safe_datetime(value):
            import pandas as pd
            from datetime import datetime as dtime
            if pd.isna(value) or value is pd.NaT:
                return None
            if isinstance(value, (pd.Timestamp, dtime)):
                return value
            if isinstance(value, str):
                v = value.strip().lower()
                if v in INVALID_DATE_STRINGS or not v:
                    return None
                try:
                    val = pd.to_datetime(value, errors="coerce")
                    if pd.isna(val):
                        return None
                    return val.to_pydatetime()
                except Exception:
                    return None
            return None

        def safe_int(val):
            try:
                return int(val)
            except (TypeError, ValueError):
                return None

        # ✅ Extract identifiers
        circle_val = str(record.get("circle")).strip() if record.get("circle") else None
        new_site_id_val = str(record.get("new_site_id")).strip() if record.get("new_site_id") else None

        if not circle_val or not new_site_id_val:
            return Response({'error': 'Circle and New Site ID are required.'}, status=400)

        # ✅ Circle restriction
        if 'CENTRAL' not in circles and circle_val not in circles:
            return Response({'error': f'You are not allowed to edit {circle_val} circle data.'}, status=403)

        # ✅ Validate columns
        required_columns = [
            col.lower()
            .strip()
            .replace(" ", "_")
            .replace("-", "_")
            .replace("4", "four_")
            .replace("5", "five_")
            .replace("/", "_")
            for col in user.columns
        ]
        allowed_data = {}

        for col in record:
            if col not in required_columns:
                continue

            val = record[col]
            if "date" in col.lower():
                val = safe_datetime(val)
            elif "ageing" in col.lower() or "count" in col.lower():
                val = safe_int(val)

            field_name = col.lower().strip().replace(" ", "_").replace("-", "_").replace("4", "four_").replace("5", "five_").replace("/", "_")
            allowed_data[col] = val

        # ✅ Add system fields
        allowed_data.update({
            "last_updated_by": userId,
            "last_updated_date": timezone.now(),
        })

        # ✅ Find existing record
        try:
            obj = AlokTrackerModel.objects.get(circle__iexact=circle_val, new_site_id__iexact=new_site_id_val)
        except AlokTrackerModel.DoesNotExist:
            return Response({'error': 'Record not found in database.'}, status=404)

        # ✅ Update object fields
        for field, value in allowed_data.items():
            setattr(obj, field, value)

        obj.save(update_fields=list(allowed_data.keys()))

        return Response({'status': True, 'message': 'Record updated successfully.'}, status=200)

    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['GET', 'POST'])
def graphs_view(request):
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
            "Allocation",
            "RFAI",
            "RFAI Survey",
            "MO Punch",
            "Material Dispatch",
            "Material Delivered",
            "Installation End",
            "Integration",
            "EMF Submission",
            "Alarm Rectification Done",
            "SCFT I-Deploy Offered",
            "RAN PAT Offer",
            "RAN SAT Offer",
            "MW PAT Offer",
            "MW SAT Offer",
            "Site ONAIR",
            "I-Deploy ONAIR",
        ]
    
    unique_data = {
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]
    }

    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        if site_tagging and "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if month_start and month_end:
            filters[f"{milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '') + '_date'}__range"] = (month_start, month_end)

        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        

        date_cols = ["integration_date", "ran_pat_offer_date", "ran_sat_offer_date", "scft_i_deploy_offered_date"]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']

        def generate_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"{end_label} Pending Count"]
            ]

            return summary

        def generate_table_summary(df, milestone1, milestone2):

            # labels = [
            #     "Integration Status",
            #     "EMF Status",
            #     "Alarm Status",
            #     "SCFT Offered",
            #     "PAT Offered",
            #     "SAT Offered",
            #     "MW PAT Offered",
            #     "MW SAT Offered",
            #     "MW MIDS MS1",
            #     "I-Deploy MS1",
            # ]

            # milestone_cols = [
            #     "integration_date",
            #     "emf_submission_date",
            #     "alarm_rectification_done_date",
            #     "scft_i_deploy_offered_date",
            #     "ran_pat_offer_date",
            #     "ran_sat_offer_date",
            #     "mw_pat_offer_date",
            #     "mw_sat_offer_date",
            #     "mw_ms1_mids_date",
            #     "i_deploy_onair_date",
            # ]

            milestone_cols = [(col.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date") for col in milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]]

            temp = df.copy()

            circles = ["ALL"] + temp["circle"].unique().tolist()

            final_df = pd.DataFrame()

            milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"
            milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"

            for circle in circles:
                if circle != "ALL":
                    temp  = df[df["circle"] == circle]

                done_row = {}
                pending_row = {}

                for col, label in zip(milestone_cols, milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]):

                    # DONE → col is not empty
                    done_row[label] = (temp[col].notna() & temp[milestone1_col].notna() ).sum()

                    # PENDING → col is empty
                    pending_row[label] = (temp[col].isna() & temp[milestone1_col].notna() ).sum()

                # Build final DataFrame exactly like your screenshot
                rows_to_add = pd.DataFrame([
                    {"Circle": circle, "Site Status": "Done", **done_row},
                    {"Circle": circle, "Site Status": "Pending", **pending_row},
                ])

                final_df = pd.concat([final_df, rows_to_add], ignore_index=True)

            # Convert to Int64
            for c in final_df.columns:
                if c != "Site Status" and c != "Circle":
                    final_df[c] = final_df[c].astype("Int64")

            return final_df

        # def generate_rfai_to_integration_table_summary(df):

        #     labels = [
        #         "Site RFAI",
        #         "Site Allocation",
        #         "Survey Status",
        #         "Material Punched",
        #         "Material Delivered",
        #         "Installation End",
        #         "Integration Status",
        #     ]

        #     milestone_cols = [
        #         "rfai_date",
        #         "allocation_date",
        #         "rfai_survey_date",
        #         "mo_punch_date",
        #         "material_delivered_date",
        #         "installation_end_date",
        #         "integration_date",
        #     ]

        #     temp = df.copy()

        #     circles = ["ALL"] + temp["circle"].unique().tolist()

        #     final_df = pd.DataFrame()

        #     for circle in circles:
        #         if circle != "ALL":
        #             temp  = df[df["circle"] == circle]

        #         done_row = {}
        #         pending_row = {}

        #         for col, label in zip(milestone_cols, labels):

        #             # DONE → col is not empty
        #             done_row[label] = (temp[col].notna() & temp["integration_date"].isna() & temp["rfai_date"].notna()).sum()

        #             # PENDING → col is empty
        #             pending_row[label] = (temp[col].isna() & temp["integration_date"].isna() & temp["rfai_date"].notna()).sum()

        #         # Build final DataFrame exactly like your screenshot
        #         rows_to_add = pd.DataFrame([
        #             {"Circle": circle, "Site Status": "Done", **done_row},
        #             {"Circle": circle, "Site Status": "Pending", **pending_row},
        #         ])

        #         final_df = pd.concat([final_df, rows_to_add], ignore_index=True)

        #     # Convert to Int64
        #     for c in final_df.columns:
        #         if c != "Site Status" and c != "Circle":
        #             final_df[c] = final_df[c].astype("Int64")

        #     return final_df

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        start_col = start_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_") + "_date"
        end_col = end_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_") + "_date"

        # ✅ Generate all three summaries
        # rfai_to_integration_summary = generate_summary(df, "RFAI", "Integration", "rfai_date", "integration_date")
        # integration_to_scft_summary = generate_summary(df, "Integration", "SCFT I-Deploy Offered", "integration_date", "scft_i_deploy_offered_date")
        # integration_to_ran_sat_offered_summary = generate_summary(df, "Integration", "RAN SAT Offered", "integration_date", "ran_sat_offer_date")
        # integration_to_ran_pat_offered_summary = generate_summary(df, "Integration", "RAN PAT Offered", "integration_date", "ran_pat_offer_date")
        # rfai_to_site_onair_summary = generate_summary(df, "RFAI", "Site ONAIR", "rfai_date", "site_onair_date")
        graph_summary = generate_summary(df, start_label, end_label, start_col, end_col)
        # integration_to_onair_table_summary = generate_table_summary(df)
        table_summary = generate_table_summary(df, start_label, end_label)

        # ✅ Convert all dataframes to json
        # rfai_to_integration_summary = rfai_to_integration_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # rfai_to_site_onair_summary = rfai_to_site_onair_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # integration_to_scft_summary = integration_to_scft_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # integration_to_ran_sat_offered_summary = integration_to_ran_sat_offered_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # integration_to_ran_pat_offered_summary = integration_to_ran_pat_offered_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        graph_summary = graph_summary.applymap(lambda x: str(x) if pd.notna(x) else "")
        # integration_to_onair_table_summary = integration_to_onair_table_summary.applymap(lambda x: str(x) if pd.notna(x) else "") 
        table_summary = table_summary.applymap(lambda x: str(x) if pd.notna(x) else "")      


        # result_json = integration_to_scft_summary.to_dict(orient="records")
        # json_data1 = json.dumps(result_json)
        
        # result_json = integration_to_ran_sat_offered_summary.to_dict(orient="records")
        # json_data2 = json.dumps(result_json)
        
        # result_json = integration_to_ran_pat_offered_summary.to_dict(orient="records")
        # json_data3 = json.dumps(result_json)

        # result_json = rfai_to_integration_summary.to_dict(orient="records")
        # json_data4 = json.dumps(result_json)
        
        # result_json = rfai_to_site_onair_summary.to_dict(orient="records")
        # json_data5 = json.dumps(result_json)

        result_json = graph_summary.to_dict(orient="records")
        json_data8 = json.dumps(result_json)

        # result_json = integration_to_onair_table_summary.to_dict(orient="records")
        # json_data6 = json.dumps(result_json)

        # result_json = rfai_to_integration_table_summary.to_dict(orient="records")
        # json_data7 = json.dumps(result_json)

        result_json = table_summary.to_dict(orient="records")
        json_data9 = json.dumps(result_json)

        json_data = {
            # "integration_to_scft_graph": json_data1,
            # "integration_to_sat_graph": json_data2,
            # "integration_to_pat_graph": json_data3,
            # "rfai_to_integration_graph": json_data4,
            # "rfai_to_site_onair_graph": json_data5,
            "graph_summary": json_data8,
            # "integration_to_onair_table": json_data6,
            # "rfai_to_integration_table": json_data7
            "table_summary": json_data9
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
 

# @api_view(['GET', 'POST'])
# def ms2_daily_waterfall(request):
#     circle = request.data.get('circle', [])
#     site_tagging = request.data.get('site_tagging', [])
#     current_status = request.data.get('relocation_method', [])
#     new_toco_name = request.data.get('new_toco_name', [])
#     start_date = request.data.get('from_date')
#     end_date = request.data.get('to_date')
#     view = request.data.get('view')

#     all_unique_circles = list(
#         AlokTrackerModel.objects.exclude(circle__isnull=True)
#         .distinct("circle")
#         .values_list("circle", flat=True)
#     )

#     all_unique_site_tagging = list(
#         AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
#         .distinct("site_tagging")
#         .values_list("site_tagging", flat=True)
#     )

#     all_unique_current_status = list(
#         AlokTrackerModel.objects.exclude(current_status__isnull=True)
#         .distinct("current_status")
#         .values_list("current_status", flat=True)
#     )

#     all_unique_new_toco_name = list(
#         AlokTrackerModel.objects.exclude(new_toco_name__isnull=True)
#         .distinct("new_toco_name")
#         .values_list("new_toco_name", flat=True)
#     )

#     # ✅ Add "ALL" at the top of each list
#     unique_data = {
#         "unique_circle": sorted(all_unique_circles),
#         "unique_site_tagging": sorted(all_unique_site_tagging),
#         "unique_relocation_method": sorted(all_unique_current_status),
#         "unique_new_toco_name": sorted(all_unique_new_toco_name),
#     }

#     # Default to 'ALL' if not provided
#     circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
#     site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
#     current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
#     new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    
#     start_date = dtime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
#     end_date = dtime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

#     try:
#         ############################################################## 🔹 Dynamic filters #########################################################
#         filters = {}
#         if "ALL" not in circle:
#             filters["circle__in"] = circle
#         if "ALL" not in site_tagging:
#             filters["site_tagging__in"] = site_tagging
#         if "ALL" not in current_status:
#             filters["current_status__in"] = current_status
#         if "ALL" not in new_toco_name:
#             filters["new_toco_name__in"] = new_toco_name

#         ############################################################## 🔹 Fetch data ###############################################################
#         obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
#         df = pd.DataFrame(obj.values())

#         if df.empty:
#             return Response({'error': 'No data found for given filters'}, status=404)

#         ############################################################### 🔹 Convert all date columns to datetime ####################################
#         for col in df.columns:
#             if "Date" in col:
#                 df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")

#         ################################################################### 🔹 Determine start_date and end_date for 26 → 25 cycle #####################
#         today = dtime.today().date()

#         if not start_date or not end_date:
#             if today.day >= 26:
#                 start_date = today.replace(day=26)
#                 if today.month == 12:
#                     end_date = today.replace(year=today.year + 1, month=1, day=25)
#                 else:
#                     end_date = today.replace(month=today.month + 1, day=25)
#             else:
#                 if today.month == 1:
#                     start_date = today.replace(year=today.year - 1, month=12, day=26)
#                 else:
#                     start_date = today.replace(month=today.month - 1, day=26)
#                 end_date = today.replace(day=25)

#         ############################################################## 🔹 Generate date range and formatted column headers ###########################
        
#         date_range = pd.date_range(start=start_date, end=min(end_date, today)).date

#         formatted_dates = [d.strftime("%d-%b-%y") for d in date_range]
#         result_columns = ["Milestone Track/Site Count", "CF"] + formatted_dates
#         result = pd.DataFrame(columns=result_columns)

#         ###################################################### 🔹 Milestone list #############################################################
#         milestones = [
#             "Site ONAIR Date",
#             "I-Deploy ONAIR Date",
#             'RAN PAT Accepted Date',
#             'RAN SAT Accepted Date',
#             'MW PAT Accepted Date',
#             'MW SAT Accepted Date',
#             'SCFT Accepted Date',
#             'KPI AT offer Date',
#             'KPI AT Accepted Date',
#             '4G MS2 Date',
#             '5G MS2 Date',
#             'Final MS2 Date'
#         ]
        
#         unique_data.update(**{"Milestone": milestones})
 
#         # #####################################################🔹 Loop through milestones and count per date ######################################
 
#         for milestone in milestones:
#             milestone_df_format = (
#                 milestone.lower()
#                 .replace(" ", "_")
#                 .replace("-", "_")
#                 .replace("(", "")
#                 .replace(")", "")
#                 .replace("/", "_")
#             )
 
#             if milestone_df_format not in df.columns:
#                 continue
 
#             df[milestone_df_format] = pd.to_datetime(df[milestone_df_format], errors="coerce").dt.date
 
#             valid_dates = df[milestone_df_format].dropna()

#             if valid_dates.empty:
#                 row = {"Milestone Track/Site Count": milestone, "CF": "-", **{d.strftime("%d-%b-%y"): "-" for d in date_range}}
#                 result.loc[len(result)] = row
#                 continue
 
#             cf_count = (valid_dates < start_date).sum()
 
#             cumulative = cf_count
#             row = {"Milestone Track/Site Count": milestone, "CF": cf_count}
#             for date in date_range:
#                 # if(date >= today):
#                 #     row[date.strftime("%d-%b-%y")] = 0
#                 #     continue
#                 count = (valid_dates == date).sum()
#                 cumulative += count
#                 if view == "Cumulative":
#                     row[date.strftime("%d-%b-%y")] = cumulative
#                 else:
#                     row[date.strftime("%d-%b-%y")] = count
 
#             result.loc[len(result)] = row
 
#         ######################################################## 🔹 Convert and format ##############################################
#         # result = result.astype(str).reset_index(drop=True)
 
#         result.columns = [
#             col.strftime("%d-%b-%y") if isinstance(col, (dtime,)) or hasattr(col, "strftime") else col
#             for col in result.columns
#         ]
 
#         # for date in formatted_dates:
#         #     result[date] = result[date].astype(int)
#         ############################################# 🔹 Arrange columns #####################################
#         result = result[["Milestone Track/Site Count", "CF"] + formatted_dates]
        
#         last_col = formatted_dates[-1]
#         dash_mask = result[last_col] == '-'
        
#         # 2️⃣ Convert to numeric where possible
#         result[last_col] = pd.to_numeric(result[last_col], errors='coerce')
        
#         # 3️⃣ Compute difference (previous - current)
#         result['Gap'] = -result[last_col].diff()
        
#         nan_mask = result['Gap'].isna()
        
#         result[last_col] = result[last_col].fillna(0).astype(int)
#         result['Gap'] = result['Gap'].fillna(0).astype(int)
        
#         result.loc[dash_mask, last_col] = '-'
#         result.loc[nan_mask, 'Gap'] = '-'
        
#         result['Gap'] = result['Gap'].astype(str)
 
        
#         result = result.astype(str).reset_index(drop=True)
        
#         result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(lambda col: col.replace(" Date", "") if " Date" in col else col)
        
#         ############################################################################################################################
#         new_result = result.copy()
        
#         for i, date in enumerate(formatted_dates, start=1):
#             new_result.rename(columns={date : f'date_{i}'}, inplace=True)
#             # new_result[f"date_{i}"] = new_result[f"date_{i}"].astype(str)
        
        
#         result_json = new_result.to_dict(orient="records")
#         json_data = json.dumps(result_json)


#         current_date = dtime.now().strftime("%Y-%m-%d")
#         current_time = dtime.now().strftime("%H-%M-%S")
 
#         BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
#         os.makedirs(BASE_URL, exist_ok=True)
#         output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
#         shutil.rmtree(output_folder, ignore_errors=True)
#         os.makedirs(output_folder, exist_ok=True)
       
#         dashboard_file_path = os.path.join(output_folder, f"MS2_DAILY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx")
        
#         with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
#             result.to_excel(writer, index=False, sheet_name='Daily Waterfall MS2')
        
#         dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
#         download_link = request.build_absolute_uri(dashboard_file_path)
    
#         return Response({'message': 'Dashboard created successfully !!!', "download_link": download_link, "data": json_data, "dates": formatted_dates, "unique_data": unique_data}, status=200)
#     except Exception as e:
#         return Response({"error": f"{str(e)}"},status=500)


@api_view(['GET', 'POST'])
def ms2_daily_waterfall(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    start_date = request.data.get('from_date')
    end_date = request.data.get('to_date')
    view = request.data.get('view')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None

    all_unique_circles = list(
        AlokTrackerModel.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        AlokTrackerModel.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )

    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]

    start_date = dtime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    end_date = dtime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

    try:
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name

        obj = AlokTrackerModel.objects.filter(**filters)
        df = pd.DataFrame(obj.values())

        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)

        for col in df.columns:
            if "Date" in col or "date" in col:
                df[col] = pd.to_datetime(df[col], errors="coerce")

        today = dtime.today().date()

        if (not month_start or not month_end) and (not start_date or not end_date):
            if today.day >= 26:
                start_date = today.replace(day=26)
                end_date = (today.replace(month=today.month + 1, day=25)
                            if today.month != 12
                            else today.replace(year=today.year + 1, month=1, day=25))
            else:
                start_date = (today.replace(month=today.month - 1, day=26)
                              if today.month != 1
                              else today.replace(year=today.year - 1, month=12, day=26))
                end_date = today.replace(day=25)

        if month_start and month_end:
            date_range = pd.date_range(start=month_start, end=min(month_end, today)).date
        else:
            date_range = pd.date_range(start=start_date, end=min(end_date, today)).date

        formatted_dates = [d.strftime("%d-%b-%y") for d in date_range]

        if month_start and month_end:
            result_columns = ["Milestone Track/Site Count", "AOP", "CF"] + formatted_dates
        else:
            result_columns = ["Milestone Track/Site Count", "AOP"] + formatted_dates

        result = pd.DataFrame(columns=result_columns)

        milestones = [
            "Site ONAIR Date",
            "I-Deploy ONAIR Date",
            'RAN PAT Accepted Date',
            'RAN SAT Accepted Date',
            'MW PAT Accepted Date',
            'MW SAT Accepted Date',
            'SCFT Accepted Date',
            'KPI AT offer Date',
            'KPI AT Accepted Date',
            'MS2 Status'
        ]

        unique_data.update({"Milestone": milestones})

        for milestone in milestones:

            if milestone == "MS2 Status":
                milestone_df_format = "final_ms2_date"
            else:
                milestone_df_format = (
                    milestone.lower()
                    .replace(" ", "_")
                    .replace("-", "_")
                    .replace("(", "")
                    .replace(")", "")
                    .replace("4", "four_")
                    .replace("5", "five_")
                )

            if milestone_df_format not in df.columns:
                continue

            df[milestone_df_format] = pd.to_datetime(
                df[milestone_df_format], errors="coerce"
            ).dt.date

            valid_dates = df[milestone_df_format].dropna()

            if valid_dates.empty:
                row = {
                    "Milestone Track/Site Count": milestone,
                    "AOP": "-"
                }
                if month_start and month_end:
                    row["CF"] = "-"

                for d in date_range:
                    row[d.strftime("%d-%b-%y")] = "-"

                result.loc[len(result)] = row
                continue

            if month_start and month_end:
                aop_count = (valid_dates < month_start).sum()
                cumulative = aop_count
                row = {
                    "Milestone Track/Site Count": milestone,
                    "AOP": aop_count,
                    "CF": aop_count
                }
            else:
                aop_count = (valid_dates < start_date).sum()
                cumulative = aop_count
                row = {
                    "Milestone Track/Site Count": milestone,
                    "AOP": aop_count
                }

            for d in date_range:
                count = (valid_dates == d).sum()
                cumulative += count

                if view == "Cumulative":
                    row[d.strftime("%d-%b-%y")] = cumulative
                else:
                    row[d.strftime("%d-%b-%y")] = count

            result.loc[len(result)] = row

        # -------- GAP CALCULATION FROM SITE ONAIR --------

        last_col = formatted_dates[-1]

        result[last_col] = pd.to_numeric(result[last_col], errors='coerce')

        site_onair_value = result.loc[
            result['Milestone Track/Site Count'] == 'Site ONAIR Date',
            last_col
        ]

        if not site_onair_value.empty:
            site_onair_value = site_onair_value.values[0]
            result['Gap'] = site_onair_value - result[last_col]
        else:
            result['Gap'] = 0

        result['Gap'] = result['Gap'].fillna(0).astype(int)
        result[last_col] = result[last_col].fillna(0).astype(int)

        result = result.astype(str).reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(
            lambda col: col.replace(" Date", "") if " Date" in col else col
        )

        new_result = result.copy()

        for i, d in enumerate(formatted_dates, start=1):
            new_result.rename(columns={d: f'date_{i}'}, inplace=True)

        result_json = new_result.to_dict(orient="records")
        json_data = json.dumps(result_json)

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)

        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)

        dashboard_file_path = os.path.join(
            output_folder,
            f"RELOCATION_TRACKING_MS2_DAILY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx"
        )

        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Daily Waterfall MS2')

        dashboard_file_path = dashboard_file_path.replace(
            settings.MEDIA_ROOT, settings.MEDIA_URL
        ).replace("\\", "/")

        download_link = request.build_absolute_uri(dashboard_file_path)

        return Response({
            'message': 'Dashboard created successfully !!!',
            "download_link": download_link,
            "data": json_data,
            "dates": formatted_dates,
            "unique_data": unique_data
        }, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)

@api_view(['GET','POST'])
def ms2_weekly_monthly_waterfall(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    year2 = request.data.get('year2')
    view = request.data.get('view')
 
    # Default to 'ALL' if not provided
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    
    if year2:
        year2 = int(year2)
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None
   
    all_unique_circles = list(
        AlokTrackerModel.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        AlokTrackerModel.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )

    # ✅ Add "ALL" at the top of each list
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    try:
        # 🔹 Dynamic filters
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name
 
        # 🔹 Fetch data
        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
       
        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")
               
        today = dtime.today().date()
 
        if year2:
            fy_start = dtime(year2, 3, 26).date()   # 26-Apr current year start
            fy_end = dtime(year2 + 1, 3, 25).date() # 25-Mar next year end
        else:
            if today.month >= 4:
                fy_start = dtime(today.year, 3, 26).date()   # 26-Apr current year start
                fy_end = dtime(today.year + 1, 3, 25).date() # 25-Mar next year end
            else:
                fy_start = dtime(today.year - 1, 3, 26).date()
                fy_end = dtime(today.year, 3, 25).date()
                
        print("2.5-------------------------------------------------------------")
 
        # 🧩 2. Determine current cycle (26th prev → 25th curr)
        if today.day >= 26:
            if today.month == 12:
                current_cycle_start = dtime(today.year, 12, 26).date()
                current_cycle_end = dtime(today.year + 1, 1, 25).date()
            else:
                current_cycle_start = dtime(today.year, today.month, 26).date()
                current_cycle_end = dtime(today.year, today.month + 1, 25).date()
        else:
            if today.month == 1:
                current_cycle_start = dtime(today.year - 1, 12, 26).date()
            else:
                current_cycle_start = dtime(today.year, today.month - 1, 26).date()
            current_cycle_end = dtime(today.year, today.month, 25).date()
            
        if year2:
            current_cycle_start = min(dtime(year2+1, 2, 26).date(), current_cycle_start)
            current_cycle_end = min(dtime(year2+1, 3, 25).date(), current_cycle_end)
 
        # 🗓️ 3. Create monthly periods (Apr → month before current)
        months = []
        start = fy_start
        while start <= current_cycle_start:
            if start.month == 12:
                end = dtime(start.year + 1, 1, 25).date()
            else:
                end = dtime(start.year, start.month + 1, 25).date()
            months.append((start, end))
            start = end + timedelta(days=1)
 
        # 🗓️ 4. Create weekly periods for current month
        weeks = []
        
        if month_start and month_end:
            current_cycle_start = month_start
            current_cycle_end = month_end
        
        week_start = current_cycle_start
 
        while week_start <= current_cycle_end:
            week_end = week_start + timedelta(days=6)
            if week_end > current_cycle_end:
                week_end = current_cycle_end
            weeks.append((week_start, week_end))
            week_start = week_end + timedelta(days=1)
 
 
        # 📊 5. Prepare result DataFrame
        result = pd.DataFrame()
       
       
        # result.set_index("Milestone Track/Site Count", inplace=True)
 
        milestones = [
            "Site ONAIR Date",
            "I-Deploy ONAIR Date",
            'RAN PAT Accepted Date',
            'RAN SAT Accepted Date',
            'MW PAT Accepted Date',
            'MW SAT Accepted Date',
            'SCFT Accepted Date',
            'KPI AT offer Date',
            'KPI AT Accepted Date',
            '4G MS2 Date',
            '5G MS2 Date',
            'Final MS2 Date'
        ]
 
        data = []
        for milestone in milestones:
            
            milestone_df_format = milestone.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_").replace("4", "four_").replace("5", "five_")
            if milestone_df_format not in df.columns:
                continue
            
            milestone_data = pd.to_datetime(df[milestone_df_format], errors='coerce').dt.date
            row = {"Milestone Track/Site Count": milestone}
            if milestone_data.dropna().empty:
                row["CF"] = "-"
                # Set all month and week columns to "-"
                for _, end in months:
                    month_name = end.strftime("%b-%y")
                    row[month_name] = "-"
                for i, _ in enumerate(weeks, 1):
                    week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                    row[week_name] = "-"
                data.append(row)
                continue
 
 
            # CF → sites before financial year start
            row["CF"] = (milestone_data < fy_start).sum()
 
            cumulative_month = row["CF"]
            for start, end in months:
                month_name = end.strftime("%b-%y")
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_month += count
                if view == 'Cumulative':
                    row[month_name] = cumulative_month
                else:
                    row[month_name] = count
 
            cumulative_week = 0
            for i, (start, end) in enumerate(weeks, 1):
                week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                if(start >= today):
                    row[week_name] = 0
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_week += count
                if view == 'Cumulative':
                    row[week_name] = cumulative_week
                else:
                    row[week_name] = count
 
            data.append(row)
 

        result = pd.DataFrame(data)
 
        for col in result.columns:
            if col != "Milestone Track/Site Count":
                result[col] = result[col].astype(str)
 
        result = result.reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(lambda col: col.replace(" Date", "") if " Date" in col else col)
 
        new_data = result.copy()
        months_columns = [col for col in new_data.columns.to_list()[2:] if " W" not in col]
        months_data = new_data[new_data.columns.to_list()[:2] + months_columns].copy()
        
        weeks_columns = [col for col in new_data.columns.to_list()[2:] if " W" in col]
        
        week_data = new_data[new_data.columns.to_list()[:1] + weeks_columns].copy()
        
        months_data.rename(columns={
            col : f"Month-{i}" for i, col in enumerate(months_columns, start=1)
        }, inplace=True)
        
        week_data.rename(columns={
            col : f"Month_Week-{i}" for i, col in enumerate(weeks_columns, start=1)
        }, inplace=True)
        
        unique_data.update(**{"month_columns": months_columns, "week_columns" : weeks_columns})
        

        month_dict_data = months_data.to_dict(orient="records")
        month_json_data = json.dumps(month_dict_data)
        week_dict_data = week_data.to_dict(orient="records")
        week_json_data = json.dumps(week_dict_data)
        
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"MS2_WEEKLY_MONTHY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx")
       
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Monthly MS2')
       
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
   
        return Response({'message': 'Weekly and Monthly Dashboard created successfully !!!', "download_link": download_link, "unique_data": unique_data, "months_data": month_json_data, "week_data": week_json_data}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
   

@api_view(['GET', 'POST'])
def monthly_graph(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    view = request.data.get('view')
    type = request.data.get('type')
    year_filtered = request.data.get('year2')
    
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"
    milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_") + "_date"
    
    
    
    all_unique_circles = list(
        AlokTrackerModel.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        AlokTrackerModel.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )
    
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }
    
    try:
        year_filtered = int(year_filtered)
        
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name
        
            
            
        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        # print(df)
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        df = df[[milestone1_col, milestone2_col]]
       
        for col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
            
        # print(df)

        
        today = dtime.today().date()

        # # 1️⃣ Financial month (26th–25th logic)
        # df["month_num1"] = df[milestone1_col].dt.month + (df[milestone1_col].dt.day >= 26)
        # df["month_num1"] = df["month_num1"].replace({13: 1}).astype(int)

        # # 2️⃣ Financial year (FIXED)
        # df["year1"] = df[milestone1_col].dt.year - (
        #     (df[milestone1_col].dt.month < 3) |
        #     ((df[milestone1_col].dt.month == 3) & (df[milestone1_col].dt.day < 26))
        # )

        # df['month_num1'] = pd.to_numeric(df['month_num1'], errors='coerce')
        # df['year1'] = pd.to_numeric(df['year1'], errors='coerce')
        
        df["month_num1"] = (
            df[milestone1_col].dt.month + (df[milestone1_col].dt.day >= 26)
        ).replace({13: 1}).astype("Int64")

        df["year1"] = (
            df[milestone1_col].dt.year - (
                (df[milestone1_col].dt.month < 3) |
                ((df[milestone1_col].dt.month == 3) & (df[milestone1_col].dt.day < 26))
            )
        ).astype("Int64")

        # 3️⃣ Display year
        df["display_year"] = df["year1"] + (df["month_num1"] <= 3)

        # 4️⃣ Month label
        df["month_name1"] = pd.to_datetime(
            df["month_num1"].astype(str) + "-" + df["display_year"].astype(str),
            format="%m-%Y",
            errors="coerce"
        ).dt.strftime("%b-%y")

        # 5️⃣ Current FY
        current_fy = year_filtered

        # 6️⃣ CF logic
        df.loc[df["year1"] < current_fy, "month_name1"] = "CF"



        df['month_num2'] = df[milestone2_col].dt.month + (df[milestone2_col].dt.day >= 26)
        df['month_num2'] = df['month_num2'].replace({13: 1})

        # 2️⃣ Financial year (derived ONLY from original date)
        df['year2'] = df[milestone2_col].dt.year - (
            (df[milestone2_col].dt.month < 3) |
            ((df[milestone2_col].dt.month == 3) & (df[milestone2_col].dt.day < 26))
        )

        # 3️⃣ Safe numeric conversion (IMPORTANT)
        df['month_num2'] = pd.to_numeric(df['month_num2'], errors='coerce')
        df['year2'] = pd.to_numeric(df['year2'], errors='coerce')

        # 4️⃣ Display year
        df['display_year2'] = df['year2'] + (df['month_num2'] <= 3)

        # 5️⃣ Month label (NaT-safe)
        df['month_name2'] = pd.to_datetime(
            df['month_num2'].astype('Int64').astype(str) + '-' +
            df['display_year2'].astype('Int64').astype(str),
            format='%m-%Y',
            errors='coerce'
        ).dt.strftime('%b-%y')

        # 6️⃣ Current financial year
        current_fy = year_filtered

        # 7️⃣ Carry Forward (only where year2 exists)
        df.loc[df['year2'].notna() & (df['year2'] < current_fy), 'month_name2'] = 'CF'
        
        print(df['month_name1'].dropna().unique().tolist())

        
        print(df['year1'])
        
        print(df['month_name2'])
        
        print(df['year2'])
        
        df.loc[df['year1'].notna() & (df['year1'] > current_fy), ["month_name1"]] = "Future"
        df.loc[df['year2'].notna() & (df['year2'] > current_fy), ["month_name2"]] = "Future"

        def sort_financial_year(summary):
            # Extract sorted unique Month-Year (excluding CF)
            months = [
                m for m in summary['month_name'].unique()
                if m not in ["CF", "Future"] and pd.notna(m)
            ]
            
            # Convert to datetime to sort correctly
            months_sorted = sorted(
                months,
                key=lambda x: pd.to_datetime(x, format='%b-%y')
            )

            # CF should always be first
            ordered = []

            if "CF" in summary['month_name'].values:
                ordered.append("CF")

            ordered += months_sorted

            if "Future" in summary['month_name'].values:
                ordered.append("Future")
            
            # Convert into ordered categorical for final sort
            summary['month_name'] = pd.Categorical(summary['month_name'], ordered, ordered=True)

            return summary.sort_values('month_name')
        
        def generate_type1_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()

            # Start counts by month_name
            start_counts = (
                temp[temp[start_col].notna()]
                .groupby("month_name1")
                .size()
                .reset_index(name=f"{start_label} Done Count")
                .rename(columns={"month_name1": "month_name"})
            )

            # End counts by month_name2
            end_counts = (
                temp[temp[end_col].notna()]
                .groupby("month_name2")
                .size()
                .reset_index(name=f"{end_label} Done Count")
                .rename(columns={"month_name2": "month_name"})
            )

            # Merge both results
            summary = pd.merge(start_counts, end_counts, on="month_name", how="outer")

            # Convert to Int64 nullable type
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")
            summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].astype("Int64")

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            return summary.reset_index(drop=True)
        
        def generate_type2_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("month_name1").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    # f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
                .rename(columns={"month_name1": "month_name"})
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")

            summary = summary[
                ["month_name", f"{start_label} Done Count", f"{end_label} Done Count"]
            ]

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            return summary.reset_index(drop=True)
        
        summary = generate_type1_summary(df, milestone1, milestone2, milestone1_col, milestone2_col) if type == 'type1' else generate_type2_summary(df, milestone1, milestone2, milestone1_col, milestone2_col)

        # numeric_cols = [col for col in summary.columns if "Done Count" in col]
        # summary[numeric_cols] = summary[numeric_cols].fillna(0).astype(int)
        
        numeric_cols = [col for col in summary.columns if "Done Count" in col]

        summary[numeric_cols] = summary[numeric_cols].apply(pd.to_numeric, errors="coerce")
        summary[numeric_cols] = summary[numeric_cols].fillna(0).astype(int)

        
        result_json = summary.to_dict(orient = "records")
        json_data = json.dumps(result_json)
        
        return Response({"message" : "request processed successfully ! ", "json_data": json_data, "unique_data": unique_data}, status=200)
    
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
   
    
@api_view(['GET', 'POST'])
def lifecycle_display(request):
    userId = request.data.get('userId').lower()
    siteId = request.data.get('siteId')
    circle = request.data.get('circle')

    if not userId or not siteId or not circle:
        return Response({"error": "input not provided"}, status = 400)
    
    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    circles = user.circles
    
    if 'CENTRAL' not in circles and circle not in circles:
        return Response({"error" : f"Access to {circle} sites denied"}, status = 403)
    
    try:
        filter = {
            "circle": circle,
            "new_site_id": siteId
        }

        obj = AlokTrackerModel.objects.filter(**filter).first()
        
        if obj is None:
            return Response({"error": "Site not found!"}, status=500)
        else:
            df = pd.DataFrame([model_to_dict(obj)])

        filters = {
            "circle" : circle,
            "site_id" : siteId
        }
        
        issue_obj = RelocationIssue.objects.filter(**filters)
        issue_df = pd.DataFrame(issue_obj.values())

        required_cols = ["Circle", "Site ID", "Issue Owner", "milestone", "Issue Name", "Start Date", "Close Date", "status", "Duration", "Updated_by", "Updated_at", "Created_by", "Created_at"]
        for col in required_cols:
            if col not in issue_df.columns:
                issue_df[col] = None
        
        data_cols = [
            "Allocation",
            "RFAI",
            "RFAI Survey",
            "MO Punch",
            "Material Dispatch",
            "Material Delivered",
            "Installation End",
            "Integration",
            "EMF Submission",
            "Alarm Rectification Done",
            "SCFT I-Deploy Offered",
            "RAN PAT Offer",
            "RAN SAT Offer",
            "MW PAT Offer",
            "MW SAT Offer",
            "Site ONAIR",
            "I-Deploy ONAIR"
        ]
        
        data_cols_1 = [col.lower().replace("-", "_").replace(" ", "_").replace("/", "_") + "_date" for col in data_cols] + ['circle']
        
        df_data = df[data_cols_1]
        
        issue_counts = {}
        
        for col in data_cols:
            counts = {}
            counts['total'] = (issue_df['milestone'] == col).sum()
            counts['open'] = ((issue_df['milestone'] == col) & (issue_df['status'] == 'Open')).sum()
            counts['closed'] = ((issue_df['milestone'] == col) & (issue_df['status'] == 'Closed')).sum()
            issue_counts[col] = counts
        
        print("C")
        
        if not df_data.empty:
            df_data = df_data.applymap(
                lambda x: x.strftime("%d-%b-%y") if hasattr(x, 'strftime') else x
            )
            
        
        result_json = df_data.to_dict(orient="records")
        json_data = json.dumps(result_json)
        
        return Response({'message': 'request processed successfully !!!', "json_data" : json_data, "issue_counts" : issue_counts }, status = 200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)

    
@api_view(['POST'])
def issue_timeline_display(request):
    userId = request.data.get('userId')
    if not userId:
        return Response({"error": "userId required"}, status=400)

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    
    circles = user.circles
    
    siteId = request.data.get('siteId')
    circle = request.data.get('circle')
    owner = request.data.get('owner')
    milestone = request.data.get('milestone')

    if not all([siteId, circle]):
        return Response({"message": "Circle/Site-ID not provided"}, status=400)

    if 'CENTRAL' not in circles and circle not in circles:
        return Response({"error": f"Access to {circle} sites denied!"}, status=403)

    try:
        filters = {
            "circle": circle,
            "site_id": siteId
        }
        if owner:
            filters['issue_owner'] = owner
        if milestone:
            filters['milestone'] = milestone

        obj = RelocationIssue.objects.filter(**filters)

        df = pd.DataFrame(obj.values())

        if not obj.exists():
            # required_cols = ["Circle", "Site ID", "Issue Owner", "Milestone", "Issue Name", "Start Date", "Close Date", "Status", "Duration", "Updated_by", "Updated_at", "Created_by", "Created_at"]
            # for col in required_cols:
            #     if col not in df.columns:
            #         df[col] = None
            json_data = df.to_dict(orient="records")
            json_data = json.dumps(json_data)
            return Response({"error": "Data not found!", "json_data": json_data}, status=200)

        df["start_date"] = df["start_date"].apply(
            lambda x: x.strftime("%d-%b-%y") if pd.notnull(x) else "-"
        )

        df["close_date"] = df["close_date"].apply(
            lambda x: x.strftime("%d-%b-%y") if pd.notnull(x) else "-"
        )
        
        rename_map = {
            "circle": "Circle",
            "site_id": "Site ID",
            "issue_owner": "Issue Owner",
            "milestone": "Milestone",
            "issue_name": "Issue Name",
            "start_date": "Start Date",
            "close_date": "Close Date",
            "status": "Status",
            "duration": "Duration",
            "remarks": "Remarks",
            "updated_by": "Updated_by",
            "updated_at": "Updated_at",
            "created_by": "Created_by",
            "created_at": "Created_at"
        }
        
        df = df.rename(columns=rename_map)

        json_data = df.to_dict(orient="records")

        return Response({"message": "Request processed successfully!", "json_data": json_data}, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
 
@api_view(['POST'])
def issue_timeline_add(request):
    userId = request.data.get('userId')
    if not userId:
        return Response({"error": "userId required"}, status=400)

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    if user.right == 'Read':
        return Response({"error": "Access denied!"}, status=status.HTTP_403_FORBIDDEN)

    siteId = request.data.get('siteId')
    circle = request.data.get('circle')
    owner = request.data.get('owner')
    milestone = request.data.get('milestone')
    issue = request.data.get('issue')
    start_date_str = request.data.get('start_date')   
    close_date_str = request.data.get('close_date')  
    remarks = request.data.get('remarks') 

    required_fields = [siteId, circle, owner, milestone, issue, start_date_str]
    if not all(required_fields):
        return Response({"message": "Required data not provided"}, status=status.HTTP_400_BAD_REQUEST)
    
        

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()

        close_date = None
        if close_date_str not in ["", None, "-", "nan", "undefined"]:
            close_date = datetime.strptime(close_date_str, "%Y-%m-%d").date()

    except Exception:
        return Response({"error": "Invalid date format. Expected yyyy-mm-dd"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        filters = {
            "circle": circle,
            "site_id": siteId,
            "issue_owner": owner,
            "milestone": milestone,
            "issue_name": issue,
            "status": "Open"
        }

        if RelocationIssue.objects.filter(**filters).exists():
            return Response({"error": "Open issue already exists"}, status=409)

        today = date.today()

        if close_date:
            duration = (close_date - start_date).days
        else:
            duration = (today - start_date).days

        new_issue = RelocationIssue.objects.create(
            circle=circle,
            site_id=siteId,
            issue_owner=owner,
            milestone=milestone,
            issue_name=issue,
            start_date=start_date,
            close_date=close_date,
            status="Closed" if close_date else "Open",
            duration=duration,
            remarks=remarks,
            updated_by=userId,
            created_by=userId,
        )
        
        update_ageing_new(circle, siteId)

        return Response(
            {
                "message": "Issue added successfully!",
                "json_data":{"id": new_issue.id,
                "circle": circle,
                "site": siteId,
                "milestone": milestone,
                "issue": issue,
                "start_date": start_date.strftime("%d-%b-%y"),
                "close_date": close_date.strftime("%d-%b-%y") if close_date else "-",
                "status": new_issue.status,
                "duration_days": duration,
                "remarks": remarks}
            },
            status=200
        )

    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
def issue_timeline_update(request):
    userId = request.data.get('userId')
    issue_id = request.data.get('id') 
    circle = request.data.get('circle')

    if not userId or not issue_id or not circle:
        return Response({"error": "Required data not provided"}, status=400)

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    if user.right == 'Read':
        return Response({"error": "Access denied!"}, status=status.HTTP_403_FORBIDDEN)

    # user_circle = ACCESS_RIGHTS[userId]["Circle"]
    # if user_circle != "CENTRAL" and user_circle != circle:
    #     return Response({"error": "Access denied!"}, status=403)

    try:
        issue_obj = RelocationIssue.objects.get(id=issue_id)
    except RelocationIssue.DoesNotExist:
        return Response({"error": "Issue not found"}, status=404)
    
    if issue_obj.status == 'Closed':
        return Response({"error": "Issue already closed"}, status=404)

    start_date_str = request.data.get("start_date")
    close_date_str = request.data.get("close_date")
    new_owner = request.data.get("owner")
    new_milestone = request.data.get("milestone")
    new_issue_name = request.data.get("issue")
    remarks = request.data.get('remarks')

    try:
        if start_date_str:
            issue_obj.start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()

        if close_date_str in ["", None, "-", "nan", "undefined"]:
            issue_obj.close_date = None
        else:
            issue_obj.close_date = datetime.strptime(close_date_str, "%Y-%m-%d").date()

    except Exception:
        return Response({"error": "Invalid date format. Expected yyyy-mm-dd"}, status=400)

    try:
        
        if new_owner:
            issue_obj.issue_owner = new_owner

        if new_milestone:
            issue_obj.milestone = new_milestone

        if new_issue_name:
            issue_obj.issue_name = new_issue_name
            
        if remarks:
            issue_obj.remarks = remarks

        if issue_obj.close_date:
            issue_obj.status = "Closed"
        else:
            issue_obj.status = "Open"

        today = date.today()

        if issue_obj.close_date:
            issue_obj.duration = (issue_obj.close_date - issue_obj.start_date).days
        else:
            issue_obj.duration = (today - issue_obj.start_date).days

        issue_obj.updated_by = userId

        issue_obj.save()

        out_start = issue_obj.start_date.strftime("%d-%b-%y")
        out_close = issue_obj.close_date.strftime("%d-%b-%y") if issue_obj.close_date else "-"

        response_data = {
            "id": issue_obj.id,
            "circle": issue_obj.circle,
            "site_id": issue_obj.site_id,
            "issue_owner": issue_obj.issue_owner,
            "milestone": issue_obj.milestone,
            "issue_name": issue_obj.issue_name,
            "start_date": out_start,
            "close_date": out_close,
            "status": issue_obj.status,
            "duration": issue_obj.duration,
            "remarks": remarks,
            "updated_by": userId
        }
        
        update_ageing_new(circle, issue_obj.site_id)

        return Response(
            {"message": "Issue updated successfully!", "json_data": response_data},
            status=200
        )
        
    except Exception as e:
        return Response({"Error: " : f"{str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def issue_timeline_delete(request):
    userId = request.data.get('userId')
    issue_id = request.data.get('id')
    circle = request.data.get('circle')

    if not userId or not issue_id or not circle:
        return Response(
            {"error": "userId, id and circle are required"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        user = RelocationUser.objects.filter(email=userId.lower()).first()
    except RelocationUser.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    if user.right == 'Read':
        return Response({"error": "Access denied!"}, status=status.HTTP_403_FORBIDDEN)

    # Access control
    # user_circle = ACCESS_RIGHTS[userId]["Circle"]
    # if user_circle != "CENTRAL" and user_circle != circle:
    #     return Response({"error": "Access denied!"}, status=status.HTTP_403_FORBIDDEN)

    try:
        issue_obj = RelocationIssue.objects.get(id=issue_id)
    except RelocationIssue.DoesNotExist:
        return Response({"error": "Issue not found"}, status=status.HTTP_404_NOT_FOUND)

    # Optional safety check (recommended)
    # if issue_obj.status == "Closed":
    #     return Response(
    #         {"error": "Closed issues cannot be deleted"},
    #         status=status.HTTP_400_BAD_REQUEST
    #     )
    
    update_ageing_new(circle, issue_obj.site_id)

    issue_obj.delete()

    return Response(
        {
            "message": "Issue deleted successfully!",
            "id": issue_id
        },
        status=status.HTTP_200_OK
    )


@api_view(['DELETE'])
def delete_dropped_sites(request):

    qs = AlokTrackerModel.objects.filter(current_status="Site Dropped")

    if not qs.exists():
        return Response(
            {"message": "No records found with current_status = 'Site Dropped'"},
            status=status.HTTP_404_NOT_FOUND
        )

    deleted_count, _ = qs.delete()

    return Response(
        {"message": f"{deleted_count} records deleted successfully"},
        status=status.HTTP_200_OK
    )


@api_view(['DELETE'])
def delete_sites_by_detailed_remarks(request):

    qs = AlokTrackerModel.objects.filter(detailed_remarks="database correction need to remove from tracker")

    if not qs.exists():
        return Response(
            {"message": "No records found with detailed_remarks = 'database correction need to remove from tracker'"},
            status=status.HTTP_404_NOT_FOUND
        )

    deleted_count, _ = qs.delete()

    return Response(
        {"message": f"{deleted_count} records deleted successfully"},
        status=status.HTTP_200_OK
    )
    


@api_view(['DELETE'])
def delete_specific_site(request):
    
    circle = request.data.get("circle")
    site_id = request.data.get("siteId")

    qs = AlokTrackerModel.objects.filter(circle=circle, new_site_id=site_id)

    if not qs.exists():
        return Response(
            {"message": "No record found"},
            status=status.HTTP_404_NOT_FOUND
        )

    deleted_count, _ = qs.delete()

    return Response(
        {"message": f"{deleted_count} records deleted successfully"},
        status=status.HTTP_200_OK
    )


@api_view(['POST'])
def upload_ftr_dump(request):
    userId = request.data.get('userId')
    file = request.data.get('file')

    if not userId:
        return Response({"error": "userId not provided"}, status=status.HTTP_400_BAD_REQUEST)

    if not file:
        return Response({"error": "file not provided"}, status=status.HTTP_400_BAD_REQUEST)

    user = RelocationUser.objects.filter(email=userId.lower()).first()
    if not user:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

    if user.right == 'Read':
        return Response({"error": "ACCESS DENIED"}, status=status.HTTP_403_FORBIDDEN)

    try:
        # 1️⃣ Read file
        df = (
            pd.read_csv(file, sheet_name="Sheet1")
            if file.name.endswith(".csv")
            else pd.read_excel(file, sheet_name="Sheet1")
        )

        df.columns = df.columns.str.strip()

        # 2️⃣ Validate columns
        missing_cols = [col for col in DUMP_COLUMNS if col not in df.columns]
        if missing_cols:
            return Response(
                {"error": "Missing required columns", "details": missing_cols},
                status=status.HTTP_400_BAD_REQUEST,
            )

        df = df[DUMP_COLUMNS]

        # 3️⃣ Clean values
        df = df.where(pd.notnull(df), None)
        df.replace(["", "NaN", "nan"], None, inplace=True)
        
        # df = df.where(pd.notnull(df), None)
        # df.fillna("", inplace=True)
        # df = df.replace("NaN", None)
        # df = df.replace("", None)


        df['On Air Date'] = pd.to_datetime(df['On Air Date'], errors="coerce")

        df['On Air Date'] = df['On Air Date'].apply(
            lambda x: x.date() if pd.notna(x) else None
        )
        
        
        objects = []
        for _, row in df.iterrows():
            objects.append(
                ATDumpData(
                    circle=row["Circle"],
                    site_id=row["Site ID"],
                    on_air_date=row["On Air Date"],
                    
                    band = row['Band'],

                    physical_at_status=row["Physical AT Status"],
                    physical_at_rejection_counter=row["Physical AT Rejection Counter"],

                    performance_at_status=row["Performance AT Status"],
                    performance_at_rejection_counter=row["Performance AT Rejection Counter"],

                    soft_at_status=row["Soft AT Status"],
                    soft_at_rejection_counter=row["Soft AT Rejection Counter"],

                    scft_at_status=row["SCFT AT Status"],
                    scft_at_rejection_counter=row["SCFT AT Rejection Counter"],
                )
            )

        # 5️⃣ DELETE + INSERT atomically
        with transaction.atomic():
            ATDumpData.objects.all().delete()
            ATDumpData.objects.bulk_create(objects, batch_size=1000)

        return Response(
            {
                "message": "Dump replaced successfully",
                "rows_inserted": len(objects),
            },
            status=status.HTTP_200_OK,
        )

    except Exception as e:
        return Response(
            {"error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(['POST'])
def ftr_table(request):
    year = request.data.get("year")  # e.g. "2025"
    type = request.data.get("type")

    if not year or not type:
        return Response({"error": "type and year are required"}, status=400)
    
    try:
        year = int(year)

        fy_start = pd.Timestamp(year=year, month=3, day=29)
        fy_end   = pd.Timestamp(year=year + 1, month=3, day=28)
        
        
        # 1️⃣ Fetch data
        qs = ATDumpData.objects.all().values(
            'site_id',
            'on_air_date',
            'physical_at_status',
            'physical_at_rejection_counter',
            'performance_at_status',
            'performance_at_rejection_counter',
            'soft_at_status',
            'soft_at_rejection_counter',
            'scft_at_status',
            'scft_at_rejection_counter',
        )

        df = pd.DataFrame(qs)

        if df.empty:
            return Response({"message": "No data found"}, status=200)
        
        if type == "Relocation":
            relocation_sites = (
                AlokTrackerModel.objects
                .values_list('new_site_id', flat=True)
                .distinct()
            )
            df = df[df['site_id'].isin(relocation_sites)]
        elif type != 'Overall':
            relocation_sites = (
                AlokTrackerModel.objects
                .values_list('new_site_id', flat=True)
                .distinct()
            )
            df = df[~df['site_id'].isin(relocation_sites)]

        # 2️⃣ Date handling
        df['on_air_date'] = pd.to_datetime(df['on_air_date'], errors='coerce')
        
        df = df[
            (df['on_air_date'] >= fy_start) &
            (df['on_air_date'] <= fy_end)
        ]
        
        df['FTR_Month'] = df['on_air_date'].apply(get_ftr_month)

        df = df[df['FTR_Month'].notna()]

        # 3️⃣ Total sites per month
        total = (
            df.groupby('FTR_Month')['site_id']
              .count()
              .rename('Total')
        )

        # 4️⃣ FTR counts
        phy_ftr  = ftr_count(df, 'physical_at_status', 'physical_at_rejection_counter')
        soft_ftr = ftr_count(df, 'soft_at_status', 'soft_at_rejection_counter')
        scft_ftr = ftr_count(df, 'scft_at_status', 'scft_at_rejection_counter')
        perf_ftr = ftr_count(df, 'performance_at_status', 'performance_at_rejection_counter')

        # 5️⃣ Combine
        report = pd.concat(
            [
                total,
                phy_ftr.rename('Phy AT FTR Count'),
                soft_ftr.rename('Soft AT FTR Count'),
                scft_ftr.rename('SCFT AT FTR Count'),
                perf_ftr.rename('Perf AT FTR Count')
            ],
            axis=1
        ).fillna(0)

        # 6️⃣ Percentages
        report['Phy AT FTR %'] = (report['Phy AT FTR Count'] / report['Total'] * 100).round().astype(int)
        report['Soft AT FTR %'] = (report['Soft AT FTR Count'] / report['Total'] * 100).round().astype(int)
        report['SCFT AT FTR %'] = (report['SCFT AT FTR Count'] / report['Total'] * 100).round().astype(int)
        report['Perf AT FTR %'] = (report['Perf AT FTR Count'] / report['Total'] * 100).round().astype(int)
        
        report = report.sort_index(
            key=lambda idx: idx.map(month_sort_key)
        )
        
        # Sum counts
        overall_total = report['Total'].sum()

        overall_phy  = report['Phy AT FTR Count'].sum()
        overall_soft = report['Soft AT FTR Count'].sum()
        overall_scft = report['SCFT AT FTR Count'].sum()
        overall_perf = report['Perf AT FTR Count'].sum()

        # Calculate percentages
        overall = {
            'Total': overall_total,
            'Phy AT FTR Count': overall_phy,
            'Phy AT FTR %': round(overall_phy / overall_total * 100) if overall_total else 0,
            'Soft AT FTR Count': overall_soft,
            'Soft AT FTR %': round(overall_soft / overall_total * 100) if overall_total else 0,
            'SCFT AT FTR Count': overall_scft,
            'SCFT AT FTR %': round(overall_scft / overall_total * 100) if overall_total else 0,
            'Perf AT FTR Count': overall_perf,
            'Perf AT FTR %': round(overall_perf / overall_total * 100) if overall_total else 0,
        }
        
        report.loc['Overall'] = overall

        # 7️⃣ Final shape (rows → columns like screenshot)
        final_table = report.T
        
        print(final_table)
        
        final_table = final_table.reset_index()
        final_table = final_table.rename(columns={"index": "FTR"})
        
        def format_percentage_rows(df):
            percent_rows = df['FTR'].str.contains('%', regex=False)

            month_cols = df.columns.drop('FTR')

            df.loc[percent_rows, month_cols] = (
                df.loc[percent_rows, month_cols]
                .applymap(lambda x: f"{x}%" if pd.notna(x) else x)
            )

            return df
        
        final_table = format_percentage_rows(final_table)

        return Response(
            {
                "data": final_table.to_dict(orient="records"),
                "months": list(final_table.columns)
            },
            status=200
        )

    except Exception as e:
        return Response(
            {"error": str(e)},
            status=500
        )


@api_view(['POST'])
def ftr_table_circlewise(request):
    try:
        year = request.data.get("year")
        type = request.data.get("type")
        if not year or not type:
            return Response({"error": "year/type are required"}, status=400)

        year = int(year)

        fy_start = pd.Timestamp(year=year, month=3, day=29)
        fy_end = pd.Timestamp(year=year + 1, month=3, day=28)

        qs = ATDumpData.objects.all().values(
            'circle',
            'site_id',
            'on_air_date',
            'soft_at_status',
            'soft_at_rejection_counter',
            'physical_at_status',
            'physical_at_rejection_counter',
        )

        df = pd.DataFrame(qs)

        if df.empty:
            return Response({"message": "No data found"}, status=200)
        
        if type == "Relocation":
            relocation_sites = (
                AlokTrackerModel.objects
                .values_list('new_site_id', flat=True)
                .distinct()
            )
            df = df[df['site_id'].isin(relocation_sites)]
        elif type != 'Overall':
            relocation_sites = (
                AlokTrackerModel.objects
                .values_list('new_site_id', flat=True)
                .distinct()
            )
            df = df[~df['site_id'].isin(relocation_sites)]

        df['on_air_date'] = pd.to_datetime(df['on_air_date'], errors='coerce')

        # Financial year filter
        df = df[
            (df['on_air_date'] >= fy_start) &
            (df['on_air_date'] <= fy_end)
        ]

        df['FTR_Month'] = df['on_air_date'].apply(get_ftr_month)
        df = df[df['FTR_Month'].notna()]

        # Build tables
        soft_table = build_circlewise_ftr(
            df,
            'soft_at_status',
            'soft_at_rejection_counter'
        )

        phy_table = build_circlewise_ftr(
            df,
            'physical_at_status',
            'physical_at_rejection_counter'
        )
        
        print(soft_table)
        
        print(phy_table)
        
        soft_table = soft_table.reset_index()
        soft_table = soft_table.rename(columns={"circle": "Circle"})
        phy_table = phy_table.reset_index()
        phy_table = phy_table.rename(columns={"circle": "Circle"})

        return Response(
            {
                "soft_at": soft_table.to_dict(orient="records"),
                "physical_at": phy_table.to_dict(orient="records"),
                "months": list(soft_table.columns)
            },
            status=200
        )

    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['GET', 'POST'])
def ms2_ageing_dashboard_table1(request):
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "I-Deploy ONAIR",
        'RAN PAT Accepted',
        'RAN SAT Accepted',
        'MW PAT Accepted',
        'MW SAT Accepted',
        'SCFT Accepted',
        'KPI AT offer',
        'KPI AT Accepted',
        '4G MS2',
        '5G MS2',
        'Final MS2'
    ]
    
    unique_data = {
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones
    }

    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        if site_tagging and "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if month_start and month_end:
            filters[f"{milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + '_date'}__range"] = (month_start, month_end)

        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        date_cols = [col.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + '_date' for col in milestones]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']
        
        
        def generate_table_summary(df, milestone1, milestone2):


            milestone_cols = [(col.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date") for col in milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]]

            temp = df.copy()

            circles = ["ALL"] + temp["circle"].unique().tolist()

            final_df = pd.DataFrame()

            milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
            milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"

            for circle in circles:
                if circle != "ALL":
                    temp  = df[df["circle"] == circle]

                done_row = {}
                pending_row = {}

                for col, label in zip(milestone_cols, milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]):

                    done_row[label] = (temp[col].notna() & temp[milestone1_col].notna() ).sum()

                    pending_row[label] = (temp[col].isna() & temp[milestone1_col].notna() ).sum()

                rows_to_add = pd.DataFrame([
                    {"Circle": circle, "Site Status": "Done", **done_row},
                    {"Circle": circle, "Site Status": "Pending", **pending_row},
                ])

                final_df = pd.concat([final_df, rows_to_add], ignore_index=True)

            for c in final_df.columns:
                if c != "Site Status" and c != "Circle":
                    final_df[c] = final_df[c].astype("Int64")

            return final_df

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        table_summary = generate_table_summary(df, start_label, end_label)
 
        table_summary = table_summary.applymap(lambda x: str(x) if pd.notna(x) else "")      


        result_json = table_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "table_summary": json_data1
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
  
  
@api_view(['GET', 'POST'])
def ms2_ageing_dashboard_table2(request):
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    breakpoint1 = request.data.get('breakpoint1')
    breakpoint2 = request.data.get('breakpoint2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    type = request.data.get('type')

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "I-Deploy ONAIR",
        'RAN PAT Accepted',
        'RAN SAT Accepted',
        'MW PAT Accepted',
        'MW SAT Accepted',
        'SCFT Accepted',
        'KPI AT offer',
        'KPI AT Accepted',
        '4G MS2',
        '5G MS2',
        'Final MS2'
    ]
    
    unique_data = {
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones" : milestones
    }
    

    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    
    try:

        milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
        milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"

        breakpoint1 = int(breakpoint1)
        breakpoint2 = int(breakpoint2)
        
        temp = breakpoint1
        breakpoint1 = min(breakpoint1, breakpoint2)
        breakpoint2 = max(temp, breakpoint2)

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None

        milestones = [
            "Site ONAIR",
            "I-Deploy ONAIR",
            'RAN PAT Accepted',
            'RAN SAT Accepted',
            'MW PAT Accepted',
            'MW SAT Accepted',
            'SCFT Accepted',
            'KPI AT offer',
            'KPI AT Accepted',
            '4G MS2',
            '5G MS2',
            'Final MS2'
        ]
        

        filters = {}
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
            
        if type == "type2":
            if month_start and month_end:
                filters[f"{milestone1_col}__range"] = (month_start, month_end)
                

        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        
        if type == "type1":
            if month_start and month_end:
                obj = obj.filter(
                    Q(**{f"{milestone1_col}__range": (month_start, month_end)}) |
                    Q(**{f"{milestone2_col}__range": (month_start, month_end)})
                )
        

        df = pd.DataFrame(obj.values())

        
        date_cols = [col.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + '_date' for col in milestones]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']

        if month_start is not None:
            month_start = pd.to_datetime(month_start)

        if month_end is not None:
            month_end = pd.to_datetime(month_end)

        def generate_done_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            temp = df.copy()
            
            temp["days_diff"] = (temp[end_col] - temp[start_col]).dt.days
            
            if type == "type2":
        
                summary = (
                    temp.groupby("Circle").apply(lambda g: pd.Series({
                        f"{start_label} Done Count": g[start_col].notna().sum(),
                        f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].notna(), "days_diff"].mean()
                    }))
                    .reset_index()
                )
                
                total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                total_row["Circle"] = "Total"

                overall_valid = temp.loc[
                    temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].notna(),
                    "days_diff"
                ]
                total_row["Average Days"] = overall_valid.mean()

            else:
                if month_start is not None and month_end is not None:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": (g[start_col].notna() & (g[start_col] >= month_start) & (g[start_col] <= month_end)).sum(),
                            f"{end_label} Done Count": (g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"<= {breakpoint1} days": (g["days_diff"].notna() & (g["days_diff"] <= breakpoint1) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": (g["days_diff"].notna() & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f">= {breakpoint2} days": (g["days_diff"].notna() & (g["days_diff"] >= breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna() & (temp[end_col] >= month_start) & (temp[end_col] <= month_end),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                else:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": g[start_col].notna().sum(),
                            f"{end_label} Done Count": g[end_col].notna().sum(),
                            f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & g[end_col].notna()).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna()).sum(),
                            f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & g[end_col].notna()).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna(), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna(),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                    
            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]

            return summary


        done_summary = generate_done_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        done_summary = done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        def generate_pending_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            
            today = pd.Timestamp.today()
            
            temp = df.copy()
            
            temp["days_diff"] = (today - temp[start_col]).dt.days

            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                    f"<= {breakpoint1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[start_col].notna())).sum(),
                    f"{breakpoint1+1}-{breakpoint2-1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[start_col].notna())).sum(),
                    f">= {breakpoint2} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[start_col].notna())).sum(),
                    "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].isna(), "days_diff"].mean()
                }))
                .reset_index()
            )

            total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
            total_row["Circle"] = "Total"

            overall_valid = temp.loc[
                temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].isna(),
                "days_diff"
            ]
            total_row["Average Days"] = overall_valid.mean()

            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Pending Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]
            
            return summary
        

        pending_summary = generate_pending_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        pending_summary = pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"AGEING_DASHBOARD_FILE_{site_tagging}_{current_date}_{current_time}.xlsx")

        result_json = pending_summary.to_dict(orient="records")
        json_data2 = json.dumps(result_json)

        result_json = done_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "done_summary": json_data1,
            "pending_summary": json_data2
        }
        
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            pending_summary.to_excel(writer, index=False, sheet_name='PENDING SUMMARY')
            done_summary.to_excel(writer, index=False, sheet_name='DONE SUMMARY')
        
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
        
        return Response({'message': 'request processed successfully !!!', "download_link": download_link, "json_data": json_data, "unique_data": unique_data, "breakpoint1" : breakpoint1, "breakpoint2" : breakpoint2}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
    

@api_view(['GET', 'POST'])
def ms2_graphs_view(request):
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "I-Deploy ONAIR",
        'RAN PAT Accepted',
        'RAN SAT Accepted',
        'MW PAT Accepted',
        'MW SAT Accepted',
        'SCFT Accepted',
        'KPI AT offer',
        'KPI AT Accepted',
        '4G MS2',
        '5G MS2',
        'Final MS2'
    ]
    
    unique_data = {
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones
    }

    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        if site_tagging and "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if month_start and month_end:
            filters[f"{milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + '_date'}__range"] = (month_start, month_end)

        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        df['Circle'] = df['circle']

        def generate_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"{end_label} Pending Count"]
            ]

            return summary

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        start_col = start_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
        end_col = end_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
        
        graph_summary = generate_summary(df, start_label, end_label, start_col, end_col)

        graph_summary = graph_summary.applymap(lambda x: str(x) if pd.notna(x) else "")   

        result_json = graph_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "graph_summary": json_data1,
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
 

@api_view(['GET', 'POST'])
def ms2_monthly_graph(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    view = request.data.get('view')
    type = request.data.get('type')
    year_filtered = request.data.get('year2')
    
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
    milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
    
    all_unique_circles = list(
        AlokTrackerModel.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        AlokTrackerModel.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "I-Deploy ONAIR",
        'RAN PAT Accepted',
        'RAN SAT Accepted',
        'MW PAT Accepted',
        'MW SAT Accepted',
        'SCFT Accepted',
        'KPI AT offer',
        'KPI AT Accepted',
        '4G MS2',
        '5G MS2',
        'Final MS2'
    ]
    
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
        "milestones": milestones
    }
    
    try:
        
        year_filtered = int(year_filtered)
        
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name
        
            
            
        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        # print(df)
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        df = df[[milestone1_col, milestone2_col]]
       
        for col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        
        today = dtime.today().date()
        
        df["month_num1"] = (
            df[milestone1_col].dt.month + (df[milestone1_col].dt.day >= 26)
        ).replace({13: 1}).astype("Int64")

        df["year1"] = (
            df[milestone1_col].dt.year - (
                (df[milestone1_col].dt.month < 3) |
                ((df[milestone1_col].dt.month == 3) & (df[milestone1_col].dt.day < 26))
            )
        ).astype("Int64")

        # 3️⃣ Display year
        df["display_year"] = df["year1"] + (df["month_num1"] <= 3)

        # 4️⃣ Month label
        df["month_name1"] = pd.to_datetime(
            df["month_num1"].astype(str) + "-" + df["display_year"].astype(str),
            format="%m-%Y",
            errors="coerce"
        ).dt.strftime("%b-%y")

        # 5️⃣ Current FY
        current_fy = year_filtered

        # 6️⃣ CF logic
        df.loc[df["year1"] < current_fy, "month_name1"] = "CF"



        df['month_num2'] = df[milestone2_col].dt.month + (df[milestone2_col].dt.day >= 26)
        df['month_num2'] = df['month_num2'].replace({13: 1})

        # 2️⃣ Financial year (derived ONLY from original date)
        df['year2'] = df[milestone2_col].dt.year - (
            (df[milestone2_col].dt.month < 3) |
            ((df[milestone2_col].dt.month == 3) & (df[milestone2_col].dt.day < 26))
        )

        # 3️⃣ Safe numeric conversion (IMPORTANT)
        df['month_num2'] = pd.to_numeric(df['month_num2'], errors='coerce')
        df['year2'] = pd.to_numeric(df['year2'], errors='coerce')

        # 4️⃣ Display year
        df['display_year2'] = df['year2'] + (df['month_num2'] <= 3)

        # 5️⃣ Month label (NaT-safe)
        df['month_name2'] = pd.to_datetime(
            df['month_num2'].astype('Int64').astype(str) + '-' +
            df['display_year2'].astype('Int64').astype(str),
            format='%m-%Y',
            errors='coerce'
        ).dt.strftime('%b-%y')

        # 6️⃣ Current financial year
        current_fy = year_filtered

        # 7️⃣ Carry Forward (only where year2 exists)
        df.loc[df['year2'].notna() & (df['year2'] < current_fy), 'month_name2'] = 'CF'
        
        print(df['month_name1'].dropna().unique().tolist())

        
        print(df['year1'])
        
        print(df['month_name2'])
        
        print(df['year2'])
        
        df.loc[df['year1'].notna() & (df['year1'] > current_fy), ["month_name1"]] = "Future"
        df.loc[df['year2'].notna() & (df['year2'] > current_fy), ["month_name2"]] = "Future"

        def sort_financial_year(summary):
            # Extract sorted unique Month-Year (excluding CF)
            months = [
                m for m in summary['month_name'].unique()
                if m not in ["CF", "Future"] and pd.notna(m)
            ]
            # Convert to datetime to sort correctly
            months_sorted = sorted(
                months,
                key=lambda x: pd.to_datetime(x, format='%b-%y')
            )

            # CF should always be first
            ordered = []

            if "CF" in summary['month_name'].values:
                ordered.append("CF")

            ordered += months_sorted

            if "Future" in summary['month_name'].values:
                ordered.append("Future")

            # Convert into ordered categorical for final sort
            summary['month_name'] = pd.Categorical(summary['month_name'], ordered, ordered=True)

            return summary.sort_values('month_name')
        
        def generate_type1_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()

            # Start counts by month_name
            start_counts = (
                temp[temp[start_col].notna()]
                .groupby("month_name1")
                .size()
                .reset_index(name=f"{start_label} Done Count")
                .rename(columns={"month_name1": "month_name"})
            )

            # End counts by month_name2
            end_counts = (
                temp[temp[end_col].notna()]
                .groupby("month_name2")
                .size()
                .reset_index(name=f"{end_label} Done Count")
                .rename(columns={"month_name2": "month_name"})
            )

            # Merge both results
            summary = pd.merge(start_counts, end_counts, on="month_name", how="outer")

            # Convert to Int64 nullable type
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")
            summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].astype("Int64")

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            return summary.reset_index(drop=True)
        
        def generate_type2_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("month_name1").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    # f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
                .rename(columns={"month_name1": "month_name"})
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")

            summary = summary[
                ["month_name", f"{start_label} Done Count", f"{end_label} Done Count"]
            ]

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            return summary.reset_index(drop=True)
        
        summary = generate_type1_summary(df, milestone1, milestone2, milestone1_col, milestone2_col) if type == 'type1' else generate_type2_summary(df, milestone1, milestone2, milestone1_col, milestone2_col)

        numeric_cols = [col for col in summary.columns if "Done Count" in col]

        summary[numeric_cols] = summary[numeric_cols].apply(pd.to_numeric, errors="coerce")
        summary[numeric_cols] = summary[numeric_cols].fillna(0).astype(int)

        
        result_json = summary.to_dict(orient = "records")
        json_data = json.dumps(result_json)
        
        return Response({"message" : "request processed successfully ! ", "json_data": json_data, "unique_data": unique_data}, status=200)
    
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
    


# @api_view(['GET', 'POST'])
# def dismantle_daily_dashboard_view(request):
#     circle = request.data.get('circle', [])
#     site_tagging = request.data.get('site_tagging', [])
#     current_status = request.data.get('relocation_method', [])
#     new_toco_name = request.data.get('new_toco_name', [])
#     start_date = request.data.get('from_date')
#     end_date = request.data.get('to_date')
#     view = request.data.get('view')

#     all_unique_circles = list(
#         AlokTrackerModel.objects.exclude(circle__isnull=True)
#         .distinct("circle")
#         .values_list("circle", flat=True)
#     )

#     all_unique_site_tagging = list(
#         AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
#         .distinct("site_tagging")
#         .values_list("site_tagging", flat=True)
#     )

#     all_unique_dismantling_status = list(
#         AlokTrackerModel.objects.exclude(dismantling_status__isnull=True)
#         .distinct("dismantling_status")
#         .values_list("dismantling_status", flat=True)
#     )

#     all_unique_new_toco_name = list(
#         AlokTrackerModel.objects.exclude(new_toco_name__isnull=True)
#         .distinct("new_toco_name")
#         .values_list("new_toco_name", flat=True)
#     )

#     unique_data = {
#         "unique_circle": sorted(all_unique_circles),
#         "unique_site_tagging": sorted(all_unique_site_tagging),
#         "unique_relocation_method": sorted(all_unique_dismantling_status),
#         "unique_new_toco_name": sorted(all_unique_new_toco_name),
#     }

#     circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
#     site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
#     current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
#     new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]

#     start_date = dtime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
#     end_date = dtime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

#     try:
#         filters = {}
#         if "ALL" not in circle:
#             filters["circle__in"] = circle
#         if "ALL" not in site_tagging:
#             filters["site_tagging__in"] = site_tagging
#         if "ALL" not in current_status:
#             filters["dismantling_status__in"] = current_status
#         if "ALL" not in new_toco_name:
#             filters["new_toco_name__in"] = new_toco_name

#         obj = AlokTrackerModel.objects.filter(**filters)
#         df = pd.DataFrame(obj.values())

#         if df.empty:
#             return Response({'error': 'No data found for given filters'}, status=404)
        
#         # df['rfai_date'] = df['clear_rfai_date'].where(df["clear_rfai_date"].notna(), df['rfai_date'])

#         for col in df.columns:
#             if "Date" in col:
#                 df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")

#         today = dtime.today().date()

#         if not start_date or not end_date:
#             if today.day >= 26:
#                 start_date = today.replace(day=26)
#                 if today.month == 12:
#                     end_date = today.replace(year=today.year + 1, month=1, day=25)
#                 else:
#                     end_date = today.replace(month=today.month + 1, day=25)
#             else:
#                 if today.month == 1:
#                     start_date = today.replace(year=today.year - 1, month=12, day=26)
#                 else:
#                     start_date = today.replace(month=today.month - 1, day=26)
#                 end_date = today.replace(day=25)

#         date_range = pd.date_range(start=start_date, end=min(end_date, today)).date
#         formatted_dates = [d.strftime("%d-%b-%y") for d in date_range]

#         result_columns = ["Milestone Track/Site Count", "CF"] + formatted_dates
#         result = pd.DataFrame(columns=result_columns)

#         milestones = [
#             "Site ONAIR Date",
#             "Dismantling Survey Date",
#             "SREQ/CREQ Raised Date",
#             "Dismantle Date",
#             "Material Pickup Date",
#             "Material Submission Date",
#             "OCI Done Date",
#             "Sign-off Date",
#         ]

#         unique_data.update(**{"Milestone": milestones})

#         for milestone in milestones:
#             milestone_df_format = (
#                 milestone.lower()
#                 .replace(" ", "_")
#                 .replace("-", "_")
#                 .replace("(", "")
#                 .replace(")", "")
#                 .replace("/", "_")
#             )

#             if milestone_df_format not in df.columns:
#                 continue

#             df[milestone_df_format] = pd.to_datetime(df[milestone_df_format], errors="coerce").dt.date
#             valid_dates = df[milestone_df_format].dropna()

#             if valid_dates.empty:
#                 row = {
#                     "Milestone Track/Site Count": milestone,
#                     "CF": "-",
#                     **{d.strftime("%d-%b-%y"): "-" for d in date_range}
#                 }
#                 result.loc[len(result)] = row
#                 continue

#             cf_count = (valid_dates < start_date).sum()
#             cumulative = cf_count
#             row = {"Milestone Track/Site Count": milestone, "CF": cf_count}

#             for date in date_range:
#                 count = (valid_dates == date).sum()
#                 cumulative += count

#                 if view == "Cumulative":
#                     row[date.strftime("%d-%b-%y")] = cumulative
#                 else:
#                     row[date.strftime("%d-%b-%y")] = count

#             result.loc[len(result)] = row

#         result.columns = [
#             col.strftime("%d-%b-%y") if isinstance(col, (dtime,)) or hasattr(col, "strftime") else col
#             for col in result.columns
#         ]

#         result = result[["Milestone Track/Site Count", "CF"] + formatted_dates]

#         last_col = formatted_dates[-1]
#         dash_mask = result[last_col] == '-'

#         result[last_col] = pd.to_numeric(result[last_col], errors='coerce')
#         result['Gap'] = -result[last_col].diff()

#         nan_mask = result['Gap'].isna()

#         result[last_col] = result[last_col].fillna(0).astype(int)
#         result['Gap'] = result['Gap'].fillna(0).astype(int)

#         result.loc[dash_mask, last_col] = '-'
#         result.loc[nan_mask, 'Gap'] = '-'

#         result['Gap'] = result['Gap'].astype(str)
#         result = result.astype(str).reset_index(drop=True)

#         result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(
#             lambda col: col.replace(" Date", "") if " Date" in col else col
#         )

#         new_result = result.copy()

#         for i, date in enumerate(formatted_dates, start=1):
#             new_result.rename(columns={date: f'date_{i}'}, inplace=True)

#         result_json = new_result.to_dict(orient="records")
#         json_data = json.dumps(result_json)

#         current_date = dtime.now().strftime("%Y-%m-%d")
#         current_time = dtime.now().strftime("%H-%M-%S")

#         BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
#         os.makedirs(BASE_URL, exist_ok=True)

#         output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
#         shutil.rmtree(output_folder, ignore_errors=True)
#         os.makedirs(output_folder, exist_ok=True)

#         dashboard_file_path = os.path.join(
#             output_folder,
#             f"DISMANTLE_DAILY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx"
#         )

#         with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
#             result.to_excel(writer, index=False, sheet_name='Daily Waterfall Dismantling')

#         dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
#         download_link = request.build_absolute_uri(dashboard_file_path)

#         return Response({
#             'message': 'Dashboard created successfully !!!',
#             "download_link": download_link,
#             "data": json_data,
#             "dates": formatted_dates,
#             "unique_data": unique_data
#         }, status=200)

#     except Exception as e:
#         return Response({"error": f"{str(e)}"}, status=500)


@api_view(['GET', 'POST'])
def dismantle_daily_dashboard_view(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    start_date = request.data.get('from_date')
    end_date = request.data.get('to_date')
    view = request.data.get('view')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None

    all_unique_circles = list(
        AlokTrackerModel.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(dismantling_status__isnull=True)
        .distinct("dismantling_status")
        .values_list("dismantling_status", flat=True)
    )

    all_unique_new_toco_name = list(
        AlokTrackerModel.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )

    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]

    start_date = dtime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    end_date = dtime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

    try:
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["dismantling_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name

        obj = AlokTrackerModel.objects.filter(**filters)
        df = pd.DataFrame(obj.values())
        

        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        # df['rfai_date'] = df['clear_rfai_date'].where(df["clear_rfai_date"].notna(), df['rfai_date'])

        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")

        today = dtime.today().date()

        if (not month_start or not month_end) and (not start_date or not end_date):
            if today.day >= 26:
                start_date = today.replace(day=26)
                if today.month == 12:
                    end_date = today.replace(year=today.year + 1, month=1, day=25)
                else:
                    end_date = today.replace(month=today.month + 1, day=25)
            else:
                if today.month == 1:
                    start_date = today.replace(year=today.year - 1, month=12, day=26)
                else:
                    start_date = today.replace(month=today.month - 1, day=26)
                end_date = today.replace(day=25)
                
        if month_start and month_end:
            date_range = pd.date_range(start=month_start, end=min(month_end, today)).date
        else:
            date_range = pd.date_range(start=start_date, end=min(end_date, today)).date

        formatted_dates = [d.strftime("%d-%b-%y") for d in date_range]

        if month_start and month_end:
            result_columns = ["Milestone Track/Site Count", "AOP", "CF"] + formatted_dates
        else:
            result_columns = ["Milestone Track/Site Count", "AOP"] + formatted_dates
        result = pd.DataFrame(columns=result_columns)

        milestones = [
            "Site ONAIR Date",
            "Dismantling Survey Date",
            "SREQ/CREQ Raised Date",
            "Dismantle Date",
            "Material Pickup Date",
            "Material Submission Date",
            "OCI Done Date",
            "Sign-off Date",
        ]

        unique_data.update(**{"Milestone": milestones})

        for milestone in milestones:
            milestone_df_format = (
                milestone.lower()
                .replace(" ", "_")
                .replace("-", "_")
                .replace("(", "")
                .replace(")", "")
                .replace("/", "_")
            )

            if milestone_df_format not in df.columns:
                continue

            df[milestone_df_format] = pd.to_datetime(df[milestone_df_format], errors="coerce").dt.date
            if month_start and month_end:
                mask = df['site_onair_date'].isna()
                valid_dates_cf = df.loc[mask, milestone_df_format].dropna()
            # else:
            #     valid_dates = df[milestone_df_format].dropna()
            
            valid_dates_aop = df[milestone_df_format].dropna()

            if valid_dates_aop.empty:
                if month_start and month_end:
                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": "-",
                        "CF": "-",
                        **{d.strftime("%d-%b-%y"): "-" for d in date_range}
                    }
                else:
                    row = {
                        "Milestone Track/Site Count": milestone,
                        "AOP": "-",
                        **{d.strftime("%d-%b-%y"): "-" for d in date_range}
                    }
                
                result.loc[len(result)] = row
                continue
            
            if month_start and month_end:
                cf_count = (valid_dates_cf < month_start).sum()
                aop_count = (valid_dates_aop < month_start).sum()
            else:
            #     cf_count = (valid_dates < start_date).sum()
                aop_count = (valid_dates_aop < start_date).sum()
            
            if month_start and month_end:
                cumulative = cf_count
                row = {"Milestone Track/Site Count": milestone,"AOP": aop_count, "CF": cf_count}
            else:
                cumulative = aop_count
                row = {"Milestone Track/Site Count": milestone,"AOP": aop_count}


            for d in date_range:
                count = (valid_dates_aop == d).sum()
                cumulative += count

                if view == "Cumulative":
                    row[d.strftime("%d-%b-%y")] = cumulative
                else:
                    row[d.strftime("%d-%b-%y")] = count

            result.loc[len(result)] = row

        result.columns = [
            col.strftime("%d-%b-%y") if isinstance(col, (dtime,)) or hasattr(col, "strftime") else col
            for col in result.columns
        ]
        if month_start and month_end:
            result = result[["Milestone Track/Site Count", "AOP" , "CF"] + formatted_dates]
        else:
            result = result[["Milestone Track/Site Count", "AOP"] + formatted_dates]
 
        last_col = formatted_dates[-1]
        dash_mask = result[last_col] == '-'

        result[last_col] = pd.to_numeric(result[last_col], errors='coerce')
        result['Gap'] = -result[last_col].diff()

        nan_mask = result['Gap'].isna()

        result[last_col] = result[last_col].fillna(0).astype(int)
        result['Gap'] = result['Gap'].fillna(0).astype(int)

        result.loc[dash_mask, last_col] = '-'
        result.loc[nan_mask, 'Gap'] = '-'

        result['Gap'] = result['Gap'].astype(str)
        result = result.astype(str).reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(
            lambda col: col.replace(" Date", "") if " Date" in col else col
        )

        new_result = result.copy()

        for i, d in enumerate(formatted_dates, start=1):
            new_result.rename(columns={d: f'date_{i}'}, inplace=True)

        result_json = new_result.to_dict(orient="records")
        json_data = json.dumps(result_json)

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")

        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)

        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)

        dashboard_file_path = os.path.join(
            output_folder,
            f"RELOCATION_TRACKING_MS2_DAILY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx"
        )

        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Daily Waterfall MS2')

        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)

        return Response({
            'message': 'Dashboard created successfully !!!',
            "download_link": download_link,
            "data": json_data,
            "dates": formatted_dates,
            "unique_data": unique_data
        }, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)


@api_view(['GET', 'POST'])
def dismantle_weekly_monthly_dashboard_view(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    year2 = request.data.get('year2')
    view = request.data.get('view')
 
    # Default to 'ALL' if not provided
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    
    if year2:
        year2 = int(year2)
    
    if month_filtered and year_filtered:
        month_filtered = int(month_filtered)
        year_filtered = int(year_filtered)

        month_end = date(year_filtered, month_filtered, 25)

        if month_filtered == 1:
            month_start = date(year_filtered - 1, 12, 26)
        else:
            month_start = date(year_filtered, month_filtered - 1, 26)
    else:
        month_start = None
        month_end = None
   
    all_unique_circles = list(
        AlokTrackerModel.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_dismantling_status = list(
        AlokTrackerModel.objects.exclude(dismantling_status__isnull=True)
        .distinct("dismantling_status")
        .values_list("dismantling_status", flat=True)
    )

    all_unique_new_toco_name = list(
        AlokTrackerModel.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )

    # ✅ Add "ALL" at the top of each list
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_dismantling_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
    }

    try:
        # 🔹 Dynamic filters
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["dismantling_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name
 
        # 🔹 Fetch data
        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        # df['rfai_date'] = df['re_rfai_date'].where(df["re_rfai_date"].notna(), df['rfai_date'])
       
        for col in df.columns:
            if "Date" in col:
                df[col] = pd.to_datetime(df[col], format="%d-%b-%y", errors="coerce")
               
        today = dtime.today().date()
 
        if year2:
            fy_start = dtime(year2, 3, 26).date()   # 26-Apr current year start
            fy_end = dtime(year2 + 1, 3, 25).date() # 25-Mar next year end
        else:
            if today.month >= 4:
                fy_start = dtime(today.year, 3, 26).date()   # 26-Apr current year start
                fy_end = dtime(today.year + 1, 3, 25).date() # 25-Mar next year end
            else:
                fy_start = dtime(today.year - 1, 3, 26).date()
                fy_end = dtime(today.year, 3, 25).date()
                
        print("2.5-------------------------------------------------------------")
 
        # 🧩 2. Determine current cycle (26th prev → 25th curr)
        if today.day >= 26:
            if today.month == 12:
                current_cycle_start = dtime(today.year, 12, 26).date()
                current_cycle_end = dtime(today.year + 1, 1, 25).date()
            else:
                current_cycle_start = dtime(today.year, today.month, 26).date()
                current_cycle_end = dtime(today.year, today.month + 1, 25).date()
        else:
            if today.month == 1:
                current_cycle_start = dtime(today.year - 1, 12, 26).date()
            else:
                current_cycle_start = dtime(today.year, today.month - 1, 26).date()
            current_cycle_end = dtime(today.year, today.month, 25).date()
            
        if year2:
            current_cycle_start = min(dtime(year2+1, 2, 26).date(), current_cycle_start)
            current_cycle_end = min(dtime(year2+1, 3, 25).date(), current_cycle_end)
 
        # 🗓️ 3. Create monthly periods (Apr → month before current)
        months = []
        start = fy_start
        while start <= current_cycle_start:
            if start.month == 12:
                end = dtime(start.year + 1, 1, 25).date()
            else:
                end = dtime(start.year, start.month + 1, 25).date()
            months.append((start, end))
            start = end + timedelta(days=1)
 
        # 🗓️ 4. Create weekly periods for current month
        weeks = []
        
        if month_start and month_end:
            current_cycle_start = month_start
            current_cycle_end = month_end
        
        week_start = current_cycle_start
 
        while week_start <= current_cycle_end:
            week_end = week_start + timedelta(days=6)
            if week_end > current_cycle_end:
                week_end = current_cycle_end
            weeks.append((week_start, week_end))
            week_start = week_end + timedelta(days=1)
 
 
        # 📊 5. Prepare result DataFrame
        result = pd.DataFrame()
       
       
        # result.set_index("Milestone Track/Site Count", inplace=True)
 
        milestones = [
            "Site ONAIR Date",
            "Dismantling Survey Date",
            "SREQ/CREQ Raised Date",
            "Dismantle Date",
            "Material Pickup Date",
            "Material Submission Date",
            "OCI Done Date",
            "Sign-off Date",
        ]
 
        data = []
        
        for milestone in milestones:
            
            milestone_df_format = milestone.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_")

            if milestone_df_format not in df.columns:
                continue
            
            milestone_data = pd.to_datetime(df[milestone_df_format], errors='coerce').dt.date
            row = {"Milestone Track/Site Count": milestone}
            if milestone_data.dropna().empty:
                row["CF"] = "-"
                # Set all month and week columns to "-"
                for _, end in months:
                    month_name = end.strftime("%b-%y")
                    row[month_name] = "-"
                for i, _ in enumerate(weeks, 1):
                    week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                    row[week_name] = "-"
                data.append(row)
                continue
 
 
            # CF → sites before financial year start
            row["CF"] = (milestone_data < fy_start).sum()
 
            cumulative_month = row["CF"]
            
            for start, end in months:
                month_name = end.strftime("%b-%y")
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_month += count
                if view == 'Cumulative':
                    row[month_name] = cumulative_month
                else:
                    row[month_name] = count
 
            cumulative_week = 0
            for i, (start, end) in enumerate(weeks, 1):
                week_name = f"{current_cycle_end.strftime('%b-%y')} W{i}"
                if(start >= today):
                    row[week_name] = 0
                count = ((milestone_data >= start) & (milestone_data <= end)).sum()
                cumulative_week += count
                if view == 'Cumulative':
                    row[week_name] = cumulative_week
                else:
                    row[week_name] = count
 
            data.append(row)
 

        result = pd.DataFrame(data)
 
        for col in result.columns:
            if col != "Milestone Track/Site Count":
                result[col] = result[col].astype(str)
 
        result = result.reset_index(drop=True)

        result['Milestone Track/Site Count'] = result['Milestone Track/Site Count'].apply(lambda col: col.replace(" Date", "") if " Date" in col else col)
 
        new_data = result.copy()
        months_columns = [col for col in new_data.columns.to_list()[2:] if " W" not in col]
        months_data = new_data[new_data.columns.to_list()[:2] + months_columns].copy()
        
        weeks_columns = [col for col in new_data.columns.to_list()[2:] if " W" in col]
        
        week_data = new_data[new_data.columns.to_list()[:1] + weeks_columns].copy()
        
        months_data.rename(columns={
            col : f"Month-{i}" for i, col in enumerate(months_columns, start=1)
        }, inplace=True)
        
        week_data.rename(columns={
            col : f"Month_Week-{i}" for i, col in enumerate(weeks_columns, start=1)
        }, inplace=True)
        
        unique_data.update(**{"month_columns": months_columns, "week_columns" : weeks_columns})
        

        month_dict_data = months_data.to_dict(orient="records")
        month_json_data = json.dumps(month_dict_data)
        week_dict_data = week_data.to_dict(orient="records")
        week_json_data = json.dumps(week_dict_data)
        
        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"DISMANTLE_WEEKLY_MONTHY_DASHBOARD_FILE_{current_date}_{current_time}.xlsx")
       
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            result.to_excel(writer, index=False, sheet_name='Monthly Dismantling')
       
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
   
        return Response({'message': 'Weekly and Monthly Dashboard created successfully !!!', "download_link": download_link, "unique_data": unique_data, "months_data": month_json_data, "week_data": week_json_data}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
  
   
@api_view(['GET', 'POST'])
def dismantle_ageing_dashboard_table1(request):
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "Dismantling Survey",
        "SREQ/CREQ Raised",
        "Dismantle",
        "Material Pickup",
        "Material Submission",
        "OCI Done",
        "Sign-off",
    ]
    
    unique_data = {
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones
    }

    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        if site_tagging and "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if month_start and month_end:
            filters[f"{milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_").replace("/", "_") + '_date'}__range"] = (month_start, month_end)

        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        date_cols = [col.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_").replace("/", "_") + '_date' for col in milestones]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']
        
        
        def generate_table_summary(df, milestone1, milestone2):


            milestone_cols = [(col.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date") for col in milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]]

            temp = df.copy()

            circles = ["ALL"] + temp["circle"].unique().tolist()

            final_df = pd.DataFrame()

            milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
            milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"

            for circle in circles:
                if circle != "ALL":
                    temp  = df[df["circle"] == circle]

                done_row = {}
                pending_row = {}

                for col, label in zip(milestone_cols, milestones[milestones.index(milestone1):milestones.index(milestone2) + 1]):

                    done_row[label] = (temp[col].notna() & temp[milestone1_col].notna() ).sum()

                    pending_row[label] = (temp[col].isna() & temp[milestone1_col].notna() ).sum()

                rows_to_add = pd.DataFrame([
                    {"Circle": circle, "Site Status": "Done", **done_row},
                    {"Circle": circle, "Site Status": "Pending", **pending_row},
                ])

                final_df = pd.concat([final_df, rows_to_add], ignore_index=True)

            for c in final_df.columns:
                if c != "Site Status" and c != "Circle":
                    final_df[c] = final_df[c].astype("Int64")

            return final_df

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        table_summary = generate_table_summary(df, start_label, end_label)
 
        table_summary = table_summary.applymap(lambda x: str(x) if pd.notna(x) else "")      


        result_json = table_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "table_summary": json_data1
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
  
  
@api_view(['GET', 'POST'])
def dismantle_ageing_dashboard_table2(request):
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    breakpoint1 = request.data.get('breakpoint1')
    breakpoint2 = request.data.get('breakpoint2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')
    type = request.data.get('type')

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "Dismantling Survey",
        "SREQ/CREQ Raised",
        "Dismantle",
        "Material Pickup",
        "Material Submission",
        "OCI Done",
        "Sign-off",
    ]
    
    unique_data = {
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones
    }
    

    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    
    try:

        milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
        milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"

        breakpoint1 = int(breakpoint1)
        breakpoint2 = int(breakpoint2)
        
        temp = breakpoint1
        breakpoint1 = min(breakpoint1, breakpoint2)
        breakpoint2 = max(temp, breakpoint2)

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None

        milestones = [
            "Site ONAIR",
            "Dismantling Survey",
            "SREQ/CREQ Raised",
            "Dismantle",
            "Material Pickup",
            "Material Submission",
            "OCI Done",
            "Sign-off",
        ]
        

        filters = {}
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
            
        if type == "type2":
            if month_start and month_end:
                filters[f"{milestone1_col}__range"] = (month_start, month_end)
                

        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        
        if type == "type1":
            if month_start and month_end:
                obj = obj.filter(
                    Q(**{f"{milestone1_col}__range": (month_start, month_end)}) |
                    Q(**{f"{milestone2_col}__range": (month_start, month_end)})
                )
        

        df = pd.DataFrame(obj.values())

        
        date_cols = [col.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_").replace("/", "_") + '_date' for col in milestones]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce")

        df['Circle'] = df['circle']

        if month_start is not None:
            month_start = pd.to_datetime(month_start)

        if month_end is not None:
            month_end = pd.to_datetime(month_end)

        def generate_done_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            temp = df.copy()
            
            temp["days_diff"] = (temp[end_col] - temp[start_col]).dt.days
            
            if type == "type2":
        
                summary = (
                    temp.groupby("Circle").apply(lambda g: pd.Series({
                        f"{start_label} Done Count": g[start_col].notna().sum(),
                        f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[end_col].notna()) & (g[start_col].notna())).sum(),
                        "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].notna(), "days_diff"].mean()
                    }))
                    .reset_index()
                )
                
                total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                total_row["Circle"] = "Total"

                overall_valid = temp.loc[
                    temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].notna(),
                    "days_diff"
                ]
                total_row["Average Days"] = overall_valid.mean()

            else:
                if month_start is not None and month_end is not None:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": (g[start_col].notna() & (g[start_col] >= month_start) & (g[start_col] <= month_end)).sum(),
                            f"{end_label} Done Count": (g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"<= {breakpoint1} days": (g["days_diff"].notna() & (g["days_diff"] <= breakpoint1) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": (g["days_diff"].notna() & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            f">= {breakpoint2} days": (g["days_diff"].notna() & (g["days_diff"] >= breakpoint2) & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end)).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna() & (g[end_col] >= month_start) & (g[end_col] <= month_end), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna() & (temp[end_col] >= month_start) & (temp[end_col] <= month_end),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                else:
                    summary = (
                        temp.groupby("Circle").apply(lambda g: pd.Series({
                            f"{start_label} Done Count": g[start_col].notna().sum(),
                            f"{end_label} Done Count": g[end_col].notna().sum(),
                            f"<= {breakpoint1} days": ((g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & g[end_col].notna()).sum(),
                            f"{breakpoint1+1}-{breakpoint2-1} days": ((g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & g[end_col].notna()).sum(),
                            f">= {breakpoint2} days": ((g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & g[end_col].notna()).sum(),
                            "Average Days": g.loc[g["days_diff"].notna() & g[end_col].notna(), "days_diff"].mean()
                        }))
                        .reset_index()
                    )
                    
                    total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
                    total_row["Circle"] = "Total"

                    overall_valid = temp.loc[
                        temp["days_diff"].notna() & temp[end_col].notna(),
                        "days_diff"
                    ]
                    total_row["Average Days"] = overall_valid.mean()
                    
            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]

            return summary


        done_summary = generate_done_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        done_summary = done_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        def generate_pending_summary(df, start_label, end_label, start_col, end_col, breakpoint1, breakpoint2):
            
            today = pd.Timestamp.today()
            
            temp = df.copy()
            
            temp["days_diff"] = (today - temp[start_col]).dt.days

            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                    f"<= {breakpoint1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] <= breakpoint1) & (g[start_col].notna())).sum(),
                    f"{breakpoint1+1}-{breakpoint2-1} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] > breakpoint1) & (g["days_diff"] < breakpoint2) & (g[start_col].notna())).sum(),
                    f">= {breakpoint2} days": ((g[end_col].isna()) & (g["days_diff"].notna()) & (g["days_diff"] >= breakpoint2) & (g[start_col].notna())).sum(),
                    "Average Days": g.loc[g["days_diff"].notna() & g[start_col].notna() & g[end_col].isna(), "days_diff"].mean()
                }))
                .reset_index()
            )

            total_row = summary.drop(columns=["Circle"]).sum(numeric_only=True)
            total_row["Circle"] = "Total"

            overall_valid = temp.loc[
                temp["days_diff"].notna() & temp[start_col].notna() & temp[end_col].isna(),
                "days_diff"
            ]
            total_row["Average Days"] = overall_valid.mean()

            summary = pd.concat([pd.DataFrame([total_row]), summary], ignore_index=True)

            summary["Average Days"] = summary["Average Days"].round(0).astype("Int64")
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Pending Count", f"<= {breakpoint1} days", f"{breakpoint1+1}-{breakpoint2-1} days", f">= {breakpoint2} days", "Average Days"]
            ]
            
            return summary
        

        pending_summary = generate_pending_summary(df, milestone1, milestone2, milestone1_col, milestone2_col, breakpoint1, breakpoint2)

        pending_summary = pending_summary.applymap(lambda x: str(x) if pd.notna(x) else "")

        current_date = dtime.now().strftime("%Y-%m-%d")
        current_time = dtime.now().strftime("%H-%M-%S")
 
        BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
        os.makedirs(BASE_URL, exist_ok=True)
        output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
       
        dashboard_file_path = os.path.join(output_folder, f"AGEING_DASHBOARD_FILE_{site_tagging}_{current_date}_{current_time}.xlsx")

        result_json = pending_summary.to_dict(orient="records")
        json_data2 = json.dumps(result_json)

        result_json = done_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "done_summary": json_data1,
            "pending_summary": json_data2
        }
        
        with pd.ExcelWriter(dashboard_file_path, engine='xlsxwriter') as writer:
            pending_summary.to_excel(writer, index=False, sheet_name='PENDING SUMMARY')
            done_summary.to_excel(writer, index=False, sheet_name='DONE SUMMARY')
        
        dashboard_file_path = dashboard_file_path.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
        download_link = request.build_absolute_uri(dashboard_file_path)
        
        return Response({'message': 'request processed successfully !!!', "download_link": download_link, "json_data": json_data, "unique_data": unique_data, "breakpoint1" : breakpoint1, "breakpoint2" : breakpoint2}, status=200)
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
 
   

@api_view(['GET', 'POST'])
def dismantle_graphs_view(request):
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('current_status', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    month_filtered = request.data.get('month')
    year_filtered = request.data.get('year')

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "Dismantling Survey",
        "SREQ/CREQ Raised",
        "Dismantle",
        "Material Pickup",
        "Material Submission",
        "OCI Done",
        "Sign-off",
    ]
    
    unique_data = {
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_current_status": sorted(all_unique_current_status),
        "milestones": milestones
    }

    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    
    try:

        if month_filtered and year_filtered:
            month_filtered = int(month_filtered)
            year_filtered = int(year_filtered)

            month_end = date(year_filtered, month_filtered, 25)

            if month_filtered == 1:
                month_start = date(year_filtered - 1, 12, 26)
            else:
                month_start = date(year_filtered, month_filtered - 1, 26)
        else:
            month_start = None
            month_end = None
        
        filters = {}
        if site_tagging and "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if current_status and "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if month_start and month_end:
            filters[f"{milestone1.lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '').replace("4", "four_").replace("5", "five_") + '_date'}__range"] = (month_start, month_end)

        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        
        
        
        print(list(df.columns))
        df['Circle'] = df['circle']
        print("2")

        def generate_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("Circle").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].round(0).astype("Int64")

            summary = summary[
                ["Circle", f"{start_label} Done Count", f"{end_label} Done Count", f"{end_label} Pending Count"]
            ]

            return summary

        start_label = milestone1 if milestones.index(milestone1) < milestones.index(milestone2) else milestone2
        end_label = milestone2 if milestones.index(milestone1) < milestones.index(milestone2) else milestone1

        start_col = start_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
        end_col = end_label.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
        
        graph_summary = generate_summary(df, start_label, end_label, start_col, end_col)

        graph_summary = graph_summary.applymap(lambda x: str(x) if pd.notna(x) else "")   

        result_json = graph_summary.to_dict(orient="records")
        json_data1 = json.dumps(result_json)

        json_data = {
            "graph_summary": json_data1,
        }

        return Response({'message': 'request processed successfully !!!', 'json_data': json_data, "unique_data": unique_data}, status=200)

    except Exception as e:
        return Response({"error": f"{str(e)}"}, status=500)
 

@api_view(['GET', 'POST'])
def dismantle_monthly_graph(request):
    circle = request.data.get('circle', [])
    site_tagging = request.data.get('site_tagging', [])
    current_status = request.data.get('relocation_method', [])
    new_toco_name = request.data.get('new_toco_name', [])
    milestone1 = request.data.get('milestone1')
    milestone2 = request.data.get('milestone2')
    view = request.data.get('view')
    type = request.data.get('type')
    year_filtered = request.data.get('year2')
    
    circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
    site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
    current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
    new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]
    milestone1_col = milestone1.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
    milestone2_col = milestone2.lower().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_") + "_date"
    
    all_unique_circles = list(
        AlokTrackerModel.objects.exclude(circle__isnull=True)
        .distinct("circle")
        .values_list("circle", flat=True)
    )

    all_unique_site_tagging = list(
        AlokTrackerModel.objects.exclude(site_tagging__isnull=True)
        .distinct("site_tagging")
        .values_list("site_tagging", flat=True)
    )

    all_unique_current_status = list(
        AlokTrackerModel.objects.exclude(current_status__isnull=True)
        .distinct("current_status")
        .values_list("current_status", flat=True)
    )

    all_unique_new_toco_name = list(
        AlokTrackerModel.objects.exclude(new_toco_name__isnull=True)
        .distinct("new_toco_name")
        .values_list("new_toco_name", flat=True)
    )
    
    milestones = [
        "Site ONAIR",
        "Dismantling Survey",
        "SREQ/CREQ Raised",
        "Dismantle",
        "Material Pickup",
        "Material Submission",
        "OCI Done",
        "Sign-off",
    ]
    
    unique_data = {
        "unique_circle": sorted(all_unique_circles),
        "unique_site_tagging": sorted(all_unique_site_tagging),
        "unique_relocation_method": sorted(all_unique_current_status),
        "unique_new_toco_name": sorted(all_unique_new_toco_name),
        "milestones": milestones
    }
    
    try:
        
        year_filtered = int(year_filtered)
        
        filters = {}
        if "ALL" not in circle:
            filters["circle__in"] = circle
        if "ALL" not in site_tagging:
            filters["site_tagging__in"] = site_tagging
        if "ALL" not in current_status:
            filters["current_status__in"] = current_status
        if "ALL" not in new_toco_name:
            filters["new_toco_name__in"] = new_toco_name
        
            
            
        obj = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
        df = pd.DataFrame(obj.values())
        
        # print(df)
 
        if df.empty:
            return Response({'error': 'No data found for given filters'}, status=404)
        
        df = df[[milestone1_col, milestone2_col]]
       
        for col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        
        today = dtime.today().date()
        
        df["month_num1"] = (
            df[milestone1_col].dt.month + (df[milestone1_col].dt.day >= 26)
        ).replace({13: 1}).astype("Int64")

        df["year1"] = (
            df[milestone1_col].dt.year - (
                (df[milestone1_col].dt.month < 3) |
                ((df[milestone1_col].dt.month == 3) & (df[milestone1_col].dt.day < 26))
            )
        ).astype("Int64")

        # 3️⃣ Display year
        df["display_year"] = df["year1"] + (df["month_num1"] <= 3)

        # 4️⃣ Month label
        df["month_name1"] = pd.to_datetime(
            df["month_num1"].astype(str) + "-" + df["display_year"].astype(str),
            format="%m-%Y",
            errors="coerce"
        ).dt.strftime("%b-%y")

        # 5️⃣ Current FY
        current_fy = year_filtered
        
        # 6️⃣ CF logic
        df.loc[df["year1"] < current_fy, "month_name1"] = "CF"



        df['month_num2'] = df[milestone2_col].dt.month + (df[milestone2_col].dt.day >= 26)
        df['month_num2'] = df['month_num2'].replace({13: 1})

        # 2️⃣ Financial year (derived ONLY from original date)
        df['year2'] = df[milestone2_col].dt.year - (
            (df[milestone2_col].dt.month < 3) |
            ((df[milestone2_col].dt.month == 3) & (df[milestone2_col].dt.day < 26))
        )

        # 3️⃣ Safe numeric conversion (IMPORTANT)
        df['month_num2'] = pd.to_numeric(df['month_num2'], errors='coerce')
        df['year2'] = pd.to_numeric(df['year2'], errors='coerce')

        # 4️⃣ Display year
        df['display_year2'] = df['year2'] + (df['month_num2'] <= 3)

        # 5️⃣ Month label (NaT-safe)
        df['month_name2'] = pd.to_datetime(
            df['month_num2'].astype('Int64').astype(str) + '-' +
            df['display_year2'].astype('Int64').astype(str),
            format='%m-%Y',
            errors='coerce'
        ).dt.strftime('%b-%y')

        # 6️⃣ Current financial year
        current_fy = year_filtered
        
        # 7️⃣ Carry Forward (only where year2 exists)
        df.loc[df['year2'].notna() & (df['year2'] < current_fy), 'month_name2'] = 'CF'
        
        print(df['month_name1'].dropna().unique().tolist())

        
        print(df['year1'])
        
        print(df['month_name2'])
        
        print(df['year2'])
        
        df.loc[df['year1'].notna() & (df['year1'] > current_fy), ["month_name1"]] = "Future"
        df.loc[df['year2'].notna() & (df['year2'] > current_fy), ["month_name2"]] = "Future"

        def sort_financial_year(summary):
            # Extract sorted unique Month-Year (excluding CF)
            months = [
                m for m in summary['month_name'].unique()
                if m not in ["CF", "Future"] and pd.notna(m)
            ]
            
            # Convert to datetime to sort correctly
            months_sorted = sorted(
                months,
                key=lambda x: pd.to_datetime(x, format='%b-%y')
            )

            # CF should always be first
            ordered = []

            if "CF" in summary['month_name'].values:
                ordered.append("CF")

            ordered += months_sorted

            if "Future" in summary['month_name'].values:
                ordered.append("Future")
            
            
            # Convert into ordered categorical for final sort
            summary['month_name'] = pd.Categorical(summary['month_name'], ordered, ordered=True)

            return summary.sort_values('month_name')
        
        def generate_type1_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()

            # Start counts by month_name
            start_counts = (
                temp[temp[start_col].notna()]
                .groupby("month_name1")
                .size()
                .reset_index(name=f"{start_label} Done Count")
                .rename(columns={"month_name1": "month_name"})
            )

            # End counts by month_name2
            end_counts = (
                temp[temp[end_col].notna()]
                .groupby("month_name2")
                .size()
                .reset_index(name=f"{end_label} Done Count")
                .rename(columns={"month_name2": "month_name"})
            )

            # Merge both results
            summary = pd.merge(start_counts, end_counts, on="month_name", how="outer")

            # Convert to Int64 nullable type
            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")
            summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].astype("Int64")

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            return summary.reset_index(drop=True)
        
        def generate_type2_summary(df, start_label, end_label, start_col, end_col):
            temp = df.copy()
            summary = (
                temp.groupby("month_name1").apply(lambda g: pd.Series({
                    f"{start_label} Done Count": g[start_col].notna().sum(),
                    f"{end_label} Done Count": ((g[end_col].notna()) & (g[start_col].notna())).sum(),
                    # f"{end_label} Pending Count": ((g[end_col].isna()) & (g[start_col].notna())).sum(),
                }))
                .reset_index()
                .rename(columns={"month_name1": "month_name"})
            )

            summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].astype("Int64")

            summary = summary[
                ["month_name", f"{start_label} Done Count", f"{end_label} Done Count"]
            ]

            summary = sort_financial_year(summary)

            if view == "Cumulative":
                summary[f"{start_label} Done Count"] = summary[f"{start_label} Done Count"].cumsum()
                summary[f"{end_label} Done Count"] = summary[f"{end_label} Done Count"].cumsum()

            return summary.reset_index(drop=True)
        
        summary = generate_type1_summary(df, milestone1, milestone2, milestone1_col, milestone2_col) if type == 'type1' else generate_type2_summary(df, milestone1, milestone2, milestone1_col, milestone2_col)

        numeric_cols = [col for col in summary.columns if "Done Count" in col]

        summary[numeric_cols] = summary[numeric_cols].apply(pd.to_numeric, errors="coerce")
        summary[numeric_cols] = summary[numeric_cols].fillna(0).astype(int)

        
        result_json = summary.to_dict(orient = "records")
        json_data = json.dumps(result_json)
        
        return Response({"message" : "request processed successfully ! ", "json_data": json_data, "unique_data": unique_data}, status=200)
    
    except Exception as e:
        return Response({"error": f"{str(e)}"},status=500)
   
 

# @api_view(['GET', 'POST'])
# def dismantle_gap_view(request):
    
#     userId = request.data.get('userId')
    
#     try:
#         user = RelocationUser.objects.filter(email=userId.lower()).first()
#     except RelocationUser.DoesNotExist:
#         return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
    
    
#     circles = user.circles
    
#     circle = request.data.get('circle', [])
#     site_tagging = request.data.get('site_tagging', [])
#     current_status = request.data.get('relocation_method', [])
#     new_toco_name = request.data.get('new_toco_name', [])
#     last_date = request.data.get('last_date')
#     milestone1 = request.data.get('milestone1')
#     milestone2 = request.data.get('milestone2')
#     gap = request.data.get('gap')
    
#     circle = [c.strip() for c in circle.split(',')] if circle else ["ALL"]
#     site_tagging = [s.strip() for s in site_tagging.split(',')] if site_tagging else ["ALL"]
#     current_status = [cs.strip() for cs in current_status.split(',')] if current_status else ["ALL"]
#     new_toco_name = [n.strip() for n in new_toco_name.split(',')] if new_toco_name else ["ALL"]

#     last_date = dtime.strptime(last_date, "%Y-%m-%d").date() if last_date else None
#     milestone1 = milestone1.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_")
#     milestone2 = milestone2.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "").replace("/", "_")

#     try:
        
#         today = dtime.today().date()
 
#         if not last_date:
#             if today.day >= 26:
#                 if today.month == 12:
#                     last_date = today.replace(year=today.year + 1, month=1, day=25)
#                 else:
#                     last_date = today.replace(month=today.month + 1, day=25)
#             else:
#                 last_date = today.replace(day=25)
                
#             last_date = min(today - timedelta(days=1), last_date)
                
#         if 'CENTRAL' not in circles:
#             circle = [c for c in circle if c in circles]
                
#         filters = {}
#         if "ALL" not in circle:
#             filters["circle__in"] = circle
#         if "ALL" not in site_tagging:
#             filters["site_tagging__in"] = site_tagging
#         if "ALL" not in current_status:
#             filters["current_status__in"] = current_status
#         if "ALL" not in new_toco_name:
#             filters["new_toco_name__in"] = new_toco_name
        
#         if int(gap)>0 :
#             filters[f"{milestone1}__lte"] = last_date
#         else:
#             filters[f"{milestone2}__lte"] = last_date

#         # if f"{milestone1}__lte" in filters:
#         #     del filters[f"{milestone1}__lte"]
 
#         # if f"{milestone2}__lte" in filters:
#         #     del filters[f"{milestone2}__lte"]
 
#         # 🔹 Fetch data
#         obj1 = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
#         df1 = pd.DataFrame(obj1.values())
        
#         if int(gap)>0 :
#             filters[f"{milestone2}__lte"] = last_date
#         else:
#             filters[f"{milestone1}__lte"] = last_date
        
#         obj2 = AlokTrackerModel.objects.filter(**filters)  # noqa: F405
#         df2 = pd.DataFrame(obj2.values())
        
#         df1['key'] = df1['circle'].astype(str) + "_" + df1['new_site_id'].astype(str)
#         df2['key'] = df2['circle'].astype(str) + "_" + df2['new_site_id'].astype(str)

#         # df = df1[~df1['key'].isin(df2['key'])].drop(columns=['key'])
#         # if len(df1) > len(df2):  
#         df = df1[~df1['key'].isin(df2['key'])].drop(columns=['key'])
#         # else:
#         #     df = df2[~df2['key'].isin(df1['key'])].drop(columns=['key'])
        
#         sites = df['new_site_id'].dropna().unique().tolist()

        
#         # filters = {
#         #     "site_id__in" : sites
#         # }
        
#         issue_obj = RelocationIssue.objects.filter(site_id__in=sites)
#         issue_df = pd.DataFrame(issue_obj.values())

#         rename_map = {
#             "circle": "Circle",
#             "site_id": "Site ID",
#             "issue_owner": "Issue Owner",
#             "milestone": "Milestone",
#             "issue_name": "Issue Name",
#             "start_date": "Start Date",
#             "close_date": "Close Date",
#             "status": "Status",
#             "duration": "Duration",
#             "updated_by": "Updated_by",
#             "updated_at": "Updated_at",
#             "created_by": "Created_by",
#             "created_at": "Created_at"
#         }

#         print(issue_df)
        
#         issue_df = issue_df.rename(columns=rename_map)

#         required_cols = ["Circle", "Site ID", "Issue Owner", "Milestone", "Issue Name", "Start Date", "Close Date", "Status", "Duration", "Updated_by", "Updated_at", "Created_by", "Created_at"]
#         for col in required_cols:
#             if col not in issue_df.columns:
#                 issue_df[col] = None

#         # # Convert date columns safely
#         for col in ["Start Date", "Close Date"]:
#             if col in issue_df.columns:
#                 issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
        
#         df = update_ageing(df, issue_df)
 
        
#         for col in df.columns:
#             if 'date' in col:
#                 if col != 'last_updated_date':
#                     converted = pd.to_datetime(df[col], errors='coerce')
    
#                     if converted.notna().sum() > 0:
#                         df[col] = converted.dt.strftime('%d-%b-%y')
#                 else:
#                     converted = pd.to_datetime(df[col], errors='coerce')
#                     if converted.notna().sum() > 0:
#                         df[col] = converted.dt.strftime('%d-%b-%y %H:%M:%S')
 
 
#         current_date = dtime.now().strftime("%Y-%m-%d")
#         current_time = dtime.now().strftime("%H-%M-%S")
 
#         BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
#         os.makedirs(BASE_URL, exist_ok=True)
#         output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
#         shutil.rmtree(output_folder, ignore_errors=True)
#         os.makedirs(output_folder, exist_ok=True)
       
#         tracker_file_path = os.path.join(output_folder, f"TRACKER_GAP_FILE_{milestone1}_{milestone2}_{circle}_{site_tagging}_{current_status}_{new_toco_name}_{current_date}_{current_time}.xlsx")
        
#         df.insert(0, "Unique ID", range(1, len(df) + 1))
        
       
#         df.drop(columns=['id'], inplace=True)

#         df = df[['Unique ID', 'circle', 'site_tagging', 'old_toco_name', 'old_site_id', 'new_site_id', 'new_toco_name',
#                  'sr_number', 'ran_oem', 'media_type', 'mw_oem', 'relocation_method', 'relocation_type', 'old_site_band',
#                  'new_site_band', 'rfai_date', 'allocation_date', 'rfai_survey_date', 'mo_punch_date',
#                  'material_dispatch_date', 'material_delivered_date', 'installation_start_date', 'installation_end_date',
#                  'integration_date', 'emf_submission_date', 'ran_lkf_status', 'alarm_status', 'alarm_rectification_done_date',
#                  'scft_done_date', 'scft_i_deploy_offered_date', 'ran_pat_offer_date', 'ran_sat_offer_date', 'mw_plan_id',
#                  'mw_pat_offer_date', 'rsl_value_status', 'enm_status', 'mw_lkf', 'mw_sat_offer_date', 'mw_ms1_mids_date',
#                  'site_onair_date', 'i_deploy_onair_date', 'current_status', 'detailed_remarks', 'manual_history', 'rfai_rejected_date', 
#                  'clear_rfai_date', 'pri_count', 'pri_issue_ageing', 'other_issue_ageing', 'total_issue_ageing', 
#                   'rfai_to_ms1_ageing', 'ran_pat_accepted_date', 'ran_sat_accepted_date', 
#                  'mw_pat_accepted_date', 'mw_sat_accepted_date', 'scft_accepted_date', 'kpi_at_offer_date', 'kpi_at_accepted_date',
#                  'four_g_ms2_date', 'five_g_ms2_date', 'final_ms2_date', "dismantling_survey_date", "sreq_creq_raised_date",
#                  "dismantle_date", "material_pickup_date", "material_submission_date", "oci_done_date", "sign_off_date", "toco_owner_issue",
#                  "exit_notice_issue", "commercial_issue", "workable_sites", 'last_updated_date', 'last_updated_by']]
 
 
#         # json_dict_data = df.to_dict(orient="records")
#         # json_val = json.dumps(json_dict_data)
        
#         template_path = os.path.join(BASE_URL, "template", "templateAlok_v.1.xlsx")
#         wb = load_workbook(template_path)
#         ws = wb.active  
#         date_columns = [col for col in df.columns if col.endswith("_date") and col != 'last_updated_date']

#         for col in date_columns:
#             df[col] = pd.to_datetime(df[col], errors="coerce")

#         start_row = 3
#         for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
#             for c_idx, value in enumerate(row, start=1):
#                 cell = ws.cell(row=r_idx, column=c_idx)
#                 cell.value = value

#                 # apply date format for *_date columns
#                 if df.columns[c_idx - 1].endswith("_date") and hasattr(value, "strftime") and df.columns[c_idx - 1] != 'last_updated_date':
#                     cell.number_format = "dd-mmm-yy"
 

#         # ws2 = wb.create_sheet(title="Issues Tracker")

#         # # Clean issue_df (remove id if present)
#         # if "id" in issue_df.columns:
#         #     issue_df = issue_df.drop(columns=["id"])
            
#         # # print(issue_df['updated_by'])

#         # def remove_tz_safe(x):
#         #     if isinstance(x, datetime) and x.tzinfo is not None:
#         #         return x.replace(tzinfo=None)
#         #     return x

#         # # apply to every cell in issue_df
#         # issue_df = issue_df.applymap(remove_tz_safe)
        
#         # # print(issue_df['updated_by'])


#         # # Convert date columns
#         # issue_date_columns = [col for col in issue_df.columns if "Date" in col]
#         # for col in issue_date_columns:
#         #     issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
            
#         # # print(issue_df['updated_by'])

#         # # ---- Write header ----
#         # for col_idx, column_name in enumerate(issue_df.columns, start=1):
#         #     ws2.cell(row=1, column=col_idx, value=column_name)
            
#         # # print(issue_df['updated_by'])

#         # # ---- Write data ----
#         # for r_idx, row in enumerate(dataframe_to_rows(issue_df, index=False, header=False), start=2):
#         #     for c_idx, value in enumerate(row, start=1):
#         #         cell = ws2.cell(row=r_idx, column=c_idx)
#         #         cell.value = value

#         #         # Apply date formatting
#         #         if issue_df.columns[c_idx - 1] in issue_date_columns and hasattr(value, "strftime"):
#         #             cell.number_format = "dd-mmm-yy"

#         wb.save(tracker_file_path)
        
#         relative_path = tracker_file_path.replace(settings.MEDIA_ROOT, '').lstrip(os.sep)
#         download_url = request.build_absolute_uri(
#             os.path.join(settings.MEDIA_URL, relative_path).replace('\\', '/')
#         )
    
        
        
#         return Response({'message': 'request processed successfully !!!', "download_link": download_url}, status=200)
#     except Exception as e:
#         return Response({"error": f"{str(e)}"},status=500)   



######################################################################################################################################################
####################################################################################################################################################
####################################################### OLD CODE (JUST IN CASE NEEDED) ##############################################################
#################################################################################################################################################
#####################################################################################################################################################


# def old code:
#     def rebuild_histories_from_timeline(timeline_df):
#         """
#         Takes the timeline DF created by build_issue_timeline()
#         and reconstructs pri_history and issue_history strings.
#         """

#         # Ensure sorting by Start Date for correct ordering
#         timeline_df = timeline_df.copy()
#         timeline_df["SortKey"] = pd.to_datetime(timeline_df["Start Date"], format="%d-%b-%y")
#         timeline_df = timeline_df.sort_values(by="SortKey").drop(columns=["SortKey"]).reset_index(drop=True)

#         pri_entries = []
#         issue_entries = []
#         pri_count = 0
#         issue_order_counter = {}  # for separate issue numbering

#         for _, row in timeline_df.iterrows():
#             issue_name = row["Issue Name"]
#             start = row["Start Date"]
#             close = row["Close Date"]

#             if issue_name == "PRI":
#                 pri_count += 1
#                 pri_entries.append(f"S{pri_count}: {start}, C{pri_count}: {close}")
#             else:
#                 # Maintain numbering per issue type if ever required in future logic
#                 if issue_name not in issue_order_counter:
#                     issue_order_counter[issue_name] = 1
#                 else:
#                     issue_order_counter[issue_name] += 1
                
#                 issue_entries.append(f"{issue_name}: S: {start}, C: {close}")

#         pri_history = "; ".join(pri_entries) if pri_entries else ""
#         issue_history = "; ".join(issue_entries) if issue_entries else ""

#         return pri_history, issue_history

#     def add_issue_to_timeline(timeline_df, issue, start, close):
#         today = datetime.today()

#         print("TIMELINE DF", timeline_df)

#         # Normalize inputs
#         issue = issue.strip()
#         start = normalize_date_format(start)
#         close = normalize_date_format(close)

#         # Convert to datetime
#         start_date = datetime.strptime(start, "%d-%b-%y")
#         close_date = None if close in ["", "-", "nan", None, "undefined"] else datetime.strptime(close, "%d-%b-%y")

#         if not timeline_df.empty:
#             if ((timeline_df["Issue Name"] == issue) & (timeline_df["Close Date"].astype(str).isin(["", "-", "nan", "undefined", None]))).any():
#                 raise ValueError(f"Issue '{issue}' is already Open and cannot be added again.")

#         # Add new row to DF
#         new_row = {
#             "Issue Name": issue,
#             "Start Date": start_date,
#             "Close Date": close_date,
#             # "Status": "Open" if close_date in [None, "", "-", "nan", "undefined"] else "Closed",
#             # "Duration Days": (today - start_date).days if close_date is None else (close_date - start_date).days
#         }
        
#         timeline_df = pd.concat(
#             [timeline_df, pd.DataFrame([new_row])],
#             ignore_index=True
#         )

#         print("timeline_df", timeline_df)

#         # Sort by Start Date
#         timeline_df["Start Date"] = pd.to_datetime(timeline_df["Start Date"], errors="coerce")
#         timeline_df["Close Date"] = pd.to_datetime(timeline_df["Close Date"], errors="coerce")
#         timeline_df = timeline_df.sort_values(by="Start Date").reset_index(drop=True)


#         print("timeline_df", timeline_df)

#         # Recompute Status column
#         timeline_df["Status"] = timeline_df["Close Date"].apply(
#             lambda x: "Open" if (pd.isna(x) or (x in [None, "", "-", "nan", "undefined"])) else "Closed"
#         )

#         # Recompute Duration Days
#         timeline_df["Duration Days"] = timeline_df.apply(
#             lambda row: (today - row["Start Date"]).days
#             if pd.isna(row["Close Date"])
#             else (row["Close Date"] - row["Start Date"]).days,
#             axis=1
#         )

#         # Reformat dates for final display
#         timeline_df["Start Date"] = timeline_df["Start Date"].dt.strftime("%d-%b-%y")
#         timeline_df["Close Date"] = timeline_df["Close Date"].apply(
#             lambda x: "-" if pd.isna(x) else x.strftime("%d-%b-%y")
#         )

#         return timeline_df

#     def build_issue_timeline(df):
#         rows = []
#         today = datetime.today()

#         # --- PRI History ---
#         pri_history = df.at[0, "pri_history"]
#         if pd.notna(pri_history) and pri_history.strip().lower() not in ["", "nan", "none"]:
#             pri_entries = [x.strip() for x in pri_history.split(";") if x.strip()]
#             for entry in pri_entries:
#                 parts = entry.split(",")
#                 start_str = parts[0].split(":")[1].strip()
#                 close_str = parts[1].split(":")[1].strip()

#                 start_date = datetime.strptime(start_str, "%d-%b-%y")
#                 close_date = None if close_str == "-" else datetime.strptime(close_str, "%d-%b-%y")

#                 rows.append({
#                     "Issue Name": "PRI",
#                     "Start Date": start_date,
#                     "Close Date": close_date
#                 })
                
#         print("A")

#         # --- Other Issue History ---
#         issue_history = df.at[0, "issue_history"]
#         if pd.notna(issue_history) and issue_history.strip().lower() not in ["", "nan", "none"]:
#             entries = [x.strip() for x in issue_history.split(";") if x.strip()]
#             for entry in entries:
#                 issue_name = entry.split(":")[0].strip()
#                 parts = entry.split(",")

#                 start_str = parts[0].split("S:")[1].strip()
#                 close_str = parts[1].split("C:")[1].strip()

#                 start_date = datetime.strptime(start_str, "%d-%b-%y")
#                 close_date = None if close_str == "-" else datetime.strptime(close_str, "%d-%b-%y")

#                 rows.append({
#                     "Issue Name": issue_name,
#                     "Start Date": start_date,
#                     "Close Date": close_date
#                 })
                
#         print("B")

#         # Form DataFrame
#         timeline_df = pd.DataFrame(rows)

#         if timeline_df.empty:
#             return timeline_df

#         # Sort by Start Date
#         timeline_df = timeline_df.sort_values(by="Start Date").reset_index(drop=True)

#         # Status column
#         timeline_df["Status"] = timeline_df["Close Date"].apply(
#             lambda x: "Open" if (pd.isna(x) or (x in [None, "", "-", "nan", "undefined"])) else "Closed"
#         )
        
#         # Duration days (open → till today)
#         timeline_df["Duration Days"] = timeline_df.apply(
#             lambda row: (today - row["Start Date"]).days 
#                 if pd.isna(row["Close Date"]) 
#                 else (row["Close Date"] - row["Start Date"]).days,
#             axis=1
#         )
        
#         print("C")

#         # Convert dates to required format: DD-MMM-YY
#         timeline_df["Start Date"] = timeline_df["Start Date"].dt.strftime("%d-%b-%y")
#         print("D")
        
#         print(timeline_df)
        
#         timeline_df["Close Date"] = timeline_df["Close Date"].apply(
#             lambda x: "-" if pd.isna(x) else x.strftime("%d-%b-%y")
#         )
        
#         return timeline_df

#     def normalize_date_format(date_str):
#         """
#         Accepts either format:
#         - 01-Nov-25
#         - 2025-11-01
#         Converts to: DD-MMM-YY (e.g., 01-Nov-25)
#         """
#         date_str = date_str.strip()

#         if not date_str or date_str.lower() in ["nan", "none", "-", "undefined", ""]:
#             return date_str  # leave open markers or blank untouched

#         # Try old style: 01-Nov-25
#         try:
#             dt = datetime.strptime(date_str, "%d-%b-%y")
#             return dt.strftime("%d-%b-%y")
#         except:
#             pass

#         # Try new style: 2025-11-01
#         try:
#             dt = datetime.strptime(date_str, "%Y-%m-%d")
#             return dt.strftime("%d-%b-%y")
#         except:
#             pass

#         return date_str

#     def calculate_pri_ageing(history):
#         if not history or history.lower() in ["nan", "none", ""]:
#             return 0
        
#         print("1.0.1")

#         cycles = [x.strip() for x in history.split(";") if x.strip()]
#         today = datetime.today()

#         print("1.0.2")

#         ranges = []

#         # Parse raw ranges
#         for cycle in cycles:
#             if not (cycle in ["nan", "", " ", "undefined", None]):
#                 parts = cycle.split(",")
#                 start_str = parts[0].split(":")[1].strip()
#                 close_str = parts[1].split(":")[1].strip()

#                 start_date = datetime.strptime(start_str, "%d-%b-%y")
#                 end_date = today if close_str == "-" else datetime.strptime(close_str, "%d-%b-%y")

#                 ranges.append((start_date, end_date))

#         print("1.0.3")

#         # Sort by start date
#         ranges.sort(key=lambda x: x[0])

#         print("1.0.4")

#         # Merge overlapping ranges
#         merged = []
#         current_start, current_end = ranges[0]

#         for start, end in ranges[1:]:
#             if start <= current_end:  
#                 current_end = max(current_end, end)  # extend overlap
#             else:
#                 merged.append((current_start, current_end))
#                 current_start, current_end = start, end

#         print("1.0.5")

#         merged.append((current_start, current_end))

#         # Sum merged durations
#         total_days = sum((end - start).days for start, end in merged)

#         print("1.0.6")

#         return max(total_days, 0)

#     def calculate_other_issue_ageing(issue_history: str) -> int:
#         """
#         Calculate union of ageing days (no double counting overlapping issue ranges)
#         """
#         if not issue_history or str(issue_history).strip().lower() in ["nan", "none", ""]:
#             return 0

#         intervals = []
#         today = datetime.today()

#         entries = [x.strip() for x in issue_history.split(";") if x.strip()]
        
#         for entry in entries:
#             try:
#                 parts = entry.split(",")
#                 start_str = parts[0].split("S:")[1].strip()
#                 close_str = parts[1].split("C:")[1].strip()

#                 start_date = datetime.strptime(start_str, "%d-%b-%y")

#                 if close_str == "-":
#                     end_date = today
#                 else:
#                     end_date = datetime.strptime(close_str, "%d-%b-%y")

#                 if end_date >= start_date:
#                     intervals.append((start_date, end_date))
#             except:
#                 continue
        
#         if not intervals:
#             return 0

#         # 🔥 Merge overlapping intervals for union ageing
#         intervals.sort()  # sort by start date

#         merged = [intervals[0]]
#         for current in intervals[1:]:
#             prev_start, prev_end = merged[-1]
#             cur_start, cur_end = current

#             if cur_start <= prev_end:
#                 merged[-1] = (prev_start, max(prev_end, cur_end))  # merge
#             else:
#                 merged.append(current)

#         # 🔥 Sum merged intervals
#         total_days = sum((end - start).days for start, end in merged)

#         return max(int(total_days), 0)

#     def parse_pri_intervals(history):
#         if not history or history.lower() in ["nan", "none", ""]:
#             return []

#         cycles = [x.strip() for x in history.split(";") if x.strip()]
#         today = datetime.today()
#         intervals = []

#         for cycle in cycles:
#             parts = cycle.split(",")
#             start_str = parts[0].split(":")[1].strip()
#             close_str = parts[1].split(":")[1].strip()

#             start_date = datetime.strptime(start_str, "%d-%b-%y")
#             end_date = today if close_str == "-" else datetime.strptime(close_str, "%d-%b-%y")

#             intervals.append((start_date, end_date))

#         return intervals

#     def parse_other_issue_intervals(issue_history):
#         if not issue_history or str(issue_history).strip().lower() in ["nan", "none", ""]:
#             return []

#         today = datetime.today()
#         intervals = []

#         entries = [x.strip() for x in issue_history.split(";") if x.strip()]
#         for entry in entries:
#             try:
#                 parts = entry.split(",")
#                 start_str = parts[0].split("S:")[1].strip()
#                 close_str = parts[1].split("C:")[1].strip()

#                 start_date = datetime.strptime(start_str, "%d-%b-%y")
#                 end_date = today if close_str == "-" else datetime.strptime(close_str, "%d-%b-%y")

#                 if end_date >= start_date:
#                     intervals.append((start_date, end_date))

#             except:
#                 continue

#         return intervals

#     def calculate_total_issue_ageing(pri_history, other_history):
#         # Parse both sets of intervals
#         pri_intervals = parse_pri_intervals(pri_history)
#         other_intervals = parse_other_issue_intervals(other_history)

#         # Combine
#         intervals = pri_intervals + other_intervals
#         if not intervals:
#             return 0

#         # Sort by start date
#         intervals.sort(key=lambda x: x[0])

#         # Merge all overlapping intervals
#         merged = [intervals[0]]
#         for start, end in intervals[1:]:
#             prev_start, prev_end = merged[-1]

#             if start <= prev_end:
#                 merged[-1] = (prev_start, max(prev_end, end))
#             else:
#                 merged.append((start, end))

#         # Sum unique ageing days
#         total_days = sum((end - start).days for start, end in merged)
#         return max(total_days, 0)
    

#     @api_view(['GET', 'POST'])
#     def issue_timeline_display(request):
#         userId = request.data.get('userId').lower()
#         siteId = request.data.get('siteId')
        
#         try:
#             obj = AlokTrackerModel.objects.filter(new_site_id=siteId).first()
            
#             if obj is None:
#                 return Response({"error": "Site not found!"}, status=500)
#             else:
#                 df = pd.DataFrame([model_to_dict(obj)])
            
#             circle = ACCESS_RIGHTS[userId]['Circle']
            
#             if circle != 'CENTRAL' and df.at[0, 'circle'] != circle:
#                 return Response({"error": "Access denied!"}, status=500)
            
#             df_issue_history = df[['pri_history', 'issue_history']]
            
#             df_issue_history = build_issue_timeline(df_issue_history)
            
#             print("C")

#             if not df_issue_history.empty:
#                 df_issue_history = df_issue_history.applymap(
#                     lambda x: x.strftime("%d-%b-%y") if hasattr(x, 'strftime') else x
#                 )
                
#             df_issue_history.index = range(1, len(df_issue_history) + 1)
#             df_issue_history['Index'] = df_issue_history.index
            
#             result_json = df_issue_history.to_dict(orient="records")
#             json_data = json.dumps(result_json)
            
#             return Response({'message': 'request processed successfully !!!', "json_data" : json_data }, status = 200)
#         except Exception as e:
#             return Response({"error": f"{str(e)}"},status=500)
    
    
#     @api_view(['GET', 'POST'])      
#     def issue_timeline_add(request):
#         userId = request.data.get('userId').lower()
#         siteId = request.data.get('siteId')
#         issue = request.data.get('issue')
#         start = request.data.get('start_date')
#         close = request.data.get('close_date')
#         timeline = request.data.get('timeline')

#         if not userId or not siteId or not issue or not start or not timeline:
#             return Response({'error': 'userId, siteId, issue, start_date and timeline are required.'}, status=400)
        
#         try:
#             obj = AlokTrackerModel.objects.filter(new_site_id=siteId).first()
            
#             if obj is None:
#                 return Response({"error": "Site not found!"}, status=500)
#             else:
#                 df = pd.DataFrame([model_to_dict(obj)])
                
#             issue_timeline = json.loads(timeline)
#             issue_timeline_df = pd.DataFrame(issue_timeline)
            
#             issue_timeline_df = add_issue_to_timeline(issue_timeline_df, issue, start, close)
            
#             # issue_timeline_df = issue_timeline_df.astype(str)
#             for col in ["Start Date", "Close Date"]:
#                 issue_timeline_df[col] = issue_timeline_df[col].astype(str).replace("NaT", "-")
#             for col in ["Duration Days"]:
#                 issue_timeline_df[col] = issue_timeline_df[col].apply(lambda x: int(x) if isinstance(x, (int, float)) and not pd.isna(x) else "-")
#             print("ISSUE TIMELINE DF")
#             print(issue_timeline_df)
#             print("AAAAAAAAAAAAAAAAAAAAAAAAAAA")
#             issue_timeline_df.index = range(1, len(issue_timeline_df) + 1)
#             issue_timeline_df['Index'] = issue_timeline_df.index
#             json_data = issue_timeline_df.to_dict(orient="records")
            
#             pri_history, issue_history = rebuild_histories_from_timeline(issue_timeline_df)

#             print("PRI History:", pri_history)
#             print("Issue History:", issue_history)
            
#             df.at[0, 'pri_history'] = pri_history
#             df.at[0, 'issue_history'] = issue_history
            
#             df.at[0, 'pri_count'] = (issue_timeline_df['Issue Name'] == 'PRI').sum()
            
#             df.at[0, 'pri_issue_ageing'] = calculate_pri_ageing(pri_history)
#             df.at[0, 'other_issue_ageing'] = calculate_other_issue_ageing(issue_history)
#             df.at[0, 'total_issue_ageing'] = calculate_total_issue_ageing(pri_history, issue_history)
            
#             df = df.where(pd.notnull(df), None)
#             df.fillna("", inplace=True)
#             df = df.replace("NaN", None)
#             df = df.replace("", None)
            
#             record = df.iloc[0].to_dict()
            
#             print("RECORD", record)

#             # ✅ Safe converters
#             INVALID_DATE_STRINGS = {"need to check", "pending", "na", "tbd", "n/a", "nan", "-"}

#             def safe_datetime(value):
#                 import pandas as pd
#                 from datetime import datetime as dtime
#                 if pd.isna(value) or value is pd.NaT:
#                     return None
#                 if isinstance(value, (pd.Timestamp, dtime)):
#                     return value
#                 if isinstance(value, str):
#                     v = value.strip().lower()
#                     if v in INVALID_DATE_STRINGS or not v:
#                         return None
#                     try:
#                         val = pd.to_datetime(value, errors="coerce")
#                         if pd.isna(val):
#                             return None
#                         return val.to_pydatetime()
#                     except Exception:
#                         return None
#                 return None

#             def safe_int(val):
#                 try:
#                     return int(val)
#                 except (TypeError, ValueError):
#                     return None

#             # ✅ Extract identifiers
#             circle_val = str(record.get("circle")).strip() if record.get("circle") else None
#             new_site_id_val = str(record.get("new_site_id")).strip() if record.get("new_site_id") else None

#             if not circle_val or not new_site_id_val:
#                 return Response({'error': 'Circle and New Site ID are required.'}, status=400)

#             # # ✅ Circle restriction
#             # if circle.lower() != "central" and circle_val.lower() != circle.lower():
#             #     return Response({'error': f'You are not allowed to edit {circle_val} circle data.'}, status=403)

#             # ✅ Validate columns
#             required_columns = ['pri_history', 'issue_history', 'pri_issue_ageing', 'other_issue_ageing', 'total_issue_ageing']
#             # required_columns = [
#             #     col.lower()
#             #     .strip()
#             #     .replace(" ", "_")
#             #     .replace("-", "_")
#             #     .replace("4", "four_")
#             #     .replace("5", "five_")
                    # .replace("/", "_")
#             #     for col in ACCESS_RIGHTS[userId]["Columns"]
#             # ]
#             allowed_data = {}

#             for col in record:
#                 if col not in required_columns:
#                     continue

#                 val = record[col]
#                 if "date" in col.lower():
#                     val = safe_datetime(val)
#                 elif "ageing" in col.lower() or "count" in col.lower():
#                     val = safe_int(val)

#                 field_name = col.lower().strip().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_")
#                 allowed_data[col] = val

#             # ✅ Add system fields
#             allowed_data.update({
#                 "last_updated_by": userId,
#                 "last_updated_date": timezone.now(),
#             })

#             # ✅ Find existing record
#             try:
#                 obj = AlokTrackerModel.objects.get(circle__iexact=circle_val, new_site_id__iexact=new_site_id_val)
#             except AlokTrackerModel.DoesNotExist:
#                 return Response({'error': 'Record not found in database.'}, status=404)

#             # ✅ Update object fields
#             for field, value in allowed_data.items():
#                 setattr(obj, field, value)

#             obj.save(update_fields=list(allowed_data.keys()))
            
#             return Response({"message": "request processed successfully", "json_data": json_data}, status = 200)
#         except Exception as e:
#             return Response({"error": f"{str(e)}"},status=500)
        
        
#     @api_view(['GET', 'POST'])
#     def issue_timeline_update(request):
#         userId = request.data.get('userId').lower()
#         siteId = request.data.get('siteId')
#         timeline = request.data.get('timeline')
        
#         try:
#             obj = AlokTrackerModel.objects.filter(new_site_id=siteId).first()
            
#             if obj is None:
#                 return Response({"error": "Site not found!"}, status=500)
#             else:
#                 df = pd.DataFrame([model_to_dict(obj)])
                
#             issue_timeline = json.loads(timeline)
#             issue_timeline_df = pd.DataFrame(issue_timeline)
#             issue_timeline_df["Start Date"] = issue_timeline_df["Start Date"].apply(normalize_date_format)
#             issue_timeline_df["Close Date"] = issue_timeline_df["Close Date"].apply(normalize_date_format)

#             if((issue_timeline_df["Start Date"].astype(str).isin(["", "-", "nan", "undefined", None]))).any():
#                 return Response({"error": "Start Date cannot be empty"}, status=500) 

#             # def rebuild_histories_from_timeline(timeline_df):
#             #     """
#             #     Takes the timeline DF created by build_issue_timeline()
#             #     and reconstructs pri_history and issue_history strings.
#             #     """

#             #     # Ensure sorting by Start Date for correct ordering
#             #     timeline_df = timeline_df.copy()
#             #     timeline_df["SortKey"] = pd.to_datetime(timeline_df["Start Date"], format="%d-%b-%y")
#             #     timeline_df = timeline_df.sort_values(by="SortKey").drop(columns=["SortKey"]).reset_index(drop=True)

#             #     pri_entries = []
#             #     issue_entries = []
#             #     pri_count = 0
#             #     issue_order_counter = {}  # for separate issue numbering

#             #     for _, row in timeline_df.iterrows():
#             #         issue_name = row["Issue Name"]
#             #         start = row["Start Date"]
#             #         close = row["Close Date"]

#             #         if issue_name == "PRI":
#             #             pri_count += 1
#             #             pri_entries.append(f"S{pri_count}: {start}, C{pri_count}: {close}")
#             #         else:
#             #             # Maintain numbering per issue type if ever required in future logic
#             #             if issue_name not in issue_order_counter:
#             #                 issue_order_counter[issue_name] = 1
#             #             else:
#             #                 issue_order_counter[issue_name] += 1
                        
#             #             issue_entries.append(f"{issue_name}: S: {start}, C: {close}")

#             #     pri_history = "; ".join(pri_entries) if pri_entries else ""
#             #     issue_history = "; ".join(issue_entries) if issue_entries else ""

#             #     return pri_history, issue_history
            
#             pri_history, issue_history = rebuild_histories_from_timeline(issue_timeline_df)

#             print("PRI History:", pri_history)
#             print("Issue History:", issue_history)
            
#             df.at[0, 'pri_history'] = pri_history
#             df.at[0, 'issue_history'] = issue_history
            
#             df.at[0, 'pri_count'] = (issue_timeline_df['Issue Name'] == 'PRI').sum()
            
#             df.at[0, 'pri_issue_ageing'] = calculate_pri_ageing(pri_history)
#             df.at[0, 'other_issue_ageing'] = calculate_other_issue_ageing(issue_history)
#             df.at[0, 'total_issue_ageing'] = calculate_total_issue_ageing(pri_history, issue_history)
            
#             updated_issue_timeline = build_issue_timeline(df[['pri_history', 'issue_history']])
#             updated_issue_timeline.index = range(1, len(updated_issue_timeline) + 1)
#             updated_issue_timeline['Index'] = updated_issue_timeline.index
#             json_data = updated_issue_timeline.to_dict(orient="records")
            
#             df = df.where(pd.notnull(df), None)
#             df.fillna("", inplace=True)
#             df = df.replace("NaN", None)
#             df = df.replace("", None)
            
#             record = df.iloc[0].to_dict()
            
#             print(record)

#             # ✅ Safe converters
#             INVALID_DATE_STRINGS = {"need to check", "pending", "na", "tbd", "n/a", "nan", "-"}

#             def safe_datetime(value):
#                 import pandas as pd
#                 from datetime import datetime as dtime
#                 if pd.isna(value) or value is pd.NaT:
#                     return None
#                 if isinstance(value, (pd.Timestamp, dtime)):
#                     return value
#                 if isinstance(value, str):
#                     v = value.strip().lower()
#                     if v in INVALID_DATE_STRINGS or not v:
#                         return None
#                     try:
#                         val = pd.to_datetime(value, errors="coerce")
#                         if pd.isna(val):
#                             return None
#                         return val.to_pydatetime()
#                     except Exception:
#                         return None
#                 return None

#             def safe_int(val):
#                 try:
#                     return int(val)
#                 except (TypeError, ValueError):
#                     return None

#             # ✅ Extract identifiers
#             circle_val = str(record.get("circle")).strip() if record.get("circle") else None
#             new_site_id_val = str(record.get("new_site_id")).strip() if record.get("new_site_id") else None

#             if not circle_val or not new_site_id_val:
#                 return Response({'error': 'Circle and New Site ID are required.'}, status=400)

#             # # ✅ Circle restriction
#             # if circle.lower() != "central" and circle_val.lower() != circle.lower():
#             #     return Response({'error': f'You are not allowed to edit {circle_val} circle data.'}, status=403)

#             # ✅ Validate columns
#             required_columns = ['pri_history', 'issue_history', 'pri_issue_ageing', 'other_issue_ageing', 'total_issue_ageing']
#             # required_columns = [
#             #     col.lower()
#             #     .strip()
#             #     .replace(" ", "_")
#             #     .replace("-", "_").replace("/", "_")
#             #     .replace("4", "four_")
#             #     .replace("5", "five_")
#             #     for col in ACCESS_RIGHTS[userId]["Columns"]
#             # ]
#             allowed_data = {}

#             for col in record:
#                 if col not in required_columns:
#                     continue

#                 val = record[col]
#                 if "date" in col.lower():
#                     val = safe_datetime(val)
#                 elif "ageing" in col.lower() or "count" in col.lower():
#                     val = safe_int(val)

#                 field_name = col.lower().strip().replace(" ", "_").replace("-", "_").replace("/", "_").replace("4", "four_").replace("5", "five_")
#                 allowed_data[col] = val

#             # ✅ Add system fields
#             allowed_data.update({
#                 "last_updated_by": userId,
#                 "last_updated_date": timezone.now(),
#             })

#             # ✅ Find existing record
#             try:
#                 obj = AlokTrackerModel.objects.get(circle__iexact=circle_val, new_site_id__iexact=new_site_id_val)
#             except AlokTrackerModel.DoesNotExist:
#                 return Response({'error': 'Record not found in database.'}, status=404)

#             # ✅ Update object fields
#             for field, value in allowed_data.items():
#                 setattr(obj, field, value)

#             obj.save(update_fields=list(allowed_data.keys()))
            
#             return Response({"message": "request processed successfully", "json_data": json_data}, status = 200)
#         except Exception as e:
#             return Response({"error": f"{str(e)}"},status=500)
        
        
#     ############################################################## Access Rights ##############################################################

#     ACCESS_RIGHTS = {
#         "manish.kumar4@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'Site Tagging', 'Old TOCO Name', 'Old Site Id', 'New Site ID', 'New TOCO Name', 'SR Number', 'RAN OEM', 
#                         'Media Type', 'Relocation Method', 'Relocation Type', 'OLD Site Band', 'New Site Band', 'RFAI Date', 
#                         'Allocation Date', 'RFAI Survey Date', 'MO Punch Date', 'Material Dispatch Date', 'Material Delivered Date', 
#                         'Installation Start Date', 'Installation End Date', 'EMF Submission Date', 'Site ONAIR Date', 
#                         'I-Deploy ONAIR Date', 'Current Status', 'RFAI Rejected Date', 'Re-RFAI Date', 'PRI Issue Ageing', 
#                         'Other Issue Ageing', 'Total Issue Ageing', 'PRI Count'],
#             "can_upload" : True
#         },
#         "mohsin.khan@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'Site Tagging', 'Old TOCO Name', 'Old Site Id', 'New Site ID', 'New TOCO Name', 'SR Number', 'RAN OEM', 
#                         'Media Type', 'Relocation Method', 'Relocation Type', 'OLD Site Band', 'New Site Band', 'RFAI Date', 
#                         'Allocation Date', 'RFAI Survey Date', 'MO Punch Date', 'Material Dispatch Date', 'Material Delivered Date', 
#                         'Installation Start Date', 'Installation End Date', 'EMF Submission Date', 'Site ONAIR Date', 
#                         'I-Deploy ONAIR Date', 'Current Status', 'RFAI Rejected Date', 'Re-RFAI Date'],
#             "can_upload" : True
#         },
#         "vimal.kumar@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'Site Tagging', 'Old TOCO Name', 'Old Site Id', 'New Site ID', 'New TOCO Name', 'SR Number', 'RAN OEM', 
#                         'Media Type', 'Relocation Method', 'Relocation Type', 'OLD Site Band', 'New Site Band', 'RFAI Date', 
#                         'Allocation Date', 'RFAI Survey Date', 'MO Punch Date', 'Material Dispatch Date', 'Material Delivered Date', 
#                         'Installation Start Date', 'Installation End Date', 'EMF Submission Date', 'Site ONAIR Date', 
#                         'I-Deploy ONAIR Date', 'Current Status', 'RFAI Rejected Date', 'Re-RFAI Date'],
#             "can_upload" : True
#         },




#         "shubham.mukeash@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'MW Plan ID', 'MW PAT Offer Date', 'RSL Value Status', 'ENM Status', 'MW LKF', 
#                         'MW SAT Offer Date', 'MW MS1 MIDS Date', 'MW PAT Accepted Date', 'MW SAT Accepted Date', 'MW OEM'],
#             "can_upload" : True
#         },




#         "abhishek.gupta@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'RAN SAT Offer Date', 'RAN SAT Accepted Date', 'Alarm Status', 'Alarm Rectification Done Date', 'RAN LKF Status'],
#             "can_upload" : True
#         },
#         "saurabh.sharma2@ust.com":{
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'RAN SAT Offer Date', 'RAN SAT Accepted Date'],
#             "can_upload" : True
#         },
#         "amit.rai2@mcpsinc.com" : {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'RAN SAT Offer Date', 'RAN SAT Accepted Date'],
#             "can_upload" : True
#         },


#         "mohit.batra@mcpsinc.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'RAN SAT Offer Date', 'RAN SAT Accepted Date', 'Integration Date', 'Alarm Status', 
#                         'RAN LKF Status', 'Alarm Rectification Done Date'],
#             "can_upload" : True
#         },
#         "mohit.batra@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'RAN SAT Offer Date', 'RAN SAT Accepted Date', 'Integration Date', 'Alarm Status', 
#                         'RAN LKF Status', 'Alarm Rectification Done Date'],
#             "can_upload" : True
#         },


#         "rahul.dahiya@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'Integration Date', 'RAN LKF Status', 'Alarm Status', 'Alarm Rectification Done Date'],
#             "can_upload" : True
#         },
#         "rahul.dahiya@mcpsinc.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'Integration Date', 'RAN LKF Status', 'Alarm Status', 'Alarm Rectification Done Date'],
#             "can_upload" : True
#         },
#         "chandan.kumar@mcpsinc.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'Integration Date', 'RAN LKF Status', 'Alarm Status', 'Alarm Rectification Done Date'],
#             "can_upload" : True
#         },
#         "aashish.sharma@ust.com":{
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'Integration Date', 'RAN LKF Status', 'Alarm Status', 'Alarm Rectification Done Date'],
#             "can_upload" : True
#         },
#         "narender.yadav@mcpsinc.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'Integration Date', 'RAN LKF Status', 'Alarm Status', 'Alarm Rectification Done Date'],
#             "can_upload" : True
#         },




#         "amodkumar.singh@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'SCFT Done Date', 'SCFT I-Deploy Offered Date', 'SCFT Accepted Date'],
#             "can_upload" : True
#         },




#         "chandan.rai@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'RAN PAT Offer Date', 'RAN PAT Accepted Date'],
#             "can_upload" : True
#         },
#         "vicky.kumar2@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'RAN PAT Offer Date', 'RAN PAT Accepted Date'],
#             "can_upload" : True
#         },
#         "kuldeep.charanSingh@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ['Circle', 'New Site ID', 'RAN PAT Offer Date', 'RAN PAT Accepted Date'],
#             "can_upload" : True
#         },




#         "girraj.singh@mcpsinc.in": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : True
#         },
#         "devansh.jain@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : True
#         },
#         "prerna": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : True
#         },

#         "alok.kumar2@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "deepak.kumaryadav@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "jitendra.katiyar@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "lalit.namdev2@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "ravi.dixit@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "amit.sikka@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "amit.sikka@mcpsinc.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "vipin.negi@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "vipin.negi@mcpsinc.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "rakesh.kumar3@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "rahul.dubey@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "mohit.kumar@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "saurabh.rathore@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "saurabh.rathore@mcpsinc.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "sonu.sharma@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "sonu.s@mcpsinc.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "shubham.gupta2@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "shubham.gupta@mcpsinc.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "krishna.kantverma@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "krishnakant.verma@mcpsinc.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "bharat.kamble@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "arun.kumar@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "rajeev.gandhi@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "manoj.kumarverma@ust.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "manoj.k2@mcpsinc.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : False
#         },
#         "narender.yadav@mcpsinc.com": {
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : True
#         },
#         "test1":{
#             "Circle": "CENTRAL",
#             "Columns": ALL_COLUMNS,
#             "can_upload" : True
#         }
        
#     }


#     @api_view(["POST"])
#     def upload_tracker_data_view(request):
#         userId = request.data.get('userId')
#         userId = userId.lower()
#         circle = ACCESS_RIGHTS[userId]["Circle"]
#         file = request.FILES.get("tracker_file")

#         if not userId or not circle:
#             return Response({'error': 'userId and circle are required.'}, status=400)

#         if not file:
#             return Response({"status": False, "message": "No file provided."}, status=400)

#         if not ACCESS_RIGHTS[userId]["can_upload"]:
#             return Response({'error': 'Upload Rights Not Granted'}, status=400)

#         try:
#             df = pd.read_csv(file, header=1) if file.name.endswith('.csv') else pd.read_excel(file, header=1)
#             df.columns = [col.strip() for col in df.columns]
#             df = df.where(pd.notnull(df), None)
#             df.fillna("", inplace=True)
#             df = df.replace("NaN", None)
#             df = df.replace("", None)

#             INVALID_DATE_STRINGS = {"need to check", "pending", "na", "tbd", "n/a", "nan", "-"}

#             def safe_datetime(value):
#                 import pandas as pd
#                 if pd.isna(value) or value is pd.NaT:
#                     return None
#                 if isinstance(value, (pd.Timestamp, dtime)):
#                     return value
#                 if isinstance(value, str):
#                     v = value.strip().lower()
#                     if v in INVALID_DATE_STRINGS or not v:
#                         return None
#                     try:
#                         val = pd.to_datetime(value, errors="coerce")
#                         if pd.isna(val):
#                             return None
#                         return val.to_pydatetime()
#                     except Exception:
#                         return None
#                 return None

#             for col in df.columns:
#                 if "date" in col:
#                     df[col] = df[col].apply(safe_datetime)

#             def safe_int(val):
#                 try:
#                     return int(val)
#                 except (TypeError, ValueError):
#                     return None

#             required_columns = ACCESS_RIGHTS[userId]["Columns"]
#             missing_columns = [col for col in required_columns if col not in df.columns]

#             if missing_columns:
#                 return Response(
#                     {"status": False, "error": f"Missing columns: {', '.join(missing_columns)}"},
#                     status=400
#                 )

#             df = df[required_columns]

#             df['New Site ID'] = df['New Site ID'].astype(str)
#             records = df.to_dict(orient='records')

#             existing_records = {
#                 (rec.circle.strip().lower(), str(rec.new_site_id).strip().lower()): rec
#                 for rec in AlokTrackerModel.objects.only("circle", "new_site_id")
#             }

#             objects_to_create, objects_to_update = [], []

#             with transaction.atomic():
#                 for row in records:
#                     circle_val = str(row.get("Circle")).strip() if row.get("Circle") else None
#                     new_site_id_val = str(row.get("New Site ID")).strip() if row.get("New Site ID") else None

#                     if not circle_val or not new_site_id_val:
#                         continue

#                     if circle.lower().strip() != "central" and circle_val.lower() != circle.lower().strip():
#                         continue

#                     allowed_data = {}
#                     for col in required_columns:
#                         val = row.get(col)

#                         if "date" in col.lower():
#                             val = safe_datetime(val)
#                         elif "ageing" in col.lower() or "count" in col.lower():
#                             val = safe_int(val)

#                         field_name = (
#                             col.lower()
#                             .strip()
#                             .replace(" ", "_")
#                             .replace("-", "_").replace("/", "_")
#                             .replace("4", "four_")
#                             .replace("5", "five_")
#                         )
#                         allowed_data[field_name] = val

#                     allowed_data.update({
#                         "circle": circle_val,
#                         "new_site_id": new_site_id_val,
#                         "last_updated_by": userId,
#                         "last_updated_date": timezone.now(),
#                     })

#                     key = (circle_val.lower(), new_site_id_val.lower())

#                     if key in existing_records:
#                         obj = existing_records[key]
#                         for field, value in allowed_data.items():
#                             setattr(obj, field, value)

#                         obj.last_updated_date = timezone.now()
#                         obj.last_updated_by = userId
#                         objects_to_update.append(obj)

#                     else:
#                         objects_to_create.append(AlokTrackerModel(**allowed_data))

#                 if objects_to_create:
#                     AlokTrackerModel.objects.bulk_create(objects_to_create, batch_size=500)

#                 if objects_to_update:
#                     AlokTrackerModel.objects.bulk_update(
#                         objects_to_update,
#                         fields=list({
#                             col.lower()
#                             .strip()
#                             .replace(" ", "_")
#                             .replace("-", "_").replace("/", "_")
#                             .replace("4", "four_")
#                             .replace("5", "five_")
#                             for col in required_columns
#                         }) + ['last_updated_date', 'last_updated_by'],
#                         batch_size=500,
#                     )

#             return Response({"status": True, "message": "Data inserted successfully."})

#         except Exception as e:
#             return Response({"status": False, "error": str(e)}, status=500)

#     @api_view(['POST'])
#     def download_tracker_data_view(request):
#         userId = request.data.get('userId')
#         userId = userId.lower()
#         circle = ACCESS_RIGHTS[userId]["Circle"]

#         if not userId or not circle:
#             return Response({'error': 'userId and circle are required.'}, status=400)

#         try:
#             obj = []
#             if circle == 'CENTRAL':
#                 obj = AlokTrackerModel.objects.all()
#                 issue_obj = RelocationIssue.objects.all()
#             else:
#                 obj = AlokTrackerModel.objects.filter(circle=circle)
#                 issue_obj = RelocationIssue.objects.filter(circle=circle)

#             df = pd.DataFrame(obj.values())
#             issue_df = pd.DataFrame(issue_obj.values())
            
#             rename_map = {
#                 "circle": "Circle",
#                 "site_id": "Site ID",
#                 "issue_owner": "Issue Owner",
#                 "milestone": "Milestone",
#                 "issue_name": "Issue Name",
#                 "start_date": "Start Date",
#                 "close_date": "Close Date",
#                 "status": "Status",
#                 "duration": "Duration",
#                 "updated_by": "Updated_by",
#                 "updated_at": "Updated_at",
#                 "created_by": "Created_by",
#                 "created_at": "Created_at"
#             }

#             print(issue_df)
            
#             issue_df = issue_df.rename(columns=rename_map)

#             print("1")

#             required_cols = ["Circle", "Site ID", "Issue Owner", "Milestone", "Issue Name", "Start Date", "Close Date", "Status", "Duration", "Updated_by", "Updated_at", "Created_by", "Created_at"]
#             for col in required_cols:
#                 if col not in issue_df.columns:
#                     issue_df[col] = None

#             # # Convert date columns safely
#             # for col in ["Start Date", "Close Date"]:
#             #     if col in issue_df.columns:
#             #         issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
            
#             # df = update_ageing(df, issue_df)

#             print("2")

#             for col in df.columns:
#                 if 'date' in col:
#                     if col != 'last_updated_date':
#                         converted = pd.to_datetime(df[col], errors='coerce')
#                         if converted.notna().sum() > 0:
#                             df[col] = converted.dt.strftime('%d-%b-%y')
#                     else:
#                         converted = pd.to_datetime(df[col], errors='coerce')
#                         if converted.notna().sum() > 0:
#                             df[col] = converted.dt.strftime('%d-%b-%y %H:%M:%S')

#             print("3")

#             current_date = dtime.now().strftime("%Y-%m-%d")
#             current_time = dtime.now().strftime("%H-%M-%S")

#             BASE_URL = os.path.join(settings.MEDIA_ROOT, "alok_sir_tracking")
#             os.makedirs(BASE_URL, exist_ok=True)

#             output_folder = os.path.join(BASE_URL, f"generated_files_{current_date}")
#             shutil.rmtree(output_folder, ignore_errors=True)
#             os.makedirs(output_folder, exist_ok=True)

#             tracker_file_path = os.path.join(output_folder, f"TRACKER_FILE_{current_date}_{current_time}.xlsx")
#             df.insert(0, "Unique ID", range(1, len(df) + 1))

#             df.drop(columns=['id'], inplace=True)

#             df = df[['Unique ID', 'circle', 'site_tagging', 'old_toco_name', 'old_site_id', 'new_site_id', 'new_toco_name',
#                      'sr_number', 'ran_oem', 'media_type', 'mw_oem', 'relocation_method', 'relocation_type', 'old_site_band',
#                      'new_site_band', 'rfai_date', 'allocation_date', 'rfai_survey_date', 'mo_punch_date',
#                      'material_dispatch_date', 'material_delivered_date', 'installation_start_date', 'installation_end_date',
#                      'integration_date', 'emf_submission_date', 'ran_lkf_status', 'alarm_status', 'alarm_rectification_done_date',
#                      'scft_done_date', 'scft_i_deploy_offered_date', 'ran_pat_offer_date', 'ran_sat_offer_date', 'mw_plan_id',
#                      'mw_pat_offer_date', 'rsl_value_status', 'enm_status', 'mw_lkf', 'mw_sat_offer_date', 'mw_ms1_mids_date',
#                      'site_onair_date', 'i_deploy_onair_date', 'current_status', 'detailed_remarks', 'manual_history', 'rfai_rejected_date', 
#                      'clear_rfai_date', 'pri_count', 'pri_issue_ageing', 'other_issue_ageing', 'total_issue_ageing', 
#                       'rfai_to_ms1_ageing', 'ran_pat_accepted_date', 'ran_sat_accepted_date', 
#                      'mw_pat_accepted_date', 'mw_sat_accepted_date', 'scft_accepted_date', 'kpi_at_offer_date', 'kpi_at_accepted_date',
#                      'four_g_ms2_date', 'five_g_ms2_date', 'final_ms2_date', 'last_updated_date', 'last_updated_by']]

#             template_path = os.path.join(BASE_URL, "template", "templateAlok_v.1.xlsx")
#             wb = load_workbook(template_path)
#             ws = wb.active

#             print("4")
            
#             date_columns = [col for col in df.columns if col.endswith("_date") and col != 'last_updated_date']

#             print("5")

#             for col in date_columns:
#                 df[col] = pd.to_datetime(df[col], errors="coerce")

#             start_row = 3
#             for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
#                 for c_idx, value in enumerate(row, start=1):
#                     cell = ws.cell(row=r_idx, column=c_idx)
#                     cell.value = value

#                     # print(df.columns[c_idx - 1])

#                     # apply date format for *_date columns
#                     if df.columns[c_idx - 1].endswith("_date") and hasattr(value, "strftime") and df.columns[c_idx - 1] != 'last_updated_date':
#                         cell.number_format = "dd-mmm-yy"


#             print("6")
            
#             ws2 = wb.create_sheet(title="Issues Tracker")

#             # Clean issue_df (remove id if present)
#             if "id" in issue_df.columns:
#                 issue_df = issue_df.drop(columns=["id"])
                
#             # print(issue_df['updated_by'])

#             def remove_tz_safe(x):
#                 if isinstance(x, datetime) and x.tzinfo is not None:
#                     return x.replace(tzinfo=None)
#                 return x

#             # apply to every cell in issue_df
#             issue_df = issue_df.applymap(remove_tz_safe)
            
#             # print(issue_df['updated_by'])


#             # Convert date columns
#             issue_date_columns = [col for col in issue_df.columns if "Date" in col]
#             for col in issue_date_columns:
#                 issue_df[col] = pd.to_datetime(issue_df[col], errors='coerce')
                
#             # print(issue_df['updated_by'])

#             # ---- Write header ----
#             for col_idx, column_name in enumerate(issue_df.columns, start=1):
#                 ws2.cell(row=1, column=col_idx, value=column_name)
                
#             # print(issue_df['updated_by'])

#             # ---- Write data ----
#             for r_idx, row in enumerate(dataframe_to_rows(issue_df, index=False, header=False), start=2):
#                 for c_idx, value in enumerate(row, start=1):
#                     cell = ws2.cell(row=r_idx, column=c_idx)
#                     cell.value = value

#                     # Apply date formatting
#                     if issue_df.columns[c_idx - 1] in issue_date_columns and hasattr(value, "strftime"):
#                         cell.number_format = "dd-mmm-yy"
                        
#             # print(issue_df['updated_by'])

#             wb.save(tracker_file_path)

#             relative_path = tracker_file_path.replace(settings.MEDIA_ROOT, '').lstrip(os.sep)
#             download_url = request.build_absolute_uri(
#                 os.path.join(settings.MEDIA_URL, relative_path).replace('\\', '/')
#             )

#             return Response({'message': f'Welcome {userId}! Credentials verified.', "download_link": download_url}, status=200)

#         except Exception as e:
#             return Response({"error": f"{str(e)}"}, status=500)
