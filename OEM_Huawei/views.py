from django.shortcuts import render
from django.db.models import *
from rest_framework.decorators import api_view
from rest_framework.response import Response
import pandas as pd
import os
from django.conf import settings
from rest_framework import status
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from urllib.parse import urljoin
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from .models import *
import zipfile
import shutil

# ---------------- Excel Format ----------------
def format_of_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    header_fill = PatternFill(start_color="E87529", end_color="E87529", fill_type="solid")
    header_font = Font(color="000000", bold=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Borders
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border

    # Column width
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 25

    wb.save(file_path)
@api_view(['GET', 'POST'])
def fileupload(request):

    # ---------------- GET ----------------
    if request.method == 'GET':
        alarms = HuaweiAlarm.objects.all().values(
            'oem',
            'alarm_name',
            'alarm_type',
            'sa_nsa',
            'noc_circle',
        )
        return Response(list(alarms), status=status.HTTP_200_OK)

    # ---------------- POST ----------------
    elif request.method == 'POST':

        file = request.FILES.get('file')

        if not file:
            return Response(
                {'error': 'File is required!'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Required columns (normalized uppercase)
        required_cols = {
            "OEM",
            "ALARM NAME",
            "SA/NSA",
            "NOC/CIRCLE",
            "REMARK",
            "ALARM TYPE"
        }

        try:
            valid_dfs = []

            # ---------- CSV ----------
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)

                df = df.dropna(how="all")
                df.columns = df.columns.str.strip().str.upper()

                if required_cols.issubset(set(df.columns)):
                    valid_dfs.append(df)

            # ---------- EXCEL ----------
            else:
                excel_data = pd.read_excel(file, sheet_name=None)

                for sheet_name, df in excel_data.items():

                    df = df.dropna(how="all")

                    # Normalize columns
                    df.columns = df.columns.str.strip().str.upper()

                    if required_cols.issubset(set(df.columns)):
                        valid_dfs.append(df)

            # ❌ No valid sheet found
            if not valid_dfs:
                return Response(
                    {'error': 'No sheet contains required columns!'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Clear old data
            HuaweiAlarm.objects.all().delete()

        except Exception as e:
            return Response(
                {'error': f'Error reading file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ---------- SAVE DATA ----------
        objs = []
        count = 0

        for df in valid_dfs:
            for _, row in df.iterrows():

                objs.append(
                    HuaweiAlarm(
                        oem=str(row.get("OEM", "")).strip(),
                        alarm_name=str(row.get("ALARM NAME", "")).strip(),
                        sa_nsa=str(row.get("SA/NSA", "")).strip(),
                        noc_circle=str(row.get("NOC/CIRCLE", "")).strip(),
                        remark=str(row.get("REMARK", "")).strip(),
                        alarm_type=str(row.get("ALARM TYPE", "")).strip()
                    )
                )
                count += 1

        # Bulk insert (fast)
        HuaweiAlarm.objects.bulk_create(objs)

        return Response(
            {
                "status": True,
                "message": f"{count} alarms saved successfully!"
            },
            status=status.HTTP_201_CREATED
        )

@api_view(['GET'])
def get_sites(request):
    try:
        data = list(Old_New.objects.values())

        return Response({
                "status": True,
                "data": data
        })

    except Exception as e:
        return Response({
                "status": False,
                "error": str(e)
        }, status=500)


@api_view(['DELETE'])
def delete_sites(request):
    try:
        Old_New.objects.all().delete()

        return Response({
                "status": True,
                "message": "Deleted successfully"
        })
    except Exception as e:
        return Response({
                "status": False,
                "error": str(e)
        }, status=500)


@api_view(['POST'])
def upload_site_list(request):
    file = request.FILES.get('site_file')

    if not file:
        return Response({
            "status": False,
            "error": "File required"
        }, status=400)

    if not file.name.endswith('.xlsx'):
        return Response({
            "status": False,
            "error": "Only Excel file allowed"
        }, status=400)

    try:
        Old_New.objects.all().delete()

        excel_file = pd.ExcelFile(file)
        df = None

        # Har sheet check karo
        for sheet_name in excel_file.sheet_names:
            temp_df = pd.read_excel(excel_file, sheet_name=sheet_name)

            temp_df.columns = temp_df.columns.astype(str).str.strip()

            if (
                'New SiteId' in temp_df.columns and
                'Old SiteId' in temp_df.columns
            ):
                df = temp_df
                break

        if df is None:
            return Response({
                "status": False,
                "error": "No sheet found containing 'New SiteId' and 'Old SiteId' columns"
            }, status=400)

        for _, row in df.iterrows():
            new_site = str(row.get('New SiteId', '')).strip()
            old_site = str(row.get('Old SiteId', '')).strip()

            if not new_site and not old_site:
                continue

            Old_New.objects.update_or_create(
                new_site=new_site,
                old_site=old_site
            )

        return Response({
            "status": True,
            "message": "Data uploaded successfully"
        })

    except Exception as e:
        return Response({
            "status": False,
            "error": str(e)
        }, status=500)

# ---------------- API ----------------
@api_view(['POST'])
def alarmfileUpload(request):

    alarm_files = request.FILES.getlist('alarm_file')
    mapping_file = request.FILES.get('mapping_file')


    if not alarm_files or not mapping_file:
        return Response(
                {'error': 'Both alarm_file and mapping_file are required!'},
                status=status.HTTP_400_BAD_REQUEST
        )

    # ---------------- Clean Old Output ----------------
    output_root = os.path.join(
        MEDIA_ROOT,
        "HUAWEI_OUTPUT"
    )

    final_output_folder = os.path.join(
        output_root,
        "OUTPUT"
    )

    zip_path = os.path.join(
        output_root,
        "OUTPUT.zip"
    )

    # delete old output folder
    if os.path.exists(final_output_folder):
        shutil.rmtree(final_output_folder)

    # delete old zip
    if os.path.exists(zip_path):
        os.remove(zip_path)
        
    # -------- Read files --------
    alarm_dfs = []
    for alarm_file in alarm_files:
        try:

            # ---------- Find Header Row ----------
            if alarm_file.name.endswith('.csv'):

                raw_df = pd.read_csv(
                    alarm_file,
                    header=None
                )

            else:

                raw_df = pd.read_excel(
                    alarm_file,
                    header=None
                )

            header_idx = None

            for idx, row in raw_df.iterrows():

                row_values = (
                    row.fillna("")
                    .astype(str)
                    .str.replace("\n", " ", regex=False)
                    .str.replace("\r", " ", regex=False)
                    .str.strip()
                    .tolist()
                )

                row_values_lower = [v.lower() for v in row_values]

                if (
                    "alarm source" in row_values_lower
                    and "name" in row_values_lower
                ):
                    header_idx = idx
                    break

                if (
                    "subarea" in row_values_lower
                    and "lte ne name" in row_values_lower
                    and "cell name" in row_values_lower
                ):
                    header_idx = idx
                    break 

            if header_idx is None:

                # normal status file (4G sheet etc.)
                header_idx = 0

            # reset pointer
            alarm_file.seek(0)

            # ---------- Read Again With Correct Header ----------
            if alarm_file.name.endswith('.csv'):

                df = pd.read_csv(
                    alarm_file,
                    header=header_idx
                )

            else:

                df = pd.read_excel(
                    alarm_file,
                    header=header_idx
                )

            df.columns = (
                df.columns
                .astype(str)
                .str.strip()
            )

            alarm_dfs.append(df)

        except Exception as e:
            return Response(
                {
                    'error':
                    f'Error reading {alarm_file.name}: {str(e)}'
                },
                status=status.HTTP_400_BAD_REQUEST
            )
    if not alarm_dfs:
        return Response(
            {'error': 'No valid files found'},
            status=status.HTTP_400_BAD_REQUEST
        )

    df_alarm = pd.concat(alarm_dfs, ignore_index=True)
    # ---------------- Separate 4G / 5G Status Files ----------------
    df_4g = pd.DataFrame()
    for df in alarm_dfs:
        cols = [str(c).strip() for c in df.columns]

        # ---------- 4G ----------
        if "LTE NE Name" in cols:

            temp_4g = df.copy()
            # ---------------- CLEAN COLUMNS ----------------
            temp_4g.columns = (
                temp_4g.columns
                .astype(str)
                .str.strip()
            )

            required_cols = [
                "LTE NE Name",
                "NE Connection Status",
                "Cell Name",
                "Administrative Status",
                "Operating Status"
            ]

            temp_4g = temp_4g[
                [c for c in required_cols if c in temp_4g.columns]
            ]

            # ---------------- SITE ID EXTRACTION ----------------
            temp_4g["SiteId"] = (
                temp_4g["LTE NE Name"]
                .astype(str)
                .str.split("_")
                .str[-1]
                .str.upper()
                .str.strip()
            )
            temp_4g = temp_4g[
                [
                    "SiteId",
                    "NE Connection Status",
                    "Cell Name",
                    "Administrative Status",
                    "Operating Status"
                ]
            ]
            temp_4g.rename(
                columns={
                    "Administrative Status": "Admin Status",
                    "Operating Status": "Op Status"
                },
                inplace=True
            )
            # ---------------- REMOVE NULL SITE IDs ----------------
            temp_4g = temp_4g[
                temp_4g["SiteId"].notna()
            ]

            temp_4g = temp_4g[
                temp_4g["SiteId"].astype(str).str.strip() != ""
            ]

            # ---------------- APPEND ----------------
            df_4g = pd.concat(
                [df_4g, temp_4g],
                ignore_index=True
            )
    # ---------------- Read Mapping File ----------------
    try:
        mapping_filename = mapping_file.name.lower()

        if mapping_filename.endswith('.csv'):
            df_map = pd.read_csv(mapping_file)

        elif mapping_filename.endswith(('.xlsx', '.xls')):

            excel_sheets = pd.read_excel(
                mapping_file,
                sheet_name=None
            )

            df_map = None

            for sheet_name, temp_df in excel_sheets.items():

                temp_df = temp_df.dropna(how="all")
                temp_df.columns = temp_df.columns.str.strip()

                if "SiteId" in temp_df.columns:
                    df_map = temp_df
                    print(f"Using mapping sheet: {sheet_name}")
                    break

            if df_map is None:
                return Response(
                    {
                        "error":
                        "No sheet found containing SiteId column"
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

        else:
            return Response(
                {'error': 'Unsupported mapping file format'},
                status=status.HTTP_400_BAD_REQUEST
            )

        df_map.columns = df_map.columns.str.strip()

    except Exception as e:
        return Response(
                {'error': f'Error reading mapping file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
        )

    # -------- Extract Site ID --------
    
    if "Alarm Source" in df_alarm.columns:
        df_alarm["SiteId"] = (
            df_alarm["Alarm Source"]
            .astype(str)
            .str.split("_")
            .str[-1]
            .str.strip()
        )

    else:
        df_alarm["SiteId"] = ""

    # ---------------- OSS NE Is Disconnected Handling ----------------
    
    mask = (
        df_alarm["Alarm Source"]
        .astype(str)
        .str.upper()
        .eq("OSS")
    )

    if "Location Information" in df_alarm.columns:

        df_alarm.loc[mask, "SiteId"] = (
            df_alarm.loc[mask, "Location Information"]
            .astype(str)
            .str.extract(r'neName=([^,]+)', expand=False)
            .fillna("")
            .str.replace(r'.*@', '', regex=True)
            .str.replace(r'-.*$', '', regex=True)
            .str.split("_")
            .str[-1]
            .str.strip()
            .str.upper()
        )
    print("OSS site id:",df_alarm.loc[mask, "SiteId"])
    
    needed_cols = [
                "SiteId",
                "Name",
        ]
    # -------- Keep only required columns --------
    df_alarm = df_alarm[[c for c in needed_cols if c in df_alarm.columns]]
    
    # ---------------- Load SA / NSA from DB ----------------
    alarm_master = HuaweiAlarm.objects.all()
    
    alarm_lookup = {
        str(obj.alarm_name).strip().lower(): {
            "sa_nsa": obj.sa_nsa,
            "alarm_name": obj.alarm_name,
            "alarm_type": obj.alarm_type,
            "noc_circle": obj.noc_circle
        }
        for obj in alarm_master
        if obj.alarm_name
    }

    # ---------------- Classify Alarms ----------------
    def classify_alarms(info):

        if pd.isna(info):
            return pd.Series([
                "No",
                "No Alarms",
                "No Alarms",
                "",
                ""
            ])

        alarm = str(info).strip().lower()

        matched_sa_nsa = []
        matched_alarm_names = []
        matched_alarm_types = []
        matched_noc_circles = []

        if alarm in alarm_lookup:

            details = alarm_lookup[alarm]

            if str(details["sa_nsa"]).upper() == "SA":
                matched_sa_nsa.append(
                    "SA"
                )

            elif str(details["sa_nsa"]).upper() == "NSA":
                matched_sa_nsa.append(
                    "NSA"
                )

            matched_alarm_names.append(
                details["alarm_name"]
            )

            matched_alarm_types.append(
                details["alarm_type"]
            )
            
            matched_noc_circles.append(
                details["noc_circle"]
            )


        return pd.Series([
            "Yes" if matched_alarm_names else "No",

            ", ".join(set(matched_sa_nsa))
            if matched_sa_nsa
            else "No Alarms",

            ", ".join(set(matched_alarm_names))
            if matched_alarm_names
            else "No Alarms",

            ", ".join(set(matched_alarm_types))
            if matched_alarm_types
            else "",
            ", ".join(set(matched_noc_circles))
            if matched_noc_circles
            else ""
        ])
    df_alarm[
    [
        "Alarm Status (Yes/No)",
        "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms",
        "Service Affecting Alarms 1",
        "Alarm Type",
        "NOC Circle"
    ]
    ] = (
        df_alarm["Name"]
        .apply(classify_alarms)
    )
    
    #---------------- Clean + Group ----------------

    def clean_and_merge(values, column_name=""):

        vals = [
                str(v).strip()
                for v in values
                if pd.notna(v)
                and str(v).strip()
                and str(v).strip().lower() != "nan"
        ]

        vals_set = set(vals)

        # ---------------- Alarm Status ----------------
        if column_name == "Alarm Status (Yes/No)":

                # if both Yes and No exist -> keep Yes
                if "Yes" in vals_set:
                    return "Yes"

                return "No"

        # ---------------- SA / NSA ----------------
        if column_name == "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms":

                vals_upper = {
                    str(v).strip().upper()
                    for v in vals_set
                }

                # Priority:
                # SITE DOWN > SA > NSA > No Alarms
                if "SITE DOWN" in vals_upper:
                    return "SITE DOWN"

                if "SA" in vals_upper:
                    return "SA"

                if "NSA" in vals_upper:
                    return "NSA"

                return "No Alarms"

        # ---------------- Other Columns ----------------
        cleaned = sorted(set(vals))

        return ", ".join(cleaned)
    
    df_alarm = (
            df_alarm.drop_duplicates()
            .groupby("SiteId", as_index=False)
            .agg(lambda x: clean_and_merge(x, x.name))
    )

    df_alarm = (
        df_alarm.groupby("SiteId", as_index=False)
        .agg({
            "Alarm Status (Yes/No)":
                lambda x: "Yes"
                if "Yes" in x.values
                else "No",

            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms":
                lambda x: (
                    "SITE DOWN"
                    if "SITE DOWN" in x.astype(str).values
                    else (
                        ", ".join(
                            sorted(
                                {
                                    i for i in x.dropna().astype(str)
                                    if i.strip() != "No Alarms"
                                }
                            )
                        ) or "No Alarms"
                    )
                ),
            "Service Affecting Alarms 1":
                lambda x: ", ".join(
                    sorted(
                        {
                            alarm.strip()
                            for i in x.dropna().astype(str)
                            for alarm in i.split(",")
                            if alarm.strip() and alarm.strip() != "No Alarms"
                        }
                    )
                ) or "No Alarms",

            "Alarm Type":
                lambda x: ",".join(
                    sorted(
                        set(
                            x.dropna().astype(str)
                        )
                    )
                ),
            "NOC Circle":
                lambda x: ",".join(
                    sorted(
                        set(
                            x.dropna().astype(str)
                        )
                    )
                )
        })
    )
    df_map["SiteId"] = (
        df_map["SiteId"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    df_alarm["SiteId"] = (
        df_alarm["SiteId"]
        .astype(str)
        .str.strip()
        .str.upper()
    )
    
    old_new_df = pd.DataFrame(
        list(
            Old_New.objects.values(
                "old_site",
                "new_site"
            )
        )
    )

    for col in old_new_df.columns:
        old_new_df[col] = (
            old_new_df[col]
            .astype(str)
            .str.strip()
            .str.upper()
        )

    old_site_set = set(old_new_df["old_site"])
    new_site_set = set(old_new_df["new_site"])


    def get_site_status(siteid):

        siteid = str(siteid).strip().upper()

        if siteid in new_site_set:
            return "New Site"

        elif siteid in old_site_set:
            return "OLD Site"

        return ""
    
    df_merged = pd.merge(
        df_map,
        df_alarm,
        on="SiteId",
        how="left",
        suffixes=("", "_alarm")
    )
    df_merged["Site Status"] = (
        df_merged["SiteId"]
        .apply(get_site_status)
    )
    
    cols_to_update = [
        "Alarm Status (Yes/No)",
        "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms",
        "Service Affecting Alarms 1",
        "Alarm Type",
        "NOC Circle"
    ]

    for col in cols_to_update:

        alarm_col = f"{col}_alarm"

        if alarm_col in df_merged.columns:

            df_merged[col] = df_merged[col].mask(
                df_merged[col].isna() |
                (
                    df_merged[col]
                    .astype(str)
                    .str.strip() == ""
                ),
                df_merged[alarm_col]
            )

            df_merged.drop(
                columns=[alarm_col],
                inplace=True
            )
        
    df_merged["Alarm Status (Yes/No)"] = (
        df_merged["Alarm Status (Yes/No)"]
        .fillna("No")
    )
    
    # ---------------- Fill Blank SA/NSA and alarms----------------

    df_merged[
        "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms"
    ] = (
        df_merged[
            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms"
        ]
        .fillna("No Alarms")
    )

    df_merged[
            "Service Affecting Alarms 1"
      ] = df_merged[
            "Service Affecting Alarms 1"
    ].replace(
            ["", "nan", None],
            "No Alarms" 
    )

    df_merged["Alarm Type"] = (
        df_merged["Alarm Type"]
        .fillna("")
    )
    df_merged["NOC Circle"] = (
        df_merged["NOC Circle"]
        .fillna("")
    )
    df_merged["Site Status"] = df_merged["SiteId"].apply(get_site_status)

    df_4g = df_4g[
        df_4g["SiteId"].isin(df_merged["SiteId"])
    ]

    df_4g["Site Status"] = (
        df_4g["SiteId"]
        .apply(get_site_status)
    )
    
    compare_4g_rows = []
    for _, row in old_new_df.iterrows():

        old_site = str(row["old_site"]).strip().upper()
        new_site = str(row["new_site"]).strip().upper()

        old_data = df_4g[
            df_4g["SiteId"].astype(str).str.upper() == old_site
        ]

        new_data = df_4g[
            df_4g["SiteId"].astype(str).str.upper() == new_site
        ]

        old_row = old_data.iloc[0] if not old_data.empty else {}
        new_row = new_data.iloc[0] if not new_data.empty else {}

        old_admin = str(
            old_row.get("Admin Status", "")
        ).strip().upper() if len(old_data) else ""

        new_admin = str(
            new_row.get("Admin Status", "")
        ).strip().upper() if len(new_data) else ""

        if old_admin == "UNLOCKED" and new_admin == "UNLOCKED":
            lock_status = "Both Unlocked"

        elif old_admin == "LOCKED" and new_admin == "LOCKED":
            lock_status = "Both Locked"

        elif old_admin and new_admin:
            lock_status = "Changed"

        else:
            lock_status = ""

        compare_4g_rows.append({

            "NE Connection Status_new":
                new_row.get("NE Connection Status", "")
                if len(new_data) else "",

            "Cell Name_new":
                new_row.get("Cell Name", "")
                if len(new_data) else "",

            "Admin Status_new":
                new_row.get("Admin Status", "")
                if len(new_data) else "",

            "Op Status_new":
                new_row.get("Op Status", "")
                if len(new_data) else "",

            "SiteId_new":
                new_site,

            "NE Connection Status_old":
                old_row.get("NE Connection Status", "")
                if len(old_data) else "",

            "Cell Name_old":
                old_row.get("Cell Name", "")
                if len(old_data) else "",

            "Admin Status_old":
                old_row.get("Admin Status", "")
                if len(old_data) else "",

            "Op Status_old":
                old_row.get("Op Status", "")
                if len(old_data) else "",

            "SiteId_old":
                old_site,

            "Lock Status":
                lock_status
        })

    compare_4g_df = pd.DataFrame(compare_4g_rows)
    
    print("Final Output:\n", df_alarm.head())

    # -------- Save Output --------
    output_path = os.path.join(MEDIA_ROOT, "HUAWEI_OUTPUT")
    os.makedirs(output_path, exist_ok=True)

    output_root = os.path.join(MEDIA_ROOT, "HUAWEI_OUTPUT", "OUTPUT")
    os.makedirs(output_root, exist_ok=True)

    if "Circle" not in df_merged.columns:
        return Response({"error": "Circle column not found in mapping file"}, status=400)

    for circle in df_merged["Circle"].dropna().unique():

        circle_folder = os.path.join(output_root, str(circle).strip())
        os.makedirs(circle_folder, exist_ok=True)

        df_circle = df_merged[df_merged["Circle"] == circle]
        df_4g_circle = df_4g[df_4g["SiteId"].isin(df_circle["SiteId"])]
        compare_4g_circle = compare_4g_df[
            compare_4g_df["SiteId_new"].isin(df_circle["SiteId"]) |
            compare_4g_df["SiteId_old"].isin(df_circle["SiteId"])
        ]

        file_path = os.path.join(circle_folder, f"{circle}_output.xlsx")

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_circle.to_excel(writer, sheet_name="Alarms", index=False)
            df_4g_circle.to_excel(writer, sheet_name="4G", index=False)
            compare_4g_circle.to_excel(writer, sheet_name="4G Old vs New", index=False)

        format_of_excel(file_path)

    # -------- ZIP CREATION --------
    shutil.make_archive(
        os.path.join(MEDIA_ROOT, "HUAWEI_OUTPUT", "OUTPUT"),
        'zip',
        output_root
    )

    download_link = request.build_absolute_uri(
        urljoin(MEDIA_URL, "HUAWEI_OUTPUT/OUTPUT.zip")
    )


    return Response(
        {
            "status": True,
            "message": "File processed successfully",
            "download_url": download_link
        },
        status=status.HTTP_200_OK
    )