from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from django.core.files.storage import FileSystemStorage
from datetime import date,timedelta
import openpyxl
import datetime
import pandas as pd
import os
from commom_utilities.utils import *


@api_view(["POST"])
def old_hr_trend(request):
    kpi = [
        "MV_RRC Setup Success Rate [CDBH]",
        "MV_ERAB Setup Success Rate [CDBH]",
        "MV_PS Drop Call Rate % [CDBH]",
        "MV_UL User Throughput_Mbps [CDBH]",
        "MV_DL User Throughput_Mbps [CDBH]",
        "MV_E-UTRAN Average CQI [CDBH]",
        "MV_PS handover success rate [LTE Intra System] [CDBH]",
        "MV_PS handover success rate [LTE Inter System] [CDBH]",
        "MV_CSFB Redirection Success Rate [CDBH]",
        "Paging record discarded At eNodeB [CDBH]",
        "MV_Average number of used DL PRBs [CDBH]",
        "MV_VoLTE ERAB Setup Success Rate [CBBH]",
        "MV_VoLTE DCR [CBBH]",
        "MV_VoLTE Packet Loss DL [CBBH]",
        "MV_VoLTE Packet Loss UL [CBBH]",
        "VoLTE SRVCC SR [CBBH]",
        "MV_PUSCH SINR [CBBH]",
        "MV_VoLTE IntraF HOSR Exec [CBBH]",
        "MV_VoLTE InterF HOSR Exec [CBBH]",
        "MV_VoLTE SRVCC Per Call Rate [CBBH]",
        "UL RSSI",
        "MV_4G Data Volume_GB",
        "MV_VoLTE Traffic",
        "PS handover success rate_NOM [LTE Intra System] [CDBH]",
        "PS handover success rate_DENOM [LTE Intra System] [CDBH]",
        "PS handover success rate_NOM [LTE Inter System] [CDBH]",
        "PS handover success rate_DENOM [LTE Inter System] [CDBH]",
        "MV_VoLTE IntraF HOSR Exec_Nom [CBBH]",
        "MV_VoLTE IntraF HOSR Exec_Denom [CBBH]",
        "VoLTE InterF HOSR Exec_NOM [CBBH]",
        "VoLTE InterF HOSR Exec_DENOM [CBBH]",
        "MV_VoLTE SRVCC Per Call Rate_Nom [CBBH]",
        "MV_VoLTE SRVCC Per Call Rate_Denom [CBBH]",
        "MV_Average number of used UL PRBs [CDBH]",
        "TA Sampls > 1.5 Km % [CDBH]",
        "VoLTE Packet Loss DL_NOM [CBBH]",
        "VoLTE Packet Loss UL_NOM [CBBH]",
    ]
    required_cols = [
        "MV_RRC Setup Success Rate [CDBH]",
        "MV_ERAB Setup Success Rate [CDBH]",
        "MV_PS Drop Call Rate % [CDBH]",
        "MV_UL User Throughput_Kbps [CDBH]",
        "MV_DL User Throughput_Kbps [CDBH]",
        "MV_E-UTRAN Average CQI [CDBH]",
        "MV_PS handover success rate [LTE Intra System] [CDBH]",
        "MV_PS handover success rate [LTE Inter System] [CDBH]",
        "MV_CSFB Redirection Success Rate [CDBH]",
        "Paging record discarded At eNodeB [CDBH]",
        "MV_Average number of used DL PRBs [CDBH]",
        "MV_VoLTE ERAB Setup Success Rate [CBBH]",
        "MV_VoLTE DCR [CBBH]",
        "MV_VoLTE Packet Loss DL [CBBH]",
        "MV_VoLTE Packet Loss UL [CBBH]",
        "VoLTE SRVCC SR [CBBH]",
        "MV_PUSCH SINR [CBBH]",
        "MV_VoLTE IntraF HOSR Exec [CBBH]",
        "MV_VoLTE InterF HOSR Exec [CBBH]",
        "MV_VoLTE SRVCC Per Call Rate [CBBH]",
        "UL RSSI",
        "MV_4G Data Volume_GB",
        "MV_VoLTE Traffic",
        "PS handover success rate_NOM [LTE Intra System] [CDBH]",
        "PS handover success rate_DENOM [LTE Intra System] [CDBH]",
        "PS handover success rate_NOM [LTE Inter System] [CDBH]",
        "PS handover success rate_DENOM [LTE Inter System] [CDBH]",
        "MV_VoLTE IntraF HOSR Exec_Nom [CBBH]",
        "MV_VoLTE IntraF HOSR Exec_Denom [CBBH]",
        "VoLTE InterF HOSR Exec_NOM [CBBH]",
        "VoLTE InterF HOSR Exec_DENOM [CBBH]",
        "MV_VoLTE SRVCC Per Call Rate_Nom [CBBH]",
        "MV_VoLTE SRVCC Per Call Rate_Denom [CBBH]",
        "MV_Average number of used UL PRBs [CDBH]",
        "TA Sampls > 1.5 Km % [CDBH]",
        "VoLTE Packet Loss DL_NOM [CBBH]",
        "VoLTE Packet Loss UL_NOM [CBBH]",
    ]
    gsm = [
        "SDCCH Blocking Rate [BBH]",
        "SDCCH Drop Call Rate [BBH]",
        "TCH Blocking Rate [BBH]",
        "TCH Drop Call Rate [BBH]",
        "Handover Success Rate [BBH]",
        "RX Quality [BBH]",
        "Total Voice Traffic",
        "Handover Success Rate_Nom [BBH]",
        "Handover Success Rate_Denom [BBH]",
        "Drop Call Rate_Nom [BBH]",
        "Drop Call Rate_Denom [BBH]",
        "Number of TRX [BBH]",
        "ICM%[BBH]",
        "Cell Downtime [sec] [BBH]",
        "TCH Utilization [BBH]",
        'Drop Call Rate',
    ]

    raw_kpi_4G = request.FILES["4G_raw_kpi"] if "4G_raw_kpi" in request.FILES else None
    df_raw_kpi_4G = pd.read_excel(raw_kpi_4G)

    sts,response = required_col_check(raw_kpi_4G, required_cols)

    if sts:
        return Response(response)

    ##print("procced________________________")
    response, s_l = site_list_handler_4G(request)
    ##print("_______________success_______________________", s_l)
    if s_l:
        # site_list = list(map(lambda a: str(a), s_l))  ##################### converting in str
        site_list=s_l
      
        site_list_str = [str(x) for x in site_list]
        
        ##print("Converted site_list to string:", site_list_str)
        ##print("__________site_list______", type(site_list_str[2]))

        ##print(site_list_str, "_______________", s_l)
        ##print("Checkpoint 1")
        if response:
            return Response(response)
        ##print("Checkpoint 2")

    raw_kpi_GSM = request.FILES["2G_raw_kpi"] if "2G_raw_kpi" in request.FILES else None
    raw_kpi_2G = pd.read_excel(raw_kpi_GSM)
    sts, response = required_col_check(raw_kpi_GSM, gsm)
    if sts:
        return Response(response)
    ##print("checkpoint3")

    response, s_l = site_list_handler_2G(request)
 
    if s_l:
        df_2G_site = s_l
        df_2G_site_str=[str(y) for y in df_2G_site]
        ##print("__________2g_site_______________##print", type(df_2G_site_str[0]))

        ##print(df_2G_site_str, "______2G_ROW_KPI_________", s_l)
        if response:
            return Response(response)
    ##print("checkpoint4")
    ################################################################################################
    door_path = os.path.join(MEDIA_ROOT, "trends", "hr")

    for x in required_cols:
        df_raw_kpi_4G[x] = df_raw_kpi_4G[x].replace(
            to_replace=".*", value=0, regex=True
        )
    # df_raw_kpi_4G1 = df_raw_kpi_4G.astype(str)
    ##print("_________________________df______________________", type(df_raw_kpi_4G))
    df_raw_kpi_4G["Short name"] = df_raw_kpi_4G["Short name"].fillna(method="ffill")
    df_raw_kpi_4G.columns.values[1] = "Date"

    df_raw_kpi_4G["MV_DL User Throughput_Kbps [CDBH]"] = (df_raw_kpi_4G["MV_DL User Throughput_Kbps [CDBH]"] / 1024)
    df_raw_kpi_4G.rename(columns={"MV_DL User Throughput_Kbps [CDBH]": "MV_DL User Throughput_Mbps [CDBH]"},inplace=True,)
    df_raw_kpi_4G["MV_UL User Throughput_Kbps [CDBH]"] = (df_raw_kpi_4G["MV_UL User Throughput_Kbps [CDBH]"] / 1024)
    df_raw_kpi_4G.rename(columns={"MV_UL User Throughput_Kbps [CDBH]": "MV_UL User Throughput_Mbps [CDBH]"},inplace=True,)
    # ##print(df)

    split = []
    split1 = []
    techlist = []
    for cell in df_raw_kpi_4G["Short name"]:
        if "_" in cell:
            site_id = cell.split("_")[-2][:-1]
            split.append(site_id)
        else:
            site_id = cell[:-1]
            split.append(site_id)

        if "_" in cell:
            cell_id = cell.split("_")[-2]
            split1.append(cell_id)
        else:
            cell_id = cell[:-1]
            split1.append(cell_id)

        if ("_F1_" in cell or "_F3_" in cell or "_F5_" in cell or "_T1_" in cell or "_T2_" in cell):
            if "_F1_" in cell:
                tech = "L2100"
            if "_F3_" in cell:
                tech = "L1800"
            if "_F5_" in cell:
                tech = "L900"
            if "_T1_" in cell or "_T2_" in cell:
                tech = "L2300"
            techlist.append(tech)
        else:
            cell = site_id
            techlist.append(tech)

    df_raw_kpi_4G.insert(0, "Site ID", split)

    df_raw_kpi_4G.insert(2, "cell_id", split1)
    df_raw_kpi_4G.insert(5, "Tech", techlist)

    df_raw_kpi_4G.rename(columns={"Short name": "Shortname"}, inplace=True)

    ecgi = []
    ecgi1 = []
    for cell1 in df_raw_kpi_4G["4G_ECGI"]:
        if "-" in str(cell1):
            lnbts_id = str(cell1).split("-")[-2]
            ecgi.append(lnbts_id)
        else:
            lnbts_id = str(cell1)[:-1]
            ecgi.append(lnbts_id)

        if "-" in cell1:
            lncell_id = str(cell1).split("-")[-1]
            ecgi1.append(lncell_id)
        else:
            lncell_id = str(cell1)[-1]
            ecgi1.append(lncell_id)
    df_raw_kpi_4G.insert(3, "lnbts_id", ecgi)

 
    # ##print( "_____________________lnbts_id______________________________________",lnbts_id)

    ##print("__________type_______cell1____", type(lnbts_id[0]))
    df_raw_kpi_4G.insert(4, "lncell_id", ecgi1)

    df_raw_kpi_4G.rename(columns={"Short name": "Shortname"}, inplace=True)

    df_raw_kpi_4G["Site name"] = (df_raw_kpi_4G["Site ID"].astype(str) + "(" + df_raw_kpi_4G["lnbts_id"] + ")")


    df_raw_kpi_4G.fillna(value=0, inplace=True)


    message_4G = site_comparision(ecgi, site_list_str)  # site_comparision_call
    ##print("missing sites:-----", message_4G)
    PsOsPath1 = os.path.join(door_path, "process output", "desired input.xlsx")
    df_raw_kpi_4G.to_excel(PsOsPath1, index=False)

    excel_file_1 = PsOsPath1
    # PsOsPath2=os.path.join(door_path,'project file','site.xlsx')
    # excel_file_2 = PsOsPath2

    df1 = pd.read_excel(excel_file_1)
    

    df_filter = df1[(df1.lnbts_id.isin(site_list))]
    

    ##print("______________df_filter_______________",df_filter)
    

    df_filter1 = df_filter.round(2)
    PsOsFilter = os.path.join(door_path, "process output", "filtered_df_1.xlsx")

    df_filter1.to_excel(PsOsFilter, index=False)

    df1 = pd.read_excel(PsOsFilter)

    df_pivot = df1.pivot_table(
        values=kpi,
        columns="Date",
        index=["Shortname","Site ID","cell_id","lnbts_id", "lncell_id","Site name","Tech",
            "4G_ECGI",
        ],
    )
    PsOsPivot_Path = os.path.join(door_path, "process output/pivot.xlsx")
    df_pivot.to_excel(PsOsPivot_Path)

    str_date = request.POST.get("offered_date")
    date1 = datetime.datetime.strptime(str_date, "%Y-%m-%d")
    # date1 = date(2023, 3, 9)
    d1 = date1 - timedelta(1)
    d2 = date1 - timedelta(2)
    d3 = date1 - timedelta(3)
    d4 = date1 - timedelta(4)
    d5 = date1 - timedelta(5)
    cl = [d1, d2, d3, d4, d5]
    index = df_pivot.index

    STR = os.path.join(door_path, "template", "haryanatemplate.xlsx")
    wb = openpyxl.load_workbook(STR)
    ws = wb["ALL TECH KPI"]

    ################################################
    def overwrite(kpi_name, coln1, coln2, coln3, coln4, coln5):
        dr = df_pivot[kpi_name]
        li = dr.columns.to_list()
        col1 = dr[li[0]].to_list()
        col2 = dr[li[1]].to_list()
        col3 = dr[li[2]].to_list()
        col4 = dr[li[3]].to_list()
        col5 = dr[li[4]].to_list()

        ws[coln1 + "4"].value = cl[4]
        ws[coln2 + "4"].value = cl[3]
        ws[coln3 + "4"].value = cl[2]
        ws[coln4 + "4"].value = cl[1]
        ws[coln5 + "4"].value = cl[0]

        for i, value in enumerate(index):
            j = i + 5
            ws["B" + str(j)].value = "HR"

            # ws['L'+str(j)].value=date1
            # ws['P' + str(j)].value = 'Site Type'

            ws["M" + str(j)].value = "DONE"
            ws["A" + str(j)].value = index[i][7]
            ws["C" + str(j)].value = index[i][1]
            ws["D" + str(j)].value = index[i][2]
            ws["E" + str(j)].value = index[i][3]
            ws["F" + str(j)].value = index[i][4]
            ws["G" + str(j)].value = index[i][1]
            ws["H" + str(j)].value = index[i][0]
            ws["I" + str(j)].value = index[i][5]
            ws["J" + str(j)].value = index[i][6]

            ws[coln1 + str(j)].value = col1[i]
            ws[coln2 + str(j)].value = col2[i]
            ws[coln3 + str(j)].value = col3[i]
            ws[coln4 + str(j)].value = col4[i]
            ws[coln5 + str(j)].value = col5[i]

    for kpi_name in kpi:
        if kpi_name == "MV_RRC Setup Success Rate [CDBH]":
            overwrite(kpi_name, "Q", "R", "S", "T", "U")
        if kpi_name == "MV_ERAB Setup Success Rate [CDBH]":
            overwrite(kpi_name, "X", "Y", "Z", "AA", "AB")

        if kpi_name == "MV_PS Drop Call Rate % [CDBH]":
            overwrite(kpi_name, "AE", "AF", "AG", "AH", "AI")
        if kpi_name == "MV_DL User Throughput_Mbps [CDBH]":
            overwrite(kpi_name, "AL", "AM", "AN", "AO", "AP")
        if kpi_name == "MV_UL User Throughput_Mbps [CDBH]":
            overwrite(kpi_name, "AT", "AU", "AV", "AW", "AX")
        if kpi_name == "MV_E-UTRAN Average CQI [CDBH]":
            overwrite(kpi_name, "BA", "BB", "BC", "BD", "BE")
        if kpi_name == "MV_PS handover success rate [LTE Intra System] [CDBH]":
            overwrite(kpi_name, "BH", "BI", "BJ", "BK", "BL")
        if kpi_name == "MV_PS handover success rate [LTE Inter System] [CDBH]":
            overwrite(kpi_name, "BO", "BP", "BQ", "BR", "BS")
        if kpi_name == "MV_CSFB Redirection Success Rate [CDBH]":
            overwrite(kpi_name, "BV", "BW", "BX", "BY", "BZ")
        if kpi_name == "Paging record discarded At eNodeB [CDBH]":
            overwrite(kpi_name, "CC", "CD", "CE", "CF", "CG")
        if kpi_name == "MV_Average number of used DL PRBs [CDBH]":
            overwrite(kpi_name, "CJ", "CK", "CL", "CM", "CN")
        if kpi_name == "MV_VoLTE ERAB Setup Success Rate [CBBH]":
            overwrite(kpi_name, "CQ", "CR", "CS", "CT", "CU")
        if kpi_name == "MV_VoLTE DCR [CBBH]":
            overwrite(kpi_name, "CX", "CY", "CZ", "DA", "DB")
        if kpi_name == "MV_VoLTE Packet Loss DL [CBBH]":
            overwrite(kpi_name, "DE", "DF", "DG", "DH", "DI")
        if kpi_name == "MV_VoLTE Packet Loss UL [CBBH]":
            overwrite(kpi_name, "DL", "DM", "DN", "DO", "DP")
        if kpi_name == "VoLTE SRVCC SR [CBBH]":
            overwrite(kpi_name, "DS", "DT", "DU", "DV", "DW")
        if kpi_name == "MV_PUSCH SINR [CBBH]":
            overwrite(kpi_name, "DZ", "EA", "EB", "EC", "ED")
        if kpi_name == "MV_VoLTE IntraF HOSR Exec [CBBH]":
            overwrite(kpi_name, "EG", "EH", "EI", "EJ", "EK")
        if kpi_name == "MV_VoLTE InterF HOSR Exec [CBBH]":
            overwrite(kpi_name, "EN", "EO", "EP", "EQ", "ER")
        if kpi_name == "MV_VoLTE SRVCC Per Call Rate [CBBH]":
            overwrite(kpi_name, "EU", "EV", "EW", "EX", "EY")
        if kpi_name == "UL RSSI":
            overwrite(kpi_name, "FB", "FC", "FD", "FE", "FF")
        if kpi_name == "MV_4G Data Volume_GB":
            overwrite(kpi_name, "FI", "FJ", "FK", "FL", "FM")
        if kpi_name == "MV_VoLTE Traffic":
            overwrite(kpi_name, "FP", "FQ", "FR", "FS", "FT")
        if kpi_name == "PS handover success rate_NOM [LTE Intra System] [CDBH]":
            overwrite(kpi_name, "FV", "FW", "FX", "FY", "FZ")
        if kpi_name == "PS handover success rate_DENOM [LTE Intra System] [CDBH]":
            overwrite(kpi_name, "GA", "GB", "GC", "GD", "GE")
        if kpi_name == "PS handover success rate_NOM [LTE Inter System] [CDBH]":
            overwrite(kpi_name, "GL", "GM", "GN", "GO", "GP")
        if kpi_name == "PS handover success rate_DENOM [LTE Inter System] [CDBH]":
            overwrite(kpi_name, "GQ", "GR", "GS", "GT", "GU")
        if kpi_name == "MV_VoLTE IntraF HOSR Exec_Nom [CBBH]":
            overwrite(kpi_name, "HB", "HC", "HD", "HE", "HF")
        if kpi_name == "MV_VoLTE IntraF HOSR Exec_Denom [CBBH]":
            overwrite(kpi_name, "HG", "HH", "HI", "HJ", "HK")
        if kpi_name == "VoLTE InterF HOSR Exec_NOM [CBBH]":
            overwrite(kpi_name, "HR", "HS", "HT", "HU", "HV")
        if kpi_name == "VoLTE InterF HOSR Exec_DENOM [CBBH]":
            overwrite(kpi_name, "HW", "HX", "HY", "HZ", "IA")
        if kpi_name == "MV_VoLTE SRVCC Per Call Rate_Nom [CBBH]":
            overwrite(kpi_name, "IH", "II", "IJ", "IK", "IL")
        if kpi_name == "MV_VoLTE SRVCC Per Call Rate_Denom [CBBH]":
            overwrite(kpi_name, "IM", "IN", "IO", "IP", "IQ")
        if kpi_name == "MV_Average number of used UL PRBs [CDBH]":
            overwrite(kpi_name, "IX", "IY", "IZ", "JA", "JB")
        if kpi_name == "TA Sampls > 1.5 Km % [CDBH]":
            overwrite(kpi_name, "JN", "JO", "JP", "JQ", "JR")
        if kpi_name == "VoLTE Packet Loss DL_NOM [CBBH]":
            overwrite(kpi_name, "JT", "JU", "JV", "JW", "JX")
        if kpi_name == "VoLTE Packet Loss UL_NOM [CBBH]":
            overwrite(kpi_name, "KA", "KB", "KC", "KD", "KE")

    ################################### GSM FILTE PIVOT#########################

    door_path = os.path.join(MEDIA_ROOT, "trends", "hr")
    ##print("shortname--------------------------")
    raw_kpi_2G["Short name"] = raw_kpi_2G["Short name"].fillna(method="ffill")
    raw_kpi_2G.columns.values[1] = "Date"
    a = []
    b = []
    for cell1 in raw_kpi_2G["Short name"]:
        if "-" in cell1:
            siteid = cell1.split("-")[-1][:-1]
            a.append(siteid)
        else:
            siteid = cell1[:-1]
            a.append(siteid)

        if "-" in cell1:
            cellid = cell1.split("-")[-1]
            b.append(cellid)
        else:
            cellid = cell1[:-1]

    raw_kpi_2G.insert(0, "Site ID", a)
    raw_kpi_2G.insert(3, "Cell ID", b)

    c = []
    for cell2 in raw_kpi_2G["2G CGI"]:
        if "-" in str(cell2):
            cgi = str(cell2).split("-")[-1]
            c.append(cgi)
        else:
            cgi=str(cell2)[:-1] 
            c.append(cgi)   
    raw_kpi_2G.insert(2, "cgisplit", c)
    ##print("__________type______",type(cgi[2]))
    raw_kpi_2G.fillna(value=0, inplace=True)

    message_2G = site_comparision(c, df_2G_site_str)  # site_comparision_call
    ##print("missing 2G sites:-----", message_2G)


    g2_path2 = os.path.join(door_path, "process output", "2Gfill.xlsx")
    raw_kpi_2G.to_excel(g2_path2, index=False)
    excel_1 = g2_path2

    # g2_path3=os.path.join(door_path,'project file','2gsite.xlsx')
    # excel_2 = g2_path3
    excel_2G_raw = pd.read_excel(excel_1)


    # excel_2G_raw.rename(columns={"cgisplit": "cgi_split"}, inplace=True)
  
    filter_2G = excel_2G_raw[(excel_2G_raw.cgisplit.isin(df_2G_site))]
    ##print(filter_2G)

    g2_filter = os.path.join(door_path, "process output", "2Gfilter.xlsx")
    filter_2G.to_excel(g2_filter, index=False)

    df1 = pd.read_excel(g2_filter)
    df_pivot1 = df1.pivot_table(values=gsm, columns="Date", index=["Site ID", "2G CGI", "Short name"])
    g2_pivot = os.path.join(door_path, "process output", "2gpivot.xlsx")
    df_pivot1.to_excel(g2_pivot)
    index = df_pivot1.index

    def overwrite(gsm_name, coln1, coln2, coln3, coln4, coln5, ws1):
        dr = df_pivot1[gsm_name]
        li = dr.columns
        col1 = dr[li[0]].to_list()
        col2 = dr[li[1]].to_list()
        col3 = dr[li[2]].to_list()
        col4 = dr[li[3]].to_list()
        col5 = dr[li[4]].to_list()

        ws1[coln1 + "4"].value = cl[4]
        ws1[coln2 + "4"].value = cl[3]
        ws1[coln3 + "4"].value = cl[2]
        ws1[coln4 + "4"].value = cl[1]
        ws1[coln5 + "4"].value = cl[0]

        for i, value in enumerate(index):
            j = i + 5
            ws1["F" + str(j)].value = "GSM"
            ws1["L" + str(j)].value = "Site"
            ws1["B" + str(j)].value = "HR"

            ws1["D" + str(j)].value = index[i][2]
            ws1["C" + str(j)].value = index[i][0]
            ws1["A" + str(j)].value = index[i][1]

            ws1[coln1 + str(j)].value = col1[i]
            ws1[coln2 + str(j)].value = col2[i]
            ws1[coln3 + str(j)].value = col3[i]
            ws1[coln4 + str(j)].value = col4[i]
            ws1[coln5 + str(j)].value = col5[i]

    g2_ws = wb["2G KPI"]

    for gsm_name in gsm:
        if gsm_name == "SDCCH Blocking Rate [BBH]":
            overwrite(gsm_name, "M", "N", "O", "P", "Q", g2_ws)

        if gsm_name == "SDCCH Drop Call Rate [BBH]":
            overwrite(gsm_name, "T", "U", "V", "W", "X", g2_ws)
        if gsm_name == "SDCCH Drop Call Rate [BBH]":
            overwrite(gsm_name, "T", "U", "V", "W", "X", g2_ws)
        if gsm_name == "TCH Blocking Rate [BBH]":
            overwrite(gsm_name, "AA", "AB", "AC", "AD", "AE", g2_ws)
        if gsm_name == "TCH Drop Call Rate [BBH]":
            overwrite(gsm_name, "AH", "AI", "AJ", "AK", "AL", g2_ws)
        if gsm_name == "Handover Success Rate [BBH]":
            overwrite(gsm_name, "AP", "AQ", "AR", "AS", "AT", g2_ws)

        if (gsm_name == 'Drop Call Rate'):
            overwrite(gsm_name, 'BD', 'BE', 'BF', 'BG', 'BH', g2_ws)    

        if gsm_name == "RX Quality [BBH]":
            overwrite(gsm_name, "AW", "AX", "AY", "AZ", "BA", g2_ws)
        if gsm_name == "Total Voice Traffic":
            overwrite(gsm_name, "BK", "BL", "BM", "BN", "BO", g2_ws)
        if gsm_name == "Handover Success Rate_Nom [BBH]":
            overwrite(gsm_name, "BQ", "BR", "BS", "BT", "BU", g2_ws)
        if gsm_name == "Handover Success Rate_Denom [BBH]":
            overwrite(gsm_name, "BV", "BW", "BX", "BY", "BZ", g2_ws)
        if gsm_name == "Drop Call Rate_Nom [BBH]":
            overwrite(gsm_name, "CG", "CH", "CI", "CJ", "CK", g2_ws)
        if gsm_name == "Drop Call Rate_Denom [BBH]":
            overwrite(gsm_name, "CL", "CM", "CN", "CO", "CP", g2_ws)
        if gsm_name == "Drop Call Rate_Nom [BBH]":
            overwrite(gsm_name, "CQ", "CR", "CS", "CT", "CU", g2_ws)
        if gsm_name == "Number of TRX [BBH]":
            overwrite(gsm_name, "DM", "DN", "DO", "DP", "DQ", g2_ws)
        if gsm_name == "ICM%[BBH]":
            overwrite(gsm_name, "DS", "DT", "DU", "DV", "DW", g2_ws)
        if gsm_name == "Cell Downtime [sec] [BBH]":
            overwrite(gsm_name, "DY", "DZ", "EA", "EB", "EC", g2_ws)
        if gsm_name == "TCH Utilization [BBH]":
            overwrite(gsm_name, "EE", "EF", "EG", "EH", "EI", g2_ws)
    save_output = os.path.join(door_path, "output", "Haryana_trend_output.xlsx")
    wb.save(save_output)
    print("TESTING")

    download_path = os.path.join(
        MEDIA_URL, "trends", "hr", "output", "Haryana_trend_output.xlsx"
    )
    return Response({"status": True,"message": "uploaded successfully","missing_sites_4G": message_4G,"missing_sites_2G": message_2G,"Download_url": download_path}
    )
