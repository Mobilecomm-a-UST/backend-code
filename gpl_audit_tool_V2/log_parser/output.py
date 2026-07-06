import logging
import sys

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import SHEETS, get_output_file

log = logging.getLogger(__name__)

# ── colour palette ────────────────────────────────────────────────────────────
_HDR_BG   = "1F3864"   # dark navy header
_HDR_FG   = "FFFFFF"   # white header text

_OK_BG    = "C6EFCE"   # light green
_OK_FG    = "375623"   # dark green

_NOTOK_BG = "FFC7CE"   # light red
_NOTOK_FG = "9C0006"   # dark red

_MISS_BG  = "FFEB9C"   # light amber
_MISS_FG  = "9C6500"   # dark amber

_ALT_ROW  = "EBF3FA"   # alternating row tint (every other data row)

_THIN_SIDE = Side(border_style="thin", color="C8C8C8")
_BORDER    = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)


def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


# Style objects are stateless — build each combination once and share the
# instance across every cell that uses it instead of constructing a fresh
# Font/PatternFill per cell (openpyxl dedupes shared style instances anyway;
# reusing them up front avoids the per-cell allocation and dedup-lookup cost).
_HDR_FILL   = _fill(_HDR_BG)
_HDR_FONT   = Font(color=_HDR_FG, bold=True, size=10)
_HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

_ROW_FILL_EVEN = _fill(_ALT_ROW)
_ROW_FILL_ODD  = _fill("FFFFFF")
_PLAIN_FONT    = Font(size=10)
_DATA_ALIGN    = Alignment(horizontal="center", vertical="center", wrap_text=False)

_OK_FILL    = _fill(_OK_BG)
_OK_FONT    = Font(color=_OK_FG, bold=True, size=10)
_NOTOK_FILL = _fill(_NOTOK_BG)
_NOTOK_FONT = Font(color=_NOTOK_FG, bold=True, size=10)
_MISS_FILL  = _fill(_MISS_BG)
_MISS_FONT  = Font(color=_MISS_FG, bold=True, size=10)


# Blank rows left between stacked tables on the same sheet (e.g. the Summary
# sheet's cell-mapping table followed by the parameter status roll-up).
_TABLE_GAP_ROWS = 2

_TITLE_FILL  = _fill(_HDR_BG)
_TITLE_FONT  = Font(color=_HDR_FG, bold=True, size=12)
_TITLE_ALIGN = Alignment(horizontal="center", vertical="center")


# Columns whose header AND every data cell get a fixed colour (used for the
# OK / NOT OK / Missing count columns on the Parameter Status Summary table)
# rather than the per-value status colouring used elsewhere.
_FIXED_COLUMN_STYLES: dict[str, tuple[PatternFill, Font]] = {
    "OK":      (_OK_FILL, _OK_FONT),
    "NOT OK":  (_NOTOK_FILL, _NOTOK_FONT),
    "Missing": (_MISS_FILL, _MISS_FONT),
}


def _style_block(ws, df: pd.DataFrame, start_row: int, freeze: bool = False) -> list[int]:
    """Style one header+data block starting at 1-based row `start_row`.
    Returns per-column max content length (header vs data) so callers can
    reconcile widths across multiple blocks sharing a sheet."""
    status_col_indices = {
        i + 1 for i, col in enumerate(df.columns) if "status" in col.lower()
    }
    fixed_col_styles = {
        i + 1: _FIXED_COLUMN_STYLES[col] for i, col in enumerate(df.columns) if col in _FIXED_COLUMN_STYLES
    }
    col_widths = [max(len(str(c)), 10) for c in df.columns]

    for cell in ws[start_row]:
        col_i = cell.column
        if col_i in fixed_col_styles:
            cell.fill, cell.font = fixed_col_styles[col_i]
        else:
            cell.fill, cell.font = _HDR_FILL, _HDR_FONT
        cell.alignment = _HDR_ALIGN
        cell.border    = _BORDER
    ws.row_dimensions[start_row].height = 32
    if freeze:
        ws.freeze_panes = f"A{start_row + 1}"

    data_start, data_end = start_row + 1, start_row + len(df)
    for data_row_idx, row in enumerate(ws.iter_rows(min_row=data_start, max_row=data_end), start=1):
        row_fill = _ROW_FILL_EVEN if data_row_idx % 2 == 0 else _ROW_FILL_ODD
        for cell in row:
            col_i = cell.column
            raw   = cell.value
            sval  = "" if raw is None else str(raw)

            if raw is not None:
                length = len(sval)
                if length > col_widths[col_i - 1]:
                    col_widths[col_i - 1] = length

            if col_i in fixed_col_styles:
                cell.fill, cell.font = fixed_col_styles[col_i]
            elif col_i in status_col_indices:
                val = sval.strip().upper()
                if val == "OK":
                    cell.fill, cell.font = _OK_FILL, _OK_FONT
                elif "NOT OK" in val:
                    cell.fill, cell.font = _NOTOK_FILL, _NOTOK_FONT
                elif "MISS" in val:
                    cell.fill, cell.font = _MISS_FILL, _MISS_FONT
                else:
                    cell.fill, cell.font = row_fill, _PLAIN_FONT
            else:
                cell.fill, cell.font = row_fill, _PLAIN_FONT

            cell.border    = _BORDER
            cell.alignment = _DATA_ALIGN

    return col_widths


def _format_sheet(ws, df: pd.DataFrame) -> None:
    """Apply header styling, status colours, zebra rows, borders, and auto-widths
    to a sheet holding a single table starting at row 1."""
    widths = _style_block(ws, df, start_row=1, freeze=True)
    for col_i in range(1, len(df.columns) + 1):
        letter = get_column_letter(col_i)
        width = min(widths[col_i - 1] + 3, 55)
        ws.column_dimensions[letter].width = max(width, 10)


def _write_stacked_tables(writer, sheet_name: str, tables: list[tuple[str, pd.DataFrame]]) -> None:
    """Write several titled tables top-to-bottom on one sheet, separated by a
    blank-row gap, each with its own styled header and title banner. Column
    widths are reconciled across all blocks that share a column."""
    ws = None
    row = 1
    max_cols = max(len(df.columns) for _, df in tables)
    combined_widths = [10] * max_cols

    for i, (title, df) in enumerate(tables):
        if i > 0 and title:
            # title banner directly above this block's column headers
            ws = writer.sheets[sheet_name] if ws is not None else None
            if ws is not None:
                # span the full sheet width (not just this block's own
                # columns) so the banner reads as a section divider that
                # lines up with the table above it, not a narrower strip.
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_cols)
                cell = ws.cell(row=row, column=1, value=title)
                cell.fill, cell.font, cell.alignment = _TITLE_FILL, _TITLE_FONT, _TITLE_ALIGN
                ws.row_dimensions[row].height = 24
            row += 1

        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=row - 1)
        ws = writer.sheets[sheet_name]
        widths = _style_block(ws, df, start_row=row, freeze=(i == 0))
        for c, w in enumerate(widths):
            combined_widths[c] = max(combined_widths[c], w)
        row += len(df) + 1 + _TABLE_GAP_ROWS  # header row + data rows + gap

    for col_i in range(1, max_cols + 1):
        letter = get_column_letter(col_i)
        ws.column_dimensions[letter].width = min(combined_widths[col_i - 1] + 3, 55)


def write_excel(data: dict[str, pd.DataFrame]) -> None:
    output_file = get_output_file()
    output_file.parent.mkdir(parents=True, exist_ok=True)

    try:
        sheets_written = 0
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for sheet_name, key in SHEETS:
                df = data.get(key, pd.DataFrame())
                if df.empty:
                    continue

                extra_key = f"{key}_counts"
                extra_df = data.get(extra_key, pd.DataFrame())
                if not extra_df.empty:
                    _write_stacked_tables(
                        writer, sheet_name,
                        [("", df), ("Parameter Status Summary", extra_df)],
                    )
                    log.info(
                        "Sheet '%-22s' written — %d rows x %d cols (+ %d-row summary table)",
                        sheet_name, *df.shape, len(extra_df),
                    )
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    _format_sheet(writer.sheets[sheet_name], df)
                    log.info("Sheet '%-22s' written — %d rows x %d cols", sheet_name, *df.shape)
                sheets_written += 1

        if sheets_written:
            log.info("Report saved to %s", output_file)
        else:
            log.warning("No sheets written — output file was not created.")
    except Exception as e:
        log.error("Failed to write Excel report: %s", e)
        sys.exit(1)
