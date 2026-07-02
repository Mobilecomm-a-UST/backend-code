from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side,PatternFill
from openpyxl.utils import get_column_letter
import os

def format_excel(file_path):
    wb = load_workbook(file_path)
    ws =wb.active
    header_fill = PatternFill(start_color="F4B084",end_color="F4B084",fill_type="solid")

    header_font = Font(name="Cambria", bold=True, size=9)
    center_align = Alignment(
        vertical="center",
        horizontal="center",
        wrap_text=True
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for ws in wb.worksheets:
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # Header row height
        ws.row_dimensions[1].height =24

        # ---------- Data formatting ----------
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Cambria", size=9)
                cell.alignment = Alignment(vertical="center",horizontal="center", wrap_text=True)
                cell.border = border
                
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max_length + 3, 45)
        wb.save(file_path)



def delete_existing_files(folder_path):
    if os.path.exists(folder_path):
        for file in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except Exception as e:
                print(f"Error deleting {file_path}: {e}")
