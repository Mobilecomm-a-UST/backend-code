from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from django.core.files.storage import FileSystemStorage
from datetime import date,timedelta
import openpyxl
import datetime
import pandas as pd
import os
from commom_utilities.utils import *


@api_view(['POST'])
def old_jk_trend(request):

    kpi=["RRC Setup Success Rate [CDBH]",
    "ERAB Setup Success Rate [CDBH]",
    "PS Drop Call Rate % [CDBH]",
    "E-UTRAN Average CQI [CDBH]",
    "PS handover success rate [LTE Intra System] [CDBH]",
    "PS handover success rate [LTE Inter System] [CDBH]",
    "MV_4G Data Volume_GB",
    "MV_Average number of used DL PRBs",
    "MV_UL User Throughput_Kbps [CDBH]",
    "MV_DL User Throughput_Mbps [CDBH]",
    "MV_VoLTE ERAB Setup Success Rate",
    "MV_VoLTE DCR [CBBH]",
    "MV_VoLTE Packet Loss DL [CBBH]",
    "MV_VoLTE Packet Loss UL [CBBH]",
    "MV_VoLTE IntraF HOSR Exec [CBBH]",
    "VoLTE InterF HOSR Exec",
    "MV_VoLTE Traffic [K erl]",
    "MV_VoLTE Traffic",
    "MV_CSFB Redirection Success Rate",
    "UL RSSI [CDBH]",
    "MV_Avg Connected User"]

    

    raw_kpi_4G=request.FILES['4G_raw_kpi'] if '4G_raw_kpi' in request.FILES else None
    df_jk=pd.read_excel(raw_kpi_4G)
    
    required_cols=["RRC Setup Success Rate [CDBH]",
    "ERAB Setup Success Rate [CDBH]",
    "PS Drop Call Rate % [CDBH]",
    "E-UTRAN Average CQI [CDBH]",
    "PS handover success rate [LTE Intra System] [CDBH]",
    "PS handover success rate [LTE Inter System] [CDBH]",
    "MV_4G Data Volume_GB",
    "MV_Average number of used DL PRBs",
    "MV_UL User Throughput_Kbps [CDBH]",
    "MV_DL User Throughput_Kbps [CDBH]",
    "MV_VoLTE ERAB Setup Success Rate",
    "MV_VoLTE DCR [CBBH]",
    "MV_VoLTE Packet Loss DL [CBBH]",
    "MV_VoLTE Packet Loss UL [CBBH]",
    "MV_VoLTE IntraF HOSR Exec [CBBH]",
    "VoLTE InterF HOSR Exec",
    "MV_VoLTE Traffic [K erl]",
    "MV_VoLTE Traffic",
    "MV_CSFB Redirection Success Rate",
    "UL RSSI [CDBH]",
    "MV_Avg Connected User"]

    sts,response=required_col_check(raw_kpi_4G,required_cols)
    if sts:
        return Response(response)
########################site_list#####################
    site_list_4G=request.FILES['site_list_4G'] if 'site_list_4G' in request.FILES else None
    df_site=pd.read_excel(site_list_4G)
    
    if site_list_4G:
        sts,response=required_col_check(raw_kpi_4G,required_cols)
        if sts:
            return Response(response)

  
    door_path=os.path.join(MEDIA_ROOT,'trends','jk')
   
    df_jk["Short name"]=df_jk["Short name"].fillna(method='ffill')
    df_jk.columns.values[1]='Date'
    df_jk.fillna(value=0,inplace=True)
    df_jk["MV_DL User Throughput_Kbps [CDBH]"]=df_jk["MV_DL User Throughput_Kbps [CDBH]"]/1024
    df_jk.rename(columns={"MV_DL User Throughput_Kbps [CDBH]":"MV_DL User Throughput_Mbps [CDBH]"})
    df_jk["EnodebId"]=[cell.split('-')[-2] if isinstance(cell,str) and '-' in cell else cell for cell in df_jk["4G_ECGI"]]
    df_jk['Site_id']=[cell1.split('_')[-2][:-1] if '_' in cell1 else cell1[:-1] for cell1 in df_jk["Short name"]]
     
    message=site_comparision(df_jk['Site_id'],df_site) # site_comparision_call 


    PsOsPath1=os.path.join(door_path,'process output','desired input.xlsx')
    df_jk.to_excel(PsOsPath1,index=False)
    print(df_jk)

    # PsOsPath2=os.path.join(door_path,'project file','site.xlsx')
    # df2=PsOsPath2
    # df_site=pd.read_excel(df2)
    # print(df_site)

    df_filter=df_jk[(df_jk.Site_id.isin(df_site['2G ID']))]
    print(df_filter)
    
    

    df_pivot=df_filter.pivot_table(values=kpi,columns='Date',index=['Short name','Site_id'])
    print(df_pivot)

    PsOsPath3=os.path.join(door_path,'process output','jk_pivot.xlsx')
    df_pivot.to_excel(PsOsPath3)
    # df3='jk_pivot.xlsx'
    

    STR=os.path.join(door_path,'template','UPDATE_JK.xlsx')
    wb=openpyxl.load_workbook(STR)
    ws=wb.active

    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
    # This process is similar to binary-to-
    # decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    str_date=request.POST.get("offered date")
    date1=datetime.datetime.strptime(str_date,'%Y-%m-%d')
    # date1 =date(2023,7,31)
    # date1=cal.get_date()
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]

    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        print(len(index))
        dr=df_pivot[kpi_name]
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()

        ws[coln1+"2"].value=cl[4]
        ws[coln2+"2"].value=cl[3]
        ws[coln3+"2"].value=cl[2]
        ws[coln4+"2"].value=cl[1]
        ws[coln5+"2"].value=cl[0]

        for i,value in enumerate(index):
                j=i+3
        
                ws['A'+str(j)].value=index[i][0]
                ws['B'+str(j)].value=index[i][1]
                   
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
                
    for kpi_name in kpi:
        
        if(kpi_name=='RRC Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'G') 
            
        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'M')    
            
        if(kpi_name=='PS Drop Call Rate % [CDBH]'):
            overwrite(kpi_name,'S')   
            
        if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name,'Y')
            
        if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'AE')
            
        if(kpi_name=='PS handover success rate [LTE Inter System] [CDBH]'):
            overwrite(kpi_name,'AK')
            
        if(kpi_name=='MV_4G Data Volume_GB'):
            overwrite(kpi_name,'AQ')  
            
        if(kpi_name=='MV_Average number of used DL PRBs'):
            overwrite(kpi_name,'AV')
            
        if(kpi_name=='MV_DL User Throughput_Mbps [CDBH]'):
            overwrite(kpi_name,'BB') 
            
        if(kpi_name=='MV_UL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name,'BH') 
            
        if(kpi_name=='MV_VoLTE ERAB Setup Success Rate'):
            overwrite(kpi_name,'BN')
            
        if(kpi_name=='MV_VoLTE DCR [CBBH]'):
            overwrite(kpi_name,'BT')
            
        if(kpi_name=='MV_VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name,'BZ')
            
        if(kpi_name=='MV_VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name,'CF')
            
        if(kpi_name=='MV_VoLTE IntraF HOSR Exec [CBBH]'):
            overwrite(kpi_name,'CL')
            
        if(kpi_name=='VoLTE InterF HOSR Exec'):
            overwrite(kpi_name,'CR')
            
        if(kpi_name=='MV_VoLTE Traffic [K erl]'):
            overwrite(kpi_name,'CX')
            
        if(kpi_name=='MV_VoLTE Traffic'):
            overwrite(kpi_name,'DC')
            
        if(kpi_name=='MV_CSFB Redirection Success Rate'):
            overwrite(kpi_name,'DH') 
            
        if(kpi_name=='UL RSSI [CDBH]'):
            overwrite(kpi_name,'DM')

        if(kpi_name=='MV_Avg Connected User'):
            overwrite(kpi_name,'DR')    

    save_output=os.path.join(door_path,'output','JK_Trend_Output.xlsx')    
    wb.save(save_output)
    download_path=os.path.join(MEDIA_URL,'trends','jk','output','JK_trend_output.xlsx')
    return Response({"status":True,"message":"uploaded successfully",'status':True,"missing_sites":message,'Download_url':download_path})
