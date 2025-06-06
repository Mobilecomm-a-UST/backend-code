
from django.shortcuts import render
from django.shortcuts import render,redirect
from django.http import HttpResponse
import sys
import pandas as pd
from django.contrib import messages
# import json
from django.contrib.auth.decorators import login_required
from mcom_website.settings import MEDIA_URL,BASE_DIR,ALLOWED_HOSTS
import numpy as np
import os
import datetime
from rest_framework.decorators import api_view,parser_classes
from rest_framework.response import Response
from django.http import JsonResponse

from .serializer import *

from django.forms.models import model_to_dict

from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import authentication_classes,permission_classes
from django.core.files.storage import FileSystemStorage
from mcom_website.settings import MEDIA_ROOT
from .models import *

import openpyxl
from openpyxl.styles import PatternFill
# Create your views here.
from openpyxl import Workbook,load_workbook
from statistics import mean

from django.db.models import Sum


import pandas as pd
import matplotlib  as mpl
import numpy as np

from datetime import date, timedelta


@api_view(["POST"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def old_tnch_Trend_v1(request):
        raw_kpi = request.FILES["raw_kpi"] if 'raw_kpi' in request.FILES else None
        
        site_list = request.FILES["site_list"] if 'site_list' in request.FILES else None
        offered_date = request.POST.get("offered_date")
        print(offered_date)
        offered_date=datetime.datetime.strptime(offered_date,"%Y-%m-%d").date()
        print("offered_date:",offered_date)

        location= MEDIA_ROOT + r"\Original_trend\temporary files"
        fs = FileSystemStorage(location=location)

       ####################### for converting raw kpi file to a dataframe ###########################
        raw_kpi = fs.save(raw_kpi.name, raw_kpi)
        # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
        filepath = fs.path(raw_kpi)
        print("file_path:-",filepath)
        df_raw_kpi=pd.read_excel(filepath)
        # df_raw_kpi=pd.read_csv(filepath)
        os.remove(path=filepath)
        print(filepath,"deleted...............")
        print(df_raw_kpi)

      ####################### for converting sitelist file to a dataframe ###########################
        
        site_list = fs.save(site_list.name, site_list)
        # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
        filepath = fs.path(site_list)
        print("file_path:-",filepath)
        df_site_list=pd.read_excel(filepath)
        os.remove(path=filepath)
        print(filepath,"deleted...............")
        print(df_site_list)

         # df_site=pd.read_excel("inputs/AP_SITES.xlsx")
        # df_site=pd.read_excel(site_file["text"])
        # print(df_site)
        door_root= os.path.join(MEDIA_ROOT,"Original_trend","Tnch","Trend")

        path_of_blnk_temp=os.path.join(door_root,"template","template2.xlsx")
        trend_wb=load_workbook(path_of_blnk_temp)
        print(trend_wb.sheetnames)
        print("##################################################################################")
        g2_site=[]
        f8_site=[]
        f3_site=[]
        t1t2_site=[]


        kpi=[   "MV_RRC Setup Success Rate [CDBH]",
                "MV_ERAB Setup Success Rate [CDBH]",
                "MV_PS Drop Call Rate % [CDBH]",
                "MV_DL User Throughput_Mbps [CDBH]",
                "MV_UL User Throughput_Mbps [CDBH]",
                "MV_PS handover success rate [LTE Intra System] [CDBH]",
                "MV_PS handover success rate [LTE Inter System] [CDBH]", 
                "MV_CSFB Redirection Success Rate [CDBH]",
                "MV_E-UTRAN Average CQI [CDBH]",
                "UL RSSI PUCCH",
                "MV_VoLTE ERAB Setup Success Rate [CBBH]",
                "MV_VoLTE DCR [CBBH]",
                "MV_VoLTE Packet Loss DL [CBBH]",
                "MV_VoLTE Packet Loss UL [CBBH]",
                "MV_VoLTE IntraF HOSR Exec [CBBH]",
                "MV_4G Data Volume_GB",
                "MV_VoLTE InterF HOSR Exec [CBBH]", 
                "MV_Average number of used DL PRBs [CDBH]",
                "MV_VoLTE Traffic",



            
       ]
        kpiStrToZero=[   "MV_RRC Setup Success Rate [CDBH]",
                "MV_ERAB Setup Success Rate [CDBH]",
                "MV_PS Drop Call Rate % [CDBH]",
                "MV_DL User Throughput_Kbps [CDBH]",
                "MV_UL User Throughput_Kbps [CDBH]",
                "MV_PS handover success rate [LTE Intra System] [CDBH]",
                "MV_PS handover success rate [LTE Inter System] [CDBH]", 
                "MV_CSFB Redirection Success Rate [CDBH]",
                "MV_E-UTRAN Average CQI [CDBH]",
                "UL RSSI PUCCH",
                "MV_VoLTE ERAB Setup Success Rate [CBBH]",
                "MV_VoLTE DCR [CBBH]",
                "MV_VoLTE Packet Loss DL [CBBH]",
                "MV_VoLTE Packet Loss UL [CBBH]",
                "MV_VoLTE IntraF HOSR Exec [CBBH]",
                "MV_4G Data Volume_GB",
                "MV_VoLTE InterF HOSR Exec [CBBH]", 
                "MV_Average number of used DL PRBs [CDBH]",
                "MV_VoLTE Traffic",



            
       ]
        # kpi1=[   
        #       # "MV_RRC Setup Success Rate [CDBH]",
        #         # "MV_ERAB Setup Success Rate [CDBH]",
        #         # "MV_PS Drop Call Rate % [CDBH]",
        #         "MV_DL User Throughput_Mbps [CDBH]",
        #         # "MV_UL User Throughput_Mbps [CDBH]",
        #         "MV_PS handover success rate [LTE Intra System] [CDBH]",
        #         # "MV_PS handover success rate [LTE Inter System] [CDBH]", 
        #         # "MV_CSFB Redirection Success Rate [CDBH]",
        #         "MV_E-UTRAN Average CQI [CDBH]",
        #         "UL RSSI PUCCH",
        #         # "MV_VoLTE ERAB Setup Success Rate [CBBH]",
        #         "MV_VoLTE DCR [CBBH]",
        #         "MV_VoLTE Packet Loss DL [CBBH]",
        #         "MV_VoLTE Packet Loss UL [CBBH]",
        #         "MV_VoLTE IntraF HOSR Exec [CBBH]",
        #         "MV_4G Data Volume_GB",
        #         # "MV_VoLTE InterF HOSR Exec [CBBH]", 
        #         # "MV_Average number of used DL PRBs [CDBH]",
        #         "MV_VoLTE Traffic",
        #     ]
        
        ########################## the below code is to replace every string to zero from numeric columns ##################

        for x in kpiStrToZero:
             df_raw_kpi[x] = df_raw_kpi[x].replace(to_replace='.*', value=0, regex=True)
        ######################################  * * * * * * * * * * * * * * ################################################
       
        site_list=list(df_site_list["2G ID"])

 

        print("__________________raw KPI after converting str to zero_______________")
        print(df_raw_kpi)

        df_raw_kpi["Short name"].fillna( inplace=True, method="ffill")
        df_raw_kpi["Short name"] =df_raw_kpi["Short name"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..


        df_raw_kpi.rename( columns={'Unnamed: 1':'date'}, inplace=True )

        df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"]/1024)
        df_raw_kpi.rename(columns={"MV_DL User Throughput_Kbps [CDBH]" :"MV_DL User Throughput_Mbps [CDBH]" } ,inplace = True )

        df_raw_kpi["MV_UL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_UL User Throughput_Kbps [CDBH]"]/1024)
        df_raw_kpi.rename(columns={"MV_UL User Throughput_Kbps [CDBH]" :"MV_UL User Throughput_Mbps [CDBH]" } ,inplace = True )

        print(df_raw_kpi)
        print(df_raw_kpi.columns)



        lis=list(df_raw_kpi["Short name"])
        sit_id_lis=[]
        cell_id_lis=[]
        for item in lis:
            if("_" in item):
                cell_id=item.split("_")[-2]
                ln=len(item.split("_")[-1])
                #print(ln)
                sit_id=item.split("_")[-2][:-ln]
            else:
                cell_id=item
                sit_id=item
            cell_id_lis.append(cell_id)
            sit_id_lis.append(sit_id)

        print(sit_id)
        print(cell_id_lis)

        df_raw_kpi.insert(1, "SITE_ID", sit_id_lis)
        df_raw_kpi.insert(2, "CELL_ID", cell_id_lis)

        
    
        df_raw_kpi.rename(columns={"Short name" :"Shortname" } ,inplace = True )
        df_raw_kpi.fillna(value=0,inplace=True)

        process_op_path=os.path.join(door_root,"process_outputs")
        
        savepath=os.path.join(process_op_path,"desired input.xlsx")
        df_raw_kpi.to_excel(savepath)
    
        # date1=date(2023,2,27)
        date1=offered_date
        # date1=date.today()
        dt1 = date1 - timedelta(1)
        dt2 = date1 - timedelta(2)
        dt3 = date1 - timedelta(3)
        dt4 = date1 - timedelta(4)
        dt5 = date1 - timedelta(5)
        ls=[dt1,dt2,dt3,dt4,dt5]
        only_site_fil = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.date.isin(ls))]
        
        savepath=os.path.join(process_op_path,"only_site_date_filtered_input.xlsx")
        only_site_fil.to_excel(savepath)

        def perticular_tech( tech,site_list):
            # df_filtered = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.date.isin(ls))]
            df_filtered = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.date.isin(ls)) & (df_raw_kpi.Shortname.str.contains('|'.join(tech)))]
            
            print(df_filtered)
            if not df_filtered.empty:
                address="last_filtered_input"  +str(tech)+ ".xlsx"
                savepath=os.path.join(process_op_path,address)
                df_filtered.to_excel(savepath)
                df_pivoted = df_filtered.pivot_table(index=["SITE_ID","Shortname","CELL_ID"], columns="date")
                print("technology:",tech)
                print(df_pivoted)
                address_pivot="pivoted_input"  +str(tech)+ ".xlsx"
                savepath=os.path.join(process_op_path,address_pivot)
                df_pivoted.to_excel(savepath)
                return df_pivoted
            
            return df_filtered


      
        alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        def num_hash(num):
                if num < 26:
                    return alpha[num-1]
                else:
                    q, r = num//26, num % 26
                    if r == 0:
                        if q == 1:
                            return alpha[r-1]
                        else:
                            return num_hash(q-1) + alpha[r-1]
                    else:
                        return num_hash(q) + alpha[r-1]
    
    # Driver code

    # printString(27906)

        def titleToNumber(s):
        # This process is similar to binary-to-
        # decimal conversion
            result = 0
            for B in range(len(s)):
                result *= 26
                result += ord(s[B]) - ord('A') + 1
            return result


        def overwrite(df_pivoted,kpi_name,coln1,trend_ws):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            print(kpi_name)
            index_pivot=df_pivoted.index.to_list()
            print("index ;###############################",index_pivot)
            print(len(index_pivot))
            print("index of pivoted table: ",index_pivot)
           
            dr=df_pivoted[kpi_name]
           

            print("columns of dr table",dr.columns)
            cl=dr.columns.to_list()
            print("column list",cl)
            
            # site_id=dr["SITE_ID"].to_list() 
            # cell_id=dr["CELL_ID"].to_list()    
            col1=dr[str(cl[0])].to_list()
            col2=dr[str(cl[1])].to_list()
            col3=dr[str(cl[2])].to_list()
            col4=dr[str(cl[3])].to_list()
            col5=dr[str(cl[4])].to_list()

            trend_ws[coln1+"3"].value=cl[0]
            trend_ws[coln2+"3"].value=cl[1]
            trend_ws[coln3+"3"].value=cl[2]
            trend_ws[coln4+"3"].value=cl[3]
            trend_ws[coln5+"3"].value=cl[4]

            # me=column_index_from_string(coln5)+1
            # me=get_column_letter(me)
            for i,value in enumerate(index_pivot):
                j=i+4
                trend_ws["A"+str(j)].value=index_pivot[i][1]
                trend_ws["E"+str(j)].value=index_pivot[i][0] 
                # trend_ws["C"+str(j)].value=index_pivot[i][2]
                # trend_ws["K"+str(j)].value=date1
                
                
                trend_ws[coln1+str(j)].value=col1[i]
                trend_ws[coln2+str(j)].value=col2[i]
                trend_ws[coln3+str(j)].value=col3[i]
                trend_ws[coln4+str(j)].value=col4[i]
                trend_ws[coln5+str(j)].value=col5[i]
                
                # trend_ws[me+str(j)].value='=COUNTIF(P5:T5,">=98.5")'

        # def overwrite_g2(df_pivoted,kpi_name,coln1,coln2,coln3,coln4,coln5,trend_ws):
        #     index_pivot=df_pivoted.index.to_list()
        #     print("index ;###############################",index_pivot)
        #     print(len(index_pivot))
        #     print("index of pivoted table: ",index_pivot)
        #     dr=df_pivoted[kpi_name]
        #     print("columns of dr table",dr.columns)
        #     cl=dr.columns.to_list()
        #     print("column list",cl)
            
        #     # site_id=dr["SITE_ID"].to_list() 
        #     # cell_id=dr["CELL_ID"].to_list()    
        #     col1=dr[str(cl[0])].to_list()
        #     col2=dr[str(cl[1])].to_list()
        #     col3=dr[str(cl[2])].to_list()
        #     col4=dr[str(cl[3])].to_list()
        #     col5=dr[str(cl[4])].to_list()

        #     trend_ws[coln1+"4"].value=cl[0]
        #     trend_ws[coln2+"4"].value=cl[1]
        #     trend_ws[coln3+"4"].value=cl[2]
        #     trend_ws[coln4+"4"].value=cl[3]
        #     trend_ws[coln5+"4"].value=cl[4]

        #     # me=column_index_from_string(coln5)+1
        #     # me=get_column_letter(me)
        #     for i,value in enumerate(index_pivot):
        #         j=i+5
        #         trend_ws["B"+str(j)].value=index_pivot[i][0]
        #         trend_ws["C"+str(j)].value=index_pivot[i][1]
        #         trend_ws["G"+str(j)].value=date1
                
                
                
        #         trend_ws[coln1+str(j)].value=col1[i]
        #         trend_ws[coln2+str(j)].value=col2[i]
        #         trend_ws[coln3+str(j)].value=col3[i]
        #         trend_ws[coln4+str(j)].value=col4[i]
        #         trend_ws[coln5+str(j)].value=col5[i]



        # for fdd
        pivot_fdd=perticular_tech(["_F3_"],site_list)
        trend_ws=trend_wb["L1800"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="MV_ERAB Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="MV_PS Drop Call Rate % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)

                if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)

                if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)

                if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="MV_PS handover success rate [LTE Inter System] [CDBH]"): 
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="MV_VoLTE ERAB Setup Success Rate [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)
                
                if(kpi_name=="MV_VoLTE IntraF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)

                if(kpi_name=="MV_4G Data Volume_GB"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)

                if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)
                if(kpi_name=="MV_VoLTE Traffic"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)
            
        

        # for tdd
        pivot_fdd=perticular_tech(["_T1_","_T2_"],site_list)
        trend_ws=trend_wb["L2300"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="MV_ERAB Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="MV_PS Drop Call Rate % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)

                if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)

                if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)

                if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="MV_PS handover success rate [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="MV_VoLTE ERAB Setup Success Rate [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)
                
                if(kpi_name=="MV_VoLTE IntraF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)

                if(kpi_name=="MV_4G Data Volume_GB"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)

                if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)
                if(kpi_name=="MV_VoLTE Traffic"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)
            

        pivot_fdd=perticular_tech(["_F8_"],site_list)
        trend_ws=trend_wb["L900"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="MV_ERAB Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="MV_PS Drop Call Rate % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)

                if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)

                if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)

                if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="MV_PS handover success rate [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="MV_VoLTE ERAB Setup Success Rate [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)
                
                if(kpi_name=="MV_VoLTE IntraF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)

                if(kpi_name=="MV_4G Data Volume_GB"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)

                if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)
                if(kpi_name=="MV_VoLTE Traffic"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)




        pivot_fdd=perticular_tech(["_F1_"],site_list)
        trend_ws=trend_wb["L2100"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="MV_ERAB Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="MV_PS Drop Call Rate % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)

                if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)

                if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)

                if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="MV_PS handover success rate [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="MV_VoLTE ERAB Setup Success Rate [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)
                
                if(kpi_name=="MV_VoLTE IntraF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)

                if(kpi_name=="MV_4G Data Volume_GB"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)

                if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)
                if(kpi_name=="MV_VoLTE Traffic"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)



        # # print(g2_tech(g2_site)[1])
        # pivote_g2=g2_tech(g2_site)[1]
        # # pivote_g2.to_excel("g2_pivoted.xlsx")
        # trend_ws=trend_wb["2G"]



        # for kpi_name in g2_kpi:
        #     if(kpi_name=="SDCCH Blocking Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"L","M","N","O","P",trend_ws)
            
        #     if(kpi_name=="SDCCH Drop Call Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"S","T","U","V","W",trend_ws)
            
        #     if(kpi_name=="TCH Blocking Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"Z","AA","AB","AC","AD",trend_ws)

        #     if(kpi_name=="Drop Call Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"AG","AH","AI","AJ","AK",trend_ws)
            
        #     if(kpi_name=="Handover Success Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"AO","AP","AQ","AR","AS",trend_ws)
            

        #     if(kpi_name=="RX Quality"):
        #         overwrite_g2(pivote_g2,kpi_name,"AV","AW","AX","AY","AZ",trend_ws)
    
        #     if(kpi_name=="RACH Success Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"BC","BD","BE","BF","BG",trend_ws)

            
        #     if(kpi_name=="Total Voice Traffic [BBH]"):
        #         overwrite_g2(pivote_g2,kpi_name,"BJ","BK","BL","BM","BN",trend_ws)

            
        #     if(kpi_name=="TCH Assignment Success Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"BP","BQ","BR","BS","BT",trend_ws)

            




        output_path=os.path.join(door_root,"output","Tnch_trend_output.xlsx")
        trend_wb.save(output_path)
        # load_excel_data()
        # directory = "Tnch_OUTPUT_trend"
    
        # # Parent Directory path
        # parent_dir = "C:/Users/dell7480/Desktop/" 
        # # Path
        # path = os.path.join(parent_dir, directory)
        # print(path)
        # #Create the directory
        # if(not (os.path.isdir(path))):
        #     os.makedirs(path)
        #     print("Directory '% s' created" % directory)
        # path= path + "/trend_output.xlsx"
        # trend_wb.save(path)
        download_path=os.path.join(MEDIA_URL,"Original_trend","Tnch","Trend","output","Tnch_trend_output.xlsx")
        return Response({"status":True,"message":"Succesfully uploaded","Download_url":download_path})

################################## nom denom addition and template change version 3, Note: there is no v2 ##################################
@api_view(["POST"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def old_tnch_Trend_v3(request):
        print("33333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333")
       
    
        raw_kpi = request.FILES["raw_kpi"] if 'raw_kpi' in request.FILES else None
        
        site_list = request.FILES["site_list"] if 'site_list' in request.FILES else None
        offered_date = request.POST.get("offered_date")
        print(offered_date)
        offered_date=datetime.datetime.strptime(offered_date,"%Y-%m-%d").date()
        print("offered_date:",offered_date)

        location= MEDIA_ROOT + r"\Original_trend\temporary files"
        fs = FileSystemStorage(location=location)

       ####################### for converting raw kpi file to a dataframe ###########################
        raw_kpi = fs.save(raw_kpi.name, raw_kpi)
        # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
        filepath = fs.path(raw_kpi)
        print("file_path:-",filepath)
        df_raw_kpi=pd.read_excel(filepath)
        # df_raw_kpi=pd.read_csv(filepath)
        os.remove(path=filepath)
        print(filepath,"deleted...............")
        print(df_raw_kpi)

      ####################### for converting sitelist file to a dataframe ###########################
        
        site_list = fs.save(site_list.name, site_list)
        # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
        filepath = fs.path(site_list)
        print("file_path:-",filepath)
        df_site_list=pd.read_excel(filepath)
        os.remove(path=filepath)
        print(filepath,"deleted...............")
        print(df_site_list)

         # df_site=pd.read_excel("inputs/AP_SITES.xlsx")
        # df_site=pd.read_excel(site_file["text"])
        # print(df_site)
        door_root= os.path.join(MEDIA_ROOT,"Original_trend","Tnch","Trend")

        path_of_blnk_temp=os.path.join(door_root,"template","template4.xlsx")
        trend_wb=load_workbook(path_of_blnk_temp)
        print(trend_wb.sheetnames)
        print("##################################################################################")
        g2_site=[]
        f8_site=[]
        f3_site=[]
        t1t2_site=[]


        kpi=[   "MV_RRC Setup Success Rate [CDBH]",
                "MV_ERAB Setup Success Rate [CDBH]",
                "MV_PS Drop Call Rate % [CDBH]",
                "MV_DL User Throughput_Mbps [CDBH]",
                "MV_UL User Throughput_Mbps [CDBH]",
                "MV_PS handover success rate [LTE Intra System] [CDBH]",
                "MV_PS handover success rate [LTE Inter System] [CDBH]", 
                "MV_CSFB Redirection Success Rate [CDBH]",
                "MV_E-UTRAN Average CQI [CDBH]",
                "UL RSSI PUCCH",
                "MV_VoLTE ERAB Setup Success Rate [CBBH]",
                "MV_VoLTE DCR [CBBH]",
                "MV_VoLTE Packet Loss DL [CBBH]",
                "MV_VoLTE Packet Loss UL [CBBH]",
                "MV_VoLTE IntraF HOSR Exec [CBBH]",
                "MV_4G Data Volume_GB",
                "MV_VoLTE InterF HOSR Exec [CBBH]", 
                "MV_Average number of used DL PRBs [CDBH]",
                "MV_VoLTE Traffic",

                "PS handover success rate_Nom [LTE Intra System] [CDBH]",
                "PS handover success rate_Denom [LTE Intra System] [CDBH]",
                "PS handover success rate_Nom [LTE Inter System] [CDBH]",
                "PS handover success rate_Denom [LTE Inter System] [CDBH]",
                "VoLTE DCR_Nom [CBBH]",
                "PS Drop Call Rate_Nom [CDBH]",
                "VoLTE ERAB Setup Success Rate_NOM [CBBH]",
                "VoLTE ERAB Setup Success Rate_Denom [CBBH]",
                "VoLTE IntraF HOSR Exec_Nom [CBBH]",
                "VoLTE IntraF HOSR Exec_Denom [CBBH]",
                "VoLTE InterF HOSR Exec_Nom [CBBH]",
                "VoLTE InterF HOSR Exec_Denom [CBBH]",
                "MV_Radio NW Availability"

            
       ]
        kpiStrToZero=[ "MV_RRC Setup Success Rate [CDBH]",
                "MV_ERAB Setup Success Rate [CDBH]",
                "MV_PS Drop Call Rate % [CDBH]",
                "MV_DL User Throughput_Kbps [CDBH]",
                "MV_UL User Throughput_Kbps [CDBH]",
                "MV_PS handover success rate [LTE Intra System] [CDBH]",
                "MV_PS handover success rate [LTE Inter System] [CDBH]", 
                "MV_CSFB Redirection Success Rate [CDBH]",
                "MV_E-UTRAN Average CQI [CDBH]",
                "UL RSSI PUCCH",
                "MV_VoLTE ERAB Setup Success Rate [CBBH]",
                "MV_VoLTE DCR [CBBH]",
                "MV_VoLTE Packet Loss DL [CBBH]",
                "MV_VoLTE Packet Loss UL [CBBH]",
                "MV_VoLTE IntraF HOSR Exec [CBBH]",
                "MV_4G Data Volume_GB",
                "MV_VoLTE InterF HOSR Exec [CBBH]", 
                "MV_Average number of used DL PRBs [CDBH]",
                "MV_VoLTE Traffic",

                "PS handover success rate_Nom [LTE Intra System] [CDBH]",
                "PS handover success rate_Denom [LTE Intra System] [CDBH]",
                "PS handover success rate_Nom [LTE Inter System] [CDBH]",
                "PS handover success rate_Denom [LTE Inter System] [CDBH]",
                "VoLTE DCR_Nom [CBBH]",
                "PS Drop Call Rate_Nom [CDBH]",
                "VoLTE ERAB Setup Success Rate_NOM [CBBH]",
                "VoLTE ERAB Setup Success Rate_Denom [CBBH]",
                "VoLTE IntraF HOSR Exec_Nom [CBBH]",
                "VoLTE IntraF HOSR Exec_Denom [CBBH]",
                "VoLTE InterF HOSR Exec_Nom [CBBH]",
                "VoLTE InterF HOSR Exec_Denom [CBBH]",
                "MV_Radio NW Availability"

            
       ]
        # kpi1=[   
        #       # "MV_RRC Setup Success Rate [CDBH]",
        #         # "MV_ERAB Setup Success Rate [CDBH]",
        #         # "MV_PS Drop Call Rate % [CDBH]",
        #         "MV_DL User Throughput_Mbps [CDBH]",
        #         # "MV_UL User Throughput_Mbps [CDBH]",
        #         "MV_PS handover success rate [LTE Intra System] [CDBH]",
        #         # "MV_PS handover success rate [LTE Inter System] [CDBH]", 
        #         # "MV_CSFB Redirection Success Rate [CDBH]",
        #         "MV_E-UTRAN Average CQI [CDBH]",
        #         "UL RSSI PUCCH",
        #         # "MV_VoLTE ERAB Setup Success Rate [CBBH]",
        #         "MV_VoLTE DCR [CBBH]",
        #         "MV_VoLTE Packet Loss DL [CBBH]",
        #         "MV_VoLTE Packet Loss UL [CBBH]",
        #         "MV_VoLTE IntraF HOSR Exec [CBBH]",
        #         "MV_4G Data Volume_GB",
        #         # "MV_VoLTE InterF HOSR Exec [CBBH]", 
        #         # "MV_Average number of used DL PRBs [CDBH]",
        #         "MV_VoLTE Traffic",
        #     ]
        
        ########################## the below code is to replace every string to zero from numeric columns ##################

        for x in kpiStrToZero:
             df_raw_kpi[x] = df_raw_kpi[x].replace(to_replace='.*', value=0, regex=True)
        ######################################  * * * * * * * * * * * * * * ################################################
       
        site_list=list(df_site_list["2G ID"])

 

        print("__________________raw KPI after converting str to zero_______________")
        print(df_raw_kpi)

        df_raw_kpi["Short name"].fillna( inplace=True, method="ffill")
        df_raw_kpi["Short name"] =df_raw_kpi["Short name"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..


        df_raw_kpi.rename( columns={'Unnamed: 1':'date'}, inplace=True )

        df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"]/1024)
        df_raw_kpi.rename(columns={"MV_DL User Throughput_Kbps [CDBH]" :"MV_DL User Throughput_Mbps [CDBH]" } ,inplace = True )

        df_raw_kpi["MV_UL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_UL User Throughput_Kbps [CDBH]"]/1024)
        df_raw_kpi.rename(columns={"MV_UL User Throughput_Kbps [CDBH]" :"MV_UL User Throughput_Mbps [CDBH]" } ,inplace = True )

        print(df_raw_kpi)

        
        print(df_raw_kpi.columns)



        lis=list(df_raw_kpi["Short name"])
        sit_id_lis=[]
        cell_id_lis=[]
        for item in lis:
            if("_" in item):
                cell_id=item.split("_")[-2]
                ln=len(item.split("_")[-1])
                #print(ln)
                sit_id=item.split("_")[-2][:-ln]
            else:
                cell_id=item
                sit_id=item
            cell_id_lis.append(cell_id)
            sit_id_lis.append(sit_id)

        print(sit_id)
        print(cell_id_lis)

        df_raw_kpi.insert(1, "SITE_ID", sit_id_lis)
        df_raw_kpi.insert(2, "CELL_ID", cell_id_lis)

        
    
        df_raw_kpi.rename(columns={"Short name" :"Shortname" } ,inplace = True )
        df_raw_kpi.fillna(value=0,inplace=True)

        process_op_path=os.path.join(door_root,"process_outputs")
        
        savepath=os.path.join(process_op_path,"desired input.xlsx")
        df_raw_kpi.to_excel(savepath)
    
        # date1=date(2023,2,27)
        date1=offered_date
        # date1=date.today()
        dt1 = date1 - timedelta(1)
        dt2 = date1 - timedelta(2)
        dt3 = date1 - timedelta(3)
        dt4 = date1 - timedelta(4)
        dt5 = date1 - timedelta(5)
        ls=[dt1,dt2,dt3,dt4,dt5]
        only_site_fil = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.date.isin(ls))]
        
        savepath=os.path.join(process_op_path,"only_site_date_filtered_input.xlsx")
        only_site_fil.to_excel(savepath)

        def perticular_tech( tech,site_list):
            # df_filtered = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.date.isin(ls))]
            df_filtered = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.date.isin(ls)) & (df_raw_kpi.Shortname.str.contains('|'.join(tech)))]
            
            print(df_filtered)
            if not df_filtered.empty:
                address="last_filtered_input"  + str(tech) + ".xlsx"
                savepath=os.path.join(process_op_path,address)
                df_filtered.to_excel(savepath)
                df_pivoted = df_filtered.pivot_table(index=["SITE_ID","Shortname","CELL_ID"], columns="date")
                print("technology:",tech)
                print(df_pivoted)
                address_pivot="pivoted_input"  +str(tech)+ ".xlsx"
                savepath=os.path.join(process_op_path,address_pivot)
                df_pivoted.to_excel(savepath)
                return df_pivoted
            
            return df_filtered


      
        alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        def num_hash(num):
                if num < 26:
                    return alpha[num-1]
                else:
                    q, r = num//26, num % 26
                    if r == 0:
                        if q == 1:
                            return alpha[r-1]
                        else:
                            return num_hash(q-1) + alpha[r-1]
                    else:
                        return num_hash(q) + alpha[r-1]
    
    # Driver code

    # printString(27906)

        def titleToNumber(s):
        # This process is similar to binary-to-
        # decimal conversion
            result = 0
            for B in range(len(s)):
                result *= 26
                result += ord(s[B]) - ord('A') + 1
            return result


        def overwrite(df_pivoted,kpi_name,coln1,trend_ws):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            print(kpi_name)
            index_pivot=df_pivoted.index.to_list()
            print("index ;###############################",index_pivot)
            print(len(index_pivot))
            print("index of pivoted table: ",index_pivot)
           
            dr=df_pivoted[kpi_name]
           

            print("columns of dr table",dr.columns)
            cl=dr.columns.to_list()
            print("column list",cl)
            
            # site_id=dr["SITE_ID"].to_list() 
            # cell_id=dr["CELL_ID"].to_list()    
            col1=dr[str(cl[0])].to_list()
            col2=dr[str(cl[1])].to_list()
            col3=dr[str(cl[2])].to_list()
            col4=dr[str(cl[3])].to_list()
            col5=dr[str(cl[4])].to_list()

            trend_ws[coln1+"3"].value=cl[0]
            trend_ws[coln2+"3"].value=cl[1]
            trend_ws[coln3+"3"].value=cl[2]
            trend_ws[coln4+"3"].value=cl[3]
            trend_ws[coln5+"3"].value=cl[4]

            # me=column_index_from_string(coln5)+1
            # me=get_column_letter(me)
            for i,value in enumerate(index_pivot):
                j=i+4
                trend_ws["A"+str(j)].value=index_pivot[i][1]
                trend_ws["E"+str(j)].value=index_pivot[i][0] 
                # trend_ws["C"+str(j)].value=index_pivot[i][2]
                # trend_ws["K"+str(j)].value=date1
                
                
                trend_ws[coln1+str(j)].value=col1[i]
                trend_ws[coln2+str(j)].value=col2[i]
                trend_ws[coln3+str(j)].value=col3[i]
                trend_ws[coln4+str(j)].value=col4[i]
                trend_ws[coln5+str(j)].value=col5[i]
                
                # trend_ws[me+str(j)].value='=COUNTIF(P5:T5,">=98.5")'

        # def overwrite_g2(df_pivoted,kpi_name,coln1,coln2,coln3,coln4,coln5,trend_ws):
        #     index_pivot=df_pivoted.index.to_list()
        #     print("index ;###############################",index_pivot)
        #     print(len(index_pivot))
        #     print("index of pivoted table: ",index_pivot)
        #     dr=df_pivoted[kpi_name]
        #     print("columns of dr table",dr.columns)
        #     cl=dr.columns.to_list()
        #     print("column list",cl)
            
        #     # site_id=dr["SITE_ID"].to_list() 
        #     # cell_id=dr["CELL_ID"].to_list()    
        #     col1=dr[str(cl[0])].to_list()
        #     col2=dr[str(cl[1])].to_list()
        #     col3=dr[str(cl[2])].to_list()
        #     col4=dr[str(cl[3])].to_list()
        #     col5=dr[str(cl[4])].to_list()

        #     trend_ws[coln1+"4"].value=cl[0]
        #     trend_ws[coln2+"4"].value=cl[1]
        #     trend_ws[coln3+"4"].value=cl[2]
        #     trend_ws[coln4+"4"].value=cl[3]
        #     trend_ws[coln5+"4"].value=cl[4]

        #     # me=column_index_from_string(coln5)+1
        #     # me=get_column_letter(me)
        #     for i,value in enumerate(index_pivot):
        #         j=i+5
        #         trend_ws["B"+str(j)].value=index_pivot[i][0]
        #         trend_ws["C"+str(j)].value=index_pivot[i][1]
        #         trend_ws["G"+str(j)].value=date1
                
                
                
        #         trend_ws[coln1+str(j)].value=col1[i]
        #         trend_ws[coln2+str(j)].value=col2[i]
        #         trend_ws[coln3+str(j)].value=col3[i]
        #         trend_ws[coln4+str(j)].value=col4[i]
        #         trend_ws[coln5+str(j)].value=col5[i]



        # for fdd
        pivot_fdd=perticular_tech(["_F3_"],site_list)
        trend_ws=trend_wb["L1800"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="MV_ERAB Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="MV_PS Drop Call Rate % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)

                if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)

                if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)

                if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="MV_PS handover success rate [LTE Inter System] [CDBH]"): 
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="MV_VoLTE ERAB Setup Success Rate [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)
                
                if(kpi_name=="MV_VoLTE IntraF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)

                if(kpi_name=="MV_4G Data Volume_GB"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)

                if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)
                if(kpi_name=="MV_VoLTE Traffic"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)
                if(kpi_name=="MV_Radio NW Availability"):
                    overwrite(pivot_fdd,kpi_name,"FM",trend_ws)

                #################################### nom/Denom #########################
                if(kpi_name=="PS handover success rate_Nom [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"FT",trend_ws)
                
                if(kpi_name=="PS handover success rate_Denom [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GA",trend_ws)
                
                if(kpi_name=="PS handover success rate_Nom [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GH",trend_ws)
                
                if(kpi_name=="PS handover success rate_Denom [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GO",trend_ws)
                
                if(kpi_name=="VoLTE DCR_Nom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"GV",trend_ws)
                
                if(kpi_name=="PS Drop Call Rate_Nom [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"HC",trend_ws)
                
                if(kpi_name=="VoLTE ERAB Setup Success Rate_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HJ",trend_ws)
            
                if(kpi_name=="VoLTE ERAB Setup Success Rate_Denom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HQ",trend_ws)
            
                if(kpi_name=="VoLTE IntraF HOSR Exec_Nom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HX",trend_ws)

                if(kpi_name=="VoLTE IntraF HOSR Exec_Denom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IE",trend_ws)

                if(kpi_name=="VoLTE InterF HOSR Exec_Nom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IL",trend_ws)

                if(kpi_name=="VoLTE InterF HOSR Exec_Denom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IS",trend_ws)
            
        

        # for tdd
        pivot_fdd=perticular_tech(["_T1_","_T2_"],site_list)
        trend_ws=trend_wb["L2300"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="MV_ERAB Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="MV_PS Drop Call Rate % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)

                if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)

                if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)

                if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="MV_PS handover success rate [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="MV_VoLTE ERAB Setup Success Rate [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)
                
                if(kpi_name=="MV_VoLTE IntraF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)

                if(kpi_name=="MV_4G Data Volume_GB"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)

                if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)
                if(kpi_name=="MV_VoLTE Traffic"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)
                if(kpi_name=="MV_Radio NW Availability"):
                    overwrite(pivot_fdd,kpi_name,"FM",trend_ws)

                #################################### nom/Denom #########################
                if(kpi_name=="PS handover success rate_Nom [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"FT",trend_ws)
                
                if(kpi_name=="PS handover success rate_Denom [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GA",trend_ws)
                
                if(kpi_name=="PS handover success rate_Nom [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GH",trend_ws)
                
                if(kpi_name=="PS handover success rate_Denom [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GO",trend_ws)
                
                if(kpi_name=="VoLTE DCR_Nom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"GV",trend_ws)
                
                if(kpi_name=="PS Drop Call Rate_Nom [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"HC",trend_ws)
                
                if(kpi_name=="VoLTE ERAB Setup Success Rate_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HJ",trend_ws)
            
                if(kpi_name=="VoLTE ERAB Setup Success Rate_Denom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HQ",trend_ws)
            
                if(kpi_name=="VoLTE IntraF HOSR Exec_Nom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HX",trend_ws)

                if(kpi_name=="VoLTE IntraF HOSR Exec_Denom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IE",trend_ws)

                if(kpi_name=="VoLTE InterF HOSR Exec_Nom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IL",trend_ws)

                if(kpi_name=="VoLTE InterF HOSR Exec_Denom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IS",trend_ws)

                
            

        pivot_fdd=perticular_tech(["_F8_"],site_list)
        trend_ws=trend_wb["L900"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="MV_ERAB Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="MV_PS Drop Call Rate % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)

                if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)

                if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)

                if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="MV_PS handover success rate [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="MV_VoLTE ERAB Setup Success Rate [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)
                
                if(kpi_name=="MV_VoLTE IntraF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)

                if(kpi_name=="MV_4G Data Volume_GB"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)

                if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)
                if(kpi_name=="MV_VoLTE Traffic"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)
                if(kpi_name=="MV_Radio NW Availability"):
                    overwrite(pivot_fdd,kpi_name,"FM",trend_ws)

                #################################### nom/Denom #########################
                if(kpi_name=="PS handover success rate_Nom [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"FT",trend_ws)
                
                if(kpi_name=="PS handover success rate_Denom [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GA",trend_ws)
                
                if(kpi_name=="PS handover success rate_Nom [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GH",trend_ws)
                
                if(kpi_name=="PS handover success rate_Denom [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GO",trend_ws)
                
                if(kpi_name=="VoLTE DCR_Nom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"GV",trend_ws)
                
                if(kpi_name=="PS Drop Call Rate_Nom [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"HC",trend_ws)
                
                if(kpi_name=="VoLTE ERAB Setup Success Rate_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HJ",trend_ws)
            
                if(kpi_name=="VoLTE ERAB Setup Success Rate_Denom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HQ",trend_ws)
            
                if(kpi_name=="VoLTE IntraF HOSR Exec_Nom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HX",trend_ws)

                if(kpi_name=="VoLTE IntraF HOSR Exec_Denom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IE",trend_ws)

                if(kpi_name=="VoLTE InterF HOSR Exec_Nom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IL",trend_ws)

                if(kpi_name=="VoLTE InterF HOSR Exec_Denom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IS",trend_ws)

                    




        pivot_fdd=perticular_tech(["_F1_"],site_list)
        trend_ws=trend_wb["L2100"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="MV_ERAB Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="MV_PS Drop Call Rate % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)

                if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)

                if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)

                if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="MV_PS handover success rate [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="MV_VoLTE ERAB Setup Success Rate [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)
                
                if(kpi_name=="MV_VoLTE IntraF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)

                if(kpi_name=="MV_4G Data Volume_GB"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)

                if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)
                if(kpi_name=="MV_VoLTE Traffic"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)
                if(kpi_name=="MV_Radio NW Availability"):
                    overwrite(pivot_fdd,kpi_name,"FM",trend_ws)

                #################################### nom/Denom #########################
                if(kpi_name=="PS handover success rate_Nom [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"FT",trend_ws)
                
                if(kpi_name=="PS handover success rate_Denom [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GA",trend_ws)
                
                if(kpi_name=="PS handover success rate_Nom [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GH",trend_ws)
                
                if(kpi_name=="PS handover success rate_Denom [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"GO",trend_ws)
                
                if(kpi_name=="VoLTE DCR_Nom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"GV",trend_ws)
                
                if(kpi_name=="PS Drop Call Rate_Nom [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"HC",trend_ws)
                
                if(kpi_name=="VoLTE ERAB Setup Success Rate_NOM [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HJ",trend_ws)
            
                if(kpi_name=="VoLTE ERAB Setup Success Rate_Denom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HQ",trend_ws)
            
                if(kpi_name=="VoLTE IntraF HOSR Exec_Nom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"HX",trend_ws)

                if(kpi_name=="VoLTE IntraF HOSR Exec_Denom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IE",trend_ws)

                if(kpi_name=="VoLTE InterF HOSR Exec_Nom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IL",trend_ws)

                if(kpi_name=="VoLTE InterF HOSR Exec_Denom [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"IS",trend_ws)



        # # print(g2_tech(g2_site)[1])
        # pivote_g2=g2_tech(g2_site)[1]
        # # pivote_g2.to_excel("g2_pivoted.xlsx")
        # trend_ws=trend_wb["2G"]



        # for kpi_name in g2_kpi:
        #     if(kpi_name=="SDCCH Blocking Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"L","M","N","O","P",trend_ws)
            
        #     if(kpi_name=="SDCCH Drop Call Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"S","T","U","V","W",trend_ws)
            
        #     if(kpi_name=="TCH Blocking Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"Z","AA","AB","AC","AD",trend_ws)

        #     if(kpi_name=="Drop Call Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"AG","AH","AI","AJ","AK",trend_ws)
            
        #     if(kpi_name=="Handover Success Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"AO","AP","AQ","AR","AS",trend_ws)
            

        #     if(kpi_name=="RX Quality"):
        #         overwrite_g2(pivote_g2,kpi_name,"AV","AW","AX","AY","AZ",trend_ws)
    
        #     if(kpi_name=="RACH Success Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"BC","BD","BE","BF","BG",trend_ws)

            
        #     if(kpi_name=="Total Voice Traffic [BBH]"):
        #         overwrite_g2(pivote_g2,kpi_name,"BJ","BK","BL","BM","BN",trend_ws)

            
        #     if(kpi_name=="TCH Assignment Success Rate"):
        #         overwrite_g2(pivote_g2,kpi_name,"BP","BQ","BR","BS","BT",trend_ws)

            




        output_path=os.path.join(door_root,"output","Tnch_trend_output.xlsx")
        trend_wb.save(output_path)
        # load_excel_data()
        # directory = "Tnch_OUTPUT_trend"
    
        # # Parent Directory path
        # parent_dir = "C:/Users/dell7480/Desktop/" 
        # # Path
        # path = os.path.join(parent_dir, directory)
        # print(path)
        # #Create the directory
        # if(not (os.path.isdir(path))):
        #     os.makedirs(path)
        #     print("Directory '% s' created" % directory)
        # path= path + "/trend_output.xlsx"
        # trend_wb.save(path)
        download_path=os.path.join(MEDIA_URL,"Original_trend","Tnch","Trend","output","Tnch_trend_output.xlsx")
        return Response({"status":True,"message":"Succesfully uploaded","Download_url":download_path})

@api_view(["POST"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def tnch_Trend(request):
        ############################## selected sites ########################
        selected_sites=request.POST.get("site")
        print(type(selected_sites))
        print("Selected Site :",selected_sites)
        selected_sites=selected_sites.split(",")
        print("Selected Site list :",selected_sites)
        
        ############################  site list ###############################
        exl_site_list=[]
        site_list_file = request.FILES["site_list"] if 'site_list' in request.FILES else None
        if site_list_file:
            location= MEDIA_ROOT + r"\Original_trend\temporary files"
            fs = FileSystemStorage(location=location)
            site_list_file = fs.save(site_list_file.name, site_list_file)
            # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
            filepath = fs.path(site_list_file)
            print("file_path:-",filepath)
            df_site_list=pd.read_excel(filepath)
            os.remove(path=filepath)
            print(filepath,"deleted...............")
            print(df_site_list)
            exl_site_list=list(df_site_list["2G ID"])
        
        exl_site_list.extend(selected_sites)
        set_sit=set(exl_site_list)
        ext_site_list=list(set_sit)
        
        print("Extended site list: ",ext_site_list)
        
        intg_sites=list(integrated_sites.objects.values_list("Site",flat=True))
        print("site in DB:",intg_sites)
        common_site_set=set(intg_sites) & set(ext_site_list)
        print("Common sites: ",common_site_set)
        not_available_sites=list(set(ext_site_list) - common_site_set)
        not_available_sites1=[str(X) for X in not_available_sites]
        
        print("not available sites :",not_available_sites1)
        site_list=list(common_site_set)
        
        site_list=exl_site_list
             
        
      ################################ from_date and to_date ######################
        from_date = request.POST.get("from_date")
        print("from_date: ",from_date)
        to_date = request.POST.get("to_date")
        print("to_date: ", to_date)
        
        to_date_obj=datetime.datetime.strptime(to_date,"%Y-%m-%d").date()
        print("to date obj:",to_date_obj)
        

       ############################ Offered Date calculation ##################################
        offered_date_obj=to_date_obj+ timedelta(1)
        print("offered date:",offered_date_obj)

        offered_date=offered_date_obj.strftime("%Y-%m-%d")
        print("str offered date:",offered_date)
      


      ############################ Selecting raw kpi from the DB ###############################
        try:
            raw_kpi_addr= raw_kpis.objects.get(upload_date=offered_date).file.url
            print(raw_kpi_addr)
        except Exception as e:
            print("error:",e)
            message="raw Kpi file of " + str(offered_date) + " is not available in the backend server."
            return Response({"status":False,"message":message})
        ip=ALLOWED_HOSTS[0]
        print("################### ip ############################",ip)
        addr= "http://"+ str(ip)+ ":8001" + str(raw_kpi_addr)
        print(addr) 
        df_raw_kpi= pd.read_excel(addr)
        print("############################ raw kpi file #########################################")
        print(df_raw_kpi)
       

        ##############################        ############################################


        door_root= os.path.join(MEDIA_ROOT,"Original_trend","Tnch","Trend")

        path_of_blnk_temp=os.path.join(door_root,"template","template2.xlsx")
        trend_wb=load_workbook(path_of_blnk_temp)
        print(trend_wb.sheetnames)
        print("##################################################################################")


        g2_site=[]
        f8_site=[]
        f3_site=[]
        t1t2_site=[]


        kpi=[   
              "MV_RRC Setup Success Rate [CDBH]",
                "MV_ERAB Setup Success Rate [CDBH]",
                "MV_PS Drop Call Rate % [CDBH]",
                "MV_DL User Throughput_Mbps [CDBH]",
                "MV_UL User Throughput_Mbps [CDBH]",
                "MV_PS handover success rate [LTE Intra System] [CDBH]",
                "MV_PS handover success rate [LTE Inter System] [CDBH]", 
                "MV_CSFB Redirection Success Rate [CDBH]",
                "MV_E-UTRAN Average CQI [CDBH]",
                "UL RSSI PUCCH",
                "MV_VoLTE ERAB Setup Success Rate [CBBH]",
                "MV_VoLTE DCR [CBBH]",
                "MV_VoLTE Packet Loss DL [CBBH]",
                "MV_VoLTE Packet Loss UL [CBBH]",
                "MV_VoLTE IntraF HOSR Exec [CBBH]",
                "MV_4G Data Volume_GB",
                "MV_VoLTE InterF HOSR Exec [CBBH]", 
                "MV_Average number of used DL PRBs [CDBH]",
                "MV_VoLTE Traffic",
            ]
        


       
      

        df_raw_kpi["Short name"].fillna( inplace=True, method="ffill")
        df_raw_kpi["Short name"] =df_raw_kpi["Short name"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..


        df_raw_kpi.rename( columns={'Unnamed: 1':'date'}, inplace=True )

        df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"]/1024)
        df_raw_kpi.rename(columns={"MV_DL User Throughput_Kbps [CDBH]" :"MV_DL User Throughput_Mbps [CDBH]" } ,inplace = True )

        df_raw_kpi["MV_UL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_UL User Throughput_Kbps [CDBH]"]/1024)
        df_raw_kpi.rename(columns={"MV_UL User Throughput_Kbps [CDBH]" :"MV_UL User Throughput_Mbps [CDBH]" } ,inplace = True )

        print(df_raw_kpi)
        print(df_raw_kpi.columns)



        lis=list(df_raw_kpi["Short name"])
        sit_id_lis=[]
        cell_id_lis=[]
        for item in lis:
            if("_" in item):
                cell_id=item.split("_")[-2]
                ln=len(item.split("_")[-1])
                #print(ln)
                sit_id=item.split("_")[-2][:-ln]
            else:
                cell_id=item
                sit_id=item
            cell_id_lis.append(cell_id)
            sit_id_lis.append(sit_id)

        print("Site_ID's: ",sit_id_lis)
        print("Cell ID's: ",cell_id_lis)
        
        df_raw_kpi.insert(1, "SITE_ID", sit_id_lis)
        df_raw_kpi.insert(2, "CELL_ID", cell_id_lis)

        raw_site_list=list(set(sit_id_lis))
        not_added_in_mycom_site=list(set(site_list)-set(raw_site_list))
        print("not_added_in_mycom:",not_added_in_mycom_site)

        if request.GET.get("check") == "True":
            
            return Response({"status":True,"integrated_sites":site_list,"not_avalilable_sites":not_available_sites1,"not_reflectig_in_MyCom_data":not_added_in_mycom_site}) 
        df_raw_kpi.rename(columns={"Short name" :"Shortname" } ,inplace = True )
        df_raw_kpi.fillna(value=0,inplace=True)

        process_op_path=os.path.join(door_root,"process_outputs")
        
        savepath=os.path.join(process_op_path,"desired input.xlsx")
        df_raw_kpi.to_excel(savepath)
    
        # date1=date(2023,2,27)
        date1=offered_date_obj
        # date1=date.today()
        dt1 = date1 - timedelta(1)
        dt2 = date1 - timedelta(2)
        dt3 = date1 - timedelta(3)
        dt4 = date1 - timedelta(4)
        dt5 = date1 - timedelta(5)
        ls=[dt1,dt2,dt3,dt4,dt5]
        only_site_fil = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.date.isin(ls))]
        
        savepath=os.path.join(process_op_path,"only_site_date_filtered_input.xlsx")
        only_site_fil.to_excel(savepath)
        
        def perticular_tech( tech,site_list):
            # df_filtered = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.date.isin(ls))]
            df_filtered = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.date.isin(ls)) & (df_raw_kpi.Shortname.str.contains('|'.join(tech)))]
            
            print(df_filtered)
            if not df_filtered.empty:
                address="last_filtered_input"  +str(tech)+ ".xlsx"
                savepath=os.path.join(process_op_path,address)
                df_filtered.to_excel(savepath)
                df_pivoted = df_filtered.pivot_table(index=["SITE_ID","Shortname","CELL_ID"], columns="date")
                print("technology:",tech)
                print(df_pivoted)
                address_pivot="pivoted_input"  +str(tech)+ ".xlsx"
                savepath=os.path.join(process_op_path,address_pivot)
                df_pivoted.to_excel(savepath)
                return df_pivoted
            
            return df_filtered


      

        alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        def num_hash(num):
                if num < 26:
                    return alpha[num-1]
                else:
                    q, r = num//26, num % 26
                    if r == 0:
                        if q == 1:
                            return alpha[r-1]
                        else:
                            return num_hash(q-1) + alpha[r-1]
                    else:
                        return num_hash(q) + alpha[r-1]
    
    # Driver code

    # printString(27906)

        def titleToNumber(s):
        # This process is similar to binary-to-
        # decimal conversion
            result = 0
            for B in range(len(s)):
                result *= 26
                result += ord(s[B]) - ord('A') + 1
            return result


        # def overwrite(df_pivoted,kpi_name,coln1,trend_ws):
        #     coln2=num_hash(titleToNumber(coln1)+1)
        #     coln3=num_hash(titleToNumber(coln1)+2)
        #     coln4=num_hash(titleToNumber(coln1)+3)
        #     coln5=num_hash(titleToNumber(coln1)+4)
        #     print(kpi_name)
        #     index_pivot=df_pivoted.index.to_list()
        #     print("index ;###############################",index_pivot)
        #     print(len(index_pivot))
        #     print("index of pivoted table: ",index_pivot)
        #     dr=df_pivoted[kpi_name]
        #     print("columns of dr table",dr.columns)
        #     cl=dr.columns.to_list()
        #     print("column list",cl)
            
        #     # site_id=dr["SITE_ID"].to_list() 
        #     # cell_id=dr["CELL_ID"].to_list()    
        #     col1=dr[str(cl[0])].to_list()
        #     col2=dr[str(cl[1])].to_list()
        #     col3=dr[str(cl[2])].to_list()
        #     col4=dr[str(cl[3])].to_list()
        #     col5=dr[str(cl[4])].to_list()

        #     trend_ws[coln1+"3"].value=cl[0]
        #     trend_ws[coln2+"3"].value=cl[1]
        #     trend_ws[coln3+"3"].value=cl[2]
        #     trend_ws[coln4+"3"].value=cl[3]
        #     trend_ws[coln5+"3"].value=cl[4]

        #     # me=column_index_from_string(coln5)+1
        #     # me=get_column_letter(me)
        #     for i,value in enumerate(index_pivot):
        #         j=i+4
        #         trend_ws["A"+str(j)].value=index_pivot[i][1]
        #         trend_ws["E"+str(j)].value=index_pivot[i][0] 
        #         # trend_ws["C"+str(j)].value=index_pivot[i][2]
        #         # trend_ws["K"+str(j)].value=date1
                
                
        #         trend_ws[coln1+str(j)].value=col1[i]
        #         trend_ws[coln2+str(j)].value=col2[i]
        #         trend_ws[coln3+str(j)].value=col3[i]
        #         trend_ws[coln4+str(j)].value=col4[i]
        #         trend_ws[coln5+str(j)].value=col5[i]
        
        
        def overwrite(df_pivoted,kpi_name,coln1,trend_ws):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            print(kpi_name)
            index_pivot=df_pivoted.index.to_list()
            print("index ;###############################",index_pivot)
            print(len(index_pivot))
            print("index of pivoted table: ",index_pivot)
            dr=df_pivoted[kpi_name]
            print("columns of dr table",dr.columns)
            cl=dr.columns.to_list()
            print("column list",cl)
            
            # site_id=dr["SITE_ID"].to_list() 
            # cell_id=dr["CELL_ID"].to_list()    
            col1=dr[str(cl[0])].to_list()
            col2=dr[str(cl[1])].to_list()
            col3=dr[str(cl[2])].to_list()
            col4=dr[str(cl[3])].to_list()
            col5=dr[str(cl[4])].to_list()

            trend_ws[coln1+"3"].value=cl[0]
            trend_ws[coln2+"3"].value=cl[1]
            trend_ws[coln3+"3"].value=cl[2]
            trend_ws[coln4+"3"].value=cl[3]
            trend_ws[coln5+"3"].value=cl[4]

            # me=column_index_from_string(coln5)+1
            # me=get_column_letter(me)
            for i,value in enumerate(index_pivot):
                j=i+4
                trend_ws["A"+str(j)].value=index_pivot[i][1]
                trend_ws["E"+str(j)].value=index_pivot[i][0] 
                # trend_ws["C"+str(j)].value=index_pivot[i][2]
                # trend_ws["K"+str(j)].value=date1
                
                
                trend_ws[coln1+str(j)].value=col1[i]
                trend_ws[coln2+str(j)].value=col2[i]
                trend_ws[coln3+str(j)].value=col3[i]
                trend_ws[coln4+str(j)].value=col4[i]
                trend_ws[coln5+str(j)].value=col5[i]
                
              



        # for fdd
        pivot_fdd=perticular_tech(["_F3_"],site_list)
        trend_ws=trend_wb["L1800"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="MV_ERAB Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="MV_PS Drop Call Rate % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)

                if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)

                if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)

                if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="MV_PS handover success rate [LTE Inter System] [CDBH]"): 
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="MV_VoLTE ERAB Setup Success Rate [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)
                
                if(kpi_name=="MV_VoLTE IntraF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)

                if(kpi_name=="MV_4G Data Volume_GB"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)

                if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)
                if(kpi_name=="MV_VoLTE Traffic"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)
            
        

        # for tdd
        pivot_fdd=perticular_tech(["_T1_","_T2_"],site_list)
        trend_ws=trend_wb["L2300"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="MV_ERAB Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="MV_PS Drop Call Rate % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)

                if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)

                if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)

                if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="MV_PS handover success rate [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="MV_VoLTE ERAB Setup Success Rate [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)
                
                if(kpi_name=="MV_VoLTE IntraF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)

                if(kpi_name=="MV_4G Data Volume_GB"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)

                if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)
                if(kpi_name=="MV_VoLTE Traffic"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)
            

        pivot_fdd=perticular_tech(["_F8_"],site_list)
        trend_ws=trend_wb["L900"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="MV_ERAB Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="MV_PS Drop Call Rate % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)

                if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)

                if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)

                if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="MV_PS handover success rate [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="MV_VoLTE ERAB Setup Success Rate [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)
                
                if(kpi_name=="MV_VoLTE IntraF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)

                if(kpi_name=="MV_4G Data Volume_GB"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)

                if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)
                if(kpi_name=="MV_VoLTE Traffic"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)




        pivot_fdd=perticular_tech(["_F1_"],site_list)
        trend_ws=trend_wb["L2100"]
        if not pivot_fdd.empty:
            for kpi_name in kpi:
                if(kpi_name=="MV_RRC Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AJ",trend_ws)

                if(kpi_name=="MV_ERAB Setup Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AQ",trend_ws)

                if(kpi_name=="MV_PS Drop Call Rate % [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"AX",trend_ws)

                if(kpi_name=="MV_DL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BE",trend_ws)

                if(kpi_name=="MV_UL User Throughput_Mbps [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BL",trend_ws)

                if(kpi_name=="MV_PS handover success rate [LTE Intra System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BS",trend_ws)
                
                if(kpi_name=="MV_PS handover success rate [LTE Inter System] [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"BZ",trend_ws)
                
                if(kpi_name=="MV_CSFB Redirection Success Rate [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CG",trend_ws)
                
                if(kpi_name=="MV_E-UTRAN Average CQI [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"CN",trend_ws)
                
                if(kpi_name=="UL RSSI PUCCH"):
                    overwrite(pivot_fdd,kpi_name,"CU",trend_ws)
                
                if(kpi_name=="MV_VoLTE ERAB Setup Success Rate [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DB",trend_ws)
                
                if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DI",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss DL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DP",trend_ws)
                
                if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"DW",trend_ws)
                
                if(kpi_name=="MV_VoLTE IntraF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ED",trend_ws)

                if(kpi_name=="MV_4G Data Volume_GB"):
                    overwrite(pivot_fdd,kpi_name,"EK",trend_ws)

                if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                    overwrite(pivot_fdd,kpi_name,"ER",trend_ws)

                if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                    overwrite(pivot_fdd,kpi_name,"EY",trend_ws)
                if(kpi_name=="MV_VoLTE Traffic"):
                    overwrite(pivot_fdd,kpi_name,"FF",trend_ws)



       


        
        output_path=os.path.join(door_root,"output","Tnch_trend_output.xlsx")
        print(output_path)
        trend_wb.save(output_path)
        # trend_wb.save("opopopop.xlsx")
       
        download_path=os.path.join(MEDIA_URL,"Original_trend","Tnch","Trend","output","Tnch_trend_output.xlsx")
        return Response({"status":True,"message":"Succesfully uploaded","Download_url":download_path})



@api_view(["POST"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def tnch_pre_post_upload_7(request):
  
    # pre_trend = request.FILES["pre_trend"] if 'pre_trend' in request.FILES else None
            post_file = request.FILES["post_trend"] if 'post_trend' in request.FILES else None
            pre_file = request.FILES["pre_trend"] if 'pre_trend' in request.FILES else None
            mapping_file = request.FILES["mapping"] if 'mapping' in request.FILES else None

            ######################################## for mappings ###################################################
            location= MEDIA_ROOT + r"\dpr\dpr_report_file"
            fs = FileSystemStorage(location=location)
            mapping_file = fs.save(mapping_file.name, mapping_file)
            # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
            filepath = fs.path(mapping_file)
            print("file_path:-",filepath)
            df_mapping=pd.read_excel(filepath,header=1)
            os.remove(path=filepath)
            print(filepath,"deleted...............")
            print(df_mapping)
        #     df_mapping.to_excel("mappings.xlsx")
            if not(df_mapping.empty):
                   OldSiteMapping_dic={}
                   NewSiteMapping_dic={}
                   for i,d in df_mapping.iterrows():
                          if "Old Site ID" in df_mapping.columns and  "New Site ID" in df_mapping.columns:
                                         OldSiteMapping_dic[d["Old Site ID"]]=[d["New Site ID"],d["Relocation Date"]]
                                         NewSiteMapping_dic[d["New Site ID"]]=d["Today Date"]
                                         
                          else:
                                context={"status":False,"message":"Old Site ID/New Site ID is missing or name is not correct"}
                                return Response(context)

                          
            else:
                  context={"status":False,"message":"mapping file is empty"}
                  return Response(context)       
            print(OldSiteMapping_dic)   
               

            ############################################# for Post ################################################
            location= os.path.join(MEDIA_ROOT,"Original_trend")
            fs = FileSystemStorage(location=location)
            post_file = fs.save(post_file.name, post_file)
            # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
            filepath = fs.path(post_file)
            print("file_path:-",filepath)

            df_raw_kpi=pd.read_excel(filepath)
            
            os.remove(path=filepath)
            print(filepath,"deleted...............")
            
                                            #__________ Preprocessing of the Raw KPI _____________#

            df_raw_kpi["Short name"].fillna( inplace=True, method="ffill")
            df_raw_kpi["Short name"] =df_raw_kpi["Short name"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..
    
    
            df_raw_kpi.rename( columns={'Unnamed: 1':'date'}, inplace=True )
    
            df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"]/1024)
            df_raw_kpi.rename(columns={"MV_DL User Throughput_Kbps [CDBH]" :"MV_DL User Throughput_Mbps [CDBH]" } ,inplace = True )
    
            df_raw_kpi["MV_UL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_UL User Throughput_Kbps [CDBH]"]/1024)
            df_raw_kpi.rename(columns={"MV_UL User Throughput_Kbps [CDBH]" :"MV_UL User Throughput_Mbps [CDBH]" } ,inplace = True )
    
            print(df_raw_kpi)
            print(df_raw_kpi.columns)
    
    
    
            lis=list(df_raw_kpi["Short name"])
            sit_id_lis=[]
            cell_id_lis=[]
            for item in lis:
                    if("_" in item):
                        cell_id=item.split("_")[-2]
                        ln=len(item.split("_")[-1])
                        #print(ln)
                        sit_id=item.split("_")[-2][:-ln]
                    else:
                        cell_id=item
                        sit_id=item
                    cell_id_lis.append(cell_id)
                    sit_id_lis.append(sit_id)
    
            print(sit_id)
            print(cell_id_lis)
    
            df_raw_kpi.insert(1, "SITE_ID", sit_id_lis)
            df_raw_kpi.insert(2, "CELL_ID", cell_id_lis)
    
            
            
            df_raw_kpi.rename(columns={"Short name" :"Shortname" } ,inplace = True)
            df_raw_kpi.fillna(value=0,inplace=True)

            # date1=datetime.datetime.today().date
            # date1=date.today()
            # print("<<<##############################################>>>",date1,"<<<###################################################>>>")
            date1=date(2023,1,31)
            dt1 = date1 - timedelta(1)
            dt2 = date1 - timedelta(2)
            dt3 = date1 - timedelta(3)
            dt4 = date1 - timedelta(4)
            dt5 = date1 - timedelta(5)
            dt6 = date1-  timedelta(6)
            dt7 = date1-  timedelta(7)
            print()
            ls=[dt1,dt2,dt3,dt4,dt5,dt6,dt7]
            print("Dates: ", ls)
            df_DateFiltered = df_raw_kpi[(df_raw_kpi.date.isin(ls))]
            print(df_DateFiltered)
            df_pivoted = df_DateFiltered.pivot_table(index=["Shortname","SITE_ID","CELL_ID"], columns="date",values=["MV_4G Data Volume_GB","MV_VoLTE Traffic"])
            print(df_pivoted)
            df_pivoted.to_excel("tst.xlsx")
            df=df_pivoted
        #   sys.exit("exiting...................................................")
           
                                             #_____________*************************________________# 
            print(df)
            if not(df.empty):
                dates=df["MV_4G Data Volume_GB"].columns
                print(dates)
                # print(type(dates[1]))
                date1=dates[0].day
                date2=dates[1].day
                date3=dates[2].day
                date4=dates[3].day
                date5=dates[4].day
                date6=dates[5].day
                date7=dates[6].day
                print(date1)
                print(date2)
                print(date3)
                print(date4)
                print(date5)
                print(date6)
                print(date7)
                
                # df.to_excel("post_trend_excel.xlsx") # just to visualise the data frame

               
                df.columns =df.columns.to_flat_index() # flattening the multi index header to single index
                df.to_excel("flattened_post_data_volume.xlsx", index=True)
                df.reset_index(inplace=True)
                df.to_excel("Index_Reset_flattened_post_data_volume.xlsx", index=True)
                
                
                print("############################## df post data volume ##############################")
         
               
                # df_Post_volte_traffic=df[["Unnamed: 0","PS RRC Succ Rate","Unnamed: 36","Unnamed: 37","Unnamed: 38","Unnamed: 39","Unnamed: 1","CRITERIA"]]
                # df_Post_volte_traffic.columns=df_Post_volte_traffic.iloc[1]
                # df_Post_volte_traffic = df_Post_volte_traffic.iloc[2:].reset_index(drop=True)
                # df_Post_volte_traffic.to_excel("actual_post_volte_traffic.xlsx", index=False)
                # print("################################ df post volte traffic #########################")
                # print(df_Post_volte_traffic)

                
             
             
                print(df.index)
                for d in df.values:
                        if(not (pd.isnull(d[0]))):
                            obj,created = pre_post_report2.objects.get_or_create(Post_cell_name=d[0])
                        
                            obj.Post_cell_site_id=d[1]
                            print(d)
                            obj.Today_date = NewSiteMapping_dic[d[1]]
                            
                        

                            obj.Post_Volte_Traffic_Day1=round(d[10],2)
                            obj.Post_Volte_Traffic_Day2=round(d[11],2)
                            obj.Post_Volte_Traffic_Day3=round(d[12],2)
                            obj.Post_Volte_Traffic_Day4=round(d[13],2)
                            obj.Post_Volte_Traffic_Day5=round(d[14],2)
                            obj.Post_Volte_Traffic_Day6=round(d[15],2)
                            obj.Post_Volte_Traffic_Day7=round(d[16],2)

                            obj.Post_Data_Volume_Day1=round(d[3],2)
                            obj.Post_Data_Volume_Day2=round(d[4],2)
                            obj.Post_Data_Volume_Day3=round(d[5],2)
                            obj.Post_Data_Volume_Day4=round(d[6],2)
                            obj.Post_Data_Volume_Day5=round(d[7],2)
                            obj.Post_Data_Volume_Day6=round(d[8],2)
                            obj.Post_Data_Volume_Day7=round(d[9],2)

                            lis_DV=[d[3],d[4],d[5],d[6],d[7],d[8],d[9]]
                            lis_VT=[d[10],d[11],d[12],d[13],d[14],d[15],d[16]]
                            DV_avg=round(mean(lis_DV),2)
                            VT_avg=round(mean(lis_VT),2)
                            obj.Post_Data_Volume_AVG=DV_avg
                            obj.Post_Volte_Traffic_AVG=VT_avg
 
                            obj.save()
            else:
                  context={"status":False,"message":"Post_trend is empty"}
                  return Response(context)
            ######################################### for pre #######################################               
            location= os.path.join(MEDIA_ROOT,"Original_trend")
            fs = FileSystemStorage(location=location)
            pre_file = fs.save(pre_file.name, pre_file)
            # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
            filepath = fs.path(pre_file)
            print("file_path:-",filepath)
            df_raw_kpi=pd.read_excel(filepath)
            
            os.remove(path=filepath)
            print(filepath,"deleted...............")
            
                                            #__________ Preprocessing of the Raw KPI _____________#

            df_raw_kpi["Short name"].fillna( inplace=True, method="ffill")
            df_raw_kpi["Short name"] =df_raw_kpi["Short name"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..
    
    
            df_raw_kpi.rename( columns={'Unnamed: 1':'date'}, inplace=True )
    
            df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_DL User Throughput_Kbps [CDBH]"]/1024)
            df_raw_kpi.rename(columns={"MV_DL User Throughput_Kbps [CDBH]" :"MV_DL User Throughput_Mbps [CDBH]" } ,inplace = True )
    
            df_raw_kpi["MV_UL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["MV_UL User Throughput_Kbps [CDBH]"]/1024)
            df_raw_kpi.rename(columns={"MV_UL User Throughput_Kbps [CDBH]" :"MV_UL User Throughput_Mbps [CDBH]" } ,inplace = True )
    
            print(df_raw_kpi)
            print(df_raw_kpi.columns)
    
    #pre
    
            lis=list(df_raw_kpi["Short name"])
            sit_id_lis=[]
            cell_id_lis=[]
            for item in lis:
                    if("_" in item):
                        cell_id=item.split("_")[-2]
                        ln=len(item.split("_")[-1])
                        #print(ln)
                        sit_id=item.split("_")[-2][:-ln]
                    else:
                        cell_id=item
                        sit_id=item
                    cell_id_lis.append(cell_id)
                    sit_id_lis.append(sit_id)
    
            print(sit_id)
            print(cell_id_lis)
    
            df_raw_kpi.insert(1, "SITE_ID", sit_id_lis)
            df_raw_kpi.insert(2, "CELL_ID", cell_id_lis)
    

            df_raw_kpi.rename(columns={"Short name" :"Shortname" } ,inplace = True )
            df_raw_kpi.fillna(value=0,inplace=True)
            df_raw_kpi
            # date1=datetime.datetime.today()
            print("#########################################",sit_id_lis[0],"###################################")
            date1 = OldSiteMapping_dic[sit_id_lis[0]][1]
            
            # date1=date(2023,1,24)
            dt1 = date1 - timedelta(1)
            dt2 = date1 - timedelta(2)
            dt3 = date1 - timedelta(3)
            dt4 = date1 - timedelta(4)
            dt5 = date1 - timedelta(5)
            dt6 = date1-  timedelta(6)
            dt7 = date1-  timedelta(7)
            ls=[dt1,dt2,dt3,dt4,dt5,dt6,dt7]
            print("Dates: ", ls)
            df_DateFiltered = df_raw_kpi[(df_raw_kpi.date.isin(ls))]
            print(df_DateFiltered)
            df_pivoted = df_DateFiltered.pivot_table(index=["Shortname","SITE_ID","CELL_ID"], columns="date",values=["MV_4G Data Volume_GB","MV_VoLTE Traffic"])
            print(df_pivoted)
            df_pivoted.to_excel("tst.xlsx")
            df=df_pivoted
        #   sys.exit("exiting...................................................")
           
                                             #_____________*************************________________# 
            
          
            
           
            print(df)
            
            if not(df.empty):
                dates=df["MV_4G Data Volume_GB"].columns
                print(dates)
                # print(type(dates[1]))
                print(datetime.datetime.now())
                date1=dates[0]
                date2=dates[1]
                date3=dates[2]
                date4=dates[3]
                date5=dates[4]
                print(date1)
                print(date2)
                print(date3)
                print(date4)
                print(date5)
                
                # df.to_excel("post_trend_excel.xlsx") # just to visualise the data frame

               
                df.columns =df.columns.to_flat_index() # flattening the multi index header to single index
                df.to_excel("flattened_pre_data_volume.xlsx", index=False)
                df.reset_index(inplace=True)
                df.to_excel("Index_Reset_flattened_pre_data_volume.xlsx", index=True)
                
                print("############################## df post data volume ##############################")
        
                print("Finding Exact Match....")
                print("df_col_length : ",len(df.columns.to_list()))
                for d in df.values:
                        site=d[1]
                        Pre_cell_name=d[0]
                        pre_trend_cell=d[2]
                        try:
                                objs=pre_post_report2.objects.filter(Post_cell_site_id=OldSiteMapping_dic[site][0])
                        except:
                                continue
                        ins=False
                        
                        for obj1 in objs:
                            
                              if obj1.Post_cell_name.split("_")[-1]==Pre_cell_name.split("_")[-1] and obj1.Post_cell_name.split("_")[2] == Pre_cell_name.split("_")[2] :
                                     obj=obj1
                                     ins=True
                                     print("pre-",Pre_cell_name," ","Post-",obj.Post_cell_name)
                                     break
                                  
                               
                        if ins:
                            obj.Pre_cell_name=Pre_cell_name
                            
                            obj.Pre_cell_site_id=site
                            obj.Relocation_date = OldSiteMapping_dic[site][1]
                           
                            obj.Pre_Volte_Traffic_Day1=round(d[10],2)
                            obj.Pre_Volte_Traffic_Day2=round(d[11],2)
                            obj.Pre_Volte_Traffic_Day3=round(d[12],2)
                            obj.Pre_Volte_Traffic_Day4=round(d[13],2)
                            obj.Pre_Volte_Traffic_Day5=round(d[14],2)
                            obj.Pre_Volte_Traffic_Day6=round(d[15],2)
                            obj.Pre_Volte_Traffic_Day7=round(d[16],2)

                            obj.Pre_Data_Volume_Day1=round(d[3],2)
                            obj.Pre_Data_Volume_Day2=round(d[4],2)
                            obj.Pre_Data_Volume_Day3=round(d[5],2)
                            obj.Pre_Data_Volume_Day4=round(d[6],2)
                            obj.Pre_Data_Volume_Day5=round(d[7],2)
                            obj.Pre_Data_Volume_Day6=round(d[8],2)
                            obj.Pre_Data_Volume_Day7=round(d[9],2)

                            lis_DV=[d[3],d[4],d[5],d[6],d[7],d[8],d[9]]
                            lis_VT=[d[10],d[11],d[12],d[13],d[14],d[15],d[16]]
                            DV_avg=round(mean(lis_DV),2)
                            VT_avg=round(mean(lis_VT),2)
                            obj.Pre_Data_Volume_AVG=DV_avg
                            obj.Pre_Volte_Traffic_AVG=VT_avg      


                            obj.save()
                              
            else:
                  context={"status":False,"message":"DPR Report is empty"}
                  return Response(context)


           
           
            

           ############################################ to fill average Change ##############################################
            objs=pre_post_report2.objects.all()
            for obj in objs:
                   pre_DV_avg=obj.Pre_Data_Volume_AVG
                   post_DV_avg=obj.Post_Data_Volume_AVG
                   
                   if pre_DV_avg !=0:
                        obj.Percentage_change_Data_Volume=round(((post_DV_avg-pre_DV_avg)/pre_DV_avg)*100,2)
                   else:
                        obj.Percentage_change_Data_Volume=0
                          
                   pre_VT_avg=obj.Pre_Volte_Traffic_AVG
                   post_VT_avg=obj.Post_Volte_Traffic_AVG

                   if pre_VT_avg !=0:     
                        obj.Percentage_change_Volte_Traffic=round(((post_VT_avg-pre_VT_avg)/pre_VT_avg)*100,2)
                   else:
                          obj.Percentage_change_Volte_Traffic=0
                   obj.save()
           
            
        #     dr=pd.DataFrame.from_records(pre_post_report2.objects.all().values())
        ################################ sitewise Report making ###################################################
            unique_site_list=list(pre_post_report2.objects.values_list("Post_cell_site_id",flat=True).distinct())
            print(unique_site_list)

            for site in unique_site_list:
                   SiteWise_obj,sts=pre_post_report_siteWise.objects.get_or_create(Post_cell_site_id=site)
                   obj_set=pre_post_report2.objects.filter(Post_cell_site_id=site)

                   SiteWise_obj.Today_date=obj_set[0].Today_date

                   SiteWise_obj.Relocation_date=obj_set[0].Relocation_date

                   SiteWise_obj.Pre_cell_site_id=obj_set[0].Pre_cell_site_id


                   Pre_VT_Day1=SiteWise_obj.Pre_Volte_Traffic_Day1=round(obj_set.aggregate(ar=Sum("Pre_Volte_Traffic_Day1"))["ar"],2)
                   Pre_VT_Day2=SiteWise_obj.Pre_Volte_Traffic_Day2=round(obj_set.aggregate(ar=Sum("Pre_Volte_Traffic_Day2"))["ar"],2)
                   Pre_VT_Day3=SiteWise_obj.Pre_Volte_Traffic_Day3=round(obj_set.aggregate(ar=Sum("Pre_Volte_Traffic_Day3"))["ar"],2)
                   Pre_VT_Day4=SiteWise_obj.Pre_Volte_Traffic_Day4=round(obj_set.aggregate(ar=Sum("Pre_Volte_Traffic_Day4"))["ar"],2)
                   Pre_VT_Day5=SiteWise_obj.Pre_Volte_Traffic_Day5=round(obj_set.aggregate(ar=Sum("Pre_Volte_Traffic_Day5"))["ar"],2)
                   Pre_VT_Day6=SiteWise_obj.Pre_Volte_Traffic_Day6=round(obj_set.aggregate(ar=Sum("Pre_Volte_Traffic_Day6"))["ar"],2)
                   Pre_VT_Day7=SiteWise_obj.Pre_Volte_Traffic_Day7=round(obj_set.aggregate(ar=Sum("Pre_Volte_Traffic_Day7"))["ar"],2)

                   Post_VT_Day1=SiteWise_obj.Post_Volte_Traffic_Day1=round(obj_set.aggregate(ar=Sum("Post_Volte_Traffic_Day1"))["ar"],2)
                   Post_VT_Day2=SiteWise_obj.Post_Volte_Traffic_Day2=round(obj_set.aggregate(ar=Sum("Post_Volte_Traffic_Day2"))["ar"],2)
                   Post_VT_Day3=SiteWise_obj.Post_Volte_Traffic_Day3=round(obj_set.aggregate(ar=Sum("Post_Volte_Traffic_Day3"))["ar"],2)
                   Post_VT_Day4=SiteWise_obj.Post_Volte_Traffic_Day4=round(obj_set.aggregate(ar=Sum("Post_Volte_Traffic_Day4"))["ar"],2)
                   Post_VT_Day5=SiteWise_obj.Post_Volte_Traffic_Day5=round(obj_set.aggregate(ar=Sum("Post_Volte_Traffic_Day5"))["ar"],2)
                   Post_VT_Day6=SiteWise_obj.Post_Volte_Traffic_Day6=round(obj_set.aggregate(ar=Sum("Post_Volte_Traffic_Day6"))["ar"],2)
                   Post_VT_Day7=SiteWise_obj.Post_Volte_Traffic_Day7=round(obj_set.aggregate(ar=Sum("Post_Volte_Traffic_Day7"))["ar"],2)

                   Pre_DV_Day1=SiteWise_obj.Pre_Data_Volume_Day1=round(obj_set.aggregate(ar=Sum("Pre_Data_Volume_Day1"))["ar"],2)
                   Pre_DV_Day2=SiteWise_obj.Pre_Data_Volume_Day2=round(obj_set.aggregate(ar=Sum("Pre_Data_Volume_Day2"))["ar"],2)
                   Pre_DV_Day3=SiteWise_obj.Pre_Data_Volume_Day3=round(obj_set.aggregate(ar=Sum("Pre_Data_Volume_Day3"))["ar"],2)
                   Pre_DV_Day4=SiteWise_obj.Pre_Data_Volume_Day4=round(obj_set.aggregate(ar=Sum("Pre_Data_Volume_Day4"))["ar"],2)
                   Pre_DV_Day5=SiteWise_obj.Pre_Data_Volume_Day5=round(obj_set.aggregate(ar=Sum("Pre_Data_Volume_Day5"))["ar"],2)
                   Pre_DV_Day6=SiteWise_obj.Pre_Data_Volume_Day6=round(obj_set.aggregate(ar=Sum("Pre_Data_Volume_Day6"))["ar"],2)
                   Pre_DV_Day7=SiteWise_obj.Pre_Data_Volume_Day7=round(obj_set.aggregate(ar=Sum("Pre_Data_Volume_Day7"))["ar"],2)

                   Post_DV_Day1=SiteWise_obj.Post_Data_Volume_Day1=round(obj_set.aggregate(ar=Sum("Post_Data_Volume_Day1"))["ar"],2)
                   Post_DV_Day2=SiteWise_obj.Post_Data_Volume_Day2=round(obj_set.aggregate(ar=Sum("Post_Data_Volume_Day2"))["ar"],2)
                   Post_DV_Day3=SiteWise_obj.Post_Data_Volume_Day3=round(obj_set.aggregate(ar=Sum("Post_Data_Volume_Day3"))["ar"],2)
                   Post_DV_Day4=SiteWise_obj.Post_Data_Volume_Day4=round(obj_set.aggregate(ar=Sum("Post_Data_Volume_Day4"))["ar"],2)
                   Post_DV_Day5=SiteWise_obj.Post_Data_Volume_Day5=round(obj_set.aggregate(ar=Sum("Post_Data_Volume_Day5"))["ar"],2)
                   Post_DV_Day6=SiteWise_obj.Post_Data_Volume_Day6=round(obj_set.aggregate(ar=Sum("Post_Data_Volume_Day6"))["ar"],2)
                   Post_DV_Day7=SiteWise_obj.Post_Data_Volume_Day7=round(obj_set.aggregate(ar=Sum("Post_Data_Volume_Day7"))["ar"],2)

                   Pre_VT_AVG= round((
                                Pre_VT_Day1
                              + Pre_VT_Day2 
                              + Pre_VT_Day3 
                              + Pre_VT_Day4 
                              + Pre_VT_Day5 
                              + Pre_VT_Day6 
                              + Pre_VT_Day7
                              )/7 ,2)
                  
                   Post_VT_AVG= round((Post_VT_Day1
                               + Post_VT_Day2 
                               + Post_VT_Day3 
                               + Post_VT_Day4 
                               + Post_VT_Day5 
                               + Post_VT_Day6 
                               + Post_VT_Day7
                               )/7 ,2)

                   Pre_DV_AVG= round((
                                Pre_DV_Day1
                              + Pre_DV_Day2 
                              + Pre_DV_Day3 
                              + Pre_DV_Day4 
                              + Pre_DV_Day5 
                              + Pre_DV_Day6 
                              + Pre_DV_Day7
                              )/7 ,2)
                   Post_DV_AVG=round( (Post_DV_Day1
                               + Post_DV_Day2 
                               + Post_DV_Day3 
                               + Post_DV_Day4 
                               + Post_DV_Day5 
                               + Post_DV_Day6 
                               + Post_DV_Day7
                               )/7 ,2)
                   
                   SiteWise_obj.Pre_Volte_Traffic_AVG=Pre_VT_AVG
                   SiteWise_obj.Post_Volte_Traffic_AVG=Post_VT_AVG

                   SiteWise_obj.Pre_Data_Volume_AVG=Pre_DV_AVG
                   SiteWise_obj.Post_Data_Volume_AVG=Post_DV_AVG

                   if Pre_VT_AVG !=0:    
                        SiteWise_obj.Percentage_change_Volte_Traffic= round(((Post_VT_AVG-Pre_VT_AVG)/Pre_VT_AVG)*100,2)
                   else:
                        SiteWise_obj.Percentage_change_Volte_Traffic=0
                   if Pre_DV_AVG !=0:
                        SiteWise_obj.Percentage_change_Data_Volume= round(((Post_DV_AVG-Pre_DV_AVG)/Pre_DV_AVG)*100,2)
                   else:
                          SiteWise_obj.Percentage_change_Data_Volume=0
                          
                   SiteWise_obj.save()


        ###############################################################################################################


        ########################################### to create the folder in media to save the generated record ################################
       
            directory =  "Original_trend"
            # Parent Directory path
            parent_dir =  MEDIA_ROOT 
           
            # Path
            path = os.path.join(parent_dir, directory)
            print(path)
            #Create the directory
            if(not (os.path.isdir(path))):
                os.makedirs(path)
                print("Directory '% s' created" % directory)
            
            
           
            RecordPath= os.path.join(path ,"record1.xlsx")
   
   
   
   
    ############################################# Creating DF from pre_post_report2 model (Sector wise)##################################################################       
            # we have assigned objs in the upper section code
            df=pd.DataFrame(list(objs.values()))
            df["Percentage_change_Volte_Traffic"]=df["Percentage_change_Volte_Traffic"].astype("str") +" %"
            df["Percentage_change_Data_Volume"]=df["Percentage_change_Data_Volume"].astype("str") +" %"
        #     df.to_excel("report without header.xlsx",index=True)
        #     colour_df=df.copy()
  
            tuples=[
                   ("","","Post_cell_name"),
                   ("","","Pre_cell_name"),
                   ("","","Post_cell_site_id"),
                   ("","","Pre_cell_site_id"),
                  
                   ("","","Relocation Date"),
                   ("","","Today Date"),

                   ("Volte Traffic","Pre_Volte_Traffic","Day1"),
                   ("Volte Traffic","Pre_Volte_Traffic","Day2"),
                   ("Volte Traffic","Pre_Volte_Traffic","Day3"),
                   ("Volte Traffic","Pre_Volte_Traffic","Day4"),
                   ("Volte Traffic","Pre_Volte_Traffic","Day5"),
                   ("Volte Traffic","Pre_Volte_Traffic","Day6"),
                   ("Volte Traffic","Pre_Volte_Traffic","Day7"),

                   ("Volte Traffic","Post_Volte_Traffic","Day1"),
                   ("Volte Traffic","Post_Volte_Traffic","Day2"),
                   ("Volte Traffic","Post_Volte_Traffic","Day3"),
                   ("Volte Traffic","Post_Volte_Traffic","Day4"),
                   ("Volte Traffic","Post_Volte_Traffic","Day5"),
                   ("Volte Traffic","Post_Volte_Traffic","Day6"),
                   ("Volte Traffic","Post_Volte_Traffic","Day7"),

                   ("Blank","",""),
                  

                   ("Data Volume","Pre_Data_Volume","Day1"),
                   ("Data Volume","Pre_Data_Volume","Day2"),
                   ("Data Volume","Pre_Data_Volume","Day3"),
                   ("Data Volume","Pre_Data_Volume","Day4"),
                   ("Data Volume","Pre_Data_Volume","Day5"),
                   ("Data Volume","Pre_Data_Volume","Day6"),
                   ("Data Volume","Pre_Data_Volume","Day7"),

                    ("Data Volume","Post_Data_Volume","Day1"),
                    ("Data Volume","Post_Data_Volume","Day2"),
                    ("Data Volume","Post_Data_Volume","Day3"),
                    ("Data Volume","Post_Data_Volume","Day4"),
                    ("Data Volume","Post_Data_Volume","Day5"),
                    ("Data Volume","Post_Data_Volume","Day6"),
                    ("Data Volume","Post_Data_Volume","Day7"),
                   
                    ("Comparison","","Pre_Volte_Traffic_AVG"),
                    ("Comparison","","Post_Volte_Traffic_AVG"),
                    ("Comparison","","Pre_Data_Volume_AVG"),
                    ("Comparison","","Post_Data_Volume_AVG"),
                    ("Percentage Change","","Volte Traffic"),
                    ("Percentage Change","","Data Volume"),
                   ]
            
            cols = pd.MultiIndex.from_tuples(tuples)
            df.columns=cols
            # df.to_excel("test.xlsx")
#################################################### for making colourgradient of every cell (Sector Wise) ######################################
            df_style=df.style
            idx = pd.IndexSlice
            # colors = ["red", "orange", "yellow","#ffff00", "darkgreen"]
            colors=["#cc3300","#ff9933","#ffcc33","#ffff33","#00cc33"]
           
            cmap1 = mpl.colors.LinearSegmentedColormap.from_list("mycmap",colors)
           
            index=len(df.index)
        #     df=df_style.background_gradient(cmap=cmap1,axis=1)
            for i in range(0,index):
                    slice_VT = idx[idx[i], idx["Volte Traffic",:,:]]
                    slice_DV = idx[idx[i], idx["Data Volume",:,:]]
                    df_style=df_style.background_gradient(cmap=cmap1,axis=1,subset=slice_VT)\
                                     .background_gradient(cmap=cmap1,axis=1,subset=slice_DV) #vmax=mx,vmin=mi
                    df_style.set_properties(**{'border':'1px solid black'})

                


    
###################################################################################################################################
            writer = pd.ExcelWriter(RecordPath, engine='xlsxwriter')
            df_style.to_excel(writer,sheet_name="SectorLevel",index=True)

            # writer.save()
            
############################################# Creating DF from pre_post_report2_SiteWise model (Site wise)##################################################################       
           
            objs=pre_post_report_siteWise.objects.all()
            df=pd.DataFrame(list(objs.values()))
            df["Percentage_change_Volte_Traffic"]=df["Percentage_change_Volte_Traffic"].astype("str") +" %"
            df["Percentage_change_Data_Volume"]=df["Percentage_change_Data_Volume"].astype("str") +" %"
        #     df.to_excel("report without header.xlsx",index=True)
        

            tuples=[
                   ("","ID"),
                   ("","","Post_cell_site_id"),
                   ("","","Pre_cell_site_id"),
                  
                   ("","","Relocation Date"),
                   ("","","Today Date"),

                   ("Volte Traffic","Pre_Volte_Traffic","Day1"),
                   ("Volte Traffic","Pre_Volte_Traffic","Day2"),
                   ("Volte Traffic","Pre_Volte_Traffic","Day3"),
                   ("Volte Traffic","Pre_Volte_Traffic","Day4"),
                   ("Volte Traffic","Pre_Volte_Traffic","Day5"),
                   ("Volte Traffic","Pre_Volte_Traffic","Day6"),
                   ("Volte Traffic","Pre_Volte_Traffic","Day7"),

                   ("Volte Traffic","Post_Volte_Traffic","Day1"),
                   ("Volte Traffic","Post_Volte_Traffic","Day2"),
                   ("Volte Traffic","Post_Volte_Traffic","Day3"),
                   ("Volte Traffic","Post_Volte_Traffic","Day4"),
                   ("Volte Traffic","Post_Volte_Traffic","Day5"),
                   ("Volte Traffic","Post_Volte_Traffic","Day6"),
                   ("Volte Traffic","Post_Volte_Traffic","Day7"),

                   ("Blank","",""),
                  

                   ("Data Volume","Pre_Data_Volume","Day1"),
                   ("Data Volume","Pre_Data_Volume","Day2"),
                   ("Data Volume","Pre_Data_Volume","Day3"),
                   ("Data Volume","Pre_Data_Volume","Day4"),
                   ("Data Volume","Pre_Data_Volume","Day5"),
                   ("Data Volume","Pre_Data_Volume","Day6"),
                   ("Data Volume","Pre_Data_Volume","Day7"),

                    ("Data Volume","Post_Data_Volume","Day1"),
                    ("Data Volume","Post_Data_Volume","Day2"),
                    ("Data Volume","Post_Data_Volume","Day3"),
                    ("Data Volume","Post_Data_Volume","Day4"),
                    ("Data Volume","Post_Data_Volume","Day5"),
                    ("Data Volume","Post_Data_Volume","Day6"),
                    ("Data Volume","Post_Data_Volume","Day7"),
                   
                    ("Comparison","","Pre_Volte_Traffic_AVG"),
                    ("Comparison","","Post_Volte_Traffic_AVG"),
                    ("Comparison","","Pre_Data_Volume_AVG"),
                    ("Comparison","","Post_Data_Volume_AVG"),
                    ("Percentage Change","","Volte Traffic"),
                    ("Percentage Change","","Data Volume"),
                   ]
            cols = pd.MultiIndex.from_tuples(tuples)
            df.columns=cols
            SiteWisePath= path + "/"+"SiteWise_record.xlsx"

#################################################### for making colourgradient of every cell (Site Wise) ######################################
            df_style=df.style
            idx = pd.IndexSlice
            # colors = ["red", "orange", "yellow","#ffff00", "darkgreen"]
            colors=["#cc3300","#ff9933","#ffcc33","#ffff33","#00cc33"]
            
            cmap1 = mpl.colors.LinearSegmentedColormap.from_list("mycmap",colors)
          
            index=len(df.index)
            print(df)
            for i in range(0,index):
                    # test=idx[idx[i],idx["Pre_Volte_Traffic":"Post_Volte_Traffic","Day1":"Day7"]]
                    # print(test)
                    slice_VT = idx[idx[i], idx["Volte Traffic",:,:]]
                    slice_DV = idx[idx[i], idx["Data Volume",:,:]]
                    df_style=df_style.background_gradient(cmap=cmap1,axis=1,subset=slice_VT)\
                                     .background_gradient(cmap=cmap1,axis=1,subset=slice_DV) #vmax=mx,vmin=mi
                    df_style.set_properties(**{'border':'1px solid black'})
# .background_gradient(cmap=Post_cmap1,axis=1,subset=slice_Post_VT)\
###################################################################################################################################
       
            df_style.to_excel(writer,sheet_name="SiteLevel", index=True)
            writer.save()
 
 
#################################################################### Code for Colouring the excel headers(Sector wise) #############################################################
            print("Performing excel operation.............................")
            wb = openpyxl.load_workbook(RecordPath)
            ws = wb['SectorLevel']
            ws["H1"].fill=PatternFill(patternType='solid',fgColor="ff34ebd2")
            ws["w1"].fill=PatternFill(patternType='solid',fgColor="ff34ebd2")
            ws["H2"].fill=PatternFill(patternType='solid',fgColor="ffff950a")
            ws["O2"].fill=PatternFill(patternType='solid',fgColor="ffffe414")
            ws["W2"].fill=PatternFill(patternType='solid',fgColor="ffff950a")
            ws["AD2"].fill=PatternFill(patternType='solid',fgColor="ffffe414")
            ws["AK1"].fill=PatternFill(patternType='solid',fgColor="ff148aff")
            ws["AK2"].fill=PatternFill(patternType='solid',fgColor="ff148aff")
            ws["AO1"].fill=PatternFill(patternType='solid',fgColor="ff33e0ff")
            ws["AO2"].fill=PatternFill(patternType='solid',fgColor="ff33e0ff")


            ws["AK3"].fill=PatternFill(patternType='solid',fgColor="ffff950a")
            ws["AL3"].fill=PatternFill(patternType='solid',fgColor="ffffe414")
            ws["AM3"].fill=PatternFill(patternType='solid',fgColor="ffff950a")
            ws["AN3"].fill=PatternFill(patternType='solid',fgColor="ffffe414")
          
            ws["AO3"].fill=PatternFill(patternType='solid',fgColor="ffff950a")
            ws["AP3"].fill=PatternFill(patternType='solid',fgColor="ffffe414")


 #################################################################### Code for Colouring the excel headers(Sector wise) #############################################################
       
            ws = wb["SiteLevel"]
            ws["G1"].fill=PatternFill(patternType='solid',fgColor="ff34ebd2")
            ws["V1"].fill=PatternFill(patternType='solid',fgColor="ff34ebd2")
            ws["G2"].fill=PatternFill(patternType='solid',fgColor="ffff950a")
            ws["N2"].fill=PatternFill(patternType='solid',fgColor="ffffe414")
            ws["V2"].fill=PatternFill(patternType='solid',fgColor="ffff950a")
            ws["AC2"].fill=PatternFill(patternType='solid',fgColor="ffffe414")

            ws["AJ1"].fill=PatternFill(patternType='solid',fgColor="ff148aff")
            ws["AJ2"].fill=PatternFill(patternType='solid',fgColor="ff148aff")
            ws["AN1"].fill=PatternFill(patternType='solid',fgColor="ff33e0ff")
            ws["AN2"].fill=PatternFill(patternType='solid',fgColor="ff33e0ff")

            ws["AJ3"].fill=PatternFill(patternType='solid',fgColor="ffff950a")
            ws["AK3"].fill=PatternFill(patternType='solid',fgColor="ffffe414")
            ws["AL3"].fill=PatternFill(patternType='solid',fgColor="ffff950a")
            ws["AM3"].fill=PatternFill(patternType='solid',fgColor="ffffe414")
          
            ws["AN3"].fill=PatternFill(patternType='solid',fgColor="ffff950a")
            ws["AO3"].fill=PatternFill(patternType='solid',fgColor="ffffe414")

            
            
            
            coloured_path=os.path.join( path,"coloured_record.xlsx")

            wb.save(coloured_path)
            print("excel operation done.........................")
            download_url=os.path.join(MEDIA_URL,"Original_trend","coloured_record.xlsx")
            
            return Response({"status":"true","message":"Report Generated Succesfully","Download_url":download_url})   
   




@api_view(["POST","PATCH","PUT"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])   
def raw_Kpi_upload(request):
     if request.POST.get("upload_date"):
        upload_date=request.POST["upload_date"]
     else:
          return Response({"status":"false","message":"upload_date required"})
     if request.FILES.get("file"):
         file= request.FILES["file"]
     else:
          return Response({"status":"false","message":"KPI file required"})
     obj,created=raw_kpis.objects.get_or_create(upload_date=upload_date) 
     obj.file=file
     obj.save()

     if created:
        message= "Post KPI file of " + str(upload_date) + " uploaded succesfully"
        return Response({"status":"true","message":message,"upload_date":upload_date})
     else:
        message= "Post KPI file of " + str(upload_date) + " updated succesfully"
        return Response({"status":"true","message":message,"upload_date":upload_date})
     

@api_view(["GET"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def get_integ_site_list(request):
    objs= integrated_sites.objects.all()
    serializer= ser_integrated_site(objs, many=True)

    return Response({"sites":serializer.data}) 