from rest_framework.status import HTTP_200_OK
from rest_framework.response import Response
from rest_framework.decorators import api_view
from datetime import datetime, timedelta
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import pandas as pd
import numpy as np
import os
import re



def get_circle(short_name):
    if "NE-ik" in short_name:
        if not "Sams-" in short_name:
            return short_name.split("-")[4]
        else:
            return short_name.split(",")[1][0:2]
    elif "Sams" in short_name:
        return short_name.split(",")[1][0:2]
    elif "@Nokia" in short_name:
        parts = re.split(r"(\d+)", short_name.split("-")[1])
        return parts[0] if len(parts) > 0 else None
    elif short_name.startswith("LD") or short_name.startswith("LU"):
        return "DL"
    else:
        return short_name.split("_")[0]


def num_hash(index: int) -> str:
    alphaChar = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    
    if(index < 26):
        return alphaChar[index - 1]
    else:
        q, r = index // 26, index % 26
        
        if(r == 0):
            if(q == 1):
                return alphaChar[r-1]
            else:
                return num_hash(q-1) + alphaChar[r-1]
        else:
            return num_hash(q) + alphaChar[r-1]
        

def titleToNumber(s: str) -> int:
    result = 0
    for B in range(len(s)):
        result *= 26
        result += ord(s[B]) - ord('A') + 1
    return result

def overwrite(df_pivoted: pd.DataFrame, kpi_name: str, coln1: str, trend_ws):

    index_pivot = df_pivoted.index.to_list()
    dr: pd.DataFrame = df_pivoted[kpi_name]
    
    coln2 = num_hash(titleToNumber(coln1) + 1)
    coln3 = num_hash(titleToNumber(coln1) + 2)
    coln4 = num_hash(titleToNumber(coln1) + 3)
    coln5 = num_hash(titleToNumber(coln1) + 4)
    
    cl = dr.columns.to_list()
    
    col1=dr[str(cl[0])].to_list()
    col2=dr[str(cl[1])].to_list()
    col3=dr[str(cl[2])].to_list()
    col4=dr[str(cl[3])].to_list()
    col5=dr[str(cl[4])].to_list()
    
    trend_ws[coln1+"2"].value=cl[0]
    trend_ws[coln2+"2"].value=cl[1]
    trend_ws[coln3+"2"].value=cl[2]
    trend_ws[coln4+"2"].value=cl[3]
    trend_ws[coln5+"2"].value=cl[4]
    
    for i,value in enumerate(index_pivot):
        j=i+3
        trend_ws["A"+str(j)].value=index_pivot[i][3]
        trend_ws["C"+str(j)].value=index_pivot[i][0] 
        trend_ws["D"+str(j)].value=index_pivot[i][1] 
        trend_ws["E"+str(j)].value=index_pivot[i][2] 
        trend_ws[coln1+str(j)].value="" if pd.isna(col1[i]) else col1[i]
        trend_ws[coln2+str(j)].value="" if pd.isna(col2[i]) else col2[i]
        trend_ws[coln3+str(j)].value="" if pd.isna(col3[i]) else col3[i]
        trend_ws[coln4+str(j)].value="" if pd.isna(col4[i]) else col4[i]
        trend_ws[coln5+str(j)].value="" if pd.isna(col5[i]) else col5[i]
    
    

@api_view(["POST", "GET"])
def get_kpi_monitoring_files(request):
    if request.method == "POST":
        file = request.FILES.get("kpi_report")
        
        if not file:
            return Response({"status": False, "message": "File not provided."}, status=400)

        samsung_kol_df = pd.read_csv(file)
        
        print(samsung_kol_df)

        if "Unnamed: 1" in samsung_kol_df.columns:
            samsung_kol_df.rename(columns={"Unnamed: 1": "Date"}, inplace=True)
        if "Unnamed: 2" in samsung_kol_df.columns:
            samsung_kol_df.rename(columns={"Unnamed: 2": "Date"}, inplace=True)

        samsung_kol_df['Short name'].ffill(inplace=True)

        base_url = os.path.join(MEDIA_ROOT, "KPI_MONITORING_TOOL_4G_5G_2G")
        if os.path.exists(base_url):
            os.makedirs(base_url,exist_ok=True)

        required_kpis = [
            'MV_Radio NW Availability',
            'MV_DL User Throughput_Kbps [CDBH]',
            'MV_UL User Throughput_Kbps [CDBH]',
            'MV_4G Data Volume_MB',
            'VoLTE Intra-LTE Handover Success Ratio [CBBH]',
            'MV_VoLTE Packet Loss UL [CBBH]',
            'MV_VoLTE Traffic',
            'MV_VoLTE DCR [CBBH]',
            'MV_VoLTE InterF HOSR Exec [CBBH]',
            'MV_Average number of used DL PRBs [CDBH]',
            'MV_Max Connected User [CDBH]',
            'MV_VoLTE CSSR [CBBH]'
        ]

        missing = [col for col in required_kpis if col not in samsung_kol_df.columns]
        if missing:
            return Response(
                {"status": False, "message": f"Missing KPIs: {', '.join(missing)}"},
                status=400
            )

        if '4G_ECGI' not in samsung_kol_df.columns:
            return Response(
                {"status": False, "message": "Missing column: 4G_ECGI"},
                status=400
            )
        for column in required_kpis:
            samsung_kol_df[column] = samsung_kol_df[column].astype("Float64")

        def get_tech(Short_name):
            tech = {
                '_F1_' : 'L2100',
                '_F3_' : 'L1800',
                '_F8_' : 'L900',
                '_T1_' : 'L2300',
                '_T2_' : 'L2300'
            }
            for tech,value in tech.items():
                if tech in Short_name:
                    return value
            
            return "Unknown"
        
        samsung_kol_df.insert(3, 'Tech', None)
        
        samsung_kol_df['Tech'] = samsung_kol_df['Short name'].apply(get_tech)
        samsung_kol_df.fillna(value=0, inplace=True)
        samsung_df = samsung_kol_df.pivot_table(index=['Short name', '4G_ECGI', 'Tech'],columns='Date',values=required_kpis, aggfunc=np.mean)
        
        print(samsung_df.columns)

        samsung_df[('Circle', '')] = samsung_df.index.get_level_values('Short name').map(get_circle)

        samsung_df.set_index(('Circle', ''), append=True, inplace=True)
        
        template_path = os.path.join(base_url, "templates", "KPI_4G_monitoring_tool_template_v1.1.xlsx")
        
        trend_wb = load_workbook(template_path)
        
        trend_ws=trend_wb["4G Cell Wise"]
        for kpi_name in required_kpis:
            if(kpi_name=="MV_4G Data Volume_MB"):
                overwrite(samsung_df,kpi_name,"H",trend_ws)

            if(kpi_name=="MV_VoLTE Traffic"):
                overwrite(samsung_df,kpi_name,"M",trend_ws)

            if(kpi_name=="MV_Radio NW Availability"):
                overwrite(samsung_df,kpi_name,"R",trend_ws)

            if(kpi_name=="MV_Average number of used DL PRBs [CDBH]"):
                overwrite(samsung_df,kpi_name,"W",trend_ws)

            if(kpi_name=="MV_Max Connected User [CDBH]"):
                overwrite(samsung_df,kpi_name,"AB",trend_ws)

            if(kpi_name=="MV_DL User Throughput_Kbps [CDBH]"):
                overwrite(samsung_df,kpi_name,"AG",trend_ws)

            if(kpi_name=="MV_UL User Throughput_Kbps [CDBH]"):
                overwrite(samsung_df,kpi_name,"AL",trend_ws)

            if(kpi_name=="MV_VoLTE DCR [CBBH]"):
                overwrite(samsung_df,kpi_name,"AQ",trend_ws)

            if(kpi_name=="VoLTE Intra-LTE Handover Success Ratio [CBBH]"):
                overwrite(samsung_df,kpi_name,"AV",trend_ws)

            if(kpi_name=="MV_VoLTE InterF HOSR Exec [CBBH]"):
                overwrite(samsung_df,kpi_name,"BA",trend_ws)

            if(kpi_name=="MV_VoLTE Packet Loss UL [CBBH]"):
                overwrite(samsung_df,kpi_name,"BF",trend_ws)

            if(kpi_name=="MV_VoLTE CSSR [CBBH]"):
                overwrite(samsung_df,kpi_name,"BK",trend_ws)
        
        
        output_url = os.path.join(base_url, "output_folder")
        
        if not (os.path.exists(output_url)):
            os.makedirs(output_url, exist_ok=True)
        
        current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        
        output_path = os.path.join(output_url, f"KPI_4G_monitoring_tool_output_{current_date}.xlsx")
        
        trend_wb.save(output_path)


        return Response(
            {
                "status": True,
                "message": "KPI data extracted successfully.",
            },
            status=HTTP_200_OK,
        )
