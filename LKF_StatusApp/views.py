from django.shortcuts import render
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST
from pathlib import Path
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
import os
import pandas as pd
import re
import stat
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import Counter
from datetime import datetime

#function to format and autofit excel-------------
def format_and_autofit_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Header formatting
    header_fill = PatternFill(start_color="215967", end_color="215967", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # Data cell formatting
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.font = bold_font
            if str(cell.value).strip().upper() == "NA":
                cell.fill = yellow_fill

    # Autofit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # 1-based
        column_letter = get_column_letter(column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(file_path)


#function to process sectorcarrier values----------------
def preprocess_file_content(content_lines):
    processed_lines = []
    skip_next = False

    for i, line in enumerate(content_lines):
        if skip_next:
            skip_next = False
            continue

        # Detect lines ending with '[1] ='
        if re.search(r'\[1\]\s*=\s*$', line.strip()):
            # Combine current line and the next '>>>' line
            if i + 1 < len(content_lines) and content_lines[i + 1].strip().startswith('>>>'):
                combined_line = line.strip() + ' ' + content_lines[i + 1].strip()
                processed_lines.append(combined_line)
                skip_next = True  # Skip next line
            else:
                processed_lines.append(line)
        else:
            processed_lines.append(line)

    return "\n".join(processed_lines)

def on_rm_error(func, path, exc_info):
    # Change the permission and retry
    os.chmod(path, stat.S_IWRITE)
    func(path)
    
#function to extract data from log files----------------
def explode_data_from_log(command, start_pattern, row_pattern, end_pattern, file_content):
    command_found = False
    header_found = False
    header_values = []
    data_rows = []
    node_id = None
    ip_addr = None
    timestamp = None

    # Compile patterns----------------
    command_regex = re.compile(command)
    start_regex = re.compile(start_pattern)
    row_regex = re.compile(row_pattern)
    end_regex = re.compile(end_pattern)
    log_info_regex = re.compile(
        r'(\d{6}-\d{2}:\d{2}:\d{2}\+\d{4})\s+([\da-fA-F:.]+)\s+(\d+\.\d+[a-zA-Z]*)\s+([\w.-]+)\s+(stopfile=[/\w\d]+)'
    )

    for line in file_content:
        line = line.strip()

        # Detect command
        if not command_found and command_regex.match(line):
            command_found = True
            node_id = line.split('>')[0].strip()
            continue

        # New prompt before header or end
        if command_found and not header_found and re.match(r'^[A-Z0-9_-]+>', line):
            break

        # Optional log info extraction
        log_match = log_info_regex.match(line)
        if log_match:
            timestamp, ip_addr, *_ = log_match.groups()

        # Detect header
        if command_found and not header_found:
            header_match = start_regex.match(line)
            if header_match:
                header_found = True
                header_values = list(header_match.groups()) or header_match.group(0).split()
                continue

        # Detect end of block
        if command_found and header_found and end_regex.match(line):
            break

        # Collect row data
        if command_found and header_found and not line.startswith('==='):
            row_match = row_regex.search(line)
            if row_match:
                data_rows.append(list(row_match.groups()))

    if data_rows and header_values:
        df = pd.DataFrame(data_rows, columns=header_values)
        df['Node_ID'] = node_id
        df['IP_Addr'] = ip_addr
        return df
    else:
        return pd.DataFrame()

def delete_existing_files(folder_path):
    if os.path.exists(folder_path):
        for file in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except Exception as e:
                print(f"Error deleting {file_path}: {e}")
                
#post api to upload and process files----------------
@api_view(["POST"])
def LKF_Upload(request):
    final_output_df = pd.DataFrame()

    files = request.FILES.getlist("files")
    if not files:
        return Response({"status": "ERROR", "message": "No files uploaded"}, status=HTTP_400_BAD_REQUEST)

    base_media_url = os.path.join(MEDIA_ROOT, "LKF_StatusApp")
    output_path = os.path.join(base_media_url, "LKF_Final_Output")
    log_folder = os.path.join(base_media_url, "logs_files")
    log_excel_folder = os.path.join(base_media_url, "logs_excel")
    os.makedirs(log_folder, exist_ok=True)
    os.makedirs(output_path, exist_ok=True)
    os.makedirs(log_excel_folder, exist_ok=True)

    delete_existing_files(log_folder)
    delete_existing_files(output_path)
    delete_existing_files(log_excel_folder)

    saved_files = []
    for file in files:
        file_path = os.path.join(log_folder, file.name)
        with open(file_path, "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        saved_files.append(file_path)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    # Process each uploaded file
    for uploaded_file in saved_files:
        node_name = os.path.splitext(os.path.basename(uploaded_file))[0]
        excel_filename = os.path.join(log_excel_folder, f"{node_name}_{timestamp}.xlsx")

        with open(uploaded_file, "r") as file:
            file_content = file.readlines()

        # Extract dataframes
        processed_content = preprocess_file_content(file_content).splitlines()
        st_cell_df = explode_data_from_log(
            r'[A-Z0-9_-]+>\sst\scell',
            r'(Proxy)\s+(Adm\sState)\s+(Op\.\sState)\s+(MO)',
            r'\s*(\d+)\s+(\d+\s+\((?:UNLOCKED|LOCKED)\))\s+(\d+\s+\((?:ENABLED|DISABLED)\))\s+(.*)$',
            r'^Total:\s\d+',
            file_content
        )

        hget_field_prod_df = explode_data_from_log(
            r'[A-Z0-9_-]+>\shget\sfield\sprod',
            r'MO\s+productName\s+productNumber\s+productRevision\s+productionDate\s+serialNumber',
            r'^\s*(\S+)\s+(.+)\s+(\S+/\d+)\s+(\S+)\s+(\d{8})\s+(\S+)\s*$',
            r'^Total:\s\d+',
            file_content
        )
        if not hget_field_prod_df.empty:
         hget_field_prod_df[['productName_new', 'productNumber_new']] = hget_field_prod_df['productName'].str.extract(r'^(.*\S)\s+([A-Z]{3}\s+\d+)$')

        
        st_trx_df = explode_data_from_log(
            r'[A-Z0-9_-]+>\sst\strx',
            r'(Proxy)\s+(Adm\sState)\s+(Op\.\sState)\s+(MO)',
            r'^\s*(\d+)\s+(\d+\s+\((?:UNLOCKED|LOCKED)\))\s*(\d+\s+\((?:ENABLED|DISABLED)\))?\s+(.*)$',
            r'^Total:\s\d+',
            file_content
        )

        get_maxtx_df = explode_data_from_log(
            r'[A-Z0-9_-]+>\sget\s\.?\smaxtx',
            r'MO\s+Attribute\s+Value',
            r'^\s*(\S+)\s+(\S+)\s*(.*)$',
            r'^Total:\s\d+',
            file_content
        )
        
  
        sectorcarrier_df = explode_data_from_log(
                r'[A-Z0-9_-]+>\sget\s\.\ssectorcarrier',
                r'MO\s+Attribute\s+Value',
                r'^\s*(\S+)\s+(\S+)\s+(.*)$',
                r'^Total:\s\d+',
                processed_content
            )
        if not sectorcarrier_df.empty:
         sectorcarrier_df['newValue'] = sectorcarrier_df['Value'].apply(
    lambda x: re.search(r'(SectorCarrier=\d+)', x).group(1) if re.search(r'(SectorCarrier=\d+)', x) else None)
         sectorcarrier_df=sectorcarrier_df[~sectorcarrier_df['MO'].str.contains(r'NRSectorCarrier=S\d+_N11|NRCellDU=', na=False)]
         sectorcarrier_df=sectorcarrier_df[~sectorcarrier_df['newValue'].isna()]
         
        
        fingerprint_df = explode_data_from_log(
            r'[A-Z0-9_-]+>\sget\s\.\sfingerprint',
            r'MO\s+Attribute\s+Value',
            r'^\s*(\S+)\s+(\S+)\s+(.*)$',
            r'^Total:\s\d+',
            file_content
        )
        
        sheets = {
            'st cell': st_cell_df,
            'hget field prod': hget_field_prod_df,
            'st trx': st_trx_df,
            'get maxtx': get_maxtx_df,
            'get sectorcarrier': sectorcarrier_df,
            'fingerprint': fingerprint_df,
        }

        # Save to Excel--------------------
        with pd.ExcelWriter(excel_filename, engine="xlsxwriter") as writer:
            for sheet_name, df in sheets.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Combine all processed Excel files into a final output
    final_output_df_list = []
    template_path = os.path.join(base_media_url, "template_file", "template.xlsx")

    for file in os.listdir(log_excel_folder):
        file_path = os.path.join(log_excel_folder, file)
        xls = pd.ExcelFile(file_path, engine="openpyxl")
        template_df = pd.read_excel(template_path, engine="openpyxl")
       
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
            
            if df.empty:
                continue

            
            trx=None
            for sheet_name in xls.sheet_names:
               
       #find the fingerprint --------------------     
                if sheet_name == "fingerprint" and not xls.parse(sheet_name).empty:
                    df = xls.parse(sheet_name)
                    mask = ((df['Attribute'] == 'fingerprint') & (df['Value'].notna()) & (~df['Value'].str.contains(":", na=False)))
                    fingerprint_row = df[mask]
                    finger_print = fingerprint_row['Value'].iloc[0] if not fingerprint_row.empty else "NA"
                    template_df.loc[0, 'New FINGER PRINT'] = finger_print
                   
                   
       #find the baseband and rru ---             
                if sheet_name == "hget field prod" and not xls.parse(sheet_name).empty:
                    df = xls.parse(sheet_name)
 
                    # Extract Baseband (e.g., "6651" from "RAN Processor 6651")
                    df['productName'] = df['productName'].astype(str).fillna("")
                    base = df['productName'].str.extract(r'(?:Baseband|RAN Processor)\s+(\w+)', expand=False).dropna()
                    baseband = base.iloc[0] if not base.empty else "NA"
                    template_df.loc[0, 'Baseband'] = baseband
                    template_df.loc[0, 'OLD Baseband'] = baseband
                    
                    rru_pattern = (
                        r'(AIR\s\d+\sB\d+[A-Z0-9]*|'
                        r'Radio\s(?:\d+\s*)+B\d+[A-Z0-9]*|'
                        r'RRUS\s\d+\sB\d+[A-Z0-9]*)'
                    )
                    df['productName_new'] = df['productName_new'].fillna("").astype(str) 
                    rru_df = df[df['productName_new'].str.contains(rru_pattern, na=False)]
                    model_band = rru_df['productName_new'].str.extract(rru_pattern, expand=False)

                    countRRU = model_band.value_counts()

                    result = ' + '.join(f"{item}*{count}" for item, count in countRRU.items())
                    template_df.loc[0, 'RRU Model'] = result if result else "NA"

                   
        
               
           
     
        # find the 5g count and power   -------------------------------  
                if sheet_name == "get maxtx" and not xls.parse(sheet_name).empty:
                    df = xls.parse(sheet_name)
                    count5g = df['MO'].str.extract(r'(NRSectorCarrier=S\d+_N11)')
                    template_df.loc[0, '5G'] = count5g.count()[0] if count5g.count()[0] else None
                    power5G_rows = df[df['MO'].str.contains(r'NRSectorCarrier=S\d+_N11')]
                    if not power5G_rows.empty:
                        power5G = power5G_rows['Value'].iloc[0]
                        power5G = power5G // 1000
                        template_df.loc[0, '5G Power'] = f"{power5G}W"
 
                    df_maxtx = xls.parse(sheet_name)
                    try:
                        df_sectorc = xls.parse("get sectorcarrier") 
                        
                    except Exception as e:
                        print(f" Sheet 'get . sectorcarrier' not found: {e}")
                        df_sectorc = pd.DataFrame()
                    output = ""
 
                    if not df_maxtx.empty and not df_sectorc.empty  :
                        common_mo = df_maxtx["MO"].unique().tolist()
                        df_sectorc = df_sectorc[df_sectorc["newValue"].isin(common_mo)]
                        power_tx_map = {}
                        for mo in common_mo:
                            power_vals = df_maxtx[df_maxtx["MO"] == mo]["Value"].values
                            tx_vals = df_maxtx[df_maxtx["MO"] == mo]["TxPower"].values if "TxPower" in df_maxtx.columns else [0]
                            if len(power_vals) > 0 and len(tx_vals) > 0:
                                if pd.notna(power_vals[0]) and pd.notna(tx_vals[0]):
                                    power_tx_map[mo] = (int(power_vals[0]), int(tx_vals[0]))
                        MIMO_map = {
                            "F1": "L2100",
                            "F3": "L1800",
                            "F8": "L900",
                            "T1": "L23001",
                            "T2": "L23002",
                            "F5": "L85",
                        }
                    layer_counter = Counter()
                    seen_pairs=set()
 
                    for _, row in df_sectorc.iterrows():
                        mo = row["newValue"]
                        band_info = row["MO"]
                        match = re.search(r"(F\d|T\d)", band_info)
                        if match:
                            band = match.group(1)
                            mimo_layer = MIMO_map.get(band)
                            if mimo_layer and mo in power_tx_map:
                                pair = (mo, band)
                                if pair not in seen_pairs:
                                    seen_pairs.add(pair)
                                    power, _ = power_tx_map[mo]
                                    power_watt = f"{power // 1000}W"
                                    key = f"{power_watt}{mimo_layer}"
                                    layer_counter[key] += 1
                    # Process layer_counter correctly----------
                    for key, count in layer_counter.items():
                        match = re.match(r"(\d+)W(L\d+)", key)
                        print("Layer Key:-------------------------", key)
                        if match:
                            power = match.group(1)
                            layer = match.group(2)
 
                            if layer == "L900" and count > 0:
                                template_df.loc[0, "L900"] = count
                                template_df.loc[0, "L900 Power"] = f"{power}W"
 
                            elif layer == "L1800" and count > 0:
                                template_df.loc[0, "FDD-L1800"] = count
                                template_df.loc[0, "L1800 Power"] = f"{power}W"
 
                            elif layer == "L2100" and count > 0:
                                template_df.loc[0, "L2100"] = count
                                template_df.loc[0, "L2100 Power"] = f"{power}W"
                                
                            elif layer=="L23001" and count>0:
                                template_df.loc[0, "TDD-20Mz"] = count
                                template_df.loc[0, "TDD Power"] = f"{power}W"
                                
                            elif layer=="L23002" and count>0:
                                template_df.loc[0, "TDD-10Mhz"] = count
                                template_df.loc[0, "TDD Power"] = f"{power}W"    
                                
 
                           
                                
                                   
                    layers = []
                    trx = None 
 
                
                    for sheet_name in xls.sheet_names:
                        df = xls.parse(sheet_name)
                        if df.empty:
                            continue
 
                        #Extract TRX from "st trx" sheet
                        if sheet_name == "st trx" and not xls.parse(sheet_name).empty:
                            trx_count = max(len(df) - 1, 0)
                            trx = trx_count if trx_count > 0 else None
                            template_df.loc[0, "TRX"] = trx  # optional: store for reference
 
                        elif sheet_name == "st cell" and not xls.parse(sheet_name).empty:
                            df = xls.parse(sheet_name)
               
                            df['IP_Addr'] = df['IP_Addr'].astype(str).fillna('NA')
                            ip_address = df['IP_Addr'].dropna().unique()
                            template_df.loc[0, 'Node IP'] = ip_address[0] if len(ip_address) > 0 else "NA"
 
                   
                            df['Node_ID'] = df['Node_ID'].astype(str).fillna('NA')
                            mo = df['Node_ID'].unique()
                            template_df.loc[0, 'MO Name'] = mo[0] if len(mo) > 0 else "NA"
 
                            df['site_id'] = df['Node_ID'].str.extractall(r'[A-Z]+(\d+)').groupby(level=0).last()
                            site_id = df['site_id'].iloc[0] if not df['site_id'].empty and pd.notna(df['site_id'].iloc[0]) else "NA"
                            template_df.loc[0, "Site Id"] = site_id
                            template_df.loc[0, "Activity"] = "Relocation"
                            template_df.loc[0, "Additional Remak"] = "UST"
                            template_df.loc[0, "Additional Remaks"] = "CPRI Port Expansion"
                       
                       
                           
                            # Layer extraction from MO----------
                            extracted_layers = df["MO"].str.extractall(r'_([FT]\d)_')
                            layer_matches = extracted_layers[0].dropna().unique().tolist()
 
                            layer_map = {
                                "F1": "L2100",
                                "F3": "L1800",
                                "F5": "L85",
                                "F8": "L900",
                            }
 
                            for key, value in layer_map.items():
                                if key in layer_matches:
                                    layers.append(value)
 
                            if "T1" in layer_matches or "T2" in layer_matches:
                                layers.append("TDD")
 
                            if df["MO"].str.contains("NRCellDU").any():
                                layers.append("5G")
 
                    #  Add TRX to layers if applicable
                    if pd.notna(trx) and trx > 1:
                        layers.append("TRX")
 
                    # Finalize and store Layer string------
                    layer_str = "+".join(layers)
                    print("TRX:-------------------------", trx)
                    print("Layer:", layer_str)
 
                    template_df.loc[0, "Layer"] = layer_str if layer_str else "NA"
                   
                                
         #save the output
        final_output_df_list.append(template_df)

    final_output_df = pd.concat(final_output_df_list, ignore_index=True)

    # Save final combined output
    outut_file_name = f"LKF_Status_{timestamp}.xlsx"
    final_output_path = os.path.join(output_path, outut_file_name)
    final_output_df.to_excel(final_output_path, index=False, engine="openpyxl")
    format_and_autofit_excel(final_output_path)

    relative_path = os.path.relpath(final_output_path, MEDIA_ROOT).replace("\\", "/")
    download_url = request.build_absolute_uri(MEDIA_URL + relative_path)
    print("save the fle......")
    print("End the Process-------------------------")
   

    return Response(
        {
            "status": True,
            "message": "Files processed successfully",
            "download_url": download_url,
        },
      
    )


        