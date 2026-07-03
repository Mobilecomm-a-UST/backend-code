from django.shortcuts import render
from django.db.models import *
from rest_framework.decorators import api_view
from rest_framework.response import Response
import pandas as pd
import os
from django.conf import settings
from rest_framework import status
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from urllib.parse import urljoin
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from .models import *
import zipfile
import shutil
MEDIA_ROOT = settings.MEDIA_ROOT


# ---------------- Excel Format ----------------
def format_of_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    header_fill = PatternFill(start_color="E87529", end_color="E87529", fill_type="solid")
    header_font = Font(color="000000", bold=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Borders
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border

    # Column width
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 25

    wb.save(file_path)
    
@api_view(['GET', 'POST'])
def fileupload(request):

    # ---------- GET ----------
    if request.method == 'GET':
        data = ZteAlarm.objects.all().values()
        return Response(list(data))

    # ---------- POST ----------
    elif request.method == 'POST':
        file = request.FILES.get('file')

        if not file:
            return Response({'error': 'File required'}, status=400)

        try:
            # Read file
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)

            # Clean columns
            df.columns = df.columns.str.strip()

            # Clear old data
            ZteAlarm.objects.all().delete()

            objs = []

            for _, row in df.iterrows():
                objs.append(
                    ZteAlarm(
                        alarm_name=row.get("MO"),
                        alarm_status=row.get("Alarm/No Alarm"),
                        sa_nsa=row.get("SA/NSA"),
                        alarm_bucket=row.get("Alarm Bucket"),
                        responsibility=row.get("Responsibility"),
                    )
                )

            ZteAlarm.objects.bulk_create(objs)

            return Response({"message": "Data uploaded successfully"})

        except Exception as e:
            return Response({"error": str(e)}, status=400)
        
@api_view(['GET', 'POST', 'DELETE'])
def manage_sites(request):
    try:
        # ===================== GET =====================
        if request.method == 'GET':
            data = list(Old_New.objects.values())

            return Response({
                "status": True,
                "data": data
            })

        # ===================== DELETE =====================
        elif request.method == 'DELETE':
            Old_New.objects.all().delete()

            return Response({
                "status": True,
                "message": "Deleted successfully"
            })

        # ===================== POST =====================
        elif request.method == 'POST':
            file = request.FILES.get('site_file')

            if not file:
                return Response({
                    "status": False,
                    "error": "File required"
                }, status=400)

            if not file.name.endswith('.xlsx'):
                return Response({
                    "status": False,
                    "error": "Only Excel file allowed"
                }, status=400)

            # Clear old data
            Old_New.objects.all().delete()

            excel_file = pd.ExcelFile(file)
            df = None

            # Find correct sheet
            for sheet_name in excel_file.sheet_names:
                temp_df = pd.read_excel(excel_file, sheet_name=sheet_name)
                temp_df.columns = temp_df.columns.astype(str).str.strip()

                if (
                    'New 4G NE ID' in temp_df.columns and
                    'Old 4G NE ID' in temp_df.columns
                ):
                    df = temp_df
                    break

            if df is None:
                return Response({
                    "status": False,
                    "error": "Required columns not found in any sheet"
                }, status=400)

            # Insert data
            records = []
            for _, row in df.iterrows():
                new_site = str(row.get('New 4G NE ID', '')).strip()
                old_site = str(row.get('Old 4G NE ID', '')).strip()

                if not new_site and not old_site:
                    continue

                records.append(
                    Old_New(
                        new_site=new_site,
                        old_site=old_site
                    )
                )

            # Bulk insert (faster 🔥)
            Old_New.objects.bulk_create(records)

            return Response({
                "status": True,
                "message": "Data uploaded successfully",
                "count": len(records)
            })

    except Exception as e:
        return Response({
            "status": False,
            "error": str(e)
        }, status=500)
        
mapping_path = os.path.join(settings.MEDIA_ROOT, 'mapping_file')
os.makedirs(mapping_path, exist_ok=True)
    
@api_view(['POST', 'GET', 'DELETE'])
def upload_mapping_file(request):
    # ================= POST (UPLOAD) =================
    if request.method == 'POST':

        files = request.FILES.getlist('mapping_file')

        if not files:
            return Response(
                {'error': 'No files uploaded'},
                status=status.HTTP_400_BAD_REQUEST
            )

        saved_files = []

        for f in files:
            file_path = os.path.join(mapping_path, f.name)

            with open(file_path, 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk)

            saved_files.append(f.name)

        return Response({
            'status': True,
            'message': 'Files uploaded successfully',
            'files': saved_files
        })

    # ================= GET (LIST FILES) =================
    elif request.method == 'GET':

        if not os.path.exists(mapping_path):
            return Response({'files': []})

        files = os.listdir(mapping_path)

        return Response({
            'status': True,
            'files': files
        })

    # ================= DELETE =================
    elif request.method == 'DELETE':

        if not os.path.exists(mapping_path):
            return Response({'error': 'Folder not found'}, status=404)

        deleted = []

        for file in os.listdir(mapping_path):
            file_path = os.path.join(mapping_path, file)

            if os.path.isfile(file_path):
                os.remove(file_path)
                deleted.append(file)

        return Response({
            'status': True,
            'deleted_files': deleted
        })
        
@api_view(['POST'])
def alarmfileUpload(request):

    # ================= ALARM FILES FROM REQUEST =================
    alarm_files = request.FILES.getlist('alarm_file')

    if not alarm_files:
        return Response(
            {"error": "alarm_file required"},
            status=400
        )

    # ================= READ ALARM FILES =================
    alarm_dfs = []
    for alarm_file in alarm_files:

        print("\n====================")
        print("FILE =", alarm_file.name)

        # ================= CSV =================
        if alarm_file.name.endswith(".csv"):

            temp_df = pd.read_csv(
                alarm_file,
                header=None
            )

            header_idx = 0

            for idx, row in temp_df.iterrows():

                values = (
                    row.astype(str)
                    .str.strip()
                    .tolist()
                )

                if "Severity" in values and "NE" in values:
                    header_idx = idx
                    break

                if "ManagedElement" in values and "NE_Name" in values:
                    header_idx = idx
                    break

                if "MEID" in values and "adminState" in values:
                    header_idx = idx
                    break

            alarm_file.seek(0)

            df = pd.read_csv(
                alarm_file,
                header=header_idx
            )

            df.columns = (
                df.columns.astype(str)
                .str.strip()
            )

            # print("READ COLUMNS")
            # print(df.columns.tolist())

            alarm_dfs.append(df)

        # ================= EXCEL =================
        else:
            xls = pd.ExcelFile(alarm_file)
            print("SHEETS =", xls.sheet_names)
            for sheet in xls.sheet_names:
                try:
                    alarm_file.seek(0)

                    temp_df = pd.read_excel(
                        alarm_file,
                        sheet_name=sheet,
                        header=None
                    )
                    header_idx = None
                    for idx, row in temp_df.iterrows():
                        values = (
                            row.astype(str)
                            .str.strip()
                            .tolist()
                        )
                        # Alarm Sheet
                        if "Severity" in values and "NE" in values:
                            header_idx = idx
                            break
                        # 4G TDD Sheet
                        if "ManagedElement" in values and "NE_Name" in values:
                            header_idx = idx
                            break
                        # 4G FDD Sheet
                        if "MEID" in values and "adminState" in values:
                            header_idx = idx
                            break

                    if header_idx is None:
                        continue
                    alarm_file.seek(0)
                    df = pd.read_excel(
                        alarm_file,
                        sheet_name=sheet,
                        header=header_idx
                    )
                    df.columns = (
                        df.columns.astype(str)
                        .str.strip()
                    )
                
                    alarm_dfs.append(df)
                    print("------sheets processed--------")
                except Exception as e:
                    print(
                        f"SKIPPED SHEET {sheet} : {e}"
                    )

    df_alarm = pd.concat(alarm_dfs, ignore_index=True)    
    #---------- 4g lock unlock sheet---------------
    df_4g = pd.DataFrame()
    
    for df in alarm_dfs:
        cols = [str(c).strip() for c in df.columns]
        temp_4g = df.copy()
        temp_4g.columns = temp_4g.columns.astype(str).str.strip()

        # ---------------- CASE 1: MEID ----------------
        if "MEID" in cols:
            temp_4g = temp_4g[temp_4g["MEID"].notna()]
            temp_4g["4G NE ID"] = (
                temp_4g["MEID"]
                .astype(str)
                .str.replace(r"\.0$", "", regex=True)
                .str.strip()
                .str.upper()
            )
            temp_4g = temp_4g[
                temp_4g["4G NE ID"].str.match(r"^\d+$", na=False)
            ]
        # ---------------- CASE 2: ManagedElement ----------------
        
        elif "ManagedElement" in cols:
            temp_4g = temp_4g[temp_4g["ManagedElement"].notna()]
            temp_4g["4G NE ID"] = (
                temp_4g["ManagedElement"]
                .astype(str)
                .str.replace(r"\.0$", "", regex=True)
                .str.strip()
                .str.upper()
            )
            temp_4g = temp_4g[
                    temp_4g["4G NE ID"].str.match(r"^\d+$", na=False)
                ]
        else:
            continue
        # ---------------- COMMON NORMALIZATION ----------------
        if "adminState" in temp_4g.columns:
            temp_4g["Admin Status"] = temp_4g["adminState"].astype(str).replace({
                "0": "Unlocked",
                "1": "Locked",
                "Unlocked[0]": "Unlocked",
                "Locked[0]": "Locked",
            })
        else:
            temp_4g["Admin Status"] = ""

        if "testState" in temp_4g.columns:
            temp_4g["Test Status"] = temp_4g["testState"].astype(str).replace({
                "0": "Normal",
                "2": "Testing",
                "5": "FM Testing",
                "6": "PM Testing",
                "Normal[0]": "Normal",
                "Testing[2]": "Testing",
                "FM Testing[5]": "FM Testing",
                "PM Testing[6]": "PM Testing",
            })
        else:
            temp_4g["Test Status"] = ""

        temp_4g["Cell Name"] = temp_4g["userLabel"] if "userLabel" in temp_4g.columns else ""
        df_4g = pd.concat([df_4g, temp_4g], ignore_index=True)

    # ================= LOOKUP MAP (FAST) =================
    df_4g = df_4g.drop_duplicates(subset=["4G NE ID"])
    df_4g = df_4g[
        [
            "4G NE ID",
            "Cell Name",
            "Admin Status",
            "Test Status",
        ]
    ]
    print("STEP: loading mapping file...", flush=True)
    # ================= LOAD MAPPING =================
    mapping_file = os.path.join(mapping_path, os.listdir(mapping_path)[0])
    
    if not os.path.exists(mapping_file):
        return Response({"error": "mapping file not found"}, status=400)

    df_map = pd.read_excel(mapping_file)
    df_map.columns = df_map.columns.str.strip()

    # 🔥 KEEP ONLY ZTE
    if "OEM" not in df_map.columns:
        return Response({"error": "OEM column missing in mapping file"}, status=400)

    df_map = df_map[df_map["OEM"].astype(str).str.upper() == "ZTE"]
    df_map["4G NE ID"] = df_map["4G NE ID"].astype(str).str.replace(".0", "", regex=False).str.strip().str.upper()
    print("STEP: mapping file loaded", flush=True)
    # ================= SITE ID =================
    if "ME ID" in df_alarm.columns:
        df_alarm["4G NE ID"] = df_alarm["ME ID"].astype(str).str.upper().str.strip()
    elif "Site ID(Office)" in df_alarm.columns:
        df_alarm["4G NE ID"] = df_alarm["Site ID(Office)"].astype(str).str.upper().str.strip()
    else:
        df_alarm["4G NE ID"] = ""
    
    df_alarm["4G NE ID"] = (
        df_alarm["4G NE ID"]
        .astype(str)
        .str.replace(".0", "", regex=False)
        .str.strip()
        .str.upper()
    )

    # ================= ALARM NAME =================
    if "Alarm Code" in df_alarm.columns:
        df_alarm["alarms"] = df_alarm["Alarm Code"].astype(str).str.replace(r"\(\d+\)", "", regex=True).str.strip()

    elif "Alarm title" in df_alarm.columns:
        df_alarm["alarms"] = df_alarm["Alarm title"].astype(str).str.replace(r"\(\d+\)", "", regex=True).str.strip()

    if "alarms" not in df_alarm.columns:
        df_alarm["alarms"] = ""

    needed_cols = ["4G NE ID", "alarms"]
    # -------- Keep only required columns --------

    df_alarm = df_alarm[[c for c in needed_cols if c in df_alarm.columns]]

    # ---------------- Load SA / NSA from DB ----------------
    alarm_master = ZteAlarm.objects.all()

    alarm_lookup = {
        str(obj.alarm_name).strip().lower(): {
            "sa_nsa": obj.sa_nsa,
            "alarm_name": obj.alarm_name,
            "alarm_status": obj.alarm_status,
            "alarm_bucket": obj.alarm_bucket,
            "responsibility": obj.responsibility
        }

        for obj in alarm_master if obj.alarm_name
    }

    # ---------------- Classify Alarms ----------------
    def classify_alarms(info):

        if pd.isna(info):
            return pd.Series(["No", "No Alarms", "No Alarms", "", ""])

        alarm = str(info).strip().lower()

        if alarm in alarm_lookup:

            details = alarm_lookup[alarm]

            sa_nsa = str(details["sa_nsa"]).upper()

            return pd.Series([
                "Yes",
                sa_nsa if sa_nsa in ["SA", "NSA"] else "No Alarms",
                details["alarm_name"],
                details["alarm_bucket"],
                details["responsibility"]
            ])

        return pd.Series(["No", "No Alarms", "No Alarms", "", ""])
    
    print("STEP: classifying alarms...", flush=True)
    df_alarm[
        [
            "Alarm Status (Yes/No)",
            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms",
            "Service Affecting Alarms 1",
            "alarm_bucket",
            "responsibility"
        ]
    ] = df_alarm["alarms"].apply(classify_alarms)
    print("STEP: alarms classified", flush=True)
    # ---------------- Clean + Group ----------------
    def clean_and_merge(values, column_name=""):

        vals = [
            str(v).strip()
            for v in values
            if pd.notna(v) and str(v).strip() and str(v).strip().lower() != "nan"
        ]

        vals_set = set(vals)

        if column_name == "Alarm Status (Yes/No)":
            return "Yes" if "Yes" in vals_set else "No"

        if column_name == "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms":
            vals_upper = {v.upper() for v in vals_set}

            if "SITE DOWN" in vals_upper:
                return "SITE DOWN"
            if "SA" in vals_upper:
                return "SA"
            if "NSA" in vals_upper:
                return "NSA"

            return "No Alarms"

        return ", ".join(sorted(vals_set))

    print("STEP: grouping alarms...", flush=True)
    df_alarm = (
        df_alarm.drop_duplicates()
        .groupby("4G NE ID", as_index=False)
        .agg(lambda x: clean_and_merge(x, x.name))
    )
    print("STEP: alarms grouped", flush=True)
    df_alarm = (
        df_alarm.groupby("4G NE ID", as_index=False)
        .agg({
            "Alarm Status (Yes/No)":
                lambda x: "Yes"
                if "Yes" in x.values
                else "No",

            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms":
                lambda x: (
                    "SITE DOWN"
                    if "SITE DOWN" in x.astype(str).values
                    else (
                        ", ".join(
                            sorted(
                                {
                                    i for i in x.dropna().astype(str)
                                    if i.strip() != "No Alarms"
                                }
                            )
                        ) or "No Alarms"
                    )
                ),
            "Service Affecting Alarms 1":
                lambda x: ", ".join(
                    sorted(
                        {
                            i for i in x.dropna().astype(str)
                            if i.strip() != "No Alarms"
                        }
                    )
                ) or "No Alarms",

            "alarm_bucket":
                lambda x: ",".join(
                    sorted(
                        set(
                            x.dropna().astype(str)
                        )
                    )
                ),
            "responsibility":
                lambda x: ",".join(
                    sorted(
                        set(
                            x.dropna().astype(str)
                        )
                    )
                )
        })
    )
    
    # ================= FORCE CLEAN KEYS =================

    df_alarm["4G NE ID"] = (
        df_alarm["4G NE ID"]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .str.upper()
    )

    df_map["4G NE ID"] = (
        df_map["4G NE ID"]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .str.upper()
    )
    
    old_new_df = pd.DataFrame(
        list(
            Old_New.objects.values(
                "old_site",
                "new_site"
            )
        )
    )

    for col in old_new_df.columns:
        old_new_df[col] = (
            old_new_df[col]
            .astype(str)
            .str.strip()
            .str.upper()
        )

    old_site_set = set(old_new_df["old_site"])
    new_site_set = set(old_new_df["new_site"])

    # ---------------- Site Status Dummy ----------------
    def get_site_status(siteid):

        siteid = str(siteid).strip().upper()
        print("Site ID:", siteid)

        if siteid in new_site_set:
            return "New Site"

        elif siteid in old_site_set:
            return "OLD Site"

        return ""
    print("STEP: merging map + alarm...", flush=True)
    df_merged = pd.merge(
        df_map,
        df_alarm,
        on="4G NE ID",
        how="left",
        suffixes=("", "_alarm")
    )
    print("STEP: merge done", flush=True)
    # df_merged["Site Status"] = (df_merged["4G NE ID"].apply(get_site_status))
    
    cols_to_update = [
        "Alarm Status (Yes/No)",
        "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms",
        "Service Affecting Alarms 1",
        "alarm_bucket",
        "responsibility"
    ]
    for col in cols_to_update:

        alarm_col = f"{col}_alarm"

        if alarm_col in df_merged.columns:

            df_merged[col] = df_merged[col].mask(
                df_merged[col].isna() |
                (
                    df_merged[col]
                    .astype(str)
                    .str.strip() == ""
                ),
                df_merged[alarm_col]
            )

            df_merged.drop(
                columns=[alarm_col],
                inplace=True
            )
        
    df_merged["Alarm Status (Yes/No)"] = (
        df_merged["Alarm Status (Yes/No)"]
        .fillna("No")
    )

    df_merged[
        "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms"
    ] = (
        df_merged[
            "No Alarms/Service Affecting Alarms/Non-Service Affecting Alarms"
        ]
        .fillna("No Alarms")
    )

    df_merged["Service Affecting Alarms 1"] = (
        df_merged["Service Affecting Alarms 1"]
        .fillna("No Alarms")
    )

    df_merged["alarm_bucket"] = (
        df_merged["alarm_bucket"]
        .fillna("")
    )
    df_merged["responsibility"] = (
        df_merged["responsibility"]
        .fillna("")
    )
    # df_merged["Site Status"] = df_merged["4G NE ID"].apply(get_site_status)
    df_merged["Site Status"] = df_merged["SiteId"].apply(get_site_status)
    
    if df_4g.empty:

        df_4g = pd.DataFrame(
            columns=[
                "4G NE ID",
                "Admin Status",
                "Test Status",
                "Cell Name",
                "SiteId",
                "Site Status"
                
            ]
        )
    df_4g = df_4g[
            df_4g["4G NE ID"].isin(df_merged["4G NE ID"])
        ]
        
    df_4g = df_4g.merge(
        df_map[["4G NE ID", "SiteId"]],
        on="4G NE ID",
        how="left"
    )

    df_4g["Site Status"] = df_4g["SiteId"].apply(get_site_status)
    #-----------old vs new -----------------------------------
    print("STEP: starting old vs new compare loop...", flush=True)
    compare_4g_rows = []
    for _, row in old_new_df.iterrows():

        old_site = str(row["old_site"]).strip().upper()
        new_site = str(row["new_site"]).strip().upper()

        old_data = df_4g[
            df_4g["4G NE ID"].astype(str).str.upper() == old_site
        ]

        new_data = df_4g[
            df_4g["4G NE ID"].astype(str).str.upper() == new_site
        ]

        old_row = old_data.iloc[0] if not old_data.empty else {}
        new_row = new_data.iloc[0] if not new_data.empty else {}

        old_admin = str(
            old_row.get("Admin Status", "")
        ).strip().upper() if len(old_data) else ""

        new_admin = str(
            new_row.get("Admin Status", "")
        ).strip().upper() if len(new_data) else ""

        if old_admin == "UNLOCKED" and new_admin == "UNLOCKED":
            lock_status = "Both Unlocked"

        elif old_admin == "LOCKED" and new_admin == "LOCKED":
            lock_status = "Both Locked"

        elif old_admin and new_admin:
            lock_status = "Changed"

        else:
            lock_status = ""

        compare_4g_rows.append({

            "SiteId_old": old_site,
            "Admin Status_old": old_admin,
            "Test Status_old": str(old_row.get("Test Status", "")).upper(),
            "Cell Name_old": old_row.get("Cell Name", ""),
            
            "SiteId_new": new_site,
            "Admin Status_new": new_admin,
            "Test Status_new": str(new_row.get("Test Status", "")).upper(),            
            "Cell Name_new": new_row.get("Cell Name", ""),
            
            "Lock Status":
                lock_status
        })

    compare_4g_df = pd.DataFrame(compare_4g_rows)
    print("STEP: compare loop done", flush=True)
    # ================= CLEAN OUTPUT =================
    output_root = os.path.join(settings.MEDIA_ROOT, "ZTE_OUTPUT", "OUTPUT")

    if os.path.exists(output_root):
        shutil.rmtree(output_root)

    os.makedirs(output_root, exist_ok=True)

    # ================= SAVE EXCEL =================
    if "Circle" not in df_merged.columns:
        return Response({"error": "Circle column not found in mapping file"}, status=400)

    # for Circle, df in final_outputs:
    for circle in df_merged["Circle"].dropna().unique():
        print(f"STEP: writing excel for circle = {circle}", flush=True)

        circle_folder = os.path.join(output_root, str(circle).strip())
        os.makedirs(circle_folder, exist_ok=True)
        df_circle = df_merged[df_merged["Circle"] == circle]
        df_4g_circle = df_4g[df_4g["4G NE ID"].isin(df_circle["4G NE ID"])]

        file_path = os.path.join(circle_folder, f"{circle}_output.xlsx")

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_circle.to_excel(writer, sheet_name="Alarms", index=False)
            df_4g_circle.to_excel(writer, sheet_name="4G", index=False)
            compare_4g_df.to_excel(writer, sheet_name="4G Old vs New", index=False)

        format_of_excel(file_path)
        print(f"STEP: circle {circle} done", flush=True)


    # ================= ZIP OUTPUT =================
    zip_base = os.path.join(settings.MEDIA_ROOT, "ZTE_OUTPUT", "OUTPUT")

    shutil.make_archive(zip_base, 'zip', output_root)

    download_url = request.build_absolute_uri(
        urljoin(settings.MEDIA_URL, "ZTE_OUTPUT/OUTPUT.zip")
    )

    return Response({
        "status": True,
        "message": "Processed successfully",
        "download_url": download_url
    })