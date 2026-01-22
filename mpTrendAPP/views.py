from django.shortcuts import render
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
import os
from django.shortcuts import render

from django.shortcuts import render
from django.shortcuts import render,redirect
from django.http import HttpResponse
import sys
import pandas as pd
from django.contrib import messages
# import json
from django.contrib.auth.decorators import login_required
# from mcom_websites.settings import MEDIA_URL,BASE_DIR,ALLOWED_HOSTS
import numpy as np
import os
import datetime
from rest_framework.decorators import api_view,parser_classes
from rest_framework.response import Response
from django.http import JsonResponse
# from .serializer import *
from django.forms.models import model_to_dict

from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import authentication_classes,permission_classes
from django.core.files.storage import FileSystemStorage
# from .settings import MEDIA_ROOT
from .models import *

import openpyxl
from openpyxl.styles import PatternFill
# Create your views here.
from openpyxl import Workbook,load_workbook
from statistics import mean

from django.db.models import Sum
from commom_utilities.utils import *


import pandas as pd
import matplotlib  as mpl
import numpy as np

from datetime import date, timedelta

@api_view(["POST"])
def old_mp_trend(request):
        raw_kpi = request.FILES["raw_kpi"] if 'raw_kpi' in request.FILES else None
        if raw_kpi:
             location=MEDIA_ROOT+r'\temporary_files'
             fs=FileSystemStorage(location=location)
             file=fs.save(raw_kpi.name,raw_kpi)
             file_path=fs.path(file)
             df_raw_kpi=pd.read_excel(file_path)
             print(df_raw_kpi)
            #  os.remove(path=file_path)




        required_cols = ['MV_4G Data Volume_GB', 'E-UTRAN Average CQI [CDBH]', 'Average UE Distance_KM',
            'DL User Throughput_Kbps [CDBH]', 'UL User Throughput_Kbps [CDBH]', 'Average number of used DL PRBs [CDBH]',
        'RRC Setup Success Rate [CDBH]', 'ERAB Setup Success Rate [CDBH]',
        'MV_PS Drop Call Rate % [CDBH]', 'UL RSSI [CDBH]', 'MV_CSFB Redirection Success Rate [CDBH]',
        'MV_LTE_PS_DCR', 'PS handover success rate [LTE Inter System] [CDBH]', 'PS handover success rate [LTE Intra System] [CDBH]',
        'VoLTE Traffic', 'VoLTE DCR [CBBH]',
        'VoLTE Packet Loss DL [CBBH]', 'VoLTE Packet Loss UL [CBBH]',
        'VoLTE InterF HOSR Exec [CBBH]', 'VoLTE IntraF HOSR Exec [CBBH]',
        'Radio NW Availability','TA Sample <500M','TA Sample <1KM','TA Sample <1.5KM_Nokia_1','TA Sample <2KM','RSRP Samples<-104 dBm','RSRP Samples<-110 dBm'
        ,'RSRP Samples<-116 dBm']

        sts,response=required_col_check(raw_kpi,required_cols)
        if sts:
          return Response(response)


        site_list=request.FILES['site_list'] if 'site_list' in request.FILES else None
        if site_list:
             location=MEDIA_ROOT+r'\temporary_files'
             fs=FileSystemStorage(location=location)
             file=fs.save(site_list.name,site_list)
             file_path=fs.path(file)  
             df_site_list=pd.read_excel(file_path)  
             print(df_site_list)
             os.remove(path=file_path) 

        # if site_list:
        #     sts,response=required_col_check(raw_kpi,required_cols)
        #     if sts:
        #      return Response(response)
            

        # response,s_l= site_list_handler(request)
        # if s_l:
        #     df_site_list=s_l
        #     if response:
        #      return Response(response)
        response,s_l= site_list_handler(request)
        if s_l:
            df_site_list=s_l
            print("_________________________--print",type(df_site_list[0]))

            print(df_site_list,"_______________")
        if response:
            return Response(response)
   
        

      
        door_root= os.path.join(MEDIA_ROOT,'trends',"mp")
        path_of_blnk_temp=os.path.join(door_root,"template","mp_trend1.xlsx")
        wb=load_workbook(path_of_blnk_temp)
        


        df_raw_kpi["Short name"] = df_raw_kpi["Short name"] = df_raw_kpi["Short name"].fillna(method="ffill")  #######for (forward fill)ffill is used to copy  and fill
        
        print(df_raw_kpi.head())
        print(df_raw_kpi.columns)
        # df_raw_kpi.to_excel('process_output/mp.xlsx')
        df_raw_kpi.columns.values[1] = 'Date'  #####for empty colum
    # # df.rename(columns={"Cell Id" :" ECGI" } , inplace = True )


        df_raw_kpi["DL User Throughput_Kbps [CDBH]"] = (df_raw_kpi["DL User Throughput_Kbps [CDBH]"] / 1024)  #####for change kbps to mbps
    # df.rename(columns={"DL User Throughput_Kbps [CDBH]" :" DL User Throughput_Mbps [CDBH]" } , inplace = True )
     
        df_raw_kpi.columns.values[7] = 'DL User Throughput_Mbps [CDBH]'
    # df.columns.values[8] = 'DL User Throughput_Mbps [CDBH]'
    # # df.set_index("Short name",inplace=True)  ###for delete serial no in output
    # # df.fillna(value=0,inplace=True)
    # # print(df)
    # # df.to_excel('process_output/praice.xlsx',index=False)

        A = []
        B = []

        techlist = []
        for x in df_raw_kpi['Short name']:
            if("_" in str(x)):
                site_id = x.split("_")[-2][:-1]
                A.append(site_id)
            else:
                site_id = str(x)[:-1]
                A.append(site_id)
            if("_" in str(x)):
                lnbts_name = x.split(' ')[0][:-3]
                B.append(lnbts_name)
            else:
                lnbts_name = str(x)[:-3]
                B.append(lnbts_name)
            if  ('_F1_' in str(x) or '_F3_' in str(x) or '_F5_' in str(x) or '_T2_' in str(x)):
                    if ('_F1' in str(x)):
                        tech='L2100'
                        techlist.append(tech)

                    if ('_F3' in str(x)):
                        tech='L1800'
                        techlist.append(tech)

                    if ('_F5' in str(x)):
                        tech='L850'
                        techlist.append(tech)  

                    if ('_T2' in str(x)):
                        tech='L2300'
                        techlist.append(tech) 

            else:
                    tech=site_id
                    techlist.append(tech)                 


        df_raw_kpi.insert(0,'Site ID',A)
        df_raw_kpi.insert(1,'lnbts_name',B)
        df_raw_kpi.insert(2,'tech',techlist)



        print(df_raw_kpi)
        


    
        A1=[]
        B=[]
        C=[]
        # D=[]
        
        for num in df_raw_kpi['4G_ECGI']:
            if("-" in num):
                lnbts_id = num.split("-")[-2][:6]  
                A1.append(lnbts_id)
            else:
                lnbts_id = num[:-1]
                A1.append(lnbts_id)
            if("-" in num):
                lncell_id =num.split(" ")[0][7:]
                B.append(lncell_id)
            else:
                lncell_id = num[:-1]
                B.append(lncell_id)
            if("-" in num):
                mrbts_id = num.split("-")[-2][2:]
                C.append(mrbts_id)
            else:
                mrbts_id = num[:-1]
                C.append(mrbts_id)
            
            

        df_raw_kpi.insert(3,'lnbts_id',A1)
        df_raw_kpi.insert(4,'lncell_id',B)
        df_raw_kpi.insert(5,'Mrbts_id',C)
        
        
        print(df_raw_kpi)
        
        



        
        add=[]
        for i in df_raw_kpi['Mrbts_id']:
            ad=91
            add.append(ad)
        df_raw_kpi.insert(6,'add',add)
        df_raw_kpi['MRBTS_ID']=df_raw_kpi['add'].astype(str) + df_raw_kpi["Mrbts_id"]

        # message=site_comparision(A,list(df_site_list['2G ID'])) # site_comparision_call 
        message=site_comparision(A,df_site_list) # site_comparision_call 


        process_op_path=os.path.join(door_root,"process_output",'desired input.xlsx')
        
        
        # df_raw_kpi.to_excel(process_op_path)

        # path_of_blnk_temp=os.path.join(door_root,"template","mp_trend1.xlsx")
        # wb=load_workbook(path_of_blnk_temp)
        # print(wb.sheetnames)

        df_raw_kpi.to_excel(process_op_path, index=False)
        excel_file_1 = process_op_path

        df1 = pd.read_excel(excel_file_1)
        df1.rename(columns={"Site ID": "SITE_ID"}, inplace=True)

        kpii=['TA Sample <500M','TA Sample <1KM','TA Sample <1.5KM_Nokia_1','TA Sample <2KM','RSRP Samples<-104 dBm','RSRP Samples<-110 dBm'
       ,'RSRP Samples<-116 dBm']

        kpi = ['MV_4G Data Volume_GB', 'E-UTRAN Average CQI [CDBH]', 'Average UE Distance_KM',
            'DL User Throughput_Mbps [CDBH]', 'UL User Throughput_Kbps [CDBH]', 'Average number of used DL PRBs [CDBH]',
        'RRC Setup Success Rate [CDBH]', 'ERAB Setup Success Rate [CDBH]',
        'MV_PS Drop Call Rate % [CDBH]', 'UL RSSI [CDBH]', 'MV_CSFB Redirection Success Rate [CDBH]',
        'MV_LTE_PS_DCR', 'PS handover success rate [LTE Inter System] [CDBH]', 'PS handover success rate [LTE Intra System] [CDBH]',
        'VoLTE Traffic', 'VoLTE DCR [CBBH]',
        'VoLTE Packet Loss DL [CBBH]', 'VoLTE Packet Loss UL [CBBH]',
        'VoLTE InterF HOSR Exec [CBBH]', 'VoLTE IntraF HOSR Exec [CBBH]',
        'Radio NW Availability','TA Sample <500M','TA Sample <1KM','TA Sample <1.5KM_Nokia_1','TA Sample <2KM','RSRP Samples<-104 dBm','RSRP Samples<-110 dBm'
        ,'RSRP Samples<-116 dBm']

        # filtered_df_1 = df1[(df1.SITE_ID.isin(list(df_site_list['2G ID'])))]
        filtered_df_1 = df1[(df1.SITE_ID.isin(df_site_list))]


        print(filtered_df_1)
        # filtered_df_1.to_excel('process_output/filtered_df_1.xlsx')
        # df1 = pd.read_excel('process_output/filtered_df_1.xlsx')
        MpPath=os.path.join(door_root,'process_output','filtered1.xlsx')
        filtered_df_1.to_excel(MpPath,index=False)       
    # print(df)
    
        df1 = pd.read_excel(MpPath)
        df_pivot = df1.pivot_table(columns='Date', index=['Short name','SITE_ID','lnbts_id', 'lncell_id', 'lnbts_name',
            'tech','MRBTS_ID','4G_ECGI'])
        #  aggfunc=np.sum)
        # print('df_pivot')
        MpPathpivot=os.path.join(door_root,'process_output','pivot.xlsx')
        df_pivot.to_excel(MpPathpivot)
        # df_pivot.

        path_of_blnk_temp=os.path.join(door_root,"template","mp_trend1.xlsx")
        wb=load_workbook(path_of_blnk_temp)
        # print(wb.sheetnames)
        ws=wb.active

        alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
        def titleToNumber(s):
            result= 0
            for B in range(len(s)):
                result *= 26
                result += ord(s[B]) - ord('A') + 1
            return result  
            # result += ord(s[B]) - ord('A') + 1
            # return result

        str_date=request.POST.get("offered_date")
        date1 =datetime.datetime.strptime(str_date,'%Y-%m-%d')
        d1=date1-timedelta(1)
        d2=date1-timedelta(2)
        d3=date1-timedelta(3)
        d4=date1-timedelta(4)
        d5=date1-timedelta(5)
        cl=[d1,d2,d3,d4,d5]

        
        def overwrite(kpi_name,coln1):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            print(kpi_name)
            index=df_pivot.index
            # print("index ;###############################",index_pivot)
            # print(len(index_pivot))
            # 
            print(kpi_name)
            index=df_pivot.index
            print('donnnnnnnnnnnnnnnnne')
            print(len(index))
            dr=df_pivot[kpi_name]
            
            li=dr.columns

            col1=dr[li[0]].to_list()
            col2=dr[li[1]].to_list()
            col3=dr[li[2]].to_list()
            col4=dr[li[3]].to_list()
            col5=dr[li[4]].to_list()


            ws[coln1+"4"].value=cl[4]
            ws[coln2+"4"].value=cl[3]
            ws[coln3+"4"].value=cl[2]
            ws[coln4+"4"].value=cl[1]
            ws[coln5+"4"].value=cl[0]

            for i,value in enumerate(index):
                j=i+6
                ws['A'+str(j)].value='MP'
                ws['I'+str(j)].value='GOVINDPUR'
                ws['N'+str(j)].value='NOKIA'
                ws['O'+str(j)].value='Relocation'
                ws['L'+str(j)].value=date1

                ws['M'+str(j)].value='DONE'
                ws['B'+str(j)].value=index[i][1]
                ws['C'+str(j)].value=index[i][7]
                ws['D'+str(j)].value=index[i][6]
                ws['E'+str(j)].value=index[i][2]
                ws['F'+str(j)].value=index[i][3]
                ws['G'+str(j)].value=index[i][4]
                ws['H'+str(j)].value=index[i][0]
                ws['J'+str(j)].value=index[i][5]


                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]

        def overwrite_col(kpi_name,coln):
            index=df_pivot.index
            print('single celll')
            dr=df_pivot[kpi_name] 
            cl=dr.columns.to_list()
            data_list=dr[str(cl[0])].to_list()
            print(data_list)
            ws[coln+"4"].value=cl[0]

            for i,value in enumerate(index):
                j=i+6
        
                ws[coln+str(j)].value=data_list[i]
        

        for ta in kpii:
            if(ta=="TA Sample <500M"):
                overwrite_col(ta,"FZ") 
                
            if(ta=="TA Sample <1KM"):
                overwrite_col(ta,"GA")
                
            if(ta=="TA Sample <1.5KM_Nokia_1"):
                overwrite_col(ta,"GB") 

            if(ta=="TA Sample <2KM"):
                overwrite_col(ta,"GC") 

            if(ta=="RSRP Samples<-104 dBm"):
                overwrite_col(ta,"GE") 
            if(ta=="RSRP Samples<-110 dBm"):
                    overwrite_col(ta,"GF") 

            if(ta=="RSRP Samples<-116 dBm"):
                    overwrite_col(ta,"GG")                     




    # STR='l2100.xlsx'
    # wb = openpyxl.load_workbook(STR)
    # ws = wb["Sheet"]
        for kpi_name in kpi:
            if(kpi_name=='Radio NW Availability'):
                overwrite(kpi_name,'FB')

            if(kpi_name=='MV_4G Data Volume_GB'):
                overwrite(kpi_name,'EF')
            if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
                overwrite(kpi_name,'DR')

            if(kpi_name=='Average UE Distance_KM'):
                overwrite(kpi_name,'FH')

            if(kpi_name=='DL User Throughput_Mbps [CDBH]'):
                overwrite(kpi_name,'AL') 
                
            if(kpi_name=='UL User Throughput_Kbps [CDBH]'):
                overwrite(kpi_name,'AS') 

            if(kpi_name=='Average number of used DL PRBs [CDBH]'):
                overwrite(kpi_name,'EV') 

            if(kpi_name=='RRC Setup Success Rate [CDBH]'):
                overwrite(kpi_name,'Q') 

            if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
                overwrite(kpi_name,'X') 

            if(kpi_name=='MV_PS Drop Call Rate % [CDBH]'):
                overwrite(kpi_name,'AE')    

            if(kpi_name=='UL RSSI [CDBH]'):
                overwrite(kpi_name,'DY') 

            if(kpi_name=='MV_CSFB Redirection Success Rate [CDBH]'):
                overwrite(kpi_name,'BN') 

            if(kpi_name=='MV_LTE_PS_DCR'):
                overwrite(kpi_name,'FT')   

            if(kpi_name=='PS handover success rate [LTE Inter System] [CDBH]'):
                overwrite(kpi_name,'BG')  

            if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
                overwrite(kpi_name,'AZ')  

            if(kpi_name=='VoLTE Traffic'):
                overwrite(kpi_name,'EO')   

            if(kpi_name=='VoLTE DCR [CBBH]'):
                overwrite(kpi_name,'CB')  

            if(kpi_name=='VoLTE Packet Loss DL [CBBH]'):
                overwrite(kpi_name,'CI')  

            if(kpi_name=='VoLTE Packet Loss UL [CBBH]'):
                overwrite(kpi_name,'CP')    

            if(kpi_name=='VoLTE InterF HOSR Exec [CBBH]'):
                overwrite(kpi_name,'DK') 

            if(kpi_name=='VoLTE IntraF HOSR Exec [CBBH]'):
                overwrite(kpi_name,'DD') 
            if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
                overwrite(kpi_name,'BU') 

            if(kpi_name=='MV_CSFB Redirection Success Rate [CDBH]'):
                overwrite(kpi_name,'FN') 



        SaveOutput=os.path.join(door_root,'output','mptrendoutput.xlsx')        
        wb.save(SaveOutput) 
        download_path=os.path.join(MEDIA_URL,'trends','mp','output','mptrendoutput.xlsx')
        return Response({"status":True,"missing_sites":message,'Download_url':download_path}) 






@api_view(["POST"])
def old_mp2G_trend(request):
        
    raw_kpi_2G=request.FILES['raw_kpi_2G'] if 'raw_kpi_2G' in request.FILES else None
    if raw_kpi_2G:
        location=MEDIA_ROOT +r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file_2G=fs.save(raw_kpi_2G.name,raw_kpi_2G)
        file_2G_path=fs.path(file_2G)
        df_raw_kpi_2G=pd.read_excel(file_2G_path)
        print(df_raw_kpi_2G)
        os.remove(path=file_2G_path)

    required_cols = ['Drop Call Rate_Denom [BBH]', 'Total Voice Traffic [BBH]',
        'Average Traffic AMR Half Rate [BBH]', 'SDCCH Drop Call Rate [BBH]', 'TCH Drop Call Rate[BBH]',
        'SDCCH Blocking Rate [BBH]', 'TCH Blocking Rate [BBH]','TCH Availability[BBH]',
        'TCH Assignment Success Rate [BBH]', 'Handover Success Rate [BBH]', 'RX Quality [BBH]',
        'Total Data Traffic [BBH]', 'RX UL Quality [BBH]', 'RX DL Quality [BBH]',
        'TCH_TR_FAIL [BBH]']
    sts,response=required_col_check(raw_kpi_2G,required_cols)
    if sts:
        return Response(response)

    
    site_list_2G=request.FILES['site_list_2G'] if 'site_list_2G' in request.FILES else None
    if site_list_2G:
        location=MEDIA_ROOT+r'\trends\temporary_files'
        fs=FileSystemStorage(location=location)
        file_2G_sheet=fs.save(site_list_2G.name,site_list_2G)
        file_2G_path=fs.path(file_2G_sheet)
        df_2G_site=pd.read_excel(file_2G_path)
        print(df_2G_site)
        os.remove(path=file_2G_path) 

    if site_list_2G:
        sts,response=required_col_check(raw_kpi_2G,required_cols)
        if sts:
            return Response(response)

    
        door_root=os.path.join(MEDIA_ROOT,'trends','mp','MP_2G')
      

        path_of_blnk_temp=os.path.join(door_root,"template","GSM_TREND.xlsx")
        wb=openpyxl.load_workbook(path_of_blnk_temp)
        ws=wb.active

        df_raw_kpi_2G["Short name"] = df_raw_kpi_2G["Short name"].fillna(method="ffill")  #######for (forward fill)ffill is used to copy  and fill
        print(df_raw_kpi_2G.head())
        print(df_raw_kpi_2G.columns)
        # df.to_excel('process_output/mp.xlsx')
        df_raw_kpi_2G.columns.values[1] = 'Date'  #####for empty colum
       
        A=[]
        B=[]
        
         
        for x in df_raw_kpi_2G['Short name']:
       
            if(" " in str(x)):
                site_id = x.split(" ")[0][:-1]
                B.append(site_id)
            else:
                site_id = x.split(" ")[0][:-1]
                B.append(site_id)
        df_raw_kpi_2G.insert(2,'SITE_ID',B)
        # df_raw_kpi_2G.to_excel('process_output/mptool.xlsx',index=False)
        message=site_comparision(B,list(df_2G_site['SITE_ID'])) # site_comparision_call

        mpPath=os.path.join(door_root,'process_output','final.xlsx')
        df_raw_kpi_2G.to_excel(mpPath,index=False)
        
        excel_file_1 = mpPath 
        df1 = pd.read_excel(excel_file_1) 

        df1.rename(columns={"Site ID": "SITE_ID"}, inplace=True)
        df_2G_site.rename(columns={"Site ID": "SITE_ID"}, inplace=True) 


        kpi = ['Drop Call Rate_Denom [BBH]', 'Total Voice Traffic [BBH]',
        'Average Traffic AMR Half Rate [BBH]', 'SDCCH Drop Call Rate [BBH]', 'TCH Drop Call Rate[BBH]',
        'SDCCH Blocking Rate [BBH]', 'TCH Blocking Rate [BBH]','TCH Availability[BBH]',
        'TCH Assignment Success Rate [BBH]', 'Handover Success Rate [BBH]', 'RX Quality [BBH]',
        'Total Data Traffic [BBH]', 'RX UL Quality [BBH]', 'RX DL Quality [BBH]',
        'TCH_TR_FAIL [BBH]']
        filtered_df_1 = df1[(df1.SITE_ID.isin(list(df_2G_site['SITE_ID'])))]

        print(filtered_df_1)
        mp_filtr=os.path.join(door_root,'process_output','filtered_df_1.xlsx')
        filtered_df_1.to_excel(mp_filtr,index=False)
        df1 = pd.read_excel(mp_filtr)

        try:
            print('__________________________TRY_________________________________',raw_kpi_2G)
        except:
            print('an exceptional error')

        df_pivot = df1.pivot_table(columns='Date', index=['Short name', 'SITE_ID'])
        # aggfunc=np.sum)
        print('df_pivot')
   
        mp_pivot=os.path.join(door_root,'process_output','pivot.xlsx')
        df_pivot.to_excel(mp_pivot)
        
        

        alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]



        def titleToNumber(s):
    # This process is similar to binary-to-
    # decimal conversion
            result = 0
            for B in range(len(s)):
                result *= 26
                result += ord(s[B]) - ord('A') + 1
            return result
        
        str_date=request.POST.get("offered_date")
        date1=datetime.datetime.strptime(str_date,"%Y-%m-%d")
        d1=date1-timedelta(1)
        d2=date1-timedelta(2)
        d3=date1-timedelta(3)
        d4=date1-timedelta(4)
        d5=date1-timedelta(5)
        cl=[d1,d2,d3,d4,d5]

        def overwrite(kpi_name,coln1):
            coln2=num_hash(titleToNumber(coln1)+1)
            coln3=num_hash(titleToNumber(coln1)+2)
            coln4=num_hash(titleToNumber(coln1)+3)
            coln5=num_hash(titleToNumber(coln1)+4)
            print(kpi_name)
            index=df_pivot.index
            # print("index ;###############################",index_pivot)
            # print(len(index_pivot))
            print("doooooonnenn")
            dr=df_pivot[kpi_name]
            print("columns of dr tab")
            li=dr.columns
            print("column list")


            col1=dr[li[0]].to_list()
            col2=dr[li[1]].to_list()
            col3=dr[li[2]].to_list()
            col4=dr[li[3]].to_list()
            col5=dr[li[4]].to_list()


            ws[coln1+"2"].value=cl[4]
            ws[coln2+"2"].value=cl[3]
            ws[coln3+"2"].value=cl[2]
            ws[coln4+"2"].value=cl[1]
            ws[coln5+"2"].value=cl[0]

            for i,value in enumerate(index):
                j=i+3

                ws['A'+str(j)].value=index[i][0]
                ws['B'+str(j)].value=index[i][1]

                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]



        for kpi_name in kpi:
            if(kpi_name=='Drop Call Rate_Denom [BBH]'):
                overwrite(kpi_name,'M')

            if(kpi_name=='Total Voice Traffic [BBH]'):
                overwrite(kpi_name,'S')
            if(kpi_name=='Average Traffic AMR Half Rate [BBH]'):
                overwrite(kpi_name,'Y')

            if(kpi_name=='SDCCH Drop Call Rate [BBH]'):
                overwrite(kpi_name,'AE')

            if(kpi_name=='TCH Drop Call Rate[BBH]'):
                overwrite(kpi_name,'AK') 
                print('ddddddd')
                
            if(kpi_name=='SDCCH Blocking Rate [BBH]'):
                overwrite(kpi_name,'AQ') 

            if(kpi_name=='TCH Blocking Rate [BBH]'):
                overwrite(kpi_name,'AW') 

            if(kpi_name=='TCH Assignment Success Rate [BBH]'):
                overwrite(kpi_name,'BC') 

            if(kpi_name=='Handover Success Rate [BBH]'):
                overwrite(kpi_name,'BI') 

            if(kpi_name=='RX Quality [BBH]'):
                overwrite(kpi_name,'BO')    

            if(kpi_name=='TCH Availability[BBH]'):
                overwrite(kpi_name,'BU') 

            if(kpi_name=='Total Data Traffic [BBH]'):
                overwrite(kpi_name,'CA') 

            if(kpi_name=='RX UL Quality [BBH]'):
                overwrite(kpi_name,'CG') 

            if(kpi_name=='RX DL Quality [BBH]'):
                overwrite(kpi_name,'CM')

            if(kpi_name=='TCH_TR_FAIL [BBH]'):
                overwrite(kpi_name,'CS')


    SaveOutput=os.path.join(MEDIA_ROOT,'trends','mp','MP_2G','output','MP2G.xlsx')
    wb.save(SaveOutput)
  
    download_path=os.path.join(MEDIA_URL,'trends','mp','MP_2G','output','MP2G.xlsx')
    return Response({"status":True,"missing_sites":message,'Download_url':download_path,"message":"successfully uploaded"})





    

    


   



    






        



    















