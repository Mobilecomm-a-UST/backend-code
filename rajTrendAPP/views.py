import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from rest_framework.status import HTTP_200_OK
import re

##----------------- NEW CODE --------------------------------------------------------------


# Base folders---------
base_media_url = os.path.join(MEDIA_ROOT, "rajTrend")
template_dir = os.path.join(base_media_url, "template")
output_path = os.path.join(base_media_url, "OUTPUT")

os.makedirs(template_dir, exist_ok=True)
os.makedirs(output_path, exist_ok=True)

#-----------------------API-----------------
@api_view(['POST'])
def KpiTrend_4g(request):
    file = request.FILES.get('file')
    site_id = request.data.get('site_id')
    circle = request.data.get('circle')

    if not file:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    if not site_id:
        return Response({'error': 'No site_id provided'}, status=status.HTTP_400_BAD_REQUEST)

        # Normalize site_id to list
    if isinstance(site_id, str):
        site_id = re.split(r"[\s,]+", site_id.strip())
        site_id = [s.strip() for s in site_id if s.strip()]
        print("Processed site_id as string:", site_id)    
    elif isinstance(site_id, list):
        if len(site_id) == 1 and isinstance(site_id[0], str):
            site_id = re.split(r"[\s,]+", site_id[0].strip())
            site_id = [s.strip() for s in site_id if s.strip()]
        print("Processed site_id as list:", site_id)

    #save site id into df
    site_id_df = pd.DataFrame(site_id, columns=["Site ID"])
    site_id_df['Site ID'] =site_id_df['Site ID'].str.replace("-", "", regex=False)
    print("Processed site_id_df:", site_id_df)

    # Read main Excel
    if file.name.endswith('.csv'):
        df = pd.read_csv(file)
    else:
        df = pd.read_excel(file)
    print(df)
    df.columns = df.columns.str.replace(r'\t+', '', regex=True).str.strip()
    df['Short name'] = df['Short name'].ffill()    
    if 'Unnamed: 1' in df.columns:
        df.rename(columns={'Unnamed: 1': 'date'}, inplace=True)
    
    df['SITE_ID'] =df['Short name'].apply(lambda x: x.split("_")[-2][:-1] if "_" in x else x)

    
    df = df[df["SITE_ID"].isin(site_id_df['Site ID'])]
    if df.empty:
        return Response({'error': 'No matching site_id found in the file'}, status=status.HTTP_400_BAD_REQUEST)
    print(df)
    # KPI list
    kpi1 = [
        'Radio NW Availability', 'E-UTRAN Average CQI [CDBH]', 'DL User Throughput_Kbps [CDBH]', 'UL RSSI [CDBH]',
        'PS Drop Call Rate % [CDBH]', 'RRC Setup Success Rate [CDBH]', 'ERAB Setup Success Rate [CDBH]',
        'PS handover success rate [LTE Intra System] [CDBH]', 'MV_Average number of used DL PRBs [CDBH]',
        'MV_UL User Throughput_Kbps [CDBH]', 'MV_VoLTE Traffic', 'MV_VoLTE Packet Loss DL [CBBH]',
        'MV_VoLTE Packet Loss UL [CBBH]', 'MV_VoLTE DCR [CBBH]', 'MV_CSFB Redirection Success Rate',
        'MV_VoLTE ERAB Setup Success Rate', 'MV_4G Data Volume_GB',
        'PS handover success rate [LTE Inter System] [CDBH]', 'VoLTE InterF HOSR Exec [CBBH]',
        'VoLTE IntraF HOSR Exec [CBBH]', 'PS Drop Call Rate % [CBBH]', 'VoLTE DCR_Nom [CBBH]', 'CA Data Volume [MB]'
    ]

    # Convert KPI columns to numeric (invalid → 0)
    for col in kpi1:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Convert throughput
    if 'DL User Throughput_Kbps [CDBH]' in df.columns:
        df["DL User Throughput_Mbps [CDBH]"] = df["DL User Throughput_Kbps [CDBH]"] / 1024
        df.drop(columns=['DL User Throughput_Kbps [CDBH]'], inplace=True)

    # Site ID & Tech extraction
    sn, techlist = [], []
    for cell in df['Short name']:
        if "_" in cell:
            sit_id = cell.split("_")[-2][:-1]
        else:
            sit_id = cell[:-1]
        sn.append(sit_id)

        if '_F1_' in cell:
            tech = 'L2100'
        elif '_F3_' in cell:
            tech = 'L1800'
        elif '_F8_' in cell:
            tech = 'L900'
        elif '_T1_' in cell or '_T2_' in cell:
            tech = 'L2300'
        else:
            tech = sit_id
        techlist.append(tech)

    df.insert(0, 'Site ID', sn)
    df.insert(4, 'Tech', techlist)
    df["Short name"] = df["Short name"].apply(lambda x: x.strip())
    df.fillna(0, inplace=True)

    # Pivot table
    kpi_cols = [col for col in df.columns if col in kpi1 or col == 'DL User Throughput_Mbps [CDBH]']
    df_pivot = df.pivot_table(values=kpi_cols, columns='date', index=['Short name', 'Site ID', 'Tech'])

    # Load template
    template_path = os.path.join(template_dir, 'update raj trend.xlsx')
    if not os.path.exists(template_path):
        return Response({'error': 'Template file not found'}, status=status.HTTP_400_BAD_REQUEST)

    wb = load_workbook(template_path)
    ws = wb.active

    # Dates for last 5 days
    today = datetime.date.today()
    cl = [today - timedelta(days=i) for i in range(1, 6)]

    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
        if num < 26:
            return alpha[num - 1]
        else:
            q, r = divmod(num, 26)
            if r == 0:
                return num_hash(q - 1) + alpha[-1]
            else:
                return num_hash(q) + alpha[r - 1]

    def titleToNumber(s):
        result = 0
        for B in s:
            result = result * 26 + (ord(B) - ord('A') + 1)
        return result

    def overwrite(kpi_name, coln1):
        if kpi_name not in df_pivot.columns.levels[0]:
            return
        col_letters = [num_hash(titleToNumber(coln1) + i) for i in range(5)]
        dr = df_pivot[kpi_name]
        li = dr.columns
        if len(li) < 5:
            return
        col_values = [dr[li[i]].to_list() for i in range(5)]

        # Header dates
        for idx, c in enumerate(col_letters):
            ws[f"{c}4"].value = cl[4 - idx]

        # Fill data rows
        for i, value in enumerate(dr.index):
            j = i + 5
            ws['N' + str(j)].value = 'CAPACITY'
            ws['M' + str(j)].value = 'ERICSSON'
            ws['K' + str(j)].value = today
            ws['L' + str(j)].value = 'DONE'
            ws['A' + str(j)].value = circle
            ws['B' + str(j)].value = value[1]
            ws['C' + str(j)].value = value[0]
            ws['G' + str(j)].value = value[0]
            ws['I' + str(j)].value = value[2]
            ws['F' + str(j)].value = value[0]
            for idx, c in enumerate(col_letters):
                ws[f"{c}{j}"].value = col_values[idx][i]

    kpi_column_map = {
        'RRC Setup Success Rate [CDBH]': 'AN',
        'ERAB Setup Success Rate [CDBH]': 'AU',
        'PS Drop Call Rate % [CDBH]': 'BB',
        'DL User Throughput_Mbps [CDBH]': 'BI',
        'MV_UL User Throughput_Kbps [CDBH]': 'BP',
        'PS handover success rate [LTE Intra System] [CDBH]': 'BW',
        'PS handover success rate [LTE Inter System] [CDBH]': 'CD',
        'MV_CSFB Redirection Success Rate': 'CK',
        'MV_VoLTE ERAB Setup Success Rate': 'CR',
        'MV_VoLTE DCR [CBBH]': 'CY',
        'MV_VoLTE Packet Loss DL [CBBH]': 'DF',
        'MV_VoLTE Packet Loss UL [CBBH]': 'DM',
        'VoLTE IntraF HOSR Exec [CBBH]': 'EA',
        'VoLTE InterF HOSR Exec [CBBH]': 'EH',
        'E-UTRAN Average CQI [CDBH]': 'EO',
        'UL RSSI [CDBH]': 'EV',
        'MV_4G Data Volume_GB': 'FC',
        'MV_VoLTE Traffic': 'FL',
        'MV_Average number of used DL PRBs [CDBH]': 'FS',
        'Radio NW Availability': 'FY',
        'VoLTE DCR_Nom [CBBH]': 'GE',
        'CA Data Volume [MB]': 'GK',
        'pmEndcSetupUeAtt': 'GW',
        'pmEndcSetupUeSucc': 'HC'
    }

    for kpi_name, col_letter in kpi_column_map.items():
        overwrite(kpi_name, col_letter)

    output_file = os.path.join(output_path, '4G_Rajasthantrend.xlsx')
    wb.save(output_file)
    download_link = request.build_absolute_uri(MEDIA_URL + 'rajTrend/OUTPUT/4G_Rajasthantrend.xlsx') 
    print(f"Output file created: {output_path}")
    return Response(
        {
        "status": True,
        "message": "Files processed successfully",
        "download_url": download_link,
        },
        status=HTTP_200_OK,
    )



@api_view(['POST'])
def KpiTrend_2g(request):
    file = request.FILES.get('file')
    site_id = request.data.get('site_id')
    circle = request.data.get('circle', '')

    if not file:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    if not site_id:
        return Response({'error': 'No site_id provided'}, status=status.HTTP_400_BAD_REQUEST)

    # Normalize site_id to list
    if isinstance(site_id, str):
        site_id = re.split(r"[\s,]+", site_id.strip())
        site_id = [s.strip() for s in site_id if s.strip()]
        print("Processed site_id as string:", site_id)

    elif isinstance(site_id, list):
        if len(site_id) == 1 and isinstance(site_id[0], str):
            site_id = re.split(r"[\s,]+", site_id[0].strip())
            site_id = [s.strip() for s in site_id if s.strip()]
        print("Processed site_id as list:", site_id)
    else:
        return Response({'error': 'Invalid site_id format'}, status=status.HTTP_400_BAD_REQUEST)

    #save site id into df
    site_id_df = pd.DataFrame(site_id, columns=["Site ID"])
    site_id_df['Site ID'] = site_id_df['Site ID'].str.replace("-", "", regex=False)

    # Read uploaded file
    try:
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
    except Exception as e:
        return Response({'error': f'File reading failed: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

    # Fill and rename columns
    df["Short name"] = df["Short name"].fillna(method="ffill")
    df.rename(columns={'Unnamed: 1': 'Date'}, inplace=True)

    df['Date'] = pd.to_datetime(df['Date'], errors="coerce")
    unique_date = df['Date'].dropna().sort_values().unique()

    if len(unique_date) < 5:
        return Response({'error': 'Not enough unique dates in file'}, status=status.HTTP_400_BAD_REQUEST)

    # Get last 5 dates
    cl = [pd.to_datetime(val).strftime('%Y-%m-%d') for val in unique_date[-5:]]

    # Extract SITE_ID
    df['SITE_ID'] = [site[:-1] if isinstance(site, str) else '' for site in df['Short name']]
    df = df[df["SITE_ID"].isin(site_id_df['Site ID'])]

    if df.empty:
        return Response({'error': 'No matching site_id found in the file'}, status=status.HTTP_400_BAD_REQUEST)

    df.fillna(value=0, inplace=True)

    # Filter for required sites and dates
    df_filter = df[(df['SITE_ID'].isin(site_id_df['Site ID'])) & (df['Date'].isin(pd.to_datetime(cl)))]

    gsm_kpi = [
        "SDCCH Blocking Rate [BBH]",
        "TCH Blocking Rate [BBH]",
        "TCH Drop Call Rate [BBH]",
        "Handover Success Rate [BBH]",
        "Total Voice Traffic [BBH]",
        "Network availability [RNA]",
        "SDCCH Drop Call Rate_Nom [BBH]",
        "TCH Drop Call Rate_Nom [BBH]",
        "SDCCH Drop Call Rate [BBH]",
    ]

    # Pivot table
    try:
        df_pivot = df_filter.pivot_table(
            values=gsm_kpi, columns='Date', index=['SITE_ID', 'Short name']
        )
    except Exception as e:
        return Response({'error': f'Pivoting failed: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

    # Load template
    # template_dir = os.path.join(settings.BASE_DIR, 'templates')
    template_path = os.path.join(template_dir, '2G_template.xlsx')
    if not os.path.exists(template_path):
        return Response({'error': 'Template file not found'}, status=status.HTTP_400_BAD_REQUEST)

    wb = load_workbook(template_path)
    ws = wb.active
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    def num_hash(num):
        if num < 26:
            return alpha[num - 1]
        else:
            q, r = divmod(num, 26)
            if r == 0:
                if q == 1:
                    return alpha[r - 1]
                else:
                    return num_hash(q - 1) + alpha[r - 1]
            else:
                return num_hash(q) + alpha[r - 1]

    def titleToNumber(s):
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result

    def overwrite(kpi_name, coln1):
        coln2 = num_hash(titleToNumber(coln1) + 1)
        coln3 = num_hash(titleToNumber(coln1) + 2)
        coln4 = num_hash(titleToNumber(coln1) + 3)
        coln5 = num_hash(titleToNumber(coln1) + 4)

        index = df_pivot.index
        dr = df_pivot[kpi_name]

        li = dr.columns
        col1 = dr[li[0]].to_list()
        col2 = dr[li[1]].to_list()
        col3 = dr[li[2]].to_list()
        col4 = dr[li[3]].to_list()
        col5 = dr[li[4]].to_list()

        ws[coln1 + "3"].value = cl[4]
        ws[coln2 + "3"].value = cl[3]
        ws[coln3 + "3"].value = cl[2]
        ws[coln4 + "3"].value = cl[1]
        ws[coln5 + "3"].value = cl[0]

        for i, value in enumerate(index):
            j = i + 4
            ws['L' + str(j)].value = 'CAPACITY'
            ws['K' + str(j)].value = 'ERICSSON'
            ws['G' + str(j)].value = 'G900'
            ws['A' + str(j)].value = circle
            ws['I' + str(j)].value = cl[-1]   # last date
            ws['B' + str(j)].value = index[i][0]
            ws['C' + str(j)].value = index[i][1]
            ws['D' + str(j)].value = index[i][0]
            ws['E' + str(j)].value = index[i][0]

            ws[coln1 + str(j)].value = col1[i]
            ws[coln2 + str(j)].value = col2[i]
            ws[coln3 + str(j)].value = col3[i]
            ws[coln4 + str(j)].value = col4[i]
            ws[coln5 + str(j)].value = col5[i]

    # Map KPIs to Excel columns
    mapping = {
        'SDCCH Blocking Rate [BBH]': 'O',
        'TCH Blocking Rate [BBH]': 'AC',
        'TCH Drop Call Rate [BBH]': 'AJ',
        'Handover Success Rate [BBH]': 'AQ',
        'Total Voice Traffic [BBH]': 'AX',
        'Network availability [RNA]': 'BE',
        'SDCCH Drop Call Rate_Nom [BBH]': 'BL',
        'TCH Drop Call Rate_Nom [BBH]': 'BR',
        'SDCCH Drop Call Rate [BBH]': 'V',
    }

    for kpi_name, coln in mapping.items():
        if kpi_name in df_pivot.columns.levels[0]:
            overwrite(kpi_name, coln)

    output_file = os.path.join(output_path, '2G_RajasthanTrend.xlsx')
    wb.save(output_file)

    download_link = request.build_absolute_uri(MEDIA_URL + 'rajTrend/OUTPUT/2G_RajasthanTrend.xlsx') 
    print(f"Output file created: {output_path}")
    return Response(
        {
        "status": True,
        "message": "Files processed successfully",
        "download_url": download_link,
        },
        status=HTTP_200_OK,
    )




################################### old code ############################
# from django.shortcuts import render

# # Create your views here.
# from django.shortcuts import render
# from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
# from django.core.files.storage import FileSystemStorage
# from rest_framework.decorators import api_view
# from datetime import date, timedelta
# import datetime 
# from rest_framework.response import Response
# import pandas as pd
# import openpyxl
# import os
# from commom_utilities.utils import *

# @api_view(['POST'])
    
# def old_raj_trend(request):
#     kpi = ['Radio NW Availability', 'E-UTRAN Average CQI [CDBH]', 'DL User Throughput_Mbps [CDBH]', 'UL RSSI [CDBH]',
#         'PS Drop Call Rate % [CDBH]', 'RRC Setup Success Rate [CDBH]', 'ERAB Setup Success Rate [CDBH]',
#         'PS handover success rate [LTE Intra System] [CDBH]', 'MV_Average number of used DL PRBs [CDBH]',
#         'MV_UL User Throughput_Kbps [CDBH]', 'MV_VoLTE Traffic', 'MV_VoLTE Packet Loss DL [CBBH]',
#         'MV_VoLTE Packet Loss UL [CBBH]', 'MV_VoLTE DCR [CBBH]', 'MV_CSFB Redirection Success Rate',
#         'MV_VoLTE ERAB Setup Success Rate', 'MV_4G Data Volume_GB',
#         'PS handover success rate [LTE Inter System] [CDBH]', 'VoLTE InterF HOSR Exec [CBBH]',
#         'VoLTE IntraF HOSR Exec [CBBH]', 'PS Drop Call Rate % [CBBH]','VoLTE DCR_Nom [CBBH]','CA Data Volume [MB]']
    
#     kpi1 = ['Radio NW Availability', 'E-UTRAN Average CQI [CDBH]', 'DL User Throughput_Kbps [CDBH]', 'UL RSSI [CDBH]',
#         'PS Drop Call Rate % [CDBH]', 'RRC Setup Success Rate [CDBH]', 'ERAB Setup Success Rate [CDBH]',
#         'PS handover success rate [LTE Intra System] [CDBH]', 'MV_Average number of used DL PRBs [CDBH]',
#         'MV_UL User Throughput_Kbps [CDBH]', 'MV_VoLTE Traffic', 'MV_VoLTE Packet Loss DL [CBBH]',
#         'MV_VoLTE Packet Loss UL [CBBH]', 'MV_VoLTE DCR [CBBH]', 'MV_CSFB Redirection Success Rate',
#         'MV_VoLTE ERAB Setup Success Rate', 'MV_4G Data Volume_GB',
#         'PS handover success rate [LTE Inter System] [CDBH]', 'VoLTE InterF HOSR Exec [CBBH]',
#         'VoLTE IntraF HOSR Exec [CBBH]', 'PS Drop Call Rate % [CBBH]','VoLTE DCR_Nom [CBBH]','CA Data Volume [MB]',]
    
#     raw_kpi_4G=request.FILES["raw_kpi"] if "raw_kpi" in request.FILES else None
#     df=pd.read_excel(raw_kpi_4G)
    
#     required_cols = ['Radio NW Availability', 'E-UTRAN Average CQI [CDBH]', 'DL User Throughput_Kbps [CDBH]', 'UL RSSI [CDBH]',
#         'PS Drop Call Rate % [CDBH]', 'RRC Setup Success Rate [CDBH]', 'ERAB Setup Success Rate [CDBH]',
#         'PS handover success rate [LTE Intra System] [CDBH]', 'MV_Average number of used DL PRBs [CDBH]',
#         'MV_UL User Throughput_Kbps [CDBH]', 'MV_VoLTE Traffic', 'MV_VoLTE Packet Loss DL [CBBH]',
#         'MV_VoLTE Packet Loss UL [CBBH]', 'MV_VoLTE DCR [CBBH]', 'MV_CSFB Redirection Success Rate',
#         'MV_VoLTE ERAB Setup Success Rate', 'MV_4G Data Volume_GB',
#         'PS handover success rate [LTE Inter System] [CDBH]', 'VoLTE InterF HOSR Exec [CBBH]',
#         'VoLTE IntraF HOSR Exec [CBBH]', 'PS Drop Call Rate % [CBBH]','VoLTE DCR_Nom [CBBH]','CA Data Volume [MB]',]
#     sts,response=required_col_check(raw_kpi_4G,required_cols)
#     if sts:
#         return Response(response)
   
#     # os.remove(path=file_path)
   

# ################################## for site_list ############################################
#     # excel_site_list_raj=request.FILES["site_list"] if "site_list" in request.FILES else None
#     # df_site_list=pd.read_excel(excel_site_list_raj)
   
#     # if excel_site_list_raj:
#     #     sts,response=required_col_check(raw_kpi_4G,required_cols)
#     #     if sts:
#     #         return Response(response)
#     response,s_l= site_list_handler(request)
#     if s_l:
#         df_site_list=s_l
#         print("_________________________--print",type(df_site_list[0]))

#         print(df_site_list,"_______________")
#         if response:
#             return Response(response)
   

# ######################################################################        
#     door_raj_path=os.path.join(MEDIA_ROOT,'trends','raj')
#     # save_raj_path=os.path.join(door_raj_path,"actual input/RAJ_TOOL_KPI.xlsx")
#     # df = pd.read_excel(save_raj_path)
   
    
#     STR=os.path.join(door_raj_path,'template','update raj trend.xlsx')
#     wb=openpyxl.load_workbook(STR)
#     ws=wb.active

    
    
#     for x in kpi1:
#             df[x] = df[x].replace(to_replace='.*', value=0, regex=True)

#     df["Short name"] = df["Short name"].fillna(method="ffill")  #######for (forward fill)ffill is used to copy  and fill
#     df.columns.values[1] = 'Date'  #####for empty column

#     df["DL User Throughput_Kbps [CDBH]"] = (df["DL User Throughput_Kbps [CDBH]"]/1024)  #####for change kbps to mbps
#     df.columns.values[4] = 'DL User Throughput_Mbps [CDBH]'
#     # print(df)

#     sn=[]
#     techlist=[]
#     strip=[]

#     for cell in df['Short name']:
#         if("_" in cell):
#             sit_id = cell.split("_")[-2][:-1]
#             sn.append(sit_id)
        

#         else:
#             sit_id=cell[:-1]
#             sn.append(sit_id)
        
#         if('_F1_' in cell or '_F3_' in cell or '_F8_' in cell  or '_T1' in cell or '_T2_' in cell ):  
#             if('_F1_' in cell):
#                 tech='L2100'
#                 # techlist.append(tech)
#             if('_F3_' in cell):
#                 tech='L1800'
#                 # techlist.append(tech) 
#             if('_f8_' in cell):
#                 tech='L900'    
#                 # techlist.append(tech)
#             if('_T1_' in cell or '_T2_' in cell):
#                 tech='L2300'
#             techlist.append(tech)
                            
#         else:
#             cell=sit_id
#             techlist.append(cell)
        
#     # if ('' in cell):
#     #     txt=cell.strip()
#     #     strip.append(txt)
#     # else:
#     #     txt=cell()
#     #     strip.append(txt)      
        
#     # print(techlist)
#     # df.insert(2,'site type',sn)
#     df.insert(0, 'Site ID', sn)
#     df.insert(4,'Tech',techlist)  
#     # df.insert(5,'strip',strip)

#     df["Short name"] =df["Short name"].apply(lambda x: x.strip())  
#     # print("dfdfff",df_site_list['2G ID'])
#     # message=site_comparision(sn,list(df_site_list['2G ID']))# site_comparision_call 
#     message=site_comparision(sn,df_site_list)# site_comparision_call 


#     df.fillna(value=0,inplace=True)  
    
#     raj_1st_step=os.path.join(door_raj_path,'process output','desired.xlsx')
#     df.to_excel(raj_1st_step, index=False)
#     print(df)
#     excel_file_1 = raj_1st_step
#     # raj_2nd_step=os.path.join(door_raj_path,'project file','site.xlsx')
#     # excel_file_2 = raj_2nd_step

#     df1 = pd.read_excel(excel_file_1)
#     # df2 = pd.read_excel(excel_file_2)

#     df1.rename(columns={"Site ID": "Site_ID"}, inplace=True)
#     # df_site_list.rename(columns={"Site ID": "Site_ID"}, inplace=True)

   

#     filtered_df_1 = df1[(df1.Site_ID.isin(df_site_list))]

#     print(filtered_df_1)
#     raj_filtered_3rd_step=os.path.join(door_raj_path,'process output','filtered_df_1.xlsx')
#     filtered_df_1.to_excel(raj_filtered_3rd_step, index=False)

#     df1 = pd.read_excel(raj_filtered_3rd_step)
#     df_pivot = df1.pivot_table(values=kpi, columns='Date', index=['Short name', 'Site_ID','Tech'])
    
    
#     raj_pivot_4th_step=os.path.join(door_raj_path,'process output','pivot.xlsx')
#     df_pivot.to_excel(raj_pivot_4th_step)
    
#     alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
#     def num_hash(num):
#             if num < 26:
#                 return alpha[num-1]
#             else:
#                 q, r = num//26, num % 26
#                 if r == 0:
#                     if q == 1:
#                         return alpha[r-1]
#                     else:
#                         return num_hash(q-1) + alpha[r-1]
#                 else:
#                     return num_hash(q) + alpha[r-1]
            
            
#     def titleToNumber(s):
# 	# This process is similar to binary-to-
# 	# decimal conversion
#         result = 0
#         for B in range(len(s)):
#             result *= 26
#             result += ord(s[B]) - ord('A') + 1
#         return result
#     str_date=request.POST.get('offered_date')
#     date1 = datetime.datetime.strptime(str_date, '%Y-%m-%d')
#     print("Date-------------------------------",date1)
#     # date1=cal.get_date()
#     # date1=date(2023,3,2)
#     d1=date1-timedelta(1)
#     d2=date1-timedelta(2)
#     d3=date1-timedelta(3)
#     d4=date1-timedelta(4)
#     d5=date1-timedelta(5)
#     cl=[d1,d2,d3,d4,d5]
   
#     def overwrite(kpi_name,coln1):
#         coln2=num_hash(titleToNumber(coln1)+1)
#         coln3=num_hash(titleToNumber(coln1)+2)
#         coln4=num_hash(titleToNumber(coln1)+3)
#         coln5=num_hash(titleToNumber(coln1)+4)
#         print(kpi_name)
#         index=df_pivot.index
#         print('donnnnnnnnnnnnnnnnne')
#         print(len(index))
#         dr=df_pivot[kpi_name]
        
#         li=dr.columns
#         col1=dr[li[0]].to_list()
#         col2=dr[li[1]].to_list()
#         col3=dr[li[2]].to_list()
#         col4=dr[li[3]].to_list()
#         col5=dr[li[4]].to_list()


#         ws[coln1+"4"].value=cl[4]
#         ws[coln2+"4"].value=cl[3]
#         ws[coln3+"4"].value=cl[2]
#         ws[coln4+"4"].value=cl[1]
#         ws[coln5+"4"].value=cl[0]



#         for i,value in enumerate(index):
#                 j=i+6
#                 ws['N'+str(j)].value='CAPACITY'
#                 ws['M'+str(j)].value='ERICSSON'
#                 ws['K'+str(j)].value=date1

#                 ws['L'+str(j)].value='DONE'
#                 ws['B'+str(j)].value=index[i][1]
#                 ws['C'+str(j)].value=index[i][0]
#                 ws['G'+str(j)].value=index[i][0]
#                 ws['I'+str(j)].value=index[i][2]
                
#                 # ws['D'+str(j)].value=index[i][2]
#                 # ws['E'+str(j)].value=index[i][2]
#                 ws['F'+str(j)].value=index[i][0]     
#                 ws[coln1+str(j)].value=col1[i]
#                 ws[coln2+str(j)].value=col2[i]
#                 ws[coln3+str(j)].value=col3[i]
#                 ws[coln4+str(j)].value=col4[i]
#                 ws[coln5+str(j)].value=col5[i]
#     for kpi_name in kpi:
#         if(kpi_name=='RRC Setup Success Rate [CDBH]'):
#             overwrite(kpi_name,'AN')

#         if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
#             overwrite(kpi_name,'AU') 

#         if(kpi_name=='PS Drop Call Rate % [CDBH]'):
#             overwrite(kpi_name,'BB')

#         if(kpi_name=='DL User Throughput_Mbps [CDBH]'):
#             overwrite(kpi_name,'BI')

#         if(kpi_name=='MV_UL User Throughput_Kbps [CDBH]'):
#             overwrite(kpi_name,'BP') 
            
#         if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
#             overwrite(kpi_name,'BW') 

#         if(kpi_name=='PS handover success rate [LTE Inter System] [CDBH]'):
#             overwrite(kpi_name,'CD')    

#         if(kpi_name=='MV_CSFB Redirection Success Rate'):
#             overwrite(kpi_name,'CK') 

#         if(kpi_name=='MV_VoLTE ERAB Setup Success Rate'):
#             overwrite(kpi_name,'CR') 

#         if(kpi_name=='MV_VoLTE DCR [CBBH]'):
#             overwrite(kpi_name,'CY')    

#         if(kpi_name=='MV_VoLTE Packet Loss DL [CBBH]'):
#             overwrite(kpi_name,'DF') 

#         if(kpi_name=='MV_VoLTE Packet Loss UL [CBBH]'):
#             overwrite(kpi_name,'DM')  

#         if(kpi_name=='VoLTE IntraF HOSR Exec [CBBH]'):
#             overwrite(kpi_name,'EA')   

#         if(kpi_name=='VoLTE InterF HOSR Exec [CBBH]'):
#             overwrite(kpi_name,'EH')  

#         if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
#             overwrite(kpi_name,'EO')  

#         if(kpi_name=='UL RSSI [CDBH]'):
#             overwrite(kpi_name,'EV')   

#         if(kpi_name=='MV_4G Data Volume_GB'):
#             overwrite(kpi_name,'FC')  

#         if(kpi_name=='MV_VoLTE Traffic'):
#             overwrite(kpi_name,'FL')  

#         if(kpi_name=='MV_Average number of used DL PRBs [CDBH]'):
#             overwrite(kpi_name,'FS')    

#         if(kpi_name=='Radio NW Availability'):
#             overwrite(kpi_name,'FY') 
            
#         if(kpi_name=='VoLTE DCR_Nom [CBBH]'):
#             overwrite(kpi_name,'GE') 
            
#         if(kpi_name=='CA Data Volume [MB]'):
#             overwrite(kpi_name,'GK')    
#     save_raj_trend=os.path.join(door_raj_path,'output','Rajasthantrend.xlsx')    
#     print(save_raj_trend,'-----------------------------------------')
#     wb.save(save_raj_trend) 
#     Download_path=os.path.join(MEDIA_URL,'trends','raj','output','Rajasthantrend.xlsx')
#     return Response({"message":"Succesfully uploaded",'status':True,'missing_sites':message,'Download_url':Download_path})
   
       

from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from django.core.files.storage import FileSystemStorage
from rest_framework.decorators import api_view
from datetime import date, timedelta
import datetime 
from rest_framework.response import Response
import pandas as pd
import openpyxl
import os
from commom_utilities.utils import *

@api_view(['POST'])
    
def old_raj_trend(request):
    kpi = ['Radio NW Availability', 'E-UTRAN Average CQI [CDBH]', 'DL User Throughput_Mbps [CDBH]', 'UL RSSI [CDBH]',
        'PS Drop Call Rate % [CDBH]', 'RRC Setup Success Rate [CDBH]', 'ERAB Setup Success Rate [CDBH]',
        'PS handover success rate [LTE Intra System] [CDBH]', 'MV_Average number of used DL PRBs [CDBH]',
        'MV_UL User Throughput_Kbps [CDBH]', 'MV_VoLTE Traffic', 'MV_VoLTE Packet Loss DL [CBBH]',
        'MV_VoLTE Packet Loss UL [CBBH]', 'MV_VoLTE DCR [CBBH]', 'MV_CSFB Redirection Success Rate',
        'MV_VoLTE ERAB Setup Success Rate', 'MV_4G Data Volume_GB',
        'PS handover success rate [LTE Inter System] [CDBH]', 'VoLTE InterF HOSR Exec [CBBH]',
        'VoLTE IntraF HOSR Exec [CBBH]', 'PS Drop Call Rate % [CBBH]','VoLTE DCR_Nom [CBBH]','CA Data Volume [MB]']
    
    kpi1 = ['Radio NW Availability', 'E-UTRAN Average CQI [CDBH]', 'DL User Throughput_Kbps [CDBH]', 'UL RSSI [CDBH]',
        'PS Drop Call Rate % [CDBH]', 'RRC Setup Success Rate [CDBH]', 'ERAB Setup Success Rate [CDBH]',
        'PS handover success rate [LTE Intra System] [CDBH]', 'MV_Average number of used DL PRBs [CDBH]',
        'MV_UL User Throughput_Kbps [CDBH]', 'MV_VoLTE Traffic', 'MV_VoLTE Packet Loss DL [CBBH]',
        'MV_VoLTE Packet Loss UL [CBBH]', 'MV_VoLTE DCR [CBBH]', 'MV_CSFB Redirection Success Rate',
        'MV_VoLTE ERAB Setup Success Rate', 'MV_4G Data Volume_GB',
        'PS handover success rate [LTE Inter System] [CDBH]', 'VoLTE InterF HOSR Exec [CBBH]',
        'VoLTE IntraF HOSR Exec [CBBH]', 'PS Drop Call Rate % [CBBH]','VoLTE DCR_Nom [CBBH]','CA Data Volume [MB]',]
    
    raw_kpi_4G=request.FILES["raw_kpi_4G"] if "raw_kpi_4G" in request.FILES else None
    df=pd.read_excel(raw_kpi_4G)
    
    required_cols = ['Radio NW Availability', 'E-UTRAN Average CQI [CDBH]', 'DL User Throughput_Kbps [CDBH]', 'UL RSSI [CDBH]',
        'PS Drop Call Rate % [CDBH]', 'RRC Setup Success Rate [CDBH]', 'ERAB Setup Success Rate [CDBH]',
        'PS handover success rate [LTE Intra System] [CDBH]', 'MV_Average number of used DL PRBs [CDBH]',
        'MV_UL User Throughput_Kbps [CDBH]', 'MV_VoLTE Traffic', 'MV_VoLTE Packet Loss DL [CBBH]',
        'MV_VoLTE Packet Loss UL [CBBH]', 'MV_VoLTE DCR [CBBH]', 'MV_CSFB Redirection Success Rate',
        'MV_VoLTE ERAB Setup Success Rate', 'MV_4G Data Volume_GB',
        'PS handover success rate [LTE Inter System] [CDBH]', 'VoLTE InterF HOSR Exec [CBBH]',
        'VoLTE IntraF HOSR Exec [CBBH]', 'PS Drop Call Rate % [CBBH]','VoLTE DCR_Nom [CBBH]','CA Data Volume [MB]',]
    
    gsm=["SDCCH Blocking Rate [BBH]",
            "TCH Blocking Rate [BBH]",
            "TCH Drop Call Rate [BBH]",
            "Handover Success Rate [BBH]",
            "Total Voice Traffic [BBH]",
            "Network availability [RNA]",
            "SDCCH Drop Call Rate_Nom [BBH]",
            "TCH Drop Call Rate_Nom [BBH]",
            "SDCCH Drop Call Rate [BBH]"]
    sts,response=required_col_check(raw_kpi_4G,required_cols)
    if sts:
        return Response(response)
   
    # os.remove(path=file_path)
   

################################## for site_list ############################################
    # excel_site_list_raj=request.FILES["site_list"] if "site_list" in request.FILES else None
    # df_site_list=pd.read_excel(excel_site_list_raj)
   
    # if excel_site_list_raj:
    #     sts,response=required_col_check(raw_kpi_4G,required_cols)
    #     if sts:
    #         return Response(response)
    response,s_l= site_list_handler_4G(request)
    if s_l:
        df_site_list=s_l
        print("_________________________--print",type(df_site_list[0]))

        print(df_site_list,"_______________")
        if response:
            return Response(response)
        
    raw_kpi_2G=request.FILES['2G_raw_kpi'] if '2G_raw_kpi' in request.FILES else None
    df_raw_kpi_2G=pd.read_excel(raw_kpi_2G)
    sts,response=required_col_check(raw_kpi_2G,gsm)
    if sts:
        return Response(response)    
    response, s_l = site_list_handler_2G(request)
 
    if s_l:
        df_2G_site = s_l
        df_2G_site_str=[str(y) for y in df_2G_site]
        print("__________2g_site_______________print", type(df_2G_site_str[0]))

        print(df_2G_site_str, "______2G_ROW_KPI_________", s_l)
        if response:
            return Response(response)
    print("checkpoint4")

######################################################################        
    door_raj_path=os.path.join(MEDIA_ROOT,'trends','raj')
    # save_raj_path=os.path.join(door_raj_path,"actual input/RAJ_TOOL_KPI.xlsx")
    # df = pd.read_excel(save_raj_path)
   
    
    STR=os.path.join(door_raj_path,'template','update raj trend.xlsx')
    wb=openpyxl.load_workbook(STR)
    ws=wb['4G']

    
    
    for x in kpi1:
            df[x] = df[x].replace(to_replace='.*', value=0, regex=True)

    df["Short name"] = df["Short name"].fillna(method="ffill")  #######for (forward fill)ffill is used to copy  and fill
    df.columns.values[1] = 'Date'  #####for empty column

    df["DL User Throughput_Kbps [CDBH]"] = (df["DL User Throughput_Kbps [CDBH]"]/1024)  #####for change kbps to mbps
    df.columns.values[4] = 'DL User Throughput_Mbps [CDBH]'
    # print(df)

    sn=[]
    techlist=[]
    strip=[]

    for cell in df['Short name']:
        if("_" in cell):
            sit_id = cell.split("_")[-2][:-1]
            sn.append(sit_id)
        

        else:
            sit_id=cell[:-1]
            sn.append(sit_id)
        
        if('_F1_' in cell or '_F3_' in cell or '_F8_' in cell  or '_T1' in cell or '_T2_' in cell ):  
            if('_F1_' in cell):
                tech='L2100'
                # techlist.append(tech)
            if('_F3_' in cell):
                tech='L1800'
                # techlist.append(tech) 
            if('_f8_' in cell):
                tech='L900'    
                # techlist.append(tech)
            if('_T1_' in cell or '_T2_' in cell):
                tech='L2300'
            techlist.append(tech)
                            
        else:
            cell=sit_id
            techlist.append(cell)
        
    # if ('' in cell):
    #     txt=cell.strip()
    #     strip.append(txt)
    # else:
    #     txt=cell()
    #     strip.append(txt)      
        
    # print(techlist)
    # df.insert(2,'site type',sn)
    df.insert(0, 'Site ID', sn)
    df.insert(4,'Tech',techlist)  
    # df.insert(5,'strip',strip)

    df["Short name"] =df["Short name"].apply(lambda x: x.strip())  
    # print("dfdfff",df_site_list['2G ID'])
    # message=site_comparision(sn,list(df_site_list['2G ID']))# site_comparision_call 
    message_4G=site_comparision(sn,df_site_list)# site_comparision_call 


    df.fillna(value=0,inplace=True)  
    
    raj_1st_step=os.path.join(door_raj_path,'process output','desired.xlsx')
    df.to_excel(raj_1st_step, index=False)
    print(df)
    excel_file_1 = raj_1st_step
    # raj_2nd_step=os.path.join(door_raj_path,'project file','site.xlsx')
    # excel_file_2 = raj_2nd_step

    df1 = pd.read_excel(excel_file_1)
    # df2 = pd.read_excel(excel_file_2)

    df1.rename(columns={"Site ID": "Site_ID"}, inplace=True)
    # df_site_list.rename(columns={"Site ID": "Site_ID"}, inplace=True)

   

    filtered_df_1 = df1[(df1.Site_ID.isin(df_site_list))]

    print(filtered_df_1)
    raj_filtered_3rd_step=os.path.join(door_raj_path,'process output','filtered_df_1.xlsx')
    filtered_df_1.to_excel(raj_filtered_3rd_step, index=False)

    df1 = pd.read_excel(raj_filtered_3rd_step)
    df_pivot = df1.pivot_table(values=kpi, columns='Date', index=['Short name', 'Site_ID','Tech'])
    
    
    raj_pivot_4th_step=os.path.join(door_raj_path,'process output','pivot.xlsx')
    df_pivot.to_excel(raj_pivot_4th_step)
    
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
	# This process is similar to binary-to-
	# decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    str_date=request.POST.get('offered_date')
    date1 = datetime.datetime.strptime(str_date, '%Y-%m-%d')
    print("Date-------------------------------",date1)
    # date1=cal.get_date()
    # date1=date(2023,3,2)
    d1=date1-timedelta(1)
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]
   
    def overwrite(kpi_name,coln1):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        print('donnnnnnnnnnnnnnnnne')
        print(len(index))
        dr=df_pivot[kpi_name]
        
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()


        ws[coln1+"4"].value=cl[4]
        ws[coln2+"4"].value=cl[3]
        ws[coln3+"4"].value=cl[2]
        ws[coln4+"4"].value=cl[1]
        ws[coln5+"4"].value=cl[0]



        for i,value in enumerate(index):
                j=i+5
                ws['N'+str(j)].value='CAPACITY'
                ws['M'+str(j)].value='ERICSSON'
                ws['K'+str(j)].value=date1

                ws['L'+str(j)].value='DONE'
                ws['B'+str(j)].value=index[i][1]
                ws['C'+str(j)].value=index[i][0]
                ws['G'+str(j)].value=index[i][0]
                ws['I'+str(j)].value=index[i][2]
                
                # ws['D'+str(j)].value=index[i][2]
                # ws['E'+str(j)].value=index[i][2]
                ws['F'+str(j)].value=index[i][0]     
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
    for kpi_name in kpi:
        if(kpi_name=='RRC Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'AN')

        if(kpi_name=='ERAB Setup Success Rate [CDBH]'):
            overwrite(kpi_name,'AU') 

        if(kpi_name=='PS Drop Call Rate % [CDBH]'):
            overwrite(kpi_name,'BB')

        if(kpi_name=='DL User Throughput_Mbps [CDBH]'):
            overwrite(kpi_name,'BI')

        if(kpi_name=='MV_UL User Throughput_Kbps [CDBH]'):
            overwrite(kpi_name,'BP') 
            
        if(kpi_name=='PS handover success rate [LTE Intra System] [CDBH]'):
            overwrite(kpi_name,'BW') 

        if(kpi_name=='PS handover success rate [LTE Inter System] [CDBH]'):
            overwrite(kpi_name,'CD')    

        if(kpi_name=='MV_CSFB Redirection Success Rate'):
            overwrite(kpi_name,'CK') 

        if(kpi_name=='MV_VoLTE ERAB Setup Success Rate'):
            overwrite(kpi_name,'CR') 

        if(kpi_name=='MV_VoLTE DCR [CBBH]'):
            overwrite(kpi_name,'CY')    

        if(kpi_name=='MV_VoLTE Packet Loss DL [CBBH]'):
            overwrite(kpi_name,'DF') 

        if(kpi_name=='MV_VoLTE Packet Loss UL [CBBH]'):
            overwrite(kpi_name,'DM')  

        if(kpi_name=='VoLTE IntraF HOSR Exec [CBBH]'):
            overwrite(kpi_name,'EA')   

        if(kpi_name=='VoLTE InterF HOSR Exec [CBBH]'):
            overwrite(kpi_name,'EH')  

        if(kpi_name=='E-UTRAN Average CQI [CDBH]'):
            overwrite(kpi_name,'EO')  

        if(kpi_name=='UL RSSI [CDBH]'):
            overwrite(kpi_name,'EV')   

        if(kpi_name=='MV_4G Data Volume_GB'):
            overwrite(kpi_name,'FC')  

        if(kpi_name=='MV_VoLTE Traffic'):
            overwrite(kpi_name,'FL')  

        if(kpi_name=='MV_Average number of used DL PRBs [CDBH]'):
            overwrite(kpi_name,'FS')    

        if(kpi_name=='Radio NW Availability'):
            overwrite(kpi_name,'FY') 
            
        if(kpi_name=='VoLTE DCR_Nom [CBBH]'):
            overwrite(kpi_name,'GE') 
            
        if(kpi_name=='CA Data Volume [MB]'):
            overwrite(kpi_name,'GK') 



    ######################## gsm #############
    door_path=  os.path.join(MEDIA_ROOT,'trends','raj')
    df_raw_kpi_2G["Short name"] = df_raw_kpi_2G["Short name"].fillna(method="ffill") 
    df_raw_kpi_2G.columns.values[1] = 'Date' 
    df_raw_kpi_2G.fillna(value=0,inplace=True) 

    df_raw_kpi_2G['site']=[site[:-1] for site in df_raw_kpi_2G['Short name']]
    # df_dup=df.drop_duplicates('Date',keep='last')
    message_2G=site_comparision(df_raw_kpi_2G['site'],df_2G_site_str)
    print("message_2g",message_2G)
    g2_path1=os.path.join(door_path,'process output','2g_fill.xlsx')
    df_raw_kpi_2G.to_excel(g2_path1,index=False)
    excel_1=g2_path1

    # excel_file_1 = pd.read_excel(excel_1)
    # excel_file_2 = 'project file/site.xlsx'

    df1 = pd.read_excel(excel_1)
    # df2 = pd.read_excel(excel_file_2)
  
    str_date=request.POST.get('offered_date')
    date1=datetime.datetime.strptime(str_date,'%Y-%m-%d')
    # date1=date(2023,11,7)
    d1=date1-timedelta(1)                                                                                 
    d2=date1-timedelta(2)
    d3=date1-timedelta(3)
    d4=date1-timedelta(4)
    d5=date1-timedelta(5)
    cl=[d1,d2,d3,d4,d5]


    # filter_2g=df1[(df1.site.isin(list(df2['2G ID']))& df1.Date.isin(cl))]
    filter_2g=df1[(df1.site.isin(df_2G_site_str))]
    g2_filter=os.path.join(door_path,'process output','2Gfilter.xlsx')
    filter_2g.to_excel(g2_filter,index=False)
    df_filter=pd.read_excel(g2_filter)
    df_pivot=df_filter.pivot_table(values=gsm,columns='Date',index=['site','Short name'])
    g2_pivot=os.path.join(door_path,'process output','2G_pivot.xlsx')
    df_pivot.to_excel(g2_pivot)

    # STR=os.path.join(door_path,'template','2G_template.xlsx')
   
    # wb=openpyxl.load_workbook(STR)
    # ws=wb['2G']
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]
            
            
    def titleToNumber(s):
	# This process is similar to binary-to-
	# decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result
    
    
   
    def overwrite(kpi_name,coln1,ws):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index=df_pivot.index
        print('done')
        print(len(index))
        dr=df_pivot[kpi_name]
        
        li=dr.columns
        col1=dr[li[0]].to_list()
        col2=dr[li[1]].to_list()
        col3=dr[li[2]].to_list()
        col4=dr[li[3]].to_list()
        col5=dr[li[4]].to_list()


        ws[coln1+"3"].value=cl[4]
        ws[coln2+"3"].value=cl[3]
        ws[coln3+"3"].value=cl[2]
        ws[coln4+"3"].value=cl[1]
        ws[coln5+"3"].value=cl[0]



        for i,value in enumerate(index):
                j=i+4
                ws['L'+str(j)].value='CAPACITY'
                ws['K'+str(j)].value='ERICSSON'
                ws['G'+str(j)].value='G900'
                ws['A'+str(j)].value='RAJ'

                ws['I'+str(j)].value=date1

                ws['B'+str(j)].value='DONE'
             
                ws['B'+str(j)].value=index[i][0]
                ws['C'+str(j)].value=index[i][1]
                ws['D'+str(j)].value=index[i][0]
                ws['E'+str(j)].value=index[i][0]
 
                ws[coln1+str(j)].value=col1[i]
                ws[coln2+str(j)].value=col2[i]
                ws[coln3+str(j)].value=col3[i]
                ws[coln4+str(j)].value=col4[i]
                ws[coln5+str(j)].value=col5[i]
    g2_ws=wb['2G']
    for kpi_name in gsm:
        if(kpi_name=='SDCCH Blocking Rate [BBH]'):
            overwrite(kpi_name,'O',g2_ws)
        if(kpi_name=='TCH Blocking Rate [BBH]'):
            overwrite(kpi_name,'AC',g2_ws)
        if(kpi_name=='TCH Drop Call Rate [BBH]'):
            overwrite(kpi_name,'AJ',g2_ws)
        if(kpi_name=='Handover Success Rate [BBH]'):
            overwrite(kpi_name,'AQ',g2_ws)
        if(kpi_name=='Total Voice Traffic [BBH]'):
            overwrite(kpi_name,'AX',g2_ws)
        if(kpi_name=='Network availability [RNA]'):
            overwrite(kpi_name,'BE',g2_ws)
        if(kpi_name=='SDCCH Drop Call Rate_Nom [BBH]'):
            overwrite(kpi_name,'BL',g2_ws)
        if(kpi_name=='TCH Drop Call Rate_Nom [BBH]'):
            overwrite(kpi_name,'BR',g2_ws)
        if(kpi_name=='SDCCH Drop Call Rate [BBH]'):
            overwrite(kpi_name,'V',g2_ws) 
    #######################    
    save_raj_trend=os.path.join(door_raj_path,'output','Rajasthantrend.xlsx')    
    print(save_raj_trend,'-----------------------------------------')
    wb.save(save_raj_trend) 
    Download_path=os.path.join(MEDIA_URL,'trends','raj','output','Rajasthantrend.xlsx')
    return Response({"message":"Succesfully uploaded",'status':True,'missing_site_4g':message_4G,'missing_site_2g':message_2G,'Download_url':Download_path})   
