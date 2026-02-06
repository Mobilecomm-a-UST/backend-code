import pandas as pd
import os
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST 
from mcom_website.settings import MEDIA_ROOT, MEDIA_URL
from rest_framework import status
import xml.etree.ElementTree as ET
import re
import math
import gzip
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side,PatternFill
from openpyxl.utils import get_column_letter

def dumpdata_to_xml(dumy_data, output_file):

    import xml.etree.ElementTree as ET

    root = ET.Element("raml")

    # group by MO + DistName
    grouped = {}
    for row in dumy_data:
        key = (row.get("MO"), row.get("DistName"))
        grouped.setdefault(key, []).append(row)

    for (mo, dist), rows in grouped.items():

        mo_tag = ET.SubElement(root, "managedObject")
        mo_tag.set("class", mo)

        if dist:
            mo_tag.set("distName", dist)

        for r in rows:
            p = ET.SubElement(mo_tag, "p")
            p.set("name", str(r.get("Parameter")))
            val = r.get("value")
            p.text = "" if val is None else str(val)

    tree = ET.ElementTree(root)
    tree.write(output_file, encoding="utf-8", xml_declaration=True)



main_folder=os.path.join(MEDIA_ROOT, "Nokia_Slicing")
output_path = os.path.join(main_folder, "Final_Output")
dump_data_path=os.path.join(main_folder,"Dump_data")
os.makedirs(main_folder,exist_ok=True)
os.makedirs(output_path, exist_ok=True)
os.makedirs(dump_data_path, exist_ok=True)

def format_excel(file_path, sheet_name="Slicing"):
    wb = load_workbook(file_path)

    ws = wb[sheet_name] if sheet_name else wb.active
    header_fill = PatternFill(start_color="31869B",end_color="31869B",fill_type="solid")

    header_font = Font(bold=True)
    center_align = Alignment(
        vertical="center",
        horizontal="center",
        wrap_text=True
    )

    
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Header row height
    ws.row_dimensions[1].height =24

    # ---------- Data formatting ----------
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center",horizontal="center", wrap_text=True)
            cell.border = border
            
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 3, 45)
    wb.save(file_path)

#for remark--
def remark(internal, external):
    if pd.isna(internal) and pd.isna(external):
        return "No Changes in value"
    if pd.isna(internal) or pd.isna(external):
        return "No find new value"

    i = str(internal).strip().lower()
    e = str(external).strip().lower()
    if i == e:
        return "No Changes in value"
    ni = re.search(r'\d+', i)
    ne = re.search(r'\d+', e)

    if ni and ne and ni.group() == ne.group():
        return "No Changes in value"

    return "Changes in value"

# change t/f-> 0,1--
def tf_to_01(val):
    if val is None:
        return val
    val = str(val).strip().lower()
    if val == "true":
        return 1
    if val == "false":
        return 0
    return val

# normalize function-----
def normalize_id(val):
    if pd.isna(val):
        return pd.NA
    parts = [p.strip() for p in str(val).split(",") if p.strip().isdigit()]
    return ",".join(map(str, sorted(map(int, parts)))) if parts else pd.NA


def normalize(val):
    if pd.isna(val):
        return val
    v = str(val).strip()
    try:
        return int(v)
    except:
        try:
            return float(v)
        except:
            return v.lower()
        


@api_view(['POST', 'GET' , 'DELETE'])
def upload_fix_parameter(request):

    folder = os.path.join(main_folder, 'Nokia_Slicing_Fixpara')
    os.makedirs(folder, exist_ok=True)

    if request.method == 'POST':
        file = request.FILES.get('fix_para')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)

        path = os.path.join(folder, file.name)

        with open(path, 'wb+') as f:
            for chunk in file.chunks():
                f.write(chunk)

        return Response({'status': True, 'file': file.name})


    if request.method == 'GET':
        return Response({'files': os.listdir(folder)})


    if request.method == 'DELETE':
        for f in os.listdir(folder):
            os.remove(os.path.join(folder, f))
        return Response({'status': True, 'message': 'All files deleted'})



@api_view(["POST"])
def nokia_slicing_dump(request):
    excel_file=os.path.join(main_folder, 'Nokia_Slicing_Fixpara')
    if not os.path.exists(excel_file):
        return Response({"error": "Fix_Parameter folder not found"}, status=400)

    xml_files = request.FILES.getlist("xml_file")
    if not xml_files:
        return Response({"error": "Please upload XML files"}, status=HTTP_400_BAD_REQUEST)
    
    for file in xml_files:

        file_name = file.name.lower()
        if file_name.endswith(".gz"):
            xml_bytes = gzip.decompress(file.read())
        else:
            xml_bytes = file.read()
            
        root = ET.fromstring(xml_bytes)


        m = re.match(r'\{(.*)\}', root.tag)
        ns_url = m.group(1) if m else root.attrib.get("xmlns", "")
        ns = {"ns": ns_url} if ns_url else {}

        print(f"File: {file.name}, Namespace: {ns_url}")

        # ---- Try both namespace & no namespace ----
    dumy_data=[]
    managed_objects = root.findall(".//ns:managedObject", ns) if ns else []
    if not managed_objects:
        managed_objects = root.findall(".//managedObject")

    for mo in managed_objects:

        mo_class = mo.attrib.get("class", "")
        dist_name = mo.attrib.get("distName", "")

        # for NRBTS class--------
        if mo_class == "com.nokia.srbts.nrbts:NRBTS":

            required_in_nrbts = {
                "actSliceAwareScheduler",
                "actHighSliceWeightFactor",
                "actSliceAwareSchedulerUlAndEnh",
                "actSliceSwitchToDefault",
                "actAdditionalSliceSupport",
                "actSliceNumExt",
                "actExtSchedWeightAndPrefWrrAlg",
                
            }

            params_nrbts = {}

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")
                if name in required_in_nrbts:
                    params_nrbts[name] = p.text
                    dumy_data.append({
                    "MO": "NRBTS",
                    "DistName": dist_name,
                    "Parameter": name,
                    "value": tf_to_01(p.text)
                })

            # print("NRBTS FOUND ------------------")
            # print("Class:", mo_class)
            # print("DistName:", dist_name)
            # print("Params_NRBTS:", params_nrbts)

        # for NRCELL Class-----------
        elif mo_class == "com.nokia.srbts.nrbts:NRCELL":

            required_in_nrcell = {
                "actUlTxSkip",
                "srPeriodicityMin",
                "adaptiveSrLoadThresholdUp",
                "adaptiveSrLoadThresholdDown",
                "adaptiveSrBlockingMinimization",
                "maxNbrTrafficLimit",
                "preferredVoNrSrPeriod",
                "actNbrForNonGbrBearers",
                "congDetectPeriod",
                "congWeightAlg",
                "maxNumOfNbrBearers",
                "maxPrbsPerNbrUe",
                "nbrPdcchCongHandlingDl",
                "nbrPdcchCongHandlingUl",
                "nbrPdschCongHandling",
                "nbrPuschCongHandling",
                "nrResourceGroupProfileDN"
            }

            params_nrcell = {}

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")
                if name in required_in_nrcell:
                    params_nrcell[name] = p.text
                    dumy_data.append({
                    "MO": "NRCELL",
                    "DistName": dist_name,
                    "Parameter": name,
                    "value": tf_to_01(p.text)
                })

            # print("NRCELL FOUND ----------------")
            # print("Class:", mo_class)
            # print("DistName:", dist_name)
            # print("Params_NRCELL:", params_nrcell)

      
         # For NRDRB class------------------
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_id = int(dist_name.split("NRDRB-")[-1])
            except:
                continue

            group_1_ids = {5, 6, 11, 12, 21, 22, 25, 26}
            group_2_ids = {21, 22, 25, 26}
            group_3_ids = {6}
            group_4_ids = {5, 6, 7, 8, 9, 11, 12, 21, 22, 25}
            group_5_ids = {26}
            group_6_ids = {11, 21}
            group_7_ids = {12, 22}
            group_8_ids = {5, 25}
            group_9_ids = {6, 26}
            group_10_ids = {7}
            group_11_ids = {8, 9}

            group_1_params = {
                "pdcpStatRepWaitTimerOffset",
                "pdcpStatRepWaitTimer"
            }

            group_2_params = {
                "actDddsReduction",
                "actDlTxResumeOnPdcpStatRep",
                "actLostPduFastRetx",
                "actOptDataFlowRateEst",
                "reTxPrioritizationType"
            }

            group_3_params = {
                "inactPeriodX2PdcpDuplication"
            }

            group_4_params = {
                "schedulWeight"
            }

            group_5_params = {
                "schedulWeight",
                "nrDrbMacDN"
            }

            group_6_params = {"priorityLevel"}
            group_7_params = {"priorityLevel"}
            group_8_params = {"priorityLevel"}
            group_9_params = {"priorityLevel"}
            group_10_params = {"priorityLevel"}
            group_11_params = {"priorityLevel"}

            #FIX IS HERE (deep search for <p>)
            for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):
                name = p.attrib.get("name")

                if nrdrb_id in group_1_ids and name in group_1_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_2_ids and name in group_2_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_3_ids and name in group_3_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_4_ids and name in group_4_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_5_ids and name in group_5_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_6_ids and name in group_6_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_7_ids and name in group_7_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_8_ids and name in group_8_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_9_ids and name in group_9_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_10_ids and name in group_10_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

                if nrdrb_id in group_11_ids and name in group_11_params:
                    dumy_data.append({
                        "MO": "NRDRB",
                        "DistName": dist_name,
                        "ID": nrdrb_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })


        # for RDRB_5QI class---            
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB_5QI":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_5qi_id = int(dist_name.rsplit("NRDRB_5QI-", 1)[-1])
            except:
                continue

            nrdrb_5qi_param_map = {
                21: {"snssaiDN"},
                22: {"snssaiDN"},
                25: {"snssaiDN"},
                26: {"snssaiDN"},
                5:  {"snssaiDN"},
                6:  {"snssaiDN"},
                7:  {"snssaiDN"},
                8:  {"snssaiDN"},
                9:  {"snssaiDN"},
                1:  {"snssaiDN"},
                2:  {"snssaiDN"},
            }

            allowed_params = nrdrb_5qi_param_map.get(nrdrb_5qi_id)
            if not allowed_params:
                continue

            # snssaiDN is inside list/item/p
            for item in (
                mo.findall(".//ns:list[@name='snssaiList']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='snssaiList']/item")
            ):
                p = item.find("ns:p", ns) if ns else item.find("p")
                if p is None:
                    continue

                name = p.attrib.get("name")
                if name in allowed_params:
                    dumy_data.append({
                        "MO": "NRDRB_5QI",
                        "DistName": dist_name,
                        "ID": nrdrb_5qi_id,
                        "Parameter":'Item-snssaiList-snssaiDN',        
                        "value": p.text             
                    })  
        #
        # for NRDRB_MAC classs---
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB_MAC":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_mac_id = int(dist_name.rsplit("NRDRB_MAC-", 1)[-1])
            except:
                continue

            group_1_ids={1}
            group_2_ids={2}
            group_3_ids={3}
            group_4_ids={4}
            group_5_ids={4}

            group_1_params={"schedulBSD"}
            group_2_params={"schedulBSD"}
            group_3_params={"schedulBSD"}
            group_4_params={"schedulBSD"}
            group_5_params={'lcgid','maxDlHarqTxDrb','prioritisedBitRate','schedulPrio','nbrDl','nbrUl'}
            


            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

        
                if nrdrb_mac_id in group_1_ids and name in group_1_params:
                    dumy_data.append({
                        "MO": "NRDRB_MAC",
                        "DistName": dist_name,
                        "ID": nrdrb_mac_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

             
                if nrdrb_mac_id in group_2_ids and name in group_2_params:
                    dumy_data.append({
                        "MO": "NRDRB_MAC",
                        "DistName": dist_name,
                        "ID": nrdrb_mac_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })
             
                if nrdrb_mac_id in group_3_ids and name in group_3_params:
                    dumy_data.append({
                        "MO": "NRDRB_MAC",
                        "DistName": dist_name,
                        "ID": nrdrb_mac_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })  
            
                if nrdrb_mac_id in group_4_ids and name in group_4_params:
                    dumy_data.append({
                        "MO": "NRDRB_MAC",
                        "DistName": dist_name,
                        "ID": nrdrb_mac_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })  
               
                if nrdrb_mac_id in group_5_ids and name in group_5_params:
                    dumy_data.append({
                        "MO": "NRDRB_MAC",
                        "DistName": dist_name,
                        "ID": nrdrb_mac_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })


        # for NRDRB_PDCP class----------
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB_PDCP":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_pdcp_id = int(dist_name.rsplit("NRDRB_PDCP-", 1)[-1])
            except:
                continue  

            nrdrb_pdcp_param_map = {

                2:{'tReorderingDl','tReorderingUl'}
            }

            allowed_params = nrdrb_pdcp_param_map.get(nrdrb_pdcp_id)
            if not allowed_params:
                continue                
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")
                if name in allowed_params:
                    dumy_data.append({
                        "MO": "NRDRB_PDCP",
                        "DistName": dist_name,
                        "ID":nrdrb_pdcp_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })  



        #  for classs NRDRB_RLC_AM-----   
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB_RLC_AM":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_rlc_id = int(dist_name.rsplit("NRDRB_RLC_AM-", 1)[-1])
            except:
                continue


            nrdrb_rlc_param_map = {
                1:{'dlMaxRetxThreshold','dlPollByte',
                    'ulMaxRetxThreshold','ulPollByte'}
                    }

            allowed_params = nrdrb_rlc_param_map.get(nrdrb_rlc_id)
            if not allowed_params:
                continue                
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")
                if name in allowed_params:
                    dumy_data.append({
                        "MO": "NRDRB_RLC_AM",
                        "DistName": dist_name,
                        "ID":nrdrb_rlc_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })  
       

        # for classs NRDRB_RLC_UM-----   
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB_RLC_UM":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_rlcum_id = int(dist_name.rsplit("NRDRB_RLC_UM-", 1)[-1])
            except:
                continue


            nrdrb_rlcum_param_map = {
                1:{'rlcUMDrbSNLength',
                    'ulTReassemblyUm'}
                    }

            allowed_params = nrdrb_rlcum_param_map.get(nrdrb_rlcum_id)
            if not allowed_params:
                continue   
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):       
                name = p.attrib.get("name")
                if name in allowed_params:
                    dumy_data.append({
                        "MO": "NRDRB_RLC_UM",
                        "DistName": dist_name,
                        "ID":nrdrb_rlcum_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    }) 

        # for NRDRB_TCP_BOOST  
        elif mo_class == "com.nokia.srbts.nrbts:NRDRB_TCP_BOOST":
            dist_name = mo.attrib.get("distName", "")
            try:
                nrdrb_tcp_id = int(dist_name.rsplit("NRDRB_TCP_BOOST-", 1)[-1])
            except:
                continue


            nrdrb_tcp_param_map = {
                1:{'tcpBoostPugMinSrPeriodicity'}
                    }

            allowed_params = nrdrb_tcp_param_map.get(nrdrb_tcp_id)
            if not allowed_params:
                continue                
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")
                if name in allowed_params:
                    dumy_data.append({
                        "MO": "NRDRB_TCP_BOOST",
                        "DistName": dist_name,
                        "ID":nrdrb_tcp_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    }) 


        # for NRPMRNL class---
        elif mo_class == "com.nokia.srbts.nrbts:NRPMRNL":

            required_in_nrpmrl = {
               'miNrCellUtilPerNrg',
               'miNrInterRATMobilitySaPmqap',
               'miNrSaCuNrpmqap',
               'miNrNgInterfaceNrpmqap',
                'miNrPdcpCellNrpmqap',
                'miNrCellUtilization',
                'miNrHighRlcCellNrpmqap',
                'miNrUeStatNrpmqap',
                'miNrPdcpLatNrpmqap'

            }

            params_nrpmrl = {}

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")
                if name in required_in_nrpmrl :
                    params_nrpmrl[name] = p.text
                    dumy_data.append({
                    "MO": "NRPMRNL",
                    "DistName": dist_name,
                    "Parameter": name,
                    "value": tf_to_01(p.text)
                })
                    

        # for SNSSAI class---
        elif mo_class == "com.nokia.srbts.nrbts:SNSSAI":
            dist_name = mo.attrib.get("distName", "")
            try:
                snssai_id = int(dist_name.rsplit("SNSSAI-", 1)[-1])
            except:
                continue  

            group_1_ids={1}
            group_2_ids={2}

            group_1_params={'sd','sst','userLabel'}
            group_2_params={'sd','sst','userLabel'}
          

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                name = p.attrib.get("name")

        
                if snssai_id in group_1_ids and name in group_1_params:
                    dumy_data.append({
                        "MO": "SNSSAI",
                        "DistName": dist_name,
                        "ID": snssai_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    })

             
                if snssai_id in group_2_ids and name in group_2_params:
                    dumy_data.append({
                        "MO": "SNSSAI",
                        "DistName": dist_name,
                        "ID": snssai_id,
                        "Parameter": name,
                        "value": tf_to_01(p.text)
                    }) 

            



        # for TRACKINGAREA class---
        elif  mo_class == "com.nokia.srbts.nrbts:TRACKINGAREA":
            dist_name = mo.attrib.get("distName", "")

            snssai_values = []

            for item in (
                mo.findall(".//ns:list[@name='snssaiList']/ns:item", ns)
                if ns else
                mo.findall(".//list[@name='snssaiList']/item")
            ):
                p = item.find("ns:p", ns) if ns else item.find("p")
                if p is not None and p.attrib.get("name") == "snssaiDN":
                    snssai_values.append(p.text)

            if snssai_values:
                dumy_data.append({
                    "MO": "TRACKINGAREA",
                    "DistName": dist_name,
                    "Parameter": "Item-snssaiList-snssaiDN",
                    "value": ";".join(snssai_values)
                })
            
        #for NRPMQAP class-----------------------------
        elif mo_class == "com.nokia.srbts.nrbts:NRPMQAP":

            dist_name = mo.attrib.get("distName", "")
            try:
                nrpmqap_id = int(dist_name.split("NRPMQAP-")[-1])
            except:
                continue

            group_1_ids = {11, 12, 15, 16, 17, 18, 19}
            group_2_ids = {21, 22, 25, 26}

            if nrpmqap_id not in group_1_ids and nrpmqap_id not in group_2_ids:
                continue

            if nrpmqap_id in group_1_ids:

                for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):
                    if p.attrib.get("name") == "thpHistScale":
                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "DistName": dist_name,
                            "ID": nrpmqap_id,
                            "Parameter": "thpHistScale",
                            "value": tf_to_01(p.text)
                        })

                # cfgSliceId marker
                if mo.findall(".//ns:list[@name='cfgSliceId']", ns) if ns else mo.findall(".//list[@name='cfgSliceId']"):
                    dumy_data.append({
                        "MO": "NRPMQAP",
                        "DistName": dist_name,
                        "ID": nrpmqap_id,
                        "Parameter": "cfgSliceId",
                        "value": "List"
                    })

                # sd / sst
                for item in mo.findall(".//ns:list[@name='cfgSliceId']/ns:item", ns) \
                        if ns else mo.findall(".//list[@name='cfgSliceId']/item"):
                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                        if p.attrib.get("name") in {"sd", "sst"}:
                            dumy_data.append({
                                "MO": "NRPMQAP",
                                "DistName": dist_name,
                                "ID": nrpmqap_id,
                                "Parameter": f"Item-cfgSliceId-{p.attrib.get('name')}",
                                "value": tf_to_01(p.text)
                            })


            if nrpmqap_id in group_2_ids:

                simple_params = {
                    "actCplaneCounters",
                    "actL2Counters",
                    "actPacketSchedulerCounters",
                    "cfg5qiRange",
                    "cfgProfType",
                    "thpHistDownlinkMaxRange",
                    "thpHistDownlinkMinRange",
                    "thpHistScale",
                    "thpHistUplinkMaxRange",
                    "thpHistUplinkMinRange",
                }

                for p in mo.findall(".//ns:p", ns) if ns else mo.findall(".//p"):
                    name = p.attrib.get("name")
                    if name in simple_params:
                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "ID": nrpmqap_id,
                            "Parameter": name,
                            "value": tf_to_01(p.text)   
                        })

                # cfgPlmnId
                if mo.findall(".//ns:list[@name='cfgPlmnId']", ns) if ns else mo.findall(".//list[@name='cfgPlmnId']"):
                    dumy_data.append({
                        "MO": "NRPMQAP",
                        "ID": nrpmqap_id,
                        "Parameter": "cfgPlmnId",
                        "value": "List"               
                    })

                for item in mo.findall(".//ns:list[@name='cfgPlmnId']/ns:item", ns) \
                        if ns else mo.findall(".//list[@name='cfgPlmnId']/item"):
                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "ID": nrpmqap_id,
                            "Parameter": f"Item-cfgPlmnId-{p.attrib.get('name')}",
                            "value": tf_to_01(p.text)  
                        })

                # cfgSliceId
                if mo.findall(".//ns:list[@name='cfgSliceId']", ns) if ns else mo.findall(".//list[@name='cfgSliceId']"):
                    dumy_data.append({
                        "MO": "NRPMQAP",
                        "ID": nrpmqap_id,
                        "Parameter": "cfgSliceId",
                        "value": "List"               
                    })

                for item in mo.findall(".//ns:list[@name='cfgSliceId']/ns:item", ns) \
                        if ns else mo.findall(".//list[@name='cfgSliceId']/item"):
                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                        dumy_data.append({
                            "MO": "NRPMQAP",
                            "ID": nrpmqap_id,
                            "Parameter": f"Item-cfgSliceId-{p.attrib.get('name')}",
                            "value": tf_to_01(p.text)  
        
                        })


        #for ---------- NRRESOURCEGROUP ----------
        elif mo_class == "com.nokia.srbts.nrbts:NRRESOURCEGROUP":

            dist_name = mo.attrib.get("distName", "")
            try:
                rg_id = int(dist_name.split("NRRESOURCEGROUP-")[-1])
            except:
                continue

            dumy_data.append({
                "MO": "NRRESOURCEGROUP",
                "ID": rg_id,
                "Parameter": "NRRESOURCEGROUP",
                "value": rg_id
            })

            resource_dns = []

            for item in mo.findall(".//ns:list[@name='nrResourceList']/ns:item", ns) \
                    if ns else mo.findall(".//list[@name='nrResourceList']/item"):

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    if p.attrib.get("name") == "resourceDN":
                        resource_dns.append(p.text)

            if resource_dns:
                dumy_data.append({
                    "MO": "NRRESOURCEGROUP",
                    "ID": rg_id,
                    "Parameter": "nrResourceList",
                    "value": "List"
                })

                dumy_data.append({
                    "MO": "NRRESOURCEGROUP",
                    "ID": rg_id,
                    "Parameter": "Item-nrResourceList-resourceDN",
                    "value": ";".join(resource_dns)
                })

            # ---- schedulerParams ----
            sched_params = {}

            for item in mo.findall(".//ns:list[@name='schedulerParams']/ns:item", ns) \
                    if ns else mo.findall(".//list[@name='schedulerParams']/item"):

                for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                    sched_params[p.attrib.get("name")] = p.text

            if sched_params:
                dumy_data.append({
                    "MO": "NRRESOURCEGROUP",
                    "ID": rg_id,
                    "Parameter": "schedulerParams",
                    "value": "List"
                })

                for k, v in sched_params.items():
                    dumy_data.append({
                        "MO": "NRRESOURCEGROUP",
                        "ID": rg_id,
                        "Parameter": f"Item-schedulerParams-{k}",
                        "value": tf_to_01(v)
                    })

            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                if p.attrib.get("name") == "userLabel":
                    dumy_data.append({
                        "MO": "NRRESOURCEGROUP",
                        "ID": rg_id,
                        "Parameter": "userLabel",
                        "value": p.text
                    })



        # -------- NRRESOURCEGROUP_PROFILE --------
        elif mo_class == "com.nokia.srbts.nrbts:NRRESOURCEGROUP_PROFILE":

            dist_name = mo.attrib.get("distName", "")
            try:
                profile_id = int(dist_name.split("NRRESOURCEGROUP_PROFILE-")[-1])
            except:
                continue

            # MO row
            dumy_data.append({
                "MO": "NRRESOURCEGROUP_PROFILE",
                "ID": profile_id,
                "Parameter": "NRRESOURCEGROUP_PROFILE",
                "value": profile_id
            })

            # simple p
            for p in mo.findall("ns:p", ns) if ns else mo.findall("p"):
                if p.attrib.get("name") == "applyCellRacForPrivilegedUE":
                    dumy_data.append({
                        "MO": "NRRESOURCEGROUP_PROFILE",
                        "ID": profile_id,
                        "Parameter": "applyCellRacForPrivilegedUE",
                        "value": tf_to_01(p.text)
                    })

            # -------- defaultRgNsaList --------
            if mo.findall(".//ns:list[@name='defaultRgNsaList']", ns) \
                if ns else mo.findall(".//list[@name='defaultRgNsaList']"):

                dumy_data.append({
                    "MO": "NRRESOURCEGROUP_PROFILE",
                    "ID": profile_id,
                    "Parameter": "defaultRgNsaList",
                    "value": "List"
                })

                for item in mo.findall(".//ns:list[@name='defaultRgNsaList']/ns:item", ns) \
                        if ns else mo.findall(".//list[@name='defaultRgNsaList']/item"):
                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                        dumy_data.append({
                            "MO": "NRRESOURCEGROUP_PROFILE",
                            "ID": profile_id,
                            "Parameter": f"Item-defaultRgNsaList-{p.attrib.get('name')}",
                            "value": p.text
                        })

            # -------- defaultRgSaList --------
            if mo.findall(".//ns:list[@name='defaultRgSaList']", ns) \
                if ns else mo.findall(".//list[@name='defaultRgSaList']"):

                dumy_data.append({
                    "MO": "NRRESOURCEGROUP_PROFILE",
                    "ID": profile_id,
                    "Parameter": "defaultRgSaList",
                    "value": "List"
                })

                for item in mo.findall(".//ns:list[@name='defaultRgSaList']/ns:item", ns) \
                        if ns else mo.findall(".//list[@name='defaultRgSaList']/item"):
                    for p in item.findall("ns:p", ns) if ns else item.findall("p"):
                        dumy_data.append({
                            "MO": "NRRESOURCEGROUP_PROFILE",
                            "ID": profile_id,
                            "Parameter": f"Item-defaultRgSaList-{p.attrib.get('name')}",
                            "value": p.text
                        })
         
#---------------data read in fix paratmeter---------      
        df = pd.DataFrame(dumy_data)
        for c in ["MO", "ID", "Parameter", "value"]:
            if c not in df:
                df[c] = ""

        def clean_id(x):
            if pd.isna(x) or x == "":
                return ""
            return str(int(float(x)))

        df["ID"] = df["ID"].apply(clean_id)

        normal_df = (
            df[df["Parameter"] != "cfg5qiRange"]
            .groupby(["MO", "Parameter", "value"], as_index=False)
            .agg({
                "ID": lambda x: ",".join(i for i in x if i != "")
            })
        )

        cfg_df = df[df["Parameter"] == "cfg5qiRange"].copy()
        cfg_df["pair"] = cfg_df.apply(
            lambda r: f"{r['ID']}-{r['value']}" if r["ID"] != "" else "",
            axis=1
        )

        cfg_df = (
            cfg_df.groupby(["MO", "Parameter"], as_index=False)
            .agg({
                "ID": lambda x: ",".join(i for i in x if i != ""),
                "pair": lambda x: ";".join(i for i in x if i != "")
            })
            .rename(columns={"pair": "value"})
        )

        data_df = pd.concat([normal_df, cfg_df], ignore_index=True)
        data_df = data_df[["MO", "ID", "Parameter", "value"]]
        data_df.rename(columns={"value": "value(External)"}, inplace=True)

        file_name_1 = "Nokia_Slicing_dump_data.xlsx"
        dump_output_path=os.path.join(dump_data_path, file_name_1)
        data_df.to_excel(dump_output_path, index=False, engine="openpyxl")

    #matincting-----
    excel_df = pd.read_excel(excel_file, engine="openpyxl")

    excel_df["ID"] = excel_df["ID"].apply(normalize_id)
    data_df["ID"] = data_df["ID"].apply(normalize_id)


    finaldf = excel_df.merge(
        data_df[["MO", "ID", "Parameter", "value(External)"]],
        on=["MO", "ID", "Parameter"],
        how="left"
    )

    fallback_map = (
        data_df
        .dropna(subset=["value(External)"])
        .groupby(["MO", "Parameter"])["value(External)"]
        .first()
        .to_dict()
    )

    mask_na = finaldf["value(External)"].isna()

    finaldf.loc[mask_na, "value(External)"] = finaldf.loc[mask_na].apply(
        lambda r: fallback_map.get((r["MO"], r["Parameter"])),
        axis=1
    )

    finaldf["value(External)"] = finaldf["value(External)"].apply(normalize)
    finaldf["Value(Internal)"] = finaldf["Value(Internal)"].apply(normalize)

    finaldf["Final_Value"] = finaldf["Value(Internal)"]
    finaldf["Remark"] = "No Change"

    finaldf.loc[finaldf["value(External)"].isna(), "Remark"] = "Not Found"

    mask_change = (
        finaldf["value(External)"].notna() &
        (finaldf["value(External)"] != finaldf["Value(Internal)"])
    )

    finaldf.loc[mask_change, "Final_Value"] = finaldf.loc[mask_change, "value(External)"]
    finaldf.loc[mask_change, "Remark"] = "Value Changed"

    finaldf = finaldf.sort_values(["MO", "ID", "Parameter"])
    finaldf.drop(columns=["Final_Value", "Remark"], inplace=True)
    finaldf["Remarks"] = finaldf.apply(
    lambda r: remark(r["Value(Internal)"], r["value(External)"]),
    axis=1
)
    

    file_name = "Nokia_Slicing_Final_output.xlsx"
    final_output_path=os.path.join(output_path, file_name)
    finaldf.to_excel(final_output_path, index=False, engine="openpyxl",sheet_name="Slicing")
    format_excel(final_output_path,"Slicing")
    
    # Generate download URL
    relative_path = os.path.relpath(final_output_path, MEDIA_ROOT).replace("\\", "/")
    download_url = request.build_absolute_uri(MEDIA_URL + relative_path)



    print('Excel saved, end of process' )
    return Response({
        "status": True,
        "message": "Data Successfully Parsed",
        "download_url": download_url,

     
    }, status=HTTP_200_OK)

