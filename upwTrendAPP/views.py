from django.shortcuts import render
from django.http import HttpResponse
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from rest_framework.decorators import api_view
from rest_framework.response import Response
from tkinter import *
from django.core.files.storage import FileSystemStorage
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl import workbook,load_workbook
import pandas as pd
import numpy as np
from datetime import date, timedelta
import datetime
import pandas as pd
import os


from django.shortcuts import render
from django.http import HttpResponse
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from rest_framework.decorators import api_view
from rest_framework.response import Response
from tkinter import *
from django.core.files.storage import FileSystemStorage
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl import workbook,load_workbook
import pandas as pd
import numpy as np
from datetime import date, timedelta
import datetime
import pandas as pd
import os

from django.shortcuts import render
from django.http import HttpResponse
from mcom_website.settings import MEDIA_ROOT,MEDIA_URL
from rest_framework.decorators import api_view
from rest_framework.response import Response
from tkinter import *
from django.core.files.storage import FileSystemStorage
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl import workbook,load_workbook
import pandas as pd
import numpy as np
from datetime import date, timedelta
import datetime
import pandas as pd
import os

##############################################

# df_lte_bbh=pd.read_excel(header=0,usecols=["Date",
# "Time",
# "eNodeB Name",
# "Cell FDD TDD Indication",
# "Cell Name",
# "LocalCell Id",
# "eNodeB Function Name",
# "Tracking area code",
# "RRC Success Rate(%)",
# "ERAB Success Rate(%)",
# "PS Drop(%)",
# "Average User Throughput Downlink(Mbps)(Mbps)",
# "Average User Throughput Uplink(Mbps)(Mbps)",
# "Radio Network Availability Rate(%)",
# "Intra Freq HO Success Rate",
# "Inter Freq HO Success Rate",
# "CSFB Prep SR(%)",
# "Paging Discard Rate(%)",
# "PRB utilization rate(%)",
# "Total Payload_N(GB)",
# "Average CQI",
# "PS Drop(%)",
# "Inter Freq HO Success Rate",
# ])
# df_lte_bbh

# df_volte_bbh=pd.read_excel(header=0,usecols=["Cell Name","ERAB Setup Success Rate QCI1(%)",
# "VoLTE Drop Call Rate",
# "VoLTE Packet Loss DL(%)",
# "VoLTE Packet Loss UL(%)",
# "cells SRVCC success rate",
# "VoLTE UL PUSCH SINR",
# "VoLTE HO Success Rate(IntraFreq)",
# "VoLTE HO Success Rate(InterFreq)",
# "VoLTE SRVCC IRAT Per cal rate",
# "Call Setup Success Rate (VoIP)",
# "VoLTE Packet Loss DL(%)",
# "VoLTE Packet Loss UL(%)",
# "VOLTE UL BLER",
# "VoLTE HO Success Rate(IntraFreq)",

# ])
# df_volte_bbh

# df_volte_nbh=pd.read_excel(header=0,usecols=["Cell Name","VoLTE Packet Loss DL(%)"])
# df_volte_nbh

# df_merge=pd.merge(df_volte_bbh,df_volte_nbh,on="Cell Name",how='left',suffixes=('_bbh','_nbh'))
# df_merge

# df_op=pd.merge(df_merge,df_lte_bbh,on='Cell Name',how='left',suffixes=("",'_lte_bbh'))
# df_op


###################








@api_view(['POST'])
def old_upw_trend(request):
    raw_kpi_LTE_BBH = request.FILES["raw_kpi_LTE_BBH"] if 'raw_kpi_LTE_BBH' in request.FILES else None
    raw_kpi_VOLTE_BBH = request.FILES["raw_kpi_VOLTE_BBH"] if 'raw_kpi_VOLTE_BBH' in request.FILES else None
    raw_kpi_VOLTE_NBH = request.FILES["raw_kpi_VOLTE_NBH"] if 'raw_kpi_VOLTE_NBH' in request.FILES else None

######################
    
    raw_kpi_LTE_BBH=pd.read_excel(raw_kpi_LTE_BBH ,header=0,usecols=["Date",
        "Time",
        "eNodeB Name",
        "Cell FDD TDD Indication",
        "Cell Name",
        "LocalCell Id",
        "eNodeB Function Name",
        "Tracking area code",
        "RRC Success Rate(%)",
        "ERAB Success Rate(%)",
        "PS Drop(%)",
        "Average User Throughput Downlink(Mbps)(Mbps)",
        "Average User Throughput Uplink(Mbps)(Mbps)",
        "Radio Network Availability Rate(%)",
        "Intra Freq HO Success Rate",
        "Inter Freq HO Success Rate",
        "CSFB Prep SR(%)",
        "Paging Discard Rate(%)",
        "PRB utilization rate(%)",
        "Total Payload_N(GB)",
        "Average CQI",
        "PS Drop(%)",
        "Inter Freq HO Success Rate",
        ])
    raw_kpi_LTE_BBH

    raw_kpi_VOLTE_BBH=pd.read_excel(raw_kpi_VOLTE_BBH,header=0,usecols=["Cell Name","ERAB Setup Success Rate QCI1(%)",
        "VoLTE Drop Call Rate",
        "VoLTE Packet Loss DL(%)",
        "VoLTE Packet Loss UL(%)",
        "cells SRVCC success rate",
        "VoLTE UL PUSCH SINR",
        "VoLTE HO Success Rate(IntraFreq)",
        "VoLTE HO Success Rate(InterFreq)",
        "VoLTE SRVCC IRAT Per cal rate",
        "Call Setup Success Rate (VoIP)",
        "VoLTE Packet Loss DL(%)",
        "VoLTE Packet Loss UL(%)",
        "VOLTE UL BLER",
        "VoLTE HO Success Rate(IntraFreq)",

        ])
    raw_kpi_VOLTE_BBH

    raw_kpi_VOLTE_NBH=pd.read_excel(raw_kpi_VOLTE_NBH,header=0,usecols=["Cell Name","VoLTE Packet Loss DL(%)"])
    raw_kpi_VOLTE_NBH

    df_merge=pd.merge (raw_kpi_VOLTE_BBH,raw_kpi_VOLTE_NBH,on="Cell Name",how='left',suffixes=('_BBH','_NBH'))
    df_merge

    df_op=pd.merge(df_merge, raw_kpi_LTE_BBH,on='Cell Name',how='left',suffixes=("",'_LTE_BBH'))
    df_raw_kpi =df_op

    site_list = request.FILES["site_list"] if 'site_list' in request.FILES else None
    offered_date = request.POST.get("offered_date")
    print(offered_date)
    offered_date=datetime.datetime.strptime(offered_date,"%Y-%m-%d").date()
    print("offered_date:",offered_date)

    location= MEDIA_ROOT + r"\trends\temporary_files"
    fs = FileSystemStorage(location=location)

        

    site_list = fs.save(site_list.name, site_list)
        # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
    filepath = fs.path(site_list)
    print("file_path:-",filepath)
    df_site_list=pd.read_excel(filepath)
    os.remove(path=filepath)
    print(filepath,"deleted...............")
    print(df_site_list)


    door_root= os.path.join(MEDIA_ROOT,'trends',"upw")
    path_of_blnk_temp=os.path.join(door_root,"templates","UPW_TREND_TEMPLETE.xlsx")
    trend_wb=load_workbook(path_of_blnk_temp)
    print(trend_wb.sheetnames)
    print("##################################################################################")
    # g2_site=[]
    f3_site=[]
    f1_site=[]
    t1t2_site=[]
    kpi=[       "VoLTE Packet Loss UL(%)",
                "VoLTE Packet Loss DL(%)_BBH",
                "cells SRVCC success rate",
                "VoLTE Drop Call Rate",
                "VoLTE UL PUSCH SINR",
                "VoLTE SRVCC IRAT Per cal rate",
                "VOLTE UL BLER",
                "Call Setup Success Rate (VoIP)",
                "VoLTE HO Success Rate(IntraFreq)",
                "VoLTE HO Success Rate(InterFreq)",
                "VoLTE Packet Loss DL(%)_NBH",
                "RRC Success Rate(%)",
                "ERAB Success Rate(%)",
                "PS Drop(%)",
                "CSFB Prep SR(%)",
                "Total Payload_N(GB)",
                "Average CQI",
                "PRB utilization rate(%)",
                "Average User Throughput Downlink(Mbps)(Mbps)",
                "Average User Throughput Uplink(Mbps)(Mbps)",
                "Radio Network Availability Rate(%)",
                "Paging Discard Rate(%)",
                "Inter Freq HO Success Rate",
                "Intra Freq HO Success Rate",
                "ERAB Setup Success Rate QCI1(%)",
        ]

    kpi1=[      "VoLTE Packet Loss UL(%)",
                "VoLTE Packet Loss DL(%)_BBH",
                "cells SRVCC success rate",
                "VoLTE Drop Call Rate",
                "VoLTE UL PUSCH SINR",
                "VoLTE SRVCC IRAT Per cal rate",
                "VOLTE UL BLER",
                "Call Setup Success Rate (VoIP)",
                "VoLTE HO Success Rate(IntraFreq)",
                "VoLTE HO Success Rate(InterFreq)",
                "VoLTE Packet Loss DL(%)_NBH",
                "RRC Success Rate(%)",
                "ERAB Success Rate(%)",
                "PS Drop(%)",
                "CSFB Prep SR(%)",
                "Total Payload_N(GB)",
                "Average CQI",
                "PRB utilization rate(%)",
                "Average User Throughput Downlink(Mbps)(Mbps)",
                "Average User Throughput Uplink(Mbps)(Mbps)",
                "Radio Network Availability Rate(%)",
                "Paging Discard Rate(%)",
                "Inter Freq HO Success Rate",
                "Intra Freq HO Success Rate",
                "ERAB Setup Success Rate QCI1(%)",
        ]

    ########################## the below code is to replace every string to zero from numeric columns ##################

    for x in kpi1:
            df_raw_kpi[x] = df_raw_kpi[x].replace(to_replace='.*', value=0, regex=True)
    ######################################  * * * * * * * * * * * * * * ################################################

    site_list=list(df_site_list["2G ID"])
    df_raw_kpi["Cell Name"] =df_raw_kpi["Cell Name"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..


    print("__________________raw KPI after converting str to zero_______________")
    print(df_raw_kpi)

    lis=list(df_raw_kpi["Cell Name"])
    techlist=[]
    sit_id_lis=[]
    for item in lis:
        if("_" in item):
            sit_id=item.split("_")[-2][:-1]
            sit_id_lis.append(sit_id)
        if  ('_F1_' in item or '_F3_' in item  or '_T1_' in item or '_T2_' in item):
                    if ('_F1_' in item):
                        tech='L2100'
                        # techlist.append(tech)

                    if ('_F3_' in item):
                        tech='L1800'
                        # techlist.append(tech)

                    

                    if ('_T2_' in item or '_T1_' in item):
                        tech='TDD'
                    techlist.append(tech) 

        else:
                    tech=item
                    techlist.append(tech)             

    df_raw_kpi.insert(0,'Site_ID',sit_id_lis)
    df_raw_kpi.insert(1,'tech',techlist)

    df_raw_kpi.rename(columns={"Cell Name" :"CellName" } ,inplace = True )
    df_raw_kpi.fillna(value=0,inplace=True)


    process_op_path=os.path.join(door_root,"process_output")

    savepath=os.path.join(process_op_path,"desired_input.xlsx")
    df_raw_kpi.to_excel(savepath)

    date1=offered_date
    # date1=date.today()
    dt1 = date1 - timedelta(1)
    dt2 = date1 - timedelta(2)
    dt3 = date1 - timedelta(3)
    dt4 = date1 - timedelta(4)
    dt5 = date1 - timedelta(5)
    ls=[dt1,dt2,dt3,dt4,dt5]
    only_site_fil = df_raw_kpi[(df_raw_kpi.Site_ID.isin(site_list)) & (df_raw_kpi.Date.isin(ls))]

    savepath=os.path.join(process_op_path,"only_site_date_filtered_input.xlsx")
    only_site_fil.to_excel(savepath)

    def perticular_tech( tech,site_list):
        # df_filtered = df_raw_kpi[(df_raw_kpi.SITE_ID.isin(site_list)) & (df_raw_kpi.date.isin(ls))]
        df_filtered = df_raw_kpi[(df_raw_kpi.Site_ID.isin(site_list)) & (df_raw_kpi.CellName.str.contains('|'.join(tech)))]
            # print(df_filtered)
        # PsOs_Filtr=os.path.join(door_root,'process_output','filtered.xlsx')
        # df_filtered.to_excel(PsOs_Filtr,index=False)
        # df_pivoted = df_filtered.pivot_table(index=['CellName','Site_ID','tech'], columns="Date")
        # print(df_pivoted)
        # PsOs_pivot=os.path.join(door_root,'process_output','npivoted.xlsx')
        # df_pivoted.to_excel("process_output","npivoted.xlsx")
        # df_pivoted.to_excel(PsOs_pivot)
        # print('---------------filter,pivot----------------',df_pivoted)
        # return df_filtered,df_pivoted
        print(df_filtered)
        if not df_filtered.empty:
            address="last_filtered_input"  +str(tech) + ".xlsx"
            savepath=os.path.join(process_op_path,address)
            df_filtered.to_excel(savepath)
            df_pivoted = df_filtered.pivot_table(index=['CellName','Site_ID','tech'], columns="Date")
            print("technology:",tech)
            print(df_pivoted)
            address_pivot="pivoted_input"  +str(tech)+ ".xlsx"
            savepath=os.path.join(process_op_path,address_pivot)
            df_pivoted.to_excel(savepath)
            return df_pivoted
        
        return df_filtered

    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def num_hash(num):
            if num < 26:
                return alpha[num-1]
            else:
                q, r = num//26, num % 26
                if r == 0:
                    if q == 1:
                        return alpha[r-1]
                    else:
                        return num_hash(q-1) + alpha[r-1]
                else:
                    return num_hash(q) + alpha[r-1]

    # Driver code

    # printString(27906)

    def titleToNumber(s):
    # This process is similar to binary-to-
    # decimal conversion
        result = 0
        for B in range(len(s)):
            result *= 26
            result += ord(s[B]) - ord('A') + 1
        return result


    def overwrite(df_pivoted,kpi_name,coln1,trend_ws):
        coln2=num_hash(titleToNumber(coln1)+1)
        coln3=num_hash(titleToNumber(coln1)+2)
        coln4=num_hash(titleToNumber(coln1)+3)
        coln5=num_hash(titleToNumber(coln1)+4)
        print(kpi_name)
        index_pivot=df_pivoted.index.to_list()
        print("index ;###############################",index_pivot)
        print(len(index_pivot))
        print("index of pivoted table: ",index_pivot)
        
        dr=df_pivoted[kpi_name]
        

        print("columns of dr table",dr.columns)
        cl=dr.columns.to_list()
        print("column list",cl)

        col1=dr[str(cl[0])].to_list()
        col2=dr[str(cl[1])].to_list()
        col3=dr[str(cl[2])].to_list()
        col4=dr[str(cl[3])].to_list()
        col5=dr[str(cl[4])].to_list()

        trend_ws[coln1+"3"].value=cl[0]
        trend_ws[coln2+"3"].value=cl[1]
        trend_ws[coln3+"3"].value=cl[2]
        trend_ws[coln4+"3"].value=cl[3]
        trend_ws[coln5+"3"].value=cl[4]

        for i,value in enumerate(index_pivot):
            j=i+4
            trend_ws['A'+str(j)].value='UW'
            # ws['I'+str(j)].value='GOVINDPUR'
            # ws['N'+str(j)].value='NOKIA'
            # ws['O'+str(j)].value='Relocation'
            trend_ws['K'+str(j)].value=date1

            trend_ws['L'+str(j)].value='DONE'
            trend_ws['B'+str(j)].value=index_pivot[i][1]
            trend_ws['I'+str(j)].value=index_pivot[i][2]
            trend_ws['C'+str(j)].value=index_pivot[i][0]
            # ws['E'+str(j)].value=index[i][2]
            # ws['F'+str(j)].value=index[i][3]
            trend_ws['G'+str(j)].value=index_pivot[i][0]
            trend_ws['H'+str(j)].value=index_pivot[i][1]
            # ws['J'+str(j)].value=index[i][5]



            
            trend_ws[coln1+str(j)].value=col1[i]
            trend_ws[coln2+str(j)].value=col2[i]
            trend_ws[coln3+str(j)].value=col3[i]
            trend_ws[coln4+str(j)].value=col4[i]
            trend_ws[coln5+str(j)].value=col5[i]

    #for fdd
    pivot_fdd=perticular_tech(["_F3_"],site_list)
    # path_of_blnk_temp=os.path.join(door_root,'templates','UPW_TREND_TEMPLETE.xlsx')
    # trend_wb=openpyxl.load_workbook(path_of_blnk_temp) 
    trend_ws=trend_wb["FDD"]
    if not pivot_fdd.empty:
        for kpi_name in kpi:
            if(kpi_name=='RRC Success Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'P',trend_ws)

            if(kpi_name=='ERAB Success Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'W',trend_ws)
            if(kpi_name=='PS Drop(%)'):
                overwrite(pivot_fdd,kpi_name,'AD',trend_ws)

            if(kpi_name=='Average User Throughput Downlink(Mbps)(Mbps)'):
                overwrite(pivot_fdd,kpi_name,'AK',trend_ws)

            if(kpi_name=='Average User Throughput Uplink(Mbps)(Mbps)'):
                overwrite(pivot_fdd,kpi_name,'AS',trend_ws) 
                
            if(kpi_name=='Radio Network Availability Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'AZ',trend_ws) 

            if(kpi_name=='Intra Freq HO Success Rate'):
                overwrite(pivot_fdd,kpi_name,'BN',trend_ws) 

            if(kpi_name=='Inter Freq HO Success Rate'):
                overwrite(pivot_fdd,kpi_name,'BU',trend_ws) 

            if(kpi_name=='CSFB Prep SR(%)'):
                overwrite(pivot_fdd,kpi_name,'CB',trend_ws) 

            if(kpi_name=='Paging Discard Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'CI',trend_ws)    

            if(kpi_name=='PRB utilization rate(%)'):
                overwrite(pivot_fdd,kpi_name,'CP',trend_ws) 

            if(kpi_name=='Total Payload_N(GB)'):
                overwrite(pivot_fdd,kpi_name,'CW',trend_ws) 

            if(kpi_name=='ERAB Success Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'DD',trend_ws)   

            if(kpi_name=='VoLTE Drop Call Rate'):
                overwrite(pivot_fdd,kpi_name,'DK',trend_ws)  

            if(kpi_name=='VoLTE Packet Loss DL(%)_BBH'):
                overwrite(pivot_fdd,kpi_name,'DR',trend_ws)  

            if(kpi_name=='VoLTE Packet Loss UL(%)'):
                overwrite(pivot_fdd,kpi_name,'DY',trend_ws)   

            if(kpi_name=='cells SRVCC success rate'):
                overwrite(pivot_fdd,kpi_name,'EF',trend_ws)  

            if(kpi_name=='VoLTE UL PUSCH SINR'):
                overwrite(pivot_fdd,kpi_name,'EM',trend_ws)  

            if(kpi_name=='VoLTE Packet Loss DL(%)_NBH'):
                overwrite(pivot_fdd,kpi_name,'ET',trend_ws)    

            if(kpi_name=='VoLTE HO Success Rate(IntraFreq)'):
                overwrite(pivot_fdd,kpi_name,'FA',trend_ws) 

            if(kpi_name=='VoLTE HO Success Rate(InterFreq)'):
                overwrite(pivot_fdd,kpi_name,'FH',trend_ws) 
            if(kpi_name=='VoLTE SRVCC IRAT Per cal rate'):
                overwrite(pivot_fdd,kpi_name,'FO',trend_ws) 
    #for fdd
    pivot_fdd=perticular_tech(["_T1_","_T2_"],site_list)
    trend_ws=trend_wb["TDD"]
    if not pivot_fdd.empty:
        for kpi_name in kpi:
            if(kpi_name=='RRC Success Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'P',trend_ws)

            if(kpi_name=='ERAB Success Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'W',trend_ws)
            if(kpi_name=='PS Drop(%)'):
                overwrite(pivot_fdd,kpi_name,'AD',trend_ws)

            if(kpi_name=='Average User Throughput Downlink(Mbps)(Mbps)'):
                overwrite(pivot_fdd,kpi_name,'AK',trend_ws)

            if(kpi_name=='Average User Throughput Uplink(Mbps)(Mbps)'):
                overwrite(pivot_fdd,kpi_name,'AR',trend_ws) 
                
            if(kpi_name=='Radio Network Availability Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'AY',trend_ws) 

            if(kpi_name=='Average CQI'):
                overwrite(pivot_fdd,kpi_name,'BF',trend_ws) 

            if(kpi_name=='Intra Freq HO Success Rate'):
                overwrite(pivot_fdd,kpi_name,'BT',trend_ws) 

            if(kpi_name=='Inter Freq HO Success Rate'):
                overwrite(pivot_fdd,kpi_name,'CA',trend_ws) 

            if(kpi_name=='CSFB Prep SR(%)'):
                overwrite(pivot_fdd,kpi_name,'CH',trend_ws) 

            if(kpi_name=='Paging Discard Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'CO',trend_ws)    

            if(kpi_name=='PRB utilization rate(%)'):
                overwrite(pivot_fdd,kpi_name,'CV',trend_ws) 

            if(kpi_name=='Total Payload_N(GB)'):
                overwrite(pivot_fdd,kpi_name,'DC',trend_ws) 

            if(kpi_name=='ERAB Success Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'DJ',trend_ws)   

            if(kpi_name=='VoLTE Drop Call Rate'):
                overwrite(pivot_fdd,kpi_name,'DQ',trend_ws)  

            if(kpi_name=='VoLTE Packet Loss DL(%)_BBH'):
                overwrite(pivot_fdd,kpi_name,'DX',trend_ws)  

            if(kpi_name=='VoLTE Packet Loss UL(%)'):
                overwrite(pivot_fdd,kpi_name,'EE',trend_ws)   

            if(kpi_name=='cells SRVCC success rate'):
                overwrite(pivot_fdd,kpi_name,'EL',trend_ws)  

            if(kpi_name=='VoLTE UL PUSCH SINR'):
                overwrite(pivot_fdd,kpi_name,'ES',trend_ws)  

            if(kpi_name=='VoLTE Packet Loss DL(%)_NBH'):
                overwrite(pivot_fdd,kpi_name,'EZ',trend_ws)    

            if(kpi_name=='VoLTE HO Success Rate(IntraFreq)'):
                overwrite(pivot_fdd,kpi_name,'FG',trend_ws) 

            if(kpi_name=='VoLTE HO Success Rate(InterFreq)'):
                overwrite(pivot_fdd,kpi_name,'FN',trend_ws) 
            if(kpi_name=='VoLTE SRVCC IRAT Per cal rate'):
                overwrite(pivot_fdd,kpi_name,'FU',trend_ws) 

    #for fdd
    pivot_fdd=perticular_tech(["_F1_"],site_list)
    trend_ws=trend_wb["L900 L2100"]
    if not pivot_fdd.empty:
        for kpi_name in kpi:
            if(kpi_name=='RRC Success Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'P',trend_ws)

            if(kpi_name=='ERAB Success Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'W',trend_ws)
            if(kpi_name=='PS Drop(%)'):
                overwrite(pivot_fdd,kpi_name,'AD',trend_ws)

            if(kpi_name=='Intra Freq HO Success Rate'):
                overwrite(pivot_fdd,kpi_name,'AK',trend_ws)

            if(kpi_name=='Inter Freq HO Success Rate'):
                overwrite(pivot_fdd,kpi_name,'AR',trend_ws) 
            if(kpi_name=='Call Setup Success Rate (VoIP)'):
                overwrite(pivot_fdd,kpi_name,'AY',trend_ws) 
                
            if(kpi_name=='VoLTE Drop Call Rate'):
                overwrite(pivot_fdd,kpi_name,'BF',trend_ws) 

            if(kpi_name=='Average User Throughput Downlink(Mbps)(Mbps)'):
                overwrite(pivot_fdd,kpi_name,'BM',trend_ws) 

            if(kpi_name=='VoLTE Packet Loss DL(%)_BBH'):
                overwrite(pivot_fdd,kpi_name,'BT',trend_ws) 

            if(kpi_name=='VoLTE Packet Loss UL(%)'):
                overwrite(pivot_fdd,kpi_name,'CA',trend_ws) 

            if(kpi_name=='VOLTE UL BLER'):
                overwrite(pivot_fdd,kpi_name,'CH',trend_ws) 

            if(kpi_name=='VoLTE HO Success Rate(IntraFreq)'):
                overwrite(pivot_fdd,kpi_name,'CO',trend_ws)    

            if(kpi_name=='PRB utilization rate(%)'):
                overwrite(pivot_fdd,kpi_name,'CV',trend_ws) 

            if(kpi_name=='ERAB Setup Success Rate QCI1(%)'):
                overwrite(pivot_fdd,kpi_name,'DC',trend_ws) 

            if(kpi_name=='cells SRVCC success rate'):
                overwrite(pivot_fdd,kpi_name,'DJ',trend_ws)   

            if(kpi_name=='VoLTE UL PUSCH SINR'):
                overwrite(pivot_fdd,kpi_name,'DQ',trend_ws)  

            if(kpi_name=='VoLTE Packet Loss DL(%)_NBH'):
                overwrite(pivot_fdd,kpi_name,'DX',trend_ws)  

            if(kpi_name=='VoLTE HO Success Rate(InterFreq)'):
                overwrite(pivot_fdd,kpi_name,'EE',trend_ws)   

            if(kpi_name=='VoLTE SRVCC IRAT Per cal rate'):
                overwrite(pivot_fdd,kpi_name,'EL',trend_ws)  

            if(kpi_name=='Radio Network Availability Rate(%)'):
                overwrite(pivot_fdd,kpi_name,'ES',trend_ws)  

        







        output_path=os.path.join(door_root,"output","UPW_4G_TECHNOLOGYWISE_trend_output.xlsx")
        trend_wb.save(output_path)

        download_path=os.path.join(MEDIA_URL,"trends","upw","output","UPW_4G_TECHNOLOGYWISE_trend_output.xlsx")
        return Response({"status":True,"message":"Succesfully uploaded","Download_url":download_path})

       
        
    # site_list = request.FILES["site_list"] if 'site_list' in request.FILES else None
    # offered_date = request.POST.get("offered_date")
    # print(offered_date)
    # offered_date=datetime.datetime.strptime(offered_date,"%Y-%m-%d").date()
    # print("offered_date:",offered_date)

    # location= MEDIA_ROOT + r"\trends\temporary_files"
    # fs = FileSystemStorage(location=location)

    

    # site_list = fs.save(site_list.name, site_list)
    #     # the fileurl variable now contains the url to the file. This can be used to serve the file when needed.
    # filepath = fs.path(site_list)
    # print("file_path:-",filepath)
    # df_site_list=pd.read_excel(filepath)
    # os.remove(path=filepath)
    # print(filepath,"deleted...............")
    # print(df_site_list)


    # door_root= os.path.join(MEDIA_ROOT,'trends',"UP_West_4G")
    # path_of_blnk_temp=os.path.join(door_root,"templates","UPW_TREND_TEMPLETE.xlsx")
    # trend_wb=load_workbook(path_of_blnk_temp)
    # print(trend_wb.sheetnames)

    
   





    
    # g2_site=[]
    # f3_site=[]
    # t1t2_site=[]
    # f1_site=[]

    # print("##################################################################################")
    # kpi = ["VoLTE Packet Loss UL(%)",
    #        "VoLTE Packet Loss DL(%)_BBH",
    #         "cells SRVCC success rate",
    #         "VoLTE Drop Call Rate",
    #         "VoLTE UL PUSCH SINR",
    #         "VoLTE SRVCC IRAT Per cal rate",
    #         "VOLTE UL BLER",
    #         "Call Setup Success Rate (VoIP)",
    #         "VoLTE HO Success Rate(IntraFreq)",
    #         "VoLTE HO Success Rate(InterFreq)",
    #         "VoLTE Packet Loss DL(%)_NBH",
    #         "RRC Success Rate(%)",
    #         "ERAB Success Rate(%)",
    #         "PS Drop(%)",
    #         "CSFB Prep SR(%)",
    #         "Total Payload_N(GB)",
    #         "Average CQI",
    #         "PRB utilization rate(%)",
    #         "Average User Throughput Downlink(Mbps)(Mbps)",
    #         "Average User Throughput Uplink(Mbps)(Mbps)",
    #         "Radio Network Availability Rate(%)",
    #         "Paging Discard Rate(%)",
    #         "Inter Freq HO Success Rate",
    #         "Intra Freq HO Success Rate",
    #         "ERAB Setup Success Rate QCI1(%)",
    # ]

    # kpi1 = ["VoLTE Packet Loss UL(%)",
    #        "VoLTE Packet Loss DL(%)_BBH",
    #         "cells SRVCC success rate",
    #         "VoLTE Drop Call Rate",
    #         "VoLTE UL PUSCH SINR",
    #         "VoLTE SRVCC IRAT Per cal rate",
    #         "VOLTE UL BLER",
    #         "Call Setup Success Rate (VoIP)",
    #         "VoLTE HO Success Rate(IntraFreq)",
    #         "VoLTE HO Success Rate(InterFreq)",
    #         "VoLTE Packet Loss DL(%)_NBH",
    #         "RRC Success Rate(%)",
    #         "ERAB Success Rate(%)",
    #         "PS Drop(%)",
    #         "CSFB Prep SR(%)",
    #         "Total Payload_N(GB)",
    #         "Average CQI",
    #         "PRB utilization rate(%)",
    #         "Average User Throughput Downlink(Mbps)(Mbps)",
    #         "Average User Throughput Uplink(Mbps)(Mbps)",
    #         "Radio Network Availability Rate(%)",
    #         "Paging Discard Rate(%)",
    #         "Inter Freq HO Success Rate",
    #         "Intra Freq HO Success Rate",
    #         "ERAB Setup Success Rate QCI1(%)",
    # ]
    # for x in kpi1:
    #          df_raw_kpi[x] = df_raw_kpi[x].replace(to_replace='.*', value=0, regex=True)


    # site_list=list(df_site_list["Site ID"])
    # df_raw_kpi["Cell Name"] =df_raw_kpi["Cell Name"].apply(lambda x: x.strip()) # to remove all tralling spaces..... and leading spaces..




    # lis=list(df_raw_kpi["Cell Name"])
    # techlist=[]
    # sit_id_lis=[]
    # # cell_id_lis=[]
    # for item in lis:
    #     if("_" in item):
    #         # cell_id=item.split(" ")[0]
    #         # ln=len(item.split("_")[-1])
    #         #print(ln)
    #         sit_id=item.split("_")[-2][:-1]
    #         sit_id_lis.append(sit_id)

    #     if  ('_F1_' in item or '_F3_' in item  or '_T1_' in item or '_T2_' in item):
    #                 if ('_F1_' in item):
    #                     tech='L2100'
    #                     # techlist.append(tech)

    #                 if ('_F3_' in item):
    #                     tech='L1800'
    #                     # techlist.append(tech)

                    

    #                 if ('_T2_' in item or '_T1_' in item):
    #                     tech='TDD'
    #                 techlist.append(tech) 

    #     else:
    #                 tech=item
    #                 techlist.append(tech)             

    # df_raw_kpi.insert(0,'Site ID',sit_id_lis)
    # df_raw_kpi.insert(1,'tech',techlist)
    # df_raw_kpi.fillna(value=0,inplace=True)
    # # df_raw_kpi.to_excel('process_output/final.xlsx')
    # # process_op_path=os.path.join(door_root,"process_output")
        
    # # savepath=os.path.join(process_op_path,"final.xlsx")
    # # df_raw_kpi.to_excel(savepath)
    # process_op_path=os.path.join(door_root,"process_output",'desired input.xlsx')

    # df_raw_kpi.to_excel(process_op_path, index=False)
    # excel_file_1 = process_op_path

    # df1 = pd.read_excel(excel_file_1)


    # df1.rename(columns={"Site ID": "SITE_ID"}, inplace=True)
   

    # kpi = ["VoLTE Packet Loss UL(%)",
    #        "VoLTE Packet Loss DL(%)_BBH",
    #         "cells SRVCC success rate",
    #         "VoLTE Drop Call Rate",
    #         "VoLTE UL PUSCH SINR",
    #         "VoLTE SRVCC IRAT Per cal rate",
    #         "VOLTE UL BLER",
    #         "Call Setup Success Rate (VoIP)",
    #         "VoLTE HO Success Rate(IntraFreq)",
    #         "VoLTE HO Success Rate(InterFreq)",
    #         "VoLTE Packet Loss DL(%)_NBH",
    #         "RRC Success Rate(%)",
    #         "ERAB Success Rate(%)",
    #         "PS Drop(%)",
    #         "CSFB Prep SR(%)",
    #         "Total Payload_N(GB)",
    #         "Average CQI",
    #         "PRB utilization rate(%)",
    #         "Average User Throughput Downlink(Mbps)(Mbps)",
    #         "Average User Throughput Uplink(Mbps)(Mbps)",
    #         "Radio Network Availability Rate(%)",
    #         "Paging Discard Rate(%)",
    #         "Inter Freq HO Success Rate",
    #         "Intra Freq HO Success Rate",
    #         "ERAB Setup Success Rate QCI1(%)",
    #         # "Radio Network Availability Rate(%)",
    #          ]
    






    
    # filtered_df_1 = df1[(df1.SITE_ID.isin(list(df_site_list['SITE_ID'])))]

    # print(filtered_df_1)
    # upwPath=os.path.join(door_root,'process_output','filtered1.xlsx')
    
    


    # filtered_df_1.to_excel(upwPath,index=False)       
    # # print(df)
    
    # df1 = pd.read_excel(upwPath)
    
    # df_pivot = df1.pivot_table(columns='Date', index=['Cell Name','SITE_ID',
    #         'tech'])
    
    #     #  aggfunc=np.sum)
    #     # print('df_pivot')
    # upwPathPathpivot=os.path.join(door_root,'process_output','pivot.xlsx')
    # df_pivot.to_excel(upwPathPathpivot)

    





    


    # alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    # def num_hash(num):
    #             if num < 26:
    #                 return alpha[num-1]
    #             else:
    #                 q, r = num//26, num % 26
    #                 if r == 0:
    #                     if q == 1:
    #                         return alpha[r-1]
    #                     else:
    #                         return num_hash(q-1) + alpha[r-1]
    #                 else:
    #                     return num_hash(q) + alpha[r-1]
        


    # def titleToNumber(s):
    # # This process is similar to binary-to-
    # # decimal conversion
    #     result = 0
    #     for B in range(len(s)):
    #         result *= 26
    #         result += ord(s[B]) - ord('A') + 1
    #     return result


    

            
            
        





    # date1=offered_date
    # # date1=cal.get_date()
    # d1=date1-timedelta(1)
    # d2=date1-timedelta(2)
    # d3=date1-timedelta(3)
    # d4=date1-timedelta(4)
    # d5=date1-timedelta(5)
    # cl=[d1,d2,d3,d4,d5]
    # # index=df_pivot.index
    

    # def overwrite(kpi_name,coln1):
    #     coln2=num_hash(titleToNumber(coln1)+1)
    #     coln3=num_hash(titleToNumber(coln1)+2)
    #     coln4=num_hash(titleToNumber(coln1)+3)
    #     coln5=num_hash(titleToNumber(coln1)+4)
    #     print(kpi_name)
    #     index=df_pivot.index
    #     # print("index ;###############################",index_pivot)
    #     # print(len(index_pivot))
    #     print("doooooonnenn")
    #     dr=df_pivot[kpi_name]
    #     print("columns of dr tab")
    #     li=dr.columns
    #     print("column list")



    #     # dr=df_pivot[kpi_name]
    #     # li=dr.columns
    
    
        



    #     col1=dr[li[0]].to_list()
    #     col2=dr[li[1]].to_list()
    #     col3=dr[li[2]].to_list()
    #     col4=dr[li[3]].to_list()
    #     col5=dr[li[4]].to_list()


    #     trend_ws[coln1+"3"].value=cl[4]
    #     trend_ws[coln2+"3"].value=cl[3]
    #     trend_ws[coln3+"3"].value=cl[2]
    #     trend_ws[coln4+"3"].value=cl[1]
    #     trend_ws[coln5+"3"].value=cl[0]

    #     for i,value in enumerate(index):
    #         j=i+4
    #         trend_ws['A'+str(j)].value='UW'
    #         # ws['I'+str(j)].value='GOVINDPUR'
    #         # ws['N'+str(j)].value='NOKIA'
    #         # ws['O'+str(j)].value='Relocation'
    #         trend_ws['K'+str(j)].value=date1

    #         trend_ws['L'+str(j)].value='DONE'
    #         trend_ws['B'+str(j)].value=index[i][1]
    #         trend_ws['I'+str(j)].value=index[i][2]
    #         trend_ws['C'+str(j)].value=index[i][0]
    #         # ws['E'+str(j)].value=index[i][2]
    #         # ws['F'+str(j)].value=index[i][3]
    #         trend_ws['G'+str(j)].value=index[i][0]
    #         trend_ws['H'+str(j)].value=index[i][1]
    #         # ws['J'+str(j)].value=index[i][5]



            
    #         trend_ws[coln1+str(j)].value=col1[i]
    #         trend_ws[coln2+str(j)].value=col2[i]
    #         trend_ws[coln3+str(j)].value=col3[i]
    #         trend_ws[coln4+str(j)].value=col4[i]
    #         trend_ws[coln5+str(j)].value=col5[i]


    # # def overwrite_col(kpi_name,coln):
    # #     index=df_pivot.index
    # #     print('single celll')
    # #     dr=df_pivot[kpi_name] 
    # #     cl=dr.columns.to_list()
    # #     data_list=dr[str(cl[0])].to_list()
    # #     print(data_list)
    # #     ws[coln+"4"].value=cl[0]

    # #     for i,value in enumerate(index):
    # #         j=i+6
        
    #         # ws[coln+str(j)].value=data_list[i]
        

   


    # trend_ws=trend_wb["FDD"]
    # # wb = openpyxl.load_workbook(trend_ws)
    # # ws = wb["FDD"]
    # for kpi_name in kpi:
    #         if(kpi_name=='RRC Success Rate(%)'):
    #             overwrite(kpi_name,'P')

    #         if(kpi_name=='ERAB Success Rate(%)'):
    #             overwrite(kpi_name,'W')
    #         if(kpi_name=='PS Drop(%)'):
    #             overwrite(kpi_name,'AD')

    #         if(kpi_name=='Average User Throughput Downlink(Mbps)(Mbps)'):
    #             overwrite(kpi_name,'AK')

    #         if(kpi_name=='Average User Throughput Uplink(Mbps)(Mbps)'):
    #             overwrite(kpi_name,'AS') 
                
    #         if(kpi_name=='Radio Network Availability Rate(%)'):
    #             overwrite(kpi_name,'AZ') 

    #         if(kpi_name=='Intra Freq HO Success Rate'):
    #             overwrite(kpi_name,'BN') 

    #         if(kpi_name=='Inter Freq HO Success Rate'):
    #             overwrite(kpi_name,'BU') 

    #         if(kpi_name=='CSFB Prep SR(%)'):
    #             overwrite(kpi_name,'CB') 

    #         if(kpi_name=='Paging Discard Rate(%)'):
    #             overwrite(kpi_name,'CI')    

    #         if(kpi_name=='PRB utilization rate(%)'):
    #             overwrite(kpi_name,'CP') 

    #         if(kpi_name=='Total Payload_N(GB)'):
    #             overwrite(kpi_name,'CW') 

    #         if(kpi_name=='ERAB Success Rate(%)'):
    #             overwrite(kpi_name,'DD')   

    #         if(kpi_name=='VoLTE Drop Call Rate'):
    #             overwrite(kpi_name,'DK')  

    #         if(kpi_name=='VoLTE Packet Loss DL(%)_BBH'):
    #             overwrite(kpi_name,'DR')  

    #         if(kpi_name=='VoLTE Packet Loss UL(%)'):
    #             overwrite(kpi_name,'DY')   

    #         if(kpi_name=='cells SRVCC success rate'):
    #             overwrite(kpi_name,'EF')  

    #         if(kpi_name=='VoLTE UL PUSCH SINR'):
    #             overwrite(kpi_name,'EM')  

    #         if(kpi_name=='VoLTE Packet Loss DL(%)_NBH'):
    #             overwrite(kpi_name,'ET')    

    #         if(kpi_name=='VoLTE HO Success Rate(IntraFreq)'):
    #             overwrite(kpi_name,'FA') 

    #         if(kpi_name=='VoLTE HO Success Rate(InterFreq)'):
    #             overwrite(kpi_name,'FH') 
    #         if(kpi_name=='VoLTE SRVCC IRAT Per cal rate'):
    #             overwrite(kpi_name,'FO') 


    # # STR='templates/UPW_TREND_TEMPLETE.xlsx'
    # # wb = openpyxl.load_workbook(STR)
    # trend_ws=trend_wb["TDD"]
    # for kpi_name in kpi:
    #         if(kpi_name=='RRC Success Rate(%)'):
    #             overwrite(kpi_name,'P')

    #         if(kpi_name=='ERAB Success Rate(%)'):
    #             overwrite(kpi_name,'W')
    #         if(kpi_name=='PS Drop(%)'):
    #             overwrite(kpi_name,'AD')

    #         if(kpi_name=='Average User Throughput Downlink(Mbps)(Mbps)'):
    #             overwrite(kpi_name,'AK')

    #         if(kpi_name=='Average User Throughput Uplink(Mbps)(Mbps)'):
    #             overwrite(kpi_name,'AR') 
                
    #         if(kpi_name=='Radio Network Availability Rate(%)'):
    #             overwrite(kpi_name,'AY') 

    #         if(kpi_name=='Average CQI'):
    #             overwrite(kpi_name,'BF') 

    #         if(kpi_name=='Intra Freq HO Success Rate'):
    #             overwrite(kpi_name,'BT') 

    #         if(kpi_name=='Inter Freq HO Success Rate'):
    #             overwrite(kpi_name,'CA') 

    #         if(kpi_name=='CSFB Prep SR(%)'):
    #             overwrite(kpi_name,'CH') 

    #         if(kpi_name=='Paging Discard Rate(%)'):
    #             overwrite(kpi_name,'CO')    

    #         if(kpi_name=='PRB utilization rate(%)'):
    #             overwrite(kpi_name,'CV') 

    #         if(kpi_name=='Total Payload_N(GB)'):
    #             overwrite(kpi_name,'DC') 

    #         if(kpi_name=='ERAB Success Rate(%)'):
    #             overwrite(kpi_name,'DJ')   

    #         if(kpi_name=='VoLTE Drop Call Rate'):
    #             overwrite(kpi_name,'DQ')  

    #         if(kpi_name=='VoLTE Packet Loss DL(%)_BBH'):
    #             overwrite(kpi_name,'DX')  

    #         if(kpi_name=='VoLTE Packet Loss UL(%)'):
    #             overwrite(kpi_name,'EE')   

    #         if(kpi_name=='cells SRVCC success rate'):
    #             overwrite(kpi_name,'EL')  

    #         if(kpi_name=='VoLTE UL PUSCH SINR'):
    #             overwrite(kpi_name,'ES')  

    #         if(kpi_name=='VoLTE Packet Loss DL(%)_NBH'):
    #             overwrite(kpi_name,'EZ')    

    #         if(kpi_name=='VoLTE HO Success Rate(IntraFreq)'):
    #             overwrite(kpi_name,'FG') 

    #         if(kpi_name=='VoLTE HO Success Rate(InterFreq)'):
    #             overwrite(kpi_name,'FN') 
    #         if(kpi_name=='VoLTE SRVCC IRAT Per cal rate'):
    #             overwrite(kpi_name,'FU') 
 
    # trend_ws=trend_wb["L900 L2100"]
    # for kpi_name in kpi:
    #         if(kpi_name=='RRC Success Rate(%)'):
    #             overwrite(kpi_name,'P')

    #         if(kpi_name=='ERAB Success Rate(%)'):
    #             overwrite(kpi_name,'W')
    #         if(kpi_name=='PS Drop(%)'):
    #             overwrite(kpi_name,'AD')

    #         if(kpi_name=='Intra Freq HO Success Rate'):
    #             overwrite(kpi_name,'AK')

    #         if(kpi_name=='Inter Freq HO Success Rate'):
    #             overwrite(kpi_name,'AR') 
    #         if(kpi_name=='Call Setup Success Rate (VoIP)'):
    #             overwrite(kpi_name,'AY') 
                
    #         if(kpi_name=='VoLTE Drop Call Rate'):
    #             overwrite(kpi_name,'BF') 

    #         if(kpi_name=='Average User Throughput Downlink(Mbps)(Mbps)'):
    #             overwrite(kpi_name,'BM') 

    #         if(kpi_name=='VoLTE Packet Loss DL(%)_BBH'):
    #             overwrite(kpi_name,'BT') 

    #         if(kpi_name=='VoLTE Packet Loss UL(%)'):
    #             overwrite(kpi_name,'CA') 

    #         if(kpi_name=='VOLTE UL BLER'):
    #             overwrite(kpi_name,'CH') 

    #         if(kpi_name=='VoLTE HO Success Rate(IntraFreq)'):
    #             overwrite(kpi_name,'CO')    

    #         if(kpi_name=='PRB utilization rate(%)'):
    #             overwrite(kpi_name,'CV') 

    #         if(kpi_name=='ERAB Setup Success Rate QCI1(%)'):
    #             overwrite(kpi_name,'DC') 

    #         if(kpi_name=='cells SRVCC success rate'):
    #             overwrite(kpi_name,'DJ')   

    #         if(kpi_name=='VoLTE UL PUSCH SINR'):
    #             overwrite(kpi_name,'DQ')  

    #         if(kpi_name=='VoLTE Packet Loss DL(%)_NBH'):
    #             overwrite(kpi_name,'DX')  

    #         if(kpi_name=='VoLTE HO Success Rate(InterFreq)'):
    #             overwrite(kpi_name,'EE')   

    #         if(kpi_name=='VoLTE SRVCC IRAT Per cal rate'):
    #             overwrite(kpi_name,'EL')  

    #         if(kpi_name=='Radio Network Availability Rate(%)'):
    #             overwrite(kpi_name,'ES')  


    # output_path=os.path.join(door_root,"output","upw_trend_output.xlsx")
    # trend_wb.save(output_path)

    # download_path=os.path.join(MEDIA_URL,"trends","UP_West_4G","output","UPW_trend_output.xlsx")
    # return Response({"status":True,"message":"Succesfully uploaded","Download_url":download_path})